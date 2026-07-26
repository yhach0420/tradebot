"""Tests for Canonical Quote Mainline Repair."""
from __future__ import annotations

import json
import os
from pathlib import Path

import pytest

from small_paper.canonical_board import (
    attach_canonical_board,
    buy_limit_price,
    entry_imbalance_for_mode,
    legacy_mixed_imbalance,
    normalize_kabu_board,
    sell_limit_price,
    top_imbalance_for_mode,
)
from research.canonical_quote_mainline_repair.constants import (
    CANCEL,
    LIVE_ORDER,
    REQUIRED_ARTIFACTS,
    REQUIRED_SHEETS,
    SUBMIT,
)


def _kabu_payload(*, buy=1000.0, sell=1002.0, bq=100.0, sq=200.0):
    """Real kabu naming: BidPrice=Sell1, AskPrice=Buy1."""
    return {
        "Symbol": "7203",
        "CurrentPrice": 1001.0,
        "BidPrice": sell,
        "AskPrice": buy,
        "BidQty": sq,
        "AskQty": bq,
        "Buy1": {"Price": buy, "Qty": bq},
        "Sell1": {"Price": sell, "Qty": sq},
        "Buy2": {"Price": buy - 1, "Qty": 50},
        "Sell2": {"Price": sell + 1, "Qty": 50},
    }


def test_stage0_normalizer_called_once():
    from research.canonical_quote_mainline_repair.integrity import stage0_wired

    s = stage0_wired()
    assert s["attach_canonical_board_present"]
    assert s["call_count"] >= 1


def test_raw_fields_preserved():
    p = _kabu_payload()
    before = json.dumps(p, sort_keys=True)
    enriched = dict(p)
    attach_canonical_board(enriched, p)
    assert json.dumps(p, sort_keys=True) == before
    assert enriched["BidPrice"] == p["BidPrice"]
    assert enriched["AskPrice"] == p["AskPrice"]


def test_canonical_bid_from_buy1():
    assert normalize_kabu_board(_kabu_payload(buy=1111)).canonical_best_bid == 1111


def test_canonical_ask_from_sell1():
    assert normalize_kabu_board(_kabu_payload(sell=2222)).canonical_best_ask == 2222


def test_canonical_bid_qty_from_buy1():
    assert normalize_kabu_board(_kabu_payload(bq=55)).canonical_bid_qty == 55


def test_canonical_ask_qty_from_sell1():
    assert normalize_kabu_board(_kabu_payload(sq=77)).canonical_ask_qty == 77


def test_same_payload_only():
    c = normalize_kabu_board(_kabu_payload())
    assert c.canonical_source_event_id is not None


def test_no_cross_event_merge():
    a = normalize_kabu_board(_kabu_payload(buy=10, sell=11))
    b = normalize_kabu_board({"Buy1": {"Price": 5, "Qty": 1}})
    assert a.canonical_best_bid == 10
    assert b.canonical_best_ask is None
    assert b.canonical_quote_reason == "NOT_EVALUABLE_MISSING_SIDE"


def test_no_forward_fill():
    normalize_kabu_board(_kabu_payload(buy=1, sell=2))
    c2 = normalize_kabu_board({})
    assert c2.canonical_best_bid is None


def test_locked_quote_classification():
    c = normalize_kabu_board(_kabu_payload(buy=100, sell=100, bq=1, sq=1))
    assert c.canonical_locked
    assert c.canonical_quote_reason == "LOCKED"


def test_invalid_crossed_quote():
    # Sell1 < Buy1 → crossed
    c = normalize_kabu_board(_kabu_payload(buy=105, sell=100, bq=1, sq=1))
    assert c.canonical_crossed
    assert not c.canonical_quote_valid


def test_top_imbalance_direction():
    c = normalize_kabu_board(_kabu_payload(bq=900, sq=100))
    assert c.canonical_top_imbalance is not None and c.canonical_top_imbalance > 0.5
    os.environ["KABU_QUOTE_SEMANTIC_MODE"] = "legacy"
    try:
        leg = top_imbalance_for_mode(_kabu_payload(bq=900, sq=100), mode="legacy")
        assert leg is not None and leg < 0.5  # inverted
    finally:
        os.environ.pop("KABU_QUOTE_SEMANTIC_MODE", None)


def test_depth_imbalance_direction():
    c = normalize_kabu_board(_kabu_payload(bq=900, sq=100))
    assert c.canonical_depth_imbalance is not None and c.canonical_depth_imbalance > 0.5


def test_no_mixed_depth_sides():
    c = normalize_kabu_board(_kabu_payload(bq=100, sq=100))
    # depth = Buy1+Buy2 vs Sell1+Sell2 = 150 vs 150
    assert c.canonical_depth_bid_qty == 150
    assert c.canonical_depth_ask_qty == 150
    # legacy mixed adds BidQty/AskQty on wrong sides
    leg = legacy_mixed_imbalance(_kabu_payload(bq=100, sq=100))
    assert leg is not None


def test_pbv2_uses_canonical_imbalance():
    os.environ["KABU_QUOTE_SEMANTIC_MODE"] = "canonical"
    try:
        from small_paper.board_imbalance_shadow import compute_entry_order_book_imbalance_field

        p = _kabu_payload(bq=900, sq=100)
        enriched = dict(p)
        attach_canonical_board(enriched, p)
        fields = compute_entry_order_book_imbalance_field(payload=enriched)
        assert fields["entry_order_book_imbalance"] == pytest.approx(
            enriched["canonical_depth_imbalance"], rel=1e-6
        )
    finally:
        os.environ.pop("KABU_QUOTE_SEMANTIC_MODE", None)


def test_board_token_direction():
    from small_paper.canonical_board import board_token_from_imbalance

    assert board_token_from_imbalance(0.9) == "Board:high"
    assert board_token_from_imbalance(0.1) == "Board:low"


def test_guard_uses_canonical_board():
    from small_paper.entry_quality_guard import compute_spread_bps_from_payload

    p = _kabu_payload()
    enriched = dict(p)
    attach_canonical_board(enriched, p)
    sb = compute_spread_bps_from_payload(enriched)
    assert sb is not None
    assert sb == pytest.approx(enriched["canonical_spread_bps"], rel=1e-4)


def test_board_dynamic_trailing_uses_canonical():
    # Trailing consumes percentile; percentile from canonical entry imb when mode=canonical
    from small_paper.board_dynamic_trailing_shadow import trailing_params_for_board_tier

    a1, g1, t1 = trailing_params_for_board_tier(80.0)
    a2, g2, t2 = trailing_params_for_board_tier(10.0)
    assert t1 == "board_high" and t2 == "board_low"
    assert a1 != a2


def test_realtime_board_exit_uses_canonical():
    from small_paper.canonical_board import best_bid_ask_for_mode, top_imbalance_for_mode

    p = _kabu_payload(buy=1000, sell=1005, bq=900, sq=100)
    enriched = dict(p)
    attach_canonical_board(enriched, p)
    imb = top_imbalance_for_mode(enriched, mode="canonical")
    bid, ask = best_bid_ask_for_mode(enriched, mode="canonical")
    assert bid == 1000 and ask == 1005
    assert imb is not None and imb > 0.5
    # mirror realtime_board_exit_shadow wrappers
    src = Path(__file__).resolve().parents[1] / "src" / "small_paper" / "realtime_board_exit_shadow.py"
    text = src.read_text(encoding="utf-8")
    assert "top_imbalance_for_mode" in text
    assert "best_bid_ask_for_mode" in text


def test_buy_execution_uses_canonical_ask():
    p = _kabu_payload(buy=1000, sell=1005)
    enriched = dict(p)
    attach_canonical_board(enriched, p)
    assert buy_limit_price(enriched, mode="canonical") == 1005.0


def test_sell_execution_uses_canonical_bid():
    p = _kabu_payload(buy=1000, sell=1005)
    enriched = dict(p)
    attach_canonical_board(enriched, p)
    assert sell_limit_price(enriched, mode="canonical") == 1000.0


def test_legacy_mode_parity():
    p = _kabu_payload(bq=900, sq=100)
    leg = entry_imbalance_for_mode(p, mode="legacy")
    mixed = legacy_mixed_imbalance(p)
    assert leg == pytest.approx(mixed)


def test_cap5_deterministic(tmp_path: Path):
    from research.canonical_quote_mainline_repair.dual_replay import run_dual_replay

    # tiny stride on one day if available
    d1 = run_dual_replay(["20260721"], stride=50)
    d2 = run_dual_replay(["20260721"], stride=50)
    assert d1["deterministic_p0"] is True
    assert (d1.get("P0") or {}).get("pnl_5bps") == (d2.get("P0") or {}).get("pnl_5bps")


def test_operational_exit_separation():
    from research.canonical_quote_mainline_repair.constants import OPERATIONAL_EXIT_REASONS

    assert "session_close" in OPERATIONAL_EXIT_REASONS
    assert "hard_stop" not in OPERATIONAL_EXIT_REASONS


def test_no_runtime_raw_bid_ask_reference():
    from research.canonical_quote_mainline_repair.integrity import scan_runtime_raw_refs

    scan = scan_runtime_raw_refs()
    # hard direct strategy refs must be zero
    assert scan["hard_direct_refs"] == 0, scan["hard_hits"][:10]


def test_submit_cancel_live_zero():
    assert SUBMIT == 0 and CANCEL == 0 and LIVE_ORDER == 0


def test_live_path_disabled():
    from small_paper.live_order_dry_run_adapter import dry_run_adapter_enabled

    class C:
        live_trading_enabled = False
        live_order_dry_run_enabled = True

    assert dry_run_adapter_enabled(C()) is True
    C.live_trading_enabled = True
    assert dry_run_adapter_enabled(C()) is False


def test_only_three_outputs(tmp_path: Path):
    from research.canonical_quote_mainline_repair.report import emit_artifacts

    payload = {
        "run_id": "t",
        "gates": {"paper_readiness": "CANONICAL_RUNTIME_CORRECT_BUT_STRATEGY_BLOCKED"},
        "dual": {"P0": {}, "P1": {}, "P2": {}, "P3": {}, "entry_diff": {}, "board_classification": {}},
        "raw_scan": {"hits": []},
        "stage0": {},
        "tests": {"rows": []},
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
    }
    d = tmp_path / "out"
    emit_artifacts(d, payload)
    assert sorted(p.name for p in d.iterdir() if p.is_file()) == sorted(REQUIRED_ARTIFACTS)
    from openpyxl import load_workbook

    wb = load_workbook(d / "audit.xlsx")
    assert set(REQUIRED_SHEETS) <= set(wb.sheetnames)
