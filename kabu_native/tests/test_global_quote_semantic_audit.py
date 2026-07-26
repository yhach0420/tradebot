"""Tests for Global Quote Semantic Audit + canonical normalizer."""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from research.global_quote_semantic_audit.canonical import (
    CanonicalBoard,
    normalize_kabu_board,
    r0_current_from_payload,
)
from research.global_quote_semantic_audit.constants import (
    OUT_ROOT,
    REQUIRED_ARTIFACTS,
    REQUIRED_SHEETS,
    SUBMIT,
    CANCEL,
    LIVE_ORDER,
)
from research.global_quote_semantic_audit.pipeline import run_audit
from research.global_quote_semantic_audit.report import write_xlsx


def _payload(*, buy=1000.0, sell=1002.0, bq=100.0, sq=200.0, bid=None, ask=None, bid_qty=None, ask_qty=None):
    """Realistic kabu: BidPrice=Sell1, AskPrice=Buy1."""
    bid = sell if bid is None else bid
    ask = buy if ask is None else ask
    bid_qty = sq if bid_qty is None else bid_qty
    ask_qty = bq if ask_qty is None else ask_qty
    return {
        "BidPrice": bid,
        "AskPrice": ask,
        "BidQty": bid_qty,
        "AskQty": ask_qty,
        "Buy1": {"Price": buy, "Qty": bq},
        "Sell1": {"Price": sell, "Qty": sq},
        "CurrentPrice": (buy + sell) / 2,
    }


def test_kabu_bidprice_maps_to_sell1():
    p = _payload()
    assert p["BidPrice"] == p["Sell1"]["Price"]


def test_kabu_askprice_maps_to_buy1():
    p = _payload()
    assert p["AskPrice"] == p["Buy1"]["Price"]


def test_canonical_bid_from_buy1():
    c = normalize_kabu_board(_payload(buy=1111.0, sell=1113.0))
    assert c.canonical_best_bid == 1111.0


def test_canonical_ask_from_sell1():
    c = normalize_kabu_board(_payload(buy=1111.0, sell=1113.0))
    assert c.canonical_best_ask == 1113.0


def test_canonical_ask_ge_bid():
    c = normalize_kabu_board(_payload())
    assert c.canonical_best_ask is not None and c.canonical_best_bid is not None
    assert c.canonical_best_ask >= c.canonical_best_bid
    assert c.quote_valid


def test_qty_mapping():
    c = normalize_kabu_board(_payload(bq=55.0, sq=77.0))
    assert c.canonical_bid_qty == 55.0
    assert c.canonical_ask_qty == 77.0
    assert c.kabu_bid_qty_raw == 77.0  # BidQty = Sell1 qty
    assert c.kabu_ask_qty_raw == 55.0


def test_no_raw_field_overwrite():
    p = _payload()
    before = json.dumps(p, sort_keys=True)
    _ = normalize_kabu_board(p)
    after = json.dumps(p, sort_keys=True)
    assert before == after


def test_no_cross_event_merge():
    a = normalize_kabu_board(_payload(buy=10.0, sell=11.0))
    b = normalize_kabu_board(_payload(buy=20.0, sell=21.0))
    assert a.canonical_best_bid == 10.0
    assert b.canonical_best_bid == 20.0
    # missing side → NOT_EVALUABLE, not filled from other event
    c = normalize_kabu_board({"Buy1": {"Price": 5.0, "Qty": 1.0}, "BidPrice": 9.0, "AskPrice": 5.0})
    assert c.quote_reason == "NOT_EVALUABLE_MISSING_SIDE"
    assert c.canonical_best_ask is None


def test_no_forward_fill():
    c1 = normalize_kabu_board(_payload(buy=1.0, sell=2.0))
    c2 = normalize_kabu_board({})
    assert c1.canonical_best_bid == 1.0
    assert c2.canonical_best_bid is None
    assert c2.quote_reason == "NOT_EVALUABLE_MISSING_SIDE"


def test_spread_positive():
    c = normalize_kabu_board(_payload(buy=100.0, sell=101.0))
    assert c.canonical_spread == 1.0
    assert c.canonical_spread is not None and c.canonical_spread > 0


def test_imbalance_direction():
    # More bid qty → imb > 0.5
    c = normalize_kabu_board(_payload(bq=900.0, sq=100.0))
    assert c.canonical_imbalance is not None and c.canonical_imbalance > 0.5
    # R0 top uses BidQty as bid → BidQty=sq=100 → imb low (inverted)
    r0 = r0_current_from_payload(_payload(bq=900.0, sq=100.0))
    assert r0["r0_imbalance_top"] is not None and r0["r0_imbalance_top"] < 0.5


def test_board_mid_direction():
    from research.global_quote_semantic_audit.canonical import board_token
    from research.global_quote_semantic_audit.constants import BOARD_P33, BOARD_P66

    c = normalize_kabu_board(_payload(bq=900.0, sq=100.0))
    tok = board_token(c.canonical_imbalance, p33=BOARD_P33, p66=BOARD_P66)
    assert tok == "Board:high"
    r0 = r0_current_from_payload(_payload(bq=900.0, sq=100.0))
    tok0 = board_token(r0["r0_imbalance_top"], p33=BOARD_P33, p66=BOARD_P66)
    assert tok0 == "Board:low"


def test_board_dynamic_trailing_direction():
    # Structural: inverted entry imb → complementary percentile direction
    c = normalize_kabu_board(_payload(bq=800.0, sq=200.0))
    r0 = r0_current_from_payload(_payload(bq=800.0, sq=200.0))
    assert c.canonical_imbalance is not None and r0["r0_imbalance_top"] is not None
    assert abs(c.canonical_imbalance + r0["r0_imbalance_top"] - 1.0) < 1e-9


def test_buy_execution_uses_canonical_ask():
    c = normalize_kabu_board(_payload(buy=1000.0, sell=1005.0))
    buy_px = c.canonical_best_ask
    assert buy_px == 1005.0
    assert buy_px != c.kabu_ask_price_raw  # kabu AskPrice is bid


def test_sell_execution_uses_canonical_bid():
    c = normalize_kabu_board(_payload(buy=1000.0, sell=1005.0))
    sell_px = c.canonical_best_bid
    assert sell_px == 1000.0
    assert sell_px != c.kabu_bid_price_raw


def test_submit_cancel_live_zero():
    assert SUBMIT == 0 and CANCEL == 0 and LIVE_ORDER == 0


def test_mainline_unchanged_during_audit(tmp_path: Path):
    # Running audit must not write outside out_dir research artifacts
    out = run_audit(run_id="test_mainline_guard", out_root=tmp_path, days=())
    assert out["mainline_changed"] is False
    # only 3 artifacts in out dir
    files = sorted(p.name for p in (tmp_path / "test_mainline_guard").iterdir() if p.is_file())
    assert files == sorted(REQUIRED_ARTIFACTS)


def test_only_three_outputs(tmp_path: Path):
    payload = {
        "run_id": "t",
        "verdict": {"final_verdict": "QUOTE_SEMANTIC_MAINLINE_AFFECTED"},
        "static": {"search_inventory": [], "field_semantics": [], "static_references": [], "summary": {}},
        "lineage": {"runtime": [], "paper": [], "research": []},
        "impact": {"pbv2": [], "guard": [], "exit": [], "execution": [], "research": []},
        "r0_r1": {},
        "canonical_spec": [],
        "tests": {"rows": []},
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
        "mainline_changed": False,
    }
    from research.global_quote_semantic_audit.report import emit_artifacts

    d = tmp_path / "out"
    emit_artifacts(d, payload)
    names = sorted(p.name for p in d.iterdir() if p.is_file())
    assert names == sorted(REQUIRED_ARTIFACTS)
    # sheet names present
    from openpyxl import load_workbook

    wb = load_workbook(d / "audit.xlsx")
    assert set(REQUIRED_SHEETS) <= set(wb.sheetnames)


def test_normalize_returns_canonical_board_type():
    assert isinstance(normalize_kabu_board(_payload()), CanonicalBoard)
