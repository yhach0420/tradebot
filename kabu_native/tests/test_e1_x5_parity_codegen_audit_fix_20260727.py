"""Event Parity audit evidence restored — codegen_audit_fix regressions."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
PARITY = REPO / "results" / "research" / "e1_x5_runtime_offline_parity_20260727"
AUDIT_FIX = REPO / "results" / "research" / "e1_x5_runtime_offline_parity_followup_codegen_audit_fix_20260727"
CODEGEN_FIX = REPO / "results" / "research" / "e1_x5_runtime_offline_parity_followup_codegen_fix_20260727"

REQUIRED_SIDE = {
    "raw_log_path",
    "byte_size",
    "raw_file_sha256",
    "canonical_event_sha256",
    "canonicalization_schema",
    "canonicalization_version",
    "record_count",
    "sequence_min",
    "sequence_max",
    "sequence_gap_count",
    "sequence_duplicate_count",
    "sequence_inversion_count",
    "first_event_id",
    "last_event_id",
    "SCORE",
    "NO_EVALUATION",
    "NOT_DUE",
    "feature_hash_present_count",
}


@pytest.fixture(scope="module")
def audit_report():
    if not (AUDIT_FIX / "report.json").is_file():
        pytest.skip("audit_fix artifacts not generated")
    return json.loads((AUDIT_FIX / "report.json").read_text(encoding="utf-8"))


def test_report_md_has_event_parity_required_fields():
    if not (AUDIT_FIX / "report.md").is_file():
        pytest.skip("audit_fix artifacts not generated")
    md = (AUDIT_FIX / "report.md").read_text(encoding="utf-8")
    for key in (
        "raw_file_sha256",
        "canonical_event_sha256",
        "canonical_trade_ledger_sha256",
        "sequence_min",
        "sequence_max",
        "sequence_gap_count",
        "sequence_duplicate_count",
        "sequence_inversion_count",
        "first_event_id",
        "last_event_id",
        "score_mismatch_count",
        "position_mismatch_count",
        "cap_mismatch_count",
        "entry_decision_mismatch_count",
        "exit_decision_mismatch_count",
        "feature_hash_mismatch_count: `N/A`",
        "NOT_COMPARABLE_RECIPE_DIFFERENCE",
    ):
        assert key in md, f"missing in report.md: {key}"


def test_report_json_event_parity_keys(audit_report):
    ep = audit_report["event_parity"]
    assert set(REQUIRED_SIDE).issubset(ep["oracle"].keys())
    assert set(REQUIRED_SIDE).issubset(ep["runtime"].keys())
    comp = ep["comparison"]
    for k in (
        "feature_hash_comparison_status",
        "feature_hash_comparable_count",
        "feature_hash_not_comparable_count",
        "feature_hash_mismatch_count",
        "score_mismatch_count",
        "position_mismatch_count",
        "cap_mismatch_count",
        "entry_decision_mismatch_count",
        "exit_decision_mismatch_count",
        "first_mismatch",
    ):
        assert k in comp
    assert "trade_ledger_parity" in audit_report
    assert "feature_hash_comparison" in audit_report
    assert "score_availability" in audit_report
    assert "entry_funnel_exclusive" in audit_report
    assert "no_evaluation_reason_breakdown" in audit_report


def test_audit_xlsx_event_parity_all_fields():
    if not (AUDIT_FIX / "audit.xlsx").is_file():
        pytest.skip("audit_fix artifacts not generated")
    wb = load_workbook(AUDIT_FIX / "audit.xlsx")
    assert "Event Parity" in wb.sheetnames
    ws = wb["Event Parity"]
    fields = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    for key in REQUIRED_SIDE:
        assert key in fields, f"Event Parity sheet missing {key}"
    for key in (
        "comparison.score_mismatch_count",
        "comparison.position_mismatch_count",
        "comparison.cap_mismatch_count",
        "comparison.entry_decision_mismatch_count",
        "comparison.exit_decision_mismatch_count",
        "comparison.feature_hash_mismatch_count",
    ):
        assert key in fields
    # mismatch display N/A
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "comparison.feature_hash_mismatch_count":
            assert ws.cell(r, 2).value in ("N/A", None) or str(ws.cell(r, 2).value) == "N/A"
            break
    else:
        pytest.fail("comparison.feature_hash_mismatch_count row missing")
    for name in (
        "Trade Ledger Parity",
        "Feature Hash",
        "Score Availability Audit",
        "Gate Funnel",
        "NO EVALUATION Reasons",
        "Forward Provenance",
        "Safety Regression",
    ):
        assert name in wb.sheetnames


def test_event_parity_regenerated_from_logs(audit_report):
    from small_paper.e1_x5_parity_audit import compare_event_streams, build_event_parity_sections
    from small_paper.e1_x5_canonical_feature_hash import (
        LEGACY_ORACLE_FEATURE_HASH_SCHEMA,
        LEGACY_RUNTIME_FEATURE_HASH_SCHEMA,
    )

    raw = compare_event_streams(
        PARITY / "oracle_baseline" / "oracle_events.jsonl",
        PARITY / "runtime_e1_x5_event_log.jsonl",
        oracle_feature_hash_schema=LEGACY_ORACLE_FEATURE_HASH_SCHEMA,
        runtime_feature_hash_schema=LEGACY_RUNTIME_FEATURE_HASH_SCHEMA,
    )
    sections = build_event_parity_sections(raw)
    assert sections["oracle"]["raw_file_sha256"] == audit_report["event_parity"]["oracle"]["raw_file_sha256"]
    assert sections["runtime"]["canonical_event_sha256"] == audit_report["event_parity"]["runtime"][
        "canonical_event_sha256"
    ]
    assert sections["oracle"]["record_count"] == audit_report["event_parity"]["oracle"]["record_count"]


def test_sha_names_separated(audit_report):
    o = audit_report["event_parity"]["oracle"]
    t = audit_report["trade_ledger_parity"]
    assert o["raw_file_sha256"] != o["canonical_event_sha256"]
    assert "raw_file_sha256" in o and "canonical_event_sha256" in o
    assert "canonical_trade_ledger_sha256_oracle" in t
    assert t["canonical_trade_ledger_sha256_oracle"] == t["canonical_trade_ledger_sha256_runtime"]


def test_sequence_and_event_ids(audit_report):
    for side in ("oracle", "runtime"):
        s = audit_report["event_parity"][side]
        assert s["sequence_min"] is not None
        assert s["sequence_max"] is not None
        assert s["sequence_max"] >= s["sequence_min"]
        assert "sequence_gap_count" in s
        assert "sequence_duplicate_count" in s
        assert "sequence_inversion_count" in s
        assert s["first_event_id"]
        assert s["last_event_id"]


def test_feature_hash_na_and_decision_separated(audit_report):
    fh = audit_report["feature_hash_comparison"]
    comp = audit_report["event_parity"]["comparison"]
    assert fh["feature_hash_comparison_status"] == "NOT_COMPARABLE_RECIPE_DIFFERENCE"
    assert fh["feature_hash_mismatch_count"] is None
    assert fh["feature_hash_mismatch_display"] == "N/A"
    assert comp["score_mismatch_count"] == 0
    assert comp["position_mismatch_count"] == 0
    assert comp["cap_mismatch_count"] == 0
    assert comp["entry_decision_mismatch_count"] == 0
    assert comp["exit_decision_mismatch_count"] == 0
    # Must not claim feature parity PASS
    assert "PASS" not in str(fh["feature_hash_comparison_status"])


def test_blocked_when_runtime_missing(tmp_path):
    from small_paper.e1_x5_parity_audit import compare_event_streams, PARITY_AUDIT_BLOCKED

    oracle = tmp_path / "o.jsonl"
    oracle.write_text(
        '{"event_id":"a","observe_kind":"SCORE","score":1.0,"feature_hash":"x"}\n',
        encoding="utf-8",
    )
    out = compare_event_streams(oracle, tmp_path / "missing.jsonl")
    assert out["status"] == PARITY_AUDIT_BLOCKED
    assert out["score_mismatch"] is None
    assert out["feature_hash_mismatch_count"] is None


def test_funnel_and_noe(audit_report):
    funnel = audit_report["entry_funnel_exclusive"]
    assert funnel["terminal_sum"] == 17353
    assert "no_evaluation" not in funnel
    assert audit_report["no_evaluation_reason_breakdown"]["TICK_BUILD_FAILED"] == 308


def test_excel_column_widths_not_clipping_headers():
    if not (AUDIT_FIX / "audit.xlsx").is_file():
        pytest.skip("audit_fix artifacts not generated")
    wb = load_workbook(AUDIT_FIX / "audit.xlsx")
    ws = wb["Event Parity"]
    # field column should be wide enough for longest required header
    width = ws.column_dimensions["A"].width or 0
    assert width >= len("feature_hash_present_count")
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref


def test_forward_and_regressions(audit_report):
    assert audit_report["verdict_parity"] == "E1_X5_RUNTIME_OFFLINE_PARITY_FIXED"
    assert audit_report["verdict_forward"] == "E1_X5_FORWARD_DAY1_READY"
    assert audit_report["forward_provenance"]["pm_forward_status"] == "NOT_ADOPTED"
    assert audit_report["forward_provenance"]["valid_progress_sessions"] == 0
    assert audit_report["forward_provenance"]["valid_progress_trades"] == 0
    tl = audit_report["trade_ledger_parity"]
    assert tl["trades"] == 70
    assert abs(tl["net_pnl_yen_100"] - 45023.825) < 0.01
    snap = tl["snap_1240"]
    assert snap["entries"] == 19 and snap["completed"] == 15 and snap["open"] == 4
    assert abs(float(snap["pnl"]) - 17275.85) < 0.01
    assert tl["match"] is True
    assert audit_report["pbv2_impact"]["regression_diff"] == 0
    assert audit_report["submit_cancel_live"] == "0/0/0"
    assert audit_report["score_availability"]["am_score_state"] == "UNVERIFIED_PENDING_NEW_AM_PAPER"


def test_does_not_overwrite_prior_artifacts(audit_report):
    # Prior dirs still exist and were not replaced by this OUT
    assert (PARITY / "report.json").is_file()
    assert (CODEGEN_FIX / "report.json").is_file()
    assert AUDIT_FIX.resolve() != PARITY.resolve()
    assert AUDIT_FIX.resolve() != CODEGEN_FIX.resolve()
    files = {p.name for p in AUDIT_FIX.iterdir() if p.is_file()}
    assert files == {"report.md", "report.json", "audit.xlsx"}
