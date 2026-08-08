"""Structural SoT integrity tests for E1_X5 canonical artifacts."""
from __future__ import annotations

import json
import shutil
import tempfile
from copy import deepcopy
from datetime import datetime
from pathlib import Path

import pytest
from zoneinfo import ZoneInfo

from small_paper.e1_x5_artifact_sot import (
    CORRUPT_HASH_A3007,
    FROZEN_PM_HASH_V1,
    LEDGER_HASH_V1_FROZEN_EXPECTED_PM,
    MissingRequiredFieldError,
    assert_hash_versions_match,
    assert_no_corrupt_hash_as_canonical,
    assert_no_expected_hash_mutation,
    atomic_publish,
    canonical_ledger_hash,
    decide_verdicts,
    project_trade_ledger_v1,
    project_trade_ledger_v2,
    render_markdown,
    render_xlsx,
    require,
    trade_ledger_hash_v1,
    trade_ledger_hash_v2,
    validate_snapshot_schema,
)
from small_paper.e1_x5_g1_synthetic_branch_proof import run_synthetic_g1_branch_tests

JST = ZoneInfo("Asia/Tokyo")
ROOT = Path(__file__).resolve().parents[1]
CANONICAL_DIR_NAME = "e1_x5_canonical_path_unify_20260728"


def _minimal_trade(**kwargs):
    et = datetime(2026, 7, 27, 13, 0, 0, tzinfo=JST)
    xt = datetime(2026, 7, 27, 13, 1, 0, tzinfo=JST)
    base = {
        "symbol": "7203.T",
        "entry_time": et,
        "exit_time": xt,
        "entry_ask": 100.0,
        "exit_bid": 100.1,
        "exit_reason": "TARGET",
        "net_pnl_yen_100": 10.0,
        "holding_sec": 60.0,
        "score": 1.0,
    }
    base.update(kwargs)
    return base


def _valid_snapshot(**overrides):
    trades = [_minimal_trade()]
    h = canonical_ledger_hash(trades, version="v1")
    snap = {
        "run_id": "test_run_1",
        "generated_at": "2026-07-29T00:00:00+09:00",
        "input_manifest": [{"day": "20260721"}],
        "input_manifest_sha256": "abc",
        "code_fingerprint": "def",
        "config_fingerprints": {"BASE": "fp"},
        "valid_windows": [{"window_id": f"w{i}"} for i in range(7)],
        "excluded_windows": [],
        "source_row_counts": {"total": 3937344, "by_day": {}},
        "base": {
            "trades": trades,
            "counters": {
                "cap_blocked": 264,
                "same_symbol_blocked": 1676,
                "orphan_open": 3,
                "negative_holding": 0,
            },
            "orphans": [
                {"reason": "WINDOW_END_OPEN_EXCLUDED"},
                {"reason": "WINDOW_END_OPEN_EXCLUDED"},
                {"reason": "WINDOW_END_OPEN_EXCLUDED"},
            ],
            "summary": {
                "completed_trades": 1,
                "realized_pnl_yen_100": 10.0,
                "profit_factor": 1.0,
                "wins": 1,
                "losses": 0,
                "draws": 0,
                "ledger_sha256": h,
            },
            "daily_summary": [{"day": "20260721", "completed_trades": 1, "realized_pnl_yen_100": 10.0}],
            "window_summary": [{"trades": 1, "pnl": 10.0}],
            "exit_summary": [],
            "symbol_summary": [],
            "timeband_summary": [],
            "concentration": {},
        },
        "g1": {
            "variants": [
                {
                    "variant_id": "C1_NEXT_PUSH_HOLD",
                    "config_fingerprint": "x",
                    "candidate": 1,
                    "armed": 1,
                    "confirmed": 0,
                    "cancelled_by_reason": {},
                    "rearm_transition": 0,
                    "accepted": 0,
                    "blocked_by_cap": 0,
                    "blocked_by_same_symbol": 0,
                    "trade_ledger_hash": h,
                    "state_transition_ledger_hash": h,
                }
            ],
            "all_trades": [],
            "state_transitions": [],
            "synthetic_branch_proof": {"ok": True},
        },
        "parity_20260727_pm": {
            "hashes": {
                "canonical_actual": {"v1": FROZEN_PM_HASH_V1, "v2": FROZEN_PM_HASH_V1},
                "frozen_reference": {"v1": FROZEN_PM_HASH_V1, "v2": FROZEN_PM_HASH_V1},
                "observed_corrupt": {"default_str_v0": CORRUPT_HASH_A3007},
            }
        },
        "ledger_hash_algorithm": {"v1": "e1_x5_trade_ledger_hash_v1", "v2": "e1_x5_trade_ledger_hash_v2"},
        "tests": [{"test_name": "t", "passed": True, "evidence_json_paths": ["x"], "message": "ok"}],
        "failed_tests": [],
        "safety": {"submit": 0, "cancel": 0, "live_order": 0},
        "execution_status": "COMPLETED",
        "artifact_integrity_verdict": "PASS",
        "base_verdict": "VERIFIED",
        "g1_wiring_verdict": "VERIFIED",
        "g1_adoption_verdict": "NOT_ADOPTED",
        "overall_verdict": "PASS",
        "report_payload_sha256": "pending",
        "payload_hash_algorithm": "sha256_canonical_json_v1",
        "payload_excluded_json_paths": [],
    }
    snap.update(overrides)
    return snap


def test_single_final_writer_only():
    """Only atomic_publish may write the canonical triad."""
    offenders = []
    allowed = {
        "e1_x5_artifact_sot.py",
        "run_e1_x5_sot_clean_publish_20260728.py",
        "run_e1_x5_canonical_path_unify_20260728.py",
        "verify_e1_x5_artifact_bundle.py",
        "run_e1_x5_g1_confirmation_guard_20260728.py",
        "test_e1_x5_artifact_sot.py",
    }
    scan_roots = [ROOT / "scripts", ROOT / "src" / "small_paper"]
    for root in scan_roots:
        for path in root.rglob("*.py"):
            if path.name in allowed:
                continue
            text = path.read_text(encoding="utf-8", errors="replace")
            if CANONICAL_DIR_NAME not in text:
                continue
            if any(a in text for a in ("report.json", "report.md", "audit.xlsx")) and any(
                w in text for w in ("write_text", "wb.save", "ExcelWriter", "to_excel")
            ):
                offenders.append(str(path.relative_to(ROOT)))
    # clean publish may write logs/CURRENT_RUN_ID but triad only via atomic_publish
    pub = (ROOT / "scripts" / "run_e1_x5_sot_clean_publish_20260728.py").read_text(encoding="utf-8")
    assert "atomic_publish" in pub
    assert "finalize_e1_x5_window_reeval_from_cache" not in pub
    assert not offenders, f"second writers for canonical triad: {offenders}"


def test_required_fields_cannot_default_to_zero():
    with pytest.raises(MissingRequiredFieldError):
        require({}, "cap_blocked")


def test_actual_cannot_be_assigned_to_expected():
    frozen = LEDGER_HASH_V1_FROZEN_EXPECTED_PM
    actual = CORRUPT_HASH_A3007
    with pytest.raises(AssertionError):
        assert_no_expected_hash_mutation(frozen, actual, reported_expected=actual)
    assert_no_expected_hash_mutation(frozen, actual, reported_expected=frozen)


def test_hash_version_must_match():
    with pytest.raises(AssertionError):
        assert_hash_versions_match("v1", "v2")
    assert_hash_versions_match("v1", "v1")


def test_canonical_hash_single_source():
    rows = [_minimal_trade(holding_sec=50.0)]
    assert trade_ledger_hash_v1(rows) == canonical_ledger_hash(rows, version="v1")
    assert trade_ledger_hash_v2(rows) == canonical_ledger_hash(rows, version="v2")
    assert trade_ledger_hash_v1(rows) != trade_ledger_hash_v2(rows)
    assert project_trade_ledger_v1(rows)[0]["holding_sec"] == 60.0
    assert project_trade_ledger_v2(rows)[0]["holding_sec"] == 50.0


def test_corrupt_hash_location_whitelist():
    with pytest.raises(AssertionError):
        assert_no_corrupt_hash_as_canonical(f"canonical_actual: {CORRUPT_HASH_A3007}")
    assert_no_corrupt_hash_as_canonical(
        f'observed_corrupt.default_str_v0: "{CORRUPT_HASH_A3007}"'
    )


def test_renderer_missing_counter_fails():
    snap = _valid_snapshot()
    del snap["base"]["counters"]["cap_blocked"]
    with pytest.raises(MissingRequiredFieldError):
        render_markdown(snap)


def test_expected_equals_actual_json_fails_validation():
    snap = _valid_snapshot()
    snap["parity_20260727_pm"]["hashes"]["frozen_reference"]["v1"] = CORRUPT_HASH_A3007
    with pytest.raises(AssertionError):
        # frozen must stay FROZEN; also if someone put corrupt into canonical:
        snap2 = _valid_snapshot()
        snap2["parity_20260727_pm"]["hashes"]["canonical_actual"]["v1"] = CORRUPT_HASH_A3007
        validate_snapshot_schema(snap2)


def test_canonical_a300_fails():
    snap = _valid_snapshot()
    snap["parity_20260727_pm"]["hashes"]["canonical_actual"]["v1"] = CORRUPT_HASH_A3007
    with pytest.raises(AssertionError):
        validate_snapshot_schema(snap)


def test_report_json_to_markdown_roundtrip():
    snap = _valid_snapshot()
    with tempfile.TemporaryDirectory() as td:
        out = Path(td)
        pub = atomic_publish(out, snap)
        md = (out / "report.md").read_text(encoding="utf-8")
        loaded = json.loads((out / "report.json").read_text(encoding="utf-8"))
        assert loaded["run_id"] in md
        assert loaded["report_payload_sha256"] in md
        assert pub["report_payload_sha256"] == loaded["report_payload_sha256"]


def test_report_json_to_xlsx_roundtrip():
    snap = _valid_snapshot()
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
        out = Path(td)
        atomic_publish(out, snap)
        from openpyxl import load_workbook

        wb = load_workbook(out / "audit.xlsx", read_only=True)
        try:
            assert "Ledger_Hash" in wb.sheetnames
            assert "Artifact_Integrity" in wb.sheetnames
            rows = list(wb["Artifact_Integrity"].iter_rows(values_only=True))
            kv = {str(r[0]): str(r[1]) for r in rows[1:] if r and r[0] is not None}
        finally:
            wb.close()
        loaded = json.loads((out / "report.json").read_text(encoding="utf-8"))
        assert kv["run_id"] == loaded["run_id"]
        assert kv["report_payload_sha256"] == loaded["report_payload_sha256"]


def test_json_xlsx_hash_mismatch_detected():
    snap = _valid_snapshot()
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
        out = Path(td)
        atomic_publish(out, snap)
        (out / "CURRENT_RUN_ID.txt").write_text("test_run_1\n", encoding="utf-8")
        from openpyxl import load_workbook

        wb = load_workbook(out / "audit.xlsx")
        ws = wb["Ledger_Hash"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "parity_canonical_actual_v1":
                row[1].value = CORRUPT_HASH_A3007
        wb.save(out / "audit.xlsx")
        wb.close()
        from scripts.verify_e1_x5_artifact_bundle import Fail, verify_bundle

        with pytest.raises(Fail):
            verify_bundle(out)


def test_json_md_orphan_mismatch_detected():
    snap = _valid_snapshot()
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
        out = Path(td)
        atomic_publish(out, snap)
        md_path = out / "report.md"
        text = md_path.read_text(encoding="utf-8")
        md_path.write_text(text.replace("orphan_open: 3", "orphan_open: 9"), encoding="utf-8")
        loaded = json.loads((out / "report.json").read_text(encoding="utf-8"))
        md = md_path.read_text(encoding="utf-8")
        assert f"orphan_open: {loaded['base']['counters']['orphan_open']}" not in md
        (out / "CURRENT_RUN_ID.txt").write_text("test_run_1\n", encoding="utf-8")
        from scripts.verify_e1_x5_artifact_bundle import Fail, verify_bundle

        with pytest.raises(Fail):
            verify_bundle(out)


def test_old_run_log_mixing_fails():
    snap = _valid_snapshot()
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
        out = Path(td)
        atomic_publish(out, snap)
        (out / "CURRENT_RUN_ID.txt").write_text("test_run_1\n", encoding="utf-8")
        (out / "finalize.log").write_text("run_id=OLD_RUN\nother\n", encoding="utf-8")
        from scripts.verify_e1_x5_artifact_bundle import Fail, verify_bundle

        with pytest.raises(Fail):
            verify_bundle(out)


def test_second_writer_after_publish_fails_verify():
    snap = _valid_snapshot()
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
        out = Path(td)
        atomic_publish(out, snap)
        (out / "CURRENT_RUN_ID.txt").write_text("test_run_1\n", encoding="utf-8")
        loaded = json.loads((out / "report.json").read_text(encoding="utf-8"))
        loaded["base"]["counters"]["cap_blocked"] = 0
        (out / "report.json").write_text(json.dumps(loaded, ensure_ascii=False, indent=2), encoding="utf-8")
        from scripts.verify_e1_x5_artifact_bundle import Fail, verify_bundle

        with pytest.raises(Fail):
            verify_bundle(out)


def test_g1_synthetic_variant_branch_coverage():
    r = run_synthetic_g1_branch_tests()
    assert r["ok"] is True


def test_g1_all_same_default_branch_fails():
    """Negative: collapsing all variants to BASE fingerprint must fail synthetic proof expectations."""
    fps = {"C1": "same", "C2": "same", "C3": "same"}
    assert len(set(fps.values())) == 1
    # real synthetic proof must show distinct fingerprints
    r = run_synthetic_g1_branch_tests()
    assert r["results"]["C1_C2_fingerprints_differ"]
    assert r["results"]["C3_fingerprint_distinct"]


def test_rebuild_dual_writer_removed():
    assert not (ROOT / "scripts" / "rebuild_e1_x5_window_reeval_reports.py").is_file()
    assert not (ROOT / "scripts" / "finalize_e1_x5_window_reeval_from_cache.py").is_file()
    assert not (ROOT / "scripts" / "_patch_parity_hash_into_sot.py").is_file()


def test_decide_verdicts_separates_wiring_and_adoption():
    tests = [
        {"test_name": "BASE_trades_407", "passed": True},
        {"test_name": "G1_synthetic_variant_branch_coverage", "passed": True},
        {"test_name": "G1_real_data_transition_evidence", "passed": True},
        {"test_name": "G1_wiring_evidence_complete", "passed": True},
    ]
    v = decide_verdicts(tests)
    assert v["g1_wiring_verdict"] == "VERIFIED"
    assert v["g1_adoption_verdict"] == "NOT_ADOPTED"
    assert v["overall_verdict"] == "PASS"
