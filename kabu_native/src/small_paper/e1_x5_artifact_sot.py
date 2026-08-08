"""E1_X5 FinalRunSnapshot SoT — single writer, pure renderers, no default fills.

Canonical publish path: atomic_publish() ONLY for
results/research/e1_x5_canonical_path_unify_20260728/{report.json,report.md,audit.xlsx}
"""
from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence
from zoneinfo import ZoneInfo

JST = ZoneInfo("Asia/Tokyo")

MISSING_REQUIRED_FIELD = "MISSING_REQUIRED_FIELD"
CANONICAL_OUT_DIR_NAME = "e1_x5_canonical_path_unify_20260728"
ARTIFACT_NAMES = ("report.json", "report.md", "audit.xlsx")

FROZEN_PM_HASH_V1 = "b5837b4871273aad64445e76c251a3bc72ff6aa98c41107c04dffaefe04ef2d4"
CORRUPT_HASH_A3007 = "a3007cbc11ec0630645b2e89f559ae42aeb342bf840608320c427ed918b84649"
HASH_SCHEMA_V1 = "e1_x5_trade_ledger_hash_v1"
HASH_SCHEMA_V2 = "e1_x5_trade_ledger_hash_v2"

PAYLOAD_EXCLUDED_JSON_PATHS = [
    "generated_at",
    "report_payload_sha256",
    "payload_hash_algorithm",
    "payload_excluded_json_paths",
]

REQUIRED_TOP_LEVEL = [
    "run_id",
    "generated_at",
    "input_manifest",
    "input_manifest_sha256",
    "code_fingerprint",
    "config_fingerprints",
    "valid_windows",
    "excluded_windows",
    "source_row_counts",
    "base",
    "g1",
    "parity_20260727_pm",
    "ledger_hash_algorithm",
    "tests",
    "failed_tests",
    "safety",
    "execution_status",
    "artifact_integrity_verdict",
    "base_verdict",
    "g1_wiring_verdict",
    "g1_adoption_verdict",
    "overall_verdict",
    "report_payload_sha256",
    "payload_hash_algorithm",
    "payload_excluded_json_paths",
]


class MissingRequiredFieldError(ValueError):
    pass


def require(obj: Mapping[str, Any], key: str) -> Any:
    if key not in obj:
        raise MissingRequiredFieldError(f"MISSING_REQUIRED_FIELD:{key}")
    v = obj[key]
    if v is MISSING_REQUIRED_FIELD or v == MISSING_REQUIRED_FIELD:
        raise MissingRequiredFieldError(f"MISSING_REQUIRED_FIELD:{key}")
    return v


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            chunk = fh.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _iso(v: Any) -> str:
    if isinstance(v, datetime):
        dt = v if v.tzinfo else v.replace(tzinfo=JST)
        # Match frozen PM ledger: datetime.isoformat() without forced milliseconds
        # (avoids ".000" which would diverge from b5837… frozen reference).
        return dt.astimezone(JST).isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return ""
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00").replace(" ", "T", 1))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=JST)
            return dt.astimezone(JST).isoformat()
        except ValueError:
            return s.replace("Z", "+00:00")
    if v is None or v == "":
        return ""
    return str(v).replace("Z", "+00:00")


def project_trade_row(r: Mapping[str, Any], *, recompute_holding: bool) -> dict[str, Any]:
    et = r.get("entry_time")
    xt = r.get("exit_time")
    if "holding_sec" not in r:
        raise MissingRequiredFieldError("MISSING_REQUIRED_FIELD:holding_sec")
    hold = float(r["holding_sec"])
    if recompute_holding and hasattr(et, "timestamp") and hasattr(xt, "timestamp"):
        hold = (xt - et).total_seconds()
    return {
        "symbol": str(r.get("symbol") or ""),
        "entry_time": _iso(et),
        "exit_time": _iso(xt),
        "entry_ask": round(float(r.get("entry_ask") or 0), 6),
        "exit_bid": round(float(r.get("exit_bid") or 0), 6),
        "exit_reason": str(r.get("exit_reason") or ""),
        "net_pnl_yen_100": round(float(r.get("net_pnl_yen_100") or 0), 6),
        "holding_sec": round(float(hold), 6),
        "score": round(float(r.get("score") or 0), 12),
    }


def project_ledger(rows: Sequence[Mapping[str, Any]], *, version: str) -> list[dict[str, Any]]:
    recompute = version == "v1"
    canon = [project_trade_row(r, recompute_holding=recompute) for r in rows]
    canon.sort(key=lambda x: (x["entry_time"], x["symbol"], x["exit_time"]))
    return canon


def canonical_ledger_hash(rows: Sequence[Mapping[str, Any]], *, version: str = "v1") -> str:
    """Single canonical ledger hash. Never uses json default=str on raw datetimes."""
    canon = project_ledger(rows, version=version)
    raw = json.dumps(canon, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return sha256_bytes(raw.encode("utf-8"))


def strip_excluded(obj: Any, path: str = "") -> Any:
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            p = f"{path}.{k}" if path else k
            if p in PAYLOAD_EXCLUDED_JSON_PATHS or k in PAYLOAD_EXCLUDED_JSON_PATHS:
                continue
            out[k] = strip_excluded(v, p)
        return out
    if isinstance(obj, list):
        return [strip_excluded(x, path) for x in obj]
    return obj


def compute_report_payload_sha256(snapshot: Mapping[str, Any]) -> str:
    body = strip_excluded(deepcopy(dict(snapshot)))
    raw = json.dumps(body, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=_json_default)
    return sha256_bytes(raw.encode("utf-8"))


def assert_no_corrupt_hash_as_canonical(text: str) -> None:
    if CORRUPT_HASH_A3007 not in text:
        return
    for line in text.splitlines():
        if CORRUPT_HASH_A3007 not in line:
            continue
        low = line.lower()
        if "observed_corrupt" in low or "default_str_v0" in low:
            continue
        raise AssertionError(f"corrupt hash a3007 used as canonical: {line[:240]}")


def validate_snapshot_schema(snap: Mapping[str, Any]) -> None:
    for k in REQUIRED_TOP_LEVEL:
        require(snap, k)
    base = require(snap, "base")
    for k in (
        "trades",
        "counters",
        "orphans",
        "summary",
        "daily_summary",
        "window_summary",
        "exit_summary",
        "symbol_summary",
        "timeband_summary",
        "concentration",
    ):
        require(base, k)
    counters = require(base, "counters")
    for k in ("cap_blocked", "same_symbol_blocked", "orphan_open", "negative_holding"):
        require(counters, k)
    g1 = require(snap, "g1")
    for k in ("variants", "all_trades", "state_transitions", "synthetic_branch_proof"):
        require(g1, k)
    parity = require(snap, "parity_20260727_pm")
    hashes = require(parity, "hashes")
    for k in ("canonical_actual", "frozen_reference", "observed_corrupt"):
        require(hashes, k)
    ca = require(hashes, "canonical_actual")
    require(ca, "v1")
    require(ca, "v2")
    if ca["v1"] == CORRUPT_HASH_A3007 or ca["v2"] == CORRUPT_HASH_A3007:
        raise AssertionError("canonical_actual must not be corrupt a3007")
    fr = require(hashes, "frozen_reference")
    require(fr, "v1")
    oc = require(hashes, "observed_corrupt")
    require(oc, "default_str_v0")
    if oc["default_str_v0"] != CORRUPT_HASH_A3007:
        raise AssertionError("observed_corrupt.default_str_v0 must be a3007")


def assert_no_expected_hash_mutation(frozen: str, actual: str, *, reported_expected: str) -> None:
    """Forbid copying actual into expected/frozen slots."""
    if reported_expected == actual and actual != frozen:
        raise AssertionError("actual_cannot_be_assigned_to_expected")
    if reported_expected != frozen and reported_expected == CORRUPT_HASH_A3007:
        raise AssertionError("corrupt hash cannot be frozen expected")


def assert_hash_versions_match(version_a: str, version_b: str) -> None:
    if version_a != version_b:
        raise AssertionError(f"hash_version_must_match:{version_a}!={version_b}")


def decide_verdicts(tests: Sequence[Mapping[str, Any]]) -> dict[str, str]:
    failed = [t["test_name"] for t in tests if not t.get("passed")]
    integrity_names = [
        n
        for n in failed
        if not n.startswith("g1_adoption")
    ]
    artifact_integrity_verdict = "FAIL" if integrity_names else "PASS"
    base_verdict = (
        "VERIFIED"
        if not any(
            n.startswith("BASE_")
            or n in ("all_7_windows_double_replay", "standalone_BASE_equals_G1_guard_off_BASE")
            for n in failed
        )
        else "FAIL"
    )
    g1_wiring_verdict = (
        "VERIFIED"
        if not any(
            n.startswith("G1_real")
            or n.startswith("G1_synthetic")
            or n == "G1_wiring_evidence_complete"
            for n in failed
        )
        else "FAIL"
    )
    overall = "PASS" if not failed else "BLOCKED"
    return {
        "execution_status": "COMPLETED",
        "artifact_integrity_verdict": artifact_integrity_verdict,
        "base_verdict": base_verdict,
        "g1_wiring_verdict": g1_wiring_verdict,
        "g1_adoption_verdict": "NOT_ADOPTED",
        "overall_verdict": overall,
        "failed_tests": failed,
    }


def render_markdown(snap: Mapping[str, Any]) -> str:
    validate_snapshot_schema(snap)
    base = snap["base"]
    parity = snap["parity_20260727_pm"]
    lines = [
        f"# {snap['overall_verdict']}",
        "",
        f"run_id: `{snap['run_id']}`",
        f"report_payload_sha256: `{snap['report_payload_sha256']}`",
        f"generated_at: `{snap['generated_at']}`",
        f"execution_status: `{snap['execution_status']}`",
        f"artifact_integrity_verdict: `{snap['artifact_integrity_verdict']}`",
        f"base_verdict: `{snap['base_verdict']}`",
        f"g1_wiring_verdict: `{snap['g1_wiring_verdict']}`",
        f"g1_adoption_verdict: `{snap['g1_adoption_verdict']}`",
        f"overall_verdict: `{snap['overall_verdict']}`",
        "",
        "## BASE",
        f"- trades: {base['summary']['completed_trades']}",
        f"- pnl: {base['summary']['realized_pnl_yen_100']}",
        f"- pf: {base['summary']['profit_factor']}",
        f"- cap_blocked: {base['counters']['cap_blocked']}",
        f"- same_symbol_blocked: {base['counters']['same_symbol_blocked']}",
        f"- orphan_open: {base['counters']['orphan_open']}",
        "",
        "## Parity 20260727 PM hashes",
        f"- canonical_actual.v1: `{parity['hashes']['canonical_actual']['v1']}`",
        f"- canonical_actual.v2: `{parity['hashes']['canonical_actual']['v2']}`",
        f"- frozen_reference.v1: `{parity['hashes']['frozen_reference']['v1']}`",
        f"- observed_corrupt.default_str_v0: `{parity['hashes']['observed_corrupt']['default_str_v0']}`",
        "",
        "## G1 wiring",
        "```json",
        json.dumps(snap["g1"]["variants"], ensure_ascii=False, indent=2),
        "```",
        "",
        "## Tests",
        f"failed_tests={snap['failed_tests']}",
        "```json",
        json.dumps(snap["tests"], ensure_ascii=False, indent=2),
        "```",
        "",
        "## Safety",
        json.dumps(snap["safety"], ensure_ascii=False),
        "",
    ]
    return "\n".join(lines) + "\n"


def render_xlsx(path: Path, snap: Mapping[str, Any]) -> None:
    from openpyxl import Workbook

    validate_snapshot_schema(snap)

    def cell(v: Any) -> Any:
        if v is None or isinstance(v, (str, int, float, bool)):
            return v
        return json.dumps(v, ensure_ascii=False, default=str)

    wb = Workbook()
    wb.remove(wb.active)
    meta = {
        "run_id": snap["run_id"],
        "report_payload_sha256": snap["report_payload_sha256"],
        "generated_at": snap["generated_at"],
        "overall_verdict": snap["overall_verdict"],
        "artifact_integrity_verdict": snap["artifact_integrity_verdict"],
        "base_verdict": snap["base_verdict"],
        "g1_wiring_verdict": snap["g1_wiring_verdict"],
        "g1_adoption_verdict": snap["g1_adoption_verdict"],
    }
    base = require(snap, "base")
    g1 = require(snap, "g1")
    parity = require(snap, "parity_20260727_pm")
    hashes = require(parity, "hashes")
    sheets = {
        "Artifact_Integrity": meta,
        "Source_Manifest": require(snap, "input_manifest"),
        "Valid_Windows": require(snap, "valid_windows"),
        "Excluded_Windows": require(snap, "excluded_windows"),
        "BASE_Trade_Ledger": require(base, "trades"),
        "G1_Trade_Ledger": require(g1, "all_trades"),
        "G1_State_Transitions": require(g1, "state_transitions"),
        "Daily_Summary": require(base, "daily_summary"),
        "Window_Summary": require(base, "window_summary"),
        "Config_Comparison": require(g1, "variants"),
        "Exit_Summary": require(base, "exit_summary"),
        "Symbol_Summary": require(base, "symbol_summary"),
        "Timeband_Summary": require(base, "timeband_summary"),
        "Concentration": require(base, "concentration"),
        "Ledger_Hash": {
            "algorithm": require(snap, "ledger_hash_algorithm"),
            "parity_canonical_actual_v1": require(require(hashes, "canonical_actual"), "v1"),
            "parity_canonical_actual_v2": require(require(hashes, "canonical_actual"), "v2"),
            "parity_frozen_reference_v1": require(require(hashes, "frozen_reference"), "v1"),
            "parity_observed_corrupt_default_str_v0": require(require(hashes, "observed_corrupt"), "default_str_v0"),
            "BASE": require(require(base, "summary"), "ledger_sha256"),
        },
        "Tests": require(snap, "tests"),
        "Safety": require(snap, "safety"),
        "Orphans": require(base, "orphans"),
    }
    for name, data in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if isinstance(data, list):
            if not data:
                ws.append(["empty"])
                continue
            if isinstance(data[0], dict):
                keys: list[str] = []
                for row in data:
                    for k in row:
                        if k not in keys:
                            keys.append(k)
                ws.append(keys)
                for row in data:
                    ws.append([cell(row.get(k)) for k in keys])
            else:
                ws.append(["value"])
                for v in data:
                    ws.append([cell(v)])
        elif isinstance(data, dict):
            ws.append(["key", "value"])
            for k, v in data.items():
                ws.append([str(k), cell(v)])
        else:
            ws.append(["value", cell(data)])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _json_default(obj: Any) -> Any:
    if isinstance(obj, datetime):
        return _iso(obj)
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")


def atomic_publish(out_dir: Path, snapshot: dict[str, Any]) -> dict[str, Any]:
    """ONLY allowed writer for canonical OUT triad."""
    out_dir = Path(out_dir)
    snap = deepcopy(snapshot)
    snap["payload_hash_algorithm"] = "sha256_canonical_json_v1"
    snap["payload_excluded_json_paths"] = list(PAYLOAD_EXCLUDED_JSON_PATHS)
    snap["report_payload_sha256"] = "pending"
    snap["report_payload_sha256"] = compute_report_payload_sha256(snap)
    validate_snapshot_schema(snap)

    raw_json = json.dumps(snap, ensure_ascii=False, indent=2, default=_json_default) + "\n"
    assert_no_corrupt_hash_as_canonical(raw_json)

    with tempfile.TemporaryDirectory(prefix="e1x5_sot_") as td:
        tmp = Path(td)
        (tmp / "report.json").write_text(raw_json, encoding="utf-8")
        loaded = json.loads((tmp / "report.json").read_text(encoding="utf-8"))
        md = render_markdown(loaded)
        assert_no_corrupt_hash_as_canonical(md)
        (tmp / "report.md").write_text(md, encoding="utf-8")
        render_xlsx(tmp / "audit.xlsx", loaded)

        j2 = json.loads((tmp / "report.json").read_text(encoding="utf-8"))
        md2 = (tmp / "report.md").read_text(encoding="utf-8")
        assert j2["run_id"] == snap["run_id"]
        assert j2["overall_verdict"] in md2
        assert j2["report_payload_sha256"] in md2
        from openpyxl import load_workbook

        wb = load_workbook(tmp / "audit.xlsx", read_only=True)
        assert "Ledger_Hash" in wb.sheetnames
        assert "Artifact_Integrity" in wb.sheetnames
        wb.close()

        out_dir.mkdir(parents=True, exist_ok=True)
        for name in ARTIFACT_NAMES:
            shutil.copy2(tmp / name, out_dir / name)

    return {
        "run_id": snap["run_id"],
        "report_payload_sha256": snap["report_payload_sha256"],
        "generated_at": snap["generated_at"],
        "overall_verdict": snap["overall_verdict"],
        "paths": {n: str(out_dir / n) for n in ARTIFACT_NAMES},
    }


def trade_ledger_hash_v1(rows: Sequence[Mapping[str, Any]]) -> str:
    return canonical_ledger_hash(rows, version="v1")


def trade_ledger_hash_v2(rows: Sequence[Mapping[str, Any]]) -> str:
    return canonical_ledger_hash(rows, version="v2")


def project_trade_ledger_v1(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return project_ledger(rows, version="v1")


def project_trade_ledger_v2(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return project_ledger(rows, version="v2")


LEDGER_HASH_V1_FROZEN_EXPECTED_PM = FROZEN_PM_HASH_V1
