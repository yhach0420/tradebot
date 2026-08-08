"""E1_X5 parity followup helpers — score availability, SHA naming, event stream audit.

Does not mutate strategy thresholds / EXIT / CAP. Paper / observe-only evidence only.
"""
from __future__ import annotations

import hashlib
import json
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Optional

CANONICAL_EVENT_SCHEMA = "e1_x5_event_canonical_v1"
CANONICAL_TRADE_SCHEMA = "e1_x5_trade_ledger_canonical_v1"
PARITY_AUDIT_BLOCKED = "E1_X5_PARITY_AUDIT_BLOCKED"
FORWARD_DAY1_READY = "E1_X5_FORWARD_DAY1_READY"
FORWARD_DAY1_PARTIAL = "E1_X5_FORWARD_DAY1_PARTIAL"
FORWARD_DAY1_PASS = "E1_X5_FORWARD_DAY1_PASS"
VERDICT_PARITY_FIXED = "E1_X5_RUNTIME_OFFLINE_PARITY_FIXED"

# Legacy Oracle (20260727 freeze) aggregated REGULAR+STATE_CHANGE into one label.
LEGACY_REASON_SPLIT = "UNAVAILABLE_AGGREGATED"


def sha256_file(path: Path, *, chunk: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            b = fh.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def sha256_canonical(obj: Any) -> str:
    raw = json.dumps(obj, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def normalize_sample_reason(reason: Any) -> str:
    """Map legacy labels to current reason vocabulary (display/audit only)."""
    s = str(reason or "")
    mapping = {
        "periodic": "REGULAR_5S",
        "periodic_or_state_change": "LEGACY_PERIODIC_OR_STATE_CHANGE",
        "state_change": "STATE_CHANGE",
        "not_due": "NOT_DUE",
        "score_missing": "NO_EVALUATION",
        "identity_fail": "NO_EVALUATION",
        "decision_time_missing": "NO_EVALUATION",
        "REGULAR": "REGULAR_5S",
        "REGULAR_5S": "REGULAR_5S",
        "STATE_CHANGE": "STATE_CHANGE",
        "NOT_DUE": "NOT_DUE",
        "NO_EVALUATION": "NO_EVALUATION",
    }
    return mapping.get(s, s.upper() if s else "")


def canonical_event_record(row: Mapping[str, Any]) -> dict[str, Any]:
    score = row.get("score")
    if isinstance(score, float):
        score_v: Any = round(score, 12)
    else:
        score_v = score
    kind = str(row.get("observe_kind") or "")
    # Treat MISSING / MISSING_SCORE uniformly for ledger compare
    if kind in ("MISSING", "MISSING_SCORE"):
        kind = "MISSING"
    return {
        "schema": CANONICAL_EVENT_SCHEMA,
        "event_id": str(row.get("event_id") or ""),
        "observe_kind": kind,
        "score": score_v,
        "feature_hash": str(row.get("feature_hash") or ""),
        "missing_reason": row.get("missing_reason"),
        "position_before": bool(row.get("position_before")),
        "position_after": bool(row.get("position_after")),
        "cap_before": int(row.get("cap_before") or 0),
        "cap_after": int(row.get("cap_after") or 0),
        "entered": (
            bool(row.get("entered"))
            if "entered" in row
            else bool(row.get("score_evaluated") and row.get("entry_result") is None and row.get("observe_kind") == "SCORE")
        ),
        "exited": bool(row.get("exited") if "exited" in row else row.get("exit_happened")),
        "exit_reason": row.get("exit_reason"),
        "ingress_sequence": row.get("ingress_sequence"),
        "symbol": str(row.get("symbol") or ""),
    }


def iter_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            yield json.loads(line)


def score_availability_audit(
    *,
    evaluated_count: int,
    no_evaluation_count: int,
    tick_build_failed_count: int,
    missing_score_after_valid_tick: int = 0,
    score_fill_in: int = 0,
    pm_reason: str = "TICK_BUILD_FAILED",
    oracle_id: str = "E1_X5_OFFLINE_ORACLE_20260727_PM",
) -> dict[str, Any]:
    """AM unverified vs PM replay facts — never label PM TICK_BUILD as AM RESOLVED."""
    return {
        "sheet_name": "Score Availability Audit",
        "oracle_id": oracle_id,
        "am_score_state": "UNVERIFIED_PENDING_NEW_AM_PAPER",
        "pm_replay_no_evaluation": int(no_evaluation_count),
        "pm_replay_reason": str(pm_reason),
        "score_fill_in": int(score_fill_in),
        "evaluated": int(evaluated_count),
        "no_evaluation": int(no_evaluation_count),
        "missing_score_after_valid_tick": int(missing_score_after_valid_tick),
        "TICK_BUILD_FAILED": int(tick_build_failed_count),
        "topline": f"evaluated/no_evaluation: {int(evaluated_count)}/{int(no_evaluation_count)}",
        "note": (
            "Current Capture is PM-only. AM score was not observed. "
            "308 NO_EVALUATION events are PM TICK_BUILD_FAILED (5253.T); "
            "not an AM cause; no score fill-in applied."
        ),
        # Deprecated wrong label — keep explicit rejection for readers of old reports
        "deprecated_am_label_rejected": "RESOLVED_OR_EXPLICIT",
    }


def stream_side_stats(path: Path) -> dict[str, Any]:
    """Aggregate one event log without loading it fully into memory."""
    if not path.is_file():
        return {"exists": False, "path": str(path)}

    raw_sha = sha256_file(path)
    byte_size = path.stat().st_size
    kinds: Counter[str] = Counter()
    reasons: Counter[str] = Counter()
    missing_reasons: Counter[str] = Counter()
    n = 0
    seq_min: Optional[int] = None
    seq_max: Optional[int] = None
    seq_gaps = 0
    seq_dups = 0
    seq_inv = 0
    prev_seq: Optional[int] = None
    seen_seq: set[int] = set()
    first_id = ""
    last_id = ""
    feature_hash_present = 0
    canonical_hasher = hashlib.sha256()
    score_count = 0
    no_eval_count = 0
    not_due_count = 0

    for row in iter_jsonl(path):
        n += 1
        eid = str(row.get("event_id") or "")
        if n == 1:
            first_id = eid
        last_id = eid
        kind = str(row.get("observe_kind") or "")
        kinds[kind] += 1
        reasons[normalize_sample_reason(row.get("sample_reason"))] += 1
        mr = row.get("missing_reason")
        if mr:
            missing_reasons[str(mr)] += 1

        if kind in ("SCORE",):
            score_count += 1
        elif kind in ("MISSING", "MISSING_SCORE"):
            no_eval_count += 1
        elif kind in ("NO_SAMPLE",):
            not_due_count += 1

        fh = str(row.get("feature_hash") or "")
        if fh:
            feature_hash_present += 1

        seq = row.get("ingress_sequence")
        try:
            seq_i = int(seq) if seq is not None else None
        except (TypeError, ValueError):
            seq_i = None
        if seq_i is not None:
            if seq_min is None or seq_i < seq_min:
                seq_min = seq_i
            if seq_max is None or seq_i > seq_max:
                seq_max = seq_i
            if seq_i in seen_seq:
                seq_dups += 1
            else:
                seen_seq.add(seq_i)
            if prev_seq is not None:
                if seq_i < prev_seq:
                    seq_inv += 1
                elif seq_i > prev_seq + 1:
                    seq_gaps += 1
            prev_seq = seq_i

        crec = canonical_event_record(row)
        canonical_hasher.update(
            json.dumps(crec, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        )
        canonical_hasher.update(b"\n")

    return {
        "exists": True,
        "raw_log_path": str(path.resolve()),
        "byte_size": byte_size,
        "raw_file_sha256": raw_sha,
        "canonical_event_sha256": canonical_hasher.hexdigest(),
        "canonicalization_schema": CANONICAL_EVENT_SCHEMA,
        "canonicalization_version": 1,
        "record_count": n,
        "sequence_min": seq_min,
        "sequence_max": seq_max,
        "sequence_gap_count": seq_gaps,
        "sequence_duplicate_count": seq_dups,
        "sequence_inversion_count": seq_inv,
        "first_event_id": first_id,
        "last_event_id": last_id,
        "SCORE": score_count,
        "NO_EVALUATION": no_eval_count,
        "NOT_DUE": not_due_count,
        "not_due": not_due_count,  # alias
        "observe_kinds": dict(kinds),
        "sample_reasons_normalized": dict(reasons),
        "missing_reasons": dict(missing_reasons),
        "feature_hash_present": feature_hash_present,
        "feature_hash_present_count": feature_hash_present,
    }


def compare_event_streams(
    oracle_path: Path,
    runtime_path: Path,
    *,
    oracle_feature_hash_schema: Optional[str] = None,
    runtime_feature_hash_schema: Optional[str] = None,
) -> dict[str, Any]:
    """Side-by-side Oracle/Runtime event parity from Source-of-Truth JSONL."""
    from small_paper.e1_x5_canonical_feature_hash import (
        LEGACY_ORACLE_FEATURE_HASH_SCHEMA,
        LEGACY_RUNTIME_FEATURE_HASH_SCHEMA,
        compare_feature_hashes,
        infer_legacy_feature_hash_schema,
    )

    if not oracle_path.is_file() or not runtime_path.is_file():
        return {
            "status": PARITY_AUDIT_BLOCKED,
            "reason": "runtime_or_oracle_event_log_missing",
            "oracle_exists": oracle_path.is_file(),
            "runtime_exists": runtime_path.is_file(),
            "feature_hash_comparison_status": None,
            "feature_hash_comparable_count": None,
            "feature_hash_not_comparable_count": None,
            "feature_hash_mismatch_count": None,
            "feature_hash_mismatch_display": "N/A",
            "score_mismatch": None,
            "position_cap_mismatch": None,
            "entry_exit_decision_mismatch": None,
            "first_mismatch": None,
            "note": "Do not invent mismatch=0 when logs are absent.",
        }

    oracle_stats = stream_side_stats(oracle_path)
    runtime_stats = stream_side_stats(runtime_path)

    o_schema = oracle_feature_hash_schema or LEGACY_ORACLE_FEATURE_HASH_SCHEMA
    r_schema = runtime_feature_hash_schema or LEGACY_RUNTIME_FEATURE_HASH_SCHEMA
    # Prefer schema embedded in first SCORE row if present
    for row in iter_jsonl(oracle_path):
        if row.get("feature_hash_schema"):
            o_schema = str(row["feature_hash_schema"])
            break
        if str(row.get("observe_kind") or "") == "SCORE" and row.get("feature_hash"):
            break
    for row in iter_jsonl(runtime_path):
        if row.get("feature_hash_schema"):
            r_schema = str(row["feature_hash_schema"])
            break
        if str(row.get("observe_kind") or "") == "SCORE" and row.get("feature_hash"):
            break

    score_mismatch = 0
    position_mismatch = 0
    cap_mismatch = 0
    entry_decision_mismatch = 0
    exit_decision_mismatch = 0
    first_mismatch: Optional[dict[str, Any]] = None
    compared = 0
    hash_pairs: list[tuple[str, str]] = []

    o_iter = iter_jsonl(oracle_path)
    r_iter = iter_jsonl(runtime_path)
    while True:
        try:
            o = next(o_iter)
        except StopIteration:
            o = None
        try:
            r = next(r_iter)
        except StopIteration:
            r = None
        if o is None and r is None:
            break
        compared += 1
        if o is None or r is None:
            detail = {
                "i": compared - 1,
                "field": "length",
                "oracle_event_id": (o or {}).get("event_id"),
                "runtime_event_id": (r or {}).get("event_id"),
            }
            if first_mismatch is None:
                first_mismatch = detail
            entry_decision_mismatch += 1
            continue

        oc = canonical_event_record(o)
        rc = canonical_event_record(r)
        if oc["event_id"] != rc["event_id"]:
            detail = {
                "i": compared - 1,
                "field": "event_id",
                "oracle": oc["event_id"],
                "runtime": rc["event_id"],
            }
            if first_mismatch is None:
                first_mismatch = detail
            entry_decision_mismatch += 1
            continue

        if oc["score"] != rc["score"]:
            score_mismatch += 1
            if first_mismatch is None:
                first_mismatch = {
                    "i": compared - 1,
                    "field": "score",
                    "event_id": oc["event_id"],
                    "oracle": oc["score"],
                    "runtime": rc["score"],
                }

        oh = str(oc.get("feature_hash") or "")
        rh = str(rc.get("feature_hash") or "")
        if oh or rh:
            if oc["observe_kind"] == "SCORE" or rc["observe_kind"] == "SCORE":
                hash_pairs.append((oh, rh))
            if o.get("feature_hash_schema"):
                o_schema = str(o["feature_hash_schema"])
            if r.get("feature_hash_schema"):
                r_schema = str(r["feature_hash_schema"])

        pos_diff = (
            oc["position_before"] != rc["position_before"]
            or oc["position_after"] != rc["position_after"]
        )
        cap_diff = oc["cap_before"] != rc["cap_before"] or oc["cap_after"] != rc["cap_after"]
        if pos_diff:
            position_mismatch += 1
            if first_mismatch is None:
                first_mismatch = {
                    "i": compared - 1,
                    "field": "position",
                    "event_id": oc["event_id"],
                    "oracle": {
                        "position_before": oc["position_before"],
                        "position_after": oc["position_after"],
                    },
                    "runtime": {
                        "position_before": rc["position_before"],
                        "position_after": rc["position_after"],
                    },
                }
        if cap_diff:
            cap_mismatch += 1
            if first_mismatch is None:
                first_mismatch = {
                    "i": compared - 1,
                    "field": "cap",
                    "event_id": oc["event_id"],
                    "oracle": {"cap_before": oc["cap_before"], "cap_after": oc["cap_after"]},
                    "runtime": {"cap_before": rc["cap_before"], "cap_after": rc["cap_after"]},
                }

        kind_mismatch = oc["observe_kind"] != rc["observe_kind"]
        entered_mismatch = False
        if "entered" in o and "entered" in r:
            entered_mismatch = bool(o.get("entered")) != bool(r.get("entered"))
        if kind_mismatch or entered_mismatch:
            entry_decision_mismatch += 1
            if first_mismatch is None:
                first_mismatch = {
                    "i": compared - 1,
                    "field": "entry_decision",
                    "event_id": oc["event_id"],
                    "oracle": {
                        "observe_kind": oc["observe_kind"],
                        "entered": o.get("entered") if "entered" in o else oc.get("entered"),
                    },
                    "runtime": {
                        "observe_kind": rc["observe_kind"],
                        "entered": r.get("entered") if "entered" in r else rc.get("entered"),
                    },
                }

        o_ex = bool(oc["exited"])
        r_ex = bool(rc["exited"])
        o_er = oc.get("exit_reason")
        r_er = rc.get("exit_reason")
        exit_reason_mismatch = o_er is not None and r_er is not None and o_er != r_er
        if o_ex != r_ex or exit_reason_mismatch:
            exit_decision_mismatch += 1
            if first_mismatch is None:
                first_mismatch = {
                    "i": compared - 1,
                    "field": "exit_decision",
                    "event_id": oc["event_id"],
                    "oracle": {"exited": oc["exited"], "exit_reason": o_er},
                    "runtime": {"exited": rc["exited"], "exit_reason": r_er},
                }

    fh = compare_feature_hashes(
        oracle_schema=o_schema or infer_legacy_feature_hash_schema("oracle"),
        runtime_schema=r_schema or infer_legacy_feature_hash_schema("runtime"),
        pairs=hash_pairs,
    )

    status = "COMPARED"
    decision_parity_ok = (
        score_mismatch == 0
        and position_mismatch == 0
        and cap_mismatch == 0
        and entry_decision_mismatch == 0
        and exit_decision_mismatch == 0
        and oracle_stats["record_count"] == runtime_stats["record_count"]
    )
    if not decision_parity_ok:
        status = "COMPARED_WITH_MISMATCHES"
    elif fh["feature_hash_comparison_status"] == "NOT_COMPARABLE_RECIPE_DIFFERENCE":
        status = "COMPARED_FEATURE_HASH_NOT_COMPARABLE"

    return {
        "status": status,
        "decision_parity_ok": decision_parity_ok,
        "compared_records": compared,
        "oracle": oracle_stats,
        "runtime": runtime_stats,
        **fh,
        "score_mismatch": score_mismatch,
        "score_mismatch_count": score_mismatch,
        "position_mismatch": position_mismatch,
        "position_mismatch_count": position_mismatch,
        "cap_mismatch": cap_mismatch,
        "cap_mismatch_count": cap_mismatch,
        # Back-compat combined fields
        "position_cap_mismatch": position_mismatch + cap_mismatch,
        "entry_decision_mismatch": entry_decision_mismatch,
        "entry_decision_mismatch_count": entry_decision_mismatch,
        "exit_decision_mismatch": exit_decision_mismatch,
        "exit_decision_mismatch_count": exit_decision_mismatch,
        "entry_exit_decision_mismatch": entry_decision_mismatch + exit_decision_mismatch,
        "first_mismatch": first_mismatch,
        "legacy_reason_split": LEGACY_REASON_SPLIT,
        "legacy_reason_note": (
            "Frozen 20260727 Oracle aggregated REGULAR+STATE_CHANGE as "
            "periodic_or_state_change; do not rewrite the freeze. "
            f"legacy reason split: {LEGACY_REASON_SPLIT}"
        ),
        "next_paper_feature_hash": (
            "Use e1_x5_canonical_feature_hash v1 on both Runtime and Replay"
        ),
    }


def trade_ledger_sha_bundle(trades: Any) -> dict[str, str]:
    return {
        "canonical_trade_ledger_sha256": sha256_canonical(trades),
        "canonicalization_schema": CANONICAL_TRADE_SCHEMA,
    }


def funnel_exclusive_invariants(
    funnel: Mapping[str, Any],
    *,
    expected_evaluated: int,
    no_evaluation: int,
    no_evaluation_breakdown: Optional[Mapping[str, Any]] = None,
) -> dict[str, Any]:
    """Exclusive funnel must sum to evaluated; no_evaluation must not appear in funnel."""
    forbidden = ("no_evaluation" in funnel) or ("tick_build_failed" in funnel)
    keys = (
        "missing_score_after_valid_tick",
        "threshold_fail",
        "spread_fail",
        "same_symbol_blocked",
        "cap_blocked",
        "accepted_entry",
        "other_reject",
    )
    terminal_sum = int(funnel.get("terminal_sum") or sum(int(funnel.get(k) or 0) for k in keys))
    miss = int(funnel.get("missing_score_after_valid_tick") or 0)
    breakdown = dict(no_evaluation_breakdown or {})
    reason_sum = int(sum(int(v) for v in (breakdown.get("no_evaluation_reason_breakdown") or {}).values()))
    if not reason_sum and breakdown.get("TICK_BUILD_FAILED") is not None:
        reason_sum = int(breakdown.get("TICK_BUILD_FAILED") or 0)
    # Allow breakdown dict shaped as {TICK_BUILD_FAILED: 308}
    if not reason_sum and any(k.isupper() for k in breakdown.keys()):
        reason_sum = int(sum(int(v) for k, v in breakdown.items() if k.isupper()))
    return {
        "missing_score_after_valid_tick": miss,
        "terminal_sum": terminal_sum,
        "expected_evaluated": int(expected_evaluated),
        "no_evaluation_in_funnel": forbidden,
        "funnel_sum_ok": terminal_sum == int(expected_evaluated),
        "no_evaluation_reason_sum": reason_sum if reason_sum else int(no_evaluation),
        "no_evaluation_reason_sum_ok": (reason_sum == int(no_evaluation)) if reason_sum else True,
        "double_count_ok": (not forbidden) and (miss == 0) and (terminal_sum == int(expected_evaluated)),
        "double_count_forbidden": True,
    }


# Back-compat alias used by older tests / scripts
def funnel_no_double_count(funnel: Mapping[str, Any], *, no_evaluation: int) -> dict[str, Any]:
    inv = funnel_exclusive_invariants(
        funnel,
        expected_evaluated=int(funnel.get("terminal_sum") or 0) or 17353,
        no_evaluation=no_evaluation,
    )
    # Legacy shape expected no_evaluation OUTSIDE funnel
    inv["no_evaluation"] = int(no_evaluation)
    inv["tick_build_failed"] = 0 if "tick_build_failed" not in funnel else int(funnel.get("tick_build_failed") or 0)
    return inv


def rebuild_exclusive_funnel_from_prior(prior_funnel: Mapping[str, Any]) -> dict[str, int]:
    """Strip no_evaluation from a legacy prior funnel and recompute terminal_sum."""
    funnel = {
        "missing_score_after_valid_tick": int(
            prior_funnel.get("missing_score_after_valid_tick", prior_funnel.get("missing_score", 0)) or 0
        ),
        "threshold_fail": int(prior_funnel.get("threshold_fail") or 0),
        "spread_fail": int(prior_funnel.get("spread_fail") or 0),
        "same_symbol_blocked": int(prior_funnel.get("same_symbol_blocked") or 0),
        "cap_blocked": int(prior_funnel.get("cap_blocked") or 0),
        "accepted_entry": int(prior_funnel.get("accepted_entry") or 0),
        "other_reject": int(prior_funnel.get("other_reject") or 0),
    }
    funnel["terminal_sum"] = (
        funnel["missing_score_after_valid_tick"]
        + funnel["threshold_fail"]
        + funnel["spread_fail"]
        + funnel["same_symbol_blocked"]
        + funnel["cap_blocked"]
        + funnel["accepted_entry"]
        + funnel["other_reject"]
    )
    return funnel


EVENT_PARITY_SIDE_KEYS = (
    "raw_log_path",
    "byte_size",
    "raw_file_sha256",
    "canonical_event_sha256",
    "canonicalization_schema",
    "canonicalization_version",
    "record_count",
    "sequence_min",
    "sequence_max",
    "sequence_gap_count",
    "sequence_duplicate_count",
    "sequence_inversion_count",
    "first_event_id",
    "last_event_id",
    "SCORE",
    "NO_EVALUATION",
    "NOT_DUE",
    "feature_hash_present_count",
)


def side_audit_view(stats: Mapping[str, Any]) -> dict[str, Any]:
    """Normalized Oracle/Runtime side fields for report / Excel."""
    if not stats or not stats.get("exists", True):
        return {
            "exists": False,
            **{k: None for k in EVENT_PARITY_SIDE_KEYS},
        }
    return {
        "exists": True,
        "raw_log_path": stats.get("raw_log_path"),
        "byte_size": stats.get("byte_size"),
        "raw_file_sha256": stats.get("raw_file_sha256"),
        "canonical_event_sha256": stats.get("canonical_event_sha256"),
        "canonicalization_schema": stats.get("canonicalization_schema"),
        "canonicalization_version": stats.get("canonicalization_version"),
        "record_count": stats.get("record_count"),
        "sequence_min": stats.get("sequence_min"),
        "sequence_max": stats.get("sequence_max"),
        "sequence_gap_count": stats.get("sequence_gap_count"),
        "sequence_duplicate_count": stats.get("sequence_duplicate_count"),
        "sequence_inversion_count": stats.get("sequence_inversion_count"),
        "first_event_id": stats.get("first_event_id"),
        "last_event_id": stats.get("last_event_id"),
        "SCORE": stats.get("SCORE"),
        "NO_EVALUATION": stats.get("NO_EVALUATION"),
        "NOT_DUE": stats.get("NOT_DUE", stats.get("not_due")),
        "feature_hash_present_count": stats.get(
            "feature_hash_present_count", stats.get("feature_hash_present")
        ),
        "missing_reasons": stats.get("missing_reasons") or {},
    }


def build_event_parity_sections(event_parity: Mapping[str, Any]) -> dict[str, Any]:
    """Machine-readable event_parity.oracle / .runtime / .comparison."""
    status = event_parity.get("status")
    blocked = status == PARITY_AUDIT_BLOCKED
    fh_mismatch = event_parity.get("feature_hash_mismatch_count")
    fh_display = event_parity.get("feature_hash_mismatch_display")
    if fh_display is None:
        fh_display = "N/A" if fh_mismatch is None else str(fh_mismatch)

    comparison = {
        "status": status,
        "decision_parity_ok": event_parity.get("decision_parity_ok"),
        "compared_records": event_parity.get("compared_records"),
        "feature_hash_comparison_status": event_parity.get("feature_hash_comparison_status"),
        "feature_hash_comparable_count": event_parity.get("feature_hash_comparable_count"),
        "feature_hash_not_comparable_count": event_parity.get("feature_hash_not_comparable_count"),
        "feature_hash_mismatch_count": fh_mismatch,
        "feature_hash_mismatch_display": fh_display,
        "score_mismatch_count": event_parity.get("score_mismatch_count", event_parity.get("score_mismatch")),
        "position_mismatch_count": event_parity.get(
            "position_mismatch_count", event_parity.get("position_mismatch")
        ),
        "cap_mismatch_count": event_parity.get("cap_mismatch_count", event_parity.get("cap_mismatch")),
        "entry_decision_mismatch_count": event_parity.get(
            "entry_decision_mismatch_count", event_parity.get("entry_decision_mismatch")
        ),
        "exit_decision_mismatch_count": event_parity.get(
            "exit_decision_mismatch_count", event_parity.get("exit_decision_mismatch")
        ),
        "first_mismatch": event_parity.get("first_mismatch"),
        "legacy_reason_split": event_parity.get("legacy_reason_split"),
        "blocked_reason": event_parity.get("reason") if blocked else None,
        "note": event_parity.get("note"),
        "next_paper_feature_hash": event_parity.get("next_paper_feature_hash"),
    }
    return {
        "oracle": side_audit_view(event_parity.get("oracle") or {}),
        "runtime": side_audit_view(event_parity.get("runtime") or {}),
        "comparison": comparison,
    }


def event_parity_side_by_side_rows(sections: Mapping[str, Any]) -> list[dict[str, Any]]:
    """Rows for Excel Event Parity sheet: field | oracle | runtime."""
    oracle = sections.get("oracle") or {}
    runtime = sections.get("runtime") or {}
    rows: list[dict[str, Any]] = []
    for key in EVENT_PARITY_SIDE_KEYS:
        rows.append(
            {
                "field": key,
                "oracle": oracle.get(key),
                "runtime": runtime.get(key),
            }
        )
    return rows


def event_parity_comparison_rows(sections: Mapping[str, Any]) -> list[dict[str, Any]]:
    comp = sections.get("comparison") or {}
    order = [
        "status",
        "decision_parity_ok",
        "compared_records",
        "feature_hash_comparison_status",
        "feature_hash_comparable_count",
        "feature_hash_not_comparable_count",
        "feature_hash_mismatch_count",
        "feature_hash_mismatch_display",
        "score_mismatch_count",
        "position_mismatch_count",
        "cap_mismatch_count",
        "entry_decision_mismatch_count",
        "exit_decision_mismatch_count",
        "first_mismatch",
        "legacy_reason_split",
        "blocked_reason",
        "next_paper_feature_hash",
    ]
    rows = []
    for k in order:
        v = comp.get(k)
        if k == "feature_hash_mismatch_count" and v is None:
            display = "N/A"
        elif k == "first_mismatch":
            display = json.dumps(v, ensure_ascii=False, default=str) if v is not None else None
        else:
            display = v
        rows.append({"field": k, "value": display})
    return rows


def write_parity_audit_workbook(path: Path, sheets: Mapping[str, Any]) -> None:
    """Write audit.xlsx with freeze panes, autofilter, and readable column widths."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    def cell(v: Any) -> Any:
        if v is None:
            return None
        if isinstance(v, (str, int, float, bool)):
            return v
        return json.dumps(v, ensure_ascii=False, default=str)

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)

    for name, data in sheets.items():
        ws = wb.create_sheet(title=str(name)[:31])
        if isinstance(data, list) and data and isinstance(data[0], dict):
            keys = list(data[0].keys())
            ws.append(keys)
            for row in data:
                ws.append([cell(row.get(k)) for k in keys])
        elif isinstance(data, dict):
            ws.append(["field", "value"])
            for k, v in data.items():
                # Explicit N/A for null mismatch displays when key says so
                if k.endswith("mismatch_count") and v is None and "feature_hash" in k:
                    ws.append([str(k), "N/A"])
                elif k.endswith("mismatch_display") and v is None:
                    ws.append([str(k), "N/A"])
                else:
                    ws.append([str(k), cell(v)])
            keys = ["field", "value"]
        else:
            ws.append(["value"])
            ws.append([cell(data)])
            keys = ["value"]

        # Header style + freeze + filter
        for col_idx, key in enumerate(keys, start=1):
            c = ws.cell(1, col_idx)
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # Column widths: SHA / paths need width; headers must not clip
        for col_idx, key in enumerate(keys, start=1):
            key_s = str(key)
            max_len = len(key_s)
            for row_idx in range(2, min(ws.max_row, 500) + 1):
                val = ws.cell(row_idx, col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, min(len(str(val)), 80))
            # Prefer wider columns for SHA / path / field names
            if "sha" in key_s.lower() or "path" in key_s.lower() or key_s in ("field", "oracle", "runtime", "value"):
                width = max(28, min(max_len + 4, 72))
            else:
                width = max(14, min(max_len + 2, 48))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def format_event_parity_markdown(sections: Mapping[str, Any]) -> str:
    """Markdown block with Oracle/Runtime side-by-side audit values."""
    o = sections.get("oracle") or {}
    r = sections.get("runtime") or {}
    c = sections.get("comparison") or {}
    lines = [
        "## Event Parity",
        "",
        f"- status: `{c.get('status')}`",
        f"- decision_parity_ok: `{c.get('decision_parity_ok')}`",
        f"- compared_records: `{c.get('compared_records')}`",
        "",
        "| field | oracle | runtime |",
        "| --- | --- | --- |",
    ]
    for key in EVENT_PARITY_SIDE_KEYS:
        ov = o.get(key)
        rv = r.get(key)
        lines.append(f"| {key} | `{ov}` | `{rv}` |")
    fh_mm = c.get("feature_hash_mismatch_display") or "N/A"
    lines.extend(
        [
            "",
            "### Comparison (decision vs feature-hash separated)",
            f"- feature_hash_comparison_status: `{c.get('feature_hash_comparison_status')}`",
            f"- feature_hash_comparable_count: `{c.get('feature_hash_comparable_count')}`",
            f"- feature_hash_not_comparable_count: `{c.get('feature_hash_not_comparable_count')}`",
            f"- feature_hash_mismatch_count: `{fh_mm}`",
            f"- score_mismatch_count: `{c.get('score_mismatch_count')}`",
            f"- position_mismatch_count: `{c.get('position_mismatch_count')}`",
            f"- cap_mismatch_count: `{c.get('cap_mismatch_count')}`",
            f"- entry_decision_mismatch_count: `{c.get('entry_decision_mismatch_count')}`",
            f"- exit_decision_mismatch_count: `{c.get('exit_decision_mismatch_count')}`",
            f"- first_mismatch: `{json.dumps(c.get('first_mismatch'), ensure_ascii=False, default=str)}`",
            f"- legacy_reason_split: `{c.get('legacy_reason_split')}`",
        ]
    )
    return "\n".join(lines)
