"""V1R EXIT Global Search runner.

Candidate gen → Isolation → Nested CV → Full Replay → Robustness → Freeze → 8/10 holdout.
ENTRY frozen. No production mutation. 20260810 locked until freeze.
"""
from __future__ import annotations

import hashlib
import json
import math
import pickle
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

from research.e1_x22_actual_exit_factory.paths import session_end_epoch
from research.e1_x32_upstream_attribution.eval_stages import load_boards_for_symbols
from research.e1_x35_passive_exit.paths import build_path
from research.e1_x36_joint_allocator import OUTER_BLOCKS, LOT_QTY
from research.e1_x36_joint_allocator.replay import simulate_joint
from research.e1_x36r_freeze_integrity.serialize import score_fn_from_serialized
from research.e1_x37_prospective.freeze import load_model_artifact, load_v1r
from research.v1r_exit_global_search.apply_exit import apply_candidate, attach_board_series
from research.v1r_exit_global_search.candidates import family_counts, generate_all_candidates
from research.v1r_exit_global_search.learned import (
    apply_learned_exit,
    build_decision_rows,
    fit_logistic_l1,
    fit_shallow_tree,
)
from small_paper.v1r_primary_runtime import MODEL_ARTIFACT_SHA, V1R_SHA

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "v1r_exit_global_search"
PANEL_CACHE = NATIVE / "results/research/v1r_capital_sweep_0p5m_10m/_panel_cache.pkl"
FORBIDDEN = "20260810"
ANALYSIS_ID = "V1R_EXIT_GLOBAL_SEARCH"
FEATS = ("spread_bps", "imbalance", "mid_ret_60s", "mid_ret_180s", "event_rate_60s", "log_bid_qty")


def _sfn():
    ser = load_model_artifact()
    raw = score_fn_from_serialized(ser)

    def fn(e: dict) -> float:
        try:
            return float(raw({k: e.get(k) for k in FEATS}))
        except Exception:
            return float("-inf")

    return fn


def _pnl(fill_price: float, ret_bps: float) -> float:
    return float(LOT_QTY) * float(fill_price) * float(ret_bps) / 10000.0


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def write_xlsx(sheets: dict[str, list[dict]], path: Path) -> None:
    if Workbook is None:
        path.with_suffix(".sheets.json").write_text(
            json.dumps(sheets, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def summarize_pnls(pnls: list[float]) -> dict[str, Any]:
    if not pnls:
        return {
            "n": 0, "total": 0.0, "pf": 0.0, "wins": 0, "losses": 0, "flats": 0,
            "win_rate": 0.0, "avg": 0.0, "median": 0.0, "best": 0.0, "worst": 0.0,
            "gross_profit": 0.0, "gross_loss": 0.0, "max_dd": 0.0,
        }
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    flats = sum(1 for p in pnls if p == 0)
    gp = sum(p for p in pnls if p > 0)
    gl = abs(sum(p for p in pnls if p < 0))
    pf = (gp / gl) if gl > 0 else (float("inf") if gp > 0 else 0.0)
    eq = peak = 0.0
    max_dd = 0.0
    for p in pnls:
        eq += p
        peak = max(peak, eq)
        max_dd = min(max_dd, eq - peak)
    return {
        "n": len(pnls),
        "total": float(sum(pnls)),
        "pf": None if pf == float("inf") else float(pf),
        "wins": wins,
        "losses": losses,
        "flats": flats,
        "win_rate": wins / len(pnls),
        "avg": float(statistics.mean(pnls)),
        "median": float(statistics.median(pnls)),
        "best": float(max(pnls)),
        "worst": float(min(pnls)),
        "gross_profit": float(gp),
        "gross_loss": float(gl),
        "max_dd": float(max_dd),
    }


def composite_score(sm: dict[str, Any], base: dict[str, Any], *, complexity: float = 1.0) -> float:
    """Inner selection: PnL + PF + DD + worst + gross loss — not pure PnL."""
    score = 0.0
    score += (sm["total"] - base["total"]) / 1000.0
    pf = sm["pf"] if sm["pf"] is not None else 10.0
    bpf = base["pf"] if base["pf"] is not None else 10.0
    score += (pf - bpf) * 800.0
    score += (sm["max_dd"] - base["max_dd"]) / 400.0  # less negative better
    score += (sm["worst"] - base["worst"]) / 200.0
    score += (base["gross_loss"] - sm["gross_loss"]) / 500.0
    # positive-day proxy via win_rate soft
    score += (sm["win_rate"] - base["win_rate"]) * 200.0
    score -= complexity * 15.0
    return score


def support_thresholds(n_trades: int, n_days: int) -> dict[str, int]:
    """Reasonable mins from train size."""
    return {
        "min_trigger": max(12, int(0.08 * n_trades)),
        "min_days": max(4, int(0.35 * n_days)),
        "min_symbols": max(3, min(8, int(0.25 * max(n_trades // 8, 1)))),
        "n_trades": n_trades,
        "n_days": n_days,
    }


def _aggregate_trade_exits(
    rows: list[dict[str, Any]],
    *,
    cid: str,
    family: str,
) -> dict[str, Any]:
    pnls = [r["pnl"] for r in rows]
    base_pnls = [r["base_pnl"] for r in rows]
    trigger_n = 0
    trigger_holds: list[float] = []
    days = set()
    syms = set()
    winners_cut = 0
    losers_saved = 0
    mfe_caps: list[float] = []
    for r in rows:
        if r.get("triggered"):
            trigger_n += 1
            trigger_holds.append(float(r["hold_sec"]))
            days.add(r["date"])
            syms.add(r["symbol"])
            if r["base_pnl"] > 0 and r["pnl"] < r["base_pnl"]:
                winners_cut += 1
            if r["base_pnl"] < 0 and r["pnl"] > r["base_pnl"]:
                losers_saved += 1
            mfe = float(r.get("mfe_at_exit") or 0)
            if mfe > 1e-6:
                mfe_caps.append(float(r["exit_ret_bps"]) / mfe)
    sm = summarize_pnls(pnls)
    base = summarize_pnls(base_pnls)
    return {
        "id": cid,
        "family": family,
        "summary": sm,
        "base_summary": base,
        "delta_pnl": sm["total"] - base["total"],
        "delta_pf": (sm["pf"] or 0) - (base["pf"] or 0),
        "delta_worst": sm["worst"] - base["worst"],
        "delta_dd": sm["max_dd"] - base["max_dd"],
        "trigger_n": trigger_n,
        "trigger_rate": trigger_n / max(1, len(pnls)),
        "trigger_days": len(days),
        "trigger_symbols": len(syms),
        "median_hold": float(statistics.median(trigger_holds)) if trigger_holds else None,
        "winners_cut": winners_cut,
        "losers_saved": losers_saved,
        "median_mfe_capture": float(statistics.median(mfe_caps)) if mfe_caps else None,
        "pnls": pnls,
        "base_pnls": base_pnls,
        "trade_rows": rows,
    }


def eval_isolation(
    trades: list[dict[str, Any]],
    cand: dict[str, Any],
    *,
    learned_model: Optional[dict[str, Any]] = None,
    learned_thr: Optional[float] = None,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for tr in trades:
        path = tr["path"]
        fill_px = tr["fill_price"]
        base_pnl = tr["fixed600_pnl_yen"]
        if learned_model is not None and learned_thr is not None:
            ex = apply_learned_exit(path, learned_model, threshold=learned_thr)
        else:
            ex = apply_candidate(path, cand)
        if not ex.get("ok"):
            continue
        pnl = _pnl(fill_px, float(ex["exit_ret_bps"]))
        rows.append({
            "date": tr["date"],
            "symbol": tr["symbol"],
            "pnl": pnl,
            "base_pnl": base_pnl,
            "triggered": bool(ex.get("triggered")),
            "hold_sec": float(ex.get("hold_sec") or 0),
            "exit_ret_bps": float(ex["exit_ret_bps"]),
            "mfe_at_exit": float(ex.get("mfe_at_exit") or 0),
            "reason": ex.get("reason"),
        })
    return _aggregate_trade_exits(rows, cid=str(cand.get("id")), family=str(cand.get("family")))


def eval_isolation_from_cache(
    trade_rows: list[dict[str, Any]],
    *,
    cid: str,
    family: str,
    day_set: Optional[set[str]] = None,
) -> dict[str, Any]:
    rows = trade_rows if day_set is None else [r for r in trade_rows if r["date"] in day_set]
    return _aggregate_trade_exits(rows, cid=cid, family=family)


def passes_support(ev: dict[str, Any], thr: dict[str, int], *, is_time_only: bool) -> bool:
    if is_time_only or ev.get("family") in ("BASELINE", "TIME"):
        return True
    if ev["trigger_n"] < thr["min_trigger"]:
        return False
    if ev["trigger_days"] < thr["min_days"]:
        return False
    if ev["trigger_symbols"] < thr["min_symbols"]:
        return False
    return True


def patch_and_sim(
    panel_events: list[dict],
    trades_by_key: dict[tuple, dict],
    cand: dict[str, Any],
    sfn,
    *,
    learned_model: Optional[dict[str, Any]] = None,
    learned_thr: Optional[float] = None,
) -> dict[str, Any]:
    evs = [dict(e) for e in panel_events]
    reasons: Counter = Counter()
    for e in evs:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        key = (e["date"], e["symbol"], float(e["fill_time"]))
        tr = trades_by_key.get(key)
        if not tr:
            continue
        if learned_model is not None and learned_thr is not None:
            ex = apply_learned_exit(tr["path"], learned_model, threshold=learned_thr)
        else:
            ex = apply_candidate(tr["path"], cand)
        if not ex.get("ok"):
            continue
        e["canonical_exit_time"] = ex["exit_time"]
        e["canonical_exit_ret_bps"] = ex["exit_ret_bps"]
        e["canonical_hold_sec"] = ex.get("hold_sec")
        e["canonical_exit_reason"] = ex.get("reason")
        e["FIXED600_NET_BPS"] = ex["exit_ret_bps"]
        reasons[ex.get("reason")] += 1
    sim = simulate_joint(evs, score_fn=sfn)
    acc = [e for e in sim["events"] if e.get("accepted")]
    pnls = [float(e.get("realized_pnl_yen") or 0) for e in acc]
    return {
        "sim": {k: v for k, v in sim.items() if k != "events"},
        "accepted_n": len(acc),
        "summary": summarize_pnls(pnls),
        "policy_counts": dict(reasons),
        "events": sim["events"],
    }


def pareto_frontier(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """2-axis: return=total, risk=max_dd (less negative better). Keep non-dominated."""
    out = []
    for a in rows:
        dominated = False
        for b in rows:
            if a is b:
                continue
            # b better return and better (higher) DD and better worst
            better_ret = b["total"] >= a["total"] - 1e-9
            better_dd = b["max_dd"] >= a["max_dd"] - 1e-9
            better_worst = b["worst"] >= a["worst"] - 1e-9
            strictly = (
                b["total"] > a["total"] + 1e-9
                or b["max_dd"] > a["max_dd"] + 1e-9
                or b["worst"] > a["worst"] + 1e-9
            )
            if better_ret and better_dd and better_worst and strictly:
                dominated = True
                break
        if not dominated:
            out.append(a)
    return out


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_exit_gs_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    assert load_v1r().get("sha256") == V1R_SHA
    sfn = _sfn()

    cache = pickle.load(PANEL_CACHE.open("rb"))
    am_panel = [dict(e) for e in cache["am"]["panel"]]
    assert all(str(e["date"]) < FORBIDDEN for e in am_panel)

    print("  baseline FIXED600 joint...", flush=True)
    base_sim = simulate_joint([dict(e) for e in am_panel], score_fn=sfn)
    accepted = [e for e in base_sim["events"] if e.get("accepted")]
    print(f"  accepted fills={len(accepted)}", flush=True)
    base_pnls = [float(e.get("realized_pnl_yen") or 0) for e in accepted]
    base_sum = summarize_pnls(base_pnls)
    print(
        f"  FIXED600 n={base_sum['n']} PnL={base_sum['total']:.0f} "
        f"PF={base_sum['pf']} worst={base_sum['worst']:.0f}",
        flush=True,
    )

    pairs = sorted({(e["date"], e["symbol"]) for e in accepted})
    print(f"  load boards {len(pairs)}...", flush=True)
    boards = load_boards_for_symbols(pairs)

    trades: list[dict[str, Any]] = []
    trades_by_key: dict[tuple, dict] = {}
    for e in accepted:
        board = boards.get((e["date"], e["symbol"]))
        if board is None or board["t"].size == 0:
            continue
        sess_end = session_end_epoch(e["date"], e["session"])
        path = build_path(
            board,
            entry_price=float(e["fill_price"]),
            entry_t=float(e["fill_time"]),
            sess_end=sess_end,
        )
        path = attach_board_series(path, board)
        tr = {
            "date": e["date"],
            "symbol": e["symbol"],
            "session": e.get("session"),
            "fill_time": float(e["fill_time"]),
            "fill_price": float(e["fill_price"]),
            "fixed600_ret_bps": float(e.get("canonical_exit_ret_bps") or e.get("realized_ret_bps") or 0),
            "fixed600_pnl_yen": float(e.get("realized_pnl_yen") or 0),
            "path": path,
        }
        key = (tr["date"], tr["symbol"], tr["fill_time"])
        trades.append(tr)
        trades_by_key[key] = tr
    print(f"  path-ready trades={len(trades)}", flush=True)

    rule_cands = generate_all_candidates()
    counts = family_counts(rule_cands)
    print(f"  rule candidates={len(rule_cands)} families={counts}", flush=True)

    # ---- Isolation: all rule candidates ----
    print("  isolation sweep...", flush=True)
    iso_rows: list[dict[str, Any]] = []
    iso_by_id: dict[str, dict] = {}
    thr_all = support_thresholds(len(trades), len({t["date"] for t in trades}))
    for i, cand in enumerate(rule_cands):
        if i % 100 == 0:
            print(f"    {i}/{len(rule_cands)}...", flush=True)
        ev = eval_isolation(trades, cand)
        ev["support_ok"] = passes_support(
            ev, thr_all, is_time_only=cand["family"] in ("TIME", "BASELINE")
        )
        ev["complexity"] = 0.0 if cand["family"] == "BASELINE" else (
            0.5 if cand["family"] == "TIME" else 1.0 if cand["family"] in ("STOP", "TAKE") else 1.5
        )
        ev["composite"] = composite_score(ev["summary"], base_sum, complexity=ev["complexity"])
        slim = {k: v for k, v in ev.items() if k not in ("pnls", "base_pnls", "trade_rows")}
        slim.update({
            "total": ev["summary"]["total"],
            "pf": ev["summary"]["pf"],
            "worst": ev["summary"]["worst"],
            "max_dd": ev["summary"]["max_dd"],
            "gross_loss": ev["summary"]["gross_loss"],
            "n": ev["summary"]["n"],
        })
        iso_rows.append(slim)
        iso_by_id[cand["id"]] = ev
    print(f"  isolation done {len(iso_rows)}", flush=True)

    # Learned family on full hist for reporting (selection still nested)
    print("  learned EXIT fit (reporting + nested)...", flush=True)
    X_all, y_all, _ = build_decision_rows(trades)
    learned_cands_meta: list[dict[str, Any]] = []
    for C in (0.2, 0.5, 1.0):
        m = fit_logistic_l1(X_all, y_all, C=C)
        if m is None:
            continue
        for thr in (0.55, 0.65, 0.75):
            cid = f"LEARN_LOG_C{C}_t{int(thr*100)}"
            cand = {"family": "LEARNED", "id": cid}
            ev = eval_isolation(trades, cand, learned_model=m, learned_thr=thr)
            ev["support_ok"] = passes_support(ev, thr_all, is_time_only=False)
            ev["complexity"] = 2.0
            ev["composite"] = composite_score(ev["summary"], base_sum, complexity=2.0)
            slim = {k: v for k, v in ev.items() if k not in ("pnls", "base_pnls", "trade_rows")}
            slim.update({
                "total": ev["summary"]["total"], "pf": ev["summary"]["pf"],
                "worst": ev["summary"]["worst"], "max_dd": ev["summary"]["max_dd"],
                "gross_loss": ev["summary"]["gross_loss"], "n": ev["summary"]["n"],
                "learned_kind": "logistic_l1", "learned_C": C, "learned_thr": thr,
            })
            iso_rows.append(slim)
            iso_by_id[cid] = {**ev, "learned_model": m, "learned_thr": thr}
            learned_cands_meta.append(slim)
    for depth in (2, 3):
        m = fit_shallow_tree(X_all, y_all, depth=depth)
        if m is None:
            continue
        for thr in (0.55, 0.65, 0.75):
            cid = f"LEARN_TREE_d{depth}_t{int(thr*100)}"
            cand = {"family": "LEARNED", "id": cid}
            ev = eval_isolation(trades, cand, learned_model=m, learned_thr=thr)
            ev["support_ok"] = passes_support(ev, thr_all, is_time_only=False)
            ev["complexity"] = 2.2
            ev["composite"] = composite_score(ev["summary"], base_sum, complexity=2.2)
            slim = {k: v for k, v in ev.items() if k not in ("pnls", "base_pnls", "trade_rows")}
            slim.update({
                "total": ev["summary"]["total"], "pf": ev["summary"]["pf"],
                "worst": ev["summary"]["worst"], "max_dd": ev["summary"]["max_dd"],
                "gross_loss": ev["summary"]["gross_loss"], "n": ev["summary"]["n"],
                "learned_kind": "tree", "learned_depth": depth, "learned_thr": thr,
            })
            iso_rows.append(slim)
            iso_by_id[cid] = {**ev, "learned_model": m, "learned_thr": thr}
            learned_cands_meta.append(slim)
    counts["LEARNED"] = len(learned_cands_meta)
    total_cands = len(rule_cands) + len(learned_cands_meta)
    print(f"  total candidates explored={total_cands}", flush=True)

    # Family bests (isolation, support-filtered)
    family_best: dict[str, dict] = {}
    for row in iso_rows:
        if not row.get("support_ok", True) and row["family"] not in ("TIME", "BASELINE"):
            continue
        fam = row["family"]
        prev = family_best.get(fam)
        if prev is None or row["composite"] > prev["composite"]:
            family_best[fam] = row

    # ---- Nested Outer CV ----
    print("  nested outer CV...", flush=True)
    outer_rows = []
    outer_selected: list[str] = []
    cand_by_id = {c["id"]: c for c in rule_cands}
    for block, days in OUTER_BLOCKS.items():
        test_days = set(days)
        train_days = {d for ds in OUTER_BLOCKS.values() for d in ds} - test_days
        train_tr = [t for t in trades if t["date"] in train_days]
        test_tr = [t for t in trades if t["date"] in test_days]
        thr = support_thresholds(len(train_tr), len({t["date"] for t in train_tr}))
        base_train = summarize_pnls([t["fixed600_pnl_yen"] for t in train_tr])
        base_test = summarize_pnls([t["fixed600_pnl_yen"] for t in test_tr])

        best_id = "FIXED600"
        best_sc = -1e99
        best_learn = None
        # rule candidates — TRAIN-day slice of precomputed exits (grid fixed a priori)
        for cand in rule_cands:
            cached = iso_by_id[cand["id"]].get("trade_rows") or []
            ev = eval_isolation_from_cache(
                cached, cid=cand["id"], family=cand["family"], day_set=train_days
            )
            if not passes_support(ev, thr, is_time_only=cand["family"] in ("TIME", "BASELINE")):
                continue
            cx = 0.0 if cand["id"] == "FIXED600" else (
                0.5 if cand["family"] == "TIME" else 1.0
            )
            sc = composite_score(ev["summary"], base_train, complexity=cx)
            if cand["id"] == "FIXED600":
                sc += 25.0  # slight simplicity prior
            if sc > best_sc:
                best_sc = sc
                best_id = cand["id"]
                best_learn = None

        # learned on TRAIN only
        Xtr, ytr, _ = build_decision_rows(train_tr)
        for C in (0.2, 0.5, 1.0):
            m = fit_logistic_l1(Xtr, ytr, C=C)
            if m is None:
                continue
            for thr_p in (0.55, 0.65, 0.75):
                cid = f"LEARN_LOG_C{C}_t{int(thr_p*100)}"
                ev = eval_isolation(train_tr, {"id": cid, "family": "LEARNED"}, learned_model=m, learned_thr=thr_p)
                if not passes_support(ev, thr, is_time_only=False):
                    continue
                sc = composite_score(ev["summary"], base_train, complexity=2.0)
                if sc > best_sc:
                    best_sc = sc
                    best_id = cid
                    best_learn = (m, thr_p)
        for depth in (2, 3):
            m = fit_shallow_tree(Xtr, ytr, depth=depth)
            if m is None:
                continue
            for thr_p in (0.55, 0.65, 0.75):
                cid = f"LEARN_TREE_d{depth}_t{int(thr_p*100)}"
                ev = eval_isolation(train_tr, {"id": cid, "family": "LEARNED"}, learned_model=m, learned_thr=thr_p)
                if not passes_support(ev, thr, is_time_only=False):
                    continue
                sc = composite_score(ev["summary"], base_train, complexity=2.2)
                if sc > best_sc:
                    best_sc = sc
                    best_id = cid
                    best_learn = (m, thr_p)

        # evaluate once on TEST
        if best_learn is not None:
            te = eval_isolation(
                test_tr, {"id": best_id, "family": "LEARNED"},
                learned_model=best_learn[0], learned_thr=best_learn[1],
            )
        else:
            te = eval_isolation_from_cache(
                iso_by_id[best_id]["trade_rows"],
                cid=best_id, family=cand_by_id[best_id]["family"], day_set=test_days,
            )
        outer_selected.append(best_id)
        outer_rows.append({
            "block": block,
            "selected": best_id,
            "train_n": len(train_tr),
            "test_n": len(test_tr),
            "test_total": te["summary"]["total"],
            "test_base_total": base_test["total"],
            "delta_pnl": te["summary"]["total"] - base_test["total"],
            "delta_pf": (te["summary"]["pf"] or 0) - (base_test["pf"] or 0),
            "delta_worst": te["summary"]["worst"] - base_test["worst"],
            "delta_dd": te["summary"]["max_dd"] - base_test["max_dd"],
            "test_pf": te["summary"]["pf"],
            "test_worst": te["summary"]["worst"],
            "test_dd": te["summary"]["max_dd"],
            "winners_cut": te["winners_cut"],
            "losers_saved": te["losers_saved"],
        })
        print(
            f"    fold {block}: selected={best_id} delta_pnl={outer_rows[-1]['delta_pnl']:.0f}",
            flush=True,
        )

    # Global winner from isolation among support-ok + outer vote
    vote = Counter(outer_selected)
    vote_winner = vote.most_common(1)[0][0]
    # also pick isolation global by composite among support
    supported = [r for r in iso_rows if r.get("support_ok", True) or r["family"] in ("TIME", "BASELINE")]
    iso_winner = max(supported, key=lambda r: r["composite"])
    # Improvement gate vs FIXED600 on outer mean
    mean_delta = statistics.mean([r["delta_pnl"] for r in outer_rows])
    mean_pf_d = statistics.mean([r["delta_pf"] for r in outer_rows])
    mean_worst_d = statistics.mean([r["delta_worst"] for r in outer_rows])
    mean_dd_d = statistics.mean([r["delta_dd"] for r in outer_rows])
    cross_ok = (
        mean_delta >= -1.0
        and mean_pf_d >= -0.05
        and mean_dd_d >= -1.0
        and mean_worst_d >= 0.0
    )

    # Prefer FIXED600 if vote winner fails gate or is FIXED600
    frozen_id = vote_winner
    if frozen_id != "FIXED600" and not cross_ok:
        # check if FIXED600 was competitive
        frozen_id = "FIXED600"
    # if vote says FIXED600, keep it
    if vote_winner == "FIXED600":
        frozen_id = "FIXED600"

    # Top family winners for full replay (isolation-supported)
    replay_ids = ["FIXED600"]
    for fam, row in family_best.items():
        if row["id"] not in replay_ids:
            replay_ids.append(row["id"])
    # add Pareto candidates
    pareto_src = [
        {
            "id": r["id"], "family": r["family"], "total": r["total"],
            "pf": r["pf"], "max_dd": r["max_dd"], "worst": r["worst"],
            "composite": r["composite"],
        }
        for r in supported if r["family"] != "LEARNED" or r.get("support_ok")
    ]
    # limit to top 40 by composite for pareto clarity
    pareto_src = sorted(pareto_src, key=lambda x: -x["composite"])[:80]
    frontier = pareto_frontier(pareto_src)
    for p in frontier[:12]:
        if p["id"] not in replay_ids:
            replay_ids.append(p["id"])
    if frozen_id not in replay_ids:
        replay_ids.append(frozen_id)
    # cap full replays
    replay_ids = replay_ids[:25]
    print(f"  full replay {len(replay_ids)} candidates...", flush=True)

    full_replay_rows = []
    full_by_id = {}
    for rid in replay_ids:
        ev0 = iso_by_id.get(rid)
        lm = thr_l = None
        if ev0 and ev0.get("learned_model") is not None:
            lm, thr_l = ev0["learned_model"], ev0["learned_thr"]
            cand = {"id": rid, "family": "LEARNED"}
        else:
            cand = cand_by_id.get(rid) or {"id": rid, "family": "BASELINE", "fixed_hold_sec": 600.0}
        res = patch_and_sim(am_panel, trades_by_key, cand, sfn, learned_model=lm, learned_thr=thr_l)
        row = {
            "id": rid,
            "family": cand.get("family"),
            "accepted_n": res["accepted_n"],
            "total": res["summary"]["total"],
            "pf": res["summary"]["pf"],
            "worst": res["summary"]["worst"],
            "max_dd": res["summary"]["max_dd"],
            "gross_loss": res["summary"]["gross_loss"],
            "delta_pnl": res["summary"]["total"] - base_sum["total"],
            "delta_pf": (res["summary"]["pf"] or 0) - (base_sum["pf"] or 0),
            "delta_worst": res["summary"]["worst"] - base_sum["worst"],
            "delta_dd": res["summary"]["max_dd"] - base_sum["max_dd"],
            "policy_counts": res["policy_counts"],
        }
        full_replay_rows.append(row)
        full_by_id[rid] = res
        print(f"    {rid}: PnL={row['total']:.0f} PF={row['pf']} worst={row['worst']:.0f}", flush=True)

    # Re-decide freeze using full replay of vote winner vs FIXED600
    fr_win = full_by_id.get(frozen_id) or full_by_id["FIXED600"]
    fr_base = full_by_id["FIXED600"]
    improve_full = (
        fr_win["summary"]["total"] >= fr_base["summary"]["total"] - 1.0
        and (fr_win["summary"]["pf"] or 0) >= (fr_base["summary"]["pf"] or 0) - 0.05
        and fr_win["summary"]["max_dd"] >= fr_base["summary"]["max_dd"] - 1.0
        and fr_win["summary"]["worst"] >= fr_base["summary"]["worst"] - 1.0
    )
    if frozen_id != "FIXED600" and not (cross_ok and improve_full):
        # keep as Pareto note but freeze FIXED600 if not clearly better
        if not (
            fr_win["summary"]["worst"] > fr_base["summary"]["worst"] + 20000
            and fr_win["summary"]["total"] >= fr_base["summary"]["total"] * 0.9
        ):
            frozen_id = "FIXED600"

    # LODO / LOSO on frozen vs FIXED600
    print("  LODO / LOSO...", flush=True)
    all_days = sorted({t["date"] for t in trades})
    frozen_cand = cand_by_id.get(frozen_id) or {"id": frozen_id, "family": "BASELINE", "fixed_hold_sec": 600.0}
    frozen_lm = frozen_thr = None
    if frozen_id in iso_by_id and iso_by_id[frozen_id].get("learned_model") is not None:
        frozen_lm = iso_by_id[frozen_id]["learned_model"]
        frozen_thr = iso_by_id[frozen_id]["learned_thr"]

    lodo_rows = []
    for day in all_days:
        panel_wo = [e for e in am_panel if e["date"] != day]
        if frozen_id == "FIXED600":
            delta = 0.0
            ct = bt = None
        else:
            res = patch_and_sim(
                panel_wo, trades_by_key, frozen_cand, sfn,
                learned_model=frozen_lm, learned_thr=frozen_thr,
            )
            base = patch_and_sim(
                panel_wo, trades_by_key,
                {"id": "FIXED600", "family": "BASELINE", "fixed_hold_sec": 600.0}, sfn,
            )
            delta = res["summary"]["total"] - base["summary"]["total"]
            ct, bt = res["summary"]["total"], base["summary"]["total"]
        lodo_rows.append({"left_out": day, "cand_total": ct, "base_total": bt, "delta": delta})

    by_sym = defaultdict(float)
    for t in trades:
        by_sym[t["symbol"]] += t["fixed600_pnl_yen"]
    top_syms = [s for s, _ in sorted(by_sym.items(), key=lambda x: -x[1])[:6]]
    if "285A" not in top_syms:
        top_syms.append("285A")
    loso_rows = []
    for sym in top_syms:
        panel_wo = [e for e in am_panel if e["symbol"] != sym]
        if frozen_id == "FIXED600":
            loso_rows.append({"excluded": sym, "delta": 0.0, "cand_total": None, "base_total": None})
            continue
        res = patch_and_sim(
            panel_wo, trades_by_key, frozen_cand, sfn,
            learned_model=frozen_lm, learned_thr=frozen_thr,
        )
        base = patch_and_sim(
            panel_wo, trades_by_key,
            {"id": "FIXED600", "family": "BASELINE", "fixed_hold_sec": 600.0}, sfn,
        )
        loso_rows.append({
            "excluded": sym,
            "cand_total": res["summary"]["total"],
            "base_total": base["summary"]["total"],
            "delta": res["summary"]["total"] - base["summary"]["total"],
            "cand_pf": res["summary"]["pf"],
        })

    # Session AM/PM isolation split for frozen
    session_rows = []
    for sess in ("AM", "PM"):
        sub = [t for t in trades if str(t.get("session") or "").upper().startswith(sess[0]) or t.get("session") == sess]
        if not sub:
            # try date-based: most are AM in this panel
            sub = trades if sess == "AM" else []
        if not sub:
            session_rows.append({"session": sess, "n": 0})
            continue
        if frozen_id == "FIXED600":
            sm = summarize_pnls([t["fixed600_pnl_yen"] for t in sub])
            session_rows.append({"session": sess, "n": sm["n"], "total": sm["total"], "delta": 0.0})
        else:
            ev = eval_isolation(
                sub, frozen_cand, learned_model=frozen_lm, learned_thr=frozen_thr
            )
            session_rows.append({
                "session": sess, "n": ev["summary"]["n"],
                "total": ev["summary"]["total"],
                "base": ev["base_summary"]["total"],
                "delta": ev["delta_pnl"],
            })

    # Freeze BEFORE 8/10
    freeze_body = {
        "manifest_id": "V1R_EXIT_GLOBAL_SEARCH_CANDIDATE_V1",
        "kind": "research_exit_candidate_not_production",
        "entry_frozen": True,
        "v1r_sha": V1R_SHA,
        "model_sha": MODEL_ARTIFACT_SHA,
        "frozen_exit_id": frozen_id,
        "candidate_spec": frozen_cand if frozen_id in cand_by_id else {"id": frozen_id},
        "outer_vote": dict(vote),
        "outer_mean_delta_pnl": mean_delta,
        "cross_fit_gate": cross_ok,
        "selection_basis": "pre_20260810_historical_only",
        "locked_holdout": FORBIDDEN,
        "support_thresholds_full": thr_all,
        "total_candidates_explored": total_cands,
        "family_counts": counts,
    }
    freeze_sha = hashlib.sha256(
        json.dumps({k: v for k, v in freeze_body.items() if k != "sha256"}, sort_keys=True, default=str).encode()
    ).hexdigest()
    freeze_body["sha256"] = freeze_sha
    (OUT / "V1R_EXIT_GLOBAL_SEARCH_CANDIDATE_V1.json").write_text(
        json.dumps(freeze_body, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
    )
    print(f"  FROZEN exit id={frozen_id} sha={freeze_sha[:16]}...", flush=True)

    # ---- 8/10 holdout ----
    print("  8/10 locked holdout...", flush=True)
    from research.e1_x34c_passive_deployability.events import build_events
    from research.e1_x36_joint_allocator.panel import enrich_events
    from small_paper.v1r_day_engine import (
        _load_boards,
        _planned_anchors_retrospective,
        resolve_pre0905_am_universe,
        score_fn_frozen,
    )

    uni = resolve_pre0905_am_universe("20260810")
    planned = _planned_anchors_retrospective("20260810", uni["symbols"])
    boards10 = _load_boards([("20260810", s) for s in uni["symbols"]])
    raw10 = build_events(planned, boards10)
    panel10 = enrich_events(raw10, boards10)
    sim10 = simulate_joint([dict(e) for e in panel10], score_fn=score_fn_frozen())
    acc10 = [e for e in sim10["events"] if e.get("accepted")]
    tk10: dict[tuple, dict] = {}
    for e in acc10:
        b = boards10.get((e["date"], e["symbol"]))
        if b is None:
            continue
        se = session_end_epoch(e["date"], e["session"])
        path = attach_board_series(
            build_path(b, entry_price=float(e["fill_price"]), entry_t=float(e["fill_time"]), sess_end=se),
            b,
        )
        key = (e["date"], e["symbol"], float(e["fill_time"]))
        tk10[key] = {
            "date": e["date"], "symbol": e["symbol"], "session": e.get("session"),
            "fill_time": float(e["fill_time"]), "fill_price": float(e["fill_price"]),
            "fixed600_ret_bps": float(e.get("canonical_exit_ret_bps") or 0),
            "fixed600_pnl_yen": float(e.get("realized_pnl_yen") or 0),
            "path": path,
        }

    # Evaluate family bests + frozen + FIXED600 on 8/10 (reporting only)
    holdout_cands = ["FIXED600", frozen_id] + [family_best[f]["id"] for f in family_best]
    holdout_cands = list(dict.fromkeys(holdout_cands))[:30]
    holdout_rows = []
    holdout_by_id = {}
    for rid in holdout_cands:
        ev0 = iso_by_id.get(rid)
        lm = thr_l = None
        if ev0 and ev0.get("learned_model") is not None:
            # REFIT forbidden on 8/10 — use pre-8/10 full-hist model only for reporting of that id
            lm, thr_l = ev0["learned_model"], ev0["learned_thr"]
            cand = {"id": rid, "family": "LEARNED"}
        else:
            cand = cand_by_id.get(rid) or {"id": rid, "family": "BASELINE", "fixed_hold_sec": 600.0}
        res = patch_and_sim(
            panel10, tk10, cand, score_fn_frozen(), learned_model=lm, learned_thr=thr_l
        )
        holdout_rows.append({
            "id": rid, "family": cand.get("family"),
            "accepted_n": res["accepted_n"],
            "total": res["summary"]["total"], "pf": res["summary"]["pf"],
            "worst": res["summary"]["worst"], "max_dd": res["summary"]["max_dd"],
            "policy_counts": res["policy_counts"],
        })
        holdout_by_id[rid] = res

    # 5706 worst-case table
    base10 = holdout_by_id["FIXED600"]
    worst_e = None
    for e in base10["events"]:
        if not e.get("accepted"):
            continue
        pnl = float(e.get("realized_pnl_yen") or 0)
        if worst_e is None or pnl < float(worst_e.get("realized_pnl_yen") or 0):
            worst_e = e
    worst5706_rows = []
    if worst_e is not None:
        wkey = ("20260810", worst_e["symbol"], float(worst_e["fill_time"]))
        wtr = tk10.get(wkey)
        base_pnl = float(worst_e.get("realized_pnl_yen") or 0)
        for rid in holdout_cands:
            if wtr is None:
                continue
            ev0 = iso_by_id.get(rid)
            if ev0 and ev0.get("learned_model") is not None:
                ex = apply_learned_exit(wtr["path"], ev0["learned_model"], threshold=ev0["learned_thr"])
            else:
                cand = cand_by_id.get(rid) or {"id": rid, "fixed_hold_sec": 600.0}
                ex = apply_candidate(wtr["path"], cand)
            if not ex.get("ok"):
                continue
            pnl = _pnl(wtr["fill_price"], float(ex["exit_ret_bps"]))
            # buy1 at exit
            path = wtr["path"]
            buy1 = None
            if path.get("ok") and path["offs"].size:
                j = int(np.searchsorted(path["offs"], float(ex["hold_sec"]), side="left"))
                j = min(j, path["offs"].size - 1)
                # reconstruct buy1 from ret
                buy1 = wtr["fill_price"] * (1.0 + float(ex["exit_ret_bps"]) / 10000.0)
            worst5706_rows.append({
                "candidate": rid,
                "symbol": worst_e["symbol"],
                "triggered": ex.get("triggered"),
                "reason": ex.get("reason"),
                "hold_sec": ex.get("hold_sec"),
                "exit_ret_bps": ex.get("exit_ret_bps"),
                "exit_pnl": pnl,
                "baseline_pnl": base_pnl,
                "avoided_loss": pnl - base_pnl,
                "executable_buy1": buy1,
            })

    # Extreme metrics for report answers
    best_pnl = max(supported, key=lambda r: r["total"])
    best_pf = max((r for r in supported if r["pf"] is not None), key=lambda r: r["pf"])
    best_dd = max(supported, key=lambda r: r["max_dd"])
    best_worst = max(supported, key=lambda r: r["worst"])

    better_than_base = [
        r for r in supported
        if r["id"] != "FIXED600"
        and r["total"] >= base_sum["total"] - 1
        and (r["pf"] or 0) >= (base_sum["pf"] or 0) - 0.02
        and r["worst"] >= base_sum["worst"] - 1
    ]

    # Winner cut / loser saved for frozen (isolation)
    frozen_iso = iso_by_id.get(frozen_id) or iso_by_id["FIXED600"]
    # day / symbol concentration
    day_pnl = defaultdict(float)
    sym_pnl = defaultdict(float)
    for t in trades:
        day_pnl[t["date"]] += t["fixed600_pnl_yen"]
        sym_pnl[t["symbol"]] += t["fixed600_pnl_yen"]
    top_day_share = max(day_pnl.values()) / max(1.0, abs(sum(day_pnl.values())) or 1.0) if day_pnl else 0
    top_sym_share = max(sym_pnl.values()) / max(1.0, abs(sum(sym_pnl.values())) or 1.0) if sym_pnl else 0

    # Verdict
    robust_found = (
        frozen_id != "FIXED600"
        and cross_ok
        and improve_full
        and len(better_than_base) >= 1
    )
    pareto_only = (not robust_found) and len(frontier) > 1 and any(
        p["id"] != "FIXED600" and (p["worst"] > base_sum["worst"] or p["max_dd"] > base_sum["max_dd"])
        for p in frontier
    )
    if robust_found:
        verdict = "V1R_EXIT_GLOBAL_SEARCH_ROBUST_CANDIDATE_FOUND"
    elif pareto_only:
        verdict = "V1R_EXIT_GLOBAL_SEARCH_PARETO_CANDIDATES_ONLY"
    else:
        verdict = "V1R_EXIT_GLOBAL_SEARCH_FIXED600_REMAINS_BEST"

    saved_5706 = any(
        r.get("avoided_loss", 0) > 50000 and r.get("triggered") for r in worst5706_rows
    )

    # Sheets
    def fam_sheet(fam: str) -> list[dict]:
        rows = [r for r in iso_rows if r["family"] == fam]
        rows = sorted(rows, key=lambda x: -x.get("composite", -1e99))[:80]
        return [{k: v for k, v in r.items() if k != "summary" and k != "base_summary"} for r in rows]

    overview = [{
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": verdict,
        "total_candidates": total_cands,
        "rule_candidates": len(rule_cands),
        "learned_candidates": len(learned_cands_meta),
        "accepted_fills": len(trades),
        "baseline_pnl": base_sum["total"],
        "baseline_pf": base_sum["pf"],
        "baseline_worst": base_sum["worst"],
        "baseline_dd": base_sum["max_dd"],
        "frozen_exit_id": frozen_id,
        "freeze_sha": freeze_sha,
        "outer_mean_delta_pnl": mean_delta,
        "cross_fit_gate": cross_ok,
        "better_than_base_n": len(better_than_base),
        "production_mutation": "NONE",
        "entry_frozen": True,
        "holdout_locked": FORBIDDEN,
        "saved_5706_any": saved_5706,
        "support_thresholds": thr_all,
        "iso_winner": iso_winner["id"],
        "vote_winner": vote_winner,
    }]

    sheets = {
        "Overview": overview,
        "Search_Space": [{"family": k, "n": v} for k, v in sorted(counts.items())],
        "Candidate_Count": [{"metric": "total", "n": total_cands}, {"metric": "rule", "n": len(rule_cands)}, {"metric": "learned", "n": len(learned_cands_meta)}],
        "Time": fam_sheet("TIME"),
        "Stop": fam_sheet("STOP") + fam_sheet("STOP_TAKE"),
        "Take": fam_sheet("TAKE"),
        "Trailing": fam_sheet("TRAIL") + fam_sheet("ABS_GIVEBACK") + fam_sheet("RECOVERY_DONE"),
        "MAE_Recovery": fam_sheet("MAE_RECOVERY"),
        "NoProgress": fam_sheet("NO_PROGRESS"),
        "BoardFlow": fam_sheet("IMBALANCE") + fam_sheet("BID_DEPTH") + fam_sheet("SPREAD") + fam_sheet("EVENT_DECAY") + fam_sheet("MOM_FADE"),
        "StateTransition": fam_sheet("EARLY_FAIL") + fam_sheet("STATE_SEQ") + fam_sheet("SELL_FAIL"),
        "Hybrid": fam_sheet("HYBRID"),
        "Learned": fam_sheet("LEARNED"),
        "InnerCV": [{"note": "inner selection = train-only composite within each outer fold", "support": thr_all}],
        "OuterCV": outer_rows,
        "Isolation": sorted(
            [{k: v for k, v in r.items() if k not in ("summary", "base_summary")} for r in supported],
            key=lambda x: -x.get("composite", -1e99),
        )[:200],
        "FullReplay": full_replay_rows,
        "LODO": lodo_rows,
        "LOSO": loso_rows,
        "Pareto": frontier,
        "FrozenCandidates": [
            {"role": "frozen", "id": frozen_id},
            *[{"role": "family_best", "family": f, "id": family_best[f]["id"],
               "total": family_best[f]["total"], "pf": family_best[f]["pf"],
               "worst": family_best[f]["worst"], "composite": family_best[f]["composite"]}
              for f in sorted(family_best)],
        ],
        "Holdout0810": holdout_rows,
        "Worst5706": worst5706_rows,
        "Session": session_rows,
        "FamilyBest": [
            {"family": f, **{k: family_best[f][k] for k in (
                "id", "total", "pf", "worst", "max_dd", "delta_pnl", "delta_worst",
                "trigger_n", "winners_cut", "losers_saved", "composite", "support_ok",
            ) if k in family_best[f]}}
            for f in sorted(family_best)
        ],
    }
    write_xlsx(sheets, OUT / "v1r_exit_global_search.xlsx")

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": verdict,
        "answers": {
            "1_explored": {"total": total_cands, "by_family": counts},
            "2_family_best": {f: family_best[f]["id"] for f in family_best},
            "3_better_than_fixed600": [r["id"] for r in better_than_base[:20]],
            "4_pnl_max": {"id": best_pnl["id"], "total": best_pnl["total"]},
            "5_pf_max": {"id": best_pf["id"], "pf": best_pf["pf"]},
            "6_dd_min": {"id": best_dd["id"], "max_dd": best_dd["max_dd"]},
            "7_worst_improve_max": {"id": best_worst["id"], "worst": best_worst["worst"],
                                    "delta_worst": best_worst["worst"] - base_sum["worst"]},
            "8_pareto": frontier[:15],
            "9_nested_cv": outer_rows,
            "10_full_replay": full_replay_rows,
            "11_winners_cut": frozen_iso.get("winners_cut"),
            "12_losers_saved": frozen_iso.get("losers_saved"),
            "13_day_concentration": {"top_day_share_approx": top_day_share, "lodo": lodo_rows},
            "14_symbol_concentration": {"top_sym_share_approx": top_sym_share, "loso": loso_rows},
            "15_holdout_0810": holdout_rows,
            "16_saved_5706": saved_5706,
            "17_production_worthy": robust_found,
        },
        "baseline": base_sum,
        "frozen": freeze_body,
        "support_thresholds": thr_all,
        "production_mutation": "NONE",
        "worst5706": worst5706_rows,
    }
    (OUT / "report.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
    )

    md = f"""# V1R EXIT Global Search

## Verdict
`{verdict}`

## Explored
- Total candidates: **{total_cands}**
- By family: `{counts}`
- Support thresholds: `{thr_all}`

## Baseline FIXED600
- n={base_sum['n']} PnL={base_sum['total']:.0f} PF={base_sum['pf']} worst={base_sum['worst']:.0f} DD={base_sum['max_dd']:.0f}

## Family bests (isolation composite)
{chr(10).join(f"- {f}: `{family_best[f]['id']}` PnL={family_best[f]['total']:.0f} PF={family_best[f]['pf']} worst={family_best[f]['worst']:.0f}" for f in sorted(family_best))}

## Nested Outer CV
- Vote: `{dict(vote)}`
- Mean delta PnL: {mean_delta:.0f}
- Cross-fit gate: {cross_ok}

## Frozen EXIT
- id=`{frozen_id}`
- sha=`{freeze_sha}`

## Extremes (isolation, support-ok)
- PnL max: `{best_pnl['id']}` ({best_pnl['total']:.0f})
- PF max: `{best_pf['id']}` ({best_pf['pf']})
- DD best: `{best_dd['id']}` ({best_dd['max_dd']:.0f})
- Worst best: `{best_worst['id']}` ({best_worst['worst']:.0f})

## Better than FIXED600 (strict isolation gate)
- count={len(better_than_base)}
- sample: {[r['id'] for r in better_than_base[:10]]}

## 8/10 Holdout
- FIXED600 worst symbol: `{worst_e['symbol'] if worst_e else None}` pnl={base_pnl if worst_e else None}
- Any candidate saved 5706-class (>50k avoided & triggered): {saved_5706}

## Constraints
- ENTRY frozen. Production mutation: NONE.
- 20260810 used only as locked holdout after freeze.
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")
    print(json.dumps({"verdict": verdict, "frozen": frozen_id, "total_cands": total_cands}, indent=2))
    return report


if __name__ == "__main__":
    main()
