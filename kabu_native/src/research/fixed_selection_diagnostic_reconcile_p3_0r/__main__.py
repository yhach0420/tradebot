"""P3-0R runner. Reconcile selection fill diagnostic only. Clock frozen. No new strategy."""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook

NATIVE = Path(__file__).resolve().parents[3]
if str(NATIVE / "src") not in sys.path:
    sys.path.insert(0, str(NATIVE / "src"))
if str(NATIVE / "scripts") not in sys.path:
    sys.path.insert(0, str(NATIVE / "scripts"))

from research.dynamic_anchor_p2_1.inventory import build_inventory
from research.fixed_selection_diagnostic_reconcile_p3_0r import FULL14, MAX_WORKERS, P1_FILLS, VERDICT_BLOCKED
from research.fixed_selection_diagnostic_reconcile_p3_0r.publish import build_report, write_artifacts
from research.fixed_selection_diagnostic_reconcile_p3_0r.replay import replay_day

P1_REPORT = NATIVE / "results" / "research" / "current_runtime_full_capture_recalc_p1" / "report.json"
P3_0_XLSX = NATIVE / "results" / "research" / "fixed_anchor_mechanism_audit_p3_0" / "audit.xlsx"


def _load_p3_0_candidates() -> list[dict]:
    wb = load_workbook(P3_0_XLSX, read_only=True, data_only=True)
    ws = wb["Anchor_CrossSection"]
    header = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c) for c in row]
            continue
        rec = {header[j]: row[j] for j in range(len(header))}
        if str(rec.get("date")) in set(FULL14):
            rows.append(rec)
    wb.close()
    return rows


def _load_p1_trades() -> list[dict]:
    p1 = json.loads(P1_REPORT.read_text(encoding="utf-8"))
    return [t for t in (p1.get("trades") or []) if str(t.get("date")) in set(FULL14)]


def main() -> int:
    if not P1_REPORT.is_file():
        rep = build_report(diag_rows=[], recon_rows=[], failed=["NO_P1"], blocked=True, blocked_reason="NO_P1_REPORT")
        paths = write_artifacts(rep)
        print("P3_0R_BLOCKED NO_P1_REPORT", paths["report_json"], flush=True)
        return 1
    if not P3_0_XLSX.is_file():
        rep = build_report(diag_rows=[], recon_rows=[], failed=["NO_P3_0"], blocked=True, blocked_reason="NO_P3_0_XLSX")
        paths = write_artifacts(rep)
        print("P3_0R_BLOCKED NO_P3_0_XLSX", paths["report_json"], flush=True)
        return 1

    inv = {r["date"]: r for r in build_inventory()}
    cands = _load_p3_0_candidates()
    trades = _load_p1_trades()
    print(f"P3-0R candidates={len(cands)} p1_full14_trades={len(trades)} workers={MAX_WORKERS}", flush=True)
    if len(trades) != P1_FILLS:
        rep = build_report(
            diag_rows=[],
            recon_rows=[],
            failed=[f"P1_TRADE_N={len(trades)}"],
            blocked=True,
            blocked_reason=f"P1_FULL14_TRADES_{len(trades)}_NE_{P1_FILLS}",
        )
        paths = write_artifacts(rep)
        print("P3_0R_BLOCKED", rep["blocked_reason"], paths["report_json"], flush=True)
        return 1

    by_c: dict[str, list] = defaultdict(list)
    for r in cands:
        by_c[str(r["date"])].append(r)
    by_t: dict[str, list] = defaultdict(list)
    for t in trades:
        by_t[str(t["date"])].append(t)

    jobs = []
    for day in FULL14:
        row = inv.get(day)
        if not row or not row.get("capture_path") or not row.get("universe_symbols"):
            rep = build_report(
                diag_rows=[], recon_rows=[], failed=[day], blocked=True, blocked_reason=f"MISSING_{day}"
            )
            write_artifacts(rep)
            print("P3_0R_BLOCKED MISSING", day, flush=True)
            return 1
        jobs.append(
            {
                "date": day,
                "capture_path": row["capture_path"],
                "universe": row["universe_symbols"],
                "candidates": by_c.get(day) or [],
                "canonical_trades": by_t.get(day) or [],
            }
        )

    results = []
    failed = []
    with ProcessPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(replay_day, j): j["date"] for j in jobs}
        for fut in as_completed(futs):
            day = futs[fut]
            try:
                out = fut.result()
            except Exception as exc:
                print(f"FAIL {day} {exc!r}", flush=True)
                failed.append(day)
                continue
            if not out.get("ok"):
                print(f"FAIL {day} {out.get('blocker')}", flush=True)
                failed.append(day)
                continue
            n_m = sum(1 for r in out.get("reconcile_rows") or [] if r.get("klass") == "MATCH")
            n_tr = len(out.get("reconcile_rows") or [])
            print(
                f"OK {day} recon={n_m}/{n_tr} diag={len(out.get('diag_rows') or [])} sec={out.get('elapsed_sec')}",
                flush=True,
            )
            results.append(out)

    if failed or len(results) != 14:
        rep = build_report(
            diag_rows=[],
            recon_rows=[],
            failed=failed,
            blocked=True,
            blocked_reason="DAY_REPLAY_FAILED",
        )
        paths = write_artifacts(rep)
        print("P3_0R_BLOCKED DAY_REPLAY_FAILED", paths["report_json"], flush=True)
        return 1

    diag = [r for d in results for r in (d.get("diag_rows") or [])]
    recon = [r for d in results for r in (d.get("reconcile_rows") or [])]
    rep = build_report(diag_rows=diag, recon_rows=recon, failed=[], blocked=False)
    paths = write_artifacts(rep)
    print(rep["verdict"], paths["report_json"], flush=True)
    print(
        f"SELECTED_MATCH={rep['CANONICAL_SELECTED_MATCH']} "
        f"FILL_REPRODUCED={rep['CANONICAL_FILL_REPRODUCED']} "
        f"SEL={rep['SELECTION_RESULT']} MECH={rep['FIXED_ANCHOR_MECHANISM']}",
        flush=True,
    )
    return 0 if not str(rep["verdict"]).endswith("BLOCKED") else 1


if __name__ == "__main__":
    raise SystemExit(main())
