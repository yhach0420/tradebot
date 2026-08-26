"""Write P3-0R report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.fixed_selection_diagnostic_reconcile_p3_0r import (
    ANALYSIS_ID,
    CLOCK_FULL14,
    CLOCK_REST11,
    DOCUMENT_ID,
    FULL14,
    P1_FILLS,
    P3_0_INDEPENDENT_FILL_N,
    P3_0_SELECTED_N,
    P2_3_ADMITTED,
    PREDECLARED_TOP3,
    REST11,
    ROOT_CAUSE_COMPACT,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.fixed_selection_diagnostic_reconcile_p3_0r.metrics import (
    group_metrics,
    mechanism,
    mismatch_counts,
    rank_strata,
    selection_pair,
    selection_verdict,
    slice_label,
    within_anchor,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "fixed_selection_diagnostic_reconcile_p3_0r"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 88


def _filt(rows, days):
    ds = set(days)
    return [r for r in rows if str(r.get("date")) in ds]


def _sel_blocks(rows: list[dict[str, Any]]) -> dict[str, Any]:
    elig = [r for r in rows if r.get("feature_evaluable")]
    sel = [r for r in elig if r.get("selected")]
    nos = [r for r in elig if not r.get("selected")]
    sm = group_metrics(sel)
    nm = group_metrics(nos)
    return {"SELECTED": sm, "NOT_SELECTED": nm, "pair": selection_pair(sm, nm), "eligible_n": len(elig)}


def build_report(
    *,
    diag_rows: list[dict[str, Any]],
    recon_rows: list[dict[str, Any]],
    failed: list[str],
    blocked: bool,
    blocked_reason: str = "",
) -> dict[str, Any]:
    now = datetime.now(JST).isoformat(timespec="seconds")
    recon_full = [r for r in recon_rows if str(r.get("date")) in set(FULL14)]
    n_canon = len(recon_full)
    n_selected_present = sum(1 for r in recon_full if r.get("selected_present"))
    n_match = sum(1 for r in recon_full if r.get("klass") == "MATCH")
    counts = mismatch_counts(recon_full)
    missing = [r for r in recon_full if not r.get("selected_present")]

    diag_full = [r for r in diag_rows if str(r.get("date")) in set(FULL14)]
    blk_all = _sel_blocks(diag_full)
    blk_top3 = _sel_blocks(_filt(diag_full, PREDECLARED_TOP3))
    blk_rest = _sel_blocks(_filt(diag_full, REST11))
    sel_label = selection_verdict(blk_all["pair"], blk_rest["pair"])
    strata = rank_strata(diag_full)
    within = within_anchor(diag_full)
    within_top3 = within_anchor(_filt(diag_full, PREDECLARED_TOP3))
    within_rest = within_anchor(_filt(diag_full, REST11))
    mech = mechanism(sel_label)

    exit_compared = [r for r in recon_full if r.get("klass") == "MATCH" and r.get("exit_pnl_match") is not None]
    exit_pnl_ok = sum(1 for r in exit_compared if r.get("exit_pnl_match"))
    exit_t_ok = sum(1 for r in recon_full if r.get("klass") == "MATCH" and r.get("exit_time_match"))

    if blocked or n_canon != P1_FILLS:
        verdict = VERDICT_BLOCKED
        recon_status = "FAIL"
    elif n_selected_present != P1_FILLS:
        verdict = VERDICT_BLOCKED
        recon_status = "FAIL"
        blocked_reason = blocked_reason or "MISSING_SELECTED_CANONICAL_FILL"
    elif n_match != P1_FILLS:
        verdict = VERDICT_BLOCKED
        recon_status = "FAIL"
        blocked_reason = blocked_reason or f"CANONICAL_FILL_REPRODUCED={n_match}/{P1_FILLS}"
    else:
        recon_status = "PASS"
        verdict = VERDICT_OK

    old_fill = sum(1 for r in diag_full if r.get("selected") and r.get("old_independent_filled"))
    new_fill = blk_all["SELECTED"]["fill_n"]
    xlsx_sel_match = sum(1 for r in recon_full if r.get("p3_0_xlsx_selected"))
    harvest_sel_n = sum(1 for r in diag_full if r.get("selected"))

    report = {
        "task": "P3-0R",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "label": TASK_LABEL,
        "not": ["OOS", "prospective", "robust", "strategy validation"],
        "generated_at_jst": now,
        "clock_frozen": {
            "FULL14": CLOCK_FULL14,
            "REST11": CLOCK_REST11,
            "recomputed": False,
        },
        "p3_0_old": {
            "SELECTED_n": P3_0_SELECTED_N,
            "independent_fill_n": P3_0_INDEPENDENT_FILL_N,
            "xlsx_selected_among_canonical_fills": xlsx_sel_match,
            "observed_old_fill_on_harvest_selected": old_fill,
            "note": (
                "P3-0 xlsx selected is post-hoc compacted rescore. "
                "Corrected SELECTED is harvest-time snapshot.admitted (P2-3 funnel selected)."
            ),
        },
        "harvest_selected_n": harvest_sel_n,
        "p2_3_canonical_funnel": {"selected": P3_0_SELECTED_N, "admitted": P2_3_ADMITTED, "fills": P1_FILLS},
        "CANONICAL_SELECTED_MATCH": f"{n_selected_present}/{P1_FILLS}",
        "CANONICAL_FILL_REPRODUCED": f"{n_match}/{P1_FILLS}",
        "MISMATCH_COUNTS": counts,
        "DIAGNOSTIC_RECONCILE": recon_status,
        "root_cause": {
            "kind": "DIAGNOSTIC_RECONSTRUCTION_DEFECT",
            "not_runtime_defect": True,
            "detail": ROOT_CAUSE_COMPACT,
            "corrected_fill_n_selected": new_fill,
        },
        "CORRECTED_SELECTED": blk_all["SELECTED"],
        "CORRECTED_NOT_SELECTED": blk_all["NOT_SELECTED"],
        "REST11_SELECTED": blk_rest["SELECTED"],
        "REST11_NOT_SELECTED": blk_rest["NOT_SELECTED"],
        "TOP3_SELECTED": blk_top3["SELECTED"],
        "TOP3_NOT_SELECTED": blk_top3["NOT_SELECTED"],
        "SELECTION_RESULT": sel_label,
        "CLOCK_FULL14": CLOCK_FULL14,
        "CLOCK_REST11": CLOCK_REST11,
        "FIXED_ANCHOR_MECHANISM": mech,
        "WITHIN_ANCHOR": within,
        "RANK_STRATA": {s["quintile"]: s for s in strata},
        "EXIT_PARITY_MATCHED_FILLS": {
            "matched_fills": n_match,
            "exit_pnl_compared": len(exit_compared),
            "exit_pnl_match": exit_pnl_ok,
            "exit_time_match": exit_t_ok,
            "note": (
                "Independent diagnostic ignores OPEN/PENDING/CAP. "
                "If fill time/price match, Arch E path should match unless Dual Lane "
                "session-close / executable last-tick differs from full-path policy. "
                "Mismatches are documented, not forced."
            ),
        },
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "BEST_OFFSET_ADOPTED": False,
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "failed": failed,
        "blocked_reason": blocked_reason,
        "missing_selected": [
            {"date": r.get("date"), "anchor_time": r.get("anchor_time"), "symbol": r.get("symbol")}
            for r in missing
        ],
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_native_entry": _file_sha("src/small_paper/v1r_native_entry_live.py"),
        },
        "_sheets": {
            "recon_rows": recon_full,
            "diag_sel_vs_not": [
                {"group": "SELECTED", **blk_all["SELECTED"]},
                {"group": "ELIGIBLE_NOT_SELECTED", **blk_all["NOT_SELECTED"]},
            ],
            "strata": strata,
            "within_all": [within["all_eligible_anchors"]],
            "within_fill_both": [within["fill_both_groups"]],
            "top3": [
                {"group": "SELECTED", **blk_top3["SELECTED"]},
                {"group": "ELIGIBLE_NOT_SELECTED", **blk_top3["NOT_SELECTED"]},
            ],
            "rest11": [
                {"group": "SELECTED", **blk_rest["SELECTED"]},
                {"group": "ELIGIBLE_NOT_SELECTED", **blk_rest["NOT_SELECTED"]},
            ],
        },
    }
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    cs = rep["CORRECTED_SELECTED"]
    ns = rep["CORRECTED_NOT_SELECTED"]
    rs = rep["REST11_SELECTED"]
    rn = rep["REST11_NOT_SELECTED"]
    wa = rep["WITHIN_ANCHOR"]
    lines = [
        "# P3-0R Fixed selection diagnostic reconciliation",
        "",
        f"**Label:** `{rep['label']}` — not OOS / prospective / robust.",
        "",
        "Clock shift results are frozen. No new strategy. No Runtime change.",
        "",
        "Corrected SELECTED = harvest-time `ANCHOR_SYMBOL_SNAPSHOT.admitted` "
        "(same count source as P2-3 selected). P3-0 xlsx `selected` is not used "
        "as the corrected set.",
        "",
        f"CANONICAL_SELECTED_MATCH: {rep['CANONICAL_SELECTED_MATCH']}",
        f"CANONICAL_FILL_REPRODUCED: {rep['CANONICAL_FILL_REPRODUCED']}",
        f"MISMATCH_COUNTS: {json.dumps(rep['MISMATCH_COUNTS'])}",
        f"DIAGNOSTIC_RECONCILE: `{rep['DIAGNOSTIC_RECONCILE']}`",
        "",
        "## Root cause",
        "",
        str(rep["root_cause"]["detail"]),
        "",
        "Not classified as a Runtime trading defect.",
        "",
        "## CORRECTED_SELECTED",
        "",
        f"n: {cs['n']}",
        f"fill_n: {cs['fill_n']}",
        f"fill_rate: {cs['fill_rate']}",
        f"win/loss/draw: {cs.get('win')}/{cs.get('loss')}/{cs.get('draw')}",
        f"mean_pnl_per_eligible: {cs['mean_pnl_per_eligible']}",
        f"mean_pnl_per_filled: {cs['mean_pnl_per_filled']}",
        f"median_pnl_per_filled: {cs.get('median_pnl_per_filled')}",
        f"GP: {cs.get('gross_profit')}  GL: {cs.get('gross_loss')}",
        f"PF: {cs['PF']}",
        "",
        "## CORRECTED_NOT_SELECTED",
        "",
        f"n: {ns['n']}",
        f"fill_n: {ns['fill_n']}",
        f"fill_rate: {ns['fill_rate']}",
        f"win/loss/draw: {ns.get('win')}/{ns.get('loss')}/{ns.get('draw')}",
        f"mean_pnl_per_eligible: {ns['mean_pnl_per_eligible']}",
        f"mean_pnl_per_filled: {ns['mean_pnl_per_filled']}",
        f"median_pnl_per_filled: {ns.get('median_pnl_per_filled')}",
        f"GP: {ns.get('gross_profit')}  GL: {ns.get('gross_loss')}",
        f"PF: {ns['PF']}",
        "",
        "## REST11_SELECTED",
        "",
        f"n: {rs['n']}",
        f"fill_n: {rs['fill_n']}",
        f"fill_rate: {rs['fill_rate']}",
        f"mean_pnl_per_eligible: {rs['mean_pnl_per_eligible']}",
        f"mean_pnl_per_filled: {rs['mean_pnl_per_filled']}",
        f"PF: {rs['PF']}",
        "",
        "## REST11_NOT_SELECTED",
        "",
        f"n: {rn['n']}",
        f"fill_n: {rn['fill_n']}",
        f"fill_rate: {rn['fill_rate']}",
        f"mean_pnl_per_eligible: {rn['mean_pnl_per_eligible']}",
        f"mean_pnl_per_filled: {rn['mean_pnl_per_filled']}",
        f"PF: {rn['PF']}",
        "",
        f"SELECTION_RESULT: `{rep['SELECTION_RESULT']}`",
        f"CLOCK_FULL14: `{rep['CLOCK_FULL14']}`",
        f"CLOCK_REST11: `{rep['CLOCK_REST11']}`",
        f"FIXED_ANCHOR_MECHANISM: `{rep['FIXED_ANCHOR_MECHANISM']}`",
        "",
        "## WITHIN_ANCHOR",
        "",
        f"all: better={wa['all_eligible_anchors']['selected_better']} "
        f"worse={wa['all_eligible_anchors']['selected_worse']} equal={wa['all_eligible_anchors']['equal']} "
        f"median={wa['all_eligible_anchors']['median_difference']}",
        f"fill_both_groups: n={wa['fill_both_groups']['anchors_with_any_fill_both_groups']} "
        f"better={wa['fill_both_groups']['selected_better']} worse={wa['fill_both_groups']['selected_worse']} "
        f"equal={wa['fill_both_groups']['equal']} median={wa['fill_both_groups']['median_difference']}",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep['verdict']}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("CANONICAL_SELECTED_MATCH", public.get("CANONICAL_SELECTED_MATCH")),
            ("CANONICAL_FILL_REPRODUCED", public.get("CANONICAL_FILL_REPRODUCED")),
            ("MISMATCH_COUNTS", public.get("MISMATCH_COUNTS")),
            ("P3_0_XLSX_SELECTED_AMONG_CANONICAL", (public.get("p3_0_old") or {}).get("xlsx_selected_among_canonical_fills")),
            ("HARVEST_SELECTED_N", public.get("harvest_selected_n")),
            ("DIAGNOSTIC_RECONCILE", public.get("DIAGNOSTIC_RECONCILE")),
            ("CORRECTED_SELECTED_fill_n", (public.get("CORRECTED_SELECTED") or {}).get("fill_n")),
            ("CORRECTED_NOT_SELECTED_fill_n", (public.get("CORRECTED_NOT_SELECTED") or {}).get("fill_n")),
            ("SELECTION_RESULT", public.get("SELECTION_RESULT")),
            ("CLOCK_FULL14", public.get("CLOCK_FULL14")),
            ("CLOCK_REST11", public.get("CLOCK_REST11")),
            ("FIXED_ANCHOR_MECHANISM", public.get("FIXED_ANCHOR_MECHANISM")),
            ("NEW_STRATEGY_TESTED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Canonical_Fill_Reconcile"), sheets.get("recon_rows") or [])
    rca = []
    for r in sheets.get("recon_rows") or []:
        if r.get("klass") != "MATCH":
            rca.append(r)
    if not rca:
        rca = [{"klass": "MATCH", "n": public.get("CANONICAL_FILL_REPRODUCED"), "note": "no mismatches"}]
    _write_rows(wb.create_sheet("Mismatch_RCA"), rca)
    _write_rows(wb.create_sheet("Selected_vs_Not"), sheets.get("diag_sel_vs_not") or [])
    _write_rows(wb.create_sheet("Rank_Strata"), sheets.get("strata") or [])
    _write_rows(
        wb.create_sheet("Within_Anchor"),
        [
            {"subset": "all_eligible_anchors", **(sheets.get("within_all") or [{}])[0]},
            {"subset": "anchors_with_any_fill_both_groups", **(sheets.get("within_fill_both") or [{}])[0]},
        ],
    )
    _write_rows(wb.create_sheet("Top3"), sheets.get("top3") or [])
    _write_rows(wb.create_sheet("Rest11"), sheets.get("rest11") or [])
    _kv_sheet(wb.create_sheet("Identity"), list((public.get("Identity") or {}).items()) + [("label", TASK_LABEL)])
    _kv_sheet(
        wb.create_sheet("Safety"),
        [("submit", 0), ("cancel", 0), ("live", 0), ("NEW_STRATEGY_TESTED", False), ("RUNTIME_CHANGED", False)],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
