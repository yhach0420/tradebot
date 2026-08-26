"""Write P4-1 precommit (already on disk) + report.json / report.md / audit.xlsx."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.mid_hold_gate_p4_1 import (
    CANDIDATE_ID,
    CHECKPOINTS_SEC,
    DOCUMENT_ID,
    P1_PF,
    P1_PNL,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_OK,
)
from research.mid_hold_gate_p4_1.metrics import (
    classify_status,
    local_diagnostic,
    portfolio_effects,
    slice_days,
    summary_block,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "mid_hold_gate_p4_1"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
P4_0 = NATIVE / "results" / "research" / "mid_hold_state_separability_p4_0" / "report.json"
P3_4R = NATIVE / "results" / "research" / "extension_decision_alignment_p3_4r" / "report.json"


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(12, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def load_frozen() -> dict[str, Any]:
    out: dict[str, Any] = {"ok": True, "reasons": []}
    if not P4_0.is_file():
        return {"ok": False, "reasons": ["NO_P4_0_REPORT"]}
    p40 = json.loads(P4_0.read_text(encoding="utf-8"))
    if p40.get("MID_HOLD_GATEABILITY") != "MID_HOLD_STATE_SEPARABLE":
        out["ok"] = False
        out["reasons"].append("P4_0_NOT_SEPARABLE")
    fam = p40.get("CANDIDATE_STATE_FAMILIES") or []
    if "PRICE_PATH_DETERIORATION" not in fam:
        out["ok"] = False
        out["reasons"].append("P4_0_PRIMARY_FAMILY_MISSING")
    out["P4_0"] = {
        "MID_HOLD_GATEABILITY": p40.get("MID_HOLD_GATEABILITY"),
        "CANDIDATE_STATE_FAMILIES": fam,
        "verdict": p40.get("verdict"),
    }
    if not P3_4R.is_file():
        out["ok"] = False
        out["reasons"].append("NO_P3_4R_REPORT")
        return out
    p34 = json.loads(P3_4R.read_text(encoding="utf-8"))
    if p34.get("DECISION_RECONCILE") != "PASS":
        out["ok"] = False
        out["reasons"].append("P3_4R_DECISION_RECONCILE_NOT_PASS")
    if p34.get("EXTENSION_GATE_REVISED") != "EXTENSION_GATE_SUPPORTED":
        out["ok"] = False
        out["reasons"].append("P3_4R_GATE_NOT_SUPPORTED")
    out["P3_4R"] = {
        "DECISION_RECONCILE": p34.get("DECISION_RECONCILE"),
        "EXTENSION_GATE_REVISED": p34.get("EXTENSION_GATE_REVISED"),
    }
    return out


def build_report(
    *,
    precommit: dict[str, Any],
    recon: dict[str, Any],
    frozen: dict[str, Any],
    canonical: list[dict[str, Any]],
    top10_ids: list[str],
    top20_ids: list[str],
    base_days: list[dict[str, Any]],
    gate_days: list[dict[str, Any]],
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    base_trades = [t for d in base_days for t in (d.get("trades") or [])]
    gate_trades = [t for d in gate_days for t in (d.get("trades") or [])]
    base_admits = [t for d in base_days for t in (d.get("admits") or [])]
    gate_admits = [t for d in gate_days for t in (d.get("admits") or [])]
    base_fills = [t for d in base_days for t in (d.get("fills") or [])]
    gate_fills = [t for d in gate_days for t in (d.get("fills") or [])]
    mh_local = [r for d in base_days for r in (d.get("mh_records") or [])]
    leak_n = sum(int(d.get("leak_n") or 0) for d in base_days) + sum(int(d.get("leak_n") or 0) for d in gate_days)

    local = (
        local_diagnostic(
            records=mh_local,
            canonical=canonical,
            top10_ids=set(top10_ids),
            top20_ids=set(top20_ids),
        )
        if base_days
        else {}
    )
    base = summary_block(base_trades) if base_trades else {}
    gate = summary_block(gate_trades) if gate_trades else {}
    base_top = summary_block(slice_days(base_trades, PREDECLARED_TOP3)) if base_trades else {}
    gate_top = summary_block(slice_days(gate_trades, PREDECLARED_TOP3)) if gate_trades else {}
    base_rest = summary_block(slice_days(base_trades, REST11)) if base_trades else {}
    gate_rest = summary_block(slice_days(gate_trades, REST11)) if gate_trades else {}
    port = (
        portfolio_effects(
            base_trades=base_trades,
            gate_trades=gate_trades,
            base_admits=base_admits,
            gate_admits=gate_admits,
            base_fills=base_fills,
            gate_fills=gate_fills,
            base_cap=sum(int(d.get("cap_blocked") or 0) for d in base_days),
            gate_cap=sum(int(d.get("cap_blocked") or 0) for d in gate_days),
            base_same=sum(int(d.get("same_symbol_blocked") or 0) for d in base_days),
            gate_same=sum(int(d.get("same_symbol_blocked") or 0) for d in gate_days),
        )
        if base_days and gate_days
        else {}
    )

    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("BASELINE_RECONCILE_FAIL")
    if not frozen.get("ok"):
        integrity.extend(frozen.get("reasons") or ["FROZEN_MISSING"])
    if leak_n:
        integrity.append(f"FUTURE_LEAK_{leak_n}")

    status = (
        classify_status(
            integrity_flags=integrity if not blocked else integrity,
            local=local,
            base=base or {"pnl": 0, "PF": None, "maxDD": 0},
            gate=gate or {"pnl": 0, "PF": None, "maxDD": 0},
            base_rest=base_rest or {"pnl": 0, "PF": None, "maxDD": 0},
            gate_rest=gate_rest or {"pnl": 0, "PF": None, "maxDD": 0},
        )
        if not blocked and recon.get("pass") and gate_days
        else {"STATUS": None, "why": blocked_reason}
    )
    if blocked or not recon.get("pass"):
        verdict = VERDICT_BLOCKED
    else:
        verdict = VERDICT_OK

    cp_trig = [
        {
            "checkpoint": h,
            "local_first_trigger_n": ((local.get("by_checkpoint") or {}).get(h) or (local.get("by_checkpoint") or {}).get(int(h))),
            "loss48_first_trigger_n": ((local.get("loss_first_trigger_dist") or {}).get(str(h))),
        }
        for h in CHECKPOINTS_SEC
    ]
    # by_checkpoint keys are ints
    if local.get("by_checkpoint"):
        cp_trig = [
            {
                "checkpoint": h,
                "local_first_trigger_n": int((local.get("by_checkpoint") or {}).get(int(h)) or 0),
                "loss48_first_trigger_n": int((local.get("loss_first_trigger_dist") or {}).get(str(h)) or 0),
            }
            for h in CHECKPOINTS_SEC
        ]

    sheets = {
        "precommit": [{"key": k, "value": precommit.get(k)} for k in (
            "candidate_id", "SHA", "LABEL",
        )] + [{"key": "rule.gate_true", "value": ((precommit.get("rule") or {}).get("gate_true"))}],
        "local": local.get("rows") or [],
        "loss": [
            {
                "loss48_n": local.get("loss48_n"),
                "triggered": local.get("loss_triggered_n"),
                "not_triggered": local.get("loss_not_triggered_n"),
            }
        ],
        "pres": [
            {"group": g, **((local.get(g) or {}))}
            for g in ("TOP10", "TOP20", "EXTEND35")
        ],
        "baseline": [{"slice": "FULL14", **base}, {"slice": "TOP3", **base_top}, {"slice": "REST11", **base_rest}],
        "gate": [{"slice": "FULL14", **gate}, {"slice": "TOP3", **gate_top}, {"slice": "REST11", **gate_rest}],
        "port": [port] if port else [],
        "top3rest": [
            {"slice": "TOP3", "side": "baseline", **base_top},
            {"slice": "TOP3", "side": "gate_on", **gate_top},
            {"slice": "REST11", "side": "baseline", **base_rest},
            {"slice": "REST11", "side": "gate_on", **gate_rest},
        ],
        "cp": cp_trig,
    }

    report = {
        "task": "P4-1",
        "ANALYSIS_ID": "P4_1_MID_HOLD_NO_PROGRESS_GATE",
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "strategy validation", "production approval"],
        "PRECOMMIT_SHA": precommit.get("SHA"),
        "CANDIDATE": CANDIDATE_ID,
        "HISTORICAL_CUTOFF_SEARCHED": False,
        "STRUCTURAL_ZERO_BOUNDARY_PRECOMMITTED": True,
        "EVALUATION_CHECKPOINTS": list(CHECKPOINTS_SEC),
        "BASELINE_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "FROZEN": frozen,
        "LOCAL_TRIGGER": {
            "n": local.get("n"),
            "by_checkpoint": local.get("by_checkpoint"),
            "canonical_WIN": local.get("canonical_WIN"),
            "canonical_LOSS": local.get("canonical_LOSS"),
            "canonical_DRAW": local.get("canonical_DRAW"),
        },
        "LOSS_COVERAGE": {
            "triggered": local.get("loss_triggered_n"),
            "not_triggered": local.get("loss_not_triggered_n"),
            "loss48_n": local.get("loss48_n"),
            "first_trigger_distribution": local.get("loss_first_trigger_dist"),
        },
        "WINNER_PRESERVATION": {
            "TOP10_cut": (local.get("TOP10") or {}).get("cut_n"),
            "TOP20_cut": (local.get("TOP20") or {}).get("cut_n"),
            "EXTEND35_cut": (local.get("EXTEND35") or {}).get("cut_n"),
            "TOP10": local.get("TOP10"),
            "TOP20": local.get("TOP20"),
            "EXTEND35": local.get("EXTEND35"),
        },
        "BASELINE": base,
        "GATE_ON": gate,
        "TOP3": {"baseline": base_top, "gate": gate_top},
        "REST11": {
            "baseline_trades": base_rest.get("trades"),
            "gate_trades": gate_rest.get("trades"),
            "baseline_pnl": base_rest.get("pnl"),
            "gate_pnl": gate_rest.get("pnl"),
            "baseline_PF": base_rest.get("PF"),
            "gate_PF": gate_rest.get("PF"),
            "baseline_maxDD": base_rest.get("maxDD"),
            "gate_maxDD": gate_rest.get("maxDD"),
        },
        "PORTFOLIO_EFFECT": port,
        "STATUS": status.get("STATUS"),
        "STATUS_DETAIL": status,
        "NEW_EXIT_CANDIDATE_TESTED": bool(gate_days) and not blocked,
        "NEW_EXIT_ADOPTED": False,
        "RULE_CHANGED_AFTER_RESULT": False,
        "ENTRY_CHANGED": False,
        "EXISTING_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": bool(leak_n),
        "leak_n": int(leak_n),
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "PRECOMMIT_SHA": precommit.get("SHA"),
            "file_sha_p4_1_gate": _file_sha("src/research/mid_hold_gate_p4_1/gate.py"),
        },
        "PRIMARY_FULL14": {"trades": P1_TRADES, "pnl": P1_PNL, "PF": P1_PF},
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
    }
    unused = TASK_LABEL
    del unused
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    loc = rep.get("LOCAL_TRIGGER") or {}
    by = loc.get("by_checkpoint") or {}
    lines = [
        "# P4-1 Mid-Hold No-Progress Gate",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / strategy validation / production approval",
        "",
        f"PRECOMMIT_SHA: `{rep.get('PRECOMMIT_SHA')}`",
        f"BASELINE_RECONCILE: `{rep.get('BASELINE_RECONCILE')}`",
        f"CANDIDATE: `{rep.get('CANDIDATE')}`",
        "HISTORICAL_CUTOFF_SEARCHED: false",
        "STRUCTURAL_ZERO_BOUNDARY_PRECOMMITTED: true",
        f"EVALUATION_CHECKPOINTS: {json.dumps(rep.get('EVALUATION_CHECKPOINTS'))}",
        "",
        "LOCAL_TRIGGER:",
        f"n: {loc.get('n')}",
        "by_checkpoint:",
    ]
    for h in CHECKPOINTS_SEC:
        v = by.get(str(h), by.get(h))
        lines.append(f"{h}: {v}")
    lines += [
        f"canonical_WIN: {loc.get('canonical_WIN')}",
        f"canonical_LOSS: {loc.get('canonical_LOSS')}",
        f"canonical_DRAW: {loc.get('canonical_DRAW')}",
        "",
        "LOSS_COVERAGE:",
        f"triggered: {(rep.get('LOSS_COVERAGE') or {}).get('triggered')}",
        f"not_triggered: {(rep.get('LOSS_COVERAGE') or {}).get('not_triggered')}",
        "",
        "WINNER_PRESERVATION:",
        f"TOP10_cut: {(rep.get('WINNER_PRESERVATION') or {}).get('TOP10_cut')}",
        f"TOP20_cut: {(rep.get('WINNER_PRESERVATION') or {}).get('TOP20_cut')}",
        f"EXTEND35_cut: {(rep.get('WINNER_PRESERVATION') or {}).get('EXTEND35_cut')}",
        "",
        "BASELINE:",
        f"trades: {(rep.get('BASELINE') or {}).get('trades')}",
        f"pnl: {(rep.get('BASELINE') or {}).get('pnl')}",
        f"PF: {(rep.get('BASELINE') or {}).get('PF')}",
        f"maxDD: {(rep.get('BASELINE') or {}).get('maxDD')}",
        "",
        "GATE_ON:",
        f"trades: {(rep.get('GATE_ON') or {}).get('trades')}",
        f"pnl: {(rep.get('GATE_ON') or {}).get('pnl')}",
        f"PF: {(rep.get('GATE_ON') or {}).get('PF')}",
        f"maxDD: {(rep.get('GATE_ON') or {}).get('maxDD')}",
        "",
        "REST11:",
        f"baseline_pnl: {(rep.get('REST11') or {}).get('baseline_pnl')}",
        f"gate_pnl: {(rep.get('REST11') or {}).get('gate_pnl')}",
        f"baseline_PF: {(rep.get('REST11') or {}).get('baseline_PF')}",
        f"gate_PF: {(rep.get('REST11') or {}).get('gate_PF')}",
        f"baseline_maxDD: {(rep.get('REST11') or {}).get('baseline_maxDD')}",
        f"gate_maxDD: {(rep.get('REST11') or {}).get('gate_maxDD')}",
        "",
        "PORTFOLIO_EFFECT:",
        f"early_exited: {(rep.get('PORTFOLIO_EFFECT') or {}).get('early_exited')}",
        f"newly_admitted: {(rep.get('PORTFOLIO_EFFECT') or {}).get('newly_admitted')}",
        f"newly_filled: {(rep.get('PORTFOLIO_EFFECT') or {}).get('newly_filled')}",
        f"lost_admission: {(rep.get('PORTFOLIO_EFFECT') or {}).get('lost_admission')}",
        f"lost_fill: {(rep.get('PORTFOLIO_EFFECT') or {}).get('lost_fill')}",
        "",
        f"STATUS: `{rep.get('STATUS')}`",
        f"{(rep.get('STATUS_DETAIL') or {}).get('why')}",
        "",
        "NEW_EXIT_CANDIDATE_TESTED: true" if rep.get("NEW_EXIT_CANDIDATE_TESTED") else "NEW_EXIT_CANDIDATE_TESTED: false",
        "NEW_EXIT_ADOPTED: false",
        "RULE_CHANGED_AFTER_RESULT: false",
        "ENTRY_CHANGED: false",
        "EXISTING_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("PRECOMMIT_SHA", public.get("PRECOMMIT_SHA")),
            ("BASELINE_RECONCILE", public.get("BASELINE_RECONCILE")),
            ("CANDIDATE", public.get("CANDIDATE")),
            ("STATUS", public.get("STATUS")),
            ("NEW_EXIT_ADOPTED", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Precommit"), sheets.get("precommit") or [])
    _write_rows(wb.create_sheet("Local_Triggers"), sheets.get("local") or [])
    _write_rows(wb.create_sheet("Loss_Coverage"), sheets.get("loss") or [])
    _write_rows(wb.create_sheet("Winner_Preservation"), sheets.get("pres") or [])
    _write_rows(wb.create_sheet("Baseline"), sheets.get("baseline") or [])
    _write_rows(wb.create_sheet("Gate_On"), sheets.get("gate") or [])
    _write_rows(wb.create_sheet("Portfolio_Effects"), sheets.get("port") or [])
    _write_rows(wb.create_sheet("Top3_Rest11"), sheets.get("top3rest") or [])
    _write_rows(wb.create_sheet("Checkpoint_Triggers"), sheets.get("cp") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(ident, list((public.get("Identity") or {}).items()))
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("NEW_EXIT_ADOPTED", False),
            ("RULE_CHANGED_AFTER_RESULT", False),
            ("ENTRY_CHANGED", False),
            ("EXISTING_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("HISTORICAL_CUTOFF_SEARCHED", False),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
