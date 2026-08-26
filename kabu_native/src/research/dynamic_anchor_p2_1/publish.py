"""Aggregate P2-1 results and write report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import (
    ANCHOR_SHA,
    CONFIRMATION_NOT_EVALUABLE,
    CONFIRMED,
    ENTRY_SHA,
    EXIT_SHA,
    REJECTED,
    SESSION_INCOMPLETE,
    STRATEGY_SHA,
)
from research.dynamic_anchor_p2_1 import (
    CAPTURE_BOUNDARY_INCOMPLETE,
    CHECKPOINT_MISSING,
    CHECKPOINT_STALE,
    CONFIRMATION,
    DOCUMENT_ID,
    INVALID_PRICE,
    OTHER,
    P0_MISSING,
    PERIOD_END,
    PERIOD_START,
    TRIGGER,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "dynamic_anchor_event_validation_p2_1"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

P1_DISPLAY = {
    "PRIMARY_FULL": "14 days / 267 trades / +2,289,100 / PF 3.571 / MaxDD -210300",
    "REFERENCE_ALL_USABLE": "23 days / 331 trades / +2,288,850 / PF 3.098",
    "note": "Frozen P1 CURRENT_RUNTIME_REPLAY display only. Not compared to Dynamic. Dynamic rows are anchors, not trades.",
}


def _round(x: Any, n: int = 10) -> Any:
    if x is None:
        return None
    try:
        return round(float(x), n)
    except (TypeError, ValueError):
        return x


def ledger_sha(rows: list[dict[str, Any]], keys: tuple[str, ...]) -> str:
    h = hashlib.sha256()
    ordered = sorted(
        rows,
        key=lambda r: (str(r.get("date")), str(r.get("session")), str(r.get("symbol")), float(r.get("t0") or 0.0)),
    )
    for r in ordered:
        payload = {k: _round(r.get(k)) if k in ("t0", "t1", "vol_percentile_60s", "trend_slope", "endpoint_return", "p0", "p10") else r.get(k) for k in keys}
        h.update(json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True).encode("utf-8"))
        h.update(b"\n")
    return h.hexdigest()


TRIGGER_SHA_KEYS = (
    "date", "session", "symbol", "t0", "grid_index",
    "vol_percentile_60s", "peer_n", "previous_raw", "current_raw",
)
CONFIRM_SHA_KEYS = (
    "date", "session", "symbol", "t0", "t1", "status", "reason",
    "trend_slope", "endpoint_return", "fail_checkpoint",
)


def _subset(days: set[str], day_results: list[dict[str, Any]]) -> dict[str, Any]:
    trig: list[dict[str, Any]] = []
    conf: list[dict[str, Any]] = []
    grid = eval_n = raw_true = persist = dup = xs_leak = ck_leak = snap_leak = 0
    for d in day_results:
        if d["date"] not in days:
            continue
        trig.extend(d.get("triggers") or [])
        conf.extend(d.get("confirms") or [])
        grid += int(d.get("grid_evaluations") or 0)
        eval_n += int(d.get("t1_evaluable_rows") or 0)
        raw_true += int(d.get("raw_true_grid_rows") or 0)
        persist += int(d.get("TRUE_PERSISTENCE_REFIRE") or 0)
        dup += int(d.get("duplicate_edge_fires") or 0)
        xs_leak += int(d.get("cross_section_future_leak_count") or 0)
        ck_leak += int(d.get("checkpoint_future_leak_count") or 0)
        snap_leak += int(d.get("decision_snapshot_future_leak_count") or 0)
    st = Counter(c.get("status") for c in conf)
    confirmed = st.get(CONFIRMED, 0)
    complete = confirmed + st.get(REJECTED, 0)
    am_t = [t for t in trig if t.get("session") == "AM"]
    pm_t = [t for t in trig if t.get("session") == "PM"]
    am_c = [c for c in conf if c.get("session") == "AM" and c.get("status") == CONFIRMED]
    pm_c = [c for c in conf if c.get("session") == "PM" and c.get("status") == CONFIRMED]
    causes = Counter()
    fail_ck = Counter()
    for c in conf:
        if c.get("status") == CONFIRMATION_NOT_EVALUABLE:
            causes[str(c.get("reason") or OTHER)] += 1
            if c.get("fail_checkpoint") is not None:
                fail_ck[int(c["fail_checkpoint"])] += 1
    conf_slopes = [float(c["trend_slope"]) for c in conf if c.get("status") == CONFIRMED and c.get("trend_slope") is not None]
    conf_ep = [float(c["endpoint_return"]) for c in conf if c.get("status") == CONFIRMED and c.get("endpoint_return") is not None]
    rej_slopes = [float(c["trend_slope"]) for c in conf if c.get("status") == REJECTED and c.get("trend_slope") is not None]
    rej_ep = [float(c["endpoint_return"]) for c in conf if c.get("status") == REJECTED and c.get("endpoint_return") is not None]
    by_sd: dict[tuple[str, str], list[float]] = defaultdict(list)
    for t in trig:
        by_sd[(str(t["date"]), str(t["symbol"]))].append(float(t["t0"]))
    counts = [len(v) for v in by_sd.values()]
    gaps: list[float] = []
    for v in by_sd.values():
        vv = sorted(v)
        gaps.extend(b - a for a, b in zip(vv, vv[1:]))
    later = 0
    by_dss: dict[tuple[str, str, str], int] = Counter()
    for t in trig:
        by_dss[(str(t["date"]), str(t["session"]), str(t["symbol"]))] += 1
    for n in by_dss.values():
        if n >= 2:
            later += n - 1
    dens_vals: list[int] = []
    for d in day_results:
        if d["date"] not in days:
            continue
        cc = d.get("anchors_per_symbol") or {}
        n_sym = int(d.get("universe_n") or 0)
        dens_vals.extend(int(v) for v in cc.values())
        dens_vals.extend([0] * max(0, n_sym - len(cc)))
    counts = dens_vals
    n_trig = len(trig)
    return {
        "grid_evaluations": grid,
        "t1_evaluable_rows": eval_n,
        "raw_true_grid_rows": raw_true,
        "false_to_true_triggers": n_trig,
        "unique_trigger_symbols": len({t["symbol"] for t in trig}),
        "confirmation_complete": complete,
        "confirmed": confirmed,
        "rejected": st.get(REJECTED, 0),
        "not_evaluable": st.get(CONFIRMATION_NOT_EVALUABLE, 0),
        "session_incomplete": st.get(SESSION_INCOMPLETE, 0),
        "capture_boundary_incomplete": st.get(CAPTURE_BOUNDARY_INCOMPLETE, 0),
        "confirmed_rate": (confirmed / n_trig) if n_trig else None,
        "AM": {"triggers": len(am_t), "confirmed": len(am_c)},
        "PM": {"triggers": len(pm_t), "confirmed": len(pm_c)},
        "TRUE_PERSISTENCE_REFIRE": persist,
        "duplicate_edge_fires": dup,
        "second_or_later_anchors": later,
        "cross_section_future_leak_count": xs_leak,
        "checkpoint_future_leak_count": ck_leak,
        "decision_snapshot_future_leak_count": snap_leak,
        "not_evaluable_causes": {
            P0_MISSING: causes.get(P0_MISSING, 0),
            CHECKPOINT_MISSING: causes.get(CHECKPOINT_MISSING, 0),
            CHECKPOINT_STALE: causes.get(CHECKPOINT_STALE, 0),
            INVALID_PRICE: causes.get(INVALID_PRICE, 0),
            OTHER: sum(v for k, v in causes.items() if k not in {P0_MISSING, CHECKPOINT_MISSING, CHECKPOINT_STALE, INVALID_PRICE}),
        },
        "fail_checkpoint": dict(sorted(fail_ck.items())),
        "c1_shape": {
            "confirmed_median_trend_slope": statistics.median(conf_slopes) if conf_slopes else None,
            "confirmed_median_endpoint_return_10m": statistics.median(conf_ep) if conf_ep else None,
            "rejected_median_trend_slope": statistics.median(rej_slopes) if rej_slopes else None,
            "rejected_median_endpoint_return_10m": statistics.median(rej_ep) if rej_ep else None,
        },
        "density": {
            "symbol_days_with_anchor": len(counts),
            "median_anchors_per_symbol_day": statistics.median(counts) if counts else None,
            "p90_anchors_per_symbol_day": (sorted(counts)[int(0.9 * (len(counts) - 1))] if counts else None),
            "max_anchors_single_symbol_day": max(counts) if counts else 0,
            "min_inter_anchor_gap_sec": min(gaps) if gaps else None,
        },
        "triggers": trig,
        "confirms": conf,
    }


def daily_rows(day_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for d in day_results:
        conf = d.get("confirms") or []
        st = Counter(c.get("status") for c in conf)
        out.append({
            "date": d["date"],
            "capture_class": d.get("capture_class"),
            "universe_n": d.get("universe_n"),
            "universe_source": d.get("universe_source"),
            "events_scanned": d.get("events_scanned"),
            "grid_evaluations": d.get("grid_evaluations"),
            "t1_evaluable_rows": d.get("t1_evaluable_rows"),
            "raw_true_grid_rows": d.get("raw_true_grid_rows"),
            "false_to_true_triggers": d.get("false_to_true_triggers"),
            "unique_trigger_symbols": d.get("unique_trigger_symbols"),
            "confirmed": st.get(CONFIRMED, 0),
            "rejected": st.get(REJECTED, 0),
            "not_evaluable": st.get(CONFIRMATION_NOT_EVALUABLE, 0),
            "session_incomplete": st.get(SESSION_INCOMPLETE, 0),
            "capture_boundary_incomplete": st.get(CAPTURE_BOUNDARY_INCOMPLETE, 0),
            "persist_refire": d.get("TRUE_PERSISTENCE_REFIRE"),
            "duplicate_edge_fires": d.get("duplicate_edge_fires"),
            "xs_leak": d.get("cross_section_future_leak_count"),
            "ck_leak": d.get("checkpoint_future_leak_count"),
        })
    return out


def build_report(
    *,
    inventory: list[dict[str, Any]],
    day_results: list[dict[str, Any]],
    failed: list[str],
    det: dict[str, Any],
) -> dict[str, Any]:
    full_days = {r["date"] for r in inventory if r.get("capture_class") == "FULL" and r.get("replay_eligible")}
    ref_days = {r["date"] for r in inventory if r.get("replay_eligible")}
    primary = _subset(full_days, day_results)
    reference = _subset(ref_days, day_results)
    xs_all = []
    for d in day_results:
        xs_all.extend(d.get("xs_samples") or [])
    fut_xs = reference["cross_section_future_leak_count"]
    fut_ck = reference["checkpoint_future_leak_count"]
    fut_snap = reference["decision_snapshot_future_leak_count"]
    fut = fut_xs + fut_ck + fut_snap
    persist_ok = primary["TRUE_PERSISTENCE_REFIRE"] == 0 and reference["TRUE_PERSISTENCE_REFIRE"] == 0
    dup_ok = primary["duplicate_edge_fires"] == 0 and reference["duplicate_edge_fires"] == 0
    det_ok = bool(det.get("pass"))
    leak_ok = fut == 0
    completed = [d["date"] for d in day_results]
    issue = (not leak_ok) or (not persist_ok) or (not dup_ok) or (not det_ok) or bool(failed)
    if failed and not day_results:
        verdict = "P2_1_BLOCKED"
        can = "NO"
    elif issue:
        verdict = "P2_1_DYNAMIC_ANCHOR_EVENT_VALIDATION_ISSUE_FOUND"
        can = "NO"
    else:
        verdict = "P2_1_DYNAMIC_ANCHOR_EVENT_VALIDATION_PASS"
        can = "YES"
    now = datetime.now(JST).isoformat(timespec="seconds")
    n_inv = len(inventory)
    n_full = sum(1 for r in inventory if r.get("capture_class") == "FULL")
    n_part = sum(1 for r in inventory if r.get("capture_class") == "PARTIAL")
    n_elig = sum(1 for r in inventory if r.get("replay_eligible"))
    n_miss = sum(1 for r in inventory if r.get("capture_class") == "MISSING" or r["date"] in ("20260815", "20260816"))
    def _pub(block: dict[str, Any]) -> dict[str, Any]:
        skip = {"triggers", "confirms"}
        return {k: v for k, v in block.items() if k not in skip}
    return {
        "task": "P2-1",
        "analysis_id": "P2_1_DYNAMIC_ANCHOR_EVENT_VALIDATION",
        "document_id": DOCUMENT_ID,
        "generated_at_jst": now,
        "PERIOD": f"{PERIOD_START} - {PERIOD_END}",
        "INPUT_DAYS": n_inv,
        "full": n_full,
        "partial": n_part,
        "eligible": n_elig,
        "completed": len(completed),
        "failed": failed,
        "missing_excluded": ["20260815", "20260816"],
        "TRIGGER": TRIGGER,
        "CONFIRMATION": CONFIRMATION,
        "PRIMARY_FULL": _pub(primary),
        "REFERENCE_ALL_USABLE": _pub(reference),
        "REARM": {
            "second_or_later_anchors": primary["second_or_later_anchors"],
            "duplicate_edge_fires": primary["duplicate_edge_fires"] + reference["duplicate_edge_fires"],
            "TRUE_PERSISTENCE_REFIRE": primary["TRUE_PERSISTENCE_REFIRE"] + reference["TRUE_PERSISTENCE_REFIRE"],
        },
        "DENSITY": primary["density"],
        "CAUSALITY": {
            "cross_section_future_leak_count": fut_xs,
            "checkpoint_future_leak_count": fut_ck,
            "decision_snapshot_future_leak_count": fut_snap,
        },
        "FUTURE_LEAK": not leak_ok,
        "NOT_EVALUABLE_CAUSES": primary["not_evaluable_causes"],
        "C1_SHAPE_INTERNAL": {
            "PRIMARY_FULL": primary["c1_shape"],
            "note": "endpoint_return and trend_slope are C1 internals on [t0,t1]. No post-t1 return / MFE / MAE / PnL.",
        },
        "DETERMINISM": "PASS" if det_ok else "FAIL",
        "TRIGGER_LEDGER_SHA_RUN1": det.get("trig_sha1"),
        "TRIGGER_LEDGER_SHA_RUN2": det.get("trig_sha2"),
        "CONFIRM_LEDGER_SHA_RUN1": det.get("conf_sha1"),
        "CONFIRM_LEDGER_SHA_RUN2": det.get("conf_sha2"),
        "determinism_days": det.get("days"),
        "HISTORICAL_PNL_READ": False,
        "TRADE_SIMULATION_RUN": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "CAN_PROCEED_TO_P2_2_PNL_TEST": can,
        "verdict": verdict,
        "holdout_label": "NO_CLEAN_HOLDOUT — reused 20260721-20260821; not OOS / prospective",
        "p1_display_only": P1_DISPLAY,
        "identity": {
            "strategy_sha": STRATEGY_SHA,
            "entry_sha": ENTRY_SHA,
            "exit_sha": EXIT_SHA,
            "anchor_sha": ANCHOR_SHA,
        },
        "inventory": [{k: v for k, v in r.items() if k != "universe_symbols"} for r in inventory],
        "daily": daily_rows(day_results),
        "_primary_triggers": primary["triggers"],
        "_primary_confirms": primary["confirms"],
        "_reference_triggers": reference["triggers"],
        "_reference_confirms": reference["confirms"],
        "_xs_samples": xs_all,
        "missing_days": n_miss,
    }


def _write_rows(ws, rows: list[dict[str, Any]], cap: int | None = None) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    data = rows[:cap] if cap else rows
    cols = list(data[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(data, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(40, max(12, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _md(rep: dict[str, Any]) -> str:
    p = rep["PRIMARY_FULL"]
    r = rep["REFERENCE_ALL_USABLE"]
    return f"""# P2-1 Dynamic Anchor event / causality validation

- generated_at_jst: `{rep['generated_at_jst']}`
- **verdict: `{rep['verdict']}`**
- CAN_PROCEED_TO_P2_2_PNL_TEST: **{rep['CAN_PROCEED_TO_P2_2_PNL_TEST']}**
- HISTORICAL_PNL_READ: `false` · TRADE_SIMULATION_RUN: `false`
- submit/cancel/live: `0/0/0`

No ENTRY, PENDING, FILL, EXIT, or PnL. Rows below are **anchors**, not trades. P1 numbers are display-only and are not compared.

## Final report

PERIOD: `{rep['PERIOD']}`

INPUT_DAYS: {rep['INPUT_DAYS']} · full: {rep['full']} · partial: {rep['partial']} · eligible: {rep['eligible']} · completed: {rep['completed']} · failed: {rep['failed']}

TRIGGER: `{TRIGGER}`  
CONFIRMATION: `{CONFIRMATION}`

### PRIMARY_FULL

- grid_evaluations: {p['grid_evaluations']}
- t1_evaluable_rows: {p['t1_evaluable_rows']}
- raw_true_grid_rows: {p['raw_true_grid_rows']}
- false_to_true_triggers: {p['false_to_true_triggers']}
- unique_trigger_symbols: {p['unique_trigger_symbols']}
- confirmation_complete: {p['confirmation_complete']}
- confirmed: {p['confirmed']}
- rejected: {p['rejected']}
- not_evaluable: {p['not_evaluable']}
- session_incomplete: {p['session_incomplete']}
- capture_boundary_incomplete: {p['capture_boundary_incomplete']}
- confirmed_rate: {p['confirmed_rate']}
- AM triggers/confirmed: {p['AM']['triggers']} / {p['AM']['confirmed']}
- PM triggers/confirmed: {p['PM']['triggers']} / {p['PM']['confirmed']}

### REFERENCE_ALL_USABLE

- grid_evaluations: {r['grid_evaluations']}
- t1_evaluable_rows: {r['t1_evaluable_rows']}
- raw_true_grid_rows: {r['raw_true_grid_rows']}
- false_to_true_triggers: {r['false_to_true_triggers']}
- unique_trigger_symbols: {r['unique_trigger_symbols']}
- confirmation_complete: {r['confirmation_complete']}
- confirmed: {r['confirmed']}
- rejected: {r['rejected']}
- not_evaluable: {r['not_evaluable']}
- session_incomplete: {r['session_incomplete']}
- capture_boundary_incomplete: {r['capture_boundary_incomplete']}
- confirmed_rate: {r['confirmed_rate']}
- AM triggers/confirmed: {r['AM']['triggers']} / {r['AM']['confirmed']}
- PM triggers/confirmed: {r['PM']['triggers']} / {r['PM']['confirmed']}

### REARM

- second_or_later_anchors (PRIMARY): {rep['REARM']['second_or_later_anchors']}
- duplicate_edge_fires: {rep['REARM']['duplicate_edge_fires']}
- TRUE_PERSISTENCE_REFIRE: {rep['REARM']['TRUE_PERSISTENCE_REFIRE']} (required 0)

### DENSITY (PRIMARY universe symbol-days; zeros included)

- median_anchors_per_symbol_day: {p['density']['median_anchors_per_symbol_day']}
- p90_anchors_per_symbol_day: {p['density']['p90_anchors_per_symbol_day']}
- max_anchors_single_symbol_day: {p['density']['max_anchors_single_symbol_day']}
- min_inter_anchor_gap_sec: {p['density']['min_inter_anchor_gap_sec']}

### CAUSALITY

- cross_section_future_leak_count: {rep['CAUSALITY']['cross_section_future_leak_count']}
- checkpoint_future_leak_count: {rep['CAUSALITY']['checkpoint_future_leak_count']}
- decision_snapshot_future_leak_count: {rep['CAUSALITY']['decision_snapshot_future_leak_count']}
- FUTURE_LEAK: {str(rep['FUTURE_LEAK']).lower()}

### NOT_EVALUABLE_CAUSES (PRIMARY)

{json.dumps(rep['NOT_EVALUABLE_CAUSES'], ensure_ascii=False)}

Fail-checkpoint index: {json.dumps(p.get('fail_checkpoint') or {})}

C1 internal medians (CONFIRMED / REJECTED), window [t0,t1] only:  
{json.dumps(p['c1_shape'], ensure_ascii=False)}

### DETERMINISM

{rep['DETERMINISM']} days={rep.get('determinism_days')}

- TRIGGER_LEDGER_SHA_RUN1: `{rep.get('TRIGGER_LEDGER_SHA_RUN1')}`
- TRIGGER_LEDGER_SHA_RUN2: `{rep.get('TRIGGER_LEDGER_SHA_RUN2')}`
- CONFIRM_LEDGER_SHA_RUN1: `{rep.get('CONFIRM_LEDGER_SHA_RUN1')}`
- CONFIRM_LEDGER_SHA_RUN2: `{rep.get('CONFIRM_LEDGER_SHA_RUN2')}`

Holdout: `{rep['holdout_label']}`

P1 display-only (not compared): {P1_DISPLAY['PRIMARY_FULL']}

## STOP

P2-1 ends here. No P2-2 PnL / ENTRY / FILL / EXIT simulation in this task.
"""


def write_artifacts(rep: dict[str, Any]) -> dict[str, Path]:
    OUT.mkdir(parents=True, exist_ok=True)
    for p in OUT.iterdir():
        if p.is_file() and p.name not in {"report.json", "report.md", "audit.xlsx"}:
            p.unlink()
    pub = {k: v for k, v in rep.items() if not k.startswith("_")}
    (OUT / "report.json").write_text(json.dumps(pub, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT / "report.md").write_text(_md(rep), encoding="utf-8")

    wb = Workbook()

    def sh(name: str):
        if name == "Summary":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    p = rep["PRIMARY_FULL"]
    summary = [
        {"field": "PERIOD", "value": rep["PERIOD"]},
        {"field": "INPUT_DAYS", "value": rep["INPUT_DAYS"]},
        {"field": "full", "value": rep["full"]},
        {"field": "partial", "value": rep["partial"]},
        {"field": "eligible", "value": rep["eligible"]},
        {"field": "completed", "value": rep["completed"]},
        {"field": "failed", "value": json.dumps(rep["failed"])},
        {"field": "TRIGGER", "value": TRIGGER},
        {"field": "CONFIRMATION", "value": CONFIRMATION},
        {"field": "PRIMARY_false_to_true_triggers", "value": p["false_to_true_triggers"]},
        {"field": "PRIMARY_confirmed", "value": p["confirmed"]},
        {"field": "PRIMARY_rejected", "value": p["rejected"]},
        {"field": "PRIMARY_not_evaluable", "value": p["not_evaluable"]},
        {"field": "PRIMARY_session_incomplete", "value": p["session_incomplete"]},
        {"field": "PRIMARY_capture_boundary_incomplete", "value": p["capture_boundary_incomplete"]},
        {"field": "PRIMARY_confirmed_rate", "value": p["confirmed_rate"]},
        {"field": "REF_false_to_true_triggers", "value": rep["REFERENCE_ALL_USABLE"]["false_to_true_triggers"]},
        {"field": "REF_confirmed", "value": rep["REFERENCE_ALL_USABLE"]["confirmed"]},
        {"field": "TRUE_PERSISTENCE_REFIRE", "value": rep["REARM"]["TRUE_PERSISTENCE_REFIRE"]},
        {"field": "duplicate_edge_fires", "value": rep["REARM"]["duplicate_edge_fires"]},
        {"field": "FUTURE_LEAK", "value": rep["FUTURE_LEAK"]},
        {"field": "DETERMINISM", "value": rep["DETERMINISM"]},
        {"field": "HISTORICAL_PNL_READ", "value": False},
        {"field": "TRADE_SIMULATION_RUN", "value": False},
        {"field": "STRATEGY_CHANGED", "value": False},
        {"field": "ENTRY_EXIT_CHANGED", "value": False},
        {"field": "SAFETY", "value": "submit/cancel/live=0/0/0"},
        {"field": "CAN_PROCEED_TO_P2_2_PNL_TEST", "value": rep["CAN_PROCEED_TO_P2_2_PNL_TEST"]},
        {"field": "verdict", "value": rep["verdict"]},
    ]
    _write_rows(sh("Summary"), summary)
    _write_rows(sh("Daily"), rep["daily"])
    _write_rows(sh("Trigger_Events"), [
        {k: t.get(k) for k in ("date", "session", "symbol", "t0", "grid_index", "grid_time", "vol_percentile_60s", "peer_n", "previous_raw", "current_raw")}
        for t in rep["_primary_triggers"]
    ], cap=25000)
    _write_rows(sh("Confirmation_Events"), [
        {k: c.get(k) for k in ("date", "session", "symbol", "t0", "t1", "status", "reason", "trend_slope", "endpoint_return", "fail_checkpoint", "decision_fire_time", "snapshot_cutoff")}
        for c in rep["_primary_confirms"]
    ], cap=25000)
    fail_rows = [{"fail_checkpoint": k, "n": v} for k, v in (p.get("fail_checkpoint") or {}).items()]
    if not fail_rows:
        fail_rows = [{"fail_checkpoint": None, "n": 0}]
    _write_rows(sh("Checkpoint_Audit"), fail_rows)
    _write_rows(sh("CrossSection_Causality"), rep["_xs_samples"][:5000])
    _write_rows(sh("Rearm_Audit"), [
        {"field": "TRUE_PERSISTENCE_REFIRE", "value": rep["REARM"]["TRUE_PERSISTENCE_REFIRE"]},
        {"field": "duplicate_edge_fires", "value": rep["REARM"]["duplicate_edge_fires"]},
        {"field": "second_or_later_anchors_PRIMARY", "value": rep["REARM"]["second_or_later_anchors"]},
        {"field": "AM_state_carried_to_PM", "value": "forbidden; machines reset per session"},
    ])
    cov = [{"cause": k, "n": v} for k, v in rep["NOT_EVALUABLE_CAUSES"].items()]
    cov.append({"cause": "CAPTURE_BOUNDARY_INCOMPLETE", "n": p["capture_boundary_incomplete"]})
    _write_rows(sh("Coverage"), cov)
    _write_rows(sh("Determinism"), [
        {"field": "result", "value": rep["DETERMINISM"]},
        {"field": "days", "value": json.dumps(rep.get("determinism_days"))},
        {"field": "TRIGGER_LEDGER_SHA_RUN1", "value": rep.get("TRIGGER_LEDGER_SHA_RUN1")},
        {"field": "TRIGGER_LEDGER_SHA_RUN2", "value": rep.get("TRIGGER_LEDGER_SHA_RUN2")},
        {"field": "CONFIRM_LEDGER_SHA_RUN1", "value": rep.get("CONFIRM_LEDGER_SHA_RUN1")},
        {"field": "CONFIRM_LEDGER_SHA_RUN2", "value": rep.get("CONFIRM_LEDGER_SHA_RUN2")},
    ])
    _write_rows(sh("Identity"), [
        {"field": k, "value": v} for k, v in {**rep["identity"], "TRIGGER": TRIGGER, "CONFIRMATION": CONFIRMATION, "clock": "X14_CAUSAL_10S_GRID"}.items()
    ])
    _write_rows(sh("Safety"), [
        {"field": "submit/cancel/live", "value": "0/0/0"},
        {"field": "HISTORICAL_PNL_READ", "value": False},
        {"field": "TRADE_SIMULATION_RUN", "value": False},
        {"field": "STRATEGY_CHANGED", "value": False},
        {"field": "ENTRY_EXIT_CHANGED", "value": False},
        {"field": "p1_display_only", "value": P1_DISPLAY["note"]},
    ])
    wb.save(OUT / "audit.xlsx")
    return {
        "report_json": OUT / "report.json",
        "report_md": OUT / "report.md",
        "audit_xlsx": OUT / "audit.xlsx",
    }
