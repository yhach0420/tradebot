"""Write P4-2 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.mid_hold_recovery_failure_path_p4_2 import (
    CHECKPOINTS_SEC,
    DOCUMENT_ID,
    P1_PF,
    P1_PNL,
    P1_TRADES,
    PREDECLARED_TOP3,
    PRIMARY_CHECKPOINTS,
    REST11,
    SNAPSHOT_VARS,
    TASK_LABEL,
    TRAJECTORY_VARS,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.mid_hold_recovery_failure_path_p4_2.metrics import (
    at_horizon,
    casebook_trade,
    classify_mechanism,
    compare_var,
    day_stability,
    earliest_divergence,
    interval_flags,
    lodo_auc,
    persistence_block,
    pf,
    rw,
    shape_block,
    slice_days,
)
from research.mid_hold_recovery_failure_path_p4_2.path import build_trade_paths

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "mid_hold_recovery_failure_path_p4_2"
P4_0 = NATIVE / "results" / "research" / "mid_hold_state_separability_p4_0" / "report.json"
P4_1 = NATIVE / "results" / "research" / "mid_hold_gate_p4_1" / "report.json"
P4_1_XLSX = NATIVE / "results" / "research" / "mid_hold_gate_p4_1" / "audit.xlsx"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

KEY_VARS = (
    "bid_return_from_fill",
    "executable_mfe_to_t",
    "executable_mae_to_t",
    "bid_giveback_from_peak",
    "delta_bid_120_to_t",
    "delta_bid_prev_checkpoint",
    "delta_mfe_prev_checkpoint",
    "delta_mae_prev_checkpoint",
    "rebound_from_low_t",
    "underwater_checkpoint_count",
    "consecutive_underwater_count",
    "imbalance",
)


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    return hashlib.sha256(p.read_bytes()).hexdigest() if p.is_file() else ""


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(12, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def load_p41_triggers() -> dict[str, Any]:
    if not P4_1.is_file():
        return {"ok": False, "reason": "NO_P4_1_REPORT"}
    p = json.loads(P4_1.read_text(encoding="utf-8"))
    if p.get("STATUS") != "MID_HOLD_NO_PROGRESS_PORTFOLIO_HARM":
        return {"ok": False, "reason": "P4_1_NOT_CLOSED_HARM"}
    if int((p.get("LOCAL_TRIGGER") or {}).get("n") or 0) != 17:
        return {"ok": False, "reason": "P4_1_TRIGGER_N_NOT_17"}
    ids = []
    if P4_1_XLSX.is_file():
        wb = load_workbook(P4_1_XLSX, read_only=True, data_only=True)
        ws = wb["Local_Triggers"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        hdr = rows[0]
        for row in rows[1:]:
            d = dict(zip(hdr, row))
            tid = d.get("trade_id")
            if tid:
                ids.append(str(tid))
    if len(ids) != 17:
        return {"ok": False, "reason": f"P4_1_TRIGGER_IDS_{len(ids)}"}
    return {
        "ok": True,
        "STATUS": p.get("STATUS"),
        "n": 17,
        "WIN": (p.get("LOCAL_TRIGGER") or {}).get("canonical_WIN"),
        "LOSS": (p.get("LOCAL_TRIGGER") or {}).get("canonical_LOSS"),
        "trade_ids": ids,
        "TOP20_cut_ids": ((p.get("WINNER_PRESERVATION") or {}).get("TOP20") or {}).get("trade_ids") or [],
        "EXTEND_cut_ids": ((p.get("WINNER_PRESERVATION") or {}).get("EXTEND35") or {}).get("trade_ids") or [],
        "PRECOMMIT_SHA": p.get("PRECOMMIT_SHA"),
    }


def load_frozen() -> dict[str, Any]:
    out: dict[str, Any] = {"ok": True, "reasons": []}
    if not P4_0.is_file():
        return {"ok": False, "reasons": ["NO_P4_0"]}
    p0 = json.loads(P4_0.read_text(encoding="utf-8"))
    if p0.get("MID_HOLD_GATEABILITY") != "MID_HOLD_STATE_SEPARABLE":
        out["ok"] = False
        out["reasons"].append("P4_0_NOT_SEPARABLE")
    out["P4_0"] = {"MID_HOLD_GATEABILITY": p0.get("MID_HOLD_GATEABILITY"), "verdict": p0.get("verdict")}
    p41 = load_p41_triggers()
    if not p41.get("ok"):
        out["ok"] = False
        out["reasons"].append(p41.get("reason") or "P4_1_MISSING")
    out["P4_1"] = p41
    return out


def _comp_grid(trades: list[dict[str, Any]]) -> dict[int, dict[str, dict[str, Any]]]:
    return {h: {v: compare_var(trades, h, v) for v in KEY_VARS} for h in CHECKPOINTS_SEC}


def _med(comp, h, var, side):
    return ((((comp.get(h) or {}).get(var) or {}).get(side) or {}).get("median"))


def build_report(
    *,
    path_rows: list[dict[str, Any]],
    recon: dict[str, Any],
    frozen: dict[str, Any],
    leak_n: int,
    identity_n: int,
    identity_fail: int,
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    p41 = frozen.get("P4_1") or {}
    p41_ids = set(p41.get("trade_ids") or [])
    trades = build_trade_paths(path_rows, p41_ids=p41_ids) if path_rows else []
    rest = slice_days(trades, REST11)
    top3 = slice_days(trades, PREDECLARED_TOP3)

    def _cohort_n(ts):
        a = sum(1 for t in ts if t.get("cohort_A_eligible"))
        b = sum(1 for t in ts if t.get("cohort_B_adverse120"))
        rw_n = sum(1 for t in ts if rw(t))
        pf_n = sum(1 for t in ts if pf(t))
        dr = sum(1 for t in ts if t.get("ADVERSE_DRAW"))
        c = sum(1 for t in ts if t.get("cohort_C_p41_trigger"))
        return {"A": a, "B": b, "RW": rw_n, "PF": pf_n, "DRAW": dr, "C": c}

    n_all = _cohort_n(trades)
    n_rest = _cohort_n(rest)

    all_comp = _comp_grid(trades) if trades else {}
    rest_comp = _comp_grid(rest) if rest else {}

    rec_shape = {str(h): shape_block(trades, rw, h) for h in (180, 240, 300, 360)}
    pf_shape = {str(h): shape_block(trades, pf, h) for h in (180, 240, 300, 360)}
    persist = {str(h): persistence_block(trades, h) for h in PRIMARY_CHECKPOINTS}
    new_low = {
        str(h): {"RW": interval_flags(trades, h, rw), "PF": interval_flags(trades, h, pf)}
        for h in (180, 240, 300, 360, 420, 480, 540)
    }
    diverge = earliest_divergence(all_comp) if all_comp else {}

    stab_vars = ("rebound_from_low_t", "delta_bid_120_to_t", "delta_mae_prev_checkpoint", "bid_return_from_fill")
    day_stab = {v: {str(h): day_stability(trades, h, v) for h in (180, 240, 300, 360)} for v in stab_vars}
    lodo = {v: {str(h): lodo_auc(trades, h, v) for h in (180, 240, 300, 360)} for v in stab_vars}

    p41_trades = [t for t in trades if t.get("cohort_C_p41_trigger")]
    case17 = [casebook_trade(t) for t in p41_trades]
    top20_ids = set(p41.get("TOP20_cut_ids") or [])
    ext_ids = set(p41.get("EXTEND_cut_ids") or [])
    winner_cases = {
        "TOP20_cut_2": [casebook_trade(t) for t in trades if str(t.get("trade_id")) in top20_ids],
        "EXTEND_cut_2": [casebook_trade(t) for t in trades if str(t.get("trade_id")) in ext_ids],
    }

    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if not frozen.get("ok"):
        integrity.extend(frozen.get("reasons") or ["FROZEN"])
    if path_rows and leak_n:
        integrity.append(f"FUTURE_LEAK_{leak_n}")
    if path_rows and identity_fail:
        integrity.append(f"IDENTITY_FAIL_{identity_fail}/{identity_n}")

    mech = (
        classify_mechanism(
            all_comp=all_comp,
            rest_comp=rest_comp,
            n_rw_rest=n_rest["RW"],
            n_pf_rest=n_rest["PF"],
            integrity=integrity,
        )
        if trades and not blocked
        else {"MECHANISM_CLASSIFICATION": None, "CANDIDATE_PATH_FAMILIES": []}
    )

    if blocked or not recon.get("pass") or not frozen.get("ok"):
        verdict = VERDICT_BLOCKED
    elif integrity:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    def _summ(comp, var):
        return {
            str(h): {
                "RW_median": _med(comp, h, var, "RECOVERING_WINNER"),
                "PF_median": _med(comp, h, var, "PERSISTENT_FAILURE"),
                "direction": ((comp.get(h) or {}).get(var) or {}).get("direction"),
                "auc_best": ((comp.get(h) or {}).get(var) or {}).get("auc_best"),
            }
            for h in PRIMARY_CHECKPOINTS
        }

    sheets = {
        "cohorts": [
            {"slice": sl, **n}
            for sl, n in (("ALL", n_all), ("REST11", n_rest), ("TOP3", _cohort_n(top3)))
        ],
        "path": [
            {
                "slice": sl,
                "horizon_sec": h,
                "var": v,
                "rw_n": (((comp.get(h) or {}).get(v) or {}).get("RECOVERING_WINNER") or {}).get("n"),
                "rw_mean": (((comp.get(h) or {}).get(v) or {}).get("RECOVERING_WINNER") or {}).get("mean"),
                "rw_median": (((comp.get(h) or {}).get(v) or {}).get("RECOVERING_WINNER") or {}).get("median"),
                "rw_p25": (((comp.get(h) or {}).get(v) or {}).get("RECOVERING_WINNER") or {}).get("p25"),
                "rw_p75": (((comp.get(h) or {}).get(v) or {}).get("RECOVERING_WINNER") or {}).get("p75"),
                "pf_n": (((comp.get(h) or {}).get(v) or {}).get("PERSISTENT_FAILURE") or {}).get("n"),
                "pf_mean": (((comp.get(h) or {}).get(v) or {}).get("PERSISTENT_FAILURE") or {}).get("mean"),
                "pf_median": (((comp.get(h) or {}).get(v) or {}).get("PERSISTENT_FAILURE") or {}).get("median"),
                "pf_p25": (((comp.get(h) or {}).get(v) or {}).get("PERSISTENT_FAILURE") or {}).get("p25"),
                "pf_p75": (((comp.get(h) or {}).get(v) or {}).get("PERSISTENT_FAILURE") or {}).get("p75"),
                "direction": ((comp.get(h) or {}).get(v) or {}).get("direction"),
                "auc_best": ((comp.get(h) or {}).get(v) or {}).get("auc_best"),
                "cliffs_delta": ((comp.get(h) or {}).get(v) or {}).get("cliffs_delta_rw_minus_pf"),
            }
            for sl, comp in (("ALL", all_comp), ("REST11", rest_comp))
            for h in CHECKPOINTS_SEC
            for v in KEY_VARS
        ],
        "snap_traj": [
            {
                "horizon_sec": h,
                "snapshot_bid_auc_best": ((all_comp.get(h) or {}).get("bid_return_from_fill") or {}).get("auc_best"),
                "traj_delta_bid_auc_best": ((all_comp.get(h) or {}).get("delta_bid_120_to_t") or {}).get("auc_best"),
                "traj_rebound_auc_best": ((all_comp.get(h) or {}).get("rebound_from_low_t") or {}).get("auc_best"),
                "traj_delta_mae_auc_best": ((all_comp.get(h) or {}).get("delta_mae_prev_checkpoint") or {}).get("auc_best"),
                "traj_delta_mfe_auc_best": ((all_comp.get(h) or {}).get("delta_mfe_prev_checkpoint") or {}).get("auc_best"),
                "REST11_snapshot_bid_auc_best": ((rest_comp.get(h) or {}).get("bid_return_from_fill") or {}).get("auc_best"),
                "REST11_rebound_auc_best": ((rest_comp.get(h) or {}).get("rebound_from_low_t") or {}).get("auc_best"),
            }
            for h in PRIMARY_CHECKPOINTS
        ],
        "recovery": [{"horizon_end": h, **rec_shape[str(h)]} for h in (180, 240, 300, 360)],
        "persist_fail": [{"horizon_end": h, **pf_shape[str(h)]} for h in (180, 240, 300, 360)],
        "newlow": [
            {"horizon_sec": h, "group": g, **blk}
            for h, pair in new_low.items()
            for g, blk in pair.items()
        ],
        "mfe": [
            {
                "horizon_sec": h,
                "RW_new_mfe_rate": (new_low.get(str(h)) or {}).get("RW", {}).get("new_mfe_rate"),
                "PF_new_mfe_rate": (new_low.get(str(h)) or {}).get("PF", {}).get("new_mfe_rate"),
            }
            for h in (180, 240, 300, 360)
        ],
        "rebound": [
            {
                "horizon_sec": h,
                "RW_median": _med(all_comp, h, "rebound_from_low_t", "RECOVERING_WINNER"),
                "PF_median": _med(all_comp, h, "rebound_from_low_t", "PERSISTENT_FAILURE"),
                "ALL_direction": ((all_comp.get(h) or {}).get("rebound_from_low_t") or {}).get("direction"),
                "REST11_direction": ((rest_comp.get(h) or {}).get("rebound_from_low_t") or {}).get("direction"),
                "ALL_auc_best": ((all_comp.get(h) or {}).get("rebound_from_low_t") or {}).get("auc_best"),
                "REST11_auc_best": ((rest_comp.get(h) or {}).get("rebound_from_low_t") or {}).get("auc_best"),
            }
            for h in PRIMARY_CHECKPOINTS
        ],
        "case17": [
            {"trade_id": c.get("trade_id"), "date": c.get("date"), "symbol": c.get("symbol"),
             "FINAL_WIN": c.get("FINAL_WIN"), "FINAL_LOSS": c.get("FINAL_LOSS"),
             "EXTEND": c.get("EXTEND_TO_750"), "pnl": c.get("pnl_yen_100"),
             **{f"h{p['horizon_sec']}_{k}": p.get(k) for p in c.get("path") or [] for k in ("bid_return", "mfe", "mae", "giveback", "rebound")}}
            for c in case17
        ],
        "winners": [
            {"group": g, **{k: v for k, v in c.items() if k != "path"}, "path": c.get("path")}
            for g, lst in winner_cases.items()
            for c in lst
        ],
        "rest11": [
            {
                "horizon_sec": h,
                "var": v,
                "ALL_direction": ((all_comp.get(h) or {}).get(v) or {}).get("direction"),
                "REST11_direction": ((rest_comp.get(h) or {}).get(v) or {}).get("direction"),
                "ALL_auc_best": ((all_comp.get(h) or {}).get(v) or {}).get("auc_best"),
                "REST11_auc_best": ((rest_comp.get(h) or {}).get(v) or {}).get("auc_best"),
            }
            for h in PRIMARY_CHECKPOINTS
            for v in KEY_VARS
        ],
        "daystab": [
            {"var": v, "horizon_sec": h, "same": blk.get("same_direction_days"), "opp": blk.get("opposite_direction_days"), "insuff": blk.get("insufficient_days")}
            for v, byh in day_stab.items()
            for h, blk in byh.items()
        ],
        "lodo": [
            {"var": v, "horizon_sec": h, "median": blk.get("median"), "min": blk.get("min"), "max": blk.get("max"), "n_days": blk.get("n_days_scored")}
            for v, byh in lodo.items()
            for h, blk in byh.items()
        ],
    }

    unused = SNAPSHOT_VARS, TRAJECTORY_VARS, at_horizon
    del unused

    report = {
        "task": "P4-2",
        "ANALYSIS_ID": "P4_2_RECOVERY_FAILURE_PATH",
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "strategy validation", "new EXIT validation"],
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "FROZEN": frozen,
        "PRIMARY_ADVERSE120": {
            "n": n_all["B"],
            "WIN": n_all["RW"],
            "LOSS": n_all["PF"],
            "DRAW": n_all["DRAW"],
            "eligible_A": n_all["A"],
        },
        "P4_1_TRIGGER": {"n": 17, "WIN": p41.get("WIN"), "LOSS": p41.get("LOSS"), "matched": n_all["C"]},
        "SNAPSHOT": {v: _summ(all_comp, v) for v in ("bid_return_from_fill", "executable_mfe_to_t", "executable_mae_to_t", "bid_giveback_from_peak")},
        "TRAJECTORY": {v: _summ(all_comp, v) for v in TRAJECTORY_VARS},
        "EARLIEST_DIVERGENCE": diverge,
        "RECOVERY_SHAPE": rec_shape,
        "PERSISTENT_FAILURE_SHAPE": pf_shape,
        "PERSISTENCE": persist,
        "NEW_LOW": new_low,
        "REST11": {
            "cohort": n_rest,
            "directions": {
                v: {str(h): ((rest_comp.get(h) or {}).get(v) or {}).get("direction") for h in PRIMARY_CHECKPOINTS}
                for v in ("bid_return_from_fill", "rebound_from_low_t", "delta_bid_120_to_t", "delta_mae_prev_checkpoint", "delta_mfe_prev_checkpoint")
            },
        },
        "DAY_STABILITY": {v: {h: {k: blk.get(k) for k in ("same_direction_days", "opposite_direction_days", "insufficient_days")} for h, blk in byh.items()} for v, byh in day_stab.items()},
        "LODO": {v: {h: {k: blk.get(k) for k in ("median", "min", "max", "n_days_scored")} for h, blk in byh.items()} for v, byh in lodo.items()},
        "P4_1_WINNER_CASES": winner_cases,
        "P4_1_17_CASEBOOK": case17,
        "MECHANISM_CLASSIFICATION": mech.get("MECHANISM_CLASSIFICATION"),
        "MECHANISM_DETAIL": mech,
        "CANDIDATE_PATH_FAMILIES": mech.get("CANDIDATE_PATH_FAMILIES") or [],
        "THRESHOLD_SELECTED": False,
        "NEW_EXIT_TESTED": False,
        "COUNTERFACTUAL_PNL_TESTED": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": bool(leak_n) if path_rows else False,
        "leak_n": int(leak_n),
        "identity_n": int(identity_n),
        "identity_fail": int(identity_fail),
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "P4_1_PRECOMMIT_SHA": p41.get("PRECOMMIT_SHA"),
            "file_sha_p4_0_state": _file_sha("src/research/mid_hold_state_separability_p4_0/state.py"),
            "file_sha_p4_2_path": _file_sha("src/research/mid_hold_recovery_failure_path_p4_2/path.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
    }
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    adv = rep.get("PRIMARY_ADVERSE120") or {}
    p41 = rep.get("P4_1_TRIGGER") or {}
    lines = [
        "# P4-2 Recovering Winner vs Persistent Failure",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / strategy validation / new EXIT validation",
        "P4-1 MID_HOLD_NO_PROGRESS_V1 remains CLOSED. No retune.",
        "",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        "",
        "PRIMARY_ADVERSE120:",
        f"n: {adv.get('n')}",
        f"WIN: {adv.get('WIN')}",
        f"LOSS: {adv.get('LOSS')}",
        f"DRAW: {adv.get('DRAW')}",
        "",
        "P4_1_TRIGGER:",
        f"17 matched={p41.get('matched')} WIN: {p41.get('WIN')} LOSS: {p41.get('LOSS')}",
        "",
        "SNAPSHOT (RW vs PF median, primary 120–360):",
    ]
    snap = rep.get("SNAPSHOT") or {}
    for v, blk in snap.items():
        lines.append(f"{v}: {json.dumps(blk, ensure_ascii=False, default=str)}")
    lines.append("")
    lines.append("TRAJECTORY:")
    for v, blk in (rep.get("TRAJECTORY") or {}).items():
        lines.append(f"{v}: {json.dumps(blk, ensure_ascii=False, default=str)}")
    lines += [
        "",
        f"EARLIEST_DIVERGENCE: {json.dumps({k: (v or {}).get('first_nonzero_median_direction') for k, v in (rep.get('EARLIEST_DIVERGENCE') or {}).items()}, ensure_ascii=False)}",
        "",
        f"RECOVERY_SHAPE: {json.dumps(rep.get('RECOVERY_SHAPE'), ensure_ascii=False, default=str)}",
        f"PERSISTENT_FAILURE_SHAPE: {json.dumps(rep.get('PERSISTENT_FAILURE_SHAPE'), ensure_ascii=False, default=str)}",
        "",
        f"REST11: {json.dumps(rep.get('REST11'), ensure_ascii=False, default=str)}",
        f"DAY_STABILITY: {json.dumps(rep.get('DAY_STABILITY'), ensure_ascii=False, default=str)}",
        f"LODO: {json.dumps(rep.get('LODO'), ensure_ascii=False, default=str)}",
        "",
        "P4_1_WINNER_CASES / TOP20_cut_2 / EXTEND_cut_2: see audit.xlsx Winner_Cases and report.json",
        "",
        f"MECHANISM_CLASSIFICATION: `{rep.get('MECHANISM_CLASSIFICATION')}`",
        f"{(rep.get('MECHANISM_DETAIL') or {}).get('why')}",
        f"CANDIDATE_PATH_FAMILIES: {json.dumps(rep.get('CANDIDATE_PATH_FAMILIES'), ensure_ascii=False)}",
        "",
        "THRESHOLD_SELECTED: false",
        "NEW_EXIT_TESTED: false",
        "COUNTERFACTUAL_PNL_TESTED: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("MECHANISM_CLASSIFICATION", public.get("MECHANISM_CLASSIFICATION")),
            ("CANDIDATE_PATH_FAMILIES", public.get("CANDIDATE_PATH_FAMILIES")),
            ("THRESHOLD_SELECTED", False),
            ("NEW_EXIT_TESTED", False),
            ("COUNTERFACTUAL_PNL_TESTED", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Cohorts"), sheets.get("cohorts") or [])
    _write_rows(wb.create_sheet("Checkpoint_Path"), sheets.get("path") or [])
    _write_rows(wb.create_sheet("Snapshot_vs_Trajectory"), sheets.get("snap_traj") or [])
    _write_rows(wb.create_sheet("Recovery"), sheets.get("recovery") or [])
    _write_rows(wb.create_sheet("Persistent_Failure"), sheets.get("persist_fail") or [])
    _write_rows(wb.create_sheet("New_Low"), sheets.get("newlow") or [])
    _write_rows(wb.create_sheet("MFE_Development"), sheets.get("mfe") or [])
    _write_rows(wb.create_sheet("Rebound"), sheets.get("rebound") or [])
    _write_rows(wb.create_sheet("P4_1_17_Casebook"), sheets.get("case17") or [])
    _write_rows(wb.create_sheet("Winner_Cases"), sheets.get("winners") or [])
    _write_rows(wb.create_sheet("REST11"), sheets.get("rest11") or [])
    _write_rows(wb.create_sheet("Day_Stability"), sheets.get("daystab") or [])
    _write_rows(wb.create_sheet("LODO"), sheets.get("lodo") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(ident, list((public.get("Identity") or {}).items()) + [("identity_fail", public.get("identity_fail"))])
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("THRESHOLD_SELECTED", False),
            ("NEW_EXIT_TESTED", False),
            ("COUNTERFACTUAL_PNL_TESTED", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("P4_1_CANDIDATE_CLOSED", True),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
