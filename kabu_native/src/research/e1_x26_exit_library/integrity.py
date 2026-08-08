"""X25 handoff / registry integrity."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x22_actual_exit_factory.registry import (
    build_alias_groups,
    load_population_checked,
    load_x21_registry,
    rebuild_candidates_and_masks,
)

from . import (
    EXPECTED_ALIASES,
    EXPECTED_CAND_N,
    EXPECTED_POP_N,
    EXPECTED_UNIQUE_MASKS,
    SOURCE_X25,
    X25_HANDOFF_SHA,
    X25_PATH_SHA,
)

NATIVE = Path(__file__).resolve().parents[3]
X25_DIR = NATIVE / "results" / "research" / "e1_x25_long_horizon_path"


def load_x25_report() -> dict[str, Any]:
    r = json.loads((X25_DIR / "report.json").read_text(encoding="utf-8"))
    assert r["run_id"] == SOURCE_X25, f"expected {SOURCE_X25}, got {r['run_id']}"
    return r


def load_x25_handoff_rows() -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(X25_DIR / "audit.xlsx", read_only=True, data_only=True)
    ws = wb["X26Handoff"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        tags = d.get("discovery_family_tags")
        if isinstance(tags, str):
            try:
                tags = json.loads(tags)
            except Exception:
                tags = [t.strip().strip('"').strip("'") for t in tags.strip("[]").split(",") if t.strip()]
        d["discovery_family_tags"] = tags if isinstance(tags, list) else []
        aliases = d.get("aliases")
        if isinstance(aliases, str):
            try:
                aliases = json.loads(aliases)
            except Exception:
                aliases = []
        d["aliases"] = aliases if isinstance(aliases, list) else []
        out.append(d)
    return out


def verify_handoff_sha(handoff: list[dict[str, Any]]) -> dict[str, Any]:
    payload = [{
        "id": h["candidate_id"],
        "tags": h["discovery_family_tags"],
        "st": h["path_evidence_status"],
        "d900": h.get("return_900s_delta_vs_ALL"),
    } for h in handoff]
    sha = sha256_obj(payload)
    return {
        "expected": X25_HANDOFF_SHA,
        "computed": sha,
        "ok": sha == X25_HANDOFF_SHA,
        "n": len(handoff),
    }


def verify_path_sha(x25_report: dict[str, Any]) -> dict[str, Any]:
    path_sha = (
        (x25_report.get("path_meta") or {}).get("path_sha256")
        or (x25_report.get("determinism") or {}).get("path_sha_a")
    )
    return {
        "expected": X25_PATH_SHA,
        "reported": path_sha,
        "ok": path_sha == X25_PATH_SHA,
    }


def verify_registry() -> dict[str, Any]:
    rows = load_population_checked()
    cands, masks = rebuild_candidates_and_masks(rows)
    alias_rows, _, unique = build_alias_groups(cands, masks)
    alias_n = sum(1 for a in alias_rows if not a["is_representative"])
    x21 = load_x21_registry()
    ok = (
        len(rows) == EXPECTED_POP_N
        and len(cands) == EXPECTED_CAND_N
        and len(unique) == EXPECTED_UNIQUE_MASKS
        and alias_n == EXPECTED_ALIASES
        and len(x21) == EXPECTED_CAND_N
    )
    return {
        "ok": ok,
        "anchors": len(rows),
        "candidates": len(cands),
        "unique_masks": len(unique),
        "aliases": alias_n,
        "rows": rows,
        "candidates_list": cands,
        "masks": masks,
        "alias_rows": alias_rows,
        "unique_masks_map": unique,
    }
