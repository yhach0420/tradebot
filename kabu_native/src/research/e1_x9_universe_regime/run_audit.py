"""E1_X9 orchestrator."""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.failure_source.clusters import load_episodes
from research.e1_x7_pfq.candidates import passes_candidate
from research.e1_x7_pfq.config import DAYS
from research.e1_x7_pfq.feature_contract import run_phase0_audit
from research.e1_x7_pfq.run_study import _load_pullback_universe
from research.e1_x8_symbol_leverage import FROZEN as PFQ_FROZEN

from . import ANALYSIS_ID, DOCUMENT_ID, FROZEN_UPDATE_THR, SOURCE_BRIDGE, SOURCE_PFQ, SOURCE_X8, TARGET_SYMBOL
from .metadata import (
    assign_index_status,
    assign_market_segment,
    direct_ownership_status,
    load_jpx_segment_scale,
    load_turnover_20d,
    market_cap_asof_status,
    tercile_labels,
)
from .precommit import build_precommit
from .regimes import (
    compare_regime_pair,
    decide_verdict,
    high_update_regime_split,
    regime_first_touch,
    update_signal_by_regime,
    within_symbol_normalization,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
BRIDGE_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_bridge_v2"
X8_DIR = NATIVE / "results" / "research" / "e1_x8_symbol_leverage"
REV_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_exit_revision"
PUBLISH = NATIVE / "results" / "research" / "e1_x9_universe_regime"


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "pfq_revived": False,
        "entry_changed": False,
        "exit_changed": False,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
    }


def run_once(*, label: str = "A") -> dict[str, Any]:
    run_id = f"e1x9_univ_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"

    x8 = json.loads((X8_DIR / "report.json").read_text(encoding="utf-8"))
    br = json.loads((BRIDGE_DIR / "report.json").read_text(encoding="utf-8"))
    rev = json.loads((REV_DIR / "report.json").read_text(encoding="utf-8"))
    assert SOURCE_X8 in str(x8.get("run_id"))
    assert SOURCE_BRIDGE in str(br.get("run_id"))
    assert SOURCE_PFQ in str(rev.get("run_id"))

    source_shas = {
        "x8_report": sha256_file(X8_DIR / "report.json"),
        "bridge_report": sha256_file(BRIDGE_DIR / "report.json"),
        "pfq_report": sha256_file(REV_DIR / "report.json"),
        "bridge_fixed_grid": (br.get("determinism_shas") or {}).get("fixed_grid_outcome_sha"),
    }
    print(f"=== [{label}] Precommit ===", flush=True)
    precommit = build_precommit(source_shas={k: str(v) for k, v in source_shas.items()})

    print(f"=== [{label}] Load episodes + features ===", flush=True)
    sm = load_source_manifest()
    events_by_day = {day: load_day_events(day, _universe_from_manifest(sm, day)) for day in DAYS}
    universe = _load_pullback_universe()
    episodes_raw, _, _ = load_episodes()
    ep_by = {e["episode_id"]: e for e in episodes_raw}
    audits, phase0 = run_phase0_audit(universe, events_by_day, ep_by)
    fg = {r["episode_id"]: r for r in _load_sheet(BRIDGE_DIR / "audit.xlsx", "FixedGridOutcome")}
    thr = dict(PFQ_FROZEN)

    eps = []
    for a in audits:
        f = fg.get(a["episode_id"]) or {}
        path_ok = bool(f.get("evaluable"))
        eps.append({
            "episode_id": a["episode_id"],
            "cluster_id": a.get("cluster_id"),
            "day": a["day"],
            "session": a.get("session"),
            "symbol": str(a["symbol"]),
            "price_update_count_10s": a.get("price_update_count_10s"),
            "uptick_volume_ratio_30s": a.get("uptick_volume_ratio_30s"),
            "ratio_valid": a.get("ratio_valid"),
            "update_eligible_parent": a.get("price_update_count_10s") is not None and path_ok,
            "mem_UPDATE": passes_candidate(a, "PFQ_UPDATE_Q70", thr),
            "evaluable": f.get("evaluable"),
            "best_net_pnl_bps_300s": f.get("best_net_pnl_bps_300s"),
            "ft_plus5_vs_minus10": f.get("ft_plus5_vs_minus10"),
            "ft_plus5_vs_minus15": f.get("ft_plus5_vs_minus15"),
            "ft_plus10_vs_minus10": f.get("ft_plus10_vs_minus10"),
            "ft_plus10_vs_minus15": f.get("ft_plus10_vs_minus15"),
        })

    symbols = sorted({e["symbol"] for e in eps})
    print(f"=== [{label}] Metadata as-of ===", flush=True)
    jpx = load_jpx_segment_scale()
    turnover = load_turnover_20d(set(symbols))
    mcap_status = market_cap_asof_status()
    direct_status = direct_ownership_status()

    # coverage
    jpx_hit = sum(1 for s in symbols if s in jpx and jpx[s]["asof_valid"])
    turn_hit = sum(1 for s in symbols if s in turnover and turnover[s]["asof_valid"])
    ep_jpx = sum(1 for e in eps if e["symbol"] in jpx and jpx[e["symbol"]]["asof_valid"])
    ep_turn = sum(1 for e in eps if e["symbol"] in turnover and turnover[e["symbol"]]["asof_valid"])
    coverage = {
        "market_symbol": jpx_hit / len(symbols),
        "market_episode": ep_jpx / len(eps),
        "index_scale_symbol": jpx_hit / len(symbols),
        "index_scale_episode": ep_jpx / len(eps),
        "turnover_symbol": turn_hit / len(symbols),
        "turnover_episode": ep_turn / len(eps),
        "market_cap_symbol": 0.0,
        "market_cap_episode": 0.0,
        "free_float_symbol": 0.0,
        "direct_ownership_symbol": 0.0,
        "direct_ownership_episode": 0.0,
        "core_market_ok": (jpx_hit / len(symbols)) >= 0.90 and (ep_jpx / len(eps)) >= 0.90,
        "core_turnover_ok": (turn_hit / len(symbols)) >= 0.90 and (ep_turn / len(eps)) >= 0.90,
        "core_mcap_ok": False,
        "direct_ok": False,
        "any_proxy_axis_evaluable": False,
    }
    coverage["any_proxy_axis_evaluable"] = coverage["core_market_ok"] or coverage["core_turnover_ok"]
    # segment/index and/or turnover usable as proxy axes; mcap missing
    core_proxy_evaluable = bool(coverage["any_proxy_axis_evaluable"])

    # assignments
    assignments = []
    tv_map = {s: turnover[s]["avg_trading_value_20d"] for s in symbols if s in turnover and turnover[s]["asof_valid"]}
    tv_terc = tercile_labels(tv_map)
    for s in symbols:
        j = jpx.get(s)
        t = turnover.get(s)
        row = {
            "symbol": s,
            "market_segment": assign_market_segment(j["market"]) if j and j["asof_valid"] else None,
            "index_status": assign_index_status(j["scale_category"]) if j and j["asof_valid"] else None,
            "scale_category": j["scale_category"] if j and j["asof_valid"] else None,
            "mcap_tercile": None,  # not as-of available
            "turnover_tercile": (
                {"LOW": "TURNOVER_LOW", "MID": "TURNOVER_MID", "HIGH": "TURNOVER_HIGH"}[tv_terc[s]]
                if s in tv_terc else None
            ),
            "institutional_tercile": None,
            "avg_trading_value_20d": t["avg_trading_value_20d"] if t and t["asof_valid"] else None,
            "asof_market_valid": bool(j and j["asof_valid"]),
            "asof_turnover_valid": bool(t and t["asof_valid"]),
        }
        assignments.append(row)
    assign_by = {r["symbol"]: r for r in assignments}

    # attach regimes to episodes
    for e in eps:
        a = assign_by[e["symbol"]]
        e["market_segment"] = a["market_segment"]
        e["index_status"] = a["index_status"]
        e["turnover_tercile"] = a["turnover_tercile"]
        e["mcap_tercile"] = a["mcap_tercile"]

    print(f"=== [{label}] Microstructure + regimes ===", flush=True)
    # microstructure by symbol
    micro = []
    by_sym = defaultdict(list)
    for e in eps:
        by_sym[e["symbol"]].append(e)
    for s, rs in sorted(by_sym.items()):
        pu = [float(r["price_update_count_10s"]) for r in rs if r.get("price_update_count_10s") is not None]
        flow = [float(r["uptick_volume_ratio_30s"]) for r in rs if r.get("ratio_valid") and r.get("uptick_volume_ratio_30s") is not None]
        p5 = [1.0 if r.get("ft_plus5_vs_minus10") == "PLUS_FIRST" else 0.0
              for r in rs if r.get("ft_plus5_vs_minus10") not in (None, "NOT_EVALUABLE")]
        p10 = [1.0 if r.get("ft_plus10_vs_minus10") == "PLUS_FIRST" else 0.0
               for r in rs if r.get("ft_plus10_vs_minus10") not in (None, "NOT_EVALUABLE")]
        plus5 = [1.0 if (r.get("best_net_pnl_bps_300s") is not None and float(r["best_net_pnl_bps_300s"]) >= 5) else 0.0
                 for r in rs if r.get("evaluable")]
        a = assign_by[s]
        micro.append({
            "symbol": s,
            "n_episodes": len(rs),
            "median_price_update_count_10s": float(median(pu)) if pu else None,
            "median_uptick_volume_ratio_30s": float(median(flow)) if flow else None,
            "median_spread_bps": None,  # not in frozen feature table
            "median_board_age_sec": None,
            "median_price_age_sec": None,
            "fixed_grid_plus5_before_minus10_rate": float(sum(p5) / len(p5)) if p5 else None,
            "fixed_grid_plus10_before_minus10_rate": float(sum(p10) / len(p10)) if p10 else None,
            "net_plus5_rate": float(sum(plus5) / len(plus5)) if plus5 else None,
            **{k: a[k] for k in ("market_segment", "index_status", "turnover_tercile", "mcap_tercile", "avg_trading_value_20d")},
            "is_285A": s == TARGET_SYMBOL,
        })

    # regime first-touch + update signal
    regime_ft = []
    update_by_reg = []
    proxy_comps = []

    def collect(axis: str, values: list[str]):
        blocks = {}
        for val in values:
            subset = [e for e in eps if e.get(axis) == val]
            parent = [e for e in eps if e.get(axis) == val]
            ft = regime_first_touch(subset, regime_name=axis, regime_value=val)
            us = update_signal_by_regime(subset, parent, regime_name=axis, regime_value=val)
            regime_ft.append(ft)
            update_by_reg.append(us)
            blocks[val] = ft
        return blocks

    if coverage["core_market_ok"]:
        seg = collect("market_segment", ["PRIME", "STANDARD", "GROWTH", "OTHER"])
        idx = collect("index_status", ["MAJOR_INDEX_MEMBER", "NON_MAJOR_INDEX"])
        # proxy: NON_MAJOR vs MAJOR first-touch
        proxy_comps.append(compare_regime_pair(
            idx.get("NON_MAJOR_INDEX") or {"status": "X"},
            idx.get("MAJOR_INDEX_MEMBER") or {"status": "X"},
        ))
        # STANDARD+GROWTH vs PRIME as low-participation proxy (not ownership claim)
        low_mkt = [e for e in eps if e.get("market_segment") in ("STANDARD", "GROWTH")]
        high_mkt = [e for e in eps if e.get("market_segment") == "PRIME"]
        proxy_comps.append(compare_regime_pair(
            regime_first_touch(low_mkt, regime_name="market_segment_group", regime_value="STANDARD_GROWTH"),
            regime_first_touch(high_mkt, regime_name="market_segment_group", regime_value="PRIME"),
        ))

    if coverage["core_turnover_ok"]:
        turn = collect("turnover_tercile", ["TURNOVER_LOW", "TURNOVER_MID", "TURNOVER_HIGH"])
        proxy_comps.append(compare_regime_pair(
            turn.get("TURNOVER_LOW") or {"status": "X"},
            turn.get("TURNOVER_HIGH") or {"status": "X"},
        ))

    # HIGH_UPDATE split
    hu = high_update_regime_split(eps)
    heavy_eps = [e for e in eps if e["symbol"] in hu["UPDATE_HEAVY"]]
    light_eps = [e for e in eps if e["symbol"] in hu["UPDATE_LIGHT"]]
    ft_heavy = regime_first_touch(heavy_eps, regime_name="update_intensity", regime_value="UPDATE_HEAVY")
    ft_light = regime_first_touch(light_eps, regime_name="update_intensity", regime_value="UPDATE_LIGHT")
    regime_ft.extend([ft_heavy, ft_light])
    uh_comp = compare_regime_pair(ft_heavy, ft_light)
    update_by_reg.append(update_signal_by_regime(
        heavy_eps, heavy_eps, regime_name="update_intensity", regime_value="UPDATE_HEAVY",
    ))
    update_by_reg.append(update_signal_by_regime(
        light_eps, light_eps, regime_name="update_intensity", regime_value="UPDATE_LIGHT",
    ))

    # interactions limited
    interactions = []
    if coverage["core_turnover_ok"] and coverage["core_market_ok"]:
        # INDEX × cannot do MCAP — substitute INDEX × TURNOVER as closest allowed? Spec only allows MCAP×TURNOVER and INDEX×MCAP.
        # Both need MCAP which is unavailable → mark NOT_EVALUABLE
        interactions.append({
            "interaction": "MCAP tercile × TURNOVER tercile",
            "status": "NOT_EVALUABLE_SUPPORT",
            "reason": "market_cap as-of unavailable",
        })
        interactions.append({
            "interaction": "INDEX status × MCAP tercile",
            "status": "NOT_EVALUABLE_SUPPORT",
            "reason": "market_cap as-of unavailable",
        })
    else:
        interactions.append({
            "interaction": "MCAP tercile × TURNOVER tercile",
            "status": "NOT_EVALUABLE_SUPPORT",
            "reason": "mcap and/or turnover coverage",
        })
        interactions.append({
            "interaction": "INDEX status × MCAP tercile",
            "status": "NOT_EVALUABLE_SUPPORT",
            "reason": "mcap as-of unavailable",
        })

    within = within_symbol_normalization(eps)

    # economic reference
    base_tr = _load_sheet(REV_DIR / "audit.xlsx", "BaselineTrades")
    rev_tr = _load_sheet(REV_DIR / "audit.xlsx", "RevisionTrades")

    def econ_by_regime(trades: list[dict], axis: str) -> list[dict]:
        by = defaultdict(list)
        for t in trades:
            sym = str(t.get("symbol"))
            reg = (assign_by.get(sym) or {}).get(axis)
            if reg is None:
                continue
            yen = float(t.get("pnl_yen_100") if t.get("pnl_yen_100") is not None else t.get("net_pnl_yen") or 0)
            by[reg].append(yen)
        out = []
        for reg, pnls in by.items():
            out.append({
                "axis": axis,
                "regime": reg,
                "n_trades": len(pnls),
                "pnl": sum(pnls),
                "top_trade_share": (max(pnls) / sum(pnls)) if pnls and sum(pnls) != 0 else None,
                "reference_only": True,
            })
        return out

    econ_ref = {
        "baseline_by_index": econ_by_regime(base_tr, "index_status"),
        "baseline_by_segment": econ_by_regime(base_tr, "market_segment"),
        "baseline_by_turnover": econ_by_regime(base_tr, "turnover_tercile"),
        "revision_by_index": econ_by_regime(rev_tr, "index_status"),
        "note": "frozen ledger reference only; not for adoption",
    }

    # 285A profile
    kiox = next((m for m in micro if m["symbol"] == TARGET_SYMBOL), None)
    similar = [
        m for m in micro
        if m["symbol"] != TARGET_SYMBOL
        and kiox
        and m.get("market_segment") == kiox.get("market_segment")
        and m.get("index_status") == kiox.get("index_status")
        and m.get("turnover_tercile") == kiox.get("turnover_tercile")
    ]

    vd = decide_verdict(
        coverage=coverage,
        direct_status={**direct_status, "evaluable": False},
        proxy_comparisons=[c for c in proxy_comps if c.get("status") == "OK"],
        update_heavy_vs_light=uh_comp,
        core_proxy_evaluable=core_proxy_evaluable,
    )

    # metadata source registry (universe symbols only)
    meta_sources = []
    for s in symbols:
        j = jpx.get(s)
        if j:
            meta_sources.append({
                "symbol": s, "field": "market", "value": j["market"],
                "source_name": j["source_name"], "source_location": j["source_location"],
                "effective_date": j["effective_date"], "publication_date": j["publication_date"],
                "retrieved_at": j["retrieved_at"], "asof_valid": j["asof_valid"],
            })
            meta_sources.append({
                "symbol": s, "field": "scale_category", "value": j["scale_category"],
                "source_name": j["source_name"], "source_location": j["source_location"],
                "effective_date": j["effective_date"], "publication_date": j["publication_date"],
                "retrieved_at": j["retrieved_at"], "asof_valid": j["asof_valid"],
            })
        t = turnover.get(s)
        if t:
            meta_sources.append({
                "symbol": s, "field": "avg_trading_value_20d", "value": t["avg_trading_value_20d"],
                "source_name": t["source_name"], "source_location": t["source_location"],
                "effective_date": t["effective_date"], "publication_date": t["publication_date"],
                "retrieved_at": t["retrieved_at"], "asof_valid": t["asof_valid"],
            })
            meta_sources.append({
                "symbol": s, "field": "avg_volume_20d", "value": t.get("avg_volume_20d"),
                "source_name": t["source_name"], "source_location": t["source_location"],
                "effective_date": t["effective_date"], "publication_date": t["publication_date"],
                "retrieved_at": t["retrieved_at"], "asof_valid": t["asof_valid"],
            })
    meta_sources.append({
        "symbol": "*", "field": "market_cap", "value": None,
        "source_name": "none", "source_location": "",
        "effective_date": None, "publication_date": None,
        "retrieved_at": datetime.now().isoformat(timespec="seconds"),
        "asof_valid": False, "note": mcap_status.get("reason"),
    })
    meta_sources.append({
        "symbol": "*", "field": "direct_institutional_ownership", "value": None,
        "source_name": "none", "source_location": "",
        "effective_date": None, "publication_date": None,
        "retrieved_at": datetime.now().isoformat(timespec="seconds"),
        "asof_valid": False, "note": direct_status.get("reason"),
    })

    det = {
        "metadata_identity_sha": sha256_obj({"jpx_n": len(jpx), "turn_n": len(turnover), "mcap": mcap_status}),
        "asof_validity_sha": sha256_obj({
            "jpx": all(j.get("asof_valid") for j in jpx.values()) if jpx else False,
            "turn": all(t.get("asof_valid") for t in turnover.values()) if turnover else False,
            "mcap": False,
            "direct": False,
        }),
        "coverage_sha": sha256_obj(coverage),
        "regime_assignment_sha": sha256_obj([(a["symbol"], a["market_segment"], a["index_status"], a["turnover_tercile"]) for a in assignments]),
        "path_outcome_sha": sha256_obj([(r.get("regime_name"), r.get("regime_value"), r.get("status"), (r.get("metrics") or {}).get("plus5_vs_minus10", {}).get("plus_first_rate")) for r in regime_ft]),
        "update_sensitivity_sha": sha256_obj([(u.get("regime_value"), u.get("supported"), (u.get("plus5_vs_minus10") or {}).get("difference")) for u in update_by_reg]),
        "within_symbol_reference_sha": sha256_obj({
            "raw": within.get("raw_ge8_plus5_before_minus10_rate"),
            "within": within.get("within_high_p70_plus5_before_minus10_rate"),
        }),
        "economic_reference_sha": sha256_obj(econ_ref),
        "verdict": vd["verdict"],
    }

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "precommit": precommit,
        "phase0_status": phase0.get("status"),
        "identity": {"n_episodes": len(eps), "n_symbols": len(symbols), "days": list(DAYS)},
        "coverage": coverage,
        "mcap_status": mcap_status,
        "direct_ownership_status": direct_status,
        "regime_definitions": {
            "MCAP": "NOT_EVALUABLE — no as-of market cap <= 20260720",
            "TURNOVER": "terciles on 20d mean trading_value ending latest features day <= 20260720",
            "INDEX": "MAJOR if JPX scale_category contains TOPIX/Nikkei/JPX else NON_MAJOR",
            "MARKET": "PRIME/STANDARD/GROWTH/OTHER from JPX market",
            "DIRECT": "DIRECT_INSTITUTIONAL_DATA_NOT_EVALUABLE",
        },
        "kioxia_profile": {
            **(kiox or {}),
            "similar_regime_symbols": [m["symbol"] for m in similar],
            "similar_n": len(similar),
            "unique_special_case": len(similar) == 0,
        },
        "regime_first_touch": [{k: v for k, v in r.items() if k != "metrics"} | {
            "plus5_vs_minus10_rate": ((r.get("metrics") or {}).get("plus5_vs_minus10") or {}).get("plus_first_rate"),
            "plus5_vs_minus10_ci95": ((r.get("metrics") or {}).get("plus5_vs_minus10") or {}).get("ci95"),
        } for r in regime_ft],
        "update_signal_by_regime": update_by_reg,
        "proxy_comparisons": proxy_comps,
        "update_heavy_vs_light": uh_comp,
        "interactions": interactions,
        "within_symbol_reference": {k: v for k, v in within.items() if k != "sample_rows"},
        "economic_reference": econ_ref,
        "verdict": vd["verdict"],
        "verdict_detail": vd,
        "pfq_policy": precommit["pfq_policy"],
        "determinism_shas": det,
        "safety": _safety(),
        "stop": True,
        "_sheets": {
            "MetadataSources": meta_sources,
            "SymbolMetadata": assignments,
            "Microstructure": micro,
            "RegimeAssignments": assignments,
            "RegimeSupport": regime_ft,
            "FirstTouchByRegime": [
                {
                    "regime_name": r.get("regime_name"),
                    "regime_value": r.get("regime_value"),
                    "status": r.get("status"),
                    "n_symbols": r.get("n_symbols"),
                    "n_episodes": r.get("n_episodes"),
                    "n_days": r.get("n_days"),
                    "plus5_vs_minus10": ((r.get("metrics") or {}).get("plus5_vs_minus10") or {}).get("plus_first_rate"),
                    "ci95": ((r.get("metrics") or {}).get("plus5_vs_minus10") or {}).get("ci95"),
                }
                for r in regime_ft
            ],
            "UpdateSignalByRegime": update_by_reg,
            "WithinSymbolReference": within.get("sample_rows") or [],
            "Interactions": interactions,
            "EconomicReference": (
                econ_ref["baseline_by_index"] + econ_ref["baseline_by_segment"] + econ_ref["baseline_by_turnover"]
            ),
            "KioxiaProfile": [kiox] if kiox else [],
            "DirectOwnership": [direct_status],
            "AsOfValidation": [
                {"field": "jpx_segment_scale", "asof_valid": coverage["core_market_ok"], "symbol_cov": coverage["market_symbol"]},
                {"field": "turnover_20d", "asof_valid": coverage["core_turnover_ok"], "symbol_cov": coverage["turnover_symbol"]},
                {"field": "market_cap", **mcap_status},
                {"field": "direct_ownership", **direct_status},
            ],
            "Coverage": [{"metric": k, "value": v} for k, v in coverage.items()],
            "RegimeDefinitions": [{"axis": k, "definition": v} for k, v in {
                "MCAP": "NOT_EVALUABLE",
                "TURNOVER": "20d trading_value terciles",
                "INDEX": "TOPIX scale vs non",
                "MARKET": "Prime/Standard/Growth/Other",
                "DIRECT": "NOT_EVALUABLE",
            }.items()],
        },
    }
    return report
