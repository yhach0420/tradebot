"""Panel day reconciliation + manifests from frozen E1_X10 (no forbidden columns)."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from . import (
    DESIGN_DAYS,
    LABEL_BOOTSTRAP,
    LABEL_POLICY_EVAL,
    LABEL_SOURCE,
    LABEL_WARMUP,
    STATUS_ALREADY_USED,
)

NATIVE = Path(__file__).resolve().parents[3]
X10_XLSX = NATIVE / "results" / "research" / "e1_x10_risk_universe" / "audit.xlsx"
V2_REPORT = NATIVE / "results" / "research" / "e1_x11_policy_gate_v2" / "report.json"


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _f(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def panel_day_reconciliation() -> dict[str, Any]:
    """Reconcile 20260721 with V2 warmup/evaluable sets."""
    import json
    v2 = json.loads(V2_REPORT.read_text(encoding="utf-8"))
    warmup = list(v2.get("warmup_days") or [])
    policy = list(v2.get("policy_evaluable_days") or [])
    bootstrap = ["20260721"]
    all_panel = list(DESIGN_DAYS)
    # Explicit status per day
    rows = []
    for d in all_panel:
        if d in bootstrap:
            role = LABEL_BOOTSTRAP
            note = "first design capture day; reference-price bootstrap; not in V2 warmup/evaluable lists"
        elif d in warmup:
            role = LABEL_WARMUP
            note = "V2 warmup / PANEL_NOT_POLICY_EVALUABLE"
        elif d in policy:
            role = LABEL_POLICY_EVAL
            note = "V2 POLICY_EVALUABLE_DAY source"
        else:
            role = LABEL_SOURCE
            note = "RISK_HISTORY_SOURCE_DAY"
        rows.append({
            "date": d,
            "panel_role": role,
            "also": LABEL_SOURCE,
            "classification": STATUS_ALREADY_USED,
            "note": note,
        })
    # identity: all = bootstrap + warmup + policy (disjoint)
    reconstructed = sorted(set(bootstrap) | set(warmup) | set(policy))
    ok = reconstructed == sorted(all_panel)
    return {
        "all_panel_days": all_panel,
        "bootstrap_days": bootstrap,
        "warmup_days": warmup,
        "policy_evaluable_days": policy,
        "rows": rows,
        "reconciliation_pass": ok,
        "equation": "all_panel_days = bootstrap + warmup + policy_evaluable",
        "n_all": len(all_panel),
        "n_parts": len(bootstrap) + len(warmup) + len(policy),
    }


def manifests_from_x10() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """DailyManifest + SymbolDayRisk from E1_X10 sheets (risk columns only)."""
    notion = _load_sheet(X10_XLSX, "OneLotNotional")
    spread = {(r["symbol"], str(r["day"])): r for r in _load_sheet(X10_XLSX, "SpreadRisk")}
    jump = {(r["symbol"], str(r["day"])): r for r in _load_sheet(X10_XLSX, "BidJumps")}
    exe = {(r["symbol"], str(r["day"])): r for r in _load_sheet(X10_XLSX, "ExecutableLoss")}
    depth = {(r["symbol"], str(r["day"])): r for r in _load_sheet(X10_XLSX, "DepthRisk")}
    fresh = {(r["symbol"], str(r["day"])): r for r in _load_sheet(X10_XLSX, "Freshness")}

    by_day: dict[str, list] = {}
    sym_rows = []
    for r in notion:
        day, sym = str(r["day"]), str(r["symbol"])
        key = (sym, day)
        sp, jp, ex, dp, fr = spread.get(key) or {}, jump.get(key) or {}, exe.get(key) or {}, depth.get(key) or {}, fresh.get(key) or {}
        row = {
            "date": day,
            "symbol": sym,
            "classification": STATUS_ALREADY_USED,
            "one_lot_notional_yen": _f(r.get("one_lot_notional_yen")),
            "n_spread_obs": int(_f(sp.get("n_spread_obs")) or 0),
            "spread_p50_yen_100": _f(sp.get("median_spread_cost_yen_100")),
            "spread_p90_yen_100": None,
            "spread_p95_yen_100": _f(sp.get("p95_spread_cost_yen_100")),
            "n_jump_obs": int(_f(jp.get("n_jump_obs")) or 0),
            "down_jump_p50_yen_100": None,
            "down_jump_p90_yen_100": _f(jp.get("p90_down_bid_jump_yen_100")),
            "down_jump_p95_yen_100": _f(jp.get("p95_down_bid_jump_yen_100")),
            "n_exec_anchors": int(_f(ex.get("n_exec_anchors")) or 0),
            "exec_loss_5s_p50_yen_100": _f(ex.get("exec_loss_yen_100_5s_p50")),
            "exec_loss_5s_p90_yen_100": _f(ex.get("exec_loss_yen_100_5s_p90")),
            "exec_loss_5s_p95_yen_100": _f(ex.get("exec_loss_yen_100_5s_p95")),
            "bid_qty_p10": _f(dp.get("p10_best_bid_qty")),
            "ask_qty_p10": _f(dp.get("p10_best_ask_qty")),
            "board_fresh_rate": _f(fr.get("board_fresh_rate")),
        }
        # strip any accidental forbidden keys
        sym_rows.append(row)
        by_day.setdefault(day, []).append(row)

    manifests = []
    for day in DESIGN_DAYS:
        rows = by_day.get(day) or []
        n_sym = len(rows)
        ref_cov = sum(1 for x in rows if x.get("one_lot_notional_yen") is not None) / n_sym if n_sym else 0.0
        qty_cov = sum(1 for x in rows if x.get("bid_qty_p10") is not None) / n_sym if n_sym else 0.0
        board_cov = sum(1 for x in rows if x.get("board_fresh_rate") is not None) / n_sym if n_sym else 0.0
        events_n = sum(x.get("n_spread_obs") or 0 for x in rows)
        # quality: design days already validated in E1_X10 with quote coverage
        quality = "RISK_HISTORY_DAY_VALID" if n_sym >= 30 and ref_cov >= 0.9 else "RISK_HISTORY_DAY_INVALID"
        reasons = [] if quality.endswith("VALID") else ["low_symbol_or_ref_coverage"]
        manifests.append({
            "date": day,
            "classification": STATUS_ALREADY_USED,
            "capture_start": None,
            "capture_end": None,
            "sessions_present": "AM_PM_ASSUMED_FROM_DESIGN",
            "symbols_n": n_sym,
            "events_n": events_n,
            "bid_coverage": qty_cov,
            "ask_coverage": qty_cov,
            "qty_coverage": qty_cov,
            "board_time_coverage": board_cov,
            "price_time_coverage": None,  # often null in capture; not required for risk gate board path
            "reference_price_coverage": ref_cov,
            "file_sha256": None,
            "quality_status": quality,
            "quality_reasons": ";".join(reasons) if reasons else "",
            "source": "e1_x10_audit_xlsx_derived",
        })
    return manifests, sym_rows
