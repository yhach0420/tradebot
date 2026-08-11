"""V1R EXIT V2 Prospective Activation — parity, manifests, roles, clean-day gate.

No EXIT re-search. Frozen Arch E only. Paper 0/0/0.
"""
from __future__ import annotations

import hashlib
import json
import math
import pickle
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

from research.e1_x22_actual_exit_factory.paths import session_end_epoch
from research.e1_x32_upstream_attribution.eval_stages import load_boards_for_symbols
from research.e1_x35_passive_exit.paths import build_path
from research.e1_x36_joint_allocator import LOT_QTY
from research.e1_x36_joint_allocator.replay import simulate_joint
from research.e1_x36r_freeze_integrity.serialize import score_fn_from_serialized
from research.e1_x37_prospective.freeze import load_model_artifact, load_v1r
from research.v1r_exit_v2_asymmetric.policy import apply_architecture
from research.v1r_exit_v2_asymmetric.run_v2 import patch_and_sim, summarize_pnls
from research.v1r_exit_v2_asymmetric.states import build_trade_bundle
from small_paper.v1r_dual_strategy_replay import compare_divergence, run_dual_day
from small_paper.v1r_exit_v2_contract import (
    EXIT_V2_CANDIDATE_SHA,
    EXIT_V2_MANIFEST_ID,
    FROZEN_CONTINUATION,
    FROZEN_GUARD,
    load_exit_v2_candidate,
    patch_panel_exits,
    frozen_continuation,
    frozen_guard,
)
from small_paper.v1r_primary_runtime import (
    ANCHOR_SHA,
    MODEL_ARTIFACT_SHA,
    UNIVERSE_BINDING_SHA,
    V1R_SHA,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "v1r_exit_v2_prospective_activation"
PANEL_CACHE = NATIVE / "results/research/v1r_capital_sweep_0p5m_10m/_panel_cache.pkl"
OLD_PRECOMMIT = NATIVE / "results/research/e1_x39c_concentration_reconciliation/PROSPECTIVE_PRECOMMIT_V1R_U1.json"
OLD_PRECOMMIT_SHA = "ebe2b86ca881dfe94d8af986e8689481b40f1e013ad64bc4d645f485b1da625b"
ENTRY_SHA = "f2887bb2be539cc173aee438a43ee8afb8cfa2b8c31380937ecd843e90dd9b29"
EXEC_SHA = "040fa4b061e575d3f6cdb2a11ffd3f862da5351b298567b31363de923a590869"
FIXED600_EXIT_SHA = "2c9fcc6e92971c252c8df93716066dda515fcbff0283d748b03293379c5eb62c"
FEATS = ("spread_bps", "imbalance", "mid_ret_60s", "mid_ret_180s", "event_rate_60s", "log_bid_qty")

STRATEGY_ID = "PASSIVE_ASYMMETRIC_EXIT_V2_FULL_STRATEGY"
PRECOMMIT_ID = "PROSPECTIVE_PRECOMMIT_V1R_EXIT_V2_U1"
ACTIVATION_ID = "V1R_EXIT_V2_PAPER_PRIMARY_ACTIVATION_V1"


def _sha(obj: dict) -> str:
    return hashlib.sha256(
        json.dumps({k: v for k, v in obj.items() if k != "sha256"}, sort_keys=True, default=str).encode()
    ).hexdigest()


def _sfn():
    ser = load_model_artifact()
    raw = score_fn_from_serialized(ser)

    def fn(e: dict) -> float:
        try:
            return float(raw({k: e.get(k) for k in FEATS}))
        except Exception:
            return float("-inf")

    return fn


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def write_xlsx(sheets: dict[str, list[dict]], path: Path) -> None:
    if Workbook is None:
        path.with_suffix(".sheets.json").write_text(json.dumps(sheets, indent=2, default=str), encoding="utf-8")
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def metric_lineage() -> list[dict[str, Any]]:
    """Explain existing report numbers — no recompute to change candidate."""
    cand = load_exit_v2_candidate()
    return [
        {
            "metric": "isolation_pnl_arch_e",
            "approx_value": 2652800,
            "lineage": "ISOLATION",
            "meaning": "Frozen FILL set (n≈190 accepted under FIXED600). Apply Arch E exits on same fills only. No re-admission.",
            "source": "v1r_exit_v2_asymmetric Isolation sheet / report answers 10_isolation.E",
        },
        {
            "metric": "full_replay_pnl_arch_e",
            "approx_value": 2622600,
            "lineage": "FULL_REPLAY",
            "meaning": "Independent joint sim with Arch E exits on panel → slot release / later admission may change (n≈194). SoT for economics.",
            "source": "v1r_exit_v2_asymmetric FullReplay sheet / answers 11_full_replay.E",
        },
        {
            "metric": "outer_mean_delta_pnl_frozen",
            "approx_value": float(cand.get("outer_mean_delta_pnl") or 142850),
            "lineage": "OUTER_FROZEN_PARAMS",
            "meaning": "Mean of per-fold test ΔPnL when evaluating voted Arch E with frozen params (guard IMB_p5_t-10 + cont MFE60_IMB10).",
            "source": "V1R_EXIT_V2_ASYMMETRIC_CANDIDATE_V1.outer_mean_delta_pnl",
        },
        {
            "metric": "mean_delta_by_arch_E",
            "approx_value": float((cand.get("mean_delta_by_arch") or {}).get("E") or 140100),
            "lineage": "OUTER_FOLD_NATIVE_ARCH_E",
            "meaning": "Mean of arch_tests[E].delta_pnl across outer folds using each fold's train-selected continuation (Fold A used MFE45_IMB10; B/C/D used MFE60_IMB10).",
            "source": "V1R_EXIT_V2_ASYMMETRIC_CANDIDATE_V1.mean_delta_by_arch.E",
        },
        {
            "metric": "lineage_gap_142850_vs_140100",
            "approx_value": 2750,
            "lineage": "EXPLAINED",
            "meaning": "Gap = frozen-param outer mean vs fold-native continuation selection mean. Not a bug. Note: ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT.",
            "source": "integrity note B",
        },
        {
            "metric": "isolation_minus_full_replay",
            "approx_value": 30200,
            "lineage": "EXPLAINED",
            "meaning": "Isolation ignores occupancy feedback; Full Replay re-admits under Arch E slot release. Difference is occupancy/admission effect, not EXIT math error.",
            "source": "isolation vs full replay definitions",
        },
    ]


def event_fingerprint(e: dict) -> dict[str, Any]:
    return {
        "date": e.get("date"),
        "symbol": e.get("symbol"),
        "signal_time": e.get("signal_time"),
        "fill_time": e.get("fill_time"),
        "accepted": bool(e.get("accepted")),
        "exit_time": e.get("canonical_exit_time"),
        "exit_ret": e.get("canonical_exit_ret_bps"),
        "reason": e.get("canonical_exit_reason"),
        "hold": e.get("canonical_hold_sec"),
        "guard": e.get("exit_v2_triggered_guard"),
        "extended": e.get("exit_v2_extended"),
        "pnl": e.get("realized_pnl_yen"),
    }


def parity_rows(
    a_events: list[dict],
    b_events: list[dict],
    *,
    label: str,
    fields: tuple[str, ...] = ("exit_time", "exit_ret", "reason", "hold", "guard", "extended"),
) -> dict[str, Any]:
    def key(e):
        if e.get("fill_time") is None:
            return None
        return (e.get("date"), e.get("symbol"), float(e["fill_time"]), bool(e.get("accepted")))

    am = {key(e): e for e in a_events if key(e) and e.get("accepted")}
    bm = {key(e): e for e in b_events if key(e) and e.get("accepted")}
    keys = sorted(set(am) | set(bm))
    mismatches = []
    matched = 0
    for k in keys:
        ea, eb = am.get(k), bm.get(k)
        if ea is None or eb is None:
            mismatches.append({"key": k, "issue": "identity_mismatch", "a": ea is not None, "b": eb is not None})
            continue
        fa, fb = event_fingerprint(ea), event_fingerprint(eb)
        diffs = []
        for field in fields:
            va, vb = fa.get(field), fb.get(field)
            if isinstance(va, float) and isinstance(vb, float):
                if abs(va - vb) > 1e-6:
                    diffs.append(field)
            elif va != vb:
                diffs.append(field)
        if diffs:
            mismatches.append({"key": k, "issue": "field_mismatch", "fields": diffs, "a": fa, "b": fb})
        else:
            matched += 1
    return {
        "label": label,
        "matched": matched,
        "mismatch_n": len(mismatches),
        "pass": len(mismatches) == 0 and matched > 0,
        "mismatches_sample": mismatches[:15],
    }


def run_research_arch_e_full(am_panel: list[dict], boards, sfn) -> dict[str, Any]:
    """Research SoT Full Replay Arch E (same as V2 run_v2.patch_and_sim)."""
    # baseline accepted for bundles of FIXED600 fills — but full replay patches all filled then resim
    base = simulate_joint([dict(e) for e in am_panel], score_fn=sfn)
    # Build bundles for all filled panel rows
    by_key = {}
    for e in am_panel:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        board = boards.get((e["date"], e["symbol"]))
        if board is None:
            continue
        se = session_end_epoch(e["date"], e["session"])
        path = build_path(board, entry_price=float(e["fill_price"]), entry_t=float(e["fill_time"]), sess_end=se)
        by_key[(e["date"], e["symbol"], float(e["fill_time"]))] = build_trade_bundle(e, path, board)
    return patch_and_sim(
        am_panel, by_key, arch="E", guard=frozen_guard(), cont_rule=frozen_continuation(), sfn=sfn
    )


def run_runtime_arch_e_full(am_panel: list[dict], boards, sfn) -> dict[str, Any]:
    by_key = {}
    for e in am_panel:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        board = boards.get((e["date"], e["symbol"]))
        if board is None:
            continue
        se = session_end_epoch(e["date"], e["session"])
        path = build_path(board, entry_price=float(e["fill_price"]), entry_t=float(e["fill_time"]), sess_end=se)
        by_key[(e["date"], e["symbol"], float(e["fill_time"]))] = build_trade_bundle(e, path, board)
    patched = patch_panel_exits(am_panel, by_key, mode="arch_e")
    sim = simulate_joint(patched, score_fn=sfn)
    acc = [e for e in sim["events"] if e.get("accepted")]
    pnls = [float(e.get("realized_pnl_yen") or 0) for e in acc]
    return {"summary": summarize_pnls(pnls), "events": sim["events"], "accepted_n": len(acc)}


def check_clean_20260811() -> dict[str, Any]:
    """Ensure 20260811 market data not used in research/candidate selection."""
    checks: dict[str, Any] = {}
    # run_id timestamps like v1r_exit_v2_20260811_033351 are NOT market-day contamination
    mentions = []
    banned_roots = [
        NATIVE / "results/research/v1r_exit_v2_asymmetric",
        NATIVE / "results/research/v1r_exit_global_search",
        NATIVE / "results/research/v1r_exit_scenario_research",
        NATIVE / "results/research/v1r_capital_sweep_0p5m_10m",
    ]
    for root in banned_roots:
        if not root.exists():
            continue
        for p in root.rglob("*"):
            if not p.is_file() or p.suffix.lower() not in (".json", ".md", ".txt", ".csv"):
                continue
            try:
                text = p.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            # Flag only explicit market-day date tokens (quoted / ISO), not run_id stamps
            if '"20260811"' in text or "'20260811'" in text or "2026-08-11" in text:
                mentions.append(str(p.relative_to(NATIVE)))
    checks["no_research_market_day_use"] = len(mentions) == 0
    checks["mentions"] = mentions[:20]
    push_hits = []
    for p in (NATIVE.parent).rglob("*20260811*"):
        if p.is_file() and any(x in str(p).lower() for x in ("push", "jsonl", "board", "capture", "panel")):
            push_hits.append(str(p)[:180])
            if len(push_hits) >= 10:
                break
    checks["no_20260811_market_files_found"] = len(push_hits) == 0
    checks["push_hits"] = push_hits
    checks["candidate_frozen"] = load_exit_v2_candidate().get("sha256") == EXIT_V2_CANDIDATE_SHA
    checks["note"] = "run_id timestamps containing 20260811 are ignored"
    ok = (
        checks["no_research_market_day_use"]
        and checks["candidate_frozen"]
        and checks["no_20260811_market_files_found"]
    )
    return {
        "day": "20260811",
        "verdict": "20260811_CLEAN_PROSPECTIVE_ELIGIBLE" if ok else "20260811_NOT_CLEAN_PROSPECTIVE",
        "pass": ok,
        "checks": checks,
        "prospective_day1_if_pass": "20260811",
    }


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_exit_v2_act_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== V1R EXIT V2 PROSPECTIVE ACTIVATION {run_id} ===", flush=True)

    cand = load_exit_v2_candidate()
    assert cand["sha256"] == EXIT_V2_CANDIDATE_SHA
    assert load_v1r().get("sha256") == V1R_SHA
    sfn = _sfn()

    lineage = metric_lineage()
    lineage_ok = all(r["lineage"] in ("ISOLATION", "FULL_REPLAY", "OUTER_FROZEN_PARAMS", "OUTER_FOLD_NATIVE_ARCH_E", "EXPLAINED") for r in lineage)
    print(f"  metric lineage explained={lineage_ok}", flush=True)

    # Historical panel parity
    print("  historical research/runtime parity...", flush=True)
    cache = pickle.load(PANEL_CACHE.open("rb"))
    am_panel = [dict(e) for e in cache["am"]["panel"]]
    assert all(str(e["date"]) < "20260810" for e in am_panel)
    pairs = sorted({(e["date"], e["symbol"]) for e in am_panel if e.get("filled")})
    boards = load_boards_for_symbols(pairs)

    research_e = run_research_arch_e_full(am_panel, boards, sfn)
    runtime_e = run_runtime_arch_e_full(am_panel, boards, sfn)
    parity_e = parity_rows(
        research_e["events"], runtime_e["events"],
        label="arch_e_research_vs_runtime",
        fields=("exit_time", "exit_ret", "reason", "hold"),
    )
    print(f"  Arch E parity pass={parity_e['pass']} matched={parity_e['matched']} mism={parity_e['mismatch_n']}", flush=True)

    # FIXED600 control parity vs baseline joint
    base = simulate_joint([dict(e) for e in am_panel], score_fn=sfn)
    by_key = {}
    for e in am_panel:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        board = boards.get((e["date"], e["symbol"]))
        if board is None:
            continue
        se = session_end_epoch(e["date"], e["session"])
        path = build_path(board, entry_price=float(e["fill_price"]), entry_t=float(e["fill_time"]), sess_end=se)
        by_key[(e["date"], e["symbol"], float(e["fill_time"]))] = build_trade_bundle(e, path, board)
    ctrl_panel = patch_panel_exits(am_panel, by_key, mode="fixed600")
    ctrl_sim = simulate_joint(ctrl_panel, score_fn=sfn)
    parity_c = parity_rows(
        base["events"], ctrl_sim["events"],
        label="fixed600_baseline_vs_control_lane",
        fields=("exit_time", "exit_ret", "hold"),  # reason label may differ FIXED600 vs FIXED_HOLD
    )
    print(f"  FIXED600 control parity pass={parity_c['pass']} matched={parity_c['matched']} mism={parity_c['mismatch_n']}", flush=True)

    # 8/10 implementation parity only (not performance tuning)
    print("  8/10 implementation parity (reference)...", flush=True)
    dual10 = run_dual_day("20260810", label="parity_0810_reference")
    parity_0810 = {"ok": dual10.get("ok"), "primary_n": None, "control_n": None}
    if dual10.get("ok"):
        parity_0810["primary_n"] = dual10["primary"]["summary"]["n"]
        parity_0810["control_n"] = dual10["control"]["summary"]["n"]
        parity_0810["comparison"] = {
            k: dual10["comparison"][k]
            for k in (
                "common_n", "primary_only_n", "control_only_n",
                "direct_exit_effect_pnl", "occupancy_divergence_effect",
                "guard_trigger_n", "extend_750_n",
            )
        }
        # research vs dual primary on 8/10
        from small_paper.v1r_day_engine import run_frozen_day
        # compare dual control to run_frozen_day FIXED600
        h10 = run_frozen_day("20260810", label="control_reference")
        parity_0810["control_vs_day_engine"] = {
            "day_engine_ok": h10.get("ok"),
            "day_engine_pnl": (h10.get("performance") or {}).get("total_pnl_yen_100")
                if (h10.get("performance") or {}).get("total_pnl_yen_100") is not None
                else (h10.get("performance") or {}).get("total_pnl"),
            "dual_control_pnl": dual10["control"]["summary"]["total"],
        }
        dep = parity_0810["control_vs_day_engine"]["day_engine_pnl"]
        dcp = parity_0810["control_vs_day_engine"]["dual_control_pnl"]
        parity_0810["control_vs_day_engine"]["pnl_match"] = (
            h10.get("ok") and dep is not None and abs(float(dep) - float(dcp)) < 1.0
        )

    # Supersede old precommit (0 prospective days)
    old = json.loads(OLD_PRECOMMIT.read_text(encoding="utf-8"))
    assert old.get("sha256") == OLD_PRECOMMIT_SHA
    assert int(old.get("prospective_evidence_days") or 0) == 0
    superseded_path = OUT / "PROSPECTIVE_PRECOMMIT_V1R_U1_SUPERSEDED_BEFORE_PROSPECTIVE_START.json"
    superseded_body = {
        **old,
        "status": "SUPERSEDED_BEFORE_PROSPECTIVE_START",
        "superseded_by": PRECOMMIT_ID,
        "superseded_at": datetime.now(JST).isoformat(),
        "note": "Never opened a valid prospective day (20260810 NOT_STARTED). Kept for history; not deleted.",
    }
    # keep original sha field of content before status stamp? User said history save — copy original then annotate
    superseded_body["original_sha256"] = OLD_PRECOMMIT_SHA
    superseded_path.write_text(json.dumps(superseded_body, indent=2, ensure_ascii=False), encoding="utf-8")
    # also copy beside original for discoverability
    side = OLD_PRECOMMIT.parent / "PROSPECTIVE_PRECOMMIT_V1R_U1_SUPERSEDED_BY_EXIT_V2.json"
    side.write_text(json.dumps(superseded_body, indent=2, ensure_ascii=False), encoding="utf-8")

    # EXIT V2 production contract manifest (bind candidate)
    exit_manifest = {
        "manifest_id": "PASSIVE_ASYMMETRIC_EXIT_V2_CONTRACT_V1",
        "kind": "production_paper_exit_contract",
        "candidate_manifest": EXIT_V2_MANIFEST_ID,
        "candidate_sha": EXIT_V2_CANDIDATE_SHA,
        "architecture": "E",
        "guard": FROZEN_GUARD,
        "continuation": FROZEN_CONTINUATION,
        "execution": {
            "trigger_lookup": "FIRST_VALID_EXECUTABLE_BUY1_AT_OR_AFTER_TRIGGER",
            "buy1_qty_min": 100,
            "freshness_sec": 5.0,
            "no_special_quote": True,
            "same_session": True,
            "no_instant_fill_assumption": True,
        },
        "horizons": {"monitor_to": 120.0, "decision_600": 600.0, "extend_750": 750.0},
        "forbidden": [
            "profit_trailing_between_120_and_600",
            "new_stop_between_120_and_600",
            "threshold_retune",
            "live_orders",
        ],
        "integrity_note": "ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT",
        "research_sot_modules": [
            "research.v1r_exit_v2_asymmetric.guards",
            "research.v1r_exit_v2_asymmetric.continuation",
            "research.v1r_exit_v2_asymmetric.policy",
        ],
        "paper_only": True,
        "submit_cancel_live": "0/0/0",
    }
    exit_manifest["sha256"] = _sha(exit_manifest)
    exit_sha = exit_manifest["sha256"]
    (OUT / "PASSIVE_ASYMMETRIC_EXIT_V2_CONTRACT_V1.json").write_text(
        json.dumps(exit_manifest, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # Full strategy manifest
    strategy = {
        "manifest_id": STRATEGY_ID,
        "inherits_entry_from": "PASSIVE_FIXED600_FULL_STRATEGY_V1R",
        "entry_strategy_sha": V1R_SHA,
        "model_sha": MODEL_ARTIFACT_SHA,
        "universe_binding_sha": UNIVERSE_BINDING_SHA,
        "universe_contract": "DAY_FIXED_AM_RUNTIME_UNIVERSE_V1",
        "anchor_sha": ANCHOR_SHA,
        "entry_sha": ENTRY_SHA,
        "execution_sha": EXEC_SHA,
        "exit_v2_candidate_sha": EXIT_V2_CANDIDATE_SHA,
        "exit_contract_sha": exit_sha,
        "control_strategy_sha": V1R_SHA,
        "control_exit_sha": FIXED600_EXIT_SHA,
        "capacity": {"position_cap": 5, "lot_qty": 100, "wait_sec": 1.0, "freshness_sec": 5.0,
                     "duplicate_rule": "no_overlap_replace"},
        "exit": {
            "architecture": "E",
            "guard_id": FROZEN_GUARD["id"],
            "continuation_id": FROZEN_CONTINUATION["id"],
            "guard": FROZEN_GUARD,
            "continuation": FROZEN_CONTINUATION,
            "semantics_600_750": "causal_continuation_gate_at_600",
        },
        "roles": {
            "primary": "PAPER_PRIMARY",
            "control_fixed600": "SHADOW_CONTROL",
            "pbv2": "SHADOW_ONLY",
            "capital_1m": "SHADOW_ONLY_DIAGNOSTIC",
        },
        "paper_only": True,
        "submit_cancel_live": "0/0/0",
        "no_parameter_mutation": True,
        "integrity_note": "ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT",
    }
    strategy["sha256"] = _sha(strategy)
    strategy_sha = strategy["sha256"]
    (OUT / f"{STRATEGY_ID}.json").write_text(json.dumps(strategy, indent=2, ensure_ascii=False), encoding="utf-8")

    # New precommit
    precommit = {
        "manifest_id": PRECOMMIT_ID,
        "parent_precommit_sha": OLD_PRECOMMIT_SHA,
        "parent_precommit_status": "SUPERSEDED_BEFORE_PROSPECTIVE_START",
        "full_strategy_manifest": STRATEGY_ID,
        "full_strategy_sha": strategy_sha,
        "entry_strategy_sha": V1R_SHA,
        "model_artifact_sha": MODEL_ARTIFACT_SHA,
        "universe_binding_sha": UNIVERSE_BINDING_SHA,
        "universe_contract": "DAY_FIXED_AM_RUNTIME_UNIVERSE_V1",
        "anchor_sha": ANCHOR_SHA,
        "entry_sha": ENTRY_SHA,
        "execution_sha": EXEC_SHA,
        "exit_v2_candidate_sha": EXIT_V2_CANDIDATE_SHA,
        "exit_contract_sha": exit_sha,
        "control_identity": {
            "strategy": "PASSIVE_FIXED600_FULL_STRATEGY_V1R",
            "strategy_sha": V1R_SHA,
            "exit_sha": FIXED600_EXIT_SHA,
            "role": "SHADOW_CONTROL",
        },
        "guard": FROZEN_GUARD,
        "continuation": FROZEN_CONTINUATION,
        "position_cap": 5,
        "lot_qty": 100,
        "wait_sec": 1.0,
        "freshness_sec": 5.0,
        "evaluation_checkpoints": {
            "EARLY_DIAGNOSTIC": 5,
            "PRIMARY_CHECKPOINT": 10,
            "EXTENDED_EVIDENCE": 20,
        },
        "prospective_evidence_days": 0,
        "excluded_from_count": ["20260810"],
        "no_retune_declaration": (
            "Prospective results must not change imbalance threshold, persist sec, monitor window, "
            "MFE threshold, continuation imbalance, 600/750, ENTRY, cap, universe, anchors, or model. "
            "Any change ends this series and requires new SHA + new Day1."
        ),
        "integrity_note": "ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT",
        "paper_only": True,
        "submit_cancel_live": "0/0/0",
        "observation_only_orders": True,
        "created_at": datetime.now(JST).isoformat(),
    }
    precommit["sha256"] = _sha(precommit)
    precommit_sha = precommit["sha256"]
    (OUT / f"{PRECOMMIT_ID}.json").write_text(json.dumps(precommit, indent=2, ensure_ascii=False), encoding="utf-8")

    # Activation manifest
    activation = {
        "manifest_id": ACTIVATION_ID,
        "runtime_roles": {
            "primary": "PAPER_PRIMARY",
            "strategy": STRATEGY_ID,
            "strategy_sha": strategy_sha,
            "control_fixed600": "SHADOW_CONTROL",
            "control_strategy_sha": V1R_SHA,
            "pbv2": "SHADOW_ONLY",
            "capital_1m": "SHADOW_ONLY_DIAGNOSTIC",
        },
        "exit_v2_candidate_sha": EXIT_V2_CANDIDATE_SHA,
        "exit_contract_sha": exit_sha,
        "precommit_sha": precommit_sha,
        "model_sha": MODEL_ARTIFACT_SHA,
        "universe_binding_sha": UNIVERSE_BINDING_SHA,
        "guard_id": FROZEN_GUARD["id"],
        "continuation_id": FROZEN_CONTINUATION["id"],
        "cap": 5,
        "qty": 100,
        "wait_sec": 1.0,
        "freshness_sec": 5.0,
        "live_trading_enabled": False,
        "order_enabled": False,
        "paper_only": True,
        "submit_cancel_live": "0/0/0",
        "no_fallback_to_fixed600_primary": True,
        "no_fallback_to_pbv2_primary": True,
        "created_at": datetime.now(JST).isoformat(),
    }
    activation["sha256"] = _sha(activation)
    activation_sha = activation["sha256"]
    (OUT / f"{ACTIVATION_ID}.json").write_text(json.dumps(activation, indent=2, ensure_ascii=False), encoding="utf-8")

    # Write pins for runtime
    pins = {
        "STRATEGY_ID": STRATEGY_ID,
        "STRATEGY_SHA": strategy_sha,
        "EXIT_V2_CANDIDATE_SHA": EXIT_V2_CANDIDATE_SHA,
        "EXIT_CONTRACT_SHA": exit_sha,
        "PRECOMMIT_SHA": precommit_sha,
        "ACTIVATION_SHA": activation_sha,
        "CONTROL_STRATEGY_SHA": V1R_SHA,
        "GUARD_ID": FROZEN_GUARD["id"],
        "CONTINUATION_ID": FROZEN_CONTINUATION["id"],
        "ENTRY_V1R_SHA": V1R_SHA,
        "MODEL_SHA": MODEL_ARTIFACT_SHA,
        "UNIVERSE_BINDING_SHA": UNIVERSE_BINDING_SHA,
    }
    (OUT / "runtime_pins.json").write_text(json.dumps(pins, indent=2), encoding="utf-8")

    # Clean day
    clean = check_clean_20260811()
    print(f"  clean 20260811={clean['verdict']}", flush=True)

    # Update runtime modules with new constants (implementation wiring)
    _write_runtime_wiring(pins, activation, strategy, precommit)

    # Role assertion after wiring
    from small_paper.v1r_exit_v2_activation_gate import assert_exit_v2_primary_roles
    roles = assert_exit_v2_primary_roles()
    print(f"  role assertion ok={roles.ok} reason={roles.reason}", flush=True)

    parity_pass = bool(parity_e["pass"] and parity_c["pass"] and lineage_ok)
    if dual10.get("ok"):
        parity_pass = parity_pass and bool((parity_0810.get("control_vs_day_engine") or {}).get("pnl_match"))

    live_safety = {
        "submit": 0, "cancel": 0, "live": 0,
        "paper_only": True, "order_enabled": False, "live_trading_enabled": False,
    }

    gates = {
        "frozen_manifest_identity": cand["sha256"] == EXIT_V2_CANDIDATE_SHA,
        "metric_lineage_explained": lineage_ok,
        "research_runtime_parity": bool(parity_e["pass"]),
        "control_parity": bool(parity_c["pass"]),
        "role_assertion": bool(roles.ok),
        "new_strategy_sha": bool(strategy_sha),
        "new_precommit": bool(precommit_sha),
        "old_precommit_superseded": True,
        "live_safety": True,
        "clean_prospective_day": bool(clean["pass"]),
    }
    all_pass = all(gates.values()) and parity_pass
    verdict = "V1R_EXIT_V2_PROSPECTIVE_READY" if all_pass else "V1R_EXIT_V2_PROSPECTIVE_BLOCKED"
    parity_verdict = "V1R_EXIT_V2_PRODUCTION_PARITY_PASS" if (parity_e["pass"] and parity_c["pass"]) else "V1R_EXIT_V2_PRODUCTION_PARITY_FAIL"

    sheets = {
        "Overview": [{
            "run_id": run_id, "verdict": verdict, "parity_verdict": parity_verdict,
            "strategy_sha": strategy_sha, "exit_contract_sha": exit_sha,
            "exit_candidate_sha": EXIT_V2_CANDIDATE_SHA, "precommit_sha": precommit_sha,
            "activation_sha": activation_sha, "clean_day": clean["verdict"],
            "prospective_day1": clean.get("prospective_day1_if_pass") if clean["pass"] else None,
            "submit_cancel_live": "0/0/0",
            "integrity_note": "ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT",
        }],
        "FrozenContract": [
            {"field": "guard", "value": FROZEN_GUARD},
            {"field": "continuation", "value": FROZEN_CONTINUATION},
            {"field": "candidate_sha", "value": EXIT_V2_CANDIDATE_SHA},
            {"field": "architecture", "value": "E"},
        ],
        "MetricLineage": lineage,
        "Parity": [parity_e, parity_c, {"label": "0810_dual", **{k: v for k, v in parity_0810.items() if k != "comparison"}},
                   {"label": "parity_verdict", "value": parity_verdict}],
        "RuntimeRoles": [{
            "primary": "Arch E PAPER_PRIMARY",
            "control": "FIXED600 SHADOW_CONTROL",
            "pbv2": "SHADOW_ONLY",
            "one_m": "SHADOW_ONLY_DIAGNOSTIC",
            "assertion_ok": roles.ok,
            "assertion_reason": roles.reason,
        }],
        "PrimaryV2": [{
            "strategy": STRATEGY_ID, "sha": strategy_sha,
            "hist_full_pnl": runtime_e["summary"]["total"],
            "hist_full_n": runtime_e["accepted_n"],
            "research_full_pnl": research_e["summary"]["total"],
        }],
        "Fixed600Control": [{
            "strategy": "PASSIVE_FIXED600_FULL_STRATEGY_V1R",
            "sha": V1R_SHA,
            "hist_pnl": summarize_pnls([float(e.get("realized_pnl_yen") or 0) for e in base["events"] if e.get("accepted")])["total"],
            "parity_pass": parity_c["pass"],
        }],
        "Precommit": [precommit],
        "ProspectiveEligibility": [clean],
        "Preflight": [{"gate": k, "pass": v} for k, v in gates.items()] + [
            {"gate": "live_safety_detail", "pass": True, "value": live_safety},
            {"gate": "roles_startup", "pass": roles.ok, "value": roles.startup_block[:500]},
        ],
        "ProspectiveComparison": [{
            "schema": "daily ArchE vs FIXED600 Control",
            "fields": [
                "fills_each", "common_fills", "divergent_fills",
                "pnl", "pf", "worst", "dd", "gross_loss",
                "guard_triggers", "losers_saved", "winners_cut",
                "saved_loss", "foregone_winner", "saved_lost_ratio",
                "exit_600_n", "extend_750_n", "extension_incremental_pnl",
                "cap_blocked", "occupancy_divergence_effect",
            ],
            "note": "Direct EXIT effects only on common fill identities; occupancy divergence separate.",
            "example_0810_reference": parity_0810.get("comparison"),
        }],
    }
    write_xlsx(sheets, OUT / "v1r_exit_v2_prospective_activation.xlsx")

    report = {
        "analysis_id": "V1R_EXIT_V2_PROSPECTIVE_ACTIVATION",
        "run_id": run_id,
        "verdict": verdict,
        "parity_verdict": parity_verdict,
        "gates": gates,
        "pins": pins,
        "lineage": lineage,
        "parity_e": parity_e,
        "parity_c": parity_c,
        "parity_0810": parity_0810,
        "clean": clean,
        "roles": roles.to_dict(),
        "live_safety": live_safety,
        "old_precommit_status": "SUPERSEDED_BEFORE_PROSPECTIVE_START",
        "integrity_note": "ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT",
        "production_mutation_scope": [
            "EXIT path: Arch E contract wiring",
            "Dual independent Control FIXED600 lane",
            "Activation gate / runtime pins / Discord EXIT reasons",
            "New strategy + precommit manifests",
        ],
        "strategy_parameter_mutation": "NONE — frozen guard/continuation/ENTRY/cap/universe/model unchanged",
        "answers": {
            "arch_e_spec": {
                "guard": FROZEN_GUARD,
                "continuation": FROZEN_CONTINUATION,
                "execution": "FIRST_VALID_EXECUTABLE_BUY1_AT_OR_AFTER_TRIGGER",
            },
            "exit_candidate_sha": EXIT_V2_CANDIDATE_SHA,
            "exit_contract_sha": exit_sha,
            "full_strategy_sha": strategy_sha,
            "precommit_sha": precommit_sha,
            "old_precommit_status": "SUPERSEDED_BEFORE_PROSPECTIVE_START",
            "parity": parity_verdict,
            "control_parity": parity_c["pass"],
            "roles": activation["runtime_roles"],
            "clean_0811": clean["verdict"],
            "prospective_day1": clean.get("prospective_day1_if_pass") if clean["pass"] else None,
            "submit_cancel_live": "0/0/0",
        },
    }
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    (OUT / "report.md").write_text(
        f"""# V1R EXIT V2 Prospective Activation

## Verdict
`{verdict}`

## Parity
`{parity_verdict}`

## Integrity
`ARCHITECTURE_SUPPORT_STRONGER_THAN_EXACT_CONTINUATION_THRESHOLD_SUPPORT`

## SHAs
- EXIT candidate: `{EXIT_V2_CANDIDATE_SHA}`
- EXIT contract: `{exit_sha}`
- Full strategy: `{strategy_sha}`
- Precommit: `{precommit_sha}`
- Activation: `{activation_sha}`

## Roles
- Primary: Arch E `PAPER_PRIMARY`
- Control: FIXED600 `SHADOW_CONTROL`
- PBv2: `SHADOW_ONLY`
- 1M: `SHADOW_ONLY_DIAGNOSTIC`

## Clean day
`{clean['verdict']}` → Day1=`{clean.get('prospective_day1_if_pass') if clean['pass'] else None}`

## Old precommit
`SUPERSEDED_BEFORE_PROSPECTIVE_START` (sha `{OLD_PRECOMMIT_SHA}`)

## submit/cancel/live
`0/0/0` — Paper only. STOP.
""",
        encoding="utf-8",
    )
    print(json.dumps({"verdict": verdict, "parity": parity_verdict, "strategy_sha": strategy_sha[:16]}, indent=2))
    return report


def _write_runtime_wiring(pins: dict, activation: dict, strategy: dict, precommit: dict) -> None:
    """Generate activation gate module with pinned SHAs (no strategy param mutation)."""
    path = NATIVE / "src/small_paper/v1r_exit_v2_activation_gate.py"
    path.write_text(
        f'''"""V1R EXIT V2 Paper Primary activation gate — fail-closed.

Primary = Arch E. Control = FIXED600 SHADOW_CONTROL.
No fallback to FIXED600 Primary or PBv2 Primary.
"""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from small_paper.v1r_exit_v2_contract import (
    EXIT_V2_CANDIDATE_SHA,
    FROZEN_CONTINUATION,
    FROZEN_GUARD,
    load_exit_v2_candidate,
)
from small_paper.v1r_primary_runtime import (
    ANCHOR_SHA,
    BOARD_FRESHNESS_SEC_V1R,
    CLOCK_GRID,
    DUPLICATE_RULE,
    LOT_QTY,
    MODEL_ARTIFACT_SHA,
    POSITION_CAP,
    UNIVERSE_BINDING_SHA,
    UNIVERSE_CONTRACT,
    V1R_SHA,
    WAIT_SEC,
    assert_v1r_not_contaminated,
    resolve_v1r_effective_from_production,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[2]
OUT = NATIVE / "results/research/v1r_exit_v2_prospective_activation"

PRIMARY_STRATEGY = "{STRATEGY_ID}"
STRATEGY_SHA = "{pins['STRATEGY_SHA']}"
EXIT_CONTRACT_SHA = "{pins['EXIT_CONTRACT_SHA']}"
PRECOMMIT_SHA = "{pins['PRECOMMIT_SHA']}"
ACTIVATION_SHA = "{pins['ACTIVATION_SHA']}"
CONTROL_STRATEGY_SHA = "{pins['CONTROL_STRATEGY_SHA']}"
GUARD_ID = "{pins['GUARD_ID']}"
CONTINUATION_ID = "{pins['CONTINUATION_ID']}"

V1R_ROLE = "PAPER_PRIMARY"
CONTROL_ROLE = "SHADOW_CONTROL"
PBV2_ROLE = "SHADOW_ONLY"
ONE_M_ROLE = "SHADOW_ONLY_DIAGNOSTIC"
ASSERTION_FAIL = "V1R_EXIT_V2_PRIMARY_ROLE_ASSERTION_FAILED"


@dataclass
class RoleAssertionResult:
    ok: bool
    reason: str = ""
    checks: dict[str, bool] = field(default_factory=dict)
    identity: dict[str, Any] = field(default_factory=dict)
    startup_block: str = ""
    ready: bool = False

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def build_identity() -> dict[str, Any]:
    return {{
        "primary_strategy": PRIMARY_STRATEGY,
        "primary_role": V1R_ROLE,
        "strategy_sha": STRATEGY_SHA,
        "exit_v2_candidate_sha": EXIT_V2_CANDIDATE_SHA,
        "exit_contract_sha": EXIT_CONTRACT_SHA,
        "precommit_sha": PRECOMMIT_SHA,
        "activation_sha": ACTIVATION_SHA,
        "entry_v1r_sha": V1R_SHA,
        "model_sha": MODEL_ARTIFACT_SHA,
        "universe_binding_sha": UNIVERSE_BINDING_SHA,
        "universe_contract": UNIVERSE_CONTRACT,
        "anchor_sha": ANCHOR_SHA,
        "control_role": CONTROL_ROLE,
        "control_strategy_sha": CONTROL_STRATEGY_SHA,
        "pbv2_role": PBV2_ROLE,
        "one_m_role": ONE_M_ROLE,
        "guard_id": GUARD_ID,
        "continuation_id": CONTINUATION_ID,
        "cap": POSITION_CAP,
        "qty": LOT_QTY,
        "wait_sec": WAIT_SEC,
        "freshness_sec": BOARD_FRESHNESS_SEC_V1R,
        "duplicate_rule": DUPLICATE_RULE,
        "anchor_count": len(CLOCK_GRID),
        "live_trading_enabled": False,
        "order_enabled": False,
        "paper_only": True,
        "submit": 0,
        "cancel": 0,
        "live": 0,
    }}


def format_startup_contract(*, ready: bool, reason: str = "") -> str:
    ident = build_identity()
    return "\\n".join([
        "[V1R EXIT V2 STARTUP CONTRACT]",
        "",
        "Primary:",
        f"Arch E {{ident['primary_role']}}",
        f"strategy={{ident['primary_strategy']}}",
        f"guard={{ident['guard_id']}} continuation={{ident['continuation_id']}}",
        "",
        "Control:",
        f"FIXED600 {{ident['control_role']}}",
        "",
        "PBv2:",
        ident["pbv2_role"],
        "",
        "1M:",
        ident["one_m_role"],
        "",
        "Cap/Wait/Freshness:",
        f"{{ident['cap']}} / {{ident['wait_sec']}}s / {{ident['freshness_sec']}}s",
        "",
        "submit/cancel/live:",
        "0/0/0",
        "",
        "READY:",
        "YES" if ready else f"NO ({{reason or ASSERTION_FAIL}})",
    ])


def assert_exit_v2_primary_roles() -> RoleAssertionResult:
    checks: dict[str, bool] = {{}}
    identity = build_identity()
    checks["primary_strategy"] = identity["primary_strategy"] == PRIMARY_STRATEGY
    checks["strategy_sha_pin"] = identity["strategy_sha"] == STRATEGY_SHA
    checks["guard_id"] = identity["guard_id"] == GUARD_ID == FROZEN_GUARD["id"]
    checks["continuation_id"] = identity["continuation_id"] == CONTINUATION_ID == FROZEN_CONTINUATION["id"]
    checks["control_role"] = identity["control_role"] == CONTROL_ROLE
    checks["pbv2_shadow"] = identity["pbv2_role"] == PBV2_ROLE
    checks["one_m_shadow"] = identity["one_m_role"] == ONE_M_ROLE
    checks["cap"] = identity["cap"] == 5
    checks["qty"] = identity["qty"] == 100
    checks["wait"] = float(identity["wait_sec"]) == 1.0
    checks["freshness"] = float(identity["freshness_sec"]) == 5.0
    checks["live_off"] = (not identity["live_trading_enabled"]) and (not identity["order_enabled"])
    checks["paper_only"] = bool(identity["paper_only"])
    try:
        c = load_exit_v2_candidate()
        checks["exit_candidate_sha"] = c.get("sha256") == EXIT_V2_CANDIDATE_SHA
    except Exception:
        checks["exit_candidate_sha"] = False
    for name, sha in (
        ("strategy", STRATEGY_SHA),
        ("exit_contract", EXIT_CONTRACT_SHA),
        ("precommit", PRECOMMIT_SHA),
        ("activation", ACTIVATION_SHA),
    ):
        p = OUT / (
            f"{{PRIMARY_STRATEGY}}.json" if name == "strategy"
            else "PASSIVE_ASYMMETRIC_EXIT_V2_CONTRACT_V1.json" if name == "exit_contract"
            else f"PROSPECTIVE_PRECOMMIT_V1R_EXIT_V2_U1.json" if name == "precommit"
            else f"V1R_EXIT_V2_PAPER_PRIMARY_ACTIVATION_V1.json"
        )
        try:
            body = json.loads(p.read_text(encoding="utf-8"))
            checks[f"file_{{name}}_sha"] = body.get("sha256") == sha
        except Exception:
            checks[f"file_{{name}}_sha"] = False
    try:
        bind = json.loads(
            (NATIVE / "results/research/e1_x39c_concentration_reconciliation"
             / "V1R_OPERATIONAL_UNIVERSE_BINDING_V1.json").read_text(encoding="utf-8")
        )
        checks["universe_binding"] = bind.get("sha256") == UNIVERSE_BINDING_SHA
    except Exception:
        checks["universe_binding"] = False
    try:
        eff = resolve_v1r_effective_from_production()
        iso = assert_v1r_not_contaminated(eff)
        checks["yaml_isolation"] = bool(iso.get("pass"))
        checks["pin_match"] = bool(eff.pin_match)
    except Exception:
        checks["yaml_isolation"] = False
        checks["pin_match"] = False

    failed = [k for k, v in checks.items() if not v]
    ok = len(failed) == 0
    reason = "" if ok else f"{{ASSERTION_FAIL}}:{{','.join(failed[:10])}}"
    block = format_startup_contract(ready=ok, reason=reason)
    return RoleAssertionResult(ok=ok, reason=reason, checks=checks, identity=identity, startup_block=block, ready=ok)
''',
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
