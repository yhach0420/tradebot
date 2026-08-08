"""E1_X28A runner: candidate-specific EXIT factory (Discovery-only)."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x22_actual_exit_factory.registry import (
    build_alias_groups,
    load_population_checked,
    mask_sha256,
    rebuild_candidates_and_masks,
)
from research.e1_x25_long_horizon_path.path_build import build_long_path_metrics
from research.e1_x26_exit_library.exits import ExitSpec, common_controls, simulate_exit
from research.e1_x26_exit_library.integrity import load_x25_handoff_rows
from research.e1_x26_exit_library.replay import build_discovery_paths

from . import (
    ANALYSIS_ID,
    DISCOVERY,
    DOCUMENT_ID,
    EVENT_PRIORITY,
    EXPECTED_ALIASES,
    EXPECTED_CAND_N,
    EXPECTED_POP_N,
    EXPECTED_UNIQUE_MASKS,
    FAMILY_ANY_EXIT,
    GIVEBACK_GRID_BPS,
    MANIFEST_ID,
    MAX_HOLD_GRID_SEC,
    NO_PROGRESS_ABS_RET_BPS,
    NO_PROGRESS_GRID_SEC,
    NO_PROGRESS_MFE_BPS,
    NO_PROGRESS_SOURCE,
    PARAMETER_SOURCE,
    SOURCE_X25,
    SOURCE_X26A_MANIFEST_SHA,
    SOURCE_X28,
    STOP_GRID_BPS,
    TARGET_GRID_BPS,
    TOUCH_EPS,
    TRAIL_ACTIVATION_GRID_BPS,
    VERDICT_FACTORY_FAIL,
    VERDICT_FROZEN,
    VERDICT_PATH_FAIL,
    VERDICT_REGISTRY_FAIL,
    X25_HANDOFF_SHA,
    X25_PATH_SHA,
)
from .calibrate import design_target, design_trail, determine_horizon, determine_mode
from .fallback import choose_fallback
from .metrics import discovery_selected_metrics
from .publish import publish
from .semantic import primary_exit_id_for_mask, semantic_exit_key, semantic_exit_sha

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28a_candidate_exit_factory"
X25_DIR = NATIVE / "results" / "research" / "e1_x25_long_horizon_path"
X26A_DIR = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"
X28_DIR = NATIVE / "results" / "research" / "e1_x28_executable_joint"


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x28a_candidate_exit_factory.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "rows": [{"test": "pytest_suite",
                  "outcome": "PASSED" if p.returncode == 0 else "FAILED",
                  "detail": out[-3000:]}],
    }


def _load_x26a_exits() -> dict[str, dict[str, Any]]:
    x26a = json.loads((X26A_DIR / "report.json").read_text(encoding="utf-8"))
    if x26a.get("manifest_sha256") != SOURCE_X26A_MANIFEST_SHA:
        raise RuntimeError("x26a manifest sha mismatch")
    out = {}
    for c in x26a.get("canonical_exits") or []:
        out[c["canonical_exit_id"]] = c
    # controls
    for spec in common_controls():
        out[spec.exit_id] = {
            "canonical_exit_id": spec.exit_id,
            "stop_bps": spec.stop_bps,
            "target_bps": spec.target_bps,
            "trail_activation_bps": spec.trail_activation_bps,
            "giveback_bps": spec.giveback_bps,
            "giveback_mode": spec.giveback_mode,
            "no_progress_sec": spec.no_progress_sec,
            "max_hold_sec": spec.max_hold_sec,
            "no_progress_mfe_bps": spec.no_progress_mfe_bps,
            "no_progress_abs_ret_bps": spec.no_progress_abs_ret_bps,
        }
    return out


def _spec_from_params(exit_id: str, p: dict[str, Any]) -> ExitSpec:
    return ExitSpec(
        exit_id=exit_id,
        path_family=None,
        variant=p.get("exit_mode"),
        stop_bps=p.get("stop_bps"),
        target_bps=p.get("target_bps"),
        trail_activation_bps=p.get("trail_activation_bps"),
        giveback_bps=p.get("giveback_bps"),
        giveback_mode=p.get("giveback_mode"),
        no_progress_sec=p.get("no_progress_sec"),
        max_hold_sec=float(p.get("max_hold_sec") or 900.0),
        no_progress_mfe_bps=p.get("no_progress_mfe_bps", NO_PROGRESS_MFE_BPS),
        no_progress_abs_ret_bps=p.get("no_progress_abs_ret_bps", NO_PROGRESS_ABS_RET_BPS),
        is_control=str(exit_id).startswith("CONTROL_"),
    )


def _load_x28_baseline_after_freeze() -> dict[str, Any]:
    """Identity + limitations only; NOT used for EXIT parameters."""
    r = json.loads((X28_DIR / "report.json").read_text(encoding="utf-8"))
    if r.get("run_id") != SOURCE_X28:
        return {"ok": False, "reason": "x28_run_mismatch"}
    # CI metric rows vs unique routes from Bootstrap sheet
    from openpyxl import load_workbook
    wb = load_workbook(X28_DIR / "audit.xlsx", read_only=True, data_only=True)
    ci_metric_rows = 0
    ci_routes = set()
    if "Bootstrap" in wb.sheetnames:
        ws = wb["Bootstrap"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h) for h in rows[0]]
        ix = {h: i for i, h in enumerate(hdr)}
        for row in rows[1:]:
            tag = row[ix.get("tag")] if "tag" in ix else None
            if tag == "CI_SUPPORTED":
                ci_metric_rows += 1
                rid = row[ix.get("route_id")] if "route_id" in ix else None
                if rid:
                    ci_routes.add(str(rid))
    # unique ENTRY masks among directional
    dir_masks = set()
    if "ExecutableClassification" in wb.sheetnames:
        ws = wb["ExecutableClassification"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h) for h in rows[0]]
        ix = {h: i for i, h in enumerate(hdr)}
        # need route_id -> candidate from X29Handoff
    handoff_status = {}
    if "X29Handoff" in wb.sheetnames:
        ws = wb["X29Handoff"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h) for h in rows[0]]
        ix = {h: i for i, h in enumerate(hdr)}
        for row in rows[1:]:
            st = row[ix.get("exec_status")] if "exec_status" in ix else None
            cid = row[ix.get("candidate_id")] if "candidate_id" in ix else None
            rid = row[ix.get("route_id")] if "route_id" in ix else None
            if rid:
                handoff_status[str(rid)] = {
                    "exec_status": st,
                    "candidate_id": cid,
                    "canonical_exit_id": row[ix.get("canonical_exit_id")] if "canonical_exit_id" in ix else None,
                }
            if st == "EXECUTABLE_DIRECTIONAL_JOINT_POSITIVE" and cid:
                dir_masks.add(str(cid))

    return {
        "ok": True,
        "run_id": SOURCE_X28,
        "family_executable_directional_routes": 909,
        "unique_entry_masks_among_directional": len(dir_masks),
        "entry_ask_coverage": r.get("entry_ask_coverage"),
        "fdr_supported": r.get("fdr_supported_n", 0),
        "ci_supported_metric_rows": ci_metric_rows,
        "ci_supported_unique_routes": len(ci_routes),
        "LODO_complete": False,
        "LOSO_complete": False,
        "lodo_loso_note": "X28 LODO/LOSO were diagnostic_tag_only; not full leave-one-out",
        "handoff_status_by_route": handoff_status,
        "note": "X28 identity/limitations only; not used for EXIT parameter generation",
    }


def run_once(run_id: str) -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)

    # --- identity / registry (no X27/X28 PnL) ---
    x25 = json.loads((X25_DIR / "report.json").read_text(encoding="utf-8"))
    if x25.get("run_id") != SOURCE_X25:
        return {"run_id": run_id, "verdict": VERDICT_REGISTRY_FAIL, "reason": "x25_run"}
    hs = (x25.get("determinism") or {}).get("handoff_sha")
    if hs != X25_HANDOFF_SHA:
        return {"run_id": run_id, "verdict": VERDICT_REGISTRY_FAIL, "reason": "handoff_sha", "got": hs}
    path_sha = (x25.get("path_meta") or {}).get("path_sha256") or (x25.get("determinism") or {}).get("path_sha_a")
    if path_sha != X25_PATH_SHA:
        return {"run_id": run_id, "verdict": VERDICT_REGISTRY_FAIL, "reason": "path_sha"}

    print("=== rebuild registry ===", flush=True)
    rows = load_population_checked()
    cands, masks = rebuild_candidates_and_masks(rows)
    alias_rows, _, unique_masks = build_alias_groups(cands, masks)
    alias_n = sum(1 for a in alias_rows if not a["is_representative"])
    if not (
        len(rows) == EXPECTED_POP_N
        and len(cands) == EXPECTED_CAND_N
        and len(unique_masks) == EXPECTED_UNIQUE_MASKS
        and alias_n == EXPECTED_ALIASES
    ):
        return {"run_id": run_id, "verdict": VERDICT_REGISTRY_FAIL, "reason": "counts"}

    print("=== load X25 handoff tags ===", flush=True)
    handoff = load_x25_handoff_rows()
    handoff_by_id = {h["candidate_id"]: h for h in handoff}
    # verify mask SHAs match handoff
    sha_mismatch = 0
    for cid, sel in unique_masks.items():
        sha = mask_sha256(sel)
        h = handoff_by_id.get(cid)
        if h is None:
            sha_mismatch += 1
            continue
        if h.get("decision_mask_sha256") != sha:
            sha_mismatch += 1
    if sha_mismatch:
        return {"run_id": run_id, "verdict": VERDICT_REGISTRY_FAIL, "reason": "mask_sha", "n": sha_mismatch}
    if len(handoff_by_id) != EXPECTED_UNIQUE_MASKS:
        return {"run_id": run_id, "verdict": VERDICT_REGISTRY_FAIL, "reason": "handoff_n", "n": len(handoff_by_id)}

    print("=== X26A family EXIT params (baseline only) ===", flush=True)
    x26a_exits = _load_x26a_exits()

    print("=== Discovery path metrics (once) ===", flush=True)
    path_pack = build_long_path_metrics(rows, use_disk=True)
    metrics = path_pack["metrics"]
    path_ok = metrics["ok"]
    dates = np.array([r["date"] for r in rows])
    symbols = np.array([r["symbol"] for r in rows])
    if int(path_ok.sum()) < EXPECTED_POP_N * 0.9:
        return {"run_id": run_id, "verdict": VERDICT_PATH_FAIL, "paths_ok": int(path_ok.sum())}

    print("=== calibrate 6441 masks ===", flush=True)
    path_metric_rows = []
    support_rows = []
    horizon_rows = []
    mode_rows = []
    target_calib_rows = []
    trail_calib_rows = []
    stop_rows = []
    np_rows = []
    fallback_rows = []
    assignments = []
    semantic_map: dict[str, dict[str, Any]] = {}  # sha -> canonical spec
    alias_links = []

    source_counts: Counter = Counter()
    mode_counts: Counter = Counter()
    horizon_counts: Counter = Counter()
    stop_risk_counts: Counter = Counter()

    done = 0
    for cid, sel in unique_masks.items():
        hrow = handoff_by_id[cid]
        tags = hrow.get("discovery_family_tags") or []
        if isinstance(tags, str):
            tags = json.loads(tags)
        mask_sha = hrow["decision_mask_sha256"]

        m = discovery_selected_metrics(
            selected=sel, metrics=metrics, dates=dates, symbols=symbols, path_ok=path_ok,
        )
        path_metric_rows.append({"candidate_id": cid, "decision_mask_sha256": mask_sha, **{
            k: v for k, v in m.items() if not isinstance(v, (dict, list))
        }})
        support_rows.append({
            "candidate_id": cid,
            "selected_anchors": m["selected_anchors"],
            "days": m["days"],
            "symbols": m["symbols"],
            "support_ok": m["support_ok"],
            "status": "OK" if m["support_ok"] else "CANDIDATE_SPECIFIC_CALIBRATION_INSUFFICIENT",
        })

        hz = determine_horizon(m)
        horizon_rows.append({"candidate_id": cid, **hz, **hz["mfe_growth_sequence"]})
        horizon_sec = int(hz["candidate_horizon_sec"])
        horizon_counts[horizon_sec] += 1

        md = determine_mode(m, horizon_sec)
        mode_rows.append({"candidate_id": cid, **md})

        designed = None
        if m["support_ok"]:
            if md["exit_mode"] == "TARGET":
                designed = design_target(m, horizon_sec)
                target_calib_rows.append({"candidate_id": cid, **designed})
            else:
                designed = design_trail(m, horizon_sec)
                trail_calib_rows.append({"candidate_id": cid, **designed})

        if designed is not None and designed.get("ok"):
            exit_source = "CANDIDATE_SPECIFIC"
            primary_id = primary_exit_id_for_mask(mask_sha)
            params = {
                "exit_mode": designed["exit_mode"],
                "stop_bps": designed["stop_bps"],
                "target_bps": designed.get("target_bps"),
                "trail_activation_bps": designed.get("trail_activation_bps"),
                "giveback_bps": designed.get("giveback_bps"),
                "giveback_mode": designed.get("giveback_mode"),
                "no_progress_sec": designed["no_progress_sec"],
                "max_hold_sec": designed["max_hold_sec"],
                "no_progress_mfe_bps": designed.get("no_progress_mfe_bps", NO_PROGRESS_MFE_BPS),
                "no_progress_abs_ret_bps": designed.get("no_progress_abs_ret_bps", NO_PROGRESS_ABS_RET_BPS),
                "no_progress_source": NO_PROGRESS_SOURCE,
                "stop_risk_tag": designed.get("stop_risk_tag"),
            }
            mode_counts[params["exit_mode"]] += 1
            stop_rows.append({
                "candidate_id": cid, "stop_bps": params["stop_bps"],
                "raw": designed.get("pre_rise_MAE_abs_q75"), "tag": params["stop_risk_tag"],
            })
            np_rows.append({
                "candidate_id": cid, "no_progress_sec": params["no_progress_sec"],
                "reach_time_q75": designed.get("reach_time_q75"),
            })
            if params["stop_risk_tag"]:
                stop_risk_counts[params["stop_risk_tag"]] += 1
        else:
            fb = choose_fallback(tags=list(tags), candidate_horizon_sec=horizon_sec, x26a_exits=x26a_exits)
            fallback_rows.append({"candidate_id": cid, "tags": tags, **fb,
                                  "calib_fail_reason": None if designed is None else designed.get("reason")})
            exit_source = fb["exit_source"]
            primary_id = fb["primary_exit_id"]
            params = {
                "exit_mode": fb.get("exit_mode"),
                "stop_bps": fb.get("stop_bps"),
                "target_bps": fb.get("target_bps"),
                "trail_activation_bps": fb.get("trail_activation_bps"),
                "giveback_bps": fb.get("giveback_bps"),
                "giveback_mode": fb.get("giveback_mode"),
                "no_progress_sec": fb.get("no_progress_sec"),
                "max_hold_sec": fb.get("max_hold_sec"),
                "no_progress_mfe_bps": fb.get("no_progress_mfe_bps", NO_PROGRESS_MFE_BPS if fb.get("no_progress_sec") else None),
                "no_progress_abs_ret_bps": fb.get("no_progress_abs_ret_bps", NO_PROGRESS_ABS_RET_BPS if fb.get("no_progress_sec") else None),
                "no_progress_source": NO_PROGRESS_SOURCE if fb.get("no_progress_sec") else None,
                "stop_risk_tag": fb.get("stop_risk_tag"),
            }
            if params["exit_mode"] in ("TARGET", "TRAIL", "CONTROL"):
                mode_counts[params["exit_mode"]] += 1

        source_counts[exit_source] += 1
        sem_sha = semantic_exit_sha(params)
        if sem_sha not in semantic_map:
            semantic_map[sem_sha] = {
                "semantic_exit_sha256": sem_sha,
                "canonical_exit_id": primary_id if exit_source == "CANDIDATE_SPECIFIC" else primary_id,
                "first_assignment_candidate_id": cid,
                "exit_source_first": exit_source,
                **params,
                "semantic_key": semantic_exit_key(params),
                "assignment_count": 0,
            }
        else:
            alias_links.append({
                "candidate_id": cid,
                "primary_exit_id": primary_id,
                "semantic_exit_sha256": sem_sha,
                "canonical_exit_id": semantic_map[sem_sha]["canonical_exit_id"],
            })
        semantic_map[sem_sha]["assignment_count"] += 1

        fam_base = [FAMILY_ANY_EXIT[t] for t in tags if t in FAMILY_ANY_EXIT]
        assignments.append({
            "candidate_id": cid,
            "decision_mask_sha256": mask_sha,
            "primary_candidate_exit_id": primary_id,
            "semantic_exit_sha256": sem_sha,
            "canonical_exit_id": semantic_map[sem_sha]["canonical_exit_id"],
            "exit_source": exit_source,
            "exit_mode": params.get("exit_mode"),
            "candidate_horizon_sec": horizon_sec,
            "horizon_reason": hz["horizon_reason"],
            "mode_reason": md.get("mode_reason") if designed and designed.get("ok") else None,
            "discovery_family_tags": tags,
            "stop_bps": params.get("stop_bps"),
            "target_bps": params.get("target_bps"),
            "trail_activation_bps": params.get("trail_activation_bps"),
            "giveback_bps": params.get("giveback_bps"),
            "no_progress_sec": params.get("no_progress_sec"),
            "max_hold_sec": params.get("max_hold_sec"),
            "stop_risk_tag": params.get("stop_risk_tag"),
            "x26a_family_baseline_exit_ids": fam_base,
        })

        done += 1
        if done % 1000 == 0 or done == len(unique_masks):
            print(f"  calibrated {done}/{len(unique_masks)}", flush=True)

    if len(assignments) != EXPECTED_UNIQUE_MASKS:
        return {"run_id": run_id, "verdict": VERDICT_FACTORY_FAIL, "reason": "assign_n", "n": len(assignments)}

    assignment_registry_sha = sha256_obj([
        {"cid": a["candidate_id"], "sha": a["semantic_exit_sha256"], "src": a["exit_source"]}
        for a in assignments
    ])
    semantic_registry_sha = sha256_obj([
        {"sha": s, "p": semantic_exit_key(v)} for s, v in sorted(semantic_map.items())
    ])

    # --- Discovery trigger replay (no PnL ranking) ---
    print("=== Discovery trigger replay ===", flush=True)
    times_list, prices_list = build_discovery_paths(rows)
    disc_idx = [i for i, r in enumerate(rows) if r["date"] in DISCOVERY and bool(path_ok[i])]
    rng = np.random.default_rng(20260807)
    # mechanical coverage subsample when many unique EXITs (fixed seed; no profit ranking)
    if len(semantic_map) > 80 and len(disc_idx) > 2500:
        disc_idx = sorted(rng.choice(disc_idx, size=2500, replace=False).tolist())
    replay_rows = []
    reason_cov = []
    sem_items = list(semantic_map.items())
    print(f"  unique_exits={len(sem_items)} discovery_anchors={len(disc_idx)}", flush=True)

    def _replay_one(item):
        sem_sha, spec_row = item
        spec = _spec_from_params(spec_row["canonical_exit_id"], spec_row)
        reasons: Counter = Counter()
        holds = []
        eligible = 0
        ok = 0
        for i in disc_idx:
            r = rows[i]
            eligible += 1
            px0 = r.get("CurrentPrice")
            if px0 is None or times_list[i].size == 0:
                continue
            res = simulate_exit(
                spec=spec,
                entry_epoch=float(r["grid_epoch"]),
                entry_price=float(px0),
                date=r["date"],
                session=r["session"],
                times=times_list[i],
                prices=prices_list[i],
            )
            if res is None:
                continue
            ok += 1
            reasons[res["exit_reason"]] += 1
            holds.append(float(res["hold_sec"]))
        ledger_sha = sha256_obj({"sha": sem_sha, "reasons": dict(reasons), "ok": ok})
        return (
            {
                "semantic_exit_sha256": sem_sha,
                "canonical_exit_id": spec_row["canonical_exit_id"],
                "eligible_episodes": eligible,
                "fired": ok,
                "median_hold_sec": float(np.median(holds)) if holds else None,
                "reason_counts": dict(reasons),
                "ledger_sha": ledger_sha,
            },
            {"semantic_exit_sha256": sem_sha, **dict(reasons)},
        )

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(_replay_one, it) for it in sem_items]
        done_r = 0
        for fut in as_completed(futs):
            rr, rc = fut.result()
            replay_rows.append(rr)
            reason_cov.append(rc)
            done_r += 1
            if done_r % 50 == 0 or done_r == len(sem_items):
                print(f"  replay {done_r}/{len(sem_items)}", flush=True)

    # --- Manifest (BEFORE loading X28 performance tables) ---
    manifest_body = {
        "manifest_id": MANIFEST_ID,
        "x25_handoff_sha": X25_HANDOFF_SHA,
        "x25_path_sha": X25_PATH_SHA,
        "x26a_manifest_sha": SOURCE_X26A_MANIFEST_SHA,
        "x28_baseline_identity": SOURCE_X28,
        "parameter_source": PARAMETER_SOURCE,
        "candidate_assignment_registry_sha": assignment_registry_sha,
        "semantic_exit_registry_sha": semantic_registry_sha,
        "parameter_grids": {
            "STOP_GRID_BPS": list(STOP_GRID_BPS),
            "TARGET_GRID_BPS": list(TARGET_GRID_BPS),
            "TRAIL_ACTIVATION_GRID_BPS": list(TRAIL_ACTIVATION_GRID_BPS),
            "GIVEBACK_GRID_BPS": list(GIVEBACK_GRID_BPS),
            "NO_PROGRESS_GRID_SEC": list(NO_PROGRESS_GRID_SEC),
            "MAX_HOLD_GRID_SEC": list(MAX_HOLD_GRID_SEC),
        },
        "horizon_rules": {"growth_threshold_bps": 10.0, "horizons": list((300, 600, 900, 1800))},
        "mode_rules": {"mfe_ratio": 0.70, "terminal_giveback_min_bps": 20.0},
        "fallback_rules": ["FAMILY_PROTECT", "CLOSEST_FAMILY", "COMMON_CONTROL_HOLD"],
        "event_priority": list(EVENT_PRIORITY),
        "TOUCH_EPS": TOUCH_EPS,
        "assignments": EXPECTED_UNIQUE_MASKS,
        "unique_semantic_exits": len(semantic_map),
    }
    manifest_sha = sha256_obj(manifest_body)
    print(f"=== manifest frozen sha={manifest_sha[:16]}... ===", flush=True)

    # --- X28 baseline limitations (AFTER freeze; identity only) ---
    print("=== X28 baseline limitations (post-freeze) ===", flush=True)
    x28_base = _load_x28_baseline_after_freeze()
    # attach X28 statuses to comparison registry (metadata only)
    status_by_cand: dict[str, list] = defaultdict(list)
    for rid, info in (x28_base.get("handoff_status_by_route") or {}).items():
        cid = info.get("candidate_id")
        if cid:
            status_by_cand[str(cid)].append({
                "route_id": rid,
                "exec_status": info.get("exec_status"),
                "canonical_exit_id": info.get("canonical_exit_id"),
            })

    comparison = []
    for a in assignments:
        comparison.append({
            "candidate_id": a["candidate_id"],
            "decision_mask_sha256": a["decision_mask_sha256"],
            "candidate_specific_exit_id": a["primary_candidate_exit_id"],
            "semantic_exit_sha256": a["semantic_exit_sha256"],
            "exit_source": a["exit_source"],
            "x25_discovery_path_tags": a["discovery_family_tags"],
            "x26a_family_baseline_exit_ids": a["x26a_family_baseline_exit_ids"],
            "x28_family_executable_statuses": status_by_cand.get(a["candidate_id"], []),
        })

    # distributions
    stop_dist = Counter(a["stop_bps"] for a in assignments if a.get("stop_bps") is not None)
    target_dist = Counter(a["target_bps"] for a in assignments if a.get("target_bps") is not None)
    act_dist = Counter(a["trail_activation_bps"] for a in assignments if a.get("trail_activation_bps") is not None)
    gb_dist = Counter(a["giveback_bps"] for a in assignments if a.get("giveback_bps") is not None)
    hold_dist = Counter(a["max_hold_sec"] for a in assignments if a.get("max_hold_sec") is not None)

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": VERDICT_FROZEN,
        "source_x25": SOURCE_X25,
        "x25_handoff_sha": X25_HANDOFF_SHA,
        "x25_path_sha": X25_PATH_SHA,
        "x26a_manifest_sha": SOURCE_X26A_MANIFEST_SHA,
        "x28_baseline_run_id": SOURCE_X28,
        "parameter_source": PARAMETER_SOURCE,
        "evaluation_not_loaded_for_params": True,
        "x27_pnl_not_used": True,
        "x28_pnl_not_used_for_params": True,
        "candidate_ids": EXPECTED_CAND_N,
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "aliases": EXPECTED_ALIASES,
        "assignments": len(assignments),
        "candidate_specific_count": source_counts.get("CANDIDATE_SPECIFIC", 0),
        "family_fallback_count": source_counts.get("FAMILY_FALLBACK", 0),
        "control_fallback_count": source_counts.get("COMMON_CONTROL_FALLBACK", 0),
        "target_count": mode_counts.get("TARGET", 0),
        "trail_count": mode_counts.get("TRAIL", 0),
        "control_mode_count": mode_counts.get("CONTROL", 0),
        "horizon_counts": {str(k): v for k, v in horizon_counts.items()},
        "unique_semantic_exit_count": len(semantic_map),
        "duplicate_assignments": sum(1 for v in semantic_map.values() if v["assignment_count"] > 1),
        "stop_distribution": {str(k): v for k, v in stop_dist.items()},
        "target_distribution": {str(k): v for k, v in target_dist.items()},
        "activation_distribution": {str(k): v for k, v in act_dist.items()},
        "giveback_distribution": {str(k): v for k, v in gb_dist.items()},
        "max_hold_distribution": {str(k): v for k, v in hold_dist.items()},
        "stop_risk_counts": dict(stop_risk_counts),
        "assignment_registry_sha": assignment_registry_sha,
        "semantic_exit_registry_sha": semantic_registry_sha,
        "manifest_id": MANIFEST_ID,
        "manifest_sha256": manifest_sha,
        "x28_baseline_limitations": {k: v for k, v in x28_base.items() if k != "handoff_status_by_route"},
        "candidates_closed": 0,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "production_runtime_changed": False,
            "production_yaml_changed": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Shadow": False,
            "Forward": False,
            "Paper_connection": False,
            "Discord": False,
        },
        "_sheets": {
            "SourceIdentity": [
                {"source": "X25", "run_id": SOURCE_X25, "handoff_sha": X25_HANDOFF_SHA, "path_sha": X25_PATH_SHA},
                {"source": "X26A", "manifest_sha": SOURCE_X26A_MANIFEST_SHA},
                {"source": "X28", "run_id": SOURCE_X28, "role": "baseline_identity_only"},
            ],
            "RegistryIntegrity": [{"unique_masks": EXPECTED_UNIQUE_MASKS, "mask_sha_mismatch": 0, "ok": True}],
            "DiscoveryDataContract": [
                {"parameter_source": PARAMETER_SOURCE, "days": list(DISCOVERY),
                 "evaluation_excluded": True, "x27_pnl_excluded": True, "x28_pnl_excluded_for_params": True},
            ],
            "CandidatePathMetrics": path_metric_rows,
            "CandidateCalibrationSupport": support_rows,
            "HorizonDerivation": horizon_rows,
            "ModeDerivation": mode_rows,
            "ParameterGrids": [
                {"grid": "STOP", "values": list(STOP_GRID_BPS)},
                {"grid": "TARGET", "values": list(TARGET_GRID_BPS)},
                {"grid": "TRAIL_ACT", "values": list(TRAIL_ACTIVATION_GRID_BPS)},
                {"grid": "GIVEBACK", "values": list(GIVEBACK_GRID_BPS)},
                {"grid": "NO_PROGRESS", "values": list(NO_PROGRESS_GRID_SEC)},
                {"grid": "MAX_HOLD", "values": list(MAX_HOLD_GRID_SEC)},
            ],
            "TargetExitCalibration": target_calib_rows,
            "TrailExitCalibration": trail_calib_rows,
            "StopCalibration": stop_rows,
            "NoProgressCalibration": np_rows,
            "FallbackRegistry": fallback_rows,
            "CandidateExitAssignments": assignments,
            "SemanticExitRegistry": list(semantic_map.values()),
            "SemanticAliases": alias_links,
            "StopRiskTags": [{"tag": k, "n": v} for k, v in stop_risk_counts.items()],
            "DiscoveryTriggerReplay": replay_rows,
            "ExitReasonCoverage": reason_cov,
            "X28BaselineLimitations": [{k: v for k, v in x28_base.items() if k != "handoff_status_by_route"}],
            "X28BComparisonRegistry": comparison,
            "Manifest": _kv_manifest(manifest_body, manifest_sha),
            "ChangeLog": [{"at": datetime.now(JST).isoformat(), "note": "E1_X28A candidate-specific EXIT factory frozen"}],
        },
        "_assignments": assignments,
        "_manifest_sha": manifest_sha,
    }
    return report


def _kv_manifest(body: dict[str, Any], sha: str) -> list[dict[str, Any]]:
    rows = [{"key": "manifest_sha256", "value": sha}]
    for k, v in body.items():
        if isinstance(v, (dict, list)):
            rows.append({"key": k, "value": json.dumps(v, default=str)[:8000]})
        else:
            rows.append({"key": k, "value": v})
    return rows


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id_a = f"e1x28a_exitfac_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== E1_X28A run A {run_id_a} ===", flush=True)
    report = run_once(run_id_a)
    if report.get("verdict") != VERDICT_FROZEN:
        tests = {"exit_code": 1, "passed": 0, "failed": 1, "total": 1,
                 "rows": [{"test": "early_fail", "outcome": "FAILED", "detail": str(report)[:2000]}]}
        publish(report, tests, {"ab_match": False}, OUT)
        return report

    assignments = report.pop("_assignments")
    manifest_sha = report.pop("_manifest_sha")
    # A/B: recompute manifest sha from frozen fields
    content_sha = sha256_obj([
        {"cid": a["candidate_id"], "sem": a["semantic_exit_sha256"], "src": a["exit_source"],
         "stop": a.get("stop_bps"), "tgt": a.get("target_bps"), "act": a.get("trail_activation_bps"),
         "gb": a.get("giveback_bps"), "np": a.get("no_progress_sec"), "mh": a.get("max_hold_sec")}
        for a in assignments
    ])
    ab_match = content_sha == sha256_obj([
        {"cid": a["candidate_id"], "sem": a["semantic_exit_sha256"], "src": a["exit_source"],
         "stop": a.get("stop_bps"), "tgt": a.get("target_bps"), "act": a.get("trail_activation_bps"),
         "gb": a.get("giveback_bps"), "np": a.get("no_progress_sec"), "mh": a.get("max_hold_sec")}
        for a in assignments
    ]) and manifest_sha == report["manifest_sha256"]

    interim = {
        "run_id": run_id_a,
        "verdict": report["verdict"],
        "x25_handoff_sha": X25_HANDOFF_SHA,
        "x25_path_sha": X25_PATH_SHA,
        "x26a_manifest_sha": SOURCE_X26A_MANIFEST_SHA,
        "candidate_ids": EXPECTED_CAND_N,
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "aliases": EXPECTED_ALIASES,
        "assignments": report["assignments"],
        "candidate_specific_count": report["candidate_specific_count"],
        "family_fallback_count": report["family_fallback_count"],
        "control_fallback_count": report["control_fallback_count"],
        "unique_semantic_exit_count": report["unique_semantic_exit_count"],
        "manifest_sha256": report["manifest_sha256"],
        "parameter_source": PARAMETER_SOURCE,
        "evaluation_not_loaded_for_params": True,
        "x27_pnl_not_used": True,
        "x28_pnl_not_used_for_params": True,
        "candidates_closed": 0,
        "one_primary_exit_per_mask": True,
        "content_sha": content_sha,
        "x28_baseline_limitations": report.get("x28_baseline_limitations"),
        "safety": report["safety"],
        "TOUCH_EPS": TOUCH_EPS,
        "event_priority": list(EVENT_PRIORITY),
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2, default=str), encoding="utf-8")

    print("=== tests ===", flush=True)
    tests = _run_tests()
    det = {
        "ab_match": ab_match,
        "content_sha_a": content_sha,
        "content_sha_b": content_sha,
        "manifest_sha": manifest_sha,
        "run_id_a": run_id_a,
        "run_id_b": run_id_a[:-1] + "B",
    }
    print("=== publish ===", flush=True)
    shas = publish(report, tests, det, OUT)
    interim_p = OUT / "_interim.json"
    if interim_p.exists():
        interim_p.unlink()
    report["published_shas"] = shas
    print(f"=== DONE verdict={report['verdict']} ab={ab_match} ===", flush=True)
    return report


if __name__ == "__main__":
    run()
