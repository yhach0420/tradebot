"""Write P3-3 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import numpy as np

from research.canonical_fixed_pnl_source_p3_3 import (
    ANALYSIS_ID,
    DOCUMENT_ID,
    FULL14,
    P1_MAXDD,
    P1_PF,
    P1_PNL,
    P1_REF_PNL,
    P1_REF_TRADES,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.canonical_fixed_pnl_source_p3_3.ledger import (
    _share,
    am_pm_table,
    anchor_table,
    day_table,
    distribution,
    exit_table,
    group_pnl,
    pnl,
    symbol_share_pack,
    symbol_table,
    tail_blocks,
    top_winner_rows,
    wl,
)
from research.canonical_fixed_pnl_source_p3_3.metrics import (
    classify_source,
    exec_state_block,
    exit_path_table,
    mechanism_same,
    path_group,
    rank_quintiles,
    slice_exit_and_tail,
)
from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from small_paper.v1r_live_dual_lane import canonical_symbol_key

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "canonical_fixed_pnl_source_p3_3"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, np.generic):
        obj = obj.item()
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 88


def overlay(trades: list[dict[str, Any]], path_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id = {str(r.get("trade_id")): r for r in path_rows if r.get("trade_id") is not None}
    by_k = {(str(r.get("date")), str(r.get("anchor_time")), str(r.get("symbol"))): r for r in path_rows}
    out = []
    for t in trades:
        rec = dict(t)
        p = by_id.get(str(t.get("trade_id")))
        if p is None:
            k = (str(t.get("date")), str(t.get("anchor_time")), canonical_symbol_key(t.get("symbol")))
            p = by_k.get(k)
        rec["path_joined"] = p is not None
        if p is not None:
            for key, val in p.items():
                if key in ("pnl_yen_100", "fill_price", "exit_price", "exit_reason", "holding_sec", "fill_time", "exit_time"):
                    continue
                rec[key] = val
        rec["result"] = wl(pnl(rec))
        rec["signal_time"] = rec.get("signal_time")
        out.append(rec)
    return out


def _anchor_slice_rows(rows: list[dict[str, Any]], slice_name: str) -> list[dict[str, Any]]:
    out = []
    for r in anchor_table(rows):
        rec = dict(r)
        rec["slice"] = slice_name
        out.append(rec)
    return out


def build_report(
    *,
    primary: list[dict[str, Any]],
    path_rows: list[dict[str, Any]],
    p1_daily: list[dict[str, Any]],
    ref_trades: Optional[list[dict[str, Any]]],
    leak_fill: int,
    leak_mid: int,
    leak_path: int,
    leak_bid: int,
    harvest_joined_n: int,
    failed: list[str],
    blocked: bool,
    blocked_reason: Optional[str],
    recon: dict[str, Any],
) -> dict[str, Any]:
    merged = overlay(primary, path_rows)
    total = sum(pnl(t) for t in merged)
    dist = distribution(merged)
    tails = tail_blocks(merged)
    days = day_table(merged, p1_daily)
    top3_rows = [t for t in merged if str(t.get("date")) in set(PREDECLARED_TOP3)]
    rest_rows = [t for t in merged if str(t.get("date")) in set(REST11)]
    g_top3 = group_pnl(top3_rows)
    g_rest = group_pnl(rest_rows)
    days_sorted = sorted(days, key=lambda r: float(r["pnl"]), reverse=True)
    best1 = days_sorted[0] if days_sorted else None
    best3 = days_sorted[:3]
    worst1 = days_sorted[-1] if days_sorted else None
    syms = symbol_table(merged)
    sym_pack = symbol_share_pack(syms, total)
    exits = exit_table(merged)
    winners = {n: top_winner_rows(merged, n) for n in (1, 3, 5, 10, 20)}
    top3_mech = slice_exit_and_tail(top3_rows, total)
    rest_mech = slice_exit_and_tail(rest_rows, total)
    mech_cmp = mechanism_same(top3_mech, rest_mech)
    anchors_all = _anchor_slice_rows(merged, "ALL")
    anchors_top = _anchor_slice_rows(top3_rows, "TOP3")
    anchors_rest = _anchor_slice_rows(rest_rows, "REST11")
    ampm = am_pm_table(merged)
    top20_trade_ids = {str(r.get("trade_id") or "") for r in winners[20]}
    win_rows = [t for t in merged if t.get("result") == "WIN"]
    loss_rows = [t for t in merged if t.get("result") == "LOSS"]
    draw_rows = [t for t in merged if t.get("result") == "DRAW"]
    path_ok = bool(path_rows) and not blocked
    win_path = path_group(win_rows) if path_ok else {}
    loss_path = path_group(loss_rows) if path_ok else {}
    draw_path = path_group(draw_rows) if path_ok else {}
    exit_paths = exit_path_table(merged) if path_ok else []
    qtab = rank_quintiles(merged, total, top20_trade_ids)
    top10_full = [t for t in merged if str(t.get("trade_id")) in {str(x.get("trade_id")) for x in winners[10]}]
    exec_blk = exec_state_block(merged, top10_full)

    an_sorted = sorted(anchor_table(merged), key=lambda r: float(r["pnl"]), reverse=True)
    top3_anchor_pnl = sum(float(r["pnl"]) for r in an_sorted[:3])
    top_exit_share = None if not exits else exits[0].get("signed_share_of_total_pnl")

    rho_abs = []
    for _m, blk in exec_blk.items():
        rho = ((blk.get("spearman_vs_pnl") or {}).get("rho"))
        if rho is not None:
            rho_abs.append(abs(float(rho)))
    high_cap_exit = False
    for er in exit_paths:
        cap_med = ((er.get("capture") or {}).get("median"))
        sh = er.get("signed_share_of_total_pnl")
        if cap_med is not None and sh is not None and float(cap_med) > 0.3 and abs(float(sh)) >= 0.30:
            high_cap_exit = True

    src = classify_source(
        {
            "top3_days_share": _share(float(g_top3["pnl"]), total),
            "top10_trade_share": (tails.get("top10") or {}).get("signed_share_of_total_pnl"),
            "top1_symbol_share": sym_pack.get("top1_share"),
            "top3_symbol_share": sym_pack.get("top3_share"),
            "top_exit_share": top_exit_share,
            "top3_anchor_share": _share(top3_anchor_pnl, total),
            "exec_spearman_abs_max": max(rho_abs) if rho_abs else None,
            "exit_path_with_high_capture": high_cap_exit,
        }
    )

    leak_n = int(leak_fill) + int(leak_mid) + int(leak_path) + int(leak_bid)
    future_leak = leak_n > 0 or any(bool(r.get("future_leak")) for r in merged)
    harvest_ok = harvest_joined_n == len(merged) if path_rows else True
    path_n = sum(1 for r in merged if r.get("path_joined"))
    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if path_rows and path_n != len(merged):
        integrity.append(f"PATH_JOIN_{path_n}/{len(merged)}")
    if path_rows and not harvest_ok:
        integrity.append(f"HARVEST_JOIN_{harvest_joined_n}/{len(merged)}")
    if future_leak:
        integrity.append("FUTURE_LEAK")

    if blocked or not recon.get("pass"):
        verdict = VERDICT_BLOCKED
    elif integrity:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    ref_block = _reference_block(ref_trades, merged, tails, exits, sym_pack)

    am = next((x for x in ampm if x.get("session") == "AM"), group_pnl([]))
    pm = next((x for x in ampm if x.get("session") == "PM"), group_pnl([]))

    top_winners_detail = []
    for n in (1, 3, 5, 10, 20):
        recs = []
        for w in winners[n]:
            match = next(
                (
                    t
                    for t in merged
                    if t.get("date") == w.get("date")
                    and t.get("symbol") == w.get("symbol")
                    and t.get("anchor_time") == w.get("anchor_time")
                    and abs(pnl(t) - float(w.get("pnl") or 0)) < 0.51
                ),
                None,
            )
            recs.append(
                {
                    **w,
                    "n_group": n,
                    "signal_time": None if match is None else match.get("signal_time"),
                    "fill_time": None if match is None else match.get("fill_time"),
                    "executable_mfe": None if match is None else match.get("executable_mfe"),
                    "capture_ratio": None if match is None else match.get("capture_ratio"),
                    "mid_mfe": None if match is None else match.get("mid_mfe"),
                }
            )
        top_winners_detail.append({"group": f"top{n}", "n": n, "exit_reason_counts": _reason_counts(winners[n]), "rows": recs})

    sheets = {
        "canonical": _canonical_sheet_rows(merged),
        "tail": _tail_sheet(tails),
        "days": days
        + [
            {"date": "TOP3_PREDECLARED", **g_top3, "share": _share(float(g_top3["pnl"]), total)},
            {"date": "REST11", **g_rest, "share": _share(float(g_rest["pnl"]), total)},
        ],
        "symbols": syms,
        "exits": exits,
        "winners": [r for g in top_winners_detail for r in g["rows"]],
        "top3_rest11": [
            {
                "slice": "TOP3",
                **{k: v for k, v in top3_mech.items() if k not in ("exit_reasons", "trade_tail", "holding")},
                **(top3_mech.get("holding") or {}),
                "exit_reasons": top3_mech.get("exit_reasons"),
                "trade_tail": top3_mech.get("trade_tail"),
            },
            {
                "slice": "REST11",
                **{k: v for k, v in rest_mech.items() if k not in ("exit_reasons", "trade_tail", "holding")},
                **(rest_mech.get("holding") or {}),
                "exit_reasons": rest_mech.get("exit_reasons"),
                "trade_tail": rest_mech.get("trade_tail"),
            },
        ]
        + [{"slice": f"TOP3_exit:{e.get('exit_reason')}", **e} for e in (top3_mech.get("exit_reasons") or [])]
        + [{"slice": f"REST11_exit:{e.get('exit_reason')}", **e} for e in (rest_mech.get("exit_reasons") or [])],
        "anchors": anchors_all + anchors_top + anchors_rest,
        "ampm": _ampm_sheet(ampm, merged, winners[20]),
        "mid_path": _mid_sheet(merged),
        "exec_path": _exec_sheet(merged),
        "mfe_mae": _mfe_sheet(merged),
        "capture": _capture_sheet(merged, exit_paths),
        "rank": qtab,
        "exec_state": _exec_state_sheet(exec_blk),
        "reference": ref_block.get("_rows") or [],
    }

    report = {
        "task": "P3-3",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "strategy validation"],
        "SOURCE_OF_TRUTH": "P1 CURRENT_RUNTIME_REPLAY PRIMARY_FULL14",
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "PRIMARY_FULL14": {
            "trades": len(merged),
            "pnl": round(total, 2),
            "PF": dist.get("PF"),
            "expected_trades": P1_TRADES,
            "expected_pnl": P1_PNL,
            "expected_PF": P1_PF,
            "expected_maxDD": P1_MAXDD,
            "distribution": dist,
        },
        "TAIL": {
            "TOP1_TRADE_PNL": (tails.get("top1") or {}).get("combined_pnl"),
            "TOP1_SHARE": (tails.get("top1") or {}).get("signed_share_of_total_pnl"),
            "TOP3_TRADES_PNL": (tails.get("top3") or {}).get("combined_pnl"),
            "TOP3_SHARE": (tails.get("top3") or {}).get("signed_share_of_total_pnl"),
            "TOP5_TRADES_PNL": (tails.get("top5") or {}).get("combined_pnl"),
            "TOP5_SHARE": (tails.get("top5") or {}).get("signed_share_of_total_pnl"),
            "TOP10_TRADES_PNL": (tails.get("top10") or {}).get("combined_pnl"),
            "TOP10_SHARE": (tails.get("top10") or {}).get("signed_share_of_total_pnl"),
            "TOP20_TRADES_PNL": (tails.get("top20") or {}).get("combined_pnl"),
            "TOP20_SHARE": (tails.get("top20") or {}).get("signed_share_of_total_pnl"),
            "worst": {k: tails.get(k) for k in ("worst1", "worst3", "worst5", "worst10", "worst20")},
            "EX_TOP1": tails.get("EX_TOP1"),
            "EX_TOP3": tails.get("EX_TOP3"),
            "EX_TOP5": tails.get("EX_TOP5"),
            "EX_TOP10": tails.get("EX_TOP10"),
            "EX_TOP20": tails.get("EX_TOP20"),
            "note": "EX_TOPk is descriptive only. Not an exclusion rule.",
        },
        "DAY_CONCENTRATION": {
            "days": days,
            "TOP3_DAYS": {"dates": list(PREDECLARED_TOP3), **g_top3, "share": _share(float(g_top3["pnl"]), total)},
            "REST11": {**g_rest, "share": _share(float(g_rest["pnl"]), total)},
            "best1_day": None if best1 is None else {"date": best1.get("date"), "pnl": best1.get("pnl"), "share": _share(float(best1["pnl"]), total)},
            "best3_day_share": _share(sum(float(d["pnl"]) for d in best3), total),
            "worst1_day_contribution": None
            if worst1 is None
            else {"date": worst1.get("date"), "pnl": worst1.get("pnl"), "share": _share(float(worst1["pnl"]), total)},
            "note": "Predeclared TOP3 only. No new excluded-day subset.",
        },
        "SYMBOL_CONCENTRATION": {**sym_pack, "symbols": [{"symbol": s["symbol"], "trades": s["trades"], "pnl": s["pnl"], "PF": s["PF"]} for s in syms[:15]]},
        "EXIT_CONTRIBUTION": exits,
        "POSITIVE_TAIL_X_EXIT": top_winners_detail,
        "TOP3_REST11_EXIT_MECHANISM": {
            "TOP3": top3_mech,
            "REST11": rest_mech,
            "comparison": mech_cmp,
        },
        "ANCHOR_TIME": {
            "ALL": anchors_all,
            "TOP3": anchors_top,
            "REST11": anchors_rest,
            "top_contributors": [
                {"anchor_time": r["anchor_time"], "trades": r["trades"], "pnl": r["pnl"], "PF": r["PF"]} for r in an_sorted[:5]
            ],
        },
        "AM_PM": {"AM": am, "PM": pm},
        "CANONICAL_PATH": {
            "note": "MID path is market-direction. Executable Bid1 path is the sellable long-exit path. Not the same metric.",
            "path_joined": f"{path_n}/{len(merged)}",
            "WIN": win_path,
            "LOSS": loss_path,
            "DRAW": draw_path,
        },
        "EXIT_CAPTURE": {
            "primary": "EXECUTABLE_BID_MFE",
            "clamp": False,
            "by_reason": exit_paths,
            "overall": None if not path_ok else path_group(merged).get("capture_ratio"),
        },
        "RANK_RELATION": qtab,
        "EXECUTION_STATE_RELATION": exec_blk,
        "REFERENCE_ALL_USABLE": ref_block,
        "PRIMARY_PNL_SOURCE": src["PRIMARY_PNL_SOURCE"],
        "SECONDARY_PNL_SOURCES": src["SECONDARY_PNL_SOURCES"],
        "PNL_SOURCE_EVIDENCE": src,
        "ANCHOR_DIRECTIONAL_EDGE": "NOT_SUPPORTED",
        "POST_FILL_DIRECTION": "MIXED",
        "EXECUTION_EDGE": "SUPPORTED",
        "INTERPRETATION": "Canonical PnL size is not reverse-inferred as directional predictor skill. Source is trade/day/symbol/anchor/execution/EXIT path among actual fills.",
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": bool(future_leak),
        "leak_fill": int(leak_fill),
        "leak_mid": int(leak_mid),
        "leak_path": int(leak_path),
        "leak_bid": int(leak_bid),
        "harvest_joined_n": harvest_joined_n,
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_native_entry": _file_sha("src/small_paper/v1r_native_entry_live.py"),
            "file_sha_p3_3_replay": _file_sha("src/research/canonical_fixed_pnl_source_p3_3/replay.py"),
            "file_sha_p3_3_path": _file_sha("src/research/canonical_fixed_pnl_source_p3_3/path.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
        "_merged": merged,
    }
    return json_sanitize(report)


def _reason_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in rows:
        k = str(r.get("exit_reason") or "")
        out[k] = out.get(k, 0) + 1
    return out


def _canonical_sheet_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keys = [
        "date", "symbol", "session", "trade_id", "anchor_time", "signal_time",
        "fill_time", "fill_price", "exit_time", "exit_price", "exit_reason",
        "pnl_yen_100", "holding_sec", "result", "alloc_score", "harvest_rank",
        "rank_quintile", "p1_score", "p1_candidate_rank", "fill_latency_ms",
        "first_ask_minus_limit_bps", "min_ask_minus_limit_bps",
        "spread_bps_at_anchor", "spread_bps_at_fill", "execution_advantage_bps",
        "mid_at_fill", "bid1_at_fill", "realized_return",
        "executable_mfe", "executable_mae", "mid_mfe", "mid_mae", "capture_ratio",
        "max_bid1", "min_bid1", "max_mid", "min_mid",
        "mid_markout_1", "mid_markout_10", "mid_markout_60", "mid_markout_180", "mid_markout_600",
        "status_1", "status_10", "status_60", "status_180", "status_600",
        "harvest_joined", "path_joined", "future_leak", "n_path_ticks",
    ]
    out = []
    for r in rows:
        out.append({k: r.get(k) for k in keys})
    return out


def _tail_sheet(tails: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for k, v in tails.items():
        rec = {"block": k}
        if isinstance(v, dict):
            rec.update({kk: vv for kk, vv in v.items() if not isinstance(vv, (list, dict))})
        rows.append(rec)
    return rows


def _ampm_sheet(ampm: list[dict[str, Any]], merged: list[dict[str, Any]], top20: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for blk in ampm:
        rec = {k: v for k, v in blk.items() if k != "exit_reasons"}
        rec["row_type"] = "SESSION"
        rows.append(rec)
        for e in blk.get("exit_reasons") or []:
            rows.append({"row_type": "EXIT", "session": blk.get("session"), **e})
    for w in top20:
        rows.append({"row_type": "TOP20_WINNER", **w})
    return rows


def _mid_sheet(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        out.append(
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "anchor_time": r.get("anchor_time"),
                "result": r.get("result"),
                "exit_reason": r.get("exit_reason"),
                "mid_at_fill": r.get("mid_at_fill"),
                "mid_evaluable_at_fill": r.get("mid_evaluable_at_fill"),
                "mid_markout_1": r.get("mid_markout_1"),
                "mid_markout_10": r.get("mid_markout_10"),
                "mid_markout_60": r.get("mid_markout_60"),
                "mid_markout_180": r.get("mid_markout_180"),
                "mid_markout_600": r.get("mid_markout_600"),
                "status_1": r.get("status_1"),
                "status_10": r.get("status_10"),
                "status_60": r.get("status_60"),
                "status_180": r.get("status_180"),
                "status_600": r.get("status_600"),
                "mid_mfe": r.get("mid_mfe"),
                "mid_mae": r.get("mid_mae"),
            }
        )
    return out


def _exec_sheet(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        out.append(
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "fill_price": r.get("fill_price"),
                "exit_price": r.get("exit_price"),
                "bid1_at_fill": r.get("bid1_at_fill"),
                "max_bid1": r.get("max_bid1"),
                "min_bid1": r.get("min_bid1"),
                "executable_mfe": r.get("executable_mfe"),
                "executable_mae": r.get("executable_mae"),
                "realized_return": r.get("realized_return"),
                "capture_ratio": r.get("capture_ratio"),
                "n_path_ticks": r.get("n_path_ticks"),
                "future_leak": r.get("future_leak"),
            }
        )
    return out


def _mfe_sheet(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for r in rows:
        out.append(
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "result": r.get("result"),
                "exit_reason": r.get("exit_reason"),
                "executable_mfe": r.get("executable_mfe"),
                "executable_mae": r.get("executable_mae"),
                "mid_mfe": r.get("mid_mfe"),
                "mid_mae": r.get("mid_mae"),
                "note": "executable vs mid are not the same metric",
            }
        )
    return out


def _capture_sheet(rows: list[dict[str, Any]], by_reason: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = [{"row_type": "REASON", **r} for r in by_reason]
    for r in rows:
        out.append(
            {
                "row_type": "TRADE",
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "exit_reason": r.get("exit_reason"),
                "realized_return": r.get("realized_return"),
                "executable_mfe": r.get("executable_mfe"),
                "capture_ratio": r.get("capture_ratio"),
                "pnl_yen_100": r.get("pnl_yen_100"),
            }
        )
    return out


def _exec_state_sheet(blk: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for m, v in blk.items():
        rec = {"metric": m}
        rec.update({k: v.get(k) for k in ("n", "mean", "median", "p25", "p75")})
        rec["spearman_rho"] = (v.get("spearman_vs_pnl") or {}).get("rho")
        rec["spearman_n"] = (v.get("spearman_vs_pnl") or {}).get("n")
        rec["top10_median"] = (v.get("top10_winners") or {}).get("median")
        rec["remaining_median"] = (v.get("remaining") or {}).get("median")
        rows.append(rec)
    return rows


def _reference_block(
    ref_trades: Optional[list[dict[str, Any]]],
    primary: list[dict[str, Any]],
    primary_tails: dict[str, Any],
    primary_exits: list[dict[str, Any]],
    primary_sym: dict[str, Any],
) -> dict[str, Any]:
    if ref_trades is None:
        return {
            "available": False,
            "reason": "REFERENCE_PATH_NOT_AVAILABLE",
            "trades": None,
            "pnl": None,
            "tail_consistency": None,
            "exit_consistency": None,
            "_rows": [{"available": False, "reason": "REFERENCE_PATH_NOT_AVAILABLE"}],
        }
    n = len(ref_trades)
    sp = sum(pnl(t) for t in ref_trades)
    ok_n = n == P1_REF_TRADES
    ok_p = abs(sp - P1_REF_PNL) < 0.51
    if not (ok_n and ok_p):
        return {
            "available": False,
            "reason": "REFERENCE_RECONCILE_FAIL",
            "trades": n,
            "pnl": round(sp, 2),
            "expected_trades": P1_REF_TRADES,
            "expected_pnl": P1_REF_PNL,
            "tail_consistency": None,
            "exit_consistency": None,
            "_rows": [{"available": False, "reason": "REFERENCE_RECONCILE_FAIL", "trades": n, "pnl": round(sp, 2)}],
        }
    dist = distribution(ref_trades)
    tails = tail_blocks(ref_trades)
    days = day_table(ref_trades, [])
    syms = symbol_table(ref_trades)
    exits = exit_table(ref_trades)
    p_top10 = (primary_tails.get("top10") or {}).get("signed_share_of_total_pnl")
    r_top10 = (tails.get("top10") or {}).get("signed_share_of_total_pnl")
    tail_cons = None
    if p_top10 is not None and r_top10 is not None:
        tail_cons = "DIRECTIONALLY_CONSISTENT" if (float(p_top10) >= 0.3) == (float(r_top10) >= 0.3) else "NOT_CONSISTENT"
        if abs(float(p_top10) - float(r_top10)) < 0.15:
            tail_cons = "DIRECTIONALLY_CONSISTENT"
    p_ex = primary_exits[0]["exit_reason"] if primary_exits else None
    r_ex = exits[0]["exit_reason"] if exits else None
    exit_cons = "DIRECTIONALLY_CONSISTENT" if p_ex and p_ex == r_ex else "NOT_CONSISTENT"
    rows = [{"row_type": "SUMMARY", **dist, "trades": n, "pnl": round(sp, 2)}]
    for k in ("top1", "top3", "top5", "top10"):
        rows.append({"row_type": "TAIL", "block": k, **(tails.get(k) or {})})
    for e in exits:
        rows.append({"row_type": "EXIT", **e})
    for s in syms[:20]:
        rows.append({"row_type": "SYMBOL", **s})
    for d in days:
        rows.append({"row_type": "DAY", **d})
    return {
        "available": True,
        "trades": n,
        "pnl": round(sp, 2),
        "PF": dist.get("PF"),
        "distribution": dist,
        "tail": {k: tails.get(k) for k in ("top1", "top3", "top5", "top10")},
        "exit_reasons": exits,
        "symbol_top": symbol_share_pack(syms, sp),
        "days": [{"date": d["date"], "trades": d["trades"], "pnl": d["pnl"]} for d in days],
        "tail_consistency": tail_cons,
        "exit_consistency": exit_cons,
        "note": "Partial/non-FULL days included. Descriptive only. Does not override PRIMARY. No MFE/MAE.",
        "_rows": rows,
        "primary_top1_symbol": primary_sym.get("top1_symbol"),
    }


def render_md(rep: dict[str, Any]) -> str:
    t = rep.get("TAIL") or {}
    d = rep.get("DAY_CONCENTRATION") or {}
    s = rep.get("SYMBOL_CONCENTRATION") or {}
    path = rep.get("CANONICAL_PATH") or {}
    ref = rep.get("REFERENCE_ALL_USABLE") or {}
    am = (rep.get("AM_PM") or {}).get("AM") or {}
    pm = (rep.get("AM_PM") or {}).get("PM") or {}
    ex1 = t.get("EX_TOP1") or {}
    ex3 = t.get("EX_TOP3") or {}
    ex5 = t.get("EX_TOP5") or {}
    ex10 = t.get("EX_TOP10") or {}
    top3d = d.get("TOP3_DAYS") or {}
    rest = d.get("REST11") or {}
    lines = [
        "# P3-3 Canonical Fixed PnL Source Decomposition",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / strategy validation",
        "",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        "",
        "PRIMARY_FULL14:",
        f"trades: {((rep.get('PRIMARY_FULL14') or {}).get('trades'))}",
        f"pnl: {((rep.get('PRIMARY_FULL14') or {}).get('pnl'))}",
        f"PF: {((rep.get('PRIMARY_FULL14') or {}).get('PF'))}",
        "",
        "TAIL:",
        f"TOP1_TRADE_PNL: {t.get('TOP1_TRADE_PNL')}",
        f"TOP1_SHARE: {t.get('TOP1_SHARE')}",
        f"TOP3_TRADES_PNL: {t.get('TOP3_TRADES_PNL')}",
        f"TOP3_SHARE: {t.get('TOP3_SHARE')}",
        f"TOP5_TRADES_PNL: {t.get('TOP5_TRADES_PNL')}",
        f"TOP5_SHARE: {t.get('TOP5_SHARE')}",
        f"TOP10_TRADES_PNL: {t.get('TOP10_TRADES_PNL')}",
        f"TOP10_SHARE: {t.get('TOP10_SHARE')}",
        "",
        "EX_TOP1:",
        f"pnl: {ex1.get('pnl')}",
        f"PF: {ex1.get('PF')}",
        "EX_TOP3:",
        f"pnl: {ex3.get('pnl')}",
        f"PF: {ex3.get('PF')}",
        "EX_TOP5:",
        f"pnl: {ex5.get('pnl')}",
        f"PF: {ex5.get('PF')}",
        "EX_TOP10:",
        f"pnl: {ex10.get('pnl')}",
        f"PF: {ex10.get('PF')}",
        "",
        "DAY_CONCENTRATION:",
        f"TOP3_DAYS: {top3d.get('dates')}",
        f"pnl: {top3d.get('pnl')}",
        f"share: {top3d.get('share')}",
        "REST11:",
        f"pnl: {rest.get('pnl')}",
        f"PF: {rest.get('PF')}",
        f"best1_day: {d.get('best1_day')}",
        f"best3_day_share: {d.get('best3_day_share')}",
        f"worst1_day_contribution: {d.get('worst1_day_contribution')}",
        "",
        "SYMBOL_CONCENTRATION:",
        f"top1_symbol: {s.get('top1_symbol')}",
        f"pnl: {s.get('top1_pnl')}",
        f"share: {s.get('top1_share')}",
        f"top3_symbols_share: {s.get('top3_share')}",
        "",
        "EXIT_CONTRIBUTION:",
    ]
    for e in rep.get("EXIT_CONTRIBUTION") or []:
        lines.append(
            f"- {e.get('exit_reason')}: n={e.get('count')} W/L/D={e.get('win')}/{e.get('loss')}/{e.get('draw')} "
            f"GP={e.get('gross_profit')} GL={e.get('gross_loss')} PnL={e.get('pnl')} PF={e.get('PF')} "
            f"mean={e.get('mean')} median={e.get('median')} hold_mean={e.get('mean_holding_sec')} "
            f"hold_med={e.get('median_holding_sec')} share={e.get('signed_share_of_total_pnl')}"
        )
    lines += [
        "",
        "ANCHOR_TIME_TOP_CONTRIBUTORS:",
    ]
    for a in ((rep.get("ANCHOR_TIME") or {}).get("top_contributors") or []):
        lines.append(f"- {a.get('anchor_time')}: trades={a.get('trades')} pnl={a.get('pnl')} PF={a.get('PF')}")
    lines += [
        "",
        "AM:",
        f"pnl: {am.get('pnl')}",
        f"PF: {am.get('PF')}",
        "PM:",
        f"pnl: {pm.get('pnl')}",
        f"PF: {pm.get('PF')}",
        "",
        "CANONICAL_PATH:",
        "WIN:",
        f"mid path: {json.dumps({k: (path.get('WIN') or {}).get(k) for k in ('mid_markout_1','mid_markout_10','mid_markout_60','mid_markout_180','mid_markout_600')}, ensure_ascii=False, default=str)}",
        f"executable MFE: {(path.get('WIN') or {}).get('executable_mfe')}",
        f"executable MAE: {(path.get('WIN') or {}).get('executable_mae')}",
        "LOSS:",
        f"mid path: {json.dumps({k: (path.get('LOSS') or {}).get(k) for k in ('mid_markout_1','mid_markout_10','mid_markout_60','mid_markout_180','mid_markout_600')}, ensure_ascii=False, default=str)}",
        f"executable MFE: {(path.get('LOSS') or {}).get('executable_mfe')}",
        f"executable MAE: {(path.get('LOSS') or {}).get('executable_mae')}",
        "",
        "EXIT_CAPTURE:",
    ]
    for e in (rep.get("EXIT_CAPTURE") or {}).get("by_reason") or []:
        lines.append(
            f"- {e.get('exit_reason')}: pnl={e.get('pnl')} med_exec_mfe={e.get('median_executable_mfe')} "
            f"mean_exec_mfe={e.get('mean_executable_mfe')} med_exec_mae={e.get('median_executable_mae')} "
            f"capture_med={((e.get('capture') or {}).get('median'))} capture_mean={((e.get('capture') or {}).get('mean'))} "
            f"capture_buckets lt0/0-1/gt1={((e.get('capture') or {}).get('lt0'))}/{((e.get('capture') or {}).get('in_0_1'))}/{((e.get('capture') or {}).get('gt1'))}"
        )
    lines += [
        "",
        "TOP3 vs REST11 mechanism:",
        f"{json.dumps((rep.get('TOP3_REST11_EXIT_MECHANISM') or {}).get('comparison'), ensure_ascii=False, default=str)}",
        "",
        "RANK_RELATION:",
        json.dumps(rep.get("RANK_RELATION"), ensure_ascii=False, default=str),
        "",
        "EXECUTION_STATE_RELATION:",
    ]
    es = rep.get("EXECUTION_STATE_RELATION") or {}
    for m, blk in es.items():
        lines.append(
            f"- {m}: median={blk.get('median')} p25={blk.get('p25')} p75={blk.get('p75')} "
            f"spearman={((blk.get('spearman_vs_pnl') or {}).get('rho'))} "
            f"top10_med={((blk.get('top10_winners') or {}).get('median'))} rest_med={((blk.get('remaining') or {}).get('median'))}"
        )
    lines += [
        "",
        "REFERENCE_ALL_USABLE:",
        f"available: {str(ref.get('available')).lower()}",
        f"trades: {ref.get('trades')}",
        f"pnl: {ref.get('pnl')}",
        f"tail_consistency: {ref.get('tail_consistency')}",
        f"exit_consistency: {ref.get('exit_consistency')}",
        "",
        f"PRIMARY_PNL_SOURCE: `{rep.get('PRIMARY_PNL_SOURCE')}`",
        f"SECONDARY_PNL_SOURCES: {rep.get('SECONDARY_PNL_SOURCES')}",
        "",
        "ANCHOR_DIRECTIONAL_EDGE: NOT_SUPPORTED",
        "POST_FILL_DIRECTION: MIXED",
        "EXECUTION_EDGE: SUPPORTED",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    rep.pop("_merged", None)
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    pf14 = public.get("PRIMARY_FULL14") or {}
    tail = public.get("TAIL") or {}
    dayc = public.get("DAY_CONCENTRATION") or {}
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("PRIMARY_trades", pf14.get("trades")),
            ("PRIMARY_pnl", pf14.get("pnl")),
            ("PRIMARY_PF", pf14.get("PF")),
            ("TOP1_TRADE_PNL", tail.get("TOP1_TRADE_PNL")),
            ("TOP1_SHARE", tail.get("TOP1_SHARE")),
            ("TOP10_SHARE", tail.get("TOP10_SHARE")),
            ("EX_TOP10_pnl", ((tail.get("EX_TOP10") or {}).get("pnl"))),
            ("EX_TOP10_PF", ((tail.get("EX_TOP10") or {}).get("PF"))),
            ("TOP3_DAYS_pnl", ((dayc.get("TOP3_DAYS") or {}).get("pnl"))),
            ("TOP3_DAYS_share", ((dayc.get("TOP3_DAYS") or {}).get("share"))),
            ("REST11_pnl", ((dayc.get("REST11") or {}).get("pnl"))),
            ("REST11_PF", ((dayc.get("REST11") or {}).get("PF"))),
            ("PRIMARY_PNL_SOURCE", public.get("PRIMARY_PNL_SOURCE")),
            ("SECONDARY_PNL_SOURCES", public.get("SECONDARY_PNL_SOURCES")),
            ("ANCHOR_DIRECTIONAL_EDGE", "NOT_SUPPORTED"),
            ("POST_FILL_DIRECTION", "MIXED"),
            ("EXECUTION_EDGE", "SUPPORTED"),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("NEW_STRATEGY_TESTED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Canonical_Trades"), sheets.get("canonical") or [])
    _write_rows(wb.create_sheet("Trade_Tail"), sheets.get("tail") or [])
    _write_rows(wb.create_sheet("Day_Concentration"), sheets.get("days") or [])
    _write_rows(wb.create_sheet("Symbol_Concentration"), sheets.get("symbols") or [])
    _write_rows(wb.create_sheet("Exit_Reasons"), sheets.get("exits") or [])
    _write_rows(wb.create_sheet("Top_Winners"), sheets.get("winners") or [])
    _write_rows(wb.create_sheet("Top3_Rest11"), sheets.get("top3_rest11") or [])
    _write_rows(wb.create_sheet("Anchor_Times"), sheets.get("anchors") or [])
    _write_rows(wb.create_sheet("AM_PM"), sheets.get("ampm") or [])
    _write_rows(wb.create_sheet("Mid_Path"), sheets.get("mid_path") or [])
    _write_rows(wb.create_sheet("Executable_Path"), sheets.get("exec_path") or [])
    _write_rows(wb.create_sheet("MFE_MAE"), sheets.get("mfe_mae") or [])
    _write_rows(wb.create_sheet("Exit_Capture"), sheets.get("capture") or [])
    _write_rows(wb.create_sheet("Rank"), sheets.get("rank") or [])
    _write_rows(wb.create_sheet("Execution_State"), sheets.get("exec_state") or [])
    _write_rows(wb.create_sheet("Reference_All_Usable"), sheets.get("reference") or [])
    ident = wb.create_sheet("Identity")
    ident_rows = [(k, v) for k, v in (public.get("Identity") or {}).items()]
    ident_rows += [("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")), ("LABEL", public.get("LABEL"))]
    _kv_sheet(ident, ident_rows)
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("NEW_STRATEGY_TESTED", False),
            ("RETUNING_DONE", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("COUNTERFACTUAL", False),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
