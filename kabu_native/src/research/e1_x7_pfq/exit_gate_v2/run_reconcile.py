"""Orchestrator for EXIT Gate Reconciliation V2."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_provisional.util import sha256_file, sha256_obj

from . import ANALYSIS_ID, CANDIDATE_ID, PAIRS, SOURCE_BRIDGE_RUN, SOURCE_EXIT_GATE_RUN, SOURCE_VERDICT
from .evaluate import check_identity, decide_verdict, evaluate_pair, expected_delta, gate_pair
from .precommit import build_precommit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[4]
BRIDGE_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_bridge_v2"
V1_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_exit_gate"
PUBLISH = NATIVE / "results" / "research" / "e1_x7_pfq_exit_gate_v2"


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "pfq_entry_changed": False,
        "pfq_exit_changed": False,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
        "exit_revision_implemented": False,
        "source_run_overwritten": False,
    }


def run_once(*, label: str = "A", ab_ok_placeholder: bool = True) -> dict[str, Any]:
    run_id = f"e1x7_pfq_exit_gate_v2_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"

    v1_report = V1_DIR / "report.json"
    v1_audit = V1_DIR / "audit.xlsx"
    bridge_report = BRIDGE_DIR / "report.json"
    bridge_audit = BRIDGE_DIR / "audit.xlsx"

    src_run_sha = sha256_file(v1_report)
    src_audit_sha = sha256_file(v1_audit)
    br_report_sha = sha256_file(bridge_report)
    br_audit_sha = sha256_file(bridge_audit)

    print(f"=== [{label}] Precommit ===", flush=True)
    precommit = build_precommit(
        source_run_sha=src_run_sha,
        source_audit_sha=src_audit_sha,
        bridge_report_sha=br_report_sha,
        bridge_audit_sha=br_audit_sha,
    )

    v1 = json.loads(v1_report.read_text(encoding="utf-8"))
    assert v1.get("run_id") == SOURCE_EXIT_GATE_RUN or SOURCE_EXIT_GATE_RUN in str(v1.get("run_id"))
    assert v1.get("verdict") == SOURCE_VERDICT

    br = json.loads(bridge_report.read_text(encoding="utf-8"))
    entry_support = (
        ((br.get("verdict_detail") or {}).get("entry_support") or {})
        .get(CANDIDATE_ID, {})
        .get("supported")
        is True
    )

    # Use frozen bridge paths/trades (same economic identity as V1)
    print(f"=== [{label}] Load bridge JointTrades / FixedGrid / CF ===", flush=True)
    trades = _load_sheet(bridge_audit, "JointTrades")
    fixed_grid = _load_sheet(bridge_audit, "FixedGridOutcome")
    counterfactual = _load_sheet(bridge_audit, "Counterfactual")
    fg_by = {r["episode_id"]: r for r in fixed_grid}
    cf_by = {(r["pair_id"], r["episode_id"]): r for r in counterfactual}

    identity = check_identity(trades)
    identity_sha = sha256_obj({
        "unique_episode_n": identity["unique_episode_n"],
        "progress": identity["progress_trade_rows"],
        "protect": identity["protect_trade_rows"],
        "dups": identity["duplicate_within_pair"],
    })
    if not identity["ok"]:
        return {
            "analysis_id": ANALYSIS_ID,
            "run_id": run_id,
            "verdict": "E1_X7_PFQ_EXIT_GATE_IDENTITY_MISMATCH",
            "identity": identity,
            "precommit": precommit,
            "safety": _safety(),
            "stop": True,
            "determinism_shas": {"identity_sha": identity_sha, "verdict": "E1_X7_PFQ_EXIT_GATE_IDENTITY_MISMATCH"},
        }

    print(f"=== [{label}] Corrected pair evaluation ===", flush=True)
    pair_results = {
        pid: evaluate_pair(trades, pair_id=pid, fixed_grid_by_eid=fg_by, cf_by_key=cf_by)
        for pid in PAIRS
    }
    gates = {
        pid: gate_pair(
            pair_results[pid],
            entry_path_support=entry_support,
            identity_ok=True,
            ab_ok=ab_ok_placeholder,
        )
        for pid in PAIRS
    }
    verdict_detail = decide_verdict(gates, pair_results)
    deltas = expected_delta(pair_results)

    denom_sha = sha256_obj({pid: pair_results[pid]["denominator_episode_ids"] for pid in PAIRS})
    repair_sha = sha256_obj({pid: pair_results[pid]["repairable_episode_ids"] for pid in PAIRS})
    opp_sha = sha256_obj({
        pid: sorted(r["episode_id"] for r in pair_results[pid]["opp_cost_rows"])
        for pid in PAIRS
    })

    pair_public = {}
    for pid, pr in pair_results.items():
        pair_public[pid] = {
            k: v for k, v in pr.items()
            if k not in ("repairable_rows", "opp_cost_rows", "integrity_rows",
                         "denominator_episode_ids", "repairable_episode_ids")
        }

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "source_exit_gate_run": SOURCE_EXIT_GATE_RUN,
        "source_verdict": SOURCE_VERDICT,
        "source_bridge_run": SOURCE_BRIDGE_RUN,
        "source_run_sha256": src_run_sha,
        "source_audit_sha256": src_audit_sha,
        "precommit": precommit,
        "identity": identity,
        "entry_path_support": {CANDIDATE_ID: entry_support},
        "pair_results": pair_public,
        "pair_gates": gates,
        "expectation_check": deltas,
        "verdict_detail": verdict_detail,
        "verdict": verdict_detail["verdict"],
        "determinism_shas": {
            "identity_sha": identity_sha,
            "denominator_sha": denom_sha,
            "repairable_subset_sha": repair_sha,
            "profitable_opportunity_cost_sha": opp_sha,
            "pair_gate": {pid: gates[pid]["pass"] for pid in PAIRS},
            "selected_baseline": verdict_detail.get("selected_baseline_pair"),
            "verdict": verdict_detail["verdict"],
        },
        "safety": _safety(),
        "stop": True,
        "_sheets": {
            "PairEpisodes": [
                r for pid in PAIRS for r in pair_results[pid]["integrity_rows"]
            ],
            "Denominators": [
                {"pair_id": pid, "episode_id": eid}
                for pid in PAIRS
                for eid in pair_results[pid]["denominator_episode_ids"]
            ],
            "RepairableLoss": [
                r for pid in PAIRS for r in pair_results[pid]["repairable_rows"]
            ],
            "ProfitableOpportunityCost": [
                r for pid in PAIRS for r in pair_results[pid]["opp_cost_rows"]
            ],
            "FailureMechanisms": [
                {
                    "pair_id": pid,
                    "mechanism": m,
                    "n": n,
                    "fraction_of_repairable": n / pair_results[pid]["repairable_in_denominator_n"]
                    if pair_results[pid]["repairable_in_denominator_n"] else 0.0,
                }
                for pid in PAIRS
                for m, n in (pair_results[pid]["failure_mechanism_counts"] or {}).items()
            ],
        },
    }
    return report


def reapply_gates_with_ab(report: dict[str, Any], *, ab_ok: bool) -> dict[str, Any]:
    repair_by = {pid: [] for pid in PAIRS}
    for r in (report.get("_sheets") or {}).get("RepairableLoss") or []:
        repair_by[r["pair_id"]].append(r)
    denom_by = {pid: [] for pid in PAIRS}
    for r in (report.get("_sheets") or {}).get("Denominators") or []:
        denom_by[r["pair_id"]].append(r["episode_id"])
    opp_by = {pid: [] for pid in PAIRS}
    for r in (report.get("_sheets") or {}).get("ProfitableOpportunityCost") or []:
        opp_by[r["pair_id"]].append(r)

    pair_results = {}
    for pid in PAIRS:
        pub = dict(report["pair_results"][pid])
        rows = repair_by[pid]
        pub["repairable_rows"] = rows
        pub["denominator_episode_ids"] = denom_by[pid]
        pub["repairable_episode_ids"] = [r["episode_id"] for r in rows]
        pub["opp_cost_rows"] = opp_by[pid]
        pub["subset_invariant_ok"] = set(pub["repairable_episode_ids"]) <= set(denom_by[pid])
        pair_results[pid] = pub

    gates = {
        pid: gate_pair(
            pair_results[pid],
            entry_path_support=report["entry_path_support"][CANDIDATE_ID],
            identity_ok=report["identity"]["ok"],
            ab_ok=ab_ok,
        )
        for pid in PAIRS
    }
    vd = decide_verdict(gates, pair_results)
    report["pair_gates"] = gates
    report["verdict_detail"] = vd
    report["verdict"] = vd["verdict"]
    report["determinism_shas"]["pair_gate"] = {pid: gates[pid]["pass"] for pid in PAIRS}
    report["determinism_shas"]["selected_baseline"] = vd.get("selected_baseline_pair")
    report["determinism_shas"]["verdict"] = vd["verdict"]
    return report
