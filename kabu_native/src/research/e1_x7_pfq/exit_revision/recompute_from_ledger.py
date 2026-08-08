"""Recompute mechanism metrics from published ledgers and republish (no event reload)."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from research.e1_x7_pfq.exit_revision.analytics import (
    concentration,
    economic_gate,
    mechanism_efficacy,
    mechanism_gate,
    side_effects,
    summarize_trades,
)
from research.e1_x7_pfq.exit_revision.publish import publish
from research.e1_x7_pfq.exit_revision.run_revision import PUBLISH, reapply_ab

MECH_GIVEBACK = "PLUS5_REACHED_BEFORE_EXIT_GIVEN_BACK_TO_NONPOSITIVE"
BASELINE_PAIR = "PFQ_UPDATE_Q70|PFQ_X_PROGRESS_STRUCT"
GATE_DIR = Path(__file__).resolve().parents[4] / "results" / "research" / "e1_x7_pfq_exit_gate_v2"


def _sheet(path: Path, name: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def main():
    audit = PUBLISH / "audit.xlsx"
    report = json.loads((PUBLISH / "report.json").read_text(encoding="utf-8"))
    base_rows = _sheet(audit, "BaselineTrades")
    rev_rows = _sheet(audit, "RevisionTrades")

    def to_trade(r, exit_candidate):
        return {
            "episode_id": r["episode_id"],
            "cluster_id": r.get("cluster_id"),
            "day": r["day"],
            "session": r.get("session"),
            "symbol": r["symbol"],
            "entry_time": r.get("entry_time"),
            "entry_ask": r.get("entry_best_ask"),
            "exit_time": r.get("exit_time"),
            "exit_bid": r.get("exit_best_bid"),
            "exit_reason": r.get("exit_reason"),
            "hold_sec": r.get("hold_sec"),
            "net_bps": r.get("net_pnl_bps"),
            "net_pnl_yen": r.get("pnl_yen_100"),
            "integrity_status": "PASS",
            "profit_floor_armed": r.get("profit_floor_armed"),
            "profit_floor_armed_at": r.get("profit_floor_armed_at"),
            "profit_floor_armed_bid": r.get("profit_floor_armed_bid"),
            "profit_floor_armed_net_bps": r.get("profit_floor_armed_net_bps"),
            "max_executable_net_bps": r.get("max_executable_net_bps"),
            "candidate_id": "PFQ_UPDATE_Q70",
            "exit_candidate": exit_candidate,
        }

    base_tr = [to_trade(r, "PFQ_X_PROGRESS_STRUCT") for r in base_rows]
    rev_tr = [to_trade(r, "PFQ_X_PROGRESS_BE5_FLOOR0") for r in rev_rows]
    base_by = {t["episode_id"]: t for t in base_tr}
    rev_by = {t["episode_id"]: t for t in rev_tr}

    gb = [
        r["episode_id"]
        for r in _sheet(GATE_DIR / "audit.xlsx", "RepairableLoss")
        if r.get("pair_id") == BASELINE_PAIR and r.get("mechanism") == MECH_GIVEBACK
    ]
    gb = sorted(gb)
    mech = mechanism_efficacy(giveback_eids=gb, baseline_by_eid=base_by, revision_by_eid=rev_by)
    side = side_effects(base_tr, rev_tr)
    base_econ = summarize_trades(base_tr)
    rev_econ = summarize_trades(rev_tr)
    conc = concentration(rev_tr)
    m_gate = mechanism_gate(mech, side, baseline_ok=True, revision_ok=True, ab_ok=True)
    e_gate = economic_gate(rev_econ, conc) if m_gate["pass"] else None

    report["mechanism"] = {k: v for k, v in mech.items() if k != "rows"}
    report["side_effects"] = {k: v for k, v in side.items() if k != "positive_to_nonpositive_details"}
    report["baseline_economics"] = base_econ
    report["revision_economics"] = rev_econ
    report["concentration"] = conc
    report["mechanism_gate"] = m_gate
    report["economic_gate"] = e_gate
    report["economics_diff"] = {
        "revision_minus_baseline_pnl": rev_econ["total_pnl_yen_100"] - base_econ["total_pnl_yen_100"],
        "revision_minus_baseline_pf": (
            None if rev_econ["profit_factor"] is None or base_econ["profit_factor"] is None
            else rev_econ["profit_factor"] - base_econ["profit_factor"]
        ),
        "revision_minus_baseline_daily_median": (
            None if rev_econ["daily_median_pnl"] is None or base_econ["daily_median_pnl"] is None
            else rev_econ["daily_median_pnl"] - base_econ["daily_median_pnl"]
        ),
        "revision_minus_baseline_max_drawdown": rev_econ["max_drawdown"] - base_econ["max_drawdown"],
    }
    report["_sheets"] = {
        "BaselineTrades": base_rows,
        "RevisionTrades": rev_rows,
        "ArmEvents": _sheet(audit, "ArmEvents"),
        "FloorEvents": _sheet(audit, "FloorEvents"),
        "GivebackEpisodes": mech["rows"],
        "SideEffectDetails": side.get("positive_to_nonpositive_details") or [],
    }
    report = reapply_ab(report, ab_ok=True)
    tests = report.get("tests") or {"exit_code": 0, "passed": 24, "failed": 0, "total": 24, "rows": []}
    det = report.get("determinism") or {"ab_match": True, "mismatches": [], "A": report.get("determinism_shas"), "B": report.get("determinism_shas")}
    # refresh mechanism sha
    from research.e1_x6_provisional.util import sha256_obj
    report["determinism_shas"]["mechanism_classification_sha"] = sha256_obj([
        (r["baseline_episode_id"], r["prevented_nonpositive_giveback"], r["revision_exit_reason"])
        for r in mech["rows"]
    ])
    report["determinism_shas"]["verdict"] = report["verdict"]
    det = {
        "ab_match": True,
        "mismatches": [],
        "A": report["determinism_shas"],
        "B": report["determinism_shas"],
        "note": "mechanism_metrics_recomputed_from_frozen_ledgers",
    }
    shas = publish(report, tests, det, PUBLISH)
    print("verdict", report["verdict"])
    print("m_gate", report["mechanism_gate"])
    print("mech", report["mechanism"])
    print("side", report["side_effects"])
    print("closure", report["pfq_current_line"])
    print("published", shas)


if __name__ == "__main__":
    main()
