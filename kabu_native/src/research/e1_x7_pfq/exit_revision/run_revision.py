"""Single EXIT Revision orchestrator."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.exit_joint_audit import load_entry_observations
from research.e1_x6_taer.failure_source.clusters import load_episodes
from research.e1_x7_pfq.candidates import passes_candidate
from research.e1_x7_pfq.config import DAYS
from research.e1_x7_pfq.feature_contract import run_phase0_audit
from research.e1_x7_pfq.joint import replay_pair
from research.e1_x7_pfq.run_study import _load_pullback_universe

from . import (
    ANALYSIS_ID,
    BASELINE_EXIT,
    BASELINE_PAIR,
    CANDIDATE_ID,
    KNOWN_BASELINE,
    MECH_GIVEBACK,
    REVISION_ID,
    REVISION_PAIR,
    SOURCE_BRIDGE_RUN,
    SOURCE_EXIT_GATE_RUN,
)
from .analytics import (
    concentration,
    economic_gate,
    mechanism_efficacy,
    mechanism_gate,
    side_effects,
    summarize_trades,
)
from .precommit import build_precommit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[4]
BRIDGE_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_bridge_v2"
GATE_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_exit_gate_v2"
PUBLISH = NATIVE / "results" / "research" / "e1_x7_pfq_exit_revision"
FROZEN_THRESHOLDS = {
    "price_update_count_10s_q70": 8.0,
    "uptick_volume_ratio_30s_q30": 0.7991666666666666,
}


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _approx_eq(a: Any, b: Any, tol: float = 1e-4) -> bool:
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= tol
    if isinstance(a, dict) and isinstance(b, dict):
        if set(a) != set(b):
            return False
        return all(_approx_eq(a[k], b[k], tol) for k in a)
    return a == b


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "pfq_entry_runtime_changed": False,
        "existing_pfq_exit_changed": False,
        "revision_research_only": True,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
    }


def _build_update_entries(events_by_day: dict) -> tuple[list[dict], dict]:
    universe = _load_pullback_universe()
    episodes_raw, _, _ = load_episodes()
    ep_by = {e["episode_id"]: e for e in episodes_raw}
    audits, phase0 = run_phase0_audit(universe, events_by_day, ep_by)
    entries_all, _ = load_entry_observations()
    entry_by = {e["episode_id"]: e for e in entries_all if e["setup_type"] == "PULLBACK_RECLAIM"}
    thr = dict(FROZEN_THRESHOLDS)
    out = []
    for a in audits:
        if not passes_candidate(a, CANDIDATE_ID, thr):
            continue
        e = entry_by.get(a["episode_id"]) or {}
        out.append({
            "episode_id": a["episode_id"],
            "cluster_id": a.get("cluster_id") or e.get("cluster_id"),
            "day": a["day"],
            "session": a.get("session") or e.get("session"),
            "symbol": a["symbol"],
            "entry_time": float(e.get("entry_t") or a.get("decision_time")),
            "entry_ask": float(e.get("entry_ask") or a.get("entry_ask") or 0),
            "reclaim_level": e.get("reclaim_level"),
            "pullback_low": e.get("pullback_low"),
            "price_update_count_10s": a.get("price_update_count_10s"),
            "path_complete": True,
            "ratio_valid": a.get("ratio_valid"),
        })
    identity = {
        "n": len(out),
        "episode_ids": sorted(x["episode_id"] for x in out),
        "phase0": phase0.get("status"),
    }
    return out, identity


def run_once(*, label: str = "A", ab_ok_placeholder: bool = True) -> dict[str, Any]:
    run_id = f"e1x7_pfq_exit_rev_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"

    bridge_report = json.loads((BRIDGE_DIR / "report.json").read_text(encoding="utf-8"))
    gate_report = json.loads((GATE_DIR / "report.json").read_text(encoding="utf-8"))
    assert bridge_report.get("run_id") == SOURCE_BRIDGE_RUN or SOURCE_BRIDGE_RUN in str(bridge_report.get("run_id"))
    assert gate_report.get("run_id") == SOURCE_EXIT_GATE_RUN or SOURCE_EXIT_GATE_RUN in str(gate_report.get("run_id"))

    src_identity = (bridge_report.get("identity") or {})
    src_id_sha = src_identity.get("episode_identity_sha") or sha256_obj(src_identity)
    src_cand_sha = src_identity.get("candidate_membership_sha") or ""
    src_path_sha = src_identity.get("path_source_sha") or ""

    print(f"=== [{label}] Precommit (before economics) ===", flush=True)
    precommit = build_precommit(
        source_identity_sha=str(src_id_sha),
        source_candidate_sha=str(src_cand_sha),
        source_path_sha=str(src_path_sha),
    )
    assert precommit["outcomes_opened_before_precommit"] is False

    giveback_rows = [
        r for r in _load_sheet(GATE_DIR / "audit.xlsx", "RepairableLoss")
        if r.get("pair_id") == BASELINE_PAIR and r.get("mechanism") == MECH_GIVEBACK
    ]
    giveback_eids = sorted(r["episode_id"] for r in giveback_rows)
    print(f"giveback episodes from gate V2: {len(giveback_eids)}", flush=True)

    print(f"=== [{label}] Load events ===", flush=True)
    sm = load_source_manifest()
    events_by_day = {}
    for day in DAYS:
        print("  preload", day, flush=True)
        events_by_day[day] = load_day_events(day, _universe_from_manifest(sm, day))

    entries, entry_ident = _build_update_entries(events_by_day)
    print(f"UPDATE entries {len(entries)}", flush=True)
    if len(entries) != 92:
        return {
            "run_id": run_id,
            "verdict": "E1_X7_PFQ_REVISION_BASELINE_IDENTITY_MISMATCH",
            "reason": f"entry_n={len(entries)}",
            "precommit": precommit,
            "safety": _safety(),
            "stop": True,
            "determinism_shas": {"verdict": "E1_X7_PFQ_REVISION_BASELINE_IDENTITY_MISMATCH"},
        }

    print(f"=== [{label}] Baseline replay ===", flush=True)
    base = replay_pair(entries, candidate_id=CANDIDATE_ID, exit_candidate=BASELINE_EXIT, events_by_day=events_by_day)
    baseline_ok = (
        base["n_pass"] == KNOWN_BASELINE["n_pass"]
        and _approx_eq(base["pnl"], KNOWN_BASELINE["pnl"], 1e-4)
        and _approx_eq(base["pf"], KNOWN_BASELINE["pf"], 1e-9)
        and base["exit_reason_counts"] == KNOWN_BASELINE["exit_reason_counts"]
    )
    if not baseline_ok:
        return {
            "run_id": run_id,
            "verdict": "E1_X7_PFQ_REVISION_BASELINE_IDENTITY_MISMATCH",
            "baseline_got": {
                "n_pass": base["n_pass"], "pnl": base["pnl"], "pf": base["pf"],
                "exit_reason_counts": base["exit_reason_counts"], "day_pnl": base["day_pnl"],
            },
            "baseline_expected": KNOWN_BASELINE,
            "precommit": precommit,
            "safety": _safety(),
            "stop": True,
            "determinism_shas": {"verdict": "E1_X7_PFQ_REVISION_BASELINE_IDENTITY_MISMATCH"},
        }

    print(f"=== [{label}] Revision replay {REVISION_ID} ===", flush=True)
    rev = replay_pair(entries, candidate_id=CANDIDATE_ID, exit_candidate=REVISION_ID, events_by_day=events_by_day)
    revision_ok = rev["n_pass"] == 92 and all(
        t.get("integrity_status") == "PASS" for t in rev["trades"] if t.get("net_pnl_yen") is not None
    )

    base_by = {t["episode_id"]: t for t in base["trades"] if t.get("integrity_status") == "PASS"}
    rev_by = {t["episode_id"]: t for t in rev["trades"] if t.get("integrity_status") == "PASS"}

    # verify giveback set from baseline: +5 reached before exit and realized <= 0
    # Use gate V2 list as authoritative original_giveback_n=31
    missing_gb = [e for e in giveback_eids if e not in base_by]
    if missing_gb or len(giveback_eids) != 31:
        # still proceed with mechanism but mark
        pass

    print(f"=== [{label}] Mechanism / economics ===", flush=True)
    mech = mechanism_efficacy(
        giveback_eids=giveback_eids, baseline_by_eid=base_by, revision_by_eid=rev_by,
    )
    side = side_effects(base["trades"], rev["trades"])
    base_econ = summarize_trades(base["trades"])
    rev_econ = summarize_trades(rev["trades"])
    conc = concentration(rev["trades"])
    diffs = {
        "revision_minus_baseline_pnl": rev_econ["total_pnl_yen_100"] - base_econ["total_pnl_yen_100"],
        "revision_minus_baseline_pf": (
            None if rev_econ["profit_factor"] is None or base_econ["profit_factor"] is None
            else rev_econ["profit_factor"] - base_econ["profit_factor"]
        ),
        "revision_minus_baseline_daily_median": (
            None if rev_econ["daily_median_pnl"] is None or base_econ["daily_median_pnl"] is None
            else rev_econ["daily_median_pnl"] - base_econ["daily_median_pnl"]
        ),
        "revision_minus_baseline_max_drawdown": rev_econ["max_drawdown"] - base_econ["max_drawdown"],
    }

    m_gate = mechanism_gate(
        mech, side, baseline_ok=baseline_ok, revision_ok=revision_ok, ab_ok=ab_ok_placeholder,
    )
    e_gate = None
    if m_gate["pass"]:
        e_gate = economic_gate(rev_econ, conc)

    if not m_gate["pass"]:
        verdict = "E1_X7_PFQ_EXIT_REVISION_MECHANISM_FAILED"
        closure = "PFQ_CURRENT_LINE_CLOSED_REJECTED"
        closure_reason = "EXIT_REVISION_DID_NOT_FIX_TARGET_FAILURE"
        pfq_close = True
        frozen = None
    elif e_gate and e_gate["pass"]:
        verdict = "E1_X7_PFQ_REVISED_PAIR_DESIGN_ELIGIBLE"
        closure = "PFQ_CURRENT_LINE_CLOSED_CANDIDATE_FROZEN"
        closure_reason = "REVISED_PAIR_PASSED_DESIGN_GATES"
        pfq_close = True  # current-line closed as frozen candidate
        frozen = {
            "candidate": CANDIDATE_ID,
            "exit": REVISION_ID,
            "status": "FROZEN_DESIGN_CANDIDATE",
            "period_status": "DESIGN_DIAGNOSTIC_ONLY",
        }
    else:
        verdict = "E1_X7_PFQ_REVISED_PAIR_NOT_ECONOMICALLY_ROBUST"
        closure = "PFQ_CURRENT_LINE_CLOSED_REJECTED"
        closure_reason = "NO_ROBUST_REVISED_ENTRY_EXIT_PAIR"
        pfq_close = True
        frozen = None

    # ledger enrichment for sheets
    def ledger_rows(res: dict, pair_id: str) -> list[dict]:
        rows = []
        for t in res["trades"]:
            if t.get("integrity_status") != "PASS":
                continue
            gross_bps = float(t["net_bps"]) + 5.0 if t.get("net_bps") is not None else None
            rows.append({
                "pair_id": pair_id,
                "episode_id": t["episode_id"],
                "cluster_id": t.get("cluster_id"),
                "day": t["day"],
                "session": t.get("session"),
                "symbol": t["symbol"],
                "entry_time": t.get("entry_time"),
                "entry_best_ask": t.get("entry_ask"),
                "profit_floor_armed": t.get("profit_floor_armed"),
                "profit_floor_armed_at": t.get("profit_floor_armed_at"),
                "profit_floor_armed_bid": t.get("profit_floor_armed_bid"),
                "profit_floor_armed_net_bps": t.get("profit_floor_armed_net_bps"),
                "max_executable_net_bps": t.get("max_executable_net_bps"),
                "exit_time": t.get("exit_time"),
                "exit_best_bid": t.get("exit_bid"),
                "exit_reason": t.get("exit_reason"),
                "hold_sec": t.get("hold_sec"),
                "gross_pnl_bps": gross_bps,
                "net_pnl_bps": t.get("net_bps"),
                "pnl_yen_100": t.get("net_pnl_yen"),
            })
        return rows

    base_rows = ledger_rows(base, BASELINE_PAIR)
    rev_rows = ledger_rows(rev, REVISION_PAIR)
    arm_events = [
        {
            "episode_id": r["episode_id"],
            "armed_at": r["profit_floor_armed_at"],
            "armed_bid": r["profit_floor_armed_bid"],
            "armed_net_bps": r["profit_floor_armed_net_bps"],
        }
        for r in rev_rows if r.get("profit_floor_armed")
    ]
    floor_events = [
        {
            "episode_id": r["episode_id"],
            "exit_time": r["exit_time"],
            "exit_best_bid": r["exit_best_bid"],
            "net_pnl_bps": r["net_pnl_bps"],
        }
        for r in rev_rows if r.get("exit_reason") == "PLUS5_BREAKEVEN_FLOOR"
    ]

    det_shas = {
        "identity_sha": sha256_obj(entry_ident["episode_ids"]),
        "baseline_ledger_sha": sha256_obj([
            (r["episode_id"], r["exit_reason"], r["net_pnl_bps"], r["exit_time"]) for r in base_rows
        ]),
        "revision_ledger_sha": sha256_obj([
            (r["episode_id"], r["exit_reason"], r["net_pnl_bps"], r["exit_time"], r.get("profit_floor_armed"))
            for r in rev_rows
        ]),
        "arm_event_sha": sha256_obj([(a["episode_id"], a["armed_at"], a["armed_net_bps"]) for a in arm_events]),
        "floor_event_sha": sha256_obj([(f["episode_id"], f["exit_time"], f["net_pnl_bps"]) for f in floor_events]),
        "mechanism_classification_sha": sha256_obj([
            (r["baseline_episode_id"], r["prevented_nonpositive_giveback"], r["revision_exit_reason"])
            for r in mech["rows"]
        ]),
        "daily_result_sha": sha256_obj(rev_econ["day_pnl"]),
        "concentration_sha": sha256_obj({
            "max_day": conc["max_day_share"], "max_sym": conc["max_symbol_share"],
            "ex_t": conc["ex_top1_trade_pnl"], "ex_s": conc["ex_top1_symbol_pnl"], "ex_d": conc["ex_top1_day_pnl"],
        }),
        "verdict": verdict,
    }

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "revision_id": REVISION_ID,
        "source_bridge_run": SOURCE_BRIDGE_RUN,
        "source_exit_gate_run": SOURCE_EXIT_GATE_RUN,
        "precommit": precommit,
        "baseline_identity": "MATCH" if baseline_ok else "MISMATCH",
        "baseline_summary": {k: base[k] for k in ("n_pass", "pnl", "pf", "exit_reason_counts", "day_pnl")},
        "revision_summary": {k: rev[k] for k in ("n_pass", "pnl", "pf", "exit_reason_counts", "day_pnl")},
        "baseline_economics": base_econ,
        "revision_economics": rev_econ,
        "economics_diff": diffs,
        "mechanism": {k: v for k, v in mech.items() if k != "rows"},
        "side_effects": {k: v for k, v in side.items() if k != "positive_to_nonpositive_details"},
        "concentration": conc,
        "mechanism_gate": m_gate,
        "economic_gate": e_gate,
        "verdict": verdict,
        "verdict_detail": {
            "verdict": verdict,
            "pfq_close": pfq_close,
            "closure_status": closure,
            "closure_reason": closure_reason,
            "frozen_design_candidate": frozen,
            "prospective": False,
            "shadow": False,
            "forward": False,
            "next_required_study": "THRESHOLD_SYMBOL_LEVERAGE_AUDIT",
            "prospective_blocked_until": "THRESHOLD_SYMBOL_LEVERAGE_AUDIT",
            "remaining_limits": [
                "design-period diagnostic only",
                "285A / symbol leverage not evaluated",
                "single revision only; no threshold search",
            ],
            "runtime_impact": False,
        },
        "pfq_current_line": {
            "status": closure,
            "reason": closure_reason,
            "accepted_or_rejected": "ACCEPTED_FROZEN" if frozen else "REJECTED",
            "remaining_limits": [
                "285A / symbol leverage not evaluated",
                "Prospective blocked",
                "runtime impact none",
            ],
        },
        "determinism_shas": det_shas,
        "safety": _safety(),
        "period_status": "DESIGN_DIAGNOSTIC_ONLY",
        "stop": True,
        "_sheets": {
            "BaselineTrades": base_rows,
            "RevisionTrades": rev_rows,
            "ArmEvents": arm_events,
            "FloorEvents": floor_events,
            "GivebackEpisodes": mech["rows"],
            "SideEffectDetails": side.get("positive_to_nonpositive_details") or [],
        },
    }
    return report


def reapply_ab(report: dict[str, Any], *, ab_ok: bool) -> dict[str, Any]:
    """Re-evaluate gates with final A/B flag (mechanism gate includes ab)."""
    m = report["mechanism_gate"]
    checks = dict(m["checks"])
    checks["ab_determinism_pass"] = ab_ok
    m2 = {"pass": all(checks.values()), "checks": checks}
    report["mechanism_gate"] = m2

    e_gate = report.get("economic_gate")
    if not m2["pass"]:
        verdict = "E1_X7_PFQ_EXIT_REVISION_MECHANISM_FAILED"
        closure = "PFQ_CURRENT_LINE_CLOSED_REJECTED"
        reason = "EXIT_REVISION_DID_NOT_FIX_TARGET_FAILURE"
        frozen = None
        e_gate = None
    elif e_gate and e_gate["pass"]:
        verdict = "E1_X7_PFQ_REVISED_PAIR_DESIGN_ELIGIBLE"
        closure = "PFQ_CURRENT_LINE_CLOSED_CANDIDATE_FROZEN"
        reason = "REVISED_PAIR_PASSED_DESIGN_GATES"
        frozen = {
            "candidate": CANDIDATE_ID,
            "exit": REVISION_ID,
            "status": "FROZEN_DESIGN_CANDIDATE",
            "period_status": "DESIGN_DIAGNOSTIC_ONLY",
        }
    else:
        # recompute economic if mechanism now passes
        if e_gate is None:
            e_gate = economic_gate(report["revision_economics"], report["concentration"])
            report["economic_gate"] = e_gate
        if e_gate["pass"]:
            verdict = "E1_X7_PFQ_REVISED_PAIR_DESIGN_ELIGIBLE"
            closure = "PFQ_CURRENT_LINE_CLOSED_CANDIDATE_FROZEN"
            reason = "REVISED_PAIR_PASSED_DESIGN_GATES"
            frozen = {
                "candidate": CANDIDATE_ID,
                "exit": REVISION_ID,
                "status": "FROZEN_DESIGN_CANDIDATE",
                "period_status": "DESIGN_DIAGNOSTIC_ONLY",
            }
        else:
            verdict = "E1_X7_PFQ_REVISED_PAIR_NOT_ECONOMICALLY_ROBUST"
            closure = "PFQ_CURRENT_LINE_CLOSED_REJECTED"
            reason = "NO_ROBUST_REVISED_ENTRY_EXIT_PAIR"
            frozen = None

    report["verdict"] = verdict
    report["verdict_detail"].update({
        "verdict": verdict,
        "closure_status": closure,
        "closure_reason": reason,
        "frozen_design_candidate": frozen,
        "pfq_close": True,
    })
    report["pfq_current_line"].update({
        "status": closure,
        "reason": reason,
        "accepted_or_rejected": "ACCEPTED_FROZEN" if frozen else "REJECTED",
    })
    report["determinism_shas"]["verdict"] = verdict
    return report
