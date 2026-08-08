"""Orchestrator for EXIT Gate Reconciliation."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_provisional.util import sha256_file, sha256_obj

from . import ANALYSIS_ID, CANDIDATE_ID, PAIRS, SOURCE_BRIDGE_RUN, SOURCE_VERDICT
from .evaluate import (
    check_identity,
    combined_reference,
    decide_verdict,
    evaluate_pair,
    gate_pair,
    reference_expectation_delta,
)
from .precommit import build_precommit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[4]
SOURCE_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_bridge_v2"
PUBLISH = NATIVE / "results" / "research" / "e1_x7_pfq_exit_gate"


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "pfq_entry_changed": False,
        "pfq_exit_changed": False,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
        "exit_revision_implemented": False,
    }


def run_once(*, label: str = "A", ab_ok_placeholder: bool = True) -> dict[str, Any]:
    run_id = f"e1x7_pfq_exit_gate_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"

    report_path = SOURCE_DIR / "report.json"
    audit_path = SOURCE_DIR / "audit.xlsx"
    source_report_sha = sha256_file(report_path)
    source_audit_sha = sha256_file(audit_path)

    print(f"=== [{label}] Precommit ===", flush=True)
    precommit = build_precommit(
        source_report_sha=source_report_sha,
        source_audit_sha=source_audit_sha,
    )

    src_report = json.loads(report_path.read_text(encoding="utf-8"))
    assert src_report.get("run_id") == SOURCE_BRIDGE_RUN or SOURCE_BRIDGE_RUN in str(src_report.get("run_id"))
    # entry support from bridge
    entry_support = (
        ((src_report.get("verdict_detail") or {}).get("entry_support") or {})
        .get(CANDIDATE_ID, {})
        .get("supported")
        is True
    )
    assert src_report.get("verdict") == SOURCE_VERDICT

    print(f"=== [{label}] Load frozen JointTrades / FixedGrid / Counterfactual ===", flush=True)
    trades = _load_sheet(audit_path, "JointTrades")
    fixed_grid = _load_sheet(audit_path, "FixedGridOutcome")
    counterfactual = _load_sheet(audit_path, "Counterfactual")
    fg_by = {r["episode_id"]: r for r in fixed_grid}
    cf_by = {(r["pair_id"], r["episode_id"]): r for r in counterfactual}

    identity = check_identity(trades)
    identity_sha = sha256_obj({
        "unique_episode_n": identity["unique_episode_n"],
        "progress": identity["progress_trade_rows"],
        "protect": identity["protect_trade_rows"],
        "dups": identity["duplicate_within_pair"],
    })
    if not identity["ok"]:
        return {
            "analysis_id": ANALYSIS_ID,
            "run_id": run_id,
            "label": label,
            "verdict": "E1_X7_PFQ_EXIT_GATE_IDENTITY_MISMATCH",
            "identity": identity,
            "precommit": precommit,
            "safety": _safety(),
            "stop": True,
            "determinism_shas": {"identity_sha": identity_sha, "verdict": "E1_X7_PFQ_EXIT_GATE_IDENTITY_MISMATCH"},
        }

    print(f"=== [{label}] Pair evaluation ===", flush=True)
    pair_results = {}
    for pid in PAIRS:
        pair_results[pid] = evaluate_pair(
            trades, pair_id=pid, fixed_grid_by_eid=fg_by, cf_by_key=cf_by,
        )

    # Gate with ab_ok_placeholder; final Gate re-applied after A/B in main
    gates = {
        pid: gate_pair(
            pair_results[pid],
            entry_path_support=entry_support,
            identity_ok=True,
            ab_ok=ab_ok_placeholder,
        )
        for pid in PAIRS
    }
    verdict_detail = decide_verdict(gates, pair_results)
    combined = combined_reference(pair_results)
    deltas = reference_expectation_delta(pair_results)

    denom_sha = sha256_obj({pid: pair_results[pid]["denominator_episode_ids"] for pid in PAIRS})
    repair_sha = sha256_obj({pid: pair_results[pid]["repairable_episode_ids"] for pid in PAIRS})
    mech_sha = sha256_obj({
        pid: sorted((r["episode_id"], r["mechanism"]) for r in pair_results[pid]["repairable_rows"])
        for pid in PAIRS
    })
    gate_sha = sha256_obj({pid: gates[pid]["pass"] for pid in PAIRS})

    # strip heavy rows from public pair summary but keep for sheets
    pair_public = {}
    for pid, pr in pair_results.items():
        pair_public[pid] = {k: v for k, v in pr.items() if k not in ("repairable_rows", "denominator_episode_ids", "repairable_episode_ids")}
        pair_public[pid]["denominator_episode_ids_sha"] = sha256_obj(pr["denominator_episode_ids"])
        pair_public[pid]["repairable_episode_ids_sha"] = sha256_obj(pr["repairable_episode_ids"])

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "source_bridge_run": SOURCE_BRIDGE_RUN,
        "source_verdict": SOURCE_VERDICT,
        "source_report_sha256": source_report_sha,
        "source_audit_sha256": source_audit_sha,
        "precommit": precommit,
        "identity": identity,
        "entry_path_support": {CANDIDATE_ID: entry_support},
        "pair_results": pair_public,
        "pair_gates": gates,
        "combined_reference": combined,
        "expectation_delta": deltas,
        "verdict_detail": verdict_detail,
        "verdict": verdict_detail["verdict"],
        "determinism_shas": {
            "identity_sha": identity_sha,
            "pair_denominator_sha": denom_sha,
            "pair_repairable_sha": repair_sha,
            "pair_mechanism_sha": mech_sha,
            "pair_gate_result": {pid: gates[pid]["pass"] for pid in PAIRS},
            "selected_baseline": verdict_detail.get("selected_baseline_pair"),
            "verdict": verdict_detail["verdict"],
        },
        "safety": _safety(),
        "stop": True,
        "_sheets": {
            "PairEpisodes": [
                {
                    "pair_id": t["pair_id"],
                    "episode_id": t["episode_id"],
                    "day": t.get("day"),
                    "symbol": t.get("symbol"),
                    "exit_reason": t.get("exit_reason"),
                    "exit_net_pnl_bps": t.get("exit_net_pnl_bps"),
                    "integrity_status": t.get("integrity_status"),
                }
                for t in trades
                if t.get("candidate_id") == CANDIDATE_ID and t.get("pair_id") in PAIRS
            ],
            "Denominators": [
                {"pair_id": pid, "episode_id": eid}
                for pid in PAIRS
                for eid in pair_results[pid]["denominator_episode_ids"]
            ],
            "Repairable": [
                r for pid in PAIRS for r in pair_results[pid]["repairable_rows"]
            ],
            "FailureMechanisms": [
                {
                    "pair_id": pid,
                    "mechanism": m,
                    "n": n,
                    "fraction_of_repairable": n / pair_results[pid]["repairable_n"]
                    if pair_results[pid]["repairable_n"] else 0.0,
                }
                for pid in PAIRS
                for m, n in (pair_results[pid]["failure_mechanism_counts"] or {}).items()
            ],
        },
    }
    return report


def reapply_gates_with_ab(report: dict[str, Any], *, ab_ok: bool) -> dict[str, Any]:
    """Recompute Gate/verdict after A/B known (ab flag only)."""
    # Reconstruct minimal pair_results from public + sheets
    repair_by_pair: dict[str, list] = {pid: [] for pid in PAIRS}
    for r in (report.get("_sheets") or {}).get("Repairable") or []:
        repair_by_pair[r["pair_id"]].append(r)
    denom_by_pair: dict[str, list] = {pid: [] for pid in PAIRS}
    for r in (report.get("_sheets") or {}).get("Denominators") or []:
        denom_by_pair[r["pair_id"]].append(r["episode_id"])

    pair_results = {}
    for pid in PAIRS:
        pub = report["pair_results"][pid]
        rows = repair_by_pair[pid]
        pair_results[pid] = {
            **pub,
            "repairable_rows": rows,
            "denominator_episode_ids": denom_by_pair[pid],
            "repairable_episode_ids": [r["episode_id"] for r in rows],
        }

    entry_support = report["entry_path_support"][CANDIDATE_ID]
    gates = {
        pid: gate_pair(
            pair_results[pid],
            entry_path_support=entry_support,
            identity_ok=report["identity"]["ok"],
            ab_ok=ab_ok,
        )
        for pid in PAIRS
    }
    verdict_detail = decide_verdict(gates, pair_results)
    report["pair_gates"] = gates
    report["verdict_detail"] = verdict_detail
    report["verdict"] = verdict_detail["verdict"]
    report["determinism_shas"]["pair_gate_result"] = {pid: gates[pid]["pass"] for pid in PAIRS}
    report["determinism_shas"]["selected_baseline"] = verdict_detail.get("selected_baseline_pair")
    report["determinism_shas"]["verdict"] = verdict_detail["verdict"]
    return report
