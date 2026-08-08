"""E1_X7 PFQ study orchestrator — Phase 0 → design joint; stop before Shadow."""
from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.failure_source.clusters import load_episodes
from research.e1_x7_pfq import DOCUMENT_ID, FAMILY_ID, SOURCE_RUN, STUDY_TYPE
from research.e1_x7_pfq.candidates import (
    assert_registry_max_three,
    candidate_registry,
    derive_thresholds,
    passes_candidate,
)
from research.e1_x7_pfq.config import (
    CANDIDATES,
    DAYS,
    EXIT_CANDIDATES,
    EXIT_THRESHOLDS,
    PERIOD_STATUS,
    PROSPECTIVE_MIN,
    p1_body,
)
from research.e1_x7_pfq.feature_contract import FEATURE_CONTRACT_DOC, run_phase0_audit
from research.e1_x7_pfq.joint import build_path_points, evaluate_reachability, path_diagnosis, replay_pair

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
V3_STORE = Path.home() / "e1x6_research_store" / "taer" / "e1x6_taer_fsa_v3_20260804_043255"
V4_STORE = Path.home() / "e1x6_research_store" / "taer" / SOURCE_RUN


def _annotate_sources() -> dict:
    body = {
        "source_analysis": "E1_X6_TAER_FAILURE_SOURCE_ANALYSIS_V4",
        "source_run": SOURCE_RUN,
        "pullback_verdict": "TAER_PULLBACK_NEW_FAMILY_HYPOTHESIS_SUPPORTED",
        "range_verdict": "TAER_RANGE_STABLE_FEATURES_MODEL_SUPPORT_INSUFFICIENT",
        "range_status": "STABLE_FEATURES_FOUND_MODEL_SUPPORT_INSUFFICIENT_REFERENCE_ONLY",
        "taer_v1_unchanged": True,
        "taer_v1_family_status": "CLOSED_NO_ROBUST_PAIR",
        "not_taer_v2": True,
        "annotated_at_jst": datetime.now(JST).isoformat(),
    }
    fp = V4_STORE / "E1_X7_SOURCE_POINTER.json"
    if not fp.exists():
        V4_STORE.mkdir(parents=True, exist_ok=True)
        fp.write_text(json.dumps(body, indent=2), encoding="utf-8")
    return body


def _load_pullback_universe() -> list[dict]:
    labels = [json.loads(l) for l in (V3_STORE / "labels.jsonl").open(encoding="utf-8") if l.strip()]
    feats = {}
    with (V3_STORE / "features.jsonl").open(encoding="utf-8") as f:
        for line in f:
            d = json.loads(line)
            feats[d["episode_id"]] = d
    rows = []
    for lb in labels:
        if lb.get("setup_type") != "PULLBACK_RECLAIM":
            continue
        f = feats.get(lb["episode_id"]) or {}
        rows.append({
            **f,
            **{k: lb[k] for k in lb if k not in f or k in (
                "net_plus_5bps", "opportunity_target_valid", "cluster_id", "best_net_pnl_bps_300s"
            )},
            "decision_time": f.get("decision_time") or lb.get("entry_time"),
            "entry_time": f.get("decision_time") or lb.get("entry_time"),
            "setup_type": "PULLBACK_RECLAIM",
            "cluster_id": lb.get("cluster_id") or f.get("cluster_id"),
            "episode_id": lb["episode_id"],
            "day": lb["day"],
            "symbol": lb["symbol"],
            "session": f.get("session") or lb.get("session"),
            "net_plus_5bps": lb.get("net_plus_5bps"),
        })
    return rows


def _run_tests() -> dict:
    env = dict(os.environ)
    env["PYTHONPATH"] = f"{NATIVE / 'src'};{NATIVE / 'research'}"
    env["PYTHONIOENCODING"] = "utf-8"
    proc = subprocess.run(
        [sys.executable, "-m", "pytest", "-q", "-rA", "--tb=line",
         "-p", "no:cacheprovider",
         str(NATIVE / "tests" / "test_e1_x7_pfq.py")],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(NATIVE), env=env, timeout=300,
    )
    rows = []
    for line in (proc.stdout or "").splitlines():
        for st in ("PASSED", "FAILED", "ERROR"):
            if line.strip().startswith(st + " "):
                rows.append({"test": line.strip().split(" ", 1)[1].split(" - ")[0], "outcome": st})
                break
    return {
        "exit_code": proc.returncode,
        "passed": sum(1 for r in rows if r["outcome"] == "PASSED"),
        "failed": sum(1 for r in rows if r["outcome"] != "PASSED"),
        "total": len(rows),
        "rows": rows,
        "tail": (proc.stdout or "")[-2000:],
    }


def _enrich_entries_with_paths(
    audits: list[dict],
    universe: list[dict],
    events_by_day: dict,
    episodes_by_id: dict,
) -> list[dict]:
    uni_by = {u["episode_id"]: u for u in universe}
    out = []
    for a in audits:
        u = uni_by.get(a["episode_id"]) or {}
        ep = episodes_by_id.get(a["episode_id"]) or {}
        entry_t = float(a.get("decision_time") or u.get("decision_time"))
        entry_ask = float(u.get("entry_ask") or (ep.get("entry_ask") if ep else 0) or 0)
        # entry ask from feature snap ask if needed
        if entry_ask <= 0:
            # will set during path from first event
            entry_ask = None
        reclaim = float((ep.get("anchor") or {}).get("reference_high") or u.get("cross_magnitude_bps") or 0) or None
        # better reclaim from path_ledger head — use anchor reference if present
        if ep.get("anchor"):
            reclaim = float(ep["anchor"].get("reference_high") or entry_ask or 0)
        pb = (ep.get("anchor") or {}).get("pullback_low")
        if pb is None:
            pb = u.get("pullback_low")
        day = a["day"]
        sym = a["symbol"]
        session = a.get("session") or u.get("session")
        points, complete, censor = [], False, None
        if entry_ask and entry_ask > 0:
            points, complete, censor = build_path_points(
                events_by_day[day],
                sym=sym,
                entry_t=entry_t,
                entry_ask=entry_ask,
                session=session,
                reclaim_level=float(reclaim or entry_ask),
                pullback_low=pb,
            )
        elif events_by_day.get(day):
            # resolve entry ask = first ask at/after entry
            for t, s, row in events_by_day[day]:
                if s == sym and float(t) >= entry_t - 1e-12:
                    entry_ask = float(row["ask"])
                    reclaim = float(reclaim or entry_ask)
                    points, complete, censor = build_path_points(
                        events_by_day[day],
                        sym=sym, entry_t=entry_t, entry_ask=entry_ask,
                        session=session, reclaim_level=reclaim, pullback_low=pb,
                    )
                    break
        diag = path_diagnosis(points)
        out.append({
            **a,
            **{k: u.get(k) for k in ("cluster_id", "symbol", "day")},
            "entry_time": entry_t,
            "entry_ask": entry_ask,
            "entry_best_ask": entry_ask,
            "reclaim_level": reclaim,
            "pullback_low": pb,
            "path_complete": complete,
            "censor_reason": censor,
            "path_n": len(points),
            "path_diagnosis": diag,
            "path_head": points[:2],
            "path_tail": points[-2:] if points else [],
        })
    return out


def run() -> dict:
    run_id = f"e1x7_pfq_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "e1_x7_pfq" / run_id
    store.mkdir(parents=True, exist_ok=True)

    print("=== Source pointers (no overwrite of V4/TAER) ===", flush=True)
    src = _annotate_sources()

    print("=== Tests ===", flush=True)
    tests = _run_tests()
    print("tests", tests["passed"], "/", tests["total"], "exit", tests["exit_code"], flush=True)
    if tests["exit_code"] != 0:
        report = {"verdict": "E1_X7_TESTS_FAILED", "run_id": run_id, "tests": tests}
        _publish(report, store, None)
        return report

    print("=== Preload events ===", flush=True)
    sm = load_source_manifest()
    events_by_day = {}
    for day in DAYS:
        print("  preload", day, flush=True)
        events_by_day[day] = load_day_events(day, _universe_from_manifest(sm, day))

    print("=== Phase 0 Feature Semantic Contract ===", flush=True)
    universe = _load_pullback_universe()
    print("pullback universe", len(universe), flush=True)
    episodes_raw, _, _ = load_episodes()
    ep_by = {e["episode_id"]: e for e in episodes_raw}

    audits, phase0 = run_phase0_audit(universe, events_by_day, ep_by)
    (store / "feature_contract.json").write_text(
        json.dumps({"doc": FEATURE_CONTRACT_DOC, "summary": phase0}, indent=2, default=str),
        encoding="utf-8",
    )
    print("Phase0", phase0["status"], "valid_rate", phase0["ratio_valid_rate"], flush=True)
    if phase0["status"] != "PFQ_FEATURE_CONTRACT_PASS":
        report = {
            "phase": "E1_X7_PFQ",
            "document_id": DOCUMENT_ID,
            "family_id": FAMILY_ID,
            "run_id": run_id,
            "verdict": "E1_X7_FEATURE_CONTRACT_INVALID",
            "phase0": phase0,
            "source": src,
            "tests": tests,
            "safety": _safety(),
            "stop": True,
        }
        _publish(report, store, {"audits": audits})
        return report

    # Enrich with entry ask from episodes
    for a in audits:
        ep = ep_by.get(a["episode_id"]) or {}
        # entry from path ledger observations
        u = next((x for x in universe if x["episode_id"] == a["episode_id"]), {})
        a["entry_ask"] = u.get("entry_ask")
        if a.get("entry_ask") is None:
            # from load_entry style — features don't have entry_ask; resolve later
            pass

    # Need entry_ask: from TAER path_ledger via load_entry_observations filtered
    from research.e1_x6_taer.exit_joint_audit import load_entry_observations
    entries_all, _ = load_entry_observations()
    entry_by = {e["episode_id"]: e for e in entries_all if e["setup_type"] == "PULLBACK_RECLAIM"}
    for a in audits:
        e = entry_by.get(a["episode_id"])
        if e:
            a["entry_ask"] = e["entry_ask"]
            a["reclaim_level"] = e.get("reclaim_level")
            a["pullback_low"] = e.get("pullback_low")
            a["entry_time"] = e["entry_t"]
            a["decision_time"] = e["entry_t"]

    print("=== P1 ENTRY PRECOMMIT (before economics) ===", flush=True)
    thr = derive_thresholds(audits)
    registry = candidate_registry(thr)
    assert_registry_max_three(registry)
    feature_contract_sha = sha256_obj({"doc": FEATURE_CONTRACT_DOC, "phase0_status": phase0["status"]})
    p1 = p1_body(feature_contract_sha=feature_contract_sha, thresholds=thr, registry=registry)
    p1["precommit_at_jst"] = datetime.now(JST).isoformat()
    p1["P1_ENTRY_PRECOMMIT_sha256"] = sha256_obj(p1)
    (store / "p1_entry_precommit.json").write_text(json.dumps(p1, indent=2), encoding="utf-8")
    print("P1 sha", p1["P1_ENTRY_PRECOMMIT_sha256"], "thr", thr, flush=True)

    print("=== Paths + Reachability ===", flush=True)
    enriched = _enrich_entries_with_paths(audits, universe, events_by_day, ep_by)
    # fix entry fields from entry_by
    for e in enriched:
        src_e = entry_by.get(e["episode_id"])
        if src_e:
            e["entry_time"] = src_e["entry_t"]
            e["entry_ask"] = src_e["entry_ask"]
            e["entry_best_ask"] = src_e["entry_ask"]
            e["reclaim_level"] = src_e.get("reclaim_level")
            e["pullback_low"] = src_e.get("pullback_low")

    cand_entries = {}
    reach = {}
    for cid in CANDIDATES:
        selected = [e for e in enriched if passes_candidate(e, cid, thr)]
        # rebuild paths if needed with correct entry_ask
        for e in selected:
            if e.get("entry_ask") and (not e.get("path_n")):
                pts, comp, cen = build_path_points(
                    events_by_day[e["day"]],
                    sym=e["symbol"], entry_t=float(e["entry_time"]),
                    entry_ask=float(e["entry_ask"]), session=e["session"],
                    reclaim_level=float(e.get("reclaim_level") or e["entry_ask"]),
                    pullback_low=e.get("pullback_low"),
                )
                e["path_complete"] = comp
                e["path_n"] = len(pts)
                e["path_diagnosis"] = path_diagnosis(pts)
        # ensure path_complete computed
        for e in selected:
            if e.get("path_n") is None or e.get("path_n") == 0:
                pts, comp, cen = build_path_points(
                    events_by_day[e["day"]],
                    sym=e["symbol"], entry_t=float(e["entry_time"]),
                    entry_ask=float(e["entry_ask"]), session=e["session"],
                    reclaim_level=float(e.get("reclaim_level") or e["entry_ask"]),
                    pullback_low=e.get("pullback_low"),
                )
                e["path_complete"] = comp
                e["censor_reason"] = cen
                e["path_n"] = len(pts)
                e["path_diagnosis"] = path_diagnosis(pts)
                e["path_head"] = pts[:2]
                e["path_tail"] = pts[-2:] if pts else []
        cand_entries[cid] = selected
        reach[cid] = evaluate_reachability(selected, candidate_id=cid)
        print("reach", cid, reach[cid]["status"], "n", len(selected), flush=True)

    reachable_ids = [cid for cid, r in reach.items() if r["reachable"]]
    if not reachable_ids:
        report = {
            "phase": "E1_X7_PFQ",
            "document_id": DOCUMENT_ID,
            "run_id": run_id,
            "verdict": "E1_X7_NO_REACHABLE_ENTRY_CANDIDATE",
            "phase0": phase0,
            "p1_sha": p1["P1_ENTRY_PRECOMMIT_sha256"],
            "thresholds": thr,
            "reachability": reach,
            "source": src,
            "tests": tests,
            "safety": _safety(),
            "stop": True,
        }
        _publish(report, store, {"audits": audits, "reach": reach})
        return report

    print("=== P2 EXIT PRECOMMIT (before pair PnL) ===", flush=True)
    # seal EXIT thresholds from path distributions (build-only)
    all_diag = []
    for cid in reachable_ids:
        for e in cand_entries[cid]:
            d = e.get("path_diagnosis") or {}
            if d.get("time_to_plus_5bps_sec") is not None:
                all_diag.append(float(d["time_to_plus_5bps_sec"]))
    # keep precommitted EXIT_THRESHOLDS (already fixed); document derivation
    p2 = {
        "precommit_type": "P2_EXIT_PRECOMMIT",
        "document_id": DOCUMENT_ID,
        "family_id": FAMILY_ID,
        "precommit_at_jst": datetime.now(JST).isoformat(),
        "exit_candidates": list(EXIT_CANDIDATES),
        "thresholds": EXIT_THRESHOLDS,
        "threshold_derivation": {
            "method": "precommitted_structural_defaults_plus_path_support_check",
            "progress_deadline_sec_ref_median_time_to_plus5": (
                sorted(all_diag)[len(all_diag) // 2] if all_diag else None
            ),
            "no_pair_pnl_selection": True,
        },
        "states": {
            "PFQ_X_PROGRESS_STRUCT": [
                "OPEN_INIT", "STRUCTURE_HOLD", "PROGRESS_CHECK", "EXIT"
            ],
            "PFQ_X_PROTECT": [
                "OPEN_INIT", "STRUCTURE_HOLD", "COST_COVERED", "PROFIT_PROTECTION", "EXIT"
            ],
        },
        "priority": ["HARD_STOP", "PULLBACK_LOW_BREAK", "RECLAIM_LEVEL_BREAK", "MAX_HOLD", "STRATEGY"],
        "cost_contract": {"bps_once": 5.0, "lot": 100},
        "max_hold_sec": 300,
        "exit_reason_registry": [
            "HARD_STOP", "PULLBACK_LOW_BREAK", "RECLAIM_LEVEL_BREAK", "MAX_HOLD",
            "NO_PROGRESS_UPDATE_DEAD", "MFE_GIVEBACK", "UPDATE_DETERIORATION",
            "RECLAIM_LEVEL_LOSS", "SESSION_END", "STREAM_END",
        ],
        "fallback": "MAX_HOLD_OR_SESSION_END",
        "economics_opened_before_precommit": False,
    }
    p2["P2_EXIT_PRECOMMIT_sha256"] = sha256_obj(p2)
    (store / "p2_exit_precommit.json").write_text(json.dumps(p2, indent=2), encoding="utf-8")
    print("P2 sha", p2["P2_EXIT_PRECOMMIT_sha256"], flush=True)

    print("=== Design-period Joint Replay A ===", flush=True)
    pairs_a = {}
    for cid in reachable_ids:
        for xc in EXIT_CANDIDATES:
            print("  pair", cid, xc, "n_entries", len(cand_entries[cid]), flush=True)
            pairs_a[f"{cid}|{xc}"] = replay_pair(
                cand_entries[cid], candidate_id=cid, exit_candidate=xc, events_by_day=events_by_day,
            )

    print("=== Design-period Joint Replay B ===", flush=True)
    pairs_b = {}
    for cid in reachable_ids:
        for xc in EXIT_CANDIDATES:
            pairs_b[f"{cid}|{xc}"] = replay_pair(
                cand_entries[cid], candidate_id=cid, exit_candidate=xc, events_by_day=events_by_day,
            )

    det = {"ab_ok": True, "detail": {}}
    for pid in pairs_a:
        a, b = pairs_a[pid], pairs_b[pid]
        match = (
            a["n_pass"] == b["n_pass"]
            and abs(float(a["pnl"]) - float(b["pnl"])) < 1e-6
            and a["exit_reason_counts"] == b["exit_reason_counts"]
        )
        det["detail"][pid] = {"match": match, "n_pass": a["n_pass"], "pnl": a["pnl"]}
        if not match:
            det["ab_ok"] = False

    # Prospective: only unused day check
    push_root = NATIVE / "data" / "push_jsonl"
    unused = []
    if push_root.exists():
        for d in sorted(push_root.iterdir()):
            if d.is_dir() and d.name.replace("-", "") not in DAYS and d.name.startswith("2026"):
                unused.append(d.name)
    prospective = {
        "status": "E1_X7_PENDING_PROSPECTIVE_EVIDENCE",
        "unused_days_found": unused,
        "unused_business_days_n": len(unused),
        "required": PROSPECTIVE_MIN,
        "note": "Design period frozen; prospective requires >=5 unused business days — not auto-started",
        "executed": False,
    }
    if len(unused) >= PROSPECTIVE_MIN["unused_business_days_min"]:
        prospective["status"] = "E1_X7_PENDING_PROSPECTIVE_EVIDENCE"
        prospective["note"] = "Unused days exist but prospective execution deferred to explicit approval after freeze review"
        prospective["executed"] = False

    # integrity / empty checks
    verdict = "E1_X7_DESIGN_PERIOD_COMPLETE_PENDING_PROSPECTIVE"
    if not det["ab_ok"]:
        verdict = "E1_X7_DETERMINISM_FAILURE"
    elif all(p["n_pass"] == 0 for p in pairs_a.values()):
        verdict = "E1_X7_ALL_PAIRS_EMPTY"
    elif len({json.dumps(p["exit_reason_counts"], sort_keys=True) for p in pairs_a.values()}) == 1 and len(pairs_a) > 1:
        # not necessarily failure — only if ledgers identical
        pass

    report = {
        "phase": "E1_X7_PFQ",
        "document_id": DOCUMENT_ID,
        "family_id": FAMILY_ID,
        "study_type": STUDY_TYPE,
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "verdict": verdict,
        "period_status": PERIOD_STATUS,
        "design_economics_status": "EXPLORATORY_DESIGN_DIAGNOSTIC_ONLY",
        "phase0": {k: v for k, v in phase0.items() if k != "contract_doc"},
        "feature_contract_status": phase0["status"],
        "p1_sha256": p1["P1_ENTRY_PRECOMMIT_sha256"],
        "p2_sha256": p2["P2_EXIT_PRECOMMIT_sha256"],
        "thresholds": thr,
        "candidate_registry": registry,
        "reachability": reach,
        "reachable_candidates": reachable_ids,
        "pairs": {
            pid: {k: v for k, v in p.items() if k != "trades"}
            for pid, p in pairs_a.items()
        },
        "determinism": det,
        "prospective": prospective,
        "source": src,
        "tests": tests,
        "safety": _safety(),
        "stop": True,
        "note": "No Shadow/Forward/Paper; no RANGE candidate; TAER V1 not resurrected",
    }

    # store pair trades
    for pid, p in pairs_a.items():
        safe = pid.replace("|", "__")
        (store / f"trades_{safe}.jsonl").write_text(
            "\n".join(json.dumps(t, default=str) for t in p["trades"]) + "\n", encoding="utf-8"
        )

    _publish(report, store, {
        "audits": audits,
        "enriched": enriched,
        "reach": reach,
        "pairs": pairs_a,
        "p1": p1,
        "p2": p2,
    })
    print("=== PUBLISHED ===", flush=True)
    print("run_id", run_id, "verdict", verdict, "ab_ok", det["ab_ok"], flush=True)
    print("STOP", flush=True)
    return report


def _safety() -> dict:
    return {
        "submit": 0, "cancel": 0, "live": 0,
        "mainline_changed": False,
        "taer_v1_unchanged": True,
        "range_candidate": False,
        "shadow": False, "forward": False, "paper": False, "discord": False,
        "taer_v2": False,
    }


def _publish(report: dict, store: Path, payload) -> None:
    out = NATIVE / "results" / "research" / "e1_x7_pfq"
    out.mkdir(parents=True, exist_ok=True)
    md = [
        f"# E1_X7 PFQ — {report.get('verdict')}",
        "",
        f"- document_id: `{DOCUMENT_ID}`",
        f"- family_id: `{FAMILY_ID}`",
        f"- run_id: `{report.get('run_id')}`",
        f"- feature_contract: `{report.get('feature_contract_status')}`",
        f"- period_status: `{PERIOD_STATUS}`",
        f"- design economics: `EXPLORATORY_DESIGN_DIAGNOSTIC_ONLY`",
        f"- ab_ok: {(report.get('determinism') or {}).get('ab_ok')}",
        f"- prospective: `{(report.get('prospective') or {}).get('status')}`",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Stop",
        "No Shadow auto-start. No RANGE candidate. TAER V1 unchanged.",
    ]
    if report.get("reachable_candidates"):
        md.append("")
        md.append("## Reachable candidates")
        for cid in report["reachable_candidates"]:
            md.append(f"- `{cid}`")
    if report.get("pairs"):
        md.append("")
        md.append("## Design pairs (diagnostic only)")
        for pid, p in report["pairs"].items():
            md.append(f"- `{pid}` n_pass={p.get('n_pass')} pnl={p.get('pnl')} pf={p.get('pf')}")

    (out / "report.md").write_text("\n".join(md), encoding="utf-8")
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    for row in (("document_id", DOCUMENT_ID), ("verdict", report.get("verdict")),
                ("run_id", report.get("run_id")), ("family_id", FAMILY_ID)):
        ws.append(list(row))

    def sheet(name, headers, rows):
        w = wb.create_sheet(name)
        w.append(headers)
        for r in rows:
            w.append([("" if v is None else v) for v in r])

    sheet("PrecommitP1", ["key", "value"], [
        ["p1_sha", report.get("p1_sha256")],
        ["thresholds", json.dumps(report.get("thresholds"))],
    ])
    sheet("FeatureContract", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("phase0") or {}).items()][:40])

    if payload and payload.get("audits"):
        fd = wb.create_sheet("FeatureDistributions")
        fd.append(["episode_id", "day", "symbol", "ratio_valid", "ratio", "classified_n",
                   "pu10", "ratio_1", "net_plus_5"])
        for a in payload["audits"][:5000]:
            fd.append([
                a.get("episode_id"), a.get("day"), a.get("symbol"), a.get("ratio_valid"),
                a.get("uptick_volume_ratio_30s"), a.get("classified_trade_count_30s"),
                a.get("price_update_count_10s"),
                a.get("uptick_volume_ratio_30s") is not None and abs(float(a.get("uptick_volume_ratio_30s") or -1) - 1) < 1e-12,
                a.get("net_plus_5bps"),
            ])
    else:
        sheet("FeatureDistributions", ["note"], [["no audits"]])

    sheet("CandidateRegistry", ["candidate_id", "rules"],
          [[c["candidate_id"], json.dumps(c["rules"])] for c in (report.get("candidate_registry") or [])])
    sheet("Reachability", ["candidate", "status", "n", "days", "max_day_share", "flow_valid"],
          [[cid, r.get("status"), r.get("gates", {}).get("entry_observation_episodes"),
            r.get("gates", {}).get("entry_days"), r.get("gates", {}).get("max_day_share"),
            r.get("gates", {}).get("flow_ratio_valid_rate")]
           for cid, r in (report.get("reachability") or {}).items()])

    if payload and payload.get("enriched"):
        eo = wb.create_sheet("EntryObservations")
        eo.append(["episode_id", "day", "symbol", "entry_time", "ratio_valid", "pu10", "ratio", "path_complete"])
        for e in payload["enriched"][:5000]:
            eo.append([e.get("episode_id"), e.get("day"), e.get("symbol"), e.get("entry_time"),
                       e.get("ratio_valid"), e.get("price_update_count_10s"),
                       e.get("uptick_volume_ratio_30s"), e.get("path_complete")])
        pl = wb.create_sheet("PathLedger")
        pl.append(["episode_id", "path_n", "complete", "t_pos", "t_plus5", "adverse", "no_progress"])
        for e in payload["enriched"][:5000]:
            d = e.get("path_diagnosis") or {}
            pl.append([e.get("episode_id"), e.get("path_n"), e.get("path_complete"),
                       d.get("time_to_net_positive_sec"), d.get("time_to_plus_5bps_sec"),
                       d.get("adverse_before_plus_5bps"), d.get("no_progress")])
    else:
        sheet("EntryObservations", ["note"], [["n/a"]])
        sheet("PathLedger", ["note"], [["n/a"]])

    sheet("PathSummary", ["note"], [["See PathLedger; scenarios diagnostic-only"]])
    sheet("PrecommitP2", ["key", "value"], [
        ["p2_sha", report.get("p2_sha256")],
        ["exits", json.dumps(EXIT_CANDIDATES)],
    ])
    sheet("ExitDefinitions", ["exit", "states"], [
        ["PFQ_X_PROGRESS_STRUCT", "OPEN_INIT>STRUCTURE_HOLD>PROGRESS_CHECK>EXIT"],
        ["PFQ_X_PROTECT", "OPEN_INIT>STRUCTURE_HOLD>COST_COVERED>PROFIT_PROTECTION>EXIT"],
    ])
    sheet("ExitTransitions", ["note"], [["See exit_sm.py; sealed in P2"]])
    sheet("JointPairs", ["pair_id", "n_pass", "pnl", "pf", "status"],
          [[pid, p.get("n_pass"), p.get("pnl"), p.get("pf"), p.get("period_status")]
           for pid, p in (report.get("pairs") or {}).items()])

    jt = wb.create_sheet("JointTrades")
    jt.append(["pair_id", "day", "symbol", "reason", "net", "integrity"])
    if payload and payload.get("pairs"):
        n = 0
        for pid, p in payload["pairs"].items():
            for t in p.get("trades") or []:
                if n >= 8000:
                    break
                if t.get("integrity_status") != "PASS":
                    continue
                jt.append([pid, t.get("day"), t.get("symbol"), t.get("exit_reason"),
                           t.get("net_pnl_yen"), t.get("integrity_status")])
                n += 1

    daily = []
    for pid, p in (report.get("pairs") or {}).items():
        for d, pnl in (p.get("day_pnl") or {}).items():
            daily.append([pid, d, pnl])
    sheet("Daily", ["pair_id", "day", "pnl"], daily)
    sheet("ProspectiveStatus", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("prospective") or {}).items()])
    sheet("Tests", ["test", "outcome"],
          [[r["test"], r["outcome"]] for r in (report.get("tests") or {}).get("rows", [])])
    sheet("Determinism", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("determinism") or {}).items()])
    sheet("Safety", ["key", "value"], [[k, v] for k, v in (report.get("safety") or {}).items()])
    sheet("ChangeLog", ["item", "note"], [
        ["source", SOURCE_RUN],
        ["family", "PFQ independent — not TAER V2"],
        ["range", "reference only; gate not relaxed"],
        ["stop", "no Shadow auto-start"],
    ])
    wb.save(out / "audit.xlsx")

    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (store / "published_shas.json").write_text(json.dumps(report["published"], indent=2), encoding="utf-8")


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    run()
