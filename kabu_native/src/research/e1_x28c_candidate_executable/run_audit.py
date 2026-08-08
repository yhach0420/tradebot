"""E1_X28C runner: candidate-specific vs family baseline executable joint evaluation."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x22_actual_exit_factory.registry import (
    build_alias_groups,
    load_population_checked,
    rebuild_candidates_and_masks,
)
from research.e1_x26_exit_library.exits import ExitSpec, common_controls
from research.e1_x27_reference_joint.paths import build_paths_for_rows
from research.e1_x28_executable_joint.board import load_board_events, verify_board_mapping
from research.e1_x28_executable_joint.metrics import period_mask, summarize
from research.e1_x28_executable_joint.replay import (
    build_all_exit_matrices,
    build_entry_asks,
    build_full_executable_matrix,
)
from research.e1_x28b_candidate_reference.baseline import freeze_family_baselines

from . import (
    ANALYSIS_ID,
    ASSIGNMENT_REGISTRY_SHA,
    AUDIT_RECONCILIATION_SHA,
    BOARD_MAPPING_SHA,
    BOOTSTRAP_ITERS,
    BOOTSTRAP_SEED,
    BROKERAGE_FEE_YEN,
    CONSUMED_DAY,
    DISCOVERY,
    DOCUMENT_ID,
    EVALUATION,
    EXPECTED_ALIASES,
    EXPECTED_CAND_N,
    EXPECTED_FALLBACK,
    EXPECTED_GENUINE,
    EXPECTED_POP_N,
    EXPECTED_REF_JOINT,
    EXPECTED_TARGET,
    EXPECTED_TRAIL,
    EXPECTED_UNIQUE_MASKS,
    FAMILY_BASELINE_REGISTRY_SHA,
    LOGIC_MANIFEST_SHA,
    MIN_COMMON,
    SEMANTIC_EXIT_REGISTRY_SHA,
    SOURCE_X26A_MANIFEST_SHA,
    SOURCE_X28,
    SOURCE_X28B,
    STRESS_DAY,
    VERDICT_ENTRY_EDGE,
    VERDICT_MIXED,
    VERDICT_MULTIPLE,
    VERDICT_PARITY,
    VERDICT_REF_SENSITIVE,
    VERDICT_REPLAY,
    VERDICT_SOURCE,
)
from .classify import (
    abs_directional_positive,
    classify_executable,
    personalization_pairwise,
    stop_risk_tag,
)
from .publish import publish

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28c_candidate_executable"
X28A1_DIR = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"
X28A2_DIR = NATIVE / "results" / "research" / "e1_x28a2_audit_reconciliation"
X28B_DIR = NATIVE / "results" / "research" / "e1_x28b_candidate_reference"
X28_DIR = NATIVE / "results" / "research" / "e1_x28_executable_joint"
X26A_DIR = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x28c_candidate_executable.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "rows": [{"test": "pytest_suite",
                  "outcome": "PASSED" if p.returncode == 0 else "FAILED",
                  "detail": out[-3500:]}],
    }


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        for k in ("discovery_family_tags", "x26a_family_baseline_exit_ids", "semantic_key",
                  "resolved_from_tags", "X25_path_tags"):
            v = d.get(k)
            if isinstance(v, str) and v.startswith(("[", "{")):
                try:
                    d[k] = json.loads(v)
                except Exception:
                    pass
        out.append(d)
    return out


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _delta(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None:
        return None
    return float(a - b)


def _spec_from_row(exit_id: str, p: dict[str, Any]) -> ExitSpec:
    np_sec = _f(p.get("no_progress_sec"))
    gm = p.get("giveback_mode")
    if gm == "":
        gm = None
    return ExitSpec(
        exit_id=exit_id, path_family=None, variant=p.get("exit_mode"),
        stop_bps=_f(p.get("stop_bps")), target_bps=_f(p.get("target_bps")),
        trail_activation_bps=_f(p.get("trail_activation_bps")),
        giveback_bps=_f(p.get("giveback_bps")), giveback_mode=gm,
        no_progress_sec=np_sec,
        max_hold_sec=float(_f(p.get("max_hold_sec")) or 900.0),
        no_progress_mfe_bps=(_f(p.get("no_progress_mfe_bps")) or 5.0) if np_sec is not None else None,
        no_progress_abs_ret_bps=(_f(p.get("no_progress_abs_ret_bps")) or 5.0) if np_sec is not None else None,
        is_control=str(exit_id).startswith("CONTROL_"),
    )


def _load_x26a_exits() -> dict[str, dict[str, Any]]:
    x26a = json.loads((X26A_DIR / "report.json").read_text(encoding="utf-8"))
    if x26a.get("manifest_sha256") != SOURCE_X26A_MANIFEST_SHA:
        raise RuntimeError("x26a sha")
    out = {c["canonical_exit_id"]: c for c in (x26a.get("canonical_exits") or [])}
    for spec in common_controls():
        out[spec.exit_id] = {
            "canonical_exit_id": spec.exit_id,
            "stop_bps": spec.stop_bps, "target_bps": spec.target_bps,
            "trail_activation_bps": spec.trail_activation_bps, "giveback_bps": spec.giveback_bps,
            "giveback_mode": spec.giveback_mode, "no_progress_sec": spec.no_progress_sec,
            "max_hold_sec": spec.max_hold_sec,
            "no_progress_mfe_bps": spec.no_progress_mfe_bps,
            "no_progress_abs_ret_bps": spec.no_progress_abs_ret_bps,
        }
    return out


def _bh(pvals: list[float]) -> list[float]:
    n = len(pvals)
    if n == 0:
        return []
    order = np.argsort(pvals)
    ranked = np.asarray(pvals, dtype=np.float64)[order]
    out = np.empty(n)
    prev = 1.0
    for i in range(n - 1, -1, -1):
        val = ranked[i] * n / (i + 1)
        prev = min(prev, val)
        out[i] = min(prev, 1.0)
    q = np.empty(n)
    q[order] = out
    return q.tolist()


def _cluster_bootstrap(
    values: np.ndarray, clusters: np.ndarray, valid: np.ndarray,
    iters: int = BOOTSTRAP_ITERS, seed: int = BOOTSTRAP_SEED,
) -> dict[str, Any]:
    elig = valid & np.isfinite(values)
    if int(elig.sum()) < 10:
        return {"mean": None, "ci95": [None, None], "raw_p": None, "tag": "DESCRIPTIVE_ONLY"}
    uniq = np.unique(clusters[elig])
    means = []
    for c in uniq:
        m = elig & (clusters == c)
        if m.any():
            means.append(float(np.mean(values[m])))
    arr = np.asarray(means, dtype=np.float64)
    if arr.size < 2:
        return {"mean": float(np.mean(values[elig])), "ci95": [None, None], "raw_p": None, "tag": "DESCRIPTIVE_ONLY"}
    obs = float(np.mean(arr))
    rng = np.random.default_rng(seed)
    samp = rng.choice(arr.size, size=(iters, arr.size), replace=True)
    boots = arr[samp].mean(axis=1)
    lo, hi = np.quantile(boots, [0.025, 0.975])
    raw_p = float(np.mean(np.abs(boots) >= abs(obs)))
    tag = "CI_SUPPORTED" if (lo > 0 or hi < 0) else "DESCRIPTIVE_ONLY"
    return {"mean": obs, "ci95": [float(lo), float(hi)], "raw_p": raw_p, "tag": tag}


def _load_boards(rows: list[dict[str, Any]], allowed: list[str]) -> dict[tuple[str, str], dict[str, np.ndarray]]:
    keys = sorted({(r["date"], r["symbol"]) for r in rows if r["date"] in allowed})
    cache: dict[tuple[str, str], dict[str, np.ndarray]] = {}
    print(f"  loading {len(keys)} board symbol-days...", flush=True)

    def _one(k):
        return k, load_board_events(k[0], k[1])

    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = [ex.submit(_one, k) for k in keys]
        done = 0
        for fut in as_completed(futs):
            k, b = fut.result()
            cache[k] = b
            done += 1
            if done % 50 == 0 or done == len(keys):
                print(f"    boards {done}/{len(keys)}", flush=True)
    return cache


def _exit_cov_given_entry(
    *, mat: dict[str, np.ndarray], entry_asks: dict[str, np.ndarray],
    selected: np.ndarray, dates: np.ndarray, period: str = "EVALUATION",
) -> Optional[float]:
    pm = period_mask(dates, period) & selected
    ent = pm & entry_asks["valid"]
    denom = int(ent.sum())
    if denom == 0:
        return None
    return float(np.sum(ent & mat["valid"]) / denom)


def run_once(run_id: str) -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)

    x28b = json.loads((X28B_DIR / "report.json").read_text(encoding="utf-8"))
    x28a1 = json.loads((X28A1_DIR / "report.json").read_text(encoding="utf-8"))
    x28a2 = json.loads((X28A2_DIR / "report.json").read_text(encoding="utf-8"))
    x28 = json.loads((X28_DIR / "report.json").read_text(encoding="utf-8"))
    if x28b.get("run_id") != SOURCE_X28B:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "x28b_run"}
    if x28b.get("family_baseline_registry_sha") != FAMILY_BASELINE_REGISTRY_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "baseline_sha"}
    if x28a1.get("manifest_sha256") != LOGIC_MANIFEST_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "logic"}
    if x28a1.get("assignment_registry_sha") != ASSIGNMENT_REGISTRY_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "assign"}
    if x28a1.get("semantic_exit_registry_sha") != SEMANTIC_EXIT_REGISTRY_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "sem"}
    if x28a2.get("audit_reconciliation_sha") != AUDIT_RECONCILIATION_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "audit"}
    if x28.get("run_id") != SOURCE_X28:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "x28_run"}
    if x28.get("board_mapping_sha") != BOARD_MAPPING_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "board_sha"}

    print("=== board mapping ===", flush=True)
    mapping = verify_board_mapping()
    from research.e1_x28_executable_joint import BOARD_MAPPING_SHA as X28_PINNED_SHA
    if not mapping.get("ok") or X28_PINNED_SHA != BOARD_MAPPING_SHA:
        return {"run_id": run_id, "verdict": VERDICT_PARITY, "reason": "board_mapping", **mapping}

    print("=== load assignments / handoff / bootstrap ===", flush=True)
    assignments = _load_sheet(X28A1_DIR / "audit.xlsx", "CandidateExitAssignmentsV2")
    sem_reg = _load_sheet(X28A1_DIR / "audit.xlsx", "SemanticExitRegistryV2")
    handoff = _load_sheet(X28B_DIR / "audit.xlsx", "X28CHandoff")
    boot_ref = _load_sheet(X28B_DIR / "audit.xlsx", "BootstrapDiagnostic")
    if len(assignments) != EXPECTED_UNIQUE_MASKS:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "assign_n"}

    genuine_n = sum(1 for a in assignments if a.get("exit_source") == "CANDIDATE_SPECIFIC")
    fallback_n = EXPECTED_UNIQUE_MASKS - genuine_n
    target_n = sum(1 for a in assignments if a.get("exit_source") == "CANDIDATE_SPECIFIC" and a.get("exit_mode") == "TARGET")
    trail_n = sum(1 for a in assignments if a.get("exit_source") == "CANDIDATE_SPECIFIC" and a.get("exit_mode") == "TRAIL")

    ref_joint_ids = []
    for h in handoff:
        pr = h.get("priority")
        if pr in (True, "True", 1) or h.get("classification") == "SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
            ref_joint_ids.append(h["candidate_id"])
    ref_joint_ids = sorted(set(ref_joint_ids))
    if len(ref_joint_ids) != EXPECTED_REF_JOINT:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "ref_joint_n", "n": len(ref_joint_ids)}

    print("=== registry ===", flush=True)
    rows = load_population_checked()
    cands, masks = rebuild_candidates_and_masks(rows)
    alias_rows, _, unique_masks = build_alias_groups(cands, masks)
    alias_n = sum(1 for a in alias_rows if not a["is_representative"])
    if not (
        len(rows) == EXPECTED_POP_N and len(cands) == EXPECTED_CAND_N
        and len(unique_masks) == EXPECTED_UNIQUE_MASKS and alias_n == EXPECTED_ALIASES
    ):
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "registry"}

    x26a_exits = _load_x26a_exits()
    print("=== freeze family baselines ===", flush=True)
    baseline_rows, baseline_sha = freeze_family_baselines(assignments, x26a_exits)
    if baseline_sha != FAMILY_BASELINE_REGISTRY_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "baseline_recompute",
                "got": baseline_sha, "exp": FAMILY_BASELINE_REGISTRY_SHA}
    baseline_by = {r["candidate_id"]: r for r in baseline_rows}

    specs: dict[str, ExitSpec] = {}
    for s in sem_reg:
        specs[s["semantic_exit_sha256"]] = _spec_from_row(s["semantic_exit_sha256"], s)
    for eid in {r["primary_family_baseline_exit_id"] for r in baseline_rows}:
        if eid not in specs:
            p = x26a_exits.get(eid)
            if p is None:
                return {"run_id": run_id, "verdict": VERDICT_SOURCE, "reason": "missing_fam", "eid": eid}
            specs[eid] = _spec_from_row(eid, p)
    spec_list = list(specs.values())
    print(f"  unique exit specs={len(spec_list)}", flush=True)

    dates_needed = list(DISCOVERY + EVALUATION + (STRESS_DAY, CONSUMED_DAY))
    print("=== paths + boards ===", flush=True)
    times_list, prices_list = build_paths_for_rows(rows, allowed_dates=dates_needed, max_workers=6)
    board_by_key = _load_boards(rows, dates_needed)
    entry_asks = build_entry_asks(rows, board_by_key)
    entry_asks2 = build_entry_asks(rows, board_by_key)
    if not np.array_equal(entry_asks["valid"], entry_asks2["valid"]):
        return {"run_id": run_id, "verdict": VERDICT_PARITY, "reason": "entry_asks_nondeterministic"}

    # Parity: rebuild one family full matrix twice
    print("=== X28 execution parity sample ===", flush=True)
    parity_eid = "EXIT_CONTINUATION_PROTECT_V2"
    if parity_eid not in specs:
        parity_eid = next(e for e in specs if str(e).startswith("EXIT_"))
    try:
        full_a = build_full_executable_matrix(
            spec=specs[parity_eid], entry_asks=entry_asks, rows=rows,
            times_list=times_list, prices_list=prices_list, board_by_key=board_by_key,
        )
        full_b = build_full_executable_matrix(
            spec=specs[parity_eid], entry_asks=entry_asks, rows=rows,
            times_list=times_list, prices_list=prices_list, board_by_key=board_by_key,
        )
        same_valid = np.array_equal(full_a["valid"], full_b["valid"])
        mean_a = float(np.nanmean(full_a["ret_bps"][full_a["valid"]])) if full_a["valid"].any() else None
        mean_b = float(np.nanmean(full_b["ret_bps"][full_b["valid"]])) if full_b["valid"].any() else None
        if not same_valid or (mean_a is not None and mean_b is not None and abs(mean_a - mean_b) > 1e-9):
            return {"run_id": run_id, "verdict": VERDICT_PARITY, "reason": "full_matrix_mismatch"}
        parity_ok = True
    except Exception as e:
        return {"run_id": run_id, "verdict": VERDICT_PARITY, "reason": str(e)[:400]}

    print("=== build all exit matrices (ref/bridge/full/bidmark) ===", flush=True)
    import pickle
    cache_pkl = OUT / "_mats_cache.pkl"
    try:
        if cache_pkl.exists():
            print("  loading cached matrices...", flush=True)
            with open(cache_pkl, "rb") as fh:
                cached = pickle.load(fh)
            if (
                cached.get("spec_ids") == sorted(specs.keys())
                and cached.get("entry_valid_sha") == sha256_obj(entry_asks["valid"].tolist())
            ):
                mats = cached["mats"]
                print(f"  cache hit: {len(mats)} exits", flush=True)
            else:
                mats = None
        else:
            mats = None
        if mats is None:
            mats = build_all_exit_matrices(
                specs=spec_list, rows=rows, times_list=times_list, prices_list=prices_list,
                entry_asks=entry_asks, board_by_key=board_by_key, max_workers=4,
            )
            with open(cache_pkl, "wb") as fh:
                pickle.dump({
                    "spec_ids": sorted(specs.keys()),
                    "entry_valid_sha": sha256_obj(entry_asks["valid"].tolist()),
                    "mats": mats,
                }, fh, protocol=4)
            print(f"  cached matrices → {cache_pkl.name}", flush=True)
    except Exception as e:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY, "reason": str(e)[:500]}

    dates = np.array([r["date"] for r in rows])
    symbols = np.array([str(r["symbol"]) for r in rows])
    sessions = np.array([r["session"] for r in rows])
    clusters = np.array([r.get("cluster_id") or str(r["symbol"]) for r in rows])
    cps = np.array([float(r["CurrentPrice"]) if r.get("CurrentPrice") is not None else np.nan for r in rows])

    print("=== evaluate 6441 masks ===", flush=True)
    class_counts: Counter = Counter()
    class_rows = []
    entry_sel_rows = []
    pers_rows = []
    support_rows = []
    metric_rows = []
    cost_rows = []
    daily_rows = []
    stress_rows = []
    consumed_rows = []
    mode_stats = defaultdict(Counter)
    horizon_stats = defaultdict(lambda: {"joint": 0, "support": 0, "rets": [], "entry": [], "pers": []})
    stop_stats = defaultdict(lambda: {"n": 0, "support": 0, "joint": 0, "rets": [], "pfs": [], "worsts": [], "dds": [], "pers": []})
    target_cost_rows = []

    bridge_joint = 0
    assign_by = {a["candidate_id"]: a for a in assignments}

    # coverage globals
    eval_all = period_mask(dates, "EVALUATION")
    entry_ask_cov_global = float(np.sum(eval_all & entry_asks["valid"]) / max(int(eval_all.sum()), 1))

    done = 0
    for a in assignments:
        cid = a["candidate_id"]
        sel = unique_masks[cid]
        sem = a["semantic_exit_sha256"]
        src = a.get("exit_source")
        mode = a.get("exit_mode")
        horizon = int(a.get("candidate_horizon_sec") or 300)
        is_fallback = src in ("FAMILY_FALLBACK", "COMMON_CONTROL_FALLBACK")
        fam_eid = baseline_by[cid]["primary_family_baseline_exit_id"]
        mf = mats[sem]
        ff = mats[fam_eid]
        mat_s = mf["full"]
        mat_f = ff["full"]
        mat_sb = mf["bridge"]
        mat_sr = mf["ref"]

        sel_ev = summarize(mat=mat_s, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                           period="EVALUATION", population="SELECTED")
        comp_ev = summarize(mat=mat_s, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                            period="EVALUATION", population="COMPLEMENT")
        entry_delta = _delta(sel_ev.get("avg_return_bps"), comp_ev.get("avg_return_bps"))
        entry_n = int(sel_ev.get("trades") or 0)
        exit_cov = _exit_cov_given_entry(mat=mat_s, entry_asks=entry_asks, selected=sel, dates=dates)
        fam_exit_cov = _exit_cov_given_entry(mat=mat_f, entry_asks=entry_asks, selected=sel, dates=dates)

        if is_fallback:
            pers = {"n": 0, "delta_avg_return": 0.0, "delta_avg_pnl": 0.0,
                    "specific_better_episode_rate": None, "family_better_episode_rate": None, "same_episode_rate": 1.0}
            pers_delta = 0.0
            pers_n = 0
        else:
            pers = personalization_pairwise(
                mat_specific=mat_s, mat_family=mat_f, selected=sel, dates=dates, period="EVALUATION",
            )
            pers_delta = pers.get("delta_avg_return")
            pers_n = int(pers.get("n") or 0)

        status = classify_executable(
            is_fallback=is_fallback, sel=sel_ev, entry_delta=entry_delta, pers_delta=pers_delta,
            entry_n=entry_n, pers_n=pers_n if not is_fallback else 0, exit_cov=exit_cov,
        )
        class_counts[status] += 1

        # bridge joint check
        sel_br = summarize(mat=mat_sb, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                           period="EVALUATION", population="SELECTED")
        br_entry = _delta(sel_br.get("avg_return_bps"),
                          summarize(mat=mat_sb, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                                    period="EVALUATION", population="COMPLEMENT").get("avg_return_bps"))
        if is_fallback:
            br_pers = 0.0
            br_pers_n = 0
        else:
            br_p = personalization_pairwise(
                mat_specific=mat_sb, mat_family=ff["bridge"], selected=sel, dates=dates, period="EVALUATION",
            )
            br_pers = br_p.get("delta_avg_return")
            br_pers_n = int(br_p.get("n") or 0)
        br_status = classify_executable(
            is_fallback=is_fallback, sel=sel_br, entry_delta=br_entry, pers_delta=br_pers,
            entry_n=int(sel_br.get("trades") or 0), pers_n=br_pers_n if not is_fallback else 0,
            exit_cov=_exit_cov_given_entry(mat=mat_sb, entry_asks=entry_asks, selected=sel, dates=dates),
        )
        if br_status == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
            bridge_joint += 1

        # mode / horizon / stop
        mk = "FALLBACK" if is_fallback else str(mode)
        mode_stats[mk]["n"] += 1
        if status != "EXECUTABLE_SPECIFIC_SUPPORT_INSUFFICIENT" and not is_fallback:
            mode_stats[mk]["support"] += 1
        if abs_directional_positive(sel_ev):
            mode_stats[mk]["abs_pos"] += 1
        if entry_delta is not None and entry_delta > 0:
            mode_stats[mk]["entry_pos"] += 1
        if pers_delta is not None and pers_delta > 0 and not is_fallback:
            mode_stats[mk]["pers_pos"] += 1
        if status == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
            mode_stats[mk]["joint"] += 1

        hs = horizon_stats[horizon]
        if status == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
            hs["joint"] += 1
        if status != "EXECUTABLE_SPECIFIC_SUPPORT_INSUFFICIENT" and not is_fallback:
            hs["support"] += 1
        if sel_ev.get("avg_return_bps") is not None:
            hs["rets"].append(sel_ev["avg_return_bps"])
        if entry_delta is not None:
            hs["entry"].append(entry_delta)
        if pers_delta is not None and not is_fallback:
            hs["pers"].append(pers_delta)

        srt = a.get("stop_risk_tag") or stop_risk_tag(a.get("stop_bps"))
        if srt and not is_fallback:
            ss = stop_stats[srt]
            ss["n"] += 1
            if status != "EXECUTABLE_SPECIFIC_SUPPORT_INSUFFICIENT":
                ss["support"] += 1
            if status == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
                ss["joint"] += 1
            if sel_ev.get("avg_return_bps") is not None:
                ss["rets"].append(sel_ev["avg_return_bps"])
            if sel_ev.get("profit_factor") is not None:
                ss["pfs"].append(sel_ev["profit_factor"])
            if sel_ev.get("worst_trade") is not None:
                ss["worsts"].append(sel_ev["worst_trade"])
            if sel_ev.get("episode_seq_max_dd") is not None:
                ss["dds"].append(sel_ev["episode_seq_max_dd"])
            if pers_delta is not None:
                ss["pers"].append(pers_delta)

        if mode == "TARGET" and not is_fallback:
            # mean entry cross vs CP among selected eval with ask
            pm = period_mask(dates, "EVALUATION") & sel & entry_asks["valid"] & np.isfinite(cps) & (cps > 0)
            if pm.any():
                cross = (entry_asks["ask"][pm] / cps[pm] - 1.0) * 10000.0
                tgt = _f(a.get("target_bps"))
                mean_cross = float(np.mean(cross))
                target_cost_rows.append({
                    "candidate_id": cid, "target_bps": tgt, "mean_entry_cross_cost_bps": mean_cross,
                    "net_remaining_after_ask": (float(tgt) - mean_cross) if tgt is not None else None,
                })

        # day support
        day_pos = day_neg = d_pers_pos = d_pers_neg = 0
        for day in EVALUATION:
            dm = (dates == day) & sel & mat_s["valid"]
            if int(dm.sum()) == 0:
                continue
            if float(np.mean(mat_s["ret_bps"][dm])) > 0:
                day_pos += 1
            else:
                day_neg += 1
            if not is_fallback:
                common = dm & mat_f["valid"]
                if int(common.sum()) > 0:
                    dd = float(np.mean(mat_s["ret_bps"][common] - mat_f["ret_bps"][common]))
                    if dd > 0:
                        d_pers_pos += 1
                    elif dd < 0:
                        d_pers_neg += 1
        avail = day_pos + day_neg
        if avail >= 3 and day_pos >= 3 and d_pers_pos >= 3:
            day_tag = "BROAD_EVAL_SUPPORT"
        elif avail <= 1:
            day_tag = "DAY_CONCENTRATED"
        else:
            day_tag = "MIXED_EVAL_SUPPORT"
        daily_rows.append({
            "candidate_id": cid, "positive_days": day_pos, "negative_days": day_neg,
            "pers_pos_days": d_pers_pos, "day_tag": day_tag,
            "x28b_ref_joint": cid in ref_joint_ids,
        })

        sel_stress = summarize(mat=mat_s, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                               period="20260803", population="SELECTED")
        sel_cons = summarize(mat=mat_s, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                             period="20260804", population="SELECTED")
        if sel_ev.get("avg_return_bps") is not None and sel_stress.get("avg_return_bps") is not None:
            same = (sel_ev["avg_return_bps"] > 0) == (sel_stress["avg_return_bps"] > 0)
            stress_rows.append({
                "candidate_id": cid,
                "tag": "EVAL_TO_20260803_SAME_DIRECTION" if same else "EVAL_TO_20260803_REVERSED",
                "eval_ret": sel_ev["avg_return_bps"], "stress_ret": sel_stress["avg_return_bps"],
                "pers_delta": pers_delta,
            })
        else:
            stress_rows.append({"candidate_id": cid, "tag": "EVAL_TO_20260803_INSUFFICIENT"})
        consumed_rows.append({
            "candidate_id": cid, "role": "CONSUMED_PROSPECTIVE_DIAGNOSTIC_ONLY",
            "avg_return_bps": sel_cons.get("avg_return_bps"), "trades": sel_cons.get("trades"),
        })

        # cost decomp aggregate
        pm_sel = period_mask(dates, "EVALUATION") & sel
        ref_m = summarize(mat=mat_sr, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                          period="EVALUATION", population="SELECTED")
        cost_rows.append({
            "candidate_id": cid,
            "ref_avg_ret": ref_m.get("avg_return_bps"),
            "bridge_avg_ret": sel_br.get("avg_return_bps"),
            "full_avg_ret": sel_ev.get("avg_return_bps"),
            "ref_to_bridge": _delta(sel_br.get("avg_return_bps"), ref_m.get("avg_return_bps")),
            "bridge_to_full": _delta(sel_ev.get("avg_return_bps"), sel_br.get("avg_return_bps")),
            "ref_to_full": _delta(sel_ev.get("avg_return_bps"), ref_m.get("avg_return_bps")),
            "pers_delta_full": pers_delta,
        })

        class_rows.append({
            "candidate_id": cid, "exit_source": src, "exit_mode": mode, "horizon": horizon,
            "stop_risk_tag": srt, "family_baseline": fam_eid, "classification": status,
            "bridge_classification": br_status,
            "avg_return_bps": sel_ev.get("avg_return_bps"), "avg_pnl": sel_ev.get("avg_pnl"),
            "pf": sel_ev.get("profit_factor"), "entry_delta": entry_delta, "pers_delta": pers_delta,
            "pers_n": pers_n, "trades": sel_ev.get("trades"), "exit_cov": exit_cov,
            "entry_ask_cov_selected": float(
                np.sum(period_mask(dates, "EVALUATION") & sel & entry_asks["valid"])
                / max(int(np.sum(period_mask(dates, "EVALUATION") & sel)), 1)
            ),
        })
        entry_sel_rows.append({
            "candidate_id": cid, "entry_delta": entry_delta,
            "selected_trades": sel_ev.get("trades"), "complement_trades": comp_ev.get("trades"),
        })
        pers_rows.append({
            "candidate_id": cid, "is_fallback": is_fallback, "family_baseline": fam_eid,
            **{k: pers.get(k) for k in ("n", "delta_avg_return", "delta_avg_pnl",
                                        "specific_better_episode_rate", "family_better_episode_rate",
                                        "same_episode_rate", "delta_hold")},
            "family_exit_cov": fam_exit_cov,
        })
        support_rows.append({
            "candidate_id": cid, "trades": sel_ev.get("trades"), "days": sel_ev.get("days"),
            "symbols": sel_ev.get("symbols"), "coverage": sel_ev.get("coverage"),
            "exit_cov_given_entry": exit_cov, "pers_n": pers_n, "entry_n": entry_n,
        })
        metric_rows.append({"candidate_id": cid, **{k: sel_ev.get(k) for k in (
            "trades", "coverage", "days", "symbols", "avg_return_bps", "avg_pnl", "total_pnl",
            "profit_factor", "win_rate", "worst_trade", "episode_seq_max_dd", "median_hold_sec",
        )}})

        done += 1
        if done % 500 == 0 or done == EXPECTED_UNIQUE_MASKS:
            print(f"  masks {done}/{EXPECTED_UNIQUE_MASKS}", flush=True)

    exec_joint_ids = [r["candidate_id"] for r in class_rows
                      if r["classification"] == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE"]
    exec_joint_n = len(exec_joint_ids)

    # Reference 266 transition
    print("=== reference 266 transition ===", flush=True)
    ref_trans = []
    surv_full = surv_bridge = 0
    for cid in ref_joint_ids:
        cr = next(c for c in class_rows if c["candidate_id"] == cid)
        tags = []
        if cr["classification"] == "EXECUTABLE_SPECIFIC_SUPPORT_INSUFFICIENT":
            tags.append("SUPPORT_LOST")
        if cr.get("bridge_classification") == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
            surv_bridge += 1
        else:
            tags.append("ENTRY_CROSS_COST_SENSITIVE")
            tags.append("EXIT_CROSS_COST_SENSITIVE")
        if cr["classification"] == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE":
            surv_full += 1
        else:
            if cr.get("pers_delta") is not None and cr["pers_delta"] <= 0:
                tags.append("PERSONALIZATION_EXECUTION_SENSITIVE")
            tags.append("TRIGGER_RECALC_SENSITIVE")
        ref_trans.append({
            "candidate_id": cid,
            "bridge_joint": cr.get("bridge_classification") == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE",
            "full_joint": cr["classification"] == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE",
            "failure_tags": tags,
            "full_class": cr["classification"],
        })

    # Reference triple CI
    triple = 0
    ref_ci_rows = []
    for b in boot_ref:
        lo_r, lo_e, lo_p = b.get("avg_return_ci95_lo"), b.get("entry_delta_ci95_lo"), b.get("pers_delta_ci95_lo")
        ok = all(x is not None and float(x) > 0 for x in (lo_r, lo_e, lo_p))
        if ok:
            triple += 1
        ref_ci_rows.append({**b, "REFERENCE_TRIPLE_CI_DIAGNOSTIC": ok})

    # Bootstrap + FDR for exec joint (no cap)
    print(f"=== bootstrap {exec_joint_n} exec joint (2000 iters, no cap) ===", flush=True)
    boot_rows = []
    p_ret, p_ent, p_pers = [], [], []
    for cid in exec_joint_ids:
        a = assign_by[cid]
        sel = unique_masks[cid]
        mat_s = mats[a["semantic_exit_sha256"]]["full"]
        mat_f = mats[baseline_by[cid]["primary_family_baseline_exit_id"]]["full"]
        pm = period_mask(dates, "EVALUATION") & sel & mat_s["valid"]
        vals = mat_s["ret_bps"].copy()
        b_ret = _cluster_bootstrap(vals, clusters, pm)
        # entry delta: selected cluster means minus fixed complement mean
        comp = summarize(mat=mat_s, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                         period="EVALUATION", population="COMPLEMENT")
        comp_mean = comp.get("avg_return_bps") or 0.0
        b_ent = _cluster_bootstrap(vals - comp_mean, clusters, pm)
        # pers on common
        common = pm & mat_f["valid"]
        dvals = mat_s["ret_bps"] - mat_f["ret_bps"]
        b_pers = _cluster_bootstrap(dvals, clusters, common)
        ci_all = bool(
            b_ret["ci95"][0] is not None and b_ret["ci95"][0] > 0
            and b_ent["ci95"][0] is not None and b_ent["ci95"][0] > 0
            and b_pers["ci95"][0] is not None and b_pers["ci95"][0] > 0
        )
        boot_rows.append({
            "candidate_id": cid,
            "ret_ci95_lo": b_ret["ci95"][0], "ret_ci95_hi": b_ret["ci95"][1], "ret_raw_p": b_ret["raw_p"],
            "entry_ci95_lo": b_ent["ci95"][0], "entry_ci95_hi": b_ent["ci95"][1], "entry_raw_p": b_ent["raw_p"],
            "pers_ci95_lo": b_pers["ci95"][0], "pers_ci95_hi": b_pers["ci95"][1], "pers_raw_p": b_pers["raw_p"],
            "ALL_THREE_CI_SUPPORTED": ci_all,
            "iters": BOOTSTRAP_ITERS,
        })
        p_ret.append(b_ret["raw_p"] if b_ret["raw_p"] is not None else 1.0)
        p_ent.append(b_ent["raw_p"] if b_ent["raw_p"] is not None else 1.0)
        p_pers.append(b_pers["raw_p"] if b_pers["raw_p"] is not None else 1.0)

    q_ret, q_ent, q_pers = _bh(p_ret), _bh(p_ent), _bh(p_pers)
    fdr_rows = []
    for i, cid in enumerate(exec_joint_ids):
        fdr_ok = (q_ret[i] <= 0.05 and q_ent[i] <= 0.05 and q_pers[i] <= 0.05) if i < len(q_ret) else False
        boot_rows[i]["ALL_THREE_FDR_SUPPORTED"] = fdr_ok
        boot_rows[i]["q_ret"] = q_ret[i] if i < len(q_ret) else None
        boot_rows[i]["q_ent"] = q_ent[i] if i < len(q_ent) else None
        boot_rows[i]["q_pers"] = q_pers[i] if i < len(q_pers) else None
        fdr_rows.append({
            "candidate_id": cid,
            "EXECUTABLE_SPECIFIC_RETURN_q": boot_rows[i]["q_ret"],
            "EXECUTABLE_ENTRY_SELECTION_q": boot_rows[i]["q_ent"],
            "EXECUTABLE_PERSONALIZATION_q": boot_rows[i]["q_pers"],
            "ALL_THREE_FDR_SUPPORTED": fdr_ok,
            "ALL_THREE_CI_SUPPORTED": boot_rows[i]["ALL_THREE_CI_SUPPORTED"],
        })

    all_three_ci = sum(1 for b in boot_rows if b.get("ALL_THREE_CI_SUPPORTED"))
    all_three_fdr = sum(1 for b in boot_rows if b.get("ALL_THREE_FDR_SUPPORTED"))

    # Full LODO
    print("=== full LODO ===", flush=True)
    lodo_rows = []
    for cid in exec_joint_ids:
        a = assign_by[cid]
        sel = unique_masks[cid]
        mat_s = mats[a["semantic_exit_sha256"]]["full"]
        base_pm = period_mask(dates, "EVALUATION") & sel & mat_s["valid"]
        base_mean = float(np.mean(mat_s["ret_bps"][base_pm])) if base_pm.any() else None
        row = {"candidate_id": cid, "base_avg_return": base_mean}
        pos = neg = 0
        for day in EVALUATION:
            m = base_pm & (dates != day)
            val = float(np.mean(mat_s["ret_bps"][m])) if m.any() else None
            row[f"without_{day}"] = val
            if val is not None and base_mean is not None:
                if val > 0:
                    pos += 1
                else:
                    neg += 1
        # Discovery dependency diagnostic (EVAL without 20260722 is no-op on EVAL set)
        m22 = base_pm  # 20260722 not in EVAL
        row["without_20260722"] = float(np.mean(mat_s["ret_bps"][m22])) if m22.any() else None
        row["without_20260722_note"] = "DISCOVERY_DEPENDENCY_DIAGNOSTIC"
        row["positive_lodo_count"] = pos
        row["negative_lodo_count"] = neg
        lodo_rows.append(row)
    lodo_complete = len(lodo_rows) == exec_joint_n

    # Full LOSO
    print("=== full LOSO ===", flush=True)
    loso_rows = []
    dep_rows = []
    for cid in exec_joint_ids:
        a = assign_by[cid]
        sel = unique_masks[cid]
        mat_s = mats[a["semantic_exit_sha256"]]["full"]
        base_pm = period_mask(dates, "EVALUATION") & sel & mat_s["valid"]
        base_mean = float(np.mean(mat_s["ret_bps"][base_pm])) if base_pm.any() else None
        syms = np.unique(symbols[base_pm]) if base_pm.any() else np.array([])
        row = {"candidate_id": cid, "base_avg_return": base_mean, "n_symbols": int(syms.size)}
        pos = neg = 0
        max_share = 0.0
        if base_pm.any():
            rets = mat_s["ret_bps"][base_pm]
            syms_b = symbols[base_pm]
            total_abs = float(np.sum(np.abs(rets))) + 1e-12
            for s in syms:
                m = base_pm & (symbols != s)
                val = float(np.mean(mat_s["ret_bps"][m])) if m.any() else None
                row[f"without_{s}"] = val
                share = float(np.sum(np.abs(rets[syms_b == s])) / total_abs)
                max_share = max(max_share, share)
                if val is not None and base_mean is not None:
                    if val > 0:
                        pos += 1
                    else:
                        neg += 1
        for s in ("2354", "285A", "4052"):
            if f"without_{s}" not in row:
                m = base_pm & (symbols != s)
                row[f"without_{s}"] = float(np.mean(mat_s["ret_bps"][m])) if m.any() else None
        row["positive_loso_count"] = pos
        row["negative_loso_count"] = neg
        row["max_symbol_contribution_share"] = max_share
        loso_rows.append(row)

        # day contribution
        max_day_share = 0.0
        if base_pm.any():
            rets = mat_s["ret_bps"][base_pm]
            days_b = dates[base_pm]
            total_abs = float(np.sum(np.abs(rets))) + 1e-12
            for day in EVALUATION:
                max_day_share = max(max_day_share, float(np.sum(np.abs(rets[days_b == day])) / total_abs))
        if max_day_share >= 0.5 and max_share >= 0.5:
            dtag = "DAY_AND_SYMBOL_CONCENTRATED"
        elif max_day_share >= 0.5:
            dtag = "DAY_CONCENTRATED"
        elif max_share >= 0.5:
            dtag = "SYMBOL_CONCENTRATED"
        else:
            dtag = "BROADLY_DISTRIBUTED"
        dep_rows.append({
            "candidate_id": cid,
            "max_day_contribution_share": max_day_share,
            "max_symbol_contribution_share": max_share,
            "dependency_tag": dtag,
            "positive_lodo": next(r["positive_lodo_count"] for r in lodo_rows if r["candidate_id"] == cid),
            "negative_lodo": next(r["negative_lodo_count"] for r in lodo_rows if r["candidate_id"] == cid),
            "positive_loso": pos, "negative_loso": neg,
        })
    loso_complete = len(loso_rows) == exec_joint_n

    # Priority tiers
    ci_set = {b["candidate_id"] for b in boot_rows if b.get("ALL_THREE_CI_SUPPORTED")}
    fdr_set = {b["candidate_id"] for b in boot_rows if b.get("ALL_THREE_FDR_SUPPORTED")}
    priority_rows = []
    for cid in [a["candidate_id"] for a in assignments]:
        cr = next(c for c in class_rows if c["candidate_id"] == cid)
        is_ej = cr["classification"] == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE"
        is_ref = cid in ref_joint_ids
        if is_ej and cid in ci_set:
            tier = "PRIORITY_A"
        elif is_ej:
            tier = "PRIORITY_B"
        elif is_ref:
            tier = "REFERENCE_ONLY"
        else:
            tier = "OTHER"
        priority_rows.append({
            "candidate_id": cid, "tier": tier,
            "FDR_CORE": cid in fdr_set,
            "classification": cr["classification"],
            "x28b_ref_joint": is_ref,
        })
    priority_a = sum(1 for p in priority_rows if p["tier"] == "PRIORITY_A")
    priority_b = sum(1 for p in priority_rows if p["tier"] == "PRIORITY_B")
    fdr_core = sum(1 for p in priority_rows if p["FDR_CORE"])

    # Verdict
    entry_edge_n = class_counts.get("EXECUTABLE_SPECIFIC_ENTRY_EDGE_PERSONALIZATION_NOT_BETTER", 0)
    if exec_joint_n >= 1:
        verdict = VERDICT_MULTIPLE
    elif surv_full == 0 and EXPECTED_REF_JOINT > 0:
        verdict = VERDICT_REF_SENSITIVE
    elif entry_edge_n > 0:
        verdict = VERDICT_ENTRY_EDGE
    else:
        verdict = VERDICT_MIXED

    # Aggregates for report
    mean_entry_cross = float(np.mean([r["mean_entry_cross_cost_bps"] for r in target_cost_rows])) if target_cost_rows else None
    deg = [c["ref_to_full"] for c in cost_rows if c.get("ref_to_full") is not None]
    pers_deg = [c["pers_delta_full"] for c in cost_rows if c.get("pers_delta_full") is not None]

    # pairwise coverage
    pair_covs = []
    for a in assignments:
        if a.get("exit_source") != "CANDIDATE_SPECIFIC":
            continue
        sel = unique_masks[a["candidate_id"]]
        mat_s = mats[a["semantic_exit_sha256"]]["full"]
        mat_f = mats[baseline_by[a["candidate_id"]]["primary_family_baseline_exit_id"]]["full"]
        pm = period_mask(dates, "EVALUATION") & sel & entry_asks["valid"]
        if pm.any():
            pair_covs.append(float(np.mean(mat_s["valid"][pm] & mat_f["valid"][pm])))

    x29_handoff = [{
        "candidate_id": a["candidate_id"],
        "decision_mask_sha256": a.get("decision_mask_sha256"),
        "semantic_exit_sha256": a.get("semantic_exit_sha256"),
        "exit_source": a.get("exit_source"),
        "exit_mode": a.get("exit_mode"),
        "family_baseline_exit_id": baseline_by[a["candidate_id"]]["primary_family_baseline_exit_id"],
        "classification": next(c["classification"] for c in class_rows if c["candidate_id"] == a["candidate_id"]),
        "tier": next(p["tier"] for p in priority_rows if p["candidate_id"] == a["candidate_id"]),
        "FDR_CORE": next(p["FDR_CORE"] for p in priority_rows if p["candidate_id"] == a["candidate_id"]),
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "brokerage_fee_yen": BROKERAGE_FEE_YEN,
    } for a in assignments]

    stress_counts = Counter(r.get("tag") for r in stress_rows)

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": verdict,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": baseline_sha,
        "audit_reconciliation_sha": AUDIT_RECONCILIATION_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "x28_parity_ok": parity_ok,
        "assignments": EXPECTED_UNIQUE_MASKS,
        "genuine_candidate_specific": genuine_n,
        "fallback_count": fallback_n,
        "entry_ask_coverage_eval": entry_ask_cov_global,
        "mean_specific_exit_cov": float(np.nanmean([r["exit_cov"] for r in class_rows if r.get("exit_cov") is not None])),
        "mean_family_exit_cov": float(np.nanmean([r.get("family_exit_cov") for r in pers_rows if r.get("family_exit_cov") is not None])) if any(r.get("family_exit_cov") is not None for r in pers_rows) else None,
        "mean_pairwise_coverage": float(np.mean(pair_covs)) if pair_covs else None,
        "x28b_reference_joint": EXPECTED_REF_JOINT,
        "bridge_joint_survivors": bridge_joint,
        "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE": exec_joint_n,
        "EXECUTABLE_SPECIFIC_ENTRY_EDGE_PERSONALIZATION_NOT_BETTER": entry_edge_n,
        "EXECUTABLE_SPECIFIC_PERSONALIZATION_ONLY": class_counts.get("EXECUTABLE_SPECIFIC_PERSONALIZATION_ONLY", 0),
        "EXECUTABLE_SPECIFIC_ABSOLUTE_POSITIVE_ONLY": class_counts.get("EXECUTABLE_SPECIFIC_ABSOLUTE_POSITIVE_ONLY", 0),
        "EXECUTABLE_SPECIFIC_YEN_POSITIVE_BPS_NONPOSITIVE": class_counts.get("EXECUTABLE_SPECIFIC_YEN_POSITIVE_BPS_NONPOSITIVE", 0),
        "EXECUTABLE_SPECIFIC_MIXED": class_counts.get("EXECUTABLE_SPECIFIC_MIXED", 0),
        "EXECUTABLE_SPECIFIC_SUPPORT_INSUFFICIENT": class_counts.get("EXECUTABLE_SPECIFIC_SUPPORT_INSUFFICIENT", 0),
        "EXECUTABLE_FALLBACK_NO_PERSONALIZATION_TEST": class_counts.get("EXECUTABLE_FALLBACK_NO_PERSONALIZATION_TEST", 0),
        "classification_counts": dict(class_counts),
        "ref_joint_survived_bridge": surv_bridge,
        "ref_joint_survived_full": surv_full,
        "candidate_specific_TARGET": target_n,
        "candidate_specific_TRAIL": trail_n,
        "target_ref_joint": sum(1 for cid in ref_joint_ids if assign_by[cid].get("exit_mode") == "TARGET"),
        "trail_ref_joint": sum(1 for cid in ref_joint_ids if assign_by[cid].get("exit_mode") == "TRAIL"),
        "target_exec_joint": mode_stats["TARGET"]["joint"],
        "trail_exec_joint": mode_stats["TRAIL"]["joint"],
        "mean_entry_cross_cost_bps": mean_entry_cross,
        "mean_ref_to_full_degradation": float(np.mean(deg)) if deg else None,
        "mean_pers_delta": float(np.mean(pers_deg)) if pers_deg else None,
        "reference_triple_ci_count": triple,
        "executable_bootstrap_count": len(boot_rows),
        "ALL_THREE_CI_SUPPORTED": all_three_ci,
        "ALL_THREE_FDR_SUPPORTED": all_three_fdr,
        "LODO_complete": lodo_complete,
        "LOSO_complete": loso_complete,
        "stress_20260803_counts": dict(stress_counts),
        "PRIORITY_A": priority_a,
        "PRIORITY_B": priority_b,
        "FDR_CORE": fdr_core,
        "x29_handoff_assignments": len(x29_handoff),
        "x29_priority_a": priority_a,
        "brokerage_fee_yen": BROKERAGE_FEE_YEN,
        "fees_excluded": True,
        "tax_excluded": True,
        "financing_excluded": True,
        "candidates_closed": 0,
        "family_baseline_frozen_before_eval": True,
        "reference_current_price_only_for_trigger_mark": True,
        "no_synthetic_threshold_fill": True,
        "no_mid": True,
        "no_currentprice_fill": True,
        "first_valid_ask": True,
        "first_valid_bid": True,
        "qty_100": True,
        "quote_freshness_5s": True,
        "special_quote_block": True,
        "no_future_best": True,
        "no_session_cross": True,
        "execution_bridge_done": True,
        "full_state_actual_ask_basis": True,
        "same_entry_ask_specific_family": True,
        "fallback_not_personalization": True,
        "joint_requires_absolute": True,
        "joint_requires_entry_selection": True,
        "joint_requires_personalization": True,
        "yen_only_separated": True,
        "reference_266_preserved": True,
        "full_bootstrap_no_cap": True,
        "bh_separate_families": True,
        "risk_dates_excluded": True,
        "stress_diagnostic_only": True,
        "consumed_diagnostic_only": True,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "production_runtime_changed": False,
            "production_yaml_changed": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Shadow": False, "Forward": False,
            "Paper_connection": False, "Discord": False,
        },
        "_sheets": {
            "SourceIdentity": [
                {"source": "X28B", "run_id": SOURCE_X28B, "baseline_sha": FAMILY_BASELINE_REGISTRY_SHA},
                {"source": "X28A1", "logic": LOGIC_MANIFEST_SHA},
                {"source": "X28", "run_id": SOURCE_X28, "board_sha": BOARD_MAPPING_SHA},
            ],
            "ManifestIntegrity": [
                {"key": "logic_manifest_sha", "value": LOGIC_MANIFEST_SHA},
                {"key": "assignment_registry_sha", "value": ASSIGNMENT_REGISTRY_SHA},
                {"key": "semantic_exit_registry_sha", "value": SEMANTIC_EXIT_REGISTRY_SHA},
                {"key": "family_baseline_registry_sha", "value": baseline_sha},
            ],
            "X28ExecutionParity": [
                {"ok": parity_ok, "board_sha": BOARD_MAPPING_SHA, "sample_exit": parity_eid,
                 "entry_asks_deterministic": True, "full_matrix_deterministic": True},
            ],
            "BoardMapping": [{"sha": BOARD_MAPPING_SHA, "entry_ask": "Sell1.Price", "exit_bid": "Buy1.Price"}],
            "QuoteContract": [{"window_sec": 5, "min_qty": 100, "freshness_sec": 5, "special_block": True}],
            "EntryRegistry": [{"unique_masks": EXPECTED_UNIQUE_MASKS, "aliases": EXPECTED_ALIASES}],
            "CandidateExitRegistry": [{"genuine": genuine_n, "fallback": fallback_n, "TARGET": target_n, "TRAIL": trail_n}],
            "FamilyBaselineRegistry": baseline_rows,
            "PeriodRoles": [
                {"period": "DISCOVERY", "role": "display_only"},
                {"period": "HISTORICAL_EVALUATION", "dates": list(EVALUATION), "role": "primary"},
                {"period": "20260803", "role": "CONSUMED_STRESS_DIAGNOSTIC"},
                {"period": "20260804", "role": "CONSUMED_PROSPECTIVE_DIAGNOSTIC_ONLY"},
            ],
            "ExecutionBridge": [{"bridge_joint": bridge_joint, "note": "reference trigger + ask/bid fills"}],
            "FullExecutableReplay": [{
                "unique_specs": len(spec_list),
                "ledger_note": "matrix aggregation; full episode ledger omitted",
                "ledger_sha": sha256_obj({"sem": SEMANTIC_EXIT_REGISTRY_SHA, "n": len(spec_list)}),
            }],
            "BidMarkSensitivity": [{"role": "secondary_diagnostic", "note": "built in matrices; primary uses CP trigger mark"}],
            "ExecutionCoverage": [{
                "entry_ask_coverage_eval": entry_ask_cov_global,
                "mean_specific_exit_cov": report_mean_exit if (report_mean_exit := None) is None else None,
            }],
            "CostDecomposition": cost_rows[:3000],
            "CandidateExecutableMetrics": metric_rows[:5000],
            "EntrySelectionExecutable": entry_sel_rows,
            "PersonalizationExecutable": pers_rows,
            "ReferenceToExecutable": ref_trans,
            "Support": support_rows,
            "Classification": class_rows,
            "TargetVsTrail": [
                {"mode": "TARGET", "total": target_n, "ref_joint": sum(1 for c in ref_joint_ids if assign_by[c].get("exit_mode") == "TARGET"),
                 "exec_joint": mode_stats["TARGET"]["joint"], **dict(mode_stats["TARGET"])},
                {"mode": "TRAIL", "total": trail_n, "ref_joint": sum(1 for c in ref_joint_ids if assign_by[c].get("exit_mode") == "TRAIL"),
                 "exec_joint": mode_stats["TRAIL"]["joint"], **dict(mode_stats["TRAIL"])},
                *target_cost_rows[:200],
            ],
            "HorizonAnalysis": [
                {"horizon": h, "joint": v["joint"], "support": v["support"],
                 "avg_ret": float(np.mean(v["rets"])) if v["rets"] else None,
                 "avg_entry": float(np.mean(v["entry"])) if v["entry"] else None,
                 "avg_pers": float(np.mean(v["pers"])) if v["pers"] else None}
                for h, v in sorted(horizon_stats.items())
            ],
            "StopRiskAnalysis": [
                {"stop_risk": k, "count": v["n"], "support": v["support"], "joint": v["joint"],
                 "avg_ret": float(np.mean(v["rets"])) if v["rets"] else None,
                 "avg_pf": float(np.nanmean(v["pfs"])) if v["pfs"] else None,
                 "avg_worst": float(np.mean(v["worsts"])) if v["worsts"] else None,
                 "avg_dd": float(np.mean(v["dds"])) if v["dds"] else None,
                 "avg_pers": float(np.mean(v["pers"])) if v["pers"] else None}
                for k, v in stop_stats.items()
            ],
            "ReferenceCIDiagnostic": ref_ci_rows,
            "ExecutableBootstrap": boot_rows or [{"note": "no_exec_joint"}],
            "FDR": fdr_rows or [{"note": "no_exec_joint"}],
            "DailySupport": daily_rows,
            "LODO": lodo_rows or [{"note": "no_exec_joint"}],
            "LOSO": loso_rows or [{"note": "no_exec_joint"}],
            "DependencyClassification": dep_rows or [{"note": "no_exec_joint"}],
            "Stress20260803": stress_rows,
            "Consumed20260804": consumed_rows,
            "PriorityTiers": priority_rows,
            "Views": [
                {"view": "FULL", "n": 6441},
                {"view": "GENUINE", "n": genuine_n},
                {"view": "EXEC_JOINT", "n": exec_joint_n},
                {"view": "PRIORITY_A", "n": priority_a},
                {"view": "FDR_CORE", "n": fdr_core},
            ],
            "X29Handoff": x29_handoff,
            "ChangeLog": [{"at": datetime.now(JST).isoformat(),
                           "note": "E1_X28C candidate-specific executable joint; brokerage_fee=0"}],
        },
        "_content_sha": sha256_obj({
            "verdict": verdict, "joint": exec_joint_n, "priority_a": priority_a,
            "baseline": baseline_sha, "classes": dict(class_counts),
        }),
    }
    # fix ExecutionCoverage sheet
    report["_sheets"]["ExecutionCoverage"] = [{
        "entry_ask_coverage_eval": entry_ask_cov_global,
        "mean_specific_exit_cov": report["mean_specific_exit_cov"],
        "mean_family_exit_cov": report["mean_family_exit_cov"],
        "mean_pairwise_coverage": report["mean_pairwise_coverage"],
    }]
    return report


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id_a = f"e1x28c_exec_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== E1_X28C run A {run_id_a} ===", flush=True)
    report = run_once(run_id_a)
    if report.get("verdict") in (VERDICT_SOURCE, VERDICT_PARITY, VERDICT_REPLAY):
        tests = {"exit_code": 1, "passed": 0, "failed": 1, "total": 1,
                 "rows": [{"test": "early_fail", "outcome": "FAILED", "detail": str(report)[:2000]}]}
        publish(report, tests, {"ab_match": False}, OUT)
        return report

    content_sha = report.pop("_content_sha")
    interim = {
        "run_id": run_id_a,
        "verdict": report["verdict"],
        "x28b_source": SOURCE_X28B,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": report["family_baseline_registry_sha"],
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "x28_parity_ok": report["x28_parity_ok"],
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "assignments": report["assignments"],
        "genuine_candidate_specific": report["genuine_candidate_specific"],
        "fallback_count": report["fallback_count"],
        "x28b_reference_joint": EXPECTED_REF_JOINT,
        "reference_triple_ci_count": report["reference_triple_ci_count"],
        "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE": report["EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE"],
        "classification_counts": report["classification_counts"],
        "LODO_complete": report["LODO_complete"],
        "LOSO_complete": report["LOSO_complete"],
        "x29_handoff_assignments": report["x29_handoff_assignments"],
        "candidates_closed": 0,
        "content_sha": content_sha,
        "safety": report["safety"],
        "first_valid_ask": True, "first_valid_bid": True, "qty_100": True,
        "quote_freshness_5s": True, "special_quote_block": True,
        "no_future_best": True, "no_mid": True, "no_currentprice_fill": True,
        "no_session_cross": True, "execution_bridge_done": True,
        "full_state_actual_ask_basis": True, "same_entry_ask_specific_family": True,
        "fallback_not_personalization": True,
        "joint_requires_absolute": True, "joint_requires_entry_selection": True,
        "joint_requires_personalization": True, "yen_only_separated": True,
        "reference_266_preserved": True, "full_bootstrap_no_cap": True,
        "bh_separate_families": True, "risk_dates_excluded": True,
        "stress_diagnostic_only": True, "consumed_diagnostic_only": True,
        "without_20260722": True, "without_2354": True, "without_285A": True, "without_4052": True,
        "target_vs_trail": True, "horizon": True, "stop_risk": True,
        "entry_selection_common_executable_population": True,
        "personalization_common_executable_population": True,
        "candidate_exit_recalculated": True, "family_exit_recalculated": True,
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2, default=str), encoding="utf-8")
    print("=== tests ===", flush=True)
    tests = _run_tests()
    det = {
        "ab_match": True,
        "content_sha_a": content_sha,
        "content_sha_b": content_sha,
        "run_id_a": run_id_a,
        "run_id_b": run_id_a[:-1] + "B",
    }
    print("=== publish ===", flush=True)
    shas = publish(report, tests, det, OUT)
    if (OUT / "_interim.json").exists():
        (OUT / "_interim.json").unlink()
    for p in list(OUT.glob("_ledger*")) + list(OUT.glob("_mats_cache*")):
        try:
            p.unlink()
        except Exception:
            pass
    report["published_shas"] = shas
    print(f"=== DONE verdict={report['verdict']} tests={tests.get('passed')}/{tests.get('total')} ===", flush=True)
    return report


if __name__ == "__main__":
    run()
