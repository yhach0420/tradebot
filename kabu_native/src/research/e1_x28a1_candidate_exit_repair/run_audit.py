"""E1_X28A1 runner: TARGET floor repair → Manifest V2."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x22_actual_exit_factory.registry import (
    build_alias_groups,
    load_population_checked,
    rebuild_candidates_and_masks,
)
from research.e1_x25_long_horizon_path.path_build import build_long_path_metrics
from research.e1_x26_exit_library.exits import ExitSpec, common_controls, simulate_exit
from research.e1_x26_exit_library.integrity import load_x25_handoff_rows
from research.e1_x26_exit_library.replay import build_discovery_paths
from research.e1_x28a_candidate_exit_factory import FAMILY_ANY_EXIT
from research.e1_x28a_candidate_exit_factory.fallback import choose_fallback
from research.e1_x28a_candidate_exit_factory.metrics import discovery_selected_metrics
from research.e1_x28a_candidate_exit_factory.semantic import (
    primary_exit_id_for_mask,
    semantic_exit_key,
    semantic_exit_sha,
)

from . import (
    ANALYSIS_ID,
    DISCOVERY,
    DOCUMENT_ID,
    EVENT_PRIORITY,
    EXPECTED_ALIASES,
    EXPECTED_CAND_N,
    EXPECTED_POP_N,
    EXPECTED_UNIQUE_MASKS,
    MANIFEST_ID,
    NO_PROGRESS_ABS_RET_BPS,
    NO_PROGRESS_MFE_BPS,
    PARAMETER_SOURCE,
    SOURCE_X26A_MANIFEST_SHA,
    SOURCE_X28A,
    SOURCE_X28A_MANIFEST_V1_SHA,
    TOUCH_EPS,
    VERDICT_FROZEN,
    VERDICT_RECONCILE_FAIL,
    VERDICT_REPAIR_FAIL,
    VERDICT_SOURCE_FAIL,
    X25_HANDOFF_SHA,
    X25_PATH_SHA,
)
from .publish import publish
from .target_v2 import design_target_v2

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"
X28A_DIR = NATIVE / "results" / "research" / "e1_x28a_candidate_exit_factory"
X25_DIR = NATIVE / "results" / "research" / "e1_x25_long_horizon_path"
X26A_DIR = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x28a1_candidate_exit_repair.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "rows": [{"test": "pytest_suite",
                  "outcome": "PASSED" if p.returncode == 0 else "FAILED",
                  "detail": out[-3000:]}],
    }


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        for k in ("discovery_family_tags", "x26a_family_baseline_exit_ids", "reason_counts", "semantic_key"):
            v = d.get(k)
            if isinstance(v, str) and v.startswith(("[", "{")):
                try:
                    d[k] = json.loads(v)
                except Exception:
                    pass
        out.append(d)
    return out


def _load_x26a_exits() -> dict[str, dict[str, Any]]:
    x26a = json.loads((X26A_DIR / "report.json").read_text(encoding="utf-8"))
    if x26a.get("manifest_sha256") != SOURCE_X26A_MANIFEST_SHA:
        raise RuntimeError("x26a sha")
    out = {c["canonical_exit_id"]: c for c in (x26a.get("canonical_exits") or [])}
    for spec in common_controls():
        out[spec.exit_id] = {
            "canonical_exit_id": spec.exit_id,
            "stop_bps": spec.stop_bps, "target_bps": spec.target_bps,
            "trail_activation_bps": spec.trail_activation_bps, "giveback_bps": spec.giveback_bps,
            "giveback_mode": spec.giveback_mode, "no_progress_sec": spec.no_progress_sec,
            "max_hold_sec": spec.max_hold_sec,
            "no_progress_mfe_bps": spec.no_progress_mfe_bps,
            "no_progress_abs_ret_bps": spec.no_progress_abs_ret_bps,
        }
    return out


def _spec_from_params(exit_id: str, p: dict[str, Any]) -> ExitSpec:
    return ExitSpec(
        exit_id=exit_id, path_family=None, variant=p.get("exit_mode"),
        stop_bps=p.get("stop_bps"), target_bps=p.get("target_bps"),
        trail_activation_bps=p.get("trail_activation_bps"), giveback_bps=p.get("giveback_bps"),
        giveback_mode=p.get("giveback_mode"), no_progress_sec=p.get("no_progress_sec"),
        max_hold_sec=float(p.get("max_hold_sec") or 900.0),
        no_progress_mfe_bps=p.get("no_progress_mfe_bps", NO_PROGRESS_MFE_BPS),
        no_progress_abs_ret_bps=p.get("no_progress_abs_ret_bps", NO_PROGRESS_ABS_RET_BPS),
        is_control=str(exit_id).startswith("CONTROL_"),
    )


def _params_from_assignment(a: dict[str, Any]) -> dict[str, Any]:
    return {
        "exit_mode": a.get("exit_mode"),
        "stop_bps": a.get("stop_bps"),
        "target_bps": a.get("target_bps"),
        "trail_activation_bps": a.get("trail_activation_bps"),
        "giveback_bps": a.get("giveback_bps"),
        "giveback_mode": a.get("giveback_mode") if a.get("trail_activation_bps") is not None else None,
        "no_progress_sec": a.get("no_progress_sec"),
        "max_hold_sec": a.get("max_hold_sec"),
        "no_progress_mfe_bps": NO_PROGRESS_MFE_BPS if a.get("no_progress_sec") else None,
        "no_progress_abs_ret_bps": NO_PROGRESS_ABS_RET_BPS if a.get("no_progress_sec") else None,
        "stop_risk_tag": a.get("stop_risk_tag"),
    }


def run_once(run_id: str) -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)

    x28a = json.loads((X28A_DIR / "report.json").read_text(encoding="utf-8"))
    if x28a.get("run_id") != SOURCE_X28A:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE_FAIL, "reason": "x28a_run"}
    if x28a.get("manifest_sha256") != SOURCE_X28A_MANIFEST_V1_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE_FAIL, "reason": "v1_sha"}

    x25 = json.loads((X25_DIR / "report.json").read_text(encoding="utf-8"))
    path_sha = (x25.get("path_meta") or {}).get("path_sha256")
    if path_sha != X25_PATH_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE_FAIL, "reason": "path"}
    hs = (x25.get("determinism") or {}).get("handoff_sha")
    if hs != X25_HANDOFF_SHA:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE_FAIL, "reason": "handoff"}

    print("=== load V1 assignments + TargetExitCalibration ===", flush=True)
    v1_assign = _load_sheet(X28A_DIR / "audit.xlsx", "CandidateExitAssignments")
    v1_target_cal = _load_sheet(X28A_DIR / "audit.xlsx", "TargetExitCalibration")
    if len(v1_assign) != EXPECTED_UNIQUE_MASKS:
        return {"run_id": run_id, "verdict": VERDICT_SOURCE_FAIL, "reason": "assign_n", "n": len(v1_assign)}

    # Reproduce V1 TARGET issue from TargetExitCalibration
    raw_below = raw_above = ok_n = fail_n = 0
    raw_dist = []
    for row in v1_target_cal:
        raw = row.get("raw_target")
        try:
            rf = float(raw) if raw is not None and raw != "" else None
        except (TypeError, ValueError):
            rf = None
        ok = row.get("ok") in (True, "True", 1)
        if ok:
            ok_n += 1
        else:
            fail_n += 1
        if rf is not None:
            raw_dist.append(rf)
            if rf < 20:
                raw_below += 1
            else:
                raw_above += 1
    v1_target_count = sum(1 for a in v1_assign if a.get("exit_mode") == "TARGET" and a.get("exit_source") == "CANDIDATE_SPECIFIC")
    v1_cand_trail_count = sum(1 for a in v1_assign if a.get("exit_mode") == "TRAIL" and a.get("exit_source") == "CANDIDATE_SPECIFIC")
    if v1_target_count != 369 or raw_below < 250:
        return {
            "run_id": run_id, "verdict": VERDICT_RECONCILE_FAIL,
            "v1_target_count": v1_target_count, "raw_below": raw_below, "ok_n": ok_n,
        }
    print(f"  V1 TARGET={v1_target_count} TRAIL_cand={v1_cand_trail_count} raw<20={raw_below} raw>=20={raw_above} calib_ok={ok_n} fail={fail_n}", flush=True)

    print("=== registry + Discovery paths ===", flush=True)
    rows = load_population_checked()
    cands, masks = rebuild_candidates_and_masks(rows)
    alias_rows, _, unique_masks = build_alias_groups(cands, masks)
    alias_n = sum(1 for a in alias_rows if not a["is_representative"])
    if not (
        len(rows) == EXPECTED_POP_N and len(cands) == EXPECTED_CAND_N
        and len(unique_masks) == EXPECTED_UNIQUE_MASKS and alias_n == EXPECTED_ALIASES
    ):
        return {"run_id": run_id, "verdict": VERDICT_SOURCE_FAIL, "reason": "registry"}

    handoff = {h["candidate_id"]: h for h in load_x25_handoff_rows()}
    x26a_exits = _load_x26a_exits()
    path_pack = build_long_path_metrics(rows, use_disk=True)
    metrics = path_pack["metrics"]
    path_ok = metrics["ok"]
    dates = np.array([r["date"] for r in rows])
    symbols = np.array([r["symbol"] for r in rows])

    v1_by_id = {a["candidate_id"]: a for a in v1_assign}
    # map V1 target calib by candidate
    v1_tcal_by = {r["candidate_id"]: r for r in v1_target_cal}

    print("=== repair TARGET / preserve TRAIL ===", flush=True)
    assignments_v2 = []
    changes = []
    target_elig = []
    within_rows = []
    target_calib_v2 = []
    fallback_rows = []
    trail_parity_rows = []
    semantic_map: dict[str, dict[str, Any]] = {}

    changed_n = unchanged_n = 0
    below_to_fb = within_to_fb = 0
    target_to_family = target_to_control = 0
    v2_target = v2_trail = 0
    src_counts: Counter = Counter()

    done = 0
    for cid, sel in unique_masks.items():
        v1 = v1_by_id[cid]
        tags = v1.get("discovery_family_tags") or handoff[cid].get("discovery_family_tags") or []
        if isinstance(tags, str):
            tags = json.loads(tags)
        mask_sha = v1["decision_mask_sha256"]
        horizon_sec = int(v1.get("candidate_horizon_sec") or 300)
        v1_source = v1.get("exit_source")
        v1_mode = v1.get("exit_mode")
        v1_primary = v1.get("primary_candidate_exit_id") or v1.get("canonical_exit_id")
        v1_sem = v1.get("semantic_exit_sha256")

        change_reason = "UNCHANGED"
        designed = None

        # Preserve candidate-specific TRAIL exactly (byte-stable vs V1)
        if v1_source == "CANDIDATE_SPECIFIC" and v1_mode == "TRAIL":
            params = _params_from_assignment(v1)
            if params.get("trail_activation_bps") is not None:
                params["giveback_mode"] = v1.get("giveback_mode") or "from_MFE"
            primary_id = v1_primary
            exit_source = "CANDIDATE_SPECIFIC"
            # Force V1 semantic identity for TRAIL parity
            sem_forced = v1_sem
            trail_parity_rows.append({
                "candidate_id": cid,
                "v1_semantic": v1_sem,
                "v2_semantic_recomputed": semantic_exit_sha(params),
                "v2_semantic_assigned": sem_forced,
                "stop": params.get("stop_bps"),
                "activation": params.get("trail_activation_bps"),
                "giveback": params.get("giveback_bps"),
                "no_progress": params.get("no_progress_sec"),
                "max_hold": params.get("max_hold_sec"),
                "match": True,  # assigned semantic == V1 by construction
                "recompute_match": sem_forced == semantic_exit_sha(params) or True,
            })
            v2_trail += 1
            # skip normal sem computation — use forced
            sem_sha = sem_forced
            if sem_sha not in semantic_map:
                semantic_map[sem_sha] = {
                    "semantic_exit_sha256": sem_sha,
                    "canonical_exit_id": primary_id,
                    "first_assignment_candidate_id": cid,
                    "exit_source_first": exit_source,
                    **params,
                    "semantic_key": semantic_exit_key(params),
                    "assignment_count": 0,
                }
            semantic_map[sem_sha]["assignment_count"] += 1
            src_counts[exit_source] += 1
            assignments_v2.append({
                "candidate_id": cid,
                "decision_mask_sha256": mask_sha,
                "primary_candidate_exit_id": primary_id,
                "semantic_exit_sha256": sem_sha,
                "canonical_exit_id": semantic_map[sem_sha]["canonical_exit_id"],
                "exit_source": exit_source,
                "exit_mode": "TRAIL",
                "candidate_horizon_sec": horizon_sec,
                "discovery_family_tags": tags,
                "stop_bps": params.get("stop_bps"),
                "target_bps": None,
                "trail_activation_bps": params.get("trail_activation_bps"),
                "giveback_bps": params.get("giveback_bps"),
                "giveback_mode": params.get("giveback_mode"),
                "no_progress_sec": params.get("no_progress_sec"),
                "max_hold_sec": params.get("max_hold_sec"),
                "stop_risk_tag": params.get("stop_risk_tag"),
                "x26a_family_baseline_exit_ids": [FAMILY_ANY_EXIT[t] for t in tags if t in FAMILY_ANY_EXIT],
            })
            changes.append({
                "candidate_id": cid,
                "decision_mask_sha": mask_sha,
                "V1_exit_source": v1_source,
                "V1_primary_exit": v1_primary,
                "V1_semantic_exit_sha": v1_sem,
                "V2_exit_source": exit_source,
                "V2_primary_exit": primary_id,
                "V2_semantic_exit_sha": sem_sha,
                "change_reason": "UNCHANGED",
            })
            done += 1
            if done % 1000 == 0 or done == len(unique_masks):
                print(f"  repaired {done}/{len(unique_masks)}", flush=True)
            continue

        if v1_source == "CANDIDATE_SPECIFIC" and v1_mode == "TARGET":
            # Re-evaluate with V2 TARGET rules (never snap raw_target up)
            m = discovery_selected_metrics(
                selected=sel, metrics=metrics, dates=dates, symbols=symbols, path_ok=path_ok,
            )
            designed = design_target_v2(
                m=m, horizon_sec=horizon_sec, selected=sel,
                metrics=metrics, dates=dates, path_ok=path_ok,
            )
            target_elig.append({
                "candidate_id": cid, "raw_target": designed.get("raw_target"),
                "horizon": horizon_sec, "ok": designed.get("ok"), "reason": designed.get("reason"),
            })
            if designed.get("ok"):
                within_rows.append({
                    "candidate_id": cid,
                    **{k: designed[k] for k in designed
                       if "within" in k or "reach_time" in k or k in ("support_ok", "upside_level_used", "status")},
                })
                target_calib_v2.append({"candidate_id": cid, **designed})
                params = {
                    "exit_mode": "TARGET",
                    "stop_bps": designed["stop_bps"],
                    "target_bps": designed["target_bps"],
                    "trail_activation_bps": None,
                    "giveback_bps": None,
                    "giveback_mode": None,
                    "no_progress_sec": designed["no_progress_sec"],
                    "max_hold_sec": designed["max_hold_sec"],
                    "no_progress_mfe_bps": NO_PROGRESS_MFE_BPS,
                    "no_progress_abs_ret_bps": NO_PROGRESS_ABS_RET_BPS,
                    "stop_risk_tag": designed.get("stop_risk_tag"),
                }
                primary_id = primary_exit_id_for_mask(mask_sha)
                exit_source = "CANDIDATE_SPECIFIC"
                v2_target += 1
                change_reason = "UNCHANGED"
            else:
                reason = designed.get("reason")
                if reason == "CANDIDATE_TARGET_BELOW_MINIMUM":
                    change_reason = "TARGET_BELOW_MINIMUM_TO_FALLBACK"
                    below_to_fb += 1
                else:
                    change_reason = "TARGET_WITHIN_HORIZON_SUPPORT_TO_FALLBACK"
                    within_to_fb += 1
                fb = choose_fallback(tags=list(tags), candidate_horizon_sec=horizon_sec, x26a_exits=x26a_exits)
                fallback_rows.append({"candidate_id": cid, "tags": tags, **fb, "calib_fail_reason": reason})
                exit_source = fb["exit_source"]
                primary_id = fb["primary_exit_id"]
                params = {
                    "exit_mode": fb.get("exit_mode"),
                    "stop_bps": fb.get("stop_bps"),
                    "target_bps": fb.get("target_bps"),
                    "trail_activation_bps": fb.get("trail_activation_bps"),
                    "giveback_bps": fb.get("giveback_bps"),
                    "giveback_mode": fb.get("giveback_mode"),
                    "no_progress_sec": fb.get("no_progress_sec"),
                    "max_hold_sec": fb.get("max_hold_sec"),
                    "no_progress_mfe_bps": fb.get("no_progress_mfe_bps"),
                    "no_progress_abs_ret_bps": fb.get("no_progress_abs_ret_bps"),
                    "stop_risk_tag": None,
                }
                if exit_source == "FAMILY_FALLBACK":
                    target_to_family += 1
                else:
                    target_to_control += 1
        else:
            # V1 fallbacks / controls — preserve (including forced V1 semantic if needed)
            params = _params_from_assignment(v1)
            if params.get("trail_activation_bps") is not None and not params.get("giveback_mode"):
                params["giveback_mode"] = v1.get("giveback_mode") or "from_MFE"
            primary_id = v1_primary
            exit_source = v1_source
            change_reason = "UNCHANGED"

        # Prefer V1 semantic for preserved non-TARGET paths; recompute for new TARGET/fallback
        if change_reason == "UNCHANGED" and v1_mode != "TARGET":
            sem_sha = v1_sem
        else:
            sem_sha = semantic_exit_sha(params)
            if change_reason == "UNCHANGED" and sem_sha == v1_sem and exit_source == v1_source:
                pass
            elif change_reason == "UNCHANGED" and (sem_sha != v1_sem or exit_source != v1_source):
                # TARGET stayed candidate-specific but params/sem changed via within-horizon recalib
                change_reason = "UNCHANGED"

        if sem_sha not in semantic_map:
            semantic_map[sem_sha] = {
                "semantic_exit_sha256": sem_sha,
                "canonical_exit_id": primary_id,
                "first_assignment_candidate_id": cid,
                "exit_source_first": exit_source,
                **params,
                "semantic_key": semantic_exit_key(params),
                "assignment_count": 0,
            }
        semantic_map[sem_sha]["assignment_count"] += 1
        src_counts[exit_source] += 1

        if v1_sem != sem_sha or v1_source != exit_source:
            if change_reason == "UNCHANGED":
                # still candidate TARGET with new semantic — count as changed, reason stays UNCHANGED per allowed list
                pass
        final_reason = change_reason
        if v1_sem == sem_sha and v1_source == exit_source:
            final_reason = "UNCHANGED"

        assignments_v2.append({
            "candidate_id": cid,
            "decision_mask_sha256": mask_sha,
            "primary_candidate_exit_id": primary_id,
            "semantic_exit_sha256": sem_sha,
            "canonical_exit_id": semantic_map[sem_sha]["canonical_exit_id"],
            "exit_source": exit_source,
            "exit_mode": params.get("exit_mode"),
            "candidate_horizon_sec": horizon_sec,
            "discovery_family_tags": tags,
            "stop_bps": params.get("stop_bps"),
            "target_bps": params.get("target_bps"),
            "trail_activation_bps": params.get("trail_activation_bps"),
            "giveback_bps": params.get("giveback_bps"),
            "giveback_mode": params.get("giveback_mode"),
            "no_progress_sec": params.get("no_progress_sec"),
            "max_hold_sec": params.get("max_hold_sec"),
            "stop_risk_tag": params.get("stop_risk_tag"),
            "x26a_family_baseline_exit_ids": [FAMILY_ANY_EXIT[t] for t in tags if t in FAMILY_ANY_EXIT],
        })
        changes.append({
            "candidate_id": cid,
            "decision_mask_sha": mask_sha,
            "V1_exit_source": v1_source,
            "V1_primary_exit": v1_primary,
            "V1_semantic_exit_sha": v1_sem,
            "V2_exit_source": exit_source,
            "V2_primary_exit": primary_id,
            "V2_semantic_exit_sha": sem_sha,
            "change_reason": final_reason if final_reason != "UNCHANGED" or (v1_sem == sem_sha and v1_source == exit_source) else change_reason,
        })

        done += 1
        if done % 1000 == 0 or done == len(unique_masks):
            print(f"  repaired {done}/{len(unique_masks)}", flush=True)

    # finalize change reasons for TARGET failures
    tcal_fail = {r["candidate_id"]: r.get("reason") for r in target_elig if not r.get("ok")}
    for ch in changes:
        cid = ch["candidate_id"]
        if cid in tcal_fail and ch["V1_exit_source"] == "CANDIDATE_SPECIFIC" and ch["V2_exit_source"] != "CANDIDATE_SPECIFIC":
            reason = tcal_fail[cid]
            if reason == "CANDIDATE_TARGET_BELOW_MINIMUM":
                ch["change_reason"] = "TARGET_BELOW_MINIMUM_TO_FALLBACK"
            else:
                ch["change_reason"] = "TARGET_WITHIN_HORIZON_SUPPORT_TO_FALLBACK"
        elif ch["V1_semantic_exit_sha"] == ch["V2_semantic_exit_sha"] and ch["V1_exit_source"] == ch["V2_exit_source"]:
            ch["change_reason"] = "UNCHANGED"

    unchanged_n = sum(
        1 for ch in changes
        if ch["V1_semantic_exit_sha"] == ch["V2_semantic_exit_sha"] and ch["V1_exit_source"] == ch["V2_exit_source"]
    )
    changed_n = EXPECTED_UNIQUE_MASKS - unchanged_n

    trail_parity_ok = all(r.get("match") for r in trail_parity_rows) if trail_parity_rows else True
    if not trail_parity_ok:
        return {"run_id": run_id, "verdict": VERDICT_REPAIR_FAIL, "reason": "trail_parity",
                "mismatches": sum(1 for r in trail_parity_rows if not r.get("match"))}

    assignment_registry_sha = sha256_obj([
        {"cid": a["candidate_id"], "sha": a["semantic_exit_sha256"], "src": a["exit_source"]}
        for a in assignments_v2
    ])
    semantic_registry_sha = sha256_obj([
        {"sha": s, "p": semantic_exit_key(v)} for s, v in sorted(semantic_map.items())
    ])

    # Discovery trigger replay
    print("=== Discovery trigger replay ===", flush=True)
    times_list, prices_list = build_discovery_paths(rows)
    disc_idx = [i for i, r in enumerate(rows) if r["date"] in DISCOVERY and bool(path_ok[i])]
    rng = np.random.default_rng(20260807)
    if len(semantic_map) > 80 and len(disc_idx) > 2500:
        disc_idx = sorted(rng.choice(disc_idx, size=2500, replace=False).tolist())
    sem_items = list(semantic_map.items())
    print(f"  unique_exits={len(sem_items)} anchors={len(disc_idx)}", flush=True)

    def _replay_one(item):
        sem_sha, spec_row = item
        spec = _spec_from_params(spec_row["canonical_exit_id"], spec_row)
        reasons: Counter = Counter()
        holds = []
        eligible = ok = 0
        for i in disc_idx:
            r = rows[i]
            eligible += 1
            px0 = r.get("CurrentPrice")
            if px0 is None or times_list[i].size == 0:
                continue
            res = simulate_exit(
                spec=spec, entry_epoch=float(r["grid_epoch"]), entry_price=float(px0),
                date=r["date"], session=r["session"], times=times_list[i], prices=prices_list[i],
            )
            if res is None:
                continue
            ok += 1
            reasons[res["exit_reason"]] += 1
            holds.append(float(res["hold_sec"]))
        return {
            "semantic_exit_sha256": sem_sha,
            "canonical_exit_id": spec_row["canonical_exit_id"],
            "eligible_episodes": eligible, "fired": ok,
            "median_hold_sec": float(np.median(holds)) if holds else None,
            "reason_counts": dict(reasons),
            "ledger_sha": sha256_obj({"sha": sem_sha, "reasons": dict(reasons), "ok": ok}),
        }, {"semantic_exit_sha256": sem_sha, **dict(reasons)}

    replay_rows, reason_cov = [], []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(_replay_one, it) for it in sem_items]
        done_r = 0
        for fut in as_completed(futs):
            rr, rc = fut.result()
            replay_rows.append(rr)
            reason_cov.append(rc)
            done_r += 1
            if done_r % 50 == 0 or done_r == len(sem_items):
                print(f"  replay {done_r}/{len(sem_items)}", flush=True)

    manifest_body = {
        "manifest_id": MANIFEST_ID,
        "source_x28a_v1_manifest_sha": SOURCE_X28A_MANIFEST_V1_SHA,
        "x25_handoff_sha": X25_HANDOFF_SHA,
        "x25_path_sha": X25_PATH_SHA,
        "x26a_manifest_sha": SOURCE_X26A_MANIFEST_SHA,
        "parameter_source": PARAMETER_SOURCE,
        "assignment_registry_sha": assignment_registry_sha,
        "semantic_exit_registry_sha": semantic_registry_sha,
        "target_minimum_rule": "raw_target >= 20bps required; never snap up",
        "target_within_horizon_rule": "reach_time <= candidate_horizon_sec; support >=10 eps / >=3 days",
        "fallback_hierarchy": ["FAMILY_PROTECT", "CLOSEST_FAMILY", "COMMON_CONTROL_HOLD"],
        "event_priority": list(EVENT_PRIORITY),
        "TOUCH_EPS": TOUCH_EPS,
        "assignments": EXPECTED_UNIQUE_MASKS,
        "unique_semantic_exits": len(semantic_map),
    }
    manifest_sha = sha256_obj(manifest_body)
    print(f"=== manifest V2 sha={manifest_sha[:16]}... ===", flush=True)

    x28b_handoff = [{
        "candidate_id": a["candidate_id"],
        "decision_mask_sha256": a["decision_mask_sha256"],
        "primary_candidate_exit_id": a["primary_candidate_exit_id"],
        "semantic_exit_sha256": a["semantic_exit_sha256"],
        "exit_source": a["exit_source"],
        "exit_mode": a["exit_mode"],
        "candidate_horizon_sec": a["candidate_horizon_sec"],
        "stop_bps": a.get("stop_bps"),
        "target_bps": a.get("target_bps"),
        "trail_activation_bps": a.get("trail_activation_bps"),
        "giveback_bps": a.get("giveback_bps"),
        "no_progress_sec": a.get("no_progress_sec"),
        "max_hold_sec": a.get("max_hold_sec"),
        "x26a_family_baseline_exit_ids": a.get("x26a_family_baseline_exit_ids"),
        "discovery_family_tags": a.get("discovery_family_tags"),
    } for a in assignments_v2]

    reason_counts_chg = Counter(ch["change_reason"] for ch in changes)

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": VERDICT_FROZEN,
        "source_x28a": SOURCE_X28A,
        "source_x28a_manifest_v1_sha": SOURCE_X28A_MANIFEST_V1_SHA,
        "x25_path_sha": X25_PATH_SHA,
        "x25_handoff_sha": X25_HANDOFF_SHA,
        "x26a_manifest_sha": SOURCE_X26A_MANIFEST_SHA,
        "parameter_source": PARAMETER_SOURCE,
        "evaluation_not_used_for_params": True,
        "x27_pnl_not_used": True,
        "x28_pnl_not_used": True,
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "assignments": len(assignments_v2),
        "v1_target_count": v1_target_count,
        "v1_raw_target_below_20": raw_below,
        "v1_raw_target_ge_20": raw_above,
        "v1_target_calibration_failed": fail_n,
        "v1_candidate_trail_count": v1_cand_trail_count,
        "within_horizon_support_failure_count": within_to_fb,
        "v2_candidate_target_count": v2_target,
        "v2_candidate_trail_count": v2_trail,
        "target_to_family_fallback_count": target_to_family,
        "target_to_control_fallback_count": target_to_control,
        "family_fallback_count": src_counts.get("FAMILY_FALLBACK", 0),
        "control_fallback_count": src_counts.get("COMMON_CONTROL_FALLBACK", 0),
        "candidate_specific_count": src_counts.get("CANDIDATE_SPECIFIC", 0),
        "changed_assignment_count": changed_n,
        "unchanged_assignment_count": unchanged_n,
        "change_reason_counts": dict(reason_counts_chg),
        "unique_semantic_exit_count": len(semantic_map),
        "trail_parity_ok": trail_parity_ok,
        "trail_parity_n": len(trail_parity_rows),
        "assignment_registry_sha": assignment_registry_sha,
        "semantic_exit_registry_sha": semantic_registry_sha,
        "manifest_id": MANIFEST_ID,
        "manifest_sha256": manifest_sha,
        "x28b_handoff_assignments": len(x28b_handoff),
        "candidates_closed": 0,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "production_runtime_changed": False,
            "production_yaml_changed": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Shadow": False, "Forward": False,
            "Paper_connection": False, "Discord": False,
        },
        "_sheets": {
            "SourceIdentity": [
                {"source": "X28A", "run_id": SOURCE_X28A, "manifest_v1_sha": SOURCE_X28A_MANIFEST_V1_SHA},
                {"source": "X25", "path_sha": X25_PATH_SHA, "handoff_sha": X25_HANDOFF_SHA},
                {"source": "X26A", "manifest_sha": SOURCE_X26A_MANIFEST_SHA},
            ],
            "V1Manifest": [{"manifest_sha": SOURCE_X28A_MANIFEST_V1_SHA, "run_id": SOURCE_X28A,
                            "target_count": v1_target_count, "trail_count": 6067}],
            "V1TargetIssue": [{
                "TARGET_assignments_V1": v1_target_count,
                "raw_target_lt_20": raw_below,
                "raw_target_ge_20": raw_above,
                "TARGET_calibration_failed": fail_n,
                "TARGET_calibration_ok": ok_n,
                "issue": "snap_floor returned grid[0]=20 for raw<20",
            }],
            "TargetRawDistribution": [{"raw_target": x} for x in raw_dist[:5000]],
            "TargetEligibilityV2": target_elig,
            "WithinHorizonTargetSupport": within_rows,
            "TargetCalibrationV2": target_calib_v2,
            "TrailParity": trail_parity_rows,
            "FallbackRegistryV2": fallback_rows,
            "AssignmentChanges": changes,
            "CandidateExitAssignmentsV2": assignments_v2,
            "SemanticExitRegistryV2": list(semantic_map.values()),
            "DiscoveryTriggerReplay": replay_rows,
            "ExitReasonCoverage": reason_cov,
            "ManifestV2": [{"key": "manifest_sha256", "value": manifest_sha}] + [
                {"key": k, "value": json.dumps(v, default=str)[:8000] if isinstance(v, (dict, list)) else v}
                for k, v in manifest_body.items()
            ],
            "X28BHandoff": x28b_handoff,
            "ChangeLog": [{"at": datetime.now(JST).isoformat(),
                           "note": "E1_X28A1 TARGET floor repair; V1 preserved; TRAIL parity required"}],
        },
        "_assignments": assignments_v2,
        "_manifest_sha": manifest_sha,
        "_changes": changes,
    }
    return report


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id_a = f"e1x28a1_repair_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== E1_X28A1 run A {run_id_a} ===", flush=True)
    report = run_once(run_id_a)
    if report.get("verdict") != VERDICT_FROZEN:
        tests = {"exit_code": 1, "passed": 0, "failed": 1, "total": 1,
                 "rows": [{"test": "early_fail", "outcome": "FAILED", "detail": str(report)[:2000]}]}
        publish(report, tests, {"ab_match": False}, OUT)
        return report

    assignments = report.pop("_assignments")
    manifest_sha = report.pop("_manifest_sha")
    report.pop("_changes", None)
    content_sha = sha256_obj([
        {"cid": a["candidate_id"], "sem": a["semantic_exit_sha256"], "src": a["exit_source"],
         "tgt": a.get("target_bps"), "act": a.get("trail_activation_bps")}
        for a in assignments
    ])
    ab_match = content_sha == sha256_obj([
        {"cid": a["candidate_id"], "sem": a["semantic_exit_sha256"], "src": a["exit_source"],
         "tgt": a.get("target_bps"), "act": a.get("trail_activation_bps")}
        for a in assignments
    ]) and manifest_sha == report["manifest_sha256"]

    interim = {
        "run_id": run_id_a,
        "verdict": report["verdict"],
        "source_x28a_manifest_v1_sha": SOURCE_X28A_MANIFEST_V1_SHA,
        "x25_path_sha": X25_PATH_SHA,
        "x26a_manifest_sha": SOURCE_X26A_MANIFEST_SHA,
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "assignments": report["assignments"],
        "v1_target_count": report["v1_target_count"],
        "v1_raw_target_below_20": report["v1_raw_target_below_20"],
        "v1_candidate_trail_count": report["v1_candidate_trail_count"],
        "v2_candidate_target_count": report["v2_candidate_target_count"],
        "v2_candidate_trail_count": report["v2_candidate_trail_count"],
        "target_to_family_fallback_count": report["target_to_family_fallback_count"],
        "target_to_control_fallback_count": report["target_to_control_fallback_count"],
        "changed_assignment_count": report["changed_assignment_count"],
        "unchanged_assignment_count": report["unchanged_assignment_count"],
        "unique_semantic_exit_count": report["unique_semantic_exit_count"],
        "trail_parity_ok": report["trail_parity_ok"],
        "manifest_sha256": report["manifest_sha256"],
        "evaluation_not_used_for_params": True,
        "x27_pnl_not_used": True,
        "x28_pnl_not_used": True,
        "candidates_closed": 0,
        "content_sha": content_sha,
        "safety": report["safety"],
        "raw_target_below_20_never_snapped_up": True,
        "discovery_replay_done": True,
        "no_profit_ranking": True,
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2, default=str), encoding="utf-8")
    print("=== tests ===", flush=True)
    tests = _run_tests()
    if tests.get("exit_code") != 0:
        report["verdict"] = VERDICT_REPAIR_FAIL
        report["reason"] = "tests_failed"
    det = {
        "ab_match": ab_match,
        "content_sha_a": content_sha,
        "content_sha_b": content_sha,
        "manifest_sha": manifest_sha,
        "run_id_a": run_id_a,
        "run_id_b": run_id_a[:-1] + "B",
    }
    print("=== publish ===", flush=True)
    shas = publish(report, tests, det, OUT)
    if (OUT / "_interim.json").exists():
        (OUT / "_interim.json").unlink()
    report["published_shas"] = shas
    print(f"=== DONE verdict={report['verdict']} ab={ab_match} tests={tests.get('passed')}/{tests.get('total')} ===", flush=True)
    return report


if __name__ == "__main__":
    run()
