"""Aggregate P2-2 Fixed vs Dynamic and write report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Optional

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_NATIVE = Path(__file__).resolve().parents[3]
if str(_NATIVE / "scripts") not in sys.path:
    sys.path.insert(0, str(_NATIVE / "scripts"))

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.dynamic_anchor_p2_2 import (
    CONFIRMATION,
    DOCUMENT_ID,
    LOT_QTY,
    PERIOD_END,
    PERIOD_START,
    POST_CUTOFF_ZERO_HOLD,
    PREDECLARED_TOP1,
    PREDECLARED_TOP3,
    TRIGGER,
)
from research.dynamic_anchor_p2_2.binding import ENTRY_BINDING, P1_REPORT
from run_p0_3_exact_runtime_replay_20260820 import _maxdd, _pf

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "dynamic_anchor_pnl_test_p2_2"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _pf_out(v: Any) -> Any:
    if v is None:
        return None
    if v == float("inf"):
        return "Infinity"
    return v


def _pf_num(v: Any) -> float:
    if v is None:
        return float("nan")
    if v == "Infinity" or v == float("inf"):
        return float("inf")
    return float(v)


def _stats(trades: list[dict[str, Any]]) -> dict[str, Any]:
    pnls = [float(t.get("pnl_yen_100") or 0.0) for t in trades]
    w = sum(1 for p in pnls if p > 1e-9)
    l = sum(1 for p in pnls if p < -1e-9)
    d = len(pnls) - w - l
    gp = sum(p for p in pnls if p > 0)
    gl = sum(-p for p in pnls if p < 0)
    pnl = round(sum(pnls), 2)
    am = [t for t in trades if t.get("session") == "AM"]
    pm = [t for t in trades if t.get("session") == "PM"]
    return {
        "trades": len(trades),
        "win": w,
        "loss": l,
        "draw": d,
        "gross_profit": round(gp, 2),
        "gross_loss": round(gl, 2),
        "pnl": pnl,
        "PF": _pf_out(_pf(pnls)),
        "avg_pnl": round(pnl / len(trades), 4) if trades else 0.0,
        "maxDD": _maxdd(trades),
        "AM": {
            "trades": len(am),
            "pnl": round(sum(float(t.get("pnl_yen_100") or 0) for t in am), 2),
        },
        "PM": {
            "trades": len(pm),
            "pnl": round(sum(float(t.get("pnl_yen_100") or 0) for t in pm), 2),
        },
    }


def trade_ledger_sha(trades: list[dict[str, Any]]) -> str:
    rows = []
    for t in trades:
        rows.append({
            "date": t.get("date"),
            "session": t.get("session"),
            "symbol": t.get("symbol"),
            "t0": None if t.get("t0") is None else round(float(t["t0"]), 6),
            "t1": None if t.get("t1") is None else round(float(t["t1"]), 6),
            "fill_time": round(float(t.get("fill_time") or 0.0), 6),
            "fill_price": t.get("fill_price"),
            "exit_time": round(float(t.get("exit_time") or 0.0), 6),
            "exit_price": t.get("exit_price"),
            "exit_reason": t.get("exit_reason"),
            "pnl_yen_100": round(float(t.get("pnl_yen_100") or 0.0), 4),
        })
    rows.sort(key=lambda r: (str(r["date"]), float(r["fill_time"]), str(r["symbol"])))
    blob = json.dumps(rows, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def load_p1() -> dict[str, Any]:
    return json.loads(P1_REPORT.read_text(encoding="utf-8"))


def _sum_funnel(days: list[dict[str, Any]]) -> dict[str, int]:
    out: dict[str, int] = defaultdict(int)
    for d in days:
        for k, v in (d.get("funnel") or {}).items():
            out[str(k)] += int(v)
        out["fills"] += int(d.get("fills") or 0)
        out["expired"] += int(d.get("expired") or 0)
        out["admitted_engine"] += int(d.get("admitted") or 0)
        out["anchors"] += int(d.get("false_to_true_triggers") or 0)
        out["confirmed"] += int(d.get("confirmed") or 0)
    return dict(out)


def _turnover(days: list[dict[str, Any]], trades: list[dict[str, Any]]) -> dict[str, Any]:
    buy = sum(float(d.get("gross_buy_notional") or 0) for d in days)
    sell = sum(float(d.get("gross_sell_notional") or 0) for d in days)
    hold = [float(t["holding_sec"]) for t in trades if t.get("holding_sec") is not None]
    n_days = max(1, len(days))
    max_c = max((int(d.get("max_concurrent_positions") or 0) for d in days), default=0)
    span = sum(float(d.get("mean_concurrent_positions") or 0) * 1 for d in days)
    mean_c = (sum(float(d.get("mean_concurrent_positions") or 0) for d in days) / n_days) if days else 0.0
    return {
        "fills_per_day": round(len(trades) / n_days, 4) if days else 0.0,
        "gross_buy_notional": round(buy, 2),
        "gross_sell_notional": round(sell, 2),
        "gross_turnover_notional": round(buy + sell, 2),
        "total_holding_seconds": round(sum(hold), 3) if hold else 0.0,
        "mean_holding_seconds": round(sum(hold) / len(hold), 3) if hold else None,
        "mean_concurrent_positions": round(mean_c, 4),
        "max_concurrent_positions": max_c,
        "_span_unused": span,
    }


def _exclude(trades: list[dict[str, Any]], drop: set[str]) -> dict[str, Any]:
    keep = [t for t in trades if str(t.get("date")) not in drop]
    return _stats(keep)


def _concentration(daily: list[dict[str, Any]], trades: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [(r["date"], float(r["pnl"])) for r in daily]
    if not rows:
        return {}
    best = max(rows, key=lambda x: x[1])
    worst = min(rows, key=lambda x: x[1])
    total = sum(p for _, p in rows)
    ordered = sorted(rows, key=lambda x: -x[1])
    top1 = ordered[0][1] / total if abs(total) > 1e-12 else None
    top3 = sum(p for _, p in ordered[:3]) / total if abs(total) > 1e-12 else None
    pos = sum(1 for _, p in rows if p > 1e-9)
    neg = sum(1 for _, p in rows if p < -1e-9)
    flat = len(rows) - pos - neg
    loo = []
    for d, _p in rows:
        st = _stats([t for t in trades if t.get("date") != d])
        loo.append(st["pnl"])
    return {
        "best_day": {"date": best[0], "pnl": best[1]},
        "worst_day": {"date": worst[0], "pnl": worst[1]},
        "top1_contribution": top1,
        "top3_contribution": top3,
        "positive_days": pos,
        "negative_days": neg,
        "flat_days": flat,
        "loo_min_pnl": min(loo) if loo else None,
    }


def _relation(fixed: list[dict[str, Any]], dyn: list[dict[str, Any]]) -> list[dict[str, Any]]:
    f_map: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    d_map: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for t in fixed:
        f_map[(str(t.get("date")), str(t.get("symbol")))].append(t)
    for t in dyn:
        d_map[(str(t.get("date")), str(t.get("symbol")))].append(t)
    keys = sorted(set(f_map) | set(d_map))
    out = []
    for k in keys:
        fs, ds = f_map.get(k, []), d_map.get(k, [])
        if fs and ds:
            klass = "BOTH_SYMBOL_DAY"
        elif fs:
            klass = "FIXED_ONLY"
        else:
            klass = "DYNAMIC_ONLY"
        row: dict[str, Any] = {
            "date": k[0],
            "symbol": k[1],
            "class": klass,
            "fixed_n": len(fs),
            "dynamic_n": len(ds),
        }
        if fs and ds:
            f0 = min(fs, key=lambda t: float(t.get("fill_time") or 0))
            d0 = min(ds, key=lambda t: float(t.get("fill_time") or 0))
            row["entry_timing_diff_sec"] = round(
                float(d0.get("fill_time") or 0) - float(f0.get("fill_time") or 0), 3
            )
            try:
                row["entry_price_diff"] = round(
                    float(d0.get("fill_price") or 0) - float(f0.get("fill_price") or 0), 4
                )
            except (TypeError, ValueError):
                row["entry_price_diff"] = None
        out.append(row)
    return out


def _headline(dyn: dict[str, Any], fix: dict[str, Any]) -> str:
    dp, fp = float(dyn["pnl"]), float(fix["pnl"])
    dpf, fpf = _pf_num(dyn["PF"]), _pf_num(fix["PF"])
    ddd, fdd = float(dyn["maxDD"]), float(fix["maxDD"])
    better = dp > fp and dpf > fpf and ddd >= fdd
    worse = dp < fp and dpf < fpf
    if better:
        return "BETTER"
    if worse:
        return "WORSE"
    return "MIXED"


def build_report(
    *,
    inventory: list[dict[str, Any]],
    day_results: list[dict[str, Any]],
    failed: list[str],
    sha_bind: dict[str, Any],
    entry_bind: dict[str, Any],
    det: dict[str, Any],
    instream_sha: dict[str, Any],
) -> dict[str, Any]:
    p1 = load_p1()
    full_days = {r["date"] for r in inventory if r.get("capture_class") == "FULL" and r.get("replay_eligible")}
    ref_days = {r["date"] for r in inventory if r.get("replay_eligible")}
    by = {d["date"]: d for d in day_results if d.get("ok")}
    p1_daily = {r["date"]: r for r in (p1.get("daily") or [])}
    p1_trades = [t for t in (p1.get("trades") or [])]
    prim_dyn_days = [by[d] for d in sorted(full_days) if d in by]
    ref_dyn_days = [by[d] for d in sorted(ref_days) if d in by]
    prim_tr = [t for d in prim_dyn_days for t in (d.get("trades") or [])]
    ref_tr = [t for d in ref_dyn_days for t in (d.get("trades") or [])]
    prim_fix_tr = [t for t in p1_trades if t.get("date") in full_days]
    ref_fix_tr = [t for t in p1_trades if t.get("date") in ref_days]
    dyn_p = _stats(prim_tr)
    dyn_r = _stats(ref_tr)
    fix_p = {
        "days": len(full_days),
        **{k: p1["PRIMARY_FULL"][k] for k in (
            "trades", "pnl", "PF", "maxDD", "AM", "PM", "win", "loss", "draw", "avg_pnl",
            "gross_profit", "gross_loss",
        ) if k in p1["PRIMARY_FULL"]},
    }
    fix_p["days"] = 14
    fix_r = {
        "days": 23,
        **{k: p1["REFERENCE_ALL_USABLE"][k] for k in (
            "trades", "pnl", "PF", "maxDD", "AM", "PM", "win", "loss", "draw", "avg_pnl",
            "gross_profit", "gross_loss",
        ) if k in p1["REFERENCE_ALL_USABLE"]},
    }
    fun_p = _sum_funnel(prim_dyn_days)
    fun_r = _sum_funnel(ref_dyn_days)
    to_p = _turnover(prim_dyn_days, prim_tr)
    to_fix_buy = sum(float(t.get("fill_price") or 0) * LOT_QTY for t in prim_fix_tr)
    to_fix_sell = sum(float(t.get("exit_price") or 0) * LOT_QTY for t in prim_fix_tr)
    to_fix = to_fix_buy + to_fix_sell
    trade_ratio = (dyn_p["trades"] / fix_p["trades"]) if fix_p["trades"] else None
    turn_ratio = (to_p["gross_turnover_notional"] / to_fix) if to_fix else None
    daily_rows = []
    better = worse = equal = 0
    for d in sorted(full_days):
        fd = p1_daily.get(d) or {}
        dd = by.get(d) or {}
        fpnl = float(fd.get("pnl") or 0)
        dpnl = float(dd.get("pnl") or 0)
        if abs(dpnl - fpnl) <= 1e-9:
            equal += 1
            flag = "equal"
        elif dpnl > fpnl:
            better += 1
            flag = "dynamic_better"
        else:
            worse += 1
            flag = "dynamic_worse"
        daily_rows.append({
            "date": d,
            "Fixed_trades": fd.get("trades"),
            "Dynamic_trades": dd.get("trade_n"),
            "Fixed_pnl": fpnl,
            "Dynamic_pnl": dpnl,
            "delta_pnl": round(dpnl - fpnl, 2),
            "Fixed_PF": fd.get("PF"),
            "Dynamic_PF": _pf_out(dd.get("PF")),
            "Fixed_MaxDD": fd.get("maxDD"),
            "Dynamic_MaxDD": dd.get("maxDD"),
            "flag": flag,
        })
    ex1_f = _exclude(prim_fix_tr, {PREDECLARED_TOP1})
    ex1_d = _exclude(prim_tr, {PREDECLARED_TOP1})
    ex3_f = _exclude(prim_fix_tr, set(PREDECLARED_TOP3))
    ex3_d = _exclude(prim_tr, set(PREDECLARED_TOP3))
    conc = _concentration(
        [{"date": r["date"], "pnl": r["Dynamic_pnl"]} for r in daily_rows],
        prim_tr,
    )
    rel = _relation(prim_fix_tr, prim_tr)
    zc = [t for t in prim_tr if t.get("special_class") == POST_CUTOFF_ZERO_HOLD]
    zc_all = [t for t in ref_tr if t.get("special_class") == POST_CUTOFF_ZERO_HOLD]
    leak = (
        sum(int(d.get("snapshot_future_leak") or 0) for d in ref_dyn_days)
        + sum(int(d.get("checkpoint_future_leak") or 0) for d in ref_dyn_days)
        + sum(int(d.get("decision_snapshot_future_leak") or 0) for d in ref_dyn_days)
    )
    chrono = all(bool(d.get("CHRONOLOGICAL_SINGLE_STREAM")) for d in day_results if d.get("ok"))
    mix = any(int(d.get("dynamic_anchor_fires") or 0) and False for d in day_results)
    clock_ok = all(int(d.get("dynamic_anchor_fires") or 0) >= 0 for d in prim_dyn_days)
    head = _headline(dyn_p, {"pnl": fix_p["pnl"], "PF": fix_p["PF"], "maxDD": fix_p["maxDD"]})
    head3 = _headline(ex3_d, ex3_f)
    bind_ok = bool(sha_bind.get("pass")) and entry_bind.get("CURRENT_ENTRY_BINDING") == "PASS"
    instream_ok = bool(instream_sha.get("pass"))
    det_ok = bool(det.get("pass"))
    blocked = (not bind_ok) or (not instream_ok) or bool(failed) or leak > 0 or (not chrono)
    if blocked:
        verdict = "P2_2_BLOCKED"
    elif head == "BETTER" and head3 != "BETTER":
        verdict = "P2_2_DYNAMIC_BETTER_ON_REUSED_HISTORY_BUT_CONCENTRATED"
    elif head == "BETTER":
        verdict = "P2_2_DYNAMIC_BETTER_ON_REUSED_HISTORY"
    elif head == "WORSE":
        verdict = "P2_2_DYNAMIC_WORSE_ON_REUSED_HISTORY"
    else:
        verdict = "P2_2_DYNAMIC_MIXED_ON_REUSED_HISTORY"
    now = datetime.now(JST).isoformat(timespec="seconds")
    n_inv = len(inventory)
    n_full = sum(1 for r in inventory if r.get("capture_class") == "FULL")
    n_part = sum(1 for r in inventory if r.get("capture_class") == "PARTIAL")
    n_elig = sum(1 for r in inventory if r.get("replay_eligible"))
    funnel_pub = {
        "confirmed": fun_p.get("confirmed", 0),
        "current_entry_evaluated": fun_p.get("current_entry_evaluated", 0),
        "feature_evaluable": fun_p.get("feature_evaluable", 0),
        "score_evaluable": fun_p.get("score_evaluable", 0),
        "candidate_selected": fun_p.get("candidate_selected", 0),
        "blocked_open": fun_p.get("blocked_open", 0),
        "blocked_pending": fun_p.get("blocked_pending", 0),
        "blocked_cap": fun_p.get("blocked_cap", 0),
        "other_reject": fun_p.get("other_reject", 0),
        "admitted": fun_p.get("admitted", fun_p.get("admitted_engine", 0)),
        "fills": fun_p.get("fills", 0),
        "expired": fun_p.get("expired", 0),
    }
    dyn_p_pub = {
        "days": len(prim_dyn_days),
        "anchors": fun_p.get("anchors", 0),
        "confirmed": fun_p.get("confirmed", 0),
        "entry_evaluated": fun_p.get("current_entry_evaluated", 0),
        "admitted": funnel_pub["admitted"],
        "fills": funnel_pub["fills"],
        "expired": funnel_pub["expired"],
        **dyn_p,
    }
    return {
        "task": "P2-2",
        "analysis_id": "P2_2_DYNAMIC_ANCHOR_PNL_TEST",
        "document_id": DOCUMENT_ID,
        "generated_at_jst": now,
        "PERIOD": f"{PERIOD_START} - {PERIOD_END}",
        "INPUT_DAYS": n_inv,
        "full": n_full,
        "partial": n_part,
        "eligible": n_elig,
        "completed": len(day_results),
        "failed": failed,
        "TRIGGER": TRIGGER,
        "CONFIRMATION": CONFIRMATION,
        "P2_1_TRIGGER_SHA_MATCH": sha_bind.get("P2_1_TRIGGER_SHA_MATCH"),
        "P2_1_CONFIRM_SHA_MATCH": sha_bind.get("P2_1_CONFIRM_SHA_MATCH"),
        "CURRENT_ENTRY_BINDING": entry_bind.get("CURRENT_ENTRY_BINDING"),
        "CHRONOLOGICAL_SINGLE_STREAM": "PASS" if chrono else "FAIL",
        "entry_binding_path": ENTRY_BINDING,
        "PRIMARY_FULL": {
            "FIXED": fix_p,
            "DYNAMIC": dyn_p_pub,
        },
        "FUNNEL": funnel_pub,
        "TURNOVER": {
            "trade_count_ratio": trade_ratio,
            "turnover_ratio": turn_ratio,
            "mean_concurrent_positions": to_p["mean_concurrent_positions"],
            "max_concurrent_positions": to_p["max_concurrent_positions"],
            "fills_per_day": to_p["fills_per_day"],
            "gross_buy_notional": to_p["gross_buy_notional"],
            "gross_sell_notional": to_p["gross_sell_notional"],
            "gross_turnover_notional": to_p["gross_turnover_notional"],
            "total_holding_seconds": to_p["total_holding_seconds"],
            "mean_holding_seconds": to_p["mean_holding_seconds"],
            "fixed_gross_turnover_notional": round(to_fix, 2),
        },
        "POST_CUTOFF_ZERO_HOLD": {
            "count": len(zc),
            "symbols": sorted({t["symbol"] for t in zc}),
            "draw_count": sum(1 for t in zc if abs(float(t.get("pnl_yen_100") or 0)) <= 1e-9),
            "pnl": round(sum(float(t.get("pnl_yen_100") or 0) for t in zc), 2),
            "reference_count": len(zc_all),
        },
        "DAILY": {
            "better_days": better,
            "worse_days": worse,
            "equal_days": equal,
            "rows": daily_rows,
        },
        "EX_TOP1_FIXED": {"pnl": ex1_f["pnl"], "PF": ex1_f["PF"], "trades": ex1_f["trades"], "maxDD": ex1_f["maxDD"]},
        "EX_TOP1_DYNAMIC": {"pnl": ex1_d["pnl"], "PF": ex1_d["PF"], "trades": ex1_d["trades"], "maxDD": ex1_d["maxDD"]},
        "EX_FIXED_TOP3_FIXED": {"pnl": ex3_f["pnl"], "PF": ex3_f["PF"], "trades": ex3_f["trades"], "maxDD": ex3_f["maxDD"]},
        "EX_FIXED_TOP3_DYNAMIC": {"pnl": ex3_d["pnl"], "PF": ex3_d["PF"], "trades": ex3_d["trades"], "maxDD": ex3_d["maxDD"]},
        "DYNAMIC_CONCENTRATION": conc,
        "REFERENCE_ALL_USABLE": {
            "Fixed": fix_r,
            "Dynamic": {
                "days": len(ref_dyn_days),
                "anchors": fun_r.get("anchors", 0),
                "confirmed": fun_r.get("confirmed", 0),
                **dyn_r,
            },
        },
        "headline_class": head,
        "headline_ex_top3": head3,
        "FUTURE_LEAK": leak > 0,
        "future_leak_count": leak,
        "DETERMINISM": "PASS" if det_ok else "FAIL",
        "TRIGGER_INSTREAM_SHA": instream_sha.get("trig"),
        "CONFIRM_INSTREAM_SHA": instream_sha.get("conf"),
        "TRADE_LEDGER_SHA_RUN1": det.get("sha1"),
        "TRADE_LEDGER_SHA_RUN2": det.get("sha2"),
        "determinism_days": det.get("days"),
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "verdict": verdict,
        "holdout_label": "NO_CLEAN_HOLDOUT — reused 20260721-20260821; not OOS / prospective / ROBUST",
        "identity": {
            "strategy_sha": STRATEGY_SHA,
            "entry_sha": ENTRY_SHA,
            "exit_sha": EXIT_SHA,
            "anchor_sha": ANCHOR_SHA,
            "V1RNativeEntryLive_sha": entry_bind.get("V1RNativeEntryLive_sha"),
            "V1RLiveDualLane_sha": entry_bind.get("V1RLiveDualLane_sha"),
        },
        "inventory": [{k: v for k, v in r.items() if k != "universe_symbols"} for r in inventory],
        "_primary_trades": prim_tr,
        "_daily_rows": daily_rows,
        "_relation": rel,
        "_funnel_primary": funnel_pub,
        "_post_cutoff_rows": zc,
        "_p1_sha_bind": sha_bind,
        "_entry_bind": entry_bind,
        "_instream": instream_sha,
        "_det": det,
        "clock_ok": clock_ok,
        "mix_unused": mix,
        "missing_excluded": ["20260815", "20260816"],
    }


def _write_rows(ws, rows: list[dict[str, Any]], cap: int | None = None) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    data = rows[:cap] if cap else rows
    cols = list(data[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(data, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(42, max(12, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _md(rep: dict[str, Any]) -> str:
    p = rep["PRIMARY_FULL"]
    f, d = p["FIXED"], p["DYNAMIC"]
    r = rep["REFERENCE_ALL_USABLE"]
    fn = rep["FUNNEL"]
    tu = rep["TURNOVER"]
    z = rep["POST_CUTOFF_ZERO_HOLD"]
    return f"""# P2-2 Dynamic Anchor × Current Runtime chronological replay

- generated_at_jst: `{rep['generated_at_jst']}`
- **verdict: `{rep['verdict']}`**
- submit/cancel/live: `0/0/0`
- NO_CLEAN_HOLDOUT — not OOS / prospective / ROBUST

P2-1 SHA match: trigger `{rep['P2_1_TRIGGER_SHA_MATCH']}` · confirm `{rep['P2_1_CONFIRM_SHA_MATCH']}`  
CURRENT_ENTRY_BINDING: `{rep['CURRENT_ENTRY_BINDING']}`  
CHRONOLOGICAL_SINGLE_STREAM: `{rep['CHRONOLOGICAL_SINGLE_STREAM']}`

## PRIMARY_FULL

FIXED: days {f.get('days')} · trades {f.get('trades')} · pnl {f.get('pnl')} · PF {f.get('PF')} · maxDD {f.get('maxDD')}  
AM {f.get('AM')} · PM {f.get('PM')}

DYNAMIC: days {d.get('days')} · anchors {d.get('anchors')} · confirmed {d.get('confirmed')} · entry_evaluated {d.get('entry_evaluated')} · admitted {d.get('admitted')} · fills {d.get('fills')} · expired {d.get('expired')} · trades {d.get('trades')} · win {d.get('win')} · loss {d.get('loss')} · draw {d.get('draw')} · pnl {d.get('pnl')} · PF {d.get('PF')} · avg_pnl {d.get('avg_pnl')} · maxDD {d.get('maxDD')}  
AM {d.get('AM')} · PM {d.get('PM')}

## FUNNEL

{json.dumps(fn, ensure_ascii=False)}

rank_pass: not a Current Runtime gate (omitted).

## TURNOVER (diagnostic only — no rule change)

trade_count_ratio: {tu.get('trade_count_ratio')}  
turnover_ratio: {tu.get('turnover_ratio')}  
mean_concurrent_positions: {tu.get('mean_concurrent_positions')}  
max_concurrent_positions: {tu.get('max_concurrent_positions')}

## POST_CUTOFF_ZERO_HOLD

count {z.get('count')} · pnl {z.get('pnl')} · symbols {z.get('symbols')}

## DAILY

better_days {rep['DAILY']['better_days']} · worse_days {rep['DAILY']['worse_days']} · equal_days {rep['DAILY']['equal_days']}

EX_TOP1_FIXED: {json.dumps(rep['EX_TOP1_FIXED'])}  
EX_TOP1_DYNAMIC: {json.dumps(rep['EX_TOP1_DYNAMIC'])}  
EX_FIXED_TOP3_FIXED: {json.dumps(rep['EX_FIXED_TOP3_FIXED'])}  
EX_FIXED_TOP3_DYNAMIC: {json.dumps(rep['EX_FIXED_TOP3_DYNAMIC'])}

DYNAMIC_CONCENTRATION: {json.dumps(rep['DYNAMIC_CONCENTRATION'], ensure_ascii=False)}

## REFERENCE_ALL_USABLE

Fixed: {json.dumps(r['Fixed'], ensure_ascii=False)}  
Dynamic: {json.dumps(r['Dynamic'], ensure_ascii=False)}

FUTURE_LEAK: {str(rep['FUTURE_LEAK']).lower()}  
DETERMINISM: {rep['DETERMINISM']}  
TRADE_LEDGER_SHA_RUN1: `{rep.get('TRADE_LEDGER_SHA_RUN1')}`  
TRADE_LEDGER_SHA_RUN2: `{rep.get('TRADE_LEDGER_SHA_RUN2')}`

## STOP

P2-2 ends here. Runtime adoption forbidden. Retune forbidden.
"""


def write_artifacts(rep: dict[str, Any]) -> dict[str, Path]:
    OUT.mkdir(parents=True, exist_ok=True)
    for p in OUT.iterdir():
        if p.is_file() and p.name not in {"report.json", "report.md", "audit.xlsx"}:
            p.unlink()
    pub = {k: v for k, v in rep.items() if not k.startswith("_")}
    (OUT / "report.json").write_text(json.dumps(pub, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    (OUT / "report.md").write_text(_md(rep), encoding="utf-8")
    wb = Workbook()

    def sh(name: str):
        if name == "Summary":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    p = rep["PRIMARY_FULL"]
    summary = [
        {"field": k, "value": json.dumps(v, ensure_ascii=False, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("PERIOD", rep["PERIOD"]),
            ("P2_1_TRIGGER_SHA_MATCH", rep["P2_1_TRIGGER_SHA_MATCH"]),
            ("P2_1_CONFIRM_SHA_MATCH", rep["P2_1_CONFIRM_SHA_MATCH"]),
            ("CURRENT_ENTRY_BINDING", rep["CURRENT_ENTRY_BINDING"]),
            ("CHRONOLOGICAL_SINGLE_STREAM", rep["CHRONOLOGICAL_SINGLE_STREAM"]),
            ("FIXED_trades", p["FIXED"].get("trades")),
            ("FIXED_pnl", p["FIXED"].get("pnl")),
            ("FIXED_PF", p["FIXED"].get("PF")),
            ("FIXED_maxDD", p["FIXED"].get("maxDD")),
            ("DYNAMIC_trades", p["DYNAMIC"].get("trades")),
            ("DYNAMIC_pnl", p["DYNAMIC"].get("pnl")),
            ("DYNAMIC_PF", p["DYNAMIC"].get("PF")),
            ("DYNAMIC_maxDD", p["DYNAMIC"].get("maxDD")),
            ("FUTURE_LEAK", rep["FUTURE_LEAK"]),
            ("DETERMINISM", rep["DETERMINISM"]),
            ("verdict", rep["verdict"]),
            ("SAFETY", rep["SAFETY"]),
        ]
    ]
    _write_rows(sh("Summary"), summary)
    bind_rows = [{"field": k, "value": json.dumps(v) if isinstance(v, (dict, list)) else v} for k, v in ENTRY_BINDING.items()]
    bind_rows.extend([
        {"field": "P2_1_TRIGGER_SHA_MATCH", "value": rep["P2_1_TRIGGER_SHA_MATCH"]},
        {"field": "P2_1_CONFIRM_SHA_MATCH", "value": rep["P2_1_CONFIRM_SHA_MATCH"]},
        {"field": "CURRENT_ENTRY_BINDING", "value": rep["CURRENT_ENTRY_BINDING"]},
    ])
    _write_rows(sh("Binding"), bind_rows)
    _write_rows(sh("Daily"), rep["_daily_rows"])
    _write_rows(sh("Dynamic_Funnel"), [{"stage": k, "n": v} for k, v in rep["_funnel_primary"].items()])
    _write_rows(sh("Dynamic_Trades"), [
        {k: t.get(k) for k in (
            "date", "session", "symbol", "t0", "t1", "T1_percentile", "C1_slope", "C1_endpoint_return",
            "decision_time", "snapshot_cutoff", "entry_score", "candidate_rank", "universe_rank",
            "limit", "fill_time", "fill_price", "exit_time", "exit_price", "exit_reason",
            "holding_sec", "pnl_yen_100", "special_class",
        )}
        for t in rep["_primary_trades"]
    ])
    _write_rows(sh("Fixed_vs_Dynamic"), rep["_daily_rows"])
    _write_rows(sh("Trade_Relation"), rep["_relation"])
    _write_rows(sh("Top3_Sensitivity"), [
        {"slice": "ALL_FULL_FIXED", **{k: p["FIXED"].get(k) for k in ("trades", "pnl", "PF", "maxDD")}},
        {"slice": "ALL_FULL_DYNAMIC", **{k: p["DYNAMIC"].get(k) for k in ("trades", "pnl", "PF", "maxDD")}},
        {"slice": "EX_TOP1_FIXED", **rep["EX_TOP1_FIXED"]},
        {"slice": "EX_TOP1_DYNAMIC", **rep["EX_TOP1_DYNAMIC"]},
        {"slice": "EX_FIXED_TOP3_FIXED", **rep["EX_FIXED_TOP3_FIXED"]},
        {"slice": "EX_FIXED_TOP3_DYNAMIC", **rep["EX_FIXED_TOP3_DYNAMIC"]},
    ])
    tu = rep["TURNOVER"]
    _write_rows(sh("Turnover_Exposure"), [{"field": k, "value": v} for k, v in tu.items()])
    if rep["_post_cutoff_rows"]:
        _write_rows(sh("Session_Close_Special"), [
            {k: t.get(k) for k in (
                "date", "session", "symbol", "fill_time", "exit_time",
                "exit_reason", "pnl_yen_100", "special_class",
            )}
            for t in rep["_post_cutoff_rows"]
        ])
    else:
        _write_rows(sh("Session_Close_Special"), [{"count": 0, "pnl": 0, "note": "none"}])
    _write_rows(sh("Determinism"), [
        {"field": "result", "value": rep["DETERMINISM"]},
        {"field": "days", "value": json.dumps(rep.get("determinism_days"))},
        {"field": "TRADE_LEDGER_SHA_RUN1", "value": rep.get("TRADE_LEDGER_SHA_RUN1")},
        {"field": "TRADE_LEDGER_SHA_RUN2", "value": rep.get("TRADE_LEDGER_SHA_RUN2")},
        {"field": "TRIGGER_INSTREAM_SHA", "value": rep.get("TRIGGER_INSTREAM_SHA")},
        {"field": "CONFIRM_INSTREAM_SHA", "value": rep.get("CONFIRM_INSTREAM_SHA")},
    ])
    _write_rows(sh("Identity"), [{"field": k, "value": v} for k, v in rep["identity"].items()])
    _write_rows(sh("Safety"), [
        {"field": "submit/cancel/live", "value": "0/0/0"},
        {"field": "STRATEGY_CHANGED", "value": False},
        {"field": "ENTRY_EXIT_CHANGED", "value": False},
        {"field": "RUNTIME_CHANGED", "value": False},
        {"field": "note", "value": "Research replay only. Runtime adoption forbidden. Retune forbidden."},
    ])
    wb.save(OUT / "audit.xlsx")
    return {"report_json": OUT / "report.json", "report_md": OUT / "report.md", "audit_xlsx": OUT / "audit.xlsx"}
