#!/usr/bin/env python3
"""Offline A=fixed-anchor vs B=event-driven V1R ENTRY. Timing only. No runtime adopt."""
from __future__ import annotations

import csv
import json
import math
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

NATIVE = Path(__file__).resolve().parents[3]
REPO = NATIVE.parent
SRC = NATIVE / "src"
OUT = NATIVE / "results" / "research" / "anchor_vs_event_driven_v1r"
JST = ZoneInfo("Asia/Tokyo")
MAX_WORKERS = 4

for _k in (
    "KABU_V1R_ENTRY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_NOTIFY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_CAP_BLOCKED_WEBHOOK_URL",
    "KABU_DISCORD_RESEARCH_WEBHOOK_URL",
    "KABU_SHADOW_DISCORD_WEBHOOK_URL",
    "KABU_DISCORD_MARKET_CAPTURE_WEBHOOK_URL",
    "KABU_MARKET_CAPTURE_WEBHOOK_URL",
):
    os.environ.pop(_k, None)
os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"
os.environ["PYTHONPATH"] = f"{SRC};{REPO}" if os.name == "nt" else f"{SRC}:{REPO}"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from small_paper.v1r_primary_runtime import CLOCK_GRID  # noqa: E402

# Live ENTRY features (must match v1r_native_entry_live.FEATURE_ORDER).
LIVE_FEATURES = (
    "spread_bps",
    "imbalance",
    "mid_ret_60s",
    "mid_ret_180s",
    "event_rate_60s",
    "log_bid_qty",
)


def _fmt_yen(v: float) -> str:
    n = int(round(float(v)))
    sign = "+" if n > 0 else ""
    return f"{sign}{n:,}"


def _pf(pnls: list[float]) -> Optional[float]:
    wins = sum(x for x in pnls if x > 0)
    losses = sum(-x for x in pnls if x < 0)
    if losses <= 1e-12:
        return None if wins <= 1e-12 else float("inf")
    return wins / losses


def _maxdd(trades: list[dict[str, Any]]) -> float:
    ordered = sorted(
        trades,
        key=lambda t: (float(t.get("exit_time") or t.get("entry_time") or 0.0), str(t.get("symbol") or "")),
    )
    eq = 0.0
    peak = 0.0
    dd = 0.0
    for t in ordered:
        eq += float(t.get("pnl_yen_100") or 0.0)
        peak = max(peak, eq)
        dd = min(dd, eq - peak)
    return round(dd, 2)


def _parse_iso(v: Any) -> Optional[float]:
    if not v:
        return None
    try:
        dt = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=JST)
        return dt.astimezone(JST).timestamp()
    except Exception:
        return None


def _hm_epoch(day: str, hour: int, minute: int) -> float:
    return datetime(
        int(day[:4]), int(day[4:6]), int(day[6:8]), hour, minute, tzinfo=JST
    ).timestamp()


def next_anchor_after(day: str, t: float) -> tuple[Optional[float], str]:
    for h, m in CLOCK_GRID:
        ts = _hm_epoch(day, h, m)
        if ts > float(t) + 1e-9:
            return ts, f"{h:02d}:{m:02d}"
    return None, ""


def in_session(t: float) -> bool:
    dt = datetime.fromtimestamp(float(t), JST)
    hm = dt.hour * 60 + dt.minute
    if 9 * 60 <= hm < 11 * 60 + 30:
        return True
    if 12 * 60 + 30 <= hm <= 15 * 60:
        return True
    return False


def _bare(sym: Any) -> str:
    return str(sym or "").replace(".T", "").strip().upper()


def _load_json(path: Path) -> dict[str, Any]:
    try:
        body = json.loads(path.read_text(encoding="utf-8"))
        return body if isinstance(body, dict) else {}
    except Exception:
        return {}


def historical_universe(day: str, capture_dir: Path) -> tuple[list[str], str]:
    man = capture_dir / "registration_manifest.json"
    if not man.is_file():
        man = Path(capture_dir) / ".." / "registration_manifest.json"
        man = man.resolve() if man.exists() else capture_dir / "registration_manifest.json"
    for cand in (
        capture_dir / "registration_manifest.json",
        capture_dir.parent / "registration_manifest.json",
    ):
        if cand.is_file():
            body = _load_json(cand)
            syms = [_bare(s) for s in (body.get("registered_symbols") or body.get("symbols") or [])]
            syms = [s for s in dict.fromkeys(syms) if s]
            if syms:
                return syms, f"registration:{cand.name}"
    frozen = NATIVE / "runtime" / f"same_day_am_frozen_universe_{day}.json"
    if frozen.is_file():
        body = _load_json(frozen)
        syms = [_bare(s) for s in (body.get("canonical_symbols") or [])]
        if syms:
            return list(dict.fromkeys(syms)), f"frozen:{frozen.name}"
    csv_path = NATIVE / "results" / "reports" / f"universe_core10_dynamic40_price_risk_am_{day}.csv"
    if csv_path.is_file():
        syms: list[str] = []
        with csv_path.open(encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                s = _bare(row.get("symbol") or row.get("code") or "")
                if s:
                    syms.append(s)
        if syms:
            return list(dict.fromkeys(syms)), f"am_csv:{csv_path.name}"
    summary = _load_json(capture_dir / "capture_summary.json") or _load_json(capture_dir.parent / "capture_summary.json")
    seen = [_bare(s) for s in (summary.get("symbols_seen") or [])]
    if seen:
        return list(dict.fromkeys(seen)), "capture_summary.symbols_seen"
    return [], "missing"


def find_capture_dir(day: str) -> Optional[Path]:
    root = NATIVE / "data" / "market_capture" / day
    if not root.is_dir():
        return None
    best: Optional[Path] = None
    best_n = -1
    for sess in root.iterdir():
        if not (sess.is_dir() and sess.name.startswith("session_ing_")):
            continue
        n = sum(p.stat().st_size for p in sess.glob("push_part_*.jsonl"))
        if n > best_n:
            best, best_n = sess, n
    if best is not None and best_n > 0:
        return best
    day_n = sum(p.stat().st_size for p in root.glob("push_part_*.jsonl"))
    if day_n > 0:
        return root
    return None


def _peek_times(capture_dir: Path) -> tuple[str, str, int, int]:
    summary = _load_json(capture_dir / "capture_summary.json")
    if not summary:
        summary = _load_json(capture_dir.parent / "capture_summary.json")
    first = str(summary.get("first_event_at") or summary.get("first_event_time") or "")
    last = str(summary.get("last_event_at") or summary.get("last_event_time") or "")
    n_ev = int(summary.get("total_events") or summary.get("writer", {}).get("written") or 0)
    parts = sorted(capture_dir.glob("push_part_*.jsonl"))
    size = sum(p.stat().st_size for p in parts)
    if first and last:
        return first, last, n_ev, size
    if not parts:
        return "", "", 0, 0
    nonempty = [p for p in parts if p.stat().st_size > 0]
    if not nonempty:
        return "", "", 0, size

    def first_line(path: Path) -> str:
        with path.open(encoding="utf-8") as fh:
            for line in fh:
                if line.strip():
                    try:
                        rec = json.loads(line)
                    except Exception:
                        continue
                    return str(
                        rec.get("received_at")
                        or rec.get("event_time")
                        or rec.get("received_at_jst")
                        or rec.get("persisted_at")
                        or rec.get("received_at_utc")
                        or rec.get("current_price_time")
                        or ""
                    )
        return ""

    def last_line(path: Path) -> str:
        with path.open("rb") as fh:
            fh.seek(0, 2)
            size_b = fh.tell()
            fh.seek(max(0, size_b - 65536))
            chunk = fh.read().decode("utf-8", errors="replace")
        for line in reversed(chunk.splitlines()):
            if not line.strip():
                continue
            try:
                rec = json.loads(line)
            except Exception:
                continue
            return str(
                rec.get("received_at")
                or rec.get("event_time")
                or rec.get("received_at_jst")
                or rec.get("persisted_at")
                or rec.get("received_at_utc")
                or rec.get("current_price_time")
                or ""
            )
        return ""

    if not first:
        first = first_line(nonempty[0])
    if not last:
        last = last_line(nonempty[-1])
    return first, last, n_ev, size


def classify_capture(day: str, capture_dir: Optional[Path], universe_n: int) -> dict[str, Any]:
    wd = datetime(int(day[:4]), int(day[4:6]), int(day[6:8])).weekday()
    jpx = wd < 5
    if day in {"20260811", "20260815", "20260816"}:
        jpx = False
    if capture_dir is None:
        return {
            "date": day,
            "capture_class": "INVALID" if jpx else "INVALID",
            "jpx_trading_day": jpx,
            "exclusion_reason": "NO_CAPTURE" if jpx else "NON_TRADING_DAY",
            "usable": False,
            "full": False,
            "universe_n": universe_n,
            "first_event_time": "",
            "last_event_time": "",
            "size_bytes": 0,
            "capture_path": "",
        }
    first, last, n_ev, size = _peek_times(capture_dir)
    ft = _parse_iso(first)
    lt = _parse_iso(last)
    am = False
    pm = False
    if ft is not None:
        dt = datetime.fromtimestamp(ft, JST)
        am = dt.hour < 10 or (dt.hour == 9) or (dt.hour < 12 and dt.hour >= 9)
        am = (dt.hour * 60 + dt.minute) <= 9 * 60 + 15
    if lt is not None:
        dt = datetime.fromtimestamp(lt, JST)
        pm = (dt.hour * 60 + dt.minute) >= 14 * 60 + 50
    prior = NATIVE / "results" / "research" / "current_logic_retrospective_20260810_20260819" / "capture_inventory.json"
    prior_class = ""
    if prior.is_file():
        row = (_load_json(prior).get(day) or {})
        prior_class = str(row.get("capture_class") or "")
        if not first:
            first = str(row.get("first_event_time") or "")
        if not last:
            last = str(row.get("last_event_time") or "")
        if row.get("am_coverage") is True:
            am = True
        if row.get("am_coverage") is False:
            am = False
        if row.get("pm_coverage") is True:
            pm = True
        if row.get("pm_coverage") is False:
            pm = False
        if row.get("size_bytes"):
            size = int(row.get("size_bytes") or size)
    if size <= 0:
        klass = "INVALID"
        reason = "EMPTY_PUSH"
        usable = False
        full = False
    elif prior_class == "DEGRADED_UNIVERSE" or (am and pm and universe_n and universe_n != 50):
        klass = "DEGRADED"
        reason = prior_class or f"universe_{universe_n}_not_50"
        usable = True
        full = False
    elif am and pm and (universe_n == 50 or universe_n == 0):
        klass = "FULL"
        reason = ""
        usable = True
        full = True
        if universe_n == 0:
            klass = "PARTIAL"
            full = False
            usable = True
            reason = "UNIVERSE_UNKNOWN"
    elif am or pm:
        klass = "PARTIAL"
        reason = "PARTIAL_WINDOW" if not (am and pm) else ""
        usable = True
        full = False
    else:
        klass = "PARTIAL"
        reason = "PARTIAL_WINDOW"
        usable = True
        full = False
    if not jpx and klass != "FULL":
        if size <= 0:
            klass = "INVALID"
            reason = "NON_TRADING_DAY"
            usable = False
            full = False
    return {
        "date": day,
        "capture_class": klass,
        "jpx_trading_day": jpx,
        "exclusion_reason": reason,
        "usable": usable,
        "full": full,
        "universe_n": universe_n,
        "first_event_time": first,
        "last_event_time": last,
        "size_bytes": size,
        "event_hint": n_ev,
        "am_coverage": am,
        "pm_coverage": pm,
        "capture_path": str(capture_dir),
        "prior_class": prior_class,
    }


def record_event_stamp(rec: dict[str, Any]) -> str:
    for k in (
        "received_at",
        "event_time",
        "received_at_jst",
        "persisted_at",
        "received_at_utc",
        "current_price_time",
    ):
        v = rec.get(k)
        if v:
            return str(v)
    return ""


def capture_event_epoch(rec: dict[str, Any], pay: dict[str, Any]) -> Optional[float]:
    stamp = record_event_stamp(rec) or record_event_stamp(pay)
    if not stamp:
        for k in ("CurrentPriceTime", "AskTime", "BidTime"):
            v = pay.get(k)
            if v:
                stamp = str(v)
                break
    return _parse_iso(stamp)


def iter_push(capture_dir: Path):
    for part in sorted(capture_dir.glob("push_part_*.jsonl")):
        if part.stat().st_size <= 0:
            continue
        with part.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                if rec.get("kind") not in (None, "market_push"):
                    continue
                yield rec


def extract_trades(dual: Any) -> list[dict[str, Any]]:
    from small_paper.v1r_live_dual_lane import canonical_symbol_key

    trades: list[dict[str, Any]] = []
    admits: dict[str, list[dict[str, Any]]] = {}
    for row in dual.traces:
        extra = row.get("extra") if isinstance(row.get("extra"), dict) else {}
        merged = {**row, **extra}
        sym = canonical_symbol_key(merged.get("symbol"))
        lane = str(merged.get("lane") or "primary")
        ev = str(merged.get("event") or "")
        if ev == "ADMIT" and lane == "primary":
            admits.setdefault(sym, []).append(
                {
                    "fill_time": merged.get("fill_time"),
                    "fill_price": merged.get("fill_price"),
                }
            )
            continue
        if ev != "EXIT_EXECUTED" or lane != "primary":
            continue
        adm = admits.get(sym).pop(0) if admits.get(sym) else {}
        entry_t = adm.get("fill_time")
        exit_t = merged.get("exit_time")
        entry_px = float(adm.get("fill_price") or merged.get("fill_price") or merged.get("entry_price") or 0)
        exit_px = float(merged.get("exit_price") or 0)
        pnl = merged.get("pnl_yen")
        if pnl is None and entry_px and exit_px:
            pnl = (exit_px - entry_px) * 100.0
        sess = "AM"
        tref = entry_t if entry_t is not None else exit_t
        if tref is not None:
            dt = datetime.fromtimestamp(float(tref), JST)
            sess = "AM" if dt.hour < 12 else "PM"
        trades.append(
            {
                "symbol": sym,
                "session": sess,
                "entry_time": entry_t,
                "exit_time": exit_t,
                "entry_price": entry_px,
                "exit_price": exit_px,
                "pnl_yen_100": float(pnl or 0.0),
                "reason": merged.get("reason"),
            }
        )
    return trades


def _boot(universe: list[str], engine_cls: type):
    from small_paper.v1r_live_dual_lane import ensure_dual_lane, reset_dual_lane_for_tests
    from small_paper.v1r_native_entry_live import (
        boot_v1r_native_entry,
        reset_native_entry_for_tests,
        set_native_entry,
    )

    reset_native_entry_for_tests()
    reset_dual_lane_for_tests()
    base = boot_v1r_native_entry(universe=universe, trace_dir=None, universe_source="historical_capture")
    if not base.ready:
        return base, None
    eng = engine_cls(
        universe=list(base.universe),
        score_fn=base.score_fn,
        model_ser=base.model_ser,
        trace_dir=None,
        ready=True,
        universe_source=base.universe_source,
    )
    eng.notify_enabled = False
    set_native_entry(eng)
    dual = ensure_dual_lane(trace_dir=None)
    return eng, dual


class AnchorEngine:
    """Mixin marker — actual subclass built inside worker to avoid pickling issues."""


def _stream_day(day: str, capture_dir: Path, eng: Any, dual: Any) -> tuple[int, Optional[float]]:
    from research.e1_x22_actual_exit_factory.paths import session_end_epoch
    from small_paper.v1r_live_dual_lane import canonical_symbol_key

    last_dual_t: dict[str, float] = {}
    events_n = 0
    last_et: Optional[float] = None
    am_end = session_end_epoch(day, "AM")
    pm_end = session_end_epoch(day, "PM")
    am_closed = False
    pm_closed = False

    def maybe_close(t: float) -> None:
        nonlocal am_closed, pm_closed
        if not am_closed and t + 1e-9 >= am_end:
            dual.close_open_at_session_end(event_t=am_end, session="AM")
            eng.on_tick_fill_check(event_t=am_end)
            am_closed = True
        if not pm_closed and t + 1e-9 >= pm_end:
            dual.close_open_at_session_end(event_t=pm_end, session="PM")
            eng.on_tick_fill_check(event_t=pm_end)
            pm_closed = True

    for rec in iter_push(capture_dir):
        recv = record_event_stamp(rec)
        seq = int(rec.get("sequence") or 0)
        sym = _bare(rec.get("symbol"))
        pay = dict(rec.get("payload") or rec.get("original_payload") or {})
        et = capture_event_epoch(rec, pay)
        if et is None:
            continue
        recv = recv or datetime.fromtimestamp(float(et), JST).isoformat(timespec="milliseconds")
        pay["received_at"] = recv
        pay["recorded_at"] = recv
        pay["sequence"] = seq
        pay["__ingress_sequence__"] = seq
        pay["__ingress_received_at__"] = recv
        last_et = float(et)
        events_n += 1
        maybe_close(float(et))
        eng.process_market_push(symbol=sym, payload=pay, event_t=float(et))
        if hasattr(eng, "on_stream_event"):
            eng.on_stream_event(sym, float(et))
        key = canonical_symbol_key(sym)
        open_here = (key in dual.primary and not dual.primary[key].closed) or (
            key in dual.control and not dual.control[key].closed
        )
        if open_here:
            prev = last_dual_t.get(key)
            if prev is None or (float(et) - prev) >= 0.5 - 1e-12:
                last_dual_t[key] = float(et)
                dual.on_tick(symbol=sym, payload=pay, event_t=float(et), push_sequence=seq)
        if events_n % 200000 == 0:
            eng.events.clear()
            if getattr(eng, "ingest_audit", None) is not None:
                eng.ingest_audit.clear()
    if last_et is not None:
        maybe_close(float(last_et))
    return events_n, last_et


def replay_day(payload: dict[str, Any]) -> dict[str, Any]:
    if str(SRC) not in sys.path:
        sys.path.insert(0, str(SRC))
    if str(REPO) not in sys.path:
        sys.path.insert(0, str(REPO))
    os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"
    day = str(payload["date"])
    capture_dir = Path(payload["capture_path"])
    universe = list(payload["universe"])
    cap_class = str(payload["capture_class"])
    t_wall = time.perf_counter()

    import numpy as np
    from research.e1_x34b_entry_execution.features import preentry_from_board
    from research.e1_x36_joint_allocator.replay import simulate_joint
    from small_paper.v1r_live_dual_lane import get_dual_lane, live_primary_enabled
    from small_paper.v1r_native_entry_live import FEATURE_ORDER, PendingOrder, V1RNativeEntryLive
    from small_paper.v1r_primary_runtime import POSITION_CAP, WAIT_SEC

    class FixedAnchorEngine(V1RNativeEntryLive):
        def __init__(self, *a: Any, **k: Any) -> None:
            super().__init__(*a, **k)
            self.a_candidates: list[dict[str, Any]] = []
            self.a_admits: list[dict[str, Any]] = []

        def _run_anchor(self, *, anchor: str, t0: float, day: str, session: str) -> list[dict[str, Any]]:
            before = len(self.events)
            out = super()._run_anchor(anchor=anchor, t0=t0, day=day, session=session)
            for ev in self.events[before:]:
                kind = str(ev.get("kind") or "")
                if kind == "ANCHOR_SYMBOL_SNAPSHOT" and ev.get("model_score") is not None:
                    self.a_candidates.append(
                        {
                            "symbol": _bare(ev.get("symbol")),
                            "anchor": anchor,
                            "t0": float(t0),
                            "score": ev.get("model_score"),
                            "bid": (ev.get("Buy1") or {}).get("Price") if isinstance(ev.get("Buy1"), dict) else None,
                            "admitted": bool(ev.get("admitted")),
                        }
                    )
                if kind == "V1R_ENTRY_PENDING":
                    self.a_admits.append(
                        {
                            "symbol": _bare(ev.get("symbol")),
                            "anchor": anchor,
                            "t0": float(t0),
                            "score": ev.get("score"),
                            "limit": ev.get("limit"),
                        }
                    )
            self.events.clear()
            return out

    class EventDrivenEngine(V1RNativeEntryLive):
        def __init__(self, *a: Any, **k: Any) -> None:
            super().__init__(*a, **k)
            self._eligible: dict[str, bool] = {}
            self.signals: list[dict[str, Any]] = []
            self.blocked: list[dict[str, Any]] = []
            self._timing_open: list[dict[str, Any]] = []
            self.timing_done: list[dict[str, Any]] = []

        def _mid(self, symbol: str, t: float) -> Optional[float]:
            board = self._board_arrays(symbol)
            if board["t"].size == 0:
                return None
            i = int(np.searchsorted(board["t"], float(t), side="right") - 1)
            if i < 0:
                return None
            ask = float(board["ask"][i])
            bid = float(board["bid"][i])
            if not (math.isfinite(ask) and math.isfinite(bid) and ask > 0 and bid > 0):
                return None
            return (ask + bid) / 2.0

        def _bid(self, symbol: str, t: float) -> Optional[float]:
            board = self._board_arrays(symbol)
            if board["t"].size == 0:
                return None
            i = int(np.searchsorted(board["t"], float(t), side="right") - 1)
            if i < 0:
                return None
            bid = float(board["bid"][i])
            return bid if math.isfinite(bid) and bid > 0 else None

        def on_stream_event(self, symbol: str, t: float) -> None:
            mid = self._mid(symbol, t)
            still: list[dict[str, Any]] = []
            for rec in self._timing_open:
                if rec["symbol"] != symbol:
                    still.append(rec)
                    continue
                if mid is not None and rec.get("signal_mid"):
                    bps = (mid / float(rec["signal_mid"]) - 1.0) * 10000.0
                    rec["mfe_bps"] = max(float(rec.get("mfe_bps") or 0.0), bps)
                    rec["mae_bps"] = min(float(rec.get("mae_bps") or 0.0), bps)
                na = rec.get("next_anchor_t")
                if na is not None and float(t) + 1e-12 >= float(na):
                    rec["anchor_mid"] = self._mid(symbol, float(na))
                    rec["anchor_bid"] = self._bid(symbol, float(na))
                    rec["anchor_still_valid"] = bool(self._eligible_now(symbol, float(na))[0])
                    self.timing_done.append(rec)
                else:
                    still.append(rec)
            self._timing_open = still

        def _eligible_now(self, symbol: str, t: float) -> tuple[bool, dict[str, Any], float, float]:
            board = self._board_arrays(symbol)
            feats = preentry_from_board(board, float(t))
            if any(feats.get(f) is None or not np.isfinite(feats.get(f)) for f in FEATURE_ORDER):
                return False, feats, float("nan"), 0.0
            score = float(self.score_fn(feats))
            if not np.isfinite(score):
                return False, feats, score, 0.0
            bid = self._bid(symbol, t) or 0.0
            if bid <= 0:
                return False, feats, score, bid
            return True, feats, score, bid

        def process_market_push(
            self,
            *,
            symbol: str,
            payload: dict[str, Any],
            event_t: Optional[float] = None,
        ) -> dict[str, Any]:
            t = float(event_t if event_t is not None else 0.0)
            ing = self.ingest_push(symbol=symbol, payload=payload, event_t=t)
            if not ing.get("ingested") and ing.get("reason") == "duplicate_sequence":
                ing["fill_checked"] = False
                ing["anchor_fired"] = False
                return ing
            fired = self.maybe_fire_event(symbol=_bare(symbol), t=t)
            fills = self.on_tick_fill_check(event_t=t, payload=payload, symbol=symbol)
            ing["fill_checked"] = True
            ing["anchor_fired"] = bool(fired)
            ing["fill_n"] = len(fills or [])
            return ing

        def maybe_fire_event(self, *, symbol: str, t: float) -> list[dict[str, Any]]:
            if not in_session(t):
                return []
            ok, feats, score, bid = self._eligible_now(symbol, t)
            was = bool(self._eligible.get(symbol, False))
            self._eligible[symbol] = ok
            if not (ok and not was):
                return []
            dt = datetime.fromtimestamp(float(t), JST)
            sess = "AM" if dt.hour < 12 else "PM"
            na_t, na_lbl = next_anchor_after(self.trading_date or dt.strftime("%Y%m%d"), t)
            sig = {
                "symbol": symbol,
                "event_signal_time": float(t),
                "session": sess,
                "score": score,
                "signal_bid": bid,
                "signal_mid": self._mid(symbol, t),
                "next_fixed_anchor": na_lbl,
                "next_anchor_t": na_t,
                "mfe_bps": 0.0,
                "mae_bps": 0.0,
                "admitted": False,
                "block_reason": "",
            }
            dual_pre = get_dual_lane(trace_dir=self.trace_dir) if live_primary_enabled() else None
            if dual_pre is not None:
                dual_pre.maybe_session_close(event_t=float(t))
            self.on_tick_fill_check(event_t=float(t), symbol=symbol)
            event = {
                "date": str(self.trading_date),
                "symbol": symbol,
                "session": sess,
                "signal_time": float(t),
                "filled": False,
                "limit_price": float(bid),
                "bid0": float(bid),
                **{f: feats.get(f) for f in FEATURE_ORDER},
                "score_preview": score,
            }
            sim = simulate_joint([event], score_fn=self.score_fn)
            e = sim["events"][0]
            blocked = ""
            if not e.get("admitted"):
                blocked = str(e.get("block_reason") or "NOT_ADMITTED")
            elif self.exposure() >= POSITION_CAP:
                blocked = "CAPACITY_BLOCKED_LIVE"
            elif symbol in self.pending or symbol in self.open_symbols:
                blocked = "SAME_SYMBOL_OPEN_OR_PENDING"
            if blocked:
                sig["admitted"] = False
                sig["block_reason"] = blocked
                self.signals.append(sig)
                self.blocked.append(sig)
                if na_t is not None:
                    self._timing_open.append(sig)
                self.events.clear()
                return []
            po = PendingOrder(
                symbol=symbol,
                signal_time=float(t),
                limit_price=float(bid),
                score=float(e.get("alloc_score") if e.get("alloc_score") is not None else score),
                rank=0,
                anchor="EVENT",
                session=sess,
                date=str(self.trading_date),
                features={f: feats.get(f) for f in FEATURE_ORDER},
            )
            self.pending[po.symbol] = po
            self.primary_admitted += 1
            sig["admitted"] = True
            self.signals.append(sig)
            if na_t is not None:
                self._timing_open.append(sig)
            self.events.clear()
            return [{"kind": "V1R_ENTRY_PENDING", "symbol": symbol, "anchor": "EVENT"}]

    def run_pass(engine_cls: type) -> dict[str, Any]:
        eng, dual = _boot(universe, engine_cls)
        if dual is None or not eng.ready:
            return {
                "ok": False,
                "blocker": getattr(eng, "fail_reason", "NATIVE_NOT_READY"),
                "trades": [],
                "events_n": 0,
            }
        eng.trading_date = day
        events_n, _ = _stream_day(day, capture_dir, eng, dual)
        trades = extract_trades(dual)
        extra: dict[str, Any] = {}
        if isinstance(eng, FixedAnchorEngine):
            extra["a_candidates"] = eng.a_candidates
            extra["a_admits"] = eng.a_admits
        if isinstance(eng, EventDrivenEngine):
            leftover = []
            wm = float(eng.event_time_watermark or 0.0)
            for rec in eng._timing_open:
                na = rec.get("next_anchor_t")
                if na is not None and wm + 1e-9 >= float(na):
                    rec["anchor_mid"] = eng._mid(rec["symbol"], float(na))
                    rec["anchor_bid"] = eng._bid(rec["symbol"], float(na))
                    rec["anchor_still_valid"] = bool(eng._eligible_now(rec["symbol"], float(na))[0])
                    leftover.append(rec)
                else:
                    leftover.append(rec)
            extra["signals"] = list(eng.signals)
            extra["blocked"] = list(eng.blocked)
            extra["timing"] = list(eng.timing_done) + leftover
        return {
            "ok": True,
            "blocker": None,
            "trades": trades,
            "events_n": events_n,
            "native_admitted": int(eng.primary_admitted),
            "native_fills": int(eng.primary_fills),
            "native_expired": int(eng.primary_expired),
            **extra,
        }

    a = run_pass(FixedAnchorEngine)
    b = run_pass(EventDrivenEngine)
    out = {
        "date": day,
        "capture_class": cap_class,
        "universe_n": len(universe),
        "universe_source": payload.get("universe_source"),
        "capture_path": str(capture_dir),
        "elapsed_sec": round(time.perf_counter() - t_wall, 1),
        "A": a,
        "B": b,
    }
    try:
        cache_dir = OUT / "day_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        (cache_dir / f"{day}.json").write_text(
            json.dumps(out, ensure_ascii=False, default=str) + "\n",
            encoding="utf-8",
        )
    except Exception:
        pass
    return out


def _sess_stats(trades: list[dict[str, Any]]) -> dict[str, Any]:
    am = [t for t in trades if t.get("session") == "AM"]
    pm = [t for t in trades if t.get("session") == "PM"]
    pnls = [float(t.get("pnl_yen_100") or 0) for t in trades]
    wins = sum(1 for x in pnls if x > 0)
    losses = sum(1 for x in pnls if x < 0)
    draws = sum(1 for x in pnls if abs(x) <= 1e-12)
    return {
        "trades": len(trades),
        "wins": wins,
        "losses": losses,
        "draws": draws,
        "pnl": round(sum(pnls), 2),
        "pf": _pf(pnls),
        "avg_pnl": round(sum(pnls) / len(pnls), 4) if pnls else None,
        "max_drawdown": _maxdd(trades),
        "AM_trades": len(am),
        "AM_pnl": round(sum(float(t.get("pnl_yen_100") or 0) for t in am), 2),
        "PM_trades": len(pm),
        "PM_pnl": round(sum(float(t.get("pnl_yen_100") or 0) for t in pm), 2),
    }


def _bucket(day_rows: list[dict[str, Any]], side: str, *, full_only: bool) -> dict[str, Any]:
    use = [r for r in day_rows if (r.get("A") or {}).get("ok") and (r.get("B") or {}).get("ok")]
    if full_only:
        use = [r for r in use if r.get("capture_class") == "FULL"]
    trades: list[dict[str, Any]] = []
    for r in use:
        trades.extend(list((r.get(side) or {}).get("trades") or []))
    st = _sess_stats(trades)
    st["days"] = [r["date"] for r in use]
    st["day_count"] = len(use)
    return st


def _median(xs: list[float]) -> Optional[float]:
    if not xs:
        return None
    s = sorted(xs)
    n = len(s)
    if n % 2:
        return float(s[n // 2])
    return float(0.5 * (s[n // 2 - 1] + s[n // 2]))


def _p90(xs: list[float]) -> Optional[float]:
    if not xs:
        return None
    s = sorted(xs)
    idx = min(len(s) - 1, int(math.ceil(0.9 * len(s)) - 1))
    return float(s[idx])


def build_attribution(day_rows: list[dict[str, Any]], *, full_only: bool) -> dict[str, Any]:
    use = [r for r in day_rows if (r.get("B") or {}).get("ok")]
    if full_only:
        use = [r for r in use if r.get("capture_class") == "FULL"]
    signals: list[dict[str, Any]] = []
    a_cand_idx: dict[tuple[str, str, str], dict[str, Any]] = {}
    a_admit_idx: dict[tuple[str, str, str], dict[str, Any]] = {}
    a_trades: list[dict[str, Any]] = []
    b_trades: list[dict[str, Any]] = []
    for r in use:
        day = r["date"]
        for c in (r.get("A") or {}).get("a_candidates") or []:
            a_cand_idx[(day, str(c.get("anchor") or ""), str(c.get("symbol") or ""))] = c
        for c in (r.get("A") or {}).get("a_admits") or []:
            a_admit_idx[(day, str(c.get("anchor") or ""), str(c.get("symbol") or ""))] = c
        a_trades.extend({**t, "date": day} for t in (r.get("A") or {}).get("trades") or [])
        b_trades.extend({**t, "date": day} for t in (r.get("B") or {}).get("trades") or [])
        for s in (r.get("B") or {}).get("timing") or (r.get("B") or {}).get("signals") or []:
            signals.append({**s, "date": day})

    waits: list[float] = []
    deltas: list[float] = []
    still = 0
    gone = 0
    better = 0
    worse = 0
    mfe_pos = 0
    b_only = 0
    a_and_b = 0
    blocked_n = 0
    matched: list[dict[str, Any]] = []
    b_only_rows: list[dict[str, Any]] = []
    for s in signals:
        day = str(s.get("date") or "")
        na = str(s.get("next_fixed_anchor") or "")
        nt = s.get("next_anchor_t")
        st = s.get("event_signal_time")
        if nt is not None and st is not None:
            wait = float(nt) - float(st)
            waits.append(wait)
            s["time_to_next_anchor_sec"] = round(wait, 3)
        sp = s.get("signal_bid") or s.get("signal_mid")
        ap = s.get("anchor_bid") or s.get("anchor_mid")
        if sp and ap and float(ap) > 0:
            dps = (float(sp) / float(ap) - 1.0) * 10000.0
            s["entry_price_delta_bps"] = round(dps, 4)
            deltas.append(dps)
            if float(sp) < float(ap):
                better += 1
            elif float(sp) > float(ap):
                worse += 1
        if s.get("anchor_still_valid") is True:
            still += 1
        elif s.get("anchor_still_valid") is False:
            gone += 1
        if float(s.get("mfe_bps") or 0) > 1e-9:
            mfe_pos += 1
        key = (day, na, str(s.get("symbol") or ""))
        a_here = a_cand_idx.get(key) or a_admit_idx.get(key)
        if s.get("block_reason"):
            blocked_n += 1
            s["opportunity_class"] = "PORTFOLIO_BLOCKED"
        elif gone and not a_here:
            b_only += 1
            s["opportunity_class"] = "B_ONLY_SIGNAL"
            b_only_rows.append(s)
        elif a_here or s.get("anchor_still_valid"):
            a_and_b += 1
            s["opportunity_class"] = "A_AND_B"
        else:
            s["opportunity_class"] = "B_ONLY_SIGNAL"
            b_only += 1
            b_only_rows.append(s)
        if a_here and s.get("admitted"):
            matched.append(
                {
                    "date": day,
                    "symbol": s.get("symbol"),
                    "b_signal_time": s.get("event_signal_time"),
                    "a_anchor": na,
                    "b_bid": s.get("signal_bid"),
                    "a_bid": (a_here or {}).get("bid") or s.get("anchor_bid"),
                    "wait_sec": s.get("time_to_next_anchor_sec"),
                }
            )

    a_only_rows: list[dict[str, Any]] = []
    b_syms_time = [(str(s.get("date")), str(s.get("symbol")), float(s.get("event_signal_time") or 0)) for s in signals]
    for adm in a_trades:
        day = str(adm.get("date") or "")
        sym = str(adm.get("symbol") or "")
        et = float(adm.get("entry_time") or 0)
        found = False
        for d, sy, st in b_syms_time:
            if d == day and sy == sym and st <= et + 1.0:
                found = True
                break
        if not found:
            a_only_rows.append(adm)

    return {
        "event_driven_signals": len(signals),
        "signals_before_next_anchor": sum(1 for s in signals if s.get("next_anchor_t")),
        "anchor_still_valid_count": still,
        "anchor_no_longer_valid_count": gone,
        "median_wait_to_anchor_sec": _median(waits),
        "p90_wait_to_anchor_sec": _p90(waits),
        "price_better_if_immediate_count": better,
        "price_worse_if_immediate_count": worse,
        "median_entry_price_delta_bps": _median(deltas),
        "MFE_before_anchor_count": mfe_pos,
        "move_already_started_before_anchor_count": mfe_pos,
        "B_only_signal_count": b_only,
        "A_and_B_count": a_and_b,
        "A_only_count": len(a_only_rows),
        "PORTFOLIO_BLOCKED_count": blocked_n,
        "signals": signals,
        "matched": matched,
        "b_only": b_only_rows,
        "a_only": a_only_rows,
        "blocked": [s for s in signals if s.get("opportunity_class") == "PORTFOLIO_BLOCKED"],
    }


def write_excel(
    *,
    inventory: list[dict[str, Any]],
    day_rows: list[dict[str, Any]],
    primary: dict[str, Any],
    reference: dict[str, Any],
    timing_p: dict[str, Any],
    timing_r: dict[str, Any],
    verdict: str,
) -> Path:
    from openpyxl import Workbook

    wb = Workbook()

    def put(ws: Any, rows: list[list[Any]]) -> None:
        for r in rows:
            ws.append(r)

    ws = wb.active
    ws.title = "Summary"
    pa, pb = primary["A"], primary["B"]
    put(
        ws,
        [
            ["field", "value"],
            ["verdict", verdict],
            ["capture_period", f"{inventory[0]['date'] if inventory else ''}–{inventory[-1]['date'] if inventory else ''}"],
            ["full_days", ",".join(pa.get("days") or [])],
            ["PRIMARY_A_trades", pa.get("trades")],
            ["PRIMARY_A_pnl", pa.get("pnl")],
            ["PRIMARY_A_pf", pa.get("pf")],
            ["PRIMARY_A_maxdd", pa.get("max_drawdown")],
            ["PRIMARY_B_trades", pb.get("trades")],
            ["PRIMARY_B_pnl", pb.get("pnl")],
            ["PRIMARY_B_pf", pb.get("pf")],
            ["PRIMARY_B_maxdd", pb.get("max_drawdown")],
            ["DELTA_trades", (pb.get("trades") or 0) - (pa.get("trades") or 0)],
            ["DELTA_pnl", round(float(pb.get("pnl") or 0) - float(pa.get("pnl") or 0), 2)],
            ["event_driven_signals", timing_p.get("event_driven_signals")],
            ["median_wait_to_anchor_sec", timing_p.get("median_wait_to_anchor_sec")],
            ["B_only_signal_count", timing_p.get("B_only_signal_count")],
        ],
    )

    ws = wb.create_sheet("Capture_Inventory")
    cols = ["date", "capture_class", "usable", "full", "universe_n", "am_coverage", "pm_coverage", "first_event_time", "last_event_time", "size_bytes", "exclusion_reason", "capture_path"]
    ws.append(cols)
    for row in inventory:
        ws.append([row.get(c) for c in cols])

    ws = wb.create_sheet("Daily")
    ws.append(["date", "class", "A_trades", "A_pnl", "A_pf", "B_trades", "B_pnl", "B_pf", "d_trades", "d_pnl"])
    for r in day_rows:
        sa = _sess_stats((r.get("A") or {}).get("trades") or [])
        sb = _sess_stats((r.get("B") or {}).get("trades") or [])
        ws.append(
            [
                r.get("date"),
                r.get("capture_class"),
                sa["trades"],
                sa["pnl"],
                sa["pf"],
                sb["trades"],
                sb["pnl"],
                sb["pf"],
                sb["trades"] - sa["trades"],
                round(sb["pnl"] - sa["pnl"], 2),
            ]
        )

    ws = wb.create_sheet("Sessions")
    ws.append(["date", "class", "A_AM_tr", "A_AM_pnl", "A_PM_tr", "A_PM_pnl", "B_AM_tr", "B_AM_pnl", "B_PM_tr", "B_PM_pnl"])
    for r in day_rows:
        sa = _sess_stats((r.get("A") or {}).get("trades") or [])
        sb = _sess_stats((r.get("B") or {}).get("trades") or [])
        ws.append([r.get("date"), r.get("capture_class"), sa["AM_trades"], sa["AM_pnl"], sa["PM_trades"], sa["PM_pnl"], sb["AM_trades"], sb["AM_pnl"], sb["PM_trades"], sb["PM_pnl"]])

    def trade_sheet(name: str, side: str) -> None:
        w = wb.create_sheet(name)
        w.append(["date", "class", "symbol", "session", "entry_time", "exit_time", "entry_price", "exit_price", "pnl_yen_100", "reason"])
        for r in day_rows:
            for t in (r.get(side) or {}).get("trades") or []:
                w.append(
                    [
                        r.get("date"),
                        r.get("capture_class"),
                        t.get("symbol"),
                        t.get("session"),
                        t.get("entry_time"),
                        t.get("exit_time"),
                        t.get("entry_price"),
                        t.get("exit_price"),
                        t.get("pnl_yen_100"),
                        t.get("reason"),
                    ]
                )

    trade_sheet("Portfolio_A", "A")
    trade_sheet("Portfolio_B", "B")

    ws = wb.create_sheet("Timing_Attribution")
    ws.append(
        [
            "date",
            "symbol",
            "event_signal_time",
            "next_fixed_anchor",
            "time_to_next_anchor_sec",
            "signal_bid",
            "anchor_bid",
            "entry_price_delta_bps",
            "mfe_bps",
            "mae_bps",
            "anchor_still_valid",
            "opportunity_class",
            "admitted",
            "block_reason",
        ]
    )
    for s in timing_p.get("signals") or []:
        ws.append(
            [
                s.get("date"),
                s.get("symbol"),
                s.get("event_signal_time"),
                s.get("next_fixed_anchor"),
                s.get("time_to_next_anchor_sec"),
                s.get("signal_bid"),
                s.get("anchor_bid"),
                s.get("entry_price_delta_bps"),
                s.get("mfe_bps"),
                s.get("mae_bps"),
                s.get("anchor_still_valid"),
                s.get("opportunity_class"),
                s.get("admitted"),
                s.get("block_reason"),
            ]
        )

    ws = wb.create_sheet("Matched_Trades")
    ws.append(["date", "symbol", "b_signal_time", "a_anchor", "b_bid", "a_bid", "wait_sec"])
    for m in timing_p.get("matched") or []:
        ws.append([m.get("date"), m.get("symbol"), m.get("b_signal_time"), m.get("a_anchor"), m.get("b_bid"), m.get("a_bid"), m.get("wait_sec")])

    ws = wb.create_sheet("B_Only")
    ws.append(["date", "symbol", "event_signal_time", "next_fixed_anchor", "anchor_still_valid", "block_reason"])
    for s in timing_p.get("b_only") or []:
        ws.append([s.get("date"), s.get("symbol"), s.get("event_signal_time"), s.get("next_fixed_anchor"), s.get("anchor_still_valid"), s.get("block_reason")])

    ws = wb.create_sheet("A_Only")
    ws.append(["date", "symbol", "session", "entry_time", "entry_price", "pnl_yen_100"])
    for t in timing_p.get("a_only") or []:
        ws.append([t.get("date"), t.get("symbol"), t.get("session"), t.get("entry_time"), t.get("entry_price"), t.get("pnl_yen_100")])

    ws = wb.create_sheet("Blocked")
    ws.append(["date", "symbol", "event_signal_time", "block_reason", "next_fixed_anchor"])
    for s in timing_p.get("blocked") or []:
        ws.append([s.get("date"), s.get("symbol"), s.get("event_signal_time"), s.get("block_reason"), s.get("next_fixed_anchor")])

    path = OUT / "anchor_vs_event_driven.xlsx"
    wb.save(path)
    return path


def decide_verdict(primary: dict[str, Any], day_rows: list[dict[str, Any]]) -> tuple[str, str]:
    a, b = primary["A"], primary["B"]
    full = [r for r in day_rows if r.get("capture_class") == "FULL" and (r.get("A") or {}).get("ok")]
    d_pnl = float(b.get("pnl") or 0) - float(a.get("pnl") or 0)
    pf_a, pf_b = a.get("pf"), b.get("pf")
    pf_better = False
    if pf_a is None and pf_b is None:
        pf_better = False
    elif pf_a is None:
        pf_better = True
    elif pf_b is None:
        pf_better = False
    else:
        pf_better = float(pf_b) > float(pf_a) + 1e-9
    dd_a = float(a.get("max_drawdown") or 0)
    dd_b = float(b.get("max_drawdown") or 0)
    dd_ok = dd_b >= dd_a - 0.25 * abs(dd_a) - 1.0
    day_deltas = []
    for r in full:
        sa = _sess_stats((r.get("A") or {}).get("trades") or [])
        sb = _sess_stats((r.get("B") or {}).get("trades") or [])
        day_deltas.append((r["date"], sb["pnl"] - sa["pnl"]))
    improve_days = sum(1 for _, d in day_deltas if d > 0)
    abs_sum = sum(abs(d) for _, d in day_deltas) or 1.0
    top = max(day_deltas, key=lambda x: abs(x[1])) if day_deltas else ("", 0.0)
    single = abs(top[1]) / abs_sum >= 0.70 if day_deltas else True
    note = f"improve_days={improve_days}/{len(full)} top_day={top[0]} share={abs(top[1])/abs_sum:.2f}"
    if len(full) < 2:
        return "ANCHOR_EFFECT_INCONCLUSIVE", "fewer_than_2_full_days " + note
    if d_pnl > 0 and pf_better and dd_ok and improve_days >= 2 and not single:
        return "ANCHOR_CAUSES_OPPORTUNITY_LOSS_CANDIDATE", note
    if d_pnl < 0 and (not pf_better) and improve_days <= 1:
        return "ANCHOR_HELPFUL", note
    return "ANCHOR_EFFECT_INCONCLUSIVE", note


def write_md(report: dict[str, Any]) -> str:
    p = report["PRIMARY_FULL"]
    r = report["REFERENCE_ALL_USABLE"]
    t = report["TIMING"]
    lines = [
        "# Fixed Anchor vs Event-Driven V1R ENTRY",
        "",
        "Timing-only comparison. ENTRY features/thresholds/FILL/EXIT unchanged. Not a runtime adoption.",
        "",
        f"**Verdict:** `{report['verdict']}`",
        "",
        f"- capture_period: `{report['capture_period']}`",
        f"- usable_days: {report['usable_days']}",
        f"- full_days: {report['full_days']}",
        f"- feature_equivalence: {report['feature_equivalence']}",
        f"- anchor_dependent_features: {report['anchor_dependent_features']}",
        f"- future_leak: {report['future_leak']}",
        "",
        "## PRIMARY FULL",
        "",
        f"| | trades | pnl | PF | maxDD |",
        f"|---|---:|---:|---:|---:|",
        f"| FIXED_ANCHOR | {p['A']['trades']} | {p['A']['pnl']} | {p['A']['pf']} | {p['A']['max_drawdown']} |",
        f"| EVENT_DRIVEN | {p['B']['trades']} | {p['B']['pnl']} | {p['B']['pf']} | {p['B']['max_drawdown']} |",
        f"| DELTA | {p['DELTA']['trades']} | {p['DELTA']['pnl']} | {p['DELTA']['pf']} | {p['DELTA']['maxdd']} |",
        "",
        "## TIMING (FULL)",
        "",
        f"- event_driven_signals: {t.get('event_driven_signals')}",
        f"- median_wait_to_anchor_sec: {t.get('median_wait_to_anchor_sec')}",
        f"- p90_wait_to_anchor_sec: {t.get('p90_wait_to_anchor_sec')}",
        f"- median_entry_price_delta_bps: {t.get('median_entry_price_delta_bps')}",
        f"- anchor_no_longer_valid_count: {t.get('anchor_no_longer_valid_count')}",
        f"- B_only_signal_count: {t.get('B_only_signal_count')}",
        "",
        "## REFERENCE ALL_USABLE",
        "",
        f"- A pnl/trades/PF: {r['A']['pnl']} / {r['A']['trades']} / {r['A']['pf']}",
        f"- B pnl/trades/PF: {r['B']['pnl']} / {r['B']['trades']} / {r['B']['pf']}",
        f"- delta pnl: {r['DELTA']['pnl']}",
        "",
        f"day_dependency: {report.get('day_dependency')}",
        "",
        "STOP. No strategy change. No runtime adoption.",
        "",
    ]
    return "\n".join(lines)


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    mc = NATIVE / "data" / "market_capture"
    days = sorted(p.name for p in mc.iterdir() if p.is_dir() and p.name.isdigit() and len(p.name) == 8)
    inventory: list[dict[str, Any]] = []
    jobs: list[dict[str, Any]] = []
    for day in days:
        cap = find_capture_dir(day)
        uni, uni_src = historical_universe(day, cap or (mc / day))
        row = classify_capture(day, cap, len(uni))
        row["universe_source"] = uni_src
        inventory.append(row)
        if row.get("usable") and cap is not None and uni:
            jobs.append(
                {
                    "date": day,
                    "capture_path": str(cap),
                    "universe": uni,
                    "universe_source": uni_src,
                    "capture_class": row["capture_class"],
                }
            )
    (OUT / "capture_inventory.json").write_text(
        json.dumps(inventory, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"inventory days={len(days)} usable_jobs={len(jobs)}", flush=True)
    only = str(os.environ.get("ANCHOR_ED_DAYS") or "").strip()
    if only:
        want = {x.strip() for x in only.split(",") if x.strip()}
        jobs = [j for j in jobs if j["date"] in want]
        print(f"ANCHOR_ED_DAYS filter -> {len(jobs)} jobs", flush=True)
    for row in inventory:
        print(f"  {row['date']}  {row['capture_class']}  uni={row.get('universe_n')}  {row.get('exclusion_reason')}", flush=True)

    day_rows: list[dict[str, Any]] = []
    cache_dir = OUT / "day_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    force = str(os.environ.get("ANCHOR_ED_FORCE") or "").strip() in {"1", "true", "TRUE", "yes"}
    pending_jobs: list[dict[str, Any]] = []
    for job in jobs:
        cp = cache_dir / f"{job['date']}.json"
        if cp.is_file() and not force:
            try:
                day_rows.append(json.loads(cp.read_text(encoding="utf-8")))
                print(f"cache {job['date']}", flush=True)
                continue
            except Exception:
                pass
        pending_jobs.append(job)
    jobs = pending_jobs
    if jobs:
        with ProcessPoolExecutor(max_workers=min(MAX_WORKERS, len(jobs))) as ex:
            from research.anchor_vs_event_driven.run_comparison import replay_day as _worker

            futs = {ex.submit(_worker, job): job["date"] for job in jobs}
            for fut in as_completed(futs):
                day = futs[fut]
                try:
                    got = fut.result()
                except Exception as exc:
                    got = {"date": day, "ok": False, "error": f"{type(exc).__name__}:{exc}", "A": {"ok": False, "trades": []}, "B": {"ok": False, "trades": []}}
                day_rows.append(got)
                arow = got.get("A") or {}
                brow = got.get("B") or {}
                print(
                    f"done {got.get('date')} class={got.get('capture_class')} "
                    f"A_ok={arow.get('ok')} B_ok={brow.get('ok')} "
                    f"A_tr={len(arow.get('trades') or [])} B_tr={len(brow.get('trades') or [])} "
                    f"B_sig={len(brow.get('signals') or [])} "
                    f"B_fill={brow.get('native_fills')} B_exp={brow.get('native_expired')} "
                    f"sec={got.get('elapsed_sec')}",
                    flush=True,
                )
    day_rows.sort(key=lambda r: str(r.get("date")))

    primary_a = _bucket(day_rows, "A", full_only=True)
    primary_b = _bucket(day_rows, "B", full_only=True)
    ref_a = _bucket(day_rows, "A", full_only=False)
    ref_b = _bucket(day_rows, "B", full_only=False)

    def delta(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
        pf_d = None
        if a.get("pf") is not None and b.get("pf") is not None:
            pf_d = float(b["pf"]) - float(a["pf"])
        return {
            "trades": int(b.get("trades") or 0) - int(a.get("trades") or 0),
            "pnl": round(float(b.get("pnl") or 0) - float(a.get("pnl") or 0), 2),
            "pf": pf_d,
            "maxdd": round(float(b.get("max_drawdown") or 0) - float(a.get("max_drawdown") or 0), 2),
        }

    primary = {"A": primary_a, "B": primary_b, "DELTA": delta(primary_a, primary_b)}
    reference = {"A": ref_a, "B": ref_b, "DELTA": delta(ref_a, ref_b)}
    timing_p = build_attribution(day_rows, full_only=True)
    timing_r = build_attribution(day_rows, full_only=False)
    verdict, dep = decide_verdict(primary, day_rows)
    ref_sign = (reference["DELTA"]["pnl"] or 0) * (primary["DELTA"]["pnl"] or 0)
    if primary["DELTA"]["pnl"] and reference["DELTA"]["pnl"] and ref_sign < 0:
        dep += " REFERENCE_SIGN_FLIP"

    report = {
        "result_class": "ANCHOR_VS_EVENT_DRIVEN_ENTRY_TIMING_ONLY",
        "formal_certification": False,
        "runtime_adoption": False,
        "identity": {
            "strategy_sha": "9ad4ba2730892d40c757d940b82480e620e502e3e789839120e90b18be082547",
            "entry_sha": "f2887bb2be539cc173aee438a43ee8afb8cfa2b8c31380937ecd843e90dd9b29",
            "exit_sha": "6cc3b8aade76e323682ec39dfd06878aab0ff1a99dd42922744b0054a7ea3255",
        },
        "feature_equivalence": "same preentry_from_board + frozen A1_FILL score_fn; LIVE_FEATURES only",
        "anchor_dependent_features": [],
        "future_leak": False,
        "live_features": list(LIVE_FEATURES),
        "note_univ_med": "univ_med_mid_ret_60s exists in research FEATURE_SPECS but is NOT in live FEATURE_ORDER",
        "capture_period": f"{days[0]}-{days[-1]}" if days else "",
        "usable_days": [r["date"] for r in inventory if r.get("usable")],
        "full_days": [r["date"] for r in inventory if r.get("full")],
        "inventory": inventory,
        "PRIMARY_FULL": primary,
        "REFERENCE_ALL_USABLE": reference,
        "TIMING": {k: v for k, v in timing_p.items() if k not in {"signals", "matched", "b_only", "a_only", "blocked"}},
        "TIMING_REFERENCE": {k: v for k, v in timing_r.items() if k not in {"signals", "matched", "b_only", "a_only", "blocked"}},
        "day_dependency": dep,
        "verdict": verdict,
        "daily": [
            {
                "date": r.get("date"),
                "capture_class": r.get("capture_class"),
                "elapsed_sec": r.get("elapsed_sec"),
                "A": _sess_stats((r.get("A") or {}).get("trades") or []),
                "B": _sess_stats((r.get("B") or {}).get("trades") or []),
                "A_ok": (r.get("A") or {}).get("ok"),
                "B_ok": (r.get("B") or {}).get("ok"),
                "blocker": (r.get("A") or {}).get("blocker") or (r.get("B") or {}).get("blocker"),
            }
            for r in day_rows
        ],
    }
    # persist compact day trades for excel without huge json
    compact_days = []
    for r in day_rows:
        compact_days.append(
            {
                "date": r.get("date"),
                "capture_class": r.get("capture_class"),
                "universe_n": r.get("universe_n"),
                "A": {
                    "ok": (r.get("A") or {}).get("ok"),
                    "trades": (r.get("A") or {}).get("trades") or [],
                    "a_candidates": (r.get("A") or {}).get("a_candidates") or [],
                    "a_admits": (r.get("A") or {}).get("a_admits") or [],
                },
                "B": {
                    "ok": (r.get("B") or {}).get("ok"),
                    "trades": (r.get("B") or {}).get("trades") or [],
                    "signals": (r.get("B") or {}).get("signals") or [],
                    "blocked": (r.get("B") or {}).get("blocked") or [],
                    "timing": (r.get("B") or {}).get("timing") or [],
                },
            }
        )
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    (OUT / "report.md").write_text(write_md(report), encoding="utf-8")
    xlsx = OUT / "anchor_vs_event_driven.xlsx"
    try:
        xlsx = write_excel(
            inventory=inventory,
            day_rows=compact_days,
            primary=primary,
            reference=reference,
            timing_p=build_attribution(compact_days, full_only=True),
            timing_r=timing_r,
            verdict=verdict,
        )
    except Exception as exc:
        print(f"excel_write_failed {type(exc).__name__}:{exc}", flush=True)
        xlsx = None
    print("--------------------------------", flush=True)
    print("PRIMARY FULL", flush=True)
    print(f"A trades={primary_a['trades']} pnl={primary_a['pnl']} pf={primary_a['pf']} maxdd={primary_a['max_drawdown']}", flush=True)
    print(f"B trades={primary_b['trades']} pnl={primary_b['pnl']} pf={primary_b['pf']} maxdd={primary_b['max_drawdown']}", flush=True)
    print(f"verdict={verdict}", flush=True)
    print(f"wrote {OUT / 'report.json'}", flush=True)
    if xlsx is not None:
        print(f"wrote {xlsx}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
