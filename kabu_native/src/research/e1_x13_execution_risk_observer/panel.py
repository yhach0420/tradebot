"""Load E1_X10 day panel for D−1 historical replay (no forbidden dates)."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from . import DESIGN_DAYS, FORBIDDEN_ALPHA_DAYS

NATIVE = Path(__file__).resolve().parents[3]
X10_XLSX = NATIVE / "results" / "research" / "e1_x10_risk_universe" / "audit.xlsx"


def _f(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def load_x10_panel() -> dict[tuple[str, str], dict[str, Any]]:
    """(symbol, day) → risk day row. Never includes 20260803/04/05."""
    notion = _load_sheet(X10_XLSX, "OneLotNotional")
    tick = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "TickRisk")}
    spread = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "SpreadRisk")}
    jump = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "BidJumps")}
    exec_ = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "ExecutableLoss")}
    depth = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "DepthRisk")}
    fresh = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "Freshness")}
    ref = {(str(r["symbol"]), str(r["day"])): r for r in _load_sheet(X10_XLSX, "ReferencePrices")}

    panel: dict[tuple[str, str], dict[str, Any]] = {}
    for r in notion:
        day, sym = str(r["day"]), str(r["symbol"])
        if day in FORBIDDEN_ALPHA_DAYS:
            raise RuntimeError(f"forbidden alpha day in X10 panel: {day}")
        if day not in DESIGN_DAYS:
            continue
        key = (sym, day)
        sp, jp, ex, dp, fr, tk, rf = (
            spread.get(key) or {}, jump.get(key) or {}, exec_.get(key) or {},
            depth.get(key) or {}, fresh.get(key) or {}, tick.get(key) or {}, ref.get(key) or {},
        )
        panel[key] = {
            "symbol": sym,
            "day": day,
            "reference_price": _f(rf.get("reference_price") or r.get("reference_price")),
            "reference_price_source": rf.get("reference_price_source") or "previous_session_official_close",
            "ref_status": "OK" if _f(r.get("one_lot_notional_yen")) is not None else "MISSING",
            "one_lot_notional_yen": _f(r.get("one_lot_notional_yen")),
            "tick_size_yen": _f(tk.get("tick_size_yen")),
            "one_tick_risk_yen_100": _f(tk.get("one_tick_risk_yen_100")),
            "n_spread_obs": int(_f(sp.get("n_spread_obs")) or 0),
            "median_spread_cost_yen_100": _f(sp.get("median_spread_cost_yen_100")),
            "p95_spread_cost_yen_100": _f(sp.get("p95_spread_cost_yen_100")),
            "n_jump_obs": int(_f(jp.get("n_jump_obs")) or 0),
            "p95_down_bid_jump_yen_100": _f(jp.get("p95_down_bid_jump_yen_100")),
            "n_exec_anchors": int(_f(ex.get("n_exec_anchors")) or 0),
            "exec_loss_5s_p95": _f(ex.get("exec_loss_yen_100_5s_p95")),
            "p10_best_bid_qty": _f(dp.get("p10_best_bid_qty")),
            "p10_best_ask_qty": _f(dp.get("p10_best_ask_qty")),
            "board_fresh_rate": _f(fr.get("board_fresh_rate")),
        }
    return panel


def load_x10_symbol_summary() -> dict[str, dict[str, Any]]:
    rows = _load_sheet(X10_XLSX, "SymbolRiskSummary")
    return {str(r["symbol"]): r for r in rows}
