"""Write P3-4 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.canonical_fixed_pnl_source_p3_3.ledger import _share, group_pnl, pnl
from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.fixed_winner_cluster_extension_p3_4 import (
    ANALYSIS_ID,
    DOCUMENT_ID,
    FULL14,
    P1_PF,
    P1_PNL,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    TASK_LABEL,
    TOP3_SYMBOLS,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.fixed_winner_cluster_extension_p3_4.cluster import (
    attach_flags,
    classify_cluster,
    decompose_0905,
    decompose_extend,
    decompose_symbol,
    intersections,
    overlap_16,
    residuals,
    top10_membership,
)
from research.fixed_winner_cluster_extension_p3_4.metrics import (
    compare_extend_exit,
    gate_verdict,
    incremental_block,
    interpret_incremental,
    same_anchor,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "fixed_winner_cluster_extension_p3_4"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, np.generic):
        obj = obj.item()
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def merge_path(flagged: list[dict[str, Any]], path_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id = {str(r.get("trade_id")): r for r in path_rows if r.get("trade_id") is not None}
    keep = {
        "pnl_yen_100",
        "fill_price",
        "exit_price",
        "exit_reason",
        "holding_sec",
        "fill_time",
        "exit_time",
        "date",
        "symbol",
        "anchor_time",
        "session",
        "trade_id",
    }
    out = []
    for t in flagged:
        rec = dict(t)
        p = by_id.get(str(t.get("trade_id")))
        rec["path_joined"] = p is not None
        if p is not None:
            for k, v in p.items():
                if k in keep:
                    continue
                rec[k] = v
        out.append(rec)
    return out


def _days(rows, days):
    return [r for r in rows if str(r.get("date")) in set(days)]


def build_report(
    *,
    primary: list[dict[str, Any]],
    path_rows: list[dict[str, Any]],
    recon: dict[str, Any],
    leak_decision: int,
    leak_checkpoint: int,
    identity_n: int,
    identity_fail: int,
    failed: list[str],
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    flagged = attach_flags(primary)
    merged = merge_path(flagged, path_rows)
    total = sum(pnl(t) for t in merged)
    cells = overlap_16(merged)
    inter = intersections(merged)
    resid = residuals(merged)
    t10 = top10_membership(merged)
    a0905 = decompose_0905(merged)
    syms = {s: decompose_symbol(merged, s) for s in TOP3_SYMBOLS}
    ext = decompose_extend(merged)
    cluster = classify_cluster(inter, resid, total)

    reached = [r for r in merged if r.get("reached_600")]
    if path_rows:
        n_reached = len(reached)
        n_match = sum(1 for r in reached if r.get("matched") is True)
        n_mis = sum(1 for r in reached if r.get("matched") is False)
        n_path = sum(1 for r in merged if r.get("path_joined"))
        dec_pass = n_reached > 0 and n_mis == 0 and n_path == len(merged)
    else:
        n_reached = n_match = n_mis = n_path = 0
        dec_pass = False

    all_cmp = compare_extend_exit(merged) if path_rows else {}
    top_cmp = compare_extend_exit(_days(merged, PREDECLARED_TOP3)) if path_rows else {}
    rest_cmp = compare_extend_exit(_days(merged, REST11)) if path_rows else {}
    sa_all = same_anchor(merged) if path_rows else {}
    sa_top = same_anchor(_days(merged, PREDECLARED_TOP3)) if path_rows else {}
    sa_rest = same_anchor(_days(merged, REST11)) if path_rows else {}
    gate = gate_verdict(all_cmp, rest_cmp, sa_rest, sa_all) if path_rows else {"EXTENSION_GATE": None}
    incr = incremental_block(merged) if path_rows else {}
    interp = interpret_incremental(incr, gate.get("EXTENSION_GATE")) if path_rows else {}

    leak_ok = int(leak_decision) == 0 and int(leak_checkpoint) == 0
    ident_ok = int(identity_fail) == 0
    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if path_rows and not dec_pass:
        integrity.append(f"EXTENSION_DECISION_MISMATCH_{n_match}/{n_reached}")
    if path_rows and n_path != len(merged):
        integrity.append(f"PATH_JOIN_{n_path}/{len(merged)}")
    if path_rows and not leak_ok:
        integrity.append("FUTURE_LEAK")
    if path_rows and not ident_ok:
        integrity.append(f"INCREMENTAL_IDENTITY_FAIL_{identity_fail}/{identity_n}")

    if blocked or not recon.get("pass") or (path_rows and not dec_pass):
        verdict = VERDICT_BLOCKED
    elif integrity:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    full = next((r for r in inter if str(r.get("name", "")).startswith("FULL")), {})
    none = next((r for r in resid if r.get("name") == "NONE_OF_4_FLAGS"), {})

    def resid_named(name: str) -> dict:
        return next((r for r in resid if r.get("name") == name), {})

    flag_rows = [
        {
            "trade_id": t.get("trade_id"),
            "date": t.get("date"),
            "symbol": t.get("symbol"),
            "anchor_time": t.get("anchor_time"),
            "session": t.get("session"),
            "exit_reason": t.get("exit_reason"),
            "pnl_yen_100": t.get("pnl_yen_100"),
            "IS_0905": t.get("IS_0905"),
            "IS_TOP3_DAY": t.get("IS_TOP3_DAY"),
            "IS_TOP3_SYMBOL": t.get("IS_TOP3_SYMBOL"),
            "IS_CONT_EXTEND_750": t.get("IS_CONT_EXTEND_750"),
            "IS_TOP10_WINNER": t.get("IS_TOP10_WINNER"),
            "n_mechanism_flags": t.get("n_mechanism_flags"),
        }
        for t in merged
    ]

    def flat_splits(prefix: str, blk: dict[str, Any]) -> list[dict[str, Any]]:
        rows = []
        for key, val in blk.items():
            if key in ("universe", "symbol", "pnl_tail", "rows"):
                continue
            if isinstance(val, list):
                for item in val:
                    rows.append({"group": prefix, "split_family": key, **item})
        return rows

    sheets = {
        "flags": flag_rows,
        "cells": cells,
        "inter": inter,
        "resid": resid,
        "top10": t10.get("rows") or [],
        "a0905": [{"slice": "UNIVERSE", **a0905["universe"]}] + flat_splits("09:05", a0905),
        "symbols": (
            [{"symbol": s, "slice": "UNIVERSE", **syms[s]["universe"]} for s in TOP3_SYMBOLS]
            + [r for s in TOP3_SYMBOLS for r in flat_splits(s, syms[s])]
        ),
        "extend": [{"slice": "UNIVERSE", **ext["universe"]}]
        + flat_splits("EXTEND750", ext)
        + [{"row_type": "TRADE", **r} for r in ext.get("rows") or []],
        "recon": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "anchor_time": r.get("anchor_time"),
                "exit_reason": r.get("exit_reason"),
                "reached_600": r.get("reached_600"),
                "canonical_class": r.get("canonical_class"),
                "recon_class": r.get("recon_class"),
                "matched": r.get("matched"),
                "guard_hit": r.get("independent_guard_hit"),
                "feat_ret": r.get("feat_ret"),
                "feat_mfe": r.get("feat_mfe"),
                "feat_imb": r.get("feat_imb"),
                "independent_recon_class": r.get("independent_recon_class"),
                "runtime_reason": r.get("runtime_reason"),
                "runtime_off_now": r.get("runtime_off_now"),
                "dual_lane_600_fire_after_t600": r.get("dual_lane_600_fire_after_t600"),
                "decision_future_leak": r.get("decision_future_leak"),
                "continuation_id": r.get("continuation_id"),
            }
            for r in merged
            if r.get("reached_600")
        ],
        "path": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "canonical_class": r.get("canonical_class"),
                "status_600": r.get("status_600"),
                "status_750": r.get("status_750"),
                "bid600": r.get("bid600"),
                "bid750": r.get("bid750"),
                "mid600": r.get("mid600"),
                "mid750": r.get("mid750"),
                "bid_ret_600_750": r.get("bid_ret_600_750"),
                "mid_ret_600_750": r.get("mid_ret_600_750"),
                "outcome_evaluable": r.get("outcome_evaluable"),
                "actual_exit_price": r.get("exit_price"),
            }
            for r in merged
            if r.get("reached_600")
        ],
        "incr": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "IS_TOP3_DAY": r.get("IS_TOP3_DAY"),
                "fill_price": r.get("fill_price"),
                "bid600": r.get("bid600"),
                "bid750": r.get("bid750"),
                "ret_entry_to_600": r.get("ret_entry_to_600"),
                "ret_600_to_750": r.get("ret_600_to_750"),
                "ret_entry_to_750": r.get("ret_entry_to_750"),
                "incremental_value_600_750_yen": r.get("incremental_value_600_750_yen"),
                "identity_pass": r.get("identity_pass"),
                "actual_exit_price": r.get("exit_price"),
                "pnl_yen_100": r.get("pnl_yen_100"),
            }
            for r in merged
            if r.get("canonical_class") == "EXTEND_TO_750"
        ],
        "same": (sa_all.get("cells") or [])
        + [{"slice": "TOP3", **c} for c in (sa_top.get("cells") or [])]
        + [{"slice": "REST11", **c} for c in (sa_rest.get("cells") or [])],
        "top3": _cmp_sheet("TOP3", top_cmp, sa_top, incremental_block(_days(merged, PREDECLARED_TOP3)) if path_rows else {}),
        "rest11": _cmp_sheet("REST11", rest_cmp, sa_rest, incremental_block(_days(merged, REST11)) if path_rows else {}),
    }

    report = {
        "task": "P3-4",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "strategy validation"],
        "SOURCE_OF_TRUTH": "P3-3 canonical FULL14 / P1 CURRENT_RUNTIME_REPLAY",
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "PRIMARY_FULL14": {"trades": len(merged), "pnl": round(total, 2), "PF": P1_PF, "expected_trades": P1_TRADES, "expected_pnl": P1_PNL},
        "NOTE_DO_NOT_ADD_SHARES": "Overlapping sets. Shares are not summed.",
        "OVERLAP_16CELLS": cells,
        "INTERSECTIONS": inter,
        "RESIDUALS": resid,
        "OVERLAP": {
            "FULL_INTERSECTION": {"n": full.get("n"), "pnl": full.get("pnl"), "share": full.get("share_of_total")},
            "NONE_OF_4_FLAGS": {"n": none.get("n"), "pnl": none.get("pnl"), "PF": none.get("PF")},
        },
        "RESIDUAL_SUMMARY": {
            "0905_NOT_TOP3": {"n": resid_named("09:05 AND NOT TOP3_DAY").get("n"), "pnl": resid_named("09:05 AND NOT TOP3_DAY").get("pnl")},
            "TOP3_NOT_0905": {"n": resid_named("TOP3_DAY AND NOT 09:05").get("n"), "pnl": resid_named("TOP3_DAY AND NOT 09:05").get("pnl")},
            "EXTEND_NOT_TOP3": {"n": resid_named("EXTEND750 AND NOT TOP3_DAY").get("n"), "pnl": resid_named("EXTEND750 AND NOT TOP3_DAY").get("pnl")},
            "EXTEND_NOT_0905": {"n": resid_named("EXTEND750 AND NOT 09:05").get("n"), "pnl": resid_named("EXTEND750 AND NOT 09:05").get("pnl")},
            "TOP3_SYMBOL_NOT_0905": {"n": resid_named("TOP3_SYMBOL AND NOT 09:05").get("n"), "pnl": resid_named("TOP3_SYMBOL AND NOT 09:05").get("pnl")},
        },
        "TOP10_MEMBERSHIP": t10,
        "ANCHOR_0905": a0905,
        "SYMBOLS": syms,
        "EXTEND750": ext,
        "WINNER_CONCENTRATION": cluster.get("WINNER_CONCENTRATION"),
        "WINNER_CONCENTRATION_DETAIL": cluster,
        "EXTENSION_DECISION_RECONCILE": "PASS" if dec_pass else ("FAIL" if path_rows else "NOT_RUN"),
        "extension_decision": {
            "n_reached_600": n_reached,
            "matched": n_match,
            "mismatch": n_mis,
            "continuation": "MFE60_IMB10",
            "guard": "IMB_p5_t-10",
            "note": (
                "Canonical EXTEND/EXIT600 matched Dual Lane 600_DECISION traces. "
                "That event is the freeze. Dual Lane samples ~0.5s so the first tick with "
                "off>=600 may be after fill+600; that is Runtime clock, not 750-window leak. "
                "decision_future_leak counts only 600_DECISION ticks at/after 750s. "
                "Independent full-board reconstruct at event_time<=t600 is descriptive. "
                "New features not added."
            ),
            "dual_lane_600_fire": {
                "n_reached_600": n_reached,
                "n_fire_off_gt_600": sum(1 for r in reached if r.get("dual_lane_600_fire_after_t600")),
                "max_runtime_off_now": max(
                    (float(r["runtime_off_now"]) for r in reached if r.get("runtime_off_now") is not None),
                    default=None,
                ),
                "n_off_ge_750": sum(
                    1
                    for r in reached
                    if r.get("runtime_off_now") is not None and float(r["runtime_off_now"]) + 1e-9 >= 750.0
                ),
            },
        },
        "EXTEND_VS_EXIT600": {"ALL": all_cmp, "TOP3": top_cmp, "REST11": rest_cmp},
        "SAME_ANCHOR": {"ALL": sa_all, "TOP3": sa_top, "REST11": sa_rest},
        "EXTENSION_GATE": gate.get("EXTENSION_GATE"),
        "EXTENSION_GATE_DETAIL": gate,
        "EXTEND_INCREMENTAL": incr,
        "EXTENSION_MECHANISM": interp.get("EXTENSION_MECHANISM"),
        "EXTENSION_INTERPRETATION": interp,
        "SEPARATE_ANSWERS": {
            "A_EXTEND_FINAL_PNL_HIGH": True,
            "B_IDENTIFIED_AT_600": gate.get("EXTENSION_GATE"),
            "C_INCREMENTAL_600_750": interp.get("case"),
            "note": "A / B / C are distinct. High EXTEND PnL is not evidence that holding to 750 is a strategy.",
        },
        "ANCHOR_DIRECTIONAL_EDGE": "NOT_SUPPORTED",
        "EXECUTION_EDGE": "SUPPORTED",
        "POST_FILL_DIRECTION": "MIXED",
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": not leak_ok if path_rows else False,
        "decision_future_leak": int(leak_decision),
        "checkpoint_future_leak": int(leak_checkpoint),
        "identity_n": int(identity_n),
        "identity_fail": int(identity_fail),
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_dual_lane": _file_sha("src/small_paper/v1r_live_dual_lane.py"),
            "file_sha_p3_4_decision": _file_sha("src/research/fixed_winner_cluster_extension_p3_4/decision.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
    }
    unused = group_pnl, _share
    del unused
    return json_sanitize(report)


def _cmp_sheet(slice_name: str, cmp: dict, sa: dict, incr: dict) -> list[dict[str, Any]]:
    rows = [{"slice": slice_name, "block": "COMPARE", **{k: v for k, v in cmp.items()}}]
    rows.append({"slice": slice_name, "block": "SAME_ANCHOR", **{k: v for k, v in sa.items() if k != "cells"}})
    for part, blk in (incr or {}).items():
        if isinstance(blk, dict):
            rows.append({"slice": slice_name, "block": f"INCR_{part}", **blk})
    return rows


def render_md(rep: dict[str, Any]) -> str:
    ov = rep.get("OVERLAP") or {}
    full = ov.get("FULL_INTERSECTION") or {}
    none = ov.get("NONE_OF_4_FLAGS") or {}
    rs = rep.get("RESIDUAL_SUMMARY") or {}
    vs = (rep.get("EXTEND_VS_EXIT600") or {}).get("ALL") or {}
    ext = (vs.get("EXTEND_TO_750") or {}).get("future_bid") or {}
    exi = (vs.get("EXIT_AT_600") or {}).get("future_bid") or {}
    rest = (rep.get("EXTEND_VS_EXIT600") or {}).get("REST11") or {}
    sa = (rep.get("SAME_ANCHOR") or {}).get("ALL") or {}
    incr = (rep.get("EXTEND_INCREMENTAL") or {}).get("ALL") or {}
    top_i = (rep.get("EXTEND_INCREMENTAL") or {}).get("TOP3") or {}
    rest_i = (rep.get("EXTEND_INCREMENTAL") or {}).get("REST11") or {}
    dec = rep.get("extension_decision") or {}
    lines = [
        "# P3-4 Winner Cluster & 600→750 Extension Mechanism Audit",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / strategy validation",
        "",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        "",
        "OVERLAP:",
        "FULL_INTERSECTION:",
        f"n: {full.get('n')}",
        f"pnl: {full.get('pnl')}",
        f"share: {full.get('share')}",
        "",
        "NONE_OF_4_FLAGS:",
        f"n: {none.get('n')}",
        f"pnl: {none.get('pnl')}",
        f"PF: {none.get('PF')}",
        "",
        "RESIDUALS:",
        f"0905_NOT_TOP3: n={((rs.get('0905_NOT_TOP3') or {}).get('n'))} pnl={((rs.get('0905_NOT_TOP3') or {}).get('pnl'))}",
        f"TOP3_NOT_0905: n={((rs.get('TOP3_NOT_0905') or {}).get('n'))} pnl={((rs.get('TOP3_NOT_0905') or {}).get('pnl'))}",
        f"EXTEND_NOT_TOP3: n={((rs.get('EXTEND_NOT_TOP3') or {}).get('n'))} pnl={((rs.get('EXTEND_NOT_TOP3') or {}).get('pnl'))}",
        f"EXTEND_NOT_0905: n={((rs.get('EXTEND_NOT_0905') or {}).get('n'))} pnl={((rs.get('EXTEND_NOT_0905') or {}).get('pnl'))}",
        f"TOP3_SYMBOL_NOT_0905: n={((rs.get('TOP3_SYMBOL_NOT_0905') or {}).get('n'))} pnl={((rs.get('TOP3_SYMBOL_NOT_0905') or {}).get('pnl'))}",
        "",
        "Shares of overlapping sets are not added.",
        "",
        f"WINNER_CONCENTRATION: `{rep.get('WINNER_CONCENTRATION')}`",
        f"{((rep.get('WINNER_CONCENTRATION_DETAIL') or {}).get('why'))}",
        "",
        f"EXTENSION_DECISION_RECONCILE: `{rep.get('EXTENSION_DECISION_RECONCILE')}`",
        f"n: {dec.get('n_reached_600')} matched: {dec.get('matched')} mismatch: {dec.get('mismatch')}",
        f"dual_lane_600_fire: {json.dumps(dec.get('dual_lane_600_fire'), ensure_ascii=False, default=str)}",
        "",
        "EXTEND_VS_EXIT600:",
        "EXTEND:",
        f"n: {ext.get('n')}",
        f"future_bid_mean: {ext.get('mean')}",
        f"future_bid_median: {ext.get('median')}",
        "EXIT600:",
        f"n: {exi.get('n')}",
        f"future_bid_mean: {exi.get('mean')}",
        f"future_bid_median: {exi.get('median')}",
        "",
        "REST11:",
        f"EXTEND future_bid: {json.dumps((rest.get('EXTEND_TO_750') or {}).get('future_bid'), ensure_ascii=False, default=str)}",
        f"EXIT600 future_bid: {json.dumps((rest.get('EXIT_AT_600') or {}).get('future_bid'), ensure_ascii=False, default=str)}",
        "",
        "SAME_ANCHOR:",
        f"better: {sa.get('extend_better')}",
        f"worse: {sa.get('extend_worse')}",
        f"equal: {sa.get('equal')}",
        "",
        f"EXTENSION_GATE: `{rep.get('EXTENSION_GATE')}`",
        "",
        "EXTEND_INCREMENTAL:",
        "ENTRY_TO_600:",
        f"mean: {((incr.get('ENTRY_TO_600') or {}).get('mean'))}",
        f"median: {((incr.get('ENTRY_TO_600') or {}).get('median'))}",
        "600_TO_750:",
        f"mean: {((incr.get('600_TO_750') or {}).get('mean'))}",
        f"median: {((incr.get('600_TO_750') or {}).get('median'))}",
        f"positive_rate: {((incr.get('600_TO_750') or {}).get('positive_rate'))}",
        "EXECUTABLE_INCREMENTAL_VALUE_YEN:",
        f"sum: {((incr.get('EXECUTABLE_INCREMENTAL_VALUE_YEN') or {}).get('sum'))}",
        f"mean: {((incr.get('EXECUTABLE_INCREMENTAL_VALUE_YEN') or {}).get('mean'))}",
        f"median: {((incr.get('EXECUTABLE_INCREMENTAL_VALUE_YEN') or {}).get('median'))}",
        "",
        "TOP3_EXTEND_600_TO_750:",
        json.dumps(top_i.get("600_TO_750"), ensure_ascii=False, default=str),
        "REST11_EXTEND_600_TO_750:",
        json.dumps(rest_i.get("600_TO_750"), ensure_ascii=False, default=str),
        "",
        f"EXTENSION_MECHANISM: `{rep.get('EXTENSION_MECHANISM')}`",
        f"{((rep.get('EXTENSION_INTERPRETATION') or {}).get('case_note'))}",
        "",
        "A / B / C (kept separate):",
        f"A EXTEND final PnL high: {((rep.get('SEPARATE_ANSWERS') or {}).get('A_EXTEND_FINAL_PNL_HIGH'))}",
        f"B identified at 600: {((rep.get('SEPARATE_ANSWERS') or {}).get('B_IDENTIFIED_AT_600'))}",
        f"C incremental 600→750 case: {((rep.get('SEPARATE_ANSWERS') or {}).get('C_INCREMENTAL_600_750'))}",
        "",
        "ANCHOR_DIRECTIONAL_EDGE: NOT_SUPPORTED",
        "EXECUTION_EDGE: SUPPORTED",
        "POST_FILL_DIRECTION: MIXED",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        f"decision_future_leak: {rep.get('decision_future_leak')}",
        f"checkpoint_future_leak: {rep.get('checkpoint_future_leak')}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ov = public.get("OVERLAP") or {}
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("WINNER_CONCENTRATION", public.get("WINNER_CONCENTRATION")),
            ("FULL_INTERSECTION_n", ((ov.get("FULL_INTERSECTION") or {}).get("n"))),
            ("FULL_INTERSECTION_pnl", ((ov.get("FULL_INTERSECTION") or {}).get("pnl"))),
            ("NONE_OF_4_n", ((ov.get("NONE_OF_4_FLAGS") or {}).get("n"))),
            ("NONE_OF_4_pnl", ((ov.get("NONE_OF_4_FLAGS") or {}).get("pnl"))),
            ("EXTENSION_DECISION_RECONCILE", public.get("EXTENSION_DECISION_RECONCILE")),
            ("EXTENSION_GATE", public.get("EXTENSION_GATE")),
            ("EXTENSION_MECHANISM", public.get("EXTENSION_MECHANISM")),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("decision_future_leak", public.get("decision_future_leak")),
            ("checkpoint_future_leak", public.get("checkpoint_future_leak")),
            ("NEW_STRATEGY_TESTED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Cluster_Flags"), sheets.get("flags") or [])
    _write_rows(wb.create_sheet("Overlap_16Cells"), sheets.get("cells") or [])
    _write_rows(wb.create_sheet("Intersections"), sheets.get("inter") or [])
    _write_rows(wb.create_sheet("Residuals"), sheets.get("resid") or [])
    _write_rows(wb.create_sheet("Top10_Membership"), sheets.get("top10") or [])
    _write_rows(wb.create_sheet("Anchor0905"), sheets.get("a0905") or [])
    _write_rows(wb.create_sheet("Symbols"), sheets.get("symbols") or [])
    _write_rows(wb.create_sheet("Extend750"), sheets.get("extend") or [])
    _write_rows(wb.create_sheet("Extension_Reconcile"), sheets.get("recon") or [])
    _write_rows(wb.create_sheet("Extension_600_750"), sheets.get("path") or [])
    _write_rows(wb.create_sheet("Extension_Incremental"), sheets.get("incr") or [])
    _write_rows(wb.create_sheet("Same_Anchor"), sheets.get("same") or [])
    _write_rows(wb.create_sheet("Top3"), sheets.get("top3") or [])
    _write_rows(wb.create_sheet("Rest11"), sheets.get("rest11") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(ident, list((public.get("Identity") or {}).items()) + [("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE"))])
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("NEW_STRATEGY_TESTED", False),
            ("RETUNING_DONE", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("COUNTERFACTUAL_STRATEGY", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
