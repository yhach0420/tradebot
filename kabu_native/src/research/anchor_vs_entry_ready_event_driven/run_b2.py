#!/usr/bin/env python3
"""B2 ENTRY_READY event-driven vs reused Fixed-Anchor A. Does not overwrite B1."""
from __future__ import annotations

import json
import math
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

NATIVE = Path(__file__).resolve().parents[3]
SRC = NATIVE / "src"
REPO = NATIVE.parent
B1_OUT = NATIVE / "results" / "research" / "anchor_vs_event_driven_v1r"
OUT = NATIVE / "results" / "research" / "anchor_vs_entry_ready_event_driven_v1r"
MAX_WORKERS = 4

EXPECTED_STRATEGY = "9ad4ba2730892d40c757d940b82480e620e502e3e789839120e90b18be082547"
EXPECTED_ENTRY = "f2887bb2be539cc173aee438a43ee8afb8cfa2b8c31380937ecd843e90dd9b29"
EXPECTED_EXIT = "6cc3b8aade76e323682ec39dfd06878aab0ff1a99dd42922744b0054a7ea3255"
EXPECTED_A_FULL = {
    "trades": 254,
    "pnl": 2244550.0,
    "pf": 3.8737596825,
    "maxdd": -210300.0,
    "AM_trades": 156,
    "AM_pnl": 2024700.0,
    "PM_trades": 98,
    "PM_pnl": 219850.0,
}

for _k in (
    "KABU_V1R_ENTRY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_NOTIFY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_CAP_BLOCKED_WEBHOOK_URL",
    "KABU_DISCORD_RESEARCH_WEBHOOK_URL",
    "KABU_SHADOW_DISCORD_WEBHOOK_URL",
    "KABU_DISCORD_MARKET_CAPTURE_WEBHOOK_URL",
    "KABU_MARKET_CAPTURE_WEBHOOK_URL",
):
    os.environ.pop(_k, None)
os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"
os.environ["PYTHONPATH"] = f"{SRC};{REPO}" if os.name == "nt" else f"{SRC}:{REPO}"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from research.anchor_vs_event_driven.run_comparison import (  # noqa: E402
    JST,
    LIVE_FEATURES,
    _bare,
    _boot,
    _median,
    _p90,
    _sess_stats,
    capture_event_epoch,
    extract_trades,
    in_session,
    iter_push,
    next_anchor_after,
    record_event_stamp,
)


def verify_identity() -> dict[str, Any]:
    from small_paper.v1r_exit_v2_activation_gate import ENTRY_SHA as GATE_ENTRY
    from small_paper.v1r_exit_v2_activation_gate import STRATEGY_SHA
    from small_paper.v1r_exit_v2_contract import EXIT_V2_CANDIDATE_SHA
    from small_paper.v1r_native_entry_live import ENTRY_SHA

    got = {
        "strategy_sha": str(STRATEGY_SHA),
        "entry_sha": str(ENTRY_SHA),
        "entry_sha_gate": str(GATE_ENTRY),
        "exit_sha": str(EXIT_V2_CANDIDATE_SHA),
    }
    ok = (
        got["strategy_sha"] == EXPECTED_STRATEGY
        and got["entry_sha"] == EXPECTED_ENTRY
        and got["entry_sha_gate"] == EXPECTED_ENTRY
        and got["exit_sha"] == EXPECTED_EXIT
    )
    got["ok"] = ok
    return got


def load_b1() -> dict[str, Any]:
    path = B1_OUT / "report.json"
    if not path.is_file():
        raise SystemExit(f"B1 report missing: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def load_a_day(day: str) -> dict[str, Any]:
    path = B1_OUT / "day_cache" / f"{day}.json"
    if not path.is_file():
        raise SystemExit(f"B1 day cache missing for A reuse: {path}")
    body = json.loads(path.read_text(encoding="utf-8"))
    a = body.get("A") or {}
    if not a.get("ok"):
        raise SystemExit(f"B1 A not ok for {day}")
    return {
        "ok": True,
        "trades": a.get("trades") or [],
        "blocker": None,
        "native_fills": a.get("native_fills"),
        "events_n": a.get("events_n"),
    }


def _stream_b2(day: str, capture_dir: Path, eng: Any, dual: Any) -> tuple[int, Optional[float]]:
    from research.e1_x22_actual_exit_factory.paths import session_end_epoch
    from small_paper.v1r_live_dual_lane import canonical_symbol_key

    last_dual_t: dict[str, float] = {}
    events_n = 0
    last_et: Optional[float] = None
    am_end = session_end_epoch(day, "AM")
    pm_end = session_end_epoch(day, "PM")
    am_closed = False
    pm_closed = False
    window_was = False

    def maybe_close(t: float) -> None:
        nonlocal am_closed, pm_closed
        if not am_closed and t + 1e-9 >= am_end:
            dual.close_open_at_session_end(event_t=am_end, session="AM")
            eng.on_tick_fill_check(event_t=am_end)
            am_closed = True
        if not pm_closed and t + 1e-9 >= pm_end:
            dual.close_open_at_session_end(event_t=pm_end, session="PM")
            eng.on_tick_fill_check(event_t=pm_end)
            pm_closed = True

    for rec in iter_push(capture_dir):
        pay = dict(rec.get("payload") or rec.get("original_payload") or {})
        et = capture_event_epoch(rec, pay)
        if et is None:
            continue
        recv = record_event_stamp(rec) or datetime.fromtimestamp(float(et), JST).isoformat(
            timespec="milliseconds"
        )
        seq = int(rec.get("sequence") or 0)
        sym = _bare(rec.get("symbol"))
        pay["received_at"] = recv
        pay["recorded_at"] = recv
        pay["sequence"] = seq
        pay["__ingress_sequence__"] = seq
        pay["__ingress_received_at__"] = recv
        last_et = float(et)
        events_n += 1
        t = float(et)
        window_now = in_session(t)
        window_opened = (not window_was) and window_now
        window_closed = window_was and (not window_now)
        window_was = window_now

        exp0 = int(eng.exposure())
        open0 = set(eng.open_symbols)
        pend0 = set(eng.pending.keys())

        maybe_close(t)
        eng.process_market_push(symbol=sym, payload=pay, event_t=t)
        if hasattr(eng, "on_stream_event"):
            eng.on_stream_event(sym, t)
        key = canonical_symbol_key(sym)
        open_here = (key in dual.primary and not dual.primary[key].closed) or (
            key in dual.control and not dual.control[key].closed
        )
        if open_here:
            prev = last_dual_t.get(key)
            if prev is None or (t - prev) >= 0.5 - 1e-12:
                last_dual_t[key] = t
                dual.on_tick(symbol=sym, payload=pay, event_t=t, push_sequence=seq)

        released = (open0 - set(eng.open_symbols)) | (pend0 - set(eng.pending.keys()))
        occupancy_dropped = int(eng.exposure()) < exp0 or bool(released)
        eng.on_causal_eval(
            symbol=sym,
            t=t,
            window_opened=window_opened,
            window_closed=window_closed,
            occupancy_dropped=occupancy_dropped,
            released=released,
        )
        if events_n % 200000 == 0:
            eng.events.clear()
            if getattr(eng, "ingest_audit", None) is not None:
                eng.ingest_audit.clear()
    if last_et is not None:
        maybe_close(float(last_et))
        if in_session(float(last_et)):
            eng.on_causal_eval(
                symbol="",
                t=float(last_et),
                window_opened=False,
                window_closed=False,
                occupancy_dropped=True,
                released=set(),
            )
    return events_n, last_et


def replay_b2_day(payload: dict[str, Any]) -> dict[str, Any]:
    if str(SRC) not in sys.path:
        sys.path.insert(0, str(SRC))
    if str(REPO) not in sys.path:
        sys.path.insert(0, str(REPO))
    os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"
    day = str(payload["date"])
    capture_dir = Path(payload["capture_path"])
    universe = list(payload["universe"])
    cap_class = str(payload["capture_class"])
    t_wall = time.perf_counter()

    import numpy as np
    from research.e1_x34b_entry_execution.features import preentry_from_board
    from research.e1_x36_joint_allocator.replay import simulate_joint
    from small_paper.v1r_live_dual_lane import get_dual_lane, live_primary_enabled
    from small_paper.v1r_native_entry_live import FEATURE_ORDER, PendingOrder, V1RNativeEntryLive
    from small_paper.v1r_primary_runtime import POSITION_CAP, WAIT_SEC

    class EntryReadyEngine(V1RNativeEntryLive):
        def __init__(self, *a: Any, **k: Any) -> None:
            super().__init__(*a, **k)
            self._ready: dict[str, bool] = {}
            self._block: dict[str, str] = {}
            self._raw_cache: dict[str, tuple[bool, dict[str, Any], float, float]] = {}
            self.signals: list[dict[str, Any]] = []
            self._timing_open: list[dict[str, Any]] = []
            self.timing_done: list[dict[str, Any]] = []
            self.cause_counts: dict[str, int] = {
                "RAW_CONDITION_RISE": 0,
                "SESSION_OPEN_READY": 0,
                "COOLDOWN_RELEASE_READY": 0,
                "SLOT_RELEASE_READY": 0,
                "SAME_SYMBOL_RELEASE_READY": 0,
                "OTHER_EXISTING_READY_TRANSITION": 0,
            }

        def _mid(self, symbol: str, t: float) -> Optional[float]:
            board = self._board_arrays(symbol)
            if board["t"].size == 0:
                return None
            i = int(np.searchsorted(board["t"], float(t), side="right") - 1)
            if i < 0:
                return None
            ask = float(board["ask"][i])
            bid = float(board["bid"][i])
            if not (math.isfinite(ask) and math.isfinite(bid) and ask > 0 and bid > 0):
                return None
            return (ask + bid) / 2.0

        def _bid(self, symbol: str, t: float) -> Optional[float]:
            board = self._board_arrays(symbol)
            if board["t"].size == 0:
                return None
            i = int(np.searchsorted(board["t"], float(t), side="right") - 1)
            if i < 0:
                return None
            bid = float(board["bid"][i])
            return bid if math.isfinite(bid) and bid > 0 else None

        def _raw_now(self, symbol: str, t: float, *, use_cache: bool = False) -> tuple[bool, dict[str, Any], float, float]:
            if use_cache and symbol in self._raw_cache:
                return self._raw_cache[symbol]
            board = self._board_arrays(symbol)
            feats = preentry_from_board(board, float(t))
            if any(feats.get(f) is None or not np.isfinite(feats.get(f)) for f in FEATURE_ORDER):
                out = (False, feats, float("nan"), 0.0)
                self._raw_cache[symbol] = out
                return out
            score = float(self.score_fn(feats))
            if not np.isfinite(score):
                out = (False, feats, score, 0.0)
                self._raw_cache[symbol] = out
                return out
            bid = self._bid(symbol, t) or 0.0
            if bid <= 0:
                out = (False, feats, score, bid)
                self._raw_cache[symbol] = out
                return out
            out = (True, feats, score, bid)
            self._raw_cache[symbol] = out
            return out

        def on_stream_event(self, symbol: str, t: float) -> None:
            mid = self._mid(symbol, t)
            still: list[dict[str, Any]] = []
            for rec in self._timing_open:
                if rec["symbol"] != symbol:
                    still.append(rec)
                    continue
                if mid is not None and rec.get("signal_mid"):
                    bps = (mid / float(rec["signal_mid"]) - 1.0) * 10000.0
                    rec["mfe_bps"] = max(float(rec.get("mfe_bps") or 0.0), bps)
                    rec["mae_bps"] = min(float(rec.get("mae_bps") or 0.0), bps)
                na = rec.get("next_anchor_t")
                if na is not None and float(t) + 1e-12 >= float(na):
                    rec["anchor_mid"] = self._mid(symbol, float(na))
                    rec["anchor_bid"] = self._bid(symbol, float(na))
                    rec["wait_saved_sec"] = rec.get("time_to_next_anchor_sec")
                    self.timing_done.append(rec)
                else:
                    still.append(rec)
            self._timing_open = still

        def process_market_push(
            self,
            *,
            symbol: str,
            payload: dict[str, Any],
            event_t: Optional[float] = None,
        ) -> dict[str, Any]:
            t = float(event_t if event_t is not None else 0.0)
            ing = self.ingest_push(symbol=symbol, payload=payload, event_t=t)
            if not ing.get("ingested") and ing.get("reason") == "duplicate_sequence":
                ing["fill_checked"] = False
                ing["anchor_fired"] = False
                return ing
            fills = self.on_tick_fill_check(event_t=t, payload=payload, symbol=symbol)
            ing["fill_checked"] = True
            ing["anchor_fired"] = False
            ing["fill_n"] = len(fills or [])
            return ing

        def _clear_ready_window(self) -> None:
            for s in list(self.universe):
                key = _bare(s)
                self._ready[key] = False
                self._block[key] = "window"

        def _admit_pending(
            self,
            *,
            symbol: str,
            t: float,
            feats: dict[str, Any],
            score: float,
            bid: float,
            cause: str,
        ) -> None:
            if symbol in self.pending or symbol in self.open_symbols:
                return
            if self.exposure() >= POSITION_CAP:
                self._ready[symbol] = False
                self._block[symbol] = "cap"
                return
            was = bool(self._ready.get(symbol, False))
            dt = datetime.fromtimestamp(float(t), JST)
            sess = "AM" if dt.hour < 12 else "PM"
            na_t, na_lbl = next_anchor_after(self.trading_date or dt.strftime("%Y%m%d"), t)
            wait = (float(na_t) - float(t)) if na_t is not None else None
            po = PendingOrder(
                symbol=symbol,
                signal_time=float(t),
                limit_price=float(bid),
                score=float(score),
                rank=0,
                anchor="ENTRY_READY",
                session=sess,
                date=str(self.trading_date),
                features={f: feats.get(f) for f in FEATURE_ORDER},
            )
            self.pending[po.symbol] = po
            self.primary_admitted += 1
            self._ready[symbol] = True
            self._block[symbol] = ""
            if not was:
                self.cause_counts[cause] = int(self.cause_counts.get(cause) or 0) + 1
                sig = {
                    "symbol": symbol,
                    "event_signal_time": float(t),
                    "session": sess,
                    "score": score,
                    "signal_bid": bid,
                    "signal_mid": self._mid(symbol, t),
                    "next_fixed_anchor": na_lbl,
                    "next_anchor_t": na_t,
                    "time_to_next_anchor_sec": round(wait, 3) if wait is not None else None,
                    "wait_saved_sec": round(wait, 3) if wait is not None else None,
                    "mfe_bps": 0.0,
                    "mae_bps": 0.0,
                    "ready_cause": cause,
                    "admitted": True,
                }
                self.signals.append(sig)
                if na_t is not None:
                    self._timing_open.append(sig)
            self.events.clear()

        def _admit_cohort(self, t: float, *, cause: str) -> None:
            if not in_session(t):
                self._clear_ready_window()
                return
            dual_pre = get_dual_lane(trace_dir=self.trace_dir) if live_primary_enabled() else None
            if dual_pre is not None:
                dual_pre.maybe_session_close(event_t=float(t))
            events: list[dict[str, Any]] = []
            bundle: dict[str, tuple[dict[str, Any], float, float]] = {}
            dt = datetime.fromtimestamp(float(t), JST)
            sess = "AM" if dt.hour < 12 else "PM"
            for raw_sym in list(self.universe):
                s = _bare(raw_sym)
                if s in self.pending or s in self.open_symbols:
                    self._ready[s] = True
                    continue
                ok, feats, score, bid = self._raw_now(s, t, use_cache=True)
                if not ok:
                    self._ready[s] = False
                    self._block[s] = "raw"
                    continue
                events.append(
                    {
                        "date": str(self.trading_date),
                        "symbol": s,
                        "session": sess,
                        "signal_time": float(t),
                        "filled": False,
                        "limit_price": float(bid),
                        "bid0": float(bid),
                        **{f: feats.get(f) for f in FEATURE_ORDER},
                        "score_preview": score,
                    }
                )
                bundle[s] = (feats, score, bid)
            if not events:
                return
            sim = simulate_joint([dict(e) for e in events], score_fn=self.score_fn)
            ranked = sorted(
                [e for e in sim["events"] if e.get("alloc_score") is not None],
                key=lambda e: (-float(e.get("alloc_score") or 0.0), str(e.get("symbol") or "")),
            )
            for e in ranked:
                s = _bare(e.get("symbol"))
                if not e.get("admitted"):
                    self._ready[s] = False
                    self._block[s] = str(e.get("block_reason") or "NOT_ADMITTED")
                    continue
                if self.exposure() >= POSITION_CAP:
                    self._ready[s] = False
                    self._block[s] = "cap"
                    continue
                feats, score, bid = bundle[s]
                use_score = float(e.get("alloc_score") if e.get("alloc_score") is not None else score)
                self._admit_pending(symbol=s, t=t, feats=feats, score=use_score, bid=bid, cause=cause)

        def _consider_symbol(self, symbol: str, t: float, *, cause: str) -> None:
            if not symbol:
                return
            if not in_session(t):
                self._ready[symbol] = False
                self._block[symbol] = "window"
                return
            if symbol in self.pending or symbol in self.open_symbols:
                self._ready[symbol] = True
                return
            ok, feats, score, bid = self._raw_now(symbol, t)
            if not ok:
                self._ready[symbol] = False
                self._block[symbol] = "raw"
                return
            if self.exposure() >= POSITION_CAP:
                self._ready[symbol] = False
                self._block[symbol] = "cap"
                return
            self._admit_pending(symbol=symbol, t=t, feats=feats, score=score, bid=bid, cause=cause)

        def on_causal_eval(
            self,
            *,
            symbol: str,
            t: float,
            window_opened: bool,
            window_closed: bool,
            occupancy_dropped: bool,
            released: set[str],
        ) -> None:
            if symbol:
                self._raw_now(_bare(symbol), t)
            if window_closed:
                self._clear_ready_window()
                return
            if not in_session(t):
                if symbol:
                    self._ready[_bare(symbol)] = False
                    self._block[_bare(symbol)] = "window"
                return
            if window_opened:
                self._admit_cohort(t, cause="SESSION_OPEN_READY")
                return
            if occupancy_dropped:
                for s in released:
                    self._ready[_bare(s)] = False
                    self._block[_bare(s)] = "released"
                if released and len(released) == 1 and next(iter(released)) == _bare(symbol):
                    cause = "SAME_SYMBOL_RELEASE_READY"
                else:
                    cause = "SLOT_RELEASE_READY"
                self._admit_cohort(t, cause=cause)
                return
            self._consider_symbol(_bare(symbol), t, cause="RAW_CONDITION_RISE")

    eng, dual = _boot(universe, EntryReadyEngine)
    if dual is None or not eng.ready:
        return {
            "date": day,
            "capture_class": cap_class,
            "B2": {
                "ok": False,
                "blocker": getattr(eng, "fail_reason", "NATIVE_NOT_READY"),
                "trades": [],
            },
            "elapsed_sec": round(time.perf_counter() - t_wall, 1),
        }
    eng.trading_date = day
    events_n, _ = _stream_b2(day, capture_dir, eng, dual)
    leftover = []
    wm = float(eng.event_time_watermark or 0.0)
    for rec in eng._timing_open:
        na = rec.get("next_anchor_t")
        if na is not None and wm + 1e-9 >= float(na):
            rec["anchor_mid"] = eng._mid(rec["symbol"], float(na))
            rec["anchor_bid"] = eng._bid(rec["symbol"], float(na))
        leftover.append(rec)
    b2 = {
        "ok": True,
        "blocker": None,
        "trades": extract_trades(dual),
        "events_n": events_n,
        "native_admitted": int(eng.primary_admitted),
        "native_fills": int(eng.primary_fills),
        "native_expired": int(eng.primary_expired),
        "signals": list(eng.signals),
        "timing": list(eng.timing_done) + leftover,
        "cause_counts": dict(eng.cause_counts),
    }
    out = {
        "date": day,
        "capture_class": cap_class,
        "universe_n": len(universe),
        "elapsed_sec": round(time.perf_counter() - t_wall, 1),
        "B2": b2,
    }
    try:
        cache_dir = OUT / "day_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        (cache_dir / f"{day}.json").write_text(
            json.dumps(out, ensure_ascii=False, default=str) + "\n", encoding="utf-8"
        )
    except Exception:
        pass
    return out


def build_timing(day_rows: list[dict[str, Any]], *, full_only: bool) -> dict[str, Any]:
    use = [r for r in day_rows if (r.get("B2") or {}).get("ok")]
    if full_only:
        use = [r for r in use if r.get("capture_class") == "FULL"]
    signals: list[dict[str, Any]] = []
    causes: dict[str, int] = {
        "RAW_CONDITION_RISE": 0,
        "SESSION_OPEN_READY": 0,
        "COOLDOWN_RELEASE_READY": 0,
        "SLOT_RELEASE_READY": 0,
        "SAME_SYMBOL_RELEASE_READY": 0,
        "OTHER_EXISTING_READY_TRANSITION": 0,
    }
    waits: list[float] = []
    deltas: list[float] = []
    mfe_pos = 0
    pm_signals = 0
    for r in use:
        day = r["date"]
        for k, v in ((r.get("B2") or {}).get("cause_counts") or {}).items():
            causes[k] = int(causes.get(k) or 0) + int(v or 0)
        for s in (r.get("B2") or {}).get("timing") or (r.get("B2") or {}).get("signals") or []:
            row = {**s, "date": day}
            signals.append(row)
            if row.get("session") == "PM":
                pm_signals += 1
            w = row.get("wait_saved_sec")
            if w is None and row.get("next_anchor_t") is not None and row.get("event_signal_time") is not None:
                w = float(row["next_anchor_t"]) - float(row["event_signal_time"])
                row["wait_saved_sec"] = round(w, 3)
            if w is not None:
                waits.append(float(w))
            sp = row.get("signal_bid") or row.get("signal_mid")
            ap = row.get("anchor_bid") or row.get("anchor_mid")
            if sp and ap and float(ap) > 0:
                dps = (float(sp) / float(ap) - 1.0) * 10000.0
                row["price_delta_bps"] = round(dps, 4)
                deltas.append(dps)
            if float(row.get("mfe_bps") or 0) > 1e-9:
                mfe_pos += 1
    return {
        "event_driven_signals": len(signals),
        "pm_signals": pm_signals,
        "median_wait_saved_sec": _median(waits),
        "p90_wait_saved_sec": _p90(waits),
        "median_price_delta_bps": _median(deltas),
        "MFE_before_anchor": mfe_pos,
        "READY_TRANSITIONS": causes,
        "signals": signals,
    }


def decide_verdict(primary: dict[str, Any], day_rows: list[dict[str, Any]]) -> tuple[str, str]:
    a, b = primary["A"], primary["B2"]
    full = [r for r in day_rows if r.get("capture_class") == "FULL" and (r.get("B2") or {}).get("ok")]
    d_pnl = float(b.get("pnl") or 0) - float(a.get("pnl") or 0)
    pf_a, pf_b = a.get("pf"), b.get("pf")
    pf_better = False
    if pf_a is None and pf_b is not None:
        pf_better = True
    elif pf_a is not None and pf_b is not None:
        pf_better = float(pf_b) > float(pf_a) + 1e-9
    dd_a = float(a.get("max_drawdown") or 0)
    dd_b = float(b.get("max_drawdown") or 0)
    dd_ok = dd_b >= dd_a - 0.25 * abs(dd_a) - 1.0
    day_deltas = []
    for r in full:
        sa = _sess_stats((r.get("A") or {}).get("trades") or [])
        sb = _sess_stats((r.get("B2") or {}).get("trades") or [])
        day_deltas.append((r["date"], sb["pnl"] - sa["pnl"]))
    improve_days = sum(1 for _, d in day_deltas if d > 0)
    abs_sum = sum(abs(d) for _, d in day_deltas) or 1.0
    top = max(day_deltas, key=lambda x: abs(x[1])) if day_deltas else ("", 0.0)
    single = abs(top[1]) / abs_sum >= 0.70 if day_deltas else True
    note = f"improve_days={improve_days}/{len(full)} top_day={top[0]} share={abs(top[1]) / abs_sum:.2f}"
    a_clear = d_pnl < 0 and (not pf_better) and (dd_b <= dd_a + 1.0)
    if len(full) < 2:
        return "ANCHOR_EFFECT_INCONCLUSIVE", "fewer_than_2_full_days " + note
    if d_pnl > 0 and pf_better and dd_ok and improve_days >= 2 and not single:
        return "ANCHOR_CAUSES_OPPORTUNITY_LOSS_CANDIDATE", note
    if a_clear and improve_days <= 1:
        return "ANCHOR_HELPFUL", note
    return "ANCHOR_EFFECT_INCONCLUSIVE", note


def write_md(report: dict[str, Any]) -> str:
    p = report["PRIMARY_FULL"]
    r = report["REFERENCE_ALL_USABLE"]
    t = report["TIMING"]
    c = t.get("READY_TRANSITIONS") or {}
    a, b, d = p["A"], p["B2"], p["DELTA"]
    return "\n".join(
        [
            "# Fixed Anchor vs ENTRY_READY Event-Driven B2",
            "",
            "Timing-only. Existing ENTRY_READY conjunction. B1 not overwritten. Not runtime adoption.",
            "",
            f"**Final verdict:** `{report['verdict']}`",
            "",
            f"capture_period: `{report['capture_period']}`",
            f"usable_days: {len(report['usable_days'])}",
            f"full_days: {len(report['full_days'])}",
            "",
            f"B2_definition: {report['B2_definition']}",
            f"new_features: {report['new_features']}",
            f"new_thresholds: {report['new_thresholds']}",
            f"future_leak: {report['future_leak']}",
            "",
            "## PRIMARY_FULL",
            "",
            "A_FIXED:",
            f"trades: {a['trades']}",
            f"pnl: {a['pnl']}",
            f"pf: {a['pf']}",
            f"maxdd: {a['max_drawdown']}",
            f"AM: {a['AM_trades']} / {a['AM_pnl']}",
            f"PM: {a['PM_trades']} / {a['PM_pnl']}",
            "",
            "B2_ENTRY_READY:",
            f"trades: {b['trades']}",
            f"pnl: {b['pnl']}",
            f"pf: {b['pf']}",
            f"maxdd: {b['max_drawdown']}",
            f"AM: {b['AM_trades']} / {b['AM_pnl']}",
            f"PM: {b['PM_trades']} / {b['PM_pnl']}",
            "",
            "DELTA_B2_MINUS_A:",
            f"trades: {d['trades']}",
            f"pnl: {d['pnl']}",
            f"pf: {d['pf']}",
            f"maxdd: {d['maxdd']}",
            "",
            "READY_TRANSITIONS:",
            f"total: {t.get('event_driven_signals')}",
            f"raw_condition_rise: {c.get('RAW_CONDITION_RISE')}",
            f"session_open: {c.get('SESSION_OPEN_READY')}",
            f"cooldown_release: {c.get('COOLDOWN_RELEASE_READY')}",
            f"slot_release: {c.get('SLOT_RELEASE_READY')}",
            f"same_symbol_release: {c.get('SAME_SYMBOL_RELEASE_READY')}",
            "",
            "TIMING:",
            f"median_wait_saved_sec: {t.get('median_wait_saved_sec')}",
            f"p90_wait_saved_sec: {t.get('p90_wait_saved_sec')}",
            f"median_price_delta_bps: {t.get('median_price_delta_bps')}",
            f"MFE_before_anchor: {t.get('MFE_before_anchor')}",
            "",
            "REFERENCE_ALL_USABLE:",
            f"A: trades {r['A']['trades']} pnl {r['A']['pnl']} pf {r['A']['pf']} maxdd {r['A']['max_drawdown']}",
            f"B2: trades {r['B2']['trades']} pnl {r['B2']['pnl']} pf {r['B2']['pf']} maxdd {r['B2']['max_drawdown']}",
            f"delta pnl: {r['DELTA']['pnl']}",
            "",
            f"b1_defect_check: {report.get('b1_defect_check')}",
            f"day_dependency: {report.get('day_dependency')}",
            "",
            "STOP. No strategy change. No runtime adoption.",
            "",
        ]
    )


def write_excel(
    *,
    inventory: list[dict[str, Any]],
    day_rows: list[dict[str, Any]],
    primary: dict[str, Any],
    reference: dict[str, Any],
    timing_p: dict[str, Any],
    verdict: str,
) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    pa, pb = primary["A"], primary["B2"]
    for row in [
        ["field", "value"],
        ["verdict", verdict],
        ["PRIMARY_A_trades", pa.get("trades")],
        ["PRIMARY_A_pnl", pa.get("pnl")],
        ["PRIMARY_A_pf", pa.get("pf")],
        ["PRIMARY_A_maxdd", pa.get("max_drawdown")],
        ["PRIMARY_A_AM", f"{pa.get('AM_trades')} / {pa.get('AM_pnl')}"],
        ["PRIMARY_A_PM", f"{pa.get('PM_trades')} / {pa.get('PM_pnl')}"],
        ["PRIMARY_B2_trades", pb.get("trades")],
        ["PRIMARY_B2_pnl", pb.get("pnl")],
        ["PRIMARY_B2_pf", pb.get("pf")],
        ["PRIMARY_B2_maxdd", pb.get("max_drawdown")],
        ["PRIMARY_B2_AM", f"{pb.get('AM_trades')} / {pb.get('AM_pnl')}"],
        ["PRIMARY_B2_PM", f"{pb.get('PM_trades')} / {pb.get('PM_pnl')}"],
        ["DELTA_trades", (pb.get("trades") or 0) - (pa.get("trades") or 0)],
        ["DELTA_pnl", round(float(pb.get("pnl") or 0) - float(pa.get("pnl") or 0), 2)],
        ["median_wait_saved_sec", timing_p.get("median_wait_saved_sec")],
        ["pm_signals", timing_p.get("pm_signals")],
    ]:
        ws.append(row)

    ws = wb.create_sheet("Capture_Inventory")
    cols = [
        "date",
        "capture_class",
        "usable",
        "full",
        "universe_n",
        "first_event_time",
        "last_event_time",
        "exclusion_reason",
    ]
    ws.append(cols)
    for row in inventory:
        ws.append([row.get(c) for c in cols])

    ws = wb.create_sheet("Daily")
    ws.append(["date", "class", "A_trades", "A_pnl", "B2_trades", "B2_pnl", "d_pnl"])
    for r in day_rows:
        sa = _sess_stats((r.get("A") or {}).get("trades") or [])
        sb = _sess_stats((r.get("B2") or {}).get("trades") or [])
        ws.append([r.get("date"), r.get("capture_class"), sa["trades"], sa["pnl"], sb["trades"], sb["pnl"], sb["pnl"] - sa["pnl"]])

    ws = wb.create_sheet("Sessions")
    ws.append(["date", "A_AM_tr", "A_AM_pnl", "A_PM_tr", "A_PM_pnl", "B2_AM_tr", "B2_AM_pnl", "B2_PM_tr", "B2_PM_pnl"])
    for r in day_rows:
        sa = _sess_stats((r.get("A") or {}).get("trades") or [])
        sb = _sess_stats((r.get("B2") or {}).get("trades") or [])
        ws.append([r.get("date"), sa["AM_trades"], sa["AM_pnl"], sa["PM_trades"], sa["PM_pnl"], sb["AM_trades"], sb["AM_pnl"], sb["PM_trades"], sb["PM_pnl"]])

    def trade_sheet(name: str, key: str) -> None:
        w = wb.create_sheet(name)
        w.append(["date", "symbol", "session", "entry_time", "exit_time", "entry_price", "exit_price", "pnl_yen_100", "reason"])
        for r in day_rows:
            for t in (r.get(key) or {}).get("trades") or []:
                w.append([r.get("date"), t.get("symbol"), t.get("session"), t.get("entry_time"), t.get("exit_time"), t.get("entry_price"), t.get("exit_price"), t.get("pnl_yen_100"), t.get("reason")])

    trade_sheet("Portfolio_A", "A")
    trade_sheet("Portfolio_B2", "B2")

    ws = wb.create_sheet("Timing_Attribution")
    ws.append(
        [
            "date",
            "symbol",
            "event_signal_time",
            "session",
            "next_fixed_anchor",
            "wait_saved_sec",
            "signal_bid",
            "anchor_bid",
            "price_delta_bps",
            "mfe_bps",
            "mae_bps",
            "ready_cause",
        ]
    )
    for s in timing_p.get("signals") or []:
        ws.append(
            [
                s.get("date"),
                s.get("symbol"),
                s.get("event_signal_time"),
                s.get("session"),
                s.get("next_fixed_anchor"),
                s.get("wait_saved_sec"),
                s.get("signal_bid"),
                s.get("anchor_bid"),
                s.get("price_delta_bps"),
                s.get("mfe_bps"),
                s.get("mae_bps"),
                s.get("ready_cause"),
            ]
        )

    ws = wb.create_sheet("Ready_Transitions")
    ws.append(["cause", "count"])
    for k, v in (timing_p.get("READY_TRANSITIONS") or {}).items():
        ws.append([k, v])

    path = OUT / "anchor_vs_entry_ready_event_driven.xlsx"
    wb.save(path)
    return path


def main() -> int:
    ident = verify_identity()
    print(f"identity {ident}", flush=True)
    if not ident["ok"]:
        print("IDENTITY_MISMATCH STOP", flush=True)
        return 2
    b1 = load_b1()
    a_full = (b1.get("PRIMARY_FULL") or {}).get("A") or {}
    if int(a_full.get("trades") or 0) != EXPECTED_A_FULL["trades"] or abs(float(a_full.get("pnl") or 0) - EXPECTED_A_FULL["pnl"]) > 0.5:
        print(f"A baseline mismatch trades={a_full.get('trades')} pnl={a_full.get('pnl')} STOP", flush=True)
        return 2
    inventory = list(b1.get("inventory") or [])
    jobs = []
    for row in inventory:
        if not row.get("usable"):
            continue
        cap = row.get("capture_path")
        if not cap:
            continue
        uni_src = row.get("universe_source") or ""
        from research.anchor_vs_event_driven.run_comparison import historical_universe

        uni, _ = historical_universe(row["date"], Path(cap))
        if not uni:
            continue
        jobs.append(
            {
                "date": row["date"],
                "capture_path": cap,
                "universe": uni,
                "universe_source": uni_src,
                "capture_class": row["capture_class"],
            }
        )
    only = str(os.environ.get("ANCHOR_ED_DAYS") or "").strip()
    if only:
        want = {x.strip() for x in only.split(",") if x.strip()}
        jobs = [j for j in jobs if j["date"] in want]
        print(f"ANCHOR_ED_DAYS -> {len(jobs)}", flush=True)

    OUT.mkdir(parents=True, exist_ok=True)
    cache_dir = OUT / "day_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    force = str(os.environ.get("ANCHOR_ED_FORCE") or "").strip() in {"1", "true", "TRUE"}
    day_rows: list[dict[str, Any]] = []
    pending = []
    for job in jobs:
        cp = cache_dir / f"{job['date']}.json"
        if cp.is_file() and not force:
            got = json.loads(cp.read_text(encoding="utf-8"))
            got["A"] = load_a_day(job["date"])
            day_rows.append(got)
            print(f"cache {job['date']}", flush=True)
        else:
            pending.append(job)
    print(f"b2 jobs pending={len(pending)} cached={len(day_rows)}", flush=True)
    if pending:
        with ProcessPoolExecutor(max_workers=min(MAX_WORKERS, len(pending))) as ex:
            from research.anchor_vs_entry_ready_event_driven.run_b2 import replay_b2_day as _worker

            futs = {ex.submit(_worker, job): job["date"] for job in pending}
            for fut in as_completed(futs):
                day = futs[fut]
                try:
                    got = fut.result()
                except Exception as exc:
                    got = {
                        "date": day,
                        "B2": {"ok": False, "trades": [], "blocker": f"{type(exc).__name__}:{exc}"},
                    }
                got["A"] = load_a_day(str(got.get("date") or day))
                day_rows.append(got)
                b2 = got.get("B2") or {}
                print(
                    f"done {got.get('date')} class={got.get('capture_class')} "
                    f"B2_ok={b2.get('ok')} B2_tr={len(b2.get('trades') or [])} "
                    f"B2_sig={len(b2.get('signals') or [])} "
                    f"PM_tr={_sess_stats(b2.get('trades') or [])['PM_trades']} "
                    f"sec={got.get('elapsed_sec')}",
                    flush=True,
                )
    day_rows.sort(key=lambda r: str(r.get("date")))

    def bucket_side(side: str, full_only: bool) -> dict[str, Any]:
        use = [r for r in day_rows if (r.get("A") or {}).get("ok") and (r.get("B2") or {}).get("ok")]
        if full_only:
            use = [r for r in use if r.get("capture_class") == "FULL"]
        trades: list[dict[str, Any]] = []
        for r in use:
            trades.extend(list((r.get(side) or {}).get("trades") or []))
        st = _sess_stats(trades)
        st["days"] = [r["date"] for r in use]
        st["day_count"] = len(use)
        return st

    primary_a = bucket_side("A", True)
    primary_b = bucket_side("B2", True)
    ref_a = bucket_side("A", False)
    ref_b = bucket_side("B2", False)

    def delta(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
        pf_d = None
        if a.get("pf") is not None and b.get("pf") is not None:
            pf_d = float(b["pf"]) - float(a["pf"])
        return {
            "trades": int(b.get("trades") or 0) - int(a.get("trades") or 0),
            "pnl": round(float(b.get("pnl") or 0) - float(a.get("pnl") or 0), 2),
            "pf": pf_d,
            "maxdd": round(float(b.get("max_drawdown") or 0) - float(a.get("max_drawdown") or 0), 2),
        }

    primary = {"A": primary_a, "B2": primary_b, "DELTA": delta(primary_a, primary_b)}
    reference = {"A": ref_a, "B2": ref_b, "DELTA": delta(ref_a, ref_b)}
    timing_p = build_timing(day_rows, full_only=True)
    timing_r = build_timing(day_rows, full_only=False)
    verdict, dep = decide_verdict(primary, day_rows)
    ref_sign = (reference["DELTA"]["pnl"] or 0) * (primary["DELTA"]["pnl"] or 0)
    if primary["DELTA"]["pnl"] and reference["DELTA"]["pnl"] and ref_sign < 0:
        dep += " REFERENCE_SIGN_FLIP"

    pm_tr = int(primary_b.get("PM_trades") or 0)
    b1_note = (
        "B1 defect resolved: trading-window is inside ENTRY_READY so lunch forces FALSE; "
        f"PRIMARY PM trades={pm_tr} pm_signals={timing_p.get('pm_signals')}. "
        "V1R Primary has no entry_cooldown_sec (INFRA/YAML only); cooldown_release=0."
    )
    if pm_tr == 0:
        b1_note += " PM trades=0 is not auto-fail; check session_open PM signals vs fill/cap."

    report = {
        "result_class": "ANCHOR_VS_ENTRY_READY_EVENT_DRIVEN_B2",
        "formal_certification": False,
        "runtime_adoption": False,
        "identity": {
            "strategy_sha": EXPECTED_STRATEGY,
            "entry_sha": EXPECTED_ENTRY,
            "exit_sha": EXPECTED_EXIT,
        },
        "B2_definition": (
            "ENTRY_READY = live raw candidate (finite FEATURE_ORDER + finite score + bid>0) "
            "AND existing JPX continuous window (09:00-11:30 / 12:30-15:00) "
            "AND not same-symbol open/pending AND exposure < POSITION_CAP "
            "AND existing simulate_joint admission. Signal = ENTRY_READY FALSE->TRUE. "
            "No new hysteresis/threshold/feature. CLOCK_GRID is not a B2 poll."
        ),
        "new_features": False,
        "new_thresholds": False,
        "future_leak": False,
        "live_features": list(LIVE_FEATURES),
        "b1_preserved": str(B1_OUT),
        "b1_defect_check": b1_note,
        "capture_period": b1.get("capture_period"),
        "usable_days": b1.get("usable_days"),
        "full_days": b1.get("full_days"),
        "inventory": inventory,
        "PRIMARY_FULL": primary,
        "REFERENCE_ALL_USABLE": reference,
        "TIMING": {k: v for k, v in timing_p.items() if k != "signals"},
        "TIMING_REFERENCE": {k: v for k, v in timing_r.items() if k != "signals"},
        "day_dependency": dep,
        "verdict": verdict,
        "daily": [
            {
                "date": r.get("date"),
                "capture_class": r.get("capture_class"),
                "elapsed_sec": r.get("elapsed_sec"),
                "A": _sess_stats((r.get("A") or {}).get("trades") or []),
                "B2": _sess_stats((r.get("B2") or {}).get("trades") or []),
                "B2_ok": (r.get("B2") or {}).get("ok"),
                "causes": (r.get("B2") or {}).get("cause_counts"),
            }
            for r in day_rows
        ],
    }
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    (OUT / "report.md").write_text(write_md(report), encoding="utf-8")
    xlsx = write_excel(
        inventory=inventory,
        day_rows=day_rows,
        primary=primary,
        reference=reference,
        timing_p=timing_p,
        verdict=verdict,
    )
    print("--------------------------------", flush=True)
    print("PRIMARY FULL", flush=True)
    print(
        f"A trades={primary_a['trades']} pnl={primary_a['pnl']} pf={primary_a['pf']} "
        f"maxdd={primary_a['max_drawdown']} AM={primary_a['AM_trades']} PM={primary_a['PM_trades']}",
        flush=True,
    )
    print(
        f"B2 trades={primary_b['trades']} pnl={primary_b['pnl']} pf={primary_b['pf']} "
        f"maxdd={primary_b['max_drawdown']} AM={primary_b['AM_trades']} PM={primary_b['PM_trades']}",
        flush=True,
    )
    print(f"verdict={verdict}", flush=True)
    print(f"wrote {OUT / 'report.json'}", flush=True)
    print(f"wrote {xlsx}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
