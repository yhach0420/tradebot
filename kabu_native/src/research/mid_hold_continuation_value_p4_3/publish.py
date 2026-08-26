"""Write P4-3 precommit (already on disk) + report.json / report.md / audit.xlsx."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.mid_hold_continuation_value_p4_3 import (
    DOCUMENT_ID,
    EXIT_CHECKPOINTS_SEC,
    FALSE_RECOVERY_KNOWN_ID,
    P4_1_WINNER_IDS,
    P4_2_ADVERSE_N,
    PREDECLARED_TOP3,
    PRIMARY_CHECKPOINTS,
    REST11,
    SECONDARY_CHECKPOINTS,
    STATE_NON_RECOVERING,
    STATE_RECOVERING,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.mid_hold_continuation_value_p4_3.metrics import (
    adverse_ids,
    block_for,
    classify,
    cohort_counts,
    compact_cv,
    day_stability,
    economic_mass,
    evaluable,
    false_recovery,
    lodo,
    loss_saving,
    p41_winner_rows,
    slice_days,
    strat_600,
    strat_final,
    winner_cost_rows,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "mid_hold_continuation_value_p4_3"
P4_0 = NATIVE / "results" / "research" / "mid_hold_state_separability_p4_0" / "report.json"
P4_1 = NATIVE / "results" / "research" / "mid_hold_gate_p4_1" / "report.json"
P4_2 = NATIVE / "results" / "research" / "mid_hold_recovery_failure_path_p4_2" / "report.json"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    return hashlib.sha256(p.read_bytes()).hexdigest() if p.is_file() else ""


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(12, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _flat_block(blk: dict[str, Any], extra: Optional[dict[str, Any]] = None) -> dict[str, Any]:
    cv = blk.get("continuation_value") or {}
    ex = blk.get("checkpoint_exit_pnl") or {}
    cn = blk.get("canonical_final_pnl") or {}
    d = {k: v for k, v in blk.items() if k not in ("continuation_value", "checkpoint_exit_pnl", "canonical_final_pnl")}
    d.update(
        {
            "cv_mean": cv.get("mean"),
            "cv_median": cv.get("median"),
            "cv_p25": cv.get("p25"),
            "cv_p75": cv.get("p75"),
            "exit_mean": ex.get("mean"),
            "exit_median": ex.get("median"),
            "canon_mean": cn.get("mean"),
            "canon_median": cn.get("median"),
        }
    )
    if extra:
        d.update(extra)
    return d


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def load_frozen() -> dict[str, Any]:
    out: dict[str, Any] = {"ok": True, "reasons": []}
    if not P4_0.is_file():
        return {"ok": False, "reasons": ["NO_P4_0"]}
    p0 = json.loads(P4_0.read_text(encoding="utf-8"))
    if p0.get("MID_HOLD_GATEABILITY") != "MID_HOLD_STATE_SEPARABLE":
        out["ok"] = False
        out["reasons"].append("P4_0_NOT_SEPARABLE")
    out["P4_0"] = {"MID_HOLD_GATEABILITY": p0.get("MID_HOLD_GATEABILITY"), "verdict": p0.get("verdict")}
    if not P4_1.is_file():
        out["ok"] = False
        out["reasons"].append("NO_P4_1")
        return out
    p1 = json.loads(P4_1.read_text(encoding="utf-8"))
    if p1.get("STATUS") != "MID_HOLD_NO_PROGRESS_PORTFOLIO_HARM":
        out["ok"] = False
        out["reasons"].append("P4_1_NOT_CLOSED_HARM")
    out["P4_1"] = {
        "STATUS": p1.get("STATUS"),
        "PRECOMMIT_SHA": p1.get("PRECOMMIT_SHA"),
        "LOCAL_TRIGGER_n": (p1.get("LOCAL_TRIGGER") or {}).get("n"),
    }
    if not P4_2.is_file():
        out["ok"] = False
        out["reasons"].append("NO_P4_2")
        return out
    p2 = json.loads(P4_2.read_text(encoding="utf-8"))
    if p2.get("verdict") != "P4_2_RECOVERY_FAILURE_PATH_DECOMPOSITION_COMPLETE":
        out["ok"] = False
        out["reasons"].append("P4_2_NOT_COMPLETE")
    if p2.get("MECHANISM_CLASSIFICATION") != "RECOVERY_MOMENTUM_SEPARATES":
        out["ok"] = False
        out["reasons"].append("P4_2_MECHANISM_NOT_RECOVERY_MOMENTUM")
    adv = p2.get("PRIMARY_ADVERSE120") or {}
    if int(adv.get("n") or 0) != P4_2_ADVERSE_N:
        out["ok"] = False
        out["reasons"].append("P4_2_ADVERSE_N_DRIFT")
    out["P4_2"] = {
        "verdict": p2.get("verdict"),
        "MECHANISM_CLASSIFICATION": p2.get("MECHANISM_CLASSIFICATION"),
        "CANDIDATE_PATH_FAMILIES": p2.get("CANDIDATE_PATH_FAMILIES"),
        "PRIMARY_ADVERSE120": adv,
    }
    return out


def _cv_map(rows, state):
    return {str(h): compact_cv(block_for(rows, h=h, state=state)) for h in PRIMARY_CHECKPOINTS}


def build_report(
    *,
    precommit: dict[str, Any],
    recon: dict[str, Any],
    frozen: dict[str, Any],
    rows: list[dict[str, Any]],
    leak_n: int,
    identity_n: int,
    identity_fail: int,
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    rest = slice_days(rows, REST11) if rows else []
    top3 = slice_days(rows, PREDECLARED_TOP3) if rows else []
    coh = cohort_counts(rows) if rows else {"n": 0, "WIN": 0, "LOSS": 0, "DRAW": 0, "match_p4_2": False}
    rec_all = {str(h): block_for(rows, h=h, state=STATE_RECOVERING) for h in EXIT_CHECKPOINTS_SEC} if rows else {}
    nr_all = {str(h): block_for(rows, h=h, state=STATE_NON_RECOVERING) for h in EXIT_CHECKPOINTS_SEC} if rows else {}
    rec_rest = {str(h): block_for(rest, h=h, state=STATE_RECOVERING) for h in PRIMARY_CHECKPOINTS} if rest else {}
    nr_rest = {str(h): block_for(rest, h=h, state=STATE_NON_RECOVERING) for h in PRIMARY_CHECKPOINTS} if rest else {}

    by_cp = {}
    for h in PRIMARY_CHECKPOINTS + SECONDARY_CHECKPOINTS:
        rec = rec_all.get(str(h)) or {}
        nr = nr_all.get(str(h)) or {}
        by_cp[str(h)] = {"recovering_n": rec.get("n"), "nonrecovering_n": nr.get("n")}

    e600 = [
        strat_600(rows, h=h, state=st, extend=ext)
        for h in PRIMARY_CHECKPOINTS
        for st in (STATE_RECOVERING, STATE_NON_RECOVERING)
        for ext in (False, True)
    ] if rows else []
    fin = [
        strat_final(rows, h=h, state=st, outcome=o)
        for h in PRIMARY_CHECKPOINTS
        for st in (STATE_RECOVERING, STATE_NON_RECOVERING)
        for o in ("WIN", "LOSS")
    ] if rows else []

    wc = [x for h in PRIMARY_CHECKPOINTS for x in winner_cost_rows(rows, h=h)] if rows else []
    ls = [loss_saving(rows, h=h) for h in PRIMARY_CHECKPOINTS] if rows else []
    fr = [false_recovery(rows, h=h) for h in EXIT_CHECKPOINTS_SEC] if rows else []
    p41w = p41_winner_rows(rows) if rows else []
    dstab = [day_stability(rows, h=h) for h in PRIMARY_CHECKPOINTS] if rows else []
    lodo_rows = [lodo(rows, h=h) for h in PRIMARY_CHECKPOINTS] if rows else []
    masses = [economic_mass(rows, h=h) for h in PRIMARY_CHECKPOINTS] if rows else []

    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if not frozen.get("ok"):
        integrity.extend(frozen.get("reasons") or ["FROZEN"])
    if rows and leak_n:
        integrity.append(f"FUTURE_LEAK_{leak_n}")
    if rows and identity_fail:
        integrity.append(f"IDENTITY_FAIL_{identity_fail}/{identity_n}")
    if rows and not coh.get("match_p4_2"):
        integrity.append(f"COHORT_DRIFT n={coh.get('n')} WIN={coh.get('WIN')} LOSS={coh.get('LOSS')} DRAW={coh.get('DRAW')}")
    if rows:
        missing = [tid for tid in P4_1_WINNER_IDS if tid not in adverse_ids(rows)]
        if missing:
            integrity.append(f"P4_1_WINNER_MISSING:{','.join(missing)}")

    klass = classify(all_rows=rows, rest_rows=rest, integrity=integrity) if rows and not blocked else {
        "ACTIONABILITY_CLASSIFICATION": None,
        "why": blocked_reason,
    }

    if blocked or not recon.get("pass") or not frozen.get("ok"):
        verdict = VERDICT_BLOCKED
    elif integrity:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    known9984 = [
        {
            "trade_id": r.get("trade_id"),
            "checkpoint": r.get("horizon_sec"),
            "state": r.get("state"),
            "delta_bid_120_to_t": r.get("delta_bid_120_to_t"),
            "continuation_value": r.get("continuation_value_yen_100"),
            "checkpoint_exit_pnl": r.get("checkpoint_exit_pnl_yen_100"),
            "canonical_pnl": r.get("canonical_final_pnl_yen_100"),
            "checkpoint_to_600_bid_return": r.get("checkpoint_to_600_bid_return"),
        }
        for r in rows
        if str(r.get("trade_id")) == FALSE_RECOVERY_KNOWN_ID
    ]

    sheets = {
        "cohort": [
            {"slice": sl, **cohort_counts(rs)}
            for sl, rs in (("ALL", rows), ("REST11", rest), ("TOP3", top3))
        ],
        "state": [
            {
                "slice": sl,
                "horizon_sec": h,
                "recovering_n": (rec_all if sl == "ALL" else rec_rest).get(str(h), {}).get("n") if sl != "TOP3" else block_for(top3, h=h, state=STATE_RECOVERING).get("n"),
                "nonrecovering_n": (nr_all if sl == "ALL" else nr_rest).get(str(h), {}).get("n") if sl != "TOP3" else block_for(top3, h=h, state=STATE_NON_RECOVERING).get("n"),
            }
            for sl in ("ALL", "REST11", "TOP3")
            for h in (PRIMARY_CHECKPOINTS if sl != "ALL" else EXIT_CHECKPOINTS_SEC)
        ],
        "cv": [
            _flat_block(rec_all[str(h)], {"slice": "ALL"})
            for h in EXIT_CHECKPOINTS_SEC
            if str(h) in rec_all
        ]
        + [_flat_block(nr_all[str(h)], {"slice": "ALL"}) for h in EXIT_CHECKPOINTS_SEC if str(h) in nr_all]
        + [_flat_block(rec_rest[str(h)], {"slice": "REST11"}) for h in PRIMARY_CHECKPOINTS if str(h) in rec_rest]
        + [_flat_block(nr_rest[str(h)], {"slice": "REST11"}) for h in PRIMARY_CHECKPOINTS if str(h) in nr_rest],
        "recovering": [_flat_block(rec_all[str(h)]) for h in EXIT_CHECKPOINTS_SEC if str(h) in rec_all],
        "nonrec": [_flat_block(nr_all[str(h)]) for h in EXIT_CHECKPOINTS_SEC if str(h) in nr_all],
        "e600": [_flat_block(r) for r in e600],
        "wincost": wc,
        "loss": ls,
        "p41w": p41w,
        "false": fr + [{"known_9984": True, **r} for r in known9984],
        "rest11": [
            {"horizon_sec": h, "state": STATE_RECOVERING, **compact_cv(rec_rest.get(str(h)) or {})}
            for h in PRIMARY_CHECKPOINTS
        ]
        + [
            {"horizon_sec": h, "state": STATE_NON_RECOVERING, **compact_cv(nr_rest.get(str(h)) or {})}
            for h in PRIMARY_CHECKPOINTS
        ],
        "daystab": dstab,
        "lodo": [
            {
                "horizon_sec": r.get("horizon_sec"),
                "mean_median": (r.get("mean") or {}).get("median"),
                "mean_min": (r.get("mean") or {}).get("min"),
                "mean_max": (r.get("mean") or {}).get("max"),
                "median_median": (r.get("median") or {}).get("median"),
                "median_min": (r.get("median") or {}).get("min"),
                "median_max": (r.get("median") or {}).get("max"),
            }
            for r in lodo_rows
        ],
        "raw": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "horizon_sec": r.get("horizon_sec"),
                "state": r.get("state"),
                "delta_bid_120_to_t": r.get("delta_bid_120_to_t"),
                "evaluable": r.get("evaluable"),
                "trigger_time": r.get("trigger_time"),
                "execution_time": r.get("execution_time"),
                "execution_latency": r.get("execution_latency"),
                "execution_bid": r.get("execution_bid"),
                "checkpoint_exit_pnl_yen_100": r.get("checkpoint_exit_pnl_yen_100"),
                "canonical_final_pnl_yen_100": r.get("canonical_final_pnl_yen_100"),
                "continuation_value_yen_100": r.get("continuation_value_yen_100"),
                "checkpoint_to_600_bid_return": r.get("checkpoint_to_600_bid_return"),
                "checkpoint_to_canonical_exit_bid_return": r.get("checkpoint_to_canonical_exit_bid_return"),
                "FINAL_WIN": r.get("FINAL_WIN"),
                "FINAL_LOSS": r.get("FINAL_LOSS"),
                "EXIT_AT_600": r.get("EXIT_AT_600"),
                "EXTEND_TO_750": r.get("EXTEND_TO_750"),
                "TOP10": r.get("TOP10"),
                "TOP20": r.get("TOP20"),
                "uneval_reason": r.get("uneval_reason"),
            }
            for r in rows
            if r.get("cohort_B_adverse120")
        ],
    }
    unused = evaluable, fin, masses
    del unused

    report = {
        "task": "P4-3",
        "ANALYSIS_ID": "P4_3_MID_HOLD_CONTINUATION_VALUE",
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": [
            "OOS",
            "prospective",
            "robust",
            "strategy validation",
            "new EXIT validation",
            "production approval",
        ],
        "PRECOMMIT_SHA": precommit.get("SHA"),
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "FROZEN": frozen,
        "PRIMARY_COHORT": coh,
        "BY_CHECKPOINT": by_cp,
        "NON_RECOVERING_CONTINUATION_VALUE": _cv_map(rows, STATE_NON_RECOVERING) if rows else {},
        "RECOVERING_CONTINUATION_VALUE": _cv_map(rows, STATE_RECOVERING) if rows else {},
        "NON_RECOVERING_FULL": nr_all,
        "RECOVERING_FULL": rec_all,
        "EXIT600_VS_EXTEND": e600,
        "FINAL_OUTCOME_STRAT": fin,
        "REST11": {
            "cohort": cohort_counts(rest) if rest else {},
            "RECOVERING": {str(h): compact_cv(rec_rest.get(str(h)) or {}) for h in PRIMARY_CHECKPOINTS},
            "NON_RECOVERING": {str(h): compact_cv(nr_rest.get(str(h)) or {}) for h in PRIMARY_CHECKPOINTS},
        },
        "WINNER_COST": wc,
        "LOSS_SAVING": ls,
        "ECONOMIC_MASS": masses,
        "FALSE_RECOVERY": fr,
        "FALSE_RECOVERY_9984": known9984,
        "P4_1_WINNERS": p41w,
        "DAY_STABILITY": dstab,
        "LODO": lodo_rows,
        "ACTIONABILITY_CLASSIFICATION": klass.get("ACTIONABILITY_CLASSIFICATION"),
        "ACTIONABILITY_DETAIL": klass,
        "LOCAL_COUNTERFACTUAL_EXIT_VALUE_COMPUTED": True,
        "FULL_STATE_MACHINE_REPLAYED": False,
        "NEW_GATE_CREATED": False,
        "BEST_TIME_SELECTED": False,
        "HISTORICAL_PERCENT_CUTOFF_SEARCHED": False,
        "STRUCTURAL_ZERO_BOUNDARY_PRECOMMITTED": True,
        "NEW_EXIT_ADOPTED": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": bool(leak_n) if rows else False,
        "leak_n": int(leak_n),
        "identity_n": int(identity_n),
        "identity_fail": int(identity_fail),
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "P4_3_PRECOMMIT_SHA": precommit.get("SHA"),
            "P4_1_PRECOMMIT_SHA": (frozen.get("P4_1") or {}).get("PRECOMMIT_SHA"),
            "file_sha_p4_0_state": _file_sha("src/research/mid_hold_state_separability_p4_0/state.py"),
            "file_sha_p4_3_exec": _file_sha("src/research/mid_hold_continuation_value_p4_3/exec.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
        "_precommit": precommit,
    }
    return json_sanitize(report)


def _fmt_cv(blk: dict[str, Any]) -> str:
    return (
        f"n={blk.get('n')} mean={blk.get('mean')} median={blk.get('median')} "
        f"negative_rate={blk.get('negative_rate')}"
    )


def render_md(rep: dict[str, Any]) -> str:
    coh = rep.get("PRIMARY_COHORT") or {}
    by = rep.get("BY_CHECKPOINT") or {}
    nr = rep.get("NON_RECOVERING_CONTINUATION_VALUE") or {}
    rec = rep.get("RECOVERING_CONTINUATION_VALUE") or {}
    lines = [
        "# P4-3 Mid-Hold Recovery State Continuation Value",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / strategy validation / new EXIT validation / production approval",
        "P4-1 MID_HOLD_NO_PROGRESS_V1 remains CLOSED. No new Gate. No best checkpoint. No threshold search.",
        "",
        f"PRECOMMIT_SHA: `{rep.get('PRECOMMIT_SHA')}`",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        "",
        "PRIMARY_COHORT:",
        f"n: {coh.get('n')}",
        f"WIN: {coh.get('WIN')}",
        f"LOSS: {coh.get('LOSS')}",
        f"DRAW: {coh.get('DRAW')}",
        f"match_p4_2: {coh.get('match_p4_2')}",
        "",
        "BY_CHECKPOINT:",
    ]
    for h in list(PRIMARY_CHECKPOINTS) + list(SECONDARY_CHECKPOINTS):
        blk = by.get(str(h)) or {}
        lines.append(
            f"{h}: recovering_n={blk.get('recovering_n')} nonrecovering_n={blk.get('nonrecovering_n')}"
        )
    lines += ["", "NON_RECOVERING_CONTINUATION_VALUE:"]
    for h in PRIMARY_CHECKPOINTS:
        lines.append(f"{h}: {_fmt_cv(nr.get(str(h)) or {})}")
    lines += ["", "RECOVERING_CONTINUATION_VALUE:"]
    for h in PRIMARY_CHECKPOINTS:
        lines.append(f"{h}: {_fmt_cv(rec.get(str(h)) or {})}")
    lines += [
        "",
        f"EXIT600_VS_EXTEND: {json.dumps(rep.get('EXIT600_VS_EXTEND'), ensure_ascii=False, default=str)}",
        "",
        f"REST11: {json.dumps(rep.get('REST11'), ensure_ascii=False, default=str)}",
        "",
        f"WINNER_COST: {json.dumps(rep.get('WINNER_COST'), ensure_ascii=False, default=str)}",
        "",
        f"LOSS_SAVING: {json.dumps(rep.get('LOSS_SAVING'), ensure_ascii=False, default=str)}",
        "",
        f"FALSE_RECOVERY: {json.dumps(rep.get('FALSE_RECOVERY'), ensure_ascii=False, default=str)}",
        "",
        f"P4_1_WINNERS: {json.dumps(rep.get('P4_1_WINNERS'), ensure_ascii=False, default=str)}",
        "",
        f"DAY_STABILITY: {json.dumps(rep.get('DAY_STABILITY'), ensure_ascii=False, default=str)}",
        "",
        f"LODO: {json.dumps(rep.get('LODO'), ensure_ascii=False, default=str)}",
        "",
        f"ACTIONABILITY_CLASSIFICATION: `{rep.get('ACTIONABILITY_CLASSIFICATION')}`",
        f"{(rep.get('ACTIONABILITY_DETAIL') or {}).get('why')}",
        "",
        "LOCAL_COUNTERFACTUAL_EXIT_VALUE_COMPUTED: true",
        "FULL_STATE_MACHINE_REPLAYED: false",
        "NEW_GATE_CREATED: false",
        "BEST_TIME_SELECTED: false",
        "HISTORICAL_PERCENT_CUTOFF_SEARCHED: false",
        "STRUCTURAL_ZERO_BOUNDARY_PRECOMMITTED: true",
        "NEW_EXIT_ADOPTED: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    pre = rep.pop("_precommit", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("PRECOMMIT_SHA", public.get("PRECOMMIT_SHA")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("PRIMARY_COHORT_n", (public.get("PRIMARY_COHORT") or {}).get("n")),
            ("ACTIONABILITY_CLASSIFICATION", public.get("ACTIONABILITY_CLASSIFICATION")),
            ("LOCAL_COUNTERFACTUAL_EXIT_VALUE_COMPUTED", True),
            ("FULL_STATE_MACHINE_REPLAYED", False),
            ("NEW_GATE_CREATED", False),
            ("BEST_TIME_SELECTED", False),
            ("HISTORICAL_PERCENT_CUTOFF_SEARCHED", False),
            ("STRUCTURAL_ZERO_BOUNDARY_PRECOMMITTED", True),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("verdict", public.get("verdict")),
        ],
    )
    pre_rows = [{"key": k, "value": pre.get(k)} for k in ("LABEL", "SHA", "ANALYSIS_ID")]
    pre_rows.append({"key": "state_definitions", "value": pre.get("state_definitions")})
    pre_rows.append({"key": "verdict_criteria", "value": pre.get("verdict_criteria")})
    _write_rows(wb.create_sheet("Precommit"), pre_rows)
    _write_rows(wb.create_sheet("Cohort"), sheets.get("cohort") or [])
    _write_rows(wb.create_sheet("Checkpoint_State"), sheets.get("state") or [])
    _write_rows(wb.create_sheet("Continuation_Value"), sheets.get("cv") or [])
    _write_rows(wb.create_sheet("Recovering"), sheets.get("recovering") or [])
    _write_rows(wb.create_sheet("Non_Recovering"), sheets.get("nonrec") or [])
    _write_rows(wb.create_sheet("Exit600_Extend"), sheets.get("e600") or [])
    _write_rows(wb.create_sheet("Winner_Cost"), sheets.get("wincost") or [])
    _write_rows(wb.create_sheet("Loss_Saving"), sheets.get("loss") or [])
    _write_rows(wb.create_sheet("P4_1_Winners"), sheets.get("p41w") or [])
    _write_rows(wb.create_sheet("False_Recovery"), sheets.get("false") or [])
    _write_rows(wb.create_sheet("REST11"), sheets.get("rest11") or [])
    _write_rows(wb.create_sheet("Day_Stability"), sheets.get("daystab") or [])
    _write_rows(wb.create_sheet("LODO"), sheets.get("lodo") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(ident, list((public.get("Identity") or {}).items()) + [("identity_fail", public.get("identity_fail"))])
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("LOCAL_COUNTERFACTUAL_EXIT_VALUE_COMPUTED", True),
            ("FULL_STATE_MACHINE_REPLAYED", False),
            ("NEW_GATE_CREATED", False),
            ("BEST_TIME_SELECTED", False),
            ("HISTORICAL_PERCENT_CUTOFF_SEARCHED", False),
            ("STRUCTURAL_ZERO_BOUNDARY_PRECOMMITTED", True),
            ("NEW_EXIT_ADOPTED", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("P4_1_CANDIDATE_CLOSED", True),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
        ],
    )
    unused_raw = sheets.get("raw")
    del unused_raw
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
