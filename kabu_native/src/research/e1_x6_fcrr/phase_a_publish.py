"""Publish Phase A SE reachability audit artifacts (does not overwrite frozen v1 report)."""
from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from research.e1_x6_provisional.util import sha256_file

OUT_DIR = (
    Path(__file__).resolve().parents[3]
    / "results" / "research" / "e1_x6_fcrr_phase_a_se_audit"
)


def publish_phase_a(report: dict[str, Any], store: Path) -> dict[str, str]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    tmp = OUT_DIR / f"_tmp_{datetime.now().strftime('%H%M%S')}"
    tmp.mkdir(parents=True, exist_ok=True)

    fp_json = tmp / "report.json"
    fp_json.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    sha = sha256_file(fp_json)

    ans = report.get("answers") or {}
    md = [
        f"# E1_X6_FCRR Phase A — SELLING_EXHAUSTED Reachability Audit",
        "",
        f"- plan: `{report['plan_document_id']}` {report['plan_version']}",
        f"- spec: `{report['document_id']}` {report['document_version']}",
        f"- reference_run: `{report['reference_run_id']}` ({report['reference_status']})",
        f"- phase_a_run_id: `{report['phase_a_run_id']}`",
        f"- thresholds_changed: {report['thresholds_changed']}",
        f"- economics_opened: {report['economics_opened']}",
        f"- submit/cancel/live: 0/0/0",
        f"- report.json sha: `{sha}`",
        "",
        "## Mandatory Answers",
        f"1. PULLBACK_ACTIVE=25644 meaning: **{ans.get('q1_pullback_active_25644_meaning', {}).get('answer')}**",
        f"2. unique pullback episodes: **{ans.get('q2_unique_pullback_episodes')}**",
        f"3. standalone PASS: `{json.dumps(ans.get('q3_standalone_pass_counts'), ensure_ascii=False)}`",
        f"4. cumulative AND: `{json.dumps(ans.get('q4_cumulative_and_surviving'), ensure_ascii=False)}`",
        f"5. dominant reject: `{ans.get('q5_dominant_reject')}`",
        f"6. no_new_low_30s PASS events: **{ans.get('q6_no_new_low_30s_pass_events')}**",
        f"7. missing-feature events: **{ans.get('q7_missing_feature_events')}**",
        f"8. pullback-low updated events: **{ans.get('q8_pullback_low_updated_events')}**",
        f"9. state-hold evidence n: **{len(ans.get('q9_state_hold_evidence') or [])}**",
        f"10. transitions_n=0 reason: **{ans.get('q10_transitions_n_zero_reason', {}).get('answer')}**",
        f"11. CONTEXT→PULLBACK transitions: **{ans.get('q11_context_to_pullback_transitions')}**",
        f"12. 7/21 CONTEXT=0: see answers.q12",
        "",
        "## Artifacts",
        f"- se_event_audit: `{report['artifact_paths']['se_event_audit_jsonl']}`",
        f"- state_transitions: `{report['artifact_paths']['state_transitions_jsonl']}`",
        "",
        "STOP for Phase A — proceed to Phase B only after review.",
    ]
    (tmp / "report.md").write_text("\n".join(md), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    for row in (
        ("phase", report["phase"]),
        ("phase_a_run_id", report["phase_a_run_id"]),
        ("reference_run_id", report["reference_run_id"]),
        ("report_json_sha256", sha),
        ("audit_events_n", report["audit_events_n"]),
        ("transitions_n", report["transitions_n"]),
        ("se_pass_events", report["se_pass_events"]),
    ):
        ws.append(list(row))

    w = wb.create_sheet("Answers")
    w.append(["question", "value"])
    for k, v in ans.items():
        w.append([k, json.dumps(v, ensure_ascii=False, default=str)[:32000]])

    w = wb.create_sheet("DayFunnel")
    w.append(["day", "key", "count"])
    for day, ctr in (report.get("day_funnel") or {}).items():
        for k, v in ctr.items():
            w.append([day, k, v])

    w = wb.create_sheet("FCRR_StateTransitions")
    w.append(["from_state", "to_state", "event_time", "episode_id", "symbol", "day", "trigger", "mid", "ret_15s", "ret_30s"])
    # sample first 5000 lines from jsonl
    tp = Path(report["artifact_paths"]["state_transitions_jsonl"])
    if tp.is_file():
        with tp.open(encoding="utf-8") as f:
            for i, line in enumerate(f):
                if i >= 5000:
                    break
                d = json.loads(line)
                snap = d.get("feature_snapshot") or {}
                w.append([
                    d.get("from_state"), d.get("to_state"), d.get("event_time"),
                    d.get("episode_id"), d.get("symbol"), d.get("day"), d.get("trigger"),
                    snap.get("mid"), snap.get("ret_15s"), snap.get("ret_30s"),
                ])
        w.append(["NOTE", f"full_ledger={tp}", report["transitions_n"], "", "", "", "", "", "", ""])

    w = wb.create_sheet("SE_EventAudit_Sample")
    w.append([
        "day", "symbol", "episode_id", "event_time", "no_new_low_30s_pass",
        "ret_pass", "down_pass", "spread_pass", "dominant_reject", "final_pass",
    ])
    ap = Path(report["artifact_paths"]["se_event_audit_jsonl"])
    if ap.is_file():
        with ap.open(encoding="utf-8") as f:
            for i, line in enumerate(f):
                if i >= 5000:
                    break
                d = json.loads(line)
                w.append([
                    d.get("day"), d.get("symbol"), d.get("episode_id"), d.get("event_time"),
                    d.get("no_new_low_30s_pass"), d.get("ret_15_ge_ret_30_pass"),
                    d.get("down_tick_deceleration_pass"), d.get("spread_pass"),
                    d.get("dominant_reject"), d.get("selling_exhausted_final_pass"),
                ])

    w = wb.create_sheet("Safety")
    w.append(["submit", 0])
    w.append(["cancel", 0])
    w.append(["live", 0])
    w.append(["mainline_changed", False])
    w.append(["thresholds_changed", False])
    wb.save(tmp / "audit.xlsx")

    shas = {}
    for name in ("report.json", "report.md", "audit.xlsx"):
        dst = OUT_DIR / name
        os.replace(tmp / name, dst)
        shas[name] = sha256_file(dst)
    tmp.rmdir()
    # also copy pointer into store
    (store / "published_shas.json").write_text(json.dumps(shas, indent=2), encoding="utf-8")
    return shas
