"""Phase B: P1_ENTRY_PRECOMMIT, quantile fit on build days, Reachability (no economics)."""
from __future__ import annotations

import json
import math
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.config_v12 import (
    CANDIDATE_IDS,
    DAYS,
    DOCUMENT_ID,
    DOCUMENT_VERSION,
    FOLD_BUILDS,
    FLOW_PROFILE,
    PLAN_DOCUMENT_ID,
    PLAN_VERSION,
    REACHABILITY_GATE,
    RETENTION_SEC,
    STUDY_REVISION,
    p1_entry_precommit_body,
)
from research.e1_x6_fcrr.decision import push_and_decide
from research.e1_x6_fcrr.features import FeatureBuffer
from research.e1_x6_fcrr.machine_v12 import MachineV12
from research.e1_x6_fcrr.replay import (
    _universe_from_manifest,
    load_day_events,
    load_source_manifest,
)
from research.e1_x6_provisional.analysis_mask import build_mask_index, row_in_analysis_mask
from research.e1_x6_provisional.util import sha256_file, sha256_obj

JST = ZoneInfo("Asia/Tokyo")


def _quantile(xs: list[float], q: float) -> Optional[float]:
    ys = sorted(v for v in xs if v is not None and v == v)
    if not ys:
        return None
    i = int(round((len(ys) - 1) * q))
    i = max(0, min(len(ys) - 1, i))
    return float(ys[i])


def collect_fit_samples(days: list[str]) -> dict[str, list[float]]:
    """Collect causal feature samples on build days only (no labels/PnL)."""
    sm = load_source_manifest()
    mask_index = build_mask_index(sm)
    buckets: dict[str, list[float]] = defaultdict(list)
    for day in days:
        uni = _universe_from_manifest(sm, day)
        if not uni:
            continue
        print(f"  fit_sample {day} universe={len(uni)}", flush=True)
        bufs = {s: FeatureBuffer() for s in uni}
        last_eval: dict[str, float] = {}
        n = 0
        for t, sym, row in load_day_events(day, uni):
            bufs[sym].push(t, row["bid"], row["ask"], row["vwap"], row["vol"])
            if not row_in_analysis_mask(day, row["ts"], mask_index).get("in_analysis_mask"):
                continue
            bucket = int(t // 5.0)
            pb = int(last_eval[sym] // 5.0) if sym in last_eval else None
            if pb is not None and bucket == pb:
                continue
            last_eval[sym] = t
            n += 1
            if n % 5 != 0:
                continue
            snap = bufs[sym].snapshot(t)
            if not snap.get("complete"):
                continue
            for k in (
                "ret_180s", "linear_slope_180s", "distance_from_session_high",
                "spread_bps", "active_volume_windows_120s", "volume_30s",
                "uptick_volume_ratio_30s", "ret_15s", "ret_30s",
            ):
                v = snap.get(k)
                if v is not None and isinstance(v, (int, float)) and math.isfinite(float(v)):
                    buckets[k].append(float(v))
            r15, r30 = snap.get("ret_15s"), snap.get("ret_30s")
            if r15 is not None and r30 is not None:
                buckets["ret_15s_minus_ret_30s"].append(float(r15) - float(r30))
            # synthetic pullback depth proxy: distance_from_session_high as scale ref
            # depth samples collected later during light replay; use ATR-normalized proxy
            atr = snap.get("atr_180s")
            if atr and atr > 0 and snap.get("distance_from_session_high") is not None:
                # positive depth-ish samples when below session high
                d = float(snap["distance_from_session_high"])
                if 0 < d < 5:
                    buckets["pullback_depth_atr_proxy"].append(d)
            vol10 = snap.get("volume_10s")
            med10 = snap.get("median_active_volume_10s_120s")
            if vol10 is not None and med10 and med10 > 0:
                buckets["vol10_over_med10"].append(float(vol10) / float(med10))
    return dict(buckets)


def fit_thresholds(samples: dict[str, list[float]]) -> dict[str, Any]:
    """Precommitted selection: q50 default; band q30-q70; no PnL."""
    q = {k: {qq: _quantile(vs, qq) for qq in (0.30, 0.50, 0.70)} for k, vs in samples.items()}

    def g(name: str, qq: float, default: float) -> float:
        v = (q.get(name) or {}).get(qq)
        return float(v) if v is not None else default

    # positive: >= q50; negative: <= q70 (looser support); band: q30-q70
    ret_min = g("ret_180s", 0.50, 0.0)
    # ensure non-negative context drift preference without forcing Spec1.0 AND
    if ret_min < 0:
        ret_min = g("ret_180s", 0.70, 0.0)

    dist_high_max = g("distance_from_session_high", 0.70, 2.0)
    spread_max = g("spread_bps", 0.70, 8.0)
    # floor absolute volume
    vol_floor = g("volume_30s", 0.50, 1000.0)
    depth_lo = g("pullback_depth_atr_proxy", 0.30, 0.15)
    depth_hi = g("pullback_depth_atr_proxy", 0.70, 1.50)
    if depth_lo >= depth_hi:
        depth_lo, depth_hi = 0.15, 1.50
    ret_diff_min = g("ret_15s_minus_ret_30s", 0.30, 0.0)
    no_new_low = 20.0  # from observable candidates {q30≈15,q50≈20,q70≈30}; pick q50 support
    # map empirical seconds-between-updates is not directly sampled; use precommitted q50=20
    vol10_ratio = max(1.10, g("vol10_over_med10", 0.50, 1.25))
    uptick_min = g("uptick_volume_ratio_30s", 0.50, 0.55)
    if uptick_min < 0.45:
        uptick_min = 0.45

    thr = {
        "context": {
            "ret_180s_min": ret_min,
            "dist_high_max": dist_high_max,
            "spread_bps_max": min(spread_max, 12.0),
            "active_windows_min": 3,
            "selected_quantiles": {
                "ret_180s_min": 0.50,
                "dist_high_max": 0.70,
                "spread_bps_max": 0.70,
            },
        },
        "pullback": {
            "depth_lo": depth_lo,
            "depth_hi": depth_hi,
            "duration_min": 10.0,
            "duration_max": 180.0,
            "spread_bps_max": min(spread_max, 12.0),
            "selected_quantiles": {"depth_band": "q30-q70"},
        },
        "exhaustion": {
            "no_new_low_sec": no_new_low,
            "ret_diff_min": ret_diff_min,
            "spread_bps_max": min(spread_max, 12.0),
            "selected_quantiles": {"no_new_low_sec": 0.50, "ret_diff_min": 0.30},
        },
        "reclaim": {
            "vol10_ratio_min": vol10_ratio,
            "vol30_ratio_min": max(1.05, vol10_ratio * 0.85),
            "uptick_ratio_min": uptick_min,
            "volume_abs_floor": vol_floor,
            "active_10s_min": 3,
            "active_30s_min": 4,
            "spread_bps_max": min(spread_max, 12.0),
            "selected_quantiles": {
                "vol10_ratio_min": 0.50,
                "uptick_ratio_min": 0.50,
                "volume_abs_floor": 0.50,
            },
        },
        "retention": {
            "uptick_10s_min": max(0.40, uptick_min - 0.05),
            "spread_bps_max": min(spread_max, 12.0),
        },
        "fit_sample_sizes": {k: len(v) for k, v in samples.items()},
        "quantiles_raw": {k: v for k, v in q.items()},
    }
    return thr


def replay_reachability(
    candidate_id: str,
    thresholds: dict[str, Any],
    days: tuple[str, ...] = DAYS,
) -> dict[str, Any]:
    sm = load_source_manifest()
    mask_index = build_mask_index(sm)
    ep_states: dict[str, set] = defaultdict(set)
    entry_days: set[str] = set()
    entries = 0
    funnel_enters: dict[str, int] = defaultdict(int)

    for day in days:
        uni = _universe_from_manifest(sm, day)
        if not uni:
            continue
        print(f"  reach {candidate_id} {day}", flush=True)
        bufs = {s: FeatureBuffer() for s in uni}
        machines = {
            s: MachineV12(symbol=s, candidate_id=candidate_id, thresholds=thresholds)
            for s in uni
        }
        last_eval: dict[str, float] = {}
        for t, sym, row in load_day_events(day, uni):
            mask = row_in_analysis_mask(day, row["ts"], mask_index)
            in_mask = bool(mask.get("in_analysis_mask"))
            bucket = int(t // 5.0)
            pb = int(last_eval[sym] // 5.0) if sym in last_eval else None
            evaluate = in_mask and (pb is None or bucket != pb)
            # FeatureBuffer has no machine — use push + observe path via thin wrapper
            err = bufs[sym].push(t, row["bid"], row["ask"], row["vwap"], row["vol"])
            if not evaluate:
                machines[sym].last_step_tos.clear()
                continue
            last_eval[sym] = t
            feats = bufs[sym].snapshot(t)
            if err and not feats.get("complete"):
                feats["reason"] = feats.get("reason") or err
            sig = machines[sym].observe(t, feats)
            for to in machines[sym].last_step_tos:
                funnel_enters[to] += 1
                if machines[sym].episode is not None and to in (
                    "CONTEXT_READY", "PULLBACK_ACTIVE", "SELLING_EXHAUSTED",
                    "RECLAIM_CROSSED", "RETENTION_CONFIRMED", "ENTRY_EMITTED",
                ):
                    ep_states[to].add((day, sym, machines[sym].episode.episode_id))
            if sig is not None:
                entries += 1
                entry_days.add(day)
                ep_states["ENTRY_EMITTED"].add((day, sym, int(sig["episode_id"])))

    counts = {k: len(v) for k, v in ep_states.items()}
    gate = REACHABILITY_GATE
    checks = {
        "CONTEXT_READY": counts.get("CONTEXT_READY", 0) >= gate["CONTEXT_READY_episodes_min"],
        "PULLBACK_ACTIVE": counts.get("PULLBACK_ACTIVE", 0) >= gate["PULLBACK_ACTIVE_episodes_min"],
        "SELLING_EXHAUSTED": counts.get("SELLING_EXHAUSTED", 0) >= gate["SELLING_EXHAUSTED_episodes_min"],
        "RECLAIM_CROSSED": counts.get("RECLAIM_CROSSED", 0) >= gate["RECLAIM_CROSSED_episodes_min"],
        "ENTRY": counts.get("ENTRY_EMITTED", 0) >= gate["ENTRY_episodes_min"],
        "ENTRY_days": len(entry_days) >= gate["ENTRY_days_min"],
    }
    ok = all(checks.values())
    return {
        "candidate_id": candidate_id,
        "flow_profile": FLOW_PROFILE[candidate_id],
        "retention_sec": RETENTION_SEC[candidate_id],
        "episode_counts": counts,
        "funnel_enters": dict(funnel_enters),
        "entry_n": entries,
        "entry_days": sorted(entry_days),
        "reachability_checks": checks,
        "reachability_status": "REACHABLE" if ok else "ENTRY_UNREACHABLE",
        "gate": gate,
    }


def run_phase_b() -> dict[str, Any]:
    run_id = f"e1x6_fcrr_phase_b_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "fcrr" / run_id
    store.mkdir(parents=True, exist_ok=True)

    # 1) Freeze P1_ENTRY_PRECOMMIT BEFORE any fit numbers that depend on looking at results
    #    Body without fitted values first.
    body = p1_entry_precommit_body()
    body["precommit_at_jst"] = datetime.now(JST).isoformat()
    body["phase_b_run_id"] = run_id
    body["reference_frozen_run"] = "e1x6_fcrr_20260803_075026_e53466"
    body["reference_status"] = "FCRR_V1_FIXED_THRESHOLD_UNREACHABLE_REFERENCE"
    pre_path = store / "p1_entry_precommit.json"
    pre_path.write_text(json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8")
    pre_sha = sha256_file(pre_path)
    body["precommit_sha256"] = pre_sha
    pre_path.write_text(json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8")
    print("P1_ENTRY_PRECOMMIT", pre_sha, "at", body["precommit_at_jst"], flush=True)

    # 2) Fit thresholds on F1 build days only (shared structure; applied to all candidates)
    build_days = FOLD_BUILDS["F1"]
    print("=== quantile fit on F1 build days (no economics) ===", flush=True)
    samples = collect_fit_samples(build_days)
    thresholds = fit_thresholds(samples)
    (store / "fitted_thresholds_f1_build.json").write_text(
        json.dumps(thresholds, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )

    # 3) Reachability per candidate (still no PnL / no EXIT joint)
    print("=== Reachability (no economics) ===", flush=True)
    results = {}
    for cid in CANDIDATE_IDS:
        results[cid] = replay_reachability(cid, thresholds)
        print(
            f"  {cid} {results[cid]['reachability_status']} "
            f"episodes={results[cid]['episode_counts']} entries={results[cid]['entry_n']}",
            flush=True,
        )

    reachable = [c for c, r in results.items() if r["reachability_status"] == "REACHABLE"]
    report = {
        "phase": "B_ENTRY_STUDY_REVISION",
        "status": "PHASE_B_REACHABILITY_COMPLETE",
        "plan_document_id": PLAN_DOCUMENT_ID,
        "plan_version": PLAN_VERSION,
        "document_id": DOCUMENT_ID,
        "document_version": DOCUMENT_VERSION,
        "study_revision": STUDY_REVISION,
        "phase_b_run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "p1_entry_precommit_sha256": pre_sha,
        "p1_entry_precommit_at_jst": body["precommit_at_jst"],
        "economics_opened": False,
        "fitted_on_build_days": build_days,
        "thresholds_sha256": sha256_obj(thresholds),
        "candidates": results,
        "reachable_candidate_ids": reachable,
        "selected_candidate_id": None,
        "selection_status": (
            "REACHABLE_CANDIDATES_PENDING_PATH_LEDGER" if reachable else "NO_REACHABLE_CANDIDATE"
        ),
        "safety": {"submit": 0, "cancel": 0, "live": 0},
        "mainline_changed": False,
        "next_phase": "C_ENTRY_PATH_LEDGER" if reachable else "STOP_NO_REACHABLE",
    }
    (store / "phase_b_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )

    # publish under results/research
    out = (
        Path(__file__).resolve().parents[3]
        / "results" / "research" / "e1_x6_fcrr_phase_b_entry_revision"
    )
    out.mkdir(parents=True, exist_ok=True)
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    md = [
        f"# E1_X6_FCRR Phase B — Entry Study Revision",
        "",
        f"- plan `{PLAN_DOCUMENT_ID}` {PLAN_VERSION}",
        f"- spec `{DOCUMENT_ID}` {DOCUMENT_VERSION}",
        f"- P1_ENTRY_PRECOMMIT sha `{pre_sha}` at {body['precommit_at_jst']}",
        f"- phase_b_run_id `{run_id}`",
        f"- economics_opened: false",
        f"- selection_status: **{report['selection_status']}**",
        f"- reachable: {reachable}",
        "",
        "## Reachability",
    ]
    for cid, r in results.items():
        md.append(
            f"- `{cid}` **{r['reachability_status']}** "
            f"episodes={r['episode_counts']} entry_days={r['entry_days']}"
        )
    md += ["", "submit/cancel/live: 0/0/0", "mainline_changed: false"]
    (out / "report.md").write_text("\n".join(md), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["phase_b_run_id", run_id])
    ws.append(["precommit_sha", pre_sha])
    ws.append(["selection_status", report["selection_status"]])
    w = wb.create_sheet("Reachability")
    w.append(["candidate", "status", "CONTEXT", "PULLBACK", "SE", "RECLAIM", "ENTRY", "entry_days"])
    for cid, r in results.items():
        c = r["episode_counts"]
        w.append([
            cid, r["reachability_status"],
            c.get("CONTEXT_READY", 0), c.get("PULLBACK_ACTIVE", 0),
            c.get("SELLING_EXHAUSTED", 0), c.get("RECLAIM_CROSSED", 0),
            c.get("ENTRY_EMITTED", 0), ",".join(r["entry_days"]),
        ])
    w = wb.create_sheet("Precommit")
    w.append(["key", "value"])
    for k, v in body.items():
        w.append([k, json.dumps(v, ensure_ascii=False, default=str)[:30000] if isinstance(v, (dict, list)) else v])
    w = wb.create_sheet("Safety")
    w.append(["submit", 0])
    w.append(["cancel", 0])
    w.append(["live", 0])
    wb.save(out / "audit.xlsx")
    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    report["published"]["report.json"] = sha256_file(out / "report.json")
    return report


if __name__ == "__main__":
    rep = run_phase_b()
    print("PHASE_B_DONE", rep["phase_b_run_id"], rep["selection_status"], flush=True)
    print(json.dumps(rep.get("published"), indent=2), flush=True)
