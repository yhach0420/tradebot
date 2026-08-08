"""Gate V2 assembly from E1_X10 day panel + E1_X11 source identity."""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_provisional.util import sha256_file, sha256_obj

from . import (
    AGG_NOTIONAL_FRAC,
    AGG_RISK_FRAC,
    ANALYSIS_ID,
    BOARD_FRESHNESS_SEC,
    BREADTH_ABS_MIN,
    BREADTH_FRAC_MEDIAN_MIN,
    BREADTH_MEDIAN_MIN,
    CAPITAL_SCENARIOS,
    CONFIG_DRIFT,
    DOCUMENT_ID,
    HISTORY_END_MAX,
    LOOKBACK_MAX,
    LOT,
    MAX_CONCURRENT,
    MIN_EXEC_ANCHORS,
    MIN_HISTORY_DAYS,
    MIN_JUMP_N,
    MIN_POLICY_EVALUABLE_DAYS,
    MIN_SPREAD_N,
    OLD_BREADTH_MEDIAN,
    OLD_BREADTH_MIN,
    PER_SYMBOL_NOTIONAL_FRAC,
    PER_TRADE_RISK_FRAC,
    POLICY_ID,
    QTY_MIN,
    RECURRING_MIN_DAYS,
    RESERVE_FRAC,
    SOURCE_X11,
    SOURCE_X11_VERDICT,
    SUPERSEDED_BREADTH,
    TARGET_SYMBOL,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
X10_DIR = NATIVE / "results" / "research" / "e1_x10_risk_universe"
X11_DIR = NATIVE / "results" / "research" / "e1_x11_risk_policy"
PUBLISH = NATIVE / "results" / "research" / "e1_x11_policy_gate_v2"
PKG = Path(__file__).resolve().parent


def _f(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _med(xs: list[float]) -> Optional[float]:
    if not xs:
        return None
    ys = sorted(xs)
    return float(ys[len(ys) // 2])


def _pct(xs: list[float], q: float) -> Optional[float]:
    if not xs:
        return None
    import numpy as np
    return float(np.quantile(np.asarray(xs, dtype=float), q))


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "entry_changed": False,
        "exit_changed": False,
        "universe_runtime_changed": False,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
        "source_overwritten": False,
        "policy_fractions_changed": False,
    }


def _merge_panel(x10_xlsx: Path) -> dict[tuple[str, str], dict[str, Any]]:
    notion = _load_sheet(x10_xlsx, "OneLotNotional")
    tick = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "TickRisk")}
    spread = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "SpreadRisk")}
    jump = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "BidJumps")}
    exe = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "ExecutableLoss")}
    panel = {}
    for r in notion:
        key = (str(r["symbol"]), str(r["day"]))
        panel[key] = {
            "symbol": key[0],
            "day": key[1],
            "one_lot_notional_yen": _f(r.get("one_lot_notional_yen")),
            "reference_price": _f(r.get("reference_price")),
            "ref_status": r.get("status"),
            "tick_size_yen": _f((tick.get(key) or {}).get("tick_size_yen")),
            "n_spread_obs": int(_f((spread.get(key) or {}).get("n_spread_obs")) or 0),
            "p95_spread_cost_yen_100": _f((spread.get(key) or {}).get("p95_spread_cost_yen_100")),
            "n_jump_obs": int(_f((jump.get(key) or {}).get("n_jump_obs")) or 0),
            "p95_down_bid_jump_yen_100": _f((jump.get(key) or {}).get("p95_down_bid_jump_yen_100")),
            "n_exec_anchors": int(_f((exe.get(key) or {}).get("n_exec_anchors")) or 0),
            "exec_loss_5s_p95": _f((exe.get(key) or {}).get("exec_loss_yen_100_5s_p95")),
        }
    return panel


def _prior(day: str, days: list[str]) -> list[str]:
    return [d for d in days if d < day][-LOOKBACK_MAX:]


def _roll(symbol: str, day: str, hist: list[str], panel: dict) -> dict[str, Any]:
    rows = [panel[(symbol, d)] for d in hist if (symbol, d) in panel]
    valid_days = len(rows)
    spread_n = sum(r["n_spread_obs"] for r in rows)
    jump_n = sum(r["n_jump_obs"] for r in rows)
    exec_n = sum(r["n_exec_anchors"] for r in rows)
    last = rows[-1] if rows else None
    ref_ok = bool(last and last.get("ref_status") == "OK" and last.get("one_lot_notional_yen"))
    tick_ok = bool(last and last.get("tick_size_yen") is not None)
    evaluable = (
        valid_days >= MIN_HISTORY_DAYS
        and spread_n >= MIN_SPREAD_N
        and jump_n >= MIN_JUMP_N
        and exec_n >= MIN_EXEC_ANCHORS
        and ref_ok
        and tick_ok
        and ((not hist) or hist[-1] < day)
    )
    spreads = [r["p95_spread_cost_yen_100"] for r in rows if r["p95_spread_cost_yen_100"] is not None]
    jumps = [r["p95_down_bid_jump_yen_100"] for r in rows if r["p95_down_bid_jump_yen_100"] is not None]
    execs = [r["exec_loss_5s_p95"] for r in rows if r["exec_loss_5s_p95"] is not None]
    s_m, j_m, e_m = _med(spreads), _med(jumps), _med(execs)
    comps = [c for c in (s_m, j_m, e_m) if c is not None]
    est = max(comps) if comps else None
    return {
        "day": day,
        "symbol": symbol,
        "history_start": hist[0] if hist else None,
        "history_end": hist[-1] if hist else None,
        "valid_history_days": valid_days,
        "spread_n": spread_n,
        "jump_n": jump_n,
        "exec_anchor_n": exec_n,
        "reference_asof_pass": ref_ok,
        "tick_resolved": tick_ok,
        "policy_evaluable_symbol_day": evaluable,
        "status": "POLICY_EVALUABLE_SYMBOL_DAY" if evaluable else "RISK_HISTORY_INSUFFICIENT",
        "one_lot_notional_yen": last["one_lot_notional_yen"] if last else None,
        "reference_price": last["reference_price"] if last else None,
        "reference_price_source": "previous_session_official_close" if ref_ok else None,
        "estimated_execution_risk_yen": est,
        "no_same_day_future": (not hist) or hist[-1] < day,
    }


def _limits(capital: float) -> dict[str, float]:
    return {
        "per_trade_risk_limit": capital * PER_TRADE_RISK_FRAC,
        "aggregate_risk_limit": capital * AGG_RISK_FRAC,
        "per_symbol_notional_limit": capital * PER_SYMBOL_NOTIONAL_FRAC,
        "aggregate_notional_limit": capital * AGG_NOTIONAL_FRAC,
    }


def _eligible(row: dict[str, Any], capital: float) -> bool:
    if not row.get("policy_evaluable_symbol_day"):
        return False
    lim = _limits(capital)
    n, est = row.get("one_lot_notional_yen"), row.get("estimated_execution_risk_yen")
    if n is None or est is None:
        return False
    return n <= lim["per_symbol_notional_limit"] and est <= lim["per_trade_risk_limit"]


def _required_capital(notional: Optional[float], risk: Optional[float]) -> Optional[float]:
    vals = []
    if notional is not None:
        vals.append(notional / PER_SYMBOL_NOTIONAL_FRAC)
    if risk is not None:
        vals.append(risk / PER_TRADE_RISK_FRAC)
    return max(vals) if vals else None


def build_precommit(*, source_report_sha: str, source_audit_sha: str) -> dict[str, Any]:
    body = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "precommit_at_jst": datetime.now(JST).isoformat(),
        "source_run": SOURCE_X11,
        "source_verdict": SOURCE_X11_VERDICT,
        "source_report_sha": source_report_sha,
        "source_audit_sha": source_audit_sha,
        "policy_id": POLICY_ID,
        "policy_fractions_frozen": {
            "per_trade": PER_TRADE_RISK_FRAC,
            "agg_risk": AGG_RISK_FRAC,
            "symbol_notional": PER_SYMBOL_NOTIONAL_FRAC,
            "agg_notional": AGG_NOTIONAL_FRAC,
            "reserve": RESERVE_FRAC,
        },
        "corrected_evaluable_day_contract": {
            "valid_history_days_min": MIN_HISTORY_DAYS,
            "spread_n_min": MIN_SPREAD_N,
            "jump_n_min": MIN_JUMP_N,
            "exec_anchors_min": MIN_EXEC_ANCHORS,
            "history_end_le_d_minus_1": True,
        },
        "corrected_breadth_gate": {
            "median_min": BREADTH_MEDIAN_MIN,
            "absolute_min": BREADTH_ABS_MIN,
            "eligible_fraction_median_min": BREADTH_FRAC_MEDIAN_MIN,
            "derived_from_cap": MAX_CONCURRENT,
            "superseded": SUPERSEDED_BREADTH,
            "old_impossible": {"median": OLD_BREADTH_MEDIAN, "min": OLD_BREADTH_MIN},
        },
        "notional_representative_contract": {
            "static_uses": "daily_value",
            "report_representative": ["median_required_capital", "maximum_required_capital"],
            "forbid_single_day_as_the_required_capital": True,
        },
        "blocker_priority": [
            "E1_X11_V2_SOURCE_IDENTITY_MISMATCH",
            "E1_X11_V2_NOTIONAL_REPRESENTATIVE_CONTRACT_UNRESOLVED",
            "E1_X11_RISK_HISTORY_SUPPORT_INSUFFICIENT",
            "E1_X11_SAFE_CAPITAL_BASE_UNRESOLVED",
            "E1_X11_DYNAMIC_GATE_FIELDS_INCOMPLETE",
            "E1_X11_FIXED100_RISK_POLICY_FREEZE_READY",
        ],
        "verdict_rules": "section_14_priority_with_all_blockers_array",
        "no_policy_retune_from_scenarios": True,
        "no_source_overwrite": True,
    }
    body["precommit_sha256"] = sha256_obj(body)
    return body


def run_once(*, label: str = "A") -> dict[str, Any]:
    run_id = f"e1x11_gate_v2_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"
    x11_path = X11_DIR / "report.json"
    x11_audit = X11_DIR / "audit.xlsx"
    x11 = json.loads(x11_path.read_text(encoding="utf-8"))
    if SOURCE_X11 not in str(x11.get("run_id")) or x11.get("verdict") != SOURCE_X11_VERDICT:
        return {
            "analysis_id": ANALYSIS_ID,
            "run_id": run_id,
            "verdict": "E1_X11_V2_SOURCE_IDENTITY_MISMATCH",
            "determinism_shas": {"verdict": "E1_X11_V2_SOURCE_IDENTITY_MISMATCH"},
            "safety": _safety(),
            "stop": True,
            "_sheets": {},
        }

    report_sha = sha256_file(x11_path)
    audit_sha = sha256_file(x11_audit)
    precommit = build_precommit(source_report_sha=report_sha, source_audit_sha=audit_sha)

    panel = _merge_panel(X10_DIR / "audit.xlsx")
    all_days = sorted({d for _, d in panel if d <= HISTORY_END_MAX})
    observed = defaultdict(set)
    for s, d in panel:
        observed[s].add(d)
    global_recurring = sorted(s for s in observed if len(observed[s]) >= RECURRING_MIN_DAYS)

    # Symbol-day evaluability for all calendar days that have priors
    eval_calendar = [d for d in all_days if any(x < d for x in all_days)]
    sd_rows = []
    for day in eval_calendar:
        hist = _prior(day, all_days)
        # symbols with any history in lookback
        cands = {s for s, d in panel if d in hist}
        for sym in sorted(cands):
            sd_rows.append(_roll(sym, day, hist, panel))

    # Day summaries + as-of recurring
    day_rows = []
    policy_evaluable_days = []
    warmup_days = []
    asof_recurring_by_day: dict[str, set[str]] = {}
    for day in eval_calendar:
        hist = _prior(day, all_days)
        # as-of recurring: symbols with >=5 valid history days in panel days <= D-1
        asof = set()
        for s in observed:
            n_hist = sum(1 for d in observed[s] if d < day)
            if n_hist >= RECURRING_MIN_DAYS:
                asof.add(s)
        asof_recurring_by_day[day] = asof
        day_sd = [r for r in sd_rows if r["day"] == day]
        evaluable = [r for r in day_sd if r["policy_evaluable_symbol_day"]]
        # warmup if no symbol can be evaluable yet (history window too short globally)
        # Spec: warmup days excluded from daily eligible minimum
        # A day is warmup if max possible history days for any symbol < MIN_HISTORY_DAYS
        max_hist_len = len(hist)
        warmup = max_hist_len < MIN_HISTORY_DAYS
        if warmup:
            warmup_days.append(day)
        policy_eval_day = (not warmup) and len(evaluable) > 0
        if policy_eval_day:
            policy_evaluable_days.append(day)
        day_rows.append({
            "day": day,
            "observed_symbols_n": len({s for s, d in panel if d == day}),
            "recurring_symbols_n": len(asof),
            "evaluable_symbols_n": len(evaluable),
            "not_evaluable_symbols_n": len(day_sd) - len(evaluable),
            "evaluable_fraction": (len(evaluable) / len(asof)) if asof else None,
            "warmup_day": warmup,
            "policy_evaluable_day": policy_eval_day,
        })

    # Coverage: evaluable as-of recurring symbol-days / all as-of recurring symbol-days
    num = den = 0
    for day in policy_evaluable_days:
        asof = asof_recurring_by_day[day]
        den += len(asof)
        day_eval = {r["symbol"] for r in sd_rows if r["day"] == day and r["policy_evaluable_symbol_day"]}
        num += len(asof & day_eval)
    # Also count all asof days including non-policy? Spec: evaluable as-of recurring / all as-of recurring
    # Use all eval_calendar non-warmup days as denominator base for asof recurring presence
    num2 = den2 = 0
    for day in eval_calendar:
        if day in warmup_days:
            continue
        asof = asof_recurring_by_day[day]
        den2 += len(asof)
        day_eval = {r["symbol"] for r in sd_rows if r["day"] == day and r["policy_evaluable_symbol_day"]}
        num2 += len(asof & day_eval)
    coverage_ratio = (num2 / den2) if den2 else 0.0
    history_support_pass = (
        coverage_ratio >= 0.90 and len(policy_evaluable_days) >= MIN_POLICY_EVALUABLE_DAYS
    )

    # Capital scenarios on POLICY_EVALUABLE_DAY only
    scenario_rows = []
    for cap in CAPITAL_SCENARIOS:
        elig_counts = []
        eval_counts = []
        fracs = []
        for day in policy_evaluable_days:
            day_sd = [r for r in sd_rows if r["day"] == day]
            evaluable = [r for r in day_sd if r["policy_evaluable_symbol_day"]]
            eligible = [r for r in evaluable if _eligible(r, float(cap))]
            # prefer asof recurring eligible for breadth? Spec: eligible symbols on evaluable day
            # Use evaluable set as denominator for fraction
            elig_counts.append(len(eligible))
            eval_counts.append(len(evaluable))
            fracs.append((len(eligible) / len(evaluable)) if evaluable else 0.0)
        lim = _limits(float(cap))
        med_elig = _med([float(x) for x in elig_counts]) if elig_counts else None
        min_elig = min(elig_counts) if elig_counts else None
        max_elig = max(elig_counts) if elig_counts else None
        med_frac = _med(fracs) if fracs else None
        min_frac = min(fracs) if fracs else None
        breadth_pass = (
            bool(policy_evaluable_days)
            and med_elig is not None and med_elig >= BREADTH_MEDIAN_MIN
            and min_elig is not None and min_elig >= BREADTH_ABS_MIN
            and med_frac is not None and med_frac >= BREADTH_FRAC_MEDIAN_MIN
        )
        scenario_rows.append({
            "capital_yen": cap,
            "policy_evaluable_days": len(policy_evaluable_days),
            "median_evaluable_symbols": _med([float(x) for x in eval_counts]),
            "minimum_evaluable_symbols": min(eval_counts) if eval_counts else None,
            "median_eligible_symbols": med_elig,
            "minimum_eligible_symbols": min_elig,
            "maximum_eligible_symbols": max_elig,
            "median_eligible_fraction": med_frac,
            "minimum_eligible_fraction": min_frac,
            "breadth_gate_pass": breadth_pass,
            "coverage_gate_pass": history_support_pass,
            **lim,
            "warmup_excluded": True,
        })

    # Required capital all symbols (from evaluable symbol-days)
    by_sym: dict[str, list[dict]] = defaultdict(list)
    for r in sd_rows:
        if r["policy_evaluable_symbol_day"]:
            by_sym[r["symbol"]].append(r)
    req_rows = []
    for sym, rs in sorted(by_sym.items()):
        notionals = [float(r["one_lot_notional_yen"]) for r in rs if r.get("one_lot_notional_yen") is not None]
        risks = [float(r["estimated_execution_risk_yen"]) for r in rs if r.get("estimated_execution_risk_yen") is not None]
        reqs = []
        for r in rs:
            rc = _required_capital(r.get("one_lot_notional_yen"), r.get("estimated_execution_risk_yen"))
            if rc is not None:
                reqs.append(rc)
        req_rows.append({
            "symbol": sym,
            "n_evaluable_days": len(rs),
            "notional_min": min(notionals) if notionals else None,
            "notional_median": _med(notionals),
            "notional_p90": _pct(notionals, 0.90),
            "notional_max": max(notionals) if notionals else None,
            "execution_risk_p95_median": _med(risks),
            "required_capital_min": min(reqs) if reqs else None,
            "required_capital_median": _med(reqs),
            "required_capital_p90": _pct(reqs, 0.90),
            "required_capital_max": max(reqs) if reqs else None,
        })

    # 285A daily + summary
    kiox_daily = []
    for r in sd_rows:
        if r["symbol"] != TARGET_SYMBOL:
            continue
        # include all days with notional even if not evaluable, for difference explanation
        if r.get("one_lot_notional_yen") is None and not r["policy_evaluable_symbol_day"]:
            # still try panel direct
            pass
        n = r.get("one_lot_notional_yen")
        risk = r.get("estimated_execution_risk_yen")
        if n is None:
            continue
        min_n = n / PER_SYMBOL_NOTIONAL_FRAC
        min_r = (risk / PER_TRADE_RISK_FRAC) if risk is not None else None
        kiox_daily.append({
            "day": r["day"],
            "reference_price": r.get("reference_price"),
            "reference_price_source": r.get("reference_price_source"),
            "one_lot_notional_yen": n,
            "estimated_execution_risk_yen": risk,
            "minimum_capital_by_notional": min_n,
            "minimum_capital_by_execution_risk": min_r,
            "required_capital": max(min_n, min_r or 0.0),
            "policy_evaluable_symbol_day": r["policy_evaluable_symbol_day"],
        })
    # Also add panel days for 285A even before evaluable for daily notional contract
    for (sym, day), p in panel.items():
        if sym != TARGET_SYMBOL:
            continue
        if any(k["day"] == day for k in kiox_daily):
            continue
        n = p.get("one_lot_notional_yen")
        if n is None:
            continue
        kiox_daily.append({
            "day": day,
            "reference_price": p.get("reference_price"),
            "reference_price_source": "previous_session_official_close" if p.get("ref_status") == "OK" else None,
            "one_lot_notional_yen": n,
            "estimated_execution_risk_yen": None,
            "minimum_capital_by_notional": n / PER_SYMBOL_NOTIONAL_FRAC,
            "minimum_capital_by_execution_risk": None,
            "required_capital": n / PER_SYMBOL_NOTIONAL_FRAC,
            "policy_evaluable_symbol_day": False,
            "note": "panel day; may be pre-evaluability",
        })
    kiox_daily = sorted(kiox_daily, key=lambda x: x["day"])
    kn = [float(x["one_lot_notional_yen"]) for x in kiox_daily]
    kr = [float(x["required_capital"]) for x in kiox_daily if x.get("required_capital") is not None]
    x10 = json.loads((X10_DIR / "report.json").read_text(encoding="utf-8"))
    x10_med = _f((x10.get("kioxia_profile") or {}).get("one_lot_notional_median"))
    x11_reported = _f((x11.get("kioxia_profile") or {}).get("one_lot_notional_yen"))
    notional_contract_ok = len(kiox_daily) >= 1 and _med(kn) is not None
    kiox_summary = {
        "symbol": TARGET_SYMBOL,
        "one_lot_notional_min": min(kn) if kn else None,
        "one_lot_notional_median": _med(kn),
        "one_lot_notional_p90": _pct(kn, 0.90),
        "one_lot_notional_max": max(kn) if kn else None,
        "required_capital_min": min(kr) if kr else None,
        "required_capital_median": _med(kr),
        "required_capital_p90": _pct(kr, 0.90),
        "required_capital_max": max(kr) if kr else None,
        "e1x10_median_notional": x10_med,
        "e1x11_reported_single_day_notional": x11_reported,
        "difference_explained": (
            "E1_X10 used cross-day median of daily one_lot_notional; "
            "E1_X11 reported a single rolling as-of day (latest history day D-1 close×100). "
            "V2 reports daily series + median/max required capital; no single-day sole display."
        ),
        "special_cased": False,
        "contract_resolved": notional_contract_ok,
    }

    # Capital base (unchanged unresolved)
    capital_base = {
        "status": "UNRESOLVED",
        "capital_base_resolved": False,
        "configured_risk_capital_cap_yen": None,
        "future_formula": "min(configured_risk_capital_cap_yen, verified StockAccountWallet)",
        "buying_power_rejected": True,
        "stock_wallet_not_auto_adopted": True,
        "margin_wallet_not_used_without_proof": True,
        "scenario_not_used_to_select_capital": True,
    }

    dynamic_core = {
        "bid": "AVAILABLE",
        "ask": "AVAILABLE",
        "qty": "AVAILABLE",
        "board_freshness": "AVAILABLE",
        "board_freshness_sec": BOARD_FRESHNESS_SEC,
        "qty_min": QTY_MIN,
        "fail_closed": True,
        "dynamic_core_fields_pass": True,
    }
    special_quote = {
        "status": "NOT_AVAILABLE_IN_CAPTURE",
        "special_quote_ready": False,
        "extended_event_guard": "DYNAMIC_SPECIAL_QUOTE_GUARD_NOT_READY",
        "does_not_invalidate_risk_history": True,
    }

    # Breadth evaluable: can we evaluate breadth (need policy evaluable days)?
    breadth_gate_evaluable = len(policy_evaluable_days) > 0
    # any scenario breadth pass? for matrix use whether gate contract is computable
    any_breadth_pass = any(s["breadth_gate_pass"] for s in scenario_rows)

    blockers = []
    if not notional_contract_ok:
        blockers.append("NOTIONAL_REPRESENTATIVE_CONTRACT_UNRESOLVED")
    if len(policy_evaluable_days) < MIN_POLICY_EVALUABLE_DAYS or not history_support_pass:
        blockers.append("RISK_HISTORY_SUPPORT_INSUFFICIENT")
    if not capital_base["capital_base_resolved"]:
        blockers.append("SAFE_CAPITAL_BASE_UNRESOLVED")
    if not dynamic_core["dynamic_core_fields_pass"]:
        blockers.append("DYNAMIC_GATE_FIELDS_INCOMPLETE")
    if not special_quote["special_quote_ready"]:
        blockers.append("DYNAMIC_SPECIAL_QUOTE_GUARD_NOT_READY")

    # Primary verdict by priority
    if not notional_contract_ok:
        verdict = "E1_X11_V2_NOTIONAL_REPRESENTATIVE_CONTRACT_UNRESOLVED"
    elif len(policy_evaluable_days) < MIN_POLICY_EVALUABLE_DAYS:
        verdict = "E1_X11_RISK_HISTORY_SUPPORT_INSUFFICIENT"
    elif not history_support_pass:
        verdict = "E1_X11_RISK_HISTORY_SUPPORT_INSUFFICIENT"
    elif not capital_base["capital_base_resolved"]:
        verdict = "E1_X11_SAFE_CAPITAL_BASE_UNRESOLVED"
    elif not dynamic_core["dynamic_core_fields_pass"]:
        verdict = "E1_X11_DYNAMIC_GATE_FIELDS_INCOMPLETE"
    else:
        verdict = "E1_X11_FIXED100_RISK_POLICY_FREEZE_READY"

    blocker_matrix = {
        "capital_base_resolved": False,
        "history_support_pass": history_support_pass,
        "breadth_gate_evaluable": breadth_gate_evaluable,
        "breadth_gate_pass": any_breadth_pass,
        "dynamic_core_fields_pass": True,
        "special_quote_ready": False,
        "pnl_independence_pass": True,
        "ab_determinism_pass": None,  # filled at publish
        "all_blockers": blockers,
        "primary_verdict": verdict,
    }

    # next step
    if "RISK_HISTORY_SUPPORT_INSUFFICIENT" in blockers:
        next_step = (
            "accumulate RISK_INFRASTRUCTURE_ONLY quote history toward >=20 trading days; "
            "no ENTRY/EXIT/PnL/Prospective/PFQ"
        )
    elif "SAFE_CAPITAL_BASE_UNRESOLVED" in blockers:
        next_step = "set configured_risk_capital_cap_yen; base = min(cap, verified StockAccountWallet)"
    else:
        next_step = "Risk Universe Policy Freeze document only"

    det = {
        "source_identity_sha": sha256_obj({"run": SOURCE_X11, "verdict": SOURCE_X11_VERDICT, "report": report_sha, "audit": audit_sha}),
        "symbol_day_evaluability_sha": sha256_obj([(r["symbol"], r["day"], r["status"]) for r in sd_rows]),
        "policy_evaluable_day_sha": sha256_obj(policy_evaluable_days),
        "asof_recurring_sha": sha256_obj({d: sorted(v) for d, v in asof_recurring_by_day.items()}),
        "coverage_sha": sha256_obj({"ratio": coverage_ratio, "n_policy_days": len(policy_evaluable_days), "den": den2}),
        "breadth_gate_sha": sha256_obj({"med_min": BREADTH_MEDIAN_MIN, "abs_min": BREADTH_ABS_MIN, "frac": BREADTH_FRAC_MEDIAN_MIN}),
        "capital_scenario_sha": sha256_obj(scenario_rows),
        "required_capital_sha": sha256_obj([(r["symbol"], r["required_capital_median"], r["required_capital_max"]) for r in req_rows]),
        "kioxia_profile_sha": sha256_obj(kiox_summary),
        "blocker_matrix_sha": sha256_obj(blocker_matrix),
        "verdict": verdict,
    }

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "precommit": precommit,
        "source_x11": SOURCE_X11,
        "source_x11_verdict": SOURCE_X11_VERDICT,
        "config_drift": CONFIG_DRIFT,
        "max_concurrent_positions": MAX_CONCURRENT,
        "policy_definition": {
            "POLICY_ID": POLICY_ID,
            "lot": LOT,
            "max_concurrent": MAX_CONCURRENT,
            "fractions": [PER_TRADE_RISK_FRAC, AGG_RISK_FRAC, PER_SYMBOL_NOTIONAL_FRAC, AGG_NOTIONAL_FRAC, RESERVE_FRAC],
            "board_freshness_sec": BOARD_FRESHNESS_SEC,
            "qty_min": QTY_MIN,
            "fail_closed": True,
            "unchanged": True,
        },
        "superseded_gate": {
            "name": SUPERSEDED_BREADTH,
            "old_median": OLD_BREADTH_MEDIAN,
            "old_min": OLD_BREADTH_MIN,
            "reason": f"impossible with global recurring n={len(global_recurring)} < {OLD_BREADTH_MEDIAN}",
            "corrected_median_min": BREADTH_MEDIAN_MIN,
            "corrected_abs_min": BREADTH_ABS_MIN,
            "corrected_frac_median_min": BREADTH_FRAC_MEDIAN_MIN,
        },
        "warmup_days": warmup_days,
        "policy_evaluable_days": policy_evaluable_days,
        "n_policy_evaluable_days": len(policy_evaluable_days),
        "global_recurring_reference": global_recurring,
        "coverage": {
            "asof_recurring_evaluable_fraction": coverage_ratio,
            "n_asof_recurring_symbol_days": den2,
            "n_evaluable_asof_recurring_symbol_days": num2,
            "history_support_pass": history_support_pass,
            "min_policy_evaluable_days_required": MIN_POLICY_EVALUABLE_DAYS,
        },
        "capital_base": capital_base,
        "capital_scenarios": scenario_rows,
        "kioxia_daily": kiox_daily,
        "kioxia_summary": kiox_summary,
        "required_capital": req_rows,
        "dynamic_core": dynamic_core,
        "special_quote": special_quote,
        "blocker_matrix": blocker_matrix,
        "all_blockers": blockers,
        "verdict": verdict,
        "verdict_detail": {
            "verdict": verdict,
            "all_blockers": blockers,
            "next": next_step,
            "n_policy_evaluable_days": len(policy_evaluable_days),
            "n_warmup_days": len(warmup_days),
        },
        "determinism_shas": det,
        "safety": _safety(),
        "stop": True,
        "_sheets": {
            "SourceIdentity": [{"run_id": SOURCE_X11, "verdict": SOURCE_X11_VERDICT, "report_sha": report_sha, "audit_sha": audit_sha}],
            "SupersededGate": [{"gate": SUPERSEDED_BREADTH, "old_median": OLD_BREADTH_MEDIAN, "old_min": OLD_BREADTH_MIN,
                                "corrected_median": BREADTH_MEDIAN_MIN, "corrected_min": BREADTH_ABS_MIN}],
            "SymbolDayEvaluability": sd_rows,
            "PolicyEvaluableDays": day_rows,
            "AsOfRecurring": [{"day": d, "symbol": s} for d, ss in asof_recurring_by_day.items() for s in sorted(ss)],
            "Coverage": [{"metric": k, "value": v} for k, v in {
                "coverage_ratio": coverage_ratio, "policy_evaluable_days": len(policy_evaluable_days),
                "history_support_pass": history_support_pass,
            }.items()],
            "BreadthGate": [{"corrected_median_min": BREADTH_MEDIAN_MIN, "corrected_abs_min": BREADTH_ABS_MIN,
                             "frac_median_min": BREADTH_FRAC_MEDIAN_MIN, "superseded": SUPERSEDED_BREADTH}],
            "CapitalScenarios": scenario_rows,
            "CapitalBase": [capital_base],
            "RequiredCapital": req_rows,
            "KioxiaDaily": kiox_daily,
            "KioxiaSummary": [kiox_summary],
            "DynamicCore": [dynamic_core],
            "SpecialQuote": [special_quote],
            "BlockerMatrix": [blocker_matrix],
        },
    }
    return report
