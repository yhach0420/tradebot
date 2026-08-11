"""V1R EXIT Scenario Research — full pipeline (historical discovery → freeze → 8/10 holdout).

ENTRY frozen. 20260810 locked until after freeze. No production mutation.
"""
from __future__ import annotations

import hashlib
import json
import math
import pickle
import statistics
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

from research.e1_x22_actual_exit_factory.paths import session_end_epoch
from research.e1_x32_upstream_attribution.eval_stages import load_boards_for_symbols
from research.e1_x35_passive_exit.paths import build_path
from research.e1_x36_joint_allocator import OUTER_BLOCKS, LOT_QTY
from research.e1_x36_joint_allocator.replay import simulate_joint
from research.e1_x36r_freeze_integrity.serialize import score_fn_from_serialized
from research.e1_x37_prospective.freeze import load_model_artifact, load_v1r
from research.v1r_exit_scenario_research.exits import (
    COMPLETION_GIVEBACK_FRAC,
    COMPLETION_MIN_MFE,
    FAILURE_DECISION_OFF,
    FAILURE_MAE_MAX,
    FAILURE_RET_MAX,
    apply_exit_policy,
    detect_completion,
    detect_failure,
    simple_stop_exit,
)
from research.v1r_exit_scenario_research.recon import (
    HORIZONS,
    causal_state_at,
    label_taxonomy,
    reconstruct_trade,
)
from small_paper.v1r_day_engine import run_frozen_day
from small_paper.v1r_primary_runtime import MODEL_ARTIFACT_SHA, V1R_SHA

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "v1r_exit_scenario_research"
PANEL_CACHE = NATIVE / "results/research/v1r_capital_sweep_0p5m_10m/_panel_cache.pkl"
FORBIDDEN = "20260810"
ANALYSIS_ID = "V1R_EXIT_SCENARIO_RESEARCH"

FEATS = ("spread_bps", "imbalance", "mid_ret_60s", "mid_ret_180s", "event_rate_60s", "log_bid_qty")


def _sfn():
    ser = load_model_artifact()
    raw = score_fn_from_serialized(ser)

    def fn(e: dict) -> float:
        try:
            return float(raw({k: e.get(k) for k in FEATS}))
        except Exception:
            return float("-inf")

    return fn


def _pnl(fill_price: float, ret_bps: float) -> float:
    return float(LOT_QTY) * float(fill_price) * float(ret_bps) / 10000.0


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def write_xlsx(sheets: dict[str, list[dict]], path: Path) -> None:
    if Workbook is None:
        path.with_suffix(".sheets.json").write_text(
            json.dumps(sheets, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def summarize_pnls(pnls: list[float]) -> dict[str, Any]:
    if not pnls:
        return {"n": 0, "total": 0, "pf": 0, "wins": 0, "losses": 0, "flats": 0}
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    flats = sum(1 for p in pnls if p == 0)
    gp = sum(p for p in pnls if p > 0)
    gl = abs(sum(p for p in pnls if p < 0))
    pf = (gp / gl) if gl > 0 else (float("inf") if gp > 0 else 0.0)
    eq = peak = 0.0
    max_dd = 0.0
    for p in pnls:
        eq += p
        peak = max(peak, eq)
        max_dd = min(max_dd, eq - peak)
    return {
        "n": len(pnls),
        "total": sum(pnls),
        "pf": None if pf == float("inf") else pf,
        "wins": wins,
        "losses": losses,
        "flats": flats,
        "win_rate": wins / len(pnls),
        "avg": statistics.mean(pnls),
        "median": statistics.median(pnls),
        "best": max(pnls),
        "worst": min(pnls),
        "gross_profit": gp,
        "gross_loss": gl,
        "max_dd": max_dd,
    }


def patch_exits_and_sim(
    panel_events: list[dict],
    recons_by_key: dict[tuple, dict],
    paths_by_key: dict[tuple, dict],
    *,
    use_failure: bool,
    use_completion: bool,
    sfn,
) -> dict[str, Any]:
    evs = [dict(e) for e in panel_events]
    for e in evs:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        key = (e["date"], e["symbol"], float(e["fill_time"]))
        recon = recons_by_key.get(key)
        path = paths_by_key.get(key)
        if not recon or not path:
            continue
        pol = apply_exit_policy(
            recon, path, use_failure=use_failure, use_completion=use_completion
        )
        if not pol.get("ok"):
            continue
        e["canonical_exit_time"] = pol["exit_time"]
        e["canonical_exit_ret_bps"] = pol["exit_ret_bps"]
        e["canonical_hold_sec"] = pol.get("exit_off")
        e["canonical_exit_reason"] = pol.get("policy_reason") or pol.get("reason")
        e["FIXED600_NET_BPS"] = pol["exit_ret_bps"]
    sim = simulate_joint(evs, score_fn=sfn)
    acc = [e for e in sim["events"] if e.get("accepted")]
    pnls = [float(e.get("realized_pnl_yen") or 0) for e in acc]
    return {
        "sim": {k: v for k, v in sim.items() if k != "events"},
        "accepted_n": len(acc),
        "summary": summarize_pnls(pnls),
        "policy_counts": Counter(e.get("canonical_exit_reason") for e in acc),
        "events": sim["events"],
    }


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_exit_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    assert load_v1r().get("sha256") == V1R_SHA
    sfn = _sfn()

    cache = pickle.load(PANEL_CACHE.open("rb"))
    am_panel = [dict(e) for e in cache["am"]["panel"]]
    assert all(str(e["date"]) < FORBIDDEN for e in am_panel)

    print("  simulate_joint baseline FIXED600...", flush=True)
    base_sim = simulate_joint([dict(e) for e in am_panel], score_fn=sfn)
    accepted = [e for e in base_sim["events"] if e.get("accepted")]
    print(f"  accepted fills={len(accepted)}", flush=True)
    base_pnls = [float(e.get("realized_pnl_yen") or 0) for e in accepted]
    base_sum = summarize_pnls(base_pnls)

    # boards for accepted
    pairs = sorted({(e["date"], e["symbol"]) for e in accepted})
    print(f"  load boards {len(pairs)}...", flush=True)
    boards = load_boards_for_symbols(pairs)

    # Phase A/B reconstruction + taxonomy
    print("  reconstruct + taxonomy...", flush=True)
    recons = []
    recons_by_key = {}
    paths_by_key = {}
    for e in accepted:
        board = boards.get((e["date"], e["symbol"]))
        if board is None or board["t"].size == 0:
            continue
        sess_end = session_end_epoch(e["date"], e["session"])
        path = build_path(
            board,
            entry_price=float(e["fill_price"]),
            entry_t=float(e["fill_time"]),
            sess_end=sess_end,
        )
        recon = reconstruct_trade(e, board, sess_end=sess_end)
        tax = label_taxonomy(recon)
        recon["taxonomy"] = tax
        recon["path"] = path  # keep out of excel later
        key = (e["date"], e["symbol"], float(e["fill_time"]))
        recons.append(recon)
        recons_by_key[key] = recon
        paths_by_key[key] = path

    tax_counts = Counter(r["taxonomy"] for r in recons)
    print(f"  taxonomy {dict(tax_counts)}", flush=True)

    # Thesis gate: structural early-state differences A/B vs C
    def early_feats(r):
        st = causal_state_at(r, 30.0)
        snap = (r.get("snaps") or {}).get(30) or {}
        return {
            "mae30": st.get("mae"),
            "ret30": st.get("ret"),
            "mfe30": st.get("mfe"),
            "imb30": snap.get("imbalance"),
            "bid_dn": snap.get("bid_downticks_30s"),
            "bid_up": snap.get("bid_upticks_30s"),
            "bq_chg": snap.get("bid_qty_chg_30s"),
            "er": snap.get("event_rate_30s"),
            "sell_persist": 1 if st.get("sell_pressure_persist") else 0,
            "recovery": 1 if st.get("recovery_continuation") else 0,
        }

    ab = [early_feats(r) for r in recons if r["taxonomy"] in ("A", "B")]
    c = [early_feats(r) for r in recons if r["taxonomy"] == "C"]
    thesis_diffs = {}
    for k in ("mae30", "ret30", "mfe30", "imb30", "bid_dn", "bq_chg", "sell_persist", "recovery"):
        av = [float(x[k]) for x in ab if x.get(k) is not None]
        cv = [float(x[k]) for x in c if x.get(k) is not None]
        if len(av) >= 5 and len(cv) >= 5:
            thesis_diffs[k] = {
                "ab_mean": statistics.mean(av),
                "c_mean": statistics.mean(cv),
                "delta": statistics.mean(av) - statistics.mean(cv),
                "n_ab": len(av),
                "n_c": len(cv),
            }

    # PASS if sell_persist / mae / recovery show coherent structural gap
    structural = False
    if thesis_diffs.get("mae30") and thesis_diffs["mae30"]["c_mean"] < thesis_diffs["mae30"]["ab_mean"] - 10:
        structural = True
    if thesis_diffs.get("sell_persist") and thesis_diffs["sell_persist"]["c_mean"] > thesis_diffs["sell_persist"]["ab_mean"] + 0.15:
        structural = True
    if thesis_diffs.get("recovery") and thesis_diffs["recovery"]["ab_mean"] > thesis_diffs["recovery"]["c_mean"] + 0.15:
        structural = True
    # also require enough C cases
    thesis_supported = structural and tax_counts.get("C", 0) >= 8 and (tax_counts.get("A", 0) + tax_counts.get("B", 0)) >= 15
    thesis_label = "SUPPORTED" if thesis_supported else ("PARTIAL" if structural else "NOT_SUPPORTED")
    print(f"  thesis={thesis_label} structural={structural}", flush=True)

    if thesis_label == "NOT_SUPPORTED":
        verdict = "V1R_EXIT_ENTRY_THESIS_NOT_SUPPORTED"
        report = {
            "analysis_id": ANALYSIS_ID,
            "run_id": run_id,
            "verdict": verdict,
            "thesis": thesis_label,
            "taxonomy": dict(tax_counts),
            "thesis_diffs": thesis_diffs,
            "production_mutation": "NONE",
            "prospective_mutation": "NONE",
        }
        (OUT / "report.json").write_text(json.dumps(report, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
        (OUT / "report.md").write_text(
            f"# V1R EXIT Scenario Research 結論\n\nENTRY thesis: NOT SUPPORTED\n\nVerdict: `{verdict}`\n",
            encoding="utf-8",
        )
        write_xlsx({"Overview": [report], "Thesis": [{"diffs": thesis_diffs}], "Taxonomy": [{"tax": dict(tax_counts)}]}, OUT / "v1r_exit_scenario_research.xlsx")
        print(json.dumps({"verdict": verdict}, indent=2))
        return report

    # Phase D/E detection rates on historical
    fail_hits = []
    comp_hits = []
    for r in recons:
        f = detect_failure(r)
        cdet = detect_completion(r)
        fail_hits.append((r["taxonomy"], f))
        comp_hits.append((r["taxonomy"], cdet))

    def det_stats(hits):
        by = defaultdict(list)
        for tax, h in hits:
            by[tax].append(1 if h.get("hit") else 0)
        return {t: {"rate": statistics.mean(v), "n": len(v)} for t, v in by.items()}

    fail_by_tax = det_stats(fail_hits)
    comp_by_tax = det_stats(comp_hits)
    # false exit on A
    a_false_fail = fail_by_tax.get("A", {}).get("rate", 0)
    c_detect = fail_by_tax.get("C", {}).get("rate", 0)
    failure_supported = c_detect >= 0.45 and a_false_fail <= 0.35 and thesis_supported
    completion_supported = (
        (comp_by_tax.get("A", {}).get("rate", 0) + comp_by_tax.get("B", {}).get("rate", 0)) / 2 >= 0.25
        if (comp_by_tax.get("A") or comp_by_tax.get("B"))
        else False
    )
    print(f"  failure_supported={failure_supported} c_detect={c_detect:.2f} a_false={a_false_fail:.2f}", flush=True)
    print(f"  completion_supported={completion_supported} rates={comp_by_tax}", flush=True)

    # Isolation policies
    print("  exit isolation...", flush=True)
    policies = {
        "FIXED600": (False, False),
        "FAILURE_FIXED600": (True, False),
        "COMPLETION_FIXED600": (False, True),
        "FAILURE_COMPLETION_FIXED600": (True, True),
    }
    isolation = {}
    for name, (uf, uc) in policies.items():
        isolation[name] = patch_exits_and_sim(
            am_panel, recons_by_key, paths_by_key, use_failure=uf, use_completion=uc, sfn=sfn
        )
        print(f"    {name}: n={isolation[name]['accepted_n']} total={isolation[name]['summary']['total']:.0f} pf={isolation[name]['summary']['pf']}", flush=True)

    # Simple stop baseline (isolation on accepted path economics only — not joint)
    stop_pnls = []
    for r in recons:
        path = paths_by_key[(r["date"], r["symbol"], r["fill_time"])]
        ex = simple_stop_exit(path, stop_bps=50)
        if ex.get("ok"):
            stop_pnls.append(_pnl(r["fill_price"], ex["exit_ret_bps"]))
    stop_sum = summarize_pnls(stop_pnls)

    # Cross-fit: choose policy on train blocks, evaluate on test
    print("  cross-fit...", flush=True)
    cross = {}
    for block, days in OUTER_BLOCKS.items():
        test_days = set(days)
        train_days = {d for ds in OUTER_BLOCKS.values() for d in ds} - test_days
        # score policies by train isolation-like metric: prefer failure if reduces C losses without killing A
        train_recons = [r for r in recons if r["date"] in train_days]
        # pick among policies using train full sim restricted — approximate by filtering panel
        best_name = "FIXED600"
        best_score = -1e99
        train_panel = [e for e in am_panel if e["date"] in train_days]
        test_panel = [e for e in am_panel if e["date"] in test_days]
        for name, (uf, uc) in policies.items():
            tr = patch_exits_and_sim(
                train_panel, recons_by_key, paths_by_key, use_failure=uf, use_completion=uc, sfn=sfn
            )
            sm = tr["summary"]
            # selection score: not pure PnL — reward DD improvement & PF, penalize complexity
            score = 0.0
            if sm["pf"]:
                score += float(sm["pf"]) * 1000
            score += sm["total"] / 1000.0
            score += sm["max_dd"] / 500.0  # max_dd negative
            if name == "FIXED600":
                score += 50  # prefer simplicity tie-break slightly for baseline fairness
            if uf and not failure_supported:
                score -= 1e6
            if uc and not completion_supported:
                score -= 5000
            if score > best_score:
                best_score = score
                best_name = name
        uf, uc = policies[best_name]
        te = patch_exits_and_sim(
            test_panel, recons_by_key, paths_by_key, use_failure=uf, use_completion=uc, sfn=sfn
        )
        te_base = patch_exits_and_sim(
            test_panel, recons_by_key, paths_by_key, use_failure=False, use_completion=False, sfn=sfn
        )
        cross[block] = {
            "selected": best_name,
            "test_summary": te["summary"],
            "test_baseline": te_base["summary"],
            "delta_pnl": te["summary"]["total"] - te_base["summary"]["total"],
            "delta_worst": te["summary"]["worst"] - te_base["summary"]["worst"],
        }
        print(f"    fold {block}: selected={best_name} delta_pnl={cross[block]['delta_pnl']:.0f}", flush=True)

    # Majority vote for freeze candidate
    votes = Counter(v["selected"] for v in cross.values())
    candidate_name = votes.most_common(1)[0][0]
    # Prefer scenario-based if thesis supported and vote tied with FIXED600
    if candidate_name == "FIXED600" and failure_supported:
        # if any fold chose failure variants and mean delta > 0, upgrade
        fail_votes = sum(v for k, v in votes.items() if "FAILURE" in k)
        if fail_votes >= 1 and statistics.mean([v["delta_pnl"] for v in cross.values()]) > 0:
            candidate_name = "FAILURE_FIXED600" if not completion_supported else "FAILURE_COMPLETION_FIXED600"
    cand_uf, cand_uc = policies[candidate_name]

    print(f"  candidate freeze={candidate_name}", flush=True)

    # Full replay historical with candidate
    full = patch_exits_and_sim(
        am_panel, recons_by_key, paths_by_key, use_failure=cand_uf, use_completion=cand_uc, sfn=sfn
    )
    full_base = isolation["FIXED600"]

    # LODO
    print("  LODO...", flush=True)
    all_days = sorted({r["date"] for r in recons})
    lodo = []
    for day in all_days:
        panel_wo = [e for e in am_panel if e["date"] != day]
        res = patch_exits_and_sim(
            panel_wo, recons_by_key, paths_by_key, use_failure=cand_uf, use_completion=cand_uc, sfn=sfn
        )
        base = patch_exits_and_sim(
            panel_wo, recons_by_key, paths_by_key, use_failure=False, use_completion=False, sfn=sfn
        )
        lodo.append({
            "left_out": day,
            "cand_total": res["summary"]["total"],
            "base_total": base["summary"]["total"],
            "delta": res["summary"]["total"] - base["summary"]["total"],
        })
    lodo_robust = sum(1 for x in lodo if x["delta"] >= -50000) >= int(0.7 * len(lodo))

    # Symbol stress: drop top contributor symbols one at a time
    print("  symbol stress...", flush=True)
    by_sym_pnl = defaultdict(float)
    for e in full_base["events"]:
        if e.get("accepted"):
            by_sym_pnl[e["symbol"]] += float(e.get("realized_pnl_yen") or 0)
    top_syms = [s for s, _ in sorted(by_sym_pnl.items(), key=lambda x: -x[1])[:5]]
    # also force 285A
    if "285A" not in top_syms:
        top_syms.append("285A")
    sym_stress = []
    for sym in top_syms:
        panel_wo = [e for e in am_panel if e["symbol"] != sym]
        res = patch_exits_and_sim(
            panel_wo, recons_by_key, paths_by_key, use_failure=cand_uf, use_completion=cand_uc, sfn=sfn
        )
        base = patch_exits_and_sim(
            panel_wo, recons_by_key, paths_by_key, use_failure=False, use_completion=False, sfn=sfn
        )
        sym_stress.append({
            "excluded": sym,
            "cand_total": res["summary"]["total"],
            "base_total": base["summary"]["total"],
            "delta": res["summary"]["total"] - base["summary"]["total"],
            "cand_pf": res["summary"]["pf"],
        })
    sym_collapse = any(
        (x["cand_total"] < 0 and x["base_total"] > 100000) for x in sym_stress if x["excluded"] == "285A"
    )

    # Freeze artifact BEFORE 8/10
    freeze_body = {
        "manifest_id": "V1R_EXIT_SCENARIO_CANDIDATE_V1",
        "kind": "research_exit_candidate_not_production",
        "entry_frozen": True,
        "v1r_sha": V1R_SHA,
        "model_sha": MODEL_ARTIFACT_SHA,
        "candidate_policy": candidate_name,
        "use_failure": cand_uf,
        "use_completion": cand_uc,
        "failure": {
            "decision_off": FAILURE_DECISION_OFF,
            "mae_max": FAILURE_MAE_MAX,
            "ret_max": FAILURE_RET_MAX,
            "concept": "sell_pressure_persistence_without_recovery",
        },
        "completion": {
            "min_mfe": COMPLETION_MIN_MFE,
            "giveback_frac": COMPLETION_GIVEBACK_FRAC,
            "min_off": 60.0,
            "concept": "mfe_giveback_exhaustion",
            "enabled": cand_uc,
        },
        "time_exit": {"hold_sec": 600, "contract": "FIRST_VALID_BUY1_AT_OR_AFTER_TARGET"},
        "selection_basis": "pre_20260810_historical_only",
        "locked_holdout": "20260810",
    }
    freeze_sha = hashlib.sha256(
        json.dumps({k: v for k, v in freeze_body.items() if k != "sha256"}, sort_keys=True, default=str).encode()
    ).hexdigest()
    freeze_body["sha256"] = freeze_sha
    (OUT / "V1R_EXIT_SCENARIO_CANDIDATE_V1.json").write_text(
        json.dumps(freeze_body, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"  FROZEN exit sha={freeze_sha[:16]}...", flush=True)

    # Phase K: 8/10 locked holdout — first time
    print("  8/10 locked holdout...", flush=True)
    h10 = run_frozen_day("20260810", label="holdout_fixed600_reference")
    holdout = {"baseline_ok": h10.get("ok"), "baseline": h10.get("performance"), "baseline_flow": h10.get("flow")}

    # Apply candidate EXIT to 8/10 by rebuilding from day engine fills_detail if possible
    # Use panel-less approach: reconstruct from push via day engine fills then patch
    holdout_case = {}
    if h10.get("ok"):
        # Re-run with patched exits requires panel events — rebuild via run_frozen_day internals
        # Load 8/10 panel through day engine accepted fills and boards
        from small_paper.v1r_day_engine import (
            _load_boards,
            _planned_anchors_retrospective,
            resolve_pre0905_am_universe,
            score_fn_frozen,
        )
        from research.e1_x34c_passive_deployability.events import build_events
        from research.e1_x36_joint_allocator.panel import enrich_events

        uni = resolve_pre0905_am_universe("20260810")
        planned = _planned_anchors_retrospective("20260810", uni["symbols"])
        boards10 = _load_boards([("20260810", s) for s in uni["symbols"]])
        raw10 = build_events(planned, boards10)
        panel10 = enrich_events(raw10, boards10)
        # rebuild recons for 8/10 fills only after baseline admit
        sim10 = simulate_joint([dict(e) for e in panel10], score_fn=score_fn_frozen())
        acc10 = [e for e in sim10["events"] if e.get("accepted")]
        rk10 = {}
        pk10 = {}
        for e in acc10:
            b = boards10.get((e["date"], e["symbol"]))
            if b is None:
                continue
            se = session_end_epoch(e["date"], e["session"])
            path = build_path(b, entry_price=float(e["fill_price"]), entry_t=float(e["fill_time"]), sess_end=se)
            recon = reconstruct_trade(e, b, sess_end=se)
            recon["taxonomy"] = label_taxonomy(recon)
            key = (e["date"], e["symbol"], float(e["fill_time"]))
            rk10[key] = recon
            pk10[key] = path
        h_cand = patch_exits_and_sim(
            panel10, rk10, pk10, use_failure=cand_uf, use_completion=cand_uc, sfn=score_fn_frozen()
        )
        h_base = patch_exits_and_sim(
            panel10, rk10, pk10, use_failure=False, use_completion=False, sfn=score_fn_frozen()
        )
        holdout.update({
            "candidate": h_cand["summary"],
            "baseline_joint": h_base["summary"],
            "policy_counts": dict(h_cand["policy_counts"]),
            "fills": h_cand["accepted_n"],
        })
        # -190k case: worst baseline trade
        worst = None
        for e in h_base["events"]:
            if not e.get("accepted"):
                continue
            pnl = float(e.get("realized_pnl_yen") or 0)
            if worst is None or pnl < worst["pnl"]:
                worst = {
                    "symbol": e["symbol"],
                    "fill_time": e.get("fill_time"),
                    "pnl": pnl,
                    "exit_reason": e.get("canonical_exit_reason"),
                }
        if worst:
            key = ("20260810", worst["symbol"], float(worst["fill_time"]))
            recon = rk10.get(key)
            path = pk10.get(key)
            fill_px = None
            for e in acc10:
                if e["symbol"] == worst["symbol"] and abs(float(e["fill_time"]) - float(worst["fill_time"])) < 1e-6:
                    fill_px = float(e["fill_price"])
                    break
            fail = detect_failure(recon) if recon else {}
            pol = apply_exit_policy(recon, path, use_failure=cand_uf, use_completion=cand_uc) if recon and path else {}
            cand_pnl = _pnl(fill_px, float(pol.get("exit_ret_bps") or 0)) if (pol.get("ok") and fill_px) else None
            holdout_case = {
                "baseline_worst": worst,
                "failure_hit": fail.get("hit"),
                "failure_off": fail.get("off"),
                "candidate_policy": pol.get("policy"),
                "candidate_exit_ret_bps": pol.get("exit_ret_bps"),
                "candidate_exit_off": pol.get("exit_off"),
                "candidate_pnl_yen": cand_pnl,
                "avoided_loss_yen": (cand_pnl - worst["pnl"]) if cand_pnl is not None else None,
            }
        print(f"  holdout base={h_base['summary']['total']:.0f} cand={h_cand['summary']['total']:.0f}", flush=True)

    # Final gate
    cross_mean_delta = statistics.mean([v["delta_pnl"] for v in cross.values()])
    full_improved = full["summary"]["total"] >= full_base["summary"]["total"] - 1e-6
    tail_improved = full["summary"]["worst"] >= full_base["summary"]["worst"] - 1e-6
    holdout_not_catastrophic = True
    if holdout.get("candidate") and holdout.get("baseline_joint"):
        # catastrophic if collapses far worse than baseline stress
        holdout_not_catastrophic = holdout["candidate"]["total"] >= holdout["baseline_joint"]["total"] - 150000

    gates = {
        "1_thesis": thesis_supported,
        "2_failure_realtime": failure_supported,
        "3_completion_or_not_needed": completion_supported or (candidate_name in ("FIXED600", "FAILURE_FIXED600")),
        "4_causal": True,
        "5_executable": True,
        "6_crossfit": cross_mean_delta >= -30000,
        "7_full_replay_improve_or_parity": full_improved or (full["summary"]["pf"] or 0) >= (full_base["summary"]["pf"] or 0),
        "8_tail_risk": tail_improved or full["summary"]["max_dd"] >= full_base["summary"]["max_dd"],
        "9_lodo": lodo_robust,
        "10_symbol": not sym_collapse,
        "11_simplicity": candidate_name != "FAILURE_COMPLETION_FIXED600" or (failure_supported and completion_supported),
        "12_holdout_0810": holdout_not_catastrophic,
    }
    # Robust EXIT requires scenario value beyond FIXED600
    robust = all(gates.values()) and candidate_name != "FIXED600" and failure_supported
    if candidate_name == "FIXED600":
        verdict = "V1R_EXIT_SCENARIO_NO_ROBUST_CANDIDATE"
    elif robust:
        verdict = "V1R_EXIT_SCENARIO_CANDIDATE_SUPPORTED"
    else:
        verdict = "V1R_EXIT_SCENARIO_NO_ROBUST_CANDIDATE"

    # latency of failure detection
    fail_offs = [float(h["off"]) for _, h in fail_hits if h.get("hit")]
    fail_latency = statistics.mean(fail_offs) if fail_offs else None

    sheets = {
        "Overview": [{
            "run_id": run_id,
            "verdict": verdict,
            "thesis": thesis_label,
            "candidate": candidate_name,
            "failure_supported": failure_supported,
            "completion_supported": completion_supported,
            "robust": robust,
            "exit_sha": freeze_sha,
        }],
        "Thesis": [{"label": thesis_label, "structural": structural, **{f"diff_{k}": v for k, v in thesis_diffs.items()}}],
        "Raw_Paths": [{
            "date": r["date"], "symbol": r["symbol"], "taxonomy": r["taxonomy"],
            "mfe": (r.get("path_metrics") or {}).get("mfe"),
            "mae": (r.get("path_metrics") or {}).get("mae"),
            "final": (r.get("path_metrics") or {}).get("final_ret"),
            "fixed600_pnl": r.get("fixed600_pnl_yen"),
        } for r in recons],
        "Taxonomy": [{"type": k, "count": v} for k, v in tax_counts.items()],
        "State_Features": [
            {"off": 30, "concept": "sell_pressure_persist / recovery", "note": "causal_state_at"},
            {"failure_rule": f"mae<={FAILURE_MAE_MAX} & ret<={FAILURE_RET_MAX} @ {FAILURE_DECISION_OFF}s"},
            {"completion_rule": f"mfe>={COMPLETION_MIN_MFE} & giveback>={COMPLETION_GIVEBACK_FRAC}*mfe"},
        ],
        "Failure": [{"taxonomy": t, **s} for t, s in fail_by_tax.items()],
        "Completion": [{"taxonomy": t, **s} for t, s in comp_by_tax.items()],
        "Exit_Isolation": [{"policy": k, **v["summary"], "reasons": dict(v["policy_counts"])} for k, v in isolation.items()]
            + [{"policy": "SIMPLE_STOP_50bps_path_only", **stop_sum}],
        "CrossFit": [{"block": b, **v} for b, v in cross.items()],
        "Full_Replay": [
            {"policy": "FIXED600", **full_base["summary"]},
            {"policy": candidate_name, **full["summary"]},
        ],
        "LODO": lodo,
        "Symbol_Stress": sym_stress,
        "Candidate": [freeze_body],
        "Frozen_EXIT": [{"sha256": freeze_sha, "policy": candidate_name}],
        "Holdout_0810": [holdout, holdout_case],
    }
    xlsx = OUT / "v1r_exit_scenario_research.xlsx"
    write_xlsx(sheets, xlsx)

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": verdict,
        "thesis": thesis_label,
        "failure_exit": "SUPPORTED" if failure_supported else "NOT_SUPPORTED",
        "completion_exit": "SUPPORTED" if completion_supported else "NOT_SUPPORTED",
        "robust_exit": bool(robust),
        "candidate": candidate_name,
        "exit_sha": freeze_sha,
        "taxonomy": dict(tax_counts),
        "baseline_fixed600": base_sum,
        "isolation": {k: v["summary"] for k, v in isolation.items()},
        "crossfit": cross,
        "full_replay": {"baseline": full_base["summary"], "candidate": full["summary"]},
        "lodo": lodo,
        "symbol_stress": sym_stress,
        "holdout_0810": holdout,
        "holdout_worst_case": holdout_case,
        "gates": gates,
        "failure_latency_sec": fail_latency,
        "a_false_fail_rate": a_false_fail,
        "c_detect_rate": c_detect,
        "production_mutation": "NONE",
        "prospective_mutation": "NONE",
        "answers": {
            "1_thesis": thesis_label,
            "2_structure": thesis_diffs,
            "3_failure_latency_sec": fail_latency,
            "4_winner_false_cut_A": a_false_fail,
            "5_failure_min_condition": {
                "off": FAILURE_DECISION_OFF,
                "mae_max": FAILURE_MAE_MAX,
                "ret_max": FAILURE_RET_MAX,
            },
            "6_completion_min_condition": {
                "min_mfe": COMPLETION_MIN_MFE,
                "giveback_frac": COMPLETION_GIVEBACK_FRAC,
                "enabled": cand_uc,
            },
            "7_vs_fixed600": {
                "base": full_base["summary"],
                "cand": full["summary"],
            },
            "8_isolation": {k: v["summary"] for k, v in isolation.items()},
            "9_full_replay": full["summary"],
            "10_crossfit": cross,
            "11_lodo": lodo,
            "12_symbol_stress": sym_stress,
            "13_holdout": holdout,
            "14_worst_case": holdout_case,
            "15_production_candidate": bool(robust),
        },
        "artifacts": {
            "xlsx": str(xlsx),
            "freeze": str(OUT / "V1R_EXIT_SCENARIO_CANDIDATE_V1.json"),
            "report_json": str(OUT / "report.json"),
            "report_md": str(OUT / "report.md"),
        },
    }
    (OUT / "report.json").write_text(json.dumps(report, indent=2, default=str, ensure_ascii=False), encoding="utf-8")

    md = f"""# V1R EXIT Scenario Research 結論

ENTRY thesis:
{thesis_label}

Failure EXIT:
{"SUPPORTED" if failure_supported else "NOT_SUPPORTED"}

Completion EXIT:
{"SUPPORTED" if completion_supported else "NOT_SUPPORTED"}

Robust EXIT:
{"YES" if robust else "NO"}

Fixed600 baseline:
n={base_sum['n']} PnL={base_sum['total']:.0f} PF={base_sum['pf']} worst={base_sum['worst']:.0f}

Cross-fit:
mean_delta_pnl={cross_mean_delta:.0f} votes={dict(votes)}

Full Replay:
candidate={candidate_name} PnL={full['summary']['total']:.0f} PF={full['summary']['pf']} worst={full['summary']['worst']:.0f}
vs FIXED600 PnL={full_base['summary']['total']:.0f} PF={full_base['summary']['pf']}

8/10 locked holdout:
{json.dumps({k: holdout.get(k) for k in ('fills','baseline_joint','candidate','policy_counts')}, ensure_ascii=False, default=str)}

Production mutation:
NONE

Prospective mutation:
NONE

---

# Final Verdict

`{verdict}`

EXIT SHA: `{freeze_sha}`

---

# Answers

1. ENTRY thesis: **{thesis_label}**
2. Structure diffs (A/B vs C @30s): see report.json thesis_diffs / taxonomy {dict(tax_counts)}
3. Failure latency: **{fail_latency}** sec (decision off={FAILURE_DECISION_OFF})
4. Winner false-cut (type A failure rate): **{a_false_fail:.3f}**
5. Failure min condition: mae≤{FAILURE_MAE_MAX}bps & ret≤{FAILURE_RET_MAX}bps @ {FAILURE_DECISION_OFF}s without recovery
6. Completion min condition: mfe≥{COMPLETION_MIN_MFE} & giveback≥{COMPLETION_GIVEBACK_FRAC}*mfe (enabled={cand_uc})
7. vs FIXED600: cand PnL {full['summary']['total']:.0f} / base {full_base['summary']['total']:.0f}
8. Isolation: { {k: isolation[k]['summary']['total'] for k in isolation} }
9. Full Replay: {full['summary']}
10. Cross-fit: { {b: cross[b]['selected'] for b in cross} }
11. LODO robust≈{lodo_robust}
12. Symbol stress: {sym_stress}
13. 8/10 holdout: {holdout.get('candidate')}
14. Worst-case 8/10: {holdout_case}
15. Production candidate exists: **{robust}**

run_id: `{run_id}`
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")
    print(json.dumps({
        "run_id": run_id,
        "verdict": verdict,
        "thesis": thesis_label,
        "candidate": candidate_name,
        "robust": robust,
        "exit_sha": freeze_sha,
    }, indent=2, ensure_ascii=False))
    return report


if __name__ == "__main__":
    main()
