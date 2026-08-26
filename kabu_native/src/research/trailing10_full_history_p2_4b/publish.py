"""Aggregate P2-4B TRAIL10 reused-history test. Write report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.dynamic_anchor_p2_2.binding import P1_REPORT
from research.dynamic_anchor_p2_2.publish import _stats, trade_ledger_sha
from research.dynamic_anchor_p2_3.metrics import ENTRY_ORDER, funnel_integrity, pct_block, pf_out, rate
from research.trailing10_full_history_p2_4b import (
    ANALYSIS_ID,
    CANDIDATE_ID,
    CONTAMINATION_LABEL,
    DOCUMENT_ID,
    FIXED_FILL_RATE,
    FIXED_MAXDD,
    FIXED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS,
    FIXED_PF,
    FIXED_PNL,
    FIXED_TRADES,
    FREEZE_TIMESTAMP_JST,
    FROZEN_IMPLEMENTATION_SHA,
    FROZEN_SPEC_SHA,
    FULL14,
    OLD_CANDIDATE_STATUS,
    OLD_DYNAMIC_FILL_RATE,
    OLD_DYNAMIC_MAXDD,
    OLD_DYNAMIC_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS,
    OLD_DYNAMIC_PF,
    OLD_DYNAMIC_PNL,
    OLD_DYNAMIC_TRADES,
    OLD_T1_PRIOR_COVERAGE_D,
    OLD_T1_PRIOR_COVERAGE_N,
    PERIOD_END,
    PERIOD_START,
    PREDECLARED_TOP1,
    PREDECLARED_TOP3,
    REST11,
    REST11_FIXED_PF,
    REST11_FIXED_PNL,
    REST11_OLD_DYNAMIC_PF,
    REST11_OLD_DYNAMIC_PNL,
    TASK_LABEL,
    VERDICT_BETTER,
    VERDICT_BLOCKED,
    VERDICT_NOT_SUPPORTED,
    VERDICT_PARTIAL,
    VERDICT_SUPPORTED_NOT_BETTER,
)
from research.trailing10_full_history_p2_4b.coverage import PRIOR_EDGE

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "trailing10_full_history_test_p2_4b"
P2_2_REPORT = NATIVE / "results" / "research" / "dynamic_anchor_pnl_test_p2_2" / "report.json"
P2_3_REPORT = NATIVE / "results" / "research" / "dynamic_anchor_failure_decomposition_p2_3" / "report.json"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _exclude(trades: list[dict[str, Any]], drop: set[str]) -> dict[str, Any]:
    return _stats([t for t in trades if str(t.get("date")) not in drop])


def _concentration(daily: list[dict[str, Any]], trades: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [(r["date"], float(r["TRAIL10_pnl"])) for r in daily]
    if not rows:
        return {}
    best = max(rows, key=lambda x: x[1])
    worst = min(rows, key=lambda x: x[1])
    total = sum(p for _, p in rows)
    ordered = sorted(rows, key=lambda x: -x[1])
    top1 = ordered[0][1] / total if abs(total) > 1e-12 else None
    top3 = sum(p for _, p in ordered[:3]) / total if abs(total) > 1e-12 else None
    pos = sum(1 for _, p in rows if p > 1e-9)
    neg = sum(1 for _, p in rows if p < -1e-9)
    flat = len(rows) - pos - neg
    loo = []
    for d, _p in rows:
        st = _stats([t for t in trades if t.get("date") != d])
        loo.append(st["pnl"])
    return {
        "best_day": {"date": best[0], "pnl": best[1]},
        "worst_day": {"date": worst[0], "pnl": worst[1]},
        "top1_contribution": top1,
        "top3_contribution": top3,
        "positive_days": pos,
        "negative_days": neg,
        "flat_days": flat,
        "loo_min_pnl": min(loo) if loo else None,
    }


def _anchor_density(days: list[dict[str, Any]]) -> dict[str, Any]:
    per: list[int] = []
    gaps: list[float] = []
    symbols: set[str] = set()
    n = 0
    for d in days:
        by_sym: dict[str, list[float]] = defaultdict(list)
        for a in d.get("anchors") or []:
            n += 1
            s = str(a.get("symbol"))
            symbols.add(s)
            g = a.get("g")
            if g is not None:
                by_sym[s].append(float(g))
        for s, gs in by_sym.items():
            per.append(len(gs))
            gs.sort()
            for i in range(1, len(gs)):
                gaps.append(gs[i] - gs[i - 1])
    arr = np.asarray(per, dtype=float) if per else np.asarray([], dtype=float)
    return {
        "anchors_total": n,
        "unique_symbols": len(symbols),
        "median_anchors_per_symbol_day": None if arr.size == 0 else float(np.median(arr)),
        "p90_anchors_per_symbol_day": None if arr.size == 0 else float(np.percentile(arr, 90)),
        "max_anchors_per_symbol_day": None if arr.size == 0 else int(np.max(arr)),
        "min_inter_anchor_gap_sec": None if not gaps else float(min(gaps)),
        "TRUE_persistence_refire": sum(int(d.get("TRUE_PERSISTENCE_REFIRE") or 0) for d in days),
        "duplicate_edge": sum(int(d.get("duplicate_edge") or 0) for d in days),
        "not_evaluable_created_edge": sum(int(d.get("not_evaluable_created_edge") or 0) for d in days),
    }


def _replacement(trail: dict[str, Any]) -> str:
    pnl = float(trail.get("pnl") or 0)
    pf = trail.get("PF")
    dd = trail.get("maxDD")
    pf_n = float("inf") if pf in (None, "Infinity") else float(pf)
    dd_n = float(dd) if dd is not None else float("-inf")
    better_pnl = pnl > FIXED_PNL
    better_pf = pf_n > FIXED_PF
    better_dd = dd_n >= FIXED_MAXDD
    if better_pnl and better_pf and better_dd:
        return "TRAIL10_BETTER_THAN_FIXED_ON_REUSED_HISTORY"
    if (pnl < FIXED_PNL) and (pf_n < FIXED_PF):
        return "TRAIL10_WORSE_THAN_FIXED_ON_REUSED_HISTORY"
    return "TRAIL10_MIXED_VS_FIXED_ON_REUSED_HISTORY"


def build_report(
    *,
    inventory: list[dict[str, Any]],
    day_results: list[dict[str, Any]],
    failed: list[str],
    bind: dict[str, Any],
    det: dict[str, Any],
    p1: dict[str, Any],
    p2_2: dict[str, Any],
    p2_3: dict[str, Any],
) -> dict[str, Any]:
    now = datetime.now(JST).isoformat(timespec="seconds")
    full_set = set(FULL14)
    by = {d["date"]: d for d in day_results if d.get("ok")}
    prim_days = [by[d] for d in FULL14 if d in by]
    ref_days = [d for d in day_results if d.get("ok")]
    prim_tr = [t for d in prim_days for t in (d.get("trades") or [])]
    trail = _stats(prim_tr)
    p1_daily = {r["date"]: r for r in (p1.get("daily") or [])}
    p2_daily = {r["date"]: r for r in ((p2_2.get("DAILY") or {}).get("rows") or [])}

    daily_rows = []
    for d in FULL14:
        fd = p1_daily.get(d) or {}
        od = p2_daily.get(d) or {}
        td = by.get(d) or {}
        fpnl = float(fd.get("pnl") or 0)
        opnl = float(od.get("Dynamic_pnl") or od.get("DYNAMIC_pnl") or 0)
        tpnl = float(td.get("pnl") or 0)
        daily_rows.append({
            "date": d,
            "Fixed_trades": fd.get("trades"),
            "OldDynamic_trades": od.get("Dynamic_trades"),
            "TRAIL10_trades": td.get("trade_n"),
            "Fixed_pnl": fpnl,
            "OldDynamic_pnl": opnl,
            "TRAIL10_pnl": tpnl,
            "Fixed_PF": fd.get("PF"),
            "OldDynamic_PF": od.get("Dynamic_PF"),
            "TRAIL10_PF": pf_out(td.get("PF")),
            "TRAIL10_vs_Fixed_delta": round(tpnl - fpnl, 2),
            "TRAIL10_vs_Old_delta": round(tpnl - opnl, 2),
        })

    coverage = [r for d in prim_days for r in (d.get("coverage") or [])]
    prior_n = sum(1 for r in coverage if r.get("has_prior_trail10_edge"))
    cov_counts = Counter(str(r.get("state_at_fixed_signal")) for r in coverage)
    h1 = prior_n > OLD_T1_PRIOR_COVERAGE_N

    term = [r for d in prim_days for r in (d.get("terminals") or [])]
    entry_counts = Counter(str(r.get("entry_terminal") or "") for r in term)
    fill_counts = Counter(str(r.get("fill_terminal") or "") for r in term if r.get("entry_terminal") == "ADMITTED")
    anchors_n = sum(int(d.get("false_to_true_anchors") or 0) for d in prim_days)
    admitted = int(entry_counts.get("ADMITTED") or 0)
    filled = int(fill_counts.get("FILLED") or 0)
    expired = int(fill_counts.get("EXPIRED") or 0)
    funnel = {
        "grid_evaluations": sum(int(d.get("grid_evaluations") or 0) for d in prim_days),
        "evaluable_states": sum(int(d.get("evaluable_states") or 0) for d in prim_days),
        "true_state_rows": sum(int(d.get("true_state_rows") or 0) for d in prim_days),
        "false_to_true_anchors": anchors_n,
        "current_entry_evaluated": anchors_n,
        "not_selected": int(entry_counts.get("NOT_SELECTED") or 0),
        "blocked_open": int(entry_counts.get("BLOCKED_OPEN") or 0),
        "blocked_pending": int(entry_counts.get("BLOCKED_PENDING") or 0),
        "blocked_cap": int(entry_counts.get("BLOCKED_CAP") or 0),
        "other_reject": int(entry_counts.get("OTHER_REJECT") or 0),
        "admitted": admitted,
        "fills": filled,
        "expired": expired,
        "trades": trail["trades"],
    }
    integ = funnel_integrity({
        "confirmed": anchors_n,
        **{k: funnel[k.lower()] if k.lower() in funnel else funnel.get(k) for k in ENTRY_ORDER},
        "NOT_SELECTED": funnel["not_selected"],
        "BLOCKED_OPEN": funnel["blocked_open"],
        "BLOCKED_PENDING": funnel["blocked_pending"],
        "BLOCKED_CAP": funnel["blocked_cap"],
        "OTHER_REJECT": funnel["other_reject"],
        "ADMITTED": admitted,
        "FILLED": filled,
        "EXPIRED": expired,
        "admitted": admitted,
        "fills": filled,
        "expired": expired,
    })

    ask_bps = [x for d in prim_days for x in (d.get("ask_bps") or [])]
    fill_rate = rate(filled, admitted)
    med_ask = pct_block(ask_bps).get("median")
    h2 = bool(
        fill_rate is not None
        and fill_rate > OLD_DYNAMIC_FILL_RATE
        and med_ask is not None
        and float(med_ask) < OLD_DYNAMIC_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS
    )

    trail_pf = trail.get("PF")
    trail_pf_n = float("inf") if trail_pf in (None, "Infinity") else float(trail_pf)
    h3 = bool(float(trail["pnl"]) > OLD_DYNAMIC_PNL and trail_pf_n > OLD_DYNAMIC_PF)

    leak_snap = sum(int(d.get("snapshot_future_leak") or 0) for d in prim_days)
    leak_cp = sum(int(d.get("checkpoint_future_leak") or 0) for d in prim_days)
    leak_wake = sum(int(d.get("decision_snapshot_future_leak") or 0) for d in prim_days)
    future_leak = not (leak_snap == 0 and leak_cp == 0 and leak_wake == 0)

    dens = _anchor_density(prim_days)
    contract_ok = (
        dens["TRUE_persistence_refire"] == 0
        and dens["duplicate_edge"] == 0
        and dens["not_evaluable_created_edge"] == 0
        and not future_leak
        and integ.get("pass") is True
    )

    top3_cov = [r for r in coverage if r.get("date") in PREDECLARED_TOP3]
    top3_prior = sum(1 for r in top3_cov if r.get("has_prior_trail10_edge"))
    old_top3_prior = int(((p2_3.get("TOP3") or {}).get("fixed_trades_with_prior_T1") or 27))
    old_top3_n = int(((p2_3.get("TOP3") or {}).get("fixed_trades") or 100))

    top3_fix = _stats([t for t in (p1.get("trades") or []) if t.get("date") in PREDECLARED_TOP3])
    old_top3_tr = 0
    old_top3_pnl = 0.0
    for r in ((p2_2.get("DAILY") or {}).get("rows") or []):
        if r.get("date") in PREDECLARED_TOP3:
            old_top3_tr += int(r.get("Dynamic_trades") or 0)
            old_top3_pnl += float(r.get("Dynamic_pnl") or 0)
    p2_3_top = p2_3.get("TOP3") or {}
    if p2_3_top.get("actual_dynamic_pnl") is not None:
        old_top3_pnl = float(p2_3_top["actual_dynamic_pnl"])
    trail_top3 = _stats([t for t in prim_tr if t.get("date") in PREDECLARED_TOP3])
    trail_ex1 = _exclude(prim_tr, {PREDECLARED_TOP1})
    trail_ex3 = _exclude(prim_tr, set(PREDECLARED_TOP3))
    fix_ex1 = _exclude([t for t in (p1.get("trades") or []) if t.get("date") in full_set], {PREDECLARED_TOP1})
    fix_ex3 = _exclude([t for t in (p1.get("trades") or []) if t.get("date") in full_set], set(PREDECLARED_TOP3))
    old_ex1_block = p2_2.get("EX_TOP1_DYNAMIC") or {}
    old_ex3_block = p2_2.get("EX_FIXED_TOP3_DYNAMIC") or {}

    rest_tr = [t for t in prim_tr if t.get("date") in REST11]
    rest_stats = _stats(rest_tr)

    replacement = _replacement(trail)
    n_h = sum([h1, h2, h3])
    blocked = bool(failed) or not bind.get("pass") or not det.get("pass") or not contract_ok or future_leak
    if blocked:
        verdict = VERDICT_BLOCKED
    elif n_h == 3 and replacement == "TRAIL10_BETTER_THAN_FIXED_ON_REUSED_HISTORY":
        verdict = VERDICT_BETTER
    elif n_h == 3:
        verdict = VERDICT_SUPPORTED_NOT_BETTER
    elif n_h == 0:
        verdict = VERDICT_NOT_SUPPORTED
    else:
        verdict = VERDICT_PARTIAL

    am = trail.get("AM") or {}
    pm = trail.get("PM") or {}
    dens_ok = dens["TRUE_persistence_refire"] == 0 and dens["duplicate_edge"] == 0 and dens["not_evaluable_created_edge"] == 0

    return {
        "task": "P2-4B",
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "generated_at_jst": now,
        "label": TASK_LABEL,
        "not_oos": True,
        "not_prospective": True,
        "not_clean_holdout": True,
        "not_robust": True,
        "contamination_if_mechanics": CONTAMINATION_LABEL,
        "CANDIDATE": CANDIDATE_ID,
        "SPEC_SHA_MATCH": bind.get("SPEC_SHA_MATCH"),
        "IMPLEMENTATION_SHA_MATCH": bind.get("IMPLEMENTATION_SHA_MATCH"),
        "PERIOD": f"{PERIOD_START}-{PERIOD_END}",
        "PRIMARY_FULL": {
            "FIXED": {
                "trades": FIXED_TRADES,
                "pnl": FIXED_PNL,
                "PF": FIXED_PF,
                "maxDD": FIXED_MAXDD,
            },
            "OLD_DYNAMIC": {
                "trades": OLD_DYNAMIC_TRADES,
                "pnl": OLD_DYNAMIC_PNL,
                "PF": OLD_DYNAMIC_PF,
                "maxDD": OLD_DYNAMIC_MAXDD,
                "fill_rate": OLD_DYNAMIC_FILL_RATE,
            },
            "TRAIL10": {
                "days": len(prim_days),
                "anchors": anchors_n,
                "trades": trail["trades"],
                "win": trail["win"],
                "loss": trail["loss"],
                "draw": trail["draw"],
                "gross_profit": trail["gross_profit"],
                "gross_loss": trail["gross_loss"],
                "pnl": trail["pnl"],
                "PF": pf_out(trail["PF"]),
                "avg": trail["avg_pnl"],
                "maxDD": trail["maxDD"],
                "AM": am,
                "PM": pm,
                "fill_rate": fill_rate,
                "admitted": admitted,
                "fills": filled,
                "expired": expired,
            },
        },
        "H1_TRIGGER_COVERAGE": {
            "OLD_T1_PRIOR_COVERAGE": f"{OLD_T1_PRIOR_COVERAGE_N}/{OLD_T1_PRIOR_COVERAGE_D}",
            "TRAIL10_PRIOR_EDGE_COVERAGE": f"{prior_n}/{len(coverage)}",
            "TRAIL10_PRIOR_EDGE_N": prior_n,
            "fixed_trades_analyzed": len(coverage),
            "state_counts": dict(cov_counts),
            "H1_SUPPORTED": h1,
            "top3": {
                "OLD_T1_PRIOR_COVERAGE": f"{old_top3_prior}/{old_top3_n}",
                "TRAIL10_PRIOR_EDGE_COVERAGE": f"{top3_prior}/{len(top3_cov)}",
            },
        },
        "H2_EXECUTION": {
            "OLD_DYNAMIC_FILL_RATE": OLD_DYNAMIC_FILL_RATE,
            "TRAIL10_FILL_RATE": fill_rate,
            "FIXED_FILL_RATE": FIXED_FILL_RATE,
            "OLD_DYNAMIC_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS": OLD_DYNAMIC_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS,
            "TRAIL10_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS": med_ask,
            "FIXED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS": FIXED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS,
            "ask_bps_n": len(ask_bps),
            "H2_SUPPORTED": h2,
        },
        "H3_PERFORMANCE_RECOVERY": {
            "OLD_DYNAMIC_PNL": OLD_DYNAMIC_PNL,
            "OLD_DYNAMIC_PF": OLD_DYNAMIC_PF,
            "TRAIL10_PNL": trail["pnl"],
            "TRAIL10_PF": pf_out(trail["PF"]),
            "H3_SUPPORTED": h3,
        },
        "FIXED_REPLACEMENT_RESULT": replacement,
        "TOP3": {
            "Fixed": {"trades": top3_fix["trades"], "pnl": top3_fix["pnl"], "PF": pf_out(top3_fix["PF"]), "maxDD": top3_fix["maxDD"]},
            "OldDynamic": {
                "trades": old_top3_tr,
                "pnl": old_top3_pnl,
                "PF": p2_3_top.get("dynamic_PF"),
                "maxDD": p2_3_top.get("dynamic_maxDD"),
            },
            "Trail10": {"trades": trail_top3["trades"], "pnl": trail_top3["pnl"], "PF": pf_out(trail_top3["PF"]), "maxDD": trail_top3["maxDD"]},
        },
        "EX_20260722": {
            "Fixed": {"trades": fix_ex1["trades"], "pnl": fix_ex1["pnl"], "PF": pf_out(fix_ex1["PF"]), "maxDD": fix_ex1["maxDD"]},
            "OldDynamic": old_ex1_block or None,
            "Trail10": {"trades": trail_ex1["trades"], "pnl": trail_ex1["pnl"], "PF": pf_out(trail_ex1["PF"]), "maxDD": trail_ex1["maxDD"]},
        },
        "EX_TOP3": {
            "Fixed": {"trades": fix_ex3["trades"], "pnl": fix_ex3["pnl"], "PF": pf_out(fix_ex3["PF"]), "maxDD": fix_ex3["maxDD"]},
            "OldDynamic": old_ex3_block,
            "Trail10": {"trades": trail_ex3["trades"], "pnl": trail_ex3["pnl"], "PF": pf_out(trail_ex3["PF"]), "maxDD": trail_ex3["maxDD"]},
        },
        "REST11": {
            "days": list(REST11),
            "Fixed": {"pnl": REST11_FIXED_PNL, "PF": REST11_FIXED_PF},
            "OldDynamic": {"pnl": REST11_OLD_DYNAMIC_PNL, "PF": REST11_OLD_DYNAMIC_PF},
            "Trail10": {
                "trades": rest_stats["trades"],
                "pnl": rest_stats["pnl"],
                "PF": pf_out(rest_stats["PF"]),
                "maxDD": rest_stats["maxDD"],
            },
        },
        "FUNNEL": funnel,
        "funnel_integrity": integ,
        "ANCHOR_DENSITY": dens,
        "CONCENTRATION": _concentration(daily_rows, prim_tr),
        "FUTURE_LEAK": future_leak,
        "leak_counts": {
            "checkpoint_future_leak": leak_cp,
            "snapshot_future_leak": leak_snap,
            "wake_event_future_content_leak": leak_wake,
        },
        "DETERMINISM": "PASS" if det.get("pass") else "FAIL",
        "determinism_detail": det,
        "SPEC_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "OLD_CANDIDATE_STATUS": OLD_CANDIDATE_STATUS,
        "SPEC_SHA": FROZEN_SPEC_SHA,
        "IMPLEMENTATION_SHA": FROZEN_IMPLEMENTATION_SHA,
        "FREEZE_TIMESTAMP_JST": FREEZE_TIMESTAMP_JST,
        "CURRENT_ENTRY_BINDING": bind.get("CURRENT_ENTRY_BINDING"),
        "inventory_eligible": sum(1 for r in inventory if r.get("replay_eligible")),
        "completed": len(day_results),
        "failed": failed,
        "contract_ok": contract_ok,
        "anchor_contract_ok": dens_ok,
        "CHRONOLOGICAL_SINGLE_STREAM": all(d.get("CHRONOLOGICAL_SINGLE_STREAM") for d in prim_days) if prim_days else False,
        "prospective_status": (
            "Freeze unchanged. Unseen Capture obtained after freeze_timestamp remains "
            "PROSPECTIVE_CANDIDATE_DATA. This period is REUSED_HISTORY_HYPOTHESIS_TEST only."
        ),
        "verdict": verdict,
        "identity": {
            "strategy_sha": STRATEGY_SHA,
            "entry_sha": ENTRY_SHA,
            "exit_sha": EXIT_SHA,
            "anchor_sha": ANCHOR_SHA,
            "spec_sha": FROZEN_SPEC_SHA,
            "implementation_sha": FROZEN_IMPLEMENTATION_SHA,
            "V1RNativeEntryLive_sha": bind.get("V1RNativeEntryLive_sha"),
            "V1RLiveDualLane_sha": bind.get("V1RLiveDualLane_sha"),
        },
        "_daily_rows": daily_rows,
        "_coverage": coverage,
        "_funnel": funnel,
        "_ask_pct": pct_block(ask_bps),
        "REFERENCE_DAYS": len(ref_days),
        "REFERENCE_TRAIL10_TRADES": sum(int(d.get("trade_n") or 0) for d in ref_days),
        "REFERENCE_TRAIL10_PNL": round(sum(float(d.get("pnl") or 0) for d in ref_days), 2),
    }


def _md(rep: dict[str, Any]) -> str:
    p = rep["PRIMARY_FULL"]
    f, o, t = p["FIXED"], p["OLD_DYNAMIC"], p["TRAIL10"]
    h1 = rep["H1_TRIGGER_COVERAGE"]
    h2 = rep["H2_EXECUTION"]
    h3 = rep["H3_PERFORMANCE_RECOVERY"]
    return f"""# P2-4B TRAIL10 full-history hypothesis test

Label: `{rep['label']}`. Not OOS / prospective / clean holdout / robust.

CANDIDATE:
{rep['CANDIDATE']}

SPEC_SHA_MATCH:
{rep['SPEC_SHA_MATCH']}

PERIOD:
{rep['PERIOD']}

PRIMARY_FULL:

FIXED:
trades: {f['trades']}
pnl: {f['pnl']}
PF: {f['PF']}
maxDD: {f['maxDD']}

OLD_DYNAMIC:
trades: {o['trades']}
pnl: {o['pnl']}
PF: {o['PF']}
maxDD: {o['maxDD']}
fill_rate: {o['fill_rate']}

TRAIL10:
anchors: {t['anchors']}
trades: {t['trades']}
pnl: {t['pnl']}
PF: {t['PF']}
maxDD: {t['maxDD']}
AM: {t['AM']}
PM: {t['PM']}
fill_rate: {t['fill_rate']}

H1_TRIGGER_COVERAGE:

OLD_T1_PRIOR_COVERAGE:
{h1['OLD_T1_PRIOR_COVERAGE']}

TRAIL10_PRIOR_EDGE_COVERAGE:
{h1['TRAIL10_PRIOR_EDGE_COVERAGE']}

H1_SUPPORTED:
{str(h1['H1_SUPPORTED']).lower()}

H2_EXECUTION:

OLD_DYNAMIC_FILL_RATE:
{h2['OLD_DYNAMIC_FILL_RATE']}

TRAIL10_FILL_RATE:
{h2['TRAIL10_FILL_RATE']}

FIXED_FILL_RATE:
{h2['FIXED_FILL_RATE']}

OLD_DYNAMIC_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS:
{h2['OLD_DYNAMIC_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS']}

TRAIL10_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS:
{h2['TRAIL10_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS']}

FIXED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS:
{h2['FIXED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS']}

H2_SUPPORTED:
{str(h2['H2_SUPPORTED']).lower()}

H3_PERFORMANCE_RECOVERY:

OLD_DYNAMIC_PNL:
{h3['OLD_DYNAMIC_PNL']}

OLD_DYNAMIC_PF:
{h3['OLD_DYNAMIC_PF']}

TRAIL10_PNL:
{h3['TRAIL10_PNL']}

TRAIL10_PF:
{h3['TRAIL10_PF']}

H3_SUPPORTED:
{str(h3['H3_SUPPORTED']).lower()}

FIXED_REPLACEMENT_RESULT:
{rep['FIXED_REPLACEMENT_RESULT']}

TOP3:
Fixed: {rep['TOP3']['Fixed']}
OldDynamic: {rep['TOP3']['OldDynamic']}
Trail10: {rep['TOP3']['Trail10']}

EX_TOP3:
Fixed: {rep['EX_TOP3']['Fixed']}
OldDynamic: {rep['EX_TOP3']['OldDynamic']}
Trail10: {rep['EX_TOP3']['Trail10']}

REST11:
Fixed: {rep['REST11']['Fixed']}
OldDynamic: {rep['REST11']['OldDynamic']}
Trail10: {rep['REST11']['Trail10']}

FUTURE_LEAK:
{str(rep['FUTURE_LEAK']).lower()}

DETERMINISM:
{rep['DETERMINISM']}

SPEC_CHANGED:
false

ENTRY_EXIT_CHANGED:
false

RUNTIME_CHANGED:
false

SAFETY:
submit/cancel/live=0/0/0

verdict:
{rep['verdict']}

STOP. Do not retune. Do not create a new candidate. Do not adopt into Runtime.
"""


def write_artifacts(rep: dict[str, Any]) -> dict[str, Path]:
    OUT.mkdir(parents=True, exist_ok=True)
    for p in OUT.iterdir():
        if p.is_file() and p.name not in {"report.json", "report.md", "audit.xlsx"}:
            p.unlink()
    pub = {k: v for k, v in rep.items() if not k.startswith("_")}
    (OUT / "report.json").write_text(json.dumps(pub, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    (OUT / "report.md").write_text(_md(rep), encoding="utf-8")
    wb = Workbook()

    def sh(name: str):
        if name == "Summary":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    p = rep["PRIMARY_FULL"]
    t = p["TRAIL10"]
    summary = [
        {"field": k, "value": json.dumps(v, ensure_ascii=False, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("CANDIDATE", rep["CANDIDATE"]),
            ("SPEC_SHA_MATCH", rep["SPEC_SHA_MATCH"]),
            ("PERIOD", rep["PERIOD"]),
            ("label", rep["label"]),
            ("FIXED_trades", p["FIXED"]["trades"]),
            ("FIXED_pnl", p["FIXED"]["pnl"]),
            ("FIXED_PF", p["FIXED"]["PF"]),
            ("FIXED_maxDD", p["FIXED"]["maxDD"]),
            ("OLD_DYNAMIC_trades", p["OLD_DYNAMIC"]["trades"]),
            ("OLD_DYNAMIC_pnl", p["OLD_DYNAMIC"]["pnl"]),
            ("OLD_DYNAMIC_PF", p["OLD_DYNAMIC"]["PF"]),
            ("OLD_DYNAMIC_maxDD", p["OLD_DYNAMIC"]["maxDD"]),
            ("OLD_DYNAMIC_fill_rate", p["OLD_DYNAMIC"]["fill_rate"]),
            ("TRAIL10_anchors", t["anchors"]),
            ("TRAIL10_trades", t["trades"]),
            ("TRAIL10_pnl", t["pnl"]),
            ("TRAIL10_PF", t["PF"]),
            ("TRAIL10_maxDD", t["maxDD"]),
            ("TRAIL10_AM", t["AM"]),
            ("TRAIL10_PM", t["PM"]),
            ("TRAIL10_fill_rate", t["fill_rate"]),
            ("H1_SUPPORTED", rep["H1_TRIGGER_COVERAGE"]["H1_SUPPORTED"]),
            ("H2_SUPPORTED", rep["H2_EXECUTION"]["H2_SUPPORTED"]),
            ("H3_SUPPORTED", rep["H3_PERFORMANCE_RECOVERY"]["H3_SUPPORTED"]),
            ("FIXED_REPLACEMENT_RESULT", rep["FIXED_REPLACEMENT_RESULT"]),
            ("FUTURE_LEAK", rep["FUTURE_LEAK"]),
            ("DETERMINISM", rep["DETERMINISM"]),
            ("verdict", rep["verdict"]),
            ("SAFETY", rep["SAFETY"]),
        ]
    ]
    _write_rows(sh("Summary"), summary)
    h1 = rep["H1_TRIGGER_COVERAGE"]
    h2 = rep["H2_EXECUTION"]
    h3 = rep["H3_PERFORMANCE_RECOVERY"]
    _write_rows(sh("Hypotheses"), [
        {"hypothesis": "H1_TRIGGER_COVERAGE", "supported": h1["H1_SUPPORTED"], "old": h1["OLD_T1_PRIOR_COVERAGE"], "new": h1["TRAIL10_PRIOR_EDGE_COVERAGE"]},
        {"hypothesis": "H2_EXECUTION", "supported": h2["H2_SUPPORTED"], "old": h2["OLD_DYNAMIC_FILL_RATE"], "new": h2["TRAIL10_FILL_RATE"]},
        {"hypothesis": "H3_PERFORMANCE_RECOVERY", "supported": h3["H3_SUPPORTED"], "old": h3["OLD_DYNAMIC_PNL"], "new": h3["TRAIL10_PNL"]},
        {"hypothesis": "FIXED_REPLACEMENT", "supported": None, "old": None, "new": rep["FIXED_REPLACEMENT_RESULT"]},
    ])
    _write_rows(sh("Daily"), rep["_daily_rows"])
    _write_rows(sh("Funnel"), [{"stage": k, "n": v} for k, v in rep["_funnel"].items()] + [
        {"stage": "integrity_pass", "n": rep["funnel_integrity"].get("pass")},
    ])
    _write_rows(sh("Anchor_Density"), [{"field": k, "value": v} for k, v in rep["ANCHOR_DENSITY"].items()])
    _write_rows(sh("Fixed_Coverage"), [
        {k: r.get(k) for k in (
            "date", "symbol", "session", "fixed_anchor_time", "has_prior_trail10_edge",
            "state_at_fixed_signal", "last_status", "last_trail10_state", "fixed_pnl",
        )}
        for r in rep["_coverage"]
    ] or [{"date": None}])
    _write_rows(sh("Execution"), [
        {"field": k, "value": v} for k, v in [
            *h2.items(),
            ("ask_pct", rep["_ask_pct"]),
        ]
    ])
    _write_rows(sh("Top3"), [
        {"arm": "Fixed", **rep["TOP3"]["Fixed"]},
        {"arm": "OldDynamic", **{k: v for k, v in rep["TOP3"]["OldDynamic"].items()}},
        {"arm": "Trail10", **rep["TOP3"]["Trail10"]},
        {"arm": "EX_20260722_Fixed", **rep["EX_20260722"]["Fixed"]},
        {"arm": "EX_20260722_Trail10", **rep["EX_20260722"]["Trail10"]},
        {"arm": "EX_TOP3_Fixed", **rep["EX_TOP3"]["Fixed"]},
        {"arm": "EX_TOP3_Trail10", **rep["EX_TOP3"]["Trail10"]},
    ])
    _write_rows(sh("Rest11"), [
        {"arm": k, **(v if isinstance(v, dict) else {"value": v})}
        for k, v in [
            ("Fixed", rep["REST11"]["Fixed"]),
            ("OldDynamic", rep["REST11"]["OldDynamic"]),
            ("Trail10", rep["REST11"]["Trail10"]),
        ]
    ])
    _write_rows(sh("Concentration"), [{"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v} for k, v in (rep.get("CONCENTRATION") or {}).items()])
    _write_rows(sh("Determinism"), [{"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v} for k, v in (rep.get("determinism_detail") or {}).items()])
    _write_rows(sh("Identity"), [{"field": k, "value": v} for k, v in rep["identity"].items()])
    _write_rows(sh("Safety"), [
        {"field": "submit/cancel/live", "value": "0/0/0"},
        {"field": "SPEC_CHANGED", "value": False},
        {"field": "ENTRY_EXIT_CHANGED", "value": False},
        {"field": "RUNTIME_CHANGED", "value": False},
        {"field": "label", "value": TASK_LABEL},
        {"field": "note", "value": "One-shot frozen spec. No retune. No runtime adopt."},
    ])
    wb.save(OUT / "audit.xlsx")
    return {"report_json": OUT / "report.json", "report_md": OUT / "report.md", "audit_xlsx": OUT / "audit.xlsx"}
