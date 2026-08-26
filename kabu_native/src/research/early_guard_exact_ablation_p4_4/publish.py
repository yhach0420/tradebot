"""Write P4-4 precommit + report.json / report.md / audit.xlsx."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.early_guard_exact_ablation_p4_4 import (
    CANONICAL_GUARD_N,
    DOCUMENT_ID,
    HIST_FOREGONE_WINNER,
    HIST_RATIO,
    HIST_SAVED_LOSS,
    P1_MAXDD,
    P1_PF,
    P1_PNL,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_OK,
)
from research.early_guard_exact_ablation_p4_4.metrics import (
    classify,
    daily_rows,
    guard89_off_paths,
    portfolio_effects,
    reconcile_89,
    slice_days,
    summary_block,
    tail_block,
)
from small_paper.v1r_exit_v2_contract import FROZEN_CONTINUATION, FROZEN_GUARD

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "early_guard_exact_ablation_p4_4"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    return hashlib.sha256(p.read_bytes()).hexdigest() if p.is_file() else ""


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(12, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def _delta_block(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
    def _num(x):
        if x is None or x == "Infinity":
            return None
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    def sub(k):
        xa, xb = _num(a.get(k)), _num(b.get(k))
        if xa is None or xb is None:
            return None
        return xb - xa

    return {
        "trades": (b.get("trades") or 0) - (a.get("trades") or 0),
        "pnl": round(float(b.get("pnl") or 0) - float(a.get("pnl") or 0), 2),
        "PF": sub("PF"),
        "maxDD": sub("maxDD"),
    }


def build_report(
    *,
    precommit: dict[str, Any],
    recon: dict[str, Any],
    canonical: list[dict[str, Any]],
    top10_ids: list[str],
    top20_ids: list[str],
    base_days: list[dict[str, Any]],
    off_days: list[dict[str, Any]],
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    t10 = set(top10_ids)
    t20 = set(top20_ids)
    for t in canonical:
        tid = str(t.get("trade_id") or "")
        t["TOP10"] = tid in t10
        t["TOP20"] = tid in t20

    base_trades = [t for d in base_days for t in (d.get("trades") or [])]
    off_trades = [t for d in off_days for t in (d.get("trades") or [])]
    base_admits = [t for d in base_days for t in (d.get("admits") or [])]
    off_admits = [t for d in off_days for t in (d.get("admits") or [])]
    base_fills = [t for d in base_days for t in (d.get("fills") or [])]
    off_fills = [t for d in off_days for t in (d.get("fills") or [])]

    rec89 = reconcile_89(canonical=canonical, baseline=base_trades) if base_trades else {
        "ok": False, "canonical_n": 0, "baseline_n": 0, "matched_n": 0, "rows": [],
        "missing_in_baseline": [], "extra_in_baseline": [],
    }
    for r in rec89.get("rows") or []:
        r["TOP10"] = str(r.get("canonical_trade_id") or "") in t10
        r["TOP20"] = str(r.get("canonical_trade_id") or "") in t20

    base = summary_block(base_trades) if base_trades else {}
    off = summary_block(off_trades) if off_trades else {}
    base_top = summary_block(slice_days(base_trades, PREDECLARED_TOP3)) if base_trades else {}
    off_top = summary_block(slice_days(off_trades, PREDECLARED_TOP3)) if off_trades else {}
    base_rest = summary_block(slice_days(base_trades, REST11)) if base_trades else {}
    off_rest = summary_block(slice_days(off_trades, REST11)) if off_trades else {}

    loc = (
        guard89_off_paths(guard89=rec89.get("rows") or [], off_trades=off_trades)
        if off_trades and rec89.get("rows")
        else {}
    )
    port = (
        portfolio_effects(
            base_trades=base_trades,
            off_trades=off_trades,
            base_admits=base_admits,
            off_admits=off_admits,
            base_fills=base_fills,
            off_fills=off_fills,
            base_cap=sum(int(d.get("cap_blocked") or 0) for d in base_days),
            off_cap=sum(int(d.get("cap_blocked") or 0) for d in off_days),
            base_same=sum(int(d.get("same_symbol_blocked") or 0) for d in base_days),
            off_same=sum(int(d.get("same_symbol_blocked") or 0) for d in off_days),
        )
        if base_days and off_days
        else {}
    )
    tail = (
        tail_block(
            canonical=canonical,
            guard89=rec89.get("rows") or [],
            base_trades=base_trades,
            off_trades=off_trades,
            top10_ids=t10,
            top20_ids=t20,
        )
        if base_trades
        else {}
    )
    daily = daily_rows(base_trades, off_trades) if base_trades and off_trades else []
    day_counts = {
        "GUARD_ON_BETTER": sum(1 for r in daily if r.get("winner") == "GUARD_ON_BETTER"),
        "GUARD_OFF_BETTER": sum(1 for r in daily if r.get("winner") == "GUARD_OFF_BETTER"),
        "equal": sum(1 for r in daily if r.get("winner") == "equal"),
    }

    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if base_days:
        if int(base.get("trades") or 0) != P1_TRADES:
            integrity.append(f"BASELINE_N={base.get('trades')}!={P1_TRADES}")
        if abs(float(base.get("pnl") or 0) - P1_PNL) >= 0.51:
            integrity.append(f"BASELINE_PNL={base.get('pnl')}!={P1_PNL}")
        if not rec89.get("ok"):
            integrity.append(
                f"GUARD89_MISMATCH can={rec89.get('canonical_n')} base={rec89.get('baseline_n')} "
                f"match={rec89.get('matched_n')}"
            )

    klass = classify(
        integrity=integrity if not blocked else integrity,
        base=base or {},
        off=off or {},
        base_rest=base_rest or {},
        off_rest=off_rest or {},
        net_guard_value=loc.get("net_guard_value"),
    ) if (not blocked and off_days) else {"CLASSIFICATION": None, "why": blocked_reason}

    net = loc.get("net_guard_value")
    hist_dir = None
    if net is not None:
        hist_dir = "SAME" if float(net) > 0 else "DIFFERENT"

    if blocked or not recon.get("pass") or (base_days and not rec89.get("ok")):
        verdict = VERDICT_BLOCKED
    else:
        verdict = VERDICT_OK

    delta = _delta_block(base, off) if base and off else {}
    dest = loc.get("destination") or {}

    sheets = {
        "baseline": [{"slice": "FULL14", **base}, {"slice": "TOP3", **base_top}, {"slice": "REST11", **base_rest}],
        "off": [{"slice": "FULL14", **off}, {"slice": "TOP3", **off_top}, {"slice": "REST11", **off_rest}],
        "g89": loc.get("rows") or rec89.get("rows") or [],
        "saved": [
            {
                "saved_loss_current_exact": loc.get("saved_loss"),
                "foregone_winner_current_exact": loc.get("foregone_winner"),
                "net_guard_value": loc.get("net_guard_value"),
                "ratio": loc.get("ratio"),
                "saved_n": loc.get("saved_n"),
                "foregone_n": loc.get("foregone_n"),
                "matched_n": loc.get("matched_n"),
                "historical_saved_loss": HIST_SAVED_LOSS,
                "historical_foregone_winner": HIST_FOREGONE_WINNER,
                "historical_ratio": HIST_RATIO,
                "CURRENT_VS_HISTORICAL_DIRECTION": hist_dir,
            }
        ],
        "port": [port] if port else [],
        "tail": [
            {"group": "TOP10", **((tail.get("TOP10") or {}))},
            {"group": "TOP20", **((tail.get("TOP20") or {}))},
            {"group": "CONT_EXTEND_750", **((tail.get("CONT_EXTEND_750") or {}))},
        ],
        "top3rest": [
            {"slice": "TOP3", "side": "baseline", **base_top},
            {"slice": "TOP3", "side": "guard_off", **off_top},
            {"slice": "REST11", "side": "baseline", **base_rest},
            {"slice": "REST11", "side": "guard_off", **off_rest},
        ],
        "daily": daily,
    }

    report = {
        "task": "P4-4",
        "ANALYSIS_ID": "P4_4_EARLY_GUARD_EXACT_ABLATION",
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "new strategy validation", "production approval"],
        "PRECOMMIT_SHA": precommit.get("SHA"),
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "BASELINE_RECONCILE": "PASS" if (base_days and int(base.get("trades") or 0) == P1_TRADES and abs(float(base.get("pnl") or 0) - P1_PNL) < 0.51) else ("FAIL" if base_days else None),
        "EARLY_GUARD_BASELINE_N": rec89.get("baseline_n"),
        "GUARD89_MATCH": rec89.get("ok"),
        "GUARD89_MATCHED_N": rec89.get("matched_n"),
        "reconcile": recon,
        "BASELINE": base,
        "GUARD_OFF": off,
        "DELTA": delta,
        "CURRENT_EXACT_GUARD_ECONOMICS": {
            "saved_loss": loc.get("saved_loss"),
            "foregone_winner": loc.get("foregone_winner"),
            "net_guard_value": loc.get("net_guard_value"),
            "ratio": loc.get("ratio"),
            "saved_n": loc.get("saved_n"),
            "foregone_n": loc.get("foregone_n"),
            "matched_n": loc.get("matched_n"),
        },
        "EARLY_GUARD89_DESTINATION": {
            "exit600": (dest.get("exit600") or {}).get("n"),
            "extend750": (dest.get("extend750") or {}).get("n"),
            "session_close": (dest.get("session_close") or {}).get("n"),
            "other": (dest.get("other") or {}).get("n"),
            "lost_fill": (dest.get("lost_fill") or {}).get("n"),
            "detail": dest,
        },
        "PORTFOLIO_EFFECT": port,
        "TOP3": {"baseline": base_top, "guard_off": off_top, "delta": _delta_block(base_top, off_top) if base_top and off_top else {}},
        "REST11": {
            "baseline_trades": base_rest.get("trades"),
            "guard_off_trades": off_rest.get("trades"),
            "baseline_pnl": base_rest.get("pnl"),
            "guard_off_pnl": off_rest.get("pnl"),
            "baseline_PF": base_rest.get("PF"),
            "guard_off_PF": off_rest.get("PF"),
            "baseline_maxDD": base_rest.get("maxDD"),
            "guard_off_maxDD": off_rest.get("maxDD"),
        },
        "TAIL": tail,
        "DAY_LEVEL": {"counts": day_counts, "rows": daily},
        "HISTORICAL_REFERENCE": {
            "saved_loss": HIST_SAVED_LOSS,
            "foregone_winner": HIST_FOREGONE_WINNER,
            "ratio": HIST_RATIO,
        },
        "CURRENT_VS_HISTORICAL_DIRECTION": hist_dir,
        "CLASSIFICATION": klass.get("CLASSIFICATION"),
        "CLASSIFICATION_DETAIL": klass,
        "EARLY_GUARD_CHANGED": False,
        "MID_HOLD_GATE_CREATED": False,
        "THRESHOLD_SEARCHED": False,
        "ENTRY_CHANGED": False,
        "EXISTING_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": False,
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "FROZEN_GUARD": dict(FROZEN_GUARD),
        "FROZEN_CONTINUATION": dict(FROZEN_CONTINUATION),
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "P4_4_PRECOMMIT_SHA": precommit.get("SHA"),
            "file_sha_ablation": _file_sha("src/research/early_guard_exact_ablation_p4_4/ablation.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
        "_precommit": precommit,
    }
    unused = CANONICAL_GUARD_N
    del unused
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    b = rep.get("BASELINE") or {}
    o = rep.get("GUARD_OFF") or {}
    d = rep.get("DELTA") or {}
    e = rep.get("CURRENT_EXACT_GUARD_ECONOMICS") or {}
    dest = rep.get("EARLY_GUARD89_DESTINATION") or {}
    p = rep.get("PORTFOLIO_EFFECT") or {}
    rest = rep.get("REST11") or {}
    lines = [
        "# P4-4 Early Guard exact ablation",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / new strategy validation / production approval",
        "No new Guard. No threshold search. No Mid-Hold reopen. No Runtime change.",
        "",
        f"PRECOMMIT_SHA: `{rep.get('PRECOMMIT_SHA')}`",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        f"BASELINE_RECONCILE: `{rep.get('BASELINE_RECONCILE')}`",
        f"EARLY_GUARD_BASELINE_N: {rep.get('EARLY_GUARD_BASELINE_N')}",
        f"GUARD89_MATCH: {rep.get('GUARD89_MATCH')} matched={rep.get('GUARD89_MATCHED_N')}",
        "",
        "BASELINE:",
        f"trades: {b.get('trades')}",
        f"pnl: {b.get('pnl')}",
        f"PF: {b.get('PF')}",
        f"maxDD: {b.get('maxDD')}",
        "",
        "GUARD_OFF:",
        f"trades: {o.get('trades')}",
        f"pnl: {o.get('pnl')}",
        f"PF: {o.get('PF')}",
        f"maxDD: {o.get('maxDD')}",
        "",
        "DELTA (Guard-OFF minus baseline):",
        f"trades: {d.get('trades')}",
        f"pnl: {d.get('pnl')}",
        f"PF: {d.get('PF')}",
        f"maxDD: {d.get('maxDD')}",
        "",
        "CURRENT_EXACT_GUARD_ECONOMICS:",
        f"saved_loss: {e.get('saved_loss')}",
        f"foregone_winner: {e.get('foregone_winner')}",
        f"net_guard_value: {e.get('net_guard_value')}",
        "",
        "EARLY_GUARD89_DESTINATION:",
        f"exit600: {dest.get('exit600')}",
        f"extend750: {dest.get('extend750')}",
        f"session_close: {dest.get('session_close')}",
        f"other: {dest.get('other')}",
        f"lost_fill: {dest.get('lost_fill')}",
        "",
        "PORTFOLIO_EFFECT:",
        f"newly_admitted: {p.get('newly_admitted')}",
        f"newly_filled: {p.get('newly_filled')}",
        f"lost_admission: {p.get('lost_admission')}",
        f"lost_fill: {p.get('lost_fill')}",
        f"same_symbol_changed: {p.get('same_symbol_changed')}",
        f"capacity_changed: {p.get('capacity_changed')}",
        "",
        f"TOP3: {json.dumps(rep.get('TOP3'), ensure_ascii=False, default=str)}",
        "",
        "REST11:",
        f"baseline PnL: {rest.get('baseline_pnl')}  guard-off PnL: {rest.get('guard_off_pnl')}",
        f"baseline PF: {rest.get('baseline_PF')}  guard-off PF: {rest.get('guard_off_PF')}",
        f"baseline maxDD: {rest.get('baseline_maxDD')}  guard-off maxDD: {rest.get('guard_off_maxDD')}",
        "",
        f"TAIL: {json.dumps(rep.get('TAIL'), ensure_ascii=False, default=str)}",
        "",
        "HISTORICAL_REFERENCE:",
        f"saved_loss: {HIST_SAVED_LOSS}",
        f"foregone_winner: {HIST_FOREGONE_WINNER}",
        f"ratio: {HIST_RATIO}",
        f"CURRENT_VS_HISTORICAL_DIRECTION: {rep.get('CURRENT_VS_HISTORICAL_DIRECTION')}",
        "",
        f"CLASSIFICATION: `{rep.get('CLASSIFICATION')}`",
        f"{(rep.get('CLASSIFICATION_DETAIL') or {}).get('why')}",
        "",
        "EARLY_GUARD_CHANGED: false",
        "MID_HOLD_GATE_CREATED: false",
        "THRESHOLD_SEARCHED: false",
        "ENTRY_CHANGED: false",
        "EXISTING_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        "FUTURE_LEAK: false",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    pre = rep.pop("_precommit", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("PRECOMMIT_SHA", public.get("PRECOMMIT_SHA")),
            ("BASELINE_RECONCILE", public.get("BASELINE_RECONCILE")),
            ("EARLY_GUARD_BASELINE_N", public.get("EARLY_GUARD_BASELINE_N")),
            ("GUARD89_MATCH", public.get("GUARD89_MATCH")),
            ("CLASSIFICATION", public.get("CLASSIFICATION")),
            ("EARLY_GUARD_CHANGED", False),
            ("MID_HOLD_GATE_CREATED", False),
            ("THRESHOLD_SEARCHED", False),
            ("RUNTIME_CHANGED", False),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(
        wb.create_sheet("Precommit"),
        [{"key": k, "value": pre.get(k)} for k in ("LABEL", "SHA", "ANALYSIS_ID")]
        + [{"key": "ablation", "value": pre.get("ablation")}, {"key": "classification", "value": pre.get("classification")}],
    )
    _write_rows(wb.create_sheet("Baseline"), sheets.get("baseline") or [])
    _write_rows(wb.create_sheet("Guard_Off"), sheets.get("off") or [])
    _write_rows(wb.create_sheet("EarlyGuard89"), sheets.get("g89") or [])
    _write_rows(wb.create_sheet("Saved_Foregone"), sheets.get("saved") or [])
    _write_rows(wb.create_sheet("Portfolio_Effect"), sheets.get("port") or [])
    _write_rows(wb.create_sheet("Tail"), sheets.get("tail") or [])
    _write_rows(wb.create_sheet("Top3_Rest11"), sheets.get("top3rest") or [])
    _write_rows(wb.create_sheet("Daily"), sheets.get("daily") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(ident, list((public.get("Identity") or {}).items()))
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("EARLY_GUARD_CHANGED", False),
            ("MID_HOLD_GATE_CREATED", False),
            ("THRESHOLD_SEARCHED", False),
            ("ENTRY_CHANGED", False),
            ("EXISTING_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("FUTURE_LEAK", False),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
