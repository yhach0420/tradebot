"""Write P3-1 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.fixed_selection_edge_decomposition_p3_1 import (
    ANALYSIS_ID,
    CLOCK_FULL14,
    CLOCK_REST11,
    DOCUMENT_ID,
    FULL14,
    HORIZONS_SEC,
    PREDECLARED_TOP3,
    REST11,
    SOT_NOT_SELECTED_FILL_N,
    SOT_NOT_SELECTED_N,
    SOT_SELECTED_FILL_N,
    SOT_SELECTED_N,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.fixed_selection_edge_decomposition_p3_1.metrics import (
    directional_block,
    directional_verdict,
    execution_block,
    execution_verdict,
    feature_dist,
    filled_outcome,
    rank_strata,
    same_anchor_rows,
    same_anchor_summary,
    selection_edge,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "fixed_selection_edge_decomposition_p3_1"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 88


def _filt(rows, days):
    ds = set(days)
    return [r for r in rows if str(r.get("date")) in ds]


def _split(rows):
    sel = [r for r in rows if r.get("selected")]
    nos = [r for r in rows if not r.get("selected")]
    return sel, nos


def _dir_pair(rows):
    sel, nos = _split(rows)
    out = {}
    for h in HORIZONS_SEC:
        out[int(h)] = {"SELECTED": directional_block(sel, int(h)), "NOT_SELECTED": directional_block(nos, int(h))}
    return out


def _sa_pair(rows):
    out_sum = {}
    out_det = []
    for h in HORIZONS_SEC:
        det = same_anchor_rows(rows, int(h))
        out_det.extend(det)
        out_sum[int(h)] = same_anchor_summary(det)
    return out_sum, out_det


def _horizon_line(dsel, dnos, sa) -> dict[str, Any]:
    return {
        "selected_mean": dsel.get("mean_return"),
        "not_selected_mean": dnos.get("mean_return"),
        "selected_median": dsel.get("median_return"),
        "not_selected_median": dnos.get("median_return"),
        "selected_n_evaluable": dsel.get("n_evaluable"),
        "not_selected_n_evaluable": dnos.get("n_evaluable"),
        "same_anchor_better": sa.get("selected_better_anchor_n"),
        "same_anchor_worse": sa.get("selected_worse_anchor_n"),
        "same_anchor_equal": sa.get("equal_anchor_n"),
        "same_anchor_median_difference": sa.get("median_difference"),
        "same_anchor_mean_difference": sa.get("mean_difference"),
    }


def build_report(*, rows: list[dict[str, Any]], failed: list[str], blocked: bool, blocked_reason: str = "") -> dict[str, Any]:
    now = datetime.now(JST).isoformat(timespec="seconds")
    full = [r for r in rows if str(r.get("date")) in set(FULL14)]
    top3 = _filt(full, PREDECLARED_TOP3)
    rest = _filt(full, REST11)
    sel, nos = _split(full)
    tsel, tnos = _split(top3)
    rsel, rnos = _split(rest)

    ex_all_s = execution_block(sel)
    ex_all_n = execution_block(nos)
    ex_top_s = execution_block(tsel)
    ex_top_n = execution_block(tnos)
    ex_rest_s = execution_block(rsel)
    ex_rest_n = execution_block(rnos)
    exec_v = execution_verdict(ex_all_s, ex_all_n, ex_rest_s, ex_rest_n)

    dir_all = _dir_pair(full)
    dir_top = _dir_pair(top3)
    dir_rest = _dir_pair(rest)
    sa_all, sa_all_det = _sa_pair(full)
    sa_top, _ = _sa_pair(top3)
    sa_rest, _ = _sa_pair(rest)
    dir_v = directional_verdict(all_dir=dir_all, rest_dir=dir_rest, rest_sa=sa_rest)

    fo_all_s = filled_outcome(sel)
    fo_all_n = filled_outcome(nos)
    fo_top_s = filled_outcome(tsel)
    fo_top_n = filled_outcome(tnos)
    fo_rest_s = filled_outcome(rsel)
    fo_rest_n = filled_outcome(rnos)

    strata, exec_mono, dir_mono = rank_strata(full)
    feats = (
        feature_dist(sel, "SELECTED")
        + feature_dist(nos, "NOT_SELECTED")
        + feature_dist([r for r in full if r.get("independent_filled")], "FILLED")
        + feature_dist([r for r in full if not r.get("independent_filled")], "EXPIRED")
    )

    edge = selection_edge(exec_v, dir_v)
    sot_ok = (
        ex_all_s["n"] == SOT_SELECTED_N
        and ex_all_s["fill_n"] == SOT_SELECTED_FILL_N
        and ex_all_n["n"] == SOT_NOT_SELECTED_N
        and ex_all_n["fill_n"] == SOT_NOT_SELECTED_FILL_N
    )
    if blocked:
        verdict = VERDICT_BLOCKED
    elif not sot_ok and full:
        verdict = VERDICT_BLOCKED
        blocked_reason = blocked_reason or (
            f"SOT_MISMATCH selected={ex_all_s['n']}/{ex_all_s['fill_n']} "
            f"not={ex_all_n['n']}/{ex_all_n['fill_n']}"
        )
        blocked = True
    elif failed:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    ret_pack = {}
    for h in HORIZONS_SEC:
        ret_pack[f"RET{h}"] = _horizon_line(dir_all[h]["SELECTED"], dir_all[h]["NOT_SELECTED"], sa_all[h])
    rest_dir_pack = {}
    top_dir_pack = {}
    for h in HORIZONS_SEC:
        rest_dir_pack[f"RET{h}"] = _horizon_line(dir_rest[h]["SELECTED"], dir_rest[h]["NOT_SELECTED"], sa_rest[h])
        top_dir_pack[f"RET{h}"] = _horizon_line(dir_top[h]["SELECTED"], dir_top[h]["NOT_SELECTED"], sa_top[h])

    report = {
        "task": "P3-1",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "label": TASK_LABEL,
        "not": ["OOS", "prospective", "robust", "strategy validation"],
        "generated_at_jst": now,
        "clock_frozen": {"FULL14": CLOCK_FULL14, "REST11": CLOCK_REST11, "recomputed": False},
        "sot_p3_0r": {
            "SELECTED_n": SOT_SELECTED_N,
            "SELECTED_fill_n": SOT_SELECTED_FILL_N,
            "NOT_SELECTED_n": SOT_NOT_SELECTED_N,
            "NOT_SELECTED_fill_n": SOT_NOT_SELECTED_FILL_N,
            "match": sot_ok,
        },
        "EXECUTION_EDGE": exec_v.replace("EXECUTION_EDGE_", ""),
        "EXECUTION_EDGE_FULL": exec_v,
        "SELECTED_FILL_RATE": ex_all_s["fill_rate"],
        "NOT_SELECTED_FILL_RATE": ex_all_n["fill_rate"],
        "SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS": ex_all_s["median_min_ask_minus_limit_bps"],
        "NOT_SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS": ex_all_n["median_min_ask_minus_limit_bps"],
        "EXECUTION": {
            "ALL": {"SELECTED": ex_all_s, "NOT_SELECTED": ex_all_n},
            "TOP3": {"SELECTED": ex_top_s, "NOT_SELECTED": ex_top_n},
            "REST11": {"SELECTED": ex_rest_s, "NOT_SELECTED": ex_rest_n},
        },
        "DIRECTIONAL_EDGE": dir_v.replace("DIRECTIONAL_EDGE_", ""),
        "DIRECTIONAL_EDGE_FULL": dir_v,
        "RET60": ret_pack["RET60"],
        "RET180": ret_pack["RET180"],
        "RET600": ret_pack["RET600"],
        "REST11_DIRECTIONAL": rest_dir_pack,
        "TOP3_DIRECTIONAL": top_dir_pack,
        "FILLED_OUTCOME": {
            "label": "INDEPENDENT_FILLED_ARCH_E_OUTCOME",
            "exit_caveat": "P3-0R independent Arch E pnl match 185/267, exit_time match 94/267. Descriptive only.",
            "ALL": {"SELECTED": fo_all_s, "NOT_SELECTED": fo_all_n},
            "TOP3": {"SELECTED": fo_top_s, "NOT_SELECTED": fo_top_n},
            "REST11": {"SELECTED": fo_rest_s, "NOT_SELECTED": fo_rest_n},
        },
        "EXECUTION_MONOTONICITY": exec_mono,
        "DIRECTIONAL_MONOTONICITY": dir_mono,
        "RANK_STRATA": strata,
        "SELECTION_EDGE": edge,
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "CLOCK_FULL14": CLOCK_FULL14,
        "CLOCK_REST11": CLOCK_REST11,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_native_entry": _file_sha("src/small_paper/v1r_native_entry_live.py"),
        },
        "_sheets": {
            "execution": [
                {"slice": "ALL", "group": "SELECTED", **ex_all_s},
                {"slice": "ALL", "group": "NOT_SELECTED", **ex_all_n},
                {"slice": "TOP3", "group": "SELECTED", **ex_top_s},
                {"slice": "TOP3", "group": "NOT_SELECTED", **ex_top_n},
                {"slice": "REST11", "group": "SELECTED", **ex_rest_s},
                {"slice": "REST11", "group": "NOT_SELECTED", **ex_rest_n},
            ],
            "dir60": [
                {"slice": "ALL", "group": "SELECTED", **dir_all[60]["SELECTED"]},
                {"slice": "ALL", "group": "NOT_SELECTED", **dir_all[60]["NOT_SELECTED"]},
                {"slice": "TOP3", "group": "SELECTED", **dir_top[60]["SELECTED"]},
                {"slice": "TOP3", "group": "NOT_SELECTED", **dir_top[60]["NOT_SELECTED"]},
                {"slice": "REST11", "group": "SELECTED", **dir_rest[60]["SELECTED"]},
                {"slice": "REST11", "group": "NOT_SELECTED", **dir_rest[60]["NOT_SELECTED"]},
            ],
            "dir180": [
                {"slice": "ALL", "group": "SELECTED", **dir_all[180]["SELECTED"]},
                {"slice": "ALL", "group": "NOT_SELECTED", **dir_all[180]["NOT_SELECTED"]},
                {"slice": "TOP3", "group": "SELECTED", **dir_top[180]["SELECTED"]},
                {"slice": "TOP3", "group": "NOT_SELECTED", **dir_top[180]["NOT_SELECTED"]},
                {"slice": "REST11", "group": "SELECTED", **dir_rest[180]["SELECTED"]},
                {"slice": "REST11", "group": "NOT_SELECTED", **dir_rest[180]["NOT_SELECTED"]},
            ],
            "dir600": [
                {"slice": "ALL", "group": "SELECTED", **dir_all[600]["SELECTED"]},
                {"slice": "ALL", "group": "NOT_SELECTED", **dir_all[600]["NOT_SELECTED"]},
                {"slice": "TOP3", "group": "SELECTED", **dir_top[600]["SELECTED"]},
                {"slice": "TOP3", "group": "NOT_SELECTED", **dir_top[600]["NOT_SELECTED"]},
                {"slice": "REST11", "group": "SELECTED", **dir_rest[600]["SELECTED"]},
                {"slice": "REST11", "group": "NOT_SELECTED", **dir_rest[600]["NOT_SELECTED"]},
            ],
            "same_anchor": sa_all_det,
            "filled": [
                {"slice": "ALL", "group": "SELECTED", **fo_all_s},
                {"slice": "ALL", "group": "NOT_SELECTED", **fo_all_n},
                {"slice": "TOP3", "group": "SELECTED", **fo_top_s},
                {"slice": "TOP3", "group": "NOT_SELECTED", **fo_top_n},
                {"slice": "REST11", "group": "SELECTED", **fo_rest_s},
                {"slice": "REST11", "group": "NOT_SELECTED", **fo_rest_n},
            ],
            "strata": strata,
            "features": feats,
            "top3": [
                {"part": "EXECUTION", "group": "SELECTED", **ex_top_s},
                {"part": "EXECUTION", "group": "NOT_SELECTED", **ex_top_n},
                {"part": "FILLED", "group": "SELECTED", **fo_top_s},
                {"part": "FILLED", "group": "NOT_SELECTED", **fo_top_n},
            ],
            "rest11": [
                {"part": "EXECUTION", "group": "SELECTED", **ex_rest_s},
                {"part": "EXECUTION", "group": "NOT_SELECTED", **ex_rest_n},
                {"part": "FILLED", "group": "SELECTED", **fo_rest_s},
                {"part": "FILLED", "group": "NOT_SELECTED", **fo_rest_n},
            ],
        },
    }
    return json_sanitize(report)


def _ret_md(title: str, block: dict[str, Any]) -> list[str]:
    return [
        f"{title}:",
        f"selected_mean: {block.get('selected_mean')}",
        f"not_selected_mean: {block.get('not_selected_mean')}",
        f"same_anchor_better: {block.get('same_anchor_better')}",
        f"same_anchor_worse: {block.get('same_anchor_worse')}",
        "",
    ]


def render_md(rep: dict[str, Any]) -> str:
    fo = rep["FILLED_OUTCOME"]["ALL"]
    fs = fo["SELECTED"]
    fn = fo["NOT_SELECTED"]
    lines = [
        "# P3-1 Fixed selection edge decomposition",
        "",
        f"**Label:** `{rep['label']}` — not OOS / prospective / robust / strategy validation.",
        "",
        "A = Fill?  B = fill-independent CurrentPrice direction?  C = filled Arch E outcome (secondary).",
        "Filled-trade PnL is not the primary directional evidence.",
        "",
        f"EXECUTION_EDGE: `{rep['EXECUTION_EDGE']}`",
        f"SELECTED_FILL_RATE: {rep['SELECTED_FILL_RATE']}",
        f"NOT_SELECTED_FILL_RATE: {rep['NOT_SELECTED_FILL_RATE']}",
        f"SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS: {rep['SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS']}",
        f"NOT_SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS: {rep['NOT_SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS']}",
        "",
        f"DIRECTIONAL_EDGE: `{rep['DIRECTIONAL_EDGE']}`",
        "",
        *_ret_md("RET60", rep["RET60"]),
        *_ret_md("RET180", rep["RET180"]),
        *_ret_md("RET600", rep["RET600"]),
        "FILLED_OUTCOME:",
        "",
        "SELECTED:",
        f"n: {fs.get('n')}",
        f"mean_pnl: {fs.get('mean_pnl_per_filled')}",
        f"PF: {fs.get('PF')}",
        "",
        "NOT_SELECTED:",
        f"n: {fn.get('n')}",
        f"mean_pnl: {fn.get('mean_pnl_per_filled')}",
        f"PF: {fn.get('PF')}",
        "",
        "REST11_DIRECTIONAL:",
        json.dumps(rep["REST11_DIRECTIONAL"], ensure_ascii=False),
        "",
        "TOP3_DIRECTIONAL:",
        json.dumps(rep["TOP3_DIRECTIONAL"], ensure_ascii=False),
        "",
        f"EXECUTION_MONOTONICITY: {rep['EXECUTION_MONOTONICITY']}",
        f"DIRECTIONAL_MONOTONICITY: {rep['DIRECTIONAL_MONOTONICITY']}",
        "",
        f"SELECTION_EDGE: `{rep['SELECTION_EDGE']}`",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"CLOCK_FULL14: `{rep['CLOCK_FULL14']}`",
        f"CLOCK_REST11: `{rep['CLOCK_REST11']}`",
        "",
        f"verdict: `{rep['verdict']}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("EXECUTION_EDGE", public.get("EXECUTION_EDGE")),
            ("SELECTED_FILL_RATE", public.get("SELECTED_FILL_RATE")),
            ("NOT_SELECTED_FILL_RATE", public.get("NOT_SELECTED_FILL_RATE")),
            ("SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS", public.get("SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS")),
            ("NOT_SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS", public.get("NOT_SELECTED_MEDIAN_MIN_ASK_MINUS_LIMIT_BPS")),
            ("DIRECTIONAL_EDGE", public.get("DIRECTIONAL_EDGE")),
            ("SELECTION_EDGE", public.get("SELECTION_EDGE")),
            ("EXECUTION_MONOTONICITY", public.get("EXECUTION_MONOTONICITY")),
            ("DIRECTIONAL_MONOTONICITY", public.get("DIRECTIONAL_MONOTONICITY")),
            ("CLOCK_FULL14", public.get("CLOCK_FULL14")),
            ("CLOCK_REST11", public.get("CLOCK_REST11")),
            ("NEW_STRATEGY_TESTED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Execution"), sheets.get("execution") or [])
    _write_rows(wb.create_sheet("Directional_60"), sheets.get("dir60") or [])
    _write_rows(wb.create_sheet("Directional_180"), sheets.get("dir180") or [])
    _write_rows(wb.create_sheet("Directional_600"), sheets.get("dir600") or [])
    _write_rows(wb.create_sheet("Same_Anchor"), sheets.get("same_anchor") or [])
    _write_rows(wb.create_sheet("Filled_Outcome"), sheets.get("filled") or [])
    _write_rows(wb.create_sheet("Rank_Strata"), sheets.get("strata") or [])
    _write_rows(wb.create_sheet("Feature_Distributions"), sheets.get("features") or [])
    _write_rows(wb.create_sheet("Top3"), sheets.get("top3") or [])
    _write_rows(wb.create_sheet("Rest11"), sheets.get("rest11") or [])
    _kv_sheet(wb.create_sheet("Identity"), list((public.get("Identity") or {}).items()) + [("label", TASK_LABEL)])
    _kv_sheet(
        wb.create_sheet("Safety"),
        [("submit", 0), ("cancel", 0), ("live", 0), ("NEW_STRATEGY_TESTED", False), ("RUNTIME_CHANGED", False)],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
