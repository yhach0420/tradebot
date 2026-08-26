"""Write P3-2 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.post_fill_edge_decomposition_p3_2 import (
    ANALYSIS_ID,
    CLOCK_FULL14,
    CLOCK_REST11,
    DOCUMENT_ID,
    FULL14,
    HORIZONS_SEC,
    PREDECLARED_TOP3,
    REST11,
    SOT_NOT_SELECTED_FILL_N,
    SOT_SELECTED_FILL_N,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.post_fill_edge_decomposition_p3_2.metrics import (
    dist_block,
    exec_adv_selected_better,
    markout_block,
    mechanism,
    post_fill_direction_verdict,
    rank_strata,
    same_anchor_rows,
    same_anchor_summary,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "post_fill_edge_decomposition_p3_2"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 88


def _filt(rows, days):
    return [r for r in rows if str(r.get("date")) in set(days)]


def _split(rows):
    return [r for r in rows if r.get("selected")], [r for r in rows if not r.get("selected")]


def _pair_mark(rows, field_prefix="mid_markout"):
    sel, nos = _split(rows)
    out = {}
    for h in HORIZONS_SEC:
        out[int(h)] = {
            "SELECTED": markout_block(sel, int(h), f"{field_prefix}_{h}"),
            "NOT_SELECTED": markout_block(nos, int(h), f"{field_prefix}_{h}"),
        }
    return out


def _sa(rows):
    sums, det = {}, []
    for h in HORIZONS_SEC:
        d = same_anchor_rows(rows, int(h))
        det.extend(d)
        sums[int(h)] = same_anchor_summary(d)
    return sums, det


def _line(msel, mnos) -> dict[str, Any]:
    return {
        "selected_mean": msel.get("mean"),
        "not_selected_mean": mnos.get("mean"),
        "selected_median": msel.get("median"),
        "not_selected_median": mnos.get("median"),
        "selected_n_evaluable": msel.get("n_evaluable"),
        "not_selected_n_evaluable": mnos.get("n_evaluable"),
        "selected_positive_rate": msel.get("positive_rate"),
        "not_selected_positive_rate": mnos.get("positive_rate"),
    }


def build_report(
    *,
    rows: list[dict[str, Any]],
    leak_fill: int,
    leak_mid: int,
    leak_cp: int,
    identity_n: int,
    identity_fail: int,
    harvest_selected_n: int,
    harvest_eligible_n: int,
    failed: list[str],
    blocked: bool,
    blocked_reason: str = "",
) -> dict[str, Any]:
    now = datetime.now(JST).isoformat(timespec="seconds")
    full = [r for r in rows if str(r.get("date")) in set(FULL14)]
    top3 = _filt(full, PREDECLARED_TOP3)
    rest = _filt(full, REST11)
    sel, nos = _split(full)
    tsel, tnos = _split(top3)
    rsel, rnos = _split(rest)

    ea_all_s = dist_block(sel, "execution_advantage_bps")
    ea_all_n = dist_block(nos, "execution_advantage_bps")
    ea_top_s = dist_block(tsel, "execution_advantage_bps")
    ea_top_n = dist_block(tnos, "execution_advantage_bps")
    ea_rest_s = dist_block(rsel, "execution_advantage_bps")
    ea_rest_n = dist_block(rnos, "execution_advantage_bps")
    adv_better = exec_adv_selected_better(ea_all_s, ea_all_n, ea_rest_s, ea_rest_n)

    mid_all = _pair_mark(full)
    mid_top = _pair_mark(top3)
    mid_rest = _pair_mark(rest)
    execm_all = _pair_mark(full, "execution_markout")
    sa_all, sa_det = _sa(full)
    sa_top, _ = _sa(top3)
    sa_rest, _ = _sa(rest)

    dir_v = post_fill_direction_verdict(all_mid=mid_all, rest_mid=mid_rest, rest_sa=sa_rest)
    mech = mechanism(exec_adv_better=adv_better, dir_v=dir_v, rest_mid=mid_rest)
    strata, exec_mono, dir_mono = rank_strata(full)

    ident_ok = identity_fail == 0 and identity_n > 0
    leak_ok = int(leak_fill) == 0 and int(leak_mid) == 0 and int(leak_cp) == 0
    sot_ok = len(sel) == SOT_SELECTED_FILL_N and len(nos) == SOT_NOT_SELECTED_FILL_N

    if blocked:
        verdict = VERDICT_BLOCKED
    elif full and not sot_ok:
        verdict = VERDICT_BLOCKED
        blocked_reason = blocked_reason or f"SOT_MISMATCH fills selected={len(sel)} not={len(nos)}"
        blocked = True
    elif full and not ident_ok:
        verdict = VERDICT_BLOCKED
        blocked_reason = blocked_reason or f"DECOMPOSITION_IDENTITY_FAIL n={identity_n} fail={identity_fail}"
        blocked = True
    elif full and not leak_ok:
        verdict = VERDICT_BLOCKED
        blocked_reason = blocked_reason or (
            f"FUTURE_LEAK fill={leak_fill} mid={leak_mid} cp={leak_cp}"
        )
        blocked = True
    elif failed:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    def _rest_h(h: int) -> dict[str, Any]:
        return {
            **_line(mid_rest[h]["SELECTED"], mid_rest[h]["NOT_SELECTED"]),
            "same_anchor_better": sa_rest[h].get("selected_better"),
            "same_anchor_worse": sa_rest[h].get("selected_worse"),
        }

    prefill_cols = [
        "date",
        "anchor_time",
        "symbol",
        "selected",
        "fill_latency_ms",
        "anchor_mid",
        "mid_at_fill",
        "anchor_to_fill_mid_return",
        "anchor_bid",
        "anchor_ask",
        "limit_bid",
        "first_ask_minus_limit_bps",
        "min_ask_minus_limit_bps",
        "spread_bps_at_anchor",
        "spread_bps_at_fill",
        "fill_price",
        "execution_advantage_bps",
    ]

    report = {
        "task": "P3-2",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "label": TASK_LABEL,
        "not": ["OOS", "prospective", "robust", "strategy validation"],
        "generated_at_jst": now,
        "clock_frozen": {"FULL14": CLOCK_FULL14, "REST11": CLOCK_REST11, "recomputed": False},
        "p3_1_frozen": {
            "EXECUTION_EDGE": "SUPPORTED",
            "ANCHOR_DIRECTIONAL_EDGE": "NOT_SUPPORTED",
            "SELECTION_EDGE": "EXECUTION_DOMINANT",
        },
        "sot_filled": {
            "SELECTED_FILLED": SOT_SELECTED_FILL_N,
            "NOT_SELECTED_FILLED": SOT_NOT_SELECTED_FILL_N,
            "observed_selected": len(sel),
            "observed_not_selected": len(nos),
            "harvest_selected_n": harvest_selected_n,
            "harvest_eligible_n": harvest_eligible_n,
            "match": sot_ok,
        },
        "EXECUTION_EDGE": "SUPPORTED",
        "ANCHOR_DIRECTIONAL_EDGE": "NOT_SUPPORTED",
        "EXECUTION_PRICE_ADVANTAGE": {
            "SELECTED": ea_all_s,
            "NOT_SELECTED": ea_all_n,
            "TOP3": {"SELECTED": ea_top_s, "NOT_SELECTED": ea_top_n},
            "REST11": {"SELECTED": ea_rest_s, "NOT_SELECTED": ea_rest_n},
            "selected_median_gt_not_selected_all_and_rest11": adv_better,
        },
        "POST_FILL_DIRECTION": dir_v.replace("POST_FILL_DIRECTION_", ""),
        "POST_FILL_DIRECTION_FULL": dir_v,
        "MID_MARKOUT_1S": _line(mid_all[1]["SELECTED"], mid_all[1]["NOT_SELECTED"]),
        "MID_MARKOUT_10S": _line(mid_all[10]["SELECTED"], mid_all[10]["NOT_SELECTED"]),
        "MID_MARKOUT_60S": _line(mid_all[60]["SELECTED"], mid_all[60]["NOT_SELECTED"]),
        "MID_MARKOUT_180S": _line(mid_all[180]["SELECTED"], mid_all[180]["NOT_SELECTED"]),
        "MID_MARKOUT_600S": _line(mid_all[600]["SELECTED"], mid_all[600]["NOT_SELECTED"]),
        "REST11": {"60s": _rest_h(60), "180s": _rest_h(180), "600s": _rest_h(600)},
        "SAME_ANCHOR": {
            "60s_better_worse": f"{sa_all[60].get('selected_better')}/{sa_all[60].get('selected_worse')}",
            "180s_better_worse": f"{sa_all[180].get('selected_better')}/{sa_all[180].get('selected_worse')}",
            "600s_better_worse": f"{sa_all[600].get('selected_better')}/{sa_all[600].get('selected_worse')}",
            "ALL": {str(h): sa_all[h] for h in HORIZONS_SEC},
            "TOP3": {str(h): sa_top[h] for h in HORIZONS_SEC},
            "REST11": {str(h): sa_rest[h] for h in HORIZONS_SEC},
        },
        "MICROSTRUCTURE_AFTER_FILL": {
            "1s": _line(mid_all[1]["SELECTED"], mid_all[1]["NOT_SELECTED"]),
            "10s": _line(mid_all[10]["SELECTED"], mid_all[10]["NOT_SELECTED"]),
            "note": "Descriptive only. Not a rebound/pullback rule.",
        },
        "DECOMPOSITION_IDENTITY": "PASS" if ident_ok else "FAIL",
        "identity_n": identity_n,
        "identity_fail": identity_fail,
        "EXECUTION_ADVANTAGE_MONOTONICITY": exec_mono,
        "POST_FILL_DIRECTION_MONOTONICITY": dir_mono,
        "POST_FILL_MECHANISM": mech,
        "FUTURE_LEAK": False if leak_ok else True,
        "fill_reference_future_leak": int(leak_fill),
        "future_mid_checkpoint_leak": int(leak_mid),
        "future_currentprice_checkpoint_leak": int(leak_cp),
        "ARCH_E_SECONDARY": {
            "note": "P3-1 descriptive only. Not used in P3-2 primary verdict.",
            "ALL_SELECTED_mean_pnl": 9596.32867132867,
            "ALL_SELECTED_PF": 3.1973099555662303,
            "ALL_NOT_SELECTED_mean_pnl": 6233.505154639175,
            "ALL_NOT_SELECTED_PF": 3.6317736670293796,
            "canonical_fill_reproduced": "267/267",
            "independent_arch_e_pnl_match": "185/267",
            "independent_arch_e_exit_time_match": "94/267",
        },
        "RANK_STRATA": strata,
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "CLOCK_FULL14": CLOCK_FULL14,
        "CLOCK_REST11": CLOCK_REST11,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_native_entry": _file_sha("src/small_paper/v1r_native_entry_live.py"),
        },
        "_sheets": {
            "exec_adv": [
                {"slice": "ALL", "group": "SELECTED", **ea_all_s},
                {"slice": "ALL", "group": "NOT_SELECTED", **ea_all_n},
                {"slice": "TOP3", "group": "SELECTED", **ea_top_s},
                {"slice": "TOP3", "group": "NOT_SELECTED", **ea_top_n},
                {"slice": "REST11", "group": "SELECTED", **ea_rest_s},
                {"slice": "REST11", "group": "NOT_SELECTED", **ea_rest_n},
            ],
            "m1": [
                {"slice": sl, "group": g, **_pair_mark(rs)[1][g]}
                for sl, rs in (("ALL", full), ("TOP3", top3), ("REST11", rest))
                for g in ("SELECTED", "NOT_SELECTED")
            ],
            "m10": [
                {"slice": sl, "group": g, **_pair_mark(rs)[10][g]}
                for sl, rs in (("ALL", full), ("TOP3", top3), ("REST11", rest))
                for g in ("SELECTED", "NOT_SELECTED")
            ],
            "m60": [
                {"slice": sl, "group": g, **_pair_mark(rs)[60][g]}
                for sl, rs in (("ALL", full), ("TOP3", top3), ("REST11", rest))
                for g in ("SELECTED", "NOT_SELECTED")
            ],
            "m180": [
                {"slice": sl, "group": g, **_pair_mark(rs)[180][g]}
                for sl, rs in (("ALL", full), ("TOP3", top3), ("REST11", rest))
                for g in ("SELECTED", "NOT_SELECTED")
            ],
            "m600": [
                {"slice": sl, "group": g, **_pair_mark(rs)[600][g]}
                for sl, rs in (("ALL", full), ("TOP3", top3), ("REST11", rest))
                for g in ("SELECTED", "NOT_SELECTED")
            ],
            "exec_mark": [
                {"slice": "ALL", "group": "SELECTED", "horizon": h, **execm_all[h]["SELECTED"]}
                for h in HORIZONS_SEC
            ]
            + [
                {"slice": "ALL", "group": "NOT_SELECTED", "horizon": h, **execm_all[h]["NOT_SELECTED"]}
                for h in HORIZONS_SEC
            ],
            "same_anchor": sa_det,
            "prefill": [{k: r.get(k) for k in prefill_cols} for r in full],
            "strata": strata,
            "top3": [
                {"part": "EXEC_ADV", "group": "SELECTED", **ea_top_s},
                {"part": "EXEC_ADV", "group": "NOT_SELECTED", **ea_top_n},
            ],
            "rest11": [
                {"part": "EXEC_ADV", "group": "SELECTED", **ea_rest_s},
                {"part": "EXEC_ADV", "group": "NOT_SELECTED", **ea_rest_n},
            ],
        },
    }
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    ea_s = rep["EXECUTION_PRICE_ADVANTAGE"]["SELECTED"]
    ea_n = rep["EXECUTION_PRICE_ADVANTAGE"]["NOT_SELECTED"]
    sa = rep["SAME_ANCHOR"]
    rest = rep["REST11"]

    def _m(title, b):
        return [
            f"{title}:",
            f"selected_mean: {b.get('selected_mean')}",
            f"not_selected_mean: {b.get('not_selected_mean')}",
            "",
        ]

    lines = [
        "# P3-2 Post-fill edge decomposition",
        "",
        f"**Label:** `{rep['label']}` — not OOS / prospective / robust / strategy validation.",
        "",
        "A = execution-relative markout (includes fill price).",
        "B = MID→MID post-fill direction (primary directional).",
        "C = fill vs mid at fill (execution price advantage).",
        "Filled Arch E PnL is secondary and not used in the primary verdict.",
        "",
        "EXECUTION_EDGE: SUPPORTED",
        "ANCHOR_DIRECTIONAL_EDGE: NOT_SUPPORTED",
        "",
        "EXECUTION_PRICE_ADVANTAGE:",
        "",
        "SELECTED:",
        f"median_bps: {ea_s.get('median')}",
        f"mean_bps: {ea_s.get('mean')}",
        "",
        "NOT_SELECTED:",
        f"median_bps: {ea_n.get('median')}",
        f"mean_bps: {ea_n.get('mean')}",
        "",
        f"POST_FILL_DIRECTION: `{rep['POST_FILL_DIRECTION']}`",
        "",
        *_m("MID_MARKOUT_1S", rep["MID_MARKOUT_1S"]),
        *_m("MID_MARKOUT_10S", rep["MID_MARKOUT_10S"]),
        *_m("MID_MARKOUT_60S", rep["MID_MARKOUT_60S"]),
        *_m("MID_MARKOUT_180S", rep["MID_MARKOUT_180S"]),
        *_m("MID_MARKOUT_600S", rep["MID_MARKOUT_600S"]),
        "REST11:",
        f"60s: selected_mean={rest['60s'].get('selected_mean')} not_selected_mean={rest['60s'].get('not_selected_mean')} better/worse={rest['60s'].get('same_anchor_better')}/{rest['60s'].get('same_anchor_worse')}",
        f"180s: selected_mean={rest['180s'].get('selected_mean')} not_selected_mean={rest['180s'].get('not_selected_mean')} better/worse={rest['180s'].get('same_anchor_better')}/{rest['180s'].get('same_anchor_worse')}",
        f"600s: selected_mean={rest['600s'].get('selected_mean')} not_selected_mean={rest['600s'].get('not_selected_mean')} better/worse={rest['600s'].get('same_anchor_better')}/{rest['600s'].get('same_anchor_worse')}",
        "",
        "SAME_ANCHOR:",
        f"60s better/worse: {sa.get('60s_better_worse')}",
        f"180s better/worse: {sa.get('180s_better_worse')}",
        f"600s better/worse: {sa.get('600s_better_worse')}",
        "",
        f"DECOMPOSITION_IDENTITY: `{rep['DECOMPOSITION_IDENTITY']}`",
        f"EXECUTION_ADVANTAGE_MONOTONICITY: {rep['EXECUTION_ADVANTAGE_MONOTONICITY']}",
        f"POST_FILL_DIRECTION_MONOTONICITY: {rep['POST_FILL_DIRECTION_MONOTONICITY']}",
        f"POST_FILL_MECHANISM: `{rep['POST_FILL_MECHANISM']}`",
        "",
        f"FUTURE_LEAK: {str(rep['FUTURE_LEAK']).lower()}",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep['verdict']}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("EXECUTION_EDGE", "SUPPORTED"),
            ("ANCHOR_DIRECTIONAL_EDGE", "NOT_SUPPORTED"),
            ("POST_FILL_DIRECTION", public.get("POST_FILL_DIRECTION")),
            ("POST_FILL_MECHANISM", public.get("POST_FILL_MECHANISM")),
            ("DECOMPOSITION_IDENTITY", public.get("DECOMPOSITION_IDENTITY")),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("EXECUTION_ADVANTAGE_MONOTONICITY", public.get("EXECUTION_ADVANTAGE_MONOTONICITY")),
            ("POST_FILL_DIRECTION_MONOTONICITY", public.get("POST_FILL_DIRECTION_MONOTONICITY")),
            ("NEW_STRATEGY_TESTED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Execution_Advantage"), sheets.get("exec_adv") or [])
    _write_rows(wb.create_sheet("MidMarkout_1s"), sheets.get("m1") or [])
    _write_rows(wb.create_sheet("MidMarkout_10s"), sheets.get("m10") or [])
    _write_rows(wb.create_sheet("MidMarkout_60s"), sheets.get("m60") or [])
    _write_rows(wb.create_sheet("MidMarkout_180s"), sheets.get("m180") or [])
    _write_rows(wb.create_sheet("MidMarkout_600s"), sheets.get("m600") or [])
    _write_rows(wb.create_sheet("Execution_Markout"), sheets.get("exec_mark") or [])
    _write_rows(wb.create_sheet("Same_Anchor"), sheets.get("same_anchor") or [])
    _write_rows(wb.create_sheet("PreFill_Path"), sheets.get("prefill") or [])
    _write_rows(wb.create_sheet("Rank_Strata"), sheets.get("strata") or [])
    _write_rows(wb.create_sheet("Top3"), sheets.get("top3") or [])
    _write_rows(wb.create_sheet("Rest11"), sheets.get("rest11") or [])
    _kv_sheet(wb.create_sheet("Identity"), list((public.get("Identity") or {}).items()) + [("label", TASK_LABEL)])
    _kv_sheet(
        wb.create_sheet("Safety"),
        [("submit", 0), ("cancel", 0), ("live", 0), ("NEW_STRATEGY_TESTED", False), ("RUNTIME_CHANGED", False)],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
