"""Write P2-0B precommit artifacts. No Capture. No PnL."""
from __future__ import annotations

import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import (
    ANCHOR_SHA,
    CHECKPOINT_AGE_STATUS,
    CHECKPOINT_INTERVAL_SEC,
    CHECKPOINT_N,
    CONFIRMATION_STATUS,
    DOCUMENT_ID,
    ENTRY_SHA,
    EXIT_SHA,
    GRID_SEC,
    MAX_CHECKPOINT_AGE_SEC,
    P2_0_VERDICT,
    PRICE_SOURCE,
    SELECTED_CONFIRMATION,
    SELECTED_TRIGGER,
    STRATEGY_SHA,
    THRESHOLD_SOURCE,
    TRIGGER_EDGE,
    TRIGGER_EVALUATION_CLOCK,
    VERDICT_BLOCKED,
    VERDICT_PASS,
    VOLUME_PERCENTILE_MIN,
)
from .synthetic import extra_contract_checks, run_suite

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "dynamic_anchor_confirmation_precommit_p2_0b"
JST = timezone(timedelta(hours=9))

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_report() -> dict[str, Any]:
    suite = run_suite()
    extra = extra_contract_checks()
    extra_fail = [r for r in extra if not r["ok"]]
    all_ok = suite["failed"] == 0 and not extra_fail
    now = datetime.now(JST).isoformat(timespec="seconds")
    expr = "OLS_slope(log(Pk/P0), k=0..10) > 0 AND P10 > P0"
    return {
        "task": "P2-0B",
        "analysis_id": "P2_0B_DYNAMIC_ANCHOR_CONFIRMATION_PRECOMMIT",
        "document_id": DOCUMENT_ID,
        "generated_at_jst": now,
        "p2_0_verdict_unchanged": P2_0_VERDICT,
        "SELECTED_TRIGGER": SELECTED_TRIGGER,
        "TRIGGER_EVALUATION_CLOCK": TRIGGER_EVALUATION_CLOCK,
        "TRIGGER_EDGE": TRIGGER_EDGE,
        "TRIGGER_THRESHOLD": VOLUME_PERCENTILE_MIN,
        "TRIGGER_THRESHOLD_SOURCE": THRESHOLD_SOURCE,
        "TRIGGER_STATUS": "PREVIOUSLY_RESEARCHED_BUT_OVERLAPPING",
        "TRIGGER_GRID_SEC": GRID_SEC,
        "SELECTED_CONFIRMATION": SELECTED_CONFIRMATION,
        "CONFIRMATION_STATUS": CONFIRMATION_STATUS,
        "ANCHOR_SCOPE": "SYMBOL_SPECIFIC",
        "TRIGGER_CONFIRM_ENTRY_SYMBOL_SAME": True,
        "PRICE_SOURCE": PRICE_SOURCE,
        "CHECKPOINTS": CHECKPOINT_N,
        "CHECKPOINT_INTERVAL_SEC": CHECKPOINT_INTERVAL_SEC,
        "MAX_CHECKPOINT_AGE_SEC": MAX_CHECKPOINT_AGE_SEC,
        "CHECKPOINT_AGE_STATUS": CHECKPOINT_AGE_STATUS,
        "CONFIRMATION_EXPRESSION": expr,
        "CONFIRMATION_WINDOW": "[t0, t1] with t1 = t0 + 10 continuous market minutes",
        "TUNED_THRESHOLD_USED": False,
        "HISTORICAL_CAPTURE_READ": False,
        "HISTORICAL_PNL_READ": False,
        "SYNTHETIC_TESTS": {
            "passed": suite["passed"],
            "failed": suite["failed"],
            "n": suite["n"],
            "results": suite["results"],
            "extra_contract_checks": extra,
            "extra_failed": extra_fail,
        },
        "FUTURE_LEAK": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "CAN_PROCEED_TO_P2_1_EVENT_VALIDATION": "YES" if all_ok else "NO",
        "verdict": VERDICT_PASS if all_ok else VERDICT_BLOCKED,
        "identity": {
            "strategy_sha": STRATEGY_SHA,
            "entry_sha": ENTRY_SHA,
            "exit_sha": EXIT_SHA,
            "anchor_sha": ANCHOR_SHA,
            "runtime_modules_unchanged": [
                "V1RNativeEntryLive",
                "V1RLiveDualLane",
                "Fixed Anchor",
                "ENTRY",
                "EXIT",
                "Universe",
            ],
        },
        "contamination": {
            "trigger": "PREVIOUSLY_RESEARCHED_BUT_OVERLAPPING",
            "confirmation": CONFIRMATION_STATUS,
            "complete_dynamic_rule": "NEW_PRECOMMITTED_EXPLORATORY ON REUSED HISTORICAL PERIOD",
            "do_not_call_after_0721_0821_apply": [
                "clean holdout",
                "true OOS",
                "prospective",
            ],
        },
        "p1_baseline_untouched": True,
        "forbidden_this_task": [
            "Historical Dynamic simulation",
            "Historical trigger count",
            "Historical confirmation count",
            "PnL",
            "PF",
            "Fixed vs Dynamic comparison",
            "threshold sweep",
            "grid search",
            "feature search",
            "Runtime adoption",
        ],
    }


def _md(report: dict[str, Any]) -> str:
    st = report["SYNTHETIC_TESTS"]
    rows = "\n".join(
        f"| {r['id']} | {'PASS' if r['ok'] else 'FAIL'} | {r.get('detail')} |"
        for r in st["results"]
    )
    return f"""# P2-0B 10-min pre-entry confirmation precommit

- generated_at_jst: `{report['generated_at_jst']}`
- **verdict: `{report['verdict']}`**
- CAN_PROCEED_TO_P2_1_EVENT_VALIDATION: **{report['CAN_PROCEED_TO_P2_1_EVENT_VALIDATION']}**
- submit/cancel/live: `0/0/0`
- HISTORICAL_CAPTURE_READ: `false` · HISTORICAL_PNL_READ: `false`

P2-0 remains `{P2_0_VERDICT}`. This task only precommits confirmation + ownership semantics.

## Final report

| field | value |
|---|---|
| SELECTED_TRIGGER | {SELECTED_TRIGGER} |
| TRIGGER_EVALUATION_CLOCK | {TRIGGER_EVALUATION_CLOCK} |
| TRIGGER_EDGE | {TRIGGER_EDGE} |
| SELECTED_CONFIRMATION | {SELECTED_CONFIRMATION} |
| CONFIRMATION_STATUS | {CONFIRMATION_STATUS} |
| ANCHOR_SCOPE | SYMBOL_SPECIFIC |
| TRIGGER_CONFIRM_ENTRY_SYMBOL_SAME | true |
| PRICE_SOURCE | {PRICE_SOURCE} |
| CHECKPOINTS | {CHECKPOINT_N} |
| CHECKPOINT_INTERVAL_SEC | {CHECKPOINT_INTERVAL_SEC} |
| MAX_CHECKPOINT_AGE_SEC | {MAX_CHECKPOINT_AGE_SEC} |
| CHECKPOINT_AGE_STATUS | {CHECKPOINT_AGE_STATUS} |
| CONFIRMATION_EXPRESSION | {report['CONFIRMATION_EXPRESSION']} |
| TUNED_THRESHOLD_USED | false |
| HISTORICAL_CAPTURE_READ | false |
| HISTORICAL_PNL_READ | false |
| SYNTHETIC_TESTS passed | {st['passed']} |
| SYNTHETIC_TESTS failed | {st['failed']} |
| FUTURE_LEAK | false |
| STRATEGY_CHANGED | false |
| ENTRY_EXIT_CHANGED | false |
| SAFETY | submit/cancel/live=0/0/0 |

## Trigger (unchanged from P2-0, cadence frozen)

T1 is **not** tick/event-driven.

- Clock: E1_X14 causal **10-second** grid (`GRID_SEC={GRID_SEC}`).
- `raw(g)` uses the existing X14 feature builder fields only:
  `feature_status==OK AND relative_status==OK AND rs_universe_n>=20 AND finite(volume_percentile_60s) AND volume_percentile_60s >= {VOLUME_PERCENTILE_MIN}`
- Missing → FALSE. No imputation.
- t0 = previous grid raw FALSE, current grid raw TRUE.
- Persisting TRUE does not re-fire. Peer ranking changes do not fire symbol A unless A's grid `raw` edges.

## Confirmation `C1_POSITIVE_TREND_10M_V1`

Status: **NEW_PRECOMMITTED_EXPLORATORY** (not PRE_FROZEN, not historically selected).

- Price: **CurrentPrice** only. Fill / bid executable / EXIT / PnL forbidden.
- Window `[t0, t1]`, `t1 = t0 + 600s` continuous market minutes, lunch not crossed.
- AM latest t0 **11:20** (completes 11:30). 11:21 → SESSION_INCOMPLETE.
- PM latest t0 **14:50** (completes 15:00). 14:51 → SESSION_INCOMPLETE.
- 11 checkpoints every 60s. As-of last CurrentPrice with `event_time <= checkpoint` (and `<= t1`).
- Age `checkpoint - price_event_time <= 60s`. Else CHECKPOINT_STALE → CONFIRMATION_NOT_EVALUABLE.
- 60s age cap is **NEW_PRECOMMITTED_OPERATIONAL_RULE**, not an alpha threshold tuned on PnL.
- All 11 prices valid, then:
  - `yk = log(Pk/P0)`, `k=0..10`
  - `trend_slope = OLS slope(k, yk)`
  - PASS iff `trend_slope > 0 AND P10 > P0`
- No extra bps, MFE/MAE, VWAP, or volume persistence gates.

Why both terms: endpoint-only would pass a last-tick spike after a fall; slope-only would pass a rise that finishes at or below P0.

Confirmation freezes at t1. Events after t1 are unused.

## Decision fire

First **global** market event with `event_t > t1` may schedule evaluation (same idea as Fixed Anchor `now_t > t0`). Pre-entry features for symbol A still use **state timestamp <= t1** only.

## Ownership

`dynamic_anchor.symbol` is the T1 symbol. Confirmation and ENTRY candidate are that same symbol. Universe re-ranking onto a different name is forbidden. Later admission / POSITION_CAP / PENDING / Passive Fill / EXIT remain Current Runtime (not changed here).

## Rearm

DISARMED → (raw FALSE) ARMED → (FALSE→TRUE) ANCHOR_ACTIVE → 10 min → CONFIRMED / REJECTED / SESSION_INCOMPLETE / NOT_EVALUABLE → DISARMED.

Re-arm requires TRUE→FALSE→TRUE. No time cooldown.

## Contamination

- Trigger: PREVIOUSLY_RESEARCHED_BUT_OVERLAPPING
- Confirmation: NEW_PRECOMMITTED_EXPLORATORY
- Complete rule: NEW_PRECOMMITTED_EXPLORATORY ON REUSED HISTORICAL PERIOD
- Applying later to 20260721–20260821 is **not** clean holdout / true OOS / prospective.

## Synthetic A–O

| id | result | detail |
|---|---|---|
{rows}

## STOP

P2-0B ends here. Do not apply to Historical Capture in this task. Runtime is unchanged.
"""


def write_artifacts() -> dict[str, Path]:
    OUT.mkdir(parents=True, exist_ok=True)
    # keep only the three allowed outputs
    for p in OUT.iterdir():
        if p.is_file() and p.name not in {"report.json", "report.md", "audit.xlsx"}:
            p.unlink()
    report = build_report()
    json_path = OUT / "report.json"
    md_path = OUT / "report.md"
    xlsx_path = OUT / "audit.xlsx"
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    md_path.write_text(_md(report), encoding="utf-8")

    wb = Workbook()

    def sh(name: str):
        if name == "Summary":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    summary = [
        {"field": k, "value": report[k]}
        for k in (
            "SELECTED_TRIGGER",
            "TRIGGER_EVALUATION_CLOCK",
            "TRIGGER_EDGE",
            "SELECTED_CONFIRMATION",
            "CONFIRMATION_STATUS",
            "ANCHOR_SCOPE",
            "TRIGGER_CONFIRM_ENTRY_SYMBOL_SAME",
            "PRICE_SOURCE",
            "CHECKPOINTS",
            "CHECKPOINT_INTERVAL_SEC",
            "MAX_CHECKPOINT_AGE_SEC",
            "CHECKPOINT_AGE_STATUS",
            "CONFIRMATION_EXPRESSION",
            "TUNED_THRESHOLD_USED",
            "HISTORICAL_CAPTURE_READ",
            "HISTORICAL_PNL_READ",
            "FUTURE_LEAK",
            "STRATEGY_CHANGED",
            "ENTRY_EXIT_CHANGED",
            "SAFETY",
            "CAN_PROCEED_TO_P2_1_EVENT_VALIDATION",
            "verdict",
        )
    ]
    summary.append({"field": "SYNTHETIC_TESTS.passed", "value": report["SYNTHETIC_TESTS"]["passed"]})
    summary.append({"field": "SYNTHETIC_TESTS.failed", "value": report["SYNTHETIC_TESTS"]["failed"]})
    _write_rows(sh("Summary"), summary)

    _write_rows(sh("Trigger"), [
        {"item": "candidate_id", "value": SELECTED_TRIGGER},
        {"item": "clock", "value": TRIGGER_EVALUATION_CLOCK},
        {"item": "grid_sec", "value": GRID_SEC},
        {"item": "edge", "value": TRIGGER_EDGE},
        {"item": "threshold", "value": VOLUME_PERCENTILE_MIN},
        {"item": "threshold_source", "value": THRESHOLD_SOURCE},
        {"item": "status", "value": "PREVIOUSLY_RESEARCHED_BUT_OVERLAPPING"},
        {"item": "missing", "value": "FALSE, no imputation"},
        {"item": "tick_driven_percentile", "value": "forbidden"},
        {"item": "peer_event_ranking_fire", "value": "forbidden"},
    ])
    _write_rows(sh("Confirmation"), [
        {"item": "candidate_id", "value": SELECTED_CONFIRMATION},
        {"item": "status", "value": CONFIRMATION_STATUS},
        {"item": "price_source", "value": PRICE_SOURCE},
        {"item": "window", "value": "[t0, t1]"},
        {"item": "expression", "value": report["CONFIRMATION_EXPRESSION"]},
        {"item": "tuned_threshold", "value": False},
        {"item": "age_rule", "value": CHECKPOINT_AGE_STATUS},
        {"item": "max_age_sec", "value": MAX_CHECKPOINT_AGE_SEC},
        {"item": "forbidden_inputs", "value": "fill/bid executable/EXIT/PnL/MFE/MAE"},
    ])
    _write_rows(sh("Session"), [
        {"rule": "AM_LATEST_VALID_T0", "value": "11:20"},
        {"rule": "AM_COMPLETE", "value": "11:30"},
        {"rule": "PM_LATEST_VALID_T0", "value": "14:50"},
        {"rule": "PM_COMPLETE", "value": "15:00"},
        {"rule": "LUNCH", "value": "not continuous market; no span"},
        {"rule": "horizon_sec", "value": 600},
    ])
    _write_rows(sh("Rearm"), [
        {"state": "DISARMED", "meaning": "start / after close; wait for raw FALSE"},
        {"state": "ARMED", "meaning": "raw FALSE seen; wait FALSE→TRUE"},
        {"state": "ANCHOR_ACTIVE", "meaning": "t0 fired; 10min observe; no ENTRY"},
        {"state": "CONFIRMED/REJECTED/SESSION_INCOMPLETE/NOT_EVALUABLE", "meaning": "then DISARMED"},
        {"state": "rearm", "meaning": "TRUE→FALSE→TRUE required; no time cooldown"},
    ])
    _write_rows(sh("Ownership"), [
        {"rule": "ANCHOR_SCOPE", "value": "SYMBOL_SPECIFIC"},
        {"rule": "trigger_confirm_entry", "value": "same symbol"},
        {"rule": "universe_rerank", "value": "forbidden"},
        {"rule": "later_admission", "value": "Current Runtime (unchanged this task)"},
    ])
    _write_rows(sh("DecisionFire"), [
        {"rule": "scheduler", "value": "first global event event_t > t1"},
        {"rule": "snapshot", "value": "state timestamp <= t1"},
        {"rule": "no_backflow", "value": "event_t > t1 must not enter pre-entry features"},
    ])
    _write_rows(sh("Synthetic"), [
        {"id": r["id"], "ok": r["ok"], "detail": r.get("detail")}
        for r in report["SYNTHETIC_TESTS"]["results"]
    ])
    _write_rows(sh("Contamination"), [
        {"item": k, "value": v} for k, v in report["contamination"].items()
    ])
    _write_rows(sh("Identity"), [
        {"field": k, "value": v} for k, v in report["identity"].items()
    ])
    _write_rows(sh("Safety"), [
        {"field": "submit/cancel/live", "value": "0/0/0"},
        {"field": "capture_read", "value": False},
        {"field": "pnl_read", "value": False},
        {"field": "runtime_changed", "value": False},
        {"field": "forbidden", "value": report["forbidden_this_task"]},
    ])
    wb.save(xlsx_path)
    return {"report_json": json_path, "report_md": md_path, "audit_xlsx": xlsx_path}
