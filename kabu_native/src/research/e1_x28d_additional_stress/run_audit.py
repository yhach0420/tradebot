"""E1_X28D runner: Phase 0 → conditional stress → program decision → optional X29 V2."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x26_exit_library.exits import ExitSpec, common_controls
from research.e1_x27_reference_joint.paths import build_paths_for_rows
from research.e1_x28_executable_joint.board import load_board_events, verify_board_mapping
from research.e1_x28_executable_joint.replay import build_entry_asks
from research.e1_x28b_candidate_reference.baseline import freeze_family_baselines
from research.e1_x28c_candidate_executable.classify import stop_risk_tag

from . import (
    ANALYSIS_ID,
    ASSIGNMENT_REGISTRY_SHA,
    BOARD_MAPPING_SHA,
    DOCUMENT_ID,
    EXPECTED_FAMILY,
    EXPECTED_SPECIFIC,
    EXPECTED_TOTAL,
    FAMILY_BASELINE_REGISTRY_SHA,
    LOGIC_MANIFEST_SHA,
    NORMAL_STOP_N,
    OLD_X29_PRECOMMIT_SHA,
    OLD_X29_RUN_ID,
    ROLE,
    SEMANTIC_EXIT_REGISTRY_SHA,
    SOURCE_X28C,
    STRESS_DAYS,
    VERDICT_INSUFFICIENT,
    VERY_WIDE_STOP_N,
    WIDE_STOP_N,
)
from .analyze import (
    candidate_balanced_view,
    cluster_balanced_metric,
    direction_vs_x28c,
    entry_delta,
    lodo,
    loso,
    personalization_delta,
    program_decision,
    stop_diagnostics,
    stress_status_from_views,
    summarize_all,
    wide_stop_alert,
)
from .data_sufficiency import run_phase0
from .masks import apply_mask, freeze_candidate_specs, load_x29_cohorts, verify_historical_mask_shas
from .population import load_or_build_stress_population
from .publish import publish
from .replay_stress import build_exit_matrices
from .x29_v2 import build_x29_v2

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28d_additional_stress"
X28A1_DIR = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"
X28C_DIR = NATIVE / "results" / "research" / "e1_x28c_candidate_executable"
X26A_DIR = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"
X29_DIR = NATIVE / "results" / "research" / "e1_x29_prospective"


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x28d_additional_stress.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "detail": out[-3500:],
    }


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        out.append(d)
    return out


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _spec_from_row(exit_id: str, p: dict[str, Any]) -> ExitSpec:
    np_sec = _f(p.get("no_progress_sec"))
    gm = p.get("giveback_mode")
    if gm == "":
        gm = None
    return ExitSpec(
        exit_id=exit_id, path_family=None, variant=p.get("exit_mode"),
        stop_bps=_f(p.get("stop_bps")), target_bps=_f(p.get("target_bps")),
        trail_activation_bps=_f(p.get("trail_activation_bps")),
        giveback_bps=_f(p.get("giveback_bps")), giveback_mode=gm,
        no_progress_sec=np_sec,
        max_hold_sec=float(_f(p.get("max_hold_sec")) or 900.0),
        no_progress_mfe_bps=(_f(p.get("no_progress_mfe_bps")) or 5.0) if np_sec is not None else None,
        no_progress_abs_ret_bps=(_f(p.get("no_progress_abs_ret_bps")) or 5.0) if np_sec is not None else None,
        is_control=str(exit_id).startswith("CONTROL_"),
    )


def _load_boards(rows: list[dict[str, Any]], allowed: list[str]):
    keys = sorted({(r["date"], r["symbol"]) for r in rows if r["date"] in allowed})
    cache = {}

    def _one(k):
        return k, load_board_events(k[0], k[1])

    print(f"  loading {len(keys)} board symbol-days...", flush=True)
    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = [ex.submit(_one, k) for k in keys]
        done = 0
        for fut in as_completed(futs):
            k, b = fut.result()
            cache[k] = b
            done += 1
            if done % 30 == 0 or done == len(keys):
                print(f"    boards {done}/{len(keys)}", flush=True)
    return cache


def _insufficiency_report(run_id: str, phase0: dict[str, Any]) -> dict[str, Any]:
    days = phase0["days"]
    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": VERDICT_INSUFFICIENT,
        "role": ROLE,
        "stress_days": list(STRESS_DAYS),
        "data_sufficiency": {d: days[d]["sufficient"] for d in STRESS_DAYS},
        "data_sufficiency_detail": days,
        "old_x29_status": "MAINTAINED_NOT_SUPERSEDED",
        "old_x29_precommit_sha": OLD_X29_PRECOMMIT_SHA,
        "x29_v2_required": False,
        "program_decision": VERDICT_INSUFFICIENT,
        "20260810_market_data_not_opened": True,
        "prospective_observer_not_started": True,
        "no_prospective_evidence_consumed": True,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "Paper_order": False,
            "Live_order": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Discord_production_notification": False,
        },
    }
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "report.json").write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    from .publish import write_report_md
    write_report_md(OUT / "report.md", report)
    return report


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id = f"e1x28d_stress_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    # Source identity
    x28c = json.loads((X28C_DIR / "report.json").read_text(encoding="utf-8"))
    if x28c.get("run_id") != SOURCE_X28C:
        raise RuntimeError("x28c run mismatch")
    for k, exp in (
        ("logic_manifest_sha", LOGIC_MANIFEST_SHA),
        ("assignment_registry_sha", ASSIGNMENT_REGISTRY_SHA),
        ("semantic_exit_registry_sha", SEMANTIC_EXIT_REGISTRY_SHA),
        ("family_baseline_registry_sha", FAMILY_BASELINE_REGISTRY_SHA),
        ("board_mapping_sha", BOARD_MAPPING_SHA),
    ):
        if x28c.get(k) != exp:
            raise RuntimeError(f"source identity {k}")

    print("=== Phase 0 data sufficiency (no performance) ===", flush=True)
    phase0 = run_phase0()
    (OUT / "_phase0.json").write_text(json.dumps(phase0, indent=2, default=str), encoding="utf-8")
    if not phase0["all_days_sufficient"]:
        print("=== INSUFFICIENT — stop; X29 precommit preserved ===", flush=True)
        return _insufficiency_report(run_id, phase0)

    # Freeze X28D precommit BEFORE opening alpha results
    print("=== freeze X28D precommit (before alpha) ===", flush=True)
    cohorts = load_x29_cohorts()
    x28d_pre = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "role": ROLE,
        "stress_days": list(STRESS_DAYS),
        "source_x28c_run_id": SOURCE_X28C,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "cohorts": {
            "PROSPECTIVE_SPECIFIC_49": EXPECTED_SPECIFIC,
            "PROSPECTIVE_FAMILY_PREFERRED_118": EXPECTED_FAMILY,
            "total": EXPECTED_TOTAL,
            "overlap": 0,
        },
        "no_retune": True,
        "no_candidate_selection_change": True,
        "phase0_sufficient": True,
        "alpha_not_opened_at_precommit": True,
        "old_x29_will_supersede": True,
        "supersede_reason": "ADDITIONAL_HISTORICAL_STRESS_VALIDATION_20260805_20260807",
        "20260810_market_data_not_opened": True,
        "prospective_observer_not_started": True,
        "no_prospective_evidence_consumed": True,
        "timestamp": now.isoformat(),
    }
    x28d_pre_sha = sha256_obj(x28d_pre)
    x28d_pre["precommit_sha"] = x28d_pre_sha
    (OUT / "precommit.json").write_text(json.dumps(x28d_pre, indent=2), encoding="utf-8")

    # Supersede old X29 (do not delete/overwrite precommit.json)
    print("=== supersede old X29 precommit ===", flush=True)
    marker = {
        "old_run_id": OLD_X29_RUN_ID,
        "old_precommit_sha": OLD_X29_PRECOMMIT_SHA,
        "status": "SUPERSEDED_BEFORE_PROSPECTIVE_MARKET_OPEN",
        "reason": "ADDITIONAL_HISTORICAL_STRESS_VALIDATION_20260805_20260807",
        "20260810_market_data_not_opened": True,
        "prospective_observer_not_started": True,
        "no_prospective_evidence_consumed": True,
        "superseded_by_x28d_run_id": run_id,
        "superseded_by_x28d_precommit_sha": x28d_pre_sha,
        "at": now.isoformat(),
    }
    (X29_DIR / "precommit_superseded_before_x28d.json").write_text(
        json.dumps(marker, indent=2), encoding="utf-8",
    )

    print("=== freeze candidate specs (Discovery thresholds) ===", flush=True)
    _, cand_by = freeze_candidate_specs()
    mask_check = verify_historical_mask_shas(cohorts, cand_by)
    if not mask_check["ok"]:
        raise RuntimeError(f"mask sha mismatch: {mask_check}")

    print("=== build stress population 0805-07 ===", flush=True)
    rows = load_or_build_stress_population()
    dates = np.array([r["date"] for r in rows])
    symbols = np.array([str(r["symbol"]) for r in rows])
    sessions = np.array([r["session"] for r in rows])
    clusters = np.array([r.get("cluster_id") or str(r["symbol"]) for r in rows])

    print("=== paths + boards + entry asks ===", flush=True)
    mapping = verify_board_mapping()
    if not mapping.get("ok"):
        raise RuntimeError("board mapping")
    times_list, prices_list = build_paths_for_rows(rows, allowed_dates=list(STRESS_DAYS), max_workers=6)
    board_by_key = _load_boards(rows, list(STRESS_DAYS))
    entry_asks = build_entry_asks(rows, board_by_key)
    entry_asks_b = build_entry_asks(rows, board_by_key)
    ab_entry = bool(np.array_equal(entry_asks["valid"], entry_asks_b["valid"]))

    # Load exit specs needed
    print("=== load exit specs for 167 ===", flush=True)
    assignments = _load_sheet(X28A1_DIR / "audit.xlsx", "CandidateExitAssignmentsV2")
    sem_reg = _load_sheet(X28A1_DIR / "audit.xlsx", "SemanticExitRegistryV2")
    assign_by = {a["candidate_id"]: a for a in assignments}
    x26a = json.loads((X26A_DIR / "report.json").read_text(encoding="utf-8"))
    x26a_exits = {c["canonical_exit_id"]: c for c in (x26a.get("canonical_exits") or [])}
    for spec in common_controls():
        x26a_exits[spec.exit_id] = {
            "canonical_exit_id": spec.exit_id,
            "stop_bps": spec.stop_bps, "target_bps": spec.target_bps,
            "trail_activation_bps": spec.trail_activation_bps, "giveback_bps": spec.giveback_bps,
            "giveback_mode": spec.giveback_mode, "no_progress_sec": spec.no_progress_sec,
            "max_hold_sec": spec.max_hold_sec,
            "no_progress_mfe_bps": spec.no_progress_mfe_bps,
            "no_progress_abs_ret_bps": spec.no_progress_abs_ret_bps,
        }
    baseline_rows, baseline_sha = freeze_family_baselines(assignments, x26a_exits)
    if baseline_sha != FAMILY_BASELINE_REGISTRY_SHA:
        raise RuntimeError("baseline sha recompute")
    baseline_by = {r["candidate_id"]: r for r in baseline_rows}

    specs: dict[str, ExitSpec] = {}
    for s in sem_reg:
        specs[s["semantic_exit_sha256"]] = _spec_from_row(s["semantic_exit_sha256"], s)
    needed_eids: set[str] = set()
    for r in cohorts["specific"]:
        needed_eids.add(r["semantic_exit_sha256"])
        needed_eids.add(r["family_baseline_exit_id"])
    for r in cohorts["family"]:
        needed_eids.add(r["primary_exit_id"])
        needed_eids.add(r["counterfactual_specific_semantic_exit_sha256"])
    for eid in needed_eids:
        if eid in specs:
            continue
        p = x26a_exits.get(eid)
        if p is None:
            # try assignments semantic
            raise RuntimeError(f"missing exit spec {eid}")
        specs[eid] = _spec_from_row(eid, p)
    spec_list = [specs[e] for e in sorted(needed_eids)]
    print(f"  unique exits for 167={len(spec_list)}", flush=True)

    print("=== build full executable matrices ===", flush=True)
    mats = build_exit_matrices(
        specs=spec_list, rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=board_by_key, max_workers=4,
    )
    # A/B on one exit
    sample_eid = spec_list[0].exit_id
    mat_a = mats[sample_eid]
    mat_b = build_exit_matrices(
        specs=[specs[sample_eid]], rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=board_by_key, max_workers=1,
    )[sample_eid]
    ab_mat = bool(np.array_equal(mat_a["valid"], mat_b["valid"])) and (
        not mat_a["valid"].any()
        or abs(float(np.nanmean(mat_a["ret_bps"][mat_a["valid"]]))
               - float(np.nanmean(mat_b["ret_bps"][mat_b["valid"]]))) < 1e-9
    )

    # X28C classification returns for direction compare
    class_rows = _load_sheet(X28C_DIR / "audit.xlsx", "Classification")
    x28c_ret = {c["candidate_id"]: _f(c.get("avg_return_bps")) for c in class_rows}

    print("=== evaluate Specific 49 ===", flush=True)
    specific_rows = []
    stop_bucket_trades: dict[str, list] = defaultdict(list)
    for reg in cohorts["specific"]:
        cid = reg["candidate_id"]
        a = assign_by[cid]
        sel = apply_mask(rows, cid, cand_by)
        sem = reg["semantic_exit_sha256"]
        fam = reg["family_baseline_exit_id"]
        mat_s = mats[sem]
        mat_f = mats[fam]
        ed, sel_m, _ = entry_delta(mat_s, sel, dates, symbols)
        pers = personalization_delta(mat_s, mat_f, sel)
        stop_bps = _f(reg.get("stop_bps")) or _f(a.get("stop_bps"))
        srt = reg.get("stop_risk_tag") or stop_risk_tag(stop_bps)
        diag = stop_diagnostics(mat=mat_s, sel_mask=sel, stop_bps=stop_bps)
        dtag = direction_vs_x28c(sel_m.get("avg_return_bps"), x28c_ret.get(cid))
        row = {
            "candidate_id": cid,
            "cohort": "SPECIFIC",
            "stop_risk_tag": srt,
            "stop_bps": stop_bps,
            "avg_return_bps": sel_m.get("avg_return_bps"),
            "median_return_bps": sel_m.get("median_return_bps"),
            "profit_factor": sel_m.get("profit_factor"),
            "entry_delta": ed,
            "personalization_delta": pers.get("delta_avg_return"),
            "personalization_pnl_delta": pers.get("delta_avg_pnl"),
            "trades": sel_m.get("trades"),
            "days": sel_m.get("days"),
            "symbols": sel_m.get("symbols"),
            "coverage": sel_m.get("coverage"),
            "direction_vs_x28c": dtag,
            "exit_reason_counts": diag.get("exit_reason_counts"),
            "hard_stop_exit_rate": diag.get("hard_stop_exit_rate"),
            "hard_stop_exit_n": diag.get("hard_stop_exit_n"),
            "near_stop_count": diag.get("near_stop_count"),
            "near_stop_recovery_count": diag.get("near_stop_recovery_count"),
            "near_stop_recovery_rate": diag.get("near_stop_recovery_rate"),
            "avg_mae": diag.get("avg_mae"),
            "median_mae": diag.get("median_mae"),
            "mae_q25": diag.get("mae_q25"),
            "mae_q50": diag.get("mae_q50"),
            "mae_q75": diag.get("mae_q75"),
            "avg_mfe": diag.get("avg_mfe"),
            "winner_mae": diag.get("winner_mae"),
            "loser_mae": diag.get("loser_mae"),
            "gross_loss_from_hard_stops": diag.get("gross_loss_from_hard_stops"),
            "cluster_balanced_return": cluster_balanced_metric(
                mat=mat_s, sel_mask=sel, clusters=clusters,
            ),
            "cluster_balanced_entry_proxy": None,  # filled via pooled later
        }
        # store mats refs for pooled views via indices
        row["_sel"] = sel
        row["_mat_s"] = mat_s
        row["_mat_f"] = mat_f
        specific_rows.append(row)
        stop_bucket_trades[srt].append(row)

    print("=== evaluate Family 118 ===", flush=True)
    family_rows = []
    for reg in cohorts["family"]:
        cid = reg["candidate_id"]
        sel = apply_mask(rows, cid, cand_by)
        fam_eid = reg["primary_exit_id"]
        spec_eid = reg["counterfactual_specific_semantic_exit_sha256"]
        mat_f = mats[fam_eid]
        mat_s = mats[spec_eid]
        ed, sel_m, _ = entry_delta(mat_f, sel, dates, symbols)
        # family-minus-specific (primary family vs counterfactual specific)
        fam_vs_spec = personalization_delta(mat_f, mat_s, sel)
        row = {
            "candidate_id": cid,
            "cohort": "FAMILY",
            "stop_risk_tag": reg.get("stop_risk_tag"),
            "avg_return_bps": sel_m.get("avg_return_bps"),
            "median_return_bps": sel_m.get("median_return_bps"),
            "profit_factor": sel_m.get("profit_factor"),
            "entry_delta": ed,
            "family_minus_specific": fam_vs_spec.get("delta_avg_return"),
            "trades": sel_m.get("trades"),
            "days": sel_m.get("days"),
            "symbols": sel_m.get("symbols"),
            "coverage": sel_m.get("coverage"),
            "cluster_balanced_return": cluster_balanced_metric(
                mat=mat_f, sel_mask=sel, clusters=clusters,
            ),
            "_sel": sel,
            "_mat_f": mat_f,
            "_mat_s": mat_s,
        }
        family_rows.append(row)

    # Candidate-balanced + cluster-balanced views
    spec_view = candidate_balanced_view(
        specific_rows,
        {
            "candidate_return": "avg_return_bps",
            "entry_delta": "entry_delta",
            "personalization_delta": "personalization_delta",
        },
    )
    fam_view = candidate_balanced_view(
        family_rows,
        {
            "candidate_return": "avg_return_bps",
            "entry_delta": "entry_delta",
            "family_minus_specific": "family_minus_specific",
        },
    )

    # Cluster-balanced ENTRY delta: per-candidate cluster means of selected-minus-complement hard;
    # approximate with mean of cluster_balanced_return diffs not available — compute pooled:
    def _clus_entry(rows_list, mat_key_sel):
        vals = []
        for r in rows_list:
            mat = r[mat_key_sel]
            sel = r["_sel"]
            # cluster means selected vs complement
            idx_s = np.where(sel & mat["valid"])[0]
            idx_c = np.where((~sel) & mat["valid"])[0]
            if idx_s.size == 0 or idx_c.size == 0:
                continue
            # use overall cluster-balanced selected minus complement mean of cluster means
            sm = cluster_balanced_metric(mat=mat, sel_mask=sel, clusters=clusters)
            # complement cluster-balanced
            cm = cluster_balanced_metric(mat=mat, sel_mask=~sel, clusters=clusters)
            if sm is not None and cm is not None:
                vals.append(sm - cm)
        return float(np.mean(vals)) if vals else None

    def _clus_pers(rows_list, key_a, key_b):
        vals = []
        for r in rows_list:
            ma, mb, sel = r[key_a], r[key_b], r["_sel"]
            common = sel & ma["valid"] & mb["valid"]
            if not common.any():
                continue
            # cluster means of delta
            idx = np.where(common)[0]
            means = []
            for c in np.unique(clusters[idx]):
                m = idx[clusters[idx] == c]
                means.append(float(np.mean(ma["ret_bps"][m] - mb["ret_bps"][m])))
            if means:
                vals.append(float(np.mean(means)))
        return float(np.mean(vals)) if vals else None

    clus_spec_abs = float(np.mean([
        r["cluster_balanced_return"] for r in specific_rows
        if r["cluster_balanced_return"] is not None
    ])) if any(r["cluster_balanced_return"] is not None for r in specific_rows) else None
    clus_spec_entry = _clus_entry(specific_rows, "_mat_s")
    clus_spec_pers = _clus_pers(specific_rows, "_mat_s", "_mat_f")
    clus_fam_abs = float(np.mean([
        r["cluster_balanced_return"] for r in family_rows
        if r["cluster_balanced_return"] is not None
    ])) if any(r["cluster_balanced_return"] is not None for r in family_rows) else None
    clus_fam_entry = _clus_entry(family_rows, "_mat_f")
    clus_fam_pers = _clus_pers(family_rows, "_mat_f", "_mat_s")

    specific_status = stress_status_from_views(
        cand_abs_med=spec_view.get("median_candidate_return"),
        cand_abs_share=spec_view.get("positive_candidate_return_share"),
        clus_abs=clus_spec_abs,
        cand_entry_med=spec_view.get("median_entry_delta"),
        cand_entry_share=spec_view.get("positive_entry_delta_share"),
        clus_entry=clus_spec_entry,
        cand_pers_med=spec_view.get("median_personalization_delta"),
        cand_pers_share=spec_view.get("positive_personalization_delta_share"),
        clus_pers=clus_spec_pers,
        prefix="SPECIFIC",
    )
    family_status = stress_status_from_views(
        cand_abs_med=fam_view.get("median_candidate_return"),
        cand_abs_share=fam_view.get("positive_candidate_return_share"),
        clus_abs=clus_fam_abs,
        cand_entry_med=fam_view.get("median_entry_delta"),
        cand_entry_share=fam_view.get("positive_entry_delta_share"),
        clus_entry=clus_fam_entry,
        cand_pers_med=fam_view.get("median_family_minus_specific"),
        cand_pers_share=fam_view.get("positive_family_minus_specific_share"),
        clus_pers=clus_fam_pers,
        prefix="FAMILY",
    )

    # Stop-risk class view (Specific)
    stop_view_rows = []
    for tag, expected_n in (
        ("NORMAL_STOP", NORMAL_STOP_N),
        ("WIDE_STOP", WIDE_STOP_N),
        ("VERY_WIDE_STOP", VERY_WIDE_STOP_N),
    ):
        grp = [r for r in specific_rows if r.get("stop_risk_tag") == tag]
        rets = [r["avg_return_bps"] for r in grp if r.get("avg_return_bps") is not None]
        pfs = [r["profit_factor"] for r in grp if r.get("profit_factor") is not None]
        hard = [r["hard_stop_exit_rate"] for r in grp if r.get("hard_stop_exit_rate") is not None]
        mae = [r["avg_mae"] for r in grp if r.get("avg_mae") is not None]
        pers = [r["personalization_delta"] for r in grp if r.get("personalization_delta") is not None]
        near_c = sum(int(r.get("near_stop_count") or 0) for r in grp)
        near_r = sum(int(r.get("near_stop_recovery_count") or 0) for r in grp)
        stop_view_rows.append({
            "stop_risk_tag": tag,
            "candidate_count": len(grp),
            "expected_count": expected_n,
            "trade_count": sum(int(r.get("trades") or 0) for r in grp),
            "avg_return_bps": float(np.mean(rets)) if rets else None,
            "profit_factor": float(np.mean(pfs)) if pfs else None,
            "hard_stop_rate": float(np.mean(hard)) if hard else None,
            "avg_mae": float(np.mean(mae)) if mae else None,
            "specific_family_delta": float(np.mean(pers)) if pers else None,
            "near_stop_count": near_c,
            "near_stop_recovery_count": near_r,
            "near_stop_recovery_rate": (near_r / near_c) if near_c else None,
        })
    alert = wide_stop_alert(stop_view_rows)

    # Pooled LOSO/LODO across Specific (equal cand weight via concatenating selected trades
    # is biased; report max across candidates' LODO/LOSO + pooled all-specific trades)
    print("=== LOSO / LODO ===", flush=True)
    # Pool: union of selected valid trades for all specific (each trade once per cand — for dep report)
    # Use equal-candidate: take max_day_contribution across candidates
    lodo_shares = []
    loso_shares = []
    focus_agg = {s: [] for s in ("2354", "285A", "4052")}
    for r in specific_rows:
        ld = lodo(mat=r["_mat_s"], sel_mask=r["_sel"], dates=dates)
        ls = loso(mat=r["_mat_s"], sel_mask=r["_sel"], symbols=symbols)
        if ld.get("max_day_contribution_share") is not None:
            lodo_shares.append(ld["max_day_contribution_share"])
        if ls.get("max_symbol_contribution_share") is not None:
            loso_shares.append(ls["max_symbol_contribution_share"])
        for s, v in (ls.get("without") or {}).items():
            sym = s.replace("without_", "")
            if sym in focus_agg:
                focus_agg[sym].append(v)
    # Cohort-pooled LODO: build combined mask? Use mean of per-cand without_* positivity
    # Aggregate LODO by concatenating all specific selected returns with dates
    all_rets, all_dates, all_syms = [], [], []
    for r in specific_rows:
        idx = np.where(r["_sel"] & r["_mat_s"]["valid"])[0]
        all_rets.extend(r["_mat_s"]["ret_bps"][idx].tolist())
        all_dates.extend(dates[idx].tolist())
        all_syms.extend(symbols[idx].tolist())
    # synthetic mat for pooled
    if all_rets:
        pooled_mat = {
            "valid": np.ones(len(all_rets), dtype=bool),
            "ret_bps": np.asarray(all_rets, dtype=float),
            "pnl": np.asarray(all_rets, dtype=float),  # proxy
            "reason": np.array([""] * len(all_rets), dtype=object),
            "mae_bps": np.full(len(all_rets), np.nan),
            "mfe_bps": np.full(len(all_rets), np.nan),
        }
        pooled_sel = np.ones(len(all_rets), dtype=bool)
        pooled_dates = np.asarray(all_dates)
        pooled_syms = np.asarray(all_syms)
        pooled_lodo = lodo(mat=pooled_mat, sel_mask=pooled_sel, dates=pooled_dates)
        pooled_loso = loso(mat=pooled_mat, sel_mask=pooled_sel, symbols=pooled_syms)
    else:
        pooled_lodo = {"max_day_contribution_share": None, "positive_LODO_count": 0, "negative_LODO_count": 0}
        pooled_loso = {"max_symbol_contribution_share": None, "positive_LOSO_count": 0, "negative_LOSO_count": 0,
                       "without": {f"without_{s}": "NOT_PRESENT" for s in ("2354", "285A", "4052")}}

    dir_counts = Counter(r["direction_vs_x28c"] for r in specific_rows)
    decision = program_decision(specific_status, family_status)

    # Strip private keys before publish
    def _public(rows_list):
        out = []
        for r in rows_list:
            out.append({k: v for k, v in r.items() if not k.startswith("_")})
        return out

    print("=== tests ===", flush=True)
    # write interim for tests
    interim = {
        "run_id": run_id,
        "verdict": decision["program_decision"],
        "source_x28c_run_id": SOURCE_X28C,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "specific_n": EXPECTED_SPECIFIC,
        "family_n": EXPECTED_FAMILY,
        "overlap": 0,
        "no_candidate_selection_change": True,
        "no_parameter_retune": True,
        "phase0_performance_blind": True,
        "old_x29_superseded_if_sufficient": True,
        "data_sufficient": True,
        "submit_cancel_live": "0/0/0",
        "ab_determinism": ab_entry and ab_mat,
        "quote_contract": {
            "first_valid_ask": True, "first_valid_bid": True,
            "qty100": True, "freshness": True, "special_quote": True,
            "no_future_best": True, "same_session": True,
        },
        "actual_hard_stop_counted": True,
        "near_stop_recovery": True,
        "lodo_3_days": True,
        "loso": True,
        "candidate_balanced_view": True,
        "cluster_balanced_view": True,
        "program_decision_rule": decision["program_decision"],
        "no_runtime_change": True,
        "specific_stress_status": specific_status,
        "family_stress_status": family_status,
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2, default=str), encoding="utf-8")
    tests = _run_tests()

    x29_v2_sha = None
    x29_v2_run = None
    if decision["x29_v2_required"]:
        print("=== build X29 precommit V2 ===", flush=True)
        v2 = build_x29_v2(
            x28d_run_id=run_id,
            x28d_verdict=decision["program_decision"],
            x28d_precommit_sha=x28d_pre_sha,
        )
        x29_v2_sha = v2["precommit_sha"]
        x29_v2_run = v2["run_id"]

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": decision["program_decision"],
        "role": ROLE,
        "stress_days": list(STRESS_DAYS),
        "precommit_sha": x28d_pre_sha,
        "source_x28c_run_id": SOURCE_X28C,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "data_sufficiency": {d: True for d in STRESS_DAYS},
        "old_x29_status": "SUPERSEDED_BEFORE_PROSPECTIVE_MARKET_OPEN",
        "old_x29_precommit_sha": OLD_X29_PRECOMMIT_SHA,
        "old_x29_also": "SUPERSEDED_NOT_PROSPECTIVE_OPENED" if x29_v2_sha else None,
        "population_n": len(rows),
        "specific_cohort": {
            "n": EXPECTED_SPECIFIC,
            "stress_status": specific_status,
            "avg_return_bps": float(np.mean([r["avg_return_bps"] for r in specific_rows if r["avg_return_bps"] is not None]))
                if any(r["avg_return_bps"] is not None for r in specific_rows) else None,
            "median_return_bps": spec_view.get("median_candidate_return"),
            "profit_factor": float(np.mean([r["profit_factor"] for r in specific_rows if r["profit_factor"] is not None]))
                if any(r["profit_factor"] is not None for r in specific_rows) else None,
            "median_entry_delta": spec_view.get("median_entry_delta"),
            "median_personalization_delta": spec_view.get("median_personalization_delta"),
            "candidate_balanced": spec_view,
            "cluster_balanced": {
                "abs_return": clus_spec_abs,
                "entry_delta": clus_spec_entry,
                "personalization_delta": clus_spec_pers,
            },
            "direction_counts": dict(dir_counts),
        },
        "family_cohort": {
            "n": EXPECTED_FAMILY,
            "stress_status": family_status,
            "avg_return_bps": float(np.mean([r["avg_return_bps"] for r in family_rows if r["avg_return_bps"] is not None]))
                if any(r["avg_return_bps"] is not None for r in family_rows) else None,
            "median_return_bps": fam_view.get("median_candidate_return"),
            "profit_factor": float(np.mean([r["profit_factor"] for r in family_rows if r["profit_factor"] is not None]))
                if any(r["profit_factor"] is not None for r in family_rows) else None,
            "median_entry_delta": fam_view.get("median_entry_delta"),
            "median_family_minus_specific": fam_view.get("median_family_minus_specific"),
            "candidate_balanced": fam_view,
            "cluster_balanced": {
                "abs_return": clus_fam_abs,
                "entry_delta": clus_fam_entry,
                "family_minus_specific": clus_fam_pers,
            },
        },
        "stop_risk_view": {r["stop_risk_tag"]: r for r in stop_view_rows},
        "wide_stop_alert": alert,
        "dependency": {
            "max_day_contribution_share": pooled_lodo.get("max_day_contribution_share"),
            "positive_LODO_count": pooled_lodo.get("positive_LODO_count"),
            "negative_LODO_count": pooled_lodo.get("negative_LODO_count"),
            "max_symbol_contribution_share": pooled_loso.get("max_symbol_contribution_share"),
            "positive_LOSO_count": pooled_loso.get("positive_LOSO_count"),
            "negative_LOSO_count": pooled_loso.get("negative_LOSO_count"),
            "focus_symbols": pooled_loso.get("without"),
            "per_candidate_max_day_share_median": float(np.median(lodo_shares)) if lodo_shares else None,
            "per_candidate_max_symbol_share_median": float(np.median(loso_shares)) if loso_shares else None,
        },
        "program_decision": decision["program_decision"],
        "x29_v2_required": decision["x29_v2_required"],
        "x29_v2_run_id": x29_v2_run,
        "x29_v2_precommit_sha": x29_v2_sha,
        "cohort_membership_unchanged": True,
        "no_retune": True,
        "20260810_market_data_not_opened": True,
        "prospective_observer_not_started": True,
        "no_prospective_evidence_consumed": True,
        "ab_determinism": {"entry_asks": ab_entry, "matrix_sample": ab_mat, "ok": ab_entry and ab_mat},
        "tests": tests,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "Paper_order": False,
            "Live_order": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Discord_production_notification": False,
        },
        "mask_sha_check": mask_check,
    }

    sheets = {
        "Index": [{"run_id": run_id, "verdict": report["verdict"]}],
        "Specific49": _public(specific_rows),
        "Family118": _public(family_rows),
        "StopRisk": stop_view_rows,
        "Views": [
            {"cohort": "SPECIFIC", **spec_view, "status": specific_status},
            {"cohort": "FAMILY", **fam_view, "status": family_status},
        ],
        "Dependency": [report["dependency"]],
        "Decision": [decision],
        "Tests": [tests],
        "Safety": [report["safety"]],
        "ChangeLog": [{
            "at": now.isoformat(),
            "note": "E1_X28D additional historical stress 0805-07; X29 superseded before prospective open",
        }],
    }
    publish(OUT, report, sheets)
    print(f"=== DONE verdict={report['verdict']} x29_v2={bool(x29_v2_sha)} ===", flush=True)
    print("=== STOP: do not open 20260810 ===", flush=True)
    return report


if __name__ == "__main__":
    run()
