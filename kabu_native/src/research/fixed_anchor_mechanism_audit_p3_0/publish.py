"""Write P3-0 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.fixed_anchor_mechanism_audit_p3_0 import (
    ANALYSIS_ID,
    CLOCK_OFFSETS_SEC,
    CONTAMINATION_LABEL,
    DOCUMENT_ID,
    FULL14,
    OLD_DYNAMIC_STATUS,
    P1_MAXDD,
    P1_PF,
    P1_PNL,
    P1_REST11_PF,
    P1_REST11_PNL,
    P1_REST11_TRADES,
    P1_RESULT_NAME,
    P1_TOP3_PNL,
    P1_TOP3_TRADES,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    TASK_LABEL,
    TRAIL10_STATUS,
    VERDICT_BLOCKED,
    VERDICT_COMPLETE,
    VERDICT_INTEGRITY,
)
from research.fixed_anchor_mechanism_audit_p3_0.metrics import (
    clock_label,
    filter_days,
    filter_xs,
    independent_group_metrics,
    mechanism_label,
    monotonicity,
    offset_key,
    pf_num,
    pf_out,
    rank_strata,
    selection_pair_result,
    selection_result,
    slice_pair_to_label,
    trade_stats,
    within_anchor,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "fixed_anchor_mechanism_audit_p3_0"
P1_REPORT = NATIVE / "results" / "research" / "current_runtime_full_capture_recalc_p1" / "report.json"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj == float("-inf"):
            return "-Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"


def _sel_block(rows: list[dict[str, Any]]) -> dict[str, Any]:
    elig = [r for r in rows if r.get("feature_evaluable")]
    sel = [r for r in elig if r.get("selected")]
    nos = [r for r in elig if not r.get("selected")]
    sm = independent_group_metrics(sel)
    nm = independent_group_metrics(nos)
    pair = selection_pair_result(sm, nm)
    return {"SELECTED": sm, "NOT_SELECTED": nm, "pair": pair, "eligible_n": len(elig)}


def _clock_pack(orig: dict[str, Any], shifts: dict[int, dict[str, Any]]) -> dict[str, Any]:
    ordered = [shifts[o] for o in CLOCK_OFFSETS_SEC if o in shifts]
    return {
        "original": orig,
        "shifts": {offset_key(o): shifts[o] for o in CLOCK_OFFSETS_SEC if o in shifts},
        "result": clock_label(orig, ordered),
    }


def build_report(
    *,
    inventory: list[dict[str, Any]],
    support: dict[str, Any],
    baseline_days: list[dict[str, Any]],
    cs_days: list[dict[str, Any]],
    shift_days: dict[int, list[dict[str, Any]]],
    failed: list[str],
    p1: dict[str, Any],
    leak: bool,
    blocked: bool,
    blocked_reason: str = "",
) -> dict[str, Any]:
    now = datetime.now(JST).isoformat(timespec="seconds")
    p1_daily = {str(r["date"]): r for r in (p1.get("daily") or [])}
    recon_rows = []
    recon_pass = True
    base_by = {d["date"]: d for d in baseline_days if d.get("ok")}
    for day in FULL14:
        got = base_by.get(day) or {}
        exp = p1_daily.get(day) or {}
        got_pf = got.get("PF")
        exp_pf = exp.get("PF")
        gn, en = pf_num(got_pf), pf_num(exp_pf)
        match_pf = got_pf == exp_pf or (gn == float("inf") and en == float("inf")) or (
            gn == gn and en == en and abs(gn - en) < 1e-12
        )
        row = {
            "date": day,
            "expected_trades": exp.get("trades"),
            "got_trades": got.get("trade_n"),
            "expected_pnl": exp.get("pnl"),
            "got_pnl": got.get("pnl"),
            "expected_PF": pf_out(exp_pf),
            "got_PF": pf_out(got_pf),
            "expected_maxDD": exp.get("maxDD"),
            "got_maxDD": got.get("maxDD"),
            "expected_ledger_sha": exp.get("ledger_sha"),
            "got_ledger_sha": got.get("ledger_sha"),
            "match_trades": got.get("trade_n") == exp.get("trades"),
            "match_pnl": got.get("pnl") == exp.get("pnl"),
            "match_PF": bool(match_pf),
            "match_maxDD": got.get("maxDD") == exp.get("maxDD"),
            "match_sha": got.get("ledger_sha") == exp.get("ledger_sha") and bool(exp.get("ledger_sha")),
        }
        row["match"] = all(
            row[k] for k in ("match_trades", "match_pnl", "match_PF", "match_maxDD", "match_sha")
        )
        if not row["match"]:
            recon_pass = False
        recon_rows.append(row)

    base_trades = [t for d in baseline_days if d.get("ok") for t in (d.get("trades") or [])]
    base_full = [t for t in base_trades if str(t.get("date")) in set(FULL14)]
    primary = trade_stats(base_full)

    cs_trades = [t for d in cs_days if d.get("ok") for t in (d.get("trades") or [])]
    cs_full = [t for t in cs_trades if str(t.get("date")) in set(FULL14)]
    cs_stats = trade_stats(cs_full)

    shift_stats: dict[int, dict[str, Any]] = {}
    shift_trades: dict[int, list[dict[str, Any]]] = {}
    for off in CLOCK_OFFSETS_SEC:
        days = shift_days.get(off) or []
        tr = [t for d in days if d.get("ok") for t in (d.get("trades") or [])]
        tr = [t for t in tr if str(t.get("date")) in set(FULL14)]
        shift_trades[off] = tr
        st = trade_stats(tr)
        st["offset_sec"] = off
        st["delta_trades"] = st["trades"] - cs_stats["trades"]
        st["delta_pnl"] = round(st["pnl"] - cs_stats["pnl"], 2)
        st["delta_PF"] = None
        if st["PF"] is not None and cs_stats["PF"] is not None:
            st["delta_PF"] = pf_num(st["PF"]) - pf_num(cs_stats["PF"])
            if st["delta_PF"] == float("inf") or st["delta_PF"] == float("-inf"):
                st["delta_PF"] = str(st["delta_PF"])
        shift_stats[off] = st

    def _clock_for(days_keep: tuple[str, ...]) -> dict[str, Any]:
        orig = trade_stats(filter_days(cs_full, days_keep))
        sh = {o: trade_stats(filter_days(shift_trades[o], days_keep)) for o in CLOCK_OFFSETS_SEC}
        return _clock_pack(orig, sh)

    clock_all = _clock_for(FULL14)
    clock_top3 = _clock_for(PREDECLARED_TOP3)
    clock_rest = _clock_for(REST11)

    xs_all = [r for d in baseline_days if d.get("ok") for r in (d.get("xs_rows") or [])]
    xs_all = [r for r in xs_all if str(r.get("date")) in set(FULL14)]
    market = [r for d in baseline_days if d.get("ok") for r in (d.get("market_state") or [])]

    sel_all = _sel_block(xs_all)
    sel_top3 = _sel_block(filter_xs(xs_all, PREDECLARED_TOP3))
    sel_rest = _sel_block(filter_xs(xs_all, REST11))
    sel_label = selection_result(all_pair=sel_all["pair"], rest_pair=sel_rest["pair"])
    strata = rank_strata(xs_all)
    within = within_anchor(xs_all)
    within_top3 = within_anchor(filter_xs(xs_all, PREDECLARED_TOP3))
    within_rest = within_anchor(filter_xs(xs_all, REST11))
    mono_mean = monotonicity(strata, "mean_independent_pnl")
    mono_fill = monotonicity(strata, "fill_rate")
    mono_pf = monotonicity(strata, "PF")

    mech = mechanism_label(
        clock_all=clock_all["result"],
        sel_all=sel_label,
        clock_top3=clock_top3["result"],
        sel_top3=slice_pair_to_label(sel_top3["pair"]),
        clock_rest=clock_rest["result"],
        sel_rest=slice_pair_to_label(sel_rest["pair"]),
    )

    recon_status = "PASS" if recon_pass else "FAIL"
    if blocked or not recon_pass:
        verdict = VERDICT_BLOCKED
    elif leak:
        verdict = VERDICT_INTEGRITY
    else:
        verdict = VERDICT_COMPLETE

    q1 = (
        "Yes — original common-support beats every local shift on both PnL and PF."
        if clock_all["result"] == "CLOCK_EXACT_TIME_SUPPORTED"
        else (
            "No — at least one local shift beats original common-support on both PnL and PF."
            if clock_all["result"] == "CLOCK_NOT_UNIQUELY_SPECIAL"
            else "Mixed — exact minute is not uniquely best on both PnL and PF, and no shift dominates both."
        )
    )
    q2 = (
        "Selector advantage is present on ALL FULL14 and REST11 (higher independent fill rate and mean pnl)."
        if sel_label == "SELECTION_SUPPORTED"
        else (
            "Selector is not supported (selected worse on both fill rate and mean pnl for ALL and REST11)."
            if sel_label == "SELECTION_NOT_SUPPORTED"
            else "Selector evidence is mixed; do not treat higher selected scores as proof of selector value."
        )
    )
    q3 = (
        "Yes — REST11 does not carry exact-clock or selection support; Top3 does."
        if mech == "TOP3_DOMINATED"
        else "Not only Top3 — REST11 still participates in the clock and/or selection pattern."
    )
    q4 = (
        "Both a clock and a synchronized cross-section point."
        if mech == "CLOCK_AND_SELECTION"
        else (
            "Primarily an exact clock."
            if mech == "EXACT_CLOCK_TIMING"
            else (
                "Primarily a cross-sectional synchronization / selection point."
                if mech == "CROSS_SECTIONAL_SELECTION"
                else (
                    "Strength is concentrated in predeclared Top3 days."
                    if mech == "TOP3_DOMINATED"
                    else "Evidence is mixed / no single mechanism is uniquely supported."
                )
            )
        )
    )

    report: dict[str, Any] = {
        "task": "P3-0",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "label": TASK_LABEL,
        "contamination_label": CONTAMINATION_LABEL,
        "not": ["OOS", "prospective", "robust", "strategy validation"],
        "generated_at_jst": now,
        "BASELINE_RECONCILE": recon_status,
        "BASELINE_RECONCILE_ROWS": recon_rows,
        "PRIMARY_FIXED": {
            "result_name": P1_RESULT_NAME,
            "trades": primary["trades"],
            "pnl": primary["pnl"],
            "PF": primary["PF"],
            "maxDD": primary["maxDD"],
            "p1_canonical": {
                "trades": P1_TRADES,
                "pnl": P1_PNL,
                "PF": P1_PF,
                "maxDD": P1_MAXDD,
            },
        },
        "COMMON_SUPPORT": {
            "original_anchor_count": support["original_anchor_count"],
            "anchor_count": support["common_support_anchor_count"],
            "original_grid": support["original_grid"],
            "common_support_grid": support["common_support_grid"],
            "excluded_anchor_times": support["excluded"],
            "exclusion_reasons": [e["exclusion_reasons"] for e in support["excluded"]],
            "baseline_trades": cs_stats["trades"],
            "baseline_pnl": cs_stats["pnl"],
            "baseline_PF": cs_stats["PF"],
            "baseline_maxDD": cs_stats["maxDD"],
            "note": "P1 headline is unchanged. Shift deltas use this common-support original, not P1 headline.",
        },
        "CLOCK_SHIFTS": {
            offset_key(o): {
                "trades": shift_stats[o]["trades"],
                "pnl": shift_stats[o]["pnl"],
                "PF": shift_stats[o]["PF"],
                "win": shift_stats[o]["win"],
                "loss": shift_stats[o]["loss"],
                "draw": shift_stats[o]["draw"],
                "gross_profit": shift_stats[o]["gross_profit"],
                "gross_loss": shift_stats[o]["gross_loss"],
                "avg_pnl": shift_stats[o]["avg_pnl"],
                "maxDD": shift_stats[o]["maxDD"],
                "AM_pnl": shift_stats[o]["AM_pnl"],
                "PM_pnl": shift_stats[o]["PM_pnl"],
                "delta_trades": shift_stats[o]["delta_trades"],
                "delta_pnl": shift_stats[o]["delta_pnl"],
                "delta_PF": shift_stats[o]["delta_PF"],
            }
            for o in CLOCK_OFFSETS_SEC
            if o in shift_stats
        },
        "CLOCK_RESULT": clock_all["result"],
        "CLOCK_TOP3": clock_top3,
        "CLOCK_REST11": clock_rest,
        "SELECTED_VS_NOT": {
            "SELECTED": sel_all["SELECTED"],
            "NOT_SELECTED": sel_all["NOT_SELECTED"],
            "pair": sel_all["pair"],
        },
        "WITHIN_ANCHOR": {
            "selected_better_count": within["selected_better_count"],
            "selected_worse_count": within["selected_worse_count"],
            "equal": within["equal"],
            "median_difference": within["median_difference"],
            "n_compared_anchors": within["n_compared_anchors"],
        },
        "RANK_STRATA": {s["quintile"]: s for s in strata},
        "RANK_MONOTONICITY": {
            "mean_independent_pnl": mono_mean,
            "fill_rate": mono_fill,
            "PF": mono_pf,
            "clear_q1_to_q5_monotonicity": bool(
                mono_mean.get("non_increasing") and mono_fill.get("non_increasing")
            ),
        },
        "SELECTION_RESULT": sel_label,
        "TOP3": {
            "days": list(PREDECLARED_TOP3),
            "p1_canonical": {"trades": P1_TOP3_TRADES, "pnl": P1_TOP3_PNL},
            "clock_result": clock_top3["result"],
            "selection_result": slice_pair_to_label(sel_top3["pair"]),
            "SELECTED_VS_NOT": {"SELECTED": sel_top3["SELECTED"], "NOT_SELECTED": sel_top3["NOT_SELECTED"]},
            "WITHIN_ANCHOR": {
                "selected_better_count": within_top3["selected_better_count"],
                "selected_worse_count": within_top3["selected_worse_count"],
                "equal": within_top3["equal"],
                "median_difference": within_top3["median_difference"],
            },
        },
        "REST11": {
            "days": list(REST11),
            "p1_canonical": {"trades": P1_REST11_TRADES, "pnl": P1_REST11_PNL, "PF": P1_REST11_PF},
            "clock_result": clock_rest["result"],
            "selection_result": slice_pair_to_label(sel_rest["pair"]),
            "SELECTED_VS_NOT": {"SELECTED": sel_rest["SELECTED"], "NOT_SELECTED": sel_rest["NOT_SELECTED"]},
            "WITHIN_ANCHOR": {
                "selected_better_count": within_rest["selected_better_count"],
                "selected_worse_count": within_rest["selected_worse_count"],
                "equal": within_rest["equal"],
                "median_difference": within_rest["median_difference"],
            },
        },
        "FIXED_ANCHOR_MECHANISM": mech,
        "Q1_EXACT_CLOCK_MATERIAL": q1,
        "Q2_SYNCHRONIZED_CROSS_SECTION_MATERIAL": q2,
        "Q3_TOP3_ONLY": q3,
        "Q4_CLOCK_VS_SYNC_POINT": q4,
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "DYNAMIC_RETUNED": False,
        "BEST_OFFSET_ADOPTED": False,
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "snapshot_future_leak": leak,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "old_dynamic": OLD_DYNAMIC_STATUS,
            "TRAIL10": TRAIL10_STATUS,
            "do_not_retune": True,
        },
        "file_sha": {
            "v1r_native_entry_live.py": _file_sha("src/small_paper/v1r_native_entry_live.py"),
            "v1r_live_dual_lane.py": _file_sha("src/small_paper/v1r_live_dual_lane.py"),
            "v1r_primary_runtime.py": _file_sha("src/small_paper/v1r_primary_runtime.py"),
        },
        "_sheets": {
            "recon_rows": recon_rows,
            "clock_daily": _clock_daily(cs_days, shift_days),
            "xs_rows": xs_all,
            "market_state": market,
            "within_rows": within.get("rows") or [],
            "strata": strata,
            "cs_stats": cs_stats,
            "shift_stats": shift_stats,
            "primary": primary,
            "sel_all": sel_all,
            "sel_top3": sel_top3,
            "sel_rest": sel_rest,
            "inventory": [
                {
                    "date": r.get("date"),
                    "capture_class": r.get("capture_class"),
                    "replay_eligible": r.get("replay_eligible"),
                    "universe_n": r.get("universe_n"),
                }
                for r in inventory
            ],
        },
    }
    return json_sanitize(report)


def _clock_daily(
    cs_days: list[dict[str, Any]],
    shift_days: dict[int, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    rows = []
    by_cs = {d["date"]: d for d in cs_days if d.get("ok")}
    by_sh = {
        off: {d["date"]: d for d in days if d.get("ok")} for off, days in shift_days.items()
    }
    for day in FULL14:
        cs = by_cs.get(day) or {}
        rec = {
            "date": day,
            "original_cs_trades": cs.get("trade_n"),
            "original_cs_pnl": cs.get("pnl"),
            "original_cs_PF": pf_out(cs.get("PF")),
        }
        for off in CLOCK_OFFSETS_SEC:
            d = (by_sh.get(off) or {}).get(day) or {}
            k = offset_key(off)
            rec[f"{k}_trades"] = d.get("trade_n")
            rec[f"{k}_pnl"] = d.get("pnl")
            rec[f"{k}_PF"] = pf_out(d.get("PF"))
        rows.append(rec)
    return rows


def render_md(rep: dict[str, Any]) -> str:
    cs = rep["COMMON_SUPPORT"]
    pf = rep["PRIMARY_FIXED"]
    clk = rep["CLOCK_SHIFTS"]
    sel = rep["SELECTED_VS_NOT"]
    wa = rep["WITHIN_ANCHOR"]
    rs = rep["RANK_STRATA"]
    lines = [
        "# P3-0 Fixed Anchor Mechanism Audit",
        "",
        f"**Label:** `{rep['label']}` — not OOS, not prospective, not robust, not strategy validation.",
        "",
        "New strategy tested: **false**. Retuning: **false**. Runtime changed: **false**.",
        "",
        "## BASELINE_RECONCILE",
        "",
        f"`{rep['BASELINE_RECONCILE']}`",
        "",
        "## PRIMARY_FIXED (P1 headline unchanged)",
        "",
        f"trades: {pf['trades']}",
        f"pnl: {pf['pnl']}",
        f"PF: {pf['PF']}",
        f"maxDD: {pf['maxDD']}",
        "",
        "## COMMON_SUPPORT",
        "",
        f"original_anchor_count: {cs['original_anchor_count']}",
        f"anchor_count: {cs['anchor_count']}",
        f"excluded: {json.dumps(cs['excluded_anchor_times'], ensure_ascii=False)}",
        f"baseline_trades: {cs['baseline_trades']}",
        f"baseline_pnl: {cs['baseline_pnl']}",
        f"baseline_PF: {cs['baseline_PF']}",
        "",
        "## CLOCK_SHIFTS (vs common-support original)",
        "",
    ]
    for off in CLOCK_OFFSETS_SEC:
        k = offset_key(off)
        s = clk.get(k) or {}
        lines.append(
            f"{k}: trades={s.get('trades')} pnl={s.get('pnl')} PF={s.get('PF')} "
            f"delta_trades={s.get('delta_trades')} delta_pnl={s.get('delta_pnl')} delta_PF={s.get('delta_PF')}"
        )
    lines += [
        "",
        f"CLOCK_RESULT: `{rep['CLOCK_RESULT']}`",
        "",
        "## SELECTED_VS_NOT (INDEPENDENT_DIAGNOSTIC_OUTCOME — not strategy return)",
        "",
        f"SELECTED: n={sel['SELECTED']['n']} fill_rate={sel['SELECTED']['fill_rate']} "
        f"mean_independent_pnl={sel['SELECTED']['mean_independent_pnl']} "
        f"median_independent_pnl={sel['SELECTED']['median_independent_pnl']} PF={sel['SELECTED']['PF']}",
        f"NOT_SELECTED: n={sel['NOT_SELECTED']['n']} fill_rate={sel['NOT_SELECTED']['fill_rate']} "
        f"mean_independent_pnl={sel['NOT_SELECTED']['mean_independent_pnl']} "
        f"median_independent_pnl={sel['NOT_SELECTED']['median_independent_pnl']} PF={sel['NOT_SELECTED']['PF']}",
        "",
        "## WITHIN_ANCHOR",
        "",
        f"selected_better_count: {wa['selected_better_count']}",
        f"selected_worse_count: {wa['selected_worse_count']}",
        f"equal: {wa['equal']}",
        f"median_difference: {wa['median_difference']}",
        "",
        "## RANK_STRATA",
        "",
    ]
    for q in ("Q1", "Q2", "Q3", "Q4", "Q5"):
        s = rs.get(q) or {}
        lines.append(
            f"{q}: n={s.get('n')} fill_rate={s.get('fill_rate')} "
            f"mean={s.get('mean_independent_pnl')} median={s.get('median_independent_pnl')} PF={s.get('PF')}"
        )
    lines += [
        "",
        f"Q1→Q5 clear monotonicity (fill_rate and mean pnl): {rep['RANK_MONOTONICITY']['clear_q1_to_q5_monotonicity']}",
        "",
        f"SELECTION_RESULT: `{rep['SELECTION_RESULT']}`",
        "",
        "## TOP3 / REST11",
        "",
        f"TOP3 clock_result: `{rep['TOP3']['clock_result']}`",
        f"TOP3 selection_result: `{rep['TOP3']['selection_result']}`",
        f"REST11 clock_result: `{rep['REST11']['clock_result']}`",
        f"REST11 selection_result: `{rep['REST11']['selection_result']}`",
        "",
        f"FIXED_ANCHOR_MECHANISM: `{rep['FIXED_ANCHOR_MECHANISM']}`",
        "",
        "## Questions",
        "",
        f"Q1 exact clock material? {rep['Q1_EXACT_CLOCK_MATERIAL']}",
        f"Q2 synchronized cross-section material? {rep['Q2_SYNCHRONIZED_CROSS_SECTION_MATERIAL']}",
        f"Q3 Top3 only? {rep['Q3_TOP3_ONLY']}",
        f"Q4 clock vs sync point? {rep['Q4_CLOCK_VS_SYNC_POINT']}",
        "",
        "## Safety / freeze",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep['verdict']}`",
        "",
        "STOP. Do not adopt a best offset. Do not create a new strategy. Do not change Runtime.",
        "",
    ]
    return "\n".join(lines) + "\n"


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("task", public.get("task")),
            ("label", public.get("label")),
            ("BASELINE_RECONCILE", public.get("BASELINE_RECONCILE")),
            ("PRIMARY_FIXED_trades", public["PRIMARY_FIXED"]["trades"]),
            ("PRIMARY_FIXED_pnl", public["PRIMARY_FIXED"]["pnl"]),
            ("PRIMARY_FIXED_PF", public["PRIMARY_FIXED"]["PF"]),
            ("PRIMARY_FIXED_maxDD", public["PRIMARY_FIXED"]["maxDD"]),
            ("COMMON_SUPPORT_anchor_count", public["COMMON_SUPPORT"]["anchor_count"]),
            ("COMMON_SUPPORT_baseline_trades", public["COMMON_SUPPORT"]["baseline_trades"]),
            ("COMMON_SUPPORT_baseline_pnl", public["COMMON_SUPPORT"]["baseline_pnl"]),
            ("COMMON_SUPPORT_baseline_PF", public["COMMON_SUPPORT"]["baseline_PF"]),
            ("CLOCK_RESULT", public.get("CLOCK_RESULT")),
            ("SELECTION_RESULT", public.get("SELECTION_RESULT")),
            ("FIXED_ANCHOR_MECHANISM", public.get("FIXED_ANCHOR_MECHANISM")),
            ("TOP3_clock_result", public["TOP3"]["clock_result"]),
            ("TOP3_selection_result", public["TOP3"]["selection_result"]),
            ("REST11_clock_result", public["REST11"]["clock_result"]),
            ("REST11_selection_result", public["REST11"]["selection_result"]),
            ("NEW_STRATEGY_TESTED", False),
            ("RETUNING_DONE", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Baseline_Reconcile"), sheets.get("recon_rows") or [])
    clock_rows = []
    cs_st = sheets.get("cs_stats") or {}
    clock_rows.append(
        {
            "offset_sec": 0,
            "label": "original_common_support",
            **{k: cs_st.get(k) for k in (
                "trades", "win", "loss", "draw", "gross_profit", "gross_loss",
                "pnl", "PF", "avg_pnl", "maxDD", "AM_pnl", "PM_pnl",
            )},
            "delta_trades": 0,
            "delta_pnl": 0,
            "delta_PF": 0,
        }
    )
    shst = sheets.get("shift_stats") or {}
    for off in CLOCK_OFFSETS_SEC:
        st = shst.get(off) or {}
        clock_rows.append({"offset_sec": off, "label": offset_key(off), **st})
    _write_rows(wb.create_sheet("Clock_Shifts"), clock_rows)
    _write_rows(wb.create_sheet("Clock_Shifts_Daily"), sheets.get("clock_daily") or [])
    top3rest = [
        {
            "slice": "ALL_FULL14",
            "clock_result": public["CLOCK_RESULT"],
            "selection_result": public["SELECTION_RESULT"],
        },
        {
            "slice": "TOP3",
            "clock_result": public["TOP3"]["clock_result"],
            "selection_result": public["TOP3"]["selection_result"],
            **{f"sel_{k}": public["TOP3"]["SELECTED_VS_NOT"]["SELECTED"].get(k) for k in ("n", "fill_rate", "mean_independent_pnl", "PF")},
        },
        {
            "slice": "REST11",
            "clock_result": public["REST11"]["clock_result"],
            "selection_result": public["REST11"]["selection_result"],
            **{f"sel_{k}": public["REST11"]["SELECTED_VS_NOT"]["SELECTED"].get(k) for k in ("n", "fill_rate", "mean_independent_pnl", "PF")},
        },
    ]
    _write_rows(wb.create_sheet("Clock_Top3_Rest11"), top3rest)
    xs_rows = sheets.get("xs_rows") or []
    _write_rows(
        wb.create_sheet("Anchor_CrossSection"),
        [
            {k: r.get(k) for k in (
                "date", "session", "anchor_time", "symbol", "feature_evaluable", "score",
                "alloc_score", "rank", "selected", "live_eligible", "in_open", "in_pending",
                "exposure", "cap_full", "actual_admitted", "actual_filled", "actual_trade",
                "independent_filled", "independent_pnl", "independent_exit_reason",
            )}
            for r in xs_rows
        ] or [{"note": "empty"}],
    )
    sa = sheets.get("sel_all") or {"SELECTED": {}, "NOT_SELECTED": {}}
    _write_rows(
        wb.create_sheet("Selected_vs_Not"),
        [
            {"group": "SELECTED", **sa.get("SELECTED", {})},
            {"group": "ELIGIBLE_NOT_SELECTED", **sa.get("NOT_SELECTED", {})},
        ],
    )
    _write_rows(wb.create_sheet("Rank_Strata"), sheets.get("strata") or [])
    _write_rows(wb.create_sheet("Within_Anchor"), sheets.get("within_rows") or [])
    _write_rows(wb.create_sheet("Anchor_Market_State"), sheets.get("market_state") or [])
    _write_rows(
        wb.create_sheet("Top3"),
        [
            {"group": "SELECTED", **public["TOP3"]["SELECTED_VS_NOT"]["SELECTED"]},
            {"group": "ELIGIBLE_NOT_SELECTED", **public["TOP3"]["SELECTED_VS_NOT"]["NOT_SELECTED"]},
        ],
    )
    _write_rows(
        wb.create_sheet("Rest11"),
        [
            {"group": "SELECTED", **public["REST11"]["SELECTED_VS_NOT"]["SELECTED"]},
            {"group": "ELIGIBLE_NOT_SELECTED", **public["REST11"]["SELECTED_VS_NOT"]["NOT_SELECTED"]},
        ],
    )
    _kv_sheet(
        wb.create_sheet("Identity"),
        [(k, v) for k, v in (public.get("Identity") or {}).items()]
        + [("label", TASK_LABEL), ("contamination", CONTAMINATION_LABEL)],
    )
    _kv_sheet(
        wb.create_sheet("Safety"),
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("NEW_STRATEGY_TESTED", False),
            ("RETUNING_DONE", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("BEST_OFFSET_ADOPTED", False),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
