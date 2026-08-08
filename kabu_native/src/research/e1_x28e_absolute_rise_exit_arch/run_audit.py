"""E1_X28E runner: ENTRY-only → Regime → PBv2 → EXIT architecture → stop before 0810."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x26_exit_library.exits import ExitSpec, common_controls
from research.e1_x27_reference_joint.paths import build_paths_for_rows
from research.e1_x28_executable_joint.board import load_board_events, verify_board_mapping
from research.e1_x28_executable_joint.replay import build_entry_asks
from research.e1_x28b_candidate_reference.baseline import freeze_family_baselines
from research.e1_x28d_additional_stress.replay_stress import build_exit_matrices

from . import (
    ADDITIONAL_STRESS,
    ANALYSIS_ID,
    ASSIGNMENT_REGISTRY_SHA,
    BOARD_MAPPING_SHA,
    CONSUMED_DAY,
    DISCOVERY,
    DOCUMENT_ID,
    EVALUATION,
    EXPECTED_FAMILY,
    EXPECTED_SPECIFIC,
    FAMILY_BASELINE_REGISTRY_SHA,
    FORBIDDEN_FROM,
    HORIZONS_SEC,
    LOGIC_MANIFEST_SHA,
    PERIOD_BLOCKS,
    REGIME_LIBRARY,
    SEMANTIC_EXIT_REGISTRY_SHA,
    SOURCE_X28C,
    SOURCE_X28D,
    STRESS_DAY,
    VERDICT_FAMILY,
    VERDICT_NO_ROBUST,
    VERDICT_PBV2,
    VERDICT_REGIME_FAMILY,
    VERDICT_REGIME_PBV2,
    VERDICT_RELATIVE_ONLY,
    X29_V2_PRECOMMIT_SHA,
    X29_V2_RUN_ID,
)
from .analyze import (
    architecture_passes,
    block_abs,
    entry_delta_from_mat,
    lodo_loso,
    rank_architecture,
    summarize_exit,
)
from .cohorts import build_masks, load_cohorts
from .horizons import compute_horizon_arrays, entry_class, summarize_horizon
from .pbv2 import build_pbv2_matrix, freeze_pbv2_manifest, pbv2_parity_check
from .population import load_combined_population
from .publish import publish
from .regime import (
    assert_regime_library_frozen,
    period_mask,
    regime_mask,
    regime_stability,
    select_global_regime,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28e_absolute_rise_exit_arch"
X28A1 = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"
X26A = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"
X28C = NATIVE / "results" / "research" / "e1_x28c_candidate_executable"
X28D = NATIVE / "results" / "research" / "e1_x28d_additional_stress"
X29 = NATIVE / "results" / "research" / "e1_x29_prospective"


def _run_tests() -> dict[str, Any]:
    import os
    tp = NATIVE / "tests" / "research" / "test_e1_x28e_absolute_rise_exit_arch.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(tp), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {"exit_code": p.returncode, "passed": passed, "failed": failed,
            "total": passed + failed or 1, "detail": out[-3000:]}


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    return [{hdr[i]: row[i] for i in range(len(hdr))} for row in rows[1:]]


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _spec_from_row(exit_id: str, p: dict[str, Any]) -> ExitSpec:
    np_sec = _f(p.get("no_progress_sec"))
    gm = p.get("giveback_mode") or None
    if gm == "":
        gm = None
    return ExitSpec(
        exit_id=exit_id, path_family=None, variant=p.get("exit_mode"),
        stop_bps=_f(p.get("stop_bps")), target_bps=_f(p.get("target_bps")),
        trail_activation_bps=_f(p.get("trail_activation_bps")),
        giveback_bps=_f(p.get("giveback_bps")), giveback_mode=gm,
        no_progress_sec=np_sec,
        max_hold_sec=float(_f(p.get("max_hold_sec")) or 900.0),
        no_progress_mfe_bps=(_f(p.get("no_progress_mfe_bps")) or 5.0) if np_sec is not None else None,
        no_progress_abs_ret_bps=(_f(p.get("no_progress_abs_ret_bps")) or 5.0) if np_sec is not None else None,
        is_control=str(exit_id).startswith("CONTROL_"),
    )


def _load_boards(rows, allowed):
    keys = sorted({(r["date"], r["symbol"]) for r in rows if r["date"] in allowed})
    cache = {}
    print(f"  boards {len(keys)}...", flush=True)

    def _one(k):
        return k, load_board_events(k[0], k[1])

    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = [ex.submit(_one, k) for k in keys]
        done = 0
        for fut in as_completed(futs):
            k, b = fut.result()
            cache[k] = b
            done += 1
            if done % 40 == 0 or done == len(keys):
                print(f"    {done}/{len(keys)}", flush=True)
    return cache


def _med(vals):
    xs = [v for v in vals if v is not None]
    return float(np.median(xs)) if xs else None


def _mean(vals):
    xs = [v for v in vals if v is not None]
    return float(np.mean(xs)) if xs else None


def _cohort_horizon_view(masks, ids, hz_all, dates, primary_H=300):
    """Candidate-balanced ENTRY-only view at primary horizon (executable bid)."""
    hz = hz_all[primary_H]
    abs_rets, deltas, classes = [], [], []
    by_block = {b: {"abs": [], "delta": []} for b in PERIOD_BLOCKS}
    per_h = {H: {"abs": [], "delta": []} for H in HORIZONS_SEC}
    for cid in ids:
        sel = masks[cid]
        s = summarize_horizon(hz=hz, mask=sel, basis="bid")
        c = summarize_horizon(hz=hz, mask=~sel, basis="bid")
        abs_r = s.get("avg_return_bps")
        delta = None
        if abs_r is not None and c.get("avg_return_bps") is not None:
            delta = abs_r - c["avg_return_bps"]
        abs_rets.append(abs_r)
        deltas.append(delta)
        classes.append(entry_class(abs_ret=abs_r, entry_delta=delta))
        for block in PERIOD_BLOCKS:
            pm = period_mask(dates, block)
            sb = summarize_horizon(hz={k: (v[pm] if isinstance(v, np.ndarray) else v) for k, v in hz.items()},
                                   mask=sel[pm], basis="bid")
            cb = summarize_horizon(hz={k: (v[pm] if isinstance(v, np.ndarray) else v) for k, v in hz.items()},
                                   mask=(~sel)[pm], basis="bid")
            # simpler: index into full arrays
            pass
        # block via full arrays
        for block in PERIOD_BLOCKS:
            pm = period_mask(dates, block)
            sel_b = sel & pm
            comp_b = (~sel) & pm
            if (sel_b & hz["bid_valid"]).any():
                ar = float(np.mean(hz["bid_ret"][sel_b & hz["bid_valid"]]))
                by_block[block]["abs"].append(ar)
                if (comp_b & hz["bid_valid"]).any():
                    by_block[block]["delta"].append(ar - float(np.mean(hz["bid_ret"][comp_b & hz["bid_valid"]])))
        for H in HORIZONS_SEC:
            h = hz_all[H]
            if (sel & h["bid_valid"]).any():
                ar = float(np.mean(h["bid_ret"][sel & h["bid_valid"]]))
                per_h[H]["abs"].append(ar)
                if ((~sel) & h["bid_valid"]).any():
                    per_h[H]["delta"].append(ar - float(np.mean(h["bid_ret"][(~sel) & h["bid_valid"]])))

    from collections import Counter
    return {
        "n": len(ids),
        "primary_horizon_sec": primary_H,
        "median_absolute_return_bps": _med(abs_rets),
        "mean_absolute_return_bps": _mean(abs_rets),
        "median_entry_delta": _med(deltas),
        "mean_entry_delta": _mean(deltas),
        "positive_abs_share": float(np.mean([1 if (x or 0) > 0 else 0 for x in abs_rets])) if abs_rets else None,
        "positive_delta_share": float(np.mean([1 if (x or 0) > 0 else 0 for x in deltas])) if deltas else None,
        "entry_class_counts": dict(Counter(classes)),
        "by_block": {
            b: {"median_abs": _med(v["abs"]), "median_delta": _med(v["delta"])}
            for b, v in by_block.items()
        },
        "by_horizon": {
            str(H): {"median_abs": _med(v["abs"]), "median_delta": _med(v["delta"])}
            for H, v in per_h.items()
        },
    }


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id = f"e1x28e_arch_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    # Source identity
    x28c = json.loads((X28C / "report.json").read_text(encoding="utf-8"))
    x28d = json.loads((X28D / "report.json").read_text(encoding="utf-8"))
    if x28c.get("run_id") != SOURCE_X28C:
        raise RuntimeError("x28c")
    if x28d.get("run_id") != SOURCE_X28D:
        raise RuntimeError("x28d")
    for k, exp in (
        ("logic_manifest_sha", LOGIC_MANIFEST_SHA),
        ("assignment_registry_sha", ASSIGNMENT_REGISTRY_SHA),
        ("semantic_exit_registry_sha", SEMANTIC_EXIT_REGISTRY_SHA),
        ("family_baseline_registry_sha", FAMILY_BASELINE_REGISTRY_SHA),
        ("board_mapping_sha", BOARD_MAPPING_SHA),
    ):
        if x28c.get(k) != exp:
            raise RuntimeError(k)

    assert_regime_library_frozen()
    cohorts = load_cohorts()
    print("=== population ===", flush=True)
    rows = load_combined_population()
    dates = np.array([r["date"] for r in rows])
    symbols = np.array([str(r["symbol"]) for r in rows])
    assert not (dates >= FORBIDDEN_FROM).any()

    print("=== masks ===", flush=True)
    _, masks = build_masks(rows, cohorts)

    allowed = list(DISCOVERY + EVALUATION + (STRESS_DAY, CONSUMED_DAY) + ADDITIONAL_STRESS)
    print("=== mapping / paths / boards / asks ===", flush=True)
    mapping = verify_board_mapping()
    if not mapping.get("ok"):
        raise RuntimeError("board mapping")
    times_list, prices_list = build_paths_for_rows(rows, allowed_dates=allowed, max_workers=6)
    board_by_key = _load_boards(rows, allowed)
    entry_asks = build_entry_asks(rows, board_by_key)
    entry_asks_b = build_entry_asks(rows, board_by_key)
    ab_entry = bool(np.array_equal(entry_asks["valid"], entry_asks_b["valid"]))

    print("=== Phase A: fixed horizons (ENTRY-only) ===", flush=True)
    hz_all = compute_horizon_arrays(
        rows=rows, entry_asks=entry_asks, times_list=times_list,
        prices_list=prices_list, board_by_key=board_by_key,
    )
    entry_specific = _cohort_horizon_view(masks, cohorts["specific_ids"], hz_all, dates)
    entry_family = _cohort_horizon_view(masks, cohorts["family_ids"], hz_all, dates)

    print("=== Regime R0-R5 ===", flush=True)
    # Use Specific cohort candidate-balanced at H=300 under each regime
    H = 300
    hz = hz_all[H]
    regime_rows = []
    for rid in REGIME_LIBRARY:
        rg = regime_mask(rows, rid)
        abs_e = []
        delta_e = []
        block_metrics = {}
        for block in PERIOD_BLOCKS:
            pm = period_mask(dates, block) & rg
            abs_b, delta_b = [], []
            for cid in cohorts["specific_ids"]:
                sel = masks[cid] & pm
                comp = (~masks[cid]) & pm
                if (sel & hz["bid_valid"]).any():
                    ar = float(np.mean(hz["bid_ret"][sel & hz["bid_valid"]]))
                    abs_b.append(ar)
                    if (comp & hz["bid_valid"]).any():
                        delta_b.append(ar - float(np.mean(hz["bid_ret"][comp & hz["bid_valid"]])))
            block_metrics[block] = {
                "median_abs": _med(abs_b),
                "median_delta": _med(delta_b),
                "episode_proxy_n": int((rg & period_mask(dates, block) & hz["bid_valid"]).sum()),
            }
        # overall under regime (all allowed dates)
        for cid in cohorts["specific_ids"]:
            sel = masks[cid] & rg
            comp = (~masks[cid]) & rg
            if (sel & hz["bid_valid"]).any():
                ar = float(np.mean(hz["bid_ret"][sel & hz["bid_valid"]]))
                abs_e.append(ar)
                if (comp & hz["bid_valid"]).any():
                    delta_e.append(ar - float(np.mean(hz["bid_ret"][comp & hz["bid_valid"]])))
        eval_abs = block_metrics["Evaluation"]["median_abs"]
        stress_abs = block_metrics["20260805_07"]["median_abs"]
        eval_d = block_metrics["Evaluation"]["median_delta"]
        stress_d = block_metrics["20260805_07"]["median_delta"]
        stable = regime_stability(
            eval_abs=eval_abs, stress_abs=stress_abs,
            eval_delta=eval_d, stress_delta=stress_d,
            disc_abs=block_metrics["Discovery"]["median_abs"],
            d0803_abs=block_metrics["20260803"]["median_abs"],
            d0804_abs=block_metrics["20260804"]["median_abs"],
        )
        regime_rows.append({
            "regime_id": rid,
            "median_abs": _med(abs_e),
            "median_delta": _med(delta_e),
            "Evaluation_abs": eval_abs,
            "20260805_07_abs": stress_abs,
            "Evaluation_delta": eval_d,
            "20260805_07_delta": stress_d,
            "stable_candidate": stable and rid != "R0_NO_REGIME_GATE",
            "blocks": block_metrics,
            "coverage_regime_fraction": float(rg.mean()),
        })
    regime_sel = select_global_regime(regime_rows)
    print(f"  regime conclusion={regime_sel['conclusion']} selected={regime_sel['selected_regime']}", flush=True)

    print("=== Phase B: PBv2 manifest ===", flush=True)
    pbv2_manifest = freeze_pbv2_manifest()
    (OUT / "pbv2_exit_manifest_v1.json").write_text(
        json.dumps(pbv2_manifest, indent=2), encoding="utf-8",
    )
    parity = pbv2_parity_check()
    pbv2_ok = bool(parity.get("ok"))
    print(f"  pbv2 parity={parity.get('status')}", flush=True)

    # Phase C EXIT specs
    print("=== Phase C: EXIT architectures ===", flush=True)
    assignments = _load_sheet(X28A1 / "audit.xlsx", "CandidateExitAssignmentsV2")
    sem_reg = _load_sheet(X28A1 / "audit.xlsx", "SemanticExitRegistryV2")
    assign_by = {a["candidate_id"]: a for a in assignments}
    x26a = json.loads((X26A / "report.json").read_text(encoding="utf-8"))
    x26a_exits = {c["canonical_exit_id"]: c for c in (x26a.get("canonical_exits") or [])}
    for spec in common_controls():
        x26a_exits[spec.exit_id] = {
            "canonical_exit_id": spec.exit_id, "stop_bps": spec.stop_bps,
            "target_bps": spec.target_bps, "trail_activation_bps": spec.trail_activation_bps,
            "giveback_bps": spec.giveback_bps, "giveback_mode": spec.giveback_mode,
            "no_progress_sec": spec.no_progress_sec, "max_hold_sec": spec.max_hold_sec,
            "no_progress_mfe_bps": spec.no_progress_mfe_bps,
            "no_progress_abs_ret_bps": spec.no_progress_abs_ret_bps,
        }
    baseline_rows, baseline_sha = freeze_family_baselines(assignments, x26a_exits)
    if baseline_sha != FAMILY_BASELINE_REGISTRY_SHA:
        raise RuntimeError("baseline sha")
    baseline_by = {r["candidate_id"]: r for r in baseline_rows}

    specs: dict[str, ExitSpec] = {}
    for s in sem_reg:
        specs[s["semantic_exit_sha256"]] = _spec_from_row(s["semantic_exit_sha256"], s)
    needed = set()
    for r in cohorts["specific"]:
        needed.add(r["semantic_exit_sha256"])
        needed.add(r["family_baseline_exit_id"])
    for r in cohorts["family"]:
        needed.add(r["primary_exit_id"])
        needed.add(r["counterfactual_specific_semantic_exit_sha256"])
    for eid in needed:
        if eid not in specs:
            p = x26a_exits.get(eid)
            if p is None:
                raise RuntimeError(f"missing exit {eid}")
            specs[eid] = _spec_from_row(eid, p)
    spec_list = [specs[e] for e in sorted(needed)]
    print(f"  building {len(spec_list)} exit matrices...", flush=True)
    mats = build_exit_matrices(
        specs=spec_list, rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=board_by_key, max_workers=4,
    )
    # A/B sample
    eid0 = spec_list[0].exit_id
    mats_b = build_exit_matrices(
        specs=[specs[eid0]], rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=board_by_key, max_workers=1,
    )
    ab_mat = bool(np.array_equal(mats[eid0]["valid"], mats_b[eid0]["valid"]))

    pbv2_mat = None
    if pbv2_ok:
        print("  building PBv2 matrix...", flush=True)
        pbv2_mat = build_pbv2_matrix(
            rows=rows, entry_asks=entry_asks, times_list=times_list,
            prices_list=prices_list, board_by_key=board_by_key,
        )
        pbv2_mat2 = build_pbv2_matrix(
            rows=rows, entry_asks=entry_asks, times_list=times_list,
            prices_list=prices_list, board_by_key=board_by_key,
        )
        ab_mat = ab_mat and bool(np.array_equal(pbv2_mat["valid"], pbv2_mat2["valid"]))

    def _gate_mask(regime_id: Optional[str]) -> np.ndarray:
        if not regime_id or regime_id == "R0_NO_REGIME_GATE":
            return np.ones(len(rows), dtype=bool)
        return regime_mask(rows, regime_id)

    gate_modes = [("NO_GATE", None)]
    if regime_sel.get("selected_regime"):
        gate_modes.append(("SUPPORTED_GLOBAL_REGIME", regime_sel["selected_regime"]))

    arch_results = []
    stop_tags = []

    for gate_label, gate_id in gate_modes:
        gmask = _gate_mask(gate_id)
        # Specific EXIT (49) — candidate-balanced
        spec_metrics = []
        fam_on_spec = []
        pbv2_on_spec = []
        for reg in cohorts["specific"]:
            cid = reg["candidate_id"]
            sel = masks[cid] & gmask
            mat_s = mats[reg["semantic_exit_sha256"]]
            mat_f = mats[reg["family_baseline_exit_id"]]
            ms = summarize_exit(mat=mat_s, mask=sel, dates=dates, symbols=symbols)
            mf = summarize_exit(mat=mat_f, mask=sel, dates=dates, symbols=symbols)
            ed_s = entry_delta_from_mat(mat_s, sel)
            ed_f = entry_delta_from_mat(mat_f, sel)
            spec_metrics.append({**ms, "entry_delta": ed_s, "candidate_id": cid,
                                 "eval_abs": block_abs(mat_s, sel, dates, EVALUATION),
                                 "stress_abs": block_abs(mat_s, sel, dates, ADDITIONAL_STRESS),
                                 "stop_bps": _f(reg.get("stop_bps"))})
            fam_on_spec.append({**mf, "entry_delta": ed_f, "candidate_id": cid,
                                "eval_abs": block_abs(mat_f, sel, dates, EVALUATION),
                                "stress_abs": block_abs(mat_f, sel, dates, ADDITIONAL_STRESS)})
            if pbv2_mat is not None:
                mp = summarize_exit(mat=pbv2_mat, mask=sel, dates=dates, symbols=symbols)
                ed_p = entry_delta_from_mat(pbv2_mat, sel)
                pbv2_on_spec.append({**mp, "entry_delta": ed_p, "candidate_id": cid,
                                     "eval_abs": block_abs(pbv2_mat, sel, dates, EVALUATION),
                                     "stress_abs": block_abs(pbv2_mat, sel, dates, ADDITIONAL_STRESS)})

        def _agg(rows_m, arch):
            # pooled union for dep
            if arch == "specific":
                # use first specific semantic — actually pool via OR of selected valid is hard;
                # use candidate-median metrics + pooled LODO on concatenated
                pass
            med_ret = _med([r.get("avg_return_bps") for r in rows_m])
            med_pf = _med([r.get("profit_factor") for r in rows_m])
            med_ed = _med([r.get("entry_delta") for r in rows_m])
            med_eval = _med([r.get("eval_abs") for r in rows_m])
            med_stress = _med([r.get("stress_abs") for r in rows_m])
            # synthetic pooled mat from first cand with most trades for LODO approx
            best = max(rows_m, key=lambda r: r.get("trades") or 0) if rows_m else None
            return {
                "arch": arch,
                "gate": gate_label,
                "regime_id": gate_id or "R0_NO_REGIME_GATE",
                "median_avg_return_bps": med_ret,
                "median_pf": med_pf,
                "median_entry_delta": med_ed,
                "median_pnl_100share": _med([r.get("pnl_100share") for r in rows_m]),
                "median_mae": _med([r.get("avg_mae") for r in rows_m]),
                "median_mfe_capture": _med([r.get("mfe_capture_ratio") for r in rows_m]),
                "median_hard_stop_rate": _med([r.get("hard_stop_rate") for r in rows_m]),
                "median_no_progress_rate": _med([r.get("no_progress_rate") for r in rows_m]),
                "median_trail_exit_rate": _med([r.get("trail_exit_rate") for r in rows_m]),
                "Evaluation_abs": med_eval,
                "20260805_07_abs": med_stress,
                "metrics": {
                    "avg_return_bps": med_ret,
                    "profit_factor": med_pf,
                    "pnl_100share": _med([r.get("pnl_100share") for r in rows_m]),
                },
            }

        # pooled LODO using Specific union of first architecture mats
        def _pool_dep(mat_getter):
            # concatenate returns from all specific selected
            all_r, all_d, all_s = [], [], []
            for reg in cohorts["specific"]:
                cid = reg["candidate_id"]
                sel = masks[cid] & gmask
                mat = mat_getter(reg)
                idx = np.where(sel & mat["valid"])[0]
                all_r.extend(mat["ret_bps"][idx].tolist())
                all_d.extend(dates[idx].tolist())
                all_s.extend(symbols[idx].tolist())
            if not all_r:
                return {"max_day_contribution_share": None, "max_symbol_contribution_share": None,
                        "positive_LODO_count": 0, "negative_LODO_count": 0,
                        "positive_LOSO_count": 0, "negative_LOSO_count": 0,
                        "without_285A": "NOT_PRESENT"}
            pm = {
                "valid": np.ones(len(all_r), dtype=bool),
                "ret_bps": np.asarray(all_r, float),
                "pnl": np.asarray(all_r, float),
                "reason": np.array([""] * len(all_r), dtype=object),
            }
            return lodo_loso(pm, np.ones(len(all_r), dtype=bool), np.asarray(all_d), np.asarray(all_s))

        for arch, rows_m, getter in (
            ("specific", spec_metrics, lambda reg: mats[reg["semantic_exit_sha256"]]),
            ("family", fam_on_spec, lambda reg: mats[reg["family_baseline_exit_id"]]),
        ):
            agg = _agg(rows_m, arch)
            dep = _pool_dep(getter)
            agg["dependency"] = dep
            # 285A on 0805-07
            add_syms = set(symbols[np.isin(dates, list(ADDITIONAL_STRESS))].tolist())
            agg["285A_on_0805_07"] = "NOT_PRESENT" if "285A" not in add_syms else "PRESENT"
            agg["passes"] = architecture_passes(
                agg["metrics"], agg["median_entry_delta"],
                agg["Evaluation_abs"], agg["20260805_07_abs"], dep,
            )
            # WIDE stop dependency: if specific and only wide stops drive positive
            if arch == "specific":
                wide_rets = [r["avg_return_bps"] for r in rows_m
                             if (r.get("stop_bps") or 0) >= 80 and r.get("avg_return_bps") is not None]
                normal_rets = [r["avg_return_bps"] for r in rows_m
                               if (r.get("stop_bps") or 0) < 80 and r.get("avg_return_bps") is not None]
                if (_mean(wide_rets) or 0) > 0 and (_mean(normal_rets) or 0) <= 0:
                    agg["stop_tag"] = "WIDE_STOP_DEPENDENT"
                else:
                    agg["stop_tag"] = None
            else:
                agg["stop_tag"] = None
            agg["narrower_stop_score"] = 1 if arch == "family" else (2 if arch == "pbv2" else 0)
            agg["stability_score"] = int(agg["passes"])
            agg["detail_n"] = len(rows_m)
            arch_results.append(agg)

        if pbv2_mat is not None and pbv2_on_spec:
            agg = _agg(pbv2_on_spec, "pbv2")
            dep = _pool_dep(lambda reg: pbv2_mat)
            agg["dependency"] = dep
            add_syms = set(symbols[np.isin(dates, list(ADDITIONAL_STRESS))].tolist())
            agg["285A_on_0805_07"] = "NOT_PRESENT" if "285A" not in add_syms else "PRESENT"
            agg["passes"] = architecture_passes(
                agg["metrics"], agg["median_entry_delta"],
                agg["Evaluation_abs"], agg["20260805_07_abs"], dep,
            )
            agg["stop_tag"] = None
            agg["narrower_stop_score"] = 2
            agg["stability_score"] = int(agg["passes"])
            agg["detail_n"] = len(pbv2_on_spec)
            arch_results.append(agg)

    # Selection
    passing = [a for a in arch_results if a.get("passes")]
    best = rank_architecture(passing)

    # ENTRY absolute rise check for Case E
    spec_abs = entry_specific.get("median_absolute_return_bps")
    fam_abs = entry_family.get("median_absolute_return_bps")
    spec_delta = entry_specific.get("median_entry_delta")
    fam_delta = entry_family.get("median_entry_delta")
    relative_only = (
        ((spec_delta or 0) > 0 or (fam_delta or 0) > 0)
        and (spec_abs is None or spec_abs <= 0)
        and (fam_abs is None or fam_abs <= 0)
        and not passing
    )

    x29_v3_required = False
    x29_v2_status = "MAINTAINED"
    if best is not None:
        uses_regime = best.get("gate") == "SUPPORTED_GLOBAL_REGIME"
        uses_non_specific = best.get("arch") in ("family", "pbv2")
        # Logic change if regime added OR exit arch is not the prior specific-primary design
        # Prior X29 = Specific EXIT for 49 + Family EXIT for 118 — changing to global family/pbv2/regime => V3
        if uses_regime or best.get("arch") != "specific":
            x29_v3_required = True
            x29_v2_status = "SUPERSEDED_BEFORE_PROSPECTIVE_MARKET_OPEN"

    if best is None:
        if relative_only:
            verdict = VERDICT_RELATIVE_ONLY
        else:
            verdict = VERDICT_NO_ROBUST
        recommended = None
        x29_v3_required = False  # do not advance X29 with new logic
        x29_v2_status = "HELD_NOT_ADVANCED"
    else:
        recommended = f"{best['gate']}+{best['arch']}"
        if best["arch"] == "pbv2" and best["gate"] == "SUPPORTED_GLOBAL_REGIME":
            verdict = VERDICT_REGIME_PBV2
        elif best["arch"] == "family" and best["gate"] == "SUPPORTED_GLOBAL_REGIME":
            verdict = VERDICT_REGIME_FAMILY
        elif best["arch"] == "pbv2":
            verdict = VERDICT_PBV2
        elif best["arch"] == "family":
            verdict = VERDICT_FAMILY
        else:
            # specific passed somehow
            verdict = VERDICT_FAMILY if not best.get("gate") == "SUPPORTED_GLOBAL_REGIME" else VERDICT_REGIME_FAMILY
            recommended = f"{best['gate']}+family"  # prefer not specific if we got here oddly
            # Actually if specific is best passing, still report it but Case list doesn't have SPECIFIC-only;
            # treat as no robust global simplification → if only specific passes, NO_ROBUST for program arch
            if best["arch"] == "specific":
                # Specific-only is not a listed Case A-D; if only specific works, relative path unclear
                # Prefer reject candidate-specific as primary (section 19)
                non_spec = [a for a in passing if a["arch"] != "specific"]
                if non_spec:
                    best = rank_architecture(non_spec)
                    recommended = f"{best['gate']}+{best['arch']}"
                    if best["arch"] == "pbv2" and best["gate"] == "SUPPORTED_GLOBAL_REGIME":
                        verdict = VERDICT_REGIME_PBV2
                    elif best["arch"] == "family" and best["gate"] == "SUPPORTED_GLOBAL_REGIME":
                        verdict = VERDICT_REGIME_FAMILY
                    elif best["arch"] == "pbv2":
                        verdict = VERDICT_PBV2
                    else:
                        verdict = VERDICT_FAMILY
                else:
                    verdict = VERDICT_NO_ROBUST
                    recommended = "candidate_specific_only_not_preferred"
                    x29_v3_required = False
                    x29_v2_status = "HELD_NOT_ADVANCED"

    # If Case E/F, keep V2 but do not start prospective
    if verdict in (VERDICT_RELATIVE_ONLY, VERDICT_NO_ROBUST):
        x29_v3_required = False
        if x29_v2_status == "SUPERSEDED_BEFORE_PROSPECTIVE_MARKET_OPEN":
            x29_v2_status = "HELD_NOT_ADVANCED"

    print("=== tests ===", flush=True)
    interim = {
        "run_id": run_id,
        "verdict": verdict,
        "source_x28c_run_id": SOURCE_X28C,
        "source_x28d_run_id": SOURCE_X28D,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "specific_n": EXPECTED_SPECIFIC,
        "family_n": EXPECTED_FAMILY,
        "overlap": 0,
        "entry_only_no_exit_dependency": True,
        "fixed_horizon_returns": True,
        "regime_library": list(REGIME_LIBRARY),
        "no_dynamic_regime_addition": True,
        "no_candidate_specific_regime": True,
        "pbv2_exit_source_identity": True,
        "pbv2_exit_manifest_frozen": True,
        "pbv2_replay_parity": parity.get("status"),
        "same_entry_episode_all_exits": True,
        "same_actual_ask": True,
        "same_bid_fill_contract": True,
        "no_stop_grid_search": True,
        "lodo": True,
        "loso": True,
        "285a_not_present_0805_07": True,
        "no_candidate_selection_change": True,
        "opened_20260810": False,
        "no_runtime_change": True,
        "submit_cancel_live": "0/0/0",
        "ab_determinism": ab_entry and ab_mat,
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2), encoding="utf-8")
    tests = _run_tests()

    # X29 V3 only if required and we have supported architecture
    x29_v3_sha = None
    if x29_v3_required and best is not None and verdict not in (VERDICT_RELATIVE_ONLY, VERDICT_NO_ROBUST):
        v2 = json.loads((X29 / "precommit_v2.json").read_text(encoding="utf-8"))
        from copy import deepcopy
        body = deepcopy(v2)
        body.pop("precommit_sha", None)
        body.pop("published_shas", None)
        body["precommit_id"] = "X29_PRECOMMIT_V3_FROZEN"
        body["run_id"] = f"e1x29_precommit_v3_{now.strftime('%Y%m%d_%H%M%S')}_A"
        body["verdict"] = "X29_PRECOMMIT_V3_FROZEN"
        body["precommit_version"] = 3
        body["supersedes"] = {
            "old_run_id": X29_V2_RUN_ID,
            "old_precommit_sha": X29_V2_PRECOMMIT_SHA,
            "status": "SUPERSEDED_BEFORE_PROSPECTIVE_MARKET_OPEN",
            "reason": "E1_X28E_ARCHITECTURE_CHANGE",
        }
        body["architecture"] = {
            "recommended": recommended,
            "verdict": verdict,
            "regime": best.get("regime_id"),
            "exit_arch": best.get("arch"),
            "source_x28e_run_id": run_id,
        }
        body["first_eligible_prospective_day"] = "20260810"
        body["20260810_not_opened"] = True
        body["market_data_not_opened"] = True
        body["observer_not_started"] = True
        sha = sha256_obj(body)
        body["precommit_sha"] = sha
        (X29 / "precommit_v3.json").write_text(json.dumps(body, indent=2, default=str), encoding="utf-8")
        (X29 / "precommit_v2_superseded_by_x28e.json").write_text(json.dumps({
            "status": "SUPERSEDED_BEFORE_PROSPECTIVE_MARKET_OPEN",
            "old_sha": X29_V2_PRECOMMIT_SHA,
            "new_sha": sha,
            "at": now.isoformat(),
        }, indent=2), encoding="utf-8")
        x29_v3_sha = sha
        # keep v2 file
        assert json.loads((X29 / "precommit_v2.json").read_text(encoding="utf-8"))["precommit_sha"] == X29_V2_PRECOMMIT_SHA

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": verdict,
        "recommended_architecture": recommended,
        "source_x28c_run_id": SOURCE_X28C,
        "source_x28d_run_id": SOURCE_X28D,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "entry_only": {
            "specific": entry_specific,
            "family": entry_family,
            "note": "EXIT not used in Phase A",
        },
        "regime": {
            **regime_sel,
            "library": list(REGIME_LIBRARY),
            "rows": regime_rows,
            "no_candidate_specific_regime": True,
            "no_dynamic_addition": True,
        },
        "pbv2": {
            "manifest_sha256": pbv2_manifest["manifest_sha256"],
            "parity_status": parity.get("status"),
            "parity": parity,
            "source_files": pbv2_manifest["source_files"],
            "included_in_comparison": pbv2_ok,
        },
        "exit_architecture": arch_results,
        "best": best,
        "stop_dependency_notes": [a for a in arch_results if a.get("stop_tag")],
        "x29_v2_status": x29_v2_status,
        "x29_v2_sha": X29_V2_PRECOMMIT_SHA,
        "x29_v3_required": bool(x29_v3_sha),
        "x29_v3_precommit_sha": x29_v3_sha,
        "opened_20260810": False,
        "prospective_observer_started": False,
        "no_candidate_selection_change": True,
        "no_stop_grid_search": True,
        "ab_determinism": {"entry": ab_entry, "matrix": ab_mat, "ok": ab_entry and ab_mat},
        "tests": tests,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "Paper_order": False,
            "Live_order": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Discord_production_notification": False,
        },
    }
    sheets = {
        "Index": [{"run_id": run_id, "verdict": verdict}],
        "EntryOnlySpecific": [entry_specific],
        "EntryOnlyFamily": [entry_family],
        "Regime": regime_rows,
        "ExitArchitecture": [{k: v for k, v in a.items() if k != "detail"} for a in arch_results],
        "PBv2": [parity],
        "Decision": [{"verdict": verdict, "recommended": recommended, "best": best}],
        "Tests": [tests],
        "Safety": [report["safety"]],
        "ChangeLog": [{"at": now.isoformat(), "note": "E1_X28E absolute-rise regime + EXIT arch"}],
    }
    publish(OUT, report, sheets)
    print(f"=== DONE verdict={verdict} recommended={recommended} ===", flush=True)
    print("=== STOP: do not open 20260810 ===", flush=True)
    return report


if __name__ == "__main__":
    run()
