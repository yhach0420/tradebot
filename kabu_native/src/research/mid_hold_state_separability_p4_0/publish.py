"""Write P4-0 report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.canonical_fixed_pnl_source_p3_3.ledger import pnl
from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.mid_hold_state_separability_p4_0 import (
    ANALYSIS_ID,
    CHECKPOINTS_SEC,
    DOCUMENT_ID,
    P1_PF,
    P1_PNL,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    STATE_VARS,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.mid_hold_state_separability_p4_0.metrics import (
    auc_block,
    classify_gate,
    eligible,
    preservation,
    quintiles,
    recovery,
    separation_block,
    slice_days,
    time_consistency,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "mid_hold_state_separability_p4_0"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
P3_4R_REPORT = NATIVE / "results" / "research" / "extension_decision_alignment_p3_4r" / "report.json"


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, np.generic):
        obj = obj.item()
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, max(12, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def load_frozen() -> dict[str, Any]:
    if not P3_4R_REPORT.is_file():
        return {"ok": False, "reason": "NO_P3_4R_REPORT"}
    p = json.loads(P3_4R_REPORT.read_text(encoding="utf-8"))
    if p.get("DECISION_RECONCILE") != "PASS":
        return {"ok": False, "reason": "P3_4R_DECISION_RECONCILE_NOT_PASS"}
    if p.get("EXTENSION_GATE_REVISED") != "EXTENSION_GATE_SUPPORTED":
        return {"ok": False, "reason": "P3_4R_GATE_NOT_SUPPORTED"}
    return {
        "ok": True,
        "DECISION_RECONCILE": p.get("DECISION_RECONCILE"),
        "EXTENSION_GATE_REVISED": p.get("EXTENSION_GATE_REVISED"),
        "note": "600s Continuation Gate frozen. Not retuned in P4-0.",
    }


def _pack_horizon(rows: list[dict[str, Any]], h: int) -> dict[str, Any]:
    el = eligible(rows, h)
    sep = {v: separation_block(el, v) for v in STATE_VARS}
    auc = {v: auc_block(el, v) for v in STATE_VARS}
    qnt = {v: quintiles(el, v) for v in STATE_VARS}
    return {
        "n_eligible": len(el),
        "n_win": sum(1 for r in el if r.get("CANONICAL_FINAL_WIN")),
        "n_loss": sum(1 for r in el if r.get("CANONICAL_FINAL_LOSS")),
        "n_draw": sum(1 for r in el if r.get("CANONICAL_FINAL_DRAW")),
        "n_early_fail": sum(1 for r in el if r.get("EARLY_FAILURE_BEFORE_600")),
        "n_exit600": sum(1 for r in el if r.get("REACHED_600_EXIT")),
        "n_extend": sum(1 for r in el if r.get("REACHED_600_EXTEND")),
        "separation": sep,
        "auc": auc,
        "quintiles": qnt,
        "preservation": {
            "TOP10": preservation(el, "TOP10_CANONICAL_WINNER"),
            "TOP20": preservation(el, "TOP20_CANONICAL_WINNER"),
            "EXTEND35": preservation(el, "REACHED_600_EXTEND"),
        },
        "recovery": recovery(el),
    }


def build_report(
    *,
    path_rows: list[dict[str, Any]],
    recon: dict[str, Any],
    frozen: dict[str, Any],
    leak_n: int,
    identity_n: int,
    identity_fail: int,
    failed: list[str],
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    all_h = {h: _pack_horizon(path_rows, h) for h in CHECKPOINTS_SEC} if path_rows else {}
    top_rows = slice_days(path_rows, PREDECLARED_TOP3) if path_rows else []
    rest_rows = slice_days(path_rows, REST11) if path_rows else []
    top_h = {h: _pack_horizon(top_rows, h) for h in CHECKPOINTS_SEC} if path_rows else {}
    rest_h = {h: _pack_horizon(rest_rows, h) for h in CHECKPOINTS_SEC} if path_rows else {}

    consist_all = {}
    consist_rest = {}
    if path_rows:
        sep_all = {h: all_h[h]["separation"] for h in CHECKPOINTS_SEC}
        sep_rest = {h: rest_h[h]["separation"] for h in CHECKPOINTS_SEC}
        consist_all = {v: time_consistency(sep_all, v) for v in STATE_VARS}
        consist_rest = {v: time_consistency(sep_rest, v) for v in STATE_VARS}

    auc_rest = {h: rest_h[h]["auc"] for h in CHECKPOINTS_SEC} if path_rows else {}
    pres_t10 = {h: all_h[h]["preservation"]["TOP10"] for h in CHECKPOINTS_SEC} if path_rows else {}
    recov_all = {h: all_h[h]["recovery"] for h in CHECKPOINTS_SEC} if path_rows else {}
    gate = (
        classify_gate(rest_consist=consist_rest, auc_rest=auc_rest, pres_top10=pres_t10, recov=recov_all)
        if path_rows
        else {"MID_HOLD_GATEABILITY": None, "CANDIDATE_STATE_FAMILIES": []}
    )

    leak_ok = int(leak_n) == 0
    ident_ok = int(identity_fail) == 0
    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if not frozen.get("ok"):
        integrity.append(frozen.get("reason") or "P3_4R_FROZEN_MISSING")
    if path_rows and not leak_ok:
        integrity.append("FUTURE_LEAK")
    if path_rows and not ident_ok:
        integrity.append(f"IDENTITY_FAIL_{identity_fail}/{identity_n}")
    if path_rows and len({str(r.get("trade_id")) for r in path_rows if r.get("horizon_sec") == 120}) != P1_TRADES:
        n120 = len({str(r.get("trade_id")) for r in path_rows if r.get("horizon_sec") == 120})
        integrity.append(f"TRADE_GRID_{n120}/{P1_TRADES}")

    if blocked or not recon.get("pass") or not frozen.get("ok"):
        verdict = VERDICT_BLOCKED
    elif integrity:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    elig_n = {str(h): (all_h.get(h) or {}).get("n_eligible") for h in CHECKPOINTS_SEC}

    def _med(h, var, side):
        return ((((all_h.get(h) or {}).get("separation") or {}).get(var) or {}).get(side) or {}).get("median")

    state_sep_summary = {}
    for v in STATE_VARS:
        state_sep_summary[v] = {
            str(h): {
                "WIN_median": _med(h, v, "WIN"),
                "LOSS_median": _med(h, v, "LOSS"),
                "direction": (((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("direction_win_vs_loss"),
            }
            for h in CHECKPOINTS_SEC
        }

    auc_sum = {
        str(h): {
            v: {
                "fail_vs_win_auc_best": ((((all_h.get(h) or {}).get("auc") or {}).get(v) or {}).get("fail_vs_win") or {}).get(
                    "auc_best"
                ),
                "extend_vs_exit_auc_best": (
                    (((all_h.get(h) or {}).get("auc") or {}).get(v) or {}).get("extend_vs_exit600") or {}
                ).get("auc_best"),
            }
            for v in STATE_VARS
        }
        for h in CHECKPOINTS_SEC
    }

    sheets = {
        "eligible": [
            {
                "horizon_sec": h,
                "n_rows": sum(1 for r in path_rows if int(r.get("horizon_sec") or 0) == h),
                "n_eligible": (all_h.get(h) or {}).get("n_eligible"),
                "n_already_exited": sum(
                    1
                    for r in path_rows
                    if int(r.get("horizon_sec") or 0) == h and r.get("uneval_reason") == "ALREADY_EXITED"
                ),
                "n_path_not_evaluable": sum(
                    1
                    for r in path_rows
                    if int(r.get("horizon_sec") or 0) == h and str(r.get("status")) == "PATH_NOT_EVALUABLE"
                ),
                "n_win": (all_h.get(h) or {}).get("n_win"),
                "n_loss": (all_h.get(h) or {}).get("n_loss"),
                "n_extend": (all_h.get(h) or {}).get("n_extend"),
                "n_exit600": (all_h.get(h) or {}).get("n_exit600"),
            }
            for h in CHECKPOINTS_SEC
        ],
        "state": [
            {
                "horizon_sec": h,
                "var": v,
                "win_median": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("WIN") or {}).get("median"),
                "win_p25": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("WIN") or {}).get("p25"),
                "win_p75": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("WIN") or {}).get("p75"),
                "loss_median": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("LOSS") or {}).get("median"),
                "loss_p25": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("LOSS") or {}).get("p25"),
                "loss_p75": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("LOSS") or {}).get("p75"),
                "direction": (((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("direction_win_vs_loss"),
            }
            for h in CHECKPOINTS_SEC
            for v in STATE_VARS
        ],
        "winloss": [
            {
                "horizon_sec": h,
                "side": side,
                **((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get(side) or {}),
                "var": v,
            }
            for h in CHECKPOINTS_SEC
            for v in STATE_VARS
            for side in ("WIN", "LOSS")
        ],
        "early": [
            {
                "horizon_sec": h,
                "n_early_fail_eligible": (all_h.get(h) or {}).get("n_early_fail"),
                "n_eligible": (all_h.get(h) or {}).get("n_eligible"),
            }
            for h in CHECKPOINTS_SEC
        ],
        "ex600": [
            {
                "horizon_sec": h,
                "var": v,
                "extend_median": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("EXTEND") or {}).get(
                    "median"
                ),
                "exit600_median": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("EXIT600") or {}).get(
                    "median"
                ),
                "extend_n": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("EXTEND") or {}).get("n"),
                "exit600_n": ((((all_h.get(h) or {}).get("separation") or {}).get(v) or {}).get("EXIT600") or {}).get("n"),
            }
            for h in CHECKPOINTS_SEC
            for v in STATE_VARS
        ],
        "auc": [
            {
                "horizon_sec": h,
                "var": v,
                "fail_vs_win_auc": ((((all_h.get(h) or {}).get("auc") or {}).get(v) or {}).get("fail_vs_win") or {}).get("auc"),
                "fail_vs_win_auc_best": (
                    (((all_h.get(h) or {}).get("auc") or {}).get(v) or {}).get("fail_vs_win") or {}
                ).get("auc_best"),
                "extend_vs_exit_auc": (
                    (((all_h.get(h) or {}).get("auc") or {}).get(v) or {}).get("extend_vs_exit600") or {}
                ).get("auc"),
                "extend_vs_exit_auc_best": (
                    (((all_h.get(h) or {}).get("auc") or {}).get(v) or {}).get("extend_vs_exit600") or {}
                ).get("auc_best"),
            }
            for h in CHECKPOINTS_SEC
            for v in STATE_VARS
        ],
        "qnt": [
            {
                "horizon_sec": h,
                "var": v,
                "q": b.get("q"),
                "lo": b.get("lo"),
                "hi": b.get("hi"),
                "n": b.get("n"),
                "loss_rate": b.get("loss_rate"),
                "win_rate": b.get("win_rate"),
                "extend_rate": b.get("extend_rate"),
                "monotonic_loss": (((all_h.get(h) or {}).get("quintiles") or {}).get(v) or {}).get("monotonic_loss"),
            }
            for h in CHECKPOINTS_SEC
            for v in STATE_VARS
            for b in ((((all_h.get(h) or {}).get("quintiles") or {}).get(v) or {}).get("bins") or [])
        ],
        "pres": [
            {
                "horizon_sec": h,
                "group": g,
                **((((all_h.get(h) or {}).get("preservation") or {}).get(g) or {})),
            }
            for h in CHECKPOINTS_SEC
            for g in ("TOP10", "TOP20", "EXTEND35")
        ],
        "recov": [{"horizon_sec": h, **((all_h.get(h) or {}).get("recovery") or {})} for h in CHECKPOINTS_SEC],
        "top3rest": [
            {
                "slice": sl,
                "horizon_sec": h,
                "var": v,
                "direction": (((blk.get(h) or {}).get("separation") or {}).get(v) or {}).get("direction_win_vs_loss"),
                "win_median": ((((blk.get(h) or {}).get("separation") or {}).get(v) or {}).get("WIN") or {}).get("median"),
                "loss_median": ((((blk.get(h) or {}).get("separation") or {}).get(v) or {}).get("LOSS") or {}).get("median"),
                "fail_vs_win_auc_best": (
                    (((blk.get(h) or {}).get("auc") or {}).get(v) or {}).get("fail_vs_win") or {}
                ).get("auc_best"),
            }
            for sl, blk in (("TOP3", top_h), ("REST11", rest_h))
            for h in CHECKPOINTS_SEC
            for v in STATE_VARS
        ],
        "time": [
            {"slice": sl, "var": v, **(blk.get(v) or {})}
            for sl, blk in (("ALL", consist_all), ("REST11", consist_rest))
            for v in STATE_VARS
        ],
    }

    report = {
        "task": "P4-0",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "strategy validation", "new EXIT validation"],
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "PRIMARY_FULL14": {"trades": P1_TRADES, "pnl": P1_PNL, "PF": P1_PF},
        "FROZEN_P3_4R": frozen,
        "ELIGIBLE_BY_CHECKPOINT": elig_n,
        "STATE_SEPARATION": state_sep_summary,
        "AUC_SUMMARY": auc_sum,
        "WINNER_PRESERVATION": {str(h): (all_h.get(h) or {}).get("preservation") for h in CHECKPOINTS_SEC},
        "RECOVERY": {str(h): (all_h.get(h) or {}).get("recovery") for h in CHECKPOINTS_SEC},
        "REST11": {
            "time_consistency": consist_rest,
            "eligible": {str(h): (rest_h.get(h) or {}).get("n_eligible") for h in CHECKPOINTS_SEC},
        },
        "TIME_CONSISTENCY": {"ALL": consist_all, "REST11": consist_rest},
        "MID_HOLD_GATEABILITY": gate.get("MID_HOLD_GATEABILITY"),
        "MID_HOLD_GATEABILITY_DETAIL": gate,
        "CANDIDATE_STATE_FAMILIES": gate.get("CANDIDATE_STATE_FAMILIES") or [],
        "THRESHOLD_SELECTED": False,
        "NEW_EXIT_TESTED": False,
        "COUNTERFACTUAL_PNL_TESTED": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "FUTURE_LEAK": (not leak_ok) if path_rows else False,
        "leak_n": int(leak_n),
        "identity_n": int(identity_n),
        "identity_fail": int(identity_fail),
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "by_horizon": {str(h): {k: v for k, v in blk.items() if k != "quintiles"} for h, blk in all_h.items()},
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_p4_0_state": _file_sha("src/research/mid_hold_state_separability_p4_0/state.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
    }
    unused = pnl
    del unused
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    el = rep.get("ELIGIBLE_BY_CHECKPOINT") or {}
    gate = rep.get("MID_HOLD_GATEABILITY_DETAIL") or {}
    pres = rep.get("WINNER_PRESERVATION") or {}
    rec = rep.get("RECOVERY") or {}
    lines = [
        "# P4-0 Mid-Hold State Separability",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "Not: OOS / prospective / robust / strategy validation / new EXIT validation",
        "",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        "",
        "ELIGIBLE_BY_CHECKPOINT:",
    ]
    for h in CHECKPOINTS_SEC:
        lines.append(f"{h}: {el.get(str(h))}")
    sep = rep.get("STATE_SEPARATION") or {}
    auc = rep.get("AUC_SUMMARY") or {}
    lines += ["", "STATE_SEPARATION:"]
    for v in STATE_VARS:
        lines.append(f"{v}: {json.dumps(sep.get(v), ensure_ascii=False, default=str)}")
    lines += ["", "AUC_SUMMARY:"]
    for h in CHECKPOINTS_SEC:
        br = ((auc.get(str(h)) or {}).get("bid_return_from_fill") or {})
        lines.append(f"{h} bid_return fail_vs_win_auc_best={br.get('fail_vs_win_auc_best')} extend_vs_exit={br.get('extend_vs_exit_auc_best')}")
    lines += ["", "WINNER_PRESERVATION:"]
    for g in ("TOP10", "TOP20", "EXTEND35"):
        freqs = {
            str(h): ((pres.get(str(h)) or {}).get(g) or {}).get("adverse_state_frequency")
            for h in CHECKPOINTS_SEC
        }
        lines.append(f"{g} adverse_state_frequency: {json.dumps(freqs, ensure_ascii=False)}")
    lines += ["", "RECOVERY:"]
    for h in CHECKPOINTS_SEC:
        r = rec.get(str(h)) or {}
        lines.append(
            f"{h}: adverse_then_win={r.get('RECOVERED_TO_WIN')} "
            f"adverse_then_extend={r.get('RECOVERED_TO_EXTEND')} "
            f"adverse_then_fail={r.get('FAILED')}"
        )
    rest_c = ((rep.get("TIME_CONSISTENCY") or {}).get("REST11") or {}).get("bid_return_from_fill") or {}
    lines += [
        "",
        f"REST11 bid_return time_consistency: {json.dumps(rest_c, ensure_ascii=False, default=str)}",
        f"TIME_CONSISTENCY: {json.dumps({k: {vv: (blk.get(vv) or {}).get('majority') for vv in STATE_VARS} for k, blk in ((rep.get('TIME_CONSISTENCY') or {}).items())}, ensure_ascii=False)}",
        "",
        f"MID_HOLD_GATEABILITY: `{rep.get('MID_HOLD_GATEABILITY')}`",
        f"{gate.get('why')}",
        "",
        f"CANDIDATE_STATE_FAMILIES: {json.dumps(rep.get('CANDIDATE_STATE_FAMILIES'), ensure_ascii=False)}",
        "",
        "THRESHOLD_SELECTED: false",
        "NEW_EXIT_TESTED: false",
        "COUNTERFACTUAL_PNL_TESTED: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("MID_HOLD_GATEABILITY", public.get("MID_HOLD_GATEABILITY")),
            ("CANDIDATE_STATE_FAMILIES", public.get("CANDIDATE_STATE_FAMILIES")),
            ("THRESHOLD_SELECTED", False),
            ("NEW_EXIT_TESTED", False),
            ("COUNTERFACTUAL_PNL_TESTED", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Eligible"), sheets.get("eligible") or [])
    _write_rows(wb.create_sheet("Checkpoint_State"), sheets.get("state") or [])
    _write_rows(wb.create_sheet("Win_Loss"), sheets.get("winloss") or [])
    _write_rows(wb.create_sheet("EarlyFailure"), sheets.get("early") or [])
    _write_rows(wb.create_sheet("Exit600_Extend"), sheets.get("ex600") or [])
    _write_rows(wb.create_sheet("AUC"), sheets.get("auc") or [])
    _write_rows(wb.create_sheet("Quintiles"), sheets.get("qnt") or [])
    _write_rows(wb.create_sheet("Winner_Preservation"), sheets.get("pres") or [])
    _write_rows(wb.create_sheet("Recovery"), sheets.get("recov") or [])
    _write_rows(wb.create_sheet("Top3_Rest11"), sheets.get("top3rest") or [])
    _write_rows(wb.create_sheet("Time_Consistency"), sheets.get("time") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(
        ident,
        list((public.get("Identity") or {}).items())
        + [("identity_n", public.get("identity_n")), ("identity_fail", public.get("identity_fail"))],
    )
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("THRESHOLD_SELECTED", False),
            ("NEW_EXIT_TESTED", False),
            ("COUNTERFACTUAL_PNL_TESTED", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
