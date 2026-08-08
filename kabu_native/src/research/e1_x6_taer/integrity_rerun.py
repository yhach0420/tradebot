"""TAER economic integrity fix + same-condition 6-pair A/B rerun."""
from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.config import DAYS
from research.e1_x6_taer.exit_joint_audit import (
    PRIOR_STORE as ENTRY_PATH_STORE,
    load_entry_observations,
    decompose_s7,
    setup_path_summary,
)
from research.e1_x6_taer.exit_sm import EXIT_THRESHOLDS
from research.e1_x6_taer.integrity_replay import (
    LOCKED_P1,
    LOCKED_P2,
    PAIRS,
    noncore_gates,
    replay_pair_integrity,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
INVALID_RUN = "e1x6_taer_exit_joint_20260804_001315"


def _run_tests() -> dict:
    env = dict(os.environ)
    env["PYTHONPATH"] = f"{NATIVE / 'src'};{NATIVE / 'research'}"
    env["PYTHONIOENCODING"] = "utf-8"
    proc = subprocess.run(
        [sys.executable, "-m", "pytest", "-q", "-rA", "--tb=line",
         "-p", "no:cacheprovider",
         str(NATIVE / "tests" / "test_e1_x6_taer_integrity.py")],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(NATIVE), env=env, timeout=300,
    )
    rows = []
    for line in (proc.stdout or "").splitlines():
        for st in ("PASSED", "FAILED", "ERROR"):
            if line.strip().startswith(st + " "):
                rows.append({"test": line.strip().split(" ", 1)[1].split(" - ")[0], "outcome": st})
                break
    return {
        "exit_code": proc.returncode,
        "passed": sum(1 for r in rows if r["outcome"] == "PASSED"),
        "failed": sum(1 for r in rows if r["outcome"] != "PASSED"),
        "total": len(rows),
        "rows": rows,
        "tail": (proc.stdout or "")[-2500:],
    }


def run() -> dict:
    run_id = f"e1x6_taer_integrity_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "taer" / run_id
    store.mkdir(parents=True, exist_ok=True)

    print("=== Phase0 freeze reference ===", flush=True)
    freeze = {
        "invalid_run": INVALID_RUN,
        "status": "TAER_V1_JOINT_INVALID_ECONOMIC_INTEGRITY",
        "locked_p1": LOCKED_P1,
        "locked_p2": LOCKED_P2,
        "artifact_shas_path": str(
            Path.home() / "e1x6_research_store" / "taer" / INVALID_RUN / "ARTIFACT_SHAS.json"
        ),
    }

    print("=== Tests (must pass before economics) ===", flush=True)
    tests = _run_tests()
    print("tests", tests["passed"], "/", tests["total"], "exit", tests["exit_code"], flush=True)
    if tests["exit_code"] != 0 or tests["failed"] > 0:
        report = {
            "verdict": "TAER_V1_JOINT_INVALID_ECONOMIC_INTEGRITY",
            "run_id": run_id,
            "reason": "regression_tests_failed",
            "tests": tests,
            "safety": {"submit": 0, "cancel": 0, "live": 0},
            "mainline_changed": False,
        }
        _publish(report, store)
        return report

    # P2 reference (locked SHA — do not mutate thresholds)
    p2_ref = {
        "precommit_type": "P2_EXIT_PRECOMMIT_LOCKED_REFERENCE",
        "locked_sha256": LOCKED_P2,
        "thresholds_unchanged": EXIT_THRESHOLDS,
        "precommit_at_jst": datetime.now(JST).isoformat(),
        "note": "Same thresholds as frozen P2; new integrity replay only",
        "economics_opened_before_precommit": False,
    }
    p2_path = store / "p2_exit_precommit_lock.json"
    p2_path.write_text(json.dumps(p2_ref, ensure_ascii=False, indent=2), encoding="utf-8")
    p2_ref["file_sha256"] = sha256_file(p2_path)
    print("P2 lock ref", p2_ref["file_sha256"], "locked", LOCKED_P2, flush=True)

    entries, entry_meta = load_entry_observations()
    s7 = decompose_s7(entries)
    setup_sum = setup_path_summary(entries)

    sm = load_source_manifest()
    day_events = {}
    for day in DAYS:
        uni = _universe_from_manifest(sm, day)
        print(f"  preload {day}", flush=True)
        day_events[day] = load_day_events(day, uni)

    def run_lane(tag: str) -> dict:
        out = {}
        for _, setup, _, xc in PAIRS:
            print(f"=== {tag} {setup} x {xc} ===", flush=True)
            res = replay_pair_integrity(
                entries, setup_type=setup, exit_candidate=xc, day_events=day_events,
            )
            fp = store / f"ledger_{tag}_{setup}_{xc}.jsonl"
            with fp.open("w", encoding="utf-8") as f:
                for tr in res["trades"]:
                    f.write(json.dumps(tr, ensure_ascii=False, default=str) + "\n")
            slim = {k: v for k, v in res.items() if k != "trades"}
            slim["gates"] = noncore_gates(slim)
            slim["ledger_path"] = str(fp)
            out[res["pair_id"]] = slim
            print(
                f"  n_pass={slim['n_pass']} n_fail={slim['n_fail']} pnl={slim['pnl']:.1f} "
                f"pf={slim['pf']} sha={slim['ledger_sha256'][:12]} "
                f"integrity={slim['integrity_counts']}",
                flush=True,
            )
        return out

    print("=== Economics A ===", flush=True)
    lane_a = run_lane("A")
    print("=== Economics B (determinism) ===", flush=True)
    lane_b = run_lane("B")

    ab_ok = True
    ab_detail = {}
    for pid in lane_a:
        a, b = lane_a[pid], lane_b[pid]
        match = (
            a["ledger_sha256"] == b["ledger_sha256"]
            and a["n_all"] == b["n_all"]
            and abs(float(a["pnl"]) - float(b["pnl"])) < 1e-6
            and a["exit_reason_counts"] == b["exit_reason_counts"]
        )
        ab_detail[pid] = {
            "match": match,
            "sha_a": a["ledger_sha256"],
            "sha_b": b["ledger_sha256"],
            "n_a": a["n_all"],
            "n_b": b["n_all"],
        }
        if not match:
            ab_ok = False

    # Integrity summary across pairs (lane A)
    inv_counts = {
        "invalid_trade_count": 0,
        "cross_symbol_count": 0,
        "cross_day_count": 0,
        "cross_session_count": 0,
        "mfe_mae_violation_count": 0,
        "max_hold_violation_count": 0,
        "session_end_price_violation_count": 0,
    }
    # scan ledgers
    for _, setup, _, xc in PAIRS:
        fp = store / f"ledger_A_{setup}_{xc}.jsonl"
        with fp.open(encoding="utf-8") as f:
            for line in f:
                t = json.loads(line)
                if t["integrity_status"] == "FAIL":
                    inv_counts["invalid_trade_count"] += 1
                for reason in t.get("integrity_failure_reasons") or []:
                    if "CROSS_SYMBOL" in reason:
                        inv_counts["cross_symbol_count"] += 1
                    if "CROSS_DAY" in reason:
                        inv_counts["cross_day_count"] += 1
                    if "CROSS_SESSION" in reason:
                        inv_counts["cross_session_count"] += 1
                    if "MFE_MAE" in reason:
                        inv_counts["mfe_mae_violation_count"] += 1
                    if "MAX_HOLD" in reason:
                        inv_counts["max_hold_violation_count"] += 1
                    if "SESSION_END" in reason or reason == "NOT_EVALUABLE_SESSION_END_EXIT_PRICE":
                        inv_counts["session_end_price_violation_count"] += 1

    integrity_ok = inv_counts["invalid_trade_count"] == 0 and ab_ok
    economic_integrity_status = "PASS" if integrity_ok else "FAIL"

    any_pair_gate = any(p["gates"]["all_pass"] for p in lane_a.values())
    if not integrity_ok:
        verdict = "TAER_V1_JOINT_INVALID_ECONOMIC_INTEGRITY"
    elif not any_pair_gate:
        verdict = "E1_X6_NO_ROBUST_ENTRY_EXIT_PAIR"
    else:
        verdict = "E1_X6_RESEARCH_PAIR_PENDING_CORE_EVIDENCE"

    report = {
        "phase": "TAER_ECONOMIC_INTEGRITY_RERUN",
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "verdict": verdict,
        "freeze": freeze,
        "prior_identity_status": "EXIT_CANDIDATES_NOT_DISTINCT",
        "current_identity_status": "EXIT_PAIRS_DISTINCT",
        "economic_integrity_status": economic_integrity_status,
        "integrity_counts": inv_counts,
        "locked_p1_sha256": LOCKED_P1,
        "locked_p2_sha256": LOCKED_P2,
        "p2_lock_file_sha256": p2_ref["file_sha256"],
        "entry_meta": entry_meta,
        "s7_decomposition": s7,
        "setup_path_summary": setup_sum,
        "pairs": lane_a,
        "determinism": {"ab_ok": ab_ok, "detail": ab_detail},
        "tests": tests,
        "core_valid": 0,
        "rolling_origin_renamed_to": "retrospective_confirm_day_slice",
        "safety": {"submit": 0, "cancel": 0, "live": 0},
        "mainline_changed": False,
        "SHADOW_STARTED": False,
        "FORWARD_STARTED": False,
        "PAPER_STARTED": False,
        "DISCORD_SENT": False,
        "TAER_V2_STARTED": False,
    }
    _publish(report, store)
    return report


def _publish(report: dict, store: Path) -> None:
    out = NATIVE / "results" / "research" / "e1_x6_taer_integrity_rerun"
    out.mkdir(parents=True, exist_ok=True)
    (store / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    md = [
        f"# TAER Integrity Rerun — {report['verdict']}",
        "",
        f"- run_id: `{report.get('run_id')}`",
        f"- economic_integrity_status: `{report.get('economic_integrity_status')}`",
        f"- prior_identity_status: `{report.get('prior_identity_status')}`",
        f"- current_identity_status: `{report.get('current_identity_status')}`",
        f"- locked P1: `{report.get('locked_p1_sha256')}`",
        f"- locked P2: `{report.get('locked_p2_sha256')}`",
        f"- ab_ok: {((report.get('determinism') or {}).get('ab_ok'))}",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Pairs (PASS trades only for PnL)",
    ]
    for pid, p in (report.get("pairs") or {}).items():
        md.append(
            f"- `{pid}` pass={p.get('n_pass')} fail={p.get('n_fail')} "
            f"pnl={p.get('pnl')} pf={p.get('pf')} gates={p.get('gates', {}).get('failed')}"
        )
    (out / "report.md").write_text("\n".join(md), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    for row in (
        ("verdict", report.get("verdict")),
        ("economic_integrity_status", report.get("economic_integrity_status")),
        ("run_id", report.get("run_id")),
        ("locked_p2", report.get("locked_p2_sha256")),
    ):
        ws.append(list(row))

    def sheet(name, headers, rows):
        w = wb.create_sheet(name)
        w.append(headers)
        for r in rows:
            w.append(r)

    sheet("Precommit", ["key", "value"], [
        ["locked_p1", report.get("locked_p1_sha256")],
        ["locked_p2", report.get("locked_p2_sha256")],
        ["thresholds_unchanged", True],
    ])
    ic = report.get("integrity_counts") or {}
    sheet("IntegritySummary", ["metric", "value"], [[k, v] for k, v in {
        "economic_integrity_status": report.get("economic_integrity_status"),
        "prior_identity_status": report.get("prior_identity_status"),
        "current_identity_status": report.get("current_identity_status"),
        **ic,
    }.items()])

    # IntegrityTrades sample
    it = wb.create_sheet("IntegrityTrades")
    it.append(["pair_id", "day", "symbol", "hold_sec", "exit_reason", "entry_ask", "exit_bid",
               "mfe", "mae", "realized", "net", "integrity_status", "failures"])
    n = 0
    for _, setup, _, xc in PAIRS:
        fp = store / f"ledger_A_{setup}_{xc}.jsonl"
        if not fp.exists():
            continue
        with fp.open(encoding="utf-8") as f:
            for line in f:
                if n >= 8000:
                    break
                t = json.loads(line)
                it.append([
                    t.get("pair_id"), t.get("day"), t.get("symbol"), t.get("hold_sec"),
                    t.get("exit_reason"), t.get("entry_price_used"), t.get("exit_price_used"),
                    t.get("mfe_price_delta"), t.get("mae_price_delta"), t.get("realized_price_delta"),
                    t.get("net_pnl_yen"), t.get("integrity_status"),
                    json.dumps(t.get("integrity_failure_reasons")),
                ])
                n += 1
        if n >= 8000:
            break

    se = wb.create_sheet("SessionEndAudit")
    se.append(["pair_id", "episode_id", "symbol", "exit_bid", "boundary", "sec_to_boundary", "status"])
    mh = wb.create_sheet("MaxHoldAudit")
    mh.append(["pair_id", "episode_id", "hold_sec", "gap_sec", "reason", "status"])
    for _, setup, _, xc in PAIRS:
        fp = store / f"ledger_A_{setup}_{xc}.jsonl"
        if not fp.exists():
            continue
        with fp.open(encoding="utf-8") as f:
            for line in f:
                t = json.loads(line)
                if t.get("exit_reason") in ("SESSION_END", "NOT_EVALUABLE_SESSION_END_EXIT_PRICE"):
                    se.append([t.get("pair_id"), t.get("episode_id"), t.get("symbol"),
                               t.get("selected_exit_bid"), t.get("session_boundary_time"),
                               t.get("seconds_from_selected_event_to_boundary"),
                               t.get("integrity_status")])
                if t.get("exit_reason") in ("MAX_HOLD", "MAX_HOLD_GAP_EXIT") or float(t.get("hold_sec") or 0) > 350:
                    mh.append([t.get("pair_id"), t.get("episode_id"), t.get("hold_sec"),
                               t.get("max_hold_gap_sec"), t.get("exit_reason"),
                               t.get("integrity_status")])

    sheet("ExitDefinitions", ["note"], [["Unchanged EXIT_THRESHOLDS from locked P2"]])
    sheet("ExitTransitions", ["note"], [["see integrity ledgers"]])
    sheet("JointPairs",
          ["pair_id", "n_pass", "n_fail", "pnl", "pf", "gates_pass", "failed", "sha"],
          [[pid, p["n_pass"], p["n_fail"], p["pnl"], p["pf"], p["gates"]["all_pass"],
            json.dumps(p["gates"]["failed"]), p["ledger_sha256"]]
           for pid, p in (report.get("pairs") or {}).items()])

    jt = wb.create_sheet("JointTrades")
    jt.append(["pair_id", "day", "symbol", "entry_t", "exit_t", "reason", "net", "integrity"])
    n = 0
    for _, setup, _, xc in PAIRS:
        fp = store / f"ledger_A_{setup}_{xc}.jsonl"
        if not fp.exists():
            continue
        with fp.open(encoding="utf-8") as f:
            for line in f:
                if n >= 5000:
                    break
                t = json.loads(line)
                if t.get("integrity_status") != "PASS":
                    continue
                jt.append([t["pair_id"], t["day"], t["symbol"], t["entry_time"], t["exit_time"],
                           t["exit_reason"], t["net_pnl_yen"], t["integrity_status"]])
                n += 1

    daily = []
    for pid, p in (report.get("pairs") or {}).items():
        for d, pnl in (p.get("day_pnl") or {}).items():
            daily.append([pid, d, pnl])
    sheet("Daily", ["pair_id", "day", "pnl"], daily)
    cds = []
    for pid, p in (report.get("pairs") or {}).items():
        for s in p.get("retrospective_confirm_day_slice") or []:
            cds.append([pid, s["fold"], s["confirm"], s["confirm_pnl"]])
    sheet("ConfirmDaySlices", ["pair_id", "fold", "confirm", "pnl"], cds)
    dd = []
    for pid, p in (report.get("pairs") or {}).items():
        for r in p.get("day_deletion") or []:
            dd.append([pid, r["held_out_day"], r["remaining_pnl"]])
    sheet("DayDeletion", ["pair_id", "held_out", "remaining"], dd)
    sheet("Concentration", ["pair_id", "top1_day", "top1_symbol", "top1_trade"],
          [[pid, json.dumps(p.get("top1_day")), json.dumps(p.get("top1_symbol")), p.get("top1_trade_pnl")]
           for pid, p in (report.get("pairs") or {}).items()])
    sheet("Tests", ["test", "outcome"],
          [[r["test"], r["outcome"]] for r in (report.get("tests") or {}).get("rows", [])])
    sheet("Determinism", ["pair_id", "match", "sha_a", "sha_b"],
          [[pid, d.get("match"), d.get("sha_a"), d.get("sha_b")]
           for pid, d in ((report.get("determinism") or {}).get("detail") or {}).items()])
    sheet("Safety", ["key", "value"],
          [["submit", 0], ["cancel", 0], ["live", 0], ["mainline_changed", False],
           ["SHADOW", False], ["FORWARD", False], ["PAPER", False], ["DISCORD", False],
           ["TAER_V2", False]])
    sheet("ChangeLog", ["item", "note"], [
        ["invalid_prior", INVALID_RUN],
        ["fix", "SESSION_END same-symbol/session bid; MAX_HOLD gap separation; MFE/MAE envelope"],
        ["rolling_origin", "renamed retrospective_confirm_day_slice"],
    ])
    wb.save(out / "audit.xlsx")

    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "published_shas.json").write_text(json.dumps(report["published"], indent=2), encoding="utf-8")


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    rep = run()
    print("INTEGRITY_RERUN_DONE", rep.get("run_id"), rep.get("verdict"), flush=True)
    print(json.dumps(rep.get("published"), indent=2), flush=True)
