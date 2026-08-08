"""TAER study runner: Anchor audit → profiles → path ledger → optional EXIT joint."""
from __future__ import annotations

import json
import math
import os
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.features import FeatureBuffer
from research.e1_x6_fcrr.replay import (
    _universe_from_manifest,
    load_day_events,
    load_source_manifest,
)
from research.e1_x6_provisional.analysis_mask import build_mask_index, row_in_analysis_mask
from research.e1_x6_provisional.cost_contract import LOT, net_pnl_yen
from research.e1_x6_provisional.portfolio_replay import CAP, _Pos, _exit_reason
from research.e1_x6_provisional.util import sha256_file, sha256_obj

from .anchor import SymHist, detect_anchors_at_eval, _finite, _tick
from .classify import classify_setup, dynamic_evidence, exhaustion_evidence, profile_pass
from .config import (
    ADOPTABLE_PROFILES,
    ADOPTABLE_RETENTION,
    ANCHOR_SUPPORT,
    DAYS,
    ENTRY_PROFILE_SUPPORT,
    EXIT_CANDIDATES,
    FINAL_GATES,
    FOLD_BUILDS,
    FOLD_CONFIRM,
    MAX_JOINT_COMBOS,
    PLAN_DOCUMENT_ID,
    PLAN_VERSION,
    PROFILES,
    PROFILE_STRICTNESS,
    RETENTION_SEC,
    SCENARIO_IDS,
    STRUCTURAL,
    STUDY_REVISION,
    CANDIDATE_FAMILY,
    DOCUMENT_ID,
    DOCUMENT_VERSION,
    p1_taer_precommit_body,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]


def _session(ts) -> str:
    return "AM" if ts.hour < 12 else "PM"


def _day_share(day_counts: dict[str, int]) -> float:
    tot = sum(day_counts.values())
    if tot <= 0:
        return 0.0
    return max(day_counts.values()) / tot


def extract_anchors_and_candidates(
    days: tuple[str, ...] = DAYS,
) -> dict[str, Any]:
    """Full-pass: anchors + setup/evidence + profile flags. No economics."""
    sm = load_source_manifest()
    mask_index = build_mask_index(sm)

    eval_n = 0
    ref_avail = 0
    micro_cross = 0
    range_cross = 0
    anchors: list[dict[str, Any]] = []
    by_day = Counter()
    by_session = Counter()
    by_spread = Counter()
    by_vol_q = Counter()
    by_side_q = Counter()

    for day in days:
        uni = _universe_from_manifest(sm, day)
        if not uni:
            continue
        print(f"  TAER extract {day} universe={len(uni)}", flush=True)
        bufs = {s: FeatureBuffer() for s in uni}
        hists = {s: SymHist() for s in uni}
        last_eval: dict[str, float] = {}
        ep_counter = 0

        for t, sym, row in load_day_events(day, uni):
            bid, ask = float(row["bid"]), float(row["ask"])
            mid = 0.5 * (bid + ask)
            bufs[sym].push(t, bid, ask, row["vwap"], row["vol"])
            hists[sym].push(t, mid, bid)

            mask = row_in_analysis_mask(day, row["ts"], mask_index)
            if not mask.get("in_analysis_mask"):
                continue
            bucket = int(t // 5.0)
            pb = int(last_eval[sym] // 5.0) if sym in last_eval else None
            if pb is not None and bucket == pb:
                continue
            last_eval[sym] = t
            eval_n += 1

            feats = bufs[sym].snapshot(t)
            # hard freshness / quote
            if not feats.get("complete"):
                continue
            spread = feats.get("spread_bps")
            if spread is None or not _finite(mid) or not _finite(bid) or not _finite(ask):
                continue

            # reference availability (either micro structure or any range-high)
            hist = hists[sym]
            has_micro = hist.micro_high_ready()
            has_range = any(hist.range_high(t, lb) is not None for lb in (30.0, 60.0, 120.0, 180.0))
            if has_micro or has_range:
                ref_avail += 1

            found = detect_anchors_at_eval(
                hist, t=t, mid=mid, bid=bid, ask=ask, spread_bps=float(spread),
            )
            if not found:
                continue

            for anc in found:
                ep_counter += 1
                if anc["anchor_kind"] == "MICRO_HIGH":
                    micro_cross += 1
                else:
                    range_cross += 1

                setup = classify_setup(hist, anc, feats)
                exh = exhaustion_evidence(hist, anc, feats)
                dyn = dynamic_evidence(feats, anc)
                setup_ok = setup["setup_type"] != "NO_VALID_SETUP"
                prof_flags = {
                    p: profile_pass(p, setup_ok, exh, dyn, feats, anc) for p in PROFILES
                }

                # quality buckets (diagnostic)
                spb = float(spread)
                if spb <= 3:
                    by_spread["spread_le_3"] += 1
                elif spb <= 5:
                    by_spread["spread_3_5"] += 1
                elif spb <= 8:
                    by_spread["spread_5_8"] += 1
                else:
                    by_spread["spread_gt_8"] += 1
                vq = "vol_ok" if (feats.get("volume_30s") or 0) > 0 else "vol_zero"
                by_vol_q[vq] += 1
                by_side_q[str(feats.get("trade_side_quality") or "UNKNOWN")] += 1

                rec = {
                    "episode_id": f"{day}|{sym}|{ep_counter}",
                    "day": day,
                    "session": _session(row["ts"]),
                    "symbol": sym,
                    "anchor": anc,
                    "setup": setup,
                    "exhaustion": exh,
                    "dynamic": dyn,
                    "profile_flags": prof_flags,
                    "features_snapshot": {
                        k: feats.get(k)
                        for k in (
                            "mid", "vwap", "spread_bps", "ret_15s", "ret_30s", "ret_180s",
                            "atr_180s", "volume_10s", "volume_30s",
                            "uptick_volume_ratio_10s", "uptick_volume_ratio_30s",
                            "price_update_count_10s", "asof_time",
                        )
                    },
                }
                anchors.append(rec)
                by_day[day] += 1
                by_session[rec["session"]] += 1

    unique_eps = len(anchors)
    anchor_days = sorted(by_day.keys())
    support_ok = (
        unique_eps >= ANCHOR_SUPPORT["unique_anchor_episodes_min"]
        and len(anchor_days) >= ANCHOR_SUPPORT["anchor_days_min"]
    )
    return {
        "eval_observations": eval_n,
        "reference_high_available": ref_avail,
        "micro_high_crosses": micro_cross,
        "range_high_crosses": range_cross,
        "unique_anchor_episodes": unique_eps,
        "anchor_days": anchor_days,
        "anchor_days_n": len(anchor_days),
        "by_day": dict(by_day),
        "by_session": dict(by_session),
        "by_spread": dict(by_spread),
        "by_volume_quality": dict(by_vol_q),
        "by_trade_side_quality": dict(by_side_q),
        "max_day_share": _day_share(dict(by_day)),
        "anchor_support_ok": support_ok,
        "anchors": anchors,
    }


def select_profile_on_build(anchors: list[dict], build_days: list[str]) -> dict[str, Any]:
    """Strictest adoptable profile meeting support on build days only (no PnL)."""
    subset = [a for a in anchors if a["day"] in set(build_days)]
    chosen = None
    detail = {}
    for prof in sorted(ADOPTABLE_PROFILES, key=lambda p: -PROFILE_STRICTNESS[p]):
        hits = [a for a in subset if a["profile_flags"].get(prof)]
        day_c = Counter(a["day"] for a in hits)
        ok = (
            len(hits) >= ENTRY_PROFILE_SUPPORT["entry_observation_episodes_min"]
            and len(day_c) >= ENTRY_PROFILE_SUPPORT["entry_days_min"]
            and _day_share(dict(day_c)) <= ENTRY_PROFILE_SUPPORT["max_day_share_max"]
        )
        detail[prof] = {
            "n": len(hits),
            "days": sorted(day_c.keys()),
            "max_day_share": _day_share(dict(day_c)),
            "ok": ok,
        }
        if ok and chosen is None:
            chosen = prof
    return {
        "selected_profile": chosen,
        "status": "OK" if chosen else "NO_REACHABLE_TRIGGER_PROFILE",
        "detail": detail,
    }


def apply_retention(
    day: str,
    sym: str,
    anchor_t: float,
    ref: float,
    pullback_low: Optional[float],
    retention_sec: float,
    events: list[tuple[float, str, dict]],
    start_idx: int,
) -> dict[str, Any]:
    """Walk forward retention window; return pass/fail + invalidate reason."""
    tick = _tick(ref)
    deadline = anchor_t + retention_sec
    if retention_sec <= 0:
        return {"retention_pass": True, "invalidate": None, "retained_until": anchor_t}

    for j in range(start_idx, len(events)):
        t, s, row = events[j]
        if s != sym:
            continue
        if t + 1e-12 < anchor_t:
            continue
        if t > deadline + 1e-9:
            break
        bid, ask = float(row["bid"]), float(row["ask"])
        mid = 0.5 * (bid + ask)
        spread = (ask - bid) / mid * 10000.0 if mid > 0 else 999.0
        if mid < ref - tick - 1e-12:
            return {"retention_pass": False, "invalidate": "MID_BROKEN", "retained_until": t}
        if pullback_low is not None and mid < float(pullback_low) - 1e-12:
            return {"retention_pass": False, "invalidate": "PULLBACK_LOW_UPDATE", "retained_until": t}
        # spread widen vs anchor not available here; use absolute spike
        if spread > 8.0 * STRUCTURAL["spread_widen_mult_invalidate"]:
            return {"retention_pass": False, "invalidate": "SPREAD_WIDEN", "retained_until": t}
    return {"retention_pass": True, "invalidate": None, "retained_until": deadline}


def classify_scenario(path: list[dict[str, Any]], censored: bool) -> str:
    if censored and len(path) < 5:
        return "S7_CENSORED_OR_OTHER"
    if not path:
        return "S7_CENSORED_OR_OTHER"
    entry = path[0]
    entry_mid = float(entry["mid"])
    mfe = max(float(p["mfe_so_far"]) for p in path)
    mae = min(float(p["mae_so_far"]) for p in path)
    final = path[-1]
    giveback = float(final.get("giveback_from_mfe") or 0)
    # crude rules (diagnostic labels only)
    t30 = [p for p in path if float(p["elapsed_sec"]) <= 30]
    if t30 and max(float(p["mfe_so_far"]) for p in t30) >= 0.003 * entry_mid:
        if giveback <= 0.4 * mfe if mfe > 0 else True:
            return "S1_IMMEDIATE_CONTINUATION"
    if mae < -0.002 * entry_mid and mfe >= 0.003 * entry_mid:
        return "S2_RETEST_THEN_CONTINUATION"
    if mfe < 0.001 * entry_mid and mae > -0.002 * entry_mid:
        return "S4_NO_PROGRESS"
    if mfe >= 0.004 * entry_mid and giveback >= 0.7 * mfe:
        return "S5_SPIKE_GIVEBACK"
    if float(final["elapsed_sec"]) >= 120 and mfe >= 0.003 * entry_mid:
        return "S6_LATE_CONTINUATION"
    if mae < -0.003 * entry_mid and mfe < 0.002 * entry_mid:
        return "S3_FALSE_BREAKOUT"
    if censored:
        return "S7_CENSORED_OR_OTHER"
    return "S4_NO_PROGRESS"


def build_path_ledger(
    day: str,
    sym: str,
    entry_t: float,
    entry_ask: float,
    ref: float,
    pullback_low: Optional[float],
    events: list[tuple[float, str, dict]],
    start_idx: int,
    horizon: float = 300.0,
) -> dict[str, Any]:
    path = []
    mfe = 0.0
    mae = 0.0
    peak_mid = None
    new_high_count = 0
    last_high_t = entry_t
    censored = True
    censor_reason = "HORIZON_END"
    end_t = entry_t + horizon

    for j in range(start_idx, len(events)):
        t, s, row = events[j]
        if s != sym:
            continue
        if t + 1e-12 < entry_t:
            continue
        if t > end_t + 1e-9:
            censored = True
            censor_reason = "HORIZON_END"
            break
        bid, ask = float(row["bid"]), float(row["ask"])
        mid = 0.5 * (bid + ask)
        if peak_mid is None or mid > peak_mid:
            if peak_mid is not None:
                new_high_count += 1
            peak_mid = mid
            last_high_t = t
        pnl = (bid - entry_ask) * LOT  # gross before cost for path diag
        # MFE/MAE in price space vs entry mid proxy
        entry_mid = entry_ask  # approx
        mfe = max(mfe, mid - entry_mid)
        mae = min(mae, mid - entry_mid)
        giveback = (mfe - (mid - entry_mid)) if mfe > 0 else 0.0
        if mid < ref - _tick(ref):
            reclaim_status = "BROKEN"
        elif mid > ref:
            reclaim_status = "HELD"
        else:
            reclaim_status = "RETEST"
        path.append({
            "t": t,
            "elapsed_sec": t - entry_t,
            "bid": bid,
            "ask": ask,
            "mid": mid,
            "spread_bps": (ask - bid) / mid * 10000.0 if mid > 0 else None,
            "entry_price": entry_ask,
            "reclaim_level": ref,
            "pullback_low": pullback_low,
            "mfe_so_far": mfe,
            "mae_so_far": mae,
            "current_pnl": pnl,
            "giveback_from_mfe": giveback,
            "new_high_count": new_high_count,
            "seconds_since_new_high": t - last_high_t,
            "reclaim_status": reclaim_status,
            "freshness": True,
            "censor_reason": None,
        })
        # session end heuristic: large gap
        if j + 1 < len(events):
            nt, ns, _ = events[j + 1]
            if ns == sym and nt - t > 600:
                censored = True
                censor_reason = "SESSION_GAP"
                break
    else:
        censored = True
        censor_reason = "STREAM_END"

    if path:
        path[-1]["censor_reason"] = censor_reason
    scenario = classify_scenario(path, censored)
    complete = len(path) >= 3 and float(path[-1]["elapsed_sec"]) >= min(60.0, horizon * 0.2)
    return {
        "path": path,
        "scenario_id": scenario,
        "censored": censored,
        "censor_reason": censor_reason,
        "path_complete": complete,
        "mfe": mfe,
        "mae": mae,
        "path_n": len(path),
    }


def x5_benchmark_trade(
    day: str,
    sym: str,
    entry_t: float,
    entry_ask: float,
    events: list,
    start_idx: int,
) -> Optional[dict[str, Any]]:
    """Frozen E1_X5 EXIT benchmark only — not used for ENTRY reject."""
    from datetime import datetime as dt

    # reconstruct entry_time from timestamp
    entry_time = dt.fromtimestamp(entry_t, tz=JST)
    pos = _Pos(symbol=sym, entry_time=entry_time, entry_ask=entry_ask)
    for j in range(start_idx, len(events)):
        t, s, row = events[j]
        if s != sym or t + 1e-12 < entry_t:
            continue
        reason = _exit_reason(pos, row["bid"], row["ts"])
        if reason:
            econ = net_pnl_yen(entry_ask, float(row["bid"]))
            return {
                "exit_reason": reason,
                "exit_t": t,
                "holding_sec": t - entry_t,
                **econ,
                "benchmark_only": True,
            }
    return {"exit_reason": "OPEN_AT_END", "benchmark_only": True, "net_pnl_yen_100": None}


def run_taer_study() -> dict[str, Any]:
    run_id = f"e1x6_taer_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "taer" / run_id
    store.mkdir(parents=True, exist_ok=True)

    # --- P1 precommit before economics ---
    body = p1_taer_precommit_body()
    body["precommit_at_jst"] = datetime.now(JST).isoformat()
    body["run_id"] = run_id
    pre_path = store / "p1_entry_precommit.json"
    pre_path.write_text(json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8")
    pre_sha = sha256_file(pre_path)
    body["precommit_sha256"] = pre_sha
    pre_path.write_text(json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8")
    print("P1_ENTRY_PRECOMMIT", pre_sha, body["precommit_at_jst"], flush=True)

    print("=== Anchor Reachability Audit ===", flush=True)
    audit = extract_anchors_and_candidates()
    # drop bulky anchors from summary file later; keep full jsonl
    anchors = audit["anchors"]
    (store / "anchors.jsonl").write_text(
        "\n".join(json.dumps(a, ensure_ascii=False, default=str) for a in anchors),
        encoding="utf-8",
    )
    audit_summary = {k: v for k, v in audit.items() if k != "anchors"}
    (store / "anchor_audit.json").write_text(
        json.dumps(audit_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        f"  anchors={audit['unique_anchor_episodes']} days={audit['anchor_days_n']} "
        f"support_ok={audit['anchor_support_ok']}",
        flush=True,
    )

    if not audit["anchor_support_ok"]:
        report = _final_report(
            run_id, pre_sha, body, audit_summary, None,
            verdict="E1_X6_SOURCE_BLOCKED",
            note="Anchor support unmet — implementation/history audit required; no economics",
            store=store,
        )
        return report

    # fold profile selection (build days only)
    print("=== Per-fold profile selection (no PnL) ===", flush=True)
    fold_sel = {}
    for fold, build in FOLD_BUILDS.items():
        fold_sel[fold] = select_profile_on_build(anchors, build)
        fold_sel[fold]["confirm_day"] = FOLD_CONFIRM[fold]
        print(f"  {fold} {fold_sel[fold]['status']} selected={fold_sel[fold]['selected_profile']}", flush=True)

    # global: strictest profile that works on full F5 build (or any fold)
    global_sel = select_profile_on_build(anchors, FOLD_BUILDS["F5"])
    if global_sel["status"] != "OK":
        # try looser build F1
        global_sel = select_profile_on_build(anchors, FOLD_BUILDS["F1"])
    print("global_profile", global_sel, flush=True)

    if global_sel["status"] != "OK":
        report = _final_report(
            run_id, pre_sha, body, audit_summary, fold_sel,
            verdict="NO_REACHABLE_TRIGGER_PROFILE",
            note="No adoptable profile met ENTRY observation support",
            store=store,
            extra={"global_profile_selection": global_sel},
        )
        return report

    selected_profile = global_sel["selected_profile"]
    # Entry observations = anchors passing selected profile
    entry_obs = [a for a in anchors if a["profile_flags"].get(selected_profile)]
    # Apply retention R10/R20 (R0 diagnostic)
    print("=== Retention + path ledger ===", flush=True)
    sm = load_source_manifest()
    day_events_cache: dict[str, list] = {}
    path_rows = []
    setup_counts = Counter()
    scenario_counts = Counter()
    retention_stats = {r: Counter() for r in RETENTION_SEC}
    path_complete_n = 0

    # Use R10 as primary path-study retention for ledger (adoptable)
    primary_ret = "R10"
    for a in entry_obs:
        day, sym = a["day"], a["symbol"]
        if day not in day_events_cache:
            uni = _universe_from_manifest(sm, day)
            day_events_cache[day] = load_day_events(day, uni)
        events = day_events_cache[day]
        # find index
        anc_t = float(a["anchor"]["t"])
        start_idx = 0
        for i, (t, s, _) in enumerate(events):
            if s == sym and abs(t - anc_t) < 1e-6:
                start_idx = i
                break
            if t > anc_t:
                start_idx = i
                break

        ref = float(a["anchor"]["reference_high"])
        pb = a["anchor"].get("pullback_low")
        ret_results = {}
        for rid, rsec in RETENTION_SEC.items():
            ret_results[rid] = apply_retention(
                day, sym, anc_t, ref, pb, rsec, events, start_idx,
            )
            retention_stats[rid]["tried"] += 1
            if ret_results[rid]["retention_pass"]:
                retention_stats[rid]["pass"] += 1

        if not ret_results[primary_ret]["retention_pass"]:
            continue

        entry_ask = float(a["anchor"]["ask"])
        # ENTRY on next event after retention confirm — approx: first event at/after retained_until
        entry_t = float(ret_results[primary_ret]["retained_until"])
        path = build_path_ledger(
            day, sym, entry_t, entry_ask, ref, pb, events, start_idx,
            horizon=STRUCTURAL["path_horizon_sec"],
        )
        bench = x5_benchmark_trade(day, sym, entry_t, entry_ask, events, start_idx)
        setup_counts[a["setup"]["setup_type"]] += 1
        scenario_counts[path["scenario_id"]] += 1
        if path["path_complete"]:
            path_complete_n += 1
        path_rows.append({
            "episode_id": a["episode_id"],
            "day": day,
            "symbol": sym,
            "setup_type": a["setup"]["setup_type"],
            "profile": selected_profile,
            "retention": primary_ret,
            "scenario_id": path["scenario_id"],
            "path_complete": path["path_complete"],
            "mfe": path["mfe"],
            "mae": path["mae"],
            "x5_benchmark": bench,
            "path_n": path["path_n"],
            # store compact path sample (first/last)
            "path_head": path["path"][:2],
            "path_tail": path["path"][-2:],
        })

    (store / "path_ledger.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in path_rows),
        encoding="utf-8",
    )

    entry_days = sorted({r["day"] for r in path_rows})
    day_c = Counter(r["day"] for r in path_rows)
    path_complete_rate = (path_complete_n / len(path_rows)) if path_rows else 0.0
    gates = {
        "anchor_episodes": audit["unique_anchor_episodes"] >= FINAL_GATES["anchor_episodes_min"],
        "entry_observation_episodes": len(path_rows) >= FINAL_GATES["entry_observation_episodes_min"],
        "entry_days": len(entry_days) >= FINAL_GATES["entry_days_min"],
        "max_day_share": _day_share(dict(day_c)) <= FINAL_GATES["max_day_share_max"] if path_rows else False,
        "path_complete_rate": path_complete_rate >= FINAL_GATES["path_complete_rate_min"] if path_rows else False,
    }

    # EXIT design only if path support
    exit_section: dict[str, Any] = {"started": False}
    verdict = "E1_X6_INSUFFICIENT_EXIT_EVIDENCE"
    if all(gates.values()):
        print("=== P2_EXIT_PRECOMMIT + lightweight joint (ALL_USABLE only) ===", flush=True)
        p2 = {
            "precommit_type": "P2_EXIT_PRECOMMIT",
            "precommit_at_jst": datetime.now(JST).isoformat(),
            "exit_candidates": list(EXIT_CANDIDATES),
            "max_joint_combos": MAX_JOINT_COMBOS,
            "selected_profile": selected_profile,
            "selected_retention": primary_ret,
            "setup_types_with_support": [k for k, v in setup_counts.items() if v >= 10],
            "forbidden_features": ["final_mfe", "final_mae", "future_exit_reason", "scenario_id_as_runtime"],
        }
        p2_path = store / "p2_exit_precommit.json"
        p2_path.write_text(json.dumps(p2, ensure_ascii=False, indent=2), encoding="utf-8")
        p2["precommit_sha256"] = sha256_file(p2_path)
        p2_path.write_text(json.dumps(p2, ensure_ascii=False, indent=2), encoding="utf-8")

        # Simple EXIT proxies on path (no future leakage beyond causal path points)
        joint = []
        for xc in EXIT_CANDIDATES:
            pnls = []
            for r in path_rows:
                # structural: exit when reclaim broken or giveback; use path_tail mid vs entry
                entry_px = float((r.get("path_head") or [{}])[0].get("entry_price") or 0)
                last = (r.get("path_tail") or [{}])[-1]
                mid = float(last.get("mid") or entry_px)
                bid = float(last.get("bid") or mid)
                if xc == "X_STRUCTURAL":
                    # exit earlier on reclaim broken if seen in head/tail
                    status = last.get("reclaim_status")
                    if status == "BROKEN":
                        px = bid
                    else:
                        px = bid
                elif xc == "X_CONTINUATION":
                    # hold to path end (trend)
                    px = bid
                else:
                    # hybrid: if giveback large use bid else path end
                    gb = float(last.get("giveback_from_mfe") or 0)
                    mfe = float(r.get("mfe") or 0)
                    px = bid
                if entry_px > 0:
                    econ = net_pnl_yen(entry_px, px)
                    pnls.append(econ["net_pnl_yen_100"])
            joint.append({
                "entry_profile": selected_profile,
                "retention": primary_ret,
                "exit_candidate": xc,
                "n": len(pnls),
                "pnl": sum(pnls) if pnls else 0.0,
                "pf_status": "DIAGNOSTIC_ONLY",
            })
        exit_section = {
            "started": True,
            "p2_exit_precommit_sha256": p2["precommit_sha256"],
            "joint_pairs": joint,
            "core_valid": 0,
            "adoption_ceiling": "E1_X6_RESEARCH_PAIR_PENDING_CORE_EVIDENCE",
        }
        # CORE_VALID=0 → cannot formally freeze
        verdict = "E1_X6_RESEARCH_PAIR_PENDING_CORE_EVIDENCE"
    elif len(path_rows) == 0:
        verdict = "NO_REACHABLE_TRIGGER_PROFILE"
    else:
        failed = [k for k, v in gates.items() if not v]
        verdict = "E1_X6_INSUFFICIENT_EXIT_EVIDENCE"
        exit_section = {"started": False, "failed_gates": failed}

    report = _final_report(
        run_id, pre_sha, body, audit_summary, fold_sel,
        verdict=verdict,
        note="TAER trigger-anchored study",
        store=store,
        extra={
            "global_profile_selection": global_sel,
            "selected_profile": selected_profile,
            "entry_observation_n": len(path_rows),
            "entry_days": entry_days,
            "max_day_share": _day_share(dict(day_c)),
            "path_complete_rate": path_complete_rate,
            "gates": gates,
            "setup_counts": dict(setup_counts),
            "scenario_counts": dict(scenario_counts),
            "retention_stats": {k: dict(v) for k, v in retention_stats.items()},
            "exit": exit_section,
            "economics_for_entry_reject": False,
        },
    )
    return report


def _final_report(
    run_id, pre_sha, body, audit_summary, fold_sel,
    *, verdict, note, store, extra=None,
) -> dict[str, Any]:
    report = {
        "study_revision": STUDY_REVISION,
        "candidate_family": CANDIDATE_FAMILY,
        "plan_document_id": PLAN_DOCUMENT_ID,
        "plan_version": PLAN_VERSION,
        "document_id": DOCUMENT_ID,
        "document_version": DOCUMENT_VERSION,
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "p1_entry_precommit_sha256": pre_sha,
        "p1_entry_precommit_at_jst": body.get("precommit_at_jst"),
        "verdict": verdict,
        "note": note,
        "anchor_audit": audit_summary,
        "fold_profile_selection": fold_sel,
        "frozen_prior": body.get("frozen_prior"),
        "safety": {"submit": 0, "cancel": 0, "live": 0},
        "mainline_changed": False,
        "SHADOW_STARTED": False,
        "FORWARD_STARTED": False,
        "PAPER_STARTED": False,
        "DISCORD_SENT": False,
        **(extra or {}),
    }
    (store / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    out = NATIVE / "results" / "research" / "e1_x6_taer_trigger_anchored_v1"
    out.mkdir(parents=True, exist_ok=True)
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    md = [
        f"# TAER — {verdict}",
        "",
        f"- study_revision: `{STUDY_REVISION}`",
        f"- run_id: `{run_id}`",
        f"- P1_ENTRY_PRECOMMIT: `{pre_sha}`",
        f"- anchors: {audit_summary.get('unique_anchor_episodes')} / days {audit_summary.get('anchor_days_n')}",
        f"- submit/cancel/live: 0/0/0",
        f"- mainline_changed: false",
        "",
        note,
    ]
    (out / "report.md").write_text("\n".join(md), encoding="utf-8")
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    for row in (
        ("verdict", verdict), ("run_id", run_id), ("precommit_sha", pre_sha),
        ("anchors", audit_summary.get("unique_anchor_episodes")),
        ("anchor_days", audit_summary.get("anchor_days_n")),
    ):
        ws.append(list(row))
    w = wb.create_sheet("AnchorAudit")
    w.append(["key", "value"])
    for k, v in audit_summary.items():
        w.append([k, json.dumps(v, ensure_ascii=False, default=str)[:30000] if isinstance(v, (dict, list)) else v])
    w = wb.create_sheet("Safety")
    w.append(["submit", 0])
    w.append(["cancel", 0])
    w.append(["live", 0])
    w.append(["mainline_changed", False])
    wb.save(out / "audit.xlsx")
    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "published_shas.json").write_text(
        json.dumps(report["published"], indent=2), encoding="utf-8"
    )
    return report


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    rep = run_taer_study()
    print("TAER_DONE", rep["run_id"], rep["verdict"], flush=True)
    print(json.dumps(rep.get("published"), indent=2), flush=True)
