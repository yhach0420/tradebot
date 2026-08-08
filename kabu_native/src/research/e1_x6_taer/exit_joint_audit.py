"""TAER EXIT Joint Completion Audit — distinct EXIT ledgers + S7/setup analysis.

Does not overwrite e1x6_taer_20260803_232514 (frozen TAER_ENTRY_PATH_READY).
"""
from __future__ import annotations

import json
import math
import os
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.features import FeatureBuffer
from research.e1_x6_fcrr.replay import (
    _universe_from_manifest,
    load_day_events,
    load_source_manifest,
)
from research.e1_x6_provisional.cost_contract import LOT, net_pnl_yen
from research.e1_x6_provisional.util import sha256_file, sha256_obj

from .config import DAYS, EXIT_CANDIDATES, STUDY_REVISION
from .exit_sm import EXIT_THRESHOLDS, ExitPos, step_exit
from .run_study import classify_scenario  # noqa: F401 — reserved for future path reclass

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
PRIOR_RUN = "e1x6_taer_20260803_232514"
PRIOR_STORE = Path.home() / "e1x6_research_store" / "taer" / PRIOR_RUN

PAIRS = [
    ("TAER_P3", "PULLBACK_RECLAIM", "R10", "X_STRUCTURAL"),
    ("TAER_P3", "PULLBACK_RECLAIM", "R10", "X_CONTINUATION"),
    ("TAER_P3", "PULLBACK_RECLAIM", "R10", "X_HYBRID"),
    ("TAER_P3", "RANGE_BREAKOUT", "R10", "X_STRUCTURAL"),
    ("TAER_P3", "RANGE_BREAKOUT", "R10", "X_CONTINUATION"),
    ("TAER_P3", "RANGE_BREAKOUT", "R10", "X_HYBRID"),
]


def _pf(pnls: list[float]) -> tuple[Optional[float], str]:
    gains = sum(x for x in pnls if x > 0)
    losses = sum(-x for x in pnls if x < 0)
    if losses <= 1e-12 and gains > 0:
        return None, "NO_LOSS"
    if losses <= 1e-12 and gains <= 1e-12:
        return None, "EMPTY"
    return gains / losses, "OK"


def _max_dd(day_pnls: dict[str, float]) -> float:
    eq = 0.0
    peak = 0.0
    dd = 0.0
    for d in sorted(day_pnls):
        eq += day_pnls[d]
        peak = max(peak, eq)
        dd = min(dd, eq - peak)
    return dd


def load_entry_observations() -> tuple[list[dict], dict[str, Any]]:
    """Load prior path_ledger entries; flag the 3 missing entry_price rows."""
    rows = []
    excluded = []
    fp = PRIOR_STORE / "path_ledger.jsonl"
    with fp.open(encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            d = json.loads(line)
            head = (d.get("path_head") or [{}])[0]
            entry_px = head.get("entry_price")
            if entry_px is None or not (float(entry_px) > 0):
                excluded.append({
                    "episode_id": d.get("episode_id"),
                    "day": d.get("day"),
                    "symbol": d.get("symbol"),
                    "reason": "MISSING_ENTRY_PRICE",
                })
                continue
            rows.append({
                "episode_id": d["episode_id"],
                "day": d["day"],
                "symbol": d["symbol"],
                "setup_type": d["setup_type"],
                "profile": d.get("profile") or "TAER_P3",
                "retention": d.get("retention") or "R10",
                "scenario_id_prior": d.get("scenario_id"),
                "entry_t": float(head["t"]),
                "entry_ask": float(entry_px),
                "entry_mid": float(head.get("mid") or entry_px),
                "reclaim_level": float(head.get("reclaim_level") or entry_px),
                "pullback_low": head.get("pullback_low"),
                "path_n_prior": d.get("path_n"),
                "mfe_prior": d.get("mfe"),
                "mae_prior": d.get("mae"),
                "censor_reason_prior": (d.get("path_tail") or [{}])[-1].get("censor_reason")
                if d.get("path_tail") else None,
            })
    meta = {
        "path_ledger_n": len(rows) + len(excluded),
        "usable_n": len(rows),
        "excluded_n": len(excluded),
        "excluded": excluded,
        "q10_749_to_746": {
            "answer": "3 episodes lacked path_head.entry_price and were skipped in prior joint pnl loop",
            "excluded": excluded,
        },
    }
    return rows, meta


def decompose_s7(entries: list[dict]) -> dict[str, Any]:
    """Split prior S7 using path_ledger metadata (no economics)."""
    buckets: dict[str, list] = defaultdict(list)
    for e in entries:
        if e.get("scenario_id_prior") != "S7_CENSORED_OR_OTHER":
            continue
        cr = e.get("censor_reason_prior") or ""
        pn = int(e.get("path_n_prior") or 0)
        if cr in ("SESSION_GAP",) or "SESSION" in str(cr).upper():
            label = "S7_SESSION_CENSORED"
        elif pn < 5 or cr in ("STREAM_END", "HORIZON_END") and pn < 10:
            label = "S7_PATH_INCOMPLETE"
        elif pn >= 5 and (e.get("mfe_prior") or 0) == 0 and (e.get("mae_prior") or 0) == 0:
            label = "S7_NO_FIRST_TOUCH"
        elif cr in ("HORIZON_END", "STREAM_END"):
            # reached horizon but classifier fell through
            label = "S7_AMBIGUOUS"
        else:
            label = "S7_OTHER"
        buckets[label].append(e)

    out = {}
    for label, rows in buckets.items():
        by_day = Counter(r["day"] for r in rows)
        by_setup = Counter(r["setup_type"] for r in rows)
        mfes = [float(r.get("mfe_prior") or 0) for r in rows]
        maes = [float(r.get("mae_prior") or 0) for r in rows]
        # path seconds approx from path_n * ~0.5s unknown; use path_n as proxy count
        out[label] = {
            "n": len(rows),
            "by_day": dict(by_day),
            "by_setup": dict(by_setup),
            "mfe_mean": sum(mfes) / len(mfes) if mfes else None,
            "mae_mean": sum(maes) / len(maes) if maes else None,
            "path_n_mean": sum(int(r.get("path_n_prior") or 0) for r in rows) / len(rows),
            "exclude_from_exit_threshold_fit": True,
            "reason": label,
        }
    out["_total_s7"] = sum(v["n"] for k, v in out.items() if k.startswith("S7_"))
    return out


def setup_path_summary(entries: list[dict]) -> dict[str, Any]:
    out = {}
    for setup in ("PULLBACK_RECLAIM", "RANGE_BREAKOUT"):
        rows = [e for e in entries if e["setup_type"] == setup]
        sc = Counter(e.get("scenario_id_prior") for e in rows)
        mfes = [float(e.get("mfe_prior") or 0) for e in rows]
        maes = [float(e.get("mae_prior") or 0) for e in rows]
        out[setup] = {
            "n": len(rows),
            "scenario_counts": dict(sc),
            "mfe_mean": sum(mfes) / len(mfes) if mfes else None,
            "mae_mean": sum(maes) / len(maes) if maes else None,
            "mfe_median": sorted(mfes)[len(mfes) // 2] if mfes else None,
            "mae_median": sorted(maes)[len(maes) // 2] if maes else None,
            "note": "EXIT thresholds must not be auto-shared across setups",
        }
    return out


def replay_pair(
    entries: list[dict],
    *,
    setup_type: str,
    exit_candidate: str,
    day_events: dict[str, list],
) -> dict[str, Any]:
    subset = [e for e in entries if e["setup_type"] == setup_type]
    trades = []
    reason_c = Counter()
    transitions_sample = []
    s6_early_exit = 0
    s6_total = 0
    cap_blocked = 0
    follow_on = 0

    for day in sorted({e["day"] for e in subset}):
        events = day_events[day]
        day_entries = sorted(
            [e for e in subset if e["day"] == day],
            key=lambda x: x["entry_t"],
        )
        positions: dict[str, tuple[ExitPos, dict]] = {}
        entered_eps: set[str] = set()
        bufs: dict[str, FeatureBuffer] = {}
        last_eval_feat: dict[str, int] = {}

        queue: dict[str, list] = defaultdict(list)
        for e in day_entries:
            queue[e["symbol"]].append(e)
        for s in queue:
            queue[s].sort(key=lambda x: x["entry_t"])

        for t, sym, row in events:
            bid, ask = float(row["bid"]), float(row["ask"])
            mid = 0.5 * (bid + ask)
            vwap = float(row["vwap"]) if row.get("vwap") is not None else None
            spread = (ask - bid) / mid * 10000.0 if mid > 0 else None

            # update open position
            if sym in positions:
                pos, meta = positions[sym]
                if sym not in bufs:
                    bufs[sym] = FeatureBuffer()
                bufs[sym].push(t, bid, ask, row["vwap"], row["vol"])
                # evaluate features sparsely every ~5s
                feats = {}
                bucket = int(t // 5.0)
                prev_b = last_eval_feat.get(sym)
                if prev_b != bucket:
                    last_eval_feat[sym] = bucket
                    snap = bufs[sym].snapshot(t)
                    feats = snap if snap.get("complete") else {}
                hit = step_exit(
                    pos,
                    t=t, bid=bid, ask=ask, mid=mid, vwap=vwap, spread_bps=spread,
                    volume_30s=feats.get("volume_30s"),
                    price_update_count_10s=feats.get("price_update_count_10s"),
                )
                if hit:
                    econ = net_pnl_yen(pos.entry_ask, hit["exit_bid"])
                    gross = (hit["exit_bid"] - pos.entry_ask) * LOT
                    tr = {
                        "pair_id": f"TAER_P3|{setup_type}|R10|{exit_candidate}",
                        "episode_id": meta["episode_id"],
                        "day": day,
                        "am_pm": "AM" if row["ts"].hour < 12 else "PM",
                        "symbol": sym,
                        "setup_type": setup_type,
                        "exit_candidate": exit_candidate,
                        "entry_t": pos.entry_t,
                        "exit_t": hit["exit_t"],
                        "entry_ask": pos.entry_ask,
                        "exit_bid": hit["exit_bid"],
                        "exit_state": hit["exit_state"],
                        "exit_reason": hit["exit_reason"],
                        "hold_sec": hit["hold_sec"],
                        "mfe_at_exit": hit["mfe_at_exit"],
                        "mae_at_exit": hit["mae_at_exit"],
                        "giveback_at_exit": hit["giveback_at_exit"],
                        "pnl_before_cost": gross,
                        "pnl_after_5bps": econ["net_pnl_yen_100"],
                        "cost_yen_100": econ["cost_yen_100"],
                        "scenario_id_prior": meta.get("scenario_id_prior"),
                    }
                    trades.append(tr)
                    reason_c[hit["exit_reason"]] += 1
                    if len(transitions_sample) < 30:
                        transitions_sample.extend(hit.get("transitions") or [])
                    if meta.get("scenario_id_prior") == "S6_LATE_CONTINUATION":
                        s6_total += 1
                        if hit["hold_sec"] < 90.0 and hit["exit_reason"] not in (
                            "HARD_STOP", "MAX_HOLD",
                        ):
                            s6_early_exit += 1
                    del positions[sym]

            # try entries at this timestamp
            q = queue.get(sym) or []
            while q and q[0]["entry_t"] <= t + 1e-9:
                e = q.pop(0)
                if e["episode_id"] in entered_eps:
                    continue
                if sym in positions:
                    follow_on += 1
                    continue
                if len(positions) >= 5:
                    cap_blocked += 1
                    entered_eps.add(e["episode_id"])  # consume episode
                    continue
                # open
                pb = e.get("pullback_low")
                pb_f = float(pb) if pb is not None else None
                # range low approx: pullback_low or reclaim - 1 ATR proxy using mae
                range_low = pb_f
                range_high = float(e["reclaim_level"])
                pos = ExitPos(
                    symbol=sym,
                    setup_type=setup_type,
                    exit_candidate=exit_candidate,
                    entry_t=float(e["entry_t"]),
                    entry_ask=float(e["entry_ask"]),
                    entry_mid=float(e["entry_mid"]),
                    reclaim_level=float(e["reclaim_level"]),
                    pullback_low=pb_f,
                    range_high=range_high,
                    range_low=range_low,
                    vwap_at_entry=None,
                    atr=None,
                    last_progress_t=float(e["entry_t"]),
                    peak_mid=float(e["entry_mid"]),
                )
                if sym not in bufs:
                    bufs[sym] = FeatureBuffer()
                bufs[sym].push(t, bid, ask, row["vwap"], row["vol"])
                snap = bufs[sym].snapshot(t)
                if snap.get("complete"):
                    pos.atr = snap.get("atr_180s")
                    pos.vol30_at_entry = snap.get("volume_30s")
                    pos.vwap_at_entry = snap.get("vwap")
                positions[sym] = (pos, e)
                entered_eps.add(e["episode_id"])
                # force OPEN_INIT advance on entry event
                step_exit(
                    pos, t=t, bid=bid, ask=ask, mid=mid, vwap=vwap, spread_bps=spread,
                    volume_30s=pos.vol30_at_entry, price_update_count_10s=None,
                )

        # force close remaining at end
        if events:
            t, _, row = events[-1]
            bid = float(row["bid"])
            for sym, (pos, meta) in list(positions.items()):
                econ = net_pnl_yen(pos.entry_ask, bid)
                gross = (bid - pos.entry_ask) * LOT
                trades.append({
                    "pair_id": f"TAER_P3|{setup_type}|R10|{exit_candidate}",
                    "episode_id": meta["episode_id"],
                    "day": day,
                    "am_pm": "PM",
                    "symbol": sym,
                    "setup_type": setup_type,
                    "exit_candidate": exit_candidate,
                    "entry_t": pos.entry_t,
                    "exit_t": t,
                    "entry_ask": pos.entry_ask,
                    "exit_bid": bid,
                    "exit_state": pos.state,
                    "exit_reason": "SESSION_END",
                    "hold_sec": t - pos.entry_t,
                    "mfe_at_exit": pos.mfe,
                    "mae_at_exit": pos.mae,
                    "giveback_at_exit": pos.mfe - (bid - pos.entry_mid),
                    "pnl_before_cost": gross,
                    "pnl_after_5bps": econ["net_pnl_yen_100"],
                    "cost_yen_100": econ["cost_yen_100"],
                    "scenario_id_prior": meta.get("scenario_id_prior"),
                })
                reason_c["SESSION_END"] += 1
                if meta.get("scenario_id_prior") == "S6_LATE_CONTINUATION":
                    s6_total += 1

    pnls = [t["pnl_after_5bps"] for t in trades]
    day_pnl: dict[str, float] = defaultdict(float)
    for t in trades:
        day_pnl[t["day"]] += t["pnl_after_5bps"]
    pf, pf_st = _pf(pnls)
    wins = sum(1 for x in pnls if x > 0)
    losses = sum(1 for x in pnls if x < 0)
    draws = sum(1 for x in pnls if abs(x) <= 1e-12)
    ex722 = [t for t in trades if t["day"] != "20260722"]
    ex722_pnls = [t["pnl_after_5bps"] for t in ex722]
    ex722_pf, _ = _pf(ex722_pnls)

    # concentration
    by_sym = defaultdict(float)
    for t in trades:
        by_sym[t["symbol"]] += t["pnl_after_5bps"]
    top_sym = max(by_sym.items(), key=lambda x: abs(x[1])) if by_sym else (None, 0.0)
    top_trade = max(trades, key=lambda x: abs(x["pnl_after_5bps"])) if trades else None
    top_day = max(day_pnl.items(), key=lambda x: abs(x[1])) if day_pnl else (None, 0.0)

    # day deletion
    day_del = []
    for d in sorted(day_pnl):
        rem = sum(v for k, v in day_pnl.items() if k != d)
        day_del.append({"held_out_day": d, "remaining_pnl": rem})

    # rolling origin diagnostic (confirm day pnl only)
    folds = []
    confirm_days = ["20260727", "20260728", "20260729", "20260730", "20260731"]
    for i, cd in enumerate(confirm_days, 1):
        folds.append({
            "fold": f"F{i}",
            "confirm": cd,
            "confirm_pnl": day_pnl.get(cd, 0.0),
            "selection_status": "DIAGNOSTIC_ONLY_NO_CORE",
        })

    ledger_sha = sha256_obj([
        {
            "episode_id": t["episode_id"], "entry_t": t["entry_t"], "exit_t": t["exit_t"],
            "exit_reason": t["exit_reason"], "exit_bid": t["exit_bid"],
            "pnl_after_5bps": t["pnl_after_5bps"],
        }
        for t in sorted(trades, key=lambda x: (x["entry_t"], x["symbol"], x["episode_id"]))
    ])

    return {
        "pair_id": f"TAER_P3|{setup_type}|R10|{exit_candidate}",
        "entry_profile": "TAER_P3",
        "setup_type": setup_type,
        "retention": "R10",
        "exit_candidate": exit_candidate,
        "n": len(trades),
        "pnl": sum(pnls) if pnls else 0.0,
        "pf": pf,
        "pf_status": pf_st,
        "wld": {"w": wins, "l": losses, "d": draws},
        "avg_pnl": (sum(pnls) / len(pnls)) if pnls else None,
        "max_dd": _max_dd(dict(day_pnl)),
        "exit_reason_counts": dict(reason_c),
        "stop_n": reason_c.get("HARD_STOP", 0),
        "session_end_n": reason_c.get("SESSION_END", 0),
        "day_pnl": dict(day_pnl),
        "ex722_pnl": sum(ex722_pnls) if ex722_pnls else 0.0,
        "ex722_pf": ex722_pf,
        "ex722_n": len(ex722),
        "top1_day": {"day": top_day[0], "pnl": top_day[1]},
        "top1_symbol": {"symbol": top_sym[0], "pnl": top_sym[1]},
        "top1_trade_pnl": None if top_trade is None else top_trade["pnl_after_5bps"],
        "day_deletion": day_del,
        "rolling_origin": folds,
        "cap_blocked": cap_blocked,
        "follow_on_blocked": follow_on,
        "ledger_sha256": ledger_sha,
        "s6_audit": {
            "s6_trades_seen": s6_total,
            "s6_early_exit_lt_90s": s6_early_exit,
            "s6_early_exit_rate": (s6_early_exit / s6_total) if s6_total else None,
        },
        "transitions_sample": transitions_sample[:40],
        "trades": trades,
    }


def prior_exit_identity_audit() -> dict[str, Any]:
    """Answer Q1–Q10 about the frozen identical joint."""
    p2 = json.loads((PRIOR_STORE / "p2_exit_precommit.json").read_text(encoding="utf-8"))
    prior_report = json.loads((PRIOR_STORE / "report.json").read_text(encoding="utf-8"))
    # Source lines in run_study.py used path_tail bid for all three
    return {
        "q1_function_names": {
            "answer": [
                "research.e1_x6_taer.run_study.run_taer_study.<joint loop>",
                "research.e1_x6_provisional.cost_contract.net_pnl_yen",
            ],
            "detail": (
                "No distinct EXIT state-machine functions were called. "
                "X_STRUCTURAL/X_CONTINUATION/X_HYBRID only selected path_tail['bid'] identically."
            ),
        },
        "q2_settings_thresholds_transitions": {
            "answer": "NONE_DISTINCT",
            "thresholds": None,
            "state_transitions": "not implemented in prior run",
            "code_behavior": {
                "X_STRUCTURAL": "px = path_tail.bid (even if reclaim BROKEN)",
                "X_CONTINUATION": "px = path_tail.bid",
                "X_HYBRID": "px = path_tail.bid",
            },
        },
        "q3_exit_reason_counts": {
            "answer": "NOT_AVAILABLE",
            "detail": "Prior joint never recorded exit_reason per candidate; only path-end mark-to-market.",
        },
        "q4_trades_with_different_exit_times": {
            "answer": 0,
            "detail": "All candidates used the same path_tail timestamp implicitly.",
        },
        "q5_completed_ledger_sha_identical": True,
        "q6_identical_reason": (
            "EXIT_CANDIDATES_NOT_DISTINCT: joint loop assigned the same exit price "
            "(path_tail bid) for every candidate; no per-candidate event replay."
        ),
        "q7_fallback_exit_only": {
            "answer": True,
            "detail": "100% path-horizon/path_tail fallback; no live EXIT SM.",
        },
        "q8_p2_exit_precommit": {
            "precommit_at_jst": p2.get("precommit_at_jst"),
            "precommit_sha256": p2.get("precommit_sha256"),
            "body": p2,
        },
        "q9_precommit_before_economics_evidence": {
            "answer": "PARTIAL",
            "detail": (
                "P2 file timestamp precedes joint pnl aggregation in the same function, "
                "but EXIT definitions/thresholds were not fully specified — only candidate names."
            ),
            "p2_before_joint_in_code": True,
        },
        "q10_749_to_746": prior_report.get("entry_observation_n"),
        "status": "EXIT_CANDIDATES_NOT_DISTINCT",
        "prior_joint_invalidated": True,
        "corrected_verdict_for_prior_run": "E1_X6_INSUFFICIENT_EXIT_EVIDENCE",
        "prior_status": "TAER_ENTRY_PATH_READY",
    }


def run_exit_joint_audit() -> dict[str, Any]:
    run_id = f"e1x6_taer_exit_joint_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "taer" / run_id
    store.mkdir(parents=True, exist_ok=True)

    identity = prior_exit_identity_audit()
    entries, entry_meta = load_entry_observations()
    identity["q10_749_to_746"] = entry_meta["q10_749_to_746"]

    s7 = decompose_s7(entries)
    setup_sum = setup_path_summary(entries)

    # P2 precommit BEFORE pair economics
    p2 = {
        "precommit_type": "P2_EXIT_PRECOMMIT",
        "study_revision": STUDY_REVISION,
        "precommit_at_jst": datetime.now(JST).isoformat(),
        "prior_run_id": PRIOR_RUN,
        "prior_joint_invalidated": True,
        "prior_status": "EXIT_CANDIDATES_NOT_DISTINCT",
        "exit_candidates": list(EXIT_CANDIDATES),
        "pairs": [
            {"entry_profile": a, "setup_type": b, "retention": c, "exit_candidate": d}
            for a, b, c, d in PAIRS
        ],
        "thresholds": EXIT_THRESHOLDS,
        "state_machine": [
            "OPEN_INIT", "STRUCTURE_HOLD", "PROGRESS_CHECK",
            "PROFIT_PROTECTION", "TREND_MANAGEMENT", "EXIT",
        ],
        "functions": {
            "X_STRUCTURAL": "research.e1_x6_taer.exit_sm._structural",
            "X_CONTINUATION": "research.e1_x6_taer.exit_sm._continuation",
            "X_HYBRID": "research.e1_x6_taer.exit_sm._hybrid",
            "step": "research.e1_x6_taer.exit_sm.step_exit",
        },
        "setup_specific_structural": True,
        "s7_excluded_from_threshold_fit": True,
        "economics_opened_before_precommit": False,
        "max_pairs": len(PAIRS),
    }
    p2_path = store / "p2_exit_precommit.json"
    p2_path.write_text(json.dumps(p2, ensure_ascii=False, indent=2), encoding="utf-8")
    p2["precommit_sha256"] = sha256_file(p2_path)
    p2_path.write_text(json.dumps(p2, ensure_ascii=False, indent=2), encoding="utf-8")
    print("P2_EXIT_PRECOMMIT", p2["precommit_sha256"], p2["precommit_at_jst"], flush=True)

    # load events
    sm = load_source_manifest()
    day_events = {}
    for day in DAYS:
        uni = _universe_from_manifest(sm, day)
        print(f"  preload {day}", flush=True)
        day_events[day] = load_day_events(day, uni)

    pair_results = {}
    for profile, setup, ret, xc in PAIRS:
        print(f"=== joint {profile} x {setup} x {ret} x {xc} ===", flush=True)
        res = replay_pair(entries, setup_type=setup, exit_candidate=xc, day_events=day_events)
        # drop bulky trades from memory in summary; write ledger file
        ledger_fp = store / f"ledger_{setup}_{xc}.jsonl"
        with ledger_fp.open("w", encoding="utf-8") as f:
            for tr in res["trades"]:
                f.write(json.dumps(tr, ensure_ascii=False, default=str) + "\n")
        slim = {k: v for k, v in res.items() if k != "trades"}
        slim["ledger_path"] = str(ledger_fp)
        slim["ledger_n"] = res["n"]
        pair_results[res["pair_id"]] = slim
        print(
            f"  n={res['n']} pnl={res['pnl']:.1f} pf={res['pf']} sha={res['ledger_sha256'][:16]} "
            f"reasons={res['exit_reason_counts']}",
            flush=True,
        )

    shas = [pair_results[pid]["ledger_sha256"] for pid in pair_results]
    distinct = len(set(shas)) == len(shas)
    # also check exit-time diversity between structural vs continuation same setup
    time_diff_counts = {}
    for setup in ("PULLBACK_RECLAIM", "RANGE_BREAKOUT"):
        a = pair_results[f"TAER_P3|{setup}|R10|X_STRUCTURAL"]
        b = pair_results[f"TAER_P3|{setup}|R10|X_CONTINUATION"]
        # compare via re-reading ledgers
        la = [json.loads(x) for x in (store / f"ledger_{setup}_X_STRUCTURAL.jsonl").read_text(encoding="utf-8").splitlines() if x.strip()]
        lb = [json.loads(x) for x in (store / f"ledger_{setup}_X_CONTINUATION.jsonl").read_text(encoding="utf-8").splitlines() if x.strip()]
        mb = {t["episode_id"]: t for t in lb}
        diff = 0
        for t in la:
            u = mb.get(t["episode_id"])
            if u and abs(float(u["exit_t"]) - float(t["exit_t"])) > 1e-6:
                diff += 1
        time_diff_counts[setup] = {"compared": len(la), "different_exit_times": diff}

    if not distinct:
        verdict = "E1_X6_INSUFFICIENT_EXIT_EVIDENCE"
        joint_status = "EXIT_CANDIDATES_NOT_DISTINCT"
    else:
        verdict = "E1_X6_RESEARCH_PAIR_PENDING_CORE_EVIDENCE"
        joint_status = "EXIT_PAIRS_DISTINCT_DIAGNOSTIC"

    report = {
        "phase": "TAER_EXIT_JOINT_COMPLETION_AUDIT",
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "prior_run_id": PRIOR_RUN,
        "prior_status": "TAER_ENTRY_PATH_READY",
        "prior_verdict_correction": "E1_X6_INSUFFICIENT_EXIT_EVIDENCE",
        "verdict": verdict,
        "joint_status": joint_status,
        "identity_audit": identity,
        "entry_meta": entry_meta,
        "s7_decomposition": s7,
        "setup_path_summary": setup_sum,
        "p2_exit_precommit_sha256": p2["precommit_sha256"],
        "p2_exit_precommit_at_jst": p2["precommit_at_jst"],
        "p2_exit_precommit": p2,
        "pairs": pair_results,
        "ledger_sha_distinct": distinct,
        "exit_time_diff_structural_vs_continuation": time_diff_counts,
        "core_valid": 0,
        "adoption": "INTERNAL_DIAGNOSTIC_ONLY" if verdict.endswith("PENDING_CORE_EVIDENCE") else "INSUFFICIENT_EXIT",
        "safety": {"submit": 0, "cancel": 0, "live": 0},
        "mainline_changed": False,
        "SHADOW_STARTED": False,
        "FORWARD_STARTED": False,
        "PAPER_STARTED": False,
        "DISCORD_SENT": False,
    }

    _publish(report, store)
    return report


def _publish(report: dict[str, Any], store: Path) -> None:
    out = NATIVE / "results" / "research" / "e1_x6_taer_exit_joint_audit"
    out.mkdir(parents=True, exist_ok=True)
    # strip huge nested trades already slim
    (store / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    md = [
        f"# TAER EXIT Joint Audit — {report['verdict']}",
        "",
        f"- prior_run: `{report['prior_run_id']}` status=`TAER_ENTRY_PATH_READY`",
        f"- prior verdict correction: `E1_X6_INSUFFICIENT_EXIT_EVIDENCE`",
        f"- joint_status: `{report['joint_status']}`",
        f"- P2_EXIT_PRECOMMIT: `{report['p2_exit_precommit_sha256']}` at {report['p2_exit_precommit_at_jst']}",
        f"- ledger_sha_distinct: {report['ledger_sha_distinct']}",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Pairs",
    ]
    for pid, p in report["pairs"].items():
        md.append(
            f"- `{pid}` n={p['n']} pnl={p['pnl']:.1f} pf={p['pf']} "
            f"sha=`{p['ledger_sha256'][:16]}…` reasons={p['exit_reason_counts']}"
        )
    (out / "report.md").write_text("\n".join(md), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()

    def sheet(name: str, headers: list, rows: list) -> None:
        w = wb.create_sheet(name)
        w.append(headers)
        for r in rows:
            w.append(r)

    ws = wb.active
    ws.title = "Index"
    for row in (
        ("verdict", report["verdict"]),
        ("joint_status", report["joint_status"]),
        ("prior_run", report["prior_run_id"]),
        ("run_id", report["run_id"]),
        ("p2_sha", report["p2_exit_precommit_sha256"]),
    ):
        ws.append(list(row))

    sheet("Precommit", ["key", "value"],
          [[k, json.dumps(v, ensure_ascii=False, default=str)[:30000] if isinstance(v, (dict, list)) else v]
           for k, v in report["p2_exit_precommit"].items()])
    sheet("AnchorAudit", ["note"], [["see prior run anchor_audit; ENTRY path ready"]])
    sheet("EntryObservations", ["usable_n", "excluded_n"],
          [[report["entry_meta"]["usable_n"], report["entry_meta"]["excluded_n"]]])
    sheet("PathLedger", ["note"], [[f"prior path_ledger at {PRIOR_STORE/'path_ledger.jsonl'}"]])
    sheet("ScenarioSummary", ["setup", "scenario", "n"],
          [[sk, sc, n] for sk, block in report["setup_path_summary"].items()
           for sc, n in (block.get("scenario_counts") or {}).items()])
    sheet("ScenarioS7", ["label", "n", "by_day", "by_setup", "mfe_mean", "mae_mean", "path_n_mean"],
          [[k, v.get("n"), json.dumps(v.get("by_day")), json.dumps(v.get("by_setup")),
            v.get("mfe_mean"), v.get("mae_mean"), v.get("path_n_mean")]
           for k, v in report["s7_decomposition"].items() if k.startswith("S7_")])
    sheet("SetupSummary", ["setup", "n", "mfe_mean", "mae_mean"],
          [[k, v["n"], v["mfe_mean"], v["mae_mean"]] for k, v in report["setup_path_summary"].items()])
    sheet("ExitDefinitions", ["candidate", "function", "thresholds"],
          [[c, report["p2_exit_precommit"]["functions"][c],
            json.dumps(EXIT_THRESHOLDS, ensure_ascii=False)[:20000]] for c in EXIT_CANDIDATES])
    sheet("ExitTransitions", ["pair", "sample"],
          [[pid, json.dumps(p.get("transitions_sample") or [], ensure_ascii=False)[:30000]]
           for pid, p in report["pairs"].items()])
    sheet("JointPairs",
          ["pair_id", "n", "pnl", "pf", "w", "l", "d", "max_dd", "ledger_sha", "s6_early_rate"],
          [[pid, p["n"], p["pnl"], p["pf"], p["wld"]["w"], p["wld"]["l"], p["wld"]["d"],
            p["max_dd"], p["ledger_sha256"], (p.get("s6_audit") or {}).get("s6_early_exit_rate")]
           for pid, p in report["pairs"].items()])

    # JointTrades — first 5000 across pairs
    jt = wb.create_sheet("JointTrades")
    jt.append(["pair_id", "day", "symbol", "entry_t", "exit_t", "exit_reason", "hold_sec",
               "mfe", "mae", "pnl_after_5bps"])
    ntr = 0
    for setup in ("PULLBACK_RECLAIM", "RANGE_BREAKOUT"):
        for xc in EXIT_CANDIDATES:
            fp = store / f"ledger_{setup}_{xc}.jsonl"
            if not fp.exists():
                continue
            with fp.open(encoding="utf-8") as f:
                for line in f:
                    if ntr >= 5000:
                        break
                    t = json.loads(line)
                    jt.append([t["pair_id"], t["day"], t["symbol"], t["entry_t"], t["exit_t"],
                               t["exit_reason"], t["hold_sec"], t["mfe_at_exit"], t["mae_at_exit"],
                               t["pnl_after_5bps"]])
                    ntr += 1
            if ntr >= 5000:
                break
        if ntr >= 5000:
            break

    daily_rows = []
    for pid, p in report["pairs"].items():
        for d, pnl in (p.get("day_pnl") or {}).items():
            daily_rows.append([pid, d, pnl])
    sheet("Daily", ["pair_id", "day", "pnl"], daily_rows)

    wf_rows = []
    for pid, p in report["pairs"].items():
        for f in p.get("rolling_origin") or []:
            wf_rows.append([pid, f["fold"], f["confirm"], f["confirm_pnl"]])
    sheet("WalkForward", ["pair_id", "fold", "confirm", "confirm_pnl"], wf_rows)

    dd_rows = []
    for pid, p in report["pairs"].items():
        for r in p.get("day_deletion") or []:
            dd_rows.append([pid, r["held_out_day"], r["remaining_pnl"]])
    sheet("DayDeletion", ["pair_id", "held_out", "remaining_pnl"], dd_rows)

    sheet("Concentration", ["pair_id", "top1_day", "top1_symbol", "top1_trade_pnl"],
          [[pid, json.dumps(p.get("top1_day")), json.dumps(p.get("top1_symbol")), p.get("top1_trade_pnl")]
           for pid, p in report["pairs"].items()])
    sheet("Tests", ["item", "result"],
          [["ledger_sha_distinct", report["ledger_sha_distinct"]],
           ["identity_status", report["identity_audit"]["status"]]])
    sheet("Determinism", ["setup", "struct_vs_cont_diff_exit_times"],
          [[k, json.dumps(v)] for k, v in report["exit_time_diff_structural_vs_continuation"].items()])
    sheet("Safety", ["key", "value"],
          [["submit", 0], ["cancel", 0], ["live", 0], ["mainline_changed", False],
           ["SHADOW", False], ["FORWARD", False], ["PAPER", False], ["DISCORD", False]])
    sheet("ChangeLog", ["item", "note"],
          [["prior_joint", "invalidated EXIT_CANDIDATES_NOT_DISTINCT"],
           ["new_exit_sm", "setup-aware structural / continuation / hybrid"],
           ["verdict_rule", "PENDING_CORE only if ledgers distinct"]])

    # IdentityAnswers
    sheet("IdentityAnswers", ["q", "value"],
          [[k, json.dumps(v, ensure_ascii=False, default=str)[:30000]]
           for k, v in report["identity_audit"].items()])

    wb.save(out / "audit.xlsx")
    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "published_shas.json").write_text(json.dumps(report["published"], indent=2), encoding="utf-8")


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    rep = run_exit_joint_audit()
    print("EXIT_JOINT_DONE", rep["run_id"], rep["verdict"], rep["joint_status"], flush=True)
