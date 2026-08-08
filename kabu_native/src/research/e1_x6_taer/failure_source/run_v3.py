"""Orchestrate FSA V3 — label contract repair + feature stability (A/B)."""
from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.config import DAYS
from research.e1_x6_taer.failure_source.clusters import load_episodes
from research.e1_x6_taer.failure_source.v3_analysis import (
    bootstrap_v3,
    join_target_feature_rows,
    lodo_v3,
    models_v3,
    opportunity_summary_v3,
    univariate_v3,
    verdict_v3,
)
from research.e1_x6_taer.failure_source.v3_features import feature_schema_gate, rebuild_feature_table_for_reps
from research.e1_x6_taer.failure_source.v3_identity import (
    ANALYSIS_ID_V3,
    LOCKED_CLUSTER_SHA,
    LOCKED_EPISODE_SHA,
    LOCKED_OPPORTUNITY_SHA,
    PURPOSE_V3,
    V2_RUN,
    V2_STORE,
    annotate_v2_scope,
)
from research.e1_x6_taer.failure_source.v3_label import build_label_audit
from research.e1_x6_taer.failure_source.v3_precommit import TARGET_VALID_MIN_RATE, build_v3_precommit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[4]


def _load_v2_opportunity() -> list[dict]:
    rows = []
    with (V2_STORE / "opportunity.jsonl").open(encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def _verify_frozen_identity(opp: list[dict], episodes: list[dict]) -> dict:
    episode_sha = sha256_obj(sorted(e["episode_id"] for e in episodes))
    cluster_sha = sha256_obj(sorted({
        (e.get("overlap_cluster_id"), e["episode_id"], e.get("is_cluster_representative"))
        for e in episodes
    }))
    # rebuild cluster fields onto episodes from opp if needed
    opp_sha = sha256_obj(sorted(
        (r["episode_id"], r.get("best_net_pnl_bps_300s"), r.get("entry_price"), r.get("best_exit_bid"))
        for r in opp
    ))
    ok = (
        episode_sha == LOCKED_EPISODE_SHA
        and opp_sha == LOCKED_OPPORTUNITY_SHA
    )
    return {
        "ok": ok and True,  # cluster checked after merge
        "episode_identity_sha": episode_sha,
        "cluster_identity_sha": cluster_sha,
        "opportunity_table_sha": opp_sha,
        "locked_episode": LOCKED_EPISODE_SHA,
        "locked_cluster": LOCKED_CLUSTER_SHA,
        "locked_opportunity": LOCKED_OPPORTUNITY_SHA,
        "episode_match": episode_sha == LOCKED_EPISODE_SHA,
        "opportunity_match": opp_sha == LOCKED_OPPORTUNITY_SHA,
        "cluster_match": cluster_sha == LOCKED_CLUSTER_SHA,
    }


def _merge_cluster_from_opp(episodes: list[dict], opp: list[dict]) -> list[dict]:
    by = {r["episode_id"]: r for r in opp}
    out = []
    for e in episodes:
        o = by.get(e["episode_id"]) or {}
        out.append({
            **e,
            "overlap_cluster_id": o.get("overlap_cluster_id") or e.get("overlap_cluster_id"),
            "is_cluster_representative": o.get("is_cluster_representative"),
            "cluster_size": o.get("cluster_size"),
            "cluster_weight": o.get("cluster_weight"),
            "session": o.get("session") or e.get("session"),
        })
    return out


def _s7_reasons() -> dict[str, str]:
    fp = V2_STORE / "s7_audit.json"
    if not fp.exists():
        return {}
    data = json.loads(fp.read_text(encoding="utf-8"))
    out = {}
    for r in data.get("rows_sample") or []:
        out[r["episode_id"]] = r.get("reason")
    # also reconstruct from full if only sample — mark all S7 from opportunity via audit file rows_all
    # Expand: re-read opportunity and assign CONFLICTING when in sample reasons; else OTHER for S7
    return out


def _run_tests() -> dict:
    env = dict(os.environ)
    env["PYTHONPATH"] = f"{NATIVE / 'src'};{NATIVE / 'research'}"
    env["PYTHONIOENCODING"] = "utf-8"
    proc = subprocess.run(
        [sys.executable, "-m", "pytest", "-q", "-rA", "--tb=line",
         "-p", "no:cacheprovider",
         str(NATIVE / "tests" / "test_e1_x6_taer_failure_source_v3.py")],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(NATIVE), env=env, timeout=300,
    )
    rows = []
    for line in (proc.stdout or "").splitlines():
        for st in ("PASSED", "FAILED", "ERROR"):
            if line.strip().startswith(st + " "):
                rows.append({"test": line.strip().split(" ", 1)[1].split(" - ")[0], "outcome": st})
                break
    return {
        "exit_code": proc.returncode,
        "passed": sum(1 for r in rows if r["outcome"] == "PASSED"),
        "failed": sum(1 for r in rows if r["outcome"] != "PASSED"),
        "total": len(rows),
        "rows": rows,
        "tail": (proc.stdout or "")[-2000:],
    }


def _one_pass(events_by_day: dict, opp: list[dict], episodes: list[dict], s7_reasons: dict) -> dict:
    pre = build_v3_precommit()
    reps = [r for r in opp if r.get("is_cluster_representative")]
    assert len(reps) == 399

    # Expand S7 reasons for all S7 reps not in sample
    for r in reps:
        scen = str(r.get("scenario_id_prior") or "")
        if scen.startswith("S7") and r["episode_id"] not in s7_reasons:
            # default ambiguous S7 without conflicting evidence -> still scenario-valid S7
            # V2 audit classified most as CONFLICTING — use path mfe/mae if available via episode
            s7_reasons.setdefault(r["episode_id"], "OTHER")

    # Prefer full S7 classification: if prior scenario is S7 and V2 said CONFLICTING for majority,
    # recompute quickly for all S7 using same rules as audit_s7
    from research.e1_x6_taer.failure_source.analysis import audit_s7
    s7_full = audit_s7(opp, episodes)
    for r in s7_full.get("rows_sample") or []:
        s7_reasons[r["episode_id"]] = r["reason"]
    # rows_sample may be capped — re-run classification inline for all S7
    ep_by = {e["episode_id"]: e for e in episodes}
    for r in opp:
        scen = str(r.get("scenario_id_prior") or "")
        if not scen.startswith("S7"):
            continue
        if r["episode_id"] in s7_reasons and s7_reasons[r["episode_id"]] != "OTHER":
            continue
        e = ep_by.get(r["episode_id"]) or {}
        path_n = int(e.get("path_n_prior") or r.get("path_event_count") or 0)
        cr = e.get("censor_reason_prior") or ""
        reason = "OTHER"
        if cr in ("SESSION_GAP",) or "SESSION" in str(cr).upper():
            reason = "SESSION_BOUNDARY"
        elif path_n < 5 or (not r.get("path_complete") and path_n < 10):
            reason = "INSUFFICIENT_HORIZON"
        elif r.get("first_touch_plus_5_or_minus_10") == "NONE" and r.get("first_touch_plus_10_or_minus_15") == "NONE":
            reason = "NO_CLEAR_FIRST_TOUCH"
        elif not r.get("evaluable") or r.get("best_net_pnl_bps_300s") is None:
            reason = "STALE_OR_MISSING_PATH"
        elif (e.get("mfe_prior") or 0) > 0 and (e.get("mae_prior") or 0) < 0:
            reason = "CONFLICTING_SCENARIO"
        s7_reasons[r["episode_id"]] = reason

    labels, target_summary = build_label_audit(reps, s7_reasons)

    # Early stop on target quality
    if (target_summary.get("opportunity_target_valid_rate") or 0) < TARGET_VALID_MIN_RATE:
        verd = verdict_v3(
            target_summary=target_summary, schema={"status": "SKIP"},
            opp_summary={}, lodo={}, boot={}, models={},
        )
        return {
            "precommit": pre,
            "labels": labels,
            "target_summary": target_summary,
            "schema": {"status": "SKIP"},
            "feat_rows": [],
            "joined": [],
            "opp_summary": opportunity_summary_v3(labels),
            "univariate": None,
            "lodo": None,
            "bootstrap": None,
            "models": None,
            "verdict": verd,
            "shas": {
                "target_validity_sha": sha256_obj(labels),
                "feature_table_sha": None,
            },
        }

    ep_by = {e["episode_id"]: e for e in episodes}
    feat_rows, feat_meta = rebuild_feature_table_for_reps(reps, ep_by, events_by_day)
    schema = feature_schema_gate(feat_rows, labels)
    opp_summary = opportunity_summary_v3(labels)

    shas = {
        "target_validity_sha": sha256_obj([(
            r["cluster_id"], r["episode_id"], r["opportunity_target_valid"],
            r["scenario_label_valid"], r["best_net_pnl_bps_300s"],
        ) for r in labels]),
        "feature_table_sha": sha256_obj(sorted(
            (r["episode_id"], r.get("cluster_id"), r.get("missing_feature_count"),
             r.get("spread_bps"), r.get("volume_30s"), r.get("decision_time"), r.get("feature_asof_time"))
            for r in feat_rows
        )),
    }

    if schema.get("status") != "PASS":
        verd = verdict_v3(
            target_summary=target_summary, schema=schema,
            opp_summary=opp_summary, lodo={}, boot={}, models={},
        )
        return {
            "precommit": pre, "labels": labels, "target_summary": target_summary,
            "schema": schema, "feat_rows": feat_rows, "feat_meta": feat_meta,
            "joined": [], "opp_summary": opp_summary,
            "univariate": None, "lodo": None, "bootstrap": None, "models": None,
            "verdict": verd, "shas": shas,
        }

    joined = join_target_feature_rows(labels, feat_rows)
    # attach secondary from labels already in join; fix 60/120 from opp
    opp_by = {r["episode_id"]: r for r in reps}
    for j in joined:
        o = opp_by.get(j["episode_id"]) or {}
        j["best_net_pnl_bps_60s"] = o.get("best_net_pnl_bps_60s")
        j["best_net_pnl_bps_120s"] = o.get("best_net_pnl_bps_120s")

    uni = univariate_v3(joined, schema["coverage"])
    lodo = lodo_v3(joined, uni)
    boot = bootstrap_v3(joined, lodo)
    models = models_v3(joined, lodo, boot)
    verd = verdict_v3(
        target_summary=target_summary, schema=schema, opp_summary=opp_summary,
        lodo=lodo, boot=boot, models=models,
    )
    shas.update({
        "univariate_sha": sha256_obj(uni),
        "lodo_sha": sha256_obj(lodo),
        "bootstrap_sha": sha256_obj(boot),
        "model_sha": sha256_obj(models),
    })
    return {
        "precommit": pre,
        "labels": labels,
        "target_summary": target_summary,
        "schema": schema,
        "feat_rows": feat_rows,
        "feat_meta": feat_meta,
        "joined": joined,
        "opp_summary": opp_summary,
        "univariate": uni,
        "lodo": lodo,
        "bootstrap": boot,
        "models": models,
        "verdict": verd,
        "shas": shas,
    }


def run() -> dict:
    run_id = f"e1x6_taer_fsa_v3_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "taer" / run_id
    store.mkdir(parents=True, exist_ok=True)

    print("=== Annotate V2 scope (no overwrite) ===", flush=True)
    v2_scope = annotate_v2_scope()

    print("=== Precommit (before feature effects) ===", flush=True)
    pre0 = build_v3_precommit()
    (store / "precommit.json").write_text(json.dumps(pre0, indent=2), encoding="utf-8")
    print("precommit_sha", pre0["precommit_sha256"], flush=True)

    print("=== Tests ===", flush=True)
    tests = _run_tests()
    print("tests", tests["passed"], "/", tests["total"], "exit", tests["exit_code"], flush=True)
    if tests["exit_code"] != 0:
        report = {"verdict": "TAER_FAILURE_ANALYSIS_TESTS_FAILED", "tests": tests, "run_id": run_id}
        _publish(report, store, None)
        return report

    print("=== Load frozen V2 opportunity + episodes ===", flush=True)
    opp = _load_v2_opportunity()
    episodes_raw, ep_meta, _ = load_episodes()
    episodes = _merge_cluster_from_opp(episodes_raw, opp)
    ident = _verify_frozen_identity(opp, episodes)
    print("identity", {k: ident[k] for k in (
        "episode_match", "cluster_match", "opportunity_match"
    )}, flush=True)
    if not (ident["episode_match"] and ident["opportunity_match"] and ident["cluster_match"]):
        report = {
            "analysis_id": ANALYSIS_ID_V3,
            "run_id": run_id,
            "verdict": "TAER_FAILURE_ANALYSIS_IDENTITY_MISMATCH",
            "identity": ident,
            "tests": tests,
        }
        _publish(report, store, None)
        return report

    s7_reasons = _s7_reasons()

    print("=== Preload events ===", flush=True)
    sm = load_source_manifest()
    events_by_day = {}
    for day in DAYS:
        print("  preload", day, flush=True)
        events_by_day[day] = load_day_events(day, _universe_from_manifest(sm, day))

    print("=== Analysis A ===", flush=True)
    a = _one_pass(events_by_day, opp, episodes, dict(s7_reasons))
    print("=== Analysis B ===", flush=True)
    b = _one_pass(events_by_day, opp, episodes, dict(s7_reasons))

    det = {
        "ab_ok": a["shas"] == b["shas"] and a["verdict"]["verdict"] == b["verdict"]["verdict"],
        "shas_a": a["shas"],
        "shas_b": b["shas"],
        "verdict_a": a["verdict"]["verdict"],
        "verdict_b": b["verdict"]["verdict"],
    }

    report = {
        "phase": "TAER_FAILURE_SOURCE_ANALYSIS_V3",
        "analysis_id": ANALYSIS_ID_V3,
        "purpose": PURPOSE_V3,
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "verdict": a["verdict"]["verdict"],
        "verdict_detail": a["verdict"],
        "v2_scope": v2_scope,
        "precommit_sha256": a["precommit"]["precommit_sha256"],
        "frozen_identity": ident,
        "label_contract": a["target_summary"],
        "feature_schema": {
            "status": a["schema"].get("status"),
            "errors": a["schema"].get("errors"),
            "n_feature_rows": a["schema"].get("n_feature_rows"),
            "n_label_rows": a["schema"].get("n_label_rows"),
        },
        "feature_coverage": a["schema"].get("coverage"),
        "opportunity_summary": a["opp_summary"],
        "univariate": a.get("univariate"),
        "lodo": a.get("lodo"),
        "bootstrap": a.get("bootstrap"),
        "models": a.get("models"),
        "determinism": det,
        "shas": a["shas"],
        "tests": tests,
        "episode_meta": ep_meta,
        "safety": {
            "submit": 0, "cancel": 0, "live": 0,
            "mainline_changed": False,
            "taer_v1_unchanged": True,
            "taer_v2": False,
            "new_family": False,
            "shadow": False, "forward": False, "paper": False, "discord": False,
            "v2_run_not_overwritten": True,
        },
        "stop": True,
    }

    (store / "labels.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in a["labels"]) + "\n",
        encoding="utf-8",
    )
    (store / "features.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in a["feat_rows"]) + "\n",
        encoding="utf-8",
    )

    _publish(report, store, a)
    print("=== PUBLISHED ===", flush=True)
    print("run_id", run_id, flush=True)
    print("verdict", report["verdict"], flush=True)
    print("ab_ok", det["ab_ok"], flush=True)
    print("STOP", flush=True)
    return report


def _publish(report: dict, store: Path, payload) -> None:
    out = NATIVE / "results" / "research" / "e1_x6_taer_failure_source_v3"
    out.mkdir(parents=True, exist_ok=True)

    md = [
        f"# TAER Failure Source Analysis V3 — {report.get('verdict')}",
        "",
        f"- analysis_id: `{ANALYSIS_ID_V3}`",
        f"- purpose: `{PURPOSE_V3}`",
        f"- run_id: `{report.get('run_id')}`",
        f"- V2 scope: `FSA_V2_STOPPED_BY_SCENARIO_BASED_LABEL_QUALITY_GATE` (not overwritten)",
        f"- label_contract: `{((report.get('label_contract') or {}).get('opportunity_label_contract'))}`",
        f"- target_valid_rate: `{((report.get('label_contract') or {}).get('opportunity_target_valid_rate'))}`",
        f"- feature_schema: `{((report.get('feature_schema') or {}).get('status'))}`",
        f"- ab_ok: {(report.get('determinism') or {}).get('ab_ok')}",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Oracle vs realizable",
        "Opportunity envelope is oracle edge only — not proof of a realizable EXIT rule.",
        "",
        "## Stop",
        "No TAER V2, no new family implementation, no Shadow/Forward/Paper/Discord.",
    ]
    if payload and payload.get("opp_summary"):
        md.append("")
        md.append("## Opportunity (target-valid clusters; includes S7)")
        for setup, blk in (payload["opp_summary"].get("by_setup") or {}).items():
            o = blk.get("overall") or {}
            md.append(
                f"- `{setup}` n={o.get('cluster_n')} median={o.get('best_net_pnl_bps_median')} "
                f"+5rate={o.get('net_plus_5bps_rate')}"
            )

    (out / "report.md").write_text("\n".join(md), encoding="utf-8")
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    for row in (("analysis_id", ANALYSIS_ID_V3), ("verdict", report.get("verdict")),
                ("run_id", report.get("run_id")), ("v2_run", V2_RUN)):
        ws.append(list(row))

    def sheet(name, headers, rows):
        w = wb.create_sheet(name)
        w.append(headers)
        for r in rows:
            w.append([("" if v is None else v) for v in r])

    sheet("Precommit", ["key", "value"], [
        ["precommit_sha256", report.get("precommit_sha256")],
        ["purpose", PURPOSE_V3],
        ["locked_opportunity_sha", LOCKED_OPPORTUNITY_SHA],
    ])
    sheet("Identity", ["key", "value"],
          [[k, v] for k, v in (report.get("frozen_identity") or {}).items()])
    lc = report.get("label_contract") or {}
    sheet("LabelContract", ["metric", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v] for k, v in lc.items()])

    if payload:
        tv = wb.create_sheet("TargetValidity")
        tv.append(["cluster_id", "episode_id", "setup", "day", "symbol", "target_valid",
                   "best_300", "worst_300", "adverse", "integrity"])
        for r in payload.get("labels") or []:
            tv.append([r.get("cluster_id"), r.get("episode_id"), r.get("setup_type"), r.get("day"),
                       r.get("symbol"), r.get("opportunity_target_valid"),
                       r.get("best_net_pnl_bps_300s"), r.get("worst_net_pnl_bps_300s"),
                       r.get("adverse_before_best_bps"), r.get("integrity_status")])
        sv = wb.create_sheet("ScenarioValidity")
        sv.append(["cluster_id", "episode_id", "scenario_group", "scenario_valid", "invalid_reason"])
        for r in payload.get("labels") or []:
            sv.append([r.get("cluster_id"), r.get("episode_id"), r.get("scenario_group"),
                       r.get("scenario_label_valid"), r.get("scenario_invalid_reason")])
    else:
        sheet("TargetValidity", ["note"], [["no payload"]])
        sheet("ScenarioValidity", ["note"], [["no payload"]])

    sheet("EpisodeClusters", ["note"], [["Frozen from V2: overlap_cluster_n=399 CLUSTER_FIRST_EPISODE"]])
    sheet("OpportunityEnvelope", ["note"], [["Frozen V2 opportunity_table_sha; not recomputed"]])

    os_rows = []
    for setup, blk in ((report.get("opportunity_summary") or {}).get("by_setup") or {}).items():
        o = blk.get("overall") or {}
        os_rows.append([setup, "ALL", o.get("cluster_n"), o.get("best_net_pnl_bps_median"),
                        o.get("net_plus_5bps_rate")])
        for d, p in (blk.get("by_day") or {}).items():
            os_rows.append([setup, d, p.get("cluster_n"), p.get("best_net_pnl_bps_median"),
                            p.get("net_plus_5bps_rate")])
    sheet("OpportunitySummary", ["setup", "day", "n", "median_300", "plus5_rate"], os_rows)

    sheet("FeatureSchema", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("feature_schema") or {}).items()])

    if payload:
        ef = wb.create_sheet("EntryFeatures")
        hdr = ["episode_id", "cluster_id", "setup", "day", "decision_time", "feature_asof_time",
               "missing_feature_count", "spread_bps", "volume_impulse_ratio"]
        ef.append(hdr)
        for r in (payload.get("feat_rows") or [])[:5000]:
            ef.append([r.get(h) for h in hdr])
    else:
        sheet("EntryFeatures", ["note"], [["no payload"]])

    sheet("FeatureCoverage",
          ["feature", "applicable_setup", "applicable_n", "non_missing_n", "missing_rate", "eligible"],
          [[c.get("feature"), c.get("applicable_setup"), c.get("applicable_n"), c.get("non_missing_n"),
            c.get("missing_rate"), c.get("primary_candidate_eligible")]
           for c in (report.get("feature_coverage") or [])])

    uni_rows = []
    for setup, feats in ((report.get("univariate") or {}).get("by_setup") or {}).items():
        for f in feats:
            uni_rows.append([setup, f.get("feature"), f.get("n_non_missing"), f.get("spearman"),
                             f.get("median_split_effect_bps"), f.get("direction"), f.get("missing_rate")])
    sheet("FeatureUnivariate",
          ["setup", "feature", "n", "spearman", "effect", "direction", "missing_rate"], uni_rows)

    lodo_rows = []
    for setup, feats in (report.get("lodo") or {}).items():
        for f in feats:
            lodo_rows.append([setup, f.get("feature"), f.get("full_period_direction"),
                              f.get("same_direction_rate"), f.get("direction_reversal_count"),
                              f.get("evaluable_deletions"), f.get("stable_candidate")])
    sheet("FeatureLODO",
          ["setup", "feature", "full_dir", "same_rate", "reversals", "evaluable", "stable"], lodo_rows)

    boot_rows = []
    for setup, feats in (report.get("bootstrap") or {}).items():
        for f in feats:
            boot_rows.append([setup, f.get("feature"), json.dumps(f.get("effect_ci95")),
                              f.get("ci_crosses_0"), f.get("strong_stable")])
    sheet("Bootstrap", ["setup", "feature", "ci95", "crosses0", "strong"], boot_rows)

    md_rows = []
    for setup, blk in ((report.get("models") or {}).get("by_setup") or {}).items():
        md_rows.append([setup, blk.get("skipped"), blk.get("median_auc"),
                        blk.get("auc_gt_0_55_days"), json.dumps(blk.get("gates"))])
    sheet("ModelDiagnostics", ["setup", "skipped", "median_auc", "auc_gt_055", "gates"], md_rows)

    sheet("DataQuality", ["metric", "value"], [
        ["opportunity_target_valid_n", lc.get("opportunity_target_valid_n")],
        ["scenario_label_valid_n", lc.get("scenario_label_valid_n")],
        ["target_valid_but_scenario_invalid_n", lc.get("target_valid_but_scenario_invalid_n")],
    ])
    sheet("Verdict", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("verdict_detail") or {"verdict": report.get("verdict")}).items()])
    sheet("Tests", ["test", "outcome"],
          [[r["test"], r["outcome"]] for r in (report.get("tests") or {}).get("rows", [])])
    sheet("Determinism", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("determinism") or {}).items()])
    sheet("Safety", ["key", "value"], [[k, v] for k, v in (report.get("safety") or {}).items()])
    sheet("ChangeLog", ["item", "note"], [
        ["v2", "stopped by scenario-based label quality gate; not evidence of no-opportunity"],
        ["v3", "opportunity target validity separated from scenario validity; S7 kept in opportunity analysis"],
        ["stop", "no new family / no TAER V2"],
    ])
    wb.save(out / "audit.xlsx")

    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (store / "published_shas.json").write_text(json.dumps(report["published"], indent=2), encoding="utf-8")


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    run()
