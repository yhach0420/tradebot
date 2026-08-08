"""Orchestrate TAER Failure Source Analysis V2 (A/B deterministic)."""
from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.config import DAYS
from research.e1_x6_taer.failure_source import ANALYSIS_ID, CANONICAL_RUN, CANONICAL_VERDICT
from research.e1_x6_taer.failure_source.analysis import (
    audit_s7,
    bootstrap_ci,
    final_verdict,
    judge_opportunity_exists,
    label_quality,
    lodo_stability,
    model_diagnostics,
    opportunity_summary,
    primary_rows,
    univariate_analysis,
)
from research.e1_x6_taer.failure_source.clusters import build_overlap_clusters, load_episodes
from research.e1_x6_taer.failure_source.opportunity import compute_opportunity_and_features
from research.e1_x6_taer.failure_source.precommit import LOCKED_P1, LOCKED_P2, build_precommit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[4]
ENTRY_STORE = Path.home() / "e1x6_research_store" / "taer" / "e1x6_taer_20260803_232514"
INTEGRITY_STORE = Path.home() / "e1x6_research_store" / "taer" / CANONICAL_RUN


def freeze_taer_v1_closeout() -> dict:
    """Freeze canonical integrity run as CLOSED_NO_ROBUST_PAIR (do not overwrite economics)."""
    INTEGRITY_STORE.mkdir(parents=True, exist_ok=True)
    body = {
        "study_revision": "E1_X6_TRIGGER_ANCHORED_ENTRY_EXIT_JOINT_V1",
        "family_status": "CLOSED_NO_ROBUST_PAIR",
        "final_verdict": CANONICAL_VERDICT,
        "run_id": CANONICAL_RUN,
        "economic_integrity_status": "PASS",
        "period": "20260721-20260731",
        "period_status": "EXPLORATORY_FAILURE_ANALYSIS_ONLY",
        "locked_p1": LOCKED_P1,
        "locked_p2": LOCKED_P2,
        "frozen_at_jst": datetime.now(JST).isoformat(),
        "overwrite_forbidden": True,
        "not_for_holdout_forward_shadow": True,
        "no_taer_v2_from_this_closeout": True,
    }
    fp = INTEGRITY_STORE / "FAMILY_CLOSEOUT.json"
    if not fp.exists():
        fp.write_text(json.dumps(body, indent=2), encoding="utf-8")
    return body


def _run_tests() -> dict:
    env = dict(os.environ)
    env["PYTHONPATH"] = f"{NATIVE / 'src'};{NATIVE / 'research'}"
    env["PYTHONIOENCODING"] = "utf-8"
    proc = subprocess.run(
        [sys.executable, "-m", "pytest", "-q", "-rA", "--tb=line",
         "-p", "no:cacheprovider",
         str(NATIVE / "tests" / "test_e1_x6_taer_failure_source.py")],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(NATIVE), env=env, timeout=300,
    )
    rows = []
    for line in (proc.stdout or "").splitlines():
        for st in ("PASSED", "FAILED", "ERROR"):
            if line.strip().startswith(st + " "):
                rows.append({"test": line.strip().split(" ", 1)[1].split(" - ")[0], "outcome": st})
                break
    return {
        "exit_code": proc.returncode,
        "passed": sum(1 for r in rows if r["outcome"] == "PASSED"),
        "failed": sum(1 for r in rows if r["outcome"] != "PASSED"),
        "total": len(rows),
        "rows": rows,
        "tail": (proc.stdout or "")[-2000:],
    }


def _one_pass(events_by_day: dict) -> dict:
    episodes_raw, ep_meta, _ = load_episodes()
    episodes, cluster_summary = build_overlap_clusters(episodes_raw)
    pre = build_precommit(
        episode_ids=[e["episode_id"] for e in episodes],
        path_ledger_n=ep_meta["path_ledger_n"],
        usable_n=ep_meta["usable_n"],
        excluded_n=ep_meta["excluded_n"],
    )
    opp, feats, opp_meta = compute_opportunity_and_features(episodes, events_by_day)
    summary = opportunity_summary(opp, representatives_only=True)
    opp_judge = judge_opportunity_exists(summary)
    s7 = audit_s7(opp, episodes)
    prim_opp, prim_feat, prim_meta = primary_rows(opp, feats)
    lq = label_quality(cluster_summary["overlap_cluster_n"], prim_meta["primary_cluster_n"])

    # Early exits still compute identity SHAs
    identity = {
        "episode_identity_sha": sha256_obj(sorted(e["episode_id"] for e in episodes)),
        "cluster_identity_sha": sha256_obj(sorted({
            (e["overlap_cluster_id"], e["episode_id"], e["is_cluster_representative"])
            for e in episodes
        })),
        "opportunity_table_sha": sha256_obj(sorted(
            (r["episode_id"], r.get("best_net_pnl_bps_300s"), r.get("entry_price"), r.get("best_exit_bid"))
            for r in opp
        )),
        "feature_table_sha": sha256_obj(sorted(
            (r["episode_id"], r.get("missing_feature_count"), r.get("spread_bps"), r.get("volume_30s"))
            for r in feats
        )),
    }

    payload = {
        "ep_meta": ep_meta,
        "cluster_summary": cluster_summary,
        "precommit": pre,
        "opp_meta": opp_meta,
        "opp_n": len(opp),
        "feat_n": len(feats),
        "opportunity_summary": summary,
        "opportunity_judge": opp_judge,
        "s7_audit": s7,
        "primary_meta": prim_meta,
        "label_quality": lq,
        "identity": identity,
        "opp_rows": opp,
        "feat_rows": feats,
        "prim_opp": prim_opp,
        "prim_feat": prim_feat,
    }

    if lq.get("insufficient"):
        payload["univariate"] = None
        payload["lodo"] = None
        payload["bootstrap"] = None
        payload["models"] = {"ran": False, "skipped_reason": "label_quality"}
        payload["verdict"] = final_verdict(label_q=lq, opp_judge=opp_judge, lodo={}, boot={}, models={})
        return payload

    if not opp_judge.get("any_setup_opportunity_exists"):
        payload["univariate"] = None
        payload["lodo"] = None
        payload["bootstrap"] = None
        payload["models"] = {"ran": False, "skipped_reason": "no_opportunity"}
        payload["verdict"] = final_verdict(label_q=lq, opp_judge=opp_judge, lodo={}, boot={}, models={})
        return payload

    uni = univariate_analysis(prim_opp, prim_feat)
    lodo = lodo_stability(prim_opp, prim_feat, uni)
    boot = bootstrap_ci(prim_opp, prim_feat, lodo)
    models = model_diagnostics(prim_opp, prim_feat, lodo, boot, opp_judge)
    verd = final_verdict(label_q=lq, opp_judge=opp_judge, lodo=lodo, boot=boot, models=models)
    payload.update({
        "univariate": uni,
        "lodo": lodo,
        "bootstrap": boot,
        "models": models,
        "verdict": verd,
        "univariate_sha": sha256_obj(uni),
        "model_sha": sha256_obj(models),
    })
    return payload


def run() -> dict:
    run_id = f"e1x6_taer_fsa_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "taer" / run_id
    store.mkdir(parents=True, exist_ok=True)

    print("=== TAER V1 Closeout freeze ===", flush=True)
    closeout = freeze_taer_v1_closeout()

    print("=== Tests ===", flush=True)
    tests = _run_tests()
    print("tests", tests["passed"], "/", tests["total"], "exit", tests["exit_code"], flush=True)
    if tests["exit_code"] != 0 or tests["failed"] > 0:
        report = {
            "analysis_id": ANALYSIS_ID,
            "run_id": run_id,
            "verdict": "TAER_FAILURE_ANALYSIS_TESTS_FAILED",
            "tests": tests,
            "safety": {"submit": 0, "cancel": 0, "live": 0},
        }
        _publish(report, store)
        return report

    print("=== Preload events ===", flush=True)
    sm = load_source_manifest()
    events_by_day = {}
    for day in DAYS:
        print("  preload", day, flush=True)
        uni = _universe_from_manifest(sm, day)
        events_by_day[day] = load_day_events(day, uni)

    print("=== Analysis A ===", flush=True)
    a = _one_pass(events_by_day)
    # persist precommit immediately (economics already computed in same pass after precommit body built)
    (store / "precommit.json").write_text(json.dumps(a["precommit"], indent=2), encoding="utf-8")

    print("=== Analysis B (determinism) ===", flush=True)
    b = _one_pass(events_by_day)

    det = {
        "ab_ok": a["identity"] == b["identity"]
        and a["verdict"]["verdict"] == b["verdict"]["verdict"]
        and a.get("univariate_sha") == b.get("univariate_sha")
        and a.get("model_sha") == b.get("model_sha"),
        "identity_a": a["identity"],
        "identity_b": b["identity"],
        "verdict_a": a["verdict"]["verdict"],
        "verdict_b": b["verdict"]["verdict"],
        "univariate_sha_a": a.get("univariate_sha"),
        "univariate_sha_b": b.get("univariate_sha"),
        "model_sha_a": a.get("model_sha"),
        "model_sha_b": b.get("model_sha"),
    }

    report = {
        "phase": "TAER_FAILURE_SOURCE_ANALYSIS_V2",
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "verdict": a["verdict"]["verdict"],
        "verdict_detail": a["verdict"],
        "taer_v1_closeout": closeout,
        "precommit_sha256": a["precommit"]["precommit_sha256"],
        "locked_p1": LOCKED_P1,
        "locked_p2": LOCKED_P2,
        "episode_meta": a["ep_meta"],
        "clusters": a["cluster_summary"],
        "opportunity_summary": a["opportunity_summary"],
        "opportunity_judge": a["opportunity_judge"],
        "s7_audit": {
            **{k: v for k, v in a["s7_audit"].items() if k != "rows_sample"},
            "rows_sample_n": len(a["s7_audit"].get("rows_sample") or []),
        },
        "primary_meta": a["primary_meta"],
        "label_quality": a["label_quality"],
        "univariate": a.get("univariate"),
        "lodo": a.get("lodo"),
        "bootstrap": a.get("bootstrap"),
        "models": a.get("models"),
        "identity": a["identity"],
        "determinism": det,
        "tests": tests,
        "safety": {
            "submit": 0, "cancel": 0, "live": 0,
            "mainline_changed": False,
            "taer_v1_unchanged": True,
            "shadow": False, "forward": False, "paper": False, "discord": False,
            "unused_market_data": False,
            "new_family_created": False,
        },
        "stop": True,
        "note": "Review required before any new family Document/Plan/Precommit",
    }

    # store tables
    (store / "opportunity.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in a["opp_rows"]) + "\n",
        encoding="utf-8",
    )
    (store / "features.jsonl").write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in a["feat_rows"]) + "\n",
        encoding="utf-8",
    )
    (store / "s7_audit.json").write_text(json.dumps(a["s7_audit"], indent=2, default=str), encoding="utf-8")

    _publish(report, store, a)
    print("=== PUBLISHED ===", flush=True)
    print("run_id", run_id, flush=True)
    print("verdict", report["verdict"], flush=True)
    print("ab_ok", det["ab_ok"], flush=True)
    print("STOP", flush=True)
    return report


def _publish(report: dict, store: Path, payload=None) -> None:
    out = NATIVE / "results" / "research" / "e1_x6_taer_failure_source"
    out.mkdir(parents=True, exist_ok=True)
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    md = [
        f"# TAER Failure Source Analysis V2 — {report.get('verdict')}",
        "",
        f"- analysis_id: `{ANALYSIS_ID}`",
        f"- run_id: `{report.get('run_id')}`",
        f"- TAER V1 closeout: `{CANONICAL_RUN}` / `{CANONICAL_VERDICT}`",
        f"- ab_ok: {(report.get('determinism') or {}).get('ab_ok')}",
        f"- label_quality usable_frac: {(report.get('label_quality') or {}).get('usable_fraction')}",
        f"- opportunity_exists: {(report.get('opportunity_judge') or {}).get('any_setup_opportunity_exists')}",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Stop",
        "No TAER V2, no new family implementation, no Shadow/Forward/Paper/Discord.",
    ]
    (out / "report.md").write_text("\n".join(md), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    for row in (
        ("analysis_id", ANALYSIS_ID),
        ("verdict", report.get("verdict")),
        ("run_id", report.get("run_id")),
        ("canonical_run", CANONICAL_RUN),
    ):
        ws.append(list(row))

    def sheet(name, headers, rows):
        w = wb.create_sheet(name)
        w.append(headers)
        for r in rows:
            w.append([("" if v is None else v) for v in r])

    sheet("Precommit", ["key", "value"], [
        ["precommit_sha256", report.get("precommit_sha256")],
        ["locked_p1", report.get("locked_p1")],
        ["locked_p2", report.get("locked_p2")],
        ["primary_weighting", "CLUSTER_FIRST_EPISODE"],
        ["primary_outcome", "best_net_pnl_bps_300s"],
    ])
    sheet("Identity", ["key", "value"], [[k, v] for k, v in (report.get("identity") or {}).items()])
    cl = report.get("clusters") or {}
    sheet("EpisodeClusters", ["metric", "value"], [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
                                                   for k, v in cl.items()])

    # OpportunityEnvelope / Summary
    if payload:
        oe = wb.create_sheet("OpportunityEnvelope")
        oe.append(["episode_id", "setup", "day", "symbol", "best_300", "worst_300", "adverse", "ttp", "evaluable", "rep"])
        for r in (payload.get("opp_rows") or [])[:5000]:
            oe.append([
                r.get("episode_id"), r.get("setup_type"), r.get("day"), r.get("symbol"),
                r.get("best_net_pnl_bps_300s"), r.get("worst_net_pnl_bps_300s"),
                r.get("adverse_before_best_bps"), r.get("time_to_net_positive_sec"),
                r.get("evaluable"), r.get("is_cluster_representative"),
            ])
    else:
        sheet("OpportunityEnvelope", ["note"], [["no payload"]])

    os_rows = []
    for setup, blk in ((report.get("opportunity_summary") or {}).get("by_setup") or {}).items():
        o = blk.get("overall") or {}
        os_rows.append([setup, "ALL", o.get("episode_n"), o.get("best_net_pnl_bps_median"),
                        o.get("net_positive_rate"), o.get("net_plus_5bps_rate"), o.get("net_plus_10bps_rate")])
        for d, p in (blk.get("by_day") or {}).items():
            os_rows.append([setup, d, p.get("episode_n"), p.get("best_net_pnl_bps_median"),
                            p.get("net_positive_rate"), p.get("net_plus_5bps_rate"), p.get("net_plus_10bps_rate")])
    sheet("OpportunitySummary",
          ["setup", "day", "n", "median_best_300", "pos_rate", "plus5_rate", "plus10_rate"], os_rows)

    s7 = report.get("s7_audit") or {}
    sheet("S7Audit", ["reason", "n"], [[k, v] for k, v in (s7.get("by_reason") or {}).items()] or [["n", s7.get("n")]])

    if payload:
        ef = wb.create_sheet("EntryFeatures")
        headers = ["episode_id", "setup", "day", "symbol", "missing_feature_count", "spread_bps",
                   "volume_impulse_ratio", "pullback_depth_atr", "range_width_atr"]
        ef.append(headers)
        for r in (payload.get("feat_rows") or [])[:5000]:
            ef.append([r.get(h) for h in headers])
    else:
        sheet("EntryFeatures", ["note"], [["no payload"]])

    uni_rows = []
    for setup, feats in ((report.get("univariate") or {}).get("by_setup") or {}).items():
        for f in feats:
            uni_rows.append([setup, f.get("feature"), f.get("n"), f.get("spearman"),
                             f.get("median_split_effect_bps"), f.get("direction"), f.get("missing_rate")])
    sheet("FeatureUnivariate",
          ["setup", "feature", "n", "spearman", "effect", "direction", "missing_rate"], uni_rows)

    lodo_rows = []
    for setup, feats in (report.get("lodo") or {}).items():
        for f in feats:
            lodo_rows.append([setup, f.get("feature"), f.get("full_period_direction"),
                              f.get("same_direction_rate"), f.get("direction_reversal_count"),
                              f.get("evaluable_days"), f.get("stable_candidate")])
    sheet("FeatureLODO",
          ["setup", "feature", "full_dir", "same_dir_rate", "reversals", "evaluable_days", "stable"],
          lodo_rows)

    boot_rows = []
    for setup, feats in (report.get("bootstrap") or {}).items():
        for f in feats:
            boot_rows.append([setup, f.get("feature"), json.dumps(f.get("effect_ci95")),
                              f.get("ci_crosses_0"), f.get("strong_stable")])
    sheet("Bootstrap", ["setup", "feature", "ci95", "crosses_0", "strong_stable"], boot_rows)

    md_rows = []
    for setup, blk in ((report.get("models") or {}).get("by_setup") or {}).items():
        md_rows.append([setup, blk.get("skipped"), blk.get("median_auc"),
                        blk.get("auc_gt_0_55_days"), json.dumps(blk.get("gates"))])
    sheet("ModelDiagnostics", ["setup", "skipped", "median_auc", "auc_gt_055_days", "gates"], md_rows)

    sheet("DataQuality", ["metric", "value"], [[k, v] for k, v in (report.get("label_quality") or {}).items()])
    sheet("Verdict", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("verdict_detail") or {"verdict": report.get("verdict")}).items()])
    sheet("Tests", ["test", "outcome"],
          [[r["test"], r["outcome"]] for r in (report.get("tests") or {}).get("rows", [])])
    sheet("Determinism", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("determinism") or {}).items()])
    sheet("Safety", ["key", "value"], [[k, v] for k, v in (report.get("safety") or {}).items()])
    sheet("ChangeLog", ["item", "note"], [
        ["taer_v1", f"closed {CANONICAL_RUN} as CLOSED_NO_ROBUST_PAIR"],
        ["analysis", ANALYSIS_ID],
        ["no_new_family", "stop after analysis; review required"],
    ])
    wb.save(out / "audit.xlsx")

    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (store / "published_shas.json").write_text(json.dumps(report["published"], indent=2), encoding="utf-8")


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    run()
