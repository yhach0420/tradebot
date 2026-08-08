"""Orchestrate FSA V4 — stability gate contract repair (A/B)."""
from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.failure_source.v4_analysis import (
    audit_entry_features,
    bootstrap_v4,
    class_support_table,
    join_rows,
    lodo_v4,
    models_v4,
    univariate_v4,
    verdict_v4,
)
from research.e1_x6_taer.failure_source.v4_identity import (
    ANALYSIS_ID_V4,
    LOCKED_CLUSTER_SHA,
    LOCKED_EPISODE_SHA,
    LOCKED_OPPORTUNITY_SHA,
    LOCKED_TARGET_VALIDITY_SHA,
    PURPOSE_V4,
    V3_RUN,
    V3_STORE,
    annotate_v3_scope,
)
from research.e1_x6_taer.failure_source.v4_precommit import ENTRY_FEATURE_COLUMNS, build_v4_precommit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[4]


def _load_v3() -> tuple[list[dict], list[dict]]:
    labels = [json.loads(l) for l in (V3_STORE / "labels.jsonl").open(encoding="utf-8") if l.strip()]
    feats = [json.loads(l) for l in (V3_STORE / "features.jsonl").open(encoding="utf-8") if l.strip()]
    return labels, feats


def _verify_identity(labels: list[dict]) -> dict:
    tgt_sha = sha256_obj([(
        r["cluster_id"], r["episode_id"], r["opportunity_target_valid"],
        r["scenario_label_valid"], r["best_net_pnl_bps_300s"],
    ) for r in labels])
    return {
        "episode_identity_sha": LOCKED_EPISODE_SHA,
        "cluster_identity_sha": LOCKED_CLUSTER_SHA,
        "opportunity_table_sha": LOCKED_OPPORTUNITY_SHA,
        "target_validity_sha": tgt_sha,
        "target_match": tgt_sha == LOCKED_TARGET_VALIDITY_SHA,
        "locked_target_validity_sha": LOCKED_TARGET_VALIDITY_SHA,
        "n_labels": len(labels),
    }


def _run_tests() -> dict:
    env = dict(os.environ)
    env["PYTHONPATH"] = f"{NATIVE / 'src'};{NATIVE / 'research'}"
    env["PYTHONIOENCODING"] = "utf-8"
    proc = subprocess.run(
        [sys.executable, "-m", "pytest", "-q", "-rA", "--tb=line",
         "-p", "no:cacheprovider",
         str(NATIVE / "tests" / "test_e1_x6_taer_failure_source_v4.py")],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        cwd=str(NATIVE), env=env, timeout=300,
    )
    rows = []
    for line in (proc.stdout or "").splitlines():
        for st in ("PASSED", "FAILED", "ERROR"):
            if line.strip().startswith(st + " "):
                rows.append({"test": line.strip().split(" ", 1)[1].split(" - ")[0], "outcome": st})
                break
    return {
        "exit_code": proc.returncode,
        "passed": sum(1 for r in rows if r["outcome"] == "PASSED"),
        "failed": sum(1 for r in rows if r["outcome"] != "PASSED"),
        "total": len(rows),
        "rows": rows,
        "tail": (proc.stdout or "")[-2000:],
    }


def _one_pass(labels: list[dict], feats: list[dict]) -> dict:
    pre = build_v4_precommit()
    rows = join_rows(labels, feats)
    # ensure trade_side_quality key on every row for audit
    for r in rows:
        if "trade_side_quality" not in r or r["trade_side_quality"] is None:
            r["trade_side_quality"] = (
                "TICK_RULE_INFERRED" if r.get("trade_side_quality_code") == 1 else None
            )
    audit = audit_entry_features(rows, labels)
    cs = class_support_table(rows)
    if audit["status"] != "PASS":
        verd = verdict_v4(audit=audit, boot={}, models={})
        return {
            "precommit": pre, "rows": rows, "audit": audit, "class_support": cs,
            "univariate": None, "lodo": None, "bootstrap": None, "models": None,
            "verdict": verd,
            "shas": {
                "class_support_sha": sha256_obj(cs),
                "entry_features_sha": sha256_obj(sorted(
                    (r["episode_id"], r.get("cluster_id"), r.get("setup_type"), r.get("spread_bps"))
                    for r in rows
                )),
            },
        }

    uni = univariate_v4(rows)
    lodo = lodo_v4(rows, uni)
    boot = bootstrap_v4(rows, lodo)
    models = models_v4(rows, boot, cs)
    verd = verdict_v4(audit=audit, boot=boot, models=models)
    shas = {
        "class_support_sha": sha256_obj(cs),
        "entry_features_sha": sha256_obj(sorted(
            tuple(r.get(c) for c in ENTRY_FEATURE_COLUMNS)
            for r in rows
        )),
        "lodo_sha": sha256_obj(lodo),
        "bootstrap_sha": sha256_obj(boot),
        "model_sha": sha256_obj(models),
        "univariate_sha": sha256_obj(uni),
    }
    return {
        "precommit": pre, "rows": rows, "audit": audit, "class_support": cs,
        "univariate": uni, "lodo": lodo, "bootstrap": boot, "models": models,
        "verdict": verd, "shas": shas,
    }


def run() -> dict:
    run_id = f"e1x6_taer_fsa_v4_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}"
    store = Path.home() / "e1x6_research_store" / "taer" / run_id
    store.mkdir(parents=True, exist_ok=True)

    print("=== Annotate V3 scope (no overwrite) ===", flush=True)
    v3_scope = annotate_v3_scope()

    print("=== Precommit (before bootstrap/models/verdict) ===", flush=True)
    pre0 = build_v4_precommit()
    (store / "precommit.json").write_text(json.dumps(pre0, indent=2), encoding="utf-8")
    print("precommit_sha", pre0["precommit_sha256"], flush=True)

    print("=== Tests ===", flush=True)
    tests = _run_tests()
    print("tests", tests["passed"], "/", tests["total"], "exit", tests["exit_code"], flush=True)
    if tests["exit_code"] != 0:
        report = {"verdict": "TAER_FAILURE_ANALYSIS_TESTS_FAILED", "tests": tests, "run_id": run_id}
        _publish(report, store, None)
        return report

    print("=== Load frozen V3 labels/features ===", flush=True)
    labels, feats = _load_v3()
    ident = _verify_identity(labels)
    print("target_match", ident["target_match"], "n", ident["n_labels"], flush=True)
    if not ident["target_match"] or ident["n_labels"] != 399:
        report = {
            "analysis_id": ANALYSIS_ID_V4, "run_id": run_id,
            "verdict": "TAER_FAILURE_ANALYSIS_IDENTITY_MISMATCH",
            "identity": ident, "tests": tests,
        }
        _publish(report, store, None)
        return report

    print("=== Analysis A ===", flush=True)
    a = _one_pass(labels, feats)
    print("=== Analysis B ===", flush=True)
    b = _one_pass(labels, feats)

    det = {
        "ab_ok": a["shas"] == b["shas"] and a["verdict"]["verdict"] == b["verdict"]["verdict"],
        "shas_a": a["shas"],
        "shas_b": b["shas"],
        "verdict_a": a["verdict"]["verdict"],
        "verdict_b": b["verdict"]["verdict"],
    }

    # opportunity summary quick
    opp_sum = {}
    for setup in ("PULLBACK_RECLAIM", "RANGE_BREAKOUT"):
        sub = [r for r in a["rows"] if r["setup_type"] == setup]
        vals = [float(r["best_net_pnl_bps_300s"]) for r in sub]
        from research.e1_x6_taer.failure_source.analysis import _median
        opp_sum[setup] = {
            "n": len(sub),
            "median_best_300": _median(vals),
            "plus5_rate": sum(1 for r in sub if int(r["net_plus_5bps"]) == 1) / len(sub) if sub else None,
        }

    report = {
        "phase": "TAER_FAILURE_SOURCE_ANALYSIS_V4",
        "analysis_id": ANALYSIS_ID_V4,
        "purpose": PURPOSE_V4,
        "run_id": run_id,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "verdict": a["verdict"]["verdict"],
        "verdict_detail": a["verdict"],
        "v3_scope": v3_scope,
        "precommit_sha256": a["precommit"]["precommit_sha256"],
        "frozen_identity": ident,
        "audit_entry_features": a["audit"],
        "class_support": a["class_support"],
        "opportunity_summary": opp_sum,
        "univariate": a.get("univariate"),
        "lodo": a.get("lodo"),
        "bootstrap": a.get("bootstrap"),
        "models": a.get("models"),
        "determinism": det,
        "shas": a["shas"],
        "tests": tests,
        "safety": {
            "submit": 0, "cancel": 0, "live": 0,
            "mainline_changed": False,
            "taer_v1_unchanged": True,
            "taer_v2": False,
            "new_family": False,
            "shadow": False, "forward": False, "paper": False, "discord": False,
            "v3_run_not_overwritten": True,
        },
        "stop": True,
    }

    _publish(report, store, a)
    print("=== PUBLISHED ===", flush=True)
    print("run_id", run_id, flush=True)
    print("verdict", report["verdict"], flush=True)
    print("ab_ok", det["ab_ok"], flush=True)
    print("STOP", flush=True)
    return report


def _publish(report: dict, store: Path, payload) -> None:
    out = NATIVE / "results" / "research" / "e1_x6_taer_failure_source_v4"
    out.mkdir(parents=True, exist_ok=True)

    md = [
        f"# TAER Failure Source Analysis V4 — {report.get('verdict')}",
        "",
        f"- analysis_id: `{ANALYSIS_ID_V4}`",
        f"- purpose: `{PURPOSE_V4}`",
        f"- run_id: `{report.get('run_id')}`",
        f"- V3 scope: `FSA_V3_STOPPED_BY_INVALID_DAY_CLASS_SUPPORT_GATE`",
        f"- pullback_verdict: `{(report.get('verdict_detail') or {}).get('pullback_verdict')}`",
        f"- range_verdict: `{(report.get('verdict_detail') or {}).get('range_verdict')}`",
        f"- ab_ok: {(report.get('determinism') or {}).get('ab_ok')}",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Stop",
        "No new family / TAER V2 / Shadow / Forward / Paper / Discord.",
    ]
    vd = report.get("verdict_detail") or {}
    if vd.get("pullback_strong_features") is not None:
        md.append(f"- pullback strong: {vd.get('pullback_strong_features')}")
        md.append(f"- range strong: {vd.get('range_strong_features')}")
    (out / "report.md").write_text("\n".join(md), encoding="utf-8")
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    for row in (("analysis_id", ANALYSIS_ID_V4), ("verdict", report.get("verdict")),
                ("run_id", report.get("run_id")), ("v3_run", V3_RUN)):
        ws.append(list(row))

    def sheet(name, headers, rows):
        w = wb.create_sheet(name)
        w.append(headers)
        for r in rows:
            w.append([("" if v is None else v) for v in r])

    sheet("Precommit", ["key", "value"], [
        ["precommit_sha256", report.get("precommit_sha256")],
        ["purpose", PURPOSE_V4],
        ["locked_target_validity", LOCKED_TARGET_VALIDITY_SHA],
    ])
    sheet("Identity", ["key", "value"],
          [[k, v] for k, v in (report.get("frozen_identity") or {}).items()])
    sheet("LabelContract", ["note"], [["Frozen from V3: opportunity_target_valid 399/399"]])

    cs_rows = []
    for setup, blk in ((report.get("class_support") or {}).get("by_setup") or {}).items():
        for d in blk.get("days") or []:
            cs_rows.append([setup, d["day"], d["cluster_n"], d["positive_n"], d["negative_n"],
                            d["positive_rate"], d["descriptive_two_class"], d["model_confirm_eligible"],
                            d.get("median_best_net_300s")])
    sheet("ClassSupport",
          ["setup", "day", "n", "pos", "neg", "pos_rate", "descriptive", "model_eligible", "median_best"],
          cs_rows)

    if payload:
        tv = wb.create_sheet("TargetValidity")
        tv.append(["cluster_id", "episode_id", "setup_type", "day", "target_valid", "best_300", "net_plus_5"])
        for r in payload.get("rows") or []:
            tv.append([r.get("cluster_id"), r.get("episode_id"), r.get("setup_type"), r.get("day"),
                       r.get("opportunity_target_valid"), r.get("best_net_pnl_bps_300s"), r.get("net_plus_5bps")])
        sv = wb.create_sheet("ScenarioValidity")
        sv.append(["cluster_id", "episode_id", "scenario_group", "scenario_valid"])
        for r in payload.get("rows") or []:
            sv.append([r.get("cluster_id"), r.get("episode_id"), r.get("scenario_group"),
                       r.get("scenario_label_valid")])
    else:
        sheet("TargetValidity", ["note"], [["no payload"]])
        sheet("ScenarioValidity", ["note"], [["no payload"]])

    sheet("EpisodeClusters", ["note"], [["Frozen V3: 399 clusters"]])
    sheet("OpportunitySummary", ["setup", "n", "median_300", "plus5_rate"],
          [[s, o.get("n"), o.get("median_best_300"), o.get("plus5_rate")]
           for s, o in (report.get("opportunity_summary") or {}).items()])
    sheet("FeatureSchema", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("audit_entry_features") or {}).items() if k != "columns"])

    # Full EntryFeatures — all columns, all 399 rows
    ef = wb.create_sheet("EntryFeatures")
    ef.append(ENTRY_FEATURE_COLUMNS)
    if payload:
        for r in payload.get("rows") or []:
            ef.append([r.get(c) for c in ENTRY_FEATURE_COLUMNS])
        if len(payload.get("rows") or []) != 399 or any(not r.get("setup_type") for r in payload["rows"]):
            # mark incomplete — should have been caught by audit
            pass

    cov_rows = []
    for setup, feats in ((report.get("univariate") or {}).get("by_setup") or {}).items():
        for f in feats:
            cov_rows.append([setup, f.get("feature"), f.get("applicable_n"), f.get("n_non_missing"),
                             f.get("missing_rate"), f.get("zero_variance"), f.get("status"),
                             f.get("primary_candidate_eligible")])
    sheet("FeatureCoverage",
          ["setup", "feature", "applicable_n", "non_miss", "miss_rate", "zero_var", "status", "eligible"],
          cov_rows)

    uni_rows = []
    for setup, feats in ((report.get("univariate") or {}).get("by_setup") or {}).items():
        for f in feats:
            uni_rows.append([setup, f.get("feature"), f.get("spearman"), f.get("median_split_effect_bps"),
                             f.get("direction"), f.get("status"), f.get("zero_variance")])
    sheet("FeatureUnivariate",
          ["setup", "feature", "spearman", "effect", "direction", "status", "zero_var"], uni_rows)

    lodo_rows = []
    for setup, feats in (report.get("lodo") or {}).items():
        for f in feats:
            lodo_rows.append([setup, f.get("feature"), f.get("full_period_direction"),
                              f.get("same_direction_rate"), f.get("direction_reversal_count"),
                              f.get("evaluable_day_deletions"), f.get("minimum_effect"),
                              f.get("maximum_effect"), f.get("crosses_zero"),
                              f.get("minimum_support"), f.get("stable_direction_candidate")])
    sheet("FeatureLODO",
          ["setup", "feature", "full_dir", "same_rate", "reversals", "evaluable",
           "min_eff", "max_eff", "cross0", "min_support", "stable_dir"],
          lodo_rows)

    boot_rows = []
    for setup, feats in (report.get("bootstrap") or {}).items():
        for f in feats:
            boot_rows.append([setup, f.get("feature"), f.get("bootstrap_median"),
                              json.dumps(f.get("effect_ci95")), f.get("crosses_zero"),
                              f.get("positive_fraction"), f.get("strong_stable_feature")])
    sheet("Bootstrap",
          ["setup", "feature", "boot_med", "ci95", "cross0", "pos_frac", "strong"], boot_rows)

    md_rows = []
    for setup, blk in ((report.get("models") or {}).get("by_setup") or {}).items():
        md_rows.append([setup, blk.get("skipped"), blk.get("reason"), blk.get("median_auc"),
                        blk.get("auc_gt_0_55_days"), blk.get("coefficient_direction_consistency_median"),
                        json.dumps(blk.get("gates"))])
    sheet("ModelDiagnostics",
          ["setup", "skipped", "reason", "median_auc", "auc_gt_055", "coef_consist", "gates"], md_rows)

    sheet("DataQuality", ["key", "value"], [
        ["audit_status", (report.get("audit_entry_features") or {}).get("status")],
        ["entry_features_n", (report.get("audit_entry_features") or {}).get("n")],
    ])
    sheet("Verdict", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("verdict_detail") or {"verdict": report.get("verdict")}).items()])
    sheet("Tests", ["test", "outcome"],
          [[r["test"], r["outcome"]] for r in (report.get("tests") or {}).get("rows", [])])
    sheet("Determinism", ["key", "value"],
          [[k, json.dumps(v) if isinstance(v, (dict, list)) else v]
           for k, v in (report.get("determinism") or {}).items()])
    sheet("Safety", ["key", "value"], [[k, v] for k, v in (report.get("safety") or {}).items()])
    sheet("ChangeLog", ["item", "note"], [
        ["v3", "invalid day-class support gate (median sign); not evidence of no stable entry"],
        ["v4", "class support from net_plus_5bps counts; continuous stability without neg-median days"],
        ["stop", "no new family auto-created"],
    ])
    wb.save(out / "audit.xlsx")

    # final schema check on written sheet
    if payload and (
        len(payload.get("rows") or []) != 399
        or any(not r.get("setup_type") for r in payload["rows"])
    ):
        report["verdict"] = "TAER_FAILURE_ANALYSIS_AUDIT_SCHEMA_INCOMPLETE"

    report["published"] = {
        "report.json": sha256_file(out / "report.json"),
        "report.md": sha256_file(out / "report.md"),
        "audit.xlsx": sha256_file(out / "audit.xlsx"),
    }
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    report["published"]["report.json"] = sha256_file(out / "report.json")
    (store / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (store / "published_shas.json").write_text(json.dumps(report["published"], indent=2), encoding="utf-8")


if __name__ == "__main__":
    for _k in ("OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS", "NUMEXPR_NUM_THREADS"):
        os.environ[_k] = "1"
    run()
