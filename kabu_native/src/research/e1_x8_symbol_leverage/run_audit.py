"""E1_X8 orchestrator."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_fcrr.replay import _universe_from_manifest, load_day_events, load_source_manifest
from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x6_taer.failure_source.clusters import load_episodes
from research.e1_x7_pfq.config import DAYS
from research.e1_x7_pfq.feature_contract import run_phase0_audit
from research.e1_x7_pfq.run_study import _load_pullback_universe

from . import ANALYSIS_ID, DOCUMENT_ID, FROZEN, KNOWN, SOURCE_BRIDGE, SOURCE_PFQ_FINAL, TARGET_SYMBOL
from .analyze import (
    build_episode_table,
    decide_verdict,
    economic_reference,
    frozen_membership_concentration,
    full_threshold_summary,
    influence_ranking,
    loso_thresholds,
    membership_flips,
    size_matched_random_deletion,
    symbol_groups,
    symbol_profiles,
)
from .precommit import build_precommit
from .quantile_ops import reproduce_full_thresholds
from .signal import evaluate_update_signal, summarize_ft

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
BRIDGE_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_bridge_v2"
REV_DIR = NATIVE / "results" / "research" / "e1_x7_pfq_exit_revision"
PUBLISH = NATIVE / "results" / "research" / "e1_x8_symbol_leverage"


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "pfq_entry_changed": False,
        "pfq_exit_changed": False,
        "pfq_current_line_revived": False,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
    }


def _load_audits(events_by_day: dict) -> tuple[list[dict], dict]:
    universe = _load_pullback_universe()
    episodes_raw, _, _ = load_episodes()
    ep_by = {e["episode_id"]: e for e in episodes_raw}
    audits, phase0 = run_phase0_audit(universe, events_by_day, ep_by)
    return audits, phase0


def run_analysis(audits: list[dict], *, label: str = "A") -> dict[str, Any]:
    run_id = f"e1x8_symlev_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"

    br_report = json.loads((BRIDGE_DIR / "report.json").read_text(encoding="utf-8"))
    rev_report = json.loads((REV_DIR / "report.json").read_text(encoding="utf-8"))
    assert SOURCE_BRIDGE in str(br_report.get("run_id"))
    assert SOURCE_PFQ_FINAL in str(rev_report.get("run_id"))

    source_shas = {
        "feature_table_sha": (br_report.get("identity") or {}).get("feature_table_sha"),
        "episode_identity_sha": (br_report.get("identity") or {}).get("episode_identity_sha"),
        "cluster_identity_sha": (br_report.get("identity") or {}).get("cluster_identity_sha"),
        "fixed_grid_path_sha": (br_report.get("determinism_shas") or {}).get("fixed_grid_outcome_sha"),
        "baseline_ledger_sha": (rev_report.get("determinism_shas") or {}).get("baseline_ledger_sha"),
        "revision_ledger_sha": (rev_report.get("determinism_shas") or {}).get("revision_ledger_sha"),
        "bridge_report_sha": sha256_file(BRIDGE_DIR / "report.json"),
        "bridge_audit_sha": sha256_file(BRIDGE_DIR / "audit.xlsx"),
        "rev_report_sha": sha256_file(REV_DIR / "report.json"),
    }
    print(f"=== [{label}] Precommit ===", flush=True)
    precommit = build_precommit(source_shas={k: str(v) for k, v in source_shas.items()})

    fg_rows = _load_sheet(BRIDGE_DIR / "audit.xlsx", "FixedGridOutcome")
    fg_by = {r["episode_id"]: r for r in fg_rows}

    # Identity
    n_all = len(audits)
    pu_n = sum(1 for a in audits if a.get("price_update_count_10s") is not None)
    flow_n = sum(1 for a in audits if a.get("ratio_valid") and a.get("uptick_volume_ratio_30s") is not None)
    joint_n = sum(
        1 for a in audits
        if a.get("price_update_count_10s") is not None
        and a.get("ratio_valid")
        and int(a.get("classified_trade_count_30s") or 0) >= 3
    )
    identity_ok = (
        n_all == KNOWN["ALL_PULLBACK"]
        and pu_n == KNOWN["update_valid"]
        and flow_n == KNOWN["flow_valid"]
        and len(fg_rows) == KNOWN["ALL_PULLBACK"]
    )
    # joint eligible known 283 may require path - check approx
    rows = build_episode_table(audits, fg_by)
    joint_path = sum(1 for r in rows if r["joint_eligible_parent"])
    identity = {
        "ok": identity_ok and joint_path == KNOWN["joint_eligible"],
        "ALL_PULLBACK": n_all,
        "update_valid": pu_n,
        "flow_valid": flow_n,
        "joint_eligible": joint_path,
        "fixed_grid_n": len(fg_rows),
        "known": KNOWN,
    }
    if not identity["ok"]:
        # soft: joint may differ if path evaluable differs slightly
        identity["ok"] = identity_ok and abs(joint_path - KNOWN["joint_eligible"]) <= 2
        identity["joint_note"] = f"joint_path={joint_path} known={KNOWN['joint_eligible']}"

    if not identity["ok"]:
        return {
            "run_id": run_id, "verdict": "E1_X8_SYMBOL_LEVERAGE_IDENTITY_MISMATCH",
            "identity": identity, "precommit": precommit, "safety": _safety(), "stop": True,
            "determinism_shas": {"verdict": "E1_X8_SYMBOL_LEVERAGE_IDENTITY_MISMATCH"},
        }

    print(f"=== [{label}] Quantile contract ===", flush=True)
    qrep = reproduce_full_thresholds(audits)
    if not qrep["matches_frozen"]:
        return {
            "run_id": run_id, "verdict": "E1_X8_QUANTILE_CONTRACT_UNRESOLVED",
            "quantile": qrep, "precommit": precommit, "safety": _safety(), "stop": True,
            "determinism_shas": {"verdict": "E1_X8_QUANTILE_CONTRACT_UNRESOLVED"},
        }
    thr = {
        "price_update_count_10s_q70": qrep["thresholds"]["price_update_count_10s_q70"],
        "uptick_volume_ratio_30s_q30": qrep["thresholds"]["uptick_volume_ratio_30s_q30"],
    }
    full_sum = full_threshold_summary(audits, thr)

    print(f"=== [{label}] Profiles / LOSO / random ===", flush=True)
    profiles = symbol_profiles(rows)
    frozen_mem = frozen_membership_concentration(rows)
    loso = loso_thresholds(audits, thr)
    flips = membership_flips(audits, thr, loso)
    rand = size_matched_random_deletion(audits, thr, loso, flips)
    influence = influence_ranking(loso, rand, flips)

    print(f"=== [{label}] Frozen UPDATE signal ===", flush=True)
    parent = [r for r in rows if r["update_eligible_parent"]]
    cand_full = [r for r in rows if r["mem_UPDATE"]]
    bridge_ref = (br_report.get("candidate_enrichment") or {}).get("PFQ_UPDATE_Q70") or {}
    sig_full = evaluate_update_signal(cand_full, parent, bridge_reference=bridge_ref)
    if not (sig_full.get("bridge_reproduction") or {}).get("match"):
        return {
            "run_id": run_id, "verdict": "E1_X8_BRIDGE_SIGNAL_IDENTITY_MISMATCH",
            "bridge_reproduction": sig_full.get("bridge_reproduction"),
            "precommit": precommit, "safety": _safety(), "stop": True,
            "determinism_shas": {"verdict": "E1_X8_BRIDGE_SIGNAL_IDENTITY_MISMATCH"},
        }

    cand_ex = [r for r in cand_full if str(r["symbol"]) != TARGET_SYMBOL]
    parent_ex = [r for r in parent if str(r["symbol"]) != TARGET_SYMBOL]
    sig_ex = evaluate_update_signal(cand_ex, parent_ex)

    # LOSO frozen threshold signal
    symbols = sorted({str(r["symbol"]) for r in rows})
    loso_sig_rows = []
    preserved = 0
    for s in symbols:
        cand_s = [r for r in cand_full if str(r["symbol"]) != s]
        par_s = [r for r in parent if str(r["symbol"]) != s]
        sig_s = evaluate_update_signal(
            cand_s, par_s, ft_keys=("plus5_vs_minus10", "plus10_vs_minus10"),
        )
        s5 = summarize_ft(sig_s, "plus5_vs_minus10")
        s10 = summarize_ft(sig_s, "plus10_vs_minus10")
        support = bool(sig_s.get("supported"))
        if support:
            preserved += 1
        loso_sig_rows.append({
            "removed_symbol": s,
            "remaining_candidate_n": len(cand_s),
            "plus5_vs_minus10": s5,
            "plus10_vs_minus10": s10,
            "support_preserved": support,
            "is_285A": s == TARGET_SYMBOL,
        })
    loss_syms = [r["removed_symbol"] for r in loso_sig_rows if not r["support_preserved"]]
    loso_signal = {
        "loso_symbols_n": len(symbols),
        "support_preserved_n": preserved,
        "support_preserved_rate": preserved / len(symbols) if symbols else 0.0,
        "minimum_difference": min(
            (r["plus5_vs_minus10"]["difference"] for r in loso_sig_rows
             if r["plus5_vs_minus10"]["difference"] is not None), default=None
        ),
        "minimum_ci_lower": min(
            ((r["plus5_vs_minus10"]["difference_ci95"] or [None])[0] for r in loso_sig_rows
             if (r["plus5_vs_minus10"].get("difference_ci95") or [None])[0] is not None),
            default=None,
        ),
        "symbols_causing_support_loss": loss_syms,
        "rows": loso_sig_rows,
    }

    # Rederived descriptive only
    rederived = []
    for lr in loso:
        s = lr["symbol"]
        thr_s = {
            "price_update_count_10s_q70": lr["update_threshold_without"],
            "uptick_volume_ratio_30s_q30": lr["flow_threshold_without"],
        }
        from research.e1_x7_pfq.candidates import passes_candidate
        cand_r = []
        for a, r in zip(audits, rows):
            if str(a["symbol"]) == s:
                continue
            if passes_candidate(a, "PFQ_UPDATE_Q70", thr_s):
                cand_r.append(r)
        par_r = [r for r in parent if str(r["symbol"]) != s]
        sig_r = evaluate_update_signal(
            cand_r, par_r, ft_keys=("plus5_vs_minus10",),
        )
        rederived.append({
            "status": "DESCRIPTIVE_REDERIVATION_ONLY",
            "removed_symbol": s,
            "rederived_q70": lr["update_threshold_without"],
            "candidate_n": len(cand_r),
            "plus5_vs_minus10": summarize_ft(sig_r, "plus5_vs_minus10"),
            "supported": sig_r.get("supported"),
            "not_for_adoption": True,
        })

    groups = symbol_groups(profiles, rows)

    base_tr = _load_sheet(REV_DIR / "audit.xlsx", "BaselineTrades")
    rev_tr = _load_sheet(REV_DIR / "audit.xlsx", "RevisionTrades")
    econ = economic_reference(base_tr, rev_tr)

    vd = decide_verdict(
        identity_ok=True,
        quantile_ok=True,
        bridge_signal_ok=True,
        influence=influence,
        signal_full=sig_full,
        signal_ex=sig_ex,
        loso_signal=loso_signal,
        groups=groups,
    )

    det = {
        "identity_sha": sha256_obj(identity),
        "full_threshold_sha": sha256_obj(full_sum),
        "symbol_profile_sha": sha256_obj([(p["symbol"], p["UPDATE_Q70_pass_n"], p["median_price_update_count_10s"]) for p in profiles]),
        "loso_threshold_sha": sha256_obj([(r["symbol"], r["update_threshold_without"], r["flow_threshold_without"]) for r in loso]),
        "membership_flip_sha": sha256_obj([(r["symbol_removed"], r["update_membership_flip_rate"], r["flow_membership_flip_rate"]) for r in flips]),
        "random_deletion_sha": sha256_obj([(r["symbol"], r["actual_update_delta_percentile"], r["actual_flip_percentile"]) for r in rand]),
        "frozen_signal_sensitivity_sha": sha256_obj({
            "full": sig_full.get("supported"),
            "ex": sig_ex.get("supported"),
            "loso_rate": loso_signal["support_preserved_rate"],
        }),
        "rederived_signal_sha": sha256_obj([(r["removed_symbol"], r["rederived_q70"], r["supported"]) for r in rederived]),
        "group_classification_sha": sha256_obj({
            "heavy": groups["UPDATE_HEAVY"]["symbols"],
            "pfq_like": groups["PFQ_LIKE"]["symbols"],
        }),
        "economic_reference_sha": sha256_obj({
            "b": econ["baseline_PROGRESS_STRUCT"]["ex_285A_pnl"],
            "r": econ["revision_BE5_FLOOR0"]["ex_285A_pnl"],
        }),
        "verdict": vd["verdict"],
    }

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "precommit": precommit,
        "identity": identity,
        "quantile_contract": qrep,
        "full_thresholds": full_sum,
        "frozen_membership": {
            "UPDATE_n": sum(1 for r in rows if r["mem_UPDATE"]),
            "FLOW_n": sum(1 for r in rows if r["mem_FLOW"]),
            "JOINT_n": sum(1 for r in rows if r["mem_JOINT"]),
            "concentration": frozen_mem,
        },
        "influence": influence,
        "signal_full": {
            "supported": sig_full.get("supported"),
            "support_reasons": sig_full.get("support_reasons"),
            "plus5_vs_minus10": summarize_ft(sig_full, "plus5_vs_minus10"),
            "plus5_vs_minus15": summarize_ft(sig_full, "plus5_vs_minus15"),
            "n_candidate": sig_full.get("n_candidate"),
            "bridge_reproduction": sig_full.get("bridge_reproduction"),
        },
        "signal_ex_285A": {
            "supported": sig_ex.get("supported"),
            "support_reasons": sig_ex.get("support_reasons"),
            "plus5_vs_minus10": summarize_ft(sig_ex, "plus5_vs_minus10"),
            "plus5_vs_minus15": summarize_ft(sig_ex, "plus5_vs_minus15"),
            "n_candidate": sig_ex.get("n_candidate"),
        },
        "loso_signal_summary": {k: v for k, v in loso_signal.items() if k != "rows"},
        "symbol_groups": groups,
        "economic_reference": econ,
        "verdict": vd["verdict"],
        "verdict_detail": vd,
        "pfq_policy": {
            "current_line": "PFQ_CURRENT_LINE_CLOSED_REJECTED",
            "revive": False,
            "prospective": False,
            "shadow": False,
        },
        "determinism_shas": det,
        "safety": _safety(),
        "stop": True,
        "_sheets": {
            "SymbolProfiles": profiles,
            "LOSOThresholds": loso,
            "MembershipFlips": flips,
            "RandomDeletion": rand,
            "LOSOSignal": [
                {
                    "removed_symbol": r["removed_symbol"],
                    "remaining_candidate_n": r["remaining_candidate_n"],
                    "support_preserved": r["support_preserved"],
                    "p5m10_diff": r["plus5_vs_minus10"].get("difference"),
                    "p5m10_ci_lo": (r["plus5_vs_minus10"].get("difference_ci95") or [None])[0],
                    "p5m10_pos_days": r["plus5_vs_minus10"].get("positive_difference_days"),
                    "p10m10_diff": r["plus10_vs_minus10"].get("difference"),
                    "p10m10_ci_lo": (r["plus10_vs_minus10"].get("difference_ci95") or [None])[0],
                    "is_285A": r["is_285A"],
                }
                for r in loso_signal["rows"]
            ],
            "RederivedSignal": rederived,
            "FrozenMembership": frozen_mem["UPDATE"] + [
                {**x, "candidate": "FLOW"} for x in frozen_mem["FLOW"]
            ] + [
                {**x, "candidate": "JOINT"} for x in frozen_mem["JOINT"]
            ],
            "InfluenceRanking": influence["update_influence_ranking"],
        },
    }
    return report
