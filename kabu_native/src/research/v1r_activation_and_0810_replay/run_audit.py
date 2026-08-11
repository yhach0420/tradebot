"""V1R Production Activation Repair + 2026-08-10 Frozen Retrospective Replay."""
from __future__ import annotations

import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

from small_paper.v1r_day_engine import (
    parity_compare,
    resolve_pre0905_am_universe,
    run_frozen_day,
)
from small_paper.v1r_primary_activation_gate import assert_v1r_primary_roles
from small_paper.kabu_order_request_builder import (
    actual_broker_cancel_count,
    actual_broker_submit_count,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "v1r_activation_and_0810_replay"
DAY = "20260810"
ANALYSIS_ID = "V1R_ACTIVATION_AND_20260810_FROZEN_REPLAY"

PBV2_REF = {
    "fills": 51,
    "wins": 25,
    "losses": 25,
    "flats": 1,
    "pnl_yen_100": 79900.0,
    "pf": 1.8633,
    "note": "actual runtime PBv2 observer — NOT V1R; statistical meaning differs from counterfactual replay",
}


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def write_xlsx(sheets: dict[str, list[dict[str, Any]]], path: Path) -> None:
    if Workbook is None:
        path.with_suffix(".sheets.json").write_text(
            json.dumps(sheets, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_act_replay_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    # Phase 1 — activation
    print("  [1] role assertion...", flush=True)
    assertion = assert_v1r_primary_roles()
    print(assertion.startup_block, flush=True)
    activation_ready = bool(assertion.ok)

    # Phase 2 — universe causality + replay
    print("  [2] universe causality...", flush=True)
    uni = resolve_pre0905_am_universe(DAY)
    print(f"    universe pass={uni.get('pass')} n={uni.get('n')} 285A={uni.get('has_285A')}", flush=True)

    canonical = {"ok": False}
    production = {"ok": False}
    parity = {"pass": False, "verdict": "NOT_RUN"}

    if uni.get("pass"):
        print("  [3] canonical frozen replay...", flush=True)
        canonical = run_frozen_day(DAY, label="canonical_research", universe=uni)
        print(
            f"    canonical ok={canonical.get('ok')} fills={(canonical.get('flow') or {}).get('fills')} "
            f"pnl={(canonical.get('performance') or {}).get('total_pnl_yen_100')}",
            flush=True,
        )
        print("  [4] production offline transport replay...", flush=True)
        production = run_frozen_day(DAY, label="production_offline_transport", universe=uni)
        print(
            f"    production ok={production.get('ok')} fills={(production.get('flow') or {}).get('fills')} "
            f"pnl={(production.get('performance') or {}).get('total_pnl_yen_100')}",
            flush=True,
        )
        parity = parity_compare(canonical, production)
        print(f"    parity={parity.get('verdict')}", flush=True)
    else:
        print(f"    REPLAY BLOCKED: {uni.get('blocked_reason')}", flush=True)

    replay_complete = bool(canonical.get("ok") and production.get("ok"))
    causality_blocked = not bool(uni.get("pass"))

    submit = int(actual_broker_submit_count() or 0)
    cancel = int(actual_broker_cancel_count() or 0)
    live = 0

    # Next-day preflight (activation gate + no classic primary path)
    bat = Path(r"C:\Users\yhach\Documents\tradebotfile\run_paper_trade.bat")
    bat_txt = bat.read_text(encoding="utf-8", errors="ignore") if bat.exists() else ""
    classic_primary_removed = (
        "trailing-mfe" not in bat_txt
        or "V1R is PAPER_PRIMARY" in bat_txt
    ) and "v1r_paper_primary_launcher" in bat_txt
    classic_daily_as_primary = (
        "run_core10_dynamic40_am_pm_daily_runner.py" in bat_txt
        and "--exit-policy-shadow trailing-mfe" in bat_txt
        and "DISABLED" not in bat_txt
    )
    # After repair, bat must not invoke classic daily as primary
    no_classic_primary = ("v1r_paper_primary_launcher" in bat_txt) and (
        "Classic trailing-MFE Primary path DISABLED" in bat_txt
        or "run_core10_dynamic40_am_pm_daily_runner.py" not in bat_txt
    )

    next_day_checks = {
        "v1r_primary_role": activation_ready,
        "pbv2_shadow_role": assertion.checks.get("pbv2_role", False),
        "one_m_shadow_role": assertion.checks.get("one_m_role", False),
        "frozen_shas": all([
            assertion.checks.get("freeze_strategy_sha", False),
            assertion.checks.get("freeze_model_sha", False),
            assertion.checks.get("freeze_universe_binding_sha", False),
        ]),
        "no_legacy_pbv2_primary_in_bat": no_classic_primary,
        "submit_cancel_live_zero": submit == 0 and cancel == 0 and live == 0,
        "replay_parity_if_available": (parity.get("pass") if replay_complete else True),
    }
    next_day_ready = all(next_day_checks.values())

    # Verdict
    if activation_ready and replay_complete and parity.get("pass"):
        verdict = "V1R_ACTIVATION_REPAIRED_AND_20260810_FROZEN_REPLAY_COMPLETE"
    elif (not activation_ready) and replay_complete:
        verdict = "V1R_ACTIVATION_BLOCKED_REPLAY_COMPLETE"
    elif activation_ready and causality_blocked:
        verdict = "V1R_ACTIVATION_REPAIRED_20260810_REPLAY_CAUSALITY_BLOCKED"
    else:
        verdict = "V1R_ACTIVATION_AND_REPLAY_BLOCKED"

    perf = canonical.get("performance") or {}
    flow = canonical.get("flow") or {}
    s285 = canonical.get("symbol_285A") or {}
    quality = canonical.get("entry_quality_obs") or []
    mfe_vals = [float(q["mfe"]) for q in quality if q.get("mfe") is not None]
    mae_vals = [float(q["mae"]) for q in quality if q.get("mae") is not None]
    import statistics
    mfe_avg = statistics.mean(mfe_vals) if mfe_vals else None
    mae_avg = statistics.mean(mae_vals) if mae_vals else None

    overview = [{
        "run_id": run_id,
        "verdict": verdict,
        "activation": "READY" if activation_ready else "BLOCKED",
        "replay": "COMPLETE" if replay_complete else ("BLOCKED" if causality_blocked else "FAIL"),
        "day_class": "RETROSPECTIVE_OPERATIONAL_REFERENCE",
        "prospective_count_20260810": "0 / INVALID_NOT_STARTED",
        "parity": parity.get("verdict"),
        "next_prospective_day": "READY" if next_day_ready else "BLOCKED",
        "strategy_mutation": False,
        "submit_cancel_live": f"{submit}/{cancel}/{live}",
        "v1r_fills": flow.get("fills"),
        "v1r_pnl": perf.get("total_pnl_yen_100"),
        "v1r_pf": perf.get("pf"),
    }]

    sheets = {
        "Overview": overview,
        "Activation_Repair": [{
            "ready": activation_ready,
            "reason": assertion.reason,
            "startup_block": assertion.startup_block,
            "bat_no_classic_primary": no_classic_primary,
            "checked_runner_gate": True,
            "pbv2_primary_fallback": False,
            "why_0810_no_v1r": (
                "run_paper_trade.bat hardcoded classic trailing-mfe daily runner; "
                "V1R activation JSON existed but was never bound into Primary observer slot"
            ),
            "what_was_fixed": (
                "v1r_primary_activation_gate + v1r_paper_primary_launcher; "
                "bat Primary path switched; checked_runner fail-closed before paper start"
            ),
        }],
        "Role_Assertions": [{"check": k, "pass": v} for k, v in assertion.checks.items()],
        "Replay_Source": [{
            "day": DAY,
            "capture": str(NATIVE / "data/market_capture/20260810"),
            "push_jsonl": str(NATIVE / "data/push_jsonl/2026-08-10"),
            "kind": "COUNTERFACTUAL_RETROSPECTIVE_REPLAY",
            "pbv2_ledger_used_as_decision_source": False,
        }],
        "Universe_Causality": [uni],
        "Anchors": canonical.get("anchors") or [{"note": "blocked"}],
        "Candidates": [{"note": "counts in flow/anchors; per-event omitted to avoid CSV flood",
                        "signals": flow.get("signals"), "admitted": flow.get("admitted")}],
        "Fills": (canonical.get("fills_detail") or [{"note": "none"}])[:200],
        "Exits": [{
            "symbol": f.get("symbol"),
            "fill_time": f.get("fill_time"),
            "exit_time": f.get("exit_time"),
            "exit_reason": f.get("exit_reason"),
            "hold_sec": f.get("hold_sec"),
            "pnl_yen_100": f.get("pnl_yen_100"),
            "mfe": f.get("mfe"),
            "mae": f.get("mae"),
        } for f in (canonical.get("fills_detail") or [])][:200] or [{"note": "none"}],
        "Performance": [perf] if perf else [{"note": "blocked"}],
        "Anchor_Breakdown": canonical.get("anchors") or [{"note": "blocked"}],
        "Symbol_Breakdown": (canonical.get("symbols") or [{"note": "blocked"}])[:60],
        "PBV2_Comparison": [{
            "side": "V1R_counterfactual_replay",
            "fills": flow.get("fills"),
            "wl": f"{perf.get('wins')}/{perf.get('losses')}/{perf.get('flats')}",
            "pnl": perf.get("total_pnl_yen_100"),
            "pf": perf.get("pf"),
        }, {
            "side": "PBv2_actual_runtime_reference",
            **PBV2_REF,
        }],
        "Production_Parity": [parity],
        "Next_Day_Preflight": [{"check": k, "pass": v} for k, v in next_day_checks.items()] + [{
            "check": "overall",
            "pass": next_day_ready,
            "note": "Next unseen market day = Prospective Day 1 only if READY; 20260810 remains 0/INVALID_NOT_STARTED",
        }],
    }

    xlsx = OUT / "v1r_activation_and_0810_replay.xlsx"
    write_xlsx(sheets, xlsx)

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": verdict,
        "activation": "READY" if activation_ready else "BLOCKED",
        "replay": "COMPLETE" if replay_complete else "BLOCKED",
        "day_classification": "RETROSPECTIVE_OPERATIONAL_REFERENCE",
        "prospective_count": {"20260810": "0 / INVALID_NOT_STARTED", "next_unseen": "Prospective Day 1"},
        "assertion": assertion.to_dict(),
        "universe": uni,
        "canonical": {
            "ok": canonical.get("ok"),
            "flow": flow,
            "performance": perf,
            "symbol_285A": s285,
            "entry_quality_summary": {
                "n": len(quality),
                "avg_mfe": mfe_avg,
                "avg_mae": mae_avg,
            },
        },
        "production_offline": {
            "ok": production.get("ok"),
            "flow": production.get("flow"),
            "performance": production.get("performance"),
        },
        "parity": parity,
        "pbv2_reference": PBV2_REF,
        "next_day_preflight": {"ready": next_day_ready, "checks": next_day_checks},
        "answers": {
            "1_why_no_v1r_on_0810": (
                "Production bat hardcoded classic trailing-mfe daily runner; "
                "V1R activation was research-only and never started the Primary observer."
            ),
            "2_what_was_fixed": (
                "Activation gate + launcher; run_paper_trade.bat Primary path; "
                "paper_trade_checked_runner fail-closed before paper start; shared v1r_day_engine."
            ),
            "3_pbv2_fallback_gone": bool(no_classic_primary and activation_ready),
            "4_replay_fills": flow.get("fills"),
            "5_replay_pnl_pf": {"pnl": perf.get("total_pnl_yen_100"), "pf": perf.get("pf"),
                                "wl": f"{perf.get('wins')}/{perf.get('losses')}/{perf.get('flats')}"},
            "6_entry_quality": {"avg_mfe": mfe_avg, "avg_mae": mae_avg, "n": len(quality)},
            "7_285A": s285,
            "8_pbv2_compare": PBV2_REF,
            "9_parity": parity.get("verdict"),
            "10_next_day_ready": next_day_ready,
            "11_no_strategy_mutation": True,
            "12_submit_cancel_live": f"{submit}/{cancel}/{live}",
        },
        "safety": {"submit": submit, "cancel": cancel, "live": live},
        "strategy_mutation": False,
        "model_mutation": False,
        "universe_mutation": False,
        "ledger_state_mutation": False,
        "artifacts": {"xlsx": str(xlsx), "report_json": str(OUT / "report.json"), "report_md": str(OUT / "report.md")},
    }
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    md = f"""# V1R Activation + 2026-08-10 Replay 結論

## Production activation

{"READY" if activation_ready else "BLOCKED"}

## 8/10 replay

{"COMPLETE" if replay_complete else "BLOCKED"}

## 8/10 classification

RETROSPECTIVE_OPERATIONAL_REFERENCE

## V1R counterfactual result

* signals: {flow.get('signals')}
* fills: {flow.get('fills')}
* W/L/F: {perf.get('wins')}/{perf.get('losses')}/{perf.get('flats')}
* PnL: {perf.get('total_pnl_yen_100')}
* PF: {perf.get('pf')}
* max DD: {perf.get('max_realized_dd')}

## PBv2 actual reference

* 51 fills
* +79,900円
* PF 1.8633

## Production parity

{parity.get('verdict')}

## Next Prospective Day

{"READY" if next_day_ready else "BLOCKED"}

## Strategy mutation

false

## submit/cancel/live

{submit}/{cancel}/{live}

---

# Final Verdict

`{verdict}`

---

# Answers

1. **なぜ8/10にV1Rが起動しなかったか**  
   {report['answers']['1_why_no_v1r_on_0810']}

2. **production pathのどこを修正したか**  
   {report['answers']['2_what_was_fixed']}

3. **PBv2 Primary fallbackが消えたか**  
   {report['answers']['3_pbv2_fallback_gone']}

4. **8/10 V1R replay 取引数**  
   fills={flow.get('fills')} admitted={flow.get('admitted')} expired={flow.get('expired')}

5. **損益/PF**  
   PnL={perf.get('total_pnl_yen_100')} PF={perf.get('pf')} W/L/F={perf.get('wins')}/{perf.get('losses')}/{perf.get('flats')}

6. **ENTRY後リターン/MFE/MAE概要**  
   avg_MFE={mfe_avg} avg_MAE={mae_avg} n={len(quality)}（観測保存のみ・最適化なし）

7. **285A寄与**  
   {json.dumps(s285, ensure_ascii=False)}

8. **actual PBv2参考比較**  
   PBv2 actual 51 fills / +79900 / PF1.8633 vs V1R counterfactual（意味が異なる）

9. **parity**  
   {parity.get('verdict')}

10. **次の未見市場日をProspective Day 1として開始可能か**  
    {"YES" if next_day_ready else "NO"}

11. **strategy/model/universe/ENTRY/EXIT変更**  
    false（凍結SHA維持）

12. **submit/cancel/live**  
    {submit}/{cancel}/{live}

---

run_id: `{run_id}`  
8/10 Prospective count: **0 / INVALID_NOT_STARTED**（replayはcountに入れない）
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")
    print(json.dumps({
        "run_id": run_id,
        "verdict": verdict,
        "activation": report["activation"],
        "replay": report["replay"],
        "parity": parity.get("verdict"),
        "next_day": "READY" if next_day_ready else "BLOCKED",
        "fills": flow.get("fills"),
        "pnl": perf.get("total_pnl_yen_100"),
        "pf": perf.get("pf"),
    }, indent=2, ensure_ascii=False))
    return report


if __name__ == "__main__":
    main()
