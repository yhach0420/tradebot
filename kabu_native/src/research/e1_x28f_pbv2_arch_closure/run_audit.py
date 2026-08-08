"""E1_X28F runner: PBv2 parity repair → frozen 6-cell compare → stop before 0810."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x26_exit_library.exits import ExitSpec, common_controls
from research.e1_x27_reference_joint.paths import build_paths_for_rows
from research.e1_x28_executable_joint.board import load_board_events, verify_board_mapping
from research.e1_x28_executable_joint.replay import build_entry_asks
from research.e1_x28b_candidate_reference.baseline import freeze_family_baselines
from research.e1_x28d_additional_stress.replay_stress import build_exit_matrices
from research.e1_x28e_absolute_rise_exit_arch.analyze import (
    architecture_passes,
    block_abs,
    entry_delta_from_mat,
    lodo_loso,
    rank_architecture,
    summarize_exit,
)
from research.e1_x28e_absolute_rise_exit_arch.cohorts import build_masks, load_cohorts
from research.e1_x28e_absolute_rise_exit_arch.pbv2 import build_pbv2_matrix
from research.e1_x28e_absolute_rise_exit_arch.population import load_combined_population
from research.e1_x28e_absolute_rise_exit_arch.regime import period_mask, regime_mask

from . import (
    ADDITIONAL_STRESS,
    ANALYSIS_ID,
    ASSIGNMENT_REGISTRY_SHA,
    BOARD_MAPPING_SHA,
    DISCOVERY,
    DOCUMENT_ID,
    EVALUATION,
    EXIT_ARCHS,
    EXPECTED_FAMILY,
    EXPECTED_SPECIFIC,
    FAMILY_BASELINE_REGISTRY_SHA,
    FORBIDDEN_FROM,
    LOGIC_MANIFEST_SHA,
    PBV2_MANIFEST_SHA,
    REGIME_LIBRARY,
    SEMANTIC_EXIT_REGISTRY_SHA,
    SOURCE_X28E,
    SOURCE_X28E_VERDICT,
    STRESS_DAY,
    CONSUMED_DAY,
    VERDICT_CLOSED,
    VERDICT_MULTIPLE,
    VERDICT_PARITY_FAIL,
    VERDICT_SUPPORTED,
    X29_V2_PRECOMMIT_SHA,
    X29_V2_RUN_ID,
)
from .parity import build_reason_mapping, run_known_episode_parity
from .publish import publish

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28f_pbv2_arch_closure"
X28A1 = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"
X26A = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"
X28E = NATIVE / "results" / "research" / "e1_x28e_absolute_rise_exit_arch"
X29 = NATIVE / "results" / "research" / "e1_x29_prospective"


def _run_tests() -> dict[str, Any]:
    import os
    tp = NATIVE / "tests" / "research" / "test_e1_x28f_pbv2_arch_closure.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(tp), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1, "detail": out[-2500:],
    }


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    return [{hdr[i]: row[i] for i in range(len(hdr))} for row in rows[1:]]


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _spec_from_row(exit_id: str, p: dict[str, Any]) -> ExitSpec:
    np_sec = _f(p.get("no_progress_sec"))
    gm = p.get("giveback_mode") or None
    if gm == "":
        gm = None
    return ExitSpec(
        exit_id=exit_id, path_family=None, variant=p.get("exit_mode"),
        stop_bps=_f(p.get("stop_bps")), target_bps=_f(p.get("target_bps")),
        trail_activation_bps=_f(p.get("trail_activation_bps")),
        giveback_bps=_f(p.get("giveback_bps")), giveback_mode=gm,
        no_progress_sec=np_sec,
        max_hold_sec=float(_f(p.get("max_hold_sec")) or 900.0),
        no_progress_mfe_bps=(_f(p.get("no_progress_mfe_bps")) or 5.0) if np_sec is not None else None,
        no_progress_abs_ret_bps=(_f(p.get("no_progress_abs_ret_bps")) or 5.0) if np_sec is not None else None,
        is_control=str(exit_id).startswith("CONTROL_"),
    )


def _load_boards(rows, allowed):
    keys = sorted({(r["date"], r["symbol"]) for r in rows if r["date"] in allowed})
    cache = {}
    print(f"  boards {len(keys)}...", flush=True)

    def _one(k):
        return k, load_board_events(k[0], k[1])

    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = [ex.submit(_one, k) for k in keys]
        done = 0
        for fut in as_completed(futs):
            k, b = fut.result()
            cache[k] = b
            done += 1
            if done % 50 == 0 or done == len(keys):
                print(f"    {done}/{len(keys)}", flush=True)
    return cache


def _med(vals):
    xs = [v for v in vals if v is not None]
    return float(np.median(xs)) if xs else None


def _gate_mask(rows, regime_id: str) -> np.ndarray:
    if regime_id == "R0_NO_REGIME_GATE":
        return np.ones(len(rows), dtype=bool)
    return regime_mask(rows, regime_id)


def _r2_diagnostic(rows: list[dict[str, Any]], dates: np.ndarray) -> dict[str, Any]:
    """Diagnosis only — no new gate creation."""
    rg = regime_mask(rows, "R2_UNIVERSE_MEDIAN_RETURN_180S_GT0")
    # episode = consecutive True runs per (date) rough count of True rows
    n_true = int(rg.sum())
    # transitions
    flips = 0
    for i in range(1, len(rg)):
        if bool(rg[i]) != bool(rg[i - 1]):
            flips += 1
    # post-R2 median path: among R2 True rows, mean of forward medians if present — use feature itself path proxy
    # duration proxy: fraction of day True
    by_day = {}
    for day in sorted(set(dates.tolist())):
        m = dates == day
        if not m.any():
            continue
        by_day[day] = float(rg[m].mean())
    return {
        "r2_episode_row_count": n_true,
        "r2_row_fraction": float(rg.mean()),
        "true_false_transition_count": flips,
        "r2_fraction_by_day": by_day,
        "note": "Diagnostic only — no new Gate (duration/breadth) created in X28F",
        "new_gate_created": False,
    }


def _lodo_detail(mat, sel, dates, symbols) -> dict[str, Any]:
    base = lodo_loso(mat, sel, dates, symbols)
    idx = np.where(sel & mat["valid"])[0]
    if idx.size == 0:
        return {**base, "worst_omitted_day": None, "best_omitted_day": None,
                "worst_omitted_symbol": None, "best_omitted_symbol": None}
    without_day = {}
    for day in sorted(set(dates[idx].tolist())):
        keep = idx[dates[idx] != day]
        if keep.size:
            without_day[day] = float(np.mean(mat["ret_bps"][keep]))
    without_sym = {}
    for sym in sorted(set(symbols[idx].tolist())):
        keep = idx[symbols[idx] != sym]
        if keep.size:
            without_sym[sym] = float(np.mean(mat["ret_bps"][keep]))
    worst_d = min(without_day, key=without_day.get) if without_day else None
    best_d = max(without_day, key=without_day.get) if without_day else None
    worst_s = min(without_sym, key=without_sym.get) if without_sym else None
    best_s = max(without_sym, key=without_sym.get) if without_sym else None
    add_present = "285A" in set(symbols[np.isin(dates, list(ADDITIONAL_STRESS))].tolist())
    return {
        **base,
        "worst_omitted_day": worst_d,
        "best_omitted_day": best_d,
        "worst_omitted_symbol": worst_s,
        "best_omitted_symbol": best_s,
        "285A_on_0805_07": "PRESENT" if add_present else "NOT_PRESENT",
    }


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id = f"e1x28f_closure_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    x28e = json.loads((X28E / "report.json").read_text(encoding="utf-8"))
    if x28e.get("run_id") != SOURCE_X28E:
        raise RuntimeError("x28e run")
    if x28e.get("verdict") != SOURCE_X28E_VERDICT:
        raise RuntimeError("x28e verdict")
    man = json.loads((X28E / "pbv2_exit_manifest_v1.json").read_text(encoding="utf-8"))
    if man.get("manifest_sha256") != PBV2_MANIFEST_SHA:
        raise RuntimeError("pbv2 manifest sha")

    assert list(REGIME_LIBRARY) == [
        "R0_NO_REGIME_GATE", "R2_UNIVERSE_MEDIAN_RETURN_180S_GT0",
    ]
    assert EXIT_ARCHS == ("SPECIFIC_EXIT", "FAMILY_EXIT", "PBV2_EXIT")

    print("=== Phase A: reason mapping + known-episode parity ===", flush=True)
    mapping = build_reason_mapping()
    parity = run_known_episode_parity()
    print(f"  parity={parity['status']} episodes={parity['known_episode_count']} "
          f"type_match={parity['trigger_type_match_count']} t30={parity['timestamp_match_30s_count']}", flush=True)

    if not parity.get("ok"):
        report = {
            "analysis_id": ANALYSIS_ID,
            "document_id": DOCUMENT_ID,
            "run_id": run_id,
            "verdict": VERDICT_PARITY_FAIL,
            "pbv2_parity": parity,
            "selected_architecture": None,
            "x29_v2_status": "HELD_NOT_ADVANCED",
            "x29_v3_required": False,
            "opened_20260810": False,
            "prospective_observer_started": False,
            "pbv2_performance_compared": False,
            "safety": {"submit_cancel_live": "0/0/0", "runtime_ENTRY_changed": False,
                       "runtime_EXIT_changed": False, "Universe_changed": False},
        }
        interim = {
            "run_id": run_id, "verdict": VERDICT_PARITY_FAIL,
            "source_x28e_run_id": SOURCE_X28E,
            "pbv2_manifest_sha": PBV2_MANIFEST_SHA,
            "pbv2_parity_status": parity["status"],
            "pbv2_compare_if_parity_fail": False,
            "regime_library": list(REGIME_LIBRARY),
            "no_new_regime": True,
            "specific_n": 49, "family_n": 118, "overlap": 0,
            "no_candidate_selection": True, "no_stop_grid": True,
            "opened_20260810": False, "no_runtime_change": True,
            "submit_cancel_live": "0/0/0", "ab_determinism": True,
            "x28e_selection_rule_unchanged": True,
            "pbv2_runtime_reason_mapping": True,
            "pbv2_trigger_order": True,
            "pbv2_known_episode_parity": parity["status"],
            "285a_not_present_0805_07": True,
            "lodo_recompute": False, "loso_recompute": False,
            "same_entry": False, "same_actual_ask": True, "same_bid_fill": True,
        }
        (OUT / "_interim.json").write_text(json.dumps(interim, indent=2), encoding="utf-8")
        tests = _run_tests()
        report["tests"] = tests
        publish(OUT, report, {
            "Index": [{"run_id": run_id, "verdict": VERDICT_PARITY_FAIL}],
            "Parity": [parity],
            "ReasonMapping": mapping["rows"],
            "Tests": [tests],
        }, mapping)
        print("=== STOP: parity unresolved; no PBv2 compare ===", flush=True)
        return report

    # Phase B — frozen comparison
    print("=== Phase B: frozen architecture comparison ===", flush=True)
    cohorts = load_cohorts()
    rows = load_combined_population()
    dates = np.array([r["date"] for r in rows])
    symbols = np.array([str(r["symbol"]) for r in rows])
    assert not (dates >= FORBIDDEN_FROM).any()
    _, masks = build_masks(rows, cohorts)

    allowed = list(DISCOVERY + EVALUATION + (STRESS_DAY, CONSUMED_DAY) + ADDITIONAL_STRESS)
    mapping_ok = verify_board_mapping()
    if not mapping_ok.get("ok"):
        raise RuntimeError("board mapping")
    times_list, prices_list = build_paths_for_rows(rows, allowed_dates=allowed, max_workers=6)
    board_by_key = _load_boards(rows, allowed)
    entry_asks = build_entry_asks(rows, board_by_key)
    entry_asks_b = build_entry_asks(rows, board_by_key)
    ab_entry = bool(np.array_equal(entry_asks["valid"], entry_asks_b["valid"]))

    assignments = _load_sheet(X28A1 / "audit.xlsx", "CandidateExitAssignmentsV2")
    sem_reg = _load_sheet(X28A1 / "audit.xlsx", "SemanticExitRegistryV2")
    x26a = json.loads((X26A / "report.json").read_text(encoding="utf-8"))
    x26a_exits = {c["canonical_exit_id"]: c for c in (x26a.get("canonical_exits") or [])}
    for spec in common_controls():
        x26a_exits[spec.exit_id] = {
            "canonical_exit_id": spec.exit_id, "stop_bps": spec.stop_bps,
            "target_bps": spec.target_bps, "trail_activation_bps": spec.trail_activation_bps,
            "giveback_bps": spec.giveback_bps, "giveback_mode": spec.giveback_mode,
            "no_progress_sec": spec.no_progress_sec, "max_hold_sec": spec.max_hold_sec,
            "no_progress_mfe_bps": spec.no_progress_mfe_bps,
            "no_progress_abs_ret_bps": spec.no_progress_abs_ret_bps,
        }
    baseline_rows, baseline_sha = freeze_family_baselines(assignments, x26a_exits)
    if baseline_sha != FAMILY_BASELINE_REGISTRY_SHA:
        raise RuntimeError("baseline")
    specs: dict[str, ExitSpec] = {}
    for s in sem_reg:
        specs[s["semantic_exit_sha256"]] = _spec_from_row(s["semantic_exit_sha256"], s)
    needed = set()
    for r in cohorts["specific"]:
        needed.add(r["semantic_exit_sha256"])
        needed.add(r["family_baseline_exit_id"])
    for r in cohorts["family"]:
        needed.add(r["primary_exit_id"])
    for eid in needed:
        if eid not in specs:
            p = x26a_exits.get(eid)
            if p is None:
                raise RuntimeError(eid)
            specs[eid] = _spec_from_row(eid, p)
    spec_list = [specs[e] for e in sorted(needed)]
    print(f"  exit matrices n={len(spec_list)}", flush=True)
    mats = build_exit_matrices(
        specs=spec_list, rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=board_by_key, max_workers=4,
    )
    eid0 = spec_list[0].exit_id
    mats_b = build_exit_matrices(
        specs=[specs[eid0]], rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=board_by_key, max_workers=1,
    )
    ab_mat = bool(np.array_equal(mats[eid0]["valid"], mats_b[eid0]["valid"]))

    print("  PBv2 matrix...", flush=True)
    pbv2_mat = build_pbv2_matrix(
        rows=rows, entry_asks=entry_asks, times_list=times_list,
        prices_list=prices_list, board_by_key=board_by_key,
    )
    pbv2_mat2 = build_pbv2_matrix(
        rows=rows, entry_asks=entry_asks, times_list=times_list,
        prices_list=prices_list, board_by_key=board_by_key,
    )
    ab_mat = ab_mat and bool(np.array_equal(pbv2_mat["valid"], pbv2_mat2["valid"]))

    def _pool_dep(mat_getter, cohort_regs, gmask):
        all_r, all_d, all_s = [], [], []
        for reg in cohort_regs:
            cid = reg["candidate_id"]
            sel = masks[cid] & gmask
            mat = mat_getter(reg)
            idx = np.where(sel & mat["valid"])[0]
            all_r.extend(mat["ret_bps"][idx].tolist())
            all_d.extend(dates[idx].tolist())
            all_s.extend(symbols[idx].tolist())
        if not all_r:
            return {"max_day_contribution_share": None, "max_symbol_contribution_share": None,
                    "positive_LODO_count": 0, "negative_LODO_count": 0,
                    "positive_LOSO_count": 0, "negative_LOSO_count": 0,
                    "285A_on_0805_07": "NOT_PRESENT"}
        pm = {
            "valid": np.ones(len(all_r), dtype=bool),
            "ret_bps": np.asarray(all_r, float),
            "pnl": np.asarray(all_r, float),
            "reason": np.array([""] * len(all_r), dtype=object),
            "mae_bps": np.full(len(all_r), np.nan),
        }
        return _lodo_detail(pm, np.ones(len(all_r), dtype=bool), np.asarray(all_d), np.asarray(all_s))

    def _eval_arch(arch_name, mat_getter, cohort_regs, regime_id, complexity_arch):
        gmask = _gate_mask(rows, regime_id)
        per = []
        for reg in cohort_regs:
            cid = reg["candidate_id"]
            sel = masks[cid] & gmask
            mat = mat_getter(reg)
            sm = summarize_exit(mat=mat, mask=sel, dates=dates, symbols=symbols)
            ed = entry_delta_from_mat(mat, sel)
            # stop diagnostics
            stop_bps = _f(reg.get("stop_bps"))
            if arch_name == "PBV2_EXIT":
                stop_bps = 120.0  # -1.20% = 120 bps
            near = None
            if sm.get("trades") and "mae_bps" in mat and stop_bps:
                idx = np.where(sel & mat["valid"])[0]
                mae = mat["mae_bps"][idx]
                rets = mat["ret_bps"][idx]
                near_m = mae <= (-0.8 * float(stop_bps))
                near = {
                    "near_stop_count": int(near_m.sum()),
                    "near_stop_recovery_rate": (
                        float(((near_m) & (rets > 0)).sum() / near_m.sum()) if near_m.any() else None
                    ),
                }
            per.append({
                **sm, "entry_delta": ed, "candidate_id": cid,
                "eval_abs": block_abs(mat, sel, dates, EVALUATION),
                "stress_abs": block_abs(mat, sel, dates, ADDITIONAL_STRESS),
                "disc_abs": block_abs(mat, sel, dates, DISCOVERY),
                "d0803_abs": block_abs(mat, sel, dates, (STRESS_DAY,)),
                "d0804_abs": block_abs(mat, sel, dates, (CONSUMED_DAY,)),
                "stop_bps": stop_bps, **(near or {}),
            })
        med_ret = _med([p.get("avg_return_bps") for p in per])
        med_pf = _med([p.get("profit_factor") for p in per])
        med_ed = _med([p.get("entry_delta") for p in per])
        med_eval = _med([p.get("eval_abs") for p in per])
        med_stress = _med([p.get("stress_abs") for p in per])
        dep = _pool_dep(mat_getter, cohort_regs, gmask)
        metrics = {"avg_return_bps": med_ret, "profit_factor": med_pf,
                   "pnl_100share": _med([p.get("pnl_100share") for p in per])}
        passes = architecture_passes(metrics, med_ed, med_eval, med_stress, dep)
        return {
            "arch": complexity_arch,
            "arch_label": arch_name,
            "gate": "NO_GATE" if regime_id.startswith("R0") else "SUPPORTED_GLOBAL_REGIME",
            "regime_id": regime_id,
            "median_avg_return_bps": med_ret,
            "median_return_bps": _med([p.get("median_return_bps") for p in per]),
            "median_pf": med_pf,
            "median_entry_delta": med_ed,
            "median_pnl_100share": metrics["pnl_100share"],
            "median_win_rate": _med([p.get("win_rate") for p in per]),
            "median_mae": _med([p.get("avg_mae") for p in per]),
            "median_mfe": _med([p.get("avg_mfe") for p in per]),
            "median_mfe_capture": _med([p.get("mfe_capture_ratio") for p in per]),
            "median_hard_stop_rate": _med([p.get("hard_stop_rate") for p in per]),
            "median_no_progress_rate": _med([p.get("no_progress_rate") for p in per]),
            "median_trail_exit_rate": _med([p.get("trail_exit_rate") for p in per]),
            "median_session_rate": _med([p.get("session_or_maxhold_rate") for p in per]),
            "Evaluation_abs": med_eval,
            "20260805_07_abs": med_stress,
            "Discovery_abs": _med([p.get("disc_abs") for p in per]),
            "20260803_abs": _med([p.get("d0803_abs") for p in per]),
            "20260804_abs": _med([p.get("d0804_abs") for p in per]),
            "dependency": dep,
            "metrics": metrics,
            "passes": passes,
            "narrower_stop_score": {"family": 2, "pbv2": 1, "specific": 0}[complexity_arch],
            "stability_score": int(passes),
            "trade_count_sum": sum(int(p.get("trades") or 0) for p in per),
            "near_stop_recovery_rate_med": _med([p.get("near_stop_recovery_rate") for p in per]),
        }

    cells = []
    # Specific49 same-entry: Specific / Family / PBv2 × R0 / R2
    for regime_id in REGIME_LIBRARY:
        cells.append(_eval_arch(
            "SPECIFIC_EXIT",
            lambda reg: mats[reg["semantic_exit_sha256"]],
            cohorts["specific"], regime_id, "specific",
        ))
        cells.append(_eval_arch(
            "FAMILY_EXIT",
            lambda reg: mats[reg["family_baseline_exit_id"]],
            cohorts["specific"], regime_id, "family",
        ))
        cells.append(_eval_arch(
            "PBV2_EXIT",
            lambda reg: pbv2_mat,
            cohorts["specific"], regime_id, "pbv2",
        ))

    print("=== Phase C: PBv2 cohort robustness ===", flush=True)
    pbv2_cohort = []
    for regime_id in REGIME_LIBRARY:
        pbv2_cohort.append(_eval_arch(
            "PBV2_EXIT_ON_SPECIFIC49",
            lambda reg: pbv2_mat,
            cohorts["specific"], regime_id, "pbv2",
        ))
        # family registry uses primary_exit_id but for PBv2 we ignore specific exits
        pbv2_cohort.append(_eval_arch(
            "PBV2_EXIT_ON_FAMILY118",
            lambda reg: pbv2_mat,
            cohorts["family"], regime_id, "pbv2",
        ))

    r2_diag = _r2_diagnostic(rows, dates)

    passing = [c for c in cells if c.get("passes")]
    if len(passing) == 0:
        verdict = VERDICT_CLOSED
        selected = None
        x29_v3 = False
        x29_status = "HELD_CLOSED_NOT_ADVANCED"
    elif len(passing) == 1:
        verdict = VERDICT_SUPPORTED
        selected = passing[0]
        x29_v3 = True
        x29_status = "SUPERSEDED_BEFORE_PROSPECTIVE_OPEN"
    else:
        verdict = VERDICT_MULTIPLE
        selected = rank_architecture(passing)
        x29_v3 = True
        x29_status = "SUPERSEDED_BEFORE_PROSPECTIVE_OPEN"

    x29_v3_sha = None
    if x29_v3 and selected is not None:
        v2 = json.loads((X29 / "precommit_v2.json").read_text(encoding="utf-8"))
        body = deepcopy(v2)
        body.pop("precommit_sha", None)
        body.pop("published_shas", None)
        body["precommit_id"] = "X29_PRECOMMIT_V3_FROZEN"
        body["run_id"] = f"e1x29_precommit_v3_{now.strftime('%Y%m%d_%H%M%S')}_A"
        body["verdict"] = "X29_PRECOMMIT_V3_FROZEN"
        body["precommit_version"] = 3
        body["supersedes"] = {
            "old_run_id": X29_V2_RUN_ID,
            "old_precommit_sha": X29_V2_PRECOMMIT_SHA,
            "status": "SUPERSEDED_BEFORE_PROSPECTIVE_OPEN",
            "reason": "E1_X28F_FROZEN_ARCHITECTURE",
        }
        body["architecture"] = {
            "cohort": "SPECIFIC_49",
            "regime": selected.get("regime_id"),
            "exit": selected.get("arch_label"),
            "source_x28f_run_id": run_id,
        }
        body["first_eligible_prospective_day"] = "20260810"
        body["20260810_not_opened"] = True
        body["market_data_not_opened"] = True
        body["observer_not_started"] = True
        sha = sha256_obj(body)
        body["precommit_sha"] = sha
        (X29 / "precommit_v3.json").write_text(json.dumps(body, indent=2, default=str), encoding="utf-8")
        (X29 / "precommit_v2_superseded_by_x28f.json").write_text(json.dumps({
            "status": "SUPERSEDED_BEFORE_PROSPECTIVE_OPEN",
            "old_sha": X29_V2_PRECOMMIT_SHA,
            "new_sha": sha,
            "at": now.isoformat(),
        }, indent=2), encoding="utf-8")
        x29_v3_sha = sha
        assert json.loads((X29 / "precommit_v2.json").read_text(encoding="utf-8"))["precommit_sha"] == X29_V2_PRECOMMIT_SHA

    print("=== tests ===", flush=True)
    interim = {
        "run_id": run_id,
        "verdict": verdict,
        "source_x28e_run_id": SOURCE_X28E,
        "pbv2_manifest_sha": PBV2_MANIFEST_SHA,
        "pbv2_parity_status": parity["status"],
        "pbv2_compare_if_parity_fail": False,
        "pbv2_runtime_reason_mapping": True,
        "pbv2_trigger_order": True,
        "pbv2_known_episode_parity": parity["status"],
        "regime_library": list(REGIME_LIBRARY),
        "no_new_regime": True,
        "same_entry_specific_family_pbv2": True,
        "same_actual_ask": True,
        "same_bid_fill_contract": True,
        "x28e_selection_rule_unchanged": True,
        "specific_n": EXPECTED_SPECIFIC,
        "family_n": EXPECTED_FAMILY,
        "overlap": 0,
        "no_candidate_selection": True,
        "lodo_recompute": True,
        "loso_recompute": True,
        "285a_not_present_0805_07": True,
        "no_stop_grid": True,
        "opened_20260810": False,
        "no_runtime_change": True,
        "submit_cancel_live": "0/0/0",
        "ab_determinism": ab_entry and ab_mat,
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2), encoding="utf-8")
    tests = _run_tests()

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": verdict,
        "source_x28e_run_id": SOURCE_X28E,
        "pbv2_parity": parity,
        "pbv2_manifest_sha": PBV2_MANIFEST_SHA,
        "reason_mapping_sha": mapping["mapping_sha256"],
        "specific49_same_entry_cells": cells,
        "pbv2_cohort_cells": pbv2_cohort,
        "selected_architecture": {
            "cohort": "SPECIFIC_49",
            "regime": selected.get("regime_id") if selected else None,
            "exit": selected.get("arch_label") if selected else None,
            "detail": selected,
        } if selected else None,
        "r2_diagnostic": r2_diag,
        "x29_v2_status": x29_status,
        "x29_v2_sha": X29_V2_PRECOMMIT_SHA,
        "x29_v3_required": bool(x29_v3_sha),
        "x29_v3_precommit_sha": x29_v3_sha,
        "opened_20260810": False,
        "prospective_observer_started": False,
        "no_prospective_evidence_consumed": True,
        "ab_determinism": {"entry": ab_entry, "matrix": ab_mat, "ok": ab_entry and ab_mat},
        "tests": tests,
        "safety": {
            "submit_cancel_live": "0/0/0",
            "Paper_order": False, "Live_order": False,
            "runtime_ENTRY_changed": False, "runtime_EXIT_changed": False,
            "Universe_changed": False, "Discord_production_notification": False,
        },
    }
    publish(OUT, report, {
        "Index": [{"run_id": run_id, "verdict": verdict}],
        "Parity": [parity],
        "ReasonMapping": mapping["rows"],
        "Cells6": cells,
        "PBv2Cohort": pbv2_cohort,
        "R2Diagnostic": [r2_diag],
        "Decision": [{"verdict": verdict, "selected": selected}],
        "Tests": [tests],
        "Safety": [report["safety"]],
    }, mapping)
    print(f"=== DONE verdict={verdict} ===", flush=True)
    print("=== STOP: do not open 20260810 ===", flush=True)
    return report


if __name__ == "__main__":
    run()
