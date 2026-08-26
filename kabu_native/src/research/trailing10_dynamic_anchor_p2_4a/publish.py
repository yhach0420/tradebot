"""Write P2-4A precommit artifacts. No Capture. No PnL."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import (
    ANALYSIS_ID,
    ANCHOR_SCOPE,
    ANCHOR_SHA,
    CANDIDATE_ID,
    CHECKPOINT_INTERVAL_SEC,
    CHECKPOINT_N,
    CONTAMINATION_LABEL,
    CONTAMINATION_PERIOD,
    DOCUMENT_ID,
    EDGE,
    ENTRY_SHA,
    EVALUATION_CLOCK,
    EXIT_SHA,
    EXTRA_CONFIRMATION_WAIT,
    GRID_SEC,
    MAX_CHECKPOINT_AGE_SEC,
    OLD_CANDIDATE,
    OLD_CANDIDATE_STATUS,
    PRICE_SOURCE,
    STATUS,
    STRATEGY_SHA,
    VERDICT_BLOCKED,
    VERDICT_PASS,
    WINDOW_SEC,
)
from .binding import ENTRY_TIME_BINDING, verify_entry_time_binding
from .synthetic import run_suite

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "trailing10_dynamic_anchor_precommit_p2_4a"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

STATE_EXPRESSION = "OLS_slope(log(Pk/P0),k=0..10)>0 AND P10>P0"


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def spec_payload() -> dict[str, Any]:
    return {
        "candidate_id": CANDIDATE_ID,
        "evaluation_clock": EVALUATION_CLOCK,
        "grid_sec": GRID_SEC,
        "window": "[g-600s, g]",
        "window_sec": WINDOW_SEC,
        "checkpoints": CHECKPOINT_N,
        "checkpoint_interval_sec": CHECKPOINT_INTERVAL_SEC,
        "max_checkpoint_age_sec": MAX_CHECKPOINT_AGE_SEC,
        "price_source": PRICE_SOURCE,
        "state_expression": STATE_EXPRESSION,
        "edge": EDGE,
        "not_evaluable_creates_false": False,
        "anchor_scope": ANCHOR_SCOPE,
        "anchor_time": "g",
        "signal_time": "g",
        "extra_confirmation_wait": EXTRA_CONFIRMATION_WAIT,
        "lunch_bridge": False,
        "independent_t1_trigger_used": False,
    }


def spec_sha() -> str:
    blob = json.dumps(spec_payload(), ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def implementation_sha() -> str:
    parts = [
        "src/research/trailing10_dynamic_anchor_p2_4a/__init__.py",
        "src/research/trailing10_dynamic_anchor_p2_4a/contract.py",
        "src/research/trailing10_dynamic_anchor_p2_4a/synthetic.py",
        "src/research/trailing10_dynamic_anchor_p2_4a/binding.py",
    ]
    h = hashlib.sha256()
    for rel in parts:
        h.update(rel.encode("utf-8"))
        h.update(b"\n")
        h.update((NATIVE / rel).read_bytes())
    return h.hexdigest()


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols = list(rows[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_report() -> dict[str, Any]:
    suite = run_suite()
    bind = verify_entry_time_binding()
    all_ok = suite["failed"] == 0 and bind["CURRENT_ENTRY_TIME_BINDING"] == "PASS"
    now = datetime.now(JST).isoformat(timespec="seconds")
    ssha = spec_sha()
    isha = implementation_sha()
    return {
        "task": "P2-4A",
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "generated_at_jst": now,
        "CANDIDATE_ID": CANDIDATE_ID,
        "STATUS": STATUS if all_ok else "NOT_PRECOMMITTED",
        "INDEPENDENT_T1_TRIGGER_USED": False,
        "EVALUATION_CLOCK": EVALUATION_CLOCK,
        "WINDOW": "[g-600s, g]",
        "CHECKPOINTS": CHECKPOINT_N,
        "CHECKPOINT_INTERVAL_SEC": CHECKPOINT_INTERVAL_SEC,
        "MAX_CHECKPOINT_AGE_SEC": MAX_CHECKPOINT_AGE_SEC,
        "PRICE_SOURCE": PRICE_SOURCE,
        "STATE_EXPRESSION": STATE_EXPRESSION,
        "EDGE": EDGE,
        "NOT_EVALUABLE_CREATES_FALSE": False,
        "ANCHOR_SCOPE": ANCHOR_SCOPE,
        "ANCHOR_TIME": "g",
        "EXTRA_CONFIRMATION_WAIT": EXTRA_CONFIRMATION_WAIT,
        "CURRENT_ENTRY_TIME_BINDING": bind["CURRENT_ENTRY_TIME_BINDING"],
        "entry_time_binding": bind,
        "SYNTHETIC_TESTS": {
            "passed": suite["passed"],
            "failed": suite["failed"],
            "n": suite["n"],
            "results": suite["results"],
        },
        "HISTORICAL_CAPTURE_READ": False,
        "HISTORICAL_PNL_READ": False,
        "OLD_CANDIDATE": OLD_CANDIDATE,
        "OLD_CANDIDATE_STATUS": OLD_CANDIDATE_STATUS,
        "SPEC_FROZEN": bool(all_ok),
        "SPEC_SHA": ssha,
        "IMPLEMENTATION_SHA": isha,
        "FREEZE_TIMESTAMP_JST": now if all_ok else None,
        "contamination": {
            "design_after_p2_3": True,
            "period_must_not_be_called_holdout": CONTAMINATION_PERIOD,
            "forbidden_labels": ["clean holdout", "OOS", "prospective"],
            "if_later_used_for_mechanics": CONTAMINATION_LABEL,
            "prospective_candidate_data": (
                "After freeze, new unseen Capture obtained after freeze_timestamp "
                "may be reserved as PROSPECTIVE_CANDIDATE_DATA if this candidate is unchanged. "
                "Spec/implementation change after seeing new-market result resets the prospective clock."
            ),
        },
        "spec": spec_payload(),
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "CAN_PROCEED": "YES" if all_ok else "NO",
        "verdict": VERDICT_PASS if all_ok else VERDICT_BLOCKED,
        "identity": {
            "strategy_sha": STRATEGY_SHA,
            "entry_sha": ENTRY_SHA,
            "exit_sha": EXIT_SHA,
            "anchor_sha": ANCHOR_SHA,
            "spec_sha": ssha,
            "implementation_sha": isha,
            "V1RNativeEntryLive_sha": _file_sha("src/small_paper/v1r_native_entry_live.py"),
            "V1RLiveDualLane_sha": _file_sha("src/small_paper/v1r_live_dual_lane.py"),
        },
    }


def _md(rep: dict[str, Any]) -> str:
    st = rep["SYNTHETIC_TESTS"]
    failed_ids = [r.get("id") for r in st["results"] if not r.get("ok")]
    return f"""# P2-4A Trailing 10-minute state-edge precommit

Research contract + synthetic only. Historical Capture not read. Historical PnL not read. Runtime not adopted.

Old candidate `{rep['OLD_CANDIDATE']}` remains `{rep['OLD_CANDIDATE_STATUS']}` (immutable). This candidate is separate. Old C1 validation is not proof of TRAIL10.

CANDIDATE_ID:
{rep['CANDIDATE_ID']}

STATUS:
{rep['STATUS']}

INDEPENDENT_T1_TRIGGER_USED:
false

EVALUATION_CLOCK:
{rep['EVALUATION_CLOCK']}

WINDOW:
{rep['WINDOW']}

CHECKPOINTS:
{rep['CHECKPOINTS']}

CHECKPOINT_INTERVAL_SEC:
{rep['CHECKPOINT_INTERVAL_SEC']}

MAX_CHECKPOINT_AGE_SEC:
{rep['MAX_CHECKPOINT_AGE_SEC']}

PRICE_SOURCE:
{rep['PRICE_SOURCE']}

STATE_EXPRESSION:
{rep['STATE_EXPRESSION']}

EDGE:
{rep['EDGE']}

NOT_EVALUABLE_CREATES_FALSE:
false

ANCHOR_SCOPE:
{rep['ANCHOR_SCOPE']}

ANCHOR_TIME:
{rep['ANCHOR_TIME']}

EXTRA_CONFIRMATION_WAIT:
{rep['EXTRA_CONFIRMATION_WAIT']}

CURRENT_ENTRY_TIME_BINDING:
{rep['CURRENT_ENTRY_TIME_BINDING']}

SYNTHETIC_TESTS:
passed: {st['passed']}
failed: {st['failed']}
failed_ids: {failed_ids}

HISTORICAL_CAPTURE_READ:
false

HISTORICAL_PNL_READ:
false

OLD_CANDIDATE_STATUS:
{rep['OLD_CANDIDATE_STATUS']}

SPEC_FROZEN:
{str(rep['SPEC_FROZEN']).lower()}

SPEC_SHA:
{rep['SPEC_SHA']}

FREEZE_TIMESTAMP_JST:
{rep['FREEZE_TIMESTAMP_JST']}

STRATEGY_CHANGED:
false

ENTRY_EXIT_CHANGED:
false

RUNTIME_CHANGED:
false

SAFETY:
submit/cancel/live=0/0/0

CAN_PROCEED:
{rep['CAN_PROCEED']}

verdict:
{rep['verdict']}

Contamination: period {rep['contamination']['period_must_not_be_called_holdout']} must never be called clean holdout / OOS / prospective for this candidate. Later mechanics-only use: `{rep['contamination']['if_later_used_for_mechanics']}`. After freeze, unseen Capture obtained after freeze_timestamp may be reserved as PROSPECTIVE_CANDIDATE_DATA if this candidate is unchanged.

STOP. Do not apply to Historical Capture. Do not read Historical PnL. Do not adopt into Runtime.
"""


def write_artifacts(rep: dict[str, Any]) -> dict[str, Path]:
    OUT.mkdir(parents=True, exist_ok=True)
    for p in OUT.iterdir():
        if p.is_file() and p.name not in {"report.json", "report.md", "audit.xlsx"}:
            p.unlink()
    (OUT / "report.json").write_text(
        json.dumps(rep, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8"
    )
    (OUT / "report.md").write_text(_md(rep), encoding="utf-8")
    wb = Workbook()

    def sh(name: str):
        if name == "Summary":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    summary = [
        {"field": k, "value": json.dumps(v, ensure_ascii=False, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("CANDIDATE_ID", rep["CANDIDATE_ID"]),
            ("STATUS", rep["STATUS"]),
            ("INDEPENDENT_T1_TRIGGER_USED", False),
            ("EVALUATION_CLOCK", rep["EVALUATION_CLOCK"]),
            ("WINDOW", rep["WINDOW"]),
            ("CHECKPOINTS", rep["CHECKPOINTS"]),
            ("CHECKPOINT_INTERVAL_SEC", rep["CHECKPOINT_INTERVAL_SEC"]),
            ("MAX_CHECKPOINT_AGE_SEC", rep["MAX_CHECKPOINT_AGE_SEC"]),
            ("PRICE_SOURCE", rep["PRICE_SOURCE"]),
            ("STATE_EXPRESSION", rep["STATE_EXPRESSION"]),
            ("EDGE", rep["EDGE"]),
            ("NOT_EVALUABLE_CREATES_FALSE", False),
            ("ANCHOR_SCOPE", rep["ANCHOR_SCOPE"]),
            ("ANCHOR_TIME", rep["ANCHOR_TIME"]),
            ("EXTRA_CONFIRMATION_WAIT", rep["EXTRA_CONFIRMATION_WAIT"]),
            ("CURRENT_ENTRY_TIME_BINDING", rep["CURRENT_ENTRY_TIME_BINDING"]),
            ("SYNTHETIC_passed", rep["SYNTHETIC_TESTS"]["passed"]),
            ("SYNTHETIC_failed", rep["SYNTHETIC_TESTS"]["failed"]),
            ("HISTORICAL_CAPTURE_READ", False),
            ("HISTORICAL_PNL_READ", False),
            ("OLD_CANDIDATE_STATUS", rep["OLD_CANDIDATE_STATUS"]),
            ("SPEC_FROZEN", rep["SPEC_FROZEN"]),
            ("SPEC_SHA", rep["SPEC_SHA"]),
            ("FREEZE_TIMESTAMP_JST", rep["FREEZE_TIMESTAMP_JST"]),
            ("CAN_PROCEED", rep["CAN_PROCEED"]),
            ("verdict", rep["verdict"]),
            ("SAFETY", rep["SAFETY"]),
        ]
    ]
    _write_rows(sh("Summary"), summary)
    _write_rows(sh("Contract"), [{"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v} for k, v in rep["spec"].items()])
    _write_rows(sh("Synthetic_Tests"), [
        {"id": r.get("id"), "ok": r.get("ok"), "detail": json.dumps(r.get("detail"), default=str) if isinstance(r.get("detail"), (dict, list)) else r.get("detail")}
        for r in rep["SYNTHETIC_TESTS"]["results"]
    ])
    _write_rows(sh("State_Machine"), [
        {"transition": "FALSE", "effect": "ARMED"},
        {"transition": "FALSE→TRUE (both EVALUABLE)", "effect": "ANCHOR_FIRE then DISARMED"},
        {"transition": "TRUE→TRUE", "effect": "DISARMED, refire 0"},
        {"transition": "TRUE→FALSE", "effect": "ARMED"},
        {"transition": "NOT_EVALUABLE", "effect": "no ARMED/DISARMED change; previous EVALUABLE bool cleared (gap)"},
        {"transition": "FALSE→NOT_EVALUABLE→TRUE", "effect": "fire 0"},
        {"transition": "NOT_EVALUABLE→FALSE→TRUE", "effect": "fire 1"},
        {"transition": "lunch / session gap in [g-600,g]", "effect": "NOT_EVALUABLE SESSION_INVALID"},
    ])
    bind = rep["entry_time_binding"]
    _write_rows(sh("Session_Binding"), [
        {"field": "CURRENT_ENTRY_TIME_BINDING", "value": bind.get("CURRENT_ENTRY_TIME_BINDING")},
        {"field": "missing", "value": json.dumps(bind.get("missing"))},
        *[{"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v} for k, v in (bind.get("path") or {}).items()],
    ])
    _write_rows(sh("Contamination"), [
        {"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in rep["contamination"].items()
    ])
    _write_rows(sh("Identity"), [
        {"field": k, "value": v} for k, v in rep["identity"].items()
    ])
    _write_rows(sh("Safety"), [
        {"field": "submit/cancel/live", "value": "0/0/0"},
        {"field": "STRATEGY_CHANGED", "value": False},
        {"field": "ENTRY_EXIT_CHANGED", "value": False},
        {"field": "RUNTIME_CHANGED", "value": False},
        {"field": "HISTORICAL_CAPTURE_READ", "value": False},
        {"field": "HISTORICAL_PNL_READ", "value": False},
        {"field": "note", "value": "Research contract/synthetic only. Runtime adoption forbidden."},
    ])
    wb.save(OUT / "audit.xlsx")
    return {"report_json": OUT / "report.json", "report_md": OUT / "report.md", "audit_xlsx": OUT / "audit.xlsx"}
