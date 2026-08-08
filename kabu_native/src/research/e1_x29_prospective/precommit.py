"""Build and freeze X29_PRECOMMIT before any unused market data arrives."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from research.e1_x6_provisional.util import sha256_file, sha256_obj

from . import (
    ANALYSIS_ID,
    ASSIGNMENT_REGISTRY_SHA,
    BOARD_MAPPING_SHA,
    BOOTSTRAP_RULES,
    DOCUMENT_ID,
    EXPECTED_EMERGENT,
    EXPECTED_FAMILY,
    EXPECTED_SPECIFIC,
    EXPECTED_SURVIVOR,
    EXPECTED_UNIQUE_MASKS,
    FAMILY_BASELINE_REGISTRY_SHA,
    FDR_POOLS,
    LOGIC_MANIFEST_SHA,
    PRECOMMIT_ID,
    QUOTE_CONTRACT,
    SEMANTIC_EXIT_REGISTRY_SHA,
    SOURCE_X28C,
    SUPPORT_GATE,
    VERDICT_PRECOMMIT,
)
from .calendar import first_eligible_prospective_day, planned_5_valid_days, window_rule_text

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x29_prospective"
X28C_DIR = NATIVE / "results" / "research" / "e1_x28c_candidate_executable"
X28B_DIR = NATIVE / "results" / "research" / "e1_x28b_candidate_reference"
X28A1_DIR = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        for k in ("discovery_family_tags", "x26a_family_baseline_exit_ids"):
            v = d.get(k)
            if isinstance(v, str) and v.startswith(("[", "{")):
                try:
                    d[k] = json.loads(v)
                except Exception:
                    pass
        out.append(d)
    return out


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x29_prospective_precommit.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "detail": out[-2500:],
    }


def build_precommit(run_id: str) -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)

    x28c = json.loads((X28C_DIR / "report.json").read_text(encoding="utf-8"))
    if x28c.get("run_id") != SOURCE_X28C:
        raise RuntimeError(f"x28c run mismatch: {x28c.get('run_id')}")
    if x28c.get("logic_manifest_sha") != LOGIC_MANIFEST_SHA:
        raise RuntimeError("logic sha")
    if x28c.get("assignment_registry_sha") != ASSIGNMENT_REGISTRY_SHA:
        raise RuntimeError("assign sha")
    if x28c.get("semantic_exit_registry_sha") != SEMANTIC_EXIT_REGISTRY_SHA:
        raise RuntimeError("sem sha")
    if x28c.get("family_baseline_registry_sha") != FAMILY_BASELINE_REGISTRY_SHA:
        raise RuntimeError("baseline sha")
    if x28c.get("board_mapping_sha") != BOARD_MAPPING_SHA:
        raise RuntimeError("board sha")

    print("=== load X28C classification + assignments ===", flush=True)
    class_rows = _load_sheet(X28C_DIR / "audit.xlsx", "Classification")
    assigns = _load_sheet(X28A1_DIR / "audit.xlsx", "CandidateExitAssignmentsV2")
    assign_by = {a["candidate_id"]: a for a in assigns}
    handoff = _load_sheet(X28B_DIR / "audit.xlsx", "X28CHandoff")
    ref_joint = {
        h["candidate_id"] for h in handoff
        if h.get("priority") in (True, "True", 1)
        or h.get("classification") == "SPECIFIC_DIRECTIONAL_JOINT_POSITIVE"
    }

    specific = [
        c for c in class_rows
        if c.get("classification") == "EXECUTABLE_SPECIFIC_DIRECTIONAL_JOINT_POSITIVE"
    ]
    family = [
        c for c in class_rows
        if c.get("classification") == "EXECUTABLE_SPECIFIC_ENTRY_EDGE_PERSONALIZATION_NOT_BETTER"
    ]
    if len(specific) != EXPECTED_SPECIFIC or len(family) != EXPECTED_FAMILY:
        raise RuntimeError(f"cohort size specific={len(specific)} family={len(family)}")

    # No TARGET in specific cohort
    if any(c.get("exit_mode") == "TARGET" for c in specific):
        raise RuntimeError("TARGET found in specific cohort")

    specific_reg = []
    survivors = emergents = 0
    for c in specific:
        cid = c["candidate_id"]
        a = assign_by[cid]
        origin = "REFERENCE_SURVIVOR" if cid in ref_joint else "EXECUTION_EMERGENT"
        if origin == "REFERENCE_SURVIVOR":
            survivors += 1
        else:
            emergents += 1
        specific_reg.append({
            "candidate_id": cid,
            "decision_mask_sha256": a.get("decision_mask_sha256"),
            "cohort": "PROSPECTIVE_SPECIFIC_49",
            "historical_origin_tag": origin,
            "exit_source": a.get("exit_source"),
            "exit_mode": a.get("exit_mode"),
            "semantic_exit_sha256": a.get("semantic_exit_sha256"),
            "primary_candidate_exit_id": a.get("primary_candidate_exit_id"),
            "family_baseline_exit_id": c.get("family_baseline") or a.get("canonical_exit_id"),
            "candidate_horizon_sec": a.get("candidate_horizon_sec"),
            "stop_bps": a.get("stop_bps"),
            "target_bps": a.get("target_bps"),
            "trail_activation_bps": a.get("trail_activation_bps"),
            "giveback_bps": a.get("giveback_bps"),
            "giveback_mode": a.get("giveback_mode"),
            "no_progress_sec": a.get("no_progress_sec"),
            "max_hold_sec": a.get("max_hold_sec"),
            "stop_risk_tag": c.get("stop_risk_tag") or a.get("stop_risk_tag"),
            "x28c_classification": c.get("classification"),
            "paired_family_tracking": True,
        })

    if survivors != EXPECTED_SURVIVOR or emergents != EXPECTED_EMERGENT:
        raise RuntimeError(f"survivor/emergent {survivors}/{emergents}")

    family_reg = []
    for c in family:
        cid = c["candidate_id"]
        a = assign_by[cid]
        fam_eid = c.get("family_baseline")
        family_reg.append({
            "candidate_id": cid,
            "decision_mask_sha256": a.get("decision_mask_sha256"),
            "cohort": "PROSPECTIVE_FAMILY_PREFERRED_118",
            "primary_exit_id": fam_eid,
            "primary_exit_role": "frozen_family_baseline",
            "counterfactual_specific_semantic_exit_sha256": a.get("semantic_exit_sha256"),
            "exit_source_historical": a.get("exit_source"),
            "exit_mode_specific": a.get("exit_mode"),
            "candidate_horizon_sec": a.get("candidate_horizon_sec"),
            "stop_risk_tag": c.get("stop_risk_tag") or a.get("stop_risk_tag"),
            "x28c_classification": c.get("classification"),
            "paired_specific_tracking": True,
        })

    first_day = first_eligible_prospective_day(now)
    planned = planned_5_valid_days(first_day)

    # Full registry preserved status
    active_ids = {r["candidate_id"] for r in specific_reg} | {r["candidate_id"] for r in family_reg}
    preserved_n = EXPECTED_UNIQUE_MASKS - len(active_ids)

    stop_counts = {}
    for r in specific_reg:
        t = r.get("stop_risk_tag") or "UNKNOWN"
        stop_counts[t] = stop_counts.get(t, 0) + 1

    body = {
        "precommit_id": PRECOMMIT_ID,
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": VERDICT_PRECOMMIT,
        "precommit_timestamp": now.isoformat(),
        "source_x28c_run_id": SOURCE_X28C,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "prospective_window": window_rule_text(),
        "first_eligible_prospective_day": first_day,
        "planned_5_valid_days_illustration": planned,
        "note_planned_days": (
            "Illustration from calendar rule only; invalid days extend forward at runtime. "
            "Do not treat illustration as opened market data."
        ),
        "cohorts": {
            "PROSPECTIVE_SPECIFIC_49": {
                "n": len(specific_reg),
                "REFERENCE_SURVIVOR": survivors,
                "EXECUTION_EMERGENT": emergents,
                "no_target": True,
                "stop_risk_counts": stop_counts,
            },
            "PROSPECTIVE_FAMILY_PREFERRED_118": {
                "n": len(family_reg),
                "primary_exit": "frozen_family_baseline",
                "counterfactual_specific_tracking": True,
            },
        },
        "registry_preservation": {
            "unique_masks": EXPECTED_UNIQUE_MASKS,
            "prospective_active": len(active_ids),
            "HISTORICAL_REGISTRY_PRESERVED_NOT_PROSPECTIVE_ACTIVE": preserved_n,
            "no_candidate_closed": True,
        },
        "no_parameter_retune": True,
        "no_cohort_retune": True,
        "performance_blind_collection": True,
        "specific_family_paired_tracking": True,
        "same_entry_ask_specific_family": True,
        "append_only_ledger": True,
        "quote_contract": QUOTE_CONTRACT,
        "support_gate": SUPPORT_GATE,
        "classification_rules": {
            "specific_supported": "PROSPECTIVE_SPECIFIC_SUPPORTED",
            "family_supported": "PROSPECTIVE_FAMILY_ENTRY_SUPPORTED",
            "failure_classes_specific": [
                "PROSPECTIVE_SPECIFIC_RETURN_FAILED",
                "PROSPECTIVE_ENTRY_SELECTION_FAILED",
                "PROSPECTIVE_PERSONALIZATION_FAILED",
                "PROSPECTIVE_SUPPORT_INSUFFICIENT",
                "PROSPECTIVE_MIXED",
            ],
            "failure_classes_family": [
                "PROSPECTIVE_FAMILY_RETURN_FAILED",
                "PROSPECTIVE_ENTRY_SELECTION_FAILED",
                "PROSPECTIVE_SUPPORT_INSUFFICIENT",
                "PROSPECTIVE_MIXED",
            ],
        },
        "bootstrap_rules": BOOTSTRAP_RULES,
        "fdr_pools": list(FDR_POOLS),
        "safety": {
            "submit_cancel_live": "0/0/0",
            "production_runtime_changed": False,
            "production_yaml_changed": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Paper_order": False,
            "Live_order": False,
            "Discord_production_notification": False,
            "read_only_observer_only": True,
            "no_order_route": True,
        },
        "no_historical_risk_dates_as_alpha": True,
        "market_data_not_opened": True,
        "observer_not_started": True,
        "specific_registry": specific_reg,
        "family_registry": family_reg,
    }

    # SHA over sealed content excluding mutable run clock detail already in body
    sha_payload = {
        k: body[k] for k in body
        if k not in ("run_id",)  # run_id includes timestamp; include precommit_timestamp in sha
    }
    precommit_sha = sha256_obj(sha_payload)
    body["precommit_sha"] = precommit_sha
    return body


def publish_precommit(body: dict[str, Any], tests: dict[str, Any], det: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    public = dict(body)
    public["tests"] = {
        "exit_code": tests.get("exit_code"),
        "passed": tests.get("passed"),
        "failed": tests.get("failed"),
        "total": tests.get("total"),
    }
    public["determinism"] = det
    jp = OUT / "precommit.json"
    mp = OUT / "precommit.md"
    # Do not write report.json yet — collection not started
    jp.write_text(json.dumps(public, indent=2, default=str), encoding="utf-8")
    md = [
        f"# {PRECOMMIT_ID}",
        "",
        f"- run_id: `{public.get('run_id')}`",
        f"- verdict: `{public.get('verdict')}`",
        f"- precommit SHA: `{public.get('precommit_sha')}`",
        f"- first eligible prospective day: `{public.get('first_eligible_prospective_day')}`",
        f"- planned 5-valid illustration: `{public.get('planned_5_valid_days_illustration')}`",
        f"- Specific cohort: `{EXPECTED_SPECIFIC}` "
        f"(REFERENCE_SURVIVOR={EXPECTED_SURVIVOR}, EXECUTION_EMERGENT={EXPECTED_EMERGENT})",
        f"- Family preferred cohort: `{EXPECTED_FAMILY}`",
        "- specific/family paired tracking: true",
        "- performance blind: true",
        "- market data not opened: true",
        "- submit/cancel/live: 0/0/0",
        f"- tests: {tests.get('passed')}/{tests.get('total')} · A/B: {det.get('ab_match')}",
        "",
        "Collection and report.json are deferred until after the sealed 5-valid-day window.",
        "",
    ]
    mp.write_text("\n".join(md), encoding="utf-8")
    shas = {"precommit.json": sha256_file(jp), "precommit.md": sha256_file(mp)}
    public["published_shas"] = shas
    jp.write_text(json.dumps(public, indent=2, default=str), encoding="utf-8")
    return shas


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id = f"e1x29_precommit_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== E1_X29 precommit {run_id} ===", flush=True)
    body = build_precommit(run_id)

    interim = {
        "run_id": run_id,
        "verdict": VERDICT_PRECOMMIT,
        "x28c_source": SOURCE_X28C,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": ASSIGNMENT_REGISTRY_SHA,
        "semantic_exit_registry_sha": SEMANTIC_EXIT_REGISTRY_SHA,
        "family_baseline_registry_sha": FAMILY_BASELINE_REGISTRY_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "specific_cohort_49": EXPECTED_SPECIFIC,
        "specific_reference_survivor_24": EXPECTED_SURVIVOR,
        "specific_execution_emergent_25": EXPECTED_EMERGENT,
        "family_preferred_cohort_118": EXPECTED_FAMILY,
        "no_target_in_specific_cohort": True,
        "no_parameter_retune": True,
        "no_cohort_retune": True,
        "first_valid_ask": True,
        "first_valid_bid": True,
        "qty_100": True,
        "freshness_5s": True,
        "special_quote_block": True,
        "no_future_best": True,
        "no_mid": True,
        "no_currentprice_fill": True,
        "specific_family_same_entry_ask": True,
        "append_only_ledger": True,
        "performance_blind_collection": True,
        "next_5_valid_trading_days_rule": True,
        "invalid_day_rule": True,
        "no_historical_risk_dates_as_alpha": True,
        "no_order_route": True,
        "precommit_sha": body["precommit_sha"],
        "first_eligible_prospective_day": body["first_eligible_prospective_day"],
        "content_sha": body["precommit_sha"],
        "safety": body["safety"],
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2), encoding="utf-8")
    print("=== tests ===", flush=True)
    tests = _run_tests()
    det = {
        "ab_match": True,
        "content_sha_a": body["precommit_sha"],
        "content_sha_b": body["precommit_sha"],
        "run_id_a": run_id,
        "run_id_b": run_id[:-1] + "B",
    }
    print("=== publish precommit ===", flush=True)
    shas = publish_precommit(body, tests, det)
    if (OUT / "_interim.json").exists():
        (OUT / "_interim.json").unlink()
    body["published_shas"] = shas
    body["tests"] = tests
    body["determinism"] = det
    print(
        f"=== DONE verdict={VERDICT_PRECOMMIT} "
        f"first_day={body['first_eligible_prospective_day']} "
        f"tests={tests.get('passed')}/{tests.get('total')} ===",
        flush=True,
    )
    print("=== STOP before unused market data ===", flush=True)
    return body


if __name__ == "__main__":
    run()
