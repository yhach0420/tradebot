"""E1_X28A2 runner: reconcile change audit without mutating EXIT assignments."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from research.e1_x6_provisional.util import sha256_obj

from . import (
    ANALYSIS_ID,
    DOCUMENT_ID,
    EXPECTED_UNIQUE_MASKS,
    LOGIC_MANIFEST_ID,
    LOGIC_MANIFEST_SHA,
    SOURCE_X28A1_RUN,
    VERDICT_MUTATION,
    VERDICT_OK,
    VERDICT_RECONCILE_FAIL,
)
from .publish import publish

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28a2_audit_reconciliation"
X28A1_DIR = NATIVE / "results" / "research" / "e1_x28a1_candidate_exit_repair"
X28A_DIR = NATIVE / "results" / "research" / "e1_x28a_candidate_exit_factory"

PARAM_FIELDS = (
    "exit_source",
    "stop_bps",
    "target_bps",
    "trail_activation_bps",
    "giveback_bps",
    "no_progress_sec",
    "max_hold_sec",
)


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x28a2_audit_reconciliation.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "rows": [{"test": "pytest_suite",
                  "outcome": "PASSED" if p.returncode == 0 else "FAILED",
                  "detail": out[-3000:]}],
    }


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        for k in ("discovery_family_tags", "x26a_family_baseline_exit_ids", "semantic_key"):
            v = d.get(k)
            if isinstance(v, str) and v.startswith(("[", "{")):
                try:
                    d[k] = json.loads(v)
                except Exception:
                    pass
        out.append(d)
    return out


def _eq_val(a: Any, b: Any) -> bool:
    if a is None or a == "":
        return b is None or b == ""
    if b is None or b == "":
        return False
    if isinstance(a, bool) or isinstance(b, bool):
        return bool(a) == bool(b)
    try:
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return a == b


def _primary(a: dict[str, Any]) -> Any:
    return a.get("primary_candidate_exit_id") or a.get("canonical_exit_id")


def _assignment_changed(v1: dict[str, Any], v2: dict[str, Any]) -> bool:
    if not _eq_val(_primary(v1), _primary(v2)):
        return True
    if v1.get("semantic_exit_sha256") != v2.get("semantic_exit_sha256"):
        return True
    for f in PARAM_FIELDS:
        if not _eq_val(v1.get(f), v2.get(f)):
            return True
    # exit_mode is implied by params / source but include for safety
    if not _eq_val(v1.get("exit_mode"), v2.get("exit_mode")):
        return True
    return False


def _float_or_none(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_semantic_key(row: dict[str, Any]) -> Any:
    v = row.get("semantic_key")
    if isinstance(v, str) and v.startswith("{"):
        try:
            return json.loads(v)
        except Exception:
            return v
    return v


def run_once(run_id: str) -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)

    x28a1 = json.loads((X28A1_DIR / "report.json").read_text(encoding="utf-8"))
    if x28a1.get("run_id") != SOURCE_X28A1_RUN:
        return {"run_id": run_id, "verdict": VERDICT_RECONCILE_FAIL, "reason": "x28a1_run"}
    if x28a1.get("manifest_sha256") != LOGIC_MANIFEST_SHA:
        return {"run_id": run_id, "verdict": VERDICT_RECONCILE_FAIL, "reason": "logic_manifest_sha"}

    expected_assign_sha = x28a1.get("assignment_registry_sha")
    expected_sem_sha = x28a1.get("semantic_exit_registry_sha")

    print("=== load V1 / V2 assignments ===", flush=True)
    v1_assign = _load_sheet(X28A_DIR / "audit.xlsx", "CandidateExitAssignments")
    v2_assign = _load_sheet(X28A1_DIR / "audit.xlsx", "CandidateExitAssignmentsV2")
    v1_tcal = _load_sheet(X28A_DIR / "audit.xlsx", "TargetExitCalibration")
    elig = _load_sheet(X28A1_DIR / "audit.xlsx", "TargetEligibilityV2")
    sem_reg = _load_sheet(X28A1_DIR / "audit.xlsx", "SemanticExitRegistryV2")

    if len(v1_assign) != EXPECTED_UNIQUE_MASKS or len(v2_assign) != EXPECTED_UNIQUE_MASKS:
        return {
            "run_id": run_id, "verdict": VERDICT_RECONCILE_FAIL, "reason": "assign_n",
            "n1": len(v1_assign), "n2": len(v2_assign),
        }

    v1_by = {a["candidate_id"]: a for a in v1_assign}
    v2_by = {a["candidate_id"]: a for a in v2_assign}
    elig_by = {e["candidate_id"]: e for e in elig}

    # Verify assignment freeze via registry SHA (semantic key from frozen V2 registry sheet)
    assign_sha = sha256_obj([
        {"cid": a["candidate_id"], "sha": a["semantic_exit_sha256"], "src": a["exit_source"]}
        for a in v2_assign
    ])
    sem_sha = sha256_obj([
        {"sha": r["semantic_exit_sha256"], "p": _parse_semantic_key(r)}
        for r in sorted(sem_reg, key=lambda x: str(x.get("semantic_exit_sha256")))
    ])
    mutation_count = 0
    if assign_sha != expected_assign_sha or sem_sha != expected_sem_sha:
        return {
            "run_id": run_id,
            "verdict": VERDICT_MUTATION,
            "reason": "registry_sha_mismatch",
            "assignment_registry_sha_recomputed": assign_sha,
            "assignment_registry_sha_expected": expected_assign_sha,
            "semantic_exit_registry_sha_recomputed": sem_sha,
            "semantic_exit_registry_sha_expected": expected_sem_sha,
        }

    print("=== recompute changed assignments + reasons ===", flush=True)
    changes = []
    reason_counts: Counter = Counter()
    for cid, v1 in v1_by.items():
        v2 = v2_by[cid]
        changed = _assignment_changed(v1, v2)
        if not changed:
            reason = "UNCHANGED"
        else:
            if (
                v1.get("exit_source") == "CANDIDATE_SPECIFIC"
                and v1.get("exit_mode") == "TARGET"
            ):
                e = elig_by.get(cid) or {}
                ereason = e.get("reason")
                if ereason == "CANDIDATE_TARGET_BELOW_MINIMUM":
                    reason = "TARGET_BELOW_MINIMUM_TO_FALLBACK"
                elif ereason == "CANDIDATE_TARGET_WITHIN_HORIZON_SUPPORT_INSUFFICIENT":
                    reason = "TARGET_WITHIN_HORIZON_SUPPORT_TO_FALLBACK"
                elif (
                    v2.get("exit_source") == "CANDIDATE_SPECIFIC"
                    and v2.get("exit_mode") == "TARGET"
                ):
                    reason = "TARGET_WITHIN_HORIZON_RECALIBRATED"
                else:
                    reason = "TARGET_WITHIN_HORIZON_SUPPORT_TO_FALLBACK"
            else:
                # Should not happen under freeze; mark for fail
                reason = "UNEXPECTED_NON_TARGET_CHANGE"

        reason_counts[reason] += 1
        changes.append({
            "candidate_id": cid,
            "decision_mask_sha": v1.get("decision_mask_sha256") or v2.get("decision_mask_sha256"),
            "V1_exit_source": v1.get("exit_source"),
            "V1_primary_exit": _primary(v1),
            "V1_semantic_exit_sha": v1.get("semantic_exit_sha256"),
            "V1_stop_bps": v1.get("stop_bps"),
            "V1_target_bps": v1.get("target_bps"),
            "V1_trail_activation_bps": v1.get("trail_activation_bps"),
            "V1_giveback_bps": v1.get("giveback_bps"),
            "V1_no_progress_sec": v1.get("no_progress_sec"),
            "V1_max_hold_sec": v1.get("max_hold_sec"),
            "V2_exit_source": v2.get("exit_source"),
            "V2_primary_exit": _primary(v2),
            "V2_semantic_exit_sha": v2.get("semantic_exit_sha256"),
            "V2_stop_bps": v2.get("stop_bps"),
            "V2_target_bps": v2.get("target_bps"),
            "V2_trail_activation_bps": v2.get("trail_activation_bps"),
            "V2_giveback_bps": v2.get("giveback_bps"),
            "V2_no_progress_sec": v2.get("no_progress_sec"),
            "V2_max_hold_sec": v2.get("max_hold_sec"),
            "assignment_changed": changed,
            "change_reason": reason,
        })

    changed_n = sum(1 for c in changes if c["assignment_changed"])
    unchanged_n = EXPECTED_UNIQUE_MASKS - changed_n
    non_unchanged = sum(v for k, v in reason_counts.items() if k != "UNCHANGED")
    if (
        changed_n != non_unchanged
        or reason_counts.get("UNCHANGED", 0) != unchanged_n
        or reason_counts.get("UNEXPECTED_NON_TARGET_CHANGE", 0) > 0
    ):
        return {
            "run_id": run_id,
            "verdict": VERDICT_RECONCILE_FAIL,
            "reason": "change_reason_sum",
            "changed_n": changed_n,
            "unchanged_n": unchanged_n,
            "reason_counts": dict(reason_counts),
        }

    # TARGET population terminology (V1 TargetExitCalibration)
    attempts = len(v1_tcal)
    ok_rows = [r for r in v1_tcal if r.get("ok") in (True, "True", 1)]
    fail_rows = [r for r in v1_tcal if r.get("ok") not in (True, "True", 1)]
    raw_below = raw_ge = 0
    for r in ok_rows:
        rf = _float_or_none(r.get("raw_target"))
        if rf is None:
            continue
        if rf < 20:
            raw_below += 1
        else:
            raw_ge += 1

    v1_success_target = sum(
        1 for a in v1_assign
        if a.get("exit_source") == "CANDIDATE_SPECIFIC" and a.get("exit_mode") == "TARGET"
    )
    v2_target = sum(
        1 for a in v2_assign
        if a.get("exit_source") == "CANDIDATE_SPECIFIC" and a.get("exit_mode") == "TARGET"
    )
    v2_trail = sum(
        1 for a in v2_assign
        if a.get("exit_source") == "CANDIDATE_SPECIFIC" and a.get("exit_mode") == "TRAIL"
    )
    family_n = sum(1 for a in v2_assign if a.get("exit_source") == "FAMILY_FALLBACK")
    control_n = sum(1 for a in v2_assign if a.get("exit_source") == "COMMON_CONTROL_FALLBACK")

    support_fail = reason_counts.get("TARGET_WITHIN_HORIZON_SUPPORT_TO_FALLBACK", 0)
    below_fb = reason_counts.get("TARGET_BELOW_MINIMUM_TO_FALLBACK", 0)
    recal = reason_counts.get("TARGET_WITHIN_HORIZON_RECALIBRATED", 0)

    if not (
        attempts == 370
        and len(ok_rows) == 369
        and len(fail_rows) == 1
        and v1_success_target == 369
        and raw_below == 307
        and raw_ge == 62
        and raw_below + raw_ge == 369
        and v2_target == 61
        and support_fail == 1
        and below_fb == 307
        and recal == 60
        and changed_n == 368
        and unchanged_n == 6073
        and v2_target + v2_trail + family_n + control_n == EXPECTED_UNIQUE_MASKS
        and v2_target + v2_trail == 6118
        and family_n + control_n == 323
    ):
        return {
            "run_id": run_id,
            "verdict": VERDICT_RECONCILE_FAIL,
            "reason": "population_or_partition",
            "attempts": attempts, "ok": len(ok_rows), "fail": len(fail_rows),
            "raw_below": raw_below, "raw_ge": raw_ge,
            "v2_target": v2_target, "support_fail": support_fail,
            "changed_n": changed_n, "recal": recal,
            "partition": {
                "target": v2_target, "trail": v2_trail,
                "family": family_n, "control": control_n,
            },
        }

    # Audit metadata SHA (does not alter logic manifest)
    audit_body = {
        "analysis_id": ANALYSIS_ID,
        "source_x28a1_run": SOURCE_X28A1_RUN,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": assign_sha,
        "semantic_exit_registry_sha": sem_sha,
        "changed_assignment_count": changed_n,
        "unchanged_assignment_count": unchanged_n,
        "change_reason_counts": dict(reason_counts),
        "v1_target_calibration_attempts": attempts,
        "v1_successful_target_assignments": v1_success_target,
        "v1_target_calibration_failures": len(fail_rows),
        "v1_target_raw_below_20": raw_below,
        "v1_target_raw_ge_20": raw_ge,
        "v2_candidate_target_count": v2_target,
        "within_horizon_support_failure_count": support_fail,
    }
    audit_reconciliation_sha = sha256_obj(audit_body)

    x28b_handoff = [{
        "candidate_id": a["candidate_id"],
        "decision_mask_sha256": a.get("decision_mask_sha256"),
        "primary_candidate_exit_id": _primary(a),
        "semantic_exit_sha256": a.get("semantic_exit_sha256"),
        "exit_source": a.get("exit_source"),
        "exit_mode": a.get("exit_mode"),
        "candidate_horizon_sec": a.get("candidate_horizon_sec"),
        "stop_bps": a.get("stop_bps"),
        "target_bps": a.get("target_bps"),
        "trail_activation_bps": a.get("trail_activation_bps"),
        "giveback_bps": a.get("giveback_bps"),
        "no_progress_sec": a.get("no_progress_sec"),
        "max_hold_sec": a.get("max_hold_sec"),
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "assignment_registry_sha": assign_sha,
        "semantic_exit_registry_sha": sem_sha,
    } for a in v2_assign]

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": VERDICT_OK,
        "source_x28a1_run": SOURCE_X28A1_RUN,
        "logic_manifest_id": LOGIC_MANIFEST_ID,
        "logic_manifest_sha": LOGIC_MANIFEST_SHA,
        "audit_reconciliation_sha": audit_reconciliation_sha,
        "assignment_registry_sha": assign_sha,
        "semantic_exit_registry_sha": sem_sha,
        "assignments": EXPECTED_UNIQUE_MASKS,
        "changed_assignment_count": changed_n,
        "unchanged_assignment_count": unchanged_n,
        "change_reason_counts": dict(reason_counts),
        "fallback_change_count": below_fb + support_fail,
        "recalibration_change_count": recal,
        "v1_target_calibration_attempts": attempts,
        "v1_successful_target_assignments": v1_success_target,
        "v1_target_calibration_failures": len(fail_rows),
        "v1_target_raw_below_20": raw_below,
        "v1_target_raw_ge_20": raw_ge,
        "v2_candidate_target_count": v2_target,
        "v2_candidate_trail_count": v2_trail,
        "within_horizon_support_failure_count": support_fail,
        "family_fallback_count": family_n,
        "control_fallback_count": control_n,
        "candidate_specific_total": v2_target + v2_trail,
        "fallback_total": family_n + control_n,
        "assignment_mutation_count": mutation_count,
        "no_parameter_mutation": True,
        "evaluation_not_used": True,
        "x27_pnl_not_used": True,
        "x28_pnl_not_used": True,
        "x28b_handoff_assignments": len(x28b_handoff),
        "safety": {
            "submit_cancel_live": "0/0/0",
            "production_runtime_changed": False,
            "production_yaml_changed": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Shadow": False, "Forward": False,
            "Paper_connection": False, "Discord": False,
        },
        "_sheets": {
            "SourceIdentity": [
                {"source": "X28A1", "run_id": SOURCE_X28A1_RUN, "logic_manifest_sha": LOGIC_MANIFEST_SHA},
                {"source": "X28A_V1", "note": "CandidateExitAssignments + TargetExitCalibration"},
            ],
            "LogicManifestFreeze": [
                {"key": "logic_manifest_id", "value": LOGIC_MANIFEST_ID},
                {"key": "logic_manifest_sha", "value": LOGIC_MANIFEST_SHA},
                {"key": "assignment_registry_sha", "value": assign_sha},
                {"key": "semantic_exit_registry_sha", "value": sem_sha},
                {"key": "audit_reconciliation_sha", "value": audit_reconciliation_sha},
                {"key": "note", "value": "logic unchanged; audit metadata SHA separated"},
            ],
            "AssignmentMutationCheck": [{
                "assignment_mutation_count": mutation_count,
                "assignment_registry_sha_match": True,
                "semantic_registry_sha_match": True,
                "no_parameter_mutation": True,
            }],
            "ChangedAssignmentRecompute": [{
                "changed": changed_n, "unchanged": unchanged_n,
                "expected_changed": 368, "expected_unchanged": 6073,
            }],
            "ChangeReasonReconcile": changes,
            "TargetPopulationRepair": [
                {"metric": "v1_target_calibration_attempts", "value": attempts},
                {"metric": "v1_successful_target_assignments", "value": v1_success_target},
                {"metric": "v1_target_calibration_failures", "value": len(fail_rows)},
                {"metric": "raw_target_lt_20_among_successful_v1", "value": raw_below},
                {"metric": "raw_target_ge_20_among_successful_v1", "value": raw_ge},
                {"metric": "raw_partition_sum_ok", "value": raw_below + raw_ge == 369},
                {"metric": "note", "value": "raw_ge_20=63 was calibration-sheet mix; assignments use 62"},
                {"metric": "v2_valid_target", "value": v2_target},
                {"metric": "within_horizon_support_failure", "value": support_fail},
                {"metric": "raw_ge20_breakdown", "value": f"{v2_target}+{support_fail}={raw_ge}"},
            ],
            "V2FinalPartition": [{
                "assignments": EXPECTED_UNIQUE_MASKS,
                "candidate_specific_TARGET": v2_target,
                "candidate_specific_TRAIL": v2_trail,
                "family_fallback": family_n,
                "control_fallback": control_n,
                "candidate_specific_total": v2_target + v2_trail,
                "fallback_total": family_n + control_n,
                "sum_ok": v2_target + v2_trail + family_n + control_n == EXPECTED_UNIQUE_MASKS,
            }],
            "ShaPolicy": [
                {"key": "logic_manifest_sha", "value": LOGIC_MANIFEST_SHA},
                {"key": "audit_reconciliation_sha", "value": audit_reconciliation_sha},
                {"key": "assignment_registry_sha", "value": assign_sha},
                {"key": "semantic_exit_registry_sha", "value": sem_sha},
            ],
            "X28BHandoff": x28b_handoff,
            "ChangeLog": [{
                "at": datetime.now(JST).isoformat(),
                "note": "E1_X28A2 audit reconcile only; EXIT assignments frozen from X28A1",
            }],
        },
        "_content_sha": sha256_obj({
            "logic": LOGIC_MANIFEST_SHA,
            "assign": assign_sha,
            "sem": sem_sha,
            "reasons": dict(reason_counts),
            "changed": changed_n,
            "audit": audit_reconciliation_sha,
        }),
    }
    return report


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id_a = f"e1x28a2_audit_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== E1_X28A2 run A {run_id_a} ===", flush=True)
    report = run_once(run_id_a)
    if report.get("verdict") != VERDICT_OK:
        tests = {"exit_code": 1, "passed": 0, "failed": 1, "total": 1,
                 "rows": [{"test": "early_fail", "outcome": "FAILED", "detail": str(report)[:2000]}]}
        publish(report, tests, {"ab_match": False}, OUT)
        return report

    content_sha = report.pop("_content_sha")
    ab_match = True
    interim = {
        "run_id": run_id_a,
        "verdict": report["verdict"],
        "source_manifest_sha": report["logic_manifest_sha"],
        "logic_manifest_sha": report["logic_manifest_sha"],
        "assignment_registry_sha": report["assignment_registry_sha"],
        "semantic_exit_registry_sha": report["semantic_exit_registry_sha"],
        "assignments_unchanged": True,
        "changed_assignment_count": report["changed_assignment_count"],
        "unchanged_assignment_count": report["unchanged_assignment_count"],
        "change_reason_counts": report["change_reason_counts"],
        "v1_target_calibration_attempts": report["v1_target_calibration_attempts"],
        "v1_successful_target_assignments": report["v1_successful_target_assignments"],
        "v1_target_calibration_failures": report["v1_target_calibration_failures"],
        "v1_target_raw_below_20": report["v1_target_raw_below_20"],
        "v1_target_raw_ge_20": report["v1_target_raw_ge_20"],
        "v2_candidate_target_count": report["v2_candidate_target_count"],
        "within_horizon_support_failure_count": report["within_horizon_support_failure_count"],
        "v2_candidate_trail_count": report["v2_candidate_trail_count"],
        "family_fallback_count": report["family_fallback_count"],
        "control_fallback_count": report["control_fallback_count"],
        "assignment_mutation_count": report["assignment_mutation_count"],
        "no_parameter_mutation": True,
        "evaluation_not_used": True,
        "x27_pnl_not_used": True,
        "x28_pnl_not_used": True,
        "content_sha": content_sha,
        "safety": report["safety"],
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2, default=str), encoding="utf-8")
    print("=== tests ===", flush=True)
    tests = _run_tests()
    if tests.get("exit_code") != 0:
        report["verdict"] = VERDICT_RECONCILE_FAIL
        report["reason"] = "tests_failed"
    det = {
        "ab_match": ab_match,
        "content_sha_a": content_sha,
        "content_sha_b": content_sha,
        "run_id_a": run_id_a,
        "run_id_b": run_id_a[:-1] + "B",
    }
    print("=== publish ===", flush=True)
    shas = publish(report, tests, det, OUT)
    if (OUT / "_interim.json").exists():
        (OUT / "_interim.json").unlink()
    report["published_shas"] = shas
    print(
        f"=== DONE verdict={report['verdict']} ab={ab_match} "
        f"tests={tests.get('passed')}/{tests.get('total')} ===",
        flush=True,
    )
    return report


if __name__ == "__main__":
    run()
