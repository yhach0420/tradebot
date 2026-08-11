"""V1R EXIT V2 — Asymmetric Sequential EXIT State Machine.

Tests H0 (instability) vs H1 (state-dependent EXIT). ENTRY frozen.
20260810 = RETROSPECTIVE_REFERENCE_ONLY (not for selection).
Production mutation: NONE.
"""
from __future__ import annotations

import hashlib
import json
import math
import pickle
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

from research.e1_x22_actual_exit_factory.paths import session_end_epoch
from research.e1_x32_upstream_attribution.eval_stages import load_boards_for_symbols
from research.e1_x35_passive_exit.paths import build_path
from research.e1_x36_joint_allocator import OUTER_BLOCKS, LOT_QTY
from research.e1_x36_joint_allocator.replay import simulate_joint
from research.e1_x36r_freeze_integrity.serialize import score_fn_from_serialized
from research.e1_x37_prospective.freeze import load_model_artifact, load_v1r
from research.v1r_exit_v2_asymmetric.continuation import (
    continuation_supported,
    features_at_600,
    fit_continuation_models,
    generate_continuation_rules,
)
from research.v1r_exit_v2_asymmetric.guards import detect_guard_trigger, generate_guards
from research.v1r_exit_v2_asymmetric.policy import apply_architecture
from research.v1r_exit_v2_asymmetric.states import build_trade_bundle, exit_at_horizon
from small_paper.v1r_primary_runtime import MODEL_ARTIFACT_SHA, V1R_SHA

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "v1r_exit_v2_asymmetric"
PANEL_CACHE = NATIVE / "results/research/v1r_capital_sweep_0p5m_10m/_panel_cache.pkl"
FORBIDDEN = "20260810"
ANALYSIS_ID = "V1R_EXIT_V2_ASYMMETRIC"
FEATS = ("spread_bps", "imbalance", "mid_ret_60s", "mid_ret_180s", "event_rate_60s", "log_bid_qty")
CATASTROPHIC_FOLD_YEN = -150_000.0


def _sfn():
    ser = load_model_artifact()
    raw = score_fn_from_serialized(ser)

    def fn(e: dict) -> float:
        try:
            return float(raw({k: e.get(k) for k in FEATS}))
        except Exception:
            return float("-inf")

    return fn


def _pnl(fill_price: float, ret_bps: float) -> float:
    return float(LOT_QTY) * float(fill_price) * float(ret_bps) / 10000.0


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def write_xlsx(sheets: dict[str, list[dict]], path: Path) -> None:
    if Workbook is None:
        path.with_suffix(".sheets.json").write_text(
            json.dumps(sheets, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def summarize_pnls(pnls: list[float]) -> dict[str, Any]:
    if not pnls:
        return {
            "n": 0, "total": 0.0, "pf": 0.0, "wins": 0, "losses": 0, "flats": 0,
            "win_rate": 0.0, "avg": 0.0, "median": 0.0, "best": 0.0, "worst": 0.0,
            "gross_profit": 0.0, "gross_loss": 0.0, "max_dd": 0.0,
        }
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    gp = sum(p for p in pnls if p > 0)
    gl = abs(sum(p for p in pnls if p < 0))
    pf = (gp / gl) if gl > 0 else (float("inf") if gp > 0 else 0.0)
    eq = peak = 0.0
    max_dd = 0.0
    for p in pnls:
        eq += p
        peak = max(peak, eq)
        max_dd = min(max_dd, eq - peak)
    return {
        "n": len(pnls),
        "total": float(sum(pnls)),
        "pf": None if pf == float("inf") else float(pf),
        "wins": wins,
        "losses": losses,
        "flats": sum(1 for p in pnls if p == 0),
        "win_rate": wins / len(pnls),
        "avg": float(statistics.mean(pnls)),
        "median": float(statistics.median(pnls)),
        "best": float(max(pnls)),
        "worst": float(min(pnls)),
        "gross_profit": float(gp),
        "gross_loss": float(gl),
        "max_dd": float(max_dd),
    }


def signal_hhmm(fill_time: float) -> str:
    dt = datetime.fromtimestamp(float(fill_time), tz=JST)
    return f"{dt.hour:02d}:{dt.minute:02d}"


def composite(sm: dict[str, Any], base: dict[str, Any], *, complexity: float = 1.0) -> float:
    score = (sm["total"] - base["total"]) / 1000.0
    pf = sm["pf"] if sm["pf"] is not None else 10.0
    bpf = base["pf"] if base["pf"] is not None else 10.0
    score += (pf - bpf) * 600.0
    score += (sm["max_dd"] - base["max_dd"]) / 300.0
    score += (sm["worst"] - base["worst"]) / 150.0
    score += (base["gross_loss"] - sm["gross_loss"]) / 400.0
    score -= complexity * 20.0
    return score


def support_ok(trigger_n: int, days: int, syms: int, *, n_train: int, n_days: int) -> bool:
    min_trig = max(10, int(0.07 * n_train))
    min_days = max(3, int(0.3 * n_days))
    min_syms = max(3, min(6, n_train // 20))
    return trigger_n >= min_trig and days >= min_days and syms >= min_syms


def eval_policy_isolation(
    bundles: list[dict[str, Any]],
    *,
    arch: str,
    guard: Optional[dict[str, Any]] = None,
    cont_rule: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    pnls = []
    base_pnls = []
    holds = []
    guard_hits = []
    winners_cut = losers_saved = 0
    saved_loss = foregone_winner = 0.0
    ext_n = ext_win = ext_loss = 0
    ext_incr = 0.0
    days = set()
    syms = set()
    trig_holds = []

    for b in bundles:
        base_pnl = _pnl(b["fill_price"], float(b["ret600"] or 0))
        pol = apply_architecture(b, arch=arch, guard=guard, cont_rule=cont_rule)
        if not pol.get("ok"):
            continue
        pnl = _pnl(b["fill_price"], float(pol["exit_ret_bps"]))
        pnls.append(pnl)
        base_pnls.append(base_pnl)
        holds.append(float(pol["exit_off"]))

        if pol.get("triggered_guard"):
            guard_hits.append(1)
            days.add(b["date"])
            syms.add(b["symbol"])
            trig_holds.append(float(pol.get("guard_trigger_off") or pol["exit_off"]))
            delta = pnl - base_pnl
            if base_pnl < 0 and delta > 0:
                losers_saved += 1
                saved_loss += delta
            if base_pnl > 0 and delta < 0:
                winners_cut += 1
                foregone_winner += -delta
        else:
            guard_hits.append(0)

        if pol.get("extended") and not pol.get("triggered_guard"):
            ext_n += 1
            d750 = _pnl(b["fill_price"], float(b["ret750"] or 0)) - base_pnl
            ext_incr += d750
            if d750 > 0:
                ext_win += 1
            elif d750 < 0:
                ext_loss += 1

    sm = summarize_pnls(pnls)
    base = summarize_pnls(base_pnls)
    ratio = (saved_loss / foregone_winner) if foregone_winner > 1e-9 else (float("inf") if saved_loss > 0 else 0.0)
    return {
        "summary": sm,
        "base_summary": base,
        "delta_pnl": sm["total"] - base["total"],
        "delta_pf": (sm["pf"] or 0) - (base["pf"] or 0),
        "delta_worst": sm["worst"] - base["worst"],
        "delta_dd": sm["max_dd"] - base["max_dd"],
        "avg_hold": float(statistics.mean(holds)) if holds else None,
        "guard_trigger_n": int(sum(guard_hits)),
        "guard_trigger_rate": float(sum(guard_hits) / max(1, len(guard_hits))),
        "guard_days": len(days),
        "guard_symbols": len(syms),
        "median_trigger_off": float(statistics.median(trig_holds)) if trig_holds else None,
        "winners_cut": winners_cut,
        "losers_saved": losers_saved,
        "saved_loss_yen": float(saved_loss),
        "foregone_winner_yen": float(foregone_winner),
        "saved_lost_ratio": None if ratio == float("inf") else float(ratio),
        "extension_n": ext_n,
        "extension_win": ext_win,
        "extension_loss": ext_loss,
        "extension_incremental_pnl": float(ext_incr),
        "pnls": pnls,
    }


def guard_economics(bundles: list[dict[str, Any]], guard: dict[str, Any]) -> dict[str, Any]:
    """Early guard alone vs FIXED600 (survivors keep 600)."""
    return eval_policy_isolation(bundles, arch="C", guard=guard, cont_rule=None)


def heterogeneity_tests(bundles: list[dict[str, Any]]) -> dict[str, Any]:
    """Question A/B: can causal state separate treatment advantage?"""
    # A: among all trades, state@600 vs delta_750_vs_600
    rows_a = []
    for b in bundles:
        if b.get("delta_750_vs_600_bps") is None:
            continue
        f = features_at_600(b)
        d = float(b["delta_750_vs_600_bps"])
        rows_a.append({**f, "delta": d, "benefit": 1 if d > 0 else 0})

    def split_mean(rows, key, thr, direction="ge"):
        if direction == "ge":
            hi = [r["delta"] for r in rows if r[key] >= thr]
            lo = [r["delta"] for r in rows if r[key] < thr]
        else:
            hi = [r["delta"] for r in rows if r[key] <= thr]
            lo = [r["delta"] for r in rows if r[key] > thr]
        if len(hi) < 8 or len(lo) < 8:
            return None
        return {
            "key": key, "thr": thr, "direction": direction,
            "n_hi": len(hi), "n_lo": len(lo),
            "mean_hi": float(statistics.mean(hi)),
            "mean_lo": float(statistics.mean(lo)),
            "delta_means": float(statistics.mean(hi) - statistics.mean(lo)),
            "benefit_rate_hi": float(statistics.mean([1 if x > 0 else 0 for x in hi])),
            "benefit_rate_lo": float(statistics.mean([1 if x > 0 else 0 for x in lo])),
        }

    cand_a = []
    for thr in (10, 20, 30, 40):
        r = split_mean(rows_a, "ret", thr)
        if r:
            cand_a.append(r)
    for thr in (30, 45, 60):
        r = split_mean(rows_a, "mfe", thr)
        if r:
            cand_a.append(r)
    for thr in (0.4, 0.5, 0.6):
        r = split_mean(rows_a, "gb_frac", thr, direction="le")
        if r:
            cand_a.append(r)
    for thr in (0.0, 0.1):
        r = split_mean(rows_a, "imb", thr)
        if r:
            cand_a.append(r)
    best_a = max(cand_a, key=lambda x: abs(x["delta_means"])) if cand_a else None

    # B: early state@30 vs benefit of oracle early exit at 30 vs 600
    rows_b = []
    for b in bundles:
        st = b["states"].get(30) or {}
        if not st.get("ok") or b.get("ret600") is None:
            continue
        ex30 = exit_at_horizon(b["path"], 30.0)
        if not ex30.get("ok"):
            continue
        d = float(ex30["exit_ret_bps"]) - float(b["ret600"])
        rows_b.append({
            "mae": float(st.get("mae") or 0),
            "ret": float(st.get("ret") or 0),
            "mfe": float(st.get("mfe") or 0),
            "sell": 1.0 if st.get("sell_persist") else 0.0,
            "imb": float(st["imbalance"]) if st.get("imbalance") is not None else 0.0,
            "delta": d,
        })
    cand_b = []
    for thr in (-20, -30, -40, -50):
        r = split_mean(rows_b, "mae", thr, direction="le")
        if r:
            cand_b.append(r)
    for thr in (-10, -20, -30):
        r = split_mean(rows_b, "ret", thr, direction="le")
        if r:
            cand_b.append(r)
    best_b = max(cand_b, key=lambda x: abs(x["delta_means"])) if cand_b else None

    # crude separability: same-direction across day halves
    def fold_consistent(rows, split, benefit_side="hi"):
        if not split:
            return False
        days = sorted({b["date"] for b in bundles})
        mid = len(days) // 2
        d1, d2 = set(days[:mid]), set(days[mid:])
        # rebuild using date from bundles — attach date into rows via index alignment is hard;
        # approximate: overall signal only + require both means positive advantage on hi vs lo
        if benefit_side == "hi":
            return split["mean_hi"] > split["mean_lo"] + 1.0 and split["benefit_rate_hi"] >= split["benefit_rate_lo"]
        return split["mean_lo"] > split["mean_hi"] + 1.0

    a_sep = bool(best_a and best_a["delta_means"] > 2.0 and best_a["benefit_rate_hi"] > best_a["benefit_rate_lo"] + 0.05)
    b_sep = bool(best_b and best_b["mean_hi"] > best_b["mean_lo"] + 2.0)  # for mae le, hi=more negative mae group

    # Fix B interpretation: direction le → hi bucket is mae<=thr (worse MAE) should benefit more from early exit
    if best_b and best_b["direction"] == "le":
        b_sep = best_b["mean_hi"] > best_b["mean_lo"] + 2.0 and best_b["benefit_rate_hi"] >= best_b["benefit_rate_lo"]

    return {
        "n_a": len(rows_a),
        "n_b": len(rows_b),
        "best_continuation_split": best_a,
        "best_early_split": best_b,
        "question_a_separable": a_sep,
        "question_b_separable": b_sep,
        "candidates_a": cand_a[:12],
        "candidates_b": cand_b[:12],
        "mean_delta_750_vs_600_bps": float(statistics.mean([r["delta"] for r in rows_a])) if rows_a else None,
        "frac_750_helps": float(statistics.mean([r["benefit"] for r in rows_a])) if rows_a else None,
    }


def select_on_train(
    train: list[dict[str, Any]],
    guards: list[dict[str, Any]],
    cont_rules: list[dict[str, Any]],
) -> dict[str, Any]:
    """Inner selection of arch C/D/E components. A/B are fixed."""
    base = summarize_pnls([_pnl(b["fill_price"], float(b["ret600"] or 0)) for b in train])
    n_days = len({b["date"] for b in train})

    # pick best guard for C (guard+600)
    best_guard = None
    best_g_sc = -1e99
    best_g_ev = None
    for g in guards:
        ev = guard_economics(train, g)
        if not support_ok(
            ev["guard_trigger_n"], ev["guard_days"], ev["guard_symbols"],
            n_train=len(train), n_days=n_days,
        ):
            continue
        # require saved/lost ratio not terrible
        ratio = ev["saved_lost_ratio"]
        if ratio is not None and ratio < 0.8 and ev["foregone_winner_yen"] > 20000:
            continue
        sc = composite(ev["summary"], base, complexity=1.2)
        sc += (ev["saved_loss_yen"] - ev["foregone_winner_yen"]) / 2000.0
        if sc > best_g_sc:
            best_g_sc = sc
            best_guard = g
            best_g_ev = ev

    # continuation rules + learned (train survivors only for model fit)
    learned = fit_continuation_models(train)
    all_cont = list(cont_rules) + learned

    best_cont = None
    best_c_sc = -1e99
    # For E selection need a guard — use best_guard or None
    g_for_e = best_guard
    for rule in all_cont:
        ev = eval_policy_isolation(train, arch="E", guard=g_for_e, cont_rule=rule)
        # extension support
        if rule.get("kind") not in ("never",) and ev["extension_n"] < max(8, int(0.05 * len(train))):
            if rule.get("kind") != "always":
                continue
        sc = composite(ev["summary"], base, complexity=1.5 if rule.get("kind") == "learned" else 1.0)
        if sc > best_c_sc:
            best_c_sc = sc
            best_cont = rule

    # pick architecture among A–E on train
    cands = []
    for arch, guard, cont, cx in (
        ("A", None, None, 0.0),
        ("B", None, None, 0.3),
        ("C", best_guard, None, 1.0),
        ("D", best_guard, None, 1.1),
        ("E", best_guard, best_cont, 1.4),
    ):
        if arch in ("C", "D", "E") and best_guard is None and arch != "A":
            # still allow B; skip guard arches if no guard
            if arch != "B":
                continue
        ev = eval_policy_isolation(train, arch=arch, guard=guard, cont_rule=cont)
        sc = composite(ev["summary"], base, complexity=cx)
        if arch == "A":
            sc += 30.0  # simplicity prior
        cands.append({"arch": arch, "score": sc, "ev": ev, "guard": guard, "cont": cont})

    winner = max(cands, key=lambda x: x["score"])
    return {
        "selected_arch": winner["arch"],
        "guard": best_guard,
        "cont_rule": best_cont,
        "guard_train_ev": {k: v for k, v in (best_g_ev or {}).items() if k != "pnls"},
        "candidates": [
            {"arch": c["arch"], "score": c["score"], "total": c["ev"]["summary"]["total"],
             "delta_pnl": c["ev"]["delta_pnl"], "worst": c["ev"]["summary"]["worst"]}
            for c in cands
        ],
        "winner_ev": {k: v for k, v in winner["ev"].items() if k != "pnls"},
    }


def patch_and_sim(
    panel: list[dict],
    bundles_by_key: dict[tuple, dict],
    *,
    arch: str,
    guard: Optional[dict[str, Any]],
    cont_rule: Optional[dict[str, Any]],
    sfn,
) -> dict[str, Any]:
    evs = [dict(e) for e in panel]
    reasons = Counter()
    for e in evs:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        key = (e["date"], e["symbol"], float(e["fill_time"]))
        b = bundles_by_key.get(key)
        if not b:
            continue
        pol = apply_architecture(b, arch=arch, guard=guard, cont_rule=cont_rule)
        if not pol.get("ok"):
            continue
        e["canonical_exit_time"] = pol["exit_time"]
        e["canonical_exit_ret_bps"] = pol["exit_ret_bps"]
        e["canonical_hold_sec"] = pol["exit_off"]
        e["canonical_exit_reason"] = pol.get("reason")
        e["FIXED600_NET_BPS"] = pol["exit_ret_bps"]
        reasons[pol.get("reason")] += 1
    sim = simulate_joint(evs, score_fn=sfn)
    acc = [e for e in sim["events"] if e.get("accepted")]
    pnls = [float(e.get("realized_pnl_yen") or 0) for e in acc]
    return {
        "summary": summarize_pnls(pnls),
        "accepted_n": len(acc),
        "policy_counts": dict(reasons),
        "events": sim["events"],
    }


def slim_rule(rule: Optional[dict[str, Any]]) -> Optional[dict[str, Any]]:
    if rule is None:
        return None
    out = {k: v for k, v in rule.items() if k != "model"}
    if rule.get("kind") == "learned":
        out["model_kind"] = (rule.get("model") or {}).get("kind")
    return out


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_exit_v2_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)
    assert load_v1r().get("sha256") == V1R_SHA
    sfn = _sfn()

    cache = pickle.load(PANEL_CACHE.open("rb"))
    am_panel = [dict(e) for e in cache["am"]["panel"]]
    assert all(str(e["date"]) < FORBIDDEN for e in am_panel)

    print("  baseline FIXED600...", flush=True)
    base_sim = simulate_joint([dict(e) for e in am_panel], score_fn=sfn)
    accepted = [e for e in base_sim["events"] if e.get("accepted")]
    base_sum = summarize_pnls([float(e.get("realized_pnl_yen") or 0) for e in accepted])
    print(f"  fills={len(accepted)} PnL={base_sum['total']:.0f} PF={base_sum['pf']}", flush=True)

    pairs = sorted({(e["date"], e["symbol"]) for e in accepted})
    print(f"  load boards {len(pairs)}...", flush=True)
    boards = load_boards_for_symbols(pairs)

    print("  reconstruct states...", flush=True)
    bundles: list[dict[str, Any]] = []
    by_key: dict[tuple, dict] = {}
    early_rows = []
    for e in accepted:
        board = boards.get((e["date"], e["symbol"]))
        if board is None or board["t"].size == 0:
            continue
        sess_end = session_end_epoch(e["date"], e["session"])
        path = build_path(
            board,
            entry_price=float(e["fill_price"]),
            entry_t=float(e["fill_time"]),
            sess_end=sess_end,
        )
        b = build_trade_bundle(e, path, board)
        b["session"] = e.get("session")
        b["anchor_hhmm"] = signal_hhmm(float(e.get("signal_time") or e["fill_time"]))
        b["fixed600_pnl"] = _pnl(b["fill_price"], float(b["ret600"] or 0))
        b["fixed750_pnl"] = _pnl(b["fill_price"], float(b["ret750"] or 0))
        key = (b["date"], b["symbol"], b["fill_time"])
        bundles.append(b)
        by_key[key] = b
        st30 = b["states"].get(30) or {}
        early_rows.append({
            "date": b["date"], "symbol": b["symbol"],
            "ret30": st30.get("ret"), "mae30": st30.get("mae"), "mfe30": st30.get("mfe"),
            "imb30": st30.get("imbalance"), "sell30": st30.get("sell_persist"),
            "ret600": b["ret600"], "ret750": b["ret750"],
            "delta_750_vs_600_bps": b["delta_750_vs_600_bps"],
            "pnl600": b["fixed600_pnl"], "pnl750": b["fixed750_pnl"],
        })
    print(f"  bundles={len(bundles)}", flush=True)

    guards = generate_guards()
    cont_rules = generate_continuation_rules()
    print(f"  guards={len(guards)} continuation_rules={len(cont_rules)}", flush=True)

    # Heterogeneity
    print("  heterogeneity tests...", flush=True)
    het = heterogeneity_tests(bundles)
    print(
        f"  Q_A separable={het['question_a_separable']} "
        f"Q_B separable={het['question_b_separable']} "
        f"frac_750_helps={het['frac_750_helps']}",
        flush=True,
    )

    # Guard economics (full hist reporting; selection still nested)
    print("  guard economics...", flush=True)
    guard_rows = []
    for i, g in enumerate(guards):
        if i % 25 == 0:
            print(f"    guard {i}/{len(guards)}", flush=True)
        ev = guard_economics(bundles, g)
        guard_rows.append({
            "id": g["id"], "family": g["family"],
            "total": ev["summary"]["total"], "pf": ev["summary"]["pf"],
            "worst": ev["summary"]["worst"], "max_dd": ev["summary"]["max_dd"],
            "delta_pnl": ev["delta_pnl"], "delta_worst": ev["delta_worst"],
            "trigger_n": ev["guard_trigger_n"], "trigger_days": ev["guard_days"],
            "trigger_symbols": ev["guard_symbols"],
            "median_trigger_off": ev["median_trigger_off"],
            "winners_cut": ev["winners_cut"], "losers_saved": ev["losers_saved"],
            "saved_loss_yen": ev["saved_loss_yen"],
            "foregone_winner_yen": ev["foregone_winner_yen"],
            "saved_lost_ratio": ev["saved_lost_ratio"],
        })
    guard_rows_sorted = sorted(guard_rows, key=lambda r: -(r["saved_loss_yen"] - r["foregone_winner_yen"]))

    # Continuation rule isolation (no guard) as E with NEVER/ALWAYS etc — reporting
    print("  continuation isolation...", flush=True)
    cont_rows = []
    for rule in cont_rules:
        ev = eval_policy_isolation(bundles, arch="E", guard=None, cont_rule=rule)
        cont_rows.append({
            "id": rule["id"], "kind": rule["kind"],
            "total": ev["summary"]["total"], "pf": ev["summary"]["pf"],
            "worst": ev["summary"]["worst"], "delta_pnl": ev["delta_pnl"],
            "extension_n": ev["extension_n"], "extension_win": ev["extension_win"],
            "extension_loss": ev["extension_loss"],
            "extension_incremental_pnl": ev["extension_incremental_pnl"],
        })

    # Nested outer CV
    print("  nested outer CV...", flush=True)
    outer_rows = []
    outer_selected = []
    fold_specs = []
    for block, days in OUTER_BLOCKS.items():
        test_days = set(days)
        train_days = {d for ds in OUTER_BLOCKS.values() for d in ds} - test_days
        train = [b for b in bundles if b["date"] in train_days]
        test = [b for b in bundles if b["date"] in test_days]
        sel = select_on_train(train, guards, cont_rules)
        # evaluate selected on test; also evaluate all arches with train-chosen params
        fold_result = {"block": block, "selected": sel["selected_arch"], "train_n": len(train), "test_n": len(test)}
        arch_tests = {}
        for arch in ("A", "B", "C", "D", "E"):
            g = sel["guard"] if arch in ("C", "D", "E") else None
            c = sel["cont_rule"] if arch == "E" else None
            if arch in ("C", "D", "E") and g is None:
                # fall back to A metrics for missing guard
                ev = eval_policy_isolation(test, arch="A")
            else:
                ev = eval_policy_isolation(test, arch=arch, guard=g, cont_rule=c)
            base_te = summarize_pnls([b["fixed600_pnl"] for b in test])
            arch_tests[arch] = {
                "total": ev["summary"]["total"],
                "base_total": base_te["total"],
                "delta_pnl": ev["delta_pnl"],
                "delta_pf": ev["delta_pf"],
                "delta_worst": ev["delta_worst"],
                "delta_dd": ev["delta_dd"],
                "pf": ev["summary"]["pf"],
                "worst": ev["summary"]["worst"],
                "max_dd": ev["summary"]["max_dd"],
                "winners_cut": ev["winners_cut"],
                "losers_saved": ev["losers_saved"],
                "saved_loss_yen": ev["saved_loss_yen"],
                "foregone_winner_yen": ev["foregone_winner_yen"],
                "extension_n": ev["extension_n"],
            }
        fold_result.update(arch_tests[sel["selected_arch"]])
        fold_result["arch_tests"] = arch_tests
        fold_result["guard_id"] = (sel["guard"] or {}).get("id")
        fold_result["cont_id"] = (sel["cont_rule"] or {}).get("id")
        outer_rows.append(fold_result)
        outer_selected.append(sel["selected_arch"])
        fold_specs.append(sel)
        print(
            f"    fold {block}: selected={sel['selected_arch']} "
            f"guard={fold_result['guard_id']} cont={fold_result['cont_id']} "
            f"delta={fold_result['delta_pnl']:.0f}",
            flush=True,
        )

    # Heterogeneity on outer: state×treatment same direction?
    # Re-check Q_A/Q_B per fold using train-fit split applied to test
    het_fold = []
    for block, days in OUTER_BLOCKS.items():
        test_days = set(days)
        train_days = {d for ds in OUTER_BLOCKS.values() for d in ds} - test_days
        tr = [b for b in bundles if b["date"] in train_days]
        te = [b for b in bundles if b["date"] in test_days]
        htr = heterogeneity_tests(tr)
        hte = heterogeneity_tests(te)
        same_a = False
        if htr.get("best_continuation_split") and hte.get("best_continuation_split"):
            # same feature key & direction of mean advantage
            a1, a2 = htr["best_continuation_split"], hte["best_continuation_split"]
            same_a = (a1["key"] == a2["key"] and (a1["delta_means"] > 0) == (a2["delta_means"] > 0))
        same_b = False
        if htr.get("best_early_split") and hte.get("best_early_split"):
            b1, b2 = htr["best_early_split"], hte["best_early_split"]
            same_b = (b1["key"] == b2["key"] and (b1["delta_means"] > 0) == (b2["delta_means"] > 0))
        het_fold.append({"block": block, "same_dir_A": same_a, "same_dir_B": same_b,
                         "train_A": htr.get("question_a_separable"), "test_A": hte.get("question_a_separable"),
                         "train_B": htr.get("question_b_separable"), "test_B": hte.get("question_b_separable")})

    # Final historical selection: majority vote on arch; params from full-data inner (dev only, retrospective)
    vote = Counter(outer_selected)
    voted_arch = vote.most_common(1)[0][0]
    full_sel = select_on_train(bundles, guards, cont_rules)
    # Prefer outer vote for arch; params from full_sel only if same arch family — but to avoid leak,
    # freeze params as mode of fold specs when possible
    guard_votes = Counter((s["guard"] or {}).get("id") for s in fold_specs if s.get("guard"))
    cont_votes = Counter((s["cont_rule"] or {}).get("id") for s in fold_specs if s.get("cont_rule"))
    frozen_guard = None
    if guard_votes:
        gid = guard_votes.most_common(1)[0][0]
        frozen_guard = next(g for g in guards if g["id"] == gid)
    elif full_sel.get("guard"):
        frozen_guard = full_sel["guard"]
    frozen_cont = None
    if cont_votes:
        cid = cont_votes.most_common(1)[0][0]
        frozen_cont = next((r for r in cont_rules if r["id"] == cid), None)
        if frozen_cont is None:
            # learned — take from full_sel if matching id
            if (full_sel.get("cont_rule") or {}).get("id") == cid:
                frozen_cont = full_sel["cont_rule"]
    elif full_sel.get("cont_rule"):
        frozen_cont = full_sel["cont_rule"]

    # Outer acceptance for voted arch with frozen params
    outer_deltas = []
    catastrophic = False
    for row in outer_rows:
        # re-eval voted arch with frozen params on that fold's test
        test = [b for b in bundles if b["date"] in set(OUTER_BLOCKS[row["block"]])]
        g = frozen_guard if voted_arch in ("C", "D", "E") else None
        c = frozen_cont if voted_arch == "E" else None
        if voted_arch in ("C", "D", "E") and g is None:
            ev = eval_policy_isolation(test, arch="A")
        else:
            ev = eval_policy_isolation(test, arch=voted_arch, guard=g, cont_rule=c)
        outer_deltas.append(ev["delta_pnl"])
        if ev["delta_pnl"] <= CATASTROPHIC_FOLD_YEN:
            catastrophic = True
        row["voted_arch_delta"] = ev["delta_pnl"]
        row["voted_arch_worst"] = ev["summary"]["worst"]
        row["voted_arch_pf"] = ev["summary"]["pf"]

    mean_delta = float(statistics.mean(outer_deltas)) if outer_deltas else 0.0
    # also compare architecture E vs D vs A mean deltas from arch_tests
    mean_by_arch = {}
    for arch in ("A", "B", "C", "D", "E"):
        mean_by_arch[arch] = float(statistics.mean([r["arch_tests"][arch]["delta_pnl"] for r in outer_rows]))

    cross_ok = (
        mean_delta >= 0
        and not catastrophic
        and all(r.get("voted_arch_pf") is None or r["voted_arch_pf"] >= 0.5 for r in outer_rows)
    )

    # Isolation for architectures with frozen params
    print("  isolation architectures...", flush=True)
    iso = {}
    for arch in ("A", "B", "C", "D", "E"):
        g = frozen_guard if arch in ("C", "D", "E") else None
        c = frozen_cont if arch == "E" else None
        if arch in ("C", "D", "E") and g is None:
            iso[arch] = eval_policy_isolation(bundles, arch="A")
        else:
            iso[arch] = eval_policy_isolation(bundles, arch=arch, guard=g, cont_rule=c)
        print(
            f"    {arch}: PnL={iso[arch]['summary']['total']:.0f} "
            f"delta={iso[arch]['delta_pnl']:.0f} worst={iso[arch]['summary']['worst']:.0f}",
            flush=True,
        )

    # Full replay SoT
    print("  full replay...", flush=True)
    full = {}
    for arch in ("A", "B", "C", "D", "E"):
        g = frozen_guard if arch in ("C", "D", "E") else None
        c = frozen_cont if arch == "E" else None
        if arch in ("C", "D", "E") and g is None:
            res = patch_and_sim(am_panel, by_key, arch="A", guard=None, cont_rule=None, sfn=sfn)
        else:
            res = patch_and_sim(am_panel, by_key, arch=arch, guard=g, cont_rule=c, sfn=sfn)
        full[arch] = {
            "accepted_n": res["accepted_n"],
            "summary": res["summary"],
            "delta_pnl": res["summary"]["total"] - base_sum["total"],
            "delta_pf": (res["summary"]["pf"] or 0) - (base_sum["pf"] or 0),
            "delta_worst": res["summary"]["worst"] - base_sum["worst"],
            "delta_dd": res["summary"]["max_dd"] - base_sum["max_dd"],
            "policy_counts": res["policy_counts"],
            "events": res["events"],
        }
        print(f"    full {arch}: PnL={res['summary']['total']:.0f} n={res['accepted_n']}", flush=True)

    # Decide freeze candidate
    state_edge = bool(het["question_a_separable"] or het["question_b_separable"])
    # outer fold winner disagreement → instability signal
    winner_unstable = len(set(outer_selected)) >= 3 or (
        len(set(outer_selected)) >= 2 and "A" in outer_selected and mean_delta < 0
    )
    # risk pareto: C or E improves worst/dd with limited pnl damage
    risk_pareto = False
    risk_arch = None
    for arch in ("C", "E"):
        fr = full[arch]["summary"]
        if (
            fr["worst"] >= base_sum["worst"] + 20000
            and fr["max_dd"] >= base_sum["max_dd"] - 1
            and fr["total"] >= base_sum["total"] * 0.85
            and mean_by_arch.get(arch, -1e9) >= -50_000
            and not any(
                r["arch_tests"][arch]["delta_pnl"] <= CATASTROPHIC_FOLD_YEN for r in outer_rows
            )
        ):
            risk_pareto = True
            risk_arch = arch
            break

    historically_supported = (
        voted_arch != "A"
        and cross_ok
        and state_edge
        and full[voted_arch]["summary"]["total"] >= base_sum["total"] - 1
        and full[voted_arch]["summary"]["worst"] >= base_sum["worst"] - 1
        and not winner_unstable
    )

    if historically_supported:
        verdict = "V1R_EXIT_V2_ASYMMETRIC_CANDIDATE_HISTORICALLY_SUPPORTED"
        frozen_arch = voted_arch
    elif risk_pareto and state_edge:
        verdict = "V1R_EXIT_V2_RISK_PARETO_ONLY"
        frozen_arch = risk_arch or "A"
    elif state_edge and not cross_ok:
        verdict = "V1R_EXIT_V2_STATE_EDGE_NOT_ROBUST"
        frozen_arch = "A"
    else:
        verdict = "V1R_EXIT_V2_NO_STATE_DEPENDENT_EDGE"
        frozen_arch = "A"

    # If H0 wins, freeze FIXED600
    if verdict in (
        "V1R_EXIT_V2_NO_STATE_DEPENDENT_EDGE",
        "V1R_EXIT_V2_STATE_EDGE_NOT_ROBUST",
    ):
        frozen_arch = "A"

    freeze_body = {
        "manifest_id": "V1R_EXIT_V2_ASYMMETRIC_CANDIDATE_V1",
        "kind": "research_exit_candidate_not_production",
        "status": "HISTORICAL_EVIDENCE_ONLY",
        "entry_frozen": True,
        "v1r_sha": V1R_SHA,
        "model_sha": MODEL_ARTIFACT_SHA,
        "frozen_architecture": frozen_arch,
        "voted_architecture": voted_arch,
        "guard": slim_rule(frozen_guard) if frozen_arch in ("C", "D", "E") else None,
        "continuation": slim_rule(frozen_cont) if frozen_arch == "E" else None,
        "outer_vote": dict(vote),
        "outer_mean_delta_pnl": mean_delta,
        "mean_delta_by_arch": mean_by_arch,
        "catastrophic_fold": catastrophic,
        "cross_fit_gate": cross_ok,
        "heterogeneity": {
            "question_a": het["question_a_separable"],
            "question_b": het["question_b_separable"],
        },
        "selection_basis": "pre_20260810_retrospective_development_only",
        "reference_0810": "RETROSPECTIVE_REFERENCE_ONLY",
        "production_mutation": "NONE",
    }
    freeze_sha = hashlib.sha256(
        json.dumps({k: v for k, v in freeze_body.items() if k != "sha256"}, sort_keys=True, default=str).encode()
    ).hexdigest()
    freeze_body["sha256"] = freeze_sha
    (OUT / "V1R_EXIT_V2_ASYMMETRIC_CANDIDATE_V1.json").write_text(
        json.dumps(freeze_body, indent=2, ensure_ascii=False, default=str), encoding="utf-8"
    )
    print(f"  FROZEN arch={frozen_arch} sha={freeze_sha[:16]}... verdict={verdict}", flush=True)

    # LODO / LOSO / Session / Anchor on frozen arch
    print("  stability LODO/LOSO/Session/Anchor...", flush=True)
    g = frozen_guard if frozen_arch in ("C", "D", "E") else None
    c = frozen_cont if frozen_arch == "E" else None

    lodo = []
    for day in sorted({b["date"] for b in bundles}):
        sub_panel = [e for e in am_panel if e["date"] != day]
        if frozen_arch == "A":
            lodo.append({"left_out": day, "delta": 0.0})
            continue
        res = patch_and_sim(sub_panel, by_key, arch=frozen_arch, guard=g, cont_rule=c, sfn=sfn)
        base = patch_and_sim(sub_panel, by_key, arch="A", guard=None, cont_rule=None, sfn=sfn)
        lodo.append({
            "left_out": day,
            "cand_total": res["summary"]["total"],
            "base_total": base["summary"]["total"],
            "delta": res["summary"]["total"] - base["summary"]["total"],
        })

    by_sym = defaultdict(float)
    for b in bundles:
        by_sym[b["symbol"]] += b["fixed600_pnl"]
    top_syms = [s for s, _ in sorted(by_sym.items(), key=lambda x: -x[1])[:6]]
    if "285A" not in top_syms:
        top_syms.append("285A")
    loso = []
    for sym in top_syms:
        sub_panel = [e for e in am_panel if e["symbol"] != sym]
        if frozen_arch == "A":
            loso.append({"excluded": sym, "delta": 0.0})
            continue
        res = patch_and_sim(sub_panel, by_key, arch=frozen_arch, guard=g, cont_rule=c, sfn=sfn)
        base = patch_and_sim(sub_panel, by_key, arch="A", guard=None, cont_rule=None, sfn=sfn)
        loso.append({
            "excluded": sym,
            "cand_total": res["summary"]["total"],
            "base_total": base["summary"]["total"],
            "delta": res["summary"]["total"] - base["summary"]["total"],
        })

    session_rows = []
    for label, pred in (
        ("AM", lambda b: str(b.get("session") or "").upper() in ("AM", "MORNING", "A") or True),
        ("PM", lambda b: str(b.get("session") or "").upper() in ("PM", "AFTERNOON", "P")),
    ):
        sub = [b for b in bundles if pred(b)]
        if label == "AM":
            sub = bundles  # panel is AM-primary
        if label == "PM":
            sub = [b for b in bundles if str(b.get("session") or "").upper().startswith("P")]
        if not sub:
            session_rows.append({"session": label, "n": 0})
            continue
        ev = eval_policy_isolation(sub, arch=frozen_arch if frozen_arch != "A" else "A", guard=g, cont_rule=c)
        session_rows.append({
            "session": label, "n": ev["summary"]["n"], "total": ev["summary"]["total"],
            "delta_pnl": ev["delta_pnl"] if frozen_arch != "A" else 0.0,
        })

    anchor_rows = []
    by_anchor = defaultdict(list)
    for b in bundles:
        by_anchor[b["anchor_hhmm"]].append(b)
    for hhmm, sub in sorted(by_anchor.items()):
        ev = eval_policy_isolation(sub, arch=frozen_arch if frozen_arch != "A" else "A", guard=g, cont_rule=c)
        base = summarize_pnls([x["fixed600_pnl"] for x in sub])
        anchor_rows.append({
            "anchor_hhmm": hhmm, "n": len(sub),
            "cand_total": ev["summary"]["total"], "base_total": base["total"],
            "delta": ev["delta_pnl"] if frozen_arch != "A" else 0.0,
        })

    # 8/10 reference ONLY after freeze
    print("  8/10 retrospective reference...", flush=True)
    from research.e1_x34c_passive_deployability.events import build_events
    from research.e1_x36_joint_allocator.panel import enrich_events
    from small_paper.v1r_day_engine import (
        _load_boards,
        _planned_anchors_retrospective,
        resolve_pre0905_am_universe,
        score_fn_frozen,
    )

    uni = resolve_pre0905_am_universe("20260810")
    planned = _planned_anchors_retrospective("20260810", uni["symbols"])
    boards10 = _load_boards([("20260810", s) for s in uni["symbols"]])
    panel10 = enrich_events(build_events(planned, boards10), boards10)
    sim10 = simulate_joint([dict(e) for e in panel10], score_fn=score_fn_frozen())
    acc10 = [e for e in sim10["events"] if e.get("accepted")]
    by10: dict[tuple, dict] = {}
    for e in acc10:
        board = boards10.get((e["date"], e["symbol"]))
        if board is None:
            continue
        se = session_end_epoch(e["date"], e["session"])
        path = build_path(board, entry_price=float(e["fill_price"]), entry_t=float(e["fill_time"]), sess_end=se)
        b = build_trade_bundle(e, path, board)
        b["fixed600_pnl"] = _pnl(b["fill_price"], float(b["ret600"] or 0))
        by10[(b["date"], b["symbol"], b["fill_time"])] = b

    ref_rows = []
    for arch in ("A", "B", "C", "D", "E", frozen_arch):
        gg = frozen_guard if arch in ("C", "D", "E") else None
        cc = frozen_cont if arch == "E" else None
        if arch in ("C", "D", "E") and gg is None:
            continue
        res = patch_and_sim(panel10, by10, arch=arch, guard=gg, cont_rule=cc, sfn=score_fn_frozen())
        ref_rows.append({
            "arch": arch,
            "classification": "RETROSPECTIVE_REFERENCE_ONLY",
            "accepted_n": res["accepted_n"],
            "total": res["summary"]["total"],
            "pf": res["summary"]["pf"],
            "worst": res["summary"]["worst"],
            "max_dd": res["summary"]["max_dd"],
            "policy_counts": res["policy_counts"],
        })

    # 5706 reference
    base10 = next(r for r in ref_rows if r["arch"] == "A")
    # find worst from events
    a_events = patch_and_sim(panel10, by10, arch="A", guard=None, cont_rule=None, sfn=score_fn_frozen())["events"]
    worst = None
    for e in a_events:
        if not e.get("accepted"):
            continue
        pnl = float(e.get("realized_pnl_yen") or 0)
        if worst is None or pnl < worst["pnl"]:
            worst = {"symbol": e["symbol"], "fill_time": float(e["fill_time"]), "pnl": pnl, "fill_price": float(e["fill_price"])}
    ref5706 = []
    if worst:
        key = ("20260810", worst["symbol"], worst["fill_time"])
        b = by10.get(key)
        for arch in ("A", "B", "C", "D", "E"):
            if b is None:
                continue
            gg = frozen_guard if arch in ("C", "D", "E") else None
            cc = frozen_cont if arch == "E" else None
            if arch in ("C", "D", "E") and gg is None:
                continue
            pol = apply_architecture(b, arch=arch, guard=gg, cont_rule=cc)
            if not pol.get("ok"):
                continue
            pnl = _pnl(b["fill_price"], float(pol["exit_ret_bps"]))
            ref5706.append({
                "arch": arch,
                "symbol": worst["symbol"],
                "triggered_guard": pol.get("triggered_guard"),
                "extended": pol.get("extended"),
                "reason": pol.get("reason"),
                "exit_off": pol.get("exit_off"),
                "exit_ret_bps": pol.get("exit_ret_bps"),
                "exit_pnl": pnl,
                "baseline_pnl": worst["pnl"],
                "avoided_loss": pnl - worst["pnl"],
                "executable_buy1": b["fill_price"] * (1.0 + float(pol["exit_ret_bps"]) / 10000.0),
                "classification": "RETROSPECTIVE_REFERENCE_ONLY",
            })

    # Best guard economics for answers
    best_guard_econ = guard_rows_sorted[0] if guard_rows_sorted else {}
    iso_c = iso.get("C", iso["A"])
    # H0 vs H1 interpretation
    if winner_unstable and not (het["question_a_separable"] and any(h["same_dir_A"] for h in het_fold)):
        instability_vs_regime = "H0_INSTABILITY"
    elif state_edge and any(h["same_dir_A"] or h["same_dir_B"] for h in het_fold):
        instability_vs_regime = "H1_PARTIAL_STATE_EDGE"
    else:
        instability_vs_regime = "H0_INSTABILITY"

    # Pareto sheet
    pareto = []
    for arch, lab in (("A", "baseline"), ("B", "time750"), ("C", "guard600"), ("D", "guard750"), ("E", "guard_cont")):
        sm = full[arch]["summary"]
        pareto.append({
            "arch": arch, "label": lab,
            "total": sm["total"], "pf": sm["pf"], "worst": sm["worst"], "max_dd": sm["max_dd"],
            "outer_mean_delta": mean_by_arch.get(arch),
            "role": "frozen" if arch == frozen_arch else ("risk_pareto" if arch == risk_arch else "compared"),
        })

    overview = [{
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": verdict,
        "instability_vs_regime": instability_vs_regime,
        "frozen_arch": frozen_arch,
        "freeze_sha": freeze_sha,
        "baseline_pnl": base_sum["total"],
        "baseline_pf": base_sum["pf"],
        "baseline_worst": base_sum["worst"],
        "n_fills": len(bundles),
        "n_guards": len(guards),
        "n_cont_rules": len(cont_rules),
        "question_a": het["question_a_separable"],
        "question_b": het["question_b_separable"],
        "outer_mean_delta": mean_delta,
        "catastrophic_fold": catastrophic,
        "cross_fit_gate": cross_ok,
        "production_mutation": "NONE",
        "data_status": "RETROSPECTIVE_DEVELOPMENT",
        "ref_0810": "RETROSPECTIVE_REFERENCE_ONLY",
    }]

    def iso_row(arch: str) -> dict:
        ev = iso[arch]
        return {
            "arch": arch,
            "total": ev["summary"]["total"], "pf": ev["summary"]["pf"],
            "worst": ev["summary"]["worst"], "max_dd": ev["summary"]["max_dd"],
            "gross_loss": ev["summary"]["gross_loss"],
            "delta_pnl": ev["delta_pnl"], "avg_hold": ev["avg_hold"],
            "winners_cut": ev["winners_cut"], "losers_saved": ev["losers_saved"],
            "saved_loss_yen": ev["saved_loss_yen"],
            "foregone_winner_yen": ev["foregone_winner_yen"],
            "saved_lost_ratio": ev["saved_lost_ratio"],
            "extension_n": ev["extension_n"],
            "extension_incremental_pnl": ev["extension_incremental_pnl"],
            "guard_trigger_n": ev["guard_trigger_n"],
        }

    sheets = {
        "Overview": overview,
        "Data_Status": [{
            "historical": "RETROSPECTIVE_DEVELOPMENT_DATA",
            "n_days": len({b["date"] for b in bundles}),
            "n_fills": len(bundles),
            "day_0810": "RETROSPECTIVE_REFERENCE_ONLY",
            "not_fresh_holdout": True,
            "not_used_for_selection": True,
        }],
        "Early_State": early_rows[:200],
        "Failure_Guards": [{"id": g["id"], "family": g["family"], "kind": g["kind"]} for g in guards],
        "Guard_Economics": guard_rows_sorted[:80],
        "State600": [
            {"date": b["date"], "symbol": b["symbol"], **features_at_600(b),
             "delta_750_vs_600_bps": b["delta_750_vs_600_bps"]}
            for b in bundles
        ][:200],
        "Continuation": sorted(cont_rows, key=lambda r: -r["total"])[:60],
        "Heterogeneity": [
            {"metric": "question_a_separable", "value": het["question_a_separable"]},
            {"metric": "question_b_separable", "value": het["question_b_separable"]},
            {"metric": "best_A", "value": het.get("best_continuation_split")},
            {"metric": "best_B", "value": het.get("best_early_split")},
            {"metric": "frac_750_helps", "value": het.get("frac_750_helps")},
            *[{"metric": f"fold_{h['block']}", "value": h} for h in het_fold],
        ],
        "Routers": [
            {"arch": "A", "desc": "FIXED600"},
            {"arch": "B", "desc": "TIME750"},
            {"arch": "C", "desc": "EarlyGuard+FIXED600", "guard": (frozen_guard or {}).get("id")},
            {"arch": "D", "desc": "EarlyGuard+uncondTIME750", "guard": (frozen_guard or {}).get("id")},
            {"arch": "E", "desc": "EarlyGuard+ContGate600/750",
             "guard": (frozen_guard or {}).get("id"), "cont": (frozen_cont or {}).get("id")},
        ],
        "InnerCV": [{"note": "inner=train-only guard/cont/arch selection per outer fold"}],
        "OuterCV": [
            {k: v for k, v in r.items() if k != "arch_tests"}
            | {f"{a}_{kk}": vv for a, ad in r.get("arch_tests", {}).items() for kk, vv in ad.items()
               if kk in ("delta_pnl", "worst", "pf")}
            for r in outer_rows
        ],
        "Isolation": [iso_row(a) for a in ("A", "B", "C", "D", "E")],
        "FullReplay": [
            {"arch": a, "accepted_n": full[a]["accepted_n"], **full[a]["summary"],
             "delta_pnl": full[a]["delta_pnl"], "delta_worst": full[a]["delta_worst"],
             "policy_counts": full[a]["policy_counts"]}
            for a in ("A", "B", "C", "D", "E")
        ],
        "LODO": lodo,
        "LOSO": loso,
        "Session": session_rows,
        "Anchor": anchor_rows,
        "Pareto": pareto,
        "FrozenCandidate": [freeze_body],
        "Reference0810": ref_rows,
        "Worst5706Ref": ref5706,
    }
    # fix OuterCV dict merge for older python — build explicitly
    outer_sheet = []
    for r in outer_rows:
        row = {k: v for k, v in r.items() if k != "arch_tests"}
        for a, ad in r.get("arch_tests", {}).items():
            for kk in ("delta_pnl", "worst", "pf"):
                row[f"{a}_{kk}"] = ad.get(kk)
        outer_sheet.append(row)
    sheets["OuterCV"] = outer_sheet

    write_xlsx(sheets, OUT / "v1r_exit_v2_asymmetric.xlsx")

    answers = {
        "1_instability_vs_regime": instability_vs_regime,
        "2_early_failure_identifiable": het["question_b_separable"],
        "3_saved_loss_yen": iso_c.get("saved_loss_yen"),
        "4_foregone_winner_yen": iso_c.get("foregone_winner_yen"),
        "5_saved_lost_ratio": iso_c.get("saved_lost_ratio"),
        "6_extension_state_exists": het["question_a_separable"],
        "7_identifiable_at_600": het["question_a_separable"] and any(h.get("same_dir_A") for h in het_fold),
        "8_vs_uncond_750": {
            "iso_B_delta": iso["B"]["delta_pnl"],
            "iso_E_delta": iso["E"]["delta_pnl"],
            "full_B_delta": full["B"]["delta_pnl"],
            "full_E_delta": full["E"]["delta_pnl"],
            "outer_mean_B": mean_by_arch.get("B"),
            "outer_mean_E": mean_by_arch.get("E"),
            "outer_mean_D": mean_by_arch.get("D"),
        },
        "9_vs_fixed600": {a: full[a]["delta_pnl"] for a in ("B", "C", "D", "E")},
        "10_isolation": {a: iso_row(a) for a in ("A", "B", "C", "D", "E")},
        "11_full_replay": {
            a: {"total": full[a]["summary"]["total"], "pf": full[a]["summary"]["pf"],
                "worst": full[a]["summary"]["worst"], "delta_pnl": full[a]["delta_pnl"]}
            for a in ("A", "B", "C", "D", "E")
        },
        "12_outer_folds": outer_sheet,
        "13_catastrophic_fold": catastrophic,
        "14_lodo": lodo,
        "15_loso": loso,
        "16_session": session_rows,
        "17_anchor": anchor_rows,
        "18_tail_risk": {
            "baseline_worst": base_sum["worst"],
            "baseline_dd": base_sum["max_dd"],
            "C_worst": full["C"]["summary"]["worst"],
            "E_worst": full["E"]["summary"]["worst"],
            "risk_pareto_arch": risk_arch,
        },
        "19_ref_0810": ref_rows,
        "20_ref_5706": ref5706,
        "21_historical_freeze_possible": historically_supported,
        "22_production_unchanged": True,
        "best_guard_econ": best_guard_econ,
        "heterogeneity": het,
        "het_fold": het_fold,
    }

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": verdict,
        "answers": answers,
        "baseline": base_sum,
        "frozen": freeze_body,
        "production_mutation": "NONE",
        "prospective_claim": "NONE — HISTORICAL_EVIDENCE_ONLY",
    }
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    md = f"""# V1R EXIT V2 Asymmetric Sequential EXIT

## Verdict
`{verdict}`

## H0 vs H1
`{instability_vs_regime}`

- Question A (600→750 state separable): {het['question_a_separable']}
- Question B (early failure state separable): {het['question_b_separable']}
- Outer winner vote: `{dict(vote)}`
- Outer mean ΔPnL (frozen params): {mean_delta:.0f}
- Catastrophic fold: {catastrophic}
- Cross-fit gate: {cross_ok}

## Frozen
- architecture: `{frozen_arch}`
- guard: `{(frozen_guard or {}).get('id')}`
- continuation: `{(frozen_cont or {}).get('id')}`
- sha: `{freeze_sha}`
- status: HISTORICAL_EVIDENCE_ONLY / not Production

## Full Replay vs FIXED600
{chr(10).join(f"- {a}: PnL={full[a]['summary']['total']:.0f} Δ={full[a]['delta_pnl']:.0f} worst={full[a]['summary']['worst']:.0f}" for a in ('A','B','C','D','E'))}

## Early Guard economics (isolation C)
- saved_loss: {iso_c.get('saved_loss_yen')}
- foregone_winner: {iso_c.get('foregone_winner_yen')}
- ratio: {iso_c.get('saved_lost_ratio')}

## 8/10
RETROSPECTIVE_REFERENCE_ONLY (not holdout pass; not used for selection).

## Production
NONE — FIXED600 unchanged. STOP.
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")
    print(json.dumps({"verdict": verdict, "frozen_arch": frozen_arch, "h0h1": instability_vs_regime}, indent=2))
    return report


if __name__ == "__main__":
    main()
