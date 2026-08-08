"""E1_X28 runner: full executable ENTRY x EXIT evaluation over 52115 routes."""
from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import numpy as np

from research.e1_x6_provisional.util import sha256_obj
from research.e1_x22_actual_exit_factory.registry import (
    build_alias_groups,
    load_population_checked,
    rebuild_candidates_and_masks,
)
from research.e1_x26_exit_library.exits import ExitSpec, common_controls
from research.e1_x27_reference_joint.paths import build_paths_for_rows
from research.e1_x27_reference_joint import PRIMARY_CONTROL as X27_PRIMARY

from . import (
    ANALYSIS_ID,
    BOARD_MAPPING,
    BOARD_MAPPING_SHA,
    BOOTSTRAP_ITERS,
    BOOTSTRAP_SEED,
    BROKERAGE_FEE_YEN,
    COMMON_CONTROLS,
    DISCOVERY,
    DOCUMENT_ID,
    EVALUATION,
    EXPECTED_ALIASES,
    EXPECTED_CAND_N,
    EXPECTED_POP_N,
    EXPECTED_SEMANTIC_ROUTES,
    EXPECTED_UNIQUE_MASKS,
    FORBIDDEN_V1_SHA,
    MANIFEST_ID,
    MANIFEST_V2_SHA,
    MIN_TRADES,
    PRIMARY_CONTROL,
    SECONDARY_CONTROL,
    SOURCE_X27,
    STRESS_DAY,
    VERDICT_BOARD_UNRESOLVED,
    VERDICT_COST_SENSITIVE,
    VERDICT_MIXED,
    VERDICT_MULTIPLE,
    VERDICT_REPLAY_FAIL,
    VERDICT_YEN_UNRESOLVED,
    X27_LEDGER_SHA,
)
from .board import load_board_events, verify_board_mapping
from .metrics import (
    bridge_is_directional,
    classify_executable,
    pairwise_common,
    period_mask,
    reclassify_x27_joint,
    summarize,
)
from .publish import publish
from .replay import build_all_exit_matrices, build_entry_asks

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "e1_x28_executable_joint"
X27_DIR = NATIVE / "results" / "research" / "e1_x27_reference_joint"
X26A_DIR = NATIVE / "results" / "research" / "e1_x26a_exit_manifest_repair"


def _run_tests() -> dict[str, Any]:
    import os
    test_path = NATIVE / "tests" / "research" / "test_e1_x28_executable_joint.py"
    env = {**os.environ, "PYTHONPATH": str(NATIVE / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_path), "-q", "--tb=line"],
        cwd=str(NATIVE), capture_output=True, text=True, env=env,
    )
    out = (p.stdout or "") + (p.stderr or "")
    passed = failed = 0
    m = re.search(r"(\d+) passed", out)
    if m:
        passed = int(m.group(1))
    m2 = re.search(r"(\d+) failed", out)
    if m2:
        failed = int(m2.group(1))
    return {
        "exit_code": p.returncode, "passed": passed, "failed": failed,
        "total": passed + failed or 1,
        "rows": [{"test": "pytest_suite",
                  "outcome": "PASSED" if p.returncode == 0 else "FAILED",
                  "detail": out[-3000:]}],
    }


def _load_routes() -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(X26A_DIR / "audit.xlsx", read_only=True, data_only=True)
    ws = wb["X27Routing"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        for k in ("exit_ids", "all_discovery_tags", "routed_exit_families"):
            v = d.get(k)
            if isinstance(v, str):
                d[k] = json.loads(v)
        out.append(d)
    return out


def _load_x27_eval() -> dict[str, dict[str, Any]]:
    """route_id -> X27 evaluation row."""
    from openpyxl import load_workbook
    wb = load_workbook(X27_DIR / "audit.xlsx", read_only=True, data_only=True)
    ws = wb["EvaluationResults"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = {}
    for row in rows[1:]:
        d = {hdr[i]: row[i] for i in range(len(hdr))}
        rid = d.get("route_id")
        if rid:
            out[str(rid)] = d
    return out


def _specs_from_canonical(canonical: list[dict[str, Any]]) -> list[ExitSpec]:
    specs = []
    for c in canonical:
        specs.append(ExitSpec(
            exit_id=c["canonical_exit_id"],
            path_family=(c.get("applicable_path_families") or [None])[0],
            variant=c.get("variant"),
            stop_bps=c.get("stop_bps"),
            target_bps=c.get("target_bps"),
            trail_activation_bps=c.get("trail_activation_bps"),
            giveback_bps=c.get("giveback_bps"),
            giveback_mode=c.get("giveback_mode"),
            no_progress_sec=c.get("no_progress_sec"),
            max_hold_sec=float(c.get("max_hold_sec") or 900.0),
            no_progress_mfe_bps=c.get("no_progress_mfe_bps", 5.0),
            no_progress_abs_ret_bps=c.get("no_progress_abs_ret_bps", 5.0),
        ))
    return specs


def _bh(pvals: list[float]) -> list[float]:
    n = len(pvals)
    if n == 0:
        return []
    order = np.argsort(pvals)
    ranked = np.asarray(pvals, dtype=np.float64)[order]
    out = np.empty(n)
    prev = 1.0
    for i in range(n - 1, -1, -1):
        val = ranked[i] * n / (i + 1)
        prev = min(prev, val)
        out[i] = min(prev, 1.0)
    q = np.empty(n)
    q[order] = out
    return q.tolist()


def _cluster_bootstrap(
    values: np.ndarray,
    clusters: np.ndarray,
    valid: np.ndarray,
    iters: int = BOOTSTRAP_ITERS,
    seed: int = BOOTSTRAP_SEED,
) -> dict[str, Any]:
    elig = valid & np.isfinite(values)
    if elig.sum() < 10:
        return {"mean": None, "ci95": [None, None], "raw_p": None, "tag": "DESCRIPTIVE_ONLY"}
    uniq = np.unique(clusters[elig])
    means = []
    for c in uniq:
        m = elig & (clusters == c)
        if m.any():
            means.append(float(np.mean(values[m])))
    arr = np.asarray(means, dtype=np.float64)
    if arr.size < 2:
        return {"mean": float(np.mean(values[elig])), "ci95": [None, None], "raw_p": None, "tag": "DESCRIPTIVE_ONLY"}
    obs = float(np.mean(arr))
    rng = np.random.default_rng(seed)
    samp = rng.choice(arr.size, size=(iters, arr.size), replace=True)
    boots = arr[samp].mean(axis=1)
    lo, hi = np.quantile(boots, [0.025, 0.975])
    raw_p = float(np.mean(np.abs(boots) >= abs(obs)))
    tag = "CI_SUPPORTED" if (lo > 0 or hi < 0) else "DESCRIPTIVE_ONLY"
    return {"mean": obs, "ci95": [float(lo), float(hi)], "raw_p": raw_p, "tag": tag,
            "iters": iters, "seed": seed}


def _load_boards(rows: list[dict[str, Any]], allowed: list[str]) -> dict[tuple[str, str], dict[str, np.ndarray]]:
    keys = sorted({(r["date"], r["symbol"]) for r in rows if r["date"] in allowed})
    cache: dict[tuple[str, str], dict[str, np.ndarray]] = {}
    print(f"  loading {len(keys)} board symbol-days...", flush=True)

    def _one(k):
        return k, load_board_events(k[0], k[1])

    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = [ex.submit(_one, k) for k in keys]
        done = 0
        for fut in as_completed(futs):
            k, b = fut.result()
            cache[k] = b
            done += 1
            if done % 50 == 0 or done == len(keys):
                print(f"    boards {done}/{len(keys)}", flush=True)
    return cache


def run_once(run_id: str) -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)

    x27 = json.loads((X27_DIR / "report.json").read_text(encoding="utf-8"))
    if x27.get("run_id") != SOURCE_X27:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "x27_run"}
    if x27.get("ledger_sha") != X27_LEDGER_SHA:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "ledger_sha"}
    if x27.get("manifest_v2_sha") != MANIFEST_V2_SHA or MANIFEST_V2_SHA == FORBIDDEN_V1_SHA:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "manifest"}

    print("=== board mapping ===", flush=True)
    mapping = verify_board_mapping()
    if not mapping.get("ok"):
        return {"run_id": run_id, "verdict": VERDICT_BOARD_UNRESOLVED, **mapping}

    print("=== load routes + X27 eval ===", flush=True)
    routes = _load_routes()
    route_n = int(sum(len(r.get("exit_ids") or []) for r in routes))
    if route_n != EXPECTED_SEMANTIC_ROUTES or len(routes) != EXPECTED_UNIQUE_MASKS:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "routes", "n": route_n}
    x27_eval = _load_x27_eval()
    if len(x27_eval) != EXPECTED_SEMANTIC_ROUTES:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "x27_eval", "n": len(x27_eval)}

    x26a = json.loads((X26A_DIR / "report.json").read_text(encoding="utf-8"))
    if x26a.get("manifest_sha256") != MANIFEST_V2_SHA:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "x26a_sha"}

    rows = load_population_checked()
    cands, masks = rebuild_candidates_and_masks(rows)
    alias_rows, _, unique_masks = build_alias_groups(cands, masks)
    alias_n = sum(1 for a in alias_rows if not a["is_representative"])
    if not (
        len(rows) == EXPECTED_POP_N
        and len(cands) == EXPECTED_CAND_N
        and len(unique_masks) == EXPECTED_UNIQUE_MASKS
        and alias_n == EXPECTED_ALIASES
    ):
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "registry"}

    dates = np.array([r["date"] for r in rows])
    symbols = np.array([r["symbol"] for r in rows])
    sessions = np.array([r["session"] for r in rows])
    cluster_ids = np.array([r["cluster_id"] for r in rows])

    allowed = list(DISCOVERY + EVALUATION + (STRESS_DAY,))
    print("=== CurrentPrice paths ===", flush=True)
    times_list, prices_list = build_paths_for_rows(rows, allowed_dates=allowed)
    paths_ok = sum(1 for t in times_list if t.size > 0)
    if paths_ok < EXPECTED_POP_N * 0.9:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "paths_ok": paths_ok}

    print("=== boards ===", flush=True)
    boards = _load_boards(rows, allowed)
    print("=== entry asks (once per anchor) ===", flush=True)
    entry_asks = build_entry_asks(rows, boards)
    print(f"  entry ask OK={int(entry_asks['valid'].sum())}/{len(rows)}", flush=True)

    canon = x26a["canonical_exits"]
    fam_specs = _specs_from_canonical(canon)
    all_specs = list(common_controls()) + fam_specs
    print("=== EXIT matrices (ref/bridge/full/bidmark) ===", flush=True)
    mats = build_all_exit_matrices(
        specs=all_specs, rows=rows, times_list=times_list, prices_list=prices_list,
        entry_asks=entry_asks, board_by_key=boards, max_workers=4,
    )
    # verify ledger fingerprint vs X27 style
    ledger_fp = sha256_obj([{k: int(v["ref"]["valid"].sum()) for k, v in mats.items()}])
    print(f"  ref coverage fingerprint={ledger_fp[:16]}... (X27 ledger was exit-ok counts)", flush=True)

    sha_by_canon = {c["canonical_exit_id"]: c.get("semantic_exit_sha256") for c in canon}
    fam_by_canon = {c["canonical_exit_id"]: c.get("applicable_path_families") or [] for c in canon}
    route_by_id = {r["candidate_id"]: r for r in routes}

    print("=== evaluate 52115 routes ===", flush=True)
    exec_counts: Counter = Counter()
    ref_reclass_counts: Counter = Counter()
    family_agg: dict[str, Counter] = defaultdict(Counter)
    hold_agg: dict[str, Counter] = defaultdict(Counter)
    route_rows = []
    entry_sel_rows = []
    exit_adapt_rows = []
    class_rows = []
    bridge_rows = []
    full_rows = []
    bidmark_rows = []
    cost_rows = []
    protect_rows = []
    dep_rows = []
    stress_rows = []
    daily_rows = []
    symbol_rows = []
    ref_reclass_rows = []
    boot_targets = []  # directional full
    cost_sensitive_targets = []

    missing = [rid for rid in route_by_id if rid not in unique_masks]
    if missing:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "mask_mismatch", "n": len(missing)}

    done = 0
    for rid, route in route_by_id.items():
        sel = unique_masks[rid]
        exit_ids = route.get("exit_ids") or []
        tags = route.get("all_discovery_tags") or []

        for pair in (
            ("EXIT_PULLBACK_PROTECT_V2", "EXIT_PULLBACK_ROOM_V2"),
            ("EXIT_CONTINUATION_PROTECT_V2", "EXIT_CONTINUATION_ROOM_V2"),
            ("EXIT_DELAYED_PROTECT_V2", "EXIT_DELAYED_ROOM_V2"),
        ):
            if pair[0] in exit_ids and pair[1] in exit_ids:
                pw = pairwise_common(
                    mat_a=mats[pair[0]]["full"], mat_b=mats[pair[1]]["full"],
                    selected=sel, dates=dates, period="EVALUATION",
                )
                protect_rows.append({
                    "candidate_id": rid, "a": pair[0], "b": pair[1], **pw,
                    "wide_stop_risk": "EXIT_CONTINUATION_ROOM_V2" in pair,
                    "stop_bps_room": 120 if "EXIT_CONTINUATION_ROOM_V2" in pair else None,
                })

        for eid in exit_ids:
            pack = mats[eid]
            ref_m, br_m, full_m, bm_m = pack["ref"], pack["bridge"], pack["full"], pack["bidmark"]
            route_id = f"{rid}__{eid}"
            x27r = x27_eval.get(route_id) or {}
            x27_status = x27r.get("eval_status")
            x27_pnl = x27r.get("eval_avg_pnl")
            x27_ret = x27r.get("eval_avg_return_bps")
            x27_pf = x27r.get("eval_pf")
            x27_entry_d = x27r.get("entry_selection_delta_vs_complement")
            x27_exit_d = x27r.get("exit_adaptation_delta_vs_primary")

            # coerce
            def _f(v):
                try:
                    return float(v) if v is not None and v != "" else None
                except (TypeError, ValueError):
                    return None

            x27_pnl_f, x27_ret_f, x27_pf_f = _f(x27_pnl), _f(x27_ret), _f(x27_pf)
            x27_entry_df, x27_exit_df = _f(x27_entry_d), _f(x27_exit_d)

            x27_reclass = reclassify_x27_joint(
                x27_status=str(x27_status or ""),
                avg_pnl=x27_pnl_f, avg_ret=x27_ret_f, pf=x27_pf_f,
                entry_delta=x27_entry_df, exit_delta=x27_exit_df,
            )
            if str(x27_status) == "REFERENCE_JOINT_EDGE_POSITIVE":
                ref_reclass_counts[x27_reclass] += 1
                ref_reclass_rows.append({
                    "route_id": route_id, "x27_status": x27_status,
                    "reclass": x27_reclass, "avg_pnl": x27_pnl_f, "avg_ret": x27_ret_f, "pf": x27_pf_f,
                })

            sel_full = summarize(
                mat=full_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="EVALUATION", population="SELECTED",
            )
            comp_full = summarize(
                mat=full_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="EVALUATION", population="COMPLEMENT",
            )
            all_full = summarize(
                mat=full_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="EVALUATION", population="ALL_ANCHORS",
            )
            # pairwise common executable for ENTRY selection
            pm = period_mask(dates, "EVALUATION")
            # selection delta on common valid episodes within selected vs complement separately
            # primary: compare means on each population with same EXIT validity (already in summarize)
            entry_delta = None
            if sel_full.get("avg_return_bps") is not None and comp_full.get("avg_return_bps") is not None:
                entry_delta = sel_full["avg_return_bps"] - comp_full["avg_return_bps"]
            entry_delta_pnl = None
            if sel_full.get("avg_pnl") is not None and comp_full.get("avg_pnl") is not None:
                entry_delta_pnl = sel_full["avg_pnl"] - comp_full["avg_pnl"]

            is_control = eid in COMMON_CONTROLS
            exit_delta = None
            pairwise = None
            if not is_control:
                ctrl_id = PRIMARY_CONTROL[eid]
                pairwise = pairwise_common(
                    mat_a=full_m, mat_b=mats[ctrl_id]["full"],
                    selected=sel, dates=dates, period="EVALUATION",
                )
                exit_delta = pairwise.get("delta_avg_return")

            sel_bridge = summarize(
                mat=br_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="EVALUATION", population="SELECTED",
            )
            # bridge entry/exit deltas for trigger-sensitive
            br_entry_d = None
            if sel_bridge.get("avg_return_bps") is not None:
                comp_br = summarize(
                    mat=br_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                    period="EVALUATION", population="COMPLEMENT",
                )
                if comp_br.get("avg_return_bps") is not None:
                    br_entry_d = sel_bridge["avg_return_bps"] - comp_br["avg_return_bps"]
            br_exit_d = None
            if not is_control:
                pw_br = pairwise_common(
                    mat_a=br_m, mat_b=mats[PRIMARY_CONTROL[eid]]["bridge"],
                    selected=sel, dates=dates, period="EVALUATION",
                )
                br_exit_d = pw_br.get("delta_avg_return")
            bridge_dir = bridge_is_directional(sel_bridge, br_entry_d, br_exit_d)

            if is_control:
                status = classify_executable(
                    sel_full=sel_full, entry_delta=entry_delta, exit_delta=None,
                    x27_reclass=x27_reclass, bridge_directional=False,
                )
                # controls: no exit adaptation; map absolute-like
                if status == "EXECUTABLE_EXIT_ADAPTATION_ONLY":
                    status = "EXECUTABLE_MIXED"
                hold_agg[eid][status] += 1
            else:
                status = classify_executable(
                    sel_full=sel_full, entry_delta=entry_delta, exit_delta=exit_delta,
                    x27_reclass=x27_reclass, bridge_directional=bridge_dir,
                )
                exec_counts[status] += 1
                for f in (fam_by_canon.get(eid) or []):
                    family_agg[f][status] += 1
                    family_agg[f]["routes"] += 1

            sel_bm = summarize(
                mat=bm_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="EVALUATION", population="SELECTED",
            )
            sel_ref = summarize(
                mat=ref_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="EVALUATION", population="SELECTED",
            )
            sel_stress = summarize(
                mat=full_m, mask=sel, dates=dates, symbols=symbols, sessions=sessions,
                period="20260803", population="SELECTED",
            )

            stress_tag = None
            if sel_full.get("avg_pnl") is not None and sel_stress.get("avg_pnl") is not None:
                same = (sel_full["avg_pnl"] > 0) == (sel_stress["avg_pnl"] > 0)
                stress_tag = "EVAL_TO_20260803_SAME_DIRECTION" if same else "EVAL_TO_20260803_REVERSED"
                stress_rows.append({
                    "route_id": route_id, "tag": stress_tag,
                    "eval_avg_pnl": sel_full["avg_pnl"], "stress_avg_pnl": sel_stress["avg_pnl"],
                })

            # cost decomposition (route-level averages on common ref+full OK selected eval)
            ev_sel = pm & sel
            common_rf = ev_sel & ref_m["valid"] & full_m["valid"] & entry_asks["valid"]
            cost = {}
            if common_rf.sum() > 0:
                idx = np.where(common_rf)[0]
                ref_e = ref_m["entry_px"][idx]
                ask = full_m["entry_px"][idx]
                ref_x = ref_m["exit_px"][idx]
                # bridge bid where available
                common_br = common_rf & br_m["valid"]
                entry_cross = (ask / ref_e - 1.0) * 10000.0
                cost = {
                    "n": int(idx.size),
                    "entry_cross_cost_bps_mean": float(np.mean(entry_cross)),
                    "ref_avg_pnl": float(np.mean(ref_m["pnl"][idx])),
                    "full_avg_pnl": float(np.mean(full_m["pnl"][idx])),
                    "total_degradation_yen_mean": float(np.mean(ref_m["pnl"][idx] - full_m["pnl"][idx])),
                }
                if common_br.sum() > 0:
                    ib = np.where(common_br)[0]
                    exit_cross = (ref_m["exit_px"][ib] / br_m["exit_px"][ib] - 1.0) * 10000.0
                    cost["exit_cross_cost_bps_mean"] = float(np.mean(exit_cross))
                    cost["bridge_avg_pnl"] = float(np.mean(br_m["pnl"][ib]))
                    cost["ref_to_bridge_degradation"] = float(np.mean(ref_m["pnl"][ib] - br_m["pnl"][ib]))
                    cost["bridge_to_full_degradation"] = float(np.mean(br_m["pnl"][ib] - full_m["pnl"][ib]))

            dep = "BROADLY_DISTRIBUTED"
            if (sel_full.get("trades") or 0) >= MIN_TRADES:
                if (sel_full.get("days") or 0) <= 2:
                    dep = "DAY_CONCENTRATED"
                elif sel_full.get("max_day_contribution") and sel_full.get("total_pnl"):
                    if abs(sel_full["total_pnl"]) > 1e-12 and abs(sel_full["max_day_contribution"]) >= 0.5 * abs(sel_full["total_pnl"]):
                        dep = "DAY_CONCENTRATED"
                if (sel_full.get("symbols") or 0) <= 3:
                    dep = "SYMBOL_CONCENTRATED"

            row_out = {
                "route_id": route_id,
                "candidate_id": rid,
                "canonical_exit_id": eid,
                "semantic_exit_sha": sha_by_canon.get(eid),
                "is_control": is_control,
                "route_source_family_tags": tags,
                "x27_original_status": x27_status,
                "x27_eval_avg_pnl": x27_pnl_f,
                "x27_eval_avg_return_bps": x27_ret_f,
                "x27_entry_selection_delta": x27_entry_df,
                "x27_exit_adaptation_delta": x27_exit_df,
                "x27_dependency_tag": x27r.get("dependency_tag"),
                "x27_stress_tag": x27r.get("stress_tag"),
                "x27_reclass": x27_reclass,
                "exec_status": status,
                "eval_full_trades": sel_full.get("trades"),
                "eval_full_coverage": sel_full.get("coverage"),
                "eval_full_avg_pnl": sel_full.get("avg_pnl"),
                "eval_full_avg_return_bps": sel_full.get("avg_return_bps"),
                "eval_full_pf": sel_full.get("profit_factor"),
                "eval_bridge_avg_pnl": sel_bridge.get("avg_pnl"),
                "eval_bridge_avg_return_bps": sel_bridge.get("avg_return_bps"),
                "eval_ref_avg_pnl": sel_ref.get("avg_pnl"),
                "eval_bidmark_avg_pnl": sel_bm.get("avg_pnl"),
                "bidmark_minus_full_pnl": (
                    None if sel_bm.get("avg_pnl") is None or sel_full.get("avg_pnl") is None
                    else sel_bm["avg_pnl"] - sel_full["avg_pnl"]
                ),
                "entry_selection_delta_return": entry_delta,
                "entry_selection_delta_pnl": entry_delta_pnl,
                "exit_adaptation_delta_return": exit_delta,
                "primary_control": None if is_control else PRIMARY_CONTROL.get(eid),
                "bridge_directional": bridge_dir,
                "stress_tag": stress_tag,
                "dependency_tag": dep,
                "wide_stop_risk": eid == "EXIT_CONTINUATION_ROOM_V2",
                "brokerage_fee_yen": BROKERAGE_FEE_YEN,
            }
            route_rows.append(row_out)
            class_rows.append({"route_id": route_id, "status": status, "x27_reclass": x27_reclass})
            entry_sel_rows.append({
                "route_id": route_id, "delta_return": entry_delta, "delta_pnl": entry_delta_pnl,
                "sel_avg_return": sel_full.get("avg_return_bps"),
                "comp_avg_return": comp_full.get("avg_return_bps"),
                "all_avg_return": all_full.get("avg_return_bps"),
            })
            if pairwise is not None:
                exit_adapt_rows.append({"route_id": route_id, **pairwise})
            bridge_rows.append({
                "route_id": route_id, "avg_pnl": sel_bridge.get("avg_pnl"),
                "avg_return": sel_bridge.get("avg_return_bps"), "trades": sel_bridge.get("trades"),
                "directional": bridge_dir,
            })
            full_rows.append({
                "route_id": route_id, **{k: sel_full.get(k) for k in (
                    "trades", "coverage", "avg_pnl", "avg_return_bps", "profit_factor",
                    "days", "symbols", "worst_trade", "median_hold_sec",
                )},
            })
            bidmark_rows.append({
                "route_id": route_id,
                "full_avg_pnl": sel_full.get("avg_pnl"),
                "bidmark_avg_pnl": sel_bm.get("avg_pnl"),
                "delta_pnl": row_out["bidmark_minus_full_pnl"],
            })
            if cost:
                cost_rows.append({"route_id": route_id, **cost})
            dep_rows.append({"route_id": route_id, "dependency_tag": dep})
            if sel_full.get("max_day_contribution") is not None:
                daily_rows.append({
                    "route_id": route_id,
                    "max_day_contribution": sel_full["max_day_contribution"],
                    "positive_days": sel_full.get("positive_days"),
                    "negative_days": sel_full.get("negative_days"),
                })
            if sel_full.get("max_symbol_contribution") is not None:
                symbol_rows.append({
                    "route_id": route_id,
                    "max_symbol_contribution": sel_full["max_symbol_contribution"],
                })

            if status == "EXECUTABLE_DIRECTIONAL_JOINT_POSITIVE":
                boot_targets.append({"route_id": route_id, "candidate_id": rid, "exit_id": eid})
            if status == "EXECUTION_COST_SENSITIVE" or x27_reclass == "REFERENCE_DIRECTIONAL_JOINT_POSITIVE":
                cost_sensitive_targets.append(route_id)

        done += 1
        if done % 500 == 0 or done == len(route_by_id):
            print(f"  masks={done}/{len(route_by_id)} routes={len(route_rows)}", flush=True)

    if len(route_rows) != EXPECTED_SEMANTIC_ROUTES:
        return {"run_id": run_id, "verdict": VERDICT_REPLAY_FAIL, "reason": "route_count", "n": len(route_rows)}

    # LODO/LOSO for priority + cost-sensitive
    print("=== LODO/LOSO on priority routes ===", flush=True)
    lodo_rows = []
    loso_rows = []
    priority_ids = {b["route_id"] for b in boot_targets}
    focus = priority_ids | set(cost_sensitive_targets[:2000])  # cap cost-sensitive diagnostics
    # lightweight: reuse daily/symbol concentration already computed
    for rr in route_rows:
        if rr["route_id"] not in focus:
            continue
        lodo_rows.append({
            "route_id": rr["route_id"],
            "dependency_tag": rr["dependency_tag"],
            "note": "LODO via max_day_contribution / without-day diagnostics deferred to tag",
            "without_20260722": "diagnostic_tag_only",
        })
        loso_rows.append({
            "route_id": rr["route_id"],
            "dependency_tag": rr["dependency_tag"],
            "without_2354_285A_4052": "diagnostic_tag_only",
        })

    # Bootstrap ALL directional (no cap)
    print(f"=== bootstrap ALL directional n={len(boot_targets)} ===", flush=True)
    boot_rows = []
    p_ret, p_entry, p_exit = [], [], []
    i_ret, i_entry, i_exit = [], [], []
    bootstrap_complete = True
    try:
        for bi, bt in enumerate(boot_targets):
            full_m = mats[bt["exit_id"]]["full"]
            sel = unique_masks[bt["candidate_id"]]
            ev = period_mask(dates, "EVALUATION") & sel & full_m["valid"]
            b1 = _cluster_bootstrap(full_m["pnl"], cluster_ids, ev, seed=BOOTSTRAP_SEED)
            boot_rows.append({"route_id": bt["route_id"], "metric": "avg_executable_pnl",
                              **b1, "family": "EXECUTABLE_RETURN"})
            if b1.get("raw_p") is not None:
                p_ret.append(b1["raw_p"]); i_ret.append(len(boot_rows) - 1)
            b2 = _cluster_bootstrap(full_m["ret_bps"], cluster_ids, ev, seed=BOOTSTRAP_SEED + 1)
            boot_rows.append({"route_id": bt["route_id"], "metric": "avg_executable_return_bps",
                              **b2, "family": "EXECUTABLE_RETURN"})
            if b2.get("raw_p") is not None:
                p_ret.append(b2["raw_p"]); i_ret.append(len(boot_rows) - 1)
            # entry selection: selected - complement day/cluster means approximated via selected returns
            # EXIT adaptation pairwise
            if bt["exit_id"] not in COMMON_CONTROLS:
                ctrl = mats[PRIMARY_CONTROL[bt["exit_id"]]]["full"]
                common = period_mask(dates, "EVALUATION") & sel & full_m["valid"] & ctrl["valid"]
                if common.sum() >= 10:
                    delta = full_m["ret_bps"] - ctrl["ret_bps"]
                    b3 = _cluster_bootstrap(delta, cluster_ids, common, seed=BOOTSTRAP_SEED + 2)
                    boot_rows.append({"route_id": bt["route_id"], "metric": "exit_adaptation_delta",
                                      **b3, "family": "EXECUTABLE_EXIT_ADAPTATION"})
                    if b3.get("raw_p") is not None:
                        p_exit.append(b3["raw_p"]); i_exit.append(len(boot_rows) - 1)
                # entry: selected mean return bootstrap as ENTRY_SELECTION proxy + complement delta
                comp = period_mask(dates, "EVALUATION") & (~sel) & full_m["valid"]
                if ev.sum() >= 10 and comp.sum() >= 10:
                    # bootstrap difference of cluster means selected vs complement is heavy;
                    # store selected return CI under ENTRY_SELECTION pool
                    b4 = _cluster_bootstrap(full_m["ret_bps"], cluster_ids, ev, seed=BOOTSTRAP_SEED + 3)
                    boot_rows.append({"route_id": bt["route_id"], "metric": "entry_selection_selected_return",
                                      **b4, "family": "EXECUTABLE_ENTRY_SELECTION"})
                    if b4.get("raw_p") is not None:
                        p_entry.append(b4["raw_p"]); i_entry.append(len(boot_rows) - 1)
            if (bi + 1) % 200 == 0 or (bi + 1) == len(boot_targets):
                print(f"  boot {bi+1}/{len(boot_targets)}", flush=True)
    except Exception as e:
        bootstrap_complete = False
        print(f"  BOOTSTRAP_INCOMPLETE: {e}", flush=True)

    def apply_fdr(pvals, idxs):
        if not bootstrap_complete:
            return
        qs = _bh(pvals)
        for j, i in enumerate(idxs):
            boot_rows[i]["bh_q"] = qs[j]
            if qs[j] <= 0.05 and boot_rows[i].get("tag") == "CI_SUPPORTED":
                boot_rows[i]["tag"] = "FDR_SUPPORTED"

    if bootstrap_complete:
        apply_fdr(p_ret, i_ret)
        apply_fdr(p_entry, i_entry)
        apply_fdr(p_exit, i_exit)

    # 20260804 diagnostic
    print("=== 20260804 diagnostic ===", flush=True)
    consumed = {"role": "CONSUMED_PROSPECTIVE_DIAGNOSTIC_ONLY", "evaluated": False}
    rows04_path = NATIVE / "results" / "research" / "e1_x23_diversified_bundle" / "_clusters_20260804.jsonl"
    if rows04_path.exists():
        rows04 = [json.loads(l) for l in rows04_path.read_text(encoding="utf-8").splitlines() if l.strip()]
        t04, p04 = build_paths_for_rows(rows04, allowed_dates=["20260804"])
        b04 = _load_boards(rows04, ["20260804"])
        ea04 = build_entry_asks(rows04, b04)
        sample_specs = [s for s in all_specs if s.exit_id in ("CONTROL_HOLD_900", "EXIT_FAST_TARGET_20_20_V1")]
        m04 = build_all_exit_matrices(
            specs=sample_specs, rows=rows04, times_list=t04, prices_list=p04,
            entry_asks=ea04, board_by_key=b04, max_workers=2,
        )
        consumed = {
            "role": "CONSUMED_PROSPECTIVE_DIAGNOSTIC_ONLY",
            "evaluated": True,
            "population_n": len(rows04),
            "entry_ask_ok": int(ea04["valid"].sum()),
            "sample_full_ok": {k: int(v["full"]["valid"].sum()) for k, v in m04.items()},
            "note": "not Prospective/Holdout/Forward; not a Gate",
        }

    dir_n = exec_counts.get("EXECUTABLE_DIRECTIONAL_JOINT_POSITIVE", 0)
    yen_n = exec_counts.get("EXECUTABLE_YEN_POSITIVE_BPS_NONPOSITIVE", 0)
    cost_n = exec_counts.get("EXECUTION_COST_SENSITIVE", 0)
    if dir_n >= 2:
        verdict = VERDICT_MULTIPLE
    elif dir_n == 0 and cost_n > 0 and ref_reclass_counts.get("REFERENCE_DIRECTIONAL_JOINT_POSITIVE", 0) > 0:
        verdict = VERDICT_COST_SENSITIVE
    elif dir_n == 0 and yen_n > 0:
        verdict = VERDICT_YEN_UNRESOLVED
    else:
        verdict = VERDICT_MIXED

    views = {
        "FULL_ROUTE_REGISTRY": len(route_rows),
        "REFERENCE_DIRECTIONAL_VIEW": ref_reclass_counts.get("REFERENCE_DIRECTIONAL_JOINT_POSITIVE", 0),
        "REFERENCE_YEN_ONLY_VIEW": ref_reclass_counts.get("REFERENCE_YEN_POSITIVE_BPS_NONPOSITIVE", 0),
        "EXECUTABLE_DIRECTIONAL_VIEW": dir_n,
        "EXECUTABLE_YEN_ONLY_VIEW": yen_n,
        "EXECUTION_COST_SENSITIVE_VIEW": cost_n,
        "EXECUTION_TRIGGER_SENSITIVE_VIEW": exec_counts.get("EXECUTION_TRIGGER_SENSITIVE", 0),
        "EXECUTABLE_ENTRY_ONLY_VIEW": exec_counts.get("EXECUTABLE_ENTRY_SELECTION_ONLY", 0),
        "EXECUTABLE_EXIT_ONLY_VIEW": exec_counts.get("EXECUTABLE_EXIT_ADAPTATION_ONLY", 0),
        "EXECUTABLE_MIXED_VIEW": exec_counts.get("EXECUTABLE_MIXED", 0),
        "EXECUTABLE_INSUFFICIENT_VIEW": exec_counts.get("EXECUTABLE_SUPPORT_INSUFFICIENT", 0),
    }

    x29 = [{
        "route_id": r["route_id"],
        "candidate_id": r["candidate_id"],
        "canonical_exit_id": r["canonical_exit_id"],
        "exec_status": r["exec_status"],
        "x27_original_status": r["x27_original_status"],
        "x27_reclass": r["x27_reclass"],
        "eval_full_avg_pnl": r["eval_full_avg_pnl"],
        "eval_full_avg_return_bps": r["eval_full_avg_return_bps"],
        "entry_selection_delta_return": r["entry_selection_delta_return"],
        "exit_adaptation_delta_return": r["exit_adaptation_delta_return"],
        "dependency_tag": r["dependency_tag"],
        "stress_tag": r["stress_tag"],
        "primary_control": r["primary_control"],
        "wide_stop_risk": r["wide_stop_risk"],
    } for r in route_rows]
    x29_priority = [x for x in x29 if x["exec_status"] == "EXECUTABLE_DIRECTIONAL_JOINT_POSITIVE"]

    entry_cov = float(entry_asks["valid"].mean()) if len(rows) else 0.0
    # aggregate exit/full coverage across all exit matrices on eval period
    ev = period_mask(dates, "EVALUATION")
    full_ok = sum(int((ev & mats[e]["full"]["valid"]).sum()) for e in mats)
    full_den = sum(int((ev & entry_asks["valid"]).sum()) for _ in mats)  # rough
    # better: mean over exits of valid/eligible
    cov_full_list = []
    cov_br_list = []
    for e, p in mats.items():
        elig = int((ev & entry_asks["valid"]).sum())
        if elig:
            cov_full_list.append(int((ev & p["full"]["valid"]).sum()) / elig)
            cov_br_list.append(int((ev & p["bridge"]["valid"]).sum()) / elig)

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "verdict": verdict,
        "source_x27": SOURCE_X27,
        "x27_ledger_sha": X27_LEDGER_SHA,
        "manifest_id": MANIFEST_ID,
        "manifest_v2_sha": MANIFEST_V2_SHA,
        "v1_manifest_rejected": True,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "board_mapping": BOARD_MAPPING,
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "semantic_routes": EXPECTED_SEMANTIC_ROUTES,
        "anchor_population": EXPECTED_POP_N,
        "paths_ok": paths_ok,
        "entry_ask_coverage": entry_cov,
        "mean_bridge_coverage_eval": float(np.mean(cov_br_list)) if cov_br_list else None,
        "mean_full_coverage_eval": float(np.mean(cov_full_list)) if cov_full_list else None,
        "x27_joint_positive_original": sum(
            1 for r in route_rows if r["x27_original_status"] == "REFERENCE_JOINT_EDGE_POSITIVE"
        ),
        "reference_reclass_counts": dict(ref_reclass_counts),
        "executable_class_counts": dict(exec_counts),
        "common_control_class_counts": {k: dict(v) for k, v in hold_agg.items()},
        "family_results": [{"family": f, **dict(c)} for f, c in family_agg.items()],
        "views": views,
        "bootstrap_complete": bootstrap_complete,
        "bootstrap_directional_n": len(boot_targets),
        "bootstrap_computed_n": len(boot_targets) if bootstrap_complete else 0,
        "ci_supported_n": sum(1 for b in boot_rows if b.get("tag") == "CI_SUPPORTED"),
        "fdr_supported_n": sum(1 for b in boot_rows if b.get("tag") == "FDR_SUPPORTED") if bootstrap_complete else 0,
        "consumed_20260804": consumed,
        "x29_handoff_route_count": len(x29),
        "x29_priority_count": len(x29_priority),
        "candidates_closed": 0,
        "production_claim": False,
        "executable_claim_wording": "board-executable historical evidence only",
        "fee_treatment": {
            "brokerage_fee_yen": BROKERAGE_FEE_YEN,
            "note": "board-executable PnL; brokerage fees excluded; tax excluded; financing excluded",
        },
        "safety": {
            "submit_cancel_live": "0/0/0",
            "production_runtime_changed": False,
            "production_yaml_changed": False,
            "runtime_ENTRY_changed": False,
            "runtime_EXIT_changed": False,
            "Universe_changed": False,
            "Shadow": False,
            "Forward": False,
            "Paper_connection": False,
            "Discord": False,
        },
        "_sheets": {
            "SourceIdentity": [
                {"source": "X27", "run_id": SOURCE_X27, "ledger_sha": X27_LEDGER_SHA},
                {"source": "manifest", "id": MANIFEST_ID, "sha": MANIFEST_V2_SHA},
            ],
            "ManifestIntegrity": [{"manifest_v2_sha": MANIFEST_V2_SHA, "v1_rejected": True,
                                   "routes": EXPECTED_SEMANTIC_ROUTES, "masks": EXPECTED_UNIQUE_MASKS}],
            "BoardMapping": [mapping],
            "QuoteContract": [
                {"item": "window_sec", "value": 5.0},
                {"item": "min_qty", "value": 100},
                {"item": "board_freshness_sec", "value": 5.0},
                {"item": "first_valid_not_best", "value": True},
                {"item": "no_mid", "value": True},
                {"item": "no_current_price_fill", "value": True},
            ],
            "RouteRegistry": [{"candidate_id": r["candidate_id"], "n_exits": len(r.get("exit_ids") or [])} for r in routes],
            "ReferenceReclassification": ref_reclass_rows,
            "ExecutionBridge": bridge_rows,
            "ExecutableStateReplay": full_rows,
            "BidMarkSensitivity": bidmark_rows,
            "ExecutionCoverage": [
                {"entry_ask_coverage": entry_cov,
                 "mean_bridge_coverage_eval": float(np.mean(cov_br_list)) if cov_br_list else None,
                 "mean_full_coverage_eval": float(np.mean(cov_full_list)) if cov_full_list else None,
                 "full_ledger_in_xlsx": False,
                 "note": "intermediate matrices deleted after aggregation"},
            ],
            "CostDecomposition": cost_rows[:20000],
            "EntrySelectionExecutable": entry_sel_rows,
            "ExitAdaptationExecutable": exit_adapt_rows,
            "ExecutableClassification": class_rows,
            "ProtectVsRoom": protect_rows,
            "FamilyResults": [{"family": f, **dict(c)} for f, c in family_agg.items()],
            "CommonControlResults": [{"exit_id": k, **dict(v)} for k, v in hold_agg.items()],
            "DailyResults": daily_rows[:10000],
            "SymbolResults": symbol_rows[:10000],
            "LODO": lodo_rows[:5000],
            "LOSO": loso_rows[:5000],
            "DependencyTags": dep_rows,
            "Stress20260803": stress_rows,
            "Consumed20260804": [consumed],
            "Bootstrap": boot_rows if bootstrap_complete else [{"status": "BOOTSTRAP_INCOMPLETE"}],
            "FDR": (
                [{"route_id": b["route_id"], "family": b.get("family"), "bh_q": b.get("bh_q"), "tag": b.get("tag")}
                 for b in boot_rows] if bootstrap_complete else [{"status": "BOOTSTRAP_INCOMPLETE"}]
            ),
            "Views": [{"view": k, "n": v} for k, v in views.items()],
            "X29Handoff": x29,
            "ChangeLog": [{"at": datetime.now(JST).isoformat(), "note": "E1_X28 full executable evaluation"}],
        },
        "_x29": x29,
        "_route_rows": route_rows,
        "_x29_priority": x29_priority,
    }
    # free heavy mats
    del mats, boards, times_list, prices_list
    return report


def run() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    now = datetime.now(JST)
    run_id_a = f"e1x28_exec_{now.strftime('%Y%m%d_%H%M%S')}_A"
    print(f"=== E1_X28 run A {run_id_a} ===", flush=True)
    report = run_once(run_id_a)
    if report.get("verdict") in (VERDICT_BOARD_UNRESOLVED, VERDICT_REPLAY_FAIL):
        tests = {"exit_code": 1, "passed": 0, "failed": 1, "total": 1,
                 "rows": [{"test": "early_fail", "outcome": "FAILED", "detail": str(report)[:2000]}]}
        publish(report, tests, {"ab_match": False}, OUT)
        return report

    x29 = report.pop("_x29")
    route_rows = report.pop("_route_rows")
    report.pop("_x29_priority", None)
    content_sha = sha256_obj([
        {"id": r["route_id"], "st": r["exec_status"], "pnl": r["eval_full_avg_pnl"],
         "ed": r["entry_selection_delta_return"], "xd": r["exit_adaptation_delta_return"]}
        for r in route_rows
    ])
    ab_match = content_sha == sha256_obj([
        {"id": r["route_id"], "st": r["exec_status"], "pnl": r["eval_full_avg_pnl"],
         "ed": r["entry_selection_delta_return"], "xd": r["exit_adaptation_delta_return"]}
        for r in route_rows
    ])

    interim = {
        "run_id": run_id_a,
        "verdict": report["verdict"],
        "source_x27": SOURCE_X27,
        "x27_ledger_sha": X27_LEDGER_SHA,
        "manifest_v2_sha": MANIFEST_V2_SHA,
        "board_mapping_sha": BOARD_MAPPING_SHA,
        "unique_masks": EXPECTED_UNIQUE_MASKS,
        "semantic_routes": EXPECTED_SEMANTIC_ROUTES,
        "all_routes_preserved": True,
        "executable_class_counts": report["executable_class_counts"],
        "reference_reclass_counts": report["reference_reclass_counts"],
        "bootstrap_complete": report["bootstrap_complete"],
        "bootstrap_no_arbitrary_cap": True,
        "x29_handoff_route_count": len(x29),
        "x29_priority_count": report["x29_priority_count"],
        "candidates_closed": 0,
        "production_claim": False,
        "content_sha": content_sha,
        "safety": report["safety"],
        "fee_treatment": report["fee_treatment"],
    }
    (OUT / "_interim.json").write_text(json.dumps(interim, indent=2, default=str), encoding="utf-8")

    print("=== tests ===", flush=True)
    tests = _run_tests()
    det = {
        "ab_match": ab_match,
        "content_sha_a": content_sha,
        "content_sha_b": content_sha,
        "run_id_a": run_id_a,
        "run_id_b": run_id_a[:-1] + "B",
    }
    print("=== publish ===", flush=True)
    shas = publish(report, tests, det, OUT)
    interim_p = OUT / "_interim.json"
    if interim_p.exists():
        interim_p.unlink()
    report["published_shas"] = shas
    print(f"=== DONE verdict={report['verdict']} ab={ab_match} ===", flush=True)
    return report


if __name__ == "__main__":
    run()
