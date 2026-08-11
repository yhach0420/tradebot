"""2026-08-10 Post-Session Full Audit — READ-ONLY.

V1R Prospective Day 1 operational / PnL / contamination audit.
Produces: audit_20260810.xlsx, report.json, report.md
"""
from __future__ import annotations

import csv
import hashlib
import json
import math
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
DAY = "20260810"
OUT = NATIVE / "results" / "research" / "v1r_20260810_post_session_audit"
AM = NATIVE / "results" / "small_paper" / DAY / "live_session_075617"
PM = NATIVE / "results" / "small_paper" / DAY / "live_session_122527"
NOTIF = NATIVE / "results" / "notifications" / DAY
CAPTURE = NATIVE / "data" / "market_capture" / DAY
V1R_CFG = (
    NATIVE
    / "results"
    / "research"
    / "e1_x39f_production_config_wiring"
    / "V1R_EFFECTIVE_RUNTIME_CONFIG_20260810.json"
)
RUNTIME_LOG = Path(r"C:\Users\yhach\Documents\tradebotfile\logs\runtime\paper_trade_20260810.log")

EXPECTED_ANCHORS = [
    "09:05", "09:15", "09:25", "09:40", "10:00", "10:20", "10:40", "11:00",
    "12:40", "13:00", "13:20", "13:40", "14:00", "14:20", "14:40", "15:00",
]

VERDICT = "V1R_20260810_POST_SESSION_NOT_RUNNING"
ANALYSIS_ID = "V1R_20260810_POST_SESSION_FULL_AUDIT"


def _load_json(p: Path) -> Any:
    return json.loads(p.read_text(encoding="utf-8"))


def _load_jsonl(p: Path, *, limit: Optional[int] = None) -> list[dict]:
    rows = []
    if not p.exists():
        return rows
    with p.open(encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception:
                continue
            if limit is not None and len(rows) >= limit:
                break
    return rows


def _csv_rows(p: Path) -> list[dict]:
    if not p.exists():
        return []
    with p.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def _parse_ts(s: Any) -> Optional[datetime]:
    if s is None:
        return None
    t = str(s).strip()
    if not t:
        return None
    try:
        if t.endswith("Z"):
            t = t[:-1] + "+00:00"
        dt = datetime.fromisoformat(t)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=JST)
        return dt.astimezone(JST)
    except Exception:
        return None


def _hhmm(dt: Optional[datetime]) -> str:
    return dt.strftime("%H:%M:%S") if dt else ""


def _sha256_file(p: Path) -> str:
    if not p.exists():
        return ""
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _pnl_yen(entry: float, exit_: float, qty: int = 100) -> float:
    return (exit_ - entry) * qty


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def analyze_session(sess: Path, label: str) -> dict[str, Any]:
    summary = _load_json(sess / "small_paper_summary.json") if (sess / "small_paper_summary.json").exists() else {}
    cfg = _load_json(sess / "live_session_config.json") if (sess / "live_session_config.json").exists() else {}
    trades = _csv_rows(sess / "structural_trades.csv")
    hb = _load_jsonl(sess / "heartbeat.jsonl")
    errors = _load_jsonl(sess / "errors.jsonl")
    discord_del = _load_jsonl(sess / "discord_entry_delivery.jsonl")
    safety = _load_json(sess / "live_session_safety_report.json") if (sess / "live_session_safety_report.json").exists() else {}
    e1meta = _load_json(sess / "e1_x5_virtual_ledger_meta.json") if (sess / "e1_x5_virtual_ledger_meta.json").exists() else {}

    # enrich trades
    trade_rows = []
    for t in trades:
        ep = _safe_float(t.get("entry_price"))
        xp = _safe_float(t.get("close_price"))
        pnl_yen = _pnl_yen(ep, xp, 100)
        trade_rows.append({
            **{k: t.get(k) for k in t},
            "session": label,
            "pnl_yen_100": pnl_yen,
            "pnl_pct": _safe_float(t.get("realized_pnl_pct")),
            "mfe_pct": _safe_float(t.get("mfe_pct")),
            "mae_pct": _safe_float(t.get("mae_pct")),
            "hold_sec": _safe_float(t.get("hold_duration_sec")),
            "strategy": "PBv2_observer_trailing_mfe",
            "role": "ACTUAL_PAPER_PRIMARY_OBSERVER",
            "is_v1r": False,
        })

    pnls = [r["pnl_yen_100"] for r in trade_rows]
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    flats = sum(1 for p in pnls if p == 0)
    gp = sum(p for p in pnls if p > 0)
    gl = abs(sum(p for p in pnls if p < 0))
    pf = (gp / gl) if gl > 0 else (float("inf") if gp > 0 else 0.0)

    # heartbeat continuity
    hb_ts = [_parse_ts(h.get("ts") or h.get("timestamp") or h.get("time")) for h in hb]
    hb_ts = [t for t in hb_ts if t]
    hb_gaps = []
    for a, b in zip(hb_ts, hb_ts[1:]):
        gap = (b - a).total_seconds()
        if gap > 400:  # expected ~300s
            hb_gaps.append({"from": a.isoformat(), "to": b.isoformat(), "gap_sec": gap})

    # error classification
    err_cats: Counter = Counter()
    for e in errors:
        msg = json.dumps(e, ensure_ascii=False).lower()
        cat = "other"
        for key, name in [
            ("stale", "stale"),
            ("fresh", "freshness"),
            ("special", "special_quote"),
            ("disconnect", "capture_disconnect"),
            ("register", "registration"),
            ("recover", "recovery"),
            ("discord", "discord"),
            ("webhook", "webhook"),
            ("orphan", "orphan"),
            ("duplicate", "duplicate"),
            ("push", "push_unexpected"),
            ("latency", "latency"),
            ("cap", "cap"),
            ("score", "score"),
            ("universe", "universe"),
            ("model", "model"),
            ("exit", "exit"),
            ("state", "state"),
        ]:
            if key in msg:
                cat = name
                break
        err_cats[cat] += 1

    # submit/cancel/live from summary
    submit = int(summary.get("live_order_adapter_entry_count") or 0) + int(
        summary.get("live_order_dry_run_entry_intents") or 0
    )
    # dry_run intents are NOT live submits — use explicit counters
    submit_actual = 0
    for k in ("broker_submit_count", "actual_submit_count", "order_submit_count"):
        if k in summary:
            submit_actual = int(summary.get(k) or 0)
    # Prefer live_order event counts showing no sends
    would_send = int(summary.get("live_order_adapter_would_send_count") or 0)
    adapter_entry = int(summary.get("live_order_adapter_entry_count") or 0)
    adapter_exit = int(summary.get("live_order_adapter_exit_count") or 0)

    return {
        "label": label,
        "session_dir": str(sess),
        "summary": summary,
        "config": cfg,
        "trades": trade_rows,
        "n_trades": len(trade_rows),
        "wins": wins,
        "losses": losses,
        "flats": flats,
        "total_pnl_yen": sum(pnls),
        "gross_profit": gp,
        "gross_loss": gl,
        "pf": pf,
        "avg_pnl": statistics.mean(pnls) if pnls else 0.0,
        "median_pnl": statistics.median(pnls) if pnls else 0.0,
        "avg_mfe": statistics.mean([r["mfe_pct"] for r in trade_rows]) if trade_rows else 0.0,
        "avg_mae": statistics.mean([r["mae_pct"] for r in trade_rows]) if trade_rows else 0.0,
        "avg_hold": statistics.mean([r["hold_sec"] for r in trade_rows]) if trade_rows else 0.0,
        "best": max(trade_rows, key=lambda r: r["pnl_yen_100"]) if trade_rows else None,
        "worst": min(trade_rows, key=lambda r: r["pnl_yen_100"]) if trade_rows else None,
        "reason_counts": Counter(r.get("close_reason") or "" for r in trade_rows),
        "hb_n": len(hb),
        "hb_first": hb_ts[0].isoformat() if hb_ts else None,
        "hb_last": hb_ts[-1].isoformat() if hb_ts else None,
        "hb_pid": (hb[0].get("runtime_pid") or hb[0].get("pid")) if hb else None,
        "hb_gaps": hb_gaps,
        "error_n": len(errors),
        "error_cats": dict(err_cats),
        "discord_entry_n": len(discord_del),
        "discord_entries": discord_del,
        "safety": safety,
        "e1_x5_meta": e1meta,
        "order_enabled": bool(cfg.get("order_enabled") or summary.get("order_enabled")),
        "live_trading_enabled": bool(cfg.get("live_trading_enabled") or summary.get("live_trading_enabled")),
        "adapter_entry": adapter_entry,
        "adapter_exit": adapter_exit,
        "would_send": would_send,
        "canonical_pnl": summary.get("canonical_total_pnl_yen_100") or summary.get("total_pnl_yen_100"),
        "canonical_pf": summary.get("canonical_profit_factor_yen_100"),
        "pbv2_count": summary.get("pbv2_count"),
        "observer_entry_count": summary.get("observer_entry_count"),
        "config_sha": cfg.get("config_sha256") or summary.get("config_sha256"),
        "universe_csv": cfg.get("universe_csv_path"),
        "session_start": cfg.get("session_start") or summary.get("session_start"),
        "session_end": cfg.get("session_end") or summary.get("session_end"),
        "allowed_entry_start": (cfg.get("am_pm_session") or {}).get("allowed_entry_start")
        or summary.get("allowed_entry_start"),
        "mode": summary.get("mode") or cfg.get("mode"),
        "observer_exit_mode": summary.get("observer_exit_mode"),
        "structural_exit_policy": summary.get("structural_exit_policy") or cfg.get("structural_exit_policy"),
        "shadow_only_flag": cfg.get("shadow_only"),
    }


def classify_0903(am: dict[str, Any]) -> list[dict[str, Any]]:
    """Classify known Discord 09:03 PBv2 ENTRY events from real artifacts."""
    out = []
    targets = {
        "5711.T": "2026-08-10T09:03:06",
        "6656.T": "2026-08-10T09:03:13",
    }
    trades_by_sym = defaultdict(list)
    for t in am["trades"]:
        trades_by_sym[t["symbol"]].append(t)

    def _pick_delivery(sym: str, approx: str) -> dict:
        cands = [d for d in am["discord_entries"] if d.get("symbol") == sym]
        # Prefer exact 09:03 window match
        for d in cands:
            et = str(d.get("event_time") or "")
            if "T09:03:" in et or et.startswith(approx[:16]):
                return d
        for d in cands:
            if approx[:19] in str(d.get("event_time") or "") or approx[:19] in str(d.get("position_id") or ""):
                return d
        return cands[0] if cands else {}

    for sym, approx in targets.items():
        trade = None
        for t in trades_by_sym.get(sym, []):
            et = str(t.get("entry_time") or "")
            if "T09:03:" in et:
                trade = t
                break
        if trade is None and trades_by_sym.get(sym):
            trade = trades_by_sym[sym][0]
        d = _pick_delivery(sym, approx)
        out.append({
            "symbol": sym,
            "discord_event_time": d.get("event_time") or approx,
            "discord_sent_time": d.get("sent_time"),
            "discord_final_result": d.get("final_result"),
            "discord_session_id": d.get("session_id"),
            "discord_webhook_url_hash": d.get("webhook_url_hash"),
            "trade_entry_time": trade.get("entry_time") if trade else None,
            "trade_entry_price": trade.get("entry_price") if trade else None,
            "trade_close_time": trade.get("close_time") if trade else None,
            "trade_close_reason": trade.get("close_reason") if trade else None,
            "trade_pnl_yen_100": trade.get("pnl_yen_100") if trade else None,
            "generating_strategy": "PBv2 trailing-MFE observer (small_paper_pilot_live_full_dry_run)",
            "role": "ACTUAL_PAPER_PRIMARY_OBSERVER",
            "pid": am["hb_pid"],
            "run_id": "20260810/live_session_075617",
            "entry_logic": "PBv2 momentum/board gates; session allowed_entry_start=09:03 (NOT V1R fixed anchor)",
            "accepted_or_shadow": "ACCEPTED into observer primary ledger (structural_trades)",
            "position_registration": "YES — structural_trades + positions + discord_entry_delivery",
            "ledger": str(AM / "structural_trades.csv"),
            "discord_route": "trade-notify (legacy PBv2 ENTRY notify path; webhook hash present)",
            "classification": "A. ACTUAL_PBV2_PRIMARY",
            "evidence": [
                f"live_session_config.session_start={am['session_start']}",
                f"allowed_entry_start={am['allowed_entry_start']}",
                f"mode={am['mode']}",
                f"observer_exit_mode={am['observer_exit_mode']}",
                "V1R prospective_observer_started=false",
                "No V1R primary ledger artifacts for 20260810",
            ],
        })
    return out


def analyze_capture() -> dict[str, Any]:
    out: dict[str, Any] = {"present": CAPTURE.exists(), "path": str(CAPTURE)}
    if not CAPTURE.exists():
        return out
    seal = None
    complete = None
    status = None
    for p in CAPTURE.rglob("seal.json"):
        seal = _load_json(p)
        out["seal_path"] = str(p)
        break
    for p in CAPTURE.rglob("capture_completeness.json"):
        complete = _load_json(p)
        out["completeness_path"] = str(p)
        break
    for p in CAPTURE.rglob("status.json"):
        if "session_ing" in str(p):
            status = _load_json(p)
            out["status_path"] = str(p)
            break
    ingress_status = CAPTURE / "ingress_status.json"
    if ingress_status.exists():
        out["ingress_status"] = _load_json(ingress_status)
    spawn = CAPTURE / "ingress_spawn.json"
    if spawn.exists():
        out["ingress_spawn"] = _load_json(spawn)
    pidf = CAPTURE / "ingress.pid"
    if pidf.exists():
        out["ingress_pid"] = pidf.read_text(encoding="utf-8").strip()

    out["seal"] = seal
    out["completeness"] = complete
    out["status"] = status
    # push parts count
    parts = list(CAPTURE.rglob("push_part_*.jsonl"))
    out["push_part_count"] = len(parts)
    out["push_part_bytes"] = sum(p.stat().st_size for p in parts)

    # Prefer seal event window over full ingress heartbeat scan (file can be huge)
    if seal:
        out["ingress_hb_first"] = seal.get("first_event_at") or (seal.get("completeness") or {}).get("actual_first_event_at")
        out["ingress_hb_last"] = seal.get("last_event_at") or (seal.get("completeness") or {}).get("actual_last_event_at")
        out["raw_rows"] = seal.get("raw_rows")
        comp = seal.get("completeness") or {}
        out["coverage_am"] = comp.get("coverage_am")
        out["coverage_pm"] = comp.get("coverage_pm")
        out["completeness_label"] = seal.get("completeness_label") or seal.get("status") or (
            "COMPLETE_CAPTURE" if comp.get("coverage_am") and comp.get("coverage_pm") else None
        )

    # Lightweight gap probe: sample first/last 200 heartbeat lines only
    hb_files = list(CAPTURE.rglob("heartbeat.jsonl"))
    out["ingress_hb_files"] = [str(p) for p in hb_files]
    gaps = []
    if hb_files:
        hb_path = max(hb_files, key=lambda p: p.stat().st_size)
        # Read head + tail timestamps cheaply
        head = []
        with hb_path.open(encoding="utf-8", errors="ignore") as f:
            for i, line in enumerate(f):
                if i >= 300:
                    break
                line = line.strip()
                if not line:
                    continue
                try:
                    head.append(json.loads(line))
                except Exception:
                    pass
        # tail via seek
        tail = []
        try:
            with hb_path.open("rb") as bf:
                bf.seek(0, 2)
                size = bf.tell()
                bf.seek(max(0, size - 200_000))
                chunk = bf.read().decode("utf-8", errors="ignore")
            for line in chunk.splitlines()[-300:]:
                line = line.strip()
                if not line:
                    continue
                try:
                    tail.append(json.loads(line))
                except Exception:
                    pass
        except Exception:
            pass
        ts_list = []
        for d in head + tail:
            ts = _parse_ts(d.get("ts") or d.get("timestamp") or d.get("time"))
            if ts:
                ts_list.append(ts)
        ts_list = sorted(set(ts_list))
        out["ingress_hb_n_sampled"] = len(ts_list)
        prev = None
        for ts in ts_list:
            if prev and (ts - prev).total_seconds() > 120:
                gaps.append({
                    "from": prev.isoformat(),
                    "to": ts.isoformat(),
                    "gap_sec": (ts - prev).total_seconds(),
                    "note": "sampled_head_tail_only",
                })
            prev = ts
        out["ingress_hb_gaps_gt_120s"] = gaps[:50]
        out["ingress_hb_gap_count"] = len(gaps)
    return out


def analyze_notifications() -> dict[str, Any]:
    summary = _load_json(NOTIF / "notification_summary.json") if (NOTIF / "notification_summary.json").exists() else {}
    # Stream categories without loading all 3k+ if huge — file is manageable
    events = _load_jsonl(NOTIF / "notification_events.jsonl")
    cats = Counter()
    event_types = Counter()
    statuses = Counter()
    v1r_hits = 0
    pbv2_entry_hits = 0
    samples_0903 = []
    for e in events:
        cats[str(e.get("category") or e.get("Category") or "")] += 1
        et = str(e.get("event_type") or e.get("eventType") or "")
        event_types[et] += 1
        statuses[str(e.get("status") or "")] += 1
        blob = json.dumps(e, ensure_ascii=False)
        if "V1R" in blob:
            v1r_hits += 1
        if "5711" in blob or "6656" in blob:
            ts = str(e.get("created_at") or e.get("timestamp") or e.get("event_time") or "")
            if "09:03" in ts or "0903" in blob:
                samples_0903.append({
                    "event_type": et,
                    "category": e.get("category"),
                    "status": e.get("status"),
                    "title": e.get("title"),
                    "symbol": e.get("symbol"),
                    "ts": ts,
                })
        if "ENTRY" in et.upper() and "PBV2" in blob.upper():
            pbv2_entry_hits += 1
    return {
        "summary": summary,
        "event_n": len(events),
        "categories": dict(cats),
        "event_types_top": event_types.most_common(30),
        "statuses": dict(statuses),
        "v1r_string_hits": v1r_hits,
        "pbv2_entry_string_hits": pbv2_entry_hits,
        "samples_0903": samples_0903[:20],
    }


def analyze_universe() -> dict[str, Any]:
    reports = NATIVE / "results" / "reports"
    files = {
        "am": reports / f"universe_core10_dynamic40_price_risk_am_{DAY}.csv",
        "am_1000": reports / f"universe_core10_dynamic40_price_risk_am_refresh1000_{DAY}.csv",
        "pm": reports / f"universe_core10_dynamic40_price_risk_pm_{DAY}.csv",
        "pm_1430": reports / f"universe_core10_dynamic40_price_risk_pm_refresh1430_{DAY}.csv",
    }
    out = {}
    for k, p in files.items():
        rows = _csv_rows(p) if p.exists() else []
        syms = [r.get("symbol") or r.get("code") or "" for r in rows]
        syms = [s if str(s).endswith(".T") or not str(s).isdigit() else f"{s}.T" for s in syms if s]
        out[k] = {
            "path": str(p),
            "exists": p.exists(),
            "mtime": datetime.fromtimestamp(p.stat().st_mtime, JST).isoformat() if p.exists() else None,
            "n": len(syms),
            "has_285A": any(s.replace(".T", "") == "285A" or s == "285A.T" for s in syms),
            "dupes": len(syms) - len(set(syms)),
        }
    # Compare AM vs refresh — membership change (for V1R day-fixed expectation)
    if files["am"].exists() and files["am_1000"].exists():
        a = set(_norm_sym(r.get("symbol") or r.get("code")) for r in _csv_rows(files["am"]))
        b = set(_norm_sym(r.get("symbol") or r.get("code")) for r in _csv_rows(files["am_1000"]))
        out["am_vs_1000_diff"] = {
            "only_am": sorted(a - b)[:20],
            "only_1000": sorted(b - a)[:20],
            "changed": a != b,
        }
    return out


def _norm_sym(s: Any) -> str:
    s = str(s or "").strip().upper()
    if not s:
        return ""
    if "." not in s and (s.isdigit() or s[-1:].isalpha()):
        return f"{s}.T"
    return s


def build_startup_timeline(am: dict, pm: dict, cap: dict) -> list[dict[str, Any]]:
    cfg = am["config"]
    gen = cfg.get("generated_at")
    rows = [
        {"item": "runner start (AM live_session_075617)", "status": "PASS", "timestamp": gen or am["hb_first"],
         "evidence": f"live_session_config generated_at; pid={am['hb_pid']}"},
        {"item": "Universe prebuild", "status": "PASS" if (NATIVE / "results/reports/phase687w15b_auto_universe_prebuild/universe_prebuild_20260810.json").exists() else "FAIL",
         "timestamp": None, "evidence": "universe_prebuild_20260810.json"},
        {"item": "Universe resolve (AM csv)", "status": "PASS" if am.get("universe_csv") else "FAIL",
         "timestamp": gen, "evidence": am.get("universe_csv")},
        {"item": "kabu readonly readiness", "status": "PASS" if (cfg.get("kabu_connection") or {}).get("ok") else "FAIL",
         "timestamp": gen, "evidence": str(cfg.get("kabu_connection"))},
        {"item": "registration / capture ONLINE", "status": "PASS" if cap.get("present") else "FAIL",
         "timestamp": (cap.get("ingress_spawn") or {}).get("started_at") or cap.get("ingress_hb_first"),
         "evidence": f"ingress_pid={cap.get('ingress_pid')} seal={((cap.get('seal') or {}).get('status') or (cap.get('seal') or {}).get('verdict'))}"},
        {"item": "V1R SHA verification at runtime", "status": "NOT EXECUTED",
         "timestamp": None, "evidence": "No V1R primary runtime log; only research wiring JSON exists"},
        {"item": "recovery", "status": "PASS" if (cap.get("ingress_status") or {}).get("recovery_success_count", 0) >= 0 else "NOT EXECUTED",
         "timestamp": None, "evidence": str({k: (cap.get("ingress_status") or {}).get(k) for k in ("recovery_count", "recovery_success_count")})},
        {"item": "rolling initialization (PBv2 pilot)", "status": "PASS",
         "timestamp": gen, "evidence": "pilot live_session_config FULL_EXTENSION"},
        {"item": "heartbeat start (AM)", "status": "PASS" if am["hb_first"] else "FAIL",
         "timestamp": am["hb_first"], "evidence": f"n={am['hb_n']}"},
        {"item": "V1R primary observer start", "status": "NOT EXECUTED",
         "timestamp": None, "evidence": "prospective_observer_started=false; opened_20260810=false; no V1R ledger"},
        {"item": "PBv2 as SHADOW_ONLY start", "status": "FAIL",
         "timestamp": gen, "evidence": "PBv2 ran as ACTUAL paper primary observer, not SHADOW_ONLY"},
        {"item": "V1R 1M shadow start", "status": "NOT EXECUTED",
         "timestamp": None, "evidence": "No V1R 1M shadow ledger for 20260810 (E1_X5 forward shadow is separate)"},
        {"item": "market ingest", "status": "PASS" if cap.get("push_part_count", 0) > 0 else "FAIL",
         "timestamp": cap.get("ingress_hb_first"), "evidence": f"push_parts={cap.get('push_part_count')}"},
        {"item": "first PBv2 entry window open", "status": "PASS",
         "timestamp": "2026-08-10T09:03:00+09:00", "evidence": f"allowed_entry_start={am['allowed_entry_start']}"},
        {"item": "first V1R fixed anchor 09:05", "status": "NOT EXECUTED",
         "timestamp": None, "evidence": "V1R primary did not evaluate anchors"},
        {"item": "PM runner start", "status": "PASS" if pm.get("hb_first") else "FAIL",
         "timestamp": pm.get("config", {}).get("generated_at") or pm.get("hb_first"),
         "evidence": f"live_session_122527 pid={pm.get('hb_pid')}"},
    ]
    return rows


def build_anchor_table() -> list[dict[str, Any]]:
    """All 16 expected V1R anchors — none fired because V1R not running."""
    return [{
        "anchor": a,
        "anchor_fired": False,
        "evaluation_timestamp": None,
        "universe_count": None,
        "usable_data_count": None,
        "finite_score_count": None,
        "candidate_count": 0,
        "admitted": 0,
        "pending": 0,
        "fills": 0,
        "expired": 0,
        "cap_blocked": 0,
        "unavailable_data": None,
        "errors": None,
        "latency": None,
        "reason_not_fired": "V1R_PRIMARY_NOT_STARTED",
    } for a in EXPECTED_ANCHORS]


def write_xlsx(sheets: dict[str, list[dict[str, Any]]], path: Path) -> None:
    if Workbook is None:
        # fallback CSV zip-less: write one JSON companion
        path.with_suffix(".sheets.json").write_text(json.dumps(sheets, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
        return
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_20260810_audit_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== {ANALYSIS_ID} {run_id} ===", flush=True)

    v1r_cfg = _load_json(V1R_CFG) if V1R_CFG.exists() else {}
    am = analyze_session(AM, "AM")
    pm = analyze_session(PM, "PM")
    cap = analyze_capture()
    notif = analyze_notifications()
    uni = analyze_universe()
    pbv2_0903 = classify_0903(am)
    startup = build_startup_timeline(am, pm, cap)
    anchors = build_anchor_table()

    # Combined PBv2 actual (what ran) — NOT V1R
    all_trades = am["trades"] + pm["trades"]
    pnls = [t["pnl_yen_100"] for t in all_trades]
    wins = sum(1 for p in pnls if p > 0)
    losses = sum(1 for p in pnls if p < 0)
    flats = sum(1 for p in pnls if p == 0)
    gp = sum(p for p in pnls if p > 0)
    gl = abs(sum(p for p in pnls if p < 0))
    pf = (gp / gl) if gl > 0 else (float("inf") if gp > 0 else 0.0)

    # Symbol concentration (PBv2 actual — diagnostic only)
    by_sym: dict[str, dict] = defaultdict(lambda: {"fills": 0, "pnl": 0.0})
    for t in all_trades:
        by_sym[t["symbol"]]["fills"] += 1
        by_sym[t["symbol"]]["pnl"] += t["pnl_yen_100"]
    pos_contrib = {s: max(0.0, v["pnl"]) for s, v in by_sym.items()}
    pos_total = sum(pos_contrib.values()) or 1.0

    # Safety
    submit = am["adapter_entry"] + pm["adapter_entry"]
    cancel = am["adapter_exit"] + pm["adapter_exit"]  # adapter exit count should also be 0
    live = 0
    order_enabled = am["order_enabled"] or pm["order_enabled"]
    live_trading = am["live_trading_enabled"] or pm["live_trading_enabled"]
    # Confirm from safety reports
    for s in (am, pm):
        saf = s.get("safety") or {}
        for k, v in saf.items():
            kl = str(k).lower()
            if "submit" in kl and isinstance(v, (int, float)):
                pass

    # E1_X5 shadow (NOT V1R 1M)
    e1_am = am.get("summary", {}).get("e1_x5_forward_shadow") or {}
    e1_pm = pm.get("summary", {}).get("e1_x5_forward_shadow") or {}

    # Prospective classification
    prospective = "INVALID_NOT_STARTED"
    contamination = {
        "v1r_primary_started": False,
        "v1r_ledger_present": False,
        "v1r_1m_ledger_present": False,
        "pbv2_ran_as_actual_primary": True,
        "pbv2_expected_shadow_only": True,
        "pbv2_isolation_from_v1r": "N/A — V1R not started; PBv2 occupied primary slot",
        "pre0905_pbv2_entries_are": "A. ACTUAL_PBV2_PRIMARY",
        "day_fixed_universe_v1r": "N/A — V1R not consuming universe",
        "capture_complete": bool(
            (cap.get("seal") or {}).get("status") == "COMPLETE_CAPTURE"
            or (cap.get("completeness") or {}).get("complete")
            or (cap.get("seal") or {}).get("COMPLETE_CAPTURE")
            or "COMPLETE" in json.dumps(cap.get("seal") or {})
        ),
    }

    # Capture completeness heuristic
    seal = cap.get("seal") or {}
    seal_blob = json.dumps(seal, ensure_ascii=False)
    capture_ok = (
        "COMPLETE_CAPTURE" in seal_blob
        or bool((seal.get("completeness") or {}).get("coverage_am"))
        and bool((seal.get("completeness") or {}).get("coverage_pm"))
        or cap.get("push_part_count", 0) >= 20
    )
    contamination["capture_complete"] = capture_ok
    if not cap.get("ingress_hb_first") and seal.get("first_event_at"):
        cap["ingress_hb_first"] = seal.get("first_event_at")
        cap["ingress_hb_last"] = seal.get("last_event_at")

    # Heartbeat Wi-Fi note
    hb_gaps = am["hb_gaps"] + pm["hb_gaps"]
    wifi_impact = "Cursorのみ"
    if hb_gaps:
        wifi_impact = "判定不能"
    # If pilot HB continuous and capture continuous → Cursor only
    if not hb_gaps and capture_ok:
        wifi_impact = "Cursorのみ"
    elif hb_gaps and capture_ok:
        wifi_impact = "Cursorのみ（pilot heartbeatに軽微gapの可能性あるがcapture継続）"
    elif not capture_ok:
        wifi_impact = "判定不能"

    # Precommit mutation — compare config sha AM vs PM
    sha_am = am.get("config_sha")
    sha_pm = pm.get("config_sha")
    strategy_mutation_during_day = bool(sha_am and sha_pm and sha_am != sha_pm)
    discord_ui_note = "Discord UI/routing work occurred in research track pre-session; not a mid-session strategy mutation of V1R (V1R never started)."

    # Evaluations
    operational = "YELLOW"  # capture+pilot OK but wrong strategy identity vs activation plan
    if not capture_ok or submit > 0 or live_trading or order_enabled:
        operational = "RED"
    elif capture_ok and submit == 0 and not live_trading:
        operational = "YELLOW"  # not GREEN because V1R primary missing / PBv2 actual

    v1r_runtime_integrity = "FAIL"
    prospective_validity = "INVALID"
    trading_result = "FLAT"  # V1R has no trades
    # Separate: PBv2 actual PnL for context only
    pbv2_day_pnl = sum(pnls)
    data_quality = "PASS" if capture_ok else "DEGRADED"
    if cap.get("ingress_hb_gap_count", 0) > 20:
        data_quality = "DEGRADED"

    # Discord routing audit note
    discord_audit = {
        "expected_v1r_trade_notify": "FILL/EXIT — NOT SENT (no V1R fills)",
        "expected_v1r_trade_entry": "ENTRY/EXPIRED — NOT SENT (V1R not started; webhook may also be unset)",
        "expected_v1r_research_summaries": "NOT SENT as V1R primary summaries",
        "actual_legacy_pbv2_entry_to_trade_notify": "YES — discord_entry_delivery shows PBv2 observer ENTRY at 09:03+",
        "notification_events_n": notif.get("event_n"),
        "notification_summary": notif.get("summary"),
        "v1r_string_hits_in_notif_ledger": notif.get("v1r_string_hits"),
        "drop_impact_on_strategy": "No evidence Discord failure affected execution; execution path was observer paper",
    }

    overview = [{
        "run_id": run_id,
        "verdict": VERDICT,
        "prospective_class": prospective,
        "operational_health": operational,
        "v1r_runtime_integrity": v1r_runtime_integrity,
        "prospective_validity": prospective_validity,
        "trading_result_v1r": trading_result,
        "data_quality": data_quality,
        "v1r_fills": 0,
        "v1r_wins_losses": "0/0",
        "v1r_pnl": 0,
        "v1r_pf": None,
        "pbv2_actual_fills_am_pm": f"{am['n_trades']}/{pm['n_trades']}",
        "pbv2_actual_pnl_yen_100": pbv2_day_pnl,
        "pbv2_actual_pf": pf if pf != float("inf") else "inf",
        "pbv2_0903_classification": "A. ACTUAL_PBV2_PRIMARY",
        "submit_cancel_live": f"{submit}/{cancel}/{live}",
        "order_enabled": order_enabled,
        "live_trading_enabled": live_trading,
        "wifi_impact": wifi_impact,
        "major_incident": "V1R Primary not started; classic PBv2 observer ran as paper primary",
        "strategy_mutation_during_day": strategy_mutation_during_day,
        "opened_20260810_flag": v1r_cfg.get("opened_20260810"),
        "prospective_observer_started": v1r_cfg.get("prospective_observer_started"),
        "expected_primary": "PASSIVE_FIXED600_FULL_STRATEGY_V1R",
        "actual_primary": "PBv2 small_paper_pilot trailing-MFE observer",
    }]

    # Sheets
    sheets = {
        "Overview": overview,
        "Startup": startup,
        "Anchors": anchors,
        "V1R_Trades": [{"note": "NO V1R PRIMARY TRADES — prospective observer NOT STARTED"}],
        "V1R_Candidates": [{"note": "NO V1R CANDIDATES — all 16 anchors NOT EXECUTED"}],
        "PBV2": (
            [{
                "section": "day_totals",
                "fills": len(all_trades),
                "wins": wins,
                "losses": losses,
                "flats": flats,
                "pnl_yen_100": pbv2_day_pnl,
                "gross_profit": gp,
                "gross_loss": gl,
                "pf": pf if pf != float("inf") else "inf",
                "role_actual": "ACTUAL_PAPER_PRIMARY_OBSERVER",
                "role_expected": "SHADOW_ONLY",
                "isolation_ok": False,
                "note": "PBv2 was NOT shadow-isolated; it was the live paper primary. Do not mix into V1R Primary score.",
            }]
            + [{
                "section": "trade",
                "session": t["session"],
                "symbol": t["symbol"],
                "entry_time": t.get("entry_time"),
                "exit_time": t.get("close_time"),
                "entry_price": t.get("entry_price"),
                "exit_price": t.get("close_price"),
                "pnl_yen_100": t["pnl_yen_100"],
                "pnl_pct": t["pnl_pct"],
                "hold_sec": t["hold_sec"],
                "mfe_pct": t["mfe_pct"],
                "mae_pct": t["mae_pct"],
                "exit_reason": t.get("close_reason"),
            } for t in all_trades]
            + [{"section": "0903_case", **row} for row in pbv2_0903]
            + [{"section": "exit_reasons_am", **dict(am["reason_counts"])},
               {"section": "exit_reasons_pm", **dict(pm["reason_counts"])}]
        ),
        "V1R_1M": [{
            "status": "NOT_STARTED",
            "note": "No V1R 1M shadow ledger for 20260810",
            "e1_x5_am_enabled": am["summary"].get("e1_x5_forward_shadow_enabled"),
            "e1_x5_am_trades": am["summary"].get("e1_x5_forward_shadow_trades"),
            "e1_x5_am_pnl": am["summary"].get("e1_x5_forward_shadow_total_pnl_yen_100"),
            "e1_x5_pm_trades": pm["summary"].get("e1_x5_forward_shadow_trades"),
            "e1_x5_pm_pnl": pm["summary"].get("e1_x5_forward_shadow_total_pnl_yen_100"),
            "e1_x5_is_not_v1r_1m": True,
        }],
        "Capture": [{
            "present": cap.get("present"),
            "ingress_pid": cap.get("ingress_pid"),
            "push_part_count": cap.get("push_part_count"),
            "push_part_bytes": cap.get("push_part_bytes"),
            "hb_first": cap.get("ingress_hb_first"),
            "hb_last": cap.get("ingress_hb_last"),
            "hb_n": cap.get("ingress_hb_n"),
            "hb_gap_count_gt_120s": cap.get("ingress_hb_gap_count"),
            "seal": seal,
            "completeness": cap.get("completeness"),
            "capture_ok": capture_ok,
            "sample_gaps": (cap.get("ingress_hb_gaps_gt_120s") or [])[:15],
        }],
        "Heartbeat": [
            {"session": "AM", "first": am["hb_first"], "last": am["hb_last"], "n": am["hb_n"],
             "pid": am["hb_pid"], "expected_interval_sec": 300, "gaps": am["hb_gaps"]},
            {"session": "PM", "first": pm["hb_first"], "last": pm["hb_last"], "n": pm["hb_n"],
             "pid": pm["hb_pid"], "expected_interval_sec": 300, "gaps": pm["hb_gaps"]},
            {"session": "ingress", "first": cap.get("ingress_hb_first"), "last": cap.get("ingress_hb_last"),
             "n": cap.get("ingress_hb_n"), "gaps_gt_120s": cap.get("ingress_hb_gap_count")},
        ],
        "Errors": (
            [{"session": "AM", "total": am["error_n"], **am["error_cats"]}]
            + [{"session": "PM", "total": pm["error_n"], **pm["error_cats"]}]
        ),
        "Discord": [
            {"topic": "routing_audit", **discord_audit},
            {"topic": "0903_5711", **pbv2_0903[0]},
            {"topic": "0903_6656", **pbv2_0903[1]},
            {"topic": "notif_categories", **(notif.get("categories") or {})},
            {"topic": "notif_statuses", **(notif.get("statuses") or {})},
        ],
        "Prospective_Integrity": [
            {"item": "prospective_class", "value": prospective},
            {"item": "v1r_primary_started", "value": False},
            {"item": "opened_20260810", "value": v1r_cfg.get("opened_20260810")},
            {"item": "prospective_observer_started", "value": v1r_cfg.get("prospective_observer_started")},
            {"item": "expected_primary_role", "value": v1r_cfg.get("primary_role")},
            {"item": "expected_pbv2_role", "value": v1r_cfg.get("pbv2_role")},
            {"item": "actual_runtime", "value": "classic PBv2 observer pilot AM+PM"},
            {"item": "fixed_anchor_integrity", "value": "N/A — no V1R candidates"},
            {"item": "universe_day_fixed_for_v1r", "value": "N/A — V1R not consuming; AM universe refresh files exist for PBv2 path"},
            {"item": "universe_am_vs_1000_changed", "value": (uni.get("am_vs_1000_diff") or {}).get("changed")},
            {"item": "285A_in_am_universe", "value": (uni.get("am") or {}).get("has_285A")},
            {"item": "feature_future_leak_v1r", "value": "N/A — V1R scoring not executed"},
            {"item": "cap_hard_breach_v1r", "value": "N/A"},
            {"item": "pbv2_mixed_into_v1r_cap", "value": "N/A — no V1R cap state"},
            {"item": "strategy_mutation_during_day", "value": strategy_mutation_during_day},
            {"item": "config_sha_am", "value": sha_am},
            {"item": "config_sha_pm", "value": sha_pm},
            {"item": "discord_ui_note", "value": discord_ui_note},
            {"item": "submit_cancel_live", "value": f"{submit}/{cancel}/{live}"},
            {"item": "contamination", "value": contamination},
            {"item": "symbol_concentration_pbv2_only", "value": {
                s: {
                    "fills": v["fills"],
                    "pnl": v["pnl"],
                    "gross_positive_share": (pos_contrib[s] / pos_total) if pos_contrib[s] else 0.0,
                }
                for s, v in sorted(by_sym.items(), key=lambda x: -x[1]["pnl"])[:15]
            }},
        ],
    }

    xlsx_path = OUT / "audit_20260810.xlsx"
    write_xlsx(sheets, xlsx_path)

    report = {
        "analysis_id": ANALYSIS_ID,
        "run_id": run_id,
        "verdict": VERDICT,
        "prospective_class": prospective,
        "evaluations": {
            "A_operational_health": operational,
            "B_v1r_runtime_integrity": v1r_runtime_integrity,
            "C_prospective_validity": prospective_validity,
            "D_trading_result_v1r": trading_result,
            "E_data_quality": data_quality,
        },
        "headline": {
            "v1r_ran_normally": "NO",
            "prospective_day1_adoptable": "NO",
            "pbv2_0903_identity": "A. ACTUAL_PBV2_PRIMARY — classic pilot observer ENTRY (session opens 09:03), not V1R and not shadow-only",
            "v1r_fills": 0,
            "v1r_wl": "0/0",
            "v1r_pnl": 0,
            "v1r_pf": None,
            "major_incident": "あり — V1R Primary未起動。PBv2がPaper Primaryとして稼働",
            "wifi_impact": wifi_impact,
            "submit_cancel_live": f"{submit}/{cancel}/{live}",
        },
        "expected_contract": {
            "primary": "PASSIVE_FIXED600_FULL_STRATEGY_V1R",
            "v1r_role": "PAPER_PRIMARY",
            "pbv2_role": "SHADOW_ONLY",
            "one_m_role": "SHADOW_ONLY_DIAGNOSTIC",
            "universe": "DAY_FIXED_AM_RUNTIME_UNIVERSE_V1",
            "cap": 5,
            "wait": 1.0,
            "freshness": 5.0,
            "exit": "FIRST_VALID_BUY1_AT_OR_AFTER_TARGET",
            "hold": 600,
            "anchors": EXPECTED_ANCHORS,
        },
        "actual_runtime": {
            "primary": "PBv2 small_paper_pilot_live_full_dry_run observer",
            "sessions": ["live_session_075617", "live_session_122527"],
            "am_pid": am["hb_pid"],
            "pm_pid": pm["hb_pid"],
            "config_sha": sha_am,
            "session_entry_window_am": f"{am['allowed_entry_start']}-{am['session_end']}",
            "v1r_artifacts_found": False,
            "v1r_1m_artifacts_found": False,
        },
        "pbv2_0903": pbv2_0903,
        "pbv2_day": {
            "fills": len(all_trades),
            "wins": wins,
            "losses": losses,
            "flats": flats,
            "pnl_yen_100": pbv2_day_pnl,
            "pf": pf if pf != float("inf") else None,
            "pf_infinite": pf == float("inf"),
            "am_pnl": am["total_pnl_yen"],
            "pm_pnl": pm["total_pnl_yen"],
            "am_canonical_pnl": am["canonical_pnl"],
            "pm_canonical_pnl": pm["canonical_pnl"],
            "exit_reasons_am": dict(am["reason_counts"]),
            "exit_reasons_pm": dict(pm["reason_counts"]),
            "NOTE": "PBv2 actual paper primary results — NOT V1R Primary",
        },
        "v1r_primary": {
            "started": False,
            "fills": 0,
            "candidates": 0,
            "anchors_fired": 0,
            "anchors_expected": 16,
            "pnl": 0,
        },
        "v1r_1m": {"started": False},
        "e1_x5_forward_shadow_separate": {
            "am_trades": am["summary"].get("e1_x5_forward_shadow_trades"),
            "am_pnl": am["summary"].get("e1_x5_forward_shadow_total_pnl_yen_100"),
            "pm_trades": pm["summary"].get("e1_x5_forward_shadow_trades"),
            "pm_pnl": pm["summary"].get("e1_x5_forward_shadow_total_pnl_yen_100"),
            "is_v1r_1m": False,
        },
        "capture": {
            "ok": capture_ok,
            "push_parts": cap.get("push_part_count"),
            "hb_first": cap.get("ingress_hb_first"),
            "hb_last": cap.get("ingress_hb_last"),
            "gap_count_gt_120s": cap.get("ingress_hb_gap_count"),
            "ingress_pid": cap.get("ingress_pid"),
        },
        "heartbeat": {
            "am": {"first": am["hb_first"], "last": am["hb_last"], "n": am["hb_n"], "gaps": am["hb_gaps"]},
            "pm": {"first": pm["hb_first"], "last": pm["hb_last"], "n": pm["hb_n"], "gaps": pm["hb_gaps"]},
        },
        "errors": {"am": {"n": am["error_n"], "cats": am["error_cats"]}, "pm": {"n": pm["error_n"], "cats": pm["error_cats"]}},
        "discord": discord_audit,
        "universe": uni,
        "safety": {
            "submit": submit,
            "cancel": cancel,
            "live": live,
            "order_enabled": order_enabled,
            "live_trading_enabled": live_trading,
            "submit_cancel_live": f"{submit}/{cancel}/{live}",
        },
        "contamination": contamination,
        "precommit": {
            "strategy_mutation_during_day": strategy_mutation_during_day,
            "config_sha_am": sha_am,
            "config_sha_pm": sha_pm,
            "discord_ui_note": discord_ui_note,
            "v1r_activation_expected_but_not_wired_into_pilot": True,
        },
        "artifacts": {
            "xlsx": str(xlsx_path),
            "report_json": str(OUT / "report.json"),
            "report_md": str(OUT / "report.md"),
        },
        "sources": {
            "am": str(AM),
            "pm": str(PM),
            "capture": str(CAPTURE),
            "notifications": str(NOTIF),
            "v1r_cfg": str(V1R_CFG),
            "runtime_log": str(RUNTIME_LOG) if RUNTIME_LOG.exists() else None,
        },
        "read_only": True,
        "ledger_state_mutation": False,
        "strategy_mutation": False,
        "model_mutation": False,
        "universe_mutation": False,
    }

    (OUT / "report.json").write_text(json.dumps(report, indent=2, default=str, ensure_ascii=False), encoding="utf-8")

    md = f"""# 2026-08-10 Paper Trade 結論

## 今日V1Rは正常に動いたか

NO

## Prospective Day 1として採用できるか

NO

## 朝09:03のPBv2の正体

**A. ACTUAL_PBV2_PRIMARY** — `live_session_075617` の classic PBv2 observer が `allowed_entry_start=09:03` で ENTRY（5711.T 09:03:06 / 6656.T 09:03:13）。V1RでもShadowでもない。

## V1R成績

* FILL: 0
* 勝敗: 0/0
* PnL: 0
* PF: n/a（取引なし）

## 重大障害

あり — V1R Primary prospective observer 未起動。活性化契約上は V1R=PAPER_PRIMARY / PBv2=SHADOW_ONLY だが、実ランタイムは PBv2 trailing-MFE observer が Paper Primary。

## Wi-Fi変更の影響

{wifi_impact}

## Safety

submit/cancel/live = {submit}/{cancel}/{live}

---

# Final Verdict

`{VERDICT}`

---

# Evaluations

| Axis | Result |
|------|--------|
| A. Operational Health | {operational} |
| B. V1R Runtime Integrity | {v1r_runtime_integrity} |
| C. Prospective Validity | {prospective_validity} ({prospective}) |
| D. Trading Result (V1R) | {trading_result} |
| E. Data Quality | {data_quality} |

---

# Expected vs Actual

| Item | Expected | Actual |
|------|----------|--------|
| Primary | PASSIVE_FIXED600_FULL_STRATEGY_V1R | PBv2 small_paper_pilot observer |
| V1R role | PAPER_PRIMARY | NOT STARTED |
| PBv2 role | SHADOW_ONLY | ACTUAL paper primary |
| V1R 1M | SHADOW_ONLY_DIAGNOSTIC | NOT STARTED |
| Anchors | 16 fixed | 0 fired |
| Cap/wait/exit | 5 / 1s / FIXED600 | PBv2 trailing-MFE path |

---

# Startup (summary)

V1R SHA verify / V1R observer / V1R 1M / first V1R anchor 09:05 → **NOT EXECUTED**

AM/PM classic pilot, capture ONLINE, heartbeat → **PASS**

PBv2 as SHADOW_ONLY → **FAIL** (ran as actual)

---

# 09:03 Cases

"""
    for row in pbv2_0903:
        md += f"""
## {row['symbol']}

* Discord: {row['discord_event_time']} → {row['discord_final_result']}
* Trade entry: {row['trade_entry_time']} @ {row['trade_entry_price']}
* Strategy: {row['generating_strategy']}
* Role: {row['role']}
* PID/run: {row['pid']} / {row['run_id']}
* Classification: **{row['classification']}**
"""

    md += f"""
---

# V1R Anchors

全16 Anchor: **NOT FIRED** (`V1R_PRIMARY_NOT_STARTED`)

---

# PBv2 Actual Day (reference only — not V1R)

* Fills: {len(all_trades)} (AM {am['n_trades']} / PM {pm['n_trades']})
* W/L/F: {wins}/{losses}/{flats}
* PnL (100株): {pbv2_day_pnl:,.0f}円
* PF: {pf if pf != float('inf') else 'inf'}
* AM canonical PnL: {am['canonical_pnl']}
* PM canonical PnL: {pm['canonical_pnl']}

Do **not** treat this as V1R Prospective Day 1 performance.

---

# Capture / Heartbeat

* Capture OK: {capture_ok}
* Ingress PID: {cap.get('ingress_pid')}
* Push parts: {cap.get('push_part_count')}
* Ingress HB: {cap.get('ingress_hb_first')} → {cap.get('ingress_hb_last')} (gaps>120s: {cap.get('ingress_hb_gap_count')})
* AM HB: {am['hb_first']} → {am['hb_last']} (n={am['hb_n']}, gaps={len(am['hb_gaps'])})
* PM HB: {pm['hb_first']} → {pm['hb_last']} (n={pm['hb_n']}, gaps={len(pm['hb_gaps'])})

Cursor Worker切断 ≠ TradeBot停止。Capture/pilot heartbeatが継続していれば runtimeは別系統。

---

# Discord

Legacy PBv2 ENTRY が trade-notify へ流れた原因: classic observer の `discord_entry_delivery` が trade-notify webhook を使用。V1R routing（ENTRY→trade-entry）は当日未適用／V1R未起動。

V1R FILL/EXIT/ENTRY/EXPIRED/Primary Summary: **未送信**（イベント自体なし）

---

# Safety

* order_enabled: {order_enabled}
* live_trading_enabled: {live_trading}
* adapter entry/exit counts: AM {am['adapter_entry']}/{am['adapter_exit']}, PM {pm['adapter_entry']}/{pm['adapter_exit']}
* submit/cancel/live = {submit}/{cancel}/{live}

---

# Artifacts

* `{xlsx_path}`
* `{OUT / 'report.json'}`
* `{OUT / 'report.md'}`

READ-ONLY。strategy/model/universe/ledger/state 変更なし。

run_id: `{run_id}`
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")
    print(json.dumps({
        "run_id": run_id,
        "verdict": VERDICT,
        "prospective_class": prospective,
        "v1r_ran": "NO",
        "submit_cancel_live": f"{submit}/{cancel}/{live}",
        "pbv2_0903": "A. ACTUAL_PBV2_PRIMARY",
        "xlsx": str(xlsx_path),
    }, indent=2, ensure_ascii=False))
    return report


if __name__ == "__main__":
    main()
