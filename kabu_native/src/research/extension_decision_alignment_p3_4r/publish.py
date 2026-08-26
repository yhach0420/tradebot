"""Write P3-4R report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.extension_decision_alignment_p3_4r import (
    ANALYSIS_ID,
    DOCUMENT_ID,
    FROZEN_CLUSTER,
    P1_PF,
    P1_PNL,
    P1_TRADES,
    PREDECLARED_TOP3,
    REST11,
    TASK_LABEL,
    VERDICT_BLOCKED,
    VERDICT_ISSUE,
    VERDICT_OK,
)
from research.extension_decision_alignment_p3_4r.metrics import (
    compare_old,
    compare_post,
    compare_pre,
    delay_block,
    gate_verdict,
    incremental_block,
    interpret_mechanism,
    same_anchor,
    slice_days,
)

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "extension_decision_alignment_p3_4r"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
P3_4_REPORT = NATIVE / "results" / "research" / "fixed_winner_cluster_extension_p3_4" / "report.json"


def _file_sha(rel: str) -> str:
    p = NATIVE / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def json_sanitize(obj: Any) -> Any:
    if isinstance(obj, np.generic):
        obj = obj.item()
    if isinstance(obj, float):
        if obj == float("inf"):
            return "Infinity"
        if obj != obj:
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): json_sanitize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_sanitize(v) for v in obj]
    return obj


def _write_rows(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    cols: list[str] = []
    seen = set()
    for row in rows:
        for c in row.keys():
            if c not in seen:
                seen.add(c)
                cols.append(c)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            if v == float("inf"):
                v = "Infinity"
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(52, max(14, len(str(c)) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _kv_sheet(ws, rows: list[tuple[str, Any]]) -> None:
    ws.cell(1, 1, "key").fill = HEADER_FILL
    ws.cell(1, 1).font = HEADER_FONT
    ws.cell(1, 2, "value").fill = HEADER_FILL
    ws.cell(1, 2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False, default=str)
        if isinstance(v, bool):
            v = str(v).lower()
        if v == float("inf"):
            v = "Infinity"
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = WRAP
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 92


def load_frozen_p3_4() -> dict[str, Any]:
    if not P3_4_REPORT.is_file():
        return {"ok": False, "reason": "NO_P3_4_REPORT"}
    p34 = json.loads(P3_4_REPORT.read_text(encoding="utf-8"))
    if p34.get("WINNER_CONCENTRATION") != FROZEN_CLUSTER:
        return {"ok": False, "reason": "P3_4_CLUSTER_NOT_FROZEN_LABEL", "got": p34.get("WINNER_CONCENTRATION")}
    if p34.get("verdict") != "P3_4_WINNER_CLUSTER_EXTENSION_AUDIT_COMPLETE":
        return {"ok": False, "reason": "P3_4_NOT_COMPLETE", "got": p34.get("verdict")}
    return {
        "ok": True,
        "WINNER_CONCENTRATION": p34.get("WINNER_CONCENTRATION"),
        "OVERLAP": p34.get("OVERLAP"),
        "EXTENSION_GATE_P3_4": p34.get("EXTENSION_GATE"),
        "EXTENSION_MECHANISM_P3_4": p34.get("EXTENSION_MECHANISM"),
        "EXTEND_INCREMENTAL_P3_4": p34.get("EXTEND_INCREMENTAL"),
        "EXTEND_VS_EXIT600_P3_4": p34.get("EXTEND_VS_EXIT600"),
        "note": "P3-4 overlap / 16 cells / residuals / cluster not recomputed.",
    }


def build_report(
    *,
    path_rows: list[dict[str, Any]],
    recon: dict[str, Any],
    frozen: dict[str, Any],
    leak_decision: int,
    leak_checkpoint: int,
    overlap_n: int,
    n_decision_ge_750: int,
    identity_n: int,
    identity_fail: int,
    failed: list[str],
    blocked: bool,
    blocked_reason: Optional[str],
) -> dict[str, Any]:
    reached = [r for r in path_rows if r.get("reached_600")]
    n_reached = len(reached)
    n_match = sum(1 for r in reached if r.get("matched") is True)
    n_mis = sum(1 for r in reached if r.get("matched") is False)
    n_path = len(path_rows)
    dec_pass = (not blocked) and n_reached == 164 and n_mis == 0 and n_match == 164

    all_post = compare_post(path_rows) if path_rows else {}
    top_post = compare_post(slice_days(path_rows, PREDECLARED_TOP3)) if path_rows else {}
    rest_post = compare_post(slice_days(path_rows, REST11)) if path_rows else {}
    all_pre = compare_pre(path_rows) if path_rows else {}
    top_pre = compare_pre(slice_days(path_rows, PREDECLARED_TOP3)) if path_rows else {}
    rest_pre = compare_pre(slice_days(path_rows, REST11)) if path_rows else {}
    all_old = compare_old(path_rows) if path_rows else {}
    sa_all = same_anchor(path_rows) if path_rows else {}
    sa_top = same_anchor(slice_days(path_rows, PREDECLARED_TOP3)) if path_rows else {}
    sa_rest = same_anchor(slice_days(path_rows, REST11)) if path_rows else {}
    gate = gate_verdict(all_post, rest_post, sa_rest) if path_rows else {"EXTENSION_GATE_REVISED": None}
    incr = incremental_block(path_rows) if path_rows else {}
    delays = delay_block(path_rows) if path_rows else {}
    interp = interpret_mechanism(incr, gate.get("EXTENSION_GATE_REVISED"), all_pre, all_post) if path_rows else {}

    leak_ok = int(leak_decision) == 0 and int(leak_checkpoint) == 0
    overlap_ok = int(overlap_n) == 0
    ident_ok = int(identity_fail) == 0
    ge750_ok = int(n_decision_ge_750) == 0
    integrity = []
    if blocked:
        integrity.append(blocked_reason or "BLOCKED")
    if not recon.get("pass"):
        integrity.append("CANONICAL_RECONCILE_FAIL")
    if not frozen.get("ok"):
        integrity.append(frozen.get("reason") or "P3_4_FROZEN_MISSING")
    if path_rows and not dec_pass:
        integrity.append(f"DECISION_RECONCILE_{n_match}/{n_reached}")
    if path_rows and n_path != P1_TRADES:
        integrity.append(f"PATH_JOIN_{n_path}/{P1_TRADES}")
    if path_rows and not leak_ok:
        integrity.append("FUTURE_LEAK")
    if path_rows and not overlap_ok:
        integrity.append("DECISION_OUTCOME_OVERLAP")
    if path_rows and not ge750_ok:
        integrity.append(f"DECISION_GE_T750_{n_decision_ge_750}")
    if path_rows and not ident_ok:
        integrity.append(f"VALUE_IDENTITY_FAIL_{identity_fail}/{identity_n}")

    if blocked or not recon.get("pass") or not frozen.get("ok") or (path_rows and not dec_pass):
        verdict = VERDICT_BLOCKED
    elif integrity:
        verdict = VERDICT_ISSUE
    else:
        verdict = VERDICT_OK

    def bid_of(blk: dict, klass: str) -> dict:
        return ((blk.get(klass) or {}).get("decision_to_750_bid") or {})

    sheets = {
        "times": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "anchor_time": r.get("anchor_time"),
                "canonical_class": r.get("canonical_class"),
                "recon_class": r.get("recon_class"),
                "matched": r.get("matched"),
                "fill_time": r.get("fill_time"),
                "nominal_t600": r.get("nominal_t600"),
                "decision_time": r.get("decision_time"),
                "decision_offset_sec": r.get("decision_offset_sec"),
                "decision_delay_sec": r.get("decision_delay_sec"),
                "nominal_t750": r.get("nominal_t750"),
                "decision_ge_t750": r.get("decision_ge_t750"),
            }
            for r in reached
        ],
        "pre": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "canonical_class": r.get("canonical_class"),
                "bid600": r.get("bid600"),
                "bid_decision": r.get("bid_decision"),
                "predecision_bid_return": r.get("predecision_bid_return"),
                "predecision_value_yen": r.get("predecision_value_yen"),
            }
            for r in reached
        ],
        "d750": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "canonical_class": r.get("canonical_class"),
                "primary_evaluable": r.get("primary_evaluable"),
                "bid_decision": r.get("bid_decision"),
                "bid750": r.get("bid750"),
                "mid_decision": r.get("mid_decision"),
                "mid750": r.get("mid750"),
                "decision_to_750_bid_return": r.get("decision_to_750_bid_return"),
                "decision_to_750_mid_return": r.get("decision_to_750_mid_return"),
                "decision_outcome_overlap": r.get("decision_outcome_overlap"),
            }
            for r in reached
        ],
        "p150": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "canonical_class": r.get("canonical_class"),
                "plus150_evaluable": r.get("plus150_evaluable"),
                "decision_plus150_bid_return": r.get("decision_plus150_bid_return"),
                "decision_plus150_mid_return": r.get("decision_plus150_mid_return"),
                "label": "STANDARDIZED_POST_DECISION_150S_DIAGNOSTIC",
            }
            for r in reached
        ],
        "same": (sa_all.get("cells") or [])
        + [{"slice": "TOP3", **c} for c in (sa_top.get("cells") or [])]
        + [{"slice": "REST11", **c} for c in (sa_rest.get("cells") or [])],
        "incr": [
            {
                "trade_id": r.get("trade_id"),
                "date": r.get("date"),
                "symbol": r.get("symbol"),
                "canonical_class": r.get("canonical_class"),
                "old_600_750_value_yen": r.get("old_600_750_value_yen"),
                "predecision_value_yen": r.get("predecision_value_yen"),
                "post_decision_value_yen": r.get("post_decision_value_yen"),
                "identity_pass": r.get("identity_pass"),
                "old_label": "OLD_NOMINAL_600_BASED",
            }
            for r in reached
            if r.get("canonical_class") == "EXTEND_TO_750"
        ],
        "decomp": [
            {"slice": sl, **(incr.get(sl) or {})}
            for sl in ("ALL", "TOP3", "REST11")
        ],
        "top3": [
            {"block": "POST", **top_post},
            {"block": "PRE", **top_pre},
            {"block": "SAME_ANCHOR", **{k: v for k, v in sa_top.items() if k != "cells"}},
            {"block": "INCR", **(incr.get("TOP3") or {})},
        ],
        "rest11": [
            {"block": "POST", **rest_post},
            {"block": "PRE", **rest_pre},
            {"block": "SAME_ANCHOR", **{k: v for k, v in sa_rest.items() if k != "cells"}},
            {"block": "INCR", **(incr.get("REST11") or {})},
        ],
    }

    report = {
        "task": "P3-4R",
        "ANALYSIS_ID": ANALYSIS_ID,
        "DOCUMENT_ID": DOCUMENT_ID,
        "LABEL": TASK_LABEL,
        "NOT": ["OOS", "prospective", "robust", "strategy validation", "P3-4 cluster recompute"],
        "CANONICAL_RECONCILE": "PASS" if recon.get("pass") else "FAIL",
        "reconcile": recon,
        "PRIMARY_FULL14": {"trades": P1_TRADES, "pnl": P1_PNL, "PF": P1_PF},
        "WINNER_CONCENTRATION": FROZEN_CLUSTER,
        "FROZEN_P3_4": frozen,
        "DECISION_RECONCILE": "PASS" if dec_pass else ("FAIL" if path_rows else "NOT_RUN"),
        "extension_decision": {
            "n_reached_600": n_reached,
            "matched": n_match,
            "mismatch": n_mis,
            "n_decision_ge_750": int(n_decision_ge_750),
            "note": "actual decision_time = Dual Lane 600_DECISION capture event_time (pos.t[-1]).",
        },
        "DECISION_DELAY": delays,
        "OLD_600_750": all_old,
        "PREDECISION_MOVE": {"ALL": all_pre, "TOP3": top_pre, "REST11": rest_pre},
        "POST_DECISION_TO_750": {"ALL": all_post, "TOP3": top_post, "REST11": rest_post},
        "SAME_ANCHOR_POST_DECISION": {"ALL": sa_all, "TOP3": sa_top, "REST11": sa_rest},
        "EXTENSION_GATE_REVISED": gate.get("EXTENSION_GATE_REVISED"),
        "EXTENSION_GATE_DETAIL": gate,
        "EXTEND_INCREMENTAL_REVISED": incr,
        "EXTENSION_MECHANISM_REVISED": interp.get("EXTENSION_MECHANISM_REVISED"),
        "EXTENSION_INTERPRETATION": interp,
        "SEPARATE_ANSWERS": {
            "A_already_winner_at_600": interp.get("A_already_winner_at_600"),
            "B_identified_at_actual_decision": interp.get("B_identified_at_actual_decision"),
            "C_post_decision_to_750": interp.get("C_post_decision_to_750"),
            "old_374300_is_not_C": True,
        },
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "DECISION_OUTCOME_OVERLAP": (not overlap_ok) if path_rows else False,
        "FUTURE_LEAK": (not leak_ok) if path_rows else False,
        "decision_future_leak": int(leak_decision),
        "checkpoint_future_leak": int(leak_checkpoint),
        "overlap_n": int(overlap_n),
        "identity_n": int(identity_n),
        "identity_fail": int(identity_fail),
        "SAFETY": {"submit": 0, "cancel": 0, "live": 0},
        "integrity_flags": integrity,
        "failed": failed,
        "blocked_reason": blocked_reason,
        "verdict": verdict,
        "Identity": {
            "ENTRY_SHA": ENTRY_SHA,
            "ANCHOR_SHA": ANCHOR_SHA,
            "EXIT_SHA": EXIT_SHA,
            "STRATEGY_SHA": STRATEGY_SHA,
            "file_sha_dual_lane": _file_sha("src/small_paper/v1r_live_dual_lane.py"),
            "file_sha_p3_4r_replay": _file_sha("src/research/extension_decision_alignment_p3_4r/replay.py"),
        },
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "_sheets": sheets,
        "_bid_all_ext": bid_of(all_post, "EXTEND_TO_750"),
        "_bid_all_exi": bid_of(all_post, "EXIT_AT_600"),
        "_bid_rest_ext": bid_of(rest_post, "EXTEND_TO_750"),
        "_bid_rest_exi": bid_of(rest_post, "EXIT_AT_600"),
    }
    return json_sanitize(report)


def render_md(rep: dict[str, Any]) -> str:
    delay = (rep.get("DECISION_DELAY") or {}).get("ALL") or {}
    old = rep.get("OLD_600_750") or {}
    oe = old.get("EXTEND_TO_750") or {}
    ox = old.get("EXIT_AT_600") or {}
    pre = (rep.get("PREDECISION_MOVE") or {}).get("ALL") or {}
    pe = pre.get("EXTEND_TO_750") or {}
    px = pre.get("EXIT_AT_600") or {}
    post = (rep.get("POST_DECISION_TO_750") or {}).get("ALL") or {}
    e = (post.get("EXTEND_TO_750") or {}).get("decision_to_750_bid") or {}
    x = (post.get("EXIT_AT_600") or {}).get("decision_to_750_bid") or {}
    rest = (rep.get("POST_DECISION_TO_750") or {}).get("REST11") or {}
    re = (rest.get("EXTEND_TO_750") or {}).get("decision_to_750_bid") or {}
    rx = (rest.get("EXIT_AT_600") or {}).get("decision_to_750_bid") or {}
    sa = (rep.get("SAME_ANCHOR_POST_DECISION") or {}).get("ALL") or {}
    incr = (rep.get("EXTEND_INCREMENTAL_REVISED") or {}).get("ALL") or {}
    rest_i = (rep.get("EXTEND_INCREMENTAL_REVISED") or {}).get("REST11") or {}
    dec = rep.get("extension_decision") or {}
    lines = [
        "# P3-4R Extension Decision-Time Alignment",
        "",
        f"LABEL: `{rep.get('LABEL')}`",
        "P3-4 winner cluster frozen. Not recomputed.",
        "",
        f"CANONICAL_RECONCILE: `{rep.get('CANONICAL_RECONCILE')}`",
        f"DECISION_RECONCILE: `{rep.get('DECISION_RECONCILE')}`",
        f"n: {dec.get('n_reached_600')} matched: {dec.get('matched')} mismatch: {dec.get('mismatch')}",
        "",
        "DECISION_DELAY:",
        f"median: {delay.get('median')}",
        f"p90: {delay.get('p90')}",
        f"max: {delay.get('max')}",
        "",
        "OLD_600_750:",
        f"EXTEND mean: {oe.get('mean')} median: {oe.get('median')}",
        f"EXIT600 mean: {ox.get('mean')} median: {ox.get('median')}",
        "",
        "PREDECISION_MOVE:",
        f"EXTEND mean: {pe.get('mean')} median: {pe.get('median')}",
        f"EXIT600 mean: {px.get('mean')} median: {px.get('median')}",
        "",
        "POST_DECISION_TO_750:",
        f"EXTEND n: {e.get('n')} mean: {e.get('mean')} median: {e.get('median')} positive_rate: {e.get('positive_rate')}",
        f"EXIT600 n: {x.get('n')} mean: {x.get('mean')} median: {x.get('median')} positive_rate: {x.get('positive_rate')}",
        "",
        "REST11_POST_DECISION:",
        f"EXTEND: {json.dumps(re, ensure_ascii=False, default=str)}",
        f"EXIT600: {json.dumps(rx, ensure_ascii=False, default=str)}",
        "",
        "SAME_ANCHOR_POST_DECISION:",
        f"better: {sa.get('extend_better')} worse: {sa.get('extend_worse')} equal: {sa.get('equal')}",
        "",
        f"EXTENSION_GATE_REVISED: `{rep.get('EXTENSION_GATE_REVISED')}`",
        "",
        "VALUE_DECOMPOSITION:",
        f"OLD_600_750_VALUE sum: {((incr.get('OLD_NOMINAL_600_BASED') or {}).get('sum'))}",
        f"PREDECISION_VALUE sum: {((incr.get('PREDECISION_VALUE_YEN') or {}).get('sum'))}",
        f"POST_DECISION_VALUE sum: {((incr.get('POST_DECISION_INCREMENTAL_VALUE_YEN') or {}).get('sum'))}",
        f"REST11_POST_DECISION_VALUE sum: {((rest_i.get('POST_DECISION_INCREMENTAL_VALUE_YEN') or {}).get('sum'))}",
        "",
        f"EXTENSION_MECHANISM_REVISED: `{rep.get('EXTENSION_MECHANISM_REVISED')}`",
        f"WINNER_CONCENTRATION: `{rep.get('WINNER_CONCENTRATION')}`",
        "",
        "NEW_STRATEGY_TESTED: false",
        "RETUNING_DONE: false",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "RUNTIME_CHANGED: false",
        f"DECISION_OUTCOME_OVERLAP: {str(rep.get('DECISION_OUTCOME_OVERLAP')).lower()}",
        f"FUTURE_LEAK: {str(rep.get('FUTURE_LEAK')).lower()}",
        "SAFETY: submit/cancel/live=0/0/0",
        "",
        f"verdict: `{rep.get('verdict')}`",
        "",
        "STOP.",
        "",
    ]
    return "\n".join(lines)


def write_artifacts(rep: dict[str, Any]) -> dict[str, str]:
    OUT.mkdir(parents=True, exist_ok=True)
    sheets = rep.pop("_sheets", {})
    rep.pop("_bid_all_ext", None)
    rep.pop("_bid_all_exi", None)
    rep.pop("_bid_rest_ext", None)
    rep.pop("_bid_rest_exi", None)
    public = {k: v for k, v in rep.items() if not str(k).startswith("_")}
    jp = OUT / "report.json"
    mp = OUT / "report.md"
    xp = OUT / "audit.xlsx"
    jp.write_text(json.dumps(public, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    mp.write_text(render_md(public), encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    incr = public.get("EXTEND_INCREMENTAL_REVISED") or {}
    _kv_sheet(
        ws,
        [
            ("LABEL", public.get("LABEL")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("DECISION_RECONCILE", public.get("DECISION_RECONCILE")),
            ("WINNER_CONCENTRATION", public.get("WINNER_CONCENTRATION")),
            ("EXTENSION_GATE_REVISED", public.get("EXTENSION_GATE_REVISED")),
            ("EXTENSION_MECHANISM_REVISED", public.get("EXTENSION_MECHANISM_REVISED")),
            ("DECISION_OUTCOME_OVERLAP", public.get("DECISION_OUTCOME_OVERLAP")),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
            ("OLD_600_750_VALUE_sum", (((incr.get("ALL") or {}).get("OLD_NOMINAL_600_BASED") or {}).get("sum"))),
            ("PREDECISION_VALUE_sum", (((incr.get("ALL") or {}).get("PREDECISION_VALUE_YEN") or {}).get("sum"))),
            ("POST_DECISION_VALUE_sum", (((incr.get("ALL") or {}).get("POST_DECISION_INCREMENTAL_VALUE_YEN") or {}).get("sum"))),
            ("NEW_STRATEGY_TESTED", False),
            ("RUNTIME_CHANGED", False),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("verdict", public.get("verdict")),
        ],
    )
    _write_rows(wb.create_sheet("Decision_Times"), sheets.get("times") or [])
    _write_rows(wb.create_sheet("PreDecision_Move"), sheets.get("pre") or [])
    _write_rows(wb.create_sheet("Decision_to_750"), sheets.get("d750") or [])
    _write_rows(wb.create_sheet("Decision_Plus150"), sheets.get("p150") or [])
    _write_rows(wb.create_sheet("Same_Anchor"), sheets.get("same") or [])
    _write_rows(wb.create_sheet("Incremental_Value"), sheets.get("incr") or [])
    _write_rows(wb.create_sheet("Value_Decomposition"), sheets.get("decomp") or [])
    _write_rows(wb.create_sheet("Top3"), sheets.get("top3") or [])
    _write_rows(wb.create_sheet("Rest11"), sheets.get("rest11") or [])
    ident = wb.create_sheet("Identity")
    _kv_sheet(
        ident,
        list((public.get("Identity") or {}).items())
        + [
            ("identity_n", public.get("identity_n")),
            ("identity_fail", public.get("identity_fail")),
            ("CANONICAL_RECONCILE", public.get("CANONICAL_RECONCILE")),
            ("DECISION_RECONCILE", public.get("DECISION_RECONCILE")),
        ],
    )
    saf = wb.create_sheet("Safety")
    _kv_sheet(
        saf,
        [
            ("submit", 0),
            ("cancel", 0),
            ("live", 0),
            ("SAFETY", "submit/cancel/live=0/0/0"),
            ("NEW_STRATEGY_TESTED", False),
            ("RETUNING_DONE", False),
            ("STRATEGY_CHANGED", False),
            ("ENTRY_EXIT_CHANGED", False),
            ("RUNTIME_CHANGED", False),
            ("DECISION_OUTCOME_OVERLAP", public.get("DECISION_OUTCOME_OVERLAP")),
            ("FUTURE_LEAK", public.get("FUTURE_LEAK")),
        ],
    )
    wb.save(xp)
    return {"report_json": str(jp), "report_md": str(mp), "audit_xlsx": str(xp)}
