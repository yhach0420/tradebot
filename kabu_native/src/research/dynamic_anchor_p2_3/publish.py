"""Aggregate P2-3 decomposition and write report.json / report.md / audit.xlsx only."""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Optional

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_NATIVE = Path(__file__).resolve().parents[3]
if str(_NATIVE / "scripts") not in sys.path:
    sys.path.insert(0, str(_NATIVE / "scripts"))

from research.dynamic_anchor_p2_0b import ANCHOR_SHA, ENTRY_SHA, EXIT_SHA, STRATEGY_SHA
from research.dynamic_anchor_p2_2.binding import ENTRY_BINDING
from research.dynamic_anchor_p2_3 import (
    ANALYSIS_ID,
    BETTER_DAYS,
    CANDIDATE_STATUS,
    CONFIRMATION,
    DOCUMENT_ID,
    FULL14,
    HEADLINE_DELTA,
    HEADLINE_DYNAMIC_PNL,
    HEADLINE_FIXED_PNL,
    P1_CANONICAL_TRADES,
    P2_2_ADMITTED,
    P2_2_CONFIRMED,
    P2_2_EXPIRED,
    P2_2_FILLS,
    P2_2_VERDICT,
    PREDECLARED_TOP3,
    REST11_ACTUAL_DELTA,
    TOP3_ACTUAL_DELTA,
    TRIGGER,
)
from research.dynamic_anchor_p2_3.metrics import (
    ENTRY_ORDER,
    funnel_integrity,
    pct_block,
    pf_out,
    pnl_stats,
    rate,
    trade_match_key,
)
from research.dynamic_anchor_p2_3.state import classify_dynamic_state, lookup_terminal
from run_p0_3_exact_runtime_replay_20260820 import _maxdd
from small_paper.v1r_primary_runtime import WAIT_SEC

NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "dynamic_anchor_failure_decomposition_p2_3"
JST = timezone(timedelta(hours=9))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

ARCH_E_EXITS = {"CONT_EXIT_600", "CONT_EXTEND_750", "IMBALANCE", "SESSION_CLOSE"}
REST11_DAYS = tuple(d for d in FULL14 if d not in PREDECLARED_TOP3)


def _hm_epoch(day: str, hm: str) -> float:
    h, m = hm.split(":")
    return datetime(int(day[:4]), int(day[4:6]), int(day[6:]), int(h), int(m), tzinfo=JST).timestamp()


def _jsonable(v: Any) -> Any:
    if v == float("inf"):
        return "Infinity"
    if isinstance(v, dict):
        return {k: _jsonable(x) for k, x in v.items()}
    if isinstance(v, list):
        return [_jsonable(x) for x in v]
    return v


def _path_metrics(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "count": len(rows),
        "first_ask_minus_limit_bps": pct_block([r.get("first_ask_minus_limit_bps") for r in rows]),
        "min_ask_minus_limit_bps": pct_block([r.get("min_ask_minus_limit_bps") for r in rows]),
    }


def _sess_split(rows: list[dict[str, Any]], key: str = "session") -> dict[str, list[dict[str, Any]]]:
    return {
        "ALL": rows,
        "AM": [r for r in rows if r.get(key) == "AM"],
        "PM": [r for r in rows if r.get(key) == "PM"],
    }


def _fill_split(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    return {
        "ALL": rows,
        "FILLED": [r for r in rows if r.get("fill_terminal") == "FILLED"],
        "EXPIRED": [r for r in rows if r.get("fill_terminal") == "EXPIRED"],
    }


def _pos_neg_days(daily: list[dict[str, Any]], pnl_key: str) -> dict[str, int]:
    pos = sum(1 for r in daily if float(r.get(pnl_key) or 0) > 1e-9)
    neg = sum(1 for r in daily if float(r.get(pnl_key) or 0) < -1e-9)
    return {"positive_days": pos, "negative_days": neg, "flat_days": len(daily) - pos - neg}


def _drivers(
    *,
    delay: dict[str, Any],
    c1: dict[str, Any],
    no_prior: dict[str, Any],
    dyn_fill_rate: Optional[float],
    fix_fill_rate: Optional[float],
    not_selected: int,
    blocked_cap: int,
    rest11_dyn_pf: Any,
    rest11_fix_pf: Any,
) -> dict[str, Any]:
    secondary: list[str] = []
    notes: list[str] = []
    notes.append(
        f"Observed headline Fixed-Dynamic delta {HEADLINE_DELTA:.0f}; "
        f"predeclared Top3 actual delta {TOP3_ACTUAL_DELTA:.0f} "
        f"({100.0 * TOP3_ACTUAL_DELTA / HEADLINE_DELTA:.1f}% of headline). "
        f"Remaining 11 FULL days actual delta {REST11_ACTUAL_DELTA:.0f}."
    )
    primary = "FIXED_TOP3_DEPENDENCY"
    if delay.get("count", 0) >= 20:
        secondary.append("TEN_MINUTE_MATURITY_DELAY")
        notes.append(
            f"FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY n={delay.get('count')} "
            f"Fixed eventual pnl={delay.get('pnl')}."
        )
    if c1.get("count", 0) >= 10:
        secondary.append("CONFIRMATION_FILTER")
        notes.append(
            f"LAST_C1_REJECTED Fixed trades n={c1.get('count')} pnl={c1.get('pnl')} PF={c1.get('PF')}."
        )
    if no_prior.get("count", 0) >= 20:
        secondary.append("TRIGGER_COVERAGE")
        notes.append(
            f"NO_PRIOR_T1_THIS_SESSION Fixed trades n={no_prior.get('count')} pnl={no_prior.get('pnl')}."
        )
    if (
        dyn_fill_rate is not None
        and fix_fill_rate is not None
        and dyn_fill_rate + 0.02 < fix_fill_rate
    ):
        secondary.append("PASSIVE_FILL_RATE_DIFFERENCE")
        notes.append(
            f"Admitted→fill Fixed {fix_fill_rate:.4f} vs Dynamic {dyn_fill_rate:.4f} (WAIT_SEC={WAIT_SEC} unchanged)."
        )
    if not_selected + blocked_cap >= 20:
        secondary.append("CURRENT_ENTRY_SELECTION")
        notes.append(
            f"ENTRY not-admitted among confirmed: NOT_SELECTED={not_selected} BLOCKED_CAP={blocked_cap}."
        )
    if not secondary:
        secondary = ["MIXED"]
    return {
        "PRIMARY_EVIDENCE_SUPPORTED_DRIVER": primary,
        "SECONDARY_DRIVERS": secondary,
        "notes": notes,
        "rest11_pf_note": (
            f"Top3-excluded Fixed PF {pf_out(rest11_fix_pf)} vs Dynamic PF {pf_out(rest11_dyn_pf)}. "
            "Headline PnL alone does not prove general Dynamic-concept failure."
        ),
        "causal_yen_claimed": False,
    }


def build_report(
    *,
    inventory: list[dict[str, Any]],
    day_results: list[dict[str, Any]],
    failed: list[str],
    p1: dict[str, Any],
    p2_2: dict[str, Any],
    sha_bind: dict[str, Any],
    entry_bind: dict[str, Any],
) -> dict[str, Any]:
    blockers: list[str] = []
    material: list[str] = []
    by = {d["date"]: d for d in day_results if d.get("ok")}
    missing = [d for d in FULL14 if d not in by]
    if missing:
        blockers.append(f"MISSING_DAYS:{','.join(missing)}")
    if failed:
        blockers.append("DAY_FAIL:" + ",".join(failed))

    confirms: list[dict[str, Any]] = []
    terminals: list[dict[str, Any]] = []
    dyn_paths: list[dict[str, Any]] = []
    dyn_trades: list[dict[str, Any]] = []
    fixed_paths: list[dict[str, Any]] = []
    fixed_recon_trades: list[dict[str, Any]] = []
    p22_funnel = Counter()
    unresolved = []
    for d in FULL14:
        row = by.get(d)
        if not row:
            continue
        confirms.extend(row.get("confirms") or [])
        terminals.extend(row.get("dyn_terminals") or [])
        dyn_paths.extend(row.get("dyn_paths") or [])
        dyn_trades.extend(row.get("dyn_trades") or [])
        fixed_paths.extend(row.get("fixed_paths") or [])
        fixed_recon_trades.extend(row.get("fixed_trades") or [])
        for k, v in (row.get("dyn_funnel_p22_style") or {}).items():
            p22_funnel[str(k)] += int(v)
        unresolved.extend(row.get("dyn_unresolved") or [])

    p1_daily = {r["date"]: r for r in (p1.get("daily") or []) if r.get("date") in FULL14}
    p1_trades = [t for t in (p1.get("trades") or []) if t.get("date") in FULL14]
    p2_daily = {r["date"]: r for r in ((p2_2.get("DAILY") or {}).get("rows") or [])}

    exclusive = Counter()
    fill_term = Counter()
    for r in terminals:
        exclusive[str(r.get("entry_terminal") or "UNKNOWN")] += 1
        if r.get("entry_terminal") == "ADMITTED":
            fill_term[str(r.get("fill_terminal") or "UNKNOWN")] += 1
    confirmed_n = sum(1 for c in confirms if c.get("status") == "CONFIRMED")
    excl_counts = {k: int(exclusive.get(k, 0)) for k in ENTRY_ORDER}
    excl_counts["confirmed"] = confirmed_n
    excl_counts["fills"] = int(fill_term.get("FILLED", 0))
    excl_counts["expired"] = int(fill_term.get("EXPIRED", 0))
    excl_counts["admitted"] = int(exclusive.get("ADMITTED", 0))
    excl_counts["FILLED"] = int(fill_term.get("FILLED", 0))
    excl_counts["EXPIRED"] = int(fill_term.get("EXPIRED", 0))
    integ = funnel_integrity(excl_counts)
    if not integ["pass"]:
        blockers.append("FUNNEL_IDENTITY_FAIL")
    if unresolved:
        blockers.append(f"ADMITTED_UNRESOLVED:{len(unresolved)}")
    if confirmed_n != P2_2_CONFIRMED:
        blockers.append(f"CONFIRMED_MISMATCH:{confirmed_n}!={P2_2_CONFIRMED}")
    if excl_counts["admitted"] != P2_2_ADMITTED:
        blockers.append(f"ADMITTED_MISMATCH:{excl_counts['admitted']}!={P2_2_ADMITTED}")
    if excl_counts["fills"] != P2_2_FILLS or len(dyn_trades) != P2_2_FILLS:
        blockers.append(f"DYNAMIC_FILLS_MISMATCH:term={excl_counts['fills']} trades={len(dyn_trades)}")
    dyn_pnl = round(sum(float(t.get("pnl_yen_100") or 0) for t in dyn_trades), 2)
    if abs(dyn_pnl - HEADLINE_DYNAMIC_PNL) > 0.5:
        blockers.append(f"DYNAMIC_PNL_MISMATCH:{dyn_pnl}")

    p1_adm = sum(int(p1_daily[d].get("admitted") or 0) for d in FULL14 if d in p1_daily)
    p1_fills = sum(int(p1_daily[d].get("fills") or 0) for d in FULL14 if d in p1_daily)
    p1_exp = sum(int(p1_daily[d].get("expired") or 0) for d in FULL14 if d in p1_daily)
    p1_cand = sum(int(p1_daily[d].get("candidates_scored") or 0) for d in FULL14 if d in p1_daily)
    recon_adm = sum(int((by[d].get("fixed_admitted") or 0)) for d in FULL14 if d in by)
    recon_fills = sum(int((by[d].get("fixed_fills") or 0)) for d in FULL14 if d in by)
    recon_exp = sum(int((by[d].get("fixed_expired") or 0)) for d in FULL14 if d in by)
    recon_cand = sum(int((by[d].get("fixed_candidates_scored") or 0)) for d in FULL14 if d in by)
    recon_sel = sum(int((by[d].get("fixed_selected") or 0)) for d in FULL14 if d in by)

    p1_keys = {trade_match_key(t) for t in p1_trades}
    recon_keys = {trade_match_key(t) for t in fixed_recon_trades}
    fixed_match = p1_keys == recon_keys and len(p1_trades) == P1_CANONICAL_TRADES
    if len(p1_trades) != P1_CANONICAL_TRADES:
        blockers.append(f"P1_TRADE_COUNT:{len(p1_trades)}")
    if not fixed_match:
        blockers.append(
            f"FIXED_TRADE_MISMATCH:p1={len(p1_keys)} recon={len(recon_keys)} "
            f"only_p1={len(p1_keys - recon_keys)} only_recon={len(recon_keys - p1_keys)}"
        )
    if recon_fills != P1_CANONICAL_TRADES:
        blockers.append(f"FIXED_FILL_COUNT:{recon_fills}")
    if recon_adm != p1_adm or recon_exp != p1_exp:
        blockers.append(f"FIXED_FUNNEL_MISMATCH:adm {recon_adm}/{p1_adm} exp {recon_exp}/{p1_exp}")

    sha_days = []
    for d in FULL14:
        if d not in by or d not in p1_daily:
            continue
        ok_sha = str(by[d].get("fixed_ledger_sha") or "") == str(p1_daily[d].get("ledger_sha") or "")
        sha_days.append({"date": d, "match": ok_sha})
        if not ok_sha:
            blockers.append(f"FIXED_LEDGER_SHA:{d}")

    term_idx: dict[tuple[str, str, Optional[float]], dict[str, Any]] = {}
    for r in terminals:
        t1 = r.get("t1")
        if t1 is None:
            continue
        term_idx[(str(r.get("date")), str(r.get("symbol")), round(float(t1), 6))] = r

    confirms_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for c in confirms:
        confirms_by_day[str(c.get("date"))].append(c)

    state_rows: list[dict[str, Any]] = []
    for t in p1_trades:
        day = str(t.get("date"))
        hm = str(t.get("anchor_time") or "")
        try:
            sig = _hm_epoch(day, hm)
        except Exception:
            sig = float(t.get("fill_time") or 0.0)
        sess = str(t.get("session") or ("AM" if datetime.fromtimestamp(sig, tz=JST).hour < 12 else "PM"))
        st = classify_dynamic_state(
            date=day,
            symbol=str(t.get("symbol")),
            session=sess,
            signal_t=sig,
            confirms=confirms_by_day.get(day, []),
        )
        outc = lookup_terminal(
            term_idx,
            date=day,
            symbol=str(t.get("symbol")),
            t1=st.get("latest_t1") if st.get("c1_status") == "CONFIRMED" and not st.get("entry_during_anchor_active") else (
                st.get("latest_t1") if st.get("c1_status") == "CONFIRMED" else None
            ),
        )
        if st.get("entry_during_anchor_active"):
            outc = {"entry_terminal": None, "fill_terminal": None, "canonical_terminal_outcome": None}
        state_rows.append({
            "date": day,
            "symbol": t.get("symbol"),
            "session": sess,
            "fixed_anchor_time": hm,
            "fixed_signal_time": sig,
            "fixed_fill_time": t.get("fill_time"),
            "fixed_fill_price": t.get("fill_price"),
            "fixed_exit_time": t.get("exit_time"),
            "fixed_exit_price": t.get("exit_price"),
            "fixed_pnl": t.get("pnl_yen_100"),
            "dynamic_state_at_fixed_signal": st["primary_state"],
            "latest_t0": st.get("latest_t0"),
            "latest_t1": st.get("latest_t1"),
            "c1_status": st.get("c1_status"),
            "c1_reason": st.get("c1_reason"),
            "fixed_signal_minus_t0_sec": st.get("fixed_signal_minus_t0_sec"),
            "fixed_signal_minus_t1_sec": st.get("fixed_signal_minus_t1_sec"),
            "has_prior_t1": st.get("has_prior_t1"),
            "entry_during_anchor_active": st.get("entry_during_anchor_active"),
            "prior_c1_rejected": st.get("prior_c1_rejected"),
            "c1_confirmed_before_entry": st.get("c1_confirmed_before_entry"),
            "dynamic_entry_terminal": outc.get("entry_terminal"),
            "dynamic_fill_terminal": outc.get("fill_terminal"),
            "dynamic_canonical_terminal_outcome": outc.get("canonical_terminal_outcome"),
        })

    delay_rows = [r for r in state_rows if r.get("entry_during_anchor_active")]
    c1_rows = [r for r in state_rows if r.get("prior_c1_rejected")]
    no_prior_rows = [r for r in state_rows if r.get("dynamic_state_at_fixed_signal") == "NO_PRIOR_T1_THIS_SESSION"]
    delay_stats = pnl_stats(delay_rows, pnl_key="fixed_pnl")
    c1_stats = pnl_stats(c1_rows, pnl_key="fixed_pnl")
    no_prior_stats = pnl_stats(no_prior_rows, pnl_key="fixed_pnl")
    delay_stats["PF"] = pf_out(delay_stats["PF"])
    c1_stats["PF"] = pf_out(c1_stats["PF"])

    top3_state = [r for r in state_rows if r.get("date") in PREDECLARED_TOP3]
    rest_state = [r for r in state_rows if r.get("date") in REST11_DAYS]
    top3_delay = [r for r in top3_state if r.get("entry_during_anchor_active")]
    top3_c1 = [r for r in top3_state if r.get("prior_c1_rejected")]
    top3_prior = [r for r in top3_state if r.get("has_prior_t1")]
    top3_conf = [r for r in top3_state if r.get("c1_confirmed_before_entry")]
    top3_dyn_term = [r for r in terminals if r.get("date") in PREDECLARED_TOP3]
    top3_dyn_adm = [r for r in top3_dyn_term if r.get("entry_terminal") == "ADMITTED"]
    top3_dyn_fill = [r for r in top3_dyn_term if r.get("fill_terminal") == "FILLED"]
    top3_fix_tr = [t for t in p1_trades if t.get("date") in PREDECLARED_TOP3]
    top3_dyn_tr = [t for t in dyn_trades if t.get("date") in PREDECLARED_TOP3]
    rest_fix_tr = [t for t in p1_trades if t.get("date") in REST11_DAYS]
    rest_dyn_tr = [t for t in dyn_trades if t.get("date") in REST11_DAYS]
    top3_fix_pnl = round(sum(float(t.get("pnl_yen_100") or 0) for t in top3_fix_tr), 2)
    top3_dyn_pnl = round(sum(float(t.get("pnl_yen_100") or 0) for t in top3_dyn_tr), 2)
    top3_delta = round(top3_fix_pnl - top3_dyn_pnl, 2)
    if abs(top3_delta - TOP3_ACTUAL_DELTA) > 0.5:
        material.append(f"TOP3_DELTA_RECOMPUTE:{top3_delta} expected {TOP3_ACTUAL_DELTA}")

    rest_fix = pnl_stats(rest_fix_tr)
    rest_dyn = pnl_stats(rest_dyn_tr)
    rest_fix["maxDD"] = _maxdd(rest_fix_tr)
    rest_dyn["maxDD"] = _maxdd(rest_dyn_tr)
    rest_daily = []
    for d in REST11_DAYS:
        fd = p2_daily.get(d) or p1_daily.get(d) or {}
        rest_daily.append({
            "date": d,
            "Fixed_pnl": float((p2_daily.get(d) or {}).get("Fixed_pnl") or (p1_daily.get(d) or {}).get("pnl") or 0),
            "Dynamic_pnl": float((p2_daily.get(d) or {}).get("Dynamic_pnl") or 0),
        })
    rest_fix.update(_pos_neg_days(rest_daily, "Fixed_pnl"))
    rest_dyn.update(_pos_neg_days(rest_daily, "Dynamic_pnl"))

    def _am_pm_fill(paths: list[dict[str, Any]]) -> dict[str, Any]:
        out = {}
        for sess, rs in _sess_split(paths).items():
            adm = len(rs)
            fills = sum(1 for r in rs if r.get("fill_terminal") == "FILLED")
            exp = sum(1 for r in rs if r.get("fill_terminal") == "EXPIRED")
            out[sess] = {
                "admitted": adm,
                "fills": fills,
                "expired": exp,
                "fill_rate": rate(fills, adm),
                "expired_rate": rate(exp, adm),
            }
        return out

    dyn_fill_tbl = _am_pm_fill(dyn_paths)
    fix_fill_tbl = _am_pm_fill(fixed_paths)
    # Engine counters + P1 canonical fills are the fill-rate SoT.
    # Path FILLED labels can undercount when a_fills.anchor is rewritten at the next CLOCK_GRID harvest.
    am_p1 = sum(1 for t in p1_trades if t.get("session") == "AM")
    pm_p1 = sum(1 for t in p1_trades if t.get("session") == "PM")
    fix_fill_tbl["ALL"]["admitted"] = recon_adm
    fix_fill_tbl["ALL"]["fills"] = recon_fills
    fix_fill_tbl["ALL"]["expired"] = recon_exp
    fix_fill_tbl["ALL"]["fill_rate"] = rate(recon_fills, recon_adm)
    fix_fill_tbl["ALL"]["expired_rate"] = rate(recon_exp, recon_adm)
    if fix_fill_tbl["AM"]["admitted"]:
        fix_fill_tbl["AM"]["fills"] = am_p1
        fix_fill_tbl["AM"]["expired"] = int(fix_fill_tbl["AM"]["admitted"]) - am_p1
        fix_fill_tbl["AM"]["fill_rate"] = rate(am_p1, int(fix_fill_tbl["AM"]["admitted"]))
        fix_fill_tbl["AM"]["expired_rate"] = rate(fix_fill_tbl["AM"]["expired"], int(fix_fill_tbl["AM"]["admitted"]))
    if fix_fill_tbl["PM"]["admitted"]:
        fix_fill_tbl["PM"]["fills"] = pm_p1
        fix_fill_tbl["PM"]["expired"] = int(fix_fill_tbl["PM"]["admitted"]) - pm_p1
        fix_fill_tbl["PM"]["fill_rate"] = rate(pm_p1, int(fix_fill_tbl["PM"]["admitted"]))
        fix_fill_tbl["PM"]["expired_rate"] = rate(fix_fill_tbl["PM"]["expired"], int(fix_fill_tbl["PM"]["admitted"]))
    fix_fill_tbl["note"] = (
        "fill_rate uses P1 canonical fills / Current admitted. "
        "Fixed FILLED/EXPIRED ask-path split in PASSIVE_FILL may undercount FILLED labels; "
        "ALL-admitted ask-path metrics remain the wait-window description."
    )

    def _passive_block(paths: list[dict[str, Any]]) -> dict[str, Any]:
        out = {}
        for sess, rs in _sess_split(paths).items():
            fs = _fill_split(rs)
            out[sess] = {k: _path_metrics(v) for k, v in fs.items()}
        return out

    dyn_passive = _passive_block(dyn_paths)
    fix_passive = _passive_block(fixed_paths)

    def _c1_move(paths: list[dict[str, Any]]) -> dict[str, Any]:
        out = {}
        for lab, rs in _fill_split(paths).items():
            out[lab] = {
                "count": len(rs),
                "median_endpoint_return": pct_block([r.get("endpoint_return") for r in rs])["median"],
                "median_trend_slope": pct_block([r.get("trend_slope") for r in rs])["median"],
                "median_bid_move_bps": pct_block([r.get("bid_move_bps") for r in rs])["median"],
            }
        return out

    dyn_move = _c1_move(dyn_paths)

    better_rows = []
    for d in BETTER_DAYS:
        ft = [t for t in p1_trades if t.get("date") == d]
        dt = [t for t in dyn_trades if t.get("date") == d]
        f_sym = {str(t.get("symbol")) for t in ft}
        d_sym = {str(t.get("symbol")) for t in dt}
        only_f = sorted(f_sym - d_sym)
        only_d = sorted(d_sym - f_sym)
        both = sorted(f_sym & d_sym)
        fixed_losers_no_dyn = [
            {"symbol": t.get("symbol"), "pnl": t.get("pnl_yen_100"), "anchor_time": t.get("anchor_time")}
            for t in ft if str(t.get("symbol")) in set(only_f) and float(t.get("pnl_yen_100") or 0) < -1e-9
        ]
        dyn_only_winners = [
            {"symbol": t.get("symbol"), "pnl": t.get("pnl_yen_100"), "t1": t.get("t1")}
            for t in dt if str(t.get("symbol")) in set(only_d) and float(t.get("pnl_yen_100") or 0) > 1e-9
        ]
        timing = []
        for s in both:
            f0 = min((t for t in ft if str(t.get("symbol")) == s), key=lambda t: float(t.get("fill_time") or 0))
            d0 = min((t for t in dt if str(t.get("symbol")) == s), key=lambda t: float(t.get("fill_time") or 0))
            timing.append({
                "symbol": s,
                "fixed_fill_time": f0.get("fill_time"),
                "dynamic_fill_time": d0.get("fill_time"),
                "fill_time_diff_sec": round(float(d0.get("fill_time") or 0) - float(f0.get("fill_time") or 0), 3),
                "fixed_pnl": f0.get("pnl_yen_100"),
                "dynamic_pnl": d0.get("pnl_yen_100"),
            })
        p2r = p2_daily.get(d) or {}
        better_rows.append({
            "date": d,
            "Fixed_trades": len(ft),
            "Dynamic_trades": len(dt),
            "Fixed_pnl": p2r.get("Fixed_pnl"),
            "Dynamic_pnl": p2r.get("Dynamic_pnl"),
            "delta_pnl_dynamic_minus_fixed": p2r.get("delta_pnl"),
            "lower_exposure": len(dt) < len(ft),
            "fixed_only_symbols": only_f,
            "dynamic_only_symbols": only_d,
            "both_symbols": both,
            "fixed_loser_symbols_without_dynamic_fill": fixed_losers_no_dyn,
            "dynamic_only_winners": dyn_only_winners,
            "same_symbol_timing": timing,
            "note": "Descriptive ledger overlap only. Not a counterfactual yen attribution.",
        })

    exit_ctr = Counter(str(t.get("exit_reason") or "") for t in dyn_trades)
    exit_rows = []
    for reason, n in sorted(exit_ctr.items(), key=lambda x: -x[1]):
        sub = [t for t in dyn_trades if str(t.get("exit_reason") or "") == reason]
        stt = pnl_stats(sub)
        holds = [float(t["holding_sec"]) for t in sub if t.get("holding_sec") is not None]
        exit_rows.append({
            "exit_reason": reason,
            "count": n,
            "pnl": stt["pnl"],
            "median_holding_sec": pct_block(holds)["median"],
            "win": stt["win"],
            "loss": stt["loss"],
            "draw": stt["draw"],
        })
    fixed_exit_reasons = {str(t.get("exit_reason") or "") for t in p1_trades}
    dyn_exit_reasons = {str(t.get("exit_reason") or "") for t in dyn_trades}
    unknown = sorted(r for r in dyn_exit_reasons if r and r not in ARCH_E_EXITS and r not in fixed_exit_reasons)
    neg_hold = [t for t in dyn_trades if t.get("holding_sec") is not None and float(t["holding_sec"]) < -1e-6]
    if unknown:
        material.append("EXIT_UNKNOWN_REASON:" + ",".join(unknown))
    if neg_hold:
        material.append(f"EXIT_NEGATIVE_HOLDING:{len(neg_hold)}")

    same_sym_exit = []
    f_map: dict[tuple[str, str], list] = defaultdict(list)
    for t in p1_trades:
        f_map[(str(t.get("date")), str(t.get("symbol")))].append(t)
    for t in dyn_trades:
        fs = f_map.get((str(t.get("date")), str(t.get("symbol"))), [])
        same_sym_exit.append({
            "date": t.get("date"),
            "symbol": t.get("symbol"),
            "dynamic_exit_reason": t.get("exit_reason"),
            "dynamic_holding_sec": t.get("holding_sec"),
            "dynamic_pnl": t.get("pnl_yen_100"),
            "fixed_n_same_symbol_day": len(fs),
            "fixed_exit_reasons": ",".join(sorted({str(x.get("exit_reason") or "") for x in fs})) if fs else None,
            "fixed_pnl_sum": round(sum(float(x.get("pnl_yen_100") or 0) for x in fs), 2) if fs else None,
        })

    drivers = _drivers(
        delay=delay_stats,
        c1=c1_stats,
        no_prior=no_prior_stats,
        dyn_fill_rate=dyn_fill_tbl["ALL"]["fill_rate"],
        fix_fill_rate=fix_fill_tbl["ALL"]["fill_rate"],
        not_selected=int(exclusive.get("NOT_SELECTED", 0)),
        blocked_cap=int(exclusive.get("BLOCKED_CAP", 0)),
        rest11_dyn_pf=rest_dyn["PF"],
        rest11_fix_pf=rest_fix["PF"],
    )

    funnel_note = (
        "P2-2 candidate_selected counts simulate_joint admitted flags (not exclusive). "
        "P2-2 blocked_cap is the residual after admitted/pending/open/feature-fail; "
        "it mixes joint CAPACITY_BLOCKED (NOT_SELECTED here) and live CAP (BLOCKED_CAP here). "
        "EXIT policy is identical (Arch E); exit-path differences are descriptive of entry timing only."
    )

    if blockers:
        verdict = "P2_3_BLOCKED"
        funnel_acc = "FAIL"
    elif material:
        verdict = "P2_3_MATERIAL_ACCOUNTING_ISSUE_FOUND"
        funnel_acc = "PASS" if integ["pass"] else "FAIL"
    else:
        verdict = "P2_3_FAILURE_DECOMPOSITION_COMPLETE"
        funnel_acc = "PASS"

    now = datetime.now(JST).isoformat(timespec="seconds")
    state_ctr = Counter(r["dynamic_state_at_fixed_signal"] for r in state_rows)
    delay_top3 = pnl_stats(top3_delay, pnl_key="fixed_pnl")

    p22_funnel_pub = {
        "confirmed": confirmed_n,
        "candidate_selected": int(p22_funnel.get("candidate_selected", 0)),
        "blocked_open": int(p22_funnel.get("blocked_open", 0)),
        "blocked_pending": int(p22_funnel.get("blocked_pending", 0)),
        "blocked_cap": int(p22_funnel.get("blocked_cap", 0)),
        "other_reject": int(p22_funnel.get("other_reject", 0)),
        "admitted": int(p22_funnel.get("admitted", 0)),
        "fills": sum(int(by[d].get("dyn_fills_engine") or 0) for d in FULL14 if d in by),
        "expired": sum(int(by[d].get("dyn_expired_engine") or 0) for d in FULL14 if d in by),
    }

    rep = {
        "task": "P2-3",
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "generated_at_jst": now,
        "P2_2_VERDICT": P2_2_VERDICT,
        "CANDIDATE_STATUS": CANDIDATE_STATUS,
        "TRIGGER": TRIGGER,
        "CONFIRMATION": CONFIRMATION,
        "FUNNEL_ACCOUNTING": funnel_acc,
        "funnel_note": funnel_note,
        "p2_2_published_funnel": p2_2.get("FUNNEL"),
        "p2_2_style_reconstructed": p22_funnel_pub,
        "DYNAMIC_TERMINAL_OUTCOMES": {
            "confirmed": confirmed_n,
            "not_selected": int(exclusive.get("NOT_SELECTED", 0)),
            "blocked_open": int(exclusive.get("BLOCKED_OPEN", 0)),
            "blocked_pending": int(exclusive.get("BLOCKED_PENDING", 0)),
            "blocked_cap": int(exclusive.get("BLOCKED_CAP", 0)),
            "other_reject": int(exclusive.get("OTHER_REJECT", 0)),
            "admitted": int(exclusive.get("ADMITTED", 0)),
            "fills": int(fill_term.get("FILLED", 0)),
            "expired": int(fill_term.get("EXPIRED", 0)),
        },
        "funnel_integrity": integ,
        "FIXED_FUNNEL": {
            "p1_frozen": {
                "candidates_scored": p1_cand,
                "admitted": p1_adm,
                "fills": p1_fills,
                "expired": p1_exp,
            },
            "reconstructed": {
                "candidate_evaluated": recon_cand,
                "selected": recon_sel,
                "admitted": recon_adm,
                "fills": recon_fills,
                "expired": recon_exp,
                "fill_rate": rate(recon_fills, recon_adm),
                "expired_rate": rate(recon_exp, recon_adm),
            },
            "match_p1_canonical_trades": fixed_match,
            "ledger_sha_days": sha_days,
            "AM_PM": fix_fill_tbl,
        },
        "FIXED_FILL_RATE": fix_fill_tbl["ALL"]["fill_rate"],
        "DYNAMIC_FILL_RATE": dyn_fill_tbl["ALL"]["fill_rate"],
        "fill_rate_by_session": {"Fixed": fix_fill_tbl, "Dynamic": dyn_fill_tbl},
        "FIXED_TRADES_ANALYZED": len(state_rows),
        "dynamic_state_counts": dict(state_ctr),
        "FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY": {
            "count": delay_stats["count"],
            "fixed_pnl": delay_stats["pnl"],
            "PF": delay_stats["PF"],
            "gross_profit": delay_stats["gross_profit"],
            "gross_loss": delay_stats["gross_loss"],
            "win": delay_stats["win"],
            "loss": delay_stats["loss"],
            "draw": delay_stats["draw"],
            "definition": "t0 <= fixed_signal_time < t1 on the same-session P2-1 anchor. Not 'profit taken before t1'.",
            "top3": {
                "count": delay_top3["count"],
                "fixed_pnl": delay_top3["pnl"],
                "PF": pf_out(delay_top3["PF"]),
                "win": delay_top3["win"],
                "loss": delay_top3["loss"],
                "draw": delay_top3["draw"],
            },
        },
        "C1_REJECTED_FIXED_TRADES": {
            "count": c1_stats["count"],
            "fixed_pnl": c1_stats["pnl"],
            "PF": c1_stats["PF"],
            "win": c1_stats["win"],
            "loss": c1_stats["loss"],
            "draw": c1_stats["draw"],
        },
        "TRIGGER_COVERAGE_NO_PRIOR_T1": {
            "count": no_prior_stats["count"],
            "fixed_pnl": no_prior_stats["pnl"],
            "PF": pf_out(no_prior_stats["PF"]),
        },
        "TOP3": {
            "days": list(PREDECLARED_TOP3),
            "actual_fixed_pnl": top3_fix_pnl,
            "actual_dynamic_pnl": top3_dyn_pnl,
            "actual_delta": TOP3_ACTUAL_DELTA,
            "recomputed_delta": top3_delta,
            "fixed_trades": len(top3_fix_tr),
            "fixed_trades_with_prior_T1": len(top3_prior),
            "fixed_pnl_with_prior_T1": pnl_stats(top3_prior, pnl_key="fixed_pnl")["pnl"],
            "fixed_entries_during_dynamic_wait": len(top3_delay),
            "fixed_pnl_during_dynamic_wait": delay_top3["pnl"],
            "fixed_trades_prior_C1_rejected": len(top3_c1),
            "fixed_pnl_prior_C1_rejected": pnl_stats(top3_c1, pnl_key="fixed_pnl")["pnl"],
            "fixed_trades_c1_confirmed_before_entry": len(top3_conf),
            "fixed_pnl_c1_confirmed_before_entry": pnl_stats(top3_conf, pnl_key="fixed_pnl")["pnl"],
            "dynamic_admitted": len(top3_dyn_adm),
            "dynamic_filled": len(top3_dyn_fill),
            "classes_overlap": True,
            "note": "Class PnLs are not additive to TOP3_ACTUAL_DELTA.",
        },
        "REST11": {
            "days": list(REST11_DAYS),
            "Fixed": {
                "pnl": rest_fix["pnl"],
                "PF": pf_out(rest_fix["PF"]),
                "maxDD": rest_fix["maxDD"],
                "trades": rest_fix["count"],
                "gross_profit": rest_fix["gross_profit"],
                "gross_loss": rest_fix["gross_loss"],
                "positive_days": rest_fix["positive_days"],
                "negative_days": rest_fix["negative_days"],
            },
            "Dynamic": {
                "pnl": rest_dyn["pnl"],
                "PF": pf_out(rest_dyn["PF"]),
                "maxDD": rest_dyn["maxDD"],
                "trades": rest_dyn["count"],
                "gross_profit": rest_dyn["gross_profit"],
                "gross_loss": rest_dyn["gross_loss"],
                "positive_days": rest_dyn["positive_days"],
                "negative_days": rest_dyn["negative_days"],
            },
            "actual_delta": REST11_ACTUAL_DELTA,
        },
        "PASSIVE_FILL": {
            "WAIT_SEC": WAIT_SEC,
            "Fixed": fix_passive,
            "Dynamic": dyn_passive,
            "Fixed_median_min_ask_minus_limit_bps": (fix_passive.get("ALL") or {}).get("ALL", {}).get("min_ask_minus_limit_bps", {}).get("median"),
            "Dynamic_median_min_ask_minus_limit_bps": (dyn_passive.get("ALL") or {}).get("ALL", {}).get("min_ask_minus_limit_bps", {}).get("median"),
        },
        "T0_T1_MOVE": {
            "Dynamic_FILLED_median_t0_t1_return": dyn_move["FILLED"]["median_endpoint_return"],
            "Dynamic_EXPIRED_median_t0_t1_return": dyn_move["EXPIRED"]["median_endpoint_return"],
            "Dynamic_FILLED_median_slope": dyn_move["FILLED"]["median_trend_slope"],
            "Dynamic_EXPIRED_median_slope": dyn_move["EXPIRED"]["median_trend_slope"],
            "detail": dyn_move,
        },
        "BETTER_DAYS": better_rows,
        "EXIT_DESCRIPTION": {
            "policy": "EXIT policy is identical (Arch E 600/750 + IMBALANCE + SESSION_CLOSE). Not treated as a policy driver.",
            "by_reason": exit_rows,
            "unknown_dynamic_reasons": unknown,
            "negative_holding_n": len(neg_hold),
        },
        "PRIMARY_EVIDENCE_SUPPORTED_DRIVER": drivers["PRIMARY_EVIDENCE_SUPPORTED_DRIVER"],
        "SECONDARY_DRIVERS": drivers["SECONDARY_DRIVERS"],
        "FIXED_TOP3_DEPENDENCY_ROLE": {
            "total_actual_fixed_dynamic_delta": HEADLINE_DELTA,
            "predeclared_top3_actual_delta": TOP3_ACTUAL_DELTA,
            "remaining_11_full_day_actual_delta": REST11_ACTUAL_DELTA,
            "top3_share_of_headline_delta": TOP3_ACTUAL_DELTA / HEADLINE_DELTA,
            "rest11_fixed_PF": pf_out(rest_fix["PF"]),
            "rest11_dynamic_PF": pf_out(rest_dyn["PF"]),
        },
        "DYNAMIC_CONCEPT_GENERAL_FAILURE_PROVEN": False,
        "NEW_STRATEGY_TESTED": False,
        "RETUNING_DONE": False,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "RUNTIME_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "blockers": blockers,
        "material_issues": material,
        "verdict": verdict,
        "identity": {
            "strategy_sha": STRATEGY_SHA,
            "entry_sha": ENTRY_SHA,
            "exit_sha": EXIT_SHA,
            "anchor_sha": ANCHOR_SHA,
            "V1RNativeEntryLive_sha": entry_bind.get("V1RNativeEntryLive_sha"),
            "V1RLiveDualLane_sha": entry_bind.get("V1RLiveDualLane_sha"),
            "P2_1_TRIGGER_SHA_MATCH": sha_bind.get("P2_1_TRIGGER_SHA_MATCH"),
            "P2_1_CONFIRM_SHA_MATCH": sha_bind.get("P2_1_CONFIRM_SHA_MATCH"),
            "CURRENT_ENTRY_BINDING": entry_bind.get("CURRENT_ENTRY_BINDING"),
            "entry_binding_path": ENTRY_BINDING,
        },
        "_state_rows": state_rows,
        "_delay_rows": delay_rows,
        "_c1_rows": c1_rows,
        "_top3_state": top3_state,
        "_same_sym_exit": same_sym_exit,
        "_drivers": drivers,
        "_rest_state": rest_state,
        "headline_fixed_pnl": HEADLINE_FIXED_PNL,
        "headline_dynamic_pnl": HEADLINE_DYNAMIC_PNL,
        "p1_recon_trade_n": len(fixed_recon_trades),
        "dyn_recon_trade_n": len(dyn_trades),
        "inventory_full14": [r for r in inventory if r.get("date") in FULL14],
    }
    return rep


def _write_rows(ws, rows: list[dict[str, Any]], cap: int | None = None) -> None:
    if not rows:
        ws["A1"] = "(empty)"
        return
    data = rows[:cap] if cap else rows
    cols = list(data[0].keys())
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(data, 2):
        for j, c in enumerate(cols, 1):
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)
            if isinstance(v, bool):
                v = str(v).lower()
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(42, max(12, len(c) + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _md(rep: dict[str, Any]) -> str:
    t = rep["DYNAMIC_TERMINAL_OUTCOMES"]
    ff = rep["FIXED_FUNNEL"]["reconstructed"]
    dly = rep["FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY"]
    c1 = rep["C1_REJECTED_FIXED_TRADES"]
    top = rep["TOP3"]
    r11 = rep["REST11"]
    pf = rep["PASSIVE_FILL"]
    mv = rep["T0_T1_MOVE"]
    return f"""# P2-3 Dynamic Anchor failure decomposition

- generated_at_jst: `{rep['generated_at_jst']}`
- **verdict: `{rep['verdict']}`**
- P2_2_VERDICT: `{rep['P2_2_VERDICT']}`
- CANDIDATE_STATUS: `{rep['CANDIDATE_STATUS']}`
- FUNNEL_ACCOUNTING: `{rep['FUNNEL_ACCOUNTING']}`
- submit/cancel/live: `0/0/0`

Frozen candidate T1_VOLP60_GE_DESIGN_Q80 + C1_POSITIVE_TREND_10M_V1 is **REJECTED_AS_FIXED_REPLACEMENT**.
This document does not retune, does not create a new Dynamic candidate, and does not retest PnL as a strategy.

## Funnel (exclusive)

confirmed {t['confirmed']}
NOT_SELECTED {t['not_selected']}
BLOCKED_OPEN {t['blocked_open']}
BLOCKED_PENDING {t['blocked_pending']}
BLOCKED_CAP {t['blocked_cap']}
OTHER_REJECT {t['other_reject']}
ADMITTED {t['admitted']}
FILLED {t['fills']}
EXPIRED {t['expired']}

{rep['funnel_note']}

## Fixed counterpart funnel

evaluated {ff.get('candidate_evaluated')} · selected {ff.get('selected')} · admitted {ff.get('admitted')} · fills {ff.get('fills')} · expired {ff.get('expired')}
Fixed fill_rate {rep['FIXED_FILL_RATE']} · Dynamic fill_rate {rep['DYNAMIC_FILL_RATE']}

P1 canonical trades match: `{rep['FIXED_FUNNEL']['match_p1_canonical_trades']}`

## Delay / C1 / coverage

FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY: count {dly['count']} · fixed_pnl {dly['fixed_pnl']} · PF {dly['PF']} · GP {dly['gross_profit']} · GL {dly['gross_loss']} · W/L/D {dly['win']}/{dly['loss']}/{dly['draw']}
(These are Fixed trades whose signal occurred while the same-session Dynamic anchor was still in [t0, t1). PnL is realized later at Fixed exit.)

C1_REJECTED_FIXED_TRADES: count {c1['count']} · fixed_pnl {c1['fixed_pnl']} · PF {c1['PF']}

## Top3 (predeclared; classes overlap)

actual_fixed_pnl {top['actual_fixed_pnl']} · actual_dynamic_pnl {top['actual_dynamic_pnl']} · **TOP3_ACTUAL_DELTA {top['actual_delta']}**
prior_T1 {top['fixed_trades_with_prior_T1']} · during_wait {top['fixed_entries_during_dynamic_wait']} · prior_C1_rejected {top['fixed_trades_prior_C1_rejected']} · C1_confirmed_before_entry {top['fixed_trades_c1_confirmed_before_entry']}
dynamic admitted {top['dynamic_admitted']} · filled {top['dynamic_filled']}

Class PnL is **not** additive to 1,932,200.

## Rest11

Fixed pnl {r11['Fixed']['pnl']} PF {r11['Fixed']['PF']} maxDD {r11['Fixed']['maxDD']}
Dynamic pnl {r11['Dynamic']['pnl']} PF {r11['Dynamic']['PF']} maxDD {r11['Dynamic']['maxDD']}
actual_delta {r11['actual_delta']}

## Passive fill (WAIT_SEC={WAIT_SEC}, no retune)

Fixed median min_ask_minus_limit_bps {pf['Fixed_median_min_ask_minus_limit_bps']}
Dynamic median min_ask_minus_limit_bps {pf['Dynamic_median_min_ask_minus_limit_bps']}
Dynamic FILLED median t0→t1 CurrentPrice return {mv['Dynamic_FILLED_median_t0_t1_return']}
Dynamic EXPIRED median t0→t1 CurrentPrice return {mv['Dynamic_EXPIRED_median_t0_t1_return']}

## Attribution (evidence-supported, not causal yen)

PRIMARY_EVIDENCE_SUPPORTED_DRIVER: **{rep['PRIMARY_EVIDENCE_SUPPORTED_DRIVER']}**
SECONDARY_DRIVERS: {rep['SECONDARY_DRIVERS']}

Total actual Fixed−Dynamic delta: {rep['FIXED_TOP3_DEPENDENCY_ROLE']['total_actual_fixed_dynamic_delta']}
Predeclared Top3 actual delta: {rep['FIXED_TOP3_DEPENDENCY_ROLE']['predeclared_top3_actual_delta']}
Remaining 11 FULL day actual delta: {rep['FIXED_TOP3_DEPENDENCY_ROLE']['remaining_11_full_day_actual_delta']}

Top3-excluded: Fixed PF {rep['FIXED_TOP3_DEPENDENCY_ROLE']['rest11_fixed_PF']} · Dynamic PF {rep['FIXED_TOP3_DEPENDENCY_ROLE']['rest11_dynamic_PF']}

DYNAMIC_CONCEPT_GENERAL_FAILURE_PROVEN: false

EXIT policy is identical. Entry-timing differences may change which Arch E path is hit; that is descriptive only.

## STOP

NEW_STRATEGY_TESTED false · RETUNING_DONE false · STRATEGY_CHANGED false · ENTRY_EXIT_CHANGED false · RUNTIME_CHANGED false
"""


def write_artifacts(rep: dict[str, Any]) -> dict[str, Path]:
    OUT.mkdir(parents=True, exist_ok=True)
    for p in OUT.iterdir():
        if p.is_file() and p.name not in {"report.json", "report.md", "audit.xlsx"}:
            p.unlink()
    pub = {k: _jsonable(v) for k, v in rep.items() if not k.startswith("_")}
    (OUT / "report.json").write_text(
        json.dumps(pub, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8"
    )
    (OUT / "report.md").write_text(_md(rep), encoding="utf-8")
    wb = Workbook()

    def sh(name: str):
        if name == "Summary":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    t = rep["DYNAMIC_TERMINAL_OUTCOMES"]
    dly = rep["FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY"]
    c1 = rep["C1_REJECTED_FIXED_TRADES"]
    top = rep["TOP3"]
    r11 = rep["REST11"]
    summary = [
        {"field": k, "value": json.dumps(v, ensure_ascii=False, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("P2_2_VERDICT", rep["P2_2_VERDICT"]),
            ("CANDIDATE_STATUS", rep["CANDIDATE_STATUS"]),
            ("FUNNEL_ACCOUNTING", rep["FUNNEL_ACCOUNTING"]),
            ("confirmed", t["confirmed"]),
            ("not_selected", t["not_selected"]),
            ("blocked_open", t["blocked_open"]),
            ("blocked_pending", t["blocked_pending"]),
            ("blocked_cap", t["blocked_cap"]),
            ("other_reject", t["other_reject"]),
            ("admitted", t["admitted"]),
            ("fills", t["fills"]),
            ("expired", t["expired"]),
            ("FIXED_FILL_RATE", rep["FIXED_FILL_RATE"]),
            ("DYNAMIC_FILL_RATE", rep["DYNAMIC_FILL_RATE"]),
            ("FIXED_TRADES_ANALYZED", rep["FIXED_TRADES_ANALYZED"]),
            ("FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY_count", dly["count"]),
            ("FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY_pnl", dly["fixed_pnl"]),
            ("FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY_PF", dly["PF"]),
            ("C1_REJECTED_count", c1["count"]),
            ("C1_REJECTED_pnl", c1["fixed_pnl"]),
            ("C1_REJECTED_PF", c1["PF"]),
            ("TOP3_actual_delta", top["actual_delta"]),
            ("REST11_Fixed_pnl", r11["Fixed"]["pnl"]),
            ("REST11_Fixed_PF", r11["Fixed"]["PF"]),
            ("REST11_Dynamic_pnl", r11["Dynamic"]["pnl"]),
            ("REST11_Dynamic_PF", r11["Dynamic"]["PF"]),
            ("PRIMARY_EVIDENCE_SUPPORTED_DRIVER", rep["PRIMARY_EVIDENCE_SUPPORTED_DRIVER"]),
            ("SECONDARY_DRIVERS", rep["SECONDARY_DRIVERS"]),
            ("DYNAMIC_CONCEPT_GENERAL_FAILURE_PROVEN", False),
            ("verdict", rep["verdict"]),
            ("SAFETY", rep["SAFETY"]),
        ]
    ]
    _write_rows(sh("Summary"), summary)

    _write_rows(sh("Funnel_Reconciliation"), [
        {"field": "FUNNEL_ACCOUNTING", "value": rep["FUNNEL_ACCOUNTING"]},
        {"field": "note", "value": rep["funnel_note"]},
        *[{"field": f"p2_2_published.{k}", "value": v} for k, v in (rep.get("p2_2_published_funnel") or {}).items()],
        *[{"field": f"exclusive.{k}", "value": v} for k, v in t.items()],
        *[{"field": f"integrity.{k}", "value": json.dumps(v) if isinstance(v, (dict, list)) else v} for k, v in rep["funnel_integrity"].items()],
        *[{"field": f"p22_style.{k}", "value": v} for k, v in rep["p2_2_style_reconstructed"].items()],
        *[{"field": "blocker", "value": b} for b in (rep.get("blockers") or ["none"])],
    ])

    ff = rep["FIXED_FUNNEL"]
    _write_rows(sh("Fixed_Funnel"), [
        {"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            *[(f"p1_frozen.{k}", x) for k, x in ff["p1_frozen"].items()],
            *[(f"reconstructed.{k}", x) for k, x in ff["reconstructed"].items()],
            ("match_p1_canonical_trades", ff["match_p1_canonical_trades"]),
            ("FIXED_FILL_RATE", rep["FIXED_FILL_RATE"]),
            ("DYNAMIC_FILL_RATE", rep["DYNAMIC_FILL_RATE"]),
            ("AM_PM", ff["AM_PM"]),
            ("fill_rate_by_session", rep["fill_rate_by_session"]),
        ]
    ])

    _write_rows(sh("Fixed_Dynamic_State"), rep["_state_rows"])
    delay_sheet = [{
        "date": "SUMMARY",
        "symbol": "",
        "session": "",
        "fixed_anchor_time": "",
        "fixed_signal_time": None,
        "fixed_fill_time": None,
        "fixed_fill_price": None,
        "fixed_exit_time": None,
        "fixed_exit_price": None,
        "fixed_pnl": dly.get("fixed_pnl"),
        "dynamic_state_at_fixed_signal": "FIXED_ENTRY_BEFORE_DYNAMIC_MATURITY",
        "latest_t0": None,
        "latest_t1": None,
        "c1_status": None,
        "c1_reason": f"count={dly.get('count')} PF={dly.get('PF')} GP={dly.get('gross_profit')} GL={dly.get('gross_loss')} W/L/D={dly.get('win')}/{dly.get('loss')}/{dly.get('draw')}",
        "fixed_signal_minus_t0_sec": None,
        "fixed_signal_minus_t1_sec": None,
        "has_prior_t1": True,
        "entry_during_anchor_active": True,
        "prior_c1_rejected": False,
        "c1_confirmed_before_entry": False,
        "dynamic_entry_terminal": None,
        "dynamic_fill_terminal": None,
        "dynamic_canonical_terminal_outcome": dly.get("definition"),
    }] + list(rep["_delay_rows"] or [])
    _write_rows(sh("Delay_10m"), delay_sheet)
    c1_sheet = [{
        "date": "SUMMARY",
        "symbol": "",
        "session": "",
        "fixed_anchor_time": "",
        "fixed_signal_time": None,
        "fixed_fill_time": None,
        "fixed_fill_price": None,
        "fixed_exit_time": None,
        "fixed_exit_price": None,
        "fixed_pnl": c1.get("fixed_pnl"),
        "dynamic_state_at_fixed_signal": "LAST_C1_REJECTED",
        "latest_t0": None,
        "latest_t1": None,
        "c1_status": "REJECTED",
        "c1_reason": f"count={c1.get('count')} PF={c1.get('PF')} W/L/D={c1.get('win')}/{c1.get('loss')}/{c1.get('draw')}",
        "fixed_signal_minus_t0_sec": None,
        "fixed_signal_minus_t1_sec": None,
        "has_prior_t1": True,
        "entry_during_anchor_active": False,
        "prior_c1_rejected": True,
        "c1_confirmed_before_entry": False,
        "dynamic_entry_terminal": None,
        "dynamic_fill_terminal": None,
        "dynamic_canonical_terminal_outcome": None,
    }] + list(rep["_c1_rows"] or [])
    _write_rows(sh("C1_Rejection"), c1_sheet)

    pf = rep["PASSIVE_FILL"]
    _write_rows(sh("Passive_Fill"), [
        {"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("WAIT_SEC", pf["WAIT_SEC"]),
            ("Fixed_median_min_ask_minus_limit_bps", pf["Fixed_median_min_ask_minus_limit_bps"]),
            ("Dynamic_median_min_ask_minus_limit_bps", pf["Dynamic_median_min_ask_minus_limit_bps"]),
            ("Fixed", pf["Fixed"]),
            ("Dynamic", pf["Dynamic"]),
            ("T0_T1_MOVE", rep["T0_T1_MOVE"]),
        ]
    ])

    top3_sheet = [{
        **{k: None for k in (rep["_top3_state"][0].keys() if rep.get("_top3_state") else {"date": None})},
        "date": "SUMMARY",
        "symbol": "",
        "fixed_pnl": top.get("actual_fixed_pnl"),
        "dynamic_state_at_fixed_signal": "TOP3_ACTUAL_DELTA",
        "c1_reason": json.dumps({
            "actual_dynamic_pnl": top.get("actual_dynamic_pnl"),
            "actual_delta": top.get("actual_delta"),
            "fixed_trades": top.get("fixed_trades"),
            "fixed_trades_with_prior_T1": top.get("fixed_trades_with_prior_T1"),
            "fixed_entries_during_dynamic_wait": top.get("fixed_entries_during_dynamic_wait"),
            "fixed_trades_prior_C1_rejected": top.get("fixed_trades_prior_C1_rejected"),
            "fixed_trades_c1_confirmed_before_entry": top.get("fixed_trades_c1_confirmed_before_entry"),
            "dynamic_admitted": top.get("dynamic_admitted"),
            "dynamic_filled": top.get("dynamic_filled"),
            "classes_overlap": True,
        }, ensure_ascii=False),
    }] + list(rep["_top3_state"] or [])
    _write_rows(sh("Top3_Decomposition"), top3_sheet)

    _write_rows(sh("Rest11"), [
        {"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("days", r11["days"]),
            *[(f"Fixed.{k}", x) for k, x in r11["Fixed"].items()],
            *[(f"Dynamic.{k}", x) for k, x in r11["Dynamic"].items()],
            ("actual_delta", r11["actual_delta"]),
        ]
    ])

    _write_rows(sh("Better_Days"), [
        {
            "date": r["date"],
            "Fixed_trades": r["Fixed_trades"],
            "Dynamic_trades": r["Dynamic_trades"],
            "Fixed_pnl": r["Fixed_pnl"],
            "Dynamic_pnl": r["Dynamic_pnl"],
            "delta_pnl_dynamic_minus_fixed": r["delta_pnl_dynamic_minus_fixed"],
            "lower_exposure": r["lower_exposure"],
            "fixed_only_symbols": r["fixed_only_symbols"],
            "dynamic_only_symbols": r["dynamic_only_symbols"],
            "both_symbols": r["both_symbols"],
            "fixed_loser_symbols_without_dynamic_fill": r["fixed_loser_symbols_without_dynamic_fill"],
            "dynamic_only_winners": r["dynamic_only_winners"],
            "same_symbol_timing": r["same_symbol_timing"],
            "note": r["note"],
        }
        for r in rep["BETTER_DAYS"]
    ])

    _write_rows(sh("Exit_Description"), [
        {"field": "policy", "value": rep["EXIT_DESCRIPTION"]["policy"]},
        *[{"field": f"reason.{r['exit_reason']}", "value": json.dumps(r, default=str)} for r in rep["EXIT_DESCRIPTION"]["by_reason"]],
        {"field": "unknown_dynamic_reasons", "value": json.dumps(rep["EXIT_DESCRIPTION"]["unknown_dynamic_reasons"])},
        {"field": "negative_holding_n", "value": rep["EXIT_DESCRIPTION"]["negative_holding_n"]},
        *[{"field": "same_symbol_day", "value": json.dumps(r, default=str)} for r in (rep.get("_same_sym_exit") or [])],
    ])

    _write_rows(sh("Attribution"), [
        {"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in [
            ("PRIMARY_EVIDENCE_SUPPORTED_DRIVER", rep["PRIMARY_EVIDENCE_SUPPORTED_DRIVER"]),
            ("SECONDARY_DRIVERS", rep["SECONDARY_DRIVERS"]),
            *list(rep["FIXED_TOP3_DEPENDENCY_ROLE"].items()),
            ("DYNAMIC_CONCEPT_GENERAL_FAILURE_PROVEN", False),
            ("causal_yen_claimed", False),
            ("driver_notes", (rep.get("_drivers") or {}).get("notes")),
        ]
    ])

    _write_rows(sh("Identity"), [
        {"field": k, "value": json.dumps(v, default=str) if isinstance(v, (dict, list)) else v}
        for k, v in rep["identity"].items()
    ])
    _write_rows(sh("Safety"), [
        {"field": "submit/cancel/live", "value": "0/0/0"},
        {"field": "NEW_STRATEGY_TESTED", "value": False},
        {"field": "RETUNING_DONE", "value": False},
        {"field": "STRATEGY_CHANGED", "value": False},
        {"field": "ENTRY_EXIT_CHANGED", "value": False},
        {"field": "RUNTIME_CHANGED", "value": False},
        {"field": "verdict", "value": rep["verdict"]},
        {"field": "note", "value": "Research accounting only. Candidate frozen REJECTED_AS_FIXED_REPLACEMENT. No retune."},
    ])
    wb.save(OUT / "audit.xlsx")
    return {"report_json": OUT / "report.json", "report_md": OUT / "report.md", "audit_xlsx": OUT / "audit.xlsx"}
