"""E1_X11 orchestrator — calibrate FIXED100_CONSERVATIVE_V1 from E1_X10 day panel."""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from research.e1_x6_provisional.util import sha256_file, sha256_obj
from research.e1_x7_pfq.config import DAYS

from . import (
    AGG_NOTIONAL_FRAC,
    AGG_RISK_FRAC,
    ANALYSIS_ID,
    BOARD_FRESHNESS_SEC,
    CAPITAL_SCENARIOS,
    DAILY_LOSS_GUARD_PCT,
    DOCUMENT_ID,
    HISTORY_END_MAX,
    LOOKBACK_MAX_DAYS,
    LOT,
    MAX_CONCURRENT,
    MIN_EXEC_ANCHORS,
    MIN_HISTORY_DAYS,
    MIN_JUMP_N,
    MIN_SPREAD_N,
    PER_SYMBOL_NOTIONAL_FRAC,
    PER_TRADE_RISK_FRAC,
    POLICY_ID,
    PRICE_FRESHNESS_SEC,
    RECURRING_MIN_DAYS,
    RESERVE_FRAC,
    SOURCE_X10,
    SOURCE_X10_VERDICT,
    TARGET_SYMBOL,
)
from .wallet_audit import resolve_capital_base, special_quote_audit

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
X10_DIR = NATIVE / "results" / "research" / "e1_x10_risk_universe"
PUBLISH = NATIVE / "results" / "research" / "e1_x11_risk_policy"
PKG = Path(__file__).resolve().parent


def _load_sheet(path: Path, name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


def _f(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _safety() -> dict[str, Any]:
    return {
        "submit_cancel_live": "0/0/0",
        "mainline_changed": False,
        "production_yaml_changed": False,
        "entry_changed": False,
        "exit_changed": False,
        "universe_runtime_changed": False,
        "unused_data_used": False,
        "prospective": False,
        "shadow": False,
        "forward": False,
        "paper": False,
        "discord": False,
        "orders_called": False,
    }


def _pnl_independence() -> dict[str, Any]:
    bad = (
        "net_" + "pnl",
        "profit" + "_factor",
        "passes_" + "candidate",
        "PFQ_" + "UPDATE",
        "first_touch_" + "label",
    )
    hits = []
    for fp in PKG.glob("*.py"):
        if fp.name in ("__init__.py",):
            continue
        text = fp.read_text(encoding="utf-8")
        if fp.name == "run_audit.py":
            start = text.find("def _pnl_independence")
            end = text.find("\ndef _prior_days")
            if start >= 0 and end > start:
                text = text[:start] + text[end:]
        for tok in bad:
            if tok in text:
                hits.append({"file": fp.name, "token": tok})
    unused_ok = all(d <= HISTORY_END_MAX for d in DAYS) and "20260803" not in DAYS
    return {
        "status": "PASS" if not hits and unused_ok else "FAIL",
        "hits": hits,
        "unused_data_ok": unused_ok,
        "purpose": "RISK_POLICY_CALIBRATION",
        "not_purpose": "ALPHA_OPTIMIZATION",
    }


def _prior_days(day: str, all_days: list[str]) -> list[str]:
    """Trading days strictly before `day`, max LOOKBACK_MAX_DAYS."""
    prior = [d for d in all_days if d < day]
    return prior[-LOOKBACK_MAX_DAYS:]


def _merge_day_panel(x10_xlsx: Path) -> dict[tuple[str, str], dict[str, Any]]:
    notion = _load_sheet(x10_xlsx, "OneLotNotional")
    tick = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "TickRisk")}
    spread = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "SpreadRisk")}
    jump = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "BidJumps")}
    exe = {(r["symbol"], str(r["day"])): r for r in _load_sheet(x10_xlsx, "ExecutableLoss")}
    panel: dict[tuple[str, str], dict[str, Any]] = {}
    for r in notion:
        key = (str(r["symbol"]), str(r["day"]))
        panel[key] = {
            "symbol": key[0],
            "day": key[1],
            "one_lot_notional_yen": _f(r.get("one_lot_notional_yen")),
            "reference_price": _f(r.get("reference_price")),
            "ref_status": r.get("status"),
            "tick_size_yen": _f((tick.get(key) or {}).get("tick_size_yen")),
            "one_tick_risk_yen_100": _f((tick.get(key) or {}).get("one_tick_risk_yen_100")),
            "n_spread_obs": int(_f((spread.get(key) or {}).get("n_spread_obs")) or 0),
            "p95_spread_cost_yen_100": _f((spread.get(key) or {}).get("p95_spread_cost_yen_100")),
            "n_jump_obs": int(_f((jump.get(key) or {}).get("n_jump_obs")) or 0),
            "p95_down_bid_jump_yen_100": _f((jump.get(key) or {}).get("p95_down_bid_jump_yen_100")),
            "n_exec_anchors": int(_f((exe.get(key) or {}).get("n_exec_anchors")) or 0),
            "exec_loss_5s_p95": _f((exe.get(key) or {}).get("exec_loss_yen_100_5s_p95")),
        }
    return panel


def _rolling_metrics(
    symbol: str,
    day: str,
    hist_days: list[str],
    panel: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, Any]:
    rows = [panel[(symbol, d)] for d in hist_days if (symbol, d) in panel]
    valid_days = len(rows)
    spread_n = sum(r["n_spread_obs"] for r in rows)
    jump_n = sum(r["n_jump_obs"] for r in rows)
    exec_n = sum(r["n_exec_anchors"] for r in rows)
    support = (
        valid_days >= MIN_HISTORY_DAYS
        and spread_n >= MIN_SPREAD_N
        and jump_n >= MIN_JUMP_N
        and exec_n >= MIN_EXEC_ANCHORS
    )
    # rolling p95 components: median of daily p95 over history (robust), then max
    spreads = [r["p95_spread_cost_yen_100"] for r in rows if r["p95_spread_cost_yen_100"] is not None]
    jumps = [r["p95_down_bid_jump_yen_100"] for r in rows if r["p95_down_bid_jump_yen_100"] is not None]
    execs = [r["exec_loss_5s_p95"] for r in rows if r["exec_loss_5s_p95"] is not None]

    def med(xs: list[float]) -> Optional[float]:
        if not xs:
            return None
        xs = sorted(xs)
        return float(xs[len(xs) // 2])

    s_m, j_m, e_m = med(spreads), med(jumps), med(execs)
    comps = [c for c in (s_m, j_m, e_m) if c is not None]
    est = max(comps) if comps else None

    # notional / tick from latest history day (D-1) if present
    last = rows[-1] if rows else None
    return {
        "day": day,
        "symbol": symbol,
        "history_start": hist_days[0] if hist_days else None,
        "history_end": hist_days[-1] if hist_days else None,
        "valid_history_days": valid_days,
        "spread_n": spread_n,
        "jump_n": jump_n,
        "exec_anchor_n": exec_n,
        "support_status": "PASS" if support else "RISK_HISTORY_INSUFFICIENT",
        "history_support_pass": support,
        "estimated_execution_risk_yen": est,
        "spread_cost_p95_yen_100": s_m,
        "down_jump_p95_yen_100": j_m,
        "exec_loss_5s_p95_yen_100": e_m,
        "one_lot_notional_yen": last["one_lot_notional_yen"] if last else None,
        "tick_size_yen": last["tick_size_yen"] if last else None,
        "reference_asof_pass": bool(last and last.get("ref_status") == "OK" and last.get("one_lot_notional_yen")),
        # prove no same-day future: history_end < day
        "no_same_day_future": (not hist_days) or (hist_days[-1] < day),
    }


def _limits(capital: float) -> dict[str, float]:
    return {
        "risk_capital_base_yen": capital,
        "per_trade_risk_limit": capital * PER_TRADE_RISK_FRAC,
        "aggregate_risk_limit": capital * AGG_RISK_FRAC,
        "per_symbol_notional_limit": capital * PER_SYMBOL_NOTIONAL_FRAC,
        "aggregate_notional_limit": capital * AGG_NOTIONAL_FRAC,
        "reserve_yen": capital * RESERVE_FRAC,
    }


def _static_eligible(roll: dict[str, Any], capital: float) -> dict[str, Any]:
    lim = _limits(capital)
    reasons = []
    notional = roll.get("one_lot_notional_yen")
    est = roll.get("estimated_execution_risk_yen")
    notional_pass = notional is not None and notional <= lim["per_symbol_notional_limit"]
    risk_pass = est is not None and est <= lim["per_trade_risk_limit"]
    if not roll.get("history_support_pass"):
        reasons.append("RISK_HISTORY_INSUFFICIENT")
    if not roll.get("reference_asof_pass"):
        reasons.append("REFERENCE_ASOF_FAIL")
    if roll.get("tick_size_yen") is None:
        reasons.append("TICK_UNRESOLVED")
    if notional is None:
        reasons.append("NOTIONAL_MISSING")
    elif not notional_pass:
        reasons.append("NOTIONAL_ABOVE_15PCT")
    if est is None:
        reasons.append("EXEC_RISK_MISSING")
    elif not risk_pass:
        reasons.append("EXEC_RISK_ABOVE_025PCT")
    eligible = (
        roll.get("history_support_pass")
        and roll.get("reference_asof_pass")
        and roll.get("tick_size_yen") is not None
        and notional_pass
        and risk_pass
    )
    return {
        **roll,
        **lim,
        "notional_limit_yen": lim["per_symbol_notional_limit"],
        "notional_pass": notional_pass,
        "risk_limit_yen": lim["per_trade_risk_limit"],
        "risk_pass": risk_pass,
        "static_eligible": bool(eligible),
        "reason_codes": reasons if not eligible else [],
    }


def run_once(*, label: str = "A") -> dict[str, Any]:
    run_id = f"e1x11_policy_{datetime.now(JST).strftime('%Y%m%d_%H%M%S')}_{label}"

    x10 = json.loads((X10_DIR / "report.json").read_text(encoding="utf-8"))
    if SOURCE_X10 not in str(x10.get("run_id")) or x10.get("verdict") != SOURCE_X10_VERDICT:
        return {
            "analysis_id": ANALYSIS_ID,
            "document_id": DOCUMENT_ID,
            "run_id": run_id,
            "verdict": "E1_X11_SOURCE_IDENTITY_MISMATCH",
            "mismatches": {"expected_run": SOURCE_X10, "got": x10.get("run_id"),
                           "expected_verdict": SOURCE_X10_VERDICT, "got_verdict": x10.get("verdict")},
            "determinism_shas": {"verdict": "E1_X11_SOURCE_IDENTITY_MISMATCH"},
            "safety": _safety(),
            "stop": True,
            "_sheets": {},
        }

    panel = _merge_day_panel(X10_DIR / "audit.xlsx")
    all_days = sorted({d for _, d in panel.keys() if d <= HISTORY_END_MAX})
    symbols = sorted({s for s, _ in panel.keys()})
    observed_days = defaultdict(set)
    for s, d in panel.keys():
        observed_days[s].add(d)

    recurring = sorted(s for s in symbols if len(observed_days[s]) >= RECURRING_MIN_DAYS)
    all_obs = symbols

    # evaluation days: need prior history (skip first calendar day of panel)
    eval_days = [d for d in all_days if any(x < d for x in all_days)]

    print(f"=== [{label}] panel symbols={len(symbols)} recurring={len(recurring)} eval_days={len(eval_days)} ===", flush=True)

    capital_base = resolve_capital_base()
    sq = special_quote_audit()

    # rolling support rows (reference capital scenarios computed separately)
    rolling_rows = []
    for day in eval_days:
        hist = _prior_days(day, all_days)
        # symbols observed on any hist day or on day itself in universe that day
        cands = {s for s, d in panel.keys() if d in hist or d == day}
        for sym in sorted(cands):
            if not any((sym, d) in panel for d in hist):
                continue
            roll = _rolling_metrics(sym, day, hist, panel)
            rolling_rows.append(roll)

    # Recurring coverage: among recurring symbols, fraction with ≥1 PASS support on any eval day
    rec_support = defaultdict(bool)
    for r in rolling_rows:
        if r["symbol"] in recurring and r["history_support_pass"]:
            rec_support[r["symbol"]] = True
    recurring_coverage = (sum(1 for s in recurring if rec_support[s]) / len(recurring)) if recurring else 0.0
    all_support = defaultdict(bool)
    for r in rolling_rows:
        if r["history_support_pass"]:
            all_support[r["symbol"]] = True
    all_coverage = (sum(1 for s in all_obs if all_support[s]) / len(all_obs)) if all_obs else 0.0

    # Capital scenarios (sensitivity only)
    scenario_rows = []
    eligibility_by_cap: dict[int, list[dict[str, Any]]] = {}
    daily_counts_all = []
    for cap in CAPITAL_SCENARIOS:
        elig_rows = []
        daily_eligible = defaultdict(set)
        daily_recurring_eligible = defaultdict(set)
        for r in rolling_rows:
            e = _static_eligible(r, float(cap))
            elig_rows.append(e)
            if e["static_eligible"]:
                daily_eligible[e["day"]].add(e["symbol"])
                if e["symbol"] in recurring:
                    daily_recurring_eligible[e["day"]].add(e["symbol"])
        eligibility_by_cap[cap] = elig_rows
        day_counts = [len(daily_recurring_eligible[d]) for d in eval_days]
        day_counts_all_sym = [len(daily_eligible[d]) for d in eval_days]
        lim = _limits(float(cap))
        # top5 required capital by max(notional/0.15, risk/0.0025) using latest rolling per symbol
        reqs = []
        by_sym_latest = {}
        for e in elig_rows:
            by_sym_latest[(e["day"], e["symbol"])] = e
        # use last eval day snapshot
        last_day = eval_days[-1] if eval_days else None
        if last_day:
            for s in recurring:
                e = by_sym_latest.get((last_day, s))
                if not e:
                    continue
                n, risk = e.get("one_lot_notional_yen"), e.get("estimated_execution_risk_yen")
                if n is None and risk is None:
                    continue
                need_n = (n / PER_SYMBOL_NOTIONAL_FRAC) if n is not None else 0.0
                need_r = (risk / PER_TRADE_RISK_FRAC) if risk is not None else 0.0
                reqs.append({"symbol": s, "required_capital": max(need_n, need_r),
                             "need_notional": need_n, "need_risk": need_r})
        top5 = sorted(reqs, key=lambda x: -x["required_capital"])[:5]
        scenario_rows.append({
            "capital_yen": cap,
            **{k: lim[k] for k in (
                "per_trade_risk_limit", "per_symbol_notional_limit",
                "aggregate_risk_limit", "aggregate_notional_limit",
            )},
            "static_eligible_symbol_days": sum(1 for e in elig_rows if e["static_eligible"]),
            "recurring_eligible_symbol_days": sum(
                1 for e in elig_rows if e["static_eligible"] and e["symbol"] in recurring
            ),
            "median_eligible_symbols_per_day": float(sorted(day_counts)[len(day_counts) // 2]) if day_counts else 0,
            "min_eligible_symbols_per_day": min(day_counts) if day_counts else 0,
            "max_eligible_symbols_per_day": max(day_counts) if day_counts else 0,
            "median_eligible_all_symbols_per_day": float(sorted(day_counts_all_sym)[len(day_counts_all_sym) // 2]) if day_counts_all_sym else 0,
            "top5_required_capital": top5,
            "adequacy_data_gate": (
                recurring_coverage >= 0.90
                and (min(day_counts) if day_counts else 0) >= 10
                and (float(sorted(day_counts)[len(day_counts) // 2]) if day_counts else 0) >= 20
            ),
        })
        daily_counts_all.append({
            "capital_yen": cap,
            "by_day": {d: len(daily_recurring_eligible[d]) for d in eval_days},
        })

    # 285A profile — same formula
    kiox_rolls = [r for r in rolling_rows if r["symbol"] == TARGET_SYMBOL]
    kiox_latest = kiox_rolls[-1] if kiox_rolls else None
    if kiox_latest and kiox_latest.get("one_lot_notional_yen") is not None:
        n = float(kiox_latest["one_lot_notional_yen"])
        risk = float(kiox_latest["estimated_execution_risk_yen"] or 0)
        min_by_n = n / PER_SYMBOL_NOTIONAL_FRAC
        min_by_r = risk / PER_TRADE_RISK_FRAC if risk else None
        kioxia = {
            "symbol": TARGET_SYMBOL,
            "one_lot_notional_yen": n,
            "estimated_execution_risk_yen": risk if risk else None,
            "minimum_capital_by_notional": min_by_n,
            "minimum_capital_by_risk": min_by_r,
            "required_capital": max(min_by_n, min_by_r or 0.0),
            "special_cased": False,
            "history_support_pass": kiox_latest.get("history_support_pass"),
        }
    else:
        # fallback to X10 medians as reference only if rolling missing
        kx10 = (x10.get("kioxia_profile") or {})
        n = _f(kx10.get("one_lot_notional_median"))
        risk = _f(kx10.get("estimated_execution_risk_yen"))
        kioxia = {
            "symbol": TARGET_SYMBOL,
            "one_lot_notional_yen": n,
            "estimated_execution_risk_yen": risk,
            "minimum_capital_by_notional": (n / PER_SYMBOL_NOTIONAL_FRAC) if n else None,
            "minimum_capital_by_risk": (risk / PER_TRADE_RISK_FRAC) if risk else None,
            "required_capital": max(
                (n / PER_SYMBOL_NOTIONAL_FRAC) if n else 0.0,
                (risk / PER_TRADE_RISK_FRAC) if risk else 0.0,
            ) or None,
            "special_cased": False,
            "note": "from E1_X10 median reference; rolling row missing or incomplete",
        }

    # Dynamic gate feasibility
    dynamic_gate = {
        "separated_from_entry_gate": True,
        "fields": [
            {"field": "best_bid", "status": "AVAILABLE"},
            {"field": "best_ask", "status": "AVAILABLE"},
            {"field": "best_bid_qty>=100", "status": "AVAILABLE"},
            {"field": "best_ask_qty>=100", "status": "AVAILABLE"},
            {"field": "board_age<=3.0", "status": "AVAILABLE", "contract_sec": BOARD_FRESHNESS_SEC},
            {"field": "current_one_lot_notional<=15%", "status": "REQUIRES_CAPITAL_BASE"},
            {"field": "aggregate_notional<=75%", "status": "REQUIRES_CAPITAL_BASE"},
            {"field": "aggregate_risk<=1%", "status": "REQUIRES_CAPITAL_BASE"},
            {"field": "special_quote", "status": sq["status"]},
        ],
        "missing_behavior": "FAIL_CLOSED",
        "board_freshness_for_risk_gate": BOARD_FRESHNESS_SEC,
        "entry_price_freshness_unchanged": True,
        "entry_price_freshness_sec": PRICE_FRESHNESS_SEC,
        "runtime_implemented": False,
        "special_quote_guard": sq["dynamic_guard"],
        "dynamic_fields_complete": False,  # capital unresolved + special quote
    }

    price_freshness_contract = {
        "risk_gate_uses": "board timestamp (BidTime/AskTime)",
        "entry_feature_freshness": "existing contract unchanged (price 3.0s / board 3.0s)",
        "risk_gate_does_not_replace_entry_freshness": True,
        "CurrentPriceTime_often_null": True,
        "reason": "E1_X10: board freshness relatively high; CurrentPriceTime often null",
    }

    policy_def = {
        "POLICY_ID": POLICY_ID,
        "lot_size": LOT,
        "max_concurrent_positions": MAX_CONCURRENT,
        "daily_loss_guard_pct": DAILY_LOSS_GUARD_PCT,
        "per_trade_execution_risk_limit_frac": PER_TRADE_RISK_FRAC,
        "aggregate_open_risk_frac": AGG_RISK_FRAC,
        "per_symbol_notional_frac": PER_SYMBOL_NOTIONAL_FRAC,
        "aggregate_notional_frac": AGG_NOTIONAL_FRAC,
        "reserve_frac": RESERVE_FRAC,
        "board_freshness_sec": BOARD_FRESHNESS_SEC,
        "bid_ask_qty_min": LOT,
        "fail_closed": True,
        "fractions_frozen_before_outcomes": True,
    }

    config_drift = {
        "status": "CONFIG_FILENAME_CAP3_CANONICAL_CAP5_DRIFT",
        "filename_contains": "cap3",
        "canonical_key": "max_concurrent_positions",
        "canonical_value": MAX_CONCURRENT,
        "filename_changed": False,
    }

    # History inventory
    push_root = NATIVE / "data" / "push_jsonl"
    hist_dirs = sorted([
        d.name for d in push_root.iterdir()
        if d.is_dir() and d.name[:10].replace("-", "") <= HISTORY_END_MAX
    ]) if push_root.exists() else []

    # Verdict
    pnl = _pnl_independence()
    support_insufficient = recurring_coverage < 0.90 or len(recurring) == 0
    # Prefer capital unresolved when path C
    if capital_base["status"] == "UNRESOLVED":
        verdict = "E1_X11_SAFE_CAPITAL_BASE_UNRESOLVED"
        next_step = (
            "resolve safe risk_capital_base (configure cap and/or prove StockAccountWallet as equity base); "
            "do not use leveraged buying_power; then re-run calibration before freeze document"
        )
    elif sq["status"] == "NOT_AVAILABLE_IN_CAPTURE" and True:
        # also incomplete dynamic — but capital takes precedence above
        verdict = "E1_X11_DYNAMIC_GATE_FIELDS_INCOMPLETE"
        next_step = "resolve special_quote capture/API presence before freeze"
    elif support_insufficient:
        verdict = "E1_X11_RISK_HISTORY_SUPPORT_INSUFFICIENT"
        next_step = "extend risk-only history before 20260721 or lower support bars only via new precommit"
    else:
        # check restrictiveness on a hypothetical mid capital if base were set — use 5M as reference scenario only for TOO_RESTRICTIVE
        ref = next(s for s in scenario_rows if s["capital_yen"] == 5_000_000)
        if ref["min_eligible_symbols_per_day"] < 10 or ref["median_eligible_symbols_per_day"] < 20:
            verdict = "E1_X11_CONSERVATIVE_POLICY_TOO_RESTRICTIVE"
            next_step = "policy fractions fixed — do not retune from outcomes; revise only via new precommit"
        else:
            verdict = "E1_X11_FIXED100_RISK_POLICY_FREEZE_READY"
            next_step = "Risk Universe Policy Freeze document only — no runtime implementation"

    # Note: with unresolved capital we already returned UNRESOLVED; keep dynamic note in detail
    vd = {
        "verdict": verdict,
        "capital_base_status": capital_base["status"],
        "recurring_coverage": recurring_coverage,
        "all_symbol_coverage": all_coverage,
        "n_recurring": len(recurring),
        "n_all_symbols": len(all_obs),
        "special_quote": sq["dynamic_guard"],
        "dynamic_fields_complete": dynamic_gate["dynamic_fields_complete"],
        "pnl_independence": pnl["status"],
        "next": next_step,
        "policy_id": POLICY_ID,
    }

    # For sheets: use capital=None eligibility as NOT_EVALUABLE reason RISK_CAPITAL_UNRESOLVED
    static_sheet = []
    for r in rolling_rows:
        static_sheet.append({
            **{k: r.get(k) for k in (
                "day", "symbol", "history_start", "history_end", "valid_history_days",
                "spread_n", "jump_n", "exec_anchor_n", "support_status",
                "one_lot_notional_yen", "estimated_execution_risk_yen",
                "history_support_pass", "reference_asof_pass", "no_same_day_future",
            )},
            "risk_capital_base_yen": None,
            "static_eligible": False,
            "reason_codes": ["SAFE_CAPITAL_BASE_UNRESOLVED"] + (
                [] if r.get("history_support_pass") else ["RISK_HISTORY_INSUFFICIENT"]
            ),
        })

    # daily eligible counts for scenarios
    daily_elig_sheet = []
    for sc in daily_counts_all:
        for d, n in sc["by_day"].items():
            daily_elig_sheet.append({"capital_yen": sc["capital_yen"], "day": d, "recurring_eligible_n": n})

    det = {
        "source_identity_sha": sha256_obj({"run": SOURCE_X10, "verdict": SOURCE_X10_VERDICT,
                                           "report_sha": sha256_file(X10_DIR / "report.json")}),
        "policy_definition_sha": sha256_obj(policy_def),
        "capital_base_sha": sha256_obj({k: capital_base[k] for k in capital_base if k != "wallet_fields"}),
        "rolling_support_sha": sha256_obj([(r["symbol"], r["day"], r["support_status"], r["estimated_execution_risk_yen"]) for r in rolling_rows]),
        "recurring_sha": sha256_obj(recurring),
        "scenario_sha": sha256_obj([{k: s[k] for k in s if k != "top5_required_capital"} | {
            "top5": [(t["symbol"], t["required_capital"]) for t in s.get("top5_required_capital") or []]
        } for s in scenario_rows]),
        "kioxia_sha": sha256_obj(kioxia),
        "verdict": verdict,
    }

    report = {
        "analysis_id": ANALYSIS_ID,
        "document_id": DOCUMENT_ID,
        "run_id": run_id,
        "label": label,
        "generated_at_jst": datetime.now(JST).isoformat(),
        "source_x10": SOURCE_X10,
        "source_x10_verdict": SOURCE_X10_VERDICT,
        "config_drift": config_drift,
        "inherited": {
            "lot_size": LOT,
            "max_concurrent_positions": MAX_CONCURRENT,
            "daily_loss_guard_pct": DAILY_LOSS_GUARD_PCT,
            "freshness_price_sec": PRICE_FRESHNESS_SEC,
            "freshness_board_sec": BOARD_FRESHNESS_SEC,
        },
        "capital_base": capital_base,
        "policy_definition": policy_def,
        "history_inventory": {
            "push_jsonl_days_le_20260731": hist_dirs,
            "n_days": len(hist_dirs),
            "panel_days": all_days,
            "note": "no risk capture before 20260721 in push_jsonl",
        },
        "recurring_universe": {
            "n_all_observed": len(all_obs),
            "n_recurring": len(recurring),
            "recurring_min_days": RECURRING_MIN_DAYS,
            "recurring_symbols": recurring,
            "recurring_risk_metric_coverage": recurring_coverage,
            "all_symbol_coverage_reference": all_coverage,
        },
        "capital_scenarios": scenario_rows,
        "kioxia_profile": kioxia,
        "dynamic_gate": dynamic_gate,
        "price_freshness_contract": price_freshness_contract,
        "special_quote": sq,
        "pnl_independence": pnl,
        "policy_adequacy": {
            "capital_base_resolved": capital_base["status"] == "RESOLVED",
            "recurring_coverage_ge_90": recurring_coverage >= 0.90,
            "dynamic_fields_sufficient": False,
            "pnl_independence_pass": pnl["status"] == "PASS",
            "freeze_ready": False,
        },
        "verdict": verdict,
        "verdict_detail": vd,
        "determinism_shas": det,
        "safety": _safety(),
        "stop": True,
        "_sheets": {
            "SourceIdentity": [{
                "source_run": SOURCE_X10,
                "source_verdict": SOURCE_X10_VERDICT,
                "match": True,
            }],
            "ConfigDrift": [config_drift],
            "WalletFields": capital_base.get("wallet_fields") or [],
            "CapitalBase": [{k: v for k, v in capital_base.items() if k != "wallet_fields"}],
            "HistoryInventory": [{"day": d} for d in hist_dirs],
            "RollingSupport": rolling_rows,
            "RecurringUniverse": [
                {"symbol": s, "observed_days": len(observed_days[s]),
                 "class": "RECURRING" if s in recurring else "RISK_HISTORY_INSUFFICIENT_OBS"}
                for s in all_obs
            ],
            "PolicyDefinition": [{"key": k, "value": v} for k, v in policy_def.items()],
            "SymbolDayRisk": static_sheet,
            "CapitalScenarios": [
                {**{k: v for k, v in s.items() if k != "top5_required_capital"},
                 "top5_required_capital": json.dumps(s.get("top5_required_capital"))}
                for s in scenario_rows
            ],
            "StaticEligibility": static_sheet,
            "DailyEligibleCounts": daily_elig_sheet,
            "DynamicGate": dynamic_gate["fields"],
            "PriceFreshnessContract": [price_freshness_contract],
            "SpecialQuote": [sq],
            "KioxiaProfile": [kioxia],
            "PolicyAdequacy": [vd],
        },
    }
    return report
