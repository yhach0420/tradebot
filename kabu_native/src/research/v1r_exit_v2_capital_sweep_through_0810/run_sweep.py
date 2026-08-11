"""V1R EXIT V2 Capital Sweep — Arch E Full Replay × 17 capital levels × 15 days.

Paper research only. No Production / strategy / precommit mutation.
submit/cancel/live = 0/0/0.
"""
from __future__ import annotations

import json
import math
import pickle
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

from research.e1_x22_actual_exit_factory.paths import session_end_epoch
from research.e1_x32_upstream_attribution.eval_stages import load_boards_for_symbols
from research.e1_x34c_passive_deployability.events import build_events
from research.e1_x35_passive_exit.paths import build_path
from research.e1_x36_joint_allocator.panel import enrich_events
from research.e1_x36c1m_capital_diagnostic.capital_replay import simulate_joint_capital
from research.e1_x36r_freeze_integrity.serialize import score_fn_from_serialized
from research.e1_x37_prospective.freeze import load_model_artifact, load_v1r
from research.v1r_exit_v2_asymmetric.states import build_trade_bundle
from research.v1r_exit_v2_capital_sweep_through_0810 import (
    ANALYSIS_ID,
    CONTINUATION_ID,
    DAYS_15,
    DAYS_PRE_0810,
    EXIT_CANDIDATE_SHA,
    EXIT_CONTRACT_SHA,
    GUARD_ID,
    CAPITAL_LEVELS,
    REF_0810_ARCH_E_PNL,
    REF_ARCH_E_PF,
    REF_ARCH_E_PNL,
    STRATEGY_ID,
    STRATEGY_SHA,
    VERDICT_COMPLETE,
    VERDICT_IDENTITY_FAIL,
    VERDICT_INTEGRITY_FAIL,
)
from research.v1r_exit_v2_capital_sweep_through_0810.metrics import pick_candidates, summarize_capital_run
from small_paper.v1r_day_engine import (
    _load_boards,
    _planned_anchors_retrospective,
    resolve_pre0905_am_universe,
)
from small_paper.v1r_exit_v2_contract import (
    EXIT_V2_CANDIDATE_SHA,
    FROZEN_CONTINUATION,
    FROZEN_GUARD,
    load_exit_v2_candidate,
    patch_panel_exits,
)
from small_paper.v1r_primary_runtime import (
    BOARD_FRESHNESS_SEC_V1R,
    LOT_QTY,
    MODEL_ARTIFACT_SHA,
    POSITION_CAP,
    V1R_SHA,
    WAIT_SEC,
)

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[3]
OUT = NATIVE / "results" / "research" / "v1r_exit_v2_capital_sweep_through_0810"
PANEL_CACHE = NATIVE / "results/research/v1r_capital_sweep_0p5m_10m/_panel_cache.pkl"
ACTIVATION = NATIVE / "results/research/v1r_exit_v2_prospective_activation"
FEATS = ("spread_bps", "imbalance", "mid_ret_60s", "mid_ret_180s", "event_rate_60s", "log_bid_qty")


def _sfn():
    ser = load_model_artifact()
    raw = score_fn_from_serialized(ser)

    def fn(e: dict) -> float:
        try:
            return float(raw({k: e.get(k) for k in FEATS}))
        except Exception:
            return float("-inf")

    return fn


def _cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)[:32000]
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return str(v)
    return v


def write_xlsx(sheets: dict[str, list[dict]], path: Path) -> None:
    if Workbook is None:
        path.with_suffix(".sheets.json").write_text(
            json.dumps(sheets, indent=2, default=str, ensure_ascii=False), encoding="utf-8"
        )
        return
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append(["(empty)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([_cell(r.get(h)) for h in headers])
    wb.save(path)


def assert_identity() -> dict[str, Any]:
    checks: dict[str, bool] = {}
    try:
        cand = load_exit_v2_candidate()
        checks["exit_candidate_sha"] = cand.get("sha256") == EXIT_CANDIDATE_SHA == EXIT_V2_CANDIDATE_SHA
        checks["guard_id"] = (cand.get("guard") or {}).get("id") == GUARD_ID == FROZEN_GUARD["id"]
        checks["continuation_id"] = (
            (cand.get("continuation") or {}).get("id") == CONTINUATION_ID == FROZEN_CONTINUATION["id"]
        )
        checks["architecture_E"] = cand.get("frozen_architecture") == "E"
    except Exception:
        checks["exit_candidate_sha"] = False
        checks["guard_id"] = False
        checks["continuation_id"] = False
        checks["architecture_E"] = False

    strat_path = ACTIVATION / f"{STRATEGY_ID}.json"
    try:
        st = json.loads(strat_path.read_text(encoding="utf-8"))
        checks["strategy_sha"] = st.get("sha256") == STRATEGY_SHA
        checks["strategy_id"] = st.get("manifest_id") == STRATEGY_ID
        checks["strategy_binds_exit_candidate"] = st.get("exit_v2_candidate_sha") == EXIT_CANDIDATE_SHA
        checks["strategy_exit_contract"] = st.get("exit_contract_sha") == EXIT_CONTRACT_SHA
    except Exception:
        checks["strategy_sha"] = False
        checks["strategy_id"] = False
        checks["strategy_binds_exit_candidate"] = False
        checks["strategy_exit_contract"] = False

    try:
        ex = json.loads((ACTIVATION / "PASSIVE_ASYMMETRIC_EXIT_V2_CONTRACT_V1.json").read_text(encoding="utf-8"))
        checks["exit_contract_sha"] = ex.get("sha256") == EXIT_CONTRACT_SHA
    except Exception:
        checks["exit_contract_sha"] = False

    checks["entry_v1r_sha"] = load_v1r().get("sha256") == V1R_SHA
    checks["model_sha"] = load_model_artifact().get("model_artifact_sha256") == MODEL_ARTIFACT_SHA
    checks["cap"] = POSITION_CAP == 5
    checks["qty"] = LOT_QTY == 100
    checks["wait"] = float(WAIT_SEC) == 1.0
    checks["freshness"] = float(BOARD_FRESHNESS_SEC_V1R) == 5.0
    checks["guard_params"] = (
        FROZEN_GUARD["imb_threshold"] == -0.1
        and FROZEN_GUARD["persist_sec"] == 5.0
        and FROZEN_GUARD["monitor_to"] == 120.0
    )
    checks["cont_params"] = (
        FROZEN_CONTINUATION["mfe_min"] == 60.0
        and FROZEN_CONTINUATION["imb_min"] == 0.1
    )
    failed = [k for k, v in checks.items() if not v]
    return {"pass": len(failed) == 0, "checks": checks, "failed": failed}


def build_0810_panel() -> list[dict]:
    uni = resolve_pre0905_am_universe("20260810")
    if not uni.get("pass"):
        raise RuntimeError(f"0810_universe_blocked:{uni.get('blocked_reason')}")
    symbols = list(uni["symbols"])
    planned = _planned_anchors_retrospective("20260810", symbols)
    boards = _load_boards([("20260810", s) for s in symbols])
    return enrich_events(build_events(planned, boards), boards)


def apply_arch_e_panel(panel: list[dict], boards: dict) -> list[dict]:
    """Patch Arch E exits on all filled rows (pre-admission). Full replay occupancy uses these."""
    by_key = {}
    for e in panel:
        if not (e.get("filled") and e.get("fill_time") is not None):
            continue
        board = boards.get((e["date"], e["symbol"]))
        if board is None or board["t"].size == 0:
            continue
        se = session_end_epoch(e["date"], e["session"])
        path = build_path(
            board,
            entry_price=float(e["fill_price"]),
            entry_t=float(e["fill_time"]),
            sess_end=se,
        )
        by_key[(e["date"], e["symbol"], float(e["fill_time"]))] = build_trade_bundle(e, path, board)
    return patch_panel_exits(panel, by_key, mode="arch_e")


def overall_pf(events: list[dict]) -> Any:
    acc = [e for e in events if e.get("accepted")]
    gp = sum(max(float(e.get("realized_pnl_yen") or 0), 0) for e in acc)
    gl = abs(sum(min(float(e.get("realized_pnl_yen") or 0), 0) for e in acc))
    if gl > 1e-12:
        return gp / gl
    return "INF" if gp > 0 else None


def summary_row(m: dict[str, Any]) -> dict[str, Any]:
    return {
        "保証金": int(m["initial_capital"]),
        "日勝ち": m["wins"],
        "日負け": m["losses"],
        "日勝率": m["win_rate_decided"],
        "日次平均PF": m["daily_pf_mean_finite"],
        "日次PF標準偏差": m["daily_pf_std_finite"],
        "最悪日PF": m["worst_daily_pf"],
        "最大DD円": m["max_dd_yen"],
        "最大DD%": m["max_dd_pct"],
        "最大連敗": m["max_losing_day_streak"],
        "t値": m["t_value"],
        # auxiliary
        "Flat": m["flat_days"],
        "PF∞日": m["daily_pf_inf_days"],
        "最悪日": m["worst_pf_date"],
        "p値": m["two_sided_p_value"],
        "総損益": m["total_pnl"],
        "総Return": m["total_return_pct"],
        "Overall PF": m["overall_pf"],
        "Fills": m["fills"],
        "Capital Blocked": m["capital_blocked"],
    }


def main() -> dict[str, Any]:
    OUT.mkdir(parents=True, exist_ok=True)
    run_id = "v1r_exit_v2_capsweep_" + datetime.now(JST).strftime("%Y%m%d_%H%M%S")
    print(f"=== V1R EXIT V2 CAPITAL SWEEP {run_id} ===", flush=True)

    ident = assert_identity()
    print(f"  identity_pass={ident['pass']} failed={ident['failed']}", flush=True)
    if not ident["pass"]:
        report = {
            "analysis_id": ANALYSIS_ID,
            "run_id": run_id,
            "verdict": VERDICT_IDENTITY_FAIL,
            "identity": ident,
            "production_mutation": "NONE",
            "submit_cancel_live": "0/0/0",
        }
        (OUT / "report.json").write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
        (OUT / "report.md").write_text(f"# Identity FAIL\n\n`{VERDICT_IDENTITY_FAIL}`\n", encoding="utf-8")
        print(json.dumps({"verdict": VERDICT_IDENTITY_FAIL}, indent=2))
        return report

    sfn = _sfn()

    print("  load pre-0810 panel cache...", flush=True)
    cache = pickle.load(PANEL_CACHE.open("rb"))
    am14 = [dict(e) for e in cache["am"]["panel"]]
    assert all(str(e["date"]) < "20260810" for e in am14)
    assert set(e["date"] for e in am14) >= set(DAYS_PRE_0810)

    print("  build 20260810 panel (RETROSPECTIVE_REFERENCE_ONLY)...", flush=True)
    am10 = build_0810_panel()

    panel = am14 + am10
    panel = sorted(panel, key=lambda e: (e["date"], float(e["signal_time"]), str(e["symbol"])))
    print(f"  panel signals={len(panel)} days={sorted({e['date'] for e in panel})}", flush=True)

    # Boards for Arch E
    pairs = sorted({(e["date"], e["symbol"]) for e in panel if e.get("filled")})
    pairs_pre = [p for p in pairs if p[0] < "20260810"]
    pairs_10 = [p for p in pairs if p[0] == "20260810"]
    print(f"  load boards pre={len(pairs_pre)} 0810={len(pairs_10)}...", flush=True)
    boards = {}
    if pairs_pre:
        boards.update(load_boards_for_symbols(pairs_pre))
    if pairs_10:
        boards.update(_load_boards(pairs_10))

    print("  apply Arch E exits to filled panel rows...", flush=True)
    panel_e = apply_arch_e_panel(panel, boards)
    filled_with_exit = sum(
        1 for e in panel_e if e.get("filled") and e.get("canonical_exit_time") is not None
    )
    print(f"  filled_with_arch_e_exit={filled_with_exit}", flush=True)

    # Unlimited integrity: pre-0810 only
    print("  unlimited Arch E integrity (pre-0810)...", flush=True)
    pre_panel = [e for e in panel_e if e["date"] < "20260810"]
    unlim = simulate_joint_capital(pre_panel, score_fn=sfn, initial_cash=None)
    unlim_pnl = float(sum(
        float(e.get("realized_pnl_yen") or 0) for e in unlim["events"] if e.get("accepted")
    ))
    unlim_pf = overall_pf(unlim["events"])
    unlim_n = sum(1 for e in unlim["events"] if e.get("accepted"))
    # Allow tolerance — reference ~2.6226M; board/set differences may exist
    delta_pnl = unlim_pnl - REF_ARCH_E_PNL
    integrity = {
        "pre0810_unlimited_pnl": unlim_pnl,
        "pre0810_unlimited_pf": unlim_pf,
        "pre0810_unlimited_fills": unlim_n,
        "reference_pnl": REF_ARCH_E_PNL,
        "reference_pf": REF_ARCH_E_PF,
        "delta_pnl": delta_pnl,
        "hard_cap_violations": unlim.get("hard_cap_violations"),
        "pass": abs(delta_pnl) < 150_000 and int(unlim.get("hard_cap_violations") or 0) == 0,
        "note": "Tolerance for panel/board lineage; not FIXED600 sweep reuse.",
    }
    print(
        f"  unlimited pnl={unlim_pnl:.0f} (ref={REF_ARCH_E_PNL:.0f}) "
        f"fills={unlim_n} pass={integrity['pass']}",
        flush=True,
    )
    if not integrity["pass"]:
        report = {
            "run_id": run_id,
            "verdict": VERDICT_INTEGRITY_FAIL,
            "identity": ident,
            "integrity": integrity,
            "production_mutation": "NONE",
            "submit_cancel_live": "0/0/0",
        }
        (OUT / "report.json").write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
        (OUT / "report.md").write_text(f"# Integrity FAIL\n\n`{VERDICT_INTEGRITY_FAIL}`\n", encoding="utf-8")
        print(json.dumps({"verdict": VERDICT_INTEGRITY_FAIL, "integrity": integrity}, indent=2, default=str))
        return report

    # 0810 reference check (optional diagnostic)
    only10 = [e for e in panel_e if e["date"] == "20260810"]
    sim10 = simulate_joint_capital(only10, score_fn=sfn, initial_cash=None)
    pnl10 = float(sum(float(e.get("realized_pnl_yen") or 0) for e in sim10["events"] if e.get("accepted")))
    ref10 = {
        "classification": "RETROSPECTIVE_REFERENCE_ONLY",
        "unlimited_pnl": pnl10,
        "reference_approx": REF_0810_ARCH_E_PNL,
        "fills": sum(1 for e in sim10["events"] if e.get("accepted")),
    }
    print(f"  0810 ref unlimited pnl={pnl10:.0f} (approx ref {REF_0810_ARCH_E_PNL})", flush=True)

    primary: list[dict[str, Any]] = []
    daily_all: list[dict[str, Any]] = []
    equity_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []

    for cash in CAPITAL_LEVELS:
        print(f"=== capital ¥{cash:,} ===", flush=True)
        sim = simulate_joint_capital(panel_e, score_fn=sfn, initial_cash=float(cash))
        m = summarize_capital_run(sim, initial_capital=float(cash), days=DAYS_15)
        primary.append(m)
        for drow in m["daily"]:
            daily_all.append({"initial_capital": cash, **drow})
        equity_rows.append({
            "initial_capital": cash,
            **{k: m["equity_dd"][k] for k in m["equity_dd"] if k != "path_n"},
            "path_n": m["equity_dd"]["path_n"],
        })
        blocked_rows.append({
            "initial_capital": cash,
            "capital_blocked": m["capital_blocked"],
            "capacity_blocked": m["capacity_blocked"],
            "admitted": m["admitted"],
            "fills": m["fills"],
            "hard_cap_violations": m["hard_cap_violations"],
            "cash_never_negative": m["cash_never_negative"],
        })
        print(
            f"  pnl={m['total_pnl']:,.0f} fills={m['fills']} "
            f"cap_block={m['capital_blocked']} wr={m['win_rate_decided']} t={m['t_value']}",
            flush=True,
        )

    # High-capital vs unlimited explanation
    hi = primary[-1]
    hi_vs_unlim = {
        "capital": hi["initial_capital"],
        "total_pnl_15d": hi["total_pnl"],
        "pre0810_unlimited_pnl": unlim_pnl,
        "capital_blocked_at_10m": hi["capital_blocked"],
        "fills_15d": hi["fills"],
        "note": (
            "15d includes 20260810 reference day; pre-0810 unlimited is identity check only. "
            "Residual capital_blocked at 10M explained by concurrent notional vs cash."
        ),
    }

    summary_table = [summary_row(m) for m in primary]
    picks = pick_candidates(primary)
    efficiency = [
        {
            "initial_capital": m["initial_capital"],
            "total_pnl": m["total_pnl"],
            "capital_efficiency": m["capital_efficiency"],
            "total_return_pct": m["total_return_pct"],
            "fills": m["fills"],
            "capital_blocked": m["capital_blocked"],
            "t_value": m["t_value"],
            "max_dd_pct": m["max_dd_pct"],
            "win_rate_decided": m["win_rate_decided"],
        }
        for m in primary
    ]
    for a, b in zip(primary, primary[1:]):
        d_cap = float(b["initial_capital"]) - float(a["initial_capital"])
        d_pnl = float(b["total_pnl"]) - float(a["total_pnl"])
        for e in efficiency:
            if e["initial_capital"] == b["initial_capital"]:
                e["incremental_pnl"] = d_pnl
                e["incremental_capital"] = d_cap
                e["incremental_pnl_per_capital"] = (d_pnl / d_cap) if d_cap else None

    main_table = [
        {
            "保証金": r["保証金"],
            "日勝ち": r["日勝ち"],
            "日負け": r["日負け"],
            "日勝率": r["日勝率"],
            "日次平均PF": r["日次平均PF"],
            "日次PF標準偏差": r["日次PF標準偏差"],
            "最悪日PF": r["最悪日PF"],
            "最大DD円": r["最大DD円"],
            "最大DD%": r["最大DD%"],
            "最大連敗": r["最大連敗"],
            "t値": r["t値"],
        }
        for r in summary_table
    ]

    sheets = {
        "Summary": summary_table,
        "Daily": daily_all,
        "Equity_DD": equity_rows,
        "Capital_Blocked": blocked_rows,
        "Capital_Efficiency": efficiency,
        "Diagnostics": [
            {"item": "identity", "value": ident},
            {"item": "integrity_pre0810_unlimited", "value": integrity},
            {"item": "ref_0810", "value": ref10},
            {"item": "hi_vs_unlim", "value": hi_vs_unlim},
            {"item": "picks", "value": picks},
            {"item": "strategy_sha", "value": STRATEGY_SHA},
            {"item": "exit_candidate_sha", "value": EXIT_CANDIDATE_SHA},
            {"item": "exit_contract_sha", "value": EXIT_CONTRACT_SHA},
            {"item": "guard", "value": GUARD_ID},
            {"item": "continuation", "value": CONTINUATION_ID},
            {"item": "days", "value": list(DAYS_15)},
            {"item": "submit_cancel_live", "value": "0/0/0"},
            {"item": "production_mutation", "value": "NONE"},
        ],
    }
    write_xlsx(sheets, OUT / "v1r_exit_v2_capital_sweep_through_0810.xlsx")

    verdict = VERDICT_COMPLETE
    report = {
        "analysis_id": "V1R_EXIT_V2_CAPITAL_SWEEP_THROUGH_20260810",
        "run_id": run_id,
        "verdict": verdict,
        "identity": ident,
        "integrity": integrity,
        "ref_0810": ref10,
        "strategy_id": STRATEGY_ID,
        "strategy_sha": STRATEGY_SHA,
        "exit_candidate_sha": EXIT_CANDIDATE_SHA,
        "exit_contract_sha": EXIT_CONTRACT_SHA,
        "guard_id": GUARD_ID,
        "continuation_id": CONTINUATION_ID,
        "days": list(DAYS_15),
        "capital_levels": list(CAPITAL_LEVELS),
        "main_table": main_table,
        "summary_table": summary_table,
        "picks": picks,
        "hi_vs_unlim": hi_vs_unlim,
        "production_mutation": "NONE",
        "submit_cancel_live": "0/0/0",
        "answers": {
            "1_win_rate_max": picks["win_rate_max"],
            "2_t_value_max": picks["t_value_max"],
            "3_dd_pct_min": picks["dd_pct_min"],
            "4_pf_stability": picks["pf_stability_best"],
            "5_balance": picks["balance"],
        },
    }
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    def _fmt(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            if abs(v) >= 1000:
                return f"{v:,.1f}"
            return f"{v:.6g}"
        return str(v)

    md = [
        f"# V1R EXIT V2 Capital Sweep Through 2026-08-10",
        "",
        f"## Verdict",
        f"`{verdict}`",
        "",
        f"- Strategy: `{STRATEGY_ID}` / `{STRATEGY_SHA}`",
        f"- EXIT candidate: `{EXIT_CANDIDATE_SHA}`",
        f"- Guard/Cont: `{GUARD_ID}` / `{CONTINUATION_ID}`",
        f"- Days: 15 (incl 20260810 RETROSPECTIVE_REFERENCE_ONLY)",
        f"- Production mutation: NONE",
        f"- submit/cancel/live: 0/0/0",
        "",
        "## Main table",
        "",
    ]
    hdr = list(main_table[0].keys())
    md.append("| " + " | ".join(hdr) + " |")
    md.append("| " + " | ".join("---" for _ in hdr) + " |")
    for r in main_table:
        md.append("| " + " | ".join(_fmt(r.get(h)) for h in hdr) + " |")
    md += [
        "",
        "## Picks",
        f"- 日勝率最大: `{picks['win_rate_max']}`",
        f"- t値最大: `{picks['t_value_max']}`",
        f"- DD%最小: `{picks['dd_pct_min']}`",
        f"- PF安定: `{picks['pf_stability_best']}`",
        f"- Balance: `{picks['balance']}`",
        "",
        "STOP.",
    ]
    (OUT / "report.md").write_text("\n".join(md) + "\n", encoding="utf-8")
    print(json.dumps({"verdict": verdict, "picks": picks}, indent=2, default=str))
    return report


if __name__ == "__main__":
    main()
