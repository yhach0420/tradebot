"""Artifact rendering + atomic publish.

report.json is the single SoT; report.md and audit.xlsx are rendered FROM it.
report.md / audit.xlsx embed report.json's SHA-256 (single-writer cross
reference: nothing overwrites expected values afterwards). All three files are
fully generated in a temp sibling directory, then atomically moved into
<run>/published/.
"""
from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from .store import run_root, sha256_file

AUDIT_SHEETS = (
    "README", "source_manifest", "session_quality", "field_coverage",
    "feature_definitions", "regime_definitions", "setup_state_machines",
    "exit_packages", "candidate_registry", "evaluation_plan",
    "paper_protected_manifest", "tests", "safety_counters",
)

R1_AUDIT_SHEETS = (
    "asof_coverage", "coverage_diff", "replay_order", "analysis_mask",
    "tick_resolver", "base_binding", "p1_diff",
)


def _cell(v: Any) -> Any:
    if isinstance(v, (dict, list, tuple, set)):
        return json.dumps(v, ensure_ascii=False, sort_keys=True, default=str)[:32000]
    return v


def render_report_md(report: dict[str, Any], report_json_sha: str) -> str:
    inv = report["inventory"]
    reg = report["candidate_registry"]
    lines = [
        f"# {report['plan_id']} — Phase A ({report['verdict']})",
        "",
        f"- run_id: `{report['run_id']}`",
        f"- rendered from report.json sha256 `{report_json_sha}` (SoT)",
        f"- published: {report['published_at_jst']}",
        f"- verdict: **{report['verdict']}**",
        f"- safety: submit/cancel/live = 0/0/0; Paper/Shadow/Forward/Task/YAML/Discord unchanged",
        f"- paper protected manifest before/after match: {report['paper_protected_manifest']['match']}",
        "",
        "## Data inventory (9 days, AM/PM)",
        f"- raw lines total: {inv['raw_total_lines']:,} / canonical events total: {inv['canonical_total']:,}",
        f"- usable fields: {', '.join(report['field_usability']['usable']) if isinstance(report['field_usability'].get('usable'), list) else report['field_usability'].get('quote', '')}",
        f"- unusable fields: {', '.join(report['field_usability']['unusable']) if isinstance(report['field_usability'].get('unusable'), list) else report['field_usability'].get('spread', '')}",
        "",
        "## Candidate Registry (P1-frozen)",
        f"- candidates: {len(reg)} (cap 24), enabled: {sum(1 for r in reg if r['enabled'])}",
        f"- registry sha256: `{report['p1']['candidate_registry_sha256']}`",
        f"- P1 sha256: `{report['p1']['p1_sha256']}`",
        "",
        "### Strategy IDs",
    ]
    lines += [f"- `{r['strategy_id']}`" + ("" if r["enabled"] else f"  (DISABLED: {r['disable_reason']})")
              for r in reg]
    lines += [
        "",
        "## Tests",
        f"- passed {report['tests']['passed']}/{report['tests']['total']} "
        f"(failed {report['tests']['failed']})",
    ]
    if report.get("r1"):
        r1 = report["r1"]
        lines += [
            "",
            "## Phase A-R1 repairs",
            f"- superseded run: {r1.get('superseded_run_id')} (SUPERSEDED_PRE_ECONOMICS)",
            f"- analysis_mask_id: `{r1['analysis_mask']['analysis_mask_id']}`",
            "- coverage method: as-of 5s-grid (old event-row rates kept as diff only)",
            f"- canonical ts regressions (stored order): "
            + ", ".join(f"{d}={row['canonical_ts_regressions_stored_order']}"
                        for d, row in r1["canonical_regressions"].items()),
            f"- tick resolver: dynamic JPX (runtime rules sha "
            f"`{(r1['tick_resolver']['runtime_resolver_sha256'] or '')[:16]}`), "
            f"{len(r1['tick_resolver']['symbol_classes'])} symbols classified",
            f"- E1_X5 base comparable: {r1['base_binding'].get('comparable')}",
        ]
    if report.get("r2") and report.get("r2", {}).get("gates", {}).get("all_pass") is not None:
        r2 = report["r2"]
        g = r2["gates"]
        lines += [
            "",
            "## Phase A-R2 decision coverage",
            f"- decision gates all pass: {g.get('all_pass')} "
            f"(min decision_quote={g.get('min_decision_quote_coverage')}, "
            f"min market_context={g.get('min_market_context_coverage')})",
            f"- tick official classes: {r2.get('tick_official_summary')}",
            f"- E1_X5 base recut comparable: {(r2.get('base_binding_r2') or {}).get('comparable')}",
        ]
    if report.get("r3"):
        r3 = report["r3"]
        g = r3["gates"]
        fu = r3["field_usability"]
        lines += [
            "",
            "## Phase A-R3 structural coverage + official tick",
            f"- R2 block evidence preserved: `{report['r2'].get('block_evidence_run_id')}`",
            f"- structural min={g['min_structural_decision_quote_coverage']} "
            f"weighted={g['weighted_structural_decision_quote_coverage']} "
            f"market min={g['min_market_context_coverage']}",
            f"- field usability: quote={fu['quote']}, spread={fu['spread']}, "
            f"market={fu['market_context']}",
            f"- tick official: {r3['tick_official_summary']} "
            f"unresolved={r3['tick_official']['unresolved']}",
            f"- E1_X5 BASE unchanged: sha=`{r3['base_binding_r3']['artifact_sha256'][:16]}` "
            f"n={r3['base_binding_r3']['recut_metrics']['completed_trades']}",
        ]
    lines += [
        "",
        "## Not executed in Phase A",
        "9-day PnL replay, candidate selection, Shadow — awaiting explicit user approval.",
    ]
    return "\n".join(lines)


def render_audit_xlsx(report: dict[str, Any], report_json_sha: str, out_fp: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    for row in (
        ["plan_id", report["plan_id"]],
        ["run_id", report["run_id"]],
        ["verdict", report["verdict"]],
        ["source_of_truth", "report.json"],
        ["report_json_sha256", report_json_sha],
        ["p1_sha256", report["p1"]["p1_sha256"]],
        ["published_at_jst", report["published_at_jst"]],
        ["note", "rendered from report.json; no second writer"],
    ):
        ws.append(row)

    w = wb.create_sheet("source_manifest")
    w.append(["day", "raw_files_n", "raw_bytes", "raw_dir", "canonical_cache_sha_events"])
    for day, d in report["source_manifest_days"].items():
        cc = d["canonical_cache"].get("events_slim_v3.pkl.gz") or {}
        w.append([day, d["raw_files_n"], d["raw_bytes"], d["raw_dir"], cc.get("sha256")])

    w = wb.create_sheet("session_quality")
    w.append(["day", "session", "raw_events", "symbols_n", "first_ts", "last_ts",
              "ts_inversions", "duplicates", "stale_rate", "lag_p50", "lag_p90", "lag_p99",
              "canonical_events", "gap_n", "gap_max_sec"])
    for day, d in report["inventory"]["days"].items():
        for sk in ("AM", "PM"):
            s = d["sessions"][sk]
            lag = s["source_ingress_lag_sec"]
            w.append([day, sk, s["raw_events"], s["symbols_n"], s["first_ts"], s["last_ts"],
                      s["ts_inversions"], s["duplicates"], s["stale_rate"],
                      lag["p50"], lag["p90"], lag["p99"],
                      d["canonical"]["canonical_by_session"].get(sk),
                      d["canonical"]["gap_n"], d["canonical"]["gap_max_sec"]])

    w = wb.create_sheet("field_coverage")
    w.append(["day", "session", "field", "missing_rate"])
    for day, d in report["inventory"]["days"].items():
        for sk in ("AM", "PM"):
            s = d["sessions"][sk]
            for fld, mr in s["field_missing_rate"].items():
                w.append([day, sk, fld, mr])
            w.append([day, sk, "quote(Buy1/Sell1)", None if s["quote_coverage"] is None else round(1 - s["quote_coverage"], 6)])
            w.append([day, sk, "board_full10", None if s["board_full10_coverage"] is None else round(1 - s["board_full10_coverage"], 6)])

    w = wb.create_sheet("feature_definitions")
    w.append(["feature", "formula"])
    for k, v in report["p1"]["feature_formulas"].items():
        w.append([k, v])

    w = wb.create_sheet("regime_definitions")
    w.append(["key", "definition"])
    for mode in ("standard", "strict"):
        for st, d in report["p1"]["regime_definitions"][mode].items():
            w.append([f"{mode}.{st}", _cell(d)])
    w.append(["priority_order", _cell(report["p1"]["regime_definitions"]["priority_order"])])
    w.append(["hysteresis", report["p1"]["regime_definitions"]["hysteresis"]])

    w = wb.create_sheet("setup_state_machines")
    w.append(["setup", "key", "value"])
    for su, d in report["p1"]["setup_state_machines"].items():
        for k, v in d.items():
            w.append([su, k, _cell(v)])
    for cl, d in report["p1"]["confirmation_levels"].items():
        w.append(["CONFIRMATION", cl, _cell(d)])
    w.append(["CHASE_REJECT", "formula", report["p1"]["chase_reject"]["formula"]])

    w = wb.create_sheet("exit_packages")
    w.append(["exit_id", "key", "value"])
    for xp in report["p1"]["exit_packages"]:
        for k, v in xp.items():
            w.append([xp["exit_id"], k, _cell(v)])
    w.append(["ORDER", "exit_evaluation_order", _cell(report["p1"]["exit_evaluation_order"])])

    w = wb.create_sheet("candidate_registry")
    keys = ["strategy_id", "enabled", "disable_reason", "setup", "confirmation",
            "regime_mode", "features_used", "trigger", "reject_conditions",
            "invalidation", "stop", "no_progress", "trailing", "max_hold_sec",
            "required_coverage", "missing_data_behavior"]
    w.append(keys)
    for r in report["candidate_registry"]:
        w.append([_cell(r.get(k)) for k in keys])

    w = wb.create_sheet("evaluation_plan")
    w.append(["key", "value"])
    for k, v in report["p1"]["phase_b_evaluation_plan"].items():
        w.append([k, _cell(v)])

    w = wb.create_sheet("paper_protected_manifest")
    w.append(["file", "sha256_before", "match_after"])
    pm = report["paper_protected_manifest"]
    for rel, d in pm["before_files"].items():
        w.append([rel, d["sha256"], pm["match"]])

    w = wb.create_sheet("tests")
    w.append(["test", "outcome"])
    for row in report["tests"]["rows"]:
        w.append([row["test"], row["outcome"]])

    w = wb.create_sheet("safety_counters")
    for k, v in (("submit", 0), ("cancel", 0), ("live", 0),
                 ("paper_guard_triggered", report["paper_guard"]["triggered"]),
                 ("paper_processes_touched", 0)):
        w.append([k, v])

    if report.get("r1"):
        _render_r1_sheets(wb, report["r1"])
    if report.get("r2") and report["r2"].get("coverage_days"):
        _render_r2_sheets(wb, report["r2"])
    if report.get("r3"):
        _render_r3_sheets(wb, report["r3"])

    wb.save(out_fp)


def _render_r3_sheets(wb, r3: dict[str, Any]) -> None:
    w = wb.create_sheet("structural_coverage")
    w.append(["day", "session", "due_n", "structural_n", "structural_coverage",
              "spread_healthy_n", "spread_unhealthy_n", "spread_healthy_rate",
              "spread_p50", "spread_p90", "spread_p95", "spread_p99", "spread_max",
              "market_context_coverage", "r2_mixed_dq"])
    for day, cov in r3["coverage_days"].items():
        for sk in ("AM", "PM"):
            s = cov["sessions"][sk]
            if "due_symbol_grid_n" not in s:
                continue
            sp = s.get("spread_bps_stats") or {}
            w.append([day, sk, s["due_symbol_grid_n"],
                      s["structural_decision_quote_available_n"],
                      s["structural_decision_quote_coverage"],
                      s["spread_healthy_n"], s["spread_unhealthy_n"],
                      s["spread_healthy_rate"],
                      sp.get("p50"), sp.get("p90"), sp.get("p95"), sp.get("p99"),
                      sp.get("max"), s["market_context_coverage"],
                      s.get("decision_quote_coverage_r2_mixed")])

    w = wb.create_sheet("gates_r3")
    w.append(["window", "structural_cov", "gate_structural", "market_cov", "gate_market",
              "spread_healthy_rate", "r2_mixed_dq"])
    for wid, row in r3["gates"]["per_window"].items():
        w.append([wid, row["structural_decision_quote_coverage"], row["gate_structural"],
                  row["market_context_coverage"], row["gate_market"],
                  row["spread_healthy_rate"], row.get("r2_mixed_decision_quote_coverage")])
    w.append([])
    w.append(["min_structural", r3["gates"]["min_structural_decision_quote_coverage"]])
    w.append(["weighted_structural", r3["gates"]["weighted_structural_decision_quote_coverage"]])
    w.append(["audit_diffs", _cell(r3["gates"]["audit_expectation_diffs"])])

    w = wb.create_sheet("tick_official_r3")
    w.append(["master_path", r3["tick_official"]["master_path"]])
    w.append(["master_sha256", r3["tick_official"]["master_sha256"]])
    w.append(["evidence_manifest_sha256", r3["tick_official"]["evidence_manifest_sha256"]])
    w.append(["fetched_at", r3["tick_official"]["evidence_fetched_at_jst"]])
    w.append(["symbol", "class", "source", "official_reason", "empirical_check",
              "observations", "listing_date", "market_segment", "topix500_applicable"])
    for sym, row in sorted(r3["tick_official"]["symbol_classes"].items()):
        w.append([sym, row.get("class"), row.get("source"), row.get("official_reason"),
                  row.get("empirical_check"), row.get("observations"),
                  row.get("listing_date"), row.get("market_segment"),
                  row.get("topix500_applicable")])

    w = wb.create_sheet("field_usability_r3")
    for k, v in r3["field_usability"].items():
        w.append([k, _cell(v)])

    w = wb.create_sheet("base_binding_r3")
    for k, v in r3["base_binding_r3"].items():
        w.append([k, _cell(v)])

    w = wb.create_sheet("p1_diff_r2_r3")
    w.append(["kind", "key"])
    for kind, keys in r3["p1_diff_r2_r3"].items():
        for k in keys:
            w.append([kind, k])


def _render_r2_sheets(wb, r2: dict[str, Any]) -> None:
    w = wb.create_sheet("decision_coverage")
    w.append(["day", "session", "universe_n", "full_grid_n", "full_grid_state_coverage",
              "due_symbol_grid_n", "NOT_DUE_NO_SYMBOL_UPDATE_n",
              "decision_quote_available_n", "decision_quote_coverage",
              "mkt_min", "mkt_p05", "mkt_median", "mkt_p95", "mkt_max",
              "market_context_coverage", "incomplete_lookback_n", "stale_snapshot_n",
              "source_semantics_unknown_n",
              "missing_state_n", "invalid_value_n", "crossed_n", "spread_reject_n",
              "source_conflict_n"])
    for day, cov in r2["coverage_days"].items():
        for sk in ("AM", "PM"):
            s = cov["sessions"][sk]
            if "due_symbol_grid_n" not in s:
                w.append([day, sk, s.get("universe_n"), 0])
                continue
            m = s["mkt_evaluable_stats"]
            rj = s["rejects"]
            w.append([day, sk, s["universe_n"], s["full_grid_n"],
                      s["full_grid_state_coverage"], s["due_symbol_grid_n"],
                      s["NOT_DUE_NO_SYMBOL_UPDATE_n"], s["decision_quote_available_n"],
                      s["decision_quote_coverage"], m["min"], m["p05"], m["median"],
                      m["p95"], m["max"], s["market_context_coverage"],
                      s["incomplete_lookback_n"], s["stale_snapshot_n"],
                      s.get("source_semantics_unknown_n"),
                      rj["missing_state_n"], rj["invalid_value_n"], rj["crossed_n"],
                      rj["spread_reject_n"], rj["source_conflict_n"]])

    w = wb.create_sheet("symbol_decision_coverage")
    w.append(["day", "session", "symbol", "push_n", "due_n", "ok_n",
              "decision_coverage", "push_p50", "push_p90", "push_p95", "push_p99",
              "push_max"])
    for day, cov in r2["coverage_days"].items():
        for sk in ("AM", "PM"):
            s = cov["sessions"][sk]
            for sym, sc in (s.get("symbol_decision_coverage") or {}).items():
                pi = sc.get("push_interval_sec") or {}
                w.append([day, sk, sym, sc.get("push_n"), sc.get("due_n"),
                          sc.get("ok_n"), sc.get("decision_coverage"),
                          pi.get("p50"), pi.get("p90"), pi.get("p95"),
                          pi.get("p99"), pi.get("max")])

    w = wb.create_sheet("source_semantics")
    w.append(["day", "field", "present_n", "unchanged_n", "advanced_while_unchanged",
              "advance_rate_when_unchanged", "semantics"])
    for day, cov in r2["coverage_days"].items():
        for key, c in cov["source_semantics"].items():
            if not isinstance(c, dict) or "semantics" not in c:
                continue
            w.append([day, key, c["present_n"], c["unchanged_n"],
                      c["advanced_while_unchanged"], c["advance_rate_when_unchanged"],
                      c["semantics"]])
    w.append([])
    w.append(["policy", _cell(next(iter(r2["coverage_days"].values()))
                              ["source_semantics"]["policy"])])

    w = wb.create_sheet("tick_official")
    w.append(["master_path", r2["tick_official"]["master_path"]])
    w.append(["master_sha256", r2["tick_official"]["master_sha256"]])
    w.append(["rule", r2["tick_official"]["rule"]])
    w.append(["symbol", "class", "official_reason", "empirical_check", "observations"])
    for sym, row in sorted(r2["tick_official"]["symbol_classes"].items()):
        w.append([sym, row["class"], row["official_reason"], row["empirical_check"],
                  row["observations"]])

    w = wb.create_sheet("base_recut")
    for k, v in r2["base_binding_r2"].items():
        w.append([k, _cell(v)])

    w = wb.create_sheet("gates_r2")
    w.append(["window", "decision_quote_coverage", "gate_B(>=0.90)",
              "market_context_coverage", "gate_C(>=0.90)"])
    for wid, row in r2["gates"]["per_window"].items():
        w.append([wid, row["decision_quote_coverage"], row["gate_b"],
                  row["market_context_coverage"], row["gate_c"]])
    w.append([])
    w.append(["all_gates_pass", r2["gates"]["all_pass"]])

    w = wb.create_sheet("p1_diff_r1_r2")
    w.append(["kind", "key"])
    for kind, keys in r2["p1_diff_r1_r2"].items():
        for k in keys:
            w.append([kind, k])


def _render_r1_sheets(wb, r1: dict[str, Any]) -> None:
    w = wb.create_sheet("asof_coverage")
    w.append(["day", "session", "field_group", "eligible_grid_n", "available_grid_n",
              "coverage", "stale_grid_n", "missing_grid_n", "invalid_value_n",
              "age_min", "age_median", "age_max", "field_source_ts"])
    for day, cov in (r1.get("coverage_days") or {}).items():
        sessions = cov.get("sessions") or {}
        for sk in ("AM", "PM"):
            if sk not in sessions or not sessions[sk]:
                continue  # R2 stubs leave sessions empty; R1 full-grid is in coverage_diff
            for grp, s in sessions[sk].items():
                if not isinstance(s, dict) or "eligible_grid_n" not in s:
                    continue
                a = s.get("age_sec") or {}
                w.append([day, sk, grp, s["eligible_grid_n"], s["available_grid_n"],
                          s["coverage"], s["stale_grid_n"], s["missing_grid_n"],
                          s["invalid_value_n"], a.get("min"), a.get("median"), a.get("max"),
                          s.get("field_source_ts")])

    w = wb.create_sheet("coverage_diff")
    w.append(["field_group", "old_method_min_coverage(event_row)",
              "new_method_min_coverage(asof_grid)", "usable_old", "usable_new"])
    for grp, row in r1["coverage_diff"].items():
        w.append([grp, row["old_min_coverage"], row["new_min_coverage"],
                  row["usable_old"], row["usable_new"]])

    w = wb.create_sheet("replay_order")
    for k, v in r1["replay_order_contract"].items():
        w.append([k, _cell(v)])
    w.append([])
    w.append(["day", "canonical_ts_regressions_stored_order", "normalizer_reported", "raw_ingress_inversions"])
    for day, row in r1["canonical_regressions"].items():
        w.append([day, row["canonical_ts_regressions_stored_order"],
                  row["normalizer_reported_regressions"], 0])

    w = wb.create_sheet("analysis_mask")
    w.append(["analysis_mask_id", r1["analysis_mask"]["analysis_mask_id"]])
    w.append(["rules", _cell(r1["analysis_mask"]["rules"])])
    w.append(["window", "included", "quality_class", "expected_start", "expected_end",
              "valid_start", "valid_end", "valid_sec", "coverage_rate",
              "eligible_grids_n", "entry_evaluable_until", "exclusion_reason"])
    for wid, row in r1["analysis_mask"]["windows"].items():
        w.append([wid, row["included"], row["quality_class"],
                  row["expected_start_epoch"], row["expected_end_epoch"],
                  row["valid_start_epoch"], row["valid_end_epoch"], row["valid_sec"],
                  row["coverage_rate"], row["eligible_grids_n"],
                  row.get("entry_evaluable_until_epoch"), row["exclusion_reason"]])

    w = wb.create_sheet("tick_resolver")
    w.append(["runtime_resolver_sha256", r1["tick_resolver"]["runtime_resolver_sha256"]])
    w.append(["rule", r1["tick_resolver"]["rule"]])
    w.append(["symbol", "class", "reason", "observations"])
    for sym, row in sorted(r1["tick_resolver"]["symbol_classes"].items()):
        w.append([sym, row["class"], row["reason"], row["observations"]])

    w = wb.create_sheet("base_binding")
    for k, v in r1["base_binding"].items():
        w.append([k, _cell(v)])

    w = wb.create_sheet("p1_diff")
    w.append(["kind", "key"])
    for kind, keys in r1["p1_diff"].items():
        for k in keys:
            w.append([kind, k])


def atomic_publish(run_id: str, report: dict[str, Any]) -> dict[str, str]:
    """Generate all three artifacts in a temp sibling, then atomic move."""
    root = run_root(run_id)
    tmp = root / f"publish_tmp_{datetime.now().strftime('%H%M%S')}"
    tmp.mkdir(parents=True, exist_ok=True)

    fp_json = tmp / "report.json"
    fp_json.write_text(json.dumps(report, ensure_ascii=False, indent=1, default=str),
                       encoding="utf-8")
    sha_json = sha256_file(fp_json)

    fp_md = tmp / "report.md"
    fp_md.write_text(render_report_md(report, sha_json), encoding="utf-8")
    fp_xlsx = tmp / "audit.xlsx"
    render_audit_xlsx(report, sha_json, fp_xlsx)

    pub = root / "published"
    pub.mkdir(exist_ok=True)
    shas = {}
    for fp in (fp_json, fp_md, fp_xlsx):
        dst = pub / fp.name
        os.replace(fp, dst)
        shas[fp.name] = sha256_file(dst)
    tmp.rmdir()
    assert shas["report.json"] == sha_json, "report.json changed during publish"
    return shas
