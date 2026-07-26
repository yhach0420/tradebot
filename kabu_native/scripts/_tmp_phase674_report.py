#!/usr/bin/env python3
"""Phase674 report consolidator (read-only analysis output)."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[1]
EV = NATIVE / "results" / "reports" / "phase674_pm_summary_missing_evidence_20260721"
OUT = NATIVE / "results" / "research" / "phase674_pm_summary_missing_20260721"
OUT.mkdir(parents=True, exist_ok=True)

NOW = datetime.now(JST).isoformat()

report = {
    "phase": "Phase674",
    "title": "2026-07-21 PM Summary Missing Root Cause",
    "generated_at": NOW,
    "verdict": "PM_SUMMARY_ROOT_CAUSE_IDENTIFIED",
    "primary_cause": "A_WEBSOCKET_RECV_FREEZE",
    "secondary_cause": "I_RUNNER_STATE_TRANSITION_BUG",
    "confidence": 0.93,
    "am_pm_relation": "A_SAME_FAILURE",
    "am_pm_relation_detail": "B_same_websocket_origin_stop_point_equivalent",
    "required_answers": {
        "1_last_normal_push_time": "2026-07-21T15:18:35+09:00",
        "2_last_heartbeat_time": "2026-07-21T15:13:49+09:00",
        "3_last_capture_time": "2026-07-21T15:35:18+09:00 (CAPTURE_COMPLETE; last paper fanout event_count=359445 == last msg)",
        "4_runtime_pid_alive": True,
        "5_cpu_thread_state": {
            "pilot_pid": 21836,
            "pilot_cpu_sec_at_investigation": 2465.39,
            "pilot_ws_mb": 601.9,
            "pilot_threads": 26,
            "daily_pid": 29024,
            "pyspy_main": "asyncio.windows_events._poll/select idle in run_forever after reconnect",
            "pyspy_discord": "idle queue wait (not HTTP hang)",
            "pyspy_daily": "subprocess.communicate join waiting for pilot",
        },
        "6_event_loop_freeze_proven": True,
        "7_summary_function_called": False,
        "8_session_close_started": False,
        "9_active_positions_orphan": {"orphan_open": 4, "accepted": 25, "observer_exit": 21},
        "10_websocket_final_state": {
            "last_error": "push_unexpected: no close frame received or sent",
            "reconnect_count_logged": 1,
            "reconnect_success_evidence": "no reconnect_register error; process stuck in iter_messages recv",
            "push_after_reconnect": False,
        },
        "11_am_pm_same_cause": True,
        "12_w70_w71_related": False,
        "13_stale_trade_15s_related": False,
        "14_last_completed_step": "push_unexpected logged + reconnect_count=1 appended to errors.jsonl at 15:18:35; last candidate/rejected am_pm_entry_stop processed",
        "15_first_incomplete_step": "post-reconnect push.iter_messages recv/select wait — never returned to while-loop force_close/session_end/summary",
        "16_primary_cause": "A_WEBSOCKET_RECV_FREEZE",
        "17_confidence": 0.93,
        "18_repro_conditions": "Live Kabu PUSH WS unexpected close near am_pm entry_stop (AM 11:20 / PM 15:18); reconnect attempted; new/half-open socket blocks asyncio recv; force_close only polled outside async-for",
        "19_minimal_fix": [
            "Inside async for / recv wait: periodically call _maybe_am_pm_force_close + _should_stop (or recv timeout wakeup)",
            "On push_unexpected: hard deadline for reconnect; if no push within N seconds, _request_stop('push_reconnect_timeout') and fall through to summary/finalize",
            "Watchdog thread or heartbeat timer independent of WS recv to ensure session_end/force_close cannot be skipped",
        ],
        "20_runtime_logic_unchanged_this_investigation": True,
        "21_submit_cancel_zero": True,
        "22_artifact_paths": {
            "evidence": str(EV),
            "report_md": str(OUT / "report.md"),
            "report_json": str(OUT / "report.json"),
            "audit_xlsx": str(OUT / "audit.xlsx"),
        },
    },
    "session": {
        "path": str(NATIVE / "results/small_paper/20260721/live_session_124342"),
        "start": "12:33 (policy) / ready 12:43:44",
        "expected_close": "15:23 afternoon_session_close",
        "entry_stop": "15:18",
        "state": "FROZEN_ALIVE_PID — mid-session artifacts stopped; no seal/finalize",
        "summary_artifact": "small_paper_summary.json exists but is mid-run heartbeat sync (ended_at=15:13:49), NOT final PM Summary",
        "session_seal": False,
        "session_archive_backup": None,
        "session_external_backup": None,
        "finalize_flags": "absent",
    },
    "timeline_after_1500": [
        {"t": "15:03:44", "e": "observer_exit 7974.T no_progress_exit"},
        {"t": "15:03:45", "e": "ENTRY accepted 3449.T (price_age_sec=14.6 liquidity_stale_trade warning); Discord entry delivery recorded"},
        {"t": "15:03:48", "e": "Heartbeat #28 push=324750"},
        {"t": "15:08:48", "e": "Heartbeat #29"},
        {"t": "15:13:49", "e": "Heartbeat #30 push=347896; summary.json mtime sync"},
        {"t": "15:18:00+", "e": "am_pm_entry_stop rejects dominate (entry_stop reached)"},
        {"t": "15:18:35", "e": "LAST PUSH/process events; push_unexpected no close frame; reconnect_count=1"},
        {"t": "15:18:35+", "e": "NO further events/heartbeat/summary/seal"},
        {"t": "15:23:00", "e": "force_close DUE but NOT executed"},
        {"t": "15:35:18", "e": "Capture CAPTURE_COMPLETE (scheduled end); paper fanout already stopped at 359445"},
        {"t": "20:42+", "e": "Investigation: pilot PID still alive; py-spy MainThread in asyncio select"},
    ],
    "last_completed_step": "errors.jsonl: push_unexpected + reconnect(1) at 15:18:35",
    "first_incomplete_step": "return from blocking push.iter_messages after reconnect → _maybe_am_pm_force_close / session_end / final summary / Discord PM Summary / seal",
    "cause_matrix": [
        {
            "code": "A_WEBSOCKET_RECV_FREEZE",
            "role": "PRIMARY",
            "support": [
                "errors: push_unexpected no close frame at 15:18:35",
                "py-spy MainThread idle in asyncio _poll/select under run_forever/run_live_dry_run",
                "no events/heartbeat after 15:18:35 while PID alive",
            ],
            "contra": [],
            "confidence": 0.93,
        },
        {
            "code": "B_WEBSOCKET_RECONNECT_DEADLOCK",
            "role": "MECHANISM",
            "support": [
                "reconnect_count=1 logged immediately after push_unexpected",
                "no reconnect_register error; loop likely re-entered iter_messages and blocked",
            ],
            "contra": ["cannot prove register returned True vs False without live log"],
            "confidence": 0.85,
        },
        {
            "code": "I_RUNNER_STATE_TRANSITION_BUG",
            "role": "SECONDARY",
            "support": [
                "_maybe_am_pm_force_close only at while-loop top, not during blocking recv",
                "force_close 15:23 never produced afternoon_session_close exits",
            ],
            "contra": [],
            "confidence": 0.9,
        },
        {
            "code": "F_SUMMARY_GENERATION_HANG",
            "role": "REJECTED",
            "support": [],
            "contra": [
                "final summary/seal/archive keys absent",
                "code after asyncio.run(_loop()) never reached",
            ],
            "confidence": 0.05,
        },
        {
            "code": "G_DISCORD_SEND_HANG",
            "role": "REJECTED",
            "support": [],
            "contra": ["py-spy discord worker idle on queue.get"],
            "confidence": 0.02,
        },
        {
            "code": "H_ARCHIVE_BACKUP_HANG",
            "role": "REJECTED",
            "support": [],
            "contra": [
                "summary.session_archive_backup is null",
                "W70/W71 hooks only after loop exit",
            ],
            "confidence": 0.02,
        },
        {
            "code": "E_ORPHAN_CLOSE_WAIT",
            "role": "CONSEQUENCE",
            "support": ["4 orphan OPEN remain including 3449.T"],
            "contra": ["orphans exist because close never ran, not because wait-loop blocked"],
            "confidence": 0.2,
        },
        {
            "code": "J_PROCESS_CRASH",
            "role": "REJECTED",
            "support": [],
            "contra": ["PID 21836 still alive hours later"],
            "confidence": 0.01,
        },
    ],
    "w70_w71_audit": {
        "related": False,
        "evidence": [
            "session_archive_backup/session_external_backup absent in summary",
            "hooks in pilot_runner are after asyncio.run(_loop()) completes",
            "loop never completed → hooks never called",
            "py-spy not in archive/SHA256/robocopy stacks",
        ],
    },
    "stale_trade_audit": {
        "related": False,
        "entry": "3449.T 15:03:45 price_age_sec=14.6 liquidity_stale_trade",
        "evidence": [
            "PUSH continued ~15 minutes after ENTRY until 15:18:35",
            "Heartbeats 15:03/08/13 after ENTRY",
            "warning is non-reject metadata; freeze signature is WS close/reconnect",
        ],
    },
    "submit_cancel": {"submit": 0, "cancel": 0},
    "runtime_code_changed_during_investigation": False,
}

# markdown
md = f"""# Phase674 — 2026-07-21 PM Summary Missing Root Cause

**Verdict:** `{report['verdict']}`  
**Primary cause:** `{report['primary_cause']}`  
**Secondary:** `{report['secondary_cause']}`  
**Confidence:** {report['confidence']}  
**Generated:** {NOW}

## Executive conclusion

PM did **not** fail inside Summary/Discord/W70 backup.  
It froze in the live PUSH asyncio event loop after `push_unexpected` / reconnect at **15:18:35**, identical to the AM freeze.  
Because `_maybe_am_pm_force_close` / session_end are not progressed during a blocking `iter_messages` recv wait, **15:23 force_close and PM Summary never started**.

## Required answers

1. Last normal PUSH: **2026-07-21T15:18:35+09:00** (msg 359445)
2. Last Heartbeat: **2026-07-21T15:13:49+09:00** (#30)
3. Last Capture: **2026-07-21T15:35:18+09:00** CAPTURE_COMPLETE (count 359445)
4. Runtime PID alive: **YES** (pilot 21836, daily 29024)
5. CPU/thread: pilot ~2465s CPU, ~602MB, 26 threads; MainThread idle in asyncio select (py-spy)
6. Event Loop Freeze proven: **YES** (py-spy dump)
7. Summary generation function called: **NO** (final path not reached; mid-run summary.json only)
8. Session Close started: **NO** (0 afternoon_session_close exits)
9. Orphans: **4 OPEN** (6058/5016/5985/3449); accepted 25 / observer_exit 21
10. WS final: push_unexpected no close frame + reconnect=1; no push after
11. AM vs PM: **SAME cause**
12. W70/W71 related: **NO**
13. stale trade 15s related: **NO**
14. last_completed_step: `{report['last_completed_step']}`
15. first_incomplete_step: `{report['first_incomplete_step']}`
16. Primary cause: **A_WEBSOCKET_RECV_FREEZE**
17. Confidence: **0.93**
18. Repro: WS unexpected close near entry_stop → reconnect → blocking recv → force_close skipped
19. Minimal fix: recv-timeout / watchdog to call force_close & stop; reconnect deadline
20. Runtime logic unchanged this investigation: **YES**
21. submit/cancel: **0/0**
22. Paths: evidence `{EV}` ; reports `{OUT}`

## Timeline (key)

"""
for row in report["timeline_after_1500"]:
    md += f"- `{row['t']}` {row['e']}\n"

md += f"""

## AM vs PM

| Item | AM (live_session_080044) | PM (live_session_124342) |
|------|--------------------------|--------------------------|
| Error | push_unexpected no close frame | identical |
| Reconnect | count=1 @ 11:20:58 | count=1 @ 15:18:35 |
| Last HB | 11:20:04 #30 | 15:13:49 #30 |
| Near | entry_stop 11:20 | entry_stop 15:18 |
| Force close | 11:25 never ran | 15:23 never ran |
| Summary final | missing | missing (mid-run json only) |

Relation: **A same failure** (WebSocket recv freeze after unexpected close/reconnect).

## Event Loop Freeze proof

py-spy dump PID 21836 MainThread:

```
_poll (asyncio/windows_events.py)
select (asyncio/windows_events.py)
run_forever → run_live_dry_run (pilot_runner.py:7952)
```

Discord worker idle on queue (not HTTP). Daily runner blocked on subprocess.communicate waiting for pilot.

## Summary path audit

1. Final summary/seal/Discord session summary: **not called**
2. Blocked before: stuck in `push.iter_messages` after reconnect
3. `small_paper_summary.json` present = periodic mid-run write at last HB (15:13:49), not PM close summary
4. No tmp/lock/finalize flags
5. Not waiting on orphan/active_positions for summary — close never started
6. W70 archive/D-sync after loop — unreachable

## W70/W71

**Unrelated.** Archive/external backup hooks run only after `asyncio.run(_loop())` returns; never reached. py-spy stacks show no SHA256/robocopy.

## Stale trade 15s

**Unrelated.** 3449 ENTRY warning at 15:03:45; market data continued until 15:18:35.

## Minimal fix (proposal only — not applied)

1. Wake periodically from recv (timeout) and call `_maybe_am_pm_force_close` / `_should_stop`
2. Reconnect deadline: if no PUSH within N seconds after reconnect, force `_request_stop('push_reconnect_timeout')` and finalize summary
3. Optional independent watchdog heartbeat/session-close timer

## Evidence index

- `{EV}/process/`
- `{EV}/logs/`
- `{EV}/stacks/pyspy_21836.txt`
- `{EV}/artifacts/live_session_124342/`
"""

(OUT / "report.md").write_text(md, encoding="utf-8")
(OUT / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

wb = Workbook()
bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")

def sheet(name: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
        c.fill = header_fill
    for r in rows:
        ws.append(r)
    return ws

wb.remove(wb.active)
sheet(
    "Timeline",
    ["time", "event"],
    [[x["t"], x["e"]] for x in report["timeline_after_1500"]],
)
sheet(
    "Process State",
    ["pid", "role", "state", "note"],
    [
        [21836, "pm_pilot", "ALIVE_FROZEN", "py-spy asyncio select"],
        [29024, "daily_runner", "ALIVE_WAIT", "subprocess.communicate"],
        ["-", "capture", "EXITED_COMPLETE", "CAPTURE_COMPLETE 15:35:18"],
    ],
)
sheet(
    "Heartbeat",
    ["event_time", "index", "push_messages", "note"],
    [
        ["2026-07-21T15:03:48+09:00", 28, 324750, ""],
        ["2026-07-21T15:08:48+09:00", 29, 336293, ""],
        ["2026-07-21T15:13:49+09:00", 30, 347896, "LAST"],
    ],
)
sheet(
    "Push",
    ["item", "value"],
    [
        ["last_event_time", "2026-07-21T15:18:35+09:00"],
        ["last_message_index", 359445],
        ["push_unexpected", "no close frame received or sent"],
        ["reconnect_count", 1],
        ["push_after_reconnect", False],
    ],
)
sheet(
    "Capture",
    ["item", "value"],
    [
        ["status", "CAPTURE_COMPLETE"],
        ["event_count", 359445],
        ["updated_at", "2026-07-21T15:35:18+09:00"],
        ["disconnect_count", 0],
    ],
)
sheet(
    "Session State",
    ["item", "value"],
    [
        ["session", "live_session_124342"],
        ["expected_force_close", "15:23"],
        ["session_close_started", False],
        ["orphan_open", 4],
        ["accepted", 25],
        ["observer_exit", 21],
        ["final_summary", False],
        ["midrun_summary_mtime", "15:13:49"],
        ["seal", False],
    ],
)
sheet(
    "Summary Path",
    ["step", "status"],
    [
        ["WS streaming", "OK until 15:18:35"],
        ["push_unexpected", "LOGGED"],
        ["reconnect", "LOGGED count=1"],
        ["iter_messages resume", "BLOCKED (select)"],
        ["force_close 15:23", "NOT REACHED"],
        ["final summary/seal/discord", "NOT CALLED"],
        ["W70 archive", "NOT CALLED"],
    ],
)
sheet(
    "AM PM Comparison",
    ["item", "AM", "PM"],
    [
        ["error", "no close frame @11:20:58", "no close frame @15:18:35"],
        ["reconnect", 1, 1],
        ["last_hb", "11:20:04 #30", "15:13:49 #30"],
        ["entry_stop", "11:20", "15:18"],
        ["force_close_ran", False, False],
        ["same_cause", True, True],
    ],
)
sheet(
    "Thread Stacks",
    ["pid", "thread", "stack_summary"],
    [
        [21836, "MainThread", "asyncio _poll/select run_forever pilot_runner:7952"],
        [21836, "discord-notify-worker", "queue.get wait"],
        [29024, "MainThread", "subprocess.communicate join run_pilot_session"],
    ],
)
sheet(
    "W70 W71 Audit",
    ["item", "value"],
    [
        ["related", False],
        ["session_archive_backup", None],
        ["session_external_backup", None],
        ["reason", "hooks after asyncio loop; loop never exited"],
    ],
)
sheet(
    "Root Cause Matrix",
    ["code", "role", "confidence", "support", "contra"],
    [
        [
            c["code"],
            c["role"],
            c["confidence"],
            "; ".join(c["support"])[:300],
            "; ".join(c["contra"])[:200],
        ]
        for c in report["cause_matrix"]
    ],
)
sheet(
    "Evidence Index",
    ["path", "note"],
    [
        [str(EV / "process"), "PID inventory"],
        [str(EV / "logs"), "tails errors/hb/discord/events"],
        [str(EV / "stacks/pyspy_21836.txt"), "freeze proof"],
        [str(EV / "artifacts/live_session_124342"), "copied session artifacts"],
        [str(OUT / "report.md"), "this report"],
    ],
)

wb.save(OUT / "audit.xlsx")
print(json.dumps({"verdict": report["verdict"], "out": str(OUT), "cause": report["primary_cause"]}, ensure_ascii=False))
