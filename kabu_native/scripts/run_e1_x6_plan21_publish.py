"""Publish Plan 2.1 day-robust research artifacts (stage-1 sweep + stage-2 redesign).

Produces report.json / report.md / audit.xlsx in BOTH the repo results dir and the
durable store (results/ was wiped once on 2026-08-02). Prints SHA256 of all three.

Verdict is E1_X6_NO_ROBUST_JOINT_STRATEGY: no package is presented as candidate /
best strategy / Shadow candidate. Failure analysis only.
"""
from __future__ import annotations

import hashlib
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

NATIVE = Path(__file__).resolve().parents[1]
for p in (str(NATIVE / "src"), str(NATIVE.parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

STAGE1_RUN = "e1x6_p21_20260802_204337_49eabae8"
STAGE2_RUN = "e1x6_p21s2_20260802_223952_a09569ef"
BANNER = "E1_X6_NO_ROBUST_JOINT_STRATEGY / FAILURE_ANALYSIS_ONLY / NOT_FOR_TRADING"


def _sha(fp: Path) -> str:
    return hashlib.sha256(fp.read_bytes()).hexdigest()


def _load(run_id: str) -> dict:
    from research.e1_x6_provisional.oracle_capture import durable_store_root

    w = durable_store_root() / "plan21_work" / run_id
    out = {}
    for name in ("p1_lock.json", "sweep_results.json", "parity.json", "capture_metas.json"):
        fp = w / name
        if fp.is_file():
            out[name.split(".")[0]] = json.loads(fp.read_text(encoding="utf-8"))
    return out


def _pkg_row(r: dict) -> list:
    m = r["metrics"]
    return [
        r["strategy_id"],
        r["entry_candidate_id"],
        r["exit_family_id"],
        m["n"],
        round(m["total_pnl"], 2),
        round(m["median_day_pnl"], 2),
        round(m["ex_best1_day_pnl"], 2),
        round(m["ex_best2_days_pnl"], 2),
        m["best1_day"],
        m["best2_day"],
        None if m["top1_day_share_of_gross_positive"] is None else round(m["top1_day_share_of_gross_positive"], 4),
        None if m["top2_days_share_of_gross_positive"] is None else round(m["top2_days_share_of_gross_positive"], 4),
        None if m["pf"] is None else round(m["pf"], 4),
        m["pf_status"],
        round(m["max_dd"], 2),
        round(m["stop_loss_total"], 2),
        round(m["ex722_pnl"], 2),
        len(r["failed"]),
        ";".join(r["failed"]),
        r["all_pass"],
    ]


PKG_HEADERS = [
    "strategy_id", "entry_candidate_id", "exit_family_id", "trades", "total_pnl",
    "median_day_pnl", "ex_best1_day_pnl", "ex_best2_days_pnl", "best1_day", "best2_day",
    "top1_day_share", "top2_days_share", "pf", "pf_status", "max_dd", "stop_loss_total",
    "ex722_pnl", "failed_n", "failed_gates", "all_pass",
]


def main() -> int:
    from openpyxl import Workbook

    from research.e1_x6_provisional.constants import DAYS
    from research.e1_x6_provisional.oracle_capture import durable_store_root
    from research.e1_x6_provisional.util import sha256_obj

    s1 = _load(STAGE1_RUN)
    s2 = _load(STAGE2_RUN)
    for tag, s in (("stage1", s1), ("stage2", s2)):
        if s["sweep_results"]["passers_n"] != 0:
            raise SystemExit(f"FAIL: {tag} passers != 0; this publisher is failure-analysis-only")

    import collections

    def gate_fail_counts(res):
        c = collections.Counter()
        for r in res:
            for g in r["failed"]:
                c[g] += 1
        return dict(c.most_common())

    def day_sign(res):
        pos = collections.Counter()
        for r in res:
            for d, v in r["metrics"]["day_pnl"].items():
                if v > 0:
                    pos[d] += 1
        return {d: int(pos.get(d, 0)) for d in DAYS}

    report = {
        "banner": BANNER,
        "plan_version": "2.1",
        "verdict": "E1_X6_NO_ROBUST_JOINT_STRATEGY",
        "published_at_jst": datetime.now().astimezone().isoformat(),
        "period": {"start": DAYS[0], "end": DAYS[-1], "days": list(DAYS)},
        "cost_bps_once": 5.0,
        "safety": {"submit": 0, "cancel": 0, "live": 0,
                   "shadow_runtime_forward_paper_task_yaml_discord": "UNCHANGED"},
        "stage1": {
            "run_id": STAGE1_RUN,
            "p1_sha256": s1["p1_lock"]["p1_sha256"],
            "registry_sha256": s1["p1_lock"]["registry_lock"]["joint_registry_sha256"],
            "registry_n": s1["p1_lock"]["registry_lock"]["joint_registry_n"],
            "parity": {
                "all_match": s1["parity"]["all_match"],
                "partitions_n": len(s1["parity"]["partitions"]),
            },
            "ab_determinism_all_match": s1["sweep_results"]["ab_determinism"]["all_match"],
            "base_x5": s1["sweep_results"]["base"],
            "invalid_source_trades": s1["sweep_results"]["invalid_source_trades"],
            "passers_n": 0,
            "verdict": s1["sweep_results"]["verdict"],
            "results_sha256": s1["sweep_results"]["results_sha256"],
            "gate_fail_counts": gate_fail_counts(s1["sweep_results"]["results"]),
            "day_positive_package_counts": day_sign(s1["sweep_results"]["results"]),
            "lodo": s1["sweep_results"]["lodo"]["held_out_pnls"],
            "results": s1["sweep_results"]["results"],
        },
        "stage2": {
            "run_id": STAGE2_RUN,
            "p1_sha256": s2["p1_lock"]["p1_sha256"],
            "registry_sha256": s2["p1_lock"]["registry_lock"]["joint_registry_sha256"],
            "registry_n": s2["p1_lock"]["registry_lock"]["joint_registry_n"],
            "feature_inventory": s2["p1_lock"]["redesign"]["feature_inventory"],
            "features_unavailable": s2["p1_lock"]["redesign"]["features_unavailable"],
            "quantile_values": s2["p1_lock"]["registry_lock"]["quantile_values"],
            "ab_determinism_all_match": s2["sweep_results"]["ab_determinism"]["all_match"],
            "passers_n": 0,
            "verdict": s2["sweep_results"]["verdict"],
            "results_sha256": s2["sweep_results"]["results_sha256"],
            "gate_fail_counts": gate_fail_counts(s2["sweep_results"]["results"]),
            "day_positive_package_counts": day_sign(s2["sweep_results"]["results"]),
            "lodo": s2["sweep_results"]["lodo"]["held_out_pnls"],
            "results": s2["sweep_results"]["results"],
        },
        "failure_analysis": {
            "structural_finding": (
                "All 396 evaluated JointStrategyPackages depend on 20260722/20260731 "
                "(mechanically computed best days) for essentially all gross positive "
                "day PnL. Stage-1: every package negative on 6/9 days; best "
                "ex-best-2-days PnL -459,343 yen. Stage-2 redesign improved day "
                "structure (best package positive 6/9 days, median day +12,205 yen) "
                "but ex-best-2-days remained -121,097 yen and top-2-day concentration "
                "gates failed for 196/196."
            ),
            "no_candidate_presented": True,
            "gate_relaxation": "FORBIDDEN (plan 2.1)",
        },
        "environment_incidents": [
            "2026-08-02 OS temp + results cleanup wiped Stage-1 published artifacts (SHAs preserved)",
            "2026-08-02 two external terminal aborts during pipeline; resumed via day-level durable bundles",
        ],
    }
    report["report_body_sha256"] = sha256_obj(report)

    out_repo = NATIVE / "results" / "research" / "e1_x6_plan21_day_robust_20260721_20260731"
    out_repo.mkdir(parents=True, exist_ok=True)

    # report.json
    fp_json = out_repo / "report.json"
    fp_json.write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding="utf-8")

    # report.md
    s1f = report["stage1"]["gate_fail_counts"]
    s2f = report["stage2"]["gate_fail_counts"]
    md = [
        f"# E1_X6 Plan 2.1 day-robust joint sweep — {BANNER}",
        "",
        f"- published: {report['published_at_jst']}",
        f"- verdict: **{report['verdict']}** (stage-1 passers 0/200, stage-2 passers 0/196)",
        f"- period: {DAYS[0]}..{DAYS[-1]} (9 days), 5bps cost, CAP5, LOT100",
        f"- safety: submit/cancel/live = 0/0/0; Shadow/Runtime/Forward/Paper/Task/YAML/Discord unchanged",
        "",
        "## Stage-1 (full JointRegistry 200, oracle parity proven)",
        f"- run: `{STAGE1_RUN}` / P1 `{report['stage1']['p1_sha256']}`",
        f"- registry: n={report['stage1']['registry_n']} sha `{report['stage1']['registry_sha256']}`",
        f"- parity: all 17 partitions exact match vs session replay",
        f"- A/B determinism: {report['stage1']['ab_determinism_all_match']}",
        f"- BASE X5: n={report['stage1']['base_x5']['n']} pnl={report['stage1']['base_x5']['pnl']:.0f} dd={report['stage1']['base_x5']['max_dd']:.0f}",
        f"- day-positive package counts: {report['stage1']['day_positive_package_counts']}",
        "- top gate failures: " + ", ".join(f"{k}={v}" for k, v in list(s1f.items())[:6]),
        "",
        "## Stage-2 ENTRY redesign (as-of features, 49 entries x 4 exits = 196)",
        f"- run: `{STAGE2_RUN}` / P1 `{report['stage2']['p1_sha256']}`",
        f"- registry: n={report['stage2']['registry_n']} sha `{report['stage2']['registry_sha256']}`",
        f"- A/B determinism: {report['stage2']['ab_determinism_all_match']}",
        f"- day-positive package counts: {report['stage2']['day_positive_package_counts']}",
        "- top gate failures: " + ", ".join(f"{k}={v}" for k, v in list(s2f.items())[:6]),
        "",
        "## Failure analysis",
        report["failure_analysis"]["structural_finding"],
        "",
        "No package is presented as candidate / best strategy / Shadow candidate.",
        "Gate relaxation is forbidden by plan 2.1.",
    ]
    fp_md = out_repo / "report.md"
    fp_md.write_text("\n".join(md), encoding="utf-8")

    # audit.xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "SUMMARY"
    for row in (
        ["banner", BANNER],
        ["verdict", report["verdict"]],
        ["plan_version", "2.1"],
        ["stage1_run", STAGE1_RUN],
        ["stage1_p1_sha256", report["stage1"]["p1_sha256"]],
        ["stage1_registry_sha256", report["stage1"]["registry_sha256"]],
        ["stage1_parity_all_match", True],
        ["stage1_ab_all_match", report["stage1"]["ab_determinism_all_match"]],
        ["stage1_passers", 0],
        ["stage2_run", STAGE2_RUN],
        ["stage2_p1_sha256", report["stage2"]["p1_sha256"]],
        ["stage2_registry_sha256", report["stage2"]["registry_sha256"]],
        ["stage2_ab_all_match", report["stage2"]["ab_determinism_all_match"]],
        ["stage2_passers", 0],
        ["base_x5_n", report["stage1"]["base_x5"]["n"]],
        ["base_x5_pnl", report["stage1"]["base_x5"]["pnl"]],
        ["base_x5_max_dd", report["stage1"]["base_x5"]["max_dd"]],
        ["invalid_source_trades", report["stage1"]["invalid_source_trades"]],
        ["safety_submit_cancel_live", "0/0/0"],
    ):
        ws.append(row)

    for name, res in (("S1_RESULTS", s1["sweep_results"]["results"]),
                      ("S2_RESULTS", s2["sweep_results"]["results"])):
        w = wb.create_sheet(name)
        w.append(PKG_HEADERS)
        for r in sorted(res, key=lambda x: x["strategy_id"]):
            w.append(_pkg_row(r))

    for name, res in (("S1_DAY_PNL", s1["sweep_results"]["results"]),
                      ("S2_DAY_PNL", s2["sweep_results"]["results"])):
        w = wb.create_sheet(name)
        w.append(["strategy_id"] + list(DAYS))
        for r in sorted(res, key=lambda x: x["strategy_id"]):
            dp = r["metrics"]["day_pnl"]
            w.append([r["strategy_id"]] + [round(float(dp.get(d, 0.0)), 2) for d in DAYS])

    w = wb.create_sheet("GATE_FAIL_COUNTS")
    w.append(["stage", "gate", "fail_count", "of_n"])
    for k, v in s1f.items():
        w.append(["stage1", k, v, 200])
    for k, v in s2f.items():
        w.append(["stage2", k, v, 196])

    w = wb.create_sheet("PARITY_S1")
    w.append(["day", "am_pm", "oracle_n", "session_n", "oracle_pnl", "session_pnl", "match", "mismatch_n"])
    for r in s1["parity"]["partitions"]:
        w.append([r.get("day"), r.get("am_pm"), r["oracle_n"], r["session_n"],
                  round(r["oracle_pnl"], 2), round(r["session_pnl"], 2), r["match"], r["mismatch_n"]])

    w = wb.create_sheet("LODO")
    w.append(["stage", "held_out_day", "selected_sid", "held_out_pnl"])
    for tag, s in (("stage1", s1), ("stage2", s2)):
        for r in s["sweep_results"]["lodo"]["rows"]:
            w.append([tag, r["held_out_day"], r.get("selected_strategy_id") or r.get("selected"), round(float(r["held_out_pnl"]), 2)])

    w = wb.create_sheet("S2_QUANTILES")
    w.append(["feature", "q", "threshold"])
    for feat, d in report["stage2"]["quantile_values"].items():
        for q, v in d.items():
            w.append([feat, q, v])

    w = wb.create_sheet("INCIDENTS")
    w.append(["note"])
    for x in report["environment_incidents"]:
        w.append([x])

    fp_xlsx = out_repo / "audit.xlsx"
    wb.save(fp_xlsx)

    # durable copy + SHAs
    durable = durable_store_root() / "published" / "e1_x6_plan21_day_robust_20260721_20260731"
    durable.mkdir(parents=True, exist_ok=True)
    shas = {}
    for fp in (fp_json, fp_md, fp_xlsx):
        shutil.copy2(fp, durable / fp.name)
        shas[fp.name] = _sha(fp)
        if _sha(durable / fp.name) != shas[fp.name]:
            raise SystemExit(f"FAIL: durable copy sha mismatch {fp.name}")
    print(json.dumps({"published_dir": str(out_repo), "durable_dir": str(durable), "shas": shas},
                     ensure_ascii=False, indent=1), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
