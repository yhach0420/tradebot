#!/usr/bin/env python
"""P1: CURRENT_RUNTIME_REPLAY over all saved Captures. Fast path only.

Does not start Dynamic Anchor / new strategy / threshold work.
Does not rewrite Actual Paper. submit/cancel/live=0/0/0.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

for _k in (
    "KABU_V1R_ENTRY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_NOTIFY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_CAP_BLOCKED_WEBHOOK_URL",
    "KABU_DISCORD_RESEARCH_WEBHOOK_URL",
    "KABU_SHADOW_DISCORD_WEBHOOK_URL",
    "KABU_DISCORD_MARKET_CAPTURE_WEBHOOK_URL",
    "KABU_MARKET_CAPTURE_WEBHOOK_URL",
):
    os.environ.pop(_k, None)
os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"

from _p1_inventory import classify, resolve_universe  # noqa: E402
from research.anchor_vs_event_driven.run_comparison import (  # noqa: E402
    _boot,
    _load_json,
    _stream_day,
    extract_trades,
)
from run_p0_3_exact_runtime_replay_20260820 import (  # noqa: E402
    _anchor_from_fill_t,
    _iso,
    _ledger_sha,
    _maxdd,
    _pf,
    _sess_stats,
)
from run_p0_4_exact_vs_fast_parity import CollectorEngine, _Discard  # noqa: E402
from small_paper.v1r_exit_v2_activation_gate import STRATEGY_SHA  # noqa: E402
from small_paper.v1r_exit_v2_contract import EXIT_V2_CANDIDATE_SHA  # noqa: E402
from small_paper.v1r_live_dual_lane import canonical_symbol_key  # noqa: E402
from small_paper.v1r_native_entry_live import ANCHOR_SHA, ENTRY_SHA  # noqa: E402
from small_paper.v1r_primary_runtime import ANCHOR_SHA as ANCHOR_SHA_RT  # noqa: E402

OUT = ROOT / "results" / "research" / "current_runtime_full_capture_recalc_p1"
CACHE = OUT / "_day_cache"
OLD_A = ROOT / "results" / "research" / "anchor_vs_event_driven_v1r" / "report.json"
INV = OUT / "_inventory.json"

ACTUAL_PAPER = {
    "20260820": {
        "label": "ACTUAL_PAPER_WITH_RUNTIME_DEFECT",
        "trades": 5,
        "pnl": -32600.0,
    }
}


def _file_sha(rel: str) -> str:
    p = ROOT / rel
    if not p.is_file():
        return ""
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _pf_out(v: Any) -> Any:
    if v is None:
        return None
    if v == float("inf"):
        return "Infinity"
    return v


def _wl(trades: list[dict[str, Any]]) -> tuple[int, int, int]:
    w = l = d = 0
    for t in trades:
        p = float(t.get("pnl_yen_100") or 0.0)
        if p > 1e-9:
            w += 1
        elif p < -1e-9:
            l += 1
        else:
            d += 1
    return w, l, d


def _gross(trades: list[dict[str, Any]]) -> tuple[float, float]:
    gp = sum(float(t.get("pnl_yen_100") or 0.0) for t in trades if float(t.get("pnl_yen_100") or 0.0) > 0)
    gl = sum(-float(t.get("pnl_yen_100") or 0.0) for t in trades if float(t.get("pnl_yen_100") or 0.0) < 0)
    return round(gp, 2), round(gl, 2)


def replay_one(payload: dict[str, Any]) -> dict[str, Any]:
    """P0-4 Fast path. Independent engine/dual state per day."""
    for _k in (
        "KABU_V1R_ENTRY_WEBHOOK_URL",
        "KABU_SMALL_PAPER_NOTIFY_WEBHOOK_URL",
        "KABU_SMALL_PAPER_CAP_BLOCKED_WEBHOOK_URL",
        "KABU_DISCORD_RESEARCH_WEBHOOK_URL",
        "KABU_SHADOW_DISCORD_WEBHOOK_URL",
        "KABU_DISCORD_MARKET_CAPTURE_WEBHOOK_URL",
        "KABU_MARKET_CAPTURE_WEBHOOK_URL",
    ):
        os.environ.pop(_k, None)
    os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"
    if str(ROOT / "src") not in sys.path:
        sys.path.insert(0, str(ROOT / "src"))
        sys.path.insert(0, str(ROOT / "scripts"))

    day = str(payload["date"])
    capture = Path(payload["capture_path"])
    universe = list(payload["universe"])
    uni_src = str(payload["universe_source"])
    t0 = time.perf_counter()
    try:
        eng, dual = _boot(universe, CollectorEngine)
        if dual is None or not eng.ready:
            return {
                "ok": False,
                "date": day,
                "blocker": getattr(eng, "fail_reason", "dual_unavailable"),
                "elapsed_sec": round(time.perf_counter() - t0, 3),
            }
        eng.notify_enabled = False
        eng.ingest_audit = _Discard()  # type: ignore[assignment]
        events_n, last_et = _stream_day(day, capture, eng, dual)
        eng._harvest(eng.events)
        raw_trades = extract_trades(dual)
        trades: list[dict[str, Any]] = []
        for i, tr in enumerate(raw_trades, start=1):
            fill_t = float(tr.get("entry_time") or 0.0)
            exit_t = tr.get("exit_time")
            an = _anchor_from_fill_t(day, fill_t)
            sym = canonical_symbol_key(tr.get("symbol"))
            snap = eng.snapshots.get((an, sym), {})
            fill_row = next((f for f in eng.a_fills if f.get("symbol") == sym and f.get("anchor") == an), None)
            admit_row = next((a for a in eng.a_admits if a.get("symbol") == sym and a.get("anchor") == an), None)
            src = fill_row or admit_row or snap
            holding = None
            if fill_t and exit_t is not None:
                holding = round(float(exit_t) - fill_t, 3)
            trades.append(
                {
                    "date": day,
                    "session": tr.get("session"),
                    "trade_id": f"{day}|{tr.get('session')}|{an}|{sym}|{i}",
                    "symbol": sym,
                    "anchor_time": an,
                    "snapshot_sequence": snap.get("snapshot_sequence"),
                    "score": src.get("score") if src else None,
                    "candidate_rank": (admit_row or snap).get("rank") if (admit_row or snap) else None,
                    "limit": (fill_row or admit_row or {}).get("limit") or tr.get("entry_price"),
                    "fill_time": fill_t,
                    "fill_time_iso": _iso(fill_t),
                    "fill_price": tr.get("entry_price"),
                    "exit_time": exit_t,
                    "exit_time_iso": _iso(exit_t),
                    "exit_price": tr.get("exit_price"),
                    "exit_reason": tr.get("reason"),
                    "pnl_yen_100": float(tr.get("pnl_yen_100") or 0.0),
                    "holding_sec": holding,
                }
            )
        pnls = [float(t.get("pnl_yen_100") or 0.0) for t in trades]
        w, l, d = _wl(trades)
        gp, gl = _gross(trades)
        ledger_sum = round(sum(pnls), 2)
        return {
            "ok": True,
            "result_name": "CURRENT_RUNTIME_REPLAY",
            "date": day,
            "universe_n": len(universe),
            "universe_source": uni_src,
            "events_processed": events_n,
            "raw_sequence_holes": int(getattr(eng, "native_ingest_raw_sequence_holes", 0) or 0),
            "accepted_sequence_holes": int(eng.native_ingest_sequence_holes),
            "skip_universe": int(eng.native_ingest_skip_universe),
            "skip_duplicate": int(eng.native_ingest_skip_duplicate),
            "anchor_fires": int(eng.anchor_fires),
            "candidates_scored": len(eng.a_candidates),
            "admitted": int(eng.primary_admitted),
            "expired": int(eng.primary_expired),
            "fills": int(eng.primary_fills),
            "cap_blocked": int(eng.cap_blocked),
            "same_symbol_blocked": int(eng.same_symbol_blocked),
            "trades": trades,
            "trade_n": len(trades),
            "win": w,
            "loss": l,
            "draw": d,
            "pnl": ledger_sum,
            "gross_profit": gp,
            "gross_loss": gl,
            "PF": _pf(pnls),
            "avg_pnl": round(ledger_sum / len(trades), 4) if trades else 0.0,
            "maxDD": _maxdd(trades),
            "AM": _sess_stats(trades, "AM"),
            "PM": _sess_stats(trades, "PM"),
            "ledger_sha": _ledger_sha(trades),
            "ledger_pnl_sum": ledger_sum,
            "ledger_row_n": len(trades),
            "session_close_completed": True,
            "last_et": last_et,
            "elapsed_sec": round(time.perf_counter() - t0, 3),
        }
    except Exception as exc:
        return {
            "ok": False,
            "date": day,
            "blocker": f"{type(exc).__name__}:{exc}",
            "elapsed_sec": round(time.perf_counter() - t0, 3),
        }


def reclassify(inv_row: dict[str, Any]) -> dict[str, Any]:
    day = inv_row["date"]
    cap = Path(inv_row["capture_path"]) if inv_row.get("capture_path") else None
    uni = resolve_universe(day, cap)
    seq = {
        "first_seq": inv_row.get("first_seq"),
        "last_seq": inv_row.get("last_seq"),
        "line_count": inv_row.get("event_count"),
        "first_event": inv_row.get("first_event"),
        "last_event": inv_row.get("last_event"),
        "contiguous_hint": inv_row.get("sequence_continuity_hint"),
        "size_bytes": inv_row.get("size_bytes"),
    }
    klass = classify(day, cap, uni, seq)
    row = {**inv_row, **klass}
    row["universe_resolved"] = uni["resolved"]
    row["universe_source"] = uni.get("source")
    row["universe_n"] = uni.get("universe_n")
    row["universe_reason"] = uni.get("reason")
    row["universe_notes"] = uni.get("notes")
    row["universe_symbols"] = uni.get("symbols") or []
    n = int(row.get("event_count") or 0)
    last = row.get("last_seq")
    first = row.get("first_seq")
    if n and last and first == 1 and n != int(last):
        row["sequence_continuity"] = "FAIL"
        if row["capture_class"] == "FULL":
            row["capture_class"] = "DEGRADED"
            row["full"] = False
            row["exclusion_reason"] = (row.get("exclusion_reason") or "") + ";SEQ_COUNT_NE_LAST"
        else:
            extra = ";SEQ_COUNT_NE_LAST"
            if extra not in str(row.get("exclusion_reason") or ""):
                row["exclusion_reason"] = str(row.get("exclusion_reason") or "") + extra
    else:
        row["sequence_continuity"] = "PASS" if row.get("sequence_continuity_hint") else ("FAIL" if n else "N/A")
    row["replay_eligible"] = bool(
        uni["resolved"]
        and klass.get("usable")
        and row.get("capture_class") in {"FULL", "PARTIAL", "DEGRADED"}
        and cap is not None
    )
    return row


def _day_stats(trades: list[dict[str, Any]]) -> dict[str, Any]:
    w, l, d = _wl(trades)
    gp, gl = _gross(trades)
    pnl = round(sum(float(t.get("pnl_yen_100") or 0.0) for t in trades), 2)
    return {
        "trades": len(trades),
        "win": w,
        "loss": l,
        "draw": d,
        "gross_profit": gp,
        "gross_loss": gl,
        "pnl": pnl,
        "PF": _pf_out(_pf([float(t.get("pnl_yen_100") or 0.0) for t in trades])),
        "avg_pnl": round(pnl / len(trades), 4) if trades else 0.0,
        "maxDD": _maxdd(trades),
        "AM": _sess_stats(trades, "AM"),
        "PM": _sess_stats(trades, "PM"),
    }


def classify_old_delta(old: dict[str, Any], new: dict[str, Any], inv: dict[str, Any]) -> str:
    if not old:
        return "NO_OLD_A_FIXED"
    ot, nt = int(old.get("trades") or 0), int(new.get("trade_n") or 0)
    op, np_ = float(old.get("pnl") or 0.0), float(new.get("pnl") or 0.0)
    if ot == nt and abs(op - np_) < 1e-6:
        return "MATCH"
    notes = []
    old_src = str(inv.get("old_universe_source") or "")
    new_src = str(new.get("universe_source") or "")
    if old_src and new_src and old_src != new_src:
        notes.append("INPUT_BINDING_CHANGE")
    if abs(op - np_) >= 1e-6:
        notes.append("CURRENT_RUNTIME_SEMANTICS")
    if not notes:
        notes.append("UNKNOWN")
    return "+".join(notes)


def _load_old_a() -> tuple[dict[str, Any], dict[str, str]]:
    old_body = _load_json(OLD_A) if OLD_A.is_file() else {}
    old_by: dict[str, Any] = {}
    old_src: dict[str, str] = {}
    for key in ("daily", "day_rows", "daily_results", "per_day", "results"):
        for r in old_body.get(key) or []:
            if isinstance(r, dict) and r.get("date") and r.get("A"):
                old_by[str(r["date"])] = r["A"]
    for r in old_body.get("inventory") or []:
        if r.get("date"):
            old_src[str(r["date"])] = str(r.get("universe_source") or "")
    return old_by, old_src


def write_outputs(inv_rows: list[dict[str, Any]], replays: dict[str, Any], identity: dict[str, Any]) -> dict[str, Any]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    old_by, old_src = _load_old_a()
    inv_map = {r["date"]: r for r in inv_rows}
    for r in inv_rows:
        r["old_universe_source"] = old_src.get(r["date"], "")
    daily = []
    all_trades: list[dict[str, Any]] = []
    primary_trades: list[dict[str, Any]] = []
    ref_trades: list[dict[str, Any]] = []
    failed = []
    raw_gap_days = []
    for row in inv_rows:
        day = row["date"]
        rep = replays.get(day)
        actual = ACTUAL_PAPER.get(day)
        q_flags: list[str] = []
        replay_status = "SKIPPED"
        if not row.get("universe_resolved") and row.get("capture_class") != "MISSING":
            q_flags.append("UNIVERSE_BINDING_UNRESOLVED")
        if not row.get("replay_eligible"):
            replay_status = "EXCLUDED"
        elif rep is None:
            replay_status = "MISSING_RESULT"
            failed.append(day)
        elif not rep.get("ok"):
            replay_status = "FAIL"
            failed.append(day)
            q_flags.append(str(rep.get("blocker") or "REPLAY_FAIL"))
        else:
            replay_status = "OK"
            if int(rep.get("ledger_row_n") or 0) != int(rep.get("trade_n") or 0):
                replay_status = "FAIL"
                q_flags.append("LEDGER_COUNT_MISMATCH")
                failed.append(day)
            if abs(float(rep.get("pnl") or 0) - float(rep.get("ledger_pnl_sum") or 0)) > 1e-6:
                replay_status = "FAIL"
                q_flags.append("PNL_SUM_MISMATCH")
                failed.append(day)
            if int(rep.get("raw_sequence_holes") or 0) > 0:
                q_flags.append("RAW_SEQUENCE_GAP")
                raw_gap_days.append(day)
            trades = list(rep.get("trades") or [])
            if replay_status == "OK":
                all_trades.extend(trades)
                usable_class = row["capture_class"] in {"FULL", "PARTIAL", "DEGRADED"}
                primary_ok = (
                    row["capture_class"] == "FULL"
                    and "RAW_SEQUENCE_GAP" not in q_flags
                    and row.get("jpx_trading_day")
                )
                if primary_ok:
                    primary_trades.extend(trades)
                if usable_class:
                    ref_trades.extend(trades)
        old = old_by.get(day) or {}
        new_pnl = (rep or {}).get("pnl") if (rep or {}).get("ok") else None
        rec = {
            "date": day,
            "capture_class": row.get("capture_class"),
            "jpx_trading_day": row.get("jpx_trading_day"),
            "replay_eligible": row.get("replay_eligible"),
            "replay_status": replay_status,
            "QUALITY_FLAG": q_flags or (["OK"] if replay_status == "OK" else [replay_status]),
            "universe_source": row.get("universe_source"),
            "universe_n": row.get("universe_n"),
            "events_capture": row.get("event_count"),
            "events_processed": (rep or {}).get("events_processed"),
            "raw_sequence_holes": (rep or {}).get("raw_sequence_holes"),
            "accepted_sequence_holes": (rep or {}).get("accepted_sequence_holes"),
            "skip_universe": (rep or {}).get("skip_universe"),
            "anchor_fires": (rep or {}).get("anchor_fires"),
            "candidates_scored": (rep or {}).get("candidates_scored"),
            "admitted": (rep or {}).get("admitted"),
            "expired": (rep or {}).get("expired"),
            "fills": (rep or {}).get("fills"),
            "cap_blocked": (rep or {}).get("cap_blocked"),
            "same_symbol_blocked": (rep or {}).get("same_symbol_blocked"),
            "trades": (rep or {}).get("trade_n") if (rep or {}).get("ok") else None,
            "win": (rep or {}).get("win"),
            "loss": (rep or {}).get("loss"),
            "draw": (rep or {}).get("draw"),
            "pnl": new_pnl,
            "PF": _pf_out((rep or {}).get("PF")),
            "avg_pnl": (rep or {}).get("avg_pnl"),
            "maxDD": (rep or {}).get("maxDD"),
            "AM": (rep or {}).get("AM"),
            "PM": (rep or {}).get("PM"),
            "ledger_sha": (rep or {}).get("ledger_sha"),
            "actual_label": (actual or {}).get("label") if actual else "N/A",
            "actual_trades": (actual or {}).get("trades") if actual else "N/A",
            "actual_pnl": (actual or {}).get("pnl") if actual else "N/A",
            "old_trades": old.get("trades") if old else None,
            "old_pnl": old.get("pnl") if old else None,
            "elapsed_sec": (rep or {}).get("elapsed_sec"),
            "blocker": (rep or {}).get("blocker"),
        }
        daily.append(rec)

    primary = _day_stats(primary_trades)
    reference = _day_stats(ref_trades)
    primary_days = [
        r["date"]
        for r in daily
        if r["capture_class"] == "FULL"
        and r["replay_status"] == "OK"
        and "RAW_SEQUENCE_GAP" not in (r.get("QUALITY_FLAG") or [])
        and r.get("jpx_trading_day")
    ]
    ref_days = [
        r["date"]
        for r in daily
        if r["replay_status"] == "OK" and r["capture_class"] in {"FULL", "PARTIAL", "DEGRADED"}
    ]

    by_pnl = [(r["date"], float(r["pnl"])) for r in daily if r["replay_status"] == "OK" and r["date"] in primary_days]
    best = max(by_pnl, key=lambda x: x[1]) if by_pnl else ("", 0.0)
    worst = min(by_pnl, key=lambda x: x[1]) if by_pnl else ("", 0.0)
    tot = primary["pnl"] or 0.0
    ordered = sorted(by_pnl, key=lambda x: -x[1])
    top1_c = (ordered[0][1] / tot) if ordered and tot else None
    top3_c = (sum(x[1] for x in ordered[:3]) / tot) if ordered and tot else None
    pos = sum(1 for _, p in by_pnl if p > 0)
    neg = sum(1 for _, p in by_pnl if p < 0)
    flat = sum(1 for _, p in by_pnl if p == 0)

    old_cmp = []
    match_n = 0
    changed_n = 0
    for r in daily:
        old = old_by.get(r["date"]) or {}
        if not old or r["replay_status"] != "OK":
            continue
        delta_t = int(r["trades"] or 0) - int(old.get("trades") or 0)
        delta_p = float(r["pnl"] or 0) - float(old.get("pnl") or 0)
        klass = classify_old_delta(old, replays[r["date"]], inv_map[r["date"]])
        if klass == "MATCH":
            match_n += 1
        else:
            changed_n += 1
        old_cmp.append(
            {
                "date": r["date"],
                "old_trades": old.get("trades"),
                "new_trades": r["trades"],
                "old_pnl": old.get("pnl"),
                "new_pnl": r["pnl"],
                "trade_delta": delta_t,
                "pnl_delta": delta_p,
                "class": klass,
            }
        )

    classes: dict[str, int] = {}
    for r in inv_rows:
        classes[r["capture_class"]] = classes.get(r["capture_class"], 0) + 1
    unresolved_all = [r["date"] for r in inv_rows if not r.get("universe_resolved")]
    completed = [r["date"] for r in daily if r["replay_status"] == "OK"]
    eligible = [r["date"] for r in daily if r["replay_eligible"]]
    dates = [r["date"] for r in inv_rows]
    a20 = next((r for r in daily if r["date"] == "20260820"), {})

    verdict = "P1_CURRENT_RUNTIME_FULL_CAPTURE_RECALC_COMPLETE"
    if failed or len(completed) < len(eligible):
        verdict = "P1_CURRENT_RUNTIME_FULL_CAPTURE_RECALC_PARTIAL"
    if not completed:
        verdict = "P1_BLOCKED"

    report = {
        "task": "P1",
        "result_name": "CURRENT_RUNTIME_REPLAY",
        "PERIOD": f"{dates[0]} - {dates[-1]}" if dates else "",
        "INVENTORY": {
            "total_days": len(inv_rows),
            "full": classes.get("FULL", 0),
            "partial": classes.get("PARTIAL", 0),
            "degraded": classes.get("DEGRADED", 0),
            "invalid": classes.get("INVALID", 0),
            "missing": classes.get("MISSING", 0),
            "replay_eligible": len(eligible),
            "replay_completed": len(completed),
            "replay_failed": failed,
        },
        "PRIMARY_FULL": {**primary, "days": len(primary_days), "day_list": primary_days},
        "REFERENCE_ALL_USABLE": {**reference, "days": len(ref_days), "day_list": ref_days},
        "BEST_DAY": {"date": best[0], "pnl": best[1]},
        "WORST_DAY": {"date": worst[0], "pnl": worst[1]},
        "TOP1_PNL_CONTRIBUTION": top1_c,
        "TOP3_PNL_CONTRIBUTION": top3_c,
        "day_sign": {"positive_days": pos, "negative_days": neg, "flat_days": flat},
        "OLD_A_FIXED_COMPARISON": {
            "days_compared": len(old_cmp),
            "exact_match_days": match_n,
            "changed_days": changed_n,
            "rows": old_cmp,
        },
        "20260820": {
            "actual": ACTUAL_PAPER["20260820"],
            "replay": {
                "trades": a20.get("trades"),
                "pnl": a20.get("pnl"),
                "label": "CURRENT_RUNTIME_REPLAY",
            },
        },
        "RAW_SEQUENCE_GAP_DAYS": raw_gap_days,
        "UNIVERSE_BINDING_UNRESOLVED": unresolved_all,
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "identity": identity,
        "daily": daily,
        "trades": all_trades,
        "inventory": [{k: v for k, v in r.items() if k != "universe_symbols"} for r in inv_rows],
        "verdict": verdict,
    }

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

    md = f"""# P1 CURRENT_RUNTIME_REPLAY

Capture全期間に対する現行Runtime Fast Event-Time Replay。Actual Paperではない。
Dynamic Anchor / 新戦略 / threshold には進まない。submit/cancel/live=0/0/0。

PERIOD: {report['PERIOD']}

## INVENTORY
total_days: {len(inv_rows)}
full: {classes.get('FULL', 0)}
partial: {classes.get('PARTIAL', 0)}
degraded: {classes.get('DEGRADED', 0)}
invalid: {classes.get('INVALID', 0)}
missing: {classes.get('MISSING', 0)}
replay_eligible: {len(eligible)}
replay_completed: {len(completed)}
replay_failed: {failed}

## PRIMARY_FULL
days: {len(primary_days)}
trades: {primary['trades']}
win: {primary['win']}
loss: {primary['loss']}
draw: {primary['draw']}
pnl: {primary['pnl']}
PF: {primary['PF']}
avg_pnl: {primary['avg_pnl']}
maxDD: {primary['maxDD']}
AM: {primary['AM']}
PM: {primary['PM']}

## REFERENCE_ALL_USABLE
days: {len(ref_days)}
trades: {reference['trades']}
win: {reference['win']}
loss: {reference['loss']}
draw: {reference['draw']}
pnl: {reference['pnl']}
PF: {reference['PF']}
avg_pnl: {reference['avg_pnl']}
maxDD: {reference['maxDD']}
AM: {reference['AM']}
PM: {reference['PM']}

## BEST_DAY
date: {best[0]}
pnl: {best[1]}

## WORST_DAY
date: {worst[0]}
pnl: {worst[1]}

TOP1_PNL_CONTRIBUTION: {top1_c}
TOP3_PNL_CONTRIBUTION: {top3_c}
positive/negative/flat days (PRIMARY): {pos}/{neg}/{flat}

## OLD_A_FIXED_COMPARISON
days_compared: {len(old_cmp)}
exact_match_days: {match_n}
changed_days: {changed_n}

## 20260820
actual: 5 trades / -32600 ACTUAL_PAPER_WITH_RUNTIME_DEFECT
replay: {a20.get('trades')} trades / {a20.get('pnl')} CURRENT_RUNTIME_REPLAY

RAW_SEQUENCE_GAP_DAYS: {raw_gap_days}
UNIVERSE_BINDING_UNRESOLVED: {unresolved_all}

STRATEGY_CHANGED: false
ENTRY_EXIT_CHANGED: false
SAFETY: submit/cancel/live=0/0/0

verdict: **{verdict}**

P1 STOP。
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="top")

    def hdr(ws):
        for c in ws[1]:
            c.font = hf
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")

    def autosize(ws, max_w=48):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            nwid = 10
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                nwid = max(nwid, min(max_w, max((len(x) for x in v.splitlines()), default=0) + 2))
                cell.alignment = wrap
            ws.column_dimensions[letter].width = nwid

    ws = wb.active
    ws.title = "Summary"
    ws.append(["key", "value"])
    for k, v in [
        ("PERIOD", report["PERIOD"]),
        ("result_name", "CURRENT_RUNTIME_REPLAY"),
        ("PRIMARY_days", len(primary_days)),
        ("PRIMARY_trades", primary["trades"]),
        ("PRIMARY_pnl", primary["pnl"]),
        ("PRIMARY_PF", primary["PF"]),
        ("PRIMARY_maxDD", primary["maxDD"]),
        ("REFERENCE_days", len(ref_days)),
        ("REFERENCE_trades", reference["trades"]),
        ("REFERENCE_pnl", reference["pnl"]),
        ("BEST_DAY", f"{best[0]} {best[1]}"),
        ("WORST_DAY", f"{worst[0]} {worst[1]}"),
        ("20260820_actual", "5 / -32600 ACTUAL_PAPER_WITH_RUNTIME_DEFECT"),
        ("20260820_replay", f"{a20.get('trades')} / {a20.get('pnl')} CURRENT_RUNTIME_REPLAY"),
        ("verdict", verdict),
        ("SAFETY", "0/0/0"),
    ]:
        ws.append([k, v])
    hdr(ws)
    autosize(ws, 80)

    ws = wb.create_sheet("Daily")
    ws.append([
        "date", "capture_class", "replay_status", "QUALITY_FLAG", "universe_source", "universe_n",
        "events_processed", "raw_sequence_holes", "accepted_sequence_holes", "skip_universe",
        "trades", "win", "loss", "draw", "pnl", "PF", "avg_pnl", "maxDD",
        "AM_trades", "AM_pnl", "PM_trades", "PM_pnl", "anchor_fires", "admitted", "fills", "expired",
        "actual_trades", "actual_pnl", "old_trades", "old_pnl",
    ])
    for r in daily:
        ws.append([
            r.get("date"), r.get("capture_class"), r.get("replay_status"), ",".join(r.get("QUALITY_FLAG") or []),
            r.get("universe_source"), r.get("universe_n"), r.get("events_processed"),
            r.get("raw_sequence_holes"), r.get("accepted_sequence_holes"), r.get("skip_universe"),
            r.get("trades"), r.get("win"), r.get("loss"), r.get("draw"), r.get("pnl"), r.get("PF"),
            r.get("avg_pnl"), r.get("maxDD"),
            (r.get("AM") or {}).get("trades"), (r.get("AM") or {}).get("pnl"),
            (r.get("PM") or {}).get("trades"), (r.get("PM") or {}).get("pnl"),
            r.get("anchor_fires"), r.get("admitted"), r.get("fills"), r.get("expired"),
            r.get("actual_trades"), r.get("actual_pnl"), r.get("old_trades"), r.get("old_pnl"),
        ])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Trades")
    tcols = [
        "date", "session", "trade_id", "symbol", "anchor_time", "snapshot_sequence", "score",
        "candidate_rank", "limit", "fill_time_iso", "fill_price", "exit_time_iso", "exit_price",
        "exit_reason", "pnl_yen_100", "holding_sec",
    ]
    ws.append(tcols)
    for t in all_trades:
        ws.append([t.get(c) for c in tcols])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Capture_Inventory")
    icols = [
        "date", "jpx_trading_day", "capture_class", "capture_path", "status", "first_event", "last_event",
        "event_count", "first_seq", "last_seq", "sequence_continuity", "dropped_event_count",
        "am_coverage", "pm_coverage", "exclusion_reason",
    ]
    ws.append(icols)
    for r in inv_rows:
        ws.append([r.get(c) for c in icols])
    hdr(ws)
    autosize(ws, 60)

    ws = wb.create_sheet("Universe_Binding")
    ws.append(["date", "resolved", "source", "universe_n", "reason", "notes", "replay_eligible"])
    for r in inv_rows:
        ws.append([
            r["date"], r.get("universe_resolved"), r.get("universe_source"), r.get("universe_n"),
            r.get("universe_reason"), ",".join(r.get("universe_notes") or []), r.get("replay_eligible"),
        ])
    hdr(ws)
    autosize(ws, 70)

    ws = wb.create_sheet("Replay_Audit")
    ws.append([
        "date", "ok", "events_processed", "raw_holes", "accepted_holes", "skip_universe",
        "anchor_fires", "candidates", "admitted", "fills", "expired", "cap_blocked",
        "same_symbol_blocked", "ledger_sha", "elapsed_sec", "blocker",
    ])
    for r in daily:
        ws.append([
            r["date"], r["replay_status"], r.get("events_processed"), r.get("raw_sequence_holes"),
            r.get("accepted_sequence_holes"), r.get("skip_universe"), r.get("anchor_fires"),
            r.get("candidates_scored"), r.get("admitted"), r.get("fills"), r.get("expired"),
            r.get("cap_blocked"), r.get("same_symbol_blocked"), r.get("ledger_sha"),
            r.get("elapsed_sec"), r.get("blocker"),
        ])
    hdr(ws)
    autosize(ws, 70)

    ws = wb.create_sheet("Actual_vs_Replay")
    ws.append(["date", "actual_label", "actual_trades", "actual_pnl", "replay_trades", "replay_pnl", "mixed"])
    for r in daily:
        ws.append([
            r["date"], r.get("actual_label"), r.get("actual_trades"), r.get("actual_pnl"),
            r.get("trades") if r["replay_status"] == "OK" else None,
            r.get("pnl") if r["replay_status"] == "OK" else None,
            False,
        ])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Old_vs_New")
    ws.append(["date", "old_trades", "new_trades", "old_pnl", "new_pnl", "trade_delta", "pnl_delta", "class"])
    for r in old_cmp:
        ws.append([r["date"], r["old_trades"], r["new_trades"], r["old_pnl"], r["new_pnl"], r["trade_delta"], r["pnl_delta"], r["class"]])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Quality")
    ws.append(["date", "universe_resolved", "capture_readable", "seq_continuity", "raw_holes", "replay", "ledger", "QUALITY_FLAG"])
    for r in daily:
        row = inv_map[r["date"]]
        ws.append([
            r["date"], row.get("universe_resolved"), bool(row.get("capture_path")),
            row.get("sequence_continuity"), r.get("raw_sequence_holes"), r["replay_status"],
            "PASS" if r["replay_status"] == "OK" else r["replay_status"],
            ",".join(r.get("QUALITY_FLAG") or []),
        ])
    hdr(ws)
    autosize(ws, 60)

    ws = wb.create_sheet("Identity")
    ws.append(["key", "value"])
    for k, v in identity.items():
        if isinstance(v, dict):
            for k2, v2 in v.items():
                ws.append([f"{k}.{k2}", str(v2)])
        else:
            ws.append([k, v])
    hdr(ws)
    autosize(ws, 80)

    wb.save(OUT / "current_runtime_recalc.xlsx")
    return report


def main() -> int:
    if not INV.is_file():
        print("inventory missing; run scripts/_p1_inventory.py first", flush=True)
        return 2
    inv_raw = json.loads(INV.read_text(encoding="utf-8"))["days"]
    inv_rows = [reclassify(r) for r in inv_raw]
    CACHE.mkdir(parents=True, exist_ok=True)
    jobs = []
    for row in inv_rows:
        print(
            f"CLASS {row['date']} {row['capture_class']} uni={row.get('universe_source')} "
            f"elig={row['replay_eligible']} seq={row.get('sequence_continuity')}",
            flush=True,
        )
        if not row["replay_eligible"]:
            continue
        cache_p = CACHE / f"{row['date']}.json"
        if cache_p.is_file():
            continue
        jobs.append(
            {
                "date": row["date"],
                "capture_path": row["capture_path"],
                "universe": row["universe_symbols"],
                "universe_source": row["universe_source"],
            }
        )

    replays: dict[str, Any] = {}
    for p in CACHE.glob("*.json"):
        body = json.loads(p.read_text(encoding="utf-8"))
        replays[body["date"]] = body

    workers = 2
    print(f"REPLAY jobs={len(jobs)} cached={len(replays)} workers={workers}", flush=True)
    if jobs:
        with ProcessPoolExecutor(max_workers=workers) as ex:
            futs = {ex.submit(replay_one, j): j["date"] for j in jobs}
            for fut in as_completed(futs):
                day = futs[fut]
                try:
                    got = fut.result()
                except Exception as exc:
                    got = {"ok": False, "date": day, "blocker": f"{type(exc).__name__}:{exc}"}
                replays[day] = got
                (CACHE / f"{day}.json").write_text(json.dumps(got, ensure_ascii=False, default=str), encoding="utf-8")
                print(
                    f"DONE {day} ok={got.get('ok')} trades={got.get('trade_n')} pnl={got.get('pnl')} "
                    f"raw_holes={got.get('raw_sequence_holes')} sec={got.get('elapsed_sec')} {got.get('blocker') or ''}",
                    flush=True,
                )

    old_by, _old_src = _load_old_a()
    spot = []
    for day, got in replays.items():
        if not got.get("ok"):
            continue
        old = old_by.get(day) or {}
        if old and abs(float(got.get("pnl") or 0) - float(old.get("pnl") or 0)) > 1.0:
            spot.append(day)
        if day == "20260821":
            spot.append(day)
    spot = list(dict.fromkeys(spot))
    det: dict[str, Any] = {
        "rerun_days": [],
        "pass": True,
        "note": "P0-4 determinism already confirmed; rerun only changed/special days",
    }
    for day in spot[:3]:
        row = next(r for r in inv_rows if r["date"] == day)
        if not row["replay_eligible"]:
            continue
        print(f"SPOT {day}", flush=True)
        got2 = replay_one(
            {
                "date": day,
                "capture_path": row["capture_path"],
                "universe": row["universe_symbols"],
                "universe_source": row["universe_source"],
            }
        )
        sha1 = replays[day].get("ledger_sha")
        sha2 = got2.get("ledger_sha")
        ok = bool(got2.get("ok") and sha1 == sha2)
        det["rerun_days"].append({"date": day, "sha_match": ok, "sha1": sha1, "sha2": sha2})
        if not ok:
            det["pass"] = False
        print(f"  sha_match={ok}", flush=True)

    identity = {
        "result_name": "CURRENT_RUNTIME_REPLAY",
        "strategy_sha": STRATEGY_SHA,
        "entry_sha": ENTRY_SHA,
        "exit_sha": EXIT_V2_CANDIDATE_SHA,
        "anchor_sha": ANCHOR_SHA or ANCHOR_SHA_RT,
        "replay_code_sha": _file_sha("src/research/anchor_vs_event_driven/run_comparison.py"),
        "V1RNativeEntryLive_sha": _file_sha("src/small_paper/v1r_native_entry_live.py"),
        "V1RLiveDualLane_sha": _file_sha("src/small_paper/v1r_live_dual_lane.py"),
        "p1_runner_sha": _file_sha("scripts/run_p1_current_runtime_full_capture_recalc.py"),
        "fast_path": "P0-4 Fast Event-Time Replay (_stream_day + process_market_push + dual 0.5s, ingest_audit discarded)",
        "determinism_spot": det,
    }
    report = write_outputs(inv_rows, replays, identity)
    INV.unlink(missing_ok=True)
    if CACHE.is_dir():
        for p in CACHE.glob("*.json"):
            p.unlink()
        try:
            CACHE.rmdir()
        except OSError:
            pass
    print("verdict", report["verdict"], flush=True)
    return 0 if str(report["verdict"]).endswith("COMPLETE") else 1


if __name__ == "__main__":
    raise SystemExit(main())
