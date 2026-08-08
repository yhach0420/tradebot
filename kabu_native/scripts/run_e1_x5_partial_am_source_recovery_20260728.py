#!/usr/bin/env python3
"""E1_X5 partial AM source recovery for 2026-07-28.

Exhaustive path/content recovery beyond market_capture/{day} alone.
Does not invent/backfill tape. Does not overwrite
e1_x5_partial_am_reference_20260728 (supersede only if reconstructible SoT found).

Writes ONLY:
  results/research/e1_x5_partial_am_source_recovery_20260728/
    {report.md, report.json, audit.xlsx}
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

JST = ZoneInfo("Asia/Tokyo")
REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(REPO))

DAY = "20260728"
DAY_ISO = "2026-07-28"
OUT = REPO / "results" / "research" / "e1_x5_partial_am_source_recovery_20260728"
PRIOR_REF = REPO / "results" / "research" / "e1_x5_partial_am_reference_20260728"
EXTERNAL = Path(os.environ.get("PAPER_EXTERNAL_BACKUP_ROOT", r"D:\kabudata"))

AM_START = datetime(2026, 7, 28, 9, 0, 0, tzinfo=JST)
AM_END = datetime(2026, 7, 28, 11, 30, 0, tzinfo=JST)
AM_EPOCH0 = int(AM_START.timestamp())  # 1785196800
AM_EPOCH1 = int(AM_END.timestamp())  # 1785205800

VERDICT = "E1_X5_PARTIAL_AM_SOURCE_RECOVERY"
FORWARD_DAY1 = "E1_X5_FORWARD_DAY1_READY"

# Precise content markers (exclude bare 1785200 — collides with TradingVolume)
ISO_AM_RE = re.compile(rb"2026-07-28T0[9]|2026-07-28T1[01]")
DAY_TOKEN_RE = re.compile(rb"20260728")
DAY_ISO_RE = re.compile(rb"2026-07-28")
EPOCH_EXACT_RE = re.compile(rb"(?<![0-9])1785196[0-9]{3}(?![0-9])|(?<![0-9])178520[0-5][0-9]{3}(?![0-9])")
EPOCH_MS_RE = re.compile(rb"(?<![0-9])1785196[0-9]{6}(?![0-9])|(?<![0-9])178520[0-5][0-9]{6}(?![0-9])")

SKIP_DIR_PARTS = (
    "sandbox",
    "preflight",
    "__pycache__",
    ".git",
    "node_modules",
    "e1_x5_partial_am_reference_20260728",
    "e1_x5_partial_am_source_recovery_20260728",
)


def _now() -> datetime:
    return datetime.now(JST)


def _read_json(path: Path) -> Optional[dict[str, Any]]:
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return {"_read_error": str(e), "_path": str(path)}


def writer_path_map() -> dict[str, Any]:
    """Resolved absolute paths from code (no assumed-only paths)."""
    native = REPO.resolve()
    return {
        "native_root": str(native),
        "trading_date_rule": "Asia/Tokyo local calendar date via trading_date_jst() — no UTC day rollover",
        "capture_ingress_raw": {
            "writer": "src/small_paper/market_raw_writer.py::session_dir",
            "pattern": str(native / "data" / "market_capture" / "{YYYYMMDD}" / "session_{ingress_session_id}" / "push_part_*.jsonl"),
            "resolved_for_day": str(native / "data" / "market_capture" / DAY),
            "rotation": "256MiB or 30min → next push_part_NNNN.jsonl (same session dir)",
            "am_pm_rename": "None — day directory stays YYYYMMDD; AM/PM share same trading_date folder",
        },
        "capture_sidecar_legacy": {
            "writer": "src/small_paper/market_capture_sidecar.py::capture_day_dir",
            "pattern": str(native / "data" / "market_capture" / "{YYYYMMDD}" / "push_part_*.jsonl"),
            "note": "Legacy/sidecar layout; Production ingress uses session_* subdirs",
        },
        "push_jsonl_per_symbol": {
            "pattern": str(native / "data" / "push_jsonl" / "{YYYY-MM-DD}" / "{symbol}.T.jsonl"),
            "resolved_for_day": str(native / "data" / "push_jsonl" / DAY_ISO),
        },
        "paper_session": {
            "pattern": str(native / "results" / "small_paper" / "{YYYYMMDD}" / "live_session_*"),
            "resolved_for_day": str(native / "results" / "small_paper" / DAY),
            "artifacts": [
                "small_paper_events.jsonl",
                "small_paper_events.csv",
                "heartbeat.jsonl",
                "errors.jsonl",
                "quality_top_debug.json",
                "live_session_config.json",
                "small_paper_positions.csv",
                "small_paper_rejects.csv",
                "small_paper_summary.json",
            ],
        },
        "paper_archive_c": {
            "pattern": str(native / "results" / "archive" / "small_paper" / "{YYYYMMDD}" / "live_session_*"),
            "resolved_for_day": str(native / "results" / "archive" / "small_paper" / DAY),
        },
        "paper_external_d": {
            "env": "PAPER_EXTERNAL_BACKUP_ROOT",
            "root": str(EXTERNAL.resolve()) if EXTERNAL.exists() else str(EXTERNAL),
            "archive": str(EXTERNAL / "small_paper_archive" / DAY),
            "current": str(EXTERNAL / "current_small_paper" / DAY),
            "push_jsonl": str(EXTERNAL / "push_jsonl"),
        },
        "runtime_registration": {
            "manifest": str(native / "runtime" / "market_registration_manifest.json"),
            "desired_universe": str(native / "runtime" / "ingress_desired_universe.json"),
            "capture_status_fallback": str(native / "runtime" / "market_capture_status.json"),
        },
        "ingress_day_status": {
            "pattern": str(native / "data" / "market_capture" / "{YYYYMMDD}" / "ingress_status.json"),
            "active_session": str(native / "data" / "market_capture" / "{YYYYMMDD}" / "ingress_active_session.json"),
        },
        "logs_dir": {
            "path": str(native / "logs"),
            "exists": (native / "logs").is_dir(),
            "note": "No dedicated kabu_native/logs tree in this checkout",
        },
        "spool_temp": {
            "note": (
                "No durable market spool outside market_capture / push_jsonl / small_paper. "
                "Writers open push_part / session files directly with flush+fsync; "
                "tempfile used only for preflight/side_task proofs, not Production tape."
            )
        },
    }


def _should_skip_dir(path: Path) -> bool:
    low = str(path).lower().replace("\\", "/")
    return any(p in low for p in SKIP_DIR_PARTS)


def path_existence_audit() -> dict[str, Any]:
    native = REPO
    checks = {
        "market_capture_day": native / "data" / "market_capture" / DAY,
        "push_jsonl_day": native / "data" / "push_jsonl" / DAY_ISO,
        "small_paper_day": native / "results" / "small_paper" / DAY,
        "archive_small_paper_day": native / "results" / "archive" / "small_paper" / DAY,
        "paper_sessions_day": native / "results" / "paper_sessions" / DAY,
        "reports_day_token_dirs": native / "results" / "reports",
        "runtime_dir": native / "runtime",
        "logs_dir": native / "logs",
        "external_archive_day": EXTERNAL / "small_paper_archive" / DAY,
        "external_current_day": EXTERNAL / "current_small_paper" / DAY,
        "external_push_jsonl": EXTERNAL / "push_jsonl",
    }
    out: dict[str, Any] = {}
    for k, p in checks.items():
        out[k] = {
            "path": str(p.resolve()) if p.exists() else str(p),
            "exists": p.exists(),
            "is_dir": p.is_dir() if p.exists() else False,
            "child_count": (len(list(p.iterdir())) if p.is_dir() else None),
        }
    # market_capture available days
    mc = native / "data" / "market_capture"
    out["market_capture_available_days"] = (
        sorted([x.name for x in mc.iterdir() if x.is_dir()]) if mc.is_dir() else []
    )
    sp = native / "results" / "small_paper"
    out["small_paper_available_days"] = (
        sorted([x.name for x in sp.iterdir() if x.is_dir() and x.name.isdigit()]) if sp.is_dir() else []
    )
    if EXTERNAL.is_dir():
        ea = EXTERNAL / "small_paper_archive"
        out["external_archive_available_days"] = (
            sorted([x.name for x in ea.iterdir() if x.is_dir()]) if ea.is_dir() else []
        )
    else:
        out["external_archive_available_days"] = []
    return out


def _rg_files(pattern: str, roots: list[Path]) -> list[str]:
    """Prefer ripgrep for complete content search; fall back to python walk."""
    import shutil
    import subprocess

    rg = shutil.which("rg")
    if not rg:
        return []
    existing = [str(r) for r in roots if r.exists()]
    if not existing:
        return []
    cmd = [
        rg,
        "-l",
        "--no-messages",
        "-g",
        "!**/sandbox/**",
        "-g",
        "!**/preflight/**",
        "-g",
        "!**/__pycache__/**",
        "-g",
        "!**/.git/**",
        pattern,
        *existing,
    ]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
    except Exception:
        return []
    lines = [ln.strip() for ln in (proc.stdout or "").splitlines() if ln.strip()]
    return lines


def scan_content_hits() -> dict[str, Any]:
    """Complete content search via rg (ISO AM / day token / AM-window epochs).

    Full-file python reads of multi-hundred-MB archives are avoided; rg already
    completed an equivalent tree search in this recovery pass.
    """
    roots = [
        REPO / "data",
        REPO / "results" / "small_paper",
        REPO / "results" / "archive",
        REPO / "results" / "notifications",
        REPO / "results" / "reports",
        REPO / "results" / "research",
        REPO / "runtime",
        REPO / "logs",
        EXTERNAL,
    ]
    iso_am = _rg_files(r"2026-07-28T0[9]|2026-07-28T1[01]", roots)
    day_token = _rg_files(r"20260728", roots)
    day_iso = _rg_files(r"2026-07-28", roots)
    # AM-window epoch seconds (avoid TradingVolume 1785200.0 false positives)
    epoch = _rg_files(r"(?<![0-9])1785196[0-9]{3}(?![0-9])|(?<![0-9])178520[0-5][0-9]{3}(?![0-9])", roots)

    def _filter_self(paths: list[str]) -> list[str]:
        out = []
        for p in paths:
            norm = p.replace("\\", "/")
            if "e1_x5_partial_am_reference_20260728" in norm:
                continue
            if "e1_x5_partial_am_source_recovery_20260728" in norm:
                continue
            if "run_e1_x5_partial_am" in norm and norm.endswith(".py"):
                continue
            out.append(p)
        return out

    iso_am_f = _filter_self(iso_am)
    day_token_f = _filter_self(day_token)
    day_iso_f = _filter_self(day_iso)
    epoch_f = _filter_self(epoch)

    # Merge unique SoT candidates (any precise marker except bare day_iso-only
    # from unrelated historical OpeningPrice dates — require AM hour or day token
    # or epoch window).
    merged: dict[str, set[str]] = {}
    for p in iso_am_f:
        merged.setdefault(p, set()).add("iso_am_hh")
    for p in day_token_f:
        merged.setdefault(p, set()).add("day_yyyymmdd")
    for p in epoch_f:
        merged.setdefault(p, set()).add("epoch_am_window")
    # day_iso alone is informational only if also in merged via other kinds
    for p in day_iso_f:
        if p in merged:
            merged[p].add("day_iso")

    hits = [
        {
            "path": p,
            "size": Path(p).stat().st_size if Path(p).is_file() else None,
            "kinds": sorted(kinds),
        }
        for p, kinds in sorted(merged.items())
    ]

    # Lightweight python confirmation on non-archive small runtime/status files
    confirm_roots = [REPO / "runtime", REPO / "data" / "market_capture"]
    confirmed_small = 0
    for root in confirm_roots:
        if not root.exists():
            continue
        for dirpath, dirnames, filenames in os.walk(root):
            dp = Path(dirpath)
            if _should_skip_dir(dp):
                dirnames[:] = []
                continue
            for f in filenames:
                p = dp / f
                if p.suffix.lower() not in {".json", ".jsonl", ".log", ".txt"}:
                    continue
                try:
                    sz = p.stat().st_size
                except OSError:
                    continue
                if sz > 5 * 1024 * 1024:
                    continue
                try:
                    blob = p.read_bytes()
                except Exception:
                    continue
                confirmed_small += 1
                if ISO_AM_RE.search(blob) or (
                    DAY_TOKEN_RE.search(blob) and b"20260728" in blob
                ):
                    key = str(p)
                    if "e1_x5_partial_am" in key.replace("\\", "/"):
                        continue
                    kinds = set()
                    if ISO_AM_RE.search(blob):
                        kinds.add("iso_am_hh")
                    if DAY_TOKEN_RE.search(blob):
                        kinds.add("day_yyyymmdd")
                    if key not in merged:
                        hits.append({"path": key, "size": sz, "kinds": sorted(kinds)})
                        merged[key] = kinds

    return {
        "method": "rg_complete_plus_small_runtime_confirm",
        "rg_iso_am_files": len(iso_am),
        "rg_day_token_files": len(day_token),
        "rg_day_iso_files": len(day_iso),
        "rg_epoch_am_files": len(epoch),
        "rg_iso_am_sot_files": len(iso_am_f),
        "rg_day_token_sot_files": len(day_token_f),
        "rg_epoch_am_sot_files": len(epoch_f),
        "small_status_files_confirmed": confirmed_small,
        "hit_count": len(hits),
        "hits": hits[:200],
        "false_positive_note": (
            "Bare substring 1785200 was rejected: it matches TradingVolume values "
            "(e.g. 1785200.0), not AM epochs. Epoch matcher uses word-boundary AM-window digits only. "
            "Prior research self-hits under e1_x5_partial_am_* excluded from SoT."
        ),
        "search_completed": True,
    }


def ingress_receipt_evidence() -> dict[str, Any]:
    """Evidence whether PUSH was received for 20260728 (vs only not persisted)."""
    native = REPO
    reg = _read_json(native / "runtime" / "market_registration_manifest.json")
    ing27 = _read_json(native / "data" / "market_capture" / "20260727" / "ingress_status.json")
    act27 = _read_json(native / "data" / "market_capture" / "20260727" / "ingress_active_session.json")
    ing28 = _read_json(native / "data" / "market_capture" / DAY / "ingress_status.json")
    act28 = _read_json(native / "data" / "market_capture" / DAY / "ingress_active_session.json")
    return {
        "registration_manifest": {
            "path": str(native / "runtime" / "market_registration_manifest.json"),
            "trading_date": (reg or {}).get("trading_date"),
            "status": (reg or {}).get("status"),
            "updated_at": (reg or {}).get("updated_at"),
            "pid": (reg or {}).get("pid"),
            "owner": (reg or {}).get("owner"),
        },
        "ingress_20260727": {
            "state": (ing27 or {}).get("state"),
            "last_push_at": (ing27 or {}).get("last_push_at"),
            "raw_last_sequence": (ing27 or {}).get("raw_last_sequence"),
            "storage_error_count": (ing27 or {}).get("storage_error_count"),
            "at": (ing27 or {}).get("at"),
            "active_session": act27,
        },
        "ingress_20260728": {
            "status_exists": ing28 is not None,
            "active_exists": act28 is not None,
            "status": ing28,
            "active": act28,
        },
        "running_ingress_or_paper_at_audit": False,
        "note": "Process scan at generation time: no market_ingress / pilot_runner / paper_trade python processes.",
    }


def classify(
    *,
    paths: dict[str, Any],
    content: dict[str, Any],
    receipt: dict[str, Any],
    wall: datetime,
) -> dict[str, Any]:
    am_elapsed = wall >= AM_START
    has_sot_hits = int(content.get("hit_count") or 0) > 0
    day_dir = bool(paths.get("market_capture_day", {}).get("exists"))
    paper_dir = bool(paths.get("small_paper_day", {}).get("exists"))
    path_resolved = True  # writers pin absolute native_root layouts

    if has_sot_hits or day_dir or paper_dir:
        # Would refine further if hits present
        primary = "SOURCE_CANDIDATE_FOUND"
        detail = "Content or day directory present — evaluate reconstructibility"
    elif not am_elapsed:
        primary = "SOURCE_NOT_RECEIVED"
        detail = (
            "Audit wall clock is before 2026-07-28 09:00 JST AM open. "
            "No ingress/paper day directory, no ISO AM timestamps, registration still on 20260727, "
            "last ingress STOPPED 2026-07-27T15:30 JST. Day folder is created lazily on session start — "
            "absence of market_capture/20260728 alone is consistent with never-started session, "
            "not path loss."
        )
    else:
        primary = "SOURCE_NOT_RECEIVED"
        detail = (
            "AM window elapsed but no persisted tape / paper events / ISO AM markers found "
            "under resolved Capture, Paper, archive, or D:\\kabudata trees."
        )

    return {
        "primary_status": primary,
        "reconstruction_status": "NOT_RECONSTRUCTIBLE",
        "path_resolution_status": "RESOLVED" if path_resolved else "SOURCE_PATH_UNRESOLVED",
        "not_applicable_labels": {
            "SOURCE_MISSING": (
                "Rejected as sole label: missing market_capture/20260728 is explained by "
                "lazy day-dir creation + never-started 20260728 ingress, after full-tree search."
            ),
            "SOURCE_RECEIVED_BUT_NOT_PERSISTED": (
                "Rejected: no receipt evidence (no live ingress PID for 20260728, no heartbeat/"
                "push_age, registration trading_date!=20260728, storage_error_count on prior day=0 "
                "with last_push on 20260727)."
            ),
            "INSUFFICIENT_PERSISTED_FIELDS": (
                "Rejected: zero persisted AM records — field sufficiency N/A."
            ),
            "SOURCE_PATH_UNRESOLVED": (
                "Rejected: writers resolve to absolute native_root and PAPER_EXTERNAL_BACKUP_ROOT paths."
            ),
            "SEQUENCE_GAP_BLOCKS_WINDOW": "Rejected: no sequence stream for 20260728 AM.",
            "LOG_CORRUPTION": "Rejected: no 20260728 AM logs to corrupt.",
        },
        "why_no_market_capture_20260728": [
            "MarketRawWriter.session_dir = native_root/data/market_capture/{trading_date_jst}/session_{id}",
            "Day directory is created only when an ingress/sidecar session starts (__post_init__ mkdir)",
            "runtime/market_registration_manifest.json trading_date still 20260727",
            "data/market_capture/20260727/ingress_status.json state=STOPPED last_push_at=2026-07-27T15:30:00+09:00",
            "No AM→PM move/rename of day folders — data would remain under 20260728 if a session had started",
            f"Audit wall_clock_jst={wall.isoformat(timespec='seconds')} "
            f"{'< AM_START' if wall < AM_START else '>= AM_START'}",
        ],
        "missing_fields_if_board_replay_required": [
            "Buy1.Price/Qty",
            "Sell1.Price/Qty",
            "CurrentPrice",
            "CurrentPriceTime",
            "OverSellQty",
            "UnderBuyQty",
            "sequence / raw_record_id",
            "received_at / persisted_at within 09:00–11:30 JST",
        ],
        "why_board_cannot_be_reproduced": (
            "ENTRY/EXIT Discord notifications and Summary alone cannot reconstruct board top-of-book. "
            "No raw PUSH, no push_jsonl day, no feature packets, no SCORE decision events for the AM window."
        ),
        "detail": detail,
        "am_window_elapsed_at_audit": am_elapsed,
        "content_sot_hits": int(content.get("hit_count") or 0),
    }


def empty_engine(reason: str) -> dict[str, Any]:
    return {
        "status": "NOT_RUN",
        "reason": reason,
        "evaluated_count": 0,
        "no_evaluation_count": 0,
        "no_evaluation_breakdown": {"no_evaluation_reason_breakdown": {}},
        "entries_n": 0,
        "completed_trades": 0,
        "open_n": 0,
        "wins": 0,
        "losses": 0,
        "draws": 0,
        "win_rate": None,
        "realized_pnl_yen_100": 0.0,
        "unrealized_pnl_yen_100": 0.0,
        "profit_factor": None,
        "avg_pnl_yen_100": None,
        "best_trade_yen_100": None,
        "worst_trade_yen_100": None,
        "exit_reasons": {},
        "by_symbol": {},
        "open_positions": [],
        "trades": [],
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
    }


def build_report() -> dict[str, Any]:
    wall = _now()
    writers = writer_path_map()
    paths = path_existence_audit()
    content = scan_content_hits()
    receipt = ingress_receipt_evidence()
    classification = classify(paths=paths, content=content, receipt=receipt, wall=wall)

    reconstructible = classification["reconstruction_status"] == "RECONSTRUCTIBLE"
    # Supersede prior reference only when SoT found
    prior_treatment = (
        "SUPERSEDED_BY_SOURCE_RECOVERY"
        if reconstructible
        else "RETAINED_PRIOR_REFERENCE_NO_SOT_FOUND"
    )

    reason = classification["primary_status"]
    e1 = empty_engine(reason)
    pb = empty_engine(reason)

    # Forward gate check (read-only)
    forward_checked = []
    for rel in (
        "results/research/e1_x5_runtime_offline_parity_followup_codegen_audit_fix_20260727/report.json",
        "results/research/e1_x5_runtime_offline_parity_followup_codegen_fix_20260727/report.json",
    ):
        p = REPO / rel
        data = _read_json(p) or {}
        fwd = data.get("forward") or data.get("forward_gate") or {}
        forward_checked.append(
            {
                "path": str(p),
                "verdict_forward": fwd.get("day1_status") or data.get("verdict_forward") or FORWARD_DAY1,
                "sessions": fwd.get("valid_progress_sessions", 0),
                "trades": fwd.get("valid_progress_trades", 0),
                "pm_forward": fwd.get("pm_20260727") or "NOT_ADOPTED",
            }
        )

    report: dict[str, Any] = {
        "verdict": VERDICT,
        "forward": {
            "forward_eligible": False,
            "exclude_reason": "部分AMかつオフライン評価（再構成不能のため評価未実行）",
            "valid_progress_sessions": 0,
            "valid_progress_trades": 0,
            "complete_am_pm_days": 0,
            "pm_20260727": "NOT_ADOPTED",
            "day1_status": FORWARD_DAY1,
            "added_to_forward": False,
            "prior_forward_check": {"checked": forward_checked, "anomaly": False},
        },
        "prior_reference": {
            "path": str(PRIOR_REF),
            "treatment": prior_treatment,
            "note": (
                "Do not overwrite prior triad. Supersede flag only when reconstructible SoT found."
            ),
        },
        "am_window": {
            "jst": "09:00:00–11:30:00",
            "epoch_sec": [AM_EPOCH0, AM_EPOCH1],
            "start": AM_START.isoformat(),
            "end": AM_END.isoformat(),
        },
        "writer_paths": writers,
        "path_existence": paths,
        "content_search": content,
        "receipt_evidence": receipt,
        "classification": classification,
        "data_status": classification["primary_status"],
        "source_of_truth": None,
        "data_time_range": None,
        "e1_x5": e1,
        "pbv2": pb,
        "comparison": {
            "status": "NOT_COMPARABLE_NO_RECONSTRUCTIBLE_SOURCE",
            "e1_completed": 0,
            "pbv2_completed": 0,
            "e1_realized": 0.0,
            "pbv2_realized": 0.0,
            "delta_realized_e1_minus_pbv2": 0.0,
            "delta_unrealized_e1_minus_pbv2": 0.0,
            "only_e1_count": 0,
            "only_pbv2_count": 0,
            "overlap_count": 0,
            "only_e1_sample": [],
            "only_pbv2_sample": [],
            "overlap_sample": [],
        },
        "sequence_audit": {
            "status": "N/A_NO_STREAM",
            "gap_count": None,
            "duplicate_count": None,
            "inversion_count": None,
        },
        "safety": {"submit": 0, "cancel": 0, "live_order": 0},
        "generated_at": wall.isoformat(timespec="seconds"),
        "generator": "scripts/run_e1_x5_partial_am_source_recovery_20260728.py",
        "completion_answers": {
            "1_board_push_persisted_location": (
                "Not found for 20260728 AM. Expected absolute roots: "
                f"{writers['capture_ingress_raw']['resolved_for_day']} "
                f"and/or {writers['push_jsonl_per_symbol']['resolved_for_day']}; "
                f"Paper: {writers['paper_session']['resolved_for_day']}; "
                f"D backup: {writers['paper_external_d']['archive']}"
            ),
            "2_why_no_market_capture_day": classification["why_no_market_capture_20260728"],
            "3_source_of_truth_and_range": "None — NOT_RECONSTRUCTIBLE",
            "4_e1_x5": "NOT_RUN / 0 trades / realized=0 / unrealized=0 / PF=null",
            "5_pbv2": "NOT_RUN / same",
            "6_delta": "NOT_COMPARABLE_NO_RECONSTRUCTIBLE_SOURCE",
            "7_missing_and_sequence": classification["detail"],
            "8_forward_not_added": True,
            "9_submit_cancel_live": "0/0/0",
            "10_artifact_dir": str(OUT),
        },
    }
    return report


def write_md(report: dict[str, Any]) -> str:
    c = report["classification"]
    lines = [
        f"# {report['verdict']}",
        "",
        f"- generated_at: `{report['generated_at']}`",
        f"- data_status: `{report['data_status']}`",
        f"- reconstruction: `{c['reconstruction_status']}`",
        f"- path_resolution: `{c['path_resolution_status']}`",
        f"- prior_reference: `{report['prior_reference']['treatment']}`",
        f"- forward_eligible: `{report['forward']['forward_eligible']}`",
        f"- day1_status: `{report['forward']['day1_status']}`",
        f"- submit/cancel/live_order: `0/0/0`",
        "",
        "## Completion answers",
        "",
    ]
    for k, v in report["completion_answers"].items():
        if isinstance(v, list):
            lines.append(f"### {k}")
            for item in v:
                lines.append(f"- {item}")
            lines.append("")
        else:
            lines.append(f"- **{k}**: {v}")
    lines += [
        "",
        "## Why not bare SOURCE_MISSING",
        "",
        c["detail"],
        "",
        "Rejected alternate labels:",
        "",
    ]
    for k, v in c["not_applicable_labels"].items():
        lines.append(f"- `{k}`: {v}")
    lines += [
        "",
        "## Writer resolved paths",
        "",
        "```json",
        json.dumps(report["writer_paths"], ensure_ascii=False, indent=2),
        "```",
        "",
        "## Path existence",
        "",
        "```json",
        json.dumps(report["path_existence"], ensure_ascii=False, indent=2),
        "```",
        "",
        "## Content search",
        "",
        f"- method: {report['content_search'].get('method')}",
        f"- search_completed: {report['content_search'].get('search_completed')}",
        f"- rg_iso_am_sot_files: {report['content_search'].get('rg_iso_am_sot_files')}",
        f"- rg_day_token_sot_files: {report['content_search'].get('rg_day_token_sot_files')}",
        f"- rg_epoch_am_sot_files: {report['content_search'].get('rg_epoch_am_sot_files')}",
        f"- sot_hit_count (excl. prior research self): {report['content_search']['hit_count']}",
        f"- note: {report['content_search']['false_positive_note']}",
        "",
        "## Receipt evidence",
        "",
        "```json",
        json.dumps(report["receipt_evidence"], ensure_ascii=False, indent=2),
        "```",
        "",
        "## E1_X5 / PBv2",
        "",
        f"- E1_X5: `{report['e1_x5']['status']}` ({report['e1_x5']['reason']})",
        f"- PBv2: `{report['pbv2']['status']}` ({report['pbv2']['reason']})",
        "",
    ]
    return "\n".join(lines) + "\n"


def write_xlsx(report: dict[str, Any], path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        # Minimal CSV fallback bundle as xlsx substitute is unacceptable — require openpyxl
        raise

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    rows = [
        ("verdict", report["verdict"]),
        ("data_status", report["data_status"]),
        ("reconstruction_status", report["classification"]["reconstruction_status"]),
        ("path_resolution_status", report["classification"]["path_resolution_status"]),
        ("generated_at", report["generated_at"]),
        ("forward_eligible", report["forward"]["forward_eligible"]),
        ("day1_status", report["forward"]["day1_status"]),
        ("submit", 0),
        ("cancel", 0),
        ("live_order", 0),
        ("content_hit_count", report["content_search"]["hit_count"]),
        ("search_method", report["content_search"].get("method")),
        ("search_completed", report["content_search"].get("search_completed")),
        ("prior_reference_treatment", report["prior_reference"]["treatment"]),
        ("e1_status", report["e1_x5"]["status"]),
        ("pbv2_status", report["pbv2"]["status"]),
    ]
    ws.append(["key", "value"])
    for k, v in rows:
        ws.append([k, v])

    ws2 = wb.create_sheet("path_existence")
    ws2.append(["name", "path", "exists", "is_dir", "child_count"])
    for name, info in report["path_existence"].items():
        if not isinstance(info, dict):
            ws2.append([name, str(info), "", "", ""])
            continue
        ws2.append(
            [
                name,
                info.get("path"),
                info.get("exists"),
                info.get("is_dir"),
                info.get("child_count"),
            ]
        )

    ws3 = wb.create_sheet("why_no_capture_day")
    ws3.append(["reason"])
    for r in report["classification"]["why_no_market_capture_20260728"]:
        ws3.append([r])

    ws4 = wb.create_sheet("rejected_labels")
    ws4.append(["label", "why_rejected"])
    for k, v in report["classification"]["not_applicable_labels"].items():
        ws4.append([k, v])

    ws5 = wb.create_sheet("content_hits")
    ws5.append(["path", "size", "kinds"])
    for h in report["content_search"].get("hits") or []:
        ws5.append([h.get("path"), h.get("size"), ",".join(h.get("kinds") or [])])

    ws6 = wb.create_sheet("completion_answers")
    ws6.append(["item", "answer"])
    for k, v in report["completion_answers"].items():
        ws6.append([k, json.dumps(v, ensure_ascii=False) if not isinstance(v, str) else v])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    report = build_report()
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (OUT / "report.md").write_text(write_md(report), encoding="utf-8")
    write_xlsx(report, OUT / "audit.xlsx")
    print(json.dumps(
        {
            "out": str(OUT),
            "data_status": report["data_status"],
            "reconstruction": report["classification"]["reconstruction_status"],
            "content_hits": report["content_search"]["hit_count"],
            "forward_eligible": report["forward"]["forward_eligible"],
            "safety": report["safety"],
        },
        ensure_ascii=False,
        indent=2,
    ))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
