#!/usr/bin/env python3
"""Independent disk reopen verifier for E1_X5 canonical artifact triad.

Does NOT import renderer internals for values — only opens published files.
Exit 0 only when JSON / Markdown / Excel agree and integrity rules pass.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Any

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(REPO))

OUT = REPO / "results" / "research" / "e1_x5_canonical_path_unify_20260728"
CORRUPT = "a3007cbc11ec0630645b2e89f559ae42aeb342bf840608320c427ed918b84649"
FROZEN = "b5837b4871273aad64445e76c251a3bc72ff6aa98c41107c04dffaefe04ef2d4"
REQUIRED_SHEETS = [
    "Source_Manifest",
    "Valid_Windows",
    "Excluded_Windows",
    "BASE_Trade_Ledger",
    "G1_Trade_Ledger",
    "G1_State_Transitions",
    "Daily_Summary",
    "Window_Summary",
    "Config_Comparison",
    "Exit_Summary",
    "Symbol_Summary",
    "Timeband_Summary",
    "Concentration",
    "Ledger_Hash",
    "Tests",
    "Safety",
    "Artifact_Integrity",
]


class Fail(Exception):
    pass


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            chunk = fh.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def walk(obj: Any, path: str = "") -> list[tuple[str, Any]]:
    out: list[tuple[str, Any]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            p = f"{path}.{k}" if path else str(k)
            out.append((p, v))
            out.extend(walk(v, p))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            p = f"{path}[{i}]"
            out.extend(walk(v, p))
    return out


def find_hash_paths(obj: Any, needle: str) -> list[str]:
    return [p for p, v in walk(obj) if isinstance(v, str) and needle in v]


def md_field(md: str, key: str) -> str:
    m = re.search(rf"(?m)^{re.escape(key)}:\s*`([^`]+)`", md)
    if not m:
        raise Fail(f"markdown missing field {key}")
    return m.group(1)


def xlsx_kv(ws) -> dict[str, str]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    # Artifact_Integrity / Ledger_Hash / Safety style key-value
    if len(rows[0]) >= 2 and str(rows[0][0]).lower() in {"key", "keys"}:
        return {str(r[0]): ("" if r[1] is None else str(r[1])) for r in rows[1:] if r and r[0] is not None}
    # two-column without header
    if all(r and len(r) >= 2 for r in rows[: min(3, len(rows))]):
        return {str(r[0]): ("" if r[1] is None else str(r[1])) for r in rows if r and r[0] is not None}
    return {}


def independent_ledger_stats(trades: list[dict[str, Any]]) -> dict[str, Any]:
    n = len(trades)
    pnl = sum(float(t["net_pnl_yen_100"]) for t in trades)
    wins = sum(1 for t in trades if float(t["net_pnl_yen_100"]) > 0)
    losses = sum(1 for t in trades if float(t["net_pnl_yen_100"]) < 0)
    draws = sum(1 for t in trades if float(t["net_pnl_yen_100"]) == 0)
    gp = sum(float(t["net_pnl_yen_100"]) for t in trades if float(t["net_pnl_yen_100"]) > 0)
    gl = -sum(float(t["net_pnl_yen_100"]) for t in trades if float(t["net_pnl_yen_100"]) < 0)
    pf = (gp / gl) if gl > 0 else None
    return {"n": n, "pnl": pnl, "wins": wins, "losses": losses, "draws": draws, "pf": pf}


def verify_bundle(out_dir: Path = OUT) -> dict[str, Any]:
    from openpyxl import load_workbook

    errors: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            errors.append(msg)

    jp, mp, xp = out_dir / "report.json", out_dir / "report.md", out_dir / "audit.xlsx"
    for p in (jp, mp, xp):
        check(p.is_file(), f"missing {p.name}")
    if errors:
        raise Fail("; ".join(errors))

    report = json.loads(jp.read_text(encoding="utf-8"))
    md = mp.read_text(encoding="utf-8")
    wb = load_workbook(xp, read_only=True, data_only=True)

    try:
        _verify_opened(report, md, wb, out_dir, jp, mp, xp, errors, check)
    finally:
        wb.close()

    if errors:
        raise Fail("\n".join(errors))

    file_shas = {
        "report.json": sha256_file(jp),
        "report.md": sha256_file(mp),
        "audit.xlsx": sha256_file(xp),
    }
    return {
        "ok": True,
        "run_id": report["run_id"],
        "report_payload_sha256": report["report_payload_sha256"],
        "file_sha256": file_shas,
        "verdicts": {
            "execution_status": report["execution_status"],
            "artifact_integrity_verdict": report["artifact_integrity_verdict"],
            "base_verdict": report["base_verdict"],
            "g1_wiring_verdict": report["g1_wiring_verdict"],
            "g1_adoption_verdict": report["g1_adoption_verdict"],
            "overall_verdict": report["overall_verdict"],
        },
        "failed_tests": report["failed_tests"],
        "parity_canonical_actual": {
            "v1": report["parity_20260727_pm"]["hashes"]["canonical_actual"]["v1"],
            "v2": report["parity_20260727_pm"]["hashes"]["canonical_actual"]["v2"],
        },
        "base": {
            "trades": len(report["base"]["trades"]),
            "pnl": sum(float(t["net_pnl_yen_100"]) for t in report["base"]["trades"]),
            "pf": report["base"]["summary"].get("profit_factor"),
            "counters": report["base"]["counters"],
        },
    }


def _verify_opened(report, md, wb, out_dir, jp, mp, xp, errors, check) -> None:
    # required top-level
    for k in (
        "run_id",
        "report_payload_sha256",
        "failed_tests",
        "execution_status",
        "artifact_integrity_verdict",
        "base_verdict",
        "g1_wiring_verdict",
        "g1_adoption_verdict",
        "overall_verdict",
        "base",
        "g1",
        "parity_20260727_pm",
        "tests",
        "safety",
    ):
        check(k in report, f"MISSING_REQUIRED_FIELD:{k}")

    run_id = report.get("run_id")
    payload = report.get("report_payload_sha256")
    failed = report.get("failed_tests")

    check(isinstance(failed, list), "failed_tests must be list")
    check(failed == [], f"failed_tests not empty: {failed}")

    # Markdown agree
    try:
        check(md_field(md, "run_id") == run_id, "MD run_id mismatch")
        check(md_field(md, "report_payload_sha256") == payload, "MD payload mismatch")
        check(md_field(md, "overall_verdict") == report["overall_verdict"], "MD overall mismatch")
        check(md_field(md, "artifact_integrity_verdict") == report["artifact_integrity_verdict"], "MD integrity mismatch")
        check(md_field(md, "base_verdict") == report["base_verdict"], "MD base_verdict mismatch")
        check(md_field(md, "g1_wiring_verdict") == report["g1_wiring_verdict"], "MD g1_wiring mismatch")
    except Fail as e:
        errors.append(str(e))
    check(
        f"orphan_open: {report['base']['counters']['orphan_open']}" in md,
        "MD orphan_open mismatch",
    )
    check(
        f"cap_blocked: {report['base']['counters']['cap_blocked']}" in md,
        "MD cap_blocked mismatch",
    )

    # Excel sheets
    for name in REQUIRED_SHEETS:
        check(name in wb.sheetnames, f"missing sheet {name}")

    ai = xlsx_kv(wb["Artifact_Integrity"]) if "Artifact_Integrity" in wb.sheetnames else {}
    lh = xlsx_kv(wb["Ledger_Hash"]) if "Ledger_Hash" in wb.sheetnames else {}
    check(ai.get("run_id") == run_id, f"XLSX run_id mismatch: {ai.get('run_id')}")
    check(ai.get("report_payload_sha256") == payload, "XLSX payload mismatch")
    check(ai.get("overall_verdict") == str(report.get("overall_verdict")), "XLSX overall mismatch")
    check(ai.get("g1_wiring_verdict") == str(report.get("g1_wiring_verdict")), "XLSX g1_wiring mismatch")

    parity = report["parity_20260727_pm"]["hashes"]
    ca_v1 = parity["canonical_actual"]["v1"]
    ca_v2 = parity["canonical_actual"]["v2"]
    check(ca_v1 == FROZEN, f"canonical actual v1 != frozen: {ca_v1}")
    check(ca_v2 == FROZEN, f"canonical actual v2 != frozen: {ca_v2}")
    check(parity["frozen_reference"]["v1"] == FROZEN, "frozen_reference.v1 wrong")
    check(parity["observed_corrupt"]["default_str_v0"] == CORRUPT, "corrupt slot wrong")

    check(lh.get("parity_canonical_actual_v1") == ca_v1, "XLSX Ledger_Hash v1 mismatch")
    check(lh.get("parity_canonical_actual_v2") == ca_v2, "XLSX Ledger_Hash v2 mismatch")
    check(lh.get("parity_observed_corrupt_default_str_v0") == CORRUPT, "XLSX corrupt slot mismatch")

    # corrupt hash whitelist
    corrupt_paths = find_hash_paths(report, CORRUPT)
    for p in corrupt_paths:
        ok = "observed_corrupt" in p and "default_str_v0" in p
        check(ok, f"corrupt hash outside whitelist: {p}")
    if CORRUPT in md:
        for line in md.splitlines():
            if CORRUPT in line and "observed_corrupt" not in line.lower() and "default_str_v0" not in line.lower():
                errors.append(f"corrupt hash in MD non-whitelist line: {line[:200]}")

    # BASE independent recalc
    base = report["base"]
    trades = base["trades"]
    stats = independent_ledger_stats(trades)
    summary = base["summary"]
    counters = base["counters"]
    check(stats["n"] == int(summary["completed_trades"]), "independent trades != summary")
    check(abs(stats["pnl"] - float(summary["realized_pnl_yen_100"])) < 1e-6, "independent pnl != summary")
    check(stats["wins"] + stats["losses"] + stats["draws"] == stats["n"], "W+L+D != n")
    check(stats["n"] == 407, f"BASE trades != 407 ({stats['n']})")
    check(abs(stats["pnl"] - 350550.485) < 1e-3, f"BASE pnl != 350550.485 ({stats['pnl']})")
    check(stats["pf"] is not None and abs(float(stats["pf"]) - 1.390) < 0.001, f"BASE pf != 1.390 ({stats['pf']})")
    check(int(counters["cap_blocked"]) == 264, "cap_blocked != 264")
    check(int(counters["same_symbol_blocked"]) == 1676, "same != 1676")
    check(int(counters["orphan_open"]) == 3, "orphan != 3")
    check(len(base["orphans"]) == 3, "orphan ledger rows != 3")
    check(all(o.get("reason") == "WINDOW_END_OPEN_EXCLUDED" for o in base["orphans"]), "orphan reason")
    check(int(counters["negative_holding"]) == 0, "negative_holding != 0")

    day_tr = sum(int(d["completed_trades"]) for d in base["daily_summary"])
    day_pnl = sum(float(d["realized_pnl_yen_100"]) for d in base["daily_summary"])
    win_tr = sum(int(w["trades"]) for w in base["window_summary"])
    win_pnl = sum(float(w["pnl"]) for w in base["window_summary"])
    check(day_tr == stats["n"], "sum day trades != total")
    check(abs(day_pnl - stats["pnl"]) < 1e-6, "sum day pnl != total")
    check(win_tr == day_tr, "sum window trades != day")
    check(abs(win_pnl - day_pnl) < 1e-6, "sum window pnl != day")

    # G1 wiring
    check(report["g1_wiring_verdict"] == "VERIFIED", f"g1_wiring_verdict={report['g1_wiring_verdict']}")
    check(report["artifact_integrity_verdict"] == "PASS", f"integrity={report['artifact_integrity_verdict']}")
    check(report["base_verdict"] == "VERIFIED", f"base_verdict={report['base_verdict']}")
    variants = report["g1"]["variants"]
    check(len(variants) >= 6, "g1 variants incomplete")
    for v in variants:
        for k in (
            "variant_id",
            "config_fingerprint",
            "candidate",
            "armed",
            "confirmed",
            "cancelled_by_reason",
            "rearm_transition",
            "accepted",
            "blocked_by_cap",
            "blocked_by_same_symbol",
            "trade_ledger_hash",
            "state_transition_ledger_hash",
        ):
            check(k in v, f"variant missing {k} in {v.get('variant_id')}")

    # safety
    safety = report["safety"]
    check(int(safety["submit"]) == 0 and int(safety["cancel"]) == 0 and int(safety["live_order"]) == 0, "safety non-zero")

    # stale logs
    cur = out_dir / "CURRENT_RUN_ID.txt"
    check(cur.is_file(), "missing CURRENT_RUN_ID.txt")
    if cur.is_file():
        cur_id = cur.read_text(encoding="utf-8").strip()
        check(cur_id == run_id, f"CURRENT_RUN_ID stale: {cur_id} != {run_id}")
    for logp in out_dir.glob("*.log"):
        text = logp.read_text(encoding="utf-8", errors="replace")
        if run_id not in text:
            if logp.name.startswith("run_") or logp.name in {"finalize.log", "run.log"}:
                errors.append(f"stale log without run_id: {logp.name}")
        if logp.name == "finalize.log" and run_id not in text:
            errors.append("finalize.log does not match current run_id")

    for t in report["tests"]:
        check("test_name" in t and "passed" in t and "evidence_json_paths" in t, f"bad test shape: {t}")

    check(len(report["valid_windows"]) == 7, "valid_windows != 7")
    check(int(report["source_row_counts"]["total"]) == 3937344, "norm rows != 3937344")


def main(argv: list[str] | None = None) -> int:
    out = OUT
    if argv is None:
        argv = sys.argv[1:]
    if argv:
        out = Path(argv[0])
    try:
        result = verify_bundle(out)
    except Fail as e:
        print("VERIFY_FAIL", flush=True)
        print(str(e), flush=True)
        return 1
    except Exception as e:
        print("VERIFY_ERROR", flush=True)
        print(repr(e), flush=True)
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    print("VERIFY_OK exit=0", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
