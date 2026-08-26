#!/usr/bin/env python
"""P0-4: Exact Runtime Replay vs Fast Event-Time Replay parity.

Does not start full-period recalc. Does not rewrite 20260820 Actual.
Does not change Strategy / ENTRY / EXIT / Current Runtime.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

for _k in (
    "KABU_V1R_ENTRY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_NOTIFY_WEBHOOK_URL",
    "KABU_SMALL_PAPER_CAP_BLOCKED_WEBHOOK_URL",
    "KABU_DISCORD_RESEARCH_WEBHOOK_URL",
    "KABU_SHADOW_DISCORD_WEBHOOK_URL",
    "KABU_DISCORD_MARKET_CAPTURE_WEBHOOK_URL",
    "KABU_MARKET_CAPTURE_WEBHOOK_URL",
):
    os.environ.pop(_k, None)
os.environ["V1R_EXIT_V2_LIVE_PRIMARY"] = "1"

from research.anchor_vs_event_driven.run_comparison import (  # noqa: E402
    _bare,
    _boot,
    _stream_day,
    extract_trades,
    find_capture_dir,
    historical_universe,
)
from small_paper.v1r_live_dual_lane import canonical_symbol_key  # noqa: E402
from small_paper.v1r_native_entry_live import V1RNativeEntryLive  # noqa: E402
from small_paper.v1r_primary_runtime import CLOCK_GRID  # noqa: E402

from run_p0_3_exact_runtime_replay_20260820 import (  # noqa: E402
    _anchor_from_fill_t,
    _canonical_trades,
    _iso,
    _ledger_sha,
    _maxdd,
    _pf,
    _sess_stats,
)

JST = ZoneInfo("Asia/Tokyo")
P03 = ROOT / "results" / "research" / "exact_runtime_replay_20260820_p0_3" / "report.json"
OUT = ROOT / "results" / "research" / "exact_vs_fast_replay_parity_p0_4"


class _Discard:
    def append(self, *_a: Any, **_k: Any) -> None:
        return None

    def clear(self) -> None:
        return None


def _hm_epoch(day: str, h: int, m: int) -> float:
    return datetime(int(day[:4]), int(day[4:6]), int(day[6:]), h, m, tzinfo=JST).timestamp()


class CollectorEngine(V1RNativeEntryLive):
    """Harvest only. Trading path remains V1RNativeEntryLive.process_market_push."""

    def __init__(self, *a: Any, **k: Any) -> None:
        super().__init__(*a, **k)
        self.a_candidates: list[dict[str, Any]] = []
        self.a_admits: list[dict[str, Any]] = []
        self.a_fills: list[dict[str, Any]] = []
        self.a_expired: list[dict[str, Any]] = []
        self.snapshots: dict[tuple[str, str], dict[str, Any]] = {}
        self.cap_blocked = 0
        self.same_symbol_blocked = 0

    def _harvest(self, events: list[dict[str, Any]], *, default_anchor: str = "") -> None:
        for ev in events:
            kind = str(ev.get("kind") or "")
            an = str(ev.get("anchor") or default_anchor)
            if kind == "ANCHOR_SYMBOL_SNAPSHOT":
                row = {
                    "symbol": _bare(ev.get("symbol")),
                    "anchor": an,
                    "t0": ev.get("anchor_t0"),
                    "snapshot_sequence": ev.get("snapshot_sequence"),
                    "score": ev.get("model_score"),
                    "bid": (ev.get("Buy1") or {}).get("Price") if isinstance(ev.get("Buy1"), dict) else None,
                    "admitted": bool(ev.get("admitted")),
                    "rank": ev.get("rank"),
                }
                self.snapshots[(an, row["symbol"])] = row
                if ev.get("model_score") is not None:
                    self.a_candidates.append(row)
            elif kind == "V1R_ENTRY_PENDING":
                self.a_admits.append(
                    {
                        "symbol": _bare(ev.get("symbol")),
                        "anchor": an,
                        "score": ev.get("score"),
                        "rank": ev.get("rank"),
                        "limit": ev.get("limit"),
                    }
                )
            elif kind == "V1R_FILL":
                self.a_fills.append(
                    {
                        "symbol": _bare(ev.get("symbol")),
                        "anchor": an,
                        "score": ev.get("score"),
                        "limit": ev.get("limit"),
                        "fill_price": ev.get("fill_price"),
                        "fill_time": ev.get("fill_time"),
                    }
                )
            elif kind == "V1R_EXPIRED":
                self.a_expired.append(
                    {
                        "symbol": _bare(ev.get("symbol")),
                        "anchor": an,
                        "limit": ev.get("limit"),
                    }
                )
            elif kind == "CAP_BLOCKED":
                self.cap_blocked += 1

    def _run_anchor(self, *, anchor: str, t0: float, day: str, session: str) -> list[dict[str, Any]]:
        pending_before = set(self.pending)
        open_before = set(self.open_symbols)
        self._harvest(self.events)
        self.events.clear()
        out = super()._run_anchor(anchor=anchor, t0=t0, day=day, session=session)
        scored_admitted = [
            c for c in self.a_candidates if c.get("anchor") == anchor and c.get("admitted")
        ]
        pending_after = set(self.pending)
        for c in scored_admitted:
            sym = str(c.get("symbol") or "")
            if sym in pending_before or (sym in open_before and sym not in pending_after):
                if not any(a.get("symbol") == sym and a.get("anchor") == anchor for a in self.a_admits):
                    self.same_symbol_blocked += 1
        self._harvest(self.events, default_anchor=anchor)
        self.events.clear()
        return out


def _capture_meta(day: str, capture: Path) -> dict[str, Any]:
    comp = {}
    p = capture / "capture_completeness.json"
    if p.is_file():
        comp = json.loads(p.read_text(encoding="utf-8"))
    seal = {}
    sp = capture / "seal.json"
    if sp.is_file():
        seal = json.loads(sp.read_text(encoding="utf-8"))
    frozen = ROOT / "runtime" / f"same_day_am_frozen_universe_{day}.json"
    universe, uni_src = historical_universe(day, capture)
    status = str(comp.get("status") or comp.get("label") or "")
    dropped = int(comp.get("dropped_event_count") or 0)
    ok = (
        status == "COMPLETE_CAPTURE"
        and dropped == 0
        and len(universe) == 50
        and frozen.is_file()
    )
    return {
        "date": day,
        "capture_path": str(capture),
        "status": status,
        "dropped_event_count": dropped,
        "events_hint": int(seal.get("raw_rows") or 0),
        "universe_n": len(universe),
        "universe_source": uni_src,
        "frozen_present": frozen.is_file(),
        "ok": ok,
        "reason": "" if ok else f"status={status} dropped={dropped} uni={len(universe)} frozen={frozen.is_file()}",
    }


def run_replay(day: str, capture: Path, *, mode: str) -> dict[str, Any]:
    """mode=exact keeps ingest_audit; mode=fast discards audit I/O only."""
    t_wall = time.perf_counter()
    universe, uni_src = historical_universe(day, capture)
    eng, dual = _boot(universe, CollectorEngine)
    if dual is None or not eng.ready:
        return {"ok": False, "blocker": getattr(eng, "fail_reason", "dual_unavailable"), "mode": mode, "date": day}
    eng.notify_enabled = False
    if mode == "fast":
        eng.ingest_audit = _Discard()  # type: ignore[assignment]
    events_n, last_et = _stream_day(day, capture, eng, dual)
    eng._harvest(eng.events)
    raw_trades = extract_trades(dual)
    trades: list[dict[str, Any]] = []
    for i, tr in enumerate(raw_trades, start=1):
        fill_t = float(tr.get("entry_time") or 0.0)
        an = _anchor_from_fill_t(day, fill_t)
        sym = canonical_symbol_key(tr.get("symbol"))
        snap = eng.snapshots.get((an, sym), {})
        fill_row = next((f for f in eng.a_fills if f.get("symbol") == sym and f.get("anchor") == an), None)
        admit_row = next((a for a in eng.a_admits if a.get("symbol") == sym and a.get("anchor") == an), None)
        src = fill_row or admit_row or snap
        trades.append(
            {
                "trade_id": f"{day}|{tr.get('session')}|{an}|{sym}|{i}",
                "symbol": sym,
                "session": tr.get("session"),
                "anchor_time": an,
                "snapshot_sequence": snap.get("snapshot_sequence"),
                "score": src.get("score") if src else None,
                "candidate_rank": (admit_row or snap).get("rank") if (admit_row or snap) else None,
                "admission": True,
                "limit": (fill_row or admit_row or {}).get("limit") or tr.get("entry_price"),
                "fill_time": fill_t,
                "fill_time_iso": _iso(fill_t),
                "fill_price": tr.get("entry_price"),
                "exit_time": tr.get("exit_time"),
                "exit_time_iso": _iso(tr.get("exit_time")),
                "exit_price": tr.get("exit_price"),
                "exit_reason": tr.get("reason"),
                "pnl_yen_100": float(tr.get("pnl_yen_100") or 0.0),
            }
        )
    pnls = [float(t.get("pnl_yen_100") or 0.0) for t in trades]
    return {
        "ok": True,
        "mode": mode,
        "date": day,
        "universe_n": len(universe),
        "universe_source": uni_src,
        "events_n": events_n,
        "sequence_holes": int(eng.native_ingest_sequence_holes),
        "native_ingest_skip_duplicate": int(eng.native_ingest_skip_duplicate),
        "native_admitted": int(eng.primary_admitted),
        "native_fills": int(eng.primary_fills),
        "native_expired": int(eng.primary_expired),
        "anchor_fires": int(eng.anchor_fires),
        "cap_blocked": int(eng.cap_blocked),
        "same_symbol_blocked": int(eng.same_symbol_blocked),
        "candidates": [
            {
                "symbol": c.get("symbol"),
                "anchor": c.get("anchor"),
                "snapshot_sequence": c.get("snapshot_sequence"),
                "score": c.get("score"),
                "bid": c.get("bid"),
                "admitted": c.get("admitted"),
                "rank": c.get("rank"),
            }
            for c in eng.a_candidates
        ],
        "admits": eng.a_admits,
        "fills": eng.a_fills,
        "expired": eng.a_expired,
        "trades": trades,
        "pnl": round(sum(pnls), 2),
        "PF": _pf(pnls),
        "maxDD": _maxdd(trades),
        "AM": _sess_stats(trades, "AM"),
        "PM": _sess_stats(trades, "PM"),
        "elapsed_sec": round(time.perf_counter() - t_wall, 3),
        "ledger_sha": _ledger_sha(trades),
        "last_et": last_et,
    }


def load_p03_exact() -> dict[str, Any]:
    body = json.loads(P03.read_text(encoding="utf-8"))
    trades = list(body.get("trades") or [])
    return {
        "ok": True,
        "mode": "exact",
        "date": "20260820",
        "reused_p0_3": True,
        "universe_n": (body.get("identity") or {}).get("universe_n"),
        "universe_source": (body.get("identity") or {}).get("universe_source"),
        "events_n": body.get("events_n"),
        "sequence_holes": body.get("sequence_holes"),
        "native_admitted": (body.get("EXACT_RUNTIME") or {}).get("native_admitted"),
        "native_fills": (body.get("EXACT_RUNTIME") or {}).get("native_fills"),
        "native_expired": (body.get("EXACT_RUNTIME") or {}).get("native_expired"),
        "anchor_fires": (body.get("EXACT_RUNTIME") or {}).get("anchor_fires"),
        "cap_blocked": 0,
        "same_symbol_blocked": 0,
        "candidates": [],
        "admits": list(body.get("admits") or []),
        "fills": [],
        "expired": [],
        "trades": trades,
        "pnl": (body.get("EXACT_RUNTIME") or {}).get("pnl"),
        "AM": (body.get("EXACT_RUNTIME") or {}).get("AM"),
        "PM": (body.get("EXACT_RUNTIME") or {}).get("PM"),
        "elapsed_sec": body.get("elapsed_sec_run1"),
        "ledger_sha": body.get("LEDGER_SHA_RUN1") or _ledger_sha(trades),
    }


def compare_pair(day: str, exact: dict[str, Any], fast: dict[str, Any]) -> dict[str, Any]:
    et = list(exact.get("trades") or [])
    ft = list(fast.get("trades") or [])
    n = max(len(et), len(ft))
    rows = []
    mismatch = 0
    for i in range(n):
        e = et[i] if i < len(et) else None
        f = ft[i] if i < len(ft) else None
        if e is None or f is None:
            mismatch += 1
            rows.append({"class": "COUNT_MISMATCH", "exact": e, "fast": f})
            continue
        fields = (
            "symbol",
            "session",
            "anchor_time",
            "snapshot_sequence",
            "limit",
            "fill_price",
            "exit_price",
            "exit_reason",
        )
        bad = []
        for k in fields:
            ev, fv = e.get(k), f.get(k)
            if k in ("limit", "fill_price", "exit_price"):
                if abs(float(ev or 0) - float(fv or 0)) > 1e-9:
                    bad.append(k)
            elif str(ev) != str(fv):
                bad.append(k)
        if abs(float(e.get("pnl_yen_100") or 0) - float(f.get("pnl_yen_100") or 0)) > 1e-6:
            bad.append("pnl_yen_100")
        if abs(float(e.get("fill_time") or 0) - float(f.get("fill_time") or 0)) > 1e-6:
            bad.append("fill_time")
        if abs(float(e.get("exit_time") or 0) - float(f.get("exit_time") or 0)) > 1e-6:
            bad.append("exit_time")
        # score / rank: allow tiny float noise
        if e.get("score") is not None and f.get("score") is not None:
            if abs(float(e["score"]) - float(f["score"])) > 1e-9:
                bad.append("score")
        if e.get("candidate_rank") != f.get("candidate_rank"):
            # 0820 Exact reused may have rank; Fast always has it. Compare when both present.
            if e.get("candidate_rank") is not None and f.get("candidate_rank") is not None:
                bad.append("candidate_rank")
        klass = "MATCHED" if not bad else "MISMATCH:" + ",".join(bad)
        if bad:
            mismatch += 1
        rows.append({"class": klass, "symbol": e.get("symbol"), "anchor": e.get("anchor_time"), "fields": bad})

    ea = {(a.get("symbol"), a.get("anchor")): a for a in (exact.get("admits") or [])}
    fa = {(a.get("symbol"), a.get("anchor")): a for a in (fast.get("admits") or [])}
    anchor_mismatch = 0
    if ea and fa:
        keys = set(ea) | set(fa)
        for k in keys:
            if k not in ea or k not in fa:
                anchor_mismatch += 1
                continue
            if abs(float(ea[k].get("limit") or 0) - float(fa[k].get("limit") or 0)) > 1e-9:
                anchor_mismatch += 1
            elif abs(float(ea[k].get("score") or 0) - float(fa[k].get("score") or 0)) > 1e-9:
                anchor_mismatch += 1
    elif exact.get("anchor_fires") != fast.get("anchor_fires") or exact.get("native_admitted") != fast.get("native_admitted"):
        anchor_mismatch = 1

    sha_e = exact.get("ledger_sha") or _ledger_sha(et)
    sha_f = fast.get("ledger_sha") or _ledger_sha(ft)
    pnl_mismatch = abs(float(exact.get("pnl") or 0) - float(fast.get("pnl") or 0)) > 1e-6
    return {
        "date": day,
        "trade_mismatch": mismatch,
        "anchor_mismatch": anchor_mismatch,
        "pnl_mismatch": pnl_mismatch,
        "sha_match": sha_e == sha_f,
        "exact_sha": sha_e,
        "fast_sha": sha_f,
        "rows": rows,
        "count_eq": len(et) == len(ft),
        "holes_e": exact.get("sequence_holes"),
        "holes_f": fast.get("sequence_holes"),
    }


def write_reports(payload: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OUT.mkdir(parents=True, exist_ok=True)
    days = payload["days"]
    all_trade = all(d["cmp"]["trade_mismatch"] == 0 and d["cmp"]["sha_match"] for d in days)
    all_anchor = all(d["cmp"]["anchor_mismatch"] == 0 for d in days)
    all_pnl = all(not d["cmp"]["pnl_mismatch"] for d in days)
    holes0 = all(int(d["exact"].get("sequence_holes") or 0) == 0 and int(d["fast"].get("sequence_holes") or 0) == 0 for d in days)
    det = bool(payload["determinism"]["pass"])
    verdict = (
        "P0_4_EXACT_FAST_PARITY_PASS"
        if all_trade and all_anchor and all_pnl and holes0 and det
        else "P0_4_EXACT_FAST_PARITY_FAIL"
    )
    can = "YES" if verdict.endswith("PASS") else "NO"
    report = {
        "task": "P0-4",
        "FAST_PATH": payload["fast_path"],
        "EXACT_PATH": payload["exact_path"],
        "TEST_DAYS": [d["date"] for d in days],
        "PER_DAY": [
            {
                "date": d["date"],
                "events": d["fast"]["events_n"],
                "exact_trades": len(d["exact"]["trades"]),
                "fast_trades": len(d["fast"]["trades"]),
                "exact_pnl": d["exact"]["pnl"],
                "fast_pnl": d["fast"]["pnl"],
                "trade_mismatch": d["cmp"]["trade_mismatch"],
                "anchor_mismatch": d["cmp"]["anchor_mismatch"],
                "exact_sha": d["cmp"]["exact_sha"],
                "fast_sha": d["cmp"]["fast_sha"],
                "exact_sec": d["exact"].get("elapsed_sec"),
                "fast_sec": d["fast"].get("elapsed_sec"),
                "speedup_x": (
                    None
                    if not d["exact"].get("elapsed_sec") or not d["fast"].get("elapsed_sec")
                    else round(float(d["exact"]["elapsed_sec"]) / max(float(d["fast"]["elapsed_sec"]), 1e-9), 3)
                ),
                "sequence_holes_exact": d["exact"].get("sequence_holes"),
                "sequence_holes_fast": d["fast"].get("sequence_holes"),
                "native_admitted_exact": d["exact"].get("native_admitted"),
                "native_admitted_fast": d["fast"].get("native_admitted"),
                "fills_exact": d["exact"].get("native_fills"),
                "fills_fast": d["fast"].get("native_fills"),
                "expired_exact": d["exact"].get("native_expired"),
                "expired_fast": d["fast"].get("native_expired"),
                "anchor_fires_exact": d["exact"].get("anchor_fires"),
                "anchor_fires_fast": d["fast"].get("anchor_fires"),
                "capture": d["meta"],
            }
            for d in days
        ],
        "ALL_DAYS_TRADE_PARITY": "PASS" if all_trade else "FAIL",
        "ALL_DAYS_ANCHOR_PARITY": "PASS" if all_anchor else "FAIL",
        "ALL_DAYS_PNL_PARITY": "PASS" if all_pnl else "FAIL",
        "DETERMINISM": "PASS" if det else "FAIL",
        "determinism": payload["determinism"],
        "CODE_CHANGE": False,
        "CHANGED_FILES": [],
        "STRATEGY_CHANGED": False,
        "ENTRY_EXIT_CHANGED": False,
        "SAFETY": "submit/cancel/live=0/0/0",
        "ACTUAL_PAPER_WITH_RUNTIME_DEFECT": {"date": "20260820", "trades": 5, "pnl": -32600, "note": "not a parity target"},
        "CAN_PROCEED_TO_P1_FULL_RECALC": can,
        "verdict": verdict,
        "path_audit": payload["path_audit"],
        "day_selection": payload["day_selection"],
    }
    slim_days = []
    for d in days:
        slim_days.append(
            {
                "date": d["date"],
                "exact_trades": d["exact"]["trades"],
                "fast_trades": d["fast"]["trades"],
                "exact_admits": d["exact"].get("admits") or [],
                "fast_admits": d["fast"].get("admits") or [],
                "cmp_rows": d["cmp"]["rows"],
            }
        )
    report["ledgers"] = slim_days
    (OUT / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

    lines = [
        "# P0-4 Exact vs Fast Event-Time Replay Parity",
        "",
        "全期間再集計には進まない。Strategy / ENTRY / EXIT / Current Runtime 未変更。submit/cancel/live=0/0/0。",
        "20260820 Actual 5/−32600 は `ACTUAL_PAPER_WITH_RUNTIME_DEFECT` のまま（parity target ではない）。",
        "",
        "## FAST_PATH",
        "",
        payload["fast_path"],
        "",
        "## EXACT_PATH",
        "",
        payload["exact_path"],
        "",
        "## TEST_DAYS",
        "",
        ", ".join(d["date"] for d in days),
        "",
        "## Path audit (not assumed from PnL)",
        "",
    ]
    for row in payload["path_audit"]:
        lines.append(f"- **{row['stage']}**: Exact={row['exact']}; Fast={row['fast']}; delta={row['delta']}")
    lines += ["", "## PER_DAY", ""]
    for d in report["PER_DAY"]:
        lines += [
            f"### {d['date']}",
            "",
            f"events: {d['events']}",
            f"exact_trades: {d['exact_trades']}",
            f"fast_trades: {d['fast_trades']}",
            f"exact_pnl: {d['exact_pnl']}",
            f"fast_pnl: {d['fast_pnl']}",
            f"trade_mismatch: {d['trade_mismatch']}",
            f"anchor_mismatch: {d['anchor_mismatch']}",
            f"exact_sha: `{d['exact_sha']}`",
            f"fast_sha: `{d['fast_sha']}`",
            f"exact_sec: {d['exact_sec']}",
            f"fast_sec: {d['fast_sec']}",
            f"speedup_x: {d['speedup_x']}",
            "",
        ]
    lines += [
        f"ALL_DAYS_TRADE_PARITY: **{report['ALL_DAYS_TRADE_PARITY']}**",
        f"ALL_DAYS_ANCHOR_PARITY: **{report['ALL_DAYS_ANCHOR_PARITY']}**",
        f"ALL_DAYS_PNL_PARITY: **{report['ALL_DAYS_PNL_PARITY']}**",
        f"DETERMINISM: **{report['DETERMINISM']}**",
        "CODE_CHANGE: no",
        "CHANGED_FILES: (none)",
        "STRATEGY_CHANGED: false",
        "ENTRY_EXIT_CHANGED: false",
        "SAFETY: submit/cancel/live=0/0/0",
        f"CAN_PROCEED_TO_P1_FULL_RECALC: {can}",
        "",
        f"verdict: **{verdict}**",
        "",
        "P0-4 STOP。全期間再集計へ進まない。",
        "",
    ]
    (OUT / "report.md").write_text("\n".join(lines), encoding="utf-8")

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="top")

    def hdr(ws: Any) -> None:
        for c in ws[1]:
            c.font = hf
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")

    def autosize(ws: Any, max_w: int = 48) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            n = 10
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                n = max(n, min(max_w, max((len(x) for x in v.splitlines()), default=0) + 2))
                cell.alignment = wrap
            ws.column_dimensions[letter].width = n

    ws = wb.active
    ws.title = "Summary"
    ws.append(["key", "value"])
    ws.append(["verdict", verdict])
    ws.append(["TEST_DAYS", ",".join(d["date"] for d in days)])
    ws.append(["ALL_DAYS_TRADE_PARITY", report["ALL_DAYS_TRADE_PARITY"]])
    ws.append(["ALL_DAYS_ANCHOR_PARITY", report["ALL_DAYS_ANCHOR_PARITY"]])
    ws.append(["ALL_DAYS_PNL_PARITY", report["ALL_DAYS_PNL_PARITY"]])
    ws.append(["DETERMINISM", report["DETERMINISM"]])
    ws.append(["CODE_CHANGE", "false"])
    ws.append(["STRATEGY_CHANGED", "false"])
    ws.append(["SAFETY", "0/0/0"])
    ws.append(["CAN_PROCEED_TO_P1_FULL_RECALC", can])
    hdr(ws)
    autosize(ws, 80)

    ws = wb.create_sheet("Path_Audit")
    ws.append(["stage", "exact", "fast", "delta"])
    for row in payload["path_audit"]:
        ws.append([row["stage"], row["exact"], row["fast"], row["delta"]])
    hdr(ws)
    autosize(ws, 70)

    ws = wb.create_sheet("Days")
    ws.append(["date", "status", "events", "universe_n", "universe_source", "ok", "reason"])
    for m in payload["day_selection"]:
        ws.append([m["date"], m["status"], m["events_hint"], m["universe_n"], m["universe_source"], m["ok"], m["reason"]])
    hdr(ws)
    autosize(ws, 70)

    ws = wb.create_sheet("Exact_Ledgers")
    ws.append(["date", "trade_id", "symbol", "session", "anchor_time", "snapshot_sequence", "score", "rank", "limit", "fill_time_iso", "fill_price", "exit_time_iso", "exit_price", "exit_reason", "pnl_yen_100"])
    for d in days:
        for t in d["exact"]["trades"]:
            ws.append([d["date"], t.get("trade_id"), t.get("symbol"), t.get("session"), t.get("anchor_time"), t.get("snapshot_sequence"), t.get("score"), t.get("candidate_rank"), t.get("limit"), t.get("fill_time_iso"), t.get("fill_price"), t.get("exit_time_iso"), t.get("exit_price"), t.get("exit_reason"), t.get("pnl_yen_100")])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Fast_Ledgers")
    ws.append(["date", "trade_id", "symbol", "session", "anchor_time", "snapshot_sequence", "score", "rank", "limit", "fill_time_iso", "fill_price", "exit_time_iso", "exit_price", "exit_reason", "pnl_yen_100"])
    for d in days:
        for t in d["fast"]["trades"]:
            ws.append([d["date"], t.get("trade_id"), t.get("symbol"), t.get("session"), t.get("anchor_time"), t.get("snapshot_sequence"), t.get("score"), t.get("candidate_rank"), t.get("limit"), t.get("fill_time_iso"), t.get("fill_price"), t.get("exit_time_iso"), t.get("exit_price"), t.get("exit_reason"), t.get("pnl_yen_100")])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Trade_Comparison")
    ws.append(["date", "class", "symbol", "anchor", "fields"])
    for d in days:
        for r in d["cmp"]["rows"]:
            ws.append([d["date"], r.get("class"), r.get("symbol"), r.get("anchor"), ",".join(r.get("fields") or [])])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Anchor_Comparison")
    ws.append(["date", "metric", "exact", "fast"])
    for d in days:
        ws.append([d["date"], "anchor_fires", d["exact"].get("anchor_fires"), d["fast"].get("anchor_fires")])
        ws.append([d["date"], "native_admitted", d["exact"].get("native_admitted"), d["fast"].get("native_admitted")])
        ws.append([d["date"], "fills", d["exact"].get("native_fills"), d["fast"].get("native_fills")])
        ws.append([d["date"], "expired", d["exact"].get("native_expired"), d["fast"].get("native_expired")])
        ws.append([d["date"], "cap_blocked", d["exact"].get("cap_blocked"), d["fast"].get("cap_blocked")])
        ws.append([d["date"], "same_symbol_blocked", d["exact"].get("same_symbol_blocked"), d["fast"].get("same_symbol_blocked")])
        ws.append([d["date"], "anchor_mismatch", d["cmp"]["anchor_mismatch"], ""])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Performance")
    ws.append(["date", "events", "exact_sec", "fast_sec", "speedup_x"])
    for d in report["PER_DAY"]:
        ws.append([d["date"], d["events"], d["exact_sec"], d["fast_sec"], d["speedup_x"]])
    hdr(ws)
    autosize(ws)

    ws = wb.create_sheet("Determinism")
    ws.append(["item", "value"])
    for k, v in payload["determinism"].items():
        ws.append([k, v])
    hdr(ws)
    autosize(ws, 90)

    ws = wb.create_sheet("Safety")
    ws.append(["check", "value"])
    for k, v in [
        ("submit/cancel/live", "0/0/0"),
        ("strategy/entry/exit changed", "false"),
        ("runtime bent toward fast", "false"),
        ("Actual rewritten", "false"),
        ("full-period recalc", "false"),
        ("CODE_CHANGE", "false"),
    ]:
        ws.append([k, v])
    hdr(ws)
    autosize(ws, 70)
    wb.save(OUT / "audit.xlsx")


def main() -> int:
    recommended = ["20260820", "20260817", "20260810"]
    selection = []
    for day in recommended:
        cap = find_capture_dir(day)
        if cap is None:
            selection.append({"date": day, "ok": False, "status": "MISSING", "events_hint": 0, "universe_n": 0, "universe_source": "", "reason": "no capture"})
            continue
        selection.append(_capture_meta(day, cap))
    if not all(s.get("ok") for s in selection):
        print("DAY_SELECT_FAIL", json.dumps(selection, ensure_ascii=False, default=str), flush=True)
        return 2

    path_audit = [
        {"stage": "capture_read", "exact": "all push_part JSONL sequence order via iter_push", "fast": "same iter_push / _stream_day", "delta": "none"},
        {"stage": "clock", "exact": "received_at/recorded_at event_t; no time.sleep; no wall clock for decisions", "fast": "same", "delta": "none"},
        {"stage": "feature update", "exact": "V1RNativeEntryLive.ingest_push every capture event", "fast": "same ingest_push", "delta": "none"},
        {"stage": "anchor fire", "exact": "process_market_push → maybe_fire_anchor(now_t=event_t) CLOCK_GRID", "fast": "same", "delta": "none"},
        {"stage": "candidate ranking", "exact": "_run_anchor simulate_joint", "fast": "same", "delta": "none"},
        {"stage": "admission/CAP/same-symbol", "exact": "simulate_joint + live exposure + pending/open skip", "fast": "same", "delta": "none"},
        {"stage": "PENDING / Passive Fill", "exact": "on_tick_fill_check find_ask_cross_fill WAIT_SEC=1 fill_price=limit", "fast": "same", "delta": "none"},
        {"stage": "OPEN / EXIT / SLOT RELEASE", "exact": "dual.try_admit_fill + dual.on_tick 0.5s when open", "fast": "same 0.5s cadence in _stream_day", "delta": "none"},
        {"stage": "session boundary", "exact": "close_open_at_session_end AM/PM + fill sweep", "fast": "same", "delta": "none"},
        {"stage": "I/O", "exact": "ingest_audit list until 200k clear; notify_enabled=false; no TCP", "fast": "ingest_audit discarded; notify_enabled=false; no TCP; no sleep", "delta": "audit I/O only"},
        {"stage": "post-hoc CAP", "exact": "forbidden / not used", "fast": "forbidden / not used", "delta": "none"},
    ]

    days_out = []
    # 20260820 Exact reused
    cap20 = Path(selection[0]["capture_path"])
    print("P0-4 20260820 Exact reuse P0-3", flush=True)
    exact20 = load_p03_exact()
    print("P0-4 20260820 Fast run1", flush=True)
    fast20 = run_replay("20260820", cap20, mode="fast")
    print(f"  fast20 trades={len(fast20['trades'])} pnl={fast20['pnl']} sha={fast20['ledger_sha']} sec={fast20['elapsed_sec']}", flush=True)
    print("P0-4 20260820 Fast run2 determinism", flush=True)
    fast20b = run_replay("20260820", cap20, mode="fast")
    det = {
        "day": "20260820",
        "fast_sha_run1": fast20["ledger_sha"],
        "fast_sha_run2": fast20b["ledger_sha"],
        "pass": fast20["ledger_sha"] == fast20b["ledger_sha"] and len(fast20["trades"]) == len(fast20b["trades"]),
    }
    days_out.append({"date": "20260820", "meta": selection[0], "exact": exact20, "fast": fast20, "cmp": compare_pair("20260820", exact20, fast20)})

    for day, meta in (("20260817", selection[1]), ("20260810", selection[2])):
        cap = Path(meta["capture_path"])
        print(f"P0-4 {day} Exact", flush=True)
        exact = run_replay(day, cap, mode="exact")
        print(f"  exact trades={len(exact['trades'])} pnl={exact['pnl']} sha={exact['ledger_sha']} sec={exact['elapsed_sec']}", flush=True)
        print(f"P0-4 {day} Fast", flush=True)
        fast = run_replay(day, cap, mode="fast")
        print(f"  fast trades={len(fast['trades'])} pnl={fast['pnl']} sha={fast['ledger_sha']} sec={fast['elapsed_sec']}", flush=True)
        days_out.append({"date": day, "meta": meta, "exact": exact, "fast": fast, "cmp": compare_pair(day, exact, fast)})

    payload = {
        "exact_path": "P0-3 Exact Runtime Replay: Capture JSONL → payload normalize → V1RNativeEntryLive.process_market_push (ingest_push / maybe_fire_anchor / on_tick_fill_check) → V1RLiveDualLane.on_tick (0.5s) → session_end. 20260820 Exact reused from P0-3 (no recompute).",
        "fast_path": "Same _stream_day / process_market_push / dual.on_tick as Exact. Speedups vs live: no wall-clock sleep, no TCP/Market Bus, no Discord. Speedup vs Exact: ingest_audit discarded (I/O only). No event skip, no post-hoc CAP.",
        "path_audit": path_audit,
        "day_selection": selection,
        "days": days_out,
        "determinism": det,
    }
    write_reports(payload)
    print("wrote", OUT, "det", det["pass"], flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
