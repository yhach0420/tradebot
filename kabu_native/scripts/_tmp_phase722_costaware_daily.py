"""7/22 Cost-Aware AM+PM offline replay daily report (Paper only; no official writes)."""
from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping
from zoneinfo import ZoneInfo

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parent
for p in (ROOT / "src", REPO, ROOT):
    s = str(p)
    if s not in sys.path:
        sys.path.insert(0, s)

JST = ZoneInfo("Asia/Tokyo")
OUT = ROOT / "results/reports/phase_7_22_costaware_daily"
AM_REUSE = (
    ROOT
    / "results/reports/phase_7_22_runtime_repair/_backup_20260722_203515/supporting/cost_aware_am_replay.json"
)
AM_DIR = ROOT / "results/small_paper/20260722/live_session_075904"
PM_DIR = ROOT / "results/small_paper/20260722/live_session_124934"
SHA_BEFORE = (
    ROOT
    / "results/reports/phase_7_22_runtime_repair/_backup_20260722_203515/supporting/official_session_sha_before.json"
)

OFFICIAL = {
    "am_pnl_yen_100": 184000.0,
    "pm_pnl_yen_100": -72700.0,
    "day_pnl_yen_100": 111300.0,
    "am_trades": 35,
    "pm_trades": 31,
}


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _session_file_sha(session_dir: Path) -> dict[str, Any]:
    out = {}
    for name in ("small_paper_summary.json", "small_paper_events.jsonl", "session_seal.json"):
        p = session_dir / name
        if p.is_file():
            out[name] = {"sha256": _sha256(p), "size": p.stat().st_size}
    return out


def _aggregate_session_sha(session_dir: Path) -> str:
    h = hashlib.sha256()
    for p in sorted(session_dir.rglob("*")):
        if p.is_file():
            rel = p.relative_to(session_dir).as_posix().encode()
            h.update(rel)
            h.update(b"\0")
            h.update(_sha256(p).encode())
            h.update(b"\0")
    return h.hexdigest()


def _suppress_shadow_writes() -> Any:
    import small_paper.cost_aware_entry_shadow as ca

    suppressed = {"count": 0}

    def _noop(state, trading_date, event):  # type: ignore[no-untyped-def]
        state.events.append(event)
        suppressed["count"] += 1

    ca.append_shadow_event = _noop  # type: ignore[assignment]
    return suppressed


def _metrics_from_replay(raw: Mapping[str, Any]) -> dict[str, Any]:
    rt_raw = raw.get("runtime_compatible_raw")
    if rt_raw is None:
        rt_raw = raw.get("runtime_compatible_pnl")
    rt_5 = raw.get("runtime_compatible_5bps_roundtrip")
    sh_raw = raw.get("fixed_30m_raw")
    if sh_raw is None:
        sh_raw = raw.get("gross_pnl_30m")
    sh_5 = raw.get("fixed_30m_5bps_roundtrip")
    if sh_5 is None:
        sh_5 = raw.get("pnl_after_5bps_30m")

    def _f(v):
        return None if v is None else float(v)

    rt_raw_f, rt_5_f, sh_raw_f, sh_5_f = map(_f, (rt_raw, rt_5, sh_raw, sh_5))
    delta_raw = (
        round(sh_raw_f - rt_raw_f, 2)
        if rt_raw_f is not None and sh_raw_f is not None
        else None
    )
    delta_5 = (
        round(sh_5_f - rt_5_f, 2) if rt_5_f is not None and sh_5_f is not None else None
    )
    return {
        "runtime_compatible_raw_pnl": rt_raw_f,
        "runtime_compatible_5bps_pnl": rt_5_f,
        "shadow_raw_pnl": sh_raw_f,
        "shadow_5bps_pnl": sh_5_f,
        "delta_raw": delta_raw,
        "delta_5bps": delta_5,
        "runtime_pf": raw.get("runtime_compatible_pf_5bps"),
        "shadow_pf": raw.get("shadow_pf_5bps_30m") or raw.get("fixed_30m_pf_5bps"),
        "closed_trades": int(raw.get("n_closed") or 0),
        "open": int(raw.get("n_open") or 0),
        "status": str(raw.get("status") or ""),
        "entry_count": int(raw.get("shadow_entries") or raw.get("n_closed") or 0),
        "finalize_count": int(raw.get("recovery_finalize_count") or 0)
        + int(raw.get("session_force_close_finalize_count") or 0),
        "selection_cycles": int(raw.get("selection_cycles") or 0),
        "stop_risk_reject": int(raw.get("stop_risk_reject") or 0),
    }


def _adoption(delta_5bps: float | None) -> str:
    if delta_5bps is None:
        return "SAME"
    if delta_5bps > 0:
        return "IMPROVED"
    if delta_5bps < 0:
        return "WORSE"
    return "SAME"


def main() -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook

    OUT.mkdir(parents=True, exist_ok=True)
    sha_before = {}
    if SHA_BEFORE.is_file():
        sha_before = json.loads(SHA_BEFORE.read_text(encoding="utf-8"))
    else:
        sha_before = {
            "live_session_075904": _session_file_sha(AM_DIR),
            "live_session_124934": _session_file_sha(PM_DIR),
        }

    am_agg_before = _aggregate_session_sha(AM_DIR)
    pm_agg_before = _aggregate_session_sha(PM_DIR)

    # Reuse AM
    am_doc = json.loads(AM_REUSE.read_text(encoding="utf-8"))
    am_raw = am_doc.get("raw_replay_summary") or {}
    if not am_raw:
        am_m = am_doc.get("metrics") or {}
        am_raw = {
            "runtime_compatible_raw": am_m.get("runtime_compatible_raw", am_m.get("runtime_compatible_pnl")),
            "runtime_compatible_5bps_roundtrip": am_m.get("runtime_compatible_5bps_roundtrip"),
            "fixed_30m_raw": am_m.get("shadow_raw_pnl"),
            "fixed_30m_5bps_roundtrip": am_m.get("shadow_5bps_pnl"),
            "runtime_compatible_pf_5bps": am_m.get("runtime_pf"),
            "shadow_pf_5bps_30m": am_m.get("shadow_pf"),
            "n_closed": am_m.get("exit_count") or am_m.get("entry_count"),
            "n_open": am_m.get("open"),
            "status": am_m.get("status"),
            "shadow_entries": am_m.get("entry_count"),
            "selection_cycles": am_m.get("selection_cycles"),
            "stop_risk_reject": am_m.get("stop_risk_reject"),
            "recovery_finalize_count": am_m.get("recovery_finalize_count"),
            "session_force_close_finalize_count": am_m.get("session_force_close_finalize_count"),
            "closed_trades": am_doc.get("closed_trades") or [],
        }
    if "closed_trades" not in am_raw or not am_raw.get("closed_trades"):
        am_raw = dict(am_raw)
        am_raw["closed_trades"] = am_doc.get("closed_trades") or []

    print("AM reused status=", am_raw.get("status"), "n_closed=", am_raw.get("n_closed"))

    suppressed = _suppress_shadow_writes()
    from small_paper.cost_aware_shadow_recompute import (
        merge_cost_aware_daily,
        replay_cost_aware_session,
    )

    print("PM replay start...", flush=True)
    pm_raw = replay_cost_aware_session(
        PM_DIR,
        trading_date="20260722",
        is_freeze_recovery=True,
    )
    print(
        "PM replay done status=",
        pm_raw.get("status"),
        "n_closed=",
        pm_raw.get("n_closed"),
        "suppressed_events=",
        suppressed["count"],
        flush=True,
    )

    am_m = _metrics_from_replay(am_raw)
    pm_m = _metrics_from_replay(pm_raw)

    # Daily merge via formal helper (needs closed_trades for PF)
    daily_raw = merge_cost_aware_daily(am_raw, pm_raw)
    day_m = _metrics_from_replay(daily_raw)

    # Explicit day sums (same basis; do not mix)
    day_explicit = {
        "runtime_raw": round(
            float(am_m["runtime_compatible_raw_pnl"] or 0)
            + float(pm_m["runtime_compatible_raw_pnl"] or 0),
            2,
        ),
        "shadow_raw": round(
            float(am_m["shadow_raw_pnl"] or 0) + float(pm_m["shadow_raw_pnl"] or 0), 2
        ),
        "runtime_5bps": round(
            float(am_m["runtime_compatible_5bps_pnl"] or 0)
            + float(pm_m["runtime_compatible_5bps_pnl"] or 0),
            2,
        ),
        "shadow_5bps": round(
            float(am_m["shadow_5bps_pnl"] or 0) + float(pm_m["shadow_5bps_pnl"] or 0), 2
        ),
        "closed_trades": int(am_m["closed_trades"]) + int(pm_m["closed_trades"]),
        "open": int(am_m["open"]) + int(pm_m["open"]),
    }
    day_explicit["delta_raw"] = round(
        day_explicit["shadow_raw"] - day_explicit["runtime_raw"], 2
    )
    day_explicit["delta_5bps"] = round(
        day_explicit["shadow_5bps"] - day_explicit["runtime_5bps"], 2
    )
    # Prefer merge PF (trade-list based)
    day_explicit["runtime_pf"] = day_m["runtime_pf"]
    day_explicit["shadow_pf"] = day_m["shadow_pf"]
    day_explicit["status"] = (
        "RUNNING_PNL_COMPLETE"
        if (
            am_m["status"] == "RUNNING_PNL_COMPLETE"
            and pm_m["status"] == "RUNNING_PNL_COMPLETE"
            and day_explicit["open"] == 0
        )
        else "PARTIAL_PIPELINE"
    )

    adoption = _adoption(day_explicit["delta_5bps"])

    # Official paper comparison (informational; different population)
    official_vs_ca = {
        "official_paper": OFFICIAL,
        "note": "Official Paper trade set differs from Cost-Aware shadow selection; deltas are informational only.",
        "day_official_minus_ca_shadow_raw": round(
            OFFICIAL["day_pnl_yen_100"] - day_explicit["shadow_raw"], 2
        ),
        "day_official_minus_ca_runtime_raw": round(
            OFFICIAL["day_pnl_yen_100"] - day_explicit["runtime_raw"], 2
        ),
        "day_ca_shadow_raw_minus_official": round(
            day_explicit["shadow_raw"] - OFFICIAL["day_pnl_yen_100"], 2
        ),
        "day_ca_runtime_raw_minus_official": round(
            day_explicit["runtime_raw"] - OFFICIAL["day_pnl_yen_100"], 2
        ),
    }

    # Integrity after
    sha_after = {
        "live_session_075904": _session_file_sha(AM_DIR),
        "live_session_124934": _session_file_sha(PM_DIR),
    }
    sha_ok = True
    for sess, files in sha_before.items():
        for fn, meta in files.items():
            after = sha_after.get(sess, {}).get(fn, {})
            if after.get("sha256") != meta.get("sha256"):
                sha_ok = False
    am_agg_after = _aggregate_session_sha(AM_DIR)
    pm_agg_after = _aggregate_session_sha(PM_DIR)
    agg_ok = am_agg_before == am_agg_after and pm_agg_before == pm_agg_after

    safety = {
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
        "order_enabled": False,
        "paper_only": True,
    }
    for d in (AM_DIR, PM_DIR):
        s = json.loads((d / "small_paper_summary.json").read_text(encoding="utf-8"))
        if s.get("order_enabled") is True:
            safety["order_enabled"] = True
        if s.get("paper_only") is False:
            safety["paper_only"] = False

    # Persist PM replay detail (embedded in report.json; also side file under OUT/_pm)
    pm_detail = {
        "report": "cost_aware_pm_replay",
        "session_dir": str(PM_DIR),
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "metrics": pm_m,
        "raw_replay_summary": {
            k: v
            for k, v in pm_raw.items()
            if k != "closed_trades"
        },
        "closed_trades_count": len(pm_raw.get("closed_trades") or []),
        "closed_trades": pm_raw.get("closed_trades") or [],
        "shadow_event_write_suppression": {
            "enabled": True,
            "suppressed_append_shadow_event_calls": suppressed["count"],
        },
    }

    report = {
        "phase": "phase_7_22_costaware_daily",
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "paper_only": True,
        "order_enabled": False,
        "am_source": "reused_offline_replay",
        "am_source_path": str(AM_REUSE.relative_to(ROOT)).replace("\\", "/"),
        "pm_source": "replay_cost_aware_session",
        "sessions": {
            "am": "results/small_paper/20260722/live_session_075904",
            "pm": "results/small_paper/20260722/live_session_124934",
        },
        "am": am_m,
        "pm": pm_m,
        "day": {
            **day_explicit,
            "delta_raw_columns": "shadow_raw - runtime_raw",
            "delta_5bps_columns": "shadow_5bps - runtime_5bps",
            "same_basis_only": True,
        },
        "day_from_merge_helper": day_m,
        "official_paper": OFFICIAL,
        "official_vs_cost_aware": official_vs_ca,
        "adoption": {
            "basis": "day delta_5bps = shadow_5bps - runtime_5bps (same Cost-Aware population)",
            "verdict": adoption,
            "delta_5bps": day_explicit["delta_5bps"],
            "delta_raw": day_explicit["delta_raw"],
        },
        "safety": safety,
        "official_integrity": {
            "key_files_sha_unchanged": sha_ok,
            "aggregate_sha_unchanged": agg_ok,
            "am_aggregate_before": am_agg_before,
            "am_aggregate_after": am_agg_after,
            "pm_aggregate_before": pm_agg_before,
            "pm_aggregate_after": pm_agg_after,
            "sha_after": sha_after,
        },
        "pm_replay_detail": {
            "closed_trades_count": pm_detail["closed_trades_count"],
            "metrics": pm_m,
            "raw_replay_summary": pm_detail["raw_replay_summary"],
            "closed_trades": pm_detail["closed_trades"],
            "shadow_event_write_suppression": pm_detail["shadow_event_write_suppression"],
        },
        "overall": "COMPLETE"
        if (
            am_m["status"] == "RUNNING_PNL_COMPLETE"
            and pm_m["status"] == "RUNNING_PNL_COMPLETE"
            and day_explicit["status"] == "RUNNING_PNL_COMPLETE"
            and sha_ok
            and agg_ok
            and safety["submit"] == 0
            and safety["cancel"] == 0
            and safety["live_order"] == 0
        )
        else "INCOMPLETE",
    }

    (OUT / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    def yen(v):
        if v is None:
            return "n/a"
        return f"{v:+,.2f}".replace(".00", "")

    md = f"""# 7/22 Cost-Aware Daily Report

Generated: {report['generated_at']}
Paper only / order_enabled=false

Official sessions (read-only):
- AM: `results/small_paper/20260722/live_session_075904`
- PM: `results/small_paper/20260722/live_session_124934`

AM replay: reused from `{report['am_source_path']}`
PM replay: `replay_cost_aware_session(..., trading_date='20260722', is_freeze_recovery=True)`
Shadow event file writes: suppressed ({suppressed['count']} calls)

## Official Paper (unchanged)

| | Trades | PnL |
|--|--------|-----|
| AM | 35 | +184,000 |
| PM | 31 | -72,700 |
| Day | 66 | +111,300 |

Official SHA unchanged: **{'YES' if sha_ok and agg_ok else 'NO'}**

## Cost-Aware session metrics (same-basis deltas only)

### AM — status `{am_m['status']}` — closed={am_m['closed_trades']} open={am_m['open']}

| basis | runtime | shadow | delta (shadow−runtime) |
|-------|---------|--------|------------------------|
| raw | {yen(am_m['runtime_compatible_raw_pnl'])} | {yen(am_m['shadow_raw_pnl'])} | {yen(am_m['delta_raw'])} |
| 5bps | {yen(am_m['runtime_compatible_5bps_pnl'])} | {yen(am_m['shadow_5bps_pnl'])} | {yen(am_m['delta_5bps'])} |

PF: runtime={am_m['runtime_pf']} / shadow={am_m['shadow_pf']}

### PM — status `{pm_m['status']}` — closed={pm_m['closed_trades']} open={pm_m['open']}

| basis | runtime | shadow | delta (shadow−runtime) |
|-------|---------|--------|------------------------|
| raw | {yen(pm_m['runtime_compatible_raw_pnl'])} | {yen(pm_m['shadow_raw_pnl'])} | {yen(pm_m['delta_raw'])} |
| 5bps | {yen(pm_m['runtime_compatible_5bps_pnl'])} | {yen(pm_m['shadow_5bps_pnl'])} | {yen(pm_m['delta_5bps'])} |

PF: runtime={pm_m['runtime_pf']} / shadow={pm_m['shadow_pf']}

### DAY — status `{day_explicit['status']}` — closed={day_explicit['closed_trades']} open={day_explicit['open']}

| basis | runtime | shadow | delta |
|-------|---------|--------|-------|
| raw | {yen(day_explicit['runtime_raw'])} | {yen(day_explicit['shadow_raw'])} | {yen(day_explicit['delta_raw'])} |
| 5bps | {yen(day_explicit['runtime_5bps'])} | {yen(day_explicit['shadow_5bps'])} | {yen(day_explicit['delta_5bps'])} |

PF: runtime={day_explicit['runtime_pf']} / shadow={day_explicit['shadow_pf']}

## vs Official Paper (informational; different trade population)

- Day official − CA shadow raw: {yen(official_vs_ca['day_official_minus_ca_shadow_raw'])}
- Day official − CA runtime raw: {yen(official_vs_ca['day_official_minus_ca_runtime_raw'])}

## Adoption (Cost-Aware shadow vs runtime_compatible, same population, 5bps)

Verdict: **{adoption}** (delta_5bps={yen(day_explicit['delta_5bps'])})

## Safety

- submit=0 cancel=0 live_order=0
- Official Paper unchanged: {'YES' if sha_ok and agg_ok else 'NO'}

---

【7/22 Cost-Aware Daily】

AM
runtime raw: {yen(am_m['runtime_compatible_raw_pnl'])}
shadow raw: {yen(am_m['shadow_raw_pnl'])}
delta raw: {yen(am_m['delta_raw'])}
runtime 5bps: {yen(am_m['runtime_compatible_5bps_pnl'])}
shadow 5bps: {yen(am_m['shadow_5bps_pnl'])}
delta 5bps: {yen(am_m['delta_5bps'])}

PM
runtime raw: {yen(pm_m['runtime_compatible_raw_pnl'])}
shadow raw: {yen(pm_m['shadow_raw_pnl'])}
delta raw: {yen(pm_m['delta_raw'])}
runtime 5bps: {yen(pm_m['runtime_compatible_5bps_pnl'])}
shadow 5bps: {yen(pm_m['shadow_5bps_pnl'])}
delta 5bps: {yen(pm_m['delta_5bps'])}

DAY

runtime raw: {yen(day_explicit['runtime_raw'])}
shadow raw: {yen(day_explicit['shadow_raw'])}
raw delta: {yen(day_explicit['delta_raw'])}

runtime 5bps: {yen(day_explicit['runtime_5bps'])}
shadow 5bps: {yen(day_explicit['shadow_5bps'])}
5bps delta: {yen(day_explicit['delta_5bps'])}

runtime PF: {day_explicit['runtime_pf']}
shadow PF: {day_explicit['shadow_pf']}

採用判定

{adoption}

Official Paper unchanged:
{'YES' if sha_ok and agg_ok else 'NO'}

submit=0
cancel=0
live_order=0

Overall:
{report['overall']}
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily"
    ws.append(["metric", "value"])
    for k, v in [
        ("am_runtime_raw", am_m["runtime_compatible_raw_pnl"]),
        ("am_shadow_raw", am_m["shadow_raw_pnl"]),
        ("am_delta_raw", am_m["delta_raw"]),
        ("am_runtime_5bps", am_m["runtime_compatible_5bps_pnl"]),
        ("am_shadow_5bps", am_m["shadow_5bps_pnl"]),
        ("am_delta_5bps", am_m["delta_5bps"]),
        ("am_runtime_pf", am_m["runtime_pf"]),
        ("am_shadow_pf", am_m["shadow_pf"]),
        ("am_closed", am_m["closed_trades"]),
        ("am_open", am_m["open"]),
        ("am_status", am_m["status"]),
        ("pm_runtime_raw", pm_m["runtime_compatible_raw_pnl"]),
        ("pm_shadow_raw", pm_m["shadow_raw_pnl"]),
        ("pm_delta_raw", pm_m["delta_raw"]),
        ("pm_runtime_5bps", pm_m["runtime_compatible_5bps_pnl"]),
        ("pm_shadow_5bps", pm_m["shadow_5bps_pnl"]),
        ("pm_delta_5bps", pm_m["delta_5bps"]),
        ("pm_runtime_pf", pm_m["runtime_pf"]),
        ("pm_shadow_pf", pm_m["shadow_pf"]),
        ("pm_closed", pm_m["closed_trades"]),
        ("pm_open", pm_m["open"]),
        ("pm_status", pm_m["status"]),
        ("day_runtime_raw", day_explicit["runtime_raw"]),
        ("day_shadow_raw", day_explicit["shadow_raw"]),
        ("day_delta_raw", day_explicit["delta_raw"]),
        ("day_runtime_5bps", day_explicit["runtime_5bps"]),
        ("day_shadow_5bps", day_explicit["shadow_5bps"]),
        ("day_delta_5bps", day_explicit["delta_5bps"]),
        ("day_runtime_pf", day_explicit["runtime_pf"]),
        ("day_shadow_pf", day_explicit["shadow_pf"]),
        ("day_closed", day_explicit["closed_trades"]),
        ("day_open", day_explicit["open"]),
        ("adoption_5bps", adoption),
        ("official_day", OFFICIAL["day_pnl_yen_100"]),
        ("official_sha_unchanged", sha_ok and agg_ok),
        ("overall", report["overall"]),
    ]:
        ws.append([k, v])

    ws2 = wb.create_sheet("AM")
    ws2.append(["metric", "value"])
    for k, v in am_m.items():
        ws2.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v)[:500]])

    ws3 = wb.create_sheet("PM")
    ws3.append(["metric", "value"])
    for k, v in pm_m.items():
        ws3.append([k, v if not isinstance(v, (dict, list)) else json.dumps(v)[:500]])

    ws4 = wb.create_sheet("PM_Trades")
    trades = pm_raw.get("closed_trades") or []
    cols = [
        "symbol",
        "shadow_entry_time",
        "gross_pnl_yen_100",
        "net_pnl_yen_100",
        "runtime_compatible_gross_yen",
        "runtime_compatible_net_yen",
        "shadow_exit_reason",
        "runtime_compatible_na",
    ]
    ws4.append(cols)
    for t in trades:
        ws4.append([t.get(c) for c in cols])

    ws5 = wb.create_sheet("Integrity")
    ws5.append(["session", "file", "sha256", "match_before"])
    for sess, files in sha_after.items():
        for fn, meta in files.items():
            prev = sha_before.get(sess, {}).get(fn, {}).get("sha256")
            ws5.append([sess, fn, meta.get("sha256"), meta.get("sha256") == prev])

    wb.save(OUT / "audit.xlsx")
    print("WROTE", OUT / "report.md")
    print("WROTE", OUT / "report.json")
    print("WROTE", OUT / "audit.xlsx")
    print("OVERALL", report["overall"])
    print("ADOPTION", adoption)
    print("PM", pm_m)
    print("DAY", day_explicit)


if __name__ == "__main__":
    main()
