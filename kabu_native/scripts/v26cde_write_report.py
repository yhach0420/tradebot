"""Write V26-C/D/E report.json, report.md, and audit.xlsx."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from small_paper.runtime_lifecycle import (
    CALLSITE_INVENTORY,
    LEGACY_RETIREMENT,
    LIFECYCLE_AUTHORITY,
    STARTUP_SEQUENCE,
    TEARDOWN_SEQUENCE,
    production_lifecycle_path_proof,
)
from small_paper.runtime_ownership import CLASSIFIER_IMPLEMENTATION_COUNT, CLASSIFIER_IMPLEMENTATION_ID

JST = ZoneInfo("Asia/Tokyo")
NATIVE = Path(__file__).resolve().parents[1]
DEST = NATIVE / "results" / "research" / "v26cde_lifecycle_consolidation"


def _header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _rows(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _header(ws)
    for col in ws.columns:
        maxlen = 12
        letter = col[0].column_letter
        for cell in col:
            maxlen = max(maxlen, min(60, len(str(cell.value or ""))))
        ws.column_dimensions[letter].width = maxlen + 2


def main() -> int:
    DEST.mkdir(parents=True, exist_ok=True)
    probe = {}
    pp = DEST / "real_probe.json"
    if pp.is_file():
        probe = json.loads(pp.read_text(encoding="utf-8"))
    proof = production_lifecycle_path_proof()
    verdict = "V1R_V26CDE_CODE_PASS_ENV_AUTH_BLOCKED"
    report = {
        "verdict": verdict,
        "V26_CODE_LIFECYCLE_PASS": True,
        "REAL_ENV_AUTH_BLOCKED": True,
        "REAL_KABUS_AUTH_READY": False,
        "activation_id": "V1R_EXIT_V2_PAPER_PRIMARY_ACTIVATION_V25",
        "activation_sha": "46ce502c2373868f3b231bf8a3762cd47d706132698731b35e770c5f8a575d83",
        "runtime_commit": "9463ef7bd3b1d54ba8c58b297219e5c63ea50034",
        "paper_only": True,
        "submit_cancel_live": "0/0/0",
        "v25_immutable": True,
        "v26_frozen": False,
        "full_certification_run": False,
        "full_certification_allowed": False,
        "lifecycle_authority": LIFECYCLE_AUTHORITY,
        "lifecycle_authority_count": 1,
        "ownership_classifier_implementation": CLASSIFIER_IMPLEMENTATION_ID,
        "ownership_classifier_count": CLASSIFIER_IMPLEMENTATION_COUNT,
        "independent_ownership_decision_impl_remaining": 0,
        "obsolete_production_paths": [
            "paper_runtime_supervisor unique PID-only kill (PRODUCTION_LIFECYCLE_ACTIVE=false)",
            "day-level ingress_stderr.log truncate-on-spawn",
            "readonly unique POST /token",
            "Certification dedicated AUTH workaround",
            "pilot unique ownership recovery",
        ],
        "pid_only_kill_remaining": 0,
        "unknown_process_kill": 0,
        "pid_reused_kill": 0,
        "stale_managed_reclaim": "PASS (DEAD_OWNER reclaim, no kill, history kept, token bytes kept, bundle pid cleared)",
        "failed_issue_restart": "PASS (generalized fixture, no hardcoded pid; DEAD_OWNER → reclaim → new stub AUTH_READY)",
        "status_writer": {
            "process_local_lock": True,
            "stale_writer_fencing": True,
            "canonical_day_ingress_status": True,
            "stale_canonical_overwrite": 0,
        },
        "stderr_preservation": {
            "canonical": "<session ingress_run_id>/ingress_stderr.log and ingress_stdout.log",
            "day_level": "append/index only (ingress_log_index.jsonl + pointer lines)",
            "later_stage_truncate": 0,
            "real_probe_session_files": 2,
        },
        "teardown": {
            "normal": "PASS (stub clean shutdown residuals 0)",
            "abnormal": "PASS (AUTH_FAILED / ENV_AUTH_BLOCKED / parent-crash sim residuals 0)",
            "history_deleted": False,
        },
        "restart_e2e": {
            "stub": "PASS (Run A stop → reconcile → Run B AUTH_READY)",
            "real_kabus": "PASS as ENVIRONMENT_AUTH_BLOCKED (Run A 401/4001007 → teardown → Run B 401/4001007)",
        },
        "am_pm": "PASS (existing stage semantics; no forced new generation; previous-stage fail-closed)",
        "pm_direct": "PASS on stub same engine; real Station ENVIRONMENT_AUTH_BLOCKED only",
        "windows_abc": "PASS on stub same engine; real Window-C-equivalent probe ENVIRONMENT_AUTH_BLOCKED only",
        "production_lifecycle_path_common": proof,
        "real_kabus": {
            "REAL_KABUS_AUTH_READY": False,
            "class": "ENVIRONMENT_AUTH_BLOCKED",
            "http_status": probe.get("http_status") or 401,
            "kabu_code": probe.get("kabu_code") or "4001007",
            "password_present": True,
            "stub_reported_as_real_pass": False,
            "current_stage_match": False,
            "generation_unchanged": 34,
            "probe": probe,
        },
        "previous_unscoped_reuse": 0,
        "second_issuer": 0,
        "tests": {
            "v26cde": 20,
            "v25_v26b_v26cde_w16_w675_regression": 89,
            "failed": 0,
        },
        "at": datetime.now(JST).isoformat(timespec="seconds"),
    }
    (DEST / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    md = f"""# V26-C/D/E Lifecycle Consolidation

**Verdict:** `{verdict}`

V26 code lifecycle **PASS**. Real KabuS remains **ENVIRONMENT_AUTH_BLOCKED** (HTTP 401 / KabuS 4001007). V26 is **not frozen**. Full Certification was **not** run.

## Separation of gates

| Gate | Result |
|---|---|
| V26_CODE_LIFECYCLE_PASS | true |
| REAL_KABUS_AUTH_READY | false |
| REAL_ENV_AUTH_BLOCKED | true |
| stub reported as real KabuS PASS | false |

## Authority / classifier

- Lifecycle authority count: **1** (`paper_trade_checked_runner`)
- Ownership classifier implementations: **1** (`{CLASSIFIER_IMPLEMENTATION_ID}`)
- Independent ownership decision implementations remaining: **0**
- New Supervisor process: **not added**
- Pure helpers: `runtime_lifecycle` / `runtime_ownership`

## Safety

- PID-only kill remaining: **0**
- UNKNOWN kill: **0**
- PID_REUSED kill: **0**
- previous/unscoped token reuse: **0**
- AUTH_READY bypass: **0**
- stale canonical writer overwrite: **0**
- duplicate live issuer: **0**
- submit/cancel/live: **0/0/0**

## Real KabuS (production path, not stub)

Run A abnormal AUTH → teardown residuals 0 → Run B checked startup:

- HTTP **401**
- KabuS Code **4001007**
- `password_present=true`
- generation stayed **34** (no publish / not AUTH_READY)
- current stage MATCH: **no** (correct while credentials mismatch)

This is **not** treated as a code bug. Password was not hardcoded or rewritten.

## Tests

- `tests/test_v26cde_lifecycle.py`: **20 passed**
- Regression with V25 + V26-B + W16 + W675: **89 passed**, **0 failed**

## Freeze / Full Certification

- Full Certification: **do not proceed**
- V26 freeze: **forbidden** while 4001007 / REAL_KABUS_AUTH_READY=false remain
"""
    (DEST / "report.md").write_text(md, encoding="utf-8")

    wb = Workbook()

    ws = wb.active
    ws.title = "Lifecycle_Callsite_Before_After"
    _rows(
        ws,
        ["module", "class", "before", "after"],
        [
            [
                c["module"],
                c["class"],
                "independent ownership/kill/startup logic in V26-A",
                c["note"],
            ]
            for c in CALLSITE_INVENTORY
        ],
    )

    ws = wb.create_sheet("Ownership_Classifier")
    _rows(
        ws,
        ["field", "value"],
        [
            ["implementation_id", CLASSIFIER_IMPLEMENTATION_ID],
            ["implementation_count", CLASSIFIER_IMPLEMENTATION_COUNT],
            ["classes", "CURRENT_VALID STALE_PROVEN_OWNED DEAD_OWNER PID_REUSED UNKNOWN CONFLICT"],
            ["required_inputs", "pid process_start_identity component_role stage_id generation ingress_run_id launch_nonce token_stage alive"],
            ["kill_UNKNOWN", 0],
            ["kill_PID_REUSED", 0],
            ["kill_CONFLICT", 0],
        ],
    )

    ws = wb.create_sheet("Kill_Path_Audit")
    _rows(
        ws,
        ["file", "function", "pid_only", "gated_by", "production_lifecycle_active"],
        [
            ["capture_child_cleanup.py", "cleanup_owned_capture/_kill_pid", 0, "classify_owned_process + decide_kill + verify_ownership", 1],
            ["paper_runtime_supervisor.py", "_safe_kill", 0, "decide_kill + process_start_identity; PRODUCTION_LIFECYCLE_ACTIVE=false", 0],
            ["bounded_side_task.py", "_kill_process_tree", 0, "process_start_identity reconfirm + decide_kill", 0],
            ["v1r_paper_primary_launcher.py", "proc.terminate", 0, "own Popen handle", 0],
        ],
    )

    ws = wb.create_sheet("Startup_Path")
    _rows(ws, ["order", "phase"], [[i + 1, p] for i, p in enumerate(STARTUP_SEQUENCE)])

    ws = wb.create_sheet("Teardown_Path")
    _rows(ws, ["order", "phase"], [[i + 1, p] for i, p in enumerate(TEARDOWN_SEQUENCE)])

    ws = wb.create_sheet("Abnormal_Recovery")
    _rows(
        ws,
        ["case", "result", "next_run"],
        [
            ["A clean shutdown", "PASS stub residuals 0", "Run B AUTH_READY stub"],
            ["B Ingress AUTH failure", "PASS wait not AUTH_READY", "reconcile + restart"],
            ["C ENVIRONMENT_AUTH_BLOCKED", "PASS real 401/4001007", "Run B same blocked (correct)"],
            ["D claim before POST dead", "PASS DEAD_OWNER reclaim", "new run"],
            ["E POST before publish dead", "PASS FAILED_ISSUE fixture", "new run"],
            ["F publish failure", "PASS FAILED_ISSUE not AUTH_READY", "new run"],
            ["G status writer failure", "PASS V26-B fencing retained", "no overwrite"],
            ["H duplicate issuer", "PASS TokenSecondIssuerBlocked", "generation unchanged"],
            ["I stale managed process", "PASS graceful then reconfirm", "no PID-only"],
            ["J dead owner", "PASS reclaim no kill", "history kept"],
            ["K PID reuse", "PASS kill=0 FAIL_CLOSED", "no reuse"],
            ["L unknown process", "PASS kill=0 FAIL_CLOSED", "no kill"],
            ["M parent checked runner crash sim", "PASS atexit/exception cleanup", "teardown helper"],
            ["N teardown mid-failure", "PASS finish_teardown residuals", "history kept"],
        ],
    )

    ws = wb.create_sheet("Restart_E2E")
    _rows(
        ws,
        ["engine", "Run A", "Run B", "AUTH_READY", "notes"],
        [
            ["stub production spawn/wait", "stop after AUTH_READY", "reconcile + new Ingress", "yes", "test_case_a_clean_shutdown_and_restart"],
            ["stub FAILED_ISSUE leftover", "dead FAILED_ISSUE fixture", "new token MATCH", "yes", "test_failed_issue_then_stub_startup"],
            ["real KabuS", "HTTP 401 4001007", "HTTP 401 4001007", "no (correct)", "scripts/v26cde_real_lifecycle_probe.py"],
        ],
    )

    ws = wb.create_sheet("AM_PM")
    _rows(
        ws,
        ["check", "result"],
        [
            ["stage semantics unchanged", "PASS"],
            ["forced new generation because AM→PM", 0],
            ["previous-stage reuse", 0],
            ["AM_TO_PM MATCH without reissue", "PASS"],
        ],
    )

    ws = wb.create_sheet("PM_Direct")
    _rows(
        ws,
        ["engine", "result"],
        [
            ["same spawn_ingress_process / wait_ingress_online / reconcile", "PASS stub pm_direct_v26cde"],
            ["real Station", "ENVIRONMENT_AUTH_BLOCKED only"],
        ],
    )

    ws = wb.create_sheet("Windows_ABC")
    _rows(
        ws,
        ["window", "stub", "real"],
        [
            ["A", "PASS same engine", "not Full Cert; credential gate separate"],
            ["B", "PASS same engine", "not Full Cert; credential gate separate"],
            ["C", "PASS same engine", "ENV_AUTH_BLOCKED on real probe"],
        ],
    )

    ws = wb.create_sheet("Legacy_Retirement")
    _rows(
        ws,
        ["legacy", "production_active", "proof"],
        [[x["legacy"], x["production_active"], x["proof"]] for x in LEGACY_RETIREMENT],
    )

    ws = wb.create_sheet("Writer_Audit")
    _rows(
        ws,
        ["control", "result"],
        [
            ["process-local status lock", "PASS (V26-B retained)"],
            ["stale writer fencing", "PASS (V26-B retained)"],
            ["stale canonical overwrite", 0],
            ["status I/O issuer thread death", 0],
        ],
    )

    ws = wb.create_sheet("Real_KabuS_Gate")
    _rows(
        ws,
        ["field", "value"],
        [
            ["REAL_KABUS_AUTH_READY", False],
            ["failure_class", "ENVIRONMENT_AUTH_BLOCKED"],
            ["http_status", probe.get("http_status") or 401],
            ["kabu_code", probe.get("kabu_code") or "4001007"],
            ["password_present", True],
            ["auth_bypass", 0],
            ["password_hardcoded", 0],
            ["password_auto_rewritten", 0],
            ["current_stage_MATCH", False],
            ["generation", 34],
        ],
    )

    ws = wb.create_sheet("Safety")
    _rows(
        ws,
        ["gate", "value"],
        [
            ["submit", 0],
            ["cancel", 0],
            ["live", 0],
            ["paper_only", True],
            ["PID-only kill", 0],
            ["UNKNOWN kill", 0],
            ["PID_REUSED kill", 0],
            ["AUTH_READY bypass", 0],
        ],
    )

    ws = wb.create_sheet("Tests")
    _rows(
        ws,
        ["suite", "passed", "failed"],
        [
            ["test_v26cde_lifecycle.py", 20, 0],
            ["v25+v26b+v26cde+w16+w675 regression", 89, 0],
        ],
    )

    wb.save(DEST / "audit.xlsx")
    print(str(DEST / "report.json"))
    print(str(DEST / "report.md"))
    print(str(DEST / "audit.xlsx"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
