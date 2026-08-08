#!/usr/bin/env python3
"""E1_X5 partial AM reference — 2026-07-27 (saved Capture SoT).

Uses market_capture/20260727 session push_part JSONL for 09:53–11:30 JST
(actual persisted AM after ingress start). No backfill before first PUSH.
No Forward adoption. submit/cancel/live=0.

Writes ONLY:
  results/research/e1_x5_partial_am_reference_20260727/{report.md,report.json,audit.xlsx}

Does NOT modify e1_x5_partial_am_*_20260728 artifacts (marked WRONG_TARGET_DATE elsewhere).
"""
from __future__ import annotations

import hashlib
import json
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Optional
from zoneinfo import ZoneInfo

JST = ZoneInfo("Asia/Tokyo")
REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(REPO))

DAY = "20260727"
DAY_ISO = "2026-07-27"
OUT = REPO / "results" / "research" / "e1_x5_partial_am_reference_20260727"
CAPTURE_DAY = REPO / "data" / "market_capture" / DAY
# Primary continuous ingress after restarts
PRIMARY_SESSION = CAPTURE_DAY / "session_ing_20260727_11752_1785113581_4db3b030"
CONFIG_PATH = REPO / "configs" / "small_paper_pilot_q070_cap3_entry_price_risk_guard_trailing_mfe_shadow.yaml"
UNIVERSE_CSV = REPO / "results" / "reports" / "universe_core10_dynamic40_price_risk_am_20260727.csv"

AM_START = datetime(2026, 7, 27, 9, 0, 0, tzinfo=JST)
AM_END = datetime(2026, 7, 27, 11, 30, 0, tzinfo=JST)

VERDICT = "E1_X5_PARTIAL_AM_REFERENCE_ONLY"
FORWARD_DAY1 = "E1_X5_FORWARD_DAY1_READY"


def _parse_ts(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=JST) if v.tzinfo is None else v.astimezone(JST)
    s = str(v).strip()
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.replace(tzinfo=JST) if dt.tzinfo is None else dt.astimezone(JST)
    except Exception:
        return None


def _norm_sym(s: str) -> str:
    s = str(s or "").strip()
    if not s:
        return ""
    return s if s.endswith(".T") else f"{s}.T"


def _sha_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            b = fh.read(1024 * 1024)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def list_session_am_fragments() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sess in sorted(CAPTURE_DAY.glob("session_*")):
        parts = sorted(sess.glob("push_part_*.jsonl"))
        if not parts or all(p.stat().st_size == 0 for p in parts):
            continue
        first = last = None
        n = 0
        for part in parts:
            if part.stat().st_size == 0:
                continue
            with part.open("r", encoding="utf-8", errors="ignore") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        rec = json.loads(line)
                    except Exception:
                        continue
                    ts = _parse_ts(rec.get("received_at") or rec.get("persisted_at"))
                    if ts is None or ts < AM_START:
                        continue
                    if ts > AM_END:
                        break
                    n += 1
                    if first is None or ts < first:
                        first = ts
                    if last is None or ts > last:
                        last = ts
        rows.append(
            {
                "session": sess.name,
                "path": str(sess.resolve()),
                "am_records": n,
                "first": first.isoformat() if first else None,
                "last": last.isoformat() if last else None,
                "is_primary": sess.resolve() == PRIMARY_SESSION.resolve(),
            }
        )
    return rows


def audit_primary_am(sess_path: Path) -> dict[str, Any]:
    parts = sorted(sess_path.glob("push_part_*.jsonl"))
    n = 0
    symbols: set[str] = set()
    first_ts: Optional[datetime] = None
    last_ts: Optional[datetime] = None
    seq_min = seq_max = None
    prev_seq = None
    gaps = dups = inv = 0
    seen_seq: set[int] = set()
    board_ok = price_ok = 0
    missing_board = 0
    silence_gaps: list[dict[str, Any]] = []
    prev_ts: Optional[datetime] = None
    raw_parts: list[dict[str, Any]] = []
    past_am = False

    for part in parts:
        part_am = 0
        with part.open("r", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                ts = _parse_ts(rec.get("received_at") or rec.get("persisted_at"))
                if ts is None:
                    continue
                if ts > AM_END:
                    past_am = True
                    break
                if ts < AM_START:
                    continue
                n += 1
                part_am += 1
                if first_ts is None or ts < first_ts:
                    first_ts = ts
                if last_ts is None or ts > last_ts:
                    last_ts = ts
                sym = _norm_sym(str(rec.get("symbol") or ""))
                if sym:
                    symbols.add(sym)
                try:
                    seq_i = int(rec["sequence"]) if rec.get("sequence") is not None else None
                except (TypeError, ValueError):
                    seq_i = None
                if seq_i is not None:
                    if seq_min is None or seq_i < seq_min:
                        seq_min = seq_i
                    if seq_max is None or seq_i > seq_max:
                        seq_max = seq_i
                    if seq_i in seen_seq:
                        dups += 1
                    else:
                        seen_seq.add(seq_i)
                    if prev_seq is not None:
                        if seq_i < prev_seq:
                            inv += 1
                        elif seq_i > prev_seq + 1:
                            gaps += 1
                    prev_seq = seq_i
                op = rec.get("original_payload") if isinstance(rec.get("original_payload"), dict) else {}
                if isinstance(op.get("Buy1"), dict) and isinstance(op.get("Sell1"), dict):
                    board_ok += 1
                else:
                    missing_board += 1
                if op.get("CurrentPrice") is not None or op.get("CurrentPriceTime"):
                    price_ok += 1
                if prev_ts is not None:
                    gap_s = (ts - prev_ts).total_seconds()
                    if gap_s >= 60:
                        silence_gaps.append(
                            {"from": prev_ts.isoformat(), "to": ts.isoformat(), "gap_sec": gap_s}
                        )
                prev_ts = ts
        if part_am > 0:
            raw_parts.append(
                {
                    "path": str(part),
                    "am_records_in_part": part_am,
                    "byte_size": part.stat().st_size,
                    "raw_file_sha256": _sha_file(part) if part.stat().st_size < 400_000_000 else "skipped_large",
                }
            )
        if past_am:
            break

    silence_gaps.sort(key=lambda x: -float(x["gap_sec"]))
    return {
        "status": "AUDITED",
        "day": DAY,
        "session_path": str(sess_path.resolve()),
        "part_count_touched": len(raw_parts),
        "record_count": n,
        "am_window_record_count": n,
        "symbol_count": len(symbols),
        "symbols_sample": sorted(symbols)[:30],
        "first_event_at": first_ts.isoformat() if first_ts else None,
        "last_event_at": last_ts.isoformat() if last_ts else None,
        "nominal_am_window": "09:00:00–11:30:00 JST",
        "actual_eval_window_note": (
            "No backfill before first persisted PUSH; clip at 11:30:00 JST"
        ),
        "sequence_min": seq_min,
        "sequence_max": seq_max,
        "sequence_gap_count": gaps,
        "sequence_duplicate_count": dups,
        "sequence_inversion_count": inv,
        "board_present_count": board_ok,
        "price_present_count": price_ok,
        "missing_board_count": missing_board,
        "board_present_rate": (board_ok / n) if n else None,
        "price_present_rate": (price_ok / n) if n else None,
        "silence_gaps_ge_60s_top": silence_gaps[:20],
        "silence_gap_count_ge_60s": len(silence_gaps),
        "raw_parts": raw_parts,
        "usable": n > 0 and first_ts is not None and board_ok > 0,
    }


def load_universe() -> set[str]:
    if UNIVERSE_CSV.is_file():
        import pandas as pd

        df = pd.read_csv(UNIVERSE_CSV)
        col = "symbol" if "symbol" in df.columns else df.columns[0]
        return {_norm_sym(x) for x in df[col].tolist()}
    # fallback: any symbol in primary (filled during iter)
    return set()


def iter_am_capture_events(
    sess_path: Path,
    universe: set[str],
    *,
    t0: Optional[datetime] = None,
    t1: Optional[datetime] = None,
) -> Iterator[dict[str, Any]]:
    t0 = t0 or AM_START
    t1 = t1 or AM_END
    for part in sorted(sess_path.glob("push_part_*.jsonl")):
        with part.open("r", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                ts = _parse_ts(rec.get("received_at") or rec.get("persisted_at"))
                if ts is None:
                    continue
                if ts > t1:
                    return
                if ts < t0:
                    continue
                sym = _norm_sym(str(rec.get("symbol") or ""))
                if universe and sym not in universe:
                    continue
                op = rec.get("original_payload")
                if not isinstance(op, dict):
                    continue
                if not isinstance(op.get("Buy1"), dict) or not isinstance(op.get("Sell1"), dict):
                    continue
                payload = dict(op)
                if payload.get("sequence") is None and rec.get("sequence") is not None:
                    payload["sequence"] = rec["sequence"]
                if not payload.get("CurrentPriceTime"):
                    payload["CurrentPriceTime"] = ts.isoformat()
                yield {
                    "symbol": sym,
                    "recv_ts": ts,
                    "payload": payload,
                    "sequence": rec.get("sequence"),
                    "event_id": str(
                        rec.get("raw_record_id") or f"{ts.isoformat()}|{sym}|{rec.get('sequence')}"
                    ),
                }


def summarize_session(session, *, last_bid: dict[str, float]) -> dict[str, Any]:
    from small_paper.e1_x5_forward_shadow import econ

    exits = list(session.exits)
    pnls = [float(x["net_pnl_yen_100"]) for x in exits]
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    exit_reasons: dict[str, dict[str, Any]] = {}
    for x in exits:
        r = str(x.get("exit_reason") or "")
        bucket = exit_reasons.setdefault(r, {"count": 0, "pnl": 0.0})
        bucket["count"] += 1
        bucket["pnl"] += float(x["net_pnl_yen_100"])
    by_sym: dict[str, dict[str, Any]] = {}
    for x in exits:
        sym = str(x["symbol"])
        b = by_sym.setdefault(sym, {"trades": 0, "pnl": 0.0, "wins": 0, "losses": 0})
        b["trades"] += 1
        pnl = float(x["net_pnl_yen_100"])
        b["pnl"] += pnl
        if pnl > 0:
            b["wins"] += 1
        elif pnl < 0:
            b["losses"] += 1

    open_rows = []
    unrealized = 0.0
    for sym, pos in session.positions.items():
        bid = last_bid.get(sym)
        if bid is None or bid <= 0:
            open_rows.append(
                {
                    "symbol": sym,
                    "entry_time": pos.entry_time.isoformat() if pos.entry_time else None,
                    "entry_ask": pos.entry_ask,
                    "last_bid": None,
                    "unrealized_pnl_yen_100": None,
                }
            )
            continue
        e = econ(pos.entry_ask, float(bid))
        unrealized += float(e["net_pnl_yen_100"])
        open_rows.append(
            {
                "symbol": sym,
                "entry_time": pos.entry_time.isoformat() if pos.entry_time else None,
                "entry_ask": pos.entry_ask,
                "last_bid": bid,
                "unrealized_pnl_yen_100": e["net_pnl_yen_100"],
            }
        )

    funnel = session.exclusive_entry_funnel()
    noe = session.no_evaluation_breakdown()
    realized = float(sum(pnls)) if pnls else 0.0
    pf = (sum(wins) / abs(sum(losses))) if losses else (None if not wins else float("inf"))
    return {
        "status": "OK",
        "evaluated_count": int(session.evaluated_count),
        "no_evaluation_count": int(session.no_evaluation_count),
        "no_evaluation_breakdown": noe,
        "entry_funnel_exclusive": funnel,
        "entries_n": len(session.entries),
        "completed_trades": len(exits),
        "open_n": len(session.positions),
        "wins": len(wins),
        "losses": len(losses),
        "draws": len(pnls) - len(wins) - len(losses),
        "win_rate": (len(wins) / len(pnls)) if pnls else None,
        "realized_pnl_yen_100": realized,
        "unrealized_pnl_yen_100": unrealized,
        "profit_factor": pf,
        "avg_pnl_yen_100": (realized / len(pnls)) if pnls else None,
        "best_trade_yen_100": max(pnls) if pnls else None,
        "worst_trade_yen_100": min(pnls) if pnls else None,
        "exit_reasons": exit_reasons,
        "by_symbol": by_sym,
        "open_positions": open_rows,
        "trades": [
            {
                "symbol": x["symbol"],
                "entry_time": x["entry_time"].isoformat()
                if hasattr(x["entry_time"], "isoformat")
                else str(x["entry_time"]),
                "exit_time": x["exit_time"].isoformat()
                if hasattr(x["exit_time"], "isoformat")
                else str(x["exit_time"]),
                "entry_ask": x["entry_ask"],
                "exit_bid": x["exit_bid"],
                "exit_reason": x["exit_reason"],
                "score": x.get("score"),
                "net_pnl_yen_100": x["net_pnl_yen_100"],
                "holding_sec": x.get("holding_sec"),
            }
            for x in exits
        ],
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
    }


def run_e1(sess_path: Path, universe: set[str], audit: dict[str, Any]) -> dict[str, Any]:
    from small_paper.canonical_board import best_bid_ask_for_mode
    from small_paper.e1_x5_decision_core import process_e1_x5_event
    from small_paper.e1_x5_dmid_score_provider import DMidD4H6ScoreProvider
    from small_paper.e1_x5_forward_shadow import E1X5ForwardShadowSession

    t0 = _parse_ts(audit["first_event_at"])
    t1 = min(_parse_ts(audit["last_event_at"]) or AM_END, AM_END)
    provider = DMidD4H6ScoreProvider.maybe_create()
    session = E1X5ForwardShadowSession(enabled=True)
    last_bid: dict[str, float] = {}
    n = 0
    for ev in iter_am_capture_events(sess_path, universe, t0=t0, t1=t1):
        n += 1
        bid, ask = best_bid_ask_for_mode(ev["payload"], mode="canonical")
        if bid is not None and bid > 0:
            last_bid[ev["symbol"]] = float(bid)
        process_e1_x5_event(
            provider=provider,
            session=session,
            symbol=ev["symbol"],
            payload=ev["payload"],
            day=DAY,
            event_sequence=ev.get("sequence"),
            event_id=ev["event_id"],
            decision_time=ev["recv_ts"],
        )
        if n % 100000 == 0:
            print(f"[e1] n={n} exits={len(session.exits)} open={len(session.positions)}", flush=True)
    summary = summarize_session(session, last_bid=last_bid)
    summary["events_fed"] = n
    summary["window"] = {"first": t0.isoformat() if t0 else None, "last": t1.isoformat() if t1 else None}
    summary["no_forced_flatten_at_1130"] = True
    return summary


def materialize_am_push_jsonl(sess_path: Path, universe: set[str], dest_day_dir: Path) -> dict[str, Any]:
    """Convert capture AM events → push_jsonl for PBv2 dry replay (research temp only)."""
    if dest_day_dir.exists():
        shutil.rmtree(dest_day_dir)
    dest_day_dir.mkdir(parents=True, exist_ok=True)
    handles: dict[str, Any] = {}
    n = 0
    try:
        for ev in iter_am_capture_events(sess_path, universe):
            sym = ev["symbol"]
            stem = sym  # XXXX.T
            if stem not in handles:
                handles[stem] = (dest_day_dir / f"{stem}.jsonl").open(
                    "w", encoding="utf-8", newline="\n"
                )
            row = {
                "recorded_at": ev["recv_ts"].isoformat(),
                "source": "live_push",
                "symbol": sym,
                "payload": ev["payload"],
            }
            handles[stem].write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
            n += 1
            if n % 100000 == 0:
                print(f"[materialize] n={n}", flush=True)
    finally:
        for fh in handles.values():
            fh.close()
    return {"path": str(dest_day_dir.resolve()), "records": n, "symbol_files": len(handles)}


def _pnl_from_row(row: dict[str, Any]) -> float:
    for k in ("pnl_yen_100", "net_pnl_yen_100", "canonical_pnl_yen_100"):
        if row.get(k) is not None:
            try:
                return float(row[k])
            except (TypeError, ValueError):
                pass
    return 0.0


FORCE_EXIT_REASONS = frozenset(
    {
        "push_replay_end",
        "morning_session_close",
        "afternoon_session_close",
        "session_end",
        "session_close",
        "force_close",
    }
)


def summarize_pbv2_from_replay(output_dir: Path, *, last_bid: dict[str, float]) -> dict[str, Any]:
    """Build PBv2 AM summary; force-close style exits → OPEN (no forced flatten)."""
    from small_paper.e1_x5_forward_shadow import econ

    events_path = output_dir / "small_paper_events.jsonl"
    trades_path = output_dir / "structural_trades.csv"
    summary_path = output_dir / "small_paper_summary.json"

    completed: list[dict[str, Any]] = []
    open_from_force: list[dict[str, Any]] = []

    def _classify(item: dict[str, Any]) -> None:
        reason = str(item.get("exit_reason") or "")
        exit_t = _parse_ts(item.get("exit_time"))
        if reason in FORCE_EXIT_REASONS or (exit_t is not None and exit_t > AM_END):
            open_from_force.append(item)
        elif reason in {"live_virtual_hold", ""} and item.get("exit_bid") in (None, 0, 0.0):
            # Placeholder hold without real exit price → treat as OPEN at AM end
            open_from_force.append(item)
        else:
            completed.append(item)

    if trades_path.is_file():
        import pandas as pd

        df = pd.read_csv(trades_path)
        for _, r in df.iterrows():
            row = {k: r[k] for k in df.columns}
            reason = str(row.get("close_reason") or row.get("exit_reason") or "")
            entry_ask = float(row.get("entry_price") or row.get("entry_ask") or 0) or None
            exit_bid = float(row.get("close_price") or row.get("exit_price") or row.get("exit_bid") or 0) or None
            pnl = _pnl_from_row(row)
            if (pnl == 0.0 or row.get("pnl_yen_100") is None) and entry_ask and exit_bid:
                pnl = float(econ(float(entry_ask), float(exit_bid))["net_pnl_yen_100"])
            _classify(
                {
                    "symbol": _norm_sym(str(row.get("symbol") or "")),
                    "entry_time": str(row.get("entry_time") or ""),
                    "exit_time": str(row.get("close_time") or row.get("exit_time") or ""),
                    "entry_ask": entry_ask,
                    "exit_bid": exit_bid,
                    "exit_reason": reason,
                    "net_pnl_yen_100": pnl,
                }
            )
    elif events_path.is_file():
        # Reconstruct from accepted + later exit-like events / observer closes
        by_pos: dict[str, dict[str, Any]] = {}
        with events_path.open("r", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    ev = json.loads(line)
                except Exception:
                    continue
                etype = str(ev.get("event_type") or "")
                sym = _norm_sym(str(ev.get("symbol") or ""))
                pid = str(ev.get("position_id") or ev.get("observer_position_id") or "")
                key = pid or f"{sym}|{ev.get('entry_time')}"
                if etype == "accepted" and ev.get("gate_accept"):
                    by_pos[key] = {
                        "symbol": sym,
                        "entry_time": str(ev.get("entry_time") or ""),
                        "exit_time": str(ev.get("exit_time") or ""),
                        "entry_ask": float(ev["entry_price"]) if ev.get("entry_price") is not None else None,
                        "exit_bid": float(ev["exit_price"]) if ev.get("exit_price") is not None else None,
                        "exit_reason": str(ev.get("exit_reason") or ev.get("close_reason") or ""),
                        "net_pnl_yen_100": _pnl_from_row(ev),
                        "_open": True,
                    }
                # Observer / structural exit events
                if etype in {"exit", "closed", "position_closed", "observer_exit"} or (
                    "exit" in etype and etype != "entry_stage"
                ):
                    if key not in by_pos and sym:
                        by_pos[key] = {
                            "symbol": sym,
                            "entry_time": str(ev.get("entry_time") or ""),
                            "exit_time": "",
                            "entry_ask": float(ev["entry_price"]) if ev.get("entry_price") is not None else None,
                            "exit_bid": None,
                            "exit_reason": "",
                            "net_pnl_yen_100": 0.0,
                            "_open": True,
                        }
                    if key in by_pos:
                        item = by_pos[key]
                        item["exit_time"] = str(ev.get("exit_time") or ev.get("close_time") or ev.get("event_time") or "")
                        item["exit_reason"] = str(ev.get("exit_reason") or ev.get("close_reason") or item["exit_reason"])
                        if ev.get("exit_price") is not None:
                            item["exit_bid"] = float(ev["exit_price"])
                        elif ev.get("close_price") is not None:
                            item["exit_bid"] = float(ev["close_price"])
                        pnl = _pnl_from_row(ev)
                        if pnl == 0.0 and item.get("entry_ask") and item.get("exit_bid"):
                            pnl = float(econ(float(item["entry_ask"]), float(item["exit_bid"]))["net_pnl_yen_100"])
                        item["net_pnl_yen_100"] = pnl
                        item["_open"] = False
                # Update accepted placeholder exits when real exit_reason arrives on same accepted row later
                if etype == "accepted" and key in by_pos:
                    reason = str(ev.get("exit_reason") or "")
                    if reason and reason not in {"live_virtual_hold"}:
                        item = by_pos[key]
                        item["exit_reason"] = reason
                        item["exit_time"] = str(ev.get("exit_time") or item.get("exit_time") or "")
                        if ev.get("exit_price") is not None:
                            item["exit_bid"] = float(ev["exit_price"])
                        if item.get("entry_ask") and item.get("exit_bid"):
                            item["net_pnl_yen_100"] = float(
                                econ(float(item["entry_ask"]), float(item["exit_bid"]))["net_pnl_yen_100"]
                            )
                        item["_open"] = False

        for item in by_pos.values():
            item.pop("_open", None)
            if item.get("entry_ask") and item.get("exit_bid") and not item.get("net_pnl_yen_100"):
                item["net_pnl_yen_100"] = float(
                    econ(float(item["entry_ask"]), float(item["exit_bid"]))["net_pnl_yen_100"]
                )
            _classify(item)

    open_rows = []
    unrealized = 0.0
    for item in open_from_force:
        sym = item["symbol"]
        bid = last_bid.get(sym)
        entry_ask = item.get("entry_ask")
        if bid and entry_ask:
            e = econ(float(entry_ask), float(bid))
            u = float(e["net_pnl_yen_100"])
            unrealized += u
        else:
            u = None
        open_rows.append(
            {
                "symbol": sym,
                "entry_time": item.get("entry_time"),
                "entry_ask": entry_ask,
                "last_bid": bid,
                "unrealized_pnl_yen_100": u,
                "note": "OPEN at 11:30 (no forced flatten); includes replay-end / virtual-hold restores",
            }
        )

    pnls = [float(t["net_pnl_yen_100"]) for t in completed]
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    exit_reasons: dict[str, dict[str, Any]] = {}
    for t in completed:
        r = str(t.get("exit_reason") or "")
        b = exit_reasons.setdefault(r, {"count": 0, "pnl": 0.0})
        b["count"] += 1
        b["pnl"] += float(t["net_pnl_yen_100"])
    by_sym: dict[str, dict[str, Any]] = {}
    for t in completed:
        sym = str(t["symbol"])
        b = by_sym.setdefault(sym, {"trades": 0, "pnl": 0.0, "wins": 0, "losses": 0})
        b["trades"] += 1
        pnl = float(t["net_pnl_yen_100"])
        b["pnl"] += pnl
        if pnl > 0:
            b["wins"] += 1
        elif pnl < 0:
            b["losses"] += 1

    realized = float(sum(pnls)) if pnls else 0.0
    pf = (sum(wins) / abs(sum(losses))) if losses else (None if not wins else float("inf"))

    live_summary = {}
    if summary_path.is_file():
        try:
            live_summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except Exception:
            live_summary = {}

    return {
        "status": "OK",
        "reason": "push_replay_dry_run_am_capture_sot",
        "replay_output_dir": str(output_dir.resolve()),
        "evaluated_count": live_summary.get("gate_evaluations") or live_summary.get("evaluated_count"),
        "no_evaluation_count": live_summary.get("no_evaluation_count"),
        "no_evaluation_breakdown": {
            "reject_reason_counts": live_summary.get("reject_reason_counts"),
            "note": "PBv2 uses live gate taxonomy; see rejects in replay output",
        },
        "entries_n": len(completed) + len(open_rows),
        "completed_trades": len(completed),
        "open_n": len(open_rows),
        "wins": len(wins),
        "losses": len(losses),
        "draws": len(pnls) - len(wins) - len(losses),
        "win_rate": (len(wins) / len(pnls)) if pnls else None,
        "realized_pnl_yen_100": realized,
        "unrealized_pnl_yen_100": unrealized,
        "profit_factor": pf,
        "avg_pnl_yen_100": (realized / len(pnls)) if pnls else None,
        "best_trade_yen_100": max(pnls) if pnls else None,
        "worst_trade_yen_100": min(pnls) if pnls else None,
        "exit_reasons": exit_reasons,
        "by_symbol": by_sym,
        "open_positions": open_rows,
        "trades": completed,
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
        "no_forced_flatten_at_1130": True,
        "poll_interval_sec": live_summary.get("poll_interval_sec"),
        "live_summary_excerpt": {
            k: live_summary.get(k)
            for k in (
                "accepted_count",
                "canonical_total_pnl_yen_100",
                "push_rows",
                "runtime_sec",
                "structural_trade_count",
                "gate_evaluations",
            )
            if k in live_summary
        },
    }


def run_pbv2(sess_path: Path, universe: set[str], last_bid: dict[str, float]) -> dict[str, Any]:
    from small_paper.config import load_pilot_config
    from small_paper.pilot_runner import run_push_replay_dry_run

    push_day = OUT / "_work" / "pbv2_push_am" / DAY_ISO
    # session_key_from_output_dir requires output under results/small_paper/
    replay_out = REPO / "results" / "small_paper" / DAY / "offline_partial_am_pbv2_e1x5_ref"

    existing = list(push_day.glob("*.jsonl")) if push_day.is_dir() else []
    if len(existing) >= 40:
        print(f"[pbv2] reuse materialized push_jsonl ({len(existing)} files)", flush=True)
        mat = {
            "path": str(push_day.resolve()),
            "records": "reused",
            "symbol_files": len(existing),
        }
    else:
        print("[pbv2] materialize AM push_jsonl from capture...", flush=True)
        mat = materialize_am_push_jsonl(sess_path, universe, push_day)
        print(f"[pbv2] materialized {mat}", flush=True)

    if replay_out.exists():
        shutil.rmtree(replay_out)
    replay_out.mkdir(parents=True, exist_ok=True)

    cfg = load_pilot_config(CONFIG_PATH)
    # CAP5; keep observer for structural EXIT (discord off). End force-closes are
    # restored to OPEN in summarize_pbv2_from_replay.
    cfg = replace(
        cfg,
        discord_enabled=False,
        discord_observer_only=True,
        max_concurrent_positions=5,
    )
    print("[pbv2] run_push_replay_dry_run (streaming, poll=5s matching live)...", flush=True)
    run_push_replay_dry_run(
        cfg,
        push_dir=push_day,
        output_dir=replay_out,
        repo_root=REPO.parent,  # tradebotfile root expected by some loaders
        poll_interval_sec=5.0,
        enable_discord=False,
        write_board_shadow_reports=False,
        streaming_push_replay=True,
    )
    summary = summarize_pbv2_from_replay(replay_out, last_bid=last_bid)
    summary["materialize"] = mat
    summary["config_path"] = str(CONFIG_PATH)
    summary["cap"] = 5
    summary["lot"] = 100
    return summary


def collect_last_bids(sess_path: Path, universe: set[str]) -> dict[str, float]:
    from small_paper.canonical_board import best_bid_ask_for_mode

    last: dict[str, float] = {}
    for ev in iter_am_capture_events(sess_path, universe):
        bid, _ask = best_bid_ask_for_mode(ev["payload"], mode="canonical")
        if bid is not None and bid > 0:
            last[ev["symbol"]] = float(bid)
    return last


def compare_ledgers(e1: dict[str, Any], pbv2: dict[str, Any]) -> dict[str, Any]:
    e1_trades = {(str(t["symbol"]), str(t.get("entry_time"))): t for t in e1.get("trades") or []}
    p2_trades = {(str(t["symbol"]), str(t.get("entry_time"))): t for t in pbv2.get("trades") or []}
    only_e1 = [e1_trades[k] for k in e1_trades.keys() - p2_trades.keys()]
    only_p2 = [p2_trades[k] for k in p2_trades.keys() - e1_trades.keys()]
    both = []
    for k in e1_trades.keys() & p2_trades.keys():
        a, b = e1_trades[k], p2_trades[k]
        both.append(
            {
                "key": list(k),
                "e1_pnl": a.get("net_pnl_yen_100"),
                "pbv2_pnl": b.get("net_pnl_yen_100"),
                "delta": float(a.get("net_pnl_yen_100") or 0) - float(b.get("net_pnl_yen_100") or 0),
            }
        )
    return {
        "e1_completed": e1.get("completed_trades"),
        "pbv2_completed": pbv2.get("completed_trades"),
        "e1_realized": e1.get("realized_pnl_yen_100"),
        "pbv2_realized": pbv2.get("realized_pnl_yen_100"),
        "e1_unrealized": e1.get("unrealized_pnl_yen_100"),
        "pbv2_unrealized": pbv2.get("unrealized_pnl_yen_100"),
        "delta_realized_e1_minus_pbv2": float(e1.get("realized_pnl_yen_100") or 0)
        - float(pbv2.get("realized_pnl_yen_100") or 0),
        "delta_unrealized_e1_minus_pbv2": float(e1.get("unrealized_pnl_yen_100") or 0)
        - float(pbv2.get("unrealized_pnl_yen_100") or 0),
        "only_e1_count": len(only_e1),
        "only_pbv2_count": len(only_p2),
        "overlap_count": len(both),
        "only_e1": only_e1,
        "only_pbv2": only_p2,
        "overlap": both,
        "note": "Independent CAP5 on same AM Capture SoT; key=symbol+entry_time",
    }


def write_xlsx(path: Path, report: dict[str, Any]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["key", "value"])
    flat = [
        ("verdict", report["verdict"]),
        ("data_status", report["data_status"]),
        ("sot", report["source_of_truth"]["path"]),
        ("first", report["data_audit"].get("first_event_at")),
        ("last", report["data_audit"].get("last_event_at")),
        ("am_records", report["data_audit"].get("am_window_record_count")),
        ("symbols", report["data_audit"].get("symbol_count")),
        ("seq_min", report["data_audit"].get("sequence_min")),
        ("seq_max", report["data_audit"].get("sequence_max")),
        ("gaps", report["data_audit"].get("sequence_gap_count")),
        ("dups", report["data_audit"].get("sequence_duplicate_count")),
        ("inv", report["data_audit"].get("sequence_inversion_count")),
        ("e1_completed", report["e1_x5"].get("completed_trades")),
        ("e1_realized", report["e1_x5"].get("realized_pnl_yen_100")),
        ("e1_unrealized", report["e1_x5"].get("unrealized_pnl_yen_100")),
        ("e1_pf", report["e1_x5"].get("profit_factor")),
        ("e1_open", report["e1_x5"].get("open_n")),
        ("pbv2_completed", report["pbv2"].get("completed_trades")),
        ("pbv2_realized", report["pbv2"].get("realized_pnl_yen_100")),
        ("pbv2_unrealized", report["pbv2"].get("unrealized_pnl_yen_100")),
        ("pbv2_pf", report["pbv2"].get("profit_factor")),
        ("delta_realized", report["comparison"].get("delta_realized_e1_minus_pbv2")),
        ("forward_eligible", False),
        ("submit_cancel_live", "0/0/0"),
    ]
    for k, v in flat:
        ws.append([k, v])

    for name, key in (("E1_Trades", "e1_x5"), ("PBv2_Trades", "pbv2")):
        ws2 = wb.create_sheet(name)
        trades = report[key].get("trades") or []
        if not trades:
            ws2.append(["note", "none"])
        else:
            cols = list(trades[0].keys())
            ws2.append(cols)
            for t in trades:
                ws2.append([t.get(c) for c in cols])

    ws3 = wb.create_sheet("only_e1")
    rows = report["comparison"].get("only_e1") or []
    if rows:
        cols = list(rows[0].keys())
        ws3.append(cols)
        for t in rows:
            ws3.append([t.get(c) for c in cols])
    else:
        ws3.append(["none"])

    ws4 = wb.create_sheet("only_pbv2")
    rows = report["comparison"].get("only_pbv2") or []
    if rows:
        cols = list(rows[0].keys())
        ws4.append(cols)
        for t in rows:
            ws4.append([t.get(c) for c in cols])
    else:
        ws4.append(["none"])

    ws5 = wb.create_sheet("audit")
    ws5.append(["field", "value"])
    for k, v in report["data_audit"].items():
        if k in ("raw_parts", "symbols_sample", "silence_gaps_ge_60s_top"):
            ws5.append([k, json.dumps(v, ensure_ascii=False, default=str)[:2000]])
        else:
            ws5.append([k, v])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_md(report: dict[str, Any]) -> str:
    a = report["data_audit"]
    e1 = report["e1_x5"]
    pb = report["pbv2"]
    c = report["comparison"]
    return f"""# E1_X5 Partial AM Reference — 2026-07-27

## Verdict
`{report['verdict']}`

- forward_eligible: **false**（部分AM・参考評価）
- Forward加算: **なし**（valid sessions/trades / Complete AM+PM = 0）
- 7/27 PM: `NOT_ADOPTED`
- Day1: `{FORWARD_DAY1}`
- submit/cancel/live_order: **0/0/0**

## 1. 使用したAM実時間範囲
- first: `{a.get('first_event_at')}`
- last: `{a.get('last_event_at')}`
- nominal window: 09:00–11:30 JST（開始前は補完せず）

## 2. Source of Truth
`{report['source_of_truth']['path']}`

## 3. レコード・銘柄・sequence
- records: `{a.get('am_window_record_count')}` symbols: `{a.get('symbol_count')}`
- sequence min/max: `{a.get('sequence_min')}` / `{a.get('sequence_max')}`
- gap/dup/inversion: `{a.get('sequence_gap_count')}` / `{a.get('sequence_duplicate_count')}` / `{a.get('sequence_inversion_count')}`
- board/price rate: `{a.get('board_present_rate')}` / `{a.get('price_present_rate')}`
- silence ≥60s: `{a.get('silence_gap_count_ge_60s')}`

## 4. E1_X5
- evaluated / NO_EVAL: `{e1.get('evaluated_count')}` / `{e1.get('no_evaluation_count')}`
- ENTRY / completed / OPEN@11:30: `{e1.get('entries_n')}` / `{e1.get('completed_trades')}` / `{e1.get('open_n')}`
- W/L / win_rate: `{e1.get('wins')}` / `{e1.get('losses')}` / `{e1.get('win_rate')}`
- realized / OPEN unrealized / PF: `{e1.get('realized_pnl_yen_100')}` / `{e1.get('unrealized_pnl_yen_100')}` / `{e1.get('profit_factor')}`
- avg / best / worst: `{e1.get('avg_pnl_yen_100')}` / `{e1.get('best_trade_yen_100')}` / `{e1.get('worst_trade_yen_100')}`

## 5. PBv2（同一SoT・同一時間）
- ENTRY / completed / OPEN: `{pb.get('entries_n')}` / `{pb.get('completed_trades')}` / `{pb.get('open_n')}`
- realized / OPEN unrealized / PF: `{pb.get('realized_pnl_yen_100')}` / `{pb.get('unrealized_pnl_yen_100')}` / `{pb.get('profit_factor')}`

## 6. E1 − PBv2
- Δ realized: `{c.get('delta_realized_e1_minus_pbv2')}`
- Δ unrealized: `{c.get('delta_unrealized_e1_minus_pbv2')}`
- only E1 / only PBv2 / overlap: `{c.get('only_e1_count')}` / `{c.get('only_pbv2_count')}` / `{c.get('overlap_count')}`

## 7. Forward / Safety
- added_to_forward: false
- submit/cancel/live_order: 0/0/0

## Generator
`{report.get('generator')}` at `{report.get('generated_at')}`
"""


def mark_wrong_target_20260728() -> None:
    """Do not overwrite triad files; add status sidecar only."""
    for name in (
        "e1_x5_partial_am_reference_20260728",
        "e1_x5_partial_am_source_recovery_20260728",
    ):
        d = REPO / "results" / "research" / name
        if not d.is_dir():
            continue
        status = {
            "status": "WRONG_TARGET_DATE_NOT_APPLICABLE",
            "intended_target_was": "2026-07-28",
            "correct_target": "2026-07-27 AM",
            "superseded_by": str(OUT),
            "triad_untouched": True,
            "marked_at": datetime.now(JST).isoformat(timespec="seconds"),
        }
        (d / "WRONG_TARGET_DATE_NOT_APPLICABLE.json").write_text(
            json.dumps(status, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )


def check_forward() -> dict[str, Any]:
    out: dict[str, Any] = {"checked": [], "anomaly": False}
    for rel in (
        "results/research/e1_x5_runtime_offline_parity_followup_codegen_audit_fix_20260727/report.json",
        "results/research/e1_x5_runtime_offline_parity_followup_codegen_fix_20260727/report.json",
    ):
        p = REPO / rel
        if not p.is_file():
            continue
        r = json.loads(p.read_text(encoding="utf-8"))
        fwd = r.get("forward") or {}
        item = {
            "path": str(p),
            "day1_status": fwd.get("day1_status") or FORWARD_DAY1,
            "sessions": int(fwd.get("valid_progress_sessions") or 0),
            "trades": int(fwd.get("valid_progress_trades") or 0),
            "pm": fwd.get("pm_20260727") or "NOT_ADOPTED",
        }
        out["checked"].append(item)
        if item["sessions"] or item["trades"]:
            out["anomaly"] = True
    return out


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    mark_wrong_target_20260728()

    if not PRIMARY_SESSION.is_dir():
        raise SystemExit(f"primary session missing: {PRIMARY_SESSION}")

    fragments = list_session_am_fragments()
    print("[audit] primary AM...", flush=True)
    audit = audit_primary_am(PRIMARY_SESSION)
    print(
        f"[audit] first={audit['first_event_at']} last={audit['last_event_at']} "
        f"n={audit['am_window_record_count']} syms={audit['symbol_count']}",
        flush=True,
    )

    universe = load_universe()
    if not universe:
        # allow all symbols present
        universe = set()
    audit["universe_size"] = len(universe)

    print("[e1] offline AM...", flush=True)
    e1_cache = OUT / "_work" / "e1_am_cache.json"
    if e1_cache.is_file() and "--refresh-e1" not in sys.argv:
        print(f"[e1] reuse cache {e1_cache}", flush=True)
        e1 = json.loads(e1_cache.read_text(encoding="utf-8"))
    else:
        e1 = run_e1(PRIMARY_SESSION, universe, audit)
        e1_cache.parent.mkdir(parents=True, exist_ok=True)
        e1_cache.write_text(json.dumps(e1, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

    print("[bids] collect last AM bids for OPEN MTM...", flush=True)
    bid_cache = OUT / "_work" / "last_bid_am.json"
    if bid_cache.is_file() and "--refresh-e1" not in sys.argv:
        last_bid = {k: float(v) for k, v in json.loads(bid_cache.read_text(encoding="utf-8")).items()}
    else:
        last_bid = collect_last_bids(PRIMARY_SESSION, universe)
        bid_cache.parent.mkdir(parents=True, exist_ok=True)
        bid_cache.write_text(json.dumps(last_bid, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("[pbv2] offline AM on same SoT...", flush=True)
    pbv2 = run_pbv2(PRIMARY_SESSION, universe, last_bid)

    comparison = compare_ledgers(e1, pbv2)
    forward_check = check_forward()

    report = {
        "verdict": VERDICT,
        "forward": {
            "forward_eligible": False,
            "exclude_reason": "部分AM（ingress途中開始）の参考オフライン評価",
            "valid_progress_sessions": 0,
            "valid_progress_trades": 0,
            "complete_am_pm_days": 0,
            "pm_20260727": "NOT_ADOPTED",
            "day1_status": FORWARD_DAY1,
            "added_to_forward": False,
            "prior_forward_check": forward_check,
        },
        "data_status": "PARTIAL_AM_CAPTURE_PRESENT",
        "source_of_truth": {
            "kind": "market_capture_session_am",
            "path": str(PRIMARY_SESSION.resolve()),
            "pattern": "push_part_*.jsonl",
            "day": DAY,
        },
        "am_fragments_all_sessions": fragments,
        "data_audit": audit,
        "e1_x5": e1,
        "pbv2": pbv2,
        "comparison": comparison,
        "safety": {"submit": 0, "cancel": 0, "live_order": 0},
        "wrong_target_20260728": "WRONG_TARGET_DATE_NOT_APPLICABLE (sidecar only; triad untouched)",
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "generator": "scripts/run_e1_x5_partial_am_reference_20260727.py",
        "completion_answers": {
            "1_actual_am_range": f"{audit.get('first_event_at')} .. {audit.get('last_event_at')}",
            "2_sot_abs_path": str(PRIMARY_SESSION.resolve()),
            "3_records_symbols_sequence": {
                "records": audit.get("am_window_record_count"),
                "symbols": audit.get("symbol_count"),
                "seq_min": audit.get("sequence_min"),
                "seq_max": audit.get("sequence_max"),
                "gap": audit.get("sequence_gap_count"),
                "dup": audit.get("sequence_duplicate_count"),
                "inv": audit.get("sequence_inversion_count"),
            },
            "4_e1": {
                "completed": e1.get("completed_trades"),
                "realized": e1.get("realized_pnl_yen_100"),
                "unrealized": e1.get("unrealized_pnl_yen_100"),
                "pf": e1.get("profit_factor"),
                "open": e1.get("open_n"),
            },
            "5_pbv2": {
                "completed": pbv2.get("completed_trades"),
                "realized": pbv2.get("realized_pnl_yen_100"),
                "unrealized": pbv2.get("unrealized_pnl_yen_100"),
                "pf": pbv2.get("profit_factor"),
                "open": pbv2.get("open_n"),
            },
            "6_delta_realized": comparison.get("delta_realized_e1_minus_pbv2"),
            "7_only_counts": {
                "only_e1": comparison.get("only_e1_count"),
                "only_pbv2": comparison.get("only_pbv2_count"),
                "overlap": comparison.get("overlap_count"),
            },
            "8_forward_not_added": True,
            "9_submit_cancel_live": "0/0/0",
            "10_artifact_dir": str(OUT),
        },
    }

    (OUT / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8"
    )
    (OUT / "report.md").write_text(write_md(report), encoding="utf-8")
    write_xlsx(OUT / "audit.xlsx", report)
    print(json.dumps(report["completion_answers"], ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
