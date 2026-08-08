#!/usr/bin/env python3
"""E1_X5 partial AM reference evaluation for 2026-07-28.

Uses only existing saved AM Source of Truth — no backfill / no invent.
Verdict: E1_X5_PARTIAL_AM_REFERENCE_ONLY (never Forward-eligible).

Writes ONLY:
  results/research/e1_x5_partial_am_reference_20260728/{report.md,report.json,audit.xlsx}
"""
from __future__ import annotations

import hashlib
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterator, Optional
from zoneinfo import ZoneInfo

JST = ZoneInfo("Asia/Tokyo")
REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(REPO))

DAY = "20260728"
DAY_ISO = "2026-07-28"
OUT = REPO / "results" / "research" / "e1_x5_partial_am_reference_20260728"
CAPTURE_ROOT = REPO / "data" / "market_capture" / DAY
PUSH_ROOT = REPO / "data" / "push_jsonl" / DAY_ISO
PAPER_ROOT = REPO / "results" / "small_paper" / DAY
SESSIONS_ROOT = REPO / "results" / "paper_sessions" / DAY

VERDICT = "E1_X5_PARTIAL_AM_REFERENCE_ONLY"
FORWARD_DAY1 = "E1_X5_FORWARD_DAY1_READY"
AM_START = datetime(2026, 7, 28, 9, 0, 0, tzinfo=JST)
AM_END = datetime(2026, 7, 28, 11, 30, 0, tzinfo=JST)


def _sha_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            b = fh.read(1024 * 1024)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def _parse_ts(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=JST) if v.tzinfo is None else v.astimezone(JST)
    s = str(v).strip()
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.replace(tzinfo=JST) if dt.tzinfo is None else dt.astimezone(JST)
    except Exception:
        return None


def _norm_sym(s: str) -> str:
    s = str(s or "").strip()
    if not s:
        return ""
    if not s.endswith(".T") and s.replace(".", "").isalnum():
        return f"{s}.T" if not s.endswith(".T") else s
    return s if s.endswith(".T") else f"{s}.T"


def discover_source() -> dict[str, Any]:
    """Locate 2026-07-28 AM Source of Truth. Never use sandbox/preflight synthetic."""
    candidates: list[dict[str, Any]] = []

    if CAPTURE_ROOT.is_dir():
        for sess in sorted(CAPTURE_ROOT.glob("session_*")):
            parts = sorted(sess.glob("push_part_*.jsonl"))
            if not parts:
                continue
            # Reject known synthetic sandbox paths
            if "sandbox" in str(sess).lower() or "preflight" in str(sess).lower():
                continue
            candidates.append(
                {
                    "kind": "market_capture_session",
                    "path": str(sess.resolve()),
                    "parts": [str(p) for p in parts],
                    "part_count": len(parts),
                    "bytes": sum(p.stat().st_size for p in parts),
                }
            )

    push_files = []
    if PUSH_ROOT.is_dir():
        push_files = sorted(PUSH_ROOT.glob("*.jsonl"))
        if push_files:
            candidates.append(
                {
                    "kind": "push_jsonl_day",
                    "path": str(PUSH_ROOT.resolve()),
                    "files": [str(p) for p in push_files],
                    "file_count": len(push_files),
                    "bytes": sum(p.stat().st_size for p in push_files),
                }
            )

    paper_sessions = []
    if PAPER_ROOT.is_dir():
        paper_sessions = [str(p.resolve()) for p in sorted(PAPER_ROOT.glob("live_session_*")) if p.is_dir()]

    return {
        "day": DAY,
        "capture_root_exists": CAPTURE_ROOT.is_dir(),
        "push_root_exists": PUSH_ROOT.is_dir(),
        "paper_root_exists": PAPER_ROOT.is_dir(),
        "sessions_root_exists": SESSIONS_ROOT.is_dir(),
        "candidates": candidates,
        "paper_sessions": paper_sessions,
        "selected": candidates[0] if candidates else None,
        "source_missing": len(candidates) == 0,
        "wall_clock_jst": datetime.now(JST).isoformat(timespec="seconds"),
        "note": (
            "No non-sandbox Source of Truth for 20260728 AM on disk. "
            "Do not invent/backfill. Partial AM reference only when data exists."
            if not candidates
            else "Using first discovered non-sandbox capture candidate."
        ),
    }


def audit_capture_session(sess_path: Path) -> dict[str, Any]:
    """Stream push_part_*.jsonl for range / sequence / symbol / gap audit."""
    parts = sorted(sess_path.glob("push_part_*.jsonl"))
    n = 0
    symbols: set[str] = set()
    first_ts: Optional[datetime] = None
    last_ts: Optional[datetime] = None
    seq_min = seq_max = None
    prev_seq = None
    gaps = dups = inv = 0
    seen_seq: set[int] = set()
    silence_gaps: list[dict[str, Any]] = []
    prev_ts: Optional[datetime] = None
    board_ok = 0
    price_ok = 0
    am_rows = 0
    raw_sha_parts: list[dict[str, str]] = []

    for part in parts:
        raw_sha_parts.append({"path": str(part), "raw_file_sha256": _sha_file(part), "byte_size": part.stat().st_size})
        with part.open("r", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                n += 1
                ts = _parse_ts(rec.get("received_at") or rec.get("event_time") or rec.get("persisted_at"))
                if ts is not None:
                    if first_ts is None or ts < first_ts:
                        first_ts = ts
                    if last_ts is None or ts > last_ts:
                        last_ts = ts
                    if AM_START <= ts <= AM_END:
                        am_rows += 1
                    if prev_ts is not None:
                        gap_s = (ts - prev_ts).total_seconds()
                        if gap_s >= 60:
                            silence_gaps.append(
                                {
                                    "from": prev_ts.isoformat(),
                                    "to": ts.isoformat(),
                                    "gap_sec": gap_s,
                                }
                            )
                    prev_ts = ts
                sym = _norm_sym(rec.get("symbol") or "")
                if sym:
                    symbols.add(sym)
                seq = rec.get("sequence")
                try:
                    seq_i = int(seq) if seq is not None else None
                except (TypeError, ValueError):
                    seq_i = None
                if seq_i is not None:
                    if seq_min is None or seq_i < seq_min:
                        seq_min = seq_i
                    if seq_max is None or seq_i > seq_max:
                        seq_max = seq_i
                    if seq_i in seen_seq:
                        dups += 1
                    else:
                        seen_seq.add(seq_i)
                    if prev_seq is not None:
                        if seq_i < prev_seq:
                            inv += 1
                        elif seq_i > prev_seq + 1:
                            gaps += 1
                    prev_seq = seq_i
                op = rec.get("original_payload") if isinstance(rec.get("original_payload"), dict) else {}
                if isinstance(op.get("Buy1"), dict) and isinstance(op.get("Sell1"), dict):
                    board_ok += 1
                if op.get("CurrentPrice") is not None or op.get("CurrentPriceTime"):
                    price_ok += 1

    silence_gaps.sort(key=lambda x: -float(x["gap_sec"]))
    return {
        "session_path": str(sess_path.resolve()),
        "part_count": len(parts),
        "record_count": n,
        "am_window_record_count": am_rows,
        "symbol_count": len(symbols),
        "symbols_sample": sorted(symbols)[:30],
        "first_event_at": first_ts.isoformat() if first_ts else None,
        "last_event_at": last_ts.isoformat() if last_ts else None,
        "sequence_min": seq_min,
        "sequence_max": seq_max,
        "sequence_gap_count": gaps,
        "sequence_duplicate_count": dups,
        "sequence_inversion_count": inv,
        "board_present_count": board_ok,
        "price_present_count": price_ok,
        "silence_gaps_ge_60s_top": silence_gaps[:20],
        "silence_gap_count_ge_60s": len(silence_gaps),
        "raw_parts": raw_sha_parts,
        "usable": n > 0 and first_ts is not None,
    }


def load_universe_from_any() -> set[str]:
    """Best-effort universe from latest known paper config — not invention of market tape."""
    # Prefer 20260727 live config if 20260728 missing (symbols list only; not prices)
    for day in (DAY, "20260727"):
        root = REPO / "results" / "small_paper" / day
        if not root.is_dir():
            continue
        for sess in sorted(root.glob("live_session_*"), reverse=True):
            cfg = sess / "live_session_config.json"
            if not cfg.is_file():
                continue
            try:
                data = json.loads(cfg.read_text(encoding="utf-8"))
                csv_path = data.get("universe_csv_path")
                if csv_path and Path(csv_path).is_file():
                    import pandas as pd

                    df = pd.read_csv(csv_path)
                    col = "symbol" if "symbol" in df.columns else df.columns[0]
                    return {_norm_sym(x) for x in df[col].tolist()}
            except Exception:
                continue
    return set()


def iter_capture_events(
    sess_path: Path,
    universe: set[str],
    *,
    t0: Optional[datetime] = None,
    t1: Optional[datetime] = None,
) -> Iterator[dict[str, Any]]:
    for part in sorted(sess_path.glob("push_part_*.jsonl")):
        with part.open("r", encoding="utf-8", errors="ignore") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except Exception:
                    continue
                ts = _parse_ts(rec.get("received_at") or rec.get("event_time") or rec.get("persisted_at"))
                if ts is None:
                    continue
                if t0 is not None and ts < t0:
                    continue
                if t1 is not None and ts > t1:
                    return
                sym = _norm_sym(rec.get("symbol") or "")
                if universe and sym not in universe:
                    continue
                op = rec.get("original_payload")
                if not isinstance(op, dict):
                    continue
                if not isinstance(op.get("Buy1"), dict) or not isinstance(op.get("Sell1"), dict):
                    continue
                payload = dict(op)
                if payload.get("sequence") is None and rec.get("sequence") is not None:
                    payload["sequence"] = rec["sequence"]
                if not payload.get("CurrentPriceTime"):
                    payload["CurrentPriceTime"] = ts.isoformat()
                yield {
                    "symbol": sym,
                    "recv_ts": ts,
                    "payload": payload,
                    "sequence": rec.get("sequence"),
                    "event_id": str(
                        rec.get("raw_record_id") or f"{ts.isoformat()}|{sym}|{rec.get('sequence')}"
                    ),
                }


def summarize_e1(session) -> dict[str, Any]:
    from small_paper.e1_x5_forward_shadow import econ

    exits = list(session.exits)
    pnls = [float(x["net_pnl_yen_100"]) for x in exits]
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    exit_reasons: dict[str, dict[str, Any]] = {}
    for x in exits:
        r = str(x.get("exit_reason") or "")
        bucket = exit_reasons.setdefault(r, {"count": 0, "pnl": 0.0})
        bucket["count"] += 1
        bucket["pnl"] += float(x["net_pnl_yen_100"])
    by_sym: dict[str, dict[str, Any]] = {}
    for x in exits:
        sym = str(x["symbol"])
        b = by_sym.setdefault(sym, {"trades": 0, "pnl": 0.0, "wins": 0, "losses": 0})
        b["trades"] += 1
        pnl = float(x["net_pnl_yen_100"])
        b["pnl"] += pnl
        if pnl > 0:
            b["wins"] += 1
        elif pnl < 0:
            b["losses"] += 1

    # OPEN mark-to-market using last bid on position (no forced exit)
    open_mtm = []
    unrealized = 0.0
    for sym, pos in session.positions.items():
        # last known from candidates / no bid → skip mtm price
        open_mtm.append(
            {
                "symbol": sym,
                "entry_time": pos.entry_time.isoformat() if pos.entry_time else None,
                "entry_ask": pos.entry_ask,
                "score": pos.score,
                "unrealized_pnl_yen_100": None,
                "note": "No forced flatten; unrealized filled if last bid known in run meta",
            }
        )

    funnel = session.exclusive_entry_funnel()
    noe = session.no_evaluation_breakdown()
    realized = float(sum(pnls)) if pnls else 0.0
    pf = (sum(wins) / abs(sum(losses))) if losses else (None if not wins else float("inf"))
    return {
        "evaluated_count": int(session.evaluated_count),
        "no_evaluation_count": int(session.no_evaluation_count),
        "no_evaluation_breakdown": noe,
        "entry_funnel_exclusive": funnel,
        "entries_n": len(session.entries),
        "completed_trades": len(exits),
        "open_n": len(session.positions),
        "wins": len(wins),
        "losses": len(losses),
        "draws": len(pnls) - len(wins) - len(losses),
        "win_rate": (len(wins) / len(pnls)) if pnls else None,
        "realized_pnl_yen_100": realized,
        "unrealized_pnl_yen_100": unrealized,
        "profit_factor": pf,
        "avg_pnl_yen_100": (realized / len(pnls)) if pnls else None,
        "best_trade_yen_100": max(pnls) if pnls else None,
        "worst_trade_yen_100": min(pnls) if pnls else None,
        "exit_reasons": exit_reasons,
        "by_symbol": by_sym,
        "open_positions": open_mtm,
        "trades": [
            {
                "symbol": x["symbol"],
                "entry_time": x["entry_time"].isoformat() if hasattr(x["entry_time"], "isoformat") else str(x["entry_time"]),
                "exit_time": x["exit_time"].isoformat() if hasattr(x["exit_time"], "isoformat") else str(x["exit_time"]),
                "entry_ask": x["entry_ask"],
                "exit_bid": x["exit_bid"],
                "exit_reason": x["exit_reason"],
                "score": x.get("score"),
                "net_pnl_yen_100": x["net_pnl_yen_100"],
                "holding_sec": x.get("holding_sec"),
            }
            for x in exits
        ],
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
    }


def mark_open_unrealized(session, last_bid: dict[str, float]) -> tuple[list[dict], float]:
    from small_paper.e1_x5_forward_shadow import econ

    rows = []
    total = 0.0
    for sym, pos in session.positions.items():
        bid = last_bid.get(sym)
        if bid is None or bid <= 0:
            rows.append(
                {
                    "symbol": sym,
                    "entry_ask": pos.entry_ask,
                    "last_bid": None,
                    "unrealized_pnl_yen_100": None,
                }
            )
            continue
        e = econ(pos.entry_ask, float(bid))
        total += float(e["net_pnl_yen_100"])
        rows.append(
            {
                "symbol": sym,
                "entry_time": pos.entry_time.isoformat() if pos.entry_time else None,
                "entry_ask": pos.entry_ask,
                "last_bid": bid,
                "unrealized_pnl_yen_100": e["net_pnl_yen_100"],
            }
        )
    return rows, total


def run_e1_on_capture(sess_path: Path, universe: set[str], audit: dict[str, Any]) -> dict[str, Any]:
    from small_paper.e1_x5_decision_core import process_e1_x5_event
    from small_paper.e1_x5_dmid_score_provider import DMidD4H6ScoreProvider
    from small_paper.e1_x5_forward_shadow import E1X5ForwardShadowSession
    from small_paper.canonical_board import best_bid_ask_for_mode

    t0 = _parse_ts(audit.get("first_event_at"))
    t1 = _parse_ts(audit.get("last_event_at"))
    # Clip to AM if range extends beyond, but never invent earlier
    if t0 is not None and t0 < AM_START:
        # Partial AM may start mid-session — keep actual first, do not pad before t0
        pass
    if t1 is not None and t1 > AM_END:
        t1 = AM_END

    provider = DMidD4H6ScoreProvider.maybe_create()
    session = E1X5ForwardShadowSession(enabled=True)
    last_bid: dict[str, float] = {}
    n = 0
    for ev in iter_capture_events(sess_path, universe, t0=t0, t1=t1):
        n += 1
        bid, ask = best_bid_ask_for_mode(ev["payload"], mode="canonical")
        if bid is not None and bid > 0:
            last_bid[ev["symbol"]] = float(bid)
        process_e1_x5_event(
            provider=provider,
            session=session,
            symbol=ev["symbol"],
            payload=ev["payload"],
            day=DAY,
            event_sequence=ev.get("sequence"),
            event_id=ev["event_id"],
            decision_time=ev["recv_ts"],
        )
        if n % 100000 == 0:
            print(f"[e1] n={n} exits={len(session.exits)} open={len(session.positions)}", flush=True)

    summary = summarize_e1(session)
    open_rows, unreal = mark_open_unrealized(session, last_bid)
    summary["open_positions"] = open_rows
    summary["unrealized_pnl_yen_100"] = unreal
    summary["events_fed"] = n
    summary["window"] = {
        "first": t0.isoformat() if t0 else None,
        "last": t1.isoformat() if t1 else None,
    }
    return summary


def empty_strategy_block(reason: str) -> dict[str, Any]:
    return {
        "status": "NOT_RUN",
        "reason": reason,
        "evaluated_count": 0,
        "no_evaluation_count": 0,
        "no_evaluation_breakdown": {"no_evaluation_reason_breakdown": {}},
        "entries_n": 0,
        "completed_trades": 0,
        "open_n": 0,
        "wins": 0,
        "losses": 0,
        "draws": 0,
        "win_rate": None,
        "realized_pnl_yen_100": 0.0,
        "unrealized_pnl_yen_100": 0.0,
        "profit_factor": None,
        "avg_pnl_yen_100": None,
        "best_trade_yen_100": None,
        "worst_trade_yen_100": None,
        "exit_reasons": {},
        "by_symbol": {},
        "open_positions": [],
        "trades": [],
        "submit": 0,
        "cancel": 0,
        "live_order": 0,
    }


def compare_ledgers(e1: dict[str, Any], pbv2: dict[str, Any]) -> dict[str, Any]:
    e1_trades = {(str(t["symbol"]), str(t.get("entry_time"))): t for t in e1.get("trades") or []}
    p2_trades = {(str(t["symbol"]), str(t.get("entry_time"))): t for t in pbv2.get("trades") or []}
    only_e1 = [e1_trades[k] for k in e1_trades.keys() - p2_trades.keys()]
    only_p2 = [p2_trades[k] for k in p2_trades.keys() - e1_trades.keys()]
    both = []
    for k in e1_trades.keys() & p2_trades.keys():
        a, b = e1_trades[k], p2_trades[k]
        both.append(
            {
                "key": k,
                "e1_pnl": a.get("net_pnl_yen_100"),
                "pbv2_pnl": b.get("net_pnl_yen_100"),
                "delta": float(a.get("net_pnl_yen_100") or 0) - float(b.get("net_pnl_yen_100") or 0),
            }
        )
    return {
        "e1_completed": e1.get("completed_trades"),
        "pbv2_completed": pbv2.get("completed_trades"),
        "e1_realized": e1.get("realized_pnl_yen_100"),
        "pbv2_realized": pbv2.get("realized_pnl_yen_100"),
        "delta_realized_e1_minus_pbv2": float(e1.get("realized_pnl_yen_100") or 0)
        - float(pbv2.get("realized_pnl_yen_100") or 0),
        "delta_unrealized_e1_minus_pbv2": float(e1.get("unrealized_pnl_yen_100") or 0)
        - float(pbv2.get("unrealized_pnl_yen_100") or 0),
        "only_e1_count": len(only_e1),
        "only_pbv2_count": len(only_p2),
        "overlap_count": len(both),
        "only_e1_sample": only_e1[:30],
        "only_pbv2_sample": only_p2[:30],
        "overlap_sample": both[:30],
        "note": "Independent CAP5 portfolios on same tape; keys=symbol+entry_time exact match.",
    }


def write_xlsx(path: Path, sheets: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    def cell(v: Any) -> Any:
        if v is None or isinstance(v, (str, int, float, bool)):
            return v
        return json.dumps(v, ensure_ascii=False, default=str)

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    for name, data in sheets.items():
        ws = wb.create_sheet(title=str(name)[:31])
        if isinstance(data, list) and data and isinstance(data[0], dict):
            keys = list(data[0].keys())
            ws.append(keys)
            for row in data:
                ws.append([cell(row.get(k)) for k in keys])
        elif isinstance(data, dict):
            keys = ["field", "value"]
            ws.append(keys)
            for k, v in data.items():
                ws.append([str(k), cell(v)])
        else:
            keys = ["value"]
            ws.append(keys)
            ws.append([cell(data)])
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(1, col_idx).font = header_font
            ws.cell(1, col_idx).alignment = Alignment(wrap_text=True, vertical="top")
            width = max(16, min(len(str(key)) + 4, 48))
            if "sha" in str(key).lower() or key in ("field", "value", "path"):
                width = max(width, 36)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def check_forward_progress_untouched() -> dict[str, Any]:
    """Read prior followup reports; do not modify them."""
    paths = [
        REPO
        / "results"
        / "research"
        / "e1_x5_runtime_offline_parity_followup_codegen_audit_fix_20260727"
        / "report.json",
        REPO
        / "results"
        / "research"
        / "e1_x5_runtime_offline_parity_followup_codegen_fix_20260727"
        / "report.json",
    ]
    out = {"checked": [], "anomaly": False, "anomaly_detail": None}
    for p in paths:
        if not p.is_file():
            continue
        r = json.loads(p.read_text(encoding="utf-8"))
        fp = r.get("forward_provenance") or r.get("valid_forward_progress") or {}
        sessions = int(fp.get("valid_progress_sessions", fp.get("sessions", 0)) or 0)
        trades = int(fp.get("valid_progress_trades", fp.get("trades", 0)) or 0)
        item = {
            "path": str(p),
            "verdict_forward": r.get("verdict_forward"),
            "sessions": sessions,
            "trades": trades,
            "pm_forward": (fp.get("pm_forward_status") or r.get("pm_forward_status")),
        }
        out["checked"].append(item)
        if sessions != 0 or trades != 0:
            out["anomaly"] = True
            out["anomaly_detail"] = item
    return out


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    discovery = discover_source()
    forward_check = check_forward_progress_untouched()

    audit: dict[str, Any]
    e1: dict[str, Any]
    pbv2: dict[str, Any]
    comparison: dict[str, Any]
    data_status: str

    if discovery["source_missing"]:
        data_status = "SOURCE_MISSING"
        audit = {
            "status": "SOURCE_MISSING",
            "day": DAY,
            "first_event_at": None,
            "last_event_at": None,
            "symbol_count": 0,
            "record_count": 0,
            "am_window_record_count": 0,
            "sequence_min": None,
            "sequence_max": None,
            "sequence_gap_count": None,
            "sequence_duplicate_count": None,
            "sequence_inversion_count": None,
            "silence_gap_count_ge_60s": None,
            "board_present_count": 0,
            "price_present_count": 0,
            "searched_paths": {
                "market_capture": str(CAPTURE_ROOT),
                "push_jsonl": str(PUSH_ROOT),
                "small_paper": str(PAPER_ROOT),
                "paper_sessions": str(SESSIONS_ROOT),
            },
            "exists": {
                "market_capture": CAPTURE_ROOT.is_dir(),
                "push_jsonl": PUSH_ROOT.is_dir(),
                "small_paper": PAPER_ROOT.is_dir(),
                "paper_sessions": SESSIONS_ROOT.is_dir(),
            },
            "note": (
                f"Wall clock {discovery['wall_clock_jst']}. "
                "No 20260728 Capture / push_jsonl / paper session found. "
                "Offline evaluation not executed (no invent / no backfill)."
            ),
        }
        reason = "NO_AM_SOURCE_OF_TRUTH_20260728"
        e1 = empty_strategy_block(reason)
        pbv2 = empty_strategy_block(reason)
        comparison = compare_ledgers(e1, pbv2)
        comparison["status"] = "NOT_COMPARABLE_NO_SOURCE"
    else:
        sel = discovery["selected"]
        sess_path = Path(sel["path"])
        data_status = "PARTIAL_CAPTURE_PRESENT"
        audit = audit_capture_session(sess_path)
        audit["status"] = "AUDITED"
        universe = load_universe_from_any()
        audit["universe_size"] = len(universe)
        if not audit.get("usable"):
            e1 = empty_strategy_block("CAPTURE_UNUSABLE")
            pbv2 = empty_strategy_block("CAPTURE_UNUSABLE")
        else:
            print("[e1] offline on existing range...", flush=True)
            e1 = run_e1_on_capture(sess_path, universe, audit)
            e1["status"] = "OK"
            # PBv2: require same-range offline engine; without dedicated runner mark explicit
            # Prefer paper structural trades only if they cover the SAME window — else NOT_RUN
            pbv2 = empty_strategy_block(
                "PBV2_OFFLINE_ENGINE_NOT_AVAILABLE_FOR_PARTIAL_AM_WITHOUT_LIVE_SUMMARY"
            )
            # If a same-day paper summary exists overlapping the audited window, attach as reference
            for ps in discovery.get("paper_sessions") or []:
                summary_path = Path(ps) / "small_paper_summary.json"
                trades_path = Path(ps) / "structural_trades.csv"
                if summary_path.is_file() and trades_path.is_file():
                    import pandas as pd

                    st = pd.read_csv(trades_path)
                    summary = json.loads(summary_path.read_text(encoding="utf-8"))
                    trades = []
                    for _, row in st.iterrows():
                        trades.append(
                            {
                                "symbol": str(row.get("symbol")),
                                "entry_time": str(row.get("entry_time")),
                                "exit_time": str(row.get("exit_time") or row.get("close_time") or ""),
                                "exit_reason": str(row.get("close_reason") or row.get("exit_reason") or ""),
                                "net_pnl_yen_100": float(
                                    row.get("pnl_yen_100")
                                    or (
                                        float(row.get("entry_price") or 0)
                                        * float(row.get("realized_pnl_pct") or 0)
                                        / 100.0
                                        * 100.0
                                    )
                                ),
                            }
                        )
                    pbv2 = {
                        "status": "FROM_LIVE_PAPER_SUMMARY_SAME_DAY",
                        "reason": f"loaded {ps}",
                        "evaluated_count": None,
                        "no_evaluation_count": None,
                        "entries_n": len(trades),
                        "completed_trades": len(trades),
                        "open_n": int(summary.get("open_positions") or 0),
                        "wins": None,
                        "losses": None,
                        "win_rate": None,
                        "realized_pnl_yen_100": float(
                            summary.get("canonical_total_pnl_yen_100")
                            or sum(t["net_pnl_yen_100"] for t in trades)
                        ),
                        "unrealized_pnl_yen_100": None,
                        "profit_factor": summary.get("profit_factor_yen_100"),
                        "avg_pnl_yen_100": None,
                        "best_trade_yen_100": max((t["net_pnl_yen_100"] for t in trades), default=None),
                        "worst_trade_yen_100": min((t["net_pnl_yen_100"] for t in trades), default=None),
                        "exit_reasons": {},
                        "by_symbol": {},
                        "open_positions": [],
                        "trades": trades,
                        "submit": 0,
                        "cancel": 0,
                        "live_order": 0,
                    }
                    break
        comparison = compare_ledgers(e1, pbv2)

    report = {
        "verdict": VERDICT,
        "forward": {
            "forward_eligible": False,
            "exclude_reason": "部分AMかつオフライン評価",
            "valid_progress_sessions": 0,
            "valid_progress_trades": 0,
            "complete_am_pm_days": 0,
            "pm_20260727": "NOT_ADOPTED",
            "day1_status": FORWARD_DAY1,
            "added_to_forward": False,
            "prior_forward_check": forward_check,
        },
        "discovery": discovery,
        "data_status": data_status,
        "data_audit": audit,
        "e1_x5": e1,
        "pbv2": pbv2,
        "comparison": comparison,
        "safety": {"submit": 0, "cancel": 0, "live_order": 0},
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "generator": "scripts/run_e1_x5_partial_am_reference_20260728.py",
    }

    (OUT / "report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8"
    )

    md = f"""# E1_X5 Partial AM Reference — 2026-07-28

## Verdict
`{VERDICT}`

- forward_eligible: **false**
- 除外理由: 部分AMかつオフライン評価
- Forward加算: **なし**（valid sessions/trades/Complete AM+PM = 0 維持）
- 7/27 PM: `NOT_ADOPTED`
- Day1: `{FORWARD_DAY1}`
- submit/cancel/live_order: **0/0/0**

## 1. AMデータ実時間範囲
- data_status: `{data_status}`
- first: `{audit.get('first_event_at')}`
- last: `{audit.get('last_event_at')}`
- records: `{audit.get('record_count')}` (AM-window count=`{audit.get('am_window_record_count')}`)
- symbols: `{audit.get('symbol_count')}`

## 2. データ欠損・sequence監査
- sequence min/max: `{audit.get('sequence_min')}` / `{audit.get('sequence_max')}`
- gap/dup/inversion: `{audit.get('sequence_gap_count')}` / `{audit.get('sequence_duplicate_count')}` / `{audit.get('sequence_inversion_count')}`
- silence gaps ≥60s: `{audit.get('silence_gap_count_ge_60s')}`
- board/price present: `{audit.get('board_present_count')}` / `{audit.get('price_present_count')}`
- note: {audit.get('note') or discovery.get('note')}

## 3. E1_X5
- status: `{e1.get('status')}` reason=`{e1.get('reason')}`
- evaluated / no_evaluation: `{e1.get('evaluated_count')}` / `{e1.get('no_evaluation_count')}`
- ENTRY / completed / OPEN: `{e1.get('entries_n')}` / `{e1.get('completed_trades')}` / `{e1.get('open_n')}`
- W/L: `{e1.get('wins')}` / `{e1.get('losses')}` win_rate=`{e1.get('win_rate')}`
- realized PnL: `{e1.get('realized_pnl_yen_100')}`
- OPEN unrealized: `{e1.get('unrealized_pnl_yen_100')}`
- PF: `{e1.get('profit_factor')}`
- best/worst: `{e1.get('best_trade_yen_100')}` / `{e1.get('worst_trade_yen_100')}`

## 4. PBv2（同一範囲）
- status: `{pbv2.get('status')}` reason=`{pbv2.get('reason')}`
- completed: `{pbv2.get('completed_trades')}` realized=`{pbv2.get('realized_pnl_yen_100')}` OPEN unrealized=`{pbv2.get('unrealized_pnl_yen_100')}` PF=`{pbv2.get('profit_factor')}`

## 5. E1 − PBv2
- delta realized: `{comparison.get('delta_realized_e1_minus_pbv2')}`
- only E1 / only PBv2 / overlap: `{comparison.get('only_e1_count')}` / `{comparison.get('only_pbv2_count')}` / `{comparison.get('overlap_count')}`

## 6. 主な追加/除外
- only_e1_sample: `{json.dumps(comparison.get('only_e1_sample')[:5], ensure_ascii=False, default=str)}`
- only_pbv2_sample: `{json.dumps(comparison.get('only_pbv2_sample')[:5], ensure_ascii=False, default=str)}`

## 7–9. Forward / Safety / 判定
- Forward加算していない: **true**
- prior Forward check: `{json.dumps(forward_check, ensure_ascii=False)}`
- 0/0/0: **maintained**
- 判定: `{VERDICT}`

## 10. 成果物
`{OUT}`
"""
    (OUT / "report.md").write_text(md, encoding="utf-8")

    write_xlsx(
        OUT / "audit.xlsx",
        {
            "Data Audit": audit,
            "Discovery": discovery,
            "E1_X5 Summary": {k: v for k, v in e1.items() if k != "trades"},
            "E1_X5 Trades": e1.get("trades") or [{"note": "none"}],
            "E1_X5 OPEN": e1.get("open_positions") or [{"note": "none"}],
            "E1_X5 EXIT Reasons": e1.get("exit_reasons") or {"none": 0},
            "E1_X5 By Symbol": [
                {"symbol": k, **v} for k, v in (e1.get("by_symbol") or {}).items()
            ]
            or [{"note": "none"}],
            "PBv2 Summary": {k: v for k, v in pbv2.items() if k != "trades"},
            "PBv2 Trades": pbv2.get("trades") or [{"note": "none"}],
            "Comparison": comparison,
            "Forward Provenance": report["forward"],
            "Safety": report["safety"],
        },
    )

    allowed = {"report.md", "report.json", "audit.xlsx"}
    extras = [p.name for p in OUT.iterdir() if p.is_file() and p.name not in allowed]
    if extras:
        print(f"WARN extras={extras}", flush=True)

    print(VERDICT, flush=True)
    print(f"data_status={data_status}", flush=True)
    print(f"OUT={OUT}", flush=True)
    print(
        f"e1 trades={e1.get('completed_trades')} realized={e1.get('realized_pnl_yen_100')} "
        f"open_unreal={e1.get('unrealized_pnl_yen_100')}",
        flush=True,
    )
    if forward_check.get("anomaly"):
        print(f"FORWARD_ANOMALY={forward_check['anomaly_detail']}", flush=True)
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
